# This module builds and formats the Service State Page workbook (BP-2.0).
#
# The build*() methods below are unchanged business logic, transcribed from
# the root-level ServicePage.py: each pulls a table out of the ratebook data
# (via buildDataFrame's nesting waterfall) and shapes it into the DataFrame
# the rate page needs, filtered to the Service program (Class_Code_Min ==
# 70000). Same porting pattern as [[bop_hab_autoservice_port]] and
# [[bop_retail_port]]: the [[bop-nesting-order]] fix (lower-level company ->
# NGIC -> CW, not NGIC-first) and the per-company base-rate fix (root
# hardcoded buildBaseRates('NGIC') for every company's BRNACO/BRNAFF/BRNICOF
# tab — same bug already found and fixed in Hab/Auto Service/Retail).
#
# All Excel formatting (fonts, column widths, sub-header labels, page setup)
# lives in ExcelSettingsBOP.py, driven by "BOP/BOP Input File.xlsx". Most of
# Service's tables reuse table codes/layout profiles already shared with Hab,
# Auto Service and Retail (CBG/CPP/YBBG/YBPP/EBB/PDLD/LL/DO/DONM/ERP/PLUS/FR/
# AS_BR/ET/GLO/DCEQ/DC/LS_RETAIL) since the column widths and number formats
# transcribed from the root file's format*() methods matched those exactly —
# including Service's own Liability Size of Risk, whose sub-header text
# ("Building plus Business Personal Property") and widths match Retail's
# LS_RETAIL profile exactly. Service-only tables (FU/FUMP/BB) get new layout
# keys. The Dry Cleaners EXTRA Factor (DC) row grouping and the Directors &
# Officers 2-band merge (Service only has "Under 51"/"51 or More", unlike
# Hab's 5 bands) reuse the exact same post-processing as Retail's. Service
# adds four more BP-2.0-only endorsements with no Retail equivalent: Repair
# Services Specialized Endorsement (RSS, a merged single-cell "table"), Pet
# Services Specialized Endorsement (PSS, a two-table sheet like Retail's PSS),
# Pet Services Professional Liability (PSPL, which inserts a blank row 3 —
# transcribed as-is even though it looks like a cosmetic quirk in the
# original tool), and Mobile Pet and Veterinarian Services Endorsement (MPVS,
# a SIX-table sheet — the largest reconstruction of this kind so far). Like
# Retail's PSS, MPVS's underlying root-level generateWorksheet6tables/
# ExcelSettings.py module is NOT present in this repo, so _formatMPVS
# reconstructs the same semantic layout (label, header, data, repeated 6x
# with a blank separator row between blocks) from scratch using
# ws.max_row-relative appends — see that method's docstring for how the
# exact row positions were derived from the root file's insert_rows() calls.

from copy import copy

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import ExcelSettingsBOP

_THIN_BORDER = Border(
    left=Side(border_style='thin', color='C1C1C1'),
    right=Side(border_style='thin', color='C1C1C1'),
    top=Side(border_style='thin', color='C1C1C1'),
    bottom=Side(border_style='thin', color='C1C1C1'),
)


class Service:
    # Tables buildBaseRates() needs present for a given company before it can
    # be built without a KeyError.
    _BASE_RATE_TABLES = (
        'BP7_Peril_Building_Base_Rates',
        'BP7_Peril_BPP_Base_Rates',
        'BP7_Peril_Liability_Base_Rates',
    )

    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        # Individual program pages (unlike All Programs) also show the
        # "AllPeril" row/column — see [[bop_auto_service_allperil_row_fix]].
        self.perils = list(perils) + ['allperil']
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective
        self.rEffective = rEffective

        self.serviceProgramCode = 70000

    # Builds a dataframe for the given table code
    # The hierarchy matches Business Auto: lower-level company (NACO/NAFF/NICOF)
    # first, then NGIC (state-level default), then CW as the country-wide
    # fallback. See [[bop-nesting-order]] — the root-level ServicePage.py
    # checked NGIC first, which is backwards.
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if 'NACO' in self.rateTables.keys():
            if tableCode in self.rateTables['NACO'].keys():
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys():
            if tableCode in self.rateTables['NAFF'].keys():
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys():
            if tableCode in self.rateTables['NICOF'].keys():
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if tableCode in self.rateTables['NGIC'].keys():
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0])

    # Builds the dry cleaners EXTRA factor table
    # Returns a dataframe
    def buildDryCleanersFactor(self):
        dryCleanersFactor = self.buildDataFrame("BP7_Dry_Cleaners_Extra_Factor")
        filteredDryCleanersFactor = dryCleanersFactor.query(f'Class_Code_Min == {self.serviceProgramCode}')
        data = {'TierLimit': ['BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits',
                    'Tier2', 'Tier2', 'Tier2', 'Tier2', 'Tier2',
                    'Tier3', 'Tier3', 'Tier3', 'Tier3', 'Tier3',
                    'Tier4', 'Tier4', 'Tier4', 'Tier4', 'Tier4'],
                'Coverage': ['Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets'],
                'Limits (Per Item / Per Occurrence)': ['$1,000/ALS', '$1,000/$10,000', '$1,000/$5,000', '$1,000/$5,000', '$1,000/$5,000',
                    '$2,000/ALS', '$2,000/$15,000', '$2,000/$6,000', '$2,000/$6,000', '$2,000/$6,000',
                    '$3,000/ALS', '$3,000/$20,000', '$3,000/$9,000', '$3,000/$9,000', '$3,000/$9,000',
                    '$5,000/ALS', '$5,000/$25,000', '$5,000/$10,000', '$5,000/$10,000', '$5,000/$10,000']}
        dryCleaners = pd.DataFrame(data)
        finalDryCleaners = pd.merge(dryCleaners, filteredDryCleanersFactor, how='left', on=['TierLimit'])
        return finalDryCleaners.rename(columns={'TierLimit': 'Tier', 'DryCleanerExtraFactor': 'Factor'}).filter(items=['Tier', 'Coverage', 'Limits (Per Item / Per Occurrence)', 'Factor']). \
                replace({'Tier': {'BaseLimits': 'Base', 'Tier2': 'Tier 2', 'Tier3': 'Tier 3', 'Tier4': 'Tier 4'}})

    # Builds the dry cleaners EXTRA earthquake factor table
    # Returns a dataframe
    def buildDryCleanersEQFactor(self):
        miscFactors = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        dryCleanersEQFactor = miscFactors.query(f'FactorName == "DryCleanersExtraEQ"')
        return dryCleanersEQFactor.filter(items=['Factor']).rename(columns={'Factor': 'Dry Cleaners EXTRA EQ Factor'})

    # Builds the funeral home EXTRA endorsement table
    # Returns a dataframe
    def buildFuneralHomeEndorsement(self):
        funeralHomeEndorsement = self.buildDataFrame("BP7_FuneralExtraBaseRate")
        funeralHomeEndorsement['Occurence'] = funeralHomeEndorsement['LiabilityLimitOccurence'].apply(lambda x: "${0:,.0f}".format(x))
        funeralHomeEndorsement['Aggregate'] = funeralHomeEndorsement['LiabilityLimtAggregate'].apply(lambda x: "${0:,.0f}".format(x))
        funeralHomeEndorsement['Occurence / Aggregate'] = funeralHomeEndorsement['Occurence'] + ' / ' + funeralHomeEndorsement['Aggregate']
        pivotedFuneralHome = funeralHomeEndorsement.pivot(index=['LiabilityLimitOccurence', 'Occurence / Aggregate'], columns='IncrementalDescedents', values='FuneralExtraBaseRate'). \
                rename(columns={100: 'First 100 decedents', 200: 'Next 200 decedents', 300: 'Next 300 decedents', 400: 'Next 400 decedents', 1000: 'Over 1,000 decedents'}). \
                reset_index(['LiabilityLimitOccurence', 'Occurence / Aggregate']).sort_values(by=['LiabilityLimitOccurence'])
        del pivotedFuneralHome['LiabilityLimitOccurence']
        return pivotedFuneralHome

    # Builds the funeral home EXTRA endorsement minimum premium
    # Returns a dataframe
    def buildFuneralHomeMinPrem(self):
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        funeralHomeMinPrem = miscMinMaxPrem.query(f'CoverageType == "BP7FuneralDirectorsProflLiab"')
        return funeralHomeMinPrem.filter(items=['Premium'])

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building",
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction'). \
                    drop(columns=['L-Products', 'L-Violence', 'L-OtherMed', 'L-OtherPrem', 'WF', 'NC-BINC'], errors='ignore').rename(columns={'L-SlipFall': 'LIAB-Other'})
        elif coverage.casefold() == 'bpp':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction'). \
                    drop(columns=['L-Products', 'L-Violence', 'L-OtherMed', 'L-OtherPrem', 'WF', 'NC-BINC'], errors='ignore').rename(columns={'L-SlipFall': 'LIAB-Other'})

    # Builds the exclude theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'].isin(['L-Products', 'NC-BINC', 'WF'])].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'})
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop(columns=['L-Products', 'WF', 'NC-BINC'], errors='ignore')
        elif coverage.casefold() == 'bpp':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop(columns=['L-Products', 'WF', 'NC-BINC'], errors='ignore')

    # Builds the equipment breakdown base rate table
    # Returns a dataframe
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.serviceProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.serviceProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance': 'int32'})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_BLDG_BPP_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCodeMin == {self.serviceProgramCode}').rename(columns={'BldgBPPLimit_Min': 'Limit Min', 'BldgBPPLimit_Max': 'Limit Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Limit Min', 'Limit Max'], columns='Peril TypeCode', values='Factor').reset_index(['Limit Min', 'Limit Max']).fillna({'Limit Max': 'and over'})

    # Builds the general liability occupancy modifiers table
    # Returns a dataframe
    def buildGeneralOccupancyMod(self):
        generalOccupancyModifier = self.buildDataFrame("BP7_Peril_Occupant_Factor")
        filteredGeneralOccupancyMod = generalOccupancyModifier.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"'). \
                rename(columns={'OccupancyType': 'Occupancy', 'BLDGOccupantFactor': 'Building', 'BPPOccupantFactor': 'Business Personal Property'})
        return filteredGeneralOccupancyMod.replace({'Occupancy': {'Condominium': 'Condo Unit-owner', 'buildingOwnerLessorsrisk': "Lessor's Risk", 'buildingOwnerOccupant': 'Owner Occupant', 'tenant': 'Tenant'}}). \
                filter(items=['Occupancy', 'Building', 'Business Personal Property'])

    # Builds the directors and officers liability insurance table
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Service"').copy()
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 or More'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])

    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Service"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Service"').copy()
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index=filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])

    # Builds the barber or beauty shops professional liability table
    # Returns a dataframe
    def buildBarberProfLiab(self):
        barberProfLiab = self.buildDataFrame("BP7_ProfLiabarbersBeauticians_Rate")
        barberProfLiab['Occurence'] = barberProfLiab['LiabilityLimit'].apply(lambda x: "${0:,.0f}".format(x))
        barberProfLiab['Aggregate'] = barberProfLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        barberProfLiab['Occurence / Aggregate'] = barberProfLiab['Occurence'] + ' / ' + barberProfLiab['Aggregate']
        pivotedBarberProf = barberProfLiab.pivot(index=['LiabilityLimit', 'Occurence / Aggregate'], columns='ProfessionType', values='BaseRate').reset_index(['LiabilityLimit', 'Occurence / Aggregate']). \
                rename(columns={'Barber': 'Each Barber', 'Beautician': 'Each Beautician', 'Manicurist': 'Each Manicurist'}).sort_values(by=['LiabilityLimit'])
        del pivotedBarberProf['LiabilityLimit']
        return pivotedBarberProf

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.serviceProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Service premises'})

    # Builds the franchise upgrade endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.serviceProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(),
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(),
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Builds the Repair Services Specialized Endorsement table (flat premium,
    # hardcoded in the root tool rather than pulled from the ratebook)
    # Returns a dataframe
    def buildRepairSpecializedEndorsement(self):
        return pd.DataFrame({"Base premium for each Service premises": ["$300.00"]})

    # Builds the table for Pet Services Specialized Endorsement (base premium,
    # also hardcoded in the root tool)
    # Returns a dataframe
    def buildPetSpecializedEndorsement(self):
        return pd.DataFrame({"Base premium for each Service premises": ["$212.00"]})

    # Builds the Pet Services - Business Income table (the second, appended
    # block of the PSS sheet — see _formatPSSplzdEndo)
    # Returns a dataframe
    def buildServiceBusinessIncome(self):
        data = [("$25,000", "$13", "$7"), ("$50,000", "$20", "$13"), ("$100,000", "$26", "$20")]
        return pd.DataFrame(data, columns=["Limit", "1st Worker", "Each Addl"])

    # Builds the table for Pet Services Professional Liability
    # Returns a dataframe
    def buildPetServicePL(self):
        return pd.DataFrame({
            "Limits": ["$300,000", "$500,000", "$1,000,000", "$2,000,000"],
            "Rate": ["$43", "$56", "$68", "$83"],
        })

    # Builds the Pet Services block of the Mobile Pet and Veterinarian
    # Services Endorsement sheet (MPVS) — Returns a dataframe
    def buildPetServices(self):
        data = [("$15,000", "$49"), ("$25,000", "$85"), ("$50,000", "$166"), ("$100,000", "$220")]
        return pd.DataFrame(data, columns=["Limit", "Mobile Equipment"])

    # Builds the Pet Services per Customized Vehicle block of MPVS
    # Returns a dataframe
    def buildPetServicesCustomized(self):
        data = [("$25,000", "$91", "$46"), ("$50,000", "$104", "$59"), ("$100,000", "$117", "$71")]
        return pd.DataFrame(data, columns=["Limit", "1st Vehicle", "Each Addl"])

    # Builds the Veterinarian block of MPVS
    # Returns a dataframe
    def buildVet(self):
        data = [("$15,000", "$122"), ("$25,000", "$211"), ("$50,000", "$414"), ("$100,000", "$549")]
        return pd.DataFrame(data, columns=["Limits", "Mobile Equipment"])

    # Builds the Veterinarian Services per Customized Vehicle block of MPVS
    # Returns a dataframe
    def buildVetCustom(self):
        data = [("$25,000", "$227", "$113"), ("$50,000", "$260", "$146"), ("$100,000", "$293", "$179")]
        return pd.DataFrame(data, columns=["Limits", "1st Vehicle", "Each Addl"])

    # Builds the Veterinarian Services - Business Income block of MPVS
    # Returns a dataframe
    def buildVetBusinessIncome(self):
        data = [("$25,000", "$32", "$17"), ("$50,000", "$49", "$32"), ("$100,000", "$65", "$49")]
        return pd.DataFrame(data, columns=["Limit", "1st Worker", "Each Addl"])

    # Merges the "Number of Units" column of the D&O table into its 2 bands
    # ("Under 51" / "51 or More" — Service only has 2, unlike Hab's 5).
    # Handled here directly (rather than through the generic Sub Headers
    # config) because it merges DATA rows, not a header row.
    def _formatDirsOfficersLiabIns(self, ws):
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Groups the Dry Cleaners EXTRA Factor table's Tier (col A) and Factor
    # (col D) columns into 5-row blocks, one per tier — the data is always
    # exactly 20 rows (4 tiers x 5 coverage types) in this fixed order, so a
    # static merge is safe.
    def _formatDryCleanersFactor(self, ws):
        for start in (4, 9, 14, 19):
            end = start + 4
            ws.merge_cells(f'A{start}:A{end}')
            ws.merge_cells(f'D{start}:D{end}')
            ws[f'A{start}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws[f'D{start}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # The Repair Services Specialized Endorsement table is a single premium
    # value in column A, visually widened to span columns A-C (matching the
    # root tool's formatRepairSpecializedEndorsement) via a header-row and
    # data-row merge.
    def _formatRepairSpecializedEndorsement(self, ws):
        ws.merge_cells('A3:C3')
        ws.merge_cells('A4:C4')

    # Appends the Pet Services - Business Income table below the base
    # premium table on the same sheet, with its own bolded section label,
    # and widens the base premium table's header/data rows to span columns
    # A-C (same merge as _formatRepairSpecializedEndorsement). There's no
    # direct BOP equivalent of the root tool's generateWorksheet2tables (that
    # module isn't present in this repo), so this reconstructs the same
    # two-block layout by hand, same pattern as [[bop_retail_port]]'s PSS
    # sheet: the first table (buildPetSpecializedEndorsement, written
    # normally by generateWorksheet) occupies rows 3-4; this appends a
    # bolded "Pet Services - Business Income" label plus the second table's
    # header/data rows below it.
    def _formatPSSplzdEndo(self, ws, boldFont, biDf):
        ws.merge_cells('A3:C3')
        ws.merge_cells('A4:C4')
        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=1, value="Pet Services - Business Income")
        for col in range(1, len(biDf.columns) + 1):
            ws.cell(row=start_row + 1, column=col, value=biDf.columns[col - 1])
        for r, (_, row) in enumerate(biDf.iterrows(), start=start_row + 2):
            for col, val in enumerate(row, start=1):
                ws.cell(row=r, column=col, value=val)
        for row in (start_row, start_row + 1):
            for cell in ws[row]:
                cell.font = boldFont
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    # Mobile Pet and Veterinarian Services Endorsement (MPVS) — a SIX-table
    # sheet, the largest reconstruction of this kind in the BOP module.
    # generateWorksheet is called with an EMPTY dataframe for this table code
    # (just the title in A1), and this method builds the entire body from
    # row 3 down: 6 blocks, each a bolded section label + bolded column
    # header + plain data rows, separated by one blank row.
    #
    # The exact row positions below (label rows 3/10/16/22/29/35, blank
    # separator rows 9/15/21/28/34) were derived by simulating the root
    # tool's formatMPVS(), which inserts blank rows at those literal numbers
    # (each insert_rows() shifts everything below it down by one, so by the
    # time the 4th/5th/6th calls run they land past where their literal
    # argument alone would suggest) into a sheet where all 6 tables had
    # already been written back-to-back with no gaps by the (not present in
    # this repo) generateWorksheet6tables. Reconstructing forward from row 3
    # with one blank row between each block reproduces those exact same
    # positions — see [[bop_retail_port]]'s _formatPSSplzdEndo for the same
    # "no real ratebook to test against" caveat; this one is unverified
    # against a real reference even more than that one was, given its size,
    # so flag to the user to check the actual PDF output for S Table 4.H.
    # the first time real data is available. Column C's width is also left
    # unset, same as the root tool (it only ever sets A and B) — a likely
    # cosmetic gap in the original tool, kept as-is for fidelity.
    def _formatMPVS(self, ws, boldFont):
        blocks = [
            ("Pet Services", self.buildPetServices()),
            ("Pet Services per Customized Vehicle", self.buildPetServicesCustomized()),
            ("Pet Services - Business Income", self.buildServiceBusinessIncome()),
            ("Veterinarian", self.buildVet()),
            ("Veterinarian Services per Customized Vehicle", self.buildVetCustom()),
            ("Veterinarian Services - Business Income", self.buildVetBusinessIncome()),
        ]
        row = ws.max_row + 1
        for i, (label, df) in enumerate(blocks):
            if i > 0:
                row += 1  # blank separator row, left untouched (no border)
            label_row = row
            header_row = label_row + 1
            ws.cell(row=label_row, column=1, value=label)
            for col, name in enumerate(df.columns, start=1):
                ws.cell(row=header_row, column=col, value=name)
            for r_off, (_, data_row) in enumerate(df.iterrows()):
                for col, val in enumerate(data_row, start=1):
                    ws.cell(row=header_row + 1 + r_off, column=col, value=val)
            for r in (label_row, header_row):
                for cell in ws[r]:
                    cell.font = boldFont
                    cell.border = _THIN_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            row = header_row + len(df)

    # Sets up the Service Excel file and creates a separate worksheet for
    # each of the given dataframes. progress_callback (optional) is called
    # with a short message before each sheet is built.
    # Returns the Excel workbook
    def buildServicePage(self, progress_callback=None):
        companies = [c for c in self.rateTables.keys() if c != 'CW']

        Service = ExcelSettingsBOP.Excel(state=self.state, programName='Service', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        sheetSpecs = [
            ('DC', 'S Table 1.A.5.c. Dry Cleaners EXTRA Factor', self.buildDryCleanersFactor, False, True, None, self._formatDryCleanersFactor),
            ('DCEQ', 'S Table 1.A.5.d. Dry Cleaners EXTRA Earthquake Factor', self.buildDryCleanersEQFactor, False, True, None, None),
            ('FU', 'S Table 1.C.5.a. Funeral Home EXTRA Endorsement', self.buildFuneralHomeEndorsement, False, True, None, None),
            ('FUMP', 'S Table 1.C.5.c. Funeral Home EXTRA Endorsement Minimum Premium', self.buildFuneralHomeMinPrem, False, True, None, None),
        ]
        # A company can be present in rateTables (its ratebook was uploaded)
        # without having filed its own base-rate tables — a deviation
        # ratebook may only override a handful of tables. Check for the
        # specific tables buildBaseRates() needs, not just company
        # membership, or it KeyErrors on that company's missing table.
        for company, tab, label in (('NACO', 'BRNACO', 'NW Assurance'), ('NAFF', 'BRNAFF', 'NW Affinity'),
                                     ('NGIC', 'BRNGIC', 'NW General Insurance Company'), ('NICOF', 'BRNICOF', 'NICOF')):
            if company in self.rateTables and all(t in self.rateTables[company] for t in self._BASE_RATE_TABLES):
                sheetSpecs.append((tab, f'S Table 3.B.1. {label} State Base Rates', lambda c=company: self.buildBaseRates(c), False, True, 'AS_BR', None))

        sheetSpecs += [
            ('CBG', 'S Table 3.C.2.c. Construction Factor - Building', lambda: self.buildConstructionType('Building'), False, True, None, None),
            ('CPP', 'S Table 3.C.2.c. Construction Factor - BPP', lambda: self.buildConstructionType('BPP'), False, True, None, None),
            ('ET', 'S Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions, False, True, None, None),
            ('YBBG', 'S Table 3.C.2.p. Year Built Modifier - Building', lambda: self.buildYearBuiltModifier('Building'), False, True, None, None),
            ('YBPP', 'S Table 3.C.2.p. Year Built Modifier - BPP', lambda: self.buildYearBuiltModifier('BPP'), False, True, None, None),
            ('EBB', 'S Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate, False, True, None, None),
            ('PDLD', 'S Table 3.C.4.b. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount, False, True, None, None),
            ('LL', 'S Table 3.C.4.d. Liability Limit Factor', self.buildLiabilityLimitFactor, False, True, None, None),
            ('LS', 'S Table 3.C.4.e. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk, False, True, 'LS_RETAIL', None),
            ('GLO', 'S Table 3.D.1.c. General Liability Occupancy Modifiers', self.buildGeneralOccupancyMod, False, True, None, None),
            ('DO', 'S Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns, False, True, None, self._formatDirsOfficersLiabIns),
            ('DONM', 'S Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief, False, True, None, None),
            ('ERP', 'S Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods, False, True, None, None),
            ('BB', 'S Table 4.B.1.e.(1). Barber or Beauty Shops Professional Liability', self.buildBarberProfLiab, False, True, None, None),
            ('PLUS', 'S Table 4.C. Service PLUS Endorsement', self.buildEndorsementCharge, False, True, None, None),
            ('FR', 'S Table 4.D. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement, False, True, None, None),
            ('RSS', 'S Table 4.E. Repair Services Specialized Endorsement', self.buildRepairSpecializedEndorsement, False, True, None, self._formatRepairSpecializedEndorsement),
            ('PSS', 'S Table 4.F. Pet Services Specialized Endorsement', self.buildPetSpecializedEndorsement, False, True, None,
             lambda ws: self._formatPSSplzdEndo(ws, Service.fontBold, self.buildServiceBusinessIncome())),
            ('PSPL', 'S Table 4.G. Pet Services Professional Liability', self.buildPetServicePL, False, True, None, lambda ws: ws.insert_rows(3)),
            ('MPVS', 'S Table 4.H. Mobile Pet and Veterinarian Services Endorsement', lambda: pd.DataFrame(), False, False, None,
             lambda ws: self._formatMPVS(ws, Service.fontBold)),
        ]

        total = len(sheetSpecs)
        for i, (tableCode, title, build, useIndex, useHeader, layoutKey, postFormat) in enumerate(sheetSpecs, start=1):
            if progress_callback:
                progress_callback(f"Building sheet {i}/{total}: {tableCode}...")
            print(f"  [{i}/{total}] Building sheet: {tableCode}")
            ws = Service.generateWorksheet(tableCode, title, build(), useIndex, useHeader, layout_key=layoutKey)
            if postFormat:
                postFormat(ws)

        if progress_callback:
            progress_callback("Building Index sheet...")
        print(f"  [{total}/{total}] Building sheet: Index")
        Service.createIndex()
        return Service.getWB()
