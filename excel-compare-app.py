"""
Excel Comparator Pro
A professional Streamlit app to compare two Excel files at sheet and cell level.
"""

import copy
import io
import os
import re
import zipfile
from datetime import datetime
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import PageSetupProperties


def _pick_folder() -> str:
    """Open a native OS folder-picker dialog and return the chosen path.
    Works when Streamlit is running locally (not in a cloud deployment).
    Returns an empty string if the user cancels or tkinter is unavailable."""
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.wm_attributes("-topmost", 1)
        folder = filedialog.askdirectory(title="Select output folder", master=root)
        root.destroy()
        return folder or ""
    except Exception:
        return ""



# ─────────────────────────────────────────────────────────────────────────────
# Page Config
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Excel Comparator Pro",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────

CUSTOM_CSS = """
<style>
  html, body, [class*="css"] {
    font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
  }

  /* ── Hero banner ── */
  .hero {
    background: linear-gradient(135deg, #0f2942 0%, #1a4a7a 60%, #1e6091 100%);
    color: white;
    padding: 2.4rem 2rem 2rem;
    border-radius: 14px;
    margin-bottom: 1.8rem;
    text-align: center;
    box-shadow: 0 8px 32px rgba(15,41,66,0.28);
  }
  .hero h1 { margin:0; font-size:2.3rem; font-weight:800; letter-spacing:-0.5px; }
  .hero p  { margin:0.5rem 0 0; opacity:0.82; font-size:1.05rem; }

  /* ── Metric cards ── */
  .metric-grid { display:flex; gap:14px; flex-wrap:wrap; margin:1rem 0; }
  .metric-card {
    flex:1 1 130px;
    background:white;
    border-radius:12px;
    padding:1.1rem 1rem 0.9rem;
    border:1px solid #e4e8ef;
    text-align:center;
    box-shadow:0 2px 8px rgba(0,0,0,0.05);
  }
  .metric-value { font-size:2rem; font-weight:800; line-height:1.1; }
  .metric-label { font-size:11px; color:#6b7280; text-transform:uppercase;
                  letter-spacing:0.6px; margin-top:4px; }

  /* ── Sheet pill tags ── */
  .sheet-pills {
    display:flex; flex-wrap:wrap; gap:8px;
    padding:1rem 1.1rem; background:#f4f6f9;
    border-radius:10px; margin:0.8rem 0 1.2rem;
  }
  .pill {
    padding:5px 15px; border-radius:20px; font-size:13px;
    font-weight:600; border:2px solid; cursor:default;
    display:inline-flex; align-items:center; gap:5px;
  }
  .pill-unchanged { background:#edf0f4; color:#374151; border-color:#c5cdd8; }
  .pill-new       { background:#d1fae5; color:#065f46; border-color:#10b981; }
  .pill-deleted   { background:#fee2e2; color:#991b1b; border-color:#ef4444; }
  .pill-modified  { background:#fef3c7; color:#78350f; border-color:#f59e0b; }

  /* ── Legend ── */
  .legend { display:flex; flex-wrap:wrap; gap:18px; margin:0.6rem 0 1rem; }
  .legend-item { display:flex; align-items:center; gap:7px; font-size:13px; color:#374151; }
  .legend-dot  { width:15px; height:15px; border-radius:4px; flex-shrink:0; }

  /* ── Diff table ── */
  .diff-wrap {
    overflow-x:auto;
    border-radius:10px;
    border:1px solid #e4e8ef;
    box-shadow:0 2px 8px rgba(0,0,0,0.04);
    max-height:520px;
    overflow-y:auto;
  }
  table.diff {
    border-collapse:collapse;
    width:100%;
    font-size:13px;
    white-space:nowrap;
  }
  table.diff thead th {
    background:#0f2942;
    color:white;
    padding:9px 14px;
    font-weight:600;
    text-align:left;
    position:sticky;
    top:0;
    z-index:2;
    border-right:1px solid rgba(255,255,255,0.1);
  }
  table.diff thead th:first-child { width:42px; text-align:center; }
  table.diff tbody td {
    padding:7px 14px;
    border-bottom:1px solid #f0f2f5;
    border-right:1px solid #f0f2f5;
    vertical-align:top;
    max-width:260px;
    overflow:hidden;
    text-overflow:ellipsis;
  }
  table.diff tbody tr:hover td { filter:brightness(0.97); }

  /* row-level colours */
  .r-added   td { background:#ecfdf5 !important; }
  .r-deleted td { background:#fef2f2 !important; }

  /* cell-level colours */
  .c-changed { background:#fffbeb !important; }
  .c-added   { background:#ecfdf5 !important; }
  .c-deleted { background:#fef2f2 !important; }

  .rn { color:#9ca3af; font-size:11px; text-align:center; user-select:none; }

  /* inline old→new diff inside a changed cell */
  .val-old  { text-decoration:line-through; color:#dc2626; font-size:11px;
              display:block; line-height:1.3; }
  .val-new  { color:#16a34a; font-size:12px; display:block;
              font-weight:600; line-height:1.4; }
  .val-only { font-size:13px; }

  /* ── Info / instruction box ── */
  .info-box {
    background:#eff6ff; border-left:4px solid #3b82f6;
    padding:1rem 1.2rem; border-radius:0 10px 10px 0; margin:0.5rem 0;
    font-size:14px; line-height:1.7; color:#1e3a5f;
  }
  .info-box ol, .info-box ul { margin:0.4rem 0 0; padding-left:1.4rem; }
  .info-box strong { color:#0f2942; }

  /* ── Upload labels ── */
  .upload-label {
    font-size:15px; font-weight:700; color:#0f2942;
    margin-bottom:0.3rem; display:block;
  }

  /* ── Misc ── */
  .divider { border:none; border-top:1.5px solid #e4e8ef; margin:1.6rem 0; }
  .section-title { font-size:1.1rem; font-weight:700; color:#0f2942; margin:0 0 0.6rem; }
  .truncation-note {
    text-align:center; padding:0.8rem 1rem; color:#6b7280;
    font-style:italic; font-size:13px; background:#f9fafb;
  }

  /* hide streamlit default branding */
  #MainMenu { visibility:hidden; }
  footer    { visibility:hidden; }
  .block-container { padding-top:1.5rem; }
</style>
"""

st.markdown(CUSTOM_CSS, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# Data helpers
# ─────────────────────────────────────────────────────────────────────────────

def _file_bytes(uploaded_file) -> bytes:
    """Read bytes from a Streamlit UploadedFile safely."""
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    return uploaded_file.read()


def _detect_engine(raw: bytes, name: str) -> str:
    """Pick openpyxl or xlrd based on file extension."""
    return "xlrd" if name.lower().endswith(".xls") else "openpyxl"


def read_excel_sheets(raw: bytes, filename: str) -> Dict[str, pd.DataFrame]:
    """Load every sheet of an Excel file into {sheet_name: DataFrame}."""
    engine = _detect_engine(raw, filename)
    try:
        buf = io.BytesIO(raw)
        xl = pd.ExcelFile(buf, engine=engine)
        result: Dict[str, pd.DataFrame] = {}
        for sheet in xl.sheet_names:
            buf.seek(0)
            df = pd.read_excel(
                buf, sheet_name=sheet, header=None,
                dtype=str, engine=engine
            )
            df = df.fillna("")
            result[sheet] = df
        return result
    except Exception as exc:
        st.error(f"Could not read **{filename}**: {exc}")
        return {}


def cell_str(val) -> str:
    """Normalise a cell value to a clean comparable string."""
    s = str(val).strip() if val is not None else ""
    # strip trailing .0 produced by float-to-str conversion
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s[: s.index(".")]
    return s


def compare_dataframes(
    old_df: pd.DataFrame, new_df: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, dict, dict, dict]:
    """
    Align two DataFrames and return per-cell / per-row change status.

    Returns
    -------
    old_aligned, new_aligned : padded to same shape
    cell_status  : {(row, col): 'same' | 'changed' | 'added' | 'deleted'}
    row_status   : {row: 'same' | 'changed' | 'added' | 'deleted'}
    stats        : summary counts
    """
    nr = max(len(old_df), len(new_df))
    nc = max(
        len(old_df.columns) if len(old_df) else 0,
        len(new_df.columns) if len(new_df) else 0,
    )
    if nc == 0:
        empty = pd.DataFrame()
        return empty, empty, {}, {}, {
            "total_rows": 0, "total_cols": 0,
            "added_rows": 0, "deleted_rows": 0,
            "changed_rows": 0, "changed_cells": 0,
        }

    old_a = old_df.reindex(range(nr)).reindex(columns=range(nc)).fillna("")
    new_a = new_df.reindex(range(nr)).reindex(columns=range(nc)).fillna("")

    old_row_range = set(range(len(old_df)))
    new_row_range = set(range(len(new_df)))

    cell_status: dict = {}
    row_status: dict = {}
    added_rows = deleted_rows = changed_rows = changed_cells = 0

    for i in range(nr):
        in_old = i in old_row_range
        in_new = i in new_row_range

        if in_new and not in_old:
            row_status[i] = "added"
            added_rows += 1
            for j in range(nc):
                cell_status[(i, j)] = "added"

        elif in_old and not in_new:
            row_status[i] = "deleted"
            deleted_rows += 1
            for j in range(nc):
                cell_status[(i, j)] = "deleted"

        else:
            row_changed = False
            for j in range(nc):
                ov = cell_str(old_a.iat[i, j])
                nv = cell_str(new_a.iat[i, j])
                if ov != nv:
                    cell_status[(i, j)] = "changed"
                    row_changed = True
                    changed_cells += 1
                else:
                    cell_status[(i, j)] = "same"
            if row_changed:
                row_status[i] = "changed"
                changed_rows += 1
            else:
                row_status[i] = "same"

    return old_a, new_a, cell_status, row_status, {
        "total_rows":    nr,
        "total_cols":    nc,
        "added_rows":    added_rows,
        "deleted_rows":  deleted_rows,
        "changed_rows":  changed_rows,
        "changed_cells": changed_cells,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Mass-processing helpers
# ─────────────────────────────────────────────────────────────────────────────

_DATE_RE = re.compile(r"\b\d{2}-\d{2}-\d{4}\b")


def _base_name(filename: str) -> str:
    """
    Strip the date (DD-MM-YYYY) and file extension from *filename*, then
    normalise whitespace so files that differ only by date compare equal.
    """
    stem = os.path.splitext(filename)[0]
    cleaned = _DATE_RE.sub("", stem)
    cleaned = re.sub(r"[\s_\-]+", " ", cleaned).strip().lower()
    return cleaned


def _match_file_pairs(
    current_files: list, proposed_files: list
) -> Tuple[List[Tuple], List, List]:
    """
    Match uploaded files from the current and proposed folders by base name
    (name with date stripped).

    Returns
    -------
    matched  : list of (current_file, proposed_file) tuples
    unmatched_current  : files in current with no proposed counterpart
    unmatched_proposed : files in proposed with no current counterpart
    """
    cur_map:  Dict[str, object] = {}
    prop_map: Dict[str, object] = {}

    for f in current_files:
        key = _base_name(f.name)
        cur_map[key] = f

    for f in proposed_files:
        key = _base_name(f.name)
        prop_map[key] = f

    matched            = [(cur_map[k], prop_map[k]) for k in cur_map if k in prop_map]
    unmatched_current  = [cur_map[k]  for k in cur_map  if k not in prop_map]
    unmatched_proposed = [prop_map[k] for k in prop_map if k not in cur_map]

    return matched, unmatched_current, unmatched_proposed


def _process_file_pair(
    old_f, new_f, fmt: str
) -> Tuple[Optional[bytes], dict]:
    """
    Read, compare and build the Tracked Pages Excel for one file pair.

    Returns (tracked_bytes, stats_dict).  On error tracked_bytes is None
    and stats_dict contains an 'error' key with the message.
    """
    try:
        old_raw    = _file_bytes(old_f)
        new_raw    = _file_bytes(new_f)
        old_sheets = read_excel_sheets(old_raw, old_f.name)
        new_sheets = read_excel_sheets(new_raw, new_f.name)

        if not old_sheets or not new_sheets:
            return None, {"error": "Could not read one or both files"}

        old_names_s: Set[str] = set(old_sheets)
        new_names_s: Set[str] = set(new_sheets)
        new_only_s     = new_names_s - old_names_s
        deleted_only_s = old_names_s - new_names_s
        common_s       = old_names_s & new_names_s
        ordered_s      = list(old_sheets.keys()) + [
            s for s in new_sheets.keys() if s not in old_sheets
        ]

        sheet_stats_s: Dict[str, dict]  = {}
        sheet_data_s:  Dict[str, tuple] = {}
        for sname in common_s:
            oa, na, cs, rs, stats = compare_dataframes(
                old_sheets[sname], new_sheets[sname]
            )
            sheet_stats_s[sname] = stats
            sheet_data_s[sname]  = (oa, na, cs, rs)

        if fmt == "Side-by-Side":
            tracked = build_sidebyside_excel(
                old_sheets, new_sheets, new_only_s, deleted_only_s,
                sheet_stats_s, sheet_data_s, old_f.name, new_f.name,
                old_raw, new_raw,
            )
        else:
            tracked = build_inline_excel(
                old_sheets, new_sheets, new_only_s, deleted_only_s,
                sheet_stats_s, sheet_data_s, old_f.name, new_f.name,
                old_raw, new_raw,
            )

        total_chg = sum(
            v["changed_cells"] + v["added_rows"] + v["deleted_rows"]
            for v in sheet_stats_s.values()
        )
        modified_s = sum(
            1 for v in sheet_stats_s.values()
            if v["changed_cells"] + v["added_rows"] + v["deleted_rows"] > 0
        )
        return tracked, {
            "error":            None,
            "current_file":     old_f.name,
            "proposed_file":    new_f.name,
            "total_sheets":     len(old_names_s | new_names_s),
            "new_sheets":       len(new_only_s),
            "deleted_sheets":   len(deleted_only_s),
            "modified_sheets":  modified_s,
            "unchanged_sheets": len(common_s) - modified_s,
            "changed_cells":    sum(v["changed_cells"] for v in sheet_stats_s.values()),
            "added_rows":       sum(v["added_rows"]    for v in sheet_stats_s.values()),
            "deleted_rows":     sum(v["deleted_rows"]  for v in sheet_stats_s.values()),
            "total_changes":    total_chg,
        }
    except Exception as exc:
        return None, {"error": str(exc), "current_file": old_f.name, "proposed_file": new_f.name}


# ─────────────────────────────────────────────────────────────────────────────
# HTML rendering
# ─────────────────────────────────────────────────────────────────────────────

_MAX_DISPLAY_ROWS = 1000


def _esc(text: str) -> str:
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )


def render_diff_table(
    old_a: pd.DataFrame,
    new_a: pd.DataFrame,
    cell_status: dict,
    row_status: dict,
    max_rows: int = _MAX_DISPLAY_ROWS,
) -> str:
    """Return an HTML string of the colour-coded diff table."""
    nr = len(old_a)
    nc = len(old_a.columns)
    col_headers = [get_column_letter(j + 1) for j in range(nc)]

    parts = ['<div class="diff-wrap"><table class="diff"><thead><tr>']
    parts.append('<th class="rn">#</th>')
    for ch in col_headers:
        parts.append(f"<th>{ch}</th>")
    parts.append("</tr></thead><tbody>")

    show = min(nr, max_rows)
    for i in range(show):
        rs = row_status.get(i, "same")
        row_cls = {"added": "r-added", "deleted": "r-deleted"}.get(rs, "")
        parts.append(f'<tr class="{row_cls}">')
        parts.append(f'<td class="rn">{i + 1}</td>')

        for j in range(nc):
            cs = cell_status.get((i, j), "same")
            ov = _esc(cell_str(old_a.iat[i, j]))
            nv = _esc(cell_str(new_a.iat[i, j]))

            if rs == "added":
                cell_cls = "c-added"
                inner = f'<span class="val-only">{nv}</span>'
            elif rs == "deleted":
                cell_cls = "c-deleted"
                inner = f'<span class="val-only">{ov}</span>'
            elif cs == "changed":
                cell_cls = "c-changed"
                if ov and nv:
                    inner = (
                        f'<span class="val-old">{ov}</span>'
                        f'<span class="val-new">{nv}</span>'
                    )
                else:
                    inner = f'<span class="val-only">{nv or ov}</span>'
            else:
                cell_cls = ""
                inner = f'<span class="val-only">{nv}</span>'

            td_cls = f' class="{cell_cls}"' if cell_cls else ""
            parts.append(f"<td{td_cls}>{inner}</td>")

        parts.append("</tr>")

    if nr > max_rows:
        parts.append(
            f'<tr><td colspan="{nc + 1}" class="truncation-note">'
            f"⚠ Showing first {max_rows:,} of {nr:,} rows. "
            f"Download the highlighted Excel report to view all rows.</td></tr>"
        )

    parts.append("</tbody></table></div>")
    return "".join(parts)


# ─────────────────────────────────────────────────────────────────────────────
# Excel export
# ─────────────────────────────────────────────────────────────────────────────

# Colour fills
_FILL_CHANGED  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_FILL_ADDED    = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_FILL_DELETED  = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_FILL_HEADER   = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
_FILL_SUMMARY_HEADER = PatternFill(start_color="1A4A7A", end_color="1A4A7A", fill_type="solid")
_FONT_HEADER   = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
_FONT_NORMAL   = Font(name="Segoe UI", size=10)
_FONT_BOLD     = Font(name="Segoe UI", size=10, bold=True)

_TAB_COLOR = {
    "new":       "10B981",
    "deleted":   "EF4444",
    "modified":  "F59E0B",
    "unchanged": "6B7280",
}

_THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# ── Side-by-side view fills / fonts / borders ────────────────────────────────
_FILL_SBS_CHG_NEW   = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_FILL_SBS_NEW_SHEET = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_FILL_SBS_DEL_SHEET = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_FILL_SBS_SEP_DATA  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_FILL_SBS_HDR_OLD   = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
_FILL_SBS_HDR_NEW   = PatternFill(start_color="1E6091", end_color="1E6091", fill_type="solid")
_FILL_SBS_HDR_SEP   = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
_FONT_STRIKE        = Font(name="Segoe UI", size=10, strike=True, color="C00000")
_FONT_SBS_HDR       = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
_BORDER_SEP_DATA    = Border(
    left=Side(style="medium", color="595959"),
    right=Side(style="medium", color="595959"),
    top=Side(style="thin",   color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _auto_width(ws):
    """Auto-fit column widths (capped at 60)."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def _apply_border(ws, max_row: int, max_col: int):
    """Apply thin borders to the data range."""
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(r, c).border = _THIN_BORDER


def build_summary_sheet(
    wb: openpyxl.Workbook,
    ordered: List[str],
    new_only: Set[str],
    deleted_only: Set[str],
    sheet_stats: Dict[str, dict],
    old_filename: str,
    new_filename: str,
):
    """Create a formatted Summary sheet at the front of the workbook."""
    ws = wb.create_sheet("📋 Summary", 0)
    ws.sheet_properties.tabColor = "0F2942"

    # Title block
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "Excel Comparison Report"
    title_cell.font = Font(name="Segoe UI", size=16, bold=True, color="0F2942")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Meta rows
    meta = [
        ("Current Pages",  old_filename),
        ("Proposed Pages", new_filename),
        ("Generated",      datetime.now().strftime("%Y-%m-%d  %H:%M:%S")),
    ]
    for idx, (label, value) in enumerate(meta, start=2):
        ws.cell(idx, 1, label).font = _FONT_BOLD
        ws.cell(idx, 2, value).font = _FONT_NORMAL
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18
    ws.row_dimensions[4].height = 18

    # Column headers for sheet table
    headers = ["Sheet Name", "Status", "Changed Cells",
               "Added Rows", "Deleted Rows", "Changed Rows", "Notes"]
    header_row = 6
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(header_row, c, h)
        cell.fill = _FILL_SUMMARY_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[header_row].height = 22

    STATUS_FILL = {
        "Added":     PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Deleted":   PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Modified":  PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "Unchanged": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }

    for row_off, sname in enumerate(ordered, start=1):
        r = header_row + row_off
        if sname in new_only:
            status, sv = "Added", {}
            note = "New sheet in revised file"
        elif sname in deleted_only:
            status, sv = "Deleted", {}
            note = "Removed from revised file"
        else:
            sv = sheet_stats.get(sname, {})
            has_chg = sv.get("changed_cells", 0) + sv.get("added_rows", 0) + sv.get("deleted_rows", 0) > 0
            status = "Modified" if has_chg else "Unchanged"
            note = ""

        row_fill = STATUS_FILL.get(status)
        data = [
            sname,
            status,
            sv.get("changed_cells", "—"),
            sv.get("added_rows",    "—"),
            sv.get("deleted_rows",  "—"),
            sv.get("changed_rows",  "—"),
            note,
        ]
        for c, val in enumerate(data, start=1):
            cell = ws.cell(r, c, val)
            cell.font = _FONT_NORMAL
            cell.alignment = Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

    _apply_border(ws, header_row + len(ordered), len(headers))
    _auto_width(ws)


def build_highlighted_excel(
    old_sheets: Dict[str, pd.DataFrame],
    new_sheets: Dict[str, pd.DataFrame],
    ordered: List[str],
    new_only: Set[str],
    deleted_only: Set[str],
    sheet_stats: Dict[str, dict],
    old_filename: str,
    new_filename: str,
) -> bytes:
    """Build a complete highlighted Excel workbook and return its bytes."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Data sheets ──────────────────────────────────────────────────────────
    for name in ordered:
        safe_name = name[:31]
        ws = wb.create_sheet(title=safe_name)

        if name in new_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if not df.empty:
                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(i, j, None if val == "" else val)
                        c.fill = _FILL_ADDED
                        c.font = _FONT_NORMAL
                _apply_border(ws, len(df), len(df.columns))
                _auto_width(ws)

        elif name in deleted_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if not df.empty:
                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(i, j, None if val == "" else val)
                        c.fill = _FILL_DELETED
                        c.font = _FONT_NORMAL
                _apply_border(ws, len(df), len(df.columns))
                _auto_width(ws)

        else:
            old_df = old_sheets.get(name, pd.DataFrame())
            new_df = new_sheets.get(name, pd.DataFrame())
            old_a, new_a, cell_status, row_status, stats = compare_dataframes(old_df, new_df)

            has_chg = stats["changed_cells"] + stats["added_rows"] + stats["deleted_rows"] > 0
            ws.sheet_properties.tabColor = (
                _TAB_COLOR["modified"] if has_chg else _TAB_COLOR["unchanged"]
            )

            nr, nc = len(new_a), len(new_a.columns)
            for i in range(nr):
                rs = row_status.get(i, "same")
                for j in range(nc):
                    cs = cell_status.get((i, j), "same")

                    if rs == "deleted":
                        raw = old_a.iat[i, j]
                        fill = _FILL_DELETED
                    else:
                        raw = new_a.iat[i, j]
                        fill = (
                            _FILL_ADDED   if rs == "added"   else
                            _FILL_CHANGED if cs == "changed" else
                            None
                        )

                    cell = ws.cell(i + 1, j + 1, None if raw == "" else raw)
                    cell.font = _FONT_NORMAL
                    if fill:
                        cell.fill = fill

            if nr and nc:
                _apply_border(ws, nr, nc)
                _auto_width(ws)

    # ── Summary sheet ────────────────────────────────────────────────────────
    build_summary_sheet(
        wb, ordered, new_only, deleted_only, sheet_stats,
        old_filename, new_filename
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Side-by-side Excel export
# ─────────────────────────────────────────────────────────────────────────────

def _load_source_wb(raw: bytes, filename: str):
    """Load an openpyxl Workbook from raw bytes (xlsx only, full load so all
    worksheet properties including headers/footers are available)."""
    try:
        if filename.lower().endswith(".xls"):
            return None
        return openpyxl.load_workbook(io.BytesIO(raw))
    except Exception:
        return None


def _copy_print_settings(target_ws, src_wb, sheet_name: str):
    """
    Copy page setup, margins, and header/footer from source workbook.

    page_setup and page_margins are copied attribute-by-attribute to avoid
    carrying the source sheet's _parent reference into the target, which would
    cause openpyxl to silently write headers/footers against the wrong sheet.
    HeaderFooter has no _parent so a deep copy is safe there.
    """
    if src_wb is None or sheet_name not in src_wb.sheetnames:
        return
    try:
        src_ws = src_wb[sheet_name]

        # Page setup — individual attributes only
        src_ps = src_ws.page_setup
        tgt_ps = target_ws.page_setup
        for attr in (
            'orientation', 'paperSize', 'scale', 'firstPageNumber',
            'pageOrder', 'usePrinterDefaults', 'blackAndWhite', 'draft',
            'cellComments', 'useFirstPageNumber', 'horizontalDpi',
            'verticalDpi', 'copies', 'errors',
        ):
            try:
                val = getattr(src_ps, attr, None)
                if val is not None:
                    setattr(tgt_ps, attr, val)
            except Exception:
                pass

        # Page margins — individual attributes only
        src_pm = src_ws.page_margins
        tgt_pm = target_ws.page_margins
        for attr in ('left', 'right', 'top', 'bottom', 'header', 'footer'):
            try:
                val = getattr(src_pm, attr, None)
                if val is not None:
                    setattr(tgt_pm, attr, val)
            except Exception:
                pass

        # Header / footer — safe to deep copy (no _parent reference).
        # NOTE: the attribute is "HeaderFooter" (capital H and F) in openpyxl.
        target_ws.HeaderFooter = copy.deepcopy(src_ws.HeaderFooter)

    except Exception:
        pass


def _apply_border_region(ws, r1: int, r2: int, c1: int, c2: int):
    """Apply thin borders to a rectangular region [r1..r2, c1..c2] (1-indexed, inclusive)."""
    if r2 < r1 or c2 < c1:
        return
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(r, c).border = _THIN_BORDER


def _write_sbs_header(
    ws, nc_old: int, sep_col: int, nc_new: int, old_label: str, new_label: str
):
    """Write the OLD | ◄► | NEW banner (row 1) for the side-by-side sheet."""
    # OLD header block
    if nc_old > 0:
        if nc_old > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nc_old)
        c = ws.cell(1, 1)
        c.value     = old_label
        c.fill      = _FILL_SBS_HDR_OLD
        c.font      = _FONT_SBS_HDR
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Separator cell
    c = ws.cell(1, sep_col)
    c.value     = "◄  ►"
    c.fill      = _FILL_SBS_HDR_SEP
    c.font      = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = _BORDER_SEP_DATA

    # NEW header block
    new_start = sep_col + 1
    if nc_new > 0:
        end_col = new_start + nc_new - 1
        if nc_new > 1:
            ws.merge_cells(start_row=1, start_column=new_start, end_row=1, end_column=end_col)
        c = ws.cell(1, new_start)
        c.value     = new_label
        c.fill      = _FILL_SBS_HDR_NEW
        c.font      = _FONT_SBS_HDR
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 26


def build_sidebyside_excel(
    old_sheets:   Dict[str, pd.DataFrame],
    new_sheets:   Dict[str, pd.DataFrame],
    new_only:     Set[str],
    deleted_only: Set[str],
    sheet_stats:  Dict[str, dict],
    sheet_data:   Dict[str, tuple],
    old_filename: str,
    new_filename: str,
    old_raw:      bytes,
    new_raw:      bytes,
) -> bytes:
    """
    Build a professional side-by-side OLD vs NEW comparison Excel workbook.

    Layout per sheet
    ────────────────
    • Data starts at row 1 — no header banner row
    • OLD data on left | narrow grey separator column | NEW data on right

    Colour coding
    ─────────────
    • Changed cell  OLD side : strikethrough red font
    • Changed cell  NEW side : yellow background  (#FFFF00)
    • Deleted row   OLD side : light-red fill
    • Added row     NEW side : light-green fill
    • Deleted sheet OLD side : pink-red fill (#FFCCCC), right side blank
    • New sheet     NEW side : light-green fill (#CCFFCC), left side blank
    • Separator column       : mid-grey fill, no borders

    Print behaviour
    ───────────────
    • No cell gridlines shown (matches source file appearance)
    • Headers / footers copied from source workbook
    • Page margins / orientation / paper size copied from source
    • fitToWidth = 1 so both OLD and NEW columns print on the same page
    """

    old_src_wb = _load_source_wb(old_raw, old_filename)
    new_src_wb = _load_source_wb(new_raw, new_filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Sheets in ascending alphabetical order (all unique names)
    all_names = sorted(set(list(old_sheets.keys()) + list(new_sheets.keys())))

    def _finalise(ws, sep_col, src_wb, src_name):
        """Apply common post-write settings: widths, no gridlines, print setup."""
        _auto_width(ws)
        ws.column_dimensions[get_column_letter(sep_col)].width = 3
        # Hide cell grid lines to match source file appearance
        ws.sheet_view.showGridLines = False
        # Copy headers / footers / margins from source
        _copy_print_settings(ws, src_wb, src_name)
        # Force landscape + both OLD and NEW columns onto one page width
        ws.page_setup.orientation = 'landscape'
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth  = 1
        ws.page_setup.fitToHeight = 0   # unlimited rows — let it flow vertically

    for name in all_names:
        safe_name = name[:31]
        ws = wb.create_sheet(title=safe_name)

        # ── New sheet (only in revised file) ─────────────────────────────
        if name in new_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            nr        = len(df)
            nc_new    = len(df.columns)
            nc_old    = nc_new          # blank mirror on the left
            sep_col   = nc_old + 1
            new_start = sep_col + 1

            for i, row_vals in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=1
            ):
                for j in range(1, nc_old + 1):          # OLD side — empty
                    ws.cell(i, j)
                ws.cell(i, sep_col).fill = _FILL_SBS_SEP_DATA   # Separator
                for jj, val in enumerate(row_vals, start=new_start):  # NEW — green
                    c = ws.cell(i, jj, None if val == "" else val)
                    c.fill = _FILL_SBS_NEW_SHEET
                    c.font = _FONT_NORMAL

            _finalise(ws, sep_col, new_src_wb, name)

        # ── Deleted sheet (only in original file) ────────────────────────
        elif name in deleted_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            nr        = len(df)
            nc_old    = len(df.columns)
            nc_new    = nc_old          # blank mirror on the right
            sep_col   = nc_old + 1
            new_start = sep_col + 1

            for i, row_vals in enumerate(
                dataframe_to_rows(df, index=False, header=False), start=1
            ):
                for j, val in enumerate(row_vals, start=1):     # OLD side — red
                    c = ws.cell(i, j, None if val == "" else val)
                    c.fill = _FILL_SBS_DEL_SHEET
                    c.font = _FONT_NORMAL
                ws.cell(i, sep_col).fill = _FILL_SBS_SEP_DATA   # Separator
                for jj in range(new_start, new_start + nc_new): # NEW side — empty
                    ws.cell(i, jj)

            _finalise(ws, sep_col, old_src_wb, name)

        # ── Common sheet (present in both files) ─────────────────────────
        else:
            old_df = old_sheets.get(name, pd.DataFrame())
            new_df = new_sheets.get(name, pd.DataFrame())

            if name in sheet_data:
                old_a, new_a, cell_status, row_status = sheet_data[name]
            else:
                old_a, new_a, cell_status, row_status, _ = compare_dataframes(old_df, new_df)

            if old_a.empty and new_a.empty:
                continue

            nr     = max(len(old_a) if not old_a.empty else 0,
                         len(new_a) if not new_a.empty else 0)
            nc_old = len(old_a.columns) if not old_a.empty else 0
            nc_new = len(new_a.columns) if not new_a.empty else 0

            sv      = sheet_stats.get(name, {})
            has_chg = (
                sv.get("changed_cells", 0)
                + sv.get("added_rows",   0)
                + sv.get("deleted_rows", 0)
            ) > 0
            ws.sheet_properties.tabColor = (
                _TAB_COLOR["modified"] if has_chg else _TAB_COLOR["unchanged"]
            )

            sep_col   = nc_old + 1
            new_start = sep_col + 1

            for i in range(nr):
                excel_row = i + 1      # data begins at row 1, no header banner
                rs = row_status.get(i, "same")

                # OLD side ──────────────────────────────────────────────
                for j in range(nc_old):
                    val = cell_str(old_a.iat[i, j]) if i < len(old_a) else ""
                    cs  = cell_status.get((i, j), "same")
                    c   = ws.cell(excel_row, j + 1, None if val == "" else val)
                    if rs == "deleted":
                        c.fill = _FILL_DELETED
                        c.font = _FONT_NORMAL
                    elif rs == "added":
                        c.value = None
                        c.font  = _FONT_NORMAL
                    elif cs == "changed":
                        c.font = _FONT_STRIKE   # strikethrough red
                    else:
                        c.font = _FONT_NORMAL

                # Separator ─────────────────────────────────────────────
                ws.cell(excel_row, sep_col).fill = _FILL_SBS_SEP_DATA

                # NEW side ──────────────────────────────────────────────
                for j in range(nc_new):
                    val = cell_str(new_a.iat[i, j]) if i < len(new_a) else ""
                    cs  = cell_status.get((i, j), "same")
                    c   = ws.cell(excel_row, new_start + j, None if val == "" else val)
                    if rs == "added":
                        c.fill = _FILL_ADDED
                        c.font = _FONT_NORMAL
                    elif rs == "deleted":
                        c.value = None
                        c.font  = _FONT_NORMAL
                    elif cs == "changed":
                        c.fill = _FILL_SBS_CHG_NEW  # yellow
                        c.font = _FONT_NORMAL
                    else:
                        c.font = _FONT_NORMAL

            _finalise(ws, sep_col, new_src_wb, name)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Inline-diff Excel export — ZIP post-processing for rich text
# ─────────────────────────────────────────────────────────────────────────────

def _xml_esc(s: str) -> str:
    """Escape characters that are special in XML text content."""
    return (
        s.replace("&", "&amp;")
         .replace("<", "&lt;")
         .replace(">", "&gt;")
    )


def _replace_cell_rich(xml_text: str, ref: str, old_val: str, new_val: str) -> str:
    """
    In the raw worksheet XML, replace the cell at *ref* with an inlineStr
    cell that renders  ~~old_val~~  new_val  using OOXML rich text runs.

    We use a regex that matches <c r="REF" ...>...</c> and replaces the
    whole element with a hand-crafted <is> block.  This avoids using
    openpyxl's CellRichText which infects the entire sheet with inlineStr.
    """
    if old_val and new_val:
        is_xml = (
            '<r><rPr><strike/><color rgb="FFC00000"/></rPr>'
            f'<t>{_xml_esc(old_val)}</t></r>'
            f'<r><t xml:space="preserve">  {_xml_esc(new_val)}</t></r>'
        )
    elif old_val:
        is_xml = (
            '<r><rPr><strike/><color rgb="FFC00000"/></rPr>'
            f'<t>{_xml_esc(old_val)}</t></r>'
        )
    else:
        is_xml = f'<t>{_xml_esc(new_val or "")}</t>'

    pat = r'(<c r="' + re.escape(ref) + r'")((?:[^>]*))>(.*?)</c>'

    def _sub(m):
        # Strip any existing t="..." attribute from the opening tag
        attrs = re.sub(r'\s+t="[^"]*"', '', m.group(2))
        return f'{m.group(1)}{attrs} t="inlineStr"><is>{is_xml}</is></c>'

    return re.sub(pat, _sub, xml_text, flags=re.DOTALL)


def _copy_cell_style(src_cell, dst_cell) -> None:
    """
    Copy all visual formatting from *src_cell* (a source workbook cell) to
    *dst_cell* (an output workbook cell): font, alignment, fill, border,
    and number format.  Skips gracefully if src_cell is None or any attribute
    is unavailable.
    """
    if src_cell is None:
        return
    try:
        if src_cell.font is not None:
            dst_cell.font = copy.copy(src_cell.font)
    except Exception:
        pass
    try:
        if src_cell.alignment is not None:
            dst_cell.alignment = copy.copy(src_cell.alignment)
    except Exception:
        pass
    try:
        # Only copy fill when there actually is one (fill_type=None means "no fill")
        if src_cell.fill is not None and src_cell.fill.fill_type not in (None, "none"):
            dst_cell.fill = copy.copy(src_cell.fill)
    except Exception:
        pass
    try:
        if src_cell.border is not None:
            dst_cell.border = copy.copy(src_cell.border)
    except Exception:
        pass
    try:
        if src_cell.number_format and src_cell.number_format != "General":
            dst_cell.number_format = src_cell.number_format
    except Exception:
        pass


def _font_with_strike(src_font) -> Font:
    """
    Return a new Font that carries all properties of *src_font* but adds
    strike=True and dark-red colour — used for deleted rows/sheets in inline diff.
    Falls back to the global _FONT_STRIKE if src_font is None.
    """
    if src_font is None:
        return _FONT_STRIKE
    try:
        return Font(
            name=src_font.name,
            size=src_font.size,
            bold=src_font.bold,
            italic=src_font.italic,
            underline=src_font.underline,
            strike=True,
            color="C00000",
            vertAlign=src_font.vertAlign,
            charset=src_font.charset,
            scheme=src_font.scheme,
        )
    except Exception:
        return _FONT_STRIKE


def _copy_row_col_dims(src_ws, dst_ws) -> None:
    """Copy row heights and column widths from *src_ws* to *dst_ws*."""
    if src_ws is None:
        return
    try:
        for row_idx, rd in src_ws.row_dimensions.items():
            if rd.height is not None:
                dst_ws.row_dimensions[row_idx].height = rd.height
        for col_letter, cd in src_ws.column_dimensions.items():
            if cd.width is not None:
                dst_ws.column_dimensions[col_letter].width = cd.width
    except Exception:
        pass


def _patch_inline_rich_cells(
    wb_bytes: bytes,
    wb_obj: openpyxl.Workbook,
    rich_map: Dict[str, Dict[str, tuple]],
) -> bytes:
    """
    Post-process the saved workbook ZIP:  for every (sheet_name → {cell_ref →
    (old_val, new_val)}) entry in *rich_map*, open the corresponding worksheet
    XML and replace the placeholder cell with a proper inlineStr rich-text cell.

    This avoids openpyxl's CellRichText which contaminates ALL string cells
    in a sheet with t="inlineStr", producing invalid OOXML.
    """
    # Map worksheet titles to their ZIP paths (xl/worksheets/sheet1.xml etc.)
    sheet_zip_path = {
        ws.title: f"xl/worksheets/sheet{i}.xml"
        for i, ws in enumerate(wb_obj.worksheets, 1)
    }

    in_buf  = io.BytesIO(wb_bytes)
    out_buf = io.BytesIO()

    with zipfile.ZipFile(in_buf, "r") as zin:
        with zipfile.ZipFile(out_buf, "w", compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)

                # Check if this ZIP entry is a worksheet we need to patch
                for sheet_name, cells in rich_map.items():
                    if sheet_zip_path.get(sheet_name) == item.filename and cells:
                        xml_text = data.decode("utf-8")
                        for ref, (old_val, new_val) in cells.items():
                            xml_text = _replace_cell_rich(xml_text, ref, old_val, new_val)
                        data = xml_text.encode("utf-8")
                        break  # each file belongs to at most one sheet

                zout.writestr(item, data)

    return out_buf.getvalue()


def build_inline_excel(
    old_sheets:   Dict[str, pd.DataFrame],
    new_sheets:   Dict[str, pd.DataFrame],
    new_only:     Set[str],
    deleted_only: Set[str],
    sheet_stats:  Dict[str, dict],
    sheet_data:   Dict[str, tuple],
    old_filename: str,
    new_filename: str,
    old_raw:      bytes,
    new_raw:      bytes,
) -> bytes:
    """
    Build an inline-diff Excel workbook.

    Each sheet is a single table — no side-by-side doubling.

    Cell rendering
    ──────────────
    • Changed cell     : ~~old_value~~  new_value  (rich text in one cell)
    • Unchanged cell   : value (no decoration)
    • Deleted row      : every cell strikethrough
    • Added row        : every cell, light-green fill
    • Deleted sheet    : all values strikethrough, red tab
    • New sheet        : all values, green fill, green tab

    Print behaviour
    ───────────────
    • Landscape orientation, fitToWidth = 1
    • Headers / footers / margins copied from source workbook
    • No gridlines shown
    """
    old_src_wb = _load_source_wb(old_raw, old_filename)
    new_src_wb = _load_source_wb(new_raw, new_filename)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    all_names = sorted(set(list(old_sheets.keys()) + list(new_sheets.keys())))

    # rich_cells tracks cells that need post-processing:
    # { sheet_name: { "A1": (old_val, new_val), ... } }
    rich_cells: Dict[str, Dict[str, tuple]] = {}

    def _get_src_ws(wb_obj, sheet_name):
        """Return the source worksheet object, or None if unavailable."""
        if wb_obj is None:
            return None
        try:
            return wb_obj[sheet_name] if sheet_name in wb_obj.sheetnames else None
        except Exception:
            return None

    def _finalise_inline(ws, src_wb, src_name, src_ws=None):
        # Use source row/column dimensions when available; fall back to auto-width
        if src_ws is not None:
            _copy_row_col_dims(src_ws, ws)
        else:
            _auto_width(ws)
        ws.sheet_view.showGridLines = False
        # Preserve the source file's print settings exactly (orientation,
        # margins, headers, footers) — do NOT force landscape here.
        _copy_print_settings(ws, src_wb, src_name)

    for name in all_names:
        ws = wb.create_sheet(title=name[:31])

        # ── New sheet ─────────────────────────────────────────────────────
        if name in new_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            src_ws = _get_src_ws(new_src_wb, name)
            nr, nc = len(df), len(df.columns)
            for i in range(nr):
                for j in range(nc):
                    val = cell_str(df.iat[i, j])
                    c   = ws.cell(i + 1, j + 1, None if val == "" else val)
                    # Copy all source formatting first, then override fill
                    _copy_cell_style(src_ws.cell(i + 1, j + 1) if src_ws else None, c)
                    c.fill = _FILL_ADDED   # green overlay marks it as a new sheet
            _finalise_inline(ws, new_src_wb, name, src_ws)

        # ── Deleted sheet ─────────────────────────────────────────────────
        elif name in deleted_only:
            ws.sheet_properties.tabColor = _TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if df.empty:
                continue
            src_ws = _get_src_ws(old_src_wb, name)
            nr, nc = len(df), len(df.columns)
            for i in range(nr):
                for j in range(nc):
                    val     = cell_str(df.iat[i, j])
                    c       = ws.cell(i + 1, j + 1, None if val == "" else val)
                    src_c   = src_ws.cell(i + 1, j + 1) if src_ws else None
                    # Copy source formatting, then override font with strikethrough
                    _copy_cell_style(src_c, c)
                    c.font = _font_with_strike(src_c.font if src_c else None)
            _finalise_inline(ws, old_src_wb, name, src_ws)

        # ── Common sheet ──────────────────────────────────────────────────
        else:
            old_df = old_sheets.get(name, pd.DataFrame())
            new_df = new_sheets.get(name, pd.DataFrame())

            if name in sheet_data:
                old_a, new_a, cell_status, row_status = sheet_data[name]
            else:
                old_a, new_a, cell_status, row_status, _ = compare_dataframes(old_df, new_df)

            if old_a.empty and new_a.empty:
                continue

            nr = max(len(old_a) if not old_a.empty else 0,
                     len(new_a) if not new_a.empty else 0)
            nc = max(len(old_a.columns) if not old_a.empty else 0,
                     len(new_a.columns) if not new_a.empty else 0)

            sv      = sheet_stats.get(name, {})
            has_chg = (
                sv.get("changed_cells", 0)
                + sv.get("added_rows",   0)
                + sv.get("deleted_rows", 0)
            ) > 0
            ws.sheet_properties.tabColor = (
                _TAB_COLOR["modified"] if has_chg else _TAB_COLOR["unchanged"]
            )

            new_src_ws = _get_src_ws(new_src_wb, name)
            old_src_ws = _get_src_ws(old_src_wb, name)

            for i in range(nr):
                excel_row = i + 1
                rs = row_status.get(i, "same")
                for j in range(nc):
                    cs      = cell_status.get((i, j), "same")
                    old_val = cell_str(old_a.iat[i, j]) if i < len(old_a) else ""
                    new_val = cell_str(new_a.iat[i, j]) if i < len(new_a) else ""
                    c       = ws.cell(excel_row, j + 1)

                    if rs == "deleted":
                        # Row only exists in old file — use old source formatting
                        c.value  = None if old_val == "" else old_val
                        src_c    = old_src_ws.cell(excel_row, j + 1) if old_src_ws else None
                        _copy_cell_style(src_c, c)
                        c.font   = _font_with_strike(src_c.font if src_c else None)

                    elif rs == "added":
                        # Row only exists in new file — use new source formatting + green fill
                        c.value  = None if new_val == "" else new_val
                        src_c    = new_src_ws.cell(excel_row, j + 1) if new_src_ws else None
                        _copy_cell_style(src_c, c)
                        c.fill   = _FILL_ADDED   # override fill to mark added row

                    elif cs == "changed":
                        # Cell changed — copy new source formatting; ZIP patcher adds ~~old~~ new
                        src_c  = new_src_ws.cell(excel_row, j + 1) if new_src_ws else None
                        _copy_cell_style(src_c, c)
                        placeholder = new_val if new_val else old_val
                        c.value = placeholder if placeholder else None
                        if placeholder:
                            ref = f"{get_column_letter(j + 1)}{excel_row}"
                            rich_cells.setdefault(name, {})[ref] = (old_val, new_val)

                    else:
                        # Unchanged — preserve new source formatting exactly
                        c.value = None if new_val == "" else new_val
                        src_c   = new_src_ws.cell(excel_row, j + 1) if new_src_ws else None
                        _copy_cell_style(src_c, c)

            _finalise_inline(ws, new_src_wb, name, new_src_ws)

    buf = io.BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    # Post-process: patch changed cells with proper OOXML rich text
    if any(rich_cells.values()):
        raw = _patch_inline_rich_cells(raw, wb, rich_cells)

    return raw


# ─────────────────────────────────────────────────────────────────────────────
# Session-state initialisation
# ─────────────────────────────────────────────────────────────────────────────

# Individual mode
if "comp_data" not in st.session_state:
    st.session_state.comp_data = None
if "comp_file_ids" not in st.session_state:
    st.session_state.comp_file_ids = (None, None)
if "tracked_report_bytes" not in st.session_state:
    st.session_state.tracked_report_bytes = None
if "tracked_report_filename" not in st.session_state:
    st.session_state.tracked_report_filename = None
if "tracked_fmt" not in st.session_state:
    st.session_state.tracked_fmt = None

# Mass mode
if "mass_results" not in st.session_state:
    st.session_state.mass_results = None       # list of per-file stat dicts
if "mass_zip_bytes" not in st.session_state:
    st.session_state.mass_zip_bytes = None
if "mass_file_ids" not in st.session_state:
    st.session_state.mass_file_ids = (None, None)
if "mass_fmt" not in st.session_state:
    st.session_state.mass_fmt = None
if "mass_save_dir" not in st.session_state:
    st.session_state.mass_save_dir = None      # actual folder files were saved to
if "mass_output_path" not in st.session_state:
    st.session_state.mass_output_path = ""


# ─────────────────────────────────────────────────────────────────────────────
# UI — Header + Mode Tabs
# ─────────────────────────────────────────────────────────────────────────────

st.markdown(
    """
    <div class="hero">
      <h1>📊 Excel Comparator Pro</h1>
      <p>Compare Excel workbooks — individual files or an entire folder at once</p>
    </div>
    """,
    unsafe_allow_html=True,
)

_tab_ind, _tab_mass = st.tabs(["📄  Individual Pages", "📁  Mass Processing"])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — INDIVIDUAL PAGES
# ══════════════════════════════════════════════════════════════════════════════

with _tab_ind:

    # ── Upload row ────────────────────────────────────────────────────────────
    up_col1, up_col2 = st.columns(2)
    with up_col1:
        st.markdown('<span class="upload-label">📁 Current Pages</span>', unsafe_allow_html=True)
        old_file = st.file_uploader(
            "old_upload", type=["xlsx", "xls"], key="old_file",
            label_visibility="collapsed", help="The existing / current version of the pages",
        )
    with up_col2:
        st.markdown('<span class="upload-label">📁 Proposed Pages</span>', unsafe_allow_html=True)
        new_file = st.file_uploader(
            "new_upload", type=["xlsx", "xls"], key="new_file",
            label_visibility="collapsed", help="The proposed / updated version of the pages",
        )

    # ── Format selector + Generate button ────────────────────────────────────
    st.markdown(
        '<p style="font-size:14px;font-weight:700;color:#0f2942;margin:0.9rem 0 0.3rem">'
        '🔧 Tracked Pages format</p>',
        unsafe_allow_html=True,
    )
    fmt_col, btn_col = st.columns([2, 1])
    with fmt_col:
        st.radio(
            "export_format_selector",
            options=["Side-by-Side", "Inline Diff"],
            horizontal=True,
            label_visibility="collapsed",
            key="export_fmt",
            help=(
                "Side-by-Side: Current Pages on the left, Proposed Pages on the right.\n"
                "Inline Diff: single table — changed cells show ~~old~~  new in one cell."
            ),
        )
    with btn_col:
        _ind_generate = st.button(
            "⚙️ Generate Tracked Pages",
            type="primary",
            use_container_width=True,
            key="ind_generate_btn",
            disabled=not (old_file and new_file),
            help="Upload both files first, then click to create the Tracked Pages Excel",
        )

    # ── Instructions ─────────────────────────────────────────────────────────
    if not old_file or not new_file:
        st.markdown(
            """
            <div class="info-box">
              <strong>How to use — Individual mode</strong>
              <ol>
                <li>Upload your <strong>Current Pages</strong> Excel on the left.</li>
                <li>Upload your <strong>Proposed Pages</strong> Excel on the right.</li>
                <li>Choose <strong>Side-by-Side</strong> or <strong>Inline Diff</strong>.</li>
                <li>Click <strong>⚙️ Generate Tracked Pages</strong>.</li>
              </ol>
              <strong>What is detected:</strong>
              <ul>
                <li>🟢 <strong>New sheets</strong> — added in Proposed Pages</li>
                <li>🔴 <strong>Deleted sheets</strong> — removed from Current Pages</li>
                <li>🟡 <strong>Changed cells</strong> — old ↦ new value</li>
                <li>🟢 <strong>Added rows</strong> &nbsp;🔴 <strong>Deleted rows</strong></li>
              </ul>
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        # Detect file pair changes — clear stale results
        _old_fid  = f"{old_file.name}:{old_file.size}"
        _new_fid  = f"{new_file.name}:{new_file.size}"
        _cur_fids = (_old_fid, _new_fid)
        if st.session_state.comp_file_ids != _cur_fids:
            st.session_state.comp_data              = None
            st.session_state.tracked_report_bytes   = None
            st.session_state.tracked_report_filename= None
            st.session_state.tracked_fmt            = None

        # ── Generate on button click ──────────────────────────────────────────
        if _ind_generate:
            _fmt = st.session_state.get("export_fmt", "Side-by-Side")
            ts   = datetime.now().strftime("%Y%m%d_%H%M%S")

            with st.spinner("Reading workbooks…"):
                old_raw    = _file_bytes(old_file)
                new_raw    = _file_bytes(new_file)
                old_sheets = read_excel_sheets(old_raw, old_file.name)
                new_sheets = read_excel_sheets(new_raw, new_file.name)

            if old_sheets and new_sheets:
                old_names_g: Set[str] = set(old_sheets)
                new_names_g: Set[str] = set(new_sheets)
                new_only_g     = new_names_g - old_names_g
                deleted_only_g = old_names_g - new_names_g
                common_g       = old_names_g & new_names_g
                ordered_g: List[str] = list(old_sheets.keys()) + [
                    s for s in new_sheets.keys() if s not in old_sheets
                ]
                with st.spinner("Computing differences…"):
                    sheet_stats_g: Dict[str, dict] = {}
                    sheet_data_g:  Dict[str, tuple] = {}
                    for sname in common_g:
                        oa, na, cs, rs, stats = compare_dataframes(
                            old_sheets[sname], new_sheets[sname]
                        )
                        sheet_stats_g[sname] = stats
                        sheet_data_g[sname]  = (oa, na, cs, rs)

                if _fmt == "Side-by-Side":
                    with st.spinner("Generating Tracked Pages (Side-by-Side)…"):
                        tracked_bytes = build_sidebyside_excel(
                            old_sheets, new_sheets, new_only_g, deleted_only_g,
                            sheet_stats_g, sheet_data_g, old_file.name, new_file.name,
                            old_raw, new_raw,
                        )
                        tracked_fname = f"tracked_pages_sidebyside_{ts}.xlsx"
                else:
                    with st.spinner("Generating Tracked Pages (Inline Diff)…"):
                        tracked_bytes = build_inline_excel(
                            old_sheets, new_sheets, new_only_g, deleted_only_g,
                            sheet_stats_g, sheet_data_g, old_file.name, new_file.name,
                            old_raw, new_raw,
                        )
                        tracked_fname = f"tracked_pages_inline_{ts}.xlsx"

                with st.spinner("Building highlighted summary report…"):
                    report_bytes = build_highlighted_excel(
                        old_sheets, new_sheets, ordered_g, new_only_g, deleted_only_g,
                        sheet_stats_g, old_file.name, new_file.name,
                    )
                    report_fname = f"excel_diff_{ts}.xlsx"

                st.session_state.comp_data = {
                    "old_sheets": old_sheets, "new_sheets": new_sheets,
                    "ordered": ordered_g, "new_only": new_only_g,
                    "deleted_only": deleted_only_g, "common": common_g,
                    "sheet_stats": sheet_stats_g, "sheet_data": sheet_data_g,
                    "old_name": old_file.name, "new_name": new_file.name,
                    "report_bytes": report_bytes, "report_fname": report_fname,
                }
                st.session_state.tracked_report_bytes    = tracked_bytes
                st.session_state.tracked_report_filename = tracked_fname
                st.session_state.tracked_fmt             = _fmt
                st.session_state.comp_file_ids           = _cur_fids
            else:
                st.error("Could not read one or both files. Please check they are valid Excel workbooks.")

        # ── Render results ────────────────────────────────────────────────────
        _cd = st.session_state.comp_data

        if _cd is None:
            st.markdown('<hr class="divider">', unsafe_allow_html=True)
            st.info(
                "Both files are ready. Choose a format and click "
                "**⚙️ Generate Tracked Pages** to run the comparison.",
                icon="ℹ️",
            )
        else:
            # Unpack stored results
            old_sheets   = _cd["old_sheets"];  new_sheets   = _cd["new_sheets"]
            ordered      = _cd["ordered"];     new_only     = _cd["new_only"]
            deleted_only = _cd["deleted_only"]; common       = _cd["common"]
            sheet_stats  = _cd["sheet_stats"]; sheet_data   = _cd["sheet_data"]
            old_names    = set(old_sheets);    new_names    = set(new_sheets)

            total_changes = sum(
                v["changed_cells"] + v["added_rows"] + v["deleted_rows"]
                for v in sheet_stats.values()
            )
            modified_count = sum(
                1 for v in sheet_stats.values()
                if v["changed_cells"] + v["added_rows"] + v["deleted_rows"] > 0
            )

            # Summary metrics
            st.markdown('<hr class="divider">', unsafe_allow_html=True)
            st.markdown('<p class="section-title">📈 Comparison Summary</p>', unsafe_allow_html=True)
            total_sheets_seen = len(old_names | new_names)
            st.markdown(f"""
            <div class="metric-grid">
              <div class="metric-card"><div class="metric-value" style="color:#0f2942">{total_sheets_seen}</div><div class="metric-label">Total Sheets</div></div>
              <div class="metric-card"><div class="metric-value" style="color:#10b981">{len(new_only)}</div><div class="metric-label">Added Sheets</div></div>
              <div class="metric-card"><div class="metric-value" style="color:#ef4444">{len(deleted_only)}</div><div class="metric-label">Deleted Sheets</div></div>
              <div class="metric-card"><div class="metric-value" style="color:#f59e0b">{modified_count}</div><div class="metric-label">Modified Sheets</div></div>
              <div class="metric-card"><div class="metric-value" style="color:#6366f1">{len(common) - modified_count}</div><div class="metric-label">Unchanged Sheets</div></div>
              <div class="metric-card"><div class="metric-value" style="color:#dc2626">{total_changes:,}</div><div class="metric-label">Total Changes</div></div>
            </div>""", unsafe_allow_html=True)

            # Sheet pills
            st.markdown('<hr class="divider">', unsafe_allow_html=True)
            st.markdown('<p class="section-title">📋 Sheet Overview</p>', unsafe_allow_html=True)
            st.markdown("""<div class="legend">
              <div class="legend-item"><div class="legend-dot" style="background:#10b981"></div>New sheet</div>
              <div class="legend-item"><div class="legend-dot" style="background:#ef4444"></div>Deleted sheet</div>
              <div class="legend-item"><div class="legend-dot" style="background:#f59e0b"></div>Modified sheet</div>
              <div class="legend-item"><div class="legend-dot" style="background:#c5cdd8"></div>Unchanged sheet</div>
            </div>""", unsafe_allow_html=True)
            pills_html = ['<div class="sheet-pills">']
            for sname in ordered:
                if sname in new_only:
                    cls, icon = "pill pill-new", "＋"
                elif sname in deleted_only:
                    cls, icon = "pill pill-deleted", "−"
                else:
                    sv = sheet_stats.get(sname, {})
                    has_chg = sv.get("changed_cells", 0) + sv.get("added_rows", 0) + sv.get("deleted_rows", 0) > 0
                    cls  = "pill pill-modified" if has_chg else "pill pill-unchanged"
                    icon = "~" if has_chg else "✓"
                pills_html.append(f'<span class="{cls}">{_esc(f"{icon} {sname}")}</span>')
            pills_html.append("</div>")
            st.markdown("".join(pills_html), unsafe_allow_html=True)

            # Per-sheet analysis
            st.markdown('<hr class="divider">', unsafe_allow_html=True)
            st.markdown('<p class="section-title">🔍 Sheet-by-Sheet Analysis</p>', unsafe_allow_html=True)
            tab_labels: List[str] = []
            for sname in ordered:
                if sname in new_only:       tab_labels.append(f"🟢 {sname}")
                elif sname in deleted_only: tab_labels.append(f"🔴 {sname}")
                else:
                    sv = sheet_stats.get(sname, {})
                    has_chg = sv.get("changed_cells", 0) + sv.get("added_rows", 0) + sv.get("deleted_rows", 0) > 0
                    tab_labels.append(f"🟡 {sname}" if has_chg else f"⚪ {sname}")
            sheet_tabs = st.tabs(tab_labels)
            for sname, stab in zip(ordered, sheet_tabs):
                with stab:
                    if sname in new_only:
                        st.success(f"**'{sname}'** is a **new sheet** — exists only in Proposed Pages.")
                        df = new_sheets[sname]
                        st.caption(f"{len(df):,} rows × {len(df.columns):,} columns")
                        st.dataframe(df, use_container_width=True, height=380, hide_index=True)
                    elif sname in deleted_only:
                        st.error(f"**'{sname}'** was **deleted** — exists only in Current Pages.")
                        df = old_sheets[sname]
                        st.caption(f"{len(df):,} rows × {len(df.columns):,} columns")
                        st.dataframe(df, use_container_width=True, height=380, hide_index=True)
                    else:
                        old_a, new_a, cell_status, row_status = sheet_data[sname]
                        sv = sheet_stats[sname]
                        has_chg = sv["changed_cells"] + sv["added_rows"] + sv["deleted_rows"] > 0
                        if not has_chg:
                            st.info(f"✅ No changes detected in **'{sname}'**.")
                            st.dataframe(new_sheets[sname], use_container_width=True, height=320, hide_index=True)
                        else:
                            m1, m2, m3, m4 = st.columns(4)
                            m1.metric("Rows compared", f"{sv['total_rows']:,}")
                            m2.metric("Added rows",    sv["added_rows"],    delta=f"+{sv['added_rows']}"    if sv["added_rows"]    else None)
                            m3.metric("Deleted rows",  sv["deleted_rows"],  delta=f"-{sv['deleted_rows']}"  if sv["deleted_rows"]  else None, delta_color="inverse")
                            m4.metric("Changed cells", f"{sv['changed_cells']:,}")
                            st.markdown("""<div class="legend" style="margin-top:0.8rem">
                              <div class="legend-item"><div class="legend-dot" style="background:#fffbeb;border:1.5px solid #f59e0b"></div>Changed cell</div>
                              <div class="legend-item"><div class="legend-dot" style="background:#ecfdf5;border:1.5px solid #10b981"></div>Added row</div>
                              <div class="legend-item"><div class="legend-dot" style="background:#fef2f2;border:1.5px solid #ef4444"></div>Deleted row</div>
                            </div>""", unsafe_allow_html=True)
                            st.markdown(render_diff_table(old_a, new_a, cell_status, row_status), unsafe_allow_html=True)

            # Export
            st.markdown('<hr class="divider">', unsafe_allow_html=True)
            st.markdown('<p class="section-title">💾 Export</p>', unsafe_allow_html=True)
            if st.session_state.tracked_report_bytes:
                _tfmt = st.session_state.tracked_fmt or "Side-by-Side"
                st.download_button(
                    label="📋 Download Tracked Pages — Side-by-Side" if _tfmt == "Side-by-Side" else "🔀 Download Tracked Pages — Inline Diff",
                    data=st.session_state.tracked_report_bytes,
                    file_name=st.session_state.tracked_report_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary",
                )
            st.download_button(
                label="📥 Download Highlighted Excel Report",
                data=_cd["report_bytes"],
                file_name=_cd["report_fname"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                help="Colour-coded sheet tabs and highlighted cells",
            )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — MASS PROCESSING
# ══════════════════════════════════════════════════════════════════════════════

with _tab_mass:

    # ── File uploaders ────────────────────────────────────────────────────────
    mass_col1, mass_col2 = st.columns(2)

    with mass_col1:
        st.markdown('<span class="upload-label">📂 Current Pages</span>', unsafe_allow_html=True)
        mass_cur_files = st.file_uploader(
            "mass_current",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="mass_cur",
            label_visibility="collapsed",
            help="Select multiple files at once using Ctrl+click or Ctrl+A",
        )
        if mass_cur_files:
            st.caption(f"✅ {len(mass_cur_files)} file(s) selected")

    with mass_col2:
        st.markdown('<span class="upload-label">📂 Proposed Pages</span>', unsafe_allow_html=True)
        mass_prop_files = st.file_uploader(
            "mass_proposed",
            type=["xlsx", "xls"],
            accept_multiple_files=True,
            key="mass_prop",
            label_visibility="collapsed",
            help="Select multiple files at once using Ctrl+click or Ctrl+A",
        )
        if mass_prop_files:
            st.caption(f"✅ {len(mass_prop_files)} file(s) selected")

    # ── Output save location ──────────────────────────────────────────────────
    # Apply any pending Browse result BEFORE the widget is instantiated
    if "_mass_path_pending" in st.session_state:
        st.session_state.mass_output_path = st.session_state.pop("_mass_path_pending")

    st.markdown(
        '<p style="font-size:14px;font-weight:700;color:#0f2942;margin:0.9rem 0 0.3rem">'
        '💾 Output Save Location</p>',
        unsafe_allow_html=True,
    )
    path_col, browse_col = st.columns([4, 1])
    with path_col:
        st.text_input(
            "mass_output_path_input",
            label_visibility="collapsed",
            placeholder='Paste a folder path or click Browse  →  e.g.  C:\\Users\\You\\Desktop',
            help=(
                "Files will be saved to a 'Tracked Pages' subfolder here.\n"
                "Tip (Windows): right-click a folder in Explorer → Copy as path → paste above."
            ),
            key="mass_output_path",
        )
    with browse_col:
        if st.button("📂 Browse", key="mass_browse_btn", use_container_width=True,
                     help="Open a folder picker dialog (local use only)"):
            _picked = _pick_folder()
            if _picked:
                # Stage in a temp key — applied above BEFORE the widget on next run
                st.session_state["_mass_path_pending"] = _picked
                st.rerun()

    # Validate path and show resolved save folder
    _raw_path = st.session_state.mass_output_path.strip().strip('"').strip("'")
    if _raw_path:
        _resolved_save_dir = os.path.join(_raw_path, "Tracked Pages")
        if os.path.isdir(_raw_path):
            st.caption(f"✅ Files will be saved to: `{_resolved_save_dir}`")
        else:
            st.caption(f"⚠️ Folder not found: `{_raw_path}` — it will be created on Generate.")
    else:
        _resolved_save_dir = None
        st.caption("ℹ️ No save location set — files will only be available as a ZIP download.")

    # ── Format selector + Generate button ────────────────────────────────────
    st.markdown(
        '<p style="font-size:14px;font-weight:700;color:#0f2942;margin:0.9rem 0 0.3rem">'
        '🔧 Tracked Pages format</p>',
        unsafe_allow_html=True,
    )
    mfmt_col, mbtn_col = st.columns([2, 1])
    with mfmt_col:
        st.radio(
            "mass_format_selector",
            options=["Side-by-Side", "Inline Diff"],
            horizontal=True,
            label_visibility="collapsed",
            key="mass_export_fmt",
            help=(
                "Side-by-Side: Current Pages on the left, Proposed Pages on the right.\n"
                "Inline Diff: single table — changed cells show ~~old~~  new in one cell."
            ),
        )
    with mbtn_col:
        _mass_generate = st.button(
            "⚙️ Generate All Tracked Pages",
            type="primary",
            use_container_width=True,
            key="mass_generate_btn",
            disabled=not (mass_cur_files and mass_prop_files),
            help="Select both folders first, then click to batch-generate all Tracked Pages",
        )

    # ── Instructions ──────────────────────────────────────────────────────────
    if not mass_cur_files or not mass_prop_files:
        st.markdown(
            """
            <div class="info-box">
              <strong>How to use — Mass Processing mode</strong>
              <ol>
                <li>Click the <em>Current Pages</em> file picker and select all your current
                    Excel files — hold <strong>Ctrl</strong> and click each file, or press
                    <strong>Ctrl+A</strong> to select all files in the folder at once.</li>
                <li>Do the same for <em>Proposed Pages</em>.</li>
                <li>Files are matched automatically — the only difference allowed in the name is
                    the <strong>date</strong> portion (format <code>DD-MM-YYYY</code>), e.g.
                    <em>Report 01-11-2025 Final.xlsx</em> matches
                    <em>Report 15-03-2026 Final.xlsx</em>.</li>
                <li>Set an <strong>Output Save Location</strong> (optional) — files are also
                    saved directly to a <em>Tracked Pages</em> sub-folder there.</li>
                <li>Choose <strong>Side-by-Side</strong> or <strong>Inline Diff</strong> and
                    click <strong>⚙️ Generate All Tracked Pages</strong>.</li>
              </ol>
              <strong>Output file naming:</strong> each file is named after the Proposed Pages
              file with <em>" - TRACKED PAGES"</em> appended, e.g.
              <em>Report 15-03-2026 Final - TRACKED PAGES.xlsx</em>.
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        # Detect file selection changes — clear stale mass results
        _mass_cur_fid  = "|".join(sorted(f"{f.name}:{f.size}" for f in mass_cur_files))
        _mass_prop_fid = "|".join(sorted(f"{f.name}:{f.size}" for f in mass_prop_files))
        _mass_fids     = (_mass_cur_fid, _mass_prop_fid)
        if st.session_state.mass_file_ids != _mass_fids:
            st.session_state.mass_results   = None
            st.session_state.mass_zip_bytes = None
            st.session_state.mass_fmt       = None

        # File matching preview (always shown once files are uploaded)
        matched_pairs, unmatched_cur, unmatched_prop = _match_file_pairs(
            mass_cur_files, mass_prop_files
        )

        preview_col1, preview_col2, preview_col3 = st.columns(3)
        preview_col1.metric("Matched pairs",       len(matched_pairs))
        preview_col2.metric("Unmatched — Current", len(unmatched_cur))
        preview_col3.metric("Unmatched — Proposed",len(unmatched_prop))

        if unmatched_cur or unmatched_prop:
            with st.expander("⚠️ Unmatched files (will be skipped)"):
                if unmatched_cur:
                    st.markdown("**No match found in Proposed Pages:**")
                    for f in unmatched_cur:
                        st.markdown(f"&nbsp;&nbsp;• `{f.name}`")
                if unmatched_prop:
                    st.markdown("**No match found in Current Pages:**")
                    for f in unmatched_prop:
                        st.markdown(f"&nbsp;&nbsp;• `{f.name}`")

        if not matched_pairs:
            st.warning("No matching file pairs found. Check that filenames match (aside from the date portion).")
        else:
            # ── Generate on button click ──────────────────────────────────────
            if _mass_generate:
                _mfmt = st.session_state.get("mass_export_fmt", "Side-by-Side")
                ts    = datetime.now().strftime("%Y%m%d_%H%M%S")

                # Prepare output folder on disk (if path provided)
                _out_base = st.session_state.mass_output_path.strip().strip('"').strip("'")
                _disk_dir: Optional[str] = None
                if _out_base:
                    _disk_dir = os.path.join(_out_base, "Tracked Pages")
                    try:
                        os.makedirs(_disk_dir, exist_ok=True)
                    except Exception as _mkdir_err:
                        st.error(f"Could not create output folder: {_mkdir_err}")
                        _disk_dir = None

                mass_results_list = []
                zip_buf = io.BytesIO()
                total_pairs = len(matched_pairs)

                with st.status(
                    f"⚙️ Generating {total_pairs} Tracked Page(s)…",
                    expanded=True,
                ) as _status:
                    progress_bar = st.progress(0, text="Initialising…")

                    with zipfile.ZipFile(zip_buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                        for idx, (cur_f, prop_f) in enumerate(matched_pairs):
                            pct_label = f"**[{idx + 1}/{total_pairs}]** Processing `{prop_f.name}`…"
                            st.write(pct_label)
                            progress_bar.progress(
                                idx / total_pairs,
                                text=f"Processing {idx + 1} of {total_pairs}: {prop_f.name}",
                            )
                            tracked_bytes, stats = _process_file_pair(cur_f, prop_f, _mfmt)
                            mass_results_list.append(stats)
                            if tracked_bytes:
                                stem      = os.path.splitext(prop_f.name)[0]
                                out_fname = f"{stem} - TRACKED PAGES.xlsx"
                                # Add to ZIP
                                zf.writestr(out_fname, tracked_bytes)
                                # Save directly to disk (overwrite if exists)
                                if _disk_dir:
                                    try:
                                        disk_path = os.path.join(_disk_dir, out_fname)
                                        with open(disk_path, "wb") as fh:
                                            fh.write(tracked_bytes)
                                    except Exception as _write_err:
                                        stats["save_error"] = str(_write_err)

                    progress_bar.progress(1.0, text=f"✅ Done — {total_pairs} file(s) processed.")
                    _status.update(
                        label=f"✅ Complete — {total_pairs} file(s) processed.",
                        state="complete",
                        expanded=False,
                    )

                st.session_state.mass_results   = mass_results_list
                st.session_state.mass_zip_bytes = zip_buf.getvalue()
                st.session_state.mass_fmt       = _mfmt
                st.session_state.mass_file_ids  = _mass_fids
                st.session_state.mass_save_dir  = _disk_dir

            # ── Render mass results ───────────────────────────────────────────
            _mr = st.session_state.mass_results

            if _mr is None:
                st.markdown('<hr class="divider">', unsafe_allow_html=True)
                st.info(
                    f"**{len(matched_pairs)} pair(s)** ready. Choose a format and click "
                    "**⚙️ Generate All Tracked Pages** to start batch processing.",
                    icon="ℹ️",
                )
            else:
                # ── Aggregate summary cards ───────────────────────────────────
                ok_results  = [r for r in _mr if not r.get("error")]
                err_results = [r for r in _mr if r.get("error")]

                total_files_proc   = len(ok_results)
                total_chg_across   = sum(r["total_changes"]   for r in ok_results)
                total_cells_across = sum(r["changed_cells"]   for r in ok_results)
                total_added_across = sum(r["added_rows"]      for r in ok_results)
                total_del_across   = sum(r["deleted_rows"]    for r in ok_results)
                files_with_changes = sum(1 for r in ok_results if r["total_changes"] > 0)
                files_no_changes   = total_files_proc - files_with_changes

                st.markdown('<hr class="divider">', unsafe_allow_html=True)
                st.markdown('<p class="section-title">📊 Mass Processing Summary</p>', unsafe_allow_html=True)

                st.markdown(f"""
                <div class="metric-grid">
                  <div class="metric-card"><div class="metric-value" style="color:#0f2942">{total_files_proc}</div><div class="metric-label">Files Processed</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#f59e0b">{files_with_changes}</div><div class="metric-label">Files With Changes</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#10b981">{files_no_changes}</div><div class="metric-label">Files Unchanged</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#dc2626">{total_chg_across:,}</div><div class="metric-label">Total Changes</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#6366f1">{total_cells_across:,}</div><div class="metric-label">Changed Cells</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#10b981">{total_added_across:,}</div><div class="metric-label">Added Rows</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#ef4444">{total_del_across:,}</div><div class="metric-label">Deleted Rows</div></div>
                  <div class="metric-card"><div class="metric-value" style="color:#ef4444">{len(err_results)}</div><div class="metric-label">Errors</div></div>
                </div>""", unsafe_allow_html=True)

                # ── Per-file results table ────────────────────────────────────
                st.markdown('<hr class="divider">', unsafe_allow_html=True)
                st.markdown('<p class="section-title">📋 Per-File Results</p>', unsafe_allow_html=True)

                tbl_rows = []
                for r in _mr:
                    if r.get("error"):
                        tbl_rows.append({
                            "Current File":    r.get("current_file",  "—"),
                            "Proposed File":   r.get("proposed_file", "—"),
                            "Status":          f"❌ Error: {r['error']}",
                            "Sheets":          "—", "New": "—", "Deleted": "—",
                            "Modified": "—", "Changed Cells": "—",
                            "Added Rows": "—", "Deleted Rows": "—",
                        })
                    else:
                        status = "✅ No changes" if r["total_changes"] == 0 else f"🟡 {r['total_changes']:,} changes"
                        tbl_rows.append({
                            "Current File":  r["current_file"],
                            "Proposed File": r["proposed_file"],
                            "Status":        status,
                            "Sheets":        r["total_sheets"],
                            "New":           r["new_sheets"],
                            "Deleted":       r["deleted_sheets"],
                            "Modified":      r["modified_sheets"],
                            "Changed Cells": r["changed_cells"],
                            "Added Rows":    r["added_rows"],
                            "Deleted Rows":  r["deleted_rows"],
                        })

                st.dataframe(
                    pd.DataFrame(tbl_rows),
                    use_container_width=True,
                    hide_index=True,
                    height=min(400, 60 + 38 * len(tbl_rows)),
                )

                if err_results:
                    st.error(f"{len(err_results)} file(s) failed to process — see Status column above.")

                # ── Save location status ──────────────────────────────────────
                _saved_dir = st.session_state.mass_save_dir
                save_errors = [r for r in ok_results if r.get("save_error")]

                st.markdown('<hr class="divider">', unsafe_allow_html=True)
                st.markdown('<p class="section-title">💾 Output</p>', unsafe_allow_html=True)

                if _saved_dir:
                    saved_count = len(ok_results) - len(save_errors)
                    st.success(
                        f"**{saved_count} file(s) saved** to `{_saved_dir}`",
                        icon="✅",
                    )
                    if save_errors:
                        with st.expander(f"⚠️ {len(save_errors)} file(s) could not be saved to disk"):
                            for r in save_errors:
                                st.markdown(f"• `{r.get('proposed_file','?')}` — {r['save_error']}")
                else:
                    st.info("No save location was set — files are available as ZIP download only.", icon="ℹ️")

                _mfmt_label = st.session_state.get("mass_fmt", "Side-by-Side")
                st.download_button(
                    label=f"📦 Download All Tracked Pages (ZIP) — {_mfmt_label}",
                    data=st.session_state.mass_zip_bytes,
                    file_name=f"tracked_pages_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                    mime="application/zip",
                    use_container_width=True,
                    type="primary",
                    help=f"ZIP containing one Tracked Pages Excel per matched file pair ({len(ok_results)} file(s))",
                )
