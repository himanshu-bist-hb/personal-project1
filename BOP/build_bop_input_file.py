"""
build_bop_input_file.py
========================
One-time generator for "BOP/BOP Input File.xlsx" — the non-technical
person's config surface for the BOP (Business Owners Policy) LOB.

Every value written here is transcribed VERBATIM from the values that used
to be hardcoded in the root-level StatePageGenerator.py and AllProgramsPage.py
(pre-refactor). Running this script reproduces that same starting point as
an editable Excel workbook instead of buried Python literals.

Run this once to create the file:
    python BOP/build_bop_input_file.py

It is safe to re-run later to reset every tab back to these defaults —
but note that will overwrite any hand edits made directly in the workbook.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT_PATH = Path(__file__).parent / "BOP Input File.xlsx"

HEADER_FONT = Font(bold=True)


def _write_table(ws, headers, rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
    for row in rows:
        ws.append(row)
    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 10), 60)


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # =======================================================================
    # Formatting Defaults — fonts / margins / number formats / borders.
    # This is what makes header/sub-header/title/table/footer sizing
    # consistent across every BOP All Programs page: every sheet reads these
    # same values instead of each format method picking its own.
    # =======================================================================
    ws = wb.create_sheet("Formatting Defaults")
    _write_table(ws, ["Setting", "Value"], [
        ["FontName", "Arial"],
        ["FontSize", 10],
        ["HeaderFontName", "Arial"],
        ["HeaderFontSize", 10],
        ["FooterFontName", "Arial"],
        ["FooterFontSize", 10],
        ["LeftMargin", 0.25],
        ["RightMargin", 0.25],
        ["TopMargin", 1.25],
        ["BottomMargin", 0.95],
        ["HeaderMargin", 0.5],
        ["FooterMargin", 0.25],
        ["BorderColor", "C1C1C1"],
        ["CurrencyFormat", "$#,##0"],
        ["NoDecimalFormat", "#,##0"],
        ["ZipCodeFormat", "####0"],
        ["PrintTitleRows", "1:3"],
        ["PrintTitleRowsWithSubHeader", "1:4"],
    ])

    # =======================================================================
    # Header Footer Text — fixed boilerplate. {tokens} are substituted at
    # runtime with the real state/date/company values.
    # NOTE: this content is new (BOP never had a working ExcelSettings module
    # before, so there was nothing to transcribe) — edit these to the real
    # legal text your filings require.
    # =======================================================================
    ws = wb.create_sheet("Header Footer Text")
    _write_table(ws, ["Field", "Value"], [
        ["HeaderLeftText", "Commercial Lines Manual - Businessowners"],
        ["HeaderCenterTemplate", "\n\n{state} - Rate Pages"],
        ["HeaderRightTemplate", "Effective:\nNew: {n_effective}\nRenewal: {r_effective}"],
        ["FooterLeftTemplate", "{companies}"],
        # {section} resolves per page type via ExcelSettingsBOP.Excel._SECTION_CODES
        # (All Programs->PG, All Peril->AP, Habitational->H, Auto Service->AS).
        ["FooterCenterTemplate", "{section} - {state_abb} - &[Tab] - &P "],
        ["FooterRightTemplate", ""],
    ])

    # =======================================================================
    # Table Layout — column widths per table code (pixels, converted to
    # inches the same way the old pixelsToInches() helper did: px / 7).
    # ColEnd may be "REST" meaning "this column through the last column".
    # Replaces the ~30 near-duplicate format*() methods' hardcoded widths.
    # =======================================================================
    ws = wb.create_sheet("Table Layout")
    _write_table(ws, ["TableCode", "ColStart", "ColEnd", "WidthPx"], [
        ["SPR", 1, 1, 82], ["SPR", 2, 2, 68], ["SPR", 3, 3, 47],
        ["PCBG", 1, 1, 131], ["PCBG", 2, "REST", 53],
        ["PCPP", 1, 1, 131], ["PCPP", 2, "REST", 53],
        ["PCBI", 1, 1, 131], ["PCBI", 2, "REST", 53],
        ["MVBG", 1, 1, 138], ["MVBG", 2, "REST", 53],
        ["MVPP", 1, 1, 138], ["MVPP", 2, "REST", 53],
        ["BV", 1, 1, 229], ["BV", 2, 2, 54],
        ["AIBI", 1, 1, 208], ["AIBI", 2, 2, 54],
        # Cols 1-4 (Property Deductible / BPP Min / BPP Max / Building) are all
        # large whole-dollar amounts (up to $2,000,000,000) — 74px was too
        # narrow and rendered as "####" overflow; 105px (same as Building
        # already used) fits the widest value comfortably.
        ["PD", 1, 4, 105], ["PD", 5, "REST", 53],
        ["PDH", 1, 4, 105], ["PDH", 5, "REST", 53],
        # RI-only Named Storm Deductible Factor sheets — same column shape as WHOBG/WHOPP.
        ["NSPP", 1, 2, 74], ["NSPP", 3, 3, 105], ["NSPP", 4, 4, 105], ["NSPP", 5, "REST", 62],
        ["NSPPH", 1, 2, 74], ["NSPPH", 3, 3, 105], ["NSPPH", 4, 4, 105], ["NSPPH", 5, "REST", 62],
        ["NSBG", 1, 2, 74], ["NSBG", 3, 3, 105], ["NSBG", 4, 4, 105], ["NSBG", 5, "REST", 62],
        ["NSBGH", 1, 2, 74], ["NSBGH", 3, 3, 105], ["NSBGH", 4, 4, 105], ["NSBGH", 5, "REST", 62],
        # RI-only Named Storm percentage-deductible sheets — same column shape as WHPBG/WHPPP.
        ["NSPPP", 1, 1, 159], ["NSPPP", 2, 2, 105], ["NSPPP", 3, "REST", 70],
        ["NSPPPH", 1, 1, 159], ["NSPPPH", 2, 2, 105], ["NSPPPH", 3, "REST", 70],
        ["NSPBG", 1, 1, 159], ["NSPBG", 2, 2, 105], ["NSPBG", 3, "REST", 70],
        ["NSPBGH", 1, 1, 159], ["NSPBGH", 2, 2, 105], ["NSPBGH", 3, "REST", 70],
        ["WHOBG", 1, 2, 74], ["WHOBG", 3, 3, 105], ["WHOBG", 4, 4, 105], ["WHOBG", 5, "REST", 62],
        ["WHOBGH", 1, 2, 74], ["WHOBGH", 3, 3, 105], ["WHOBGH", 4, 4, 105], ["WHOBGH", 5, "REST", 62],
        ["WHOPP", 1, 2, 74], ["WHOPP", 3, 3, 105], ["WHOPP", 4, 4, 105], ["WHOPP", 5, "REST", 62],
        ["WHOPPH", 1, 2, 74], ["WHOPPH", 3, 3, 105], ["WHOPPH", 4, 4, 105], ["WHOPPH", 5, "REST", 62],
        ["WHBBG", 1, 1, 222], ["WHBBG", 2, 2, 159], ["WHBBG", 3, "REST", 70],
        ["WHBBGH", 1, 1, 222], ["WHBBGH", 2, 2, 159], ["WHBBGH", 3, "REST", 70],
        ["WHBPP", 1, 1, 222], ["WHBPP", 2, 2, 159], ["WHBPP", 3, "REST", 70],
        ["WHBPPH", 1, 1, 222], ["WHBPPH", 2, 2, 159], ["WHBPPH", 3, "REST", 70],
        ["WHPBG", 1, 1, 159], ["WHPBG", 2, 2, 105], ["WHPBG", 3, "REST", 70],
        ["WHPBGH", 1, 1, 159], ["WHPBGH", 2, 2, 105], ["WHPBGH", 3, "REST", 70],
        ["WHPPP", 1, 1, 159], ["WHPPP", 2, 2, 105], ["WHPPP", 3, "REST", 70],
        ["WHPPPH", 1, 1, 159], ["WHPPPH", 2, 2, 105], ["WHPPPH", 3, "REST", 70],
        ["BA", 1, 1, 82], ["BA", 2, 2, 166], ["BA", 3, 3, 96],
        ["CSFA", 1, "REST", 91],
        ["BABG", 1, 1, 145], ["BABG", 2, "REST", 53],
        ["BAPP", 1, 1, 145], ["BAPP", 2, "REST", 53],
        ["BABI", 1, 1, 145], ["BABI", 2, "REST", 53],
        ["AIBG", 1, 2, 82], ["AIBG", 3, "REST", 53],
        ["AIPP", 1, 2, 82], ["AIPP", 3, "REST", 53],
        ["BCEG_MULTI", 1, 1, 73], ["BCEG_MULTI", 2, 2, 66], ["BCEG_MULTI", 3, 3, 82], ["BCEG_MULTI", 4, "REST", 53],
        ["BCEG_SINGLE", 1, 1, 82], ["BCEG_SINGLE", 2, "REST", 53],
        ["TIB", 1, 1, 82], ["TIB", 2, 2, 68],
        ["EBL", 1, "REST", 180],
        ["EBD", 1, 1, 145], ["EBD", 2, 2, 68],
        ["MD", 1, 1, 180], ["MD", 2, 2, 68],
        ["TR", 1, 1, 100], ["TR", 2, 2, 150], ["TR", 3, 6, 80],
        # Pre-2.0 uses a DIFFERENT column-width profile than BP-2.0 for these
        # same two sheet codes (an extra column shifts the boundaries by one),
        # so AllProgramsPageCurrent.py passes layout_key="..._CURRENT" instead
        # of reusing the BP-2.0 rows above.
        ["WHOBG_CURRENT", 1, 3, 74], ["WHOBG_CURRENT", 4, 4, 105], ["WHOBG_CURRENT", 5, "REST", 62],
        ["WHOPP_CURRENT", 1, 3, 74], ["WHOPP_CURRENT", 4, 4, 105], ["WHOPP_CURRENT", 5, "REST", 62],
        ["WHPBG_CURRENT", 1, 1, 159], ["WHPBG_CURRENT", 2, "REST", 70],
        ["WHPPP_CURRENT", 1, 1, 159], ["WHPPP_CURRENT", 2, "REST", 70],
        # All Peril (*_AP) profiles — transcribed from the root-level
        # AllPerilPage.py format*() methods where they differ from the All
        # Programs profile of the same sheet code. Sheets without an _AP row
        # (SPR, BA, WHOBG, WHOPP) matched All Programs exactly and reuse it.
        ["TR_AP", 1, 1, 150], ["TR_AP", 2, 5, 80],
        ["PCBG_AP", 1, 1, 131], ["PCBG_AP", 2, "REST", 80],
        ["PCPP_AP", 1, 1, 131], ["PCPP_AP", 2, "REST", 80],
        ["MVBG_AP", 1, 1, 215], ["MVBG_AP", 2, 2, 100],
        ["MVPP_AP", 1, 1, 215], ["MVPP_AP", 2, 2, 100],
        ["PD_AP", 1, 4, 105], ["PD_AP", 5, "REST", 80],
        # Aligned to the exact same widths as WHBBG/WHBPP/WHPBG/WHPPP (All
        # Programs) — there was no real reason for these to differ, and the
        # user wants All Peril's Windstorm/Hail sheets to print identically.
        ["WHBBG_AP", 1, 1, 222], ["WHBBG_AP", 2, 2, 159], ["WHBBG_AP", 3, "REST", 70],
        ["WHBPP_AP", 1, 1, 222], ["WHBPP_AP", 2, 2, 159], ["WHBPP_AP", 3, "REST", 70],
        ["WHPBG_AP", 1, 1, 159], ["WHPBG_AP", 2, 2, 105], ["WHPBG_AP", 3, "REST", 70],
        ["WHPPP_AP", 1, 1, 159], ["WHPPP_AP", 2, 2, 105], ["WHPPP_AP", 3, "REST", 70],
        # RI-only Named Storm percentage-deductible sheets on the All Peril page — same
        # column shape as WHPBG_AP/WHPPP_AP (the fixed-deductible NSPP/NSBG sheets reuse
        # the All Programs NSPP/NSBG profile directly via layout_key=None, no _AP needed).
        ["NSPPP_AP", 1, 1, 159], ["NSPPP_AP", 2, 2, 105], ["NSPPP_AP", 3, "REST", 80],
        ["NSPBG_AP", 1, 1, 159], ["NSPBG_AP", 2, 2, 105], ["NSPBG_AP", 3, "REST", 80],
        ["CSFA_AP", 1, "REST", 100],
        ["BABG_AP", 1, 1, 145], ["BABG_AP", 2, "REST", 80],
        ["BAPP_AP", 1, 1, 145], ["BAPP_AP", 2, "REST", 80],
        ["BABI_AP", 1, 1, 145], ["BABI_AP", 2, "REST", 80],
        ["AIBG_AP", 1, 2, 82], ["AIBG_AP", 3, "REST", 80],
        ["AIPP_AP", 1, 2, 82], ["AIPP_AP", 3, "REST", 80],
        ["BCEG_AP_SINGLE", 1, 1, 82], ["BCEG_AP_SINGLE", 2, 2, 100],
        ["TIB_AP", 1, 1, 100], ["TIB_AP", 2, 2, 68],
        # Pre-2.0 All Peril: WH percentage sheets have one index column. Widths
        # aligned to match WHPBG_CURRENT/WHPPP_CURRENT (All Programs) exactly.
        ["WHPBG_AP_CURRENT", 1, 1, 159], ["WHPBG_AP_CURRENT", 2, "REST", 70],
        ["WHPPP_AP_CURRENT", 1, 1, 159], ["WHPPP_AP_CURRENT", 2, "REST", 70],
        # Individual Programs (Hab, Auto Service) — transcribed from the
        # root-level HabPage*.py / AutoServicePage*.py format*() methods.
        # HAB_BR / AS_BR / HAB_LA / PROGRAM_TR are shared layout_key profiles
        # (passed explicitly since each program's per-company base-rate tabs
        # use tab codes like BRNACO that vary by company); the rest reuse
        # their bare tab code as usual since it's shared across both
        # programs (and future Food/Retail/Office/Service/Wholesale ports).
        ["HAB_BR", 1, 1, 82], ["HAB_BR", 2, "REST", 120],
        ["HAB_LA", 1, 1, 150],
        ["AS_BR", 1, 1, 82], ["AS_BR", 2, "REST", 159],
        ["PROGRAM_TR", 1, 1, 70], ["PROGRAM_TR", 2, "REST", 54],
        ["CBG", 1, 1, 138], ["CBG", 2, "REST", 53],
        ["CPP", 1, 1, 138], ["CPP", 2, "REST", 53],
        ["YBBG", 1, 1, 131], ["YBBG", 2, "REST", 53],
        ["YBPP", 1, 1, 131], ["YBPP", 2, "REST", 53],
        ["NS", 1, 1, 145],
        ["PDLD", 1, 1, 187], ["PDLD", 2, 2, 54],
        # Hab's PDLD sheet needs its own Number Formats profile (dollar sign
        # on the Deductible Amount column — Auto Service's shares the "PDLD"
        # tab name but doesn't want the "$"), so it gets a separate layout
        # key too even though the column widths are identical to PDLD's.
        ["PDLD_HAB", 1, 1, 187], ["PDLD_HAB", 2, 2, 54],
        ["LL", 1, 1, 205], ["LL", 2, 2, 54],
        ["DO", 1, 1, 130], ["DO", 4, 4, 140],
        ["DONM", 1, 1, 225],
        ["ERP", 2, 2, 215],
        ["PLUS", 1, 1, 350],
        ["TO", 1, 1, 82], ["TO", 2, "REST", 80],
        ["CW", 1, 1, 124], ["CW", 2, 2, 54],
        ["LS", 1, 2, 95], ["LS", 3, "REST", 120],
        # Pre-2.0 Auto Service's Liability Size of Risk splits by AutoServType
        # (2 columns) instead of by peril, giving its data columns more room.
        ["LS_CURRENT", 1, 2, 95], ["LS_CURRENT", 3, "REST", 195],
        ["LPGE", 1, 1, 205], ["LPGE", 2, 2, 200],
        ["AIGO", 1, 1, 480],
        ["BGL", 1, 1, 100],
        ["SPD", 1, 1, 90], ["SPD", 2, 2, 54],
        ["FR", 1, 1, 145], ["FR", 2, 2, 100],
        # Retail-only tables — see BOP/RetailPage.py's module docstring for
        # which codes above (CBG/CPP/YBBG/YBPP/EBB/CW/PDLD/LL/DO/DONM/ERP/
        # LPGE/FR/AS_BR/PROGRAM_TR) Retail reuses as-is (widths transcribed
        # from the root RetailPage.py format*() methods matched exactly).
        ["DC", 2, 2, 250], ["DC", 3, 3, 245],
        ["DCEQ", 1, 1, 245],
        ["ET", 1, 1, 82], ["ET", 2, "REST", 80],
        ["GLO", 1, 1, 145], ["GLO", 2, 2, 70], ["GLO", 3, 3, 215],
        ["FL", 2, 2, 145],
        ["HE", 1, 1, 170], ["HE", 2, 2, 190],
        ["OPTI", 1, 1, 170], ["OPTI", 2, 2, 120],
        ["PED", 1, 1, 170], ["PED", 2, 2, 180],
        ["RTS", 1, 1, 350],
        # Retail's own Liability Size of Risk shape — different sub-header
        # text ("Building plus Business Personal Property" vs Auto Service's
        # "Receipts Range") and widths, so it can't reuse the shared "LS"
        # key. 2.0 and pre-2.0 use slightly different widths (140/100 vs
        # 150/125 in the root format methods), matching the LS/LS_CURRENT
        # split convention.
        ["LS_RETAIL", 1, 2, 140], ["LS_RETAIL", 3, "REST", 100],
        ["LS_RETAIL_CURRENT", 1, 2, 150], ["LS_RETAIL_CURRENT", 3, "REST", 125],
    ])

    # =======================================================================
    # Number Formats — currency / no-decimal overrides per table.
    # ColEnd may be "REST". RowStart = first data row the format applies to.
    # =======================================================================
    ws = wb.create_sheet("Number Formats")
    _write_table(ws, ["TableCode", "ColStart", "ColEnd", "RowStart", "Format"], [
        # Whole-dollar amounts, no decimals (unlike the shared "Currency" format,
        # which is "$#,##0.000" and looks awkward on whole-number amounts).
        ["PD", 1, 4, 5, "$#,##0"],
        ["PDH", 1, 4, 5, "$#,##0"],
        ["NSPP", 1, 4, 5, "$#,##0"],
        ["NSPPH", 1, 4, 5, "$#,##0"],
        ["NSBG", 1, 4, 5, "$#,##0"],
        ["NSBGH", 1, 4, 5, "$#,##0"],
        # Percentage-deductible sheets' "Amount of Insurance" column (2) — same
        # whole-dollar convention as NSPP/NSBG above, not the 3-decimal "Currency".
        ["NSPPP", 2, 2, 4, "$#,##0"],
        ["NSPPPH", 2, 2, 4, "$#,##0"],
        ["NSPBG", 2, 2, 4, "$#,##0"],
        ["NSPBGH", 2, 2, 4, "$#,##0"],
        # Same whole-dollar fix applied to every remaining dollar-amount column
        # across All Programs/All Peril (not just PD/NSPP): the Windstorm/Hail
        # AOI columns were still on the 3-decimal "Currency"/"NoDecimal"
        # formats, and the AOI (Amount of Insurance) Relativity Factor sheets
        # (AIBG/AIPP) and their siblings (EBL, EBD, MD) didn't even have a "$"
        # prefix despite holding genuine dollar amounts.
        ["WHOBG", 1, 4, 5, "$#,##0"],
        ["WHOBGH", 1, 4, 5, "$#,##0"],
        ["WHOPP", 1, 4, 5, "$#,##0"],
        ["WHOPPH", 1, 4, 5, "$#,##0"],
        ["WHBBG", 1, 2, 4, "$#,##0"],
        ["WHBBGH", 1, 2, 4, "$#,##0"],
        ["WHBPP", 1, 2, 4, "$#,##0"],
        ["WHBPPH", 1, 2, 4, "$#,##0"],
        # Building Age Years is a plain age number (0-100, plus a "101-1000"
        # text bucket) — not a dollar amount, so no "$" prefix, but the shared
        # "NoDecimal" alias is actually "#,##0.000" (misleadingly named — still
        # 3 decimals), so it needs its own literal whole-number format.
        ["BABG", 1, 1, 4, "#,##0"],
        ["BAPP", 1, 1, 4, "#,##0"],
        ["BABI", 1, 1, 4, "#,##0"],
        ["AIBG", 1, 2, 5, "$#,##0"],
        ["AIPP", 1, 2, 5, "$#,##0"],
        ["EBL", 1, 2, 5, "$#,##0"],
        ["EBD", 1, 1, 4, "$#,##0"],
        ["MD", 1, 1, 4, "$#,##0"],
        ["WHOBG_CURRENT", 1, 4, 5, "$#,##0"],
        ["WHOPP_CURRENT", 1, 4, 5, "$#,##0"],
        # All Peril (*_AP) — see the Table Layout note.
        ["PD_AP", 1, 4, 5, "$#,##0"],
        ["WHBBG_AP", 1, 2, 4, "$#,##0"],
        ["WHBPP_AP", 1, 2, 4, "$#,##0"],
        ["WHPBG_AP", 2, 2, 4, "$#,##0"],
        ["WHPPP_AP", 2, 2, 4, "$#,##0"],
        ["NSPPP_AP", 2, 2, 4, "$#,##0"],
        ["NSPBG_AP", 2, 2, 4, "$#,##0"],
        ["BABG_AP", 1, 1, 4, "#,##0"],
        ["BAPP_AP", 1, 1, 4, "#,##0"],
        ["BABI_AP", 1, 1, 4, "#,##0"],
        ["AIBG_AP", 1, 2, 5, "$#,##0"],
        ["AIPP_AP", 1, 2, 5, "$#,##0"],
        # Individual Programs (Hab, Auto Service) — see the Table Layout note.
        ["HAB_BR", 2, "REST", 4, "#,##0.0000"],
        ["HAB_LA", 2, 2, 4, "#,##0.000"],
        ["AS_BR", 2, "REST", 4, "#,##0.0000"],
        ["YBBG", 1, 1, 4, "###0"],
        ["YBPP", 1, 1, 4, "###0"],
        ["EBB", 1, 1, 4, "$#,##0.00"],
        # "NoDecimal" is itself 3-decimal (Formatting Defaults ->
        # NoDecimalFormat = "#,##0.000", despite the name) — P.D. Deductible
        # Amount / Liability Limit / D&O Limit are whole-dollar amounts, so
        # they get a literal no-decimal format instead of that
        # misleadingly-named alias.
        ["PDLD", 1, 1, 4, "#,##0"],
        # Hab wants a "$" on the numeric Deductible Amount values (the text
        # "No Deductible" row is untouched by a number format either way) —
        # Auto Service's identical-tab-name PDLD sheet does not, so this is
        # a separate layout key (see the Table Layout entry above), not an
        # edit to the shared "PDLD" row.
        ["PDLD_HAB", 1, 1, 4, "$#,##0"],
        ["LL", 1, 1, 4, "#,##0"],
        ["DO", 2, 2, 4, "#,##0"],
        ["DO", 3, 4, 4, "$#,##0.00"],
        ["DONM", 1, 1, 4, "#,##0"],
        ["DONM", 2, 2, 4, "$#,##0.00"],
        ["PLUS", 1, 1, 4, "$#,##0.00"],
        ["LS", 1, 2, 5, "NoDecimal"],
        ["LS_CURRENT", 1, 2, 5, "NoDecimal"],
        # LPG Exposures (Auto Service) — "NoDecimal"/"Currency" are actually
        # #,##0.000 / $#,##0.000 (3 decimals despite the names); both columns
        # are whole-number, so literal formats without decimals instead.
        ["LPGE", 1, 1, 4, "#,##0"],
        ["LPGE", 2, 2, 4, "$#,##0"],
        # Special Property Damage Deductible (Auto Service) — "Currency" is
        # $#,##0.000 (3 decimals); the Deductible column is a whole-dollar
        # amount, so it prints with a dollar sign but no decimals instead.
        ["SPD", 1, 1, 4, "$#,##0"],
        ["FR", 2, 2, 4, "$#,##0.00"],
        # Retail-only tables. Literal formats throughout (not the "NoDecimal"/
        # "Currency" aliases) — those are already known-mislabeled 3-decimal
        # formats (see the PDLD/LPGE comments above); no reason to introduce
        # that trap into new keys.
        ["FL", 1, 1, 4, "#,##0"],
        ["FL", 2, 2, 4, "$#,##0.00"],
        ["HE", 2, 2, 4, "$#,##0.00"],
        ["OPTI", 2, 2, 4, "$#,##0.00"],
        ["PED", 2, 2, 4, "$#,##0.00"],
        ["RTS", 1, 1, 4, "$#,##0.00"],
        ["LS_RETAIL", 1, 2, 5, "#,##0"],
        ["LS_RETAIL_CURRENT", 1, 2, 5, "#,##0"],
    ])

    # =======================================================================
    # Sub Headers — the "insert a merged label row above the column headers"
    # pattern. Up to two labeled/merged column groups per table.
    # =======================================================================
    ws = wb.create_sheet("Sub Headers")
    _write_table(ws, ["TableCode", "InsertAtRow", "PrintTitleRows",
                       "Label1Range", "Label1Text", "Label2Range", "Label2Text"], [
        ["PD", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", ""],
        ["PDH", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", ""],
        ["NSPP", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Named Storm Deductible"],
        ["NSPPH", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Named Storm Deductible"],
        ["NSBG", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Named Storm Deductible"],
        ["NSBGH", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Named Storm Deductible"],
        ["WHOBG", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Wind-Hail Deductible"],
        ["WHOBGH", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Wind-Hail Deductible"],
        ["WHOPP", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Wind-Hail Deductible"],
        ["WHOPPH", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Wind-Hail Deductible"],
        ["AIBG", 3, "1:4", "A:B", "Building Limit", "C:REST", ""],
        ["AIPP", 3, "1:4", "A:B", "Building Limit", "C:REST", ""],
        ["BCEG_SINGLE", 3, "1:4", "B:REST", "Entire State", "", ""],
        ["EBL", 3, "1:4", "A:B", "Total Property Limit", "", ""],
        ["WHOBG_CURRENT", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Wind-Hail Deductible"],
        ["WHOPP_CURRENT", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", "Wind-Hail Deductible"],
        # All Peril (*_AP) — see the Table Layout note.
        ["PD_AP", 3, "1:4", "B:D", "Amount of Insurance", "E:REST", ""],
        ["AIBG_AP", 3, "1:4", "A:B", "Building Limit", "C:REST", ""],
        ["AIPP_AP", 3, "1:4", "A:B", "Building Limit", "C:REST", ""],
        ["BCEG_AP_SINGLE", 3, "1:4", "B:REST", "Entire State", "", ""],
        # Individual Programs (Hab, Auto Service) — Liability Size of Risk's
        # "Receipts Range" label spanning the Min/Max columns.
        ["LS", 3, "1:4", "A:B", "Receipts Range", "C:REST", ""],
        ["LS_CURRENT", 3, "1:4", "A:B", "Receipts Range", "C:REST", ""],
        # Retail's own Liability Size of Risk — different label text/columns
        # than Auto Service's LS, see the Table Layout note.
        ["LS_RETAIL", 3, "1:4", "A:B", "Building plus Business Personal Property", "C:REST", ""],
        ["LS_RETAIL_CURRENT", 3, "1:4", "A:B", "Building plus Business Personal Property", "C:REST", ""],
    ])

    # =======================================================================
    # Footnotes — one-off cell text unrelated to the column/subheader system.
    # =======================================================================
    ws = wb.create_sheet("Footnotes")
    _write_table(ws, ["TableCode", "Cell", "Text"], [
        ["AIBI", "A16", "For each additional 1%, add 0.005"],
    ])

    # =======================================================================
    # Page Break Rules — sheet-name-prefix -> rule. "*" is the default rule
    # applied to every sheet that no more specific prefix matches.
    # Seeded minimally (no BOP print samples yet to know what special rules
    # are needed) — add rows here as real print issues surface, no Python
    # required.
    # =======================================================================
    ws = wb.create_sheet("Page Break Rules")
    _write_table(ws, ["SheetPrefix", "Rule"], [
        ["Index", "index"],
        # RI-only Named Storm sheets are long enough to need real pagination
        # instead of being squeezed onto one page. Deliberately not a bare "NS"
        # prefix — HabPage.py already has an unrelated tab literally named "NS"
        # ("Number of Stories Factor") that must keep the default rule.
        ["NSPP", "fit_width_only"],
        ["NSBG", "fit_width_only"],
        # PDLD ("Property Damage Liability Deductible", Hab/Auto Service programs)
        # must be listed before PD below — it would otherwise match the "PD"
        # prefix too and inherit fit_width_landscape, which isn't what it needs.
        # Explicit here so it keeps today's default (fit_single_page) either way.
        ["PDLD", "fit_single_page"],
        # Property Deductible Factor (All Programs / All Peril) — has far more
        # peril columns than the Named Storm sheets above (up to 18 vs ~9).
        # disable_fit_to_page (tried first) let it print at full scale but also
        # let it split across page-WIDTH (a stray vertical break through the
        # table) since it's too wide for one portrait page. fit_width_landscape
        # forces single-page-width like Named Storm's fit_width_only, but in
        # landscape so there's enough room that the forced scale-down stays
        # close to Named Storm's own (also non-100%) scale instead of much smaller.
        ["PD", "fit_width_landscape"],
        # Building Age Modifier (All Programs / All Peril, both BP-2.0 and
        # pre-2.0) — Building_Age_Min runs roughly year-by-year (~100 rows) and
        # has a similar peril-column count to Property Deductible, so it hits
        # the same "everything squeezed onto one page" unreadability under the
        # default rule. Same fix: single page width, landscape for room, rows
        # flow down naturally with the header row repeating on each page
        # (print_title_rows "1:3" is already the default set at sheet creation
        # in ExcelSettingsBOP._apply_page_setup — nothing else to configure).
        ["BABG", "fit_width_landscape"],
        ["BAPP", "fit_width_landscape"],
        ["BABI", "fit_width_landscape"],
        # Windstorm/Hail Deductible Factor sheets (All Programs / All Peril,
        # both BP-2.0 and pre-2.0) — same column shape/count as the Named Storm
        # sheets above (RI's Named Storm and non-RI's WH tables share the same
        # underlying builder pattern), so the same plain fit_width_only (no
        # landscape needed, unlike the wider PD/Building Age sheets) gets them
        # printing exactly like Named Storm already does. Each prefix also
        # covers its own "H" (Hab-split) sibling via startswith, e.g. "WHOBG"
        # matches both "WHOBG" and "WHOBGH" — no separate row needed.
        ["WHOBG", "fit_width_only"],
        ["WHOPP", "fit_width_only"],
        ["WHBBG", "fit_width_only"],
        ["WHBPP", "fit_width_only"],
        ["WHPBG", "fit_width_only"],
        ["WHPPP", "fit_width_only"],
        # "TRDEF" (Territory Definitions — a different, much larger grid-level
        # sheet, tens of thousands of rows, 21 columns) must be listed before
        # "TR" below — it would otherwise match the "TR" prefix too.
        # fit_single_page squeezed the whole sheet (tens of thousands of rows)
        # onto ONE page, flooring the scale at 10% and making the text
        # unreadable. Same fix as PD/Building Age above: single page width,
        # landscape for room, rows flow down naturally with the header row
        # repeating on each page.
        ["TRDEF", "fit_width_landscape"],
        # State Territory Multiplier — only 6 narrow columns (Peril, Territory,
        # Building, BPP, Liability, BI), same treatment as Named Storm/WH:
        # single page width in portrait (no landscape needed, unlike the wider
        # PD/Building Age sheets), rows flow down with the header repeating.
        ["TR", "fit_width_only"],
        ["*", "fit_single_page"],
    ])

    # =======================================================================
    # Perils By State — replaces the 4 elif-state-in-(...) blocks.
    # NOTE: states not listed here (e.g. AK, HI, LA, OK) were not supported
    # by the original BOP script either — add a row for them here whenever
    # BOP expands to those states.
    # =======================================================================
    ws = wb.create_sheet("Perils By State")
    full_18 = "allother1,cat1,cat2,cat3,cat4,fire1,fire2,fire3,fire4,liability1,liability2,liability3,liability4,theft1,water1,water2,weather1,weather2"
    no_cat3 = "allother1,cat1,cat2,cat4,fire1,fire2,fire3,fire4,liability1,liability2,liability3,liability4,theft1,water1,water2,weather1,weather2"
    no_fire2 = "allother1,cat1,cat2,cat3,cat4,fire1,fire3,fire4,liability1,liability2,liability3,liability4,theft1,water1,water2,weather1,weather2"
    no_cat3_no_fire2 = "allother1,cat1,cat2,cat4,fire1,fire3,fire4,liability1,liability2,liability3,liability4,theft1,water1,water2,weather1,weather2"
    rows = [["TX", full_18]]
    for s in ["AZ", "CA", "CO", "ID", "MT", "NM", "NV", "OR", "UT", "WA", "WY"]:
        rows.append([s, no_cat3])
    for s in ["AL", "AR", "CT", "DC", "DE", "FL", "GA", "IL", "IN", "KY", "MA", "MD", "ME",
              "MO", "MS", "NC", "NH", "NJ", "NY", "OH", "PA", "RI", "SC", "TN", "VA", "VT", "WV"]:
        rows.append([s, no_fire2])
    for s in ["IA", "KS", "MI", "MN", "ND", "NE", "SD", "WI"]:
        rows.append([s, no_cat3_no_fire2])
    _write_table(ws, ["State", "Perils"], rows)

    # =======================================================================
    # Peril Conversions — internal peril code -> display name.
    # =======================================================================
    ws = wb.create_sheet("Peril Conversions")
    _write_table(ws, ["PerilCode", "DisplayName"], [
        ["allother1", "NW-Other"], ["allperil", "AllPeril"], ["cat1", "ST"], ["cat2", "WS"],
        ["cat3", "HU"], ["cat4", "L-Products"], ["fire1", "NW-Fire"], ["fire2", "WF"],
        ["fire3", "FFEQ"], ["fire4", "NC-BINC"], ["liability1", "L-SlipFall"],
        ["liability2", "L-Violence"], ["liability3", "L-OtherMed"], ["liability4", "L-OtherPrem"],
        ["theft1", "NW-Theft"], ["water1", "NW-Water"], ["water2", "NC-Water"],
        ["weather1", "NC-Other"], ["weather2", "NC-Wind"],
    ])

    # =======================================================================
    # Protection Class Conversions — strips excess leading zeros.
    # =======================================================================
    ws = wb.create_sheet("Protection Class Conversions")
    _write_table(ws, ["Code", "DisplayValue"], [
        ["000001", "1"], ["000002", "2"], ["000003", "3"], ["000004", "4"], ["000005", "5"],
        ["000006", "6"], ["000007", "7"], ["000008", "8"], ["000009", "9"], ["000010", "10"],
        ["00001X", "1X"], ["00002X", "2X"], ["00003X", "3X"], ["00004X", "4X"], ["00005X", "5X"],
        ["00006X", "6X"], ["00007X", "7X"], ["00008X", "8X"],
        ["00001Y", "1Y"], ["00002Y", "2Y"], ["00003Y", "3Y"], ["00004Y", "4Y"], ["00005Y", "5Y"],
        ["00006Y", "6Y"], ["00007Y", "7Y"], ["00008Y", "8Y"],
        ["00001W", "1W"], ["00002W", "2W"], ["00003W", "3W"], ["00004W", "4W"], ["00005W", "5W"],
        ["00006W", "6W"], ["00007W", "7W"], ["00008W", "8W"],
        ["00008B", "8B"], ["00009E", "9E"], ["00009S", "9S"], ["00010W", "10W"],
    ])

    # =======================================================================
    # Building Codes By State — states with more than 1 BCEG group.
    # Codes is a comma-separated list of raw territory codes for that group.
    # =======================================================================
    ws = wb.create_sheet("Building Codes By State")
    _write_table(ws, ["State", "Group", "Codes"], [
        ["AL", "A", "001"], ["AL", "B", "004"], ["AL", "C", "005"], ["AL", "D", "006"],
        ["FL", "A", "011,012"], ["FL", "B", "010,015"], ["FL", "C", "002,007,008,014,016,017"], ["FL", "D", "009,013"],
        ["GA", "A", "002"], ["GA", "B", "004"], ["GA", "C", "005"], ["GA", "D", "006"],
        ["MS", "A", "002"], ["MS", "B", "003"], ["MS", "C", "004"],
        ["NC", "A", "003"], ["NC", "B", "004"], ["NC", "C", "005"], ["NC", "D", "006"],
        ["NE", "A", "701"], ["NE", "B", "703"], ["NE", "C", "704"],
        ["SC", "A", "002"], ["SC", "B", "003"], ["SC", "C", "004"],
        ["TX", "A", "004,005,006,007,008,009,015,016"], ["TX", "B", "010,011,012,013,014"],
        ["VA", "A", "001,005,006,007,008,009,012,013"], ["VA", "B", "010,011"],
        ["WY", "A", "702"], ["WY", "B", "703"],
    ])

    # =======================================================================
    # Class Codes — Class_Code_Min -> program display name (All Peril page).
    # =======================================================================
    ws = wb.create_sheet("Class Codes")
    _write_table(ws, ["ClassCodeMin", "Program"], [
        [10000, "Hab"], [20000, "Auto"], [40000, "Food"], [50000, "Retail"],
        [60000, "Office"], [70000, "Service"], [80000, "Wholesale"],
    ])

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    build()
