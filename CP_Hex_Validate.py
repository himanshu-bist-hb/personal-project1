import pandas as pd
import os.path
import numpy as np
import time
import xlsxwriter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, DEFAULT_FONT
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter

start_time = time.time()

# <editor-fold desc ="Create Comparison Dataframes">
HexbookTable = pd.ExcelFile(HexbookFile)
# FinRelTable = pd.ExcelFile(RelativitiesFile)
try:
    SMRatePageTable = pd.ExcelFile(SMPagesFile)
except:
    SMRatePageTable = pd.ExcelFile(MMPagesFile)

try:
    MMRatePageTable = pd.ExcelFile(MMPagesFile)
except:
    MMRatePageTable = pd.ExcelFile(SMPagesFile)
FinRelCols = ['GRID_ID','BGI_Final_Relativity','BGII_Final_Relativity','SCOL_Final_Relativity']
FinRelNames = ['Territory','BGI Factor','BGII Factor','SCOL Factor']

State = pd.read_excel(HexbookTable, sheet_name='BGI').iloc[3,1]
HexbookBGIGrid = pd.read_excel(HexbookTable, sheet_name='BGI', skiprows=7, usecols=[0,1],names = ['Territory', 'BGI Factor'])
HexbookBGIIGrid = pd.read_excel(HexbookTable, sheet_name='BGII', skiprows=7, usecols=[0,1], names = ['Territory','BGII Factor'])
HexbookSCOLGrid = pd.read_excel(HexbookTable, sheet_name='SCOL', skiprows=7, usecols=[0,1], names = ['Territory','SCOL Factor'])
HexbookGrid = HexbookBGIGrid.merge(HexbookBGIIGrid,on='Territory').merge(HexbookSCOLGrid,on='Territory')

# FinGrid = pd.read_excel(FinRelTable,sheet_name='Sheet1')[FinRelCols]
# FinGrid.columns = FinRelNames

SMPageGrid = pd.read_excel(SMRatePageTable, sheet_name='Territory',skiprows=2)
SMPageGrid1 = SMPageGrid.iloc[:,[0,3,4,5]]
SMPageGrid2 = SMPageGrid.iloc[:,[7,10,11,12]]
SMPageGrid2.columns = FinRelNames
SMPageGrid2.reset_index(drop=True)
SMPageGrid3 = SMPageGrid.iloc[:,[14,17,18,19]]
SMPageGrid3.columns = FinRelNames
SMPageGrid3.reset_index(drop=True)
SMPageGrid = pd.concat([SMPageGrid1,SMPageGrid2,SMPageGrid3])
SMPageGrid = SMPageGrid.dropna().reset_index(drop=True)

MMPageGrid = pd.read_excel(MMRatePageTable, sheet_name='Territory',skiprows=2)
MMPageGrid1 = MMPageGrid.iloc[:,[0,3,4,5]]
MMPageGrid2 = MMPageGrid.iloc[:,[7,10,11,12]]
MMPageGrid2.columns = FinRelNames
MMPageGrid2.reset_index(drop=True)
MMPageGrid3 = MMPageGrid.iloc[:,[14,17,18,19]]
MMPageGrid3.columns = FinRelNames
MMPageGrid3.reset_index(drop=True)
MMPageGrid = pd.concat([MMPageGrid1,MMPageGrid2,MMPageGrid3])
MMPageGrid = MMPageGrid.dropna().reset_index(drop=True)

# </editor-fold>


# <editor-fold desc ="Comparison logic">
HexGrid, SMGrid = HexbookGrid.set_index('Territory').align(SMPageGrid.set_index('Territory'))
HextoSMPageCompare = HexGrid.compare(SMGrid,result_names=('Hexbook','SM Factors')).reset_index()

HexGrid2, MMGrid = HexbookGrid.set_index('Territory').align(MMPageGrid.set_index('Territory'))
HextoMMPageCompare = HexGrid2.compare(MMGrid,result_names=('Hexbook','MM Factors')).reset_index()

SMGrid2, MMGrid2 = SMPageGrid.set_index('Territory').align(MMPageGrid.set_index('Territory'))
SMtoMMPageCompare = SMGrid2.compare(MMGrid2,result_names=('SM Factors','MM Factors')).reset_index()

# </editor-fold>

# <editor-fold desc ="Write to Excel">

comparewb = Workbook()
comparewb.active.title = "Hex Validation"
comparewb["Hex Validation"]['A1'] = "This is the output of your comparison of rate page factors to our Hexbook factors."
comparewb["Hex Validation"][
    'A2'] = "The sheets outputted from this comparison tool will only highlight the rows in your Territory factor pages that deviate from Hexbook factors."
try:
    pd.ExcelFile(SMPagesFile)
    pd.ExcelFile(MMPagesFile)
    SMtoHex = comparewb.create_sheet(title="Hex to SM Comparison")
    SMtoHex['A1'] = 'Hexbook to SM Page Comparison'
    SMtoHex['A2'] = 'Only shows differences between Hexbook & SM Dataframes.'
    SMtoHex['A3'] = 'If there are no differences, the sheet is blank.'
    if len(HextoSMPageCompare) > 0:
        for r in dataframe_to_rows(HextoSMPageCompare, False, True):
            # The header is the first row and the index is the second row, but they need to be on the same row in Excel
            if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
                SMtoHex['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
                continue
            SMtoHex.append(r)
    else:
        pass
    MMtoHex = comparewb.create_sheet(title="Hex to MM Comparison")
    MMtoHex['A1'] = 'Hexbook to MM Page Comparison'
    MMtoHex['A2'] = 'Only shows differences between Hexbook & MM Dataframes.'
    MMtoHex['A3'] = 'If there are no differences, the sheet is blank.'
    if len(HextoMMPageCompare) > 0:
        for r in dataframe_to_rows(HextoMMPageCompare, False, True):
            # The header is the first row and the index is the second row, but they need to be on the same row in Excel
            if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
                MMtoHex['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
                continue
            MMtoHex.append(r)
    else:
        pass

    MMtoSM = comparewb.create_sheet(title="SM to MM Comparison")
    MMtoSM['A1'] = 'SM to MM Page Comparison'
    MMtoSM['A2'] = 'Only shows differences between SM & MM Dataframes.'
    MMtoSM['A3'] = 'If there are no differences, the sheet is blank.'
    if len(SMtoMMPageCompare) > 0:
        for r in dataframe_to_rows(SMtoMMPageCompare, False, True):
            # The header is the first row and the index is the second row, but they need to be on the same row in Excel
            if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
                MMtoSM['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
                continue
            MMtoSM.append(r)
    else:
        pass
except:
    try:
        pd.ExcelFile(SMPagesFile)
        comparewb = Workbook()
        comparewb.active.title = "Hex Validation"
        comparewb["Hex Validation"]['A1'] = "This is the output of your comparison of rate page factors to our Hexbook factors."
        comparewb["Hex Validation"]['A2'] = "The sheets outputted from this comparison tool will only highlight the rows in your Territory factor pages that deviate from Hexbook factors."
        SMtoHex = comparewb.create_sheet(title="Hex to SM Comparison")
        SMtoHex['A1'] = 'Hexbook to SM Page Comparison'
        SMtoHex['A2'] = 'Only shows differences between Hexbook & SM Dataframes.'
        SMtoHex['A3'] = 'If there are no differences, the sheet is blank.'
        if len(HextoSMPageCompare) > 0:
            for r in dataframe_to_rows(HextoSMPageCompare, False, True):
                # The header is the first row and the index is the second row, but they need to be on the same row in Excel
                if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
                    SMtoHex['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
                    continue
                SMtoHex.append(r)
        else:
            pass
    except:
        MMtoHex = comparewb.create_sheet(title="Hex to MM Comparison")
        MMtoHex['A1'] = 'Hexbook to MM Page Comparison'
        MMtoHex['A2'] = 'Only shows differences between Hexbook & MM Dataframes.'
        MMtoHex['A3'] = 'If there are no differences, the sheet is blank.'
        if len(HextoMMPageCompare) > 0:
            for r in dataframe_to_rows(HextoMMPageCompare, False, True):
                # The header is the first row and the index is the second row, but they need to be on the same row in Excel
                if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
                    MMtoHex['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
                    continue
                MMtoHex.append(r)
        else:
            pass


comparewb.save(os.path.join(hexpath_selected, State + ' ' +'Hex Factor Comparison.xlsx'))
hexwb = xw.Book(os.path.join(hexpath_selected, State + ' ' +'Hex Factor Comparison.xlsx'))
current_label = hexwb.api.SensitivityLabel.GetLabel()
if current_label.LabelName == "":
    labelinfo = hexwb.api.SensitivityLabel.CreateLabelInfo()
    labelinfo.AssignmentMethod = 2
    labelinfo.Justification = 'init'
    labelinfo.LabelId = 'fbefcacb-1e54-47c6-a321-a2f3e970fe0d'
    labelinfo.LabelName = 'Restricted'
    hexwb.api.SensitivityLabel.SetLabel(labelinfo, labelinfo)
    hexwb.save()
# FintoPageCompare = FinGrid.compare(PageGrid)

# FintoHexCompare = FinGrid.compare(HexbookGrid)

#</editor-fold>
print(time.time() - start_time)
