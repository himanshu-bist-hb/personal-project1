# This module formats the Auto Service State Page workbook in Excel
import traceback
import warnings
from functools import reduce
import pandas as pd
import numpy as np
import ExcelSettingsBA
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.pagebreak import Break
import logging
import re
import time
import copy

from config.constants import BA_INPUT_FILE


# Define custom warning classes. Makes custom warnings when key functions fail more user friendly.
# Search up the warning names to see use cases if interested.
class PIPWarning(UserWarning):
    pass

class CSLWarning(UserWarning):
    pass

class COLLWarning(UserWarning):
    pass

class OTCWarning(UserWarning):
    pass

class MEDPAYWarning(UserWarning):
    pass

class Um_Stacked_Warning(UserWarning):
    pass

class Um_Unstacked_Warning(UserWarning):
    pass

class MMWarning(UserWarning):
    pass

# Set each warning to appear only once
warnings.simplefilter("once", PIPWarning)
warnings.simplefilter("once", CSLWarning)
warnings.simplefilter("once", COLLWarning)
warnings.simplefilter("once", OTCWarning)
warnings.simplefilter("once", MEDPAYWarning)
warnings.simplefilter("once", Um_Stacked_Warning)
warnings.simplefilter("once", Um_Unstacked_Warning)

warnings.simplefilter("ignore", DeprecationWarning)
warnings.simplefilter("ignore", FutureWarning)

def log_exceptions(func):
    def wrapper(*args, **kwargs):
        company = args[1] if len(args) > 1 else kwargs.get('company', None)

        try:
            result = func(*args, **kwargs)
            logging.info(f"Attempting company: {company} in {func.__name__}: Success")
            return result
        except Exception as e:
            logging.info(f"Attempting company: {company} in {func.__name__}: Failure")
            # Uncomment below if error checking is neccesary.
            logging.error(f"Exception in {func.__name__}: {e}", exc_info=True)
            raise
    return wrapper

class Auto:
    def __init__(self, StateAbb, State, rateTables, nEffective, rEffective, NGICRatebook, NAFFRatebook, NACORatebook, NICOFRatebook, NWAGRatebook, MMRatebook,NAICSDescriptions, SchedRatingMod) -> None:
        self.StateAbb = StateAbb
        self.State = State
        self.rateTables = rateTables
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date
        self.NGICRatebook = NGICRatebook
        self.NAFFRatebook = NAFFRatebook
        self.NACORatebook = NACORatebook
        self.NICOFRatebook = NICOFRatebook
        self.MMRatebook = MMRatebook

        self.NAICSDescriptions = NAICSDescriptions
        self.SchedRatingMod = SchedRatingMod

        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'
        self.NAICS = '#####0'
        self.Territory92a = ""

        self.self_propelled_vehicles = [
            "1",
            "2",
            "3 to 4",
            "5 to 9",
            "10 to 14",
            "15 to 19",
            "20 to 29",
            "30 to 39",
            "40 to 49",
            "50 to 59",
            "60 to 69",
            "70 to 79",
            "80 to 89",
            "90 to 99",
            "100 to 114",
            "115 to 129",
            "130 to 154",
            "155 to 194",
            "195 to 289",
            "290 or greater"
        ]
        self.no_med_states = pd.read_excel(BA_INPUT_FILE, sheet_name="No MedPay")
        self.pip_states = pd.read_excel(BA_INPUT_FILE, sheet_name="PIP States")
        self.createMM()

    def createMM(self):
        # List of company codes
        company_codes = ["AICOA", "NMIC", "NICOA", "NPCIC"]

        #check that MM is in use
        if self.MMRatebook != "Not found":
            for company_code in company_codes:
                # Copy the entire MM rate book
                company_table = copy.deepcopy(self.rateTables["MM"])

                # Extract the CompanyDeviationFactor_Ext sheet from the MM rate book
                company_dev = pd.DataFrame(company_table["CompanyDeviationFactor_Ext"][1:], index=None,
                                           columns=company_table["CompanyDeviationFactor_Ext"][0])

                # Filter the rows to only include the row with the key '<company_code>_ext' in the UnderwritingCompanyCode column
                company_specific_dev = company_dev[company_dev['UnderwritingCompanyCode'] == f'{company_code.lower()}_ext']


                # Check that code only runs for MM companies that are present
                if not company_specific_dev.empty:
                    # Update the CompanyDeviationFactor_Ext sheet in the company table
                    company_table["CompanyDeviationFactor_Ext"] = [company_table["CompanyDeviationFactor_Ext"][
                                                                       0]] + company_specific_dev.values.tolist()

                    # Store the company table in self.rateTables
                    self.rateTables[company_code] = company_table

                else:
                    self.rateTables[company_code] = None
                    warnings.warn(f"Deviation factor for {company_code} not found. Skipping")

        else:
            for company_code in company_codes:
                self.rateTables[company_code] = None


    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if table exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if 'MM' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['MM'].keys(): # Checking if table exists in NICOF
                return pd.DataFrame(data=self.rateTables['MM'][tableCode][1:], index=None, columns=self.rateTables['MM'][tableCode][0])
        #if tableCode in self.rateTables['NAICS'].keys(): # Checking if NAICS Table
        #    return pd.DataFrame(data=self.rateTables['NAICS'][tableCode][1:], index=None, columns=self.rateTables['NAICS'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company


    def process_ratebook(self, name, rateTables):
        # Nesting protocol. Assign books to levels. If a lower level doesn't have a sheet, take it from the higher level.
        # Some sheets are naturally ignored. It is important that deep copies are used.
        LEVEL1 = rateTables[name]
        LEVEL2 = rateTables["NGIC"] or rateTables["NWAG"]
        LEVEL3 = rateTables["CW"]

        # Cascade LEVEL2 → LEVEL1: skip if LEVEL2 is unavailable (None) or if
        # this company IS the LEVEL2 source (every sheet already in LEVEL1).
        if LEVEL2 is not None and LEVEL1 is not LEVEL2:
            sheets_to_ignore = [sheet for sheet in LEVEL2.keys() if
                                (sheet in ["FlowChart", "Rate Book Details"]) and (sheet.startswith("CA7"))]

            for sheet in LEVEL2.keys():

                if (sheet not in sheets_to_ignore) and (sheet not in LEVEL1.keys()):
                    LEVEL1[sheet] = copy.deepcopy(LEVEL2[sheet])

        sheets_to_ignore_2 = []
        if LEVEL2 is not None:
            sheets_to_ignore_2 = [sheet for sheet in LEVEL2.keys() if (sheet in ["FlowChart", "Rate Book Details"]) and (sheet.startswith("CA7"))]  # Check for both String and Int values of 777 in the ratebook

        if LEVEL3 is None:
            raise RuntimeError(
                "CW (Countrywide) ratebook is missing from rateTables. "
                "Ensure the CW ratebook path is accessible before running."
            )

        for sheet in LEVEL3.keys():

            data = pd.DataFrame(LEVEL3[sheet][1:], index=None, columns=LEVEL3[sheet][0])
            table_not_empty = (data.iloc[0, 1] != 777 and data.iloc[0,1] != "777")  # Flag for if table is not used.
            table_not_CA7_etc = sheet not in ["FlowChart", "Rate Book Details"] and not sheet.startswith("CA7")

            if (sheet not in sheets_to_ignore_2) and (sheet not in LEVEL1.keys()) and (table_not_empty) and (table_not_CA7_etc):
                LEVEL1[sheet] = copy.deepcopy(LEVEL3[sheet])

        # Begginning of the LCM protocol. Basically goes through each sheet in the LCM list and applies the LCM/Company dev if applicable.
        LCM_sheet_mapping = pd.read_excel(BA_INPUT_FILE, sheet_name="LCM-Comp Dev Mapping")
        LCM_map = pd.read_excel(BA_INPUT_FILE, sheet_name="LCM-Map")

        liab_LCM = pd.DataFrame(self.rateTables[name]["LCMLiability_Ext"][1:], index=None,
                                columns=self.rateTables[name]["LCMLiability_Ext"][0]).iloc[0, 2]
        phys_LCM = pd.DataFrame(self.rateTables[name]["LCMPhysical Damage_Ext"][1:], index=None,
                                columns=self.rateTables[name]["LCMPhysical Damage_Ext"][0]).iloc[0, 2]

        company_dev = pd.DataFrame(self.rateTables[name]["CompanyDeviationFactor_Ext"][1:], index=None,
                                   columns=self.rateTables[name]["CompanyDeviationFactor_Ext"][0]).iloc[0, 2]

        lcm = {
            "liab": liab_LCM,
            "phys": phys_LCM,
            "other": 1
        }

        for i in range(0, LCM_sheet_mapping.shape[0]):

            sheet = LCM_sheet_mapping.loc[i, "sheetnames"]
            coverage = LCM_sheet_mapping.loc[i, "coverage"]
            lcm_app = LCM_sheet_mapping.loc[i, "lcm"]
            company_app = LCM_sheet_mapping.loc[i, "company_dev"]

            lcm_type = LCM_map[LCM_map["coverage"] == coverage].iloc[0, 1]

            lcm_value = 1
            cmp_value = 1

            if bool(lcm_app):
                lcm_value = round(float(lcm[lcm_type]),3)
            if bool(company_app):
                cmp_value = round(float(company_dev),3)

            try:
                data = LEVEL1[sheet]
                for line in data[1:]:  # Skip the header

                    if line[-1] is not None and line[-1] != "":
                        line[-1] *= lcm_value * cmp_value  # Apply the values to the last item in the list

                LEVEL1[sheet] = data

            except (KeyError):
                continue

        return name, LEVEL1

    def nesting(self):
        """
        -- Be very careful when editing. This is a core logic proccess supporting the code.
        Nesting protocol:
            Want to make a finalized rate book for NACO, NGIC, etc at start
            Levels: 1: Current Company 2: NGIC or NMIC 3: CW
            If in migration, look at any missing tables in NGIC -> copy those down into NACO
            Similarly look into CW for remaining missing tables and copy those down into NACO.

            Notable exceptions like consolidating to 2 levels if the current company is NGIC is accounted for.
            Does not copy down notably unneeded sheets like: ["FlowChart", "Rate Book Details","CA7"]
            or if the sheet is a dummy sheet containing 777.

            LCM Protocol: Goes through list of sheets that are flagged as LCM/Company Dev applicable. Then apply the multipliers.
        """
        ratebook_names = ['NAFF', 'NACO', 'NICOF', 'CCMIC', 'HICNJ', 'NICOA', 'AICOA', 'NPCIC', 'NMIC','NWAG','NGIC'] # Level 2 company needs to be last.
        ratebooks = [self.NAFFRatebook, self.NACORatebook, self.NICOFRatebook, self.rateTables['CCMIC'], self.rateTables['HICNJ'], self.rateTables["NICOA"], self.rateTables["AICOA"], self.rateTables["NPCIC"],self.rateTables["NMIC"], self.rateTables["NWAG"], self.NGICRatebook] # level2 needs to be last

        available_names = []
        available_books = []
        for name, book in zip(ratebook_names, ratebooks):
            if book != "Not found" and book is not None:
                available_books.append(book)
                available_names.append(name)

        for name in available_names:
            name, LEVEL1 = self.process_ratebook(name, self.rateTables)
            self.rateTables[name] = LEVEL1

        # When MM runs we don't want to create the NGIC version of pages. So delete book before system recognizes it.
        # Same thing for CCMIC.
        if (self.rateTables["NMIC"] is not None) or (self.rateTables["CCMIC"] is not None):
            self.rateTables["NGIC"] = None
            self.NGICRatebook = None

    def compareCompanies(self, tableCode):
        """
        Enhanced comparison method that supports comparing multiple tableCodes between companies for clustering.
        If tableCode is a string, legacy functionality is preserved.
        If tableCode is a list, all tables must match for two companies to cluster.
        """

        CompaniesCheck = ['NGIC', 'NACO', 'NAFF', 'CCMIC', 'HICNJ', 'NICOF',
                          'NMIC', 'AICOA', 'NICOA', 'NPCIC']

        self.default_company = ["NGIC"]
        self.CompanyListDif = []
        self.RemainingCompanies = ""

        # --- Normalize input for backward compatibility ---
        if isinstance(tableCode, str):
            tableCodes = [tableCode]  # legacy mode
        else:
            tableCodes = list(tableCode)  # ensure iterable

        # ---------- Table comparison helpers ----------
        def compare_two_tables(company1, company2, code):
            """Compare a single tableCode for two companies."""
            table1 = pd.DataFrame(self.rateTables[company1][code])
            table2 = pd.DataFrame(self.rateTables[company2][code])

            table1 = table1.sort_values(by=table1.columns.tolist()).dropna().reset_index(drop=True)
            table2 = table2.sort_values(by=table2.columns.tolist()).dropna().reset_index(drop=True)

            table1 = table1.map(lambda x: round(x, 3) if isinstance(x, (int, float)) else x)
            table2 = table2.map(lambda x: round(x, 3) if isinstance(x, (int, float)) else x)

            def tables_equal(t1, t2, tol=1e-9):
                if t1.shape != t2.shape:
                    return False

                for col in t1.columns:
                    if not t1[col].equals(t2[col]):
                        # handle float columns with tolerance
                        if t1[col].dtype == 'float64' or t2[col].dtype == 'float64':
                            if not (abs(t1[col] - t2[col]) < tol).all():
                                return False
                        else:
                            return False
                return True

            return tables_equal(table1, table2)

        def compare_all_tables(company1, company2):
            """Return True only if ALL tableCodes match for the two companies."""
            for code in tableCodes:
                if not compare_two_tables(company1, company2, code):
                    return False
            return True

        # ---------- Determine which companies exist ----------
        self.existing_companies = [
            company for company in CompaniesCheck
            if company in self.rateTables and self.rateTables[company] is not None
        ]

        # Fallback logic if default company missing
        if self.default_company[0] not in self.existing_companies:
            if "NMIC" in self.existing_companies:
                self.default_company = ["NMIC"]
            elif "CCMIC" in self.existing_companies:
                self.default_company = ["CCMIC"]
            elif "NWAG" in self.existing_companies:
                self.default_company = ["NWAG"]

        # ---------- Cluster companies ----------
        groups = []

        for company in self.existing_companies:
            found_group = False
            for group in groups:
                representative = group[0]
                if compare_all_tables(company, representative):
                    group.append(company)
                    found_group = True
                    break
            if not found_group:
                groups.append([company])

        # Convert groups into output format
        result = [",".join(sorted(group)) for group in groups]

        if len(groups) == 1:
            self.CompanyListDif = self.default_company
        else:
            self.CompanyListDif = result

    # Builds the Expense Constant table
    # Returns a dataframe
    @log_exceptions
    def buildExpenseConstant(self, company):
        ExpenseConstant = pd.DataFrame(self.rateTables[company]['ExpenseConstant_Ext'][1:], index=None, columns=self.rateTables[company]['ExpenseConstant_Ext'][0])
        val = ExpenseConstant.iloc[0, 1]
        ExpenseConstant = ExpenseConstant.astype(object)   # allow mixed string/numeric
        ExpenseConstant = ExpenseConstant.astype(object)
        ExpenseConstant.iloc[0, 1] = f"${val:.0f}"

        return ExpenseConstant.filter(items=['Rate'])


    def simple_long_table_build(self, rate_tables, company, sheet_names, new_column_name, orig_values, replace_values, filter_values=None):
        """
        Processes multiple sheets from a rate table, renames the first column, replaces values, and concatenates the results.
        This function was specifically made to make stacking/formatting of tables in ratebook sheet forms of:

        Constant    Factor
        Y           X.X

        for rules 93 and 97 making them easier to implement. It will grab from a sheet only 1 line, so you may have to repeatedly visit a sheet
        to fully build a table.

        Additionally it can handle tables of the form

        Constant    Category    Factor
        Y           {C}         X.X

        by transforming this into:

        Constant            Factor
        Y + ' ' + {C}        X.X

        by concatenating the categorical columns together. View the notes section for an example of this use case.

        Parameters:
        rate_tables (dict): A dictionary containing rate tables for different companies.
        company (str): The company for which the rate tables are being processed.
        sheet_names (list of str): A list of sheet names to be processed.
        new_column_name (str): The new name for the first column. (Usually Constant)
        orig_values (list of str): A list of values to be replaced in the first column.
        replace_values (list of str): A list of values to replace in the first column.
        filter_values (list of str, optional): A list of values to filter the first column. Defaults to None.

        Notes:
            Often the new column name for the first column is left blank as it is redundant.

            Transpose is often used to flip a table for it to be in the format: Factor X

            Florida has a wide variety of example use cases
                - For simple stacking of two constant factors across two sheets: FL-293.B.2. No-Fault Factors
                - For multiple lines in a single sheet: Fl-293.C.2.a. Exclusion Of Disability Benefits Only Factors

            If there are two identifier columns it will concatenate them together with a space in between.
                - See Florida Rule 293.D.2.a for example

            For more advanced matrices as seen in rule 97 this function is not appropriate. Use pivot_table instead.

        Returns:
        pd.DataFrame: A concatenated DataFrame with the processed sheets.

        """
        tables = []
        for sheet_name, orig_value, replace_value, filter_value in zip(sheet_names, orig_values, replace_values, filter_values or [None] * len(sheet_names)):
            table = pd.DataFrame(rate_tables[company][sheet_name][1:], index=None, columns=rate_tables[company][sheet_name][0])
            first_column_name = table.columns[0]

            if len(table.columns) > 2:
                table[first_column_name] = table[first_column_name] + ' ' + table[table.columns[1]]
                table = table.drop(columns=[table.columns[1]])

            table = table.rename(columns={first_column_name: new_column_name})
            if filter_value:
                table = table[table[new_column_name] == filter_value]
            table = table.replace({new_column_name: {orig_value: replace_value}})
            tables.append(table)
        output_table = pd.concat(tables)
        return output_table


    # Builds the general Base Rate table
    # Returns a dataframe
    @log_exceptions
    def buildBaseRates(self, company, rate_type : str):
        """
        General base rate build code for rules: 222, 232, and 239.

        Description:
        The buildBaseRates method generates base rate tables for various types of vehicles and coverages.
        It processes rate tables from an Excel file and applies filters and mappings based on the state and coverage type.

        Parameters:
        - company (str): The name of the company for which the rate tables are being processed.
        - type (str): The type of vehicle for which the base rates are being built. Possible values include:
          - "TTT": Tractor-Trailer Trucks
          - "PPT": Private Passenger Types
          - "School Buses"
          - "Van Pools"
          - "Other Buses"
          - "Taxis"

        Returns:
        - output_table (pd.DataFrame): A DataFrame containing the base rates for the specified vehicle type and coverages.

        Method Overview:
        1. Mapping Sheet Names: The method uses a dictionary to map the type parameter to corresponding sheet names in the Excel file.
        2. Coverage Types: It defines a list of coverage types to be processed: ["csl", "med", "pip", "coll", "otc"].
                    a. Specified Causes Of Loss and Comprehensive are defined as groups within the OTC sheet.
        3. State Mapping: The method reads the state-specific mappings from the Excel file and applies them to filter the rate tables.
        4. Loading Rate Tables: It loads the relevant rate tables for each coverage type from the Excel file.
        5. Filtering and Grouping: The method applies filters and groups the rate tables based on the state-specific mappings.
            a. Groups are different categories within the same column of a sheet that you want seperate tables for.
            b. Filters define the level of data we want to see. Like grabbing a specific Med Pay Limit, Fault/No Fault, etc.
        6. Renaming Columns: It renames the columns of the rate tables based on predefined mappings.
        7. Merging Tables: Finally, it merges the filtered and renamed rate tables into a single output table.

        :param company: The name of the company for which the rate tables are being processed.
        :param type: The type of vehicle for which the base rates are being built.
        :return: A DataFrame containing the base rates for the specified vehicle type and coverages.
        """

        type_sheet = {
            "TTT" : "222 TTT",
            "PPT" : "232 PPT",
            "School Buses" : "239 School Buses",
            "Van Pools" : "239 Van Pools",
            "Other Buses" : "239 Other Buses",
            "Taxis" : "239 Taxis"
        }

        coverages = ["csl", "med", "pip", "ppi","coll", "otc"]

        sheet_name = type_sheet[rate_type]
        state_mapper = pd.read_excel(BA_INPUT_FILE, sheet_name=sheet_name)

        mappers = {}
        for coverage in coverages:
            filtered_map = state_mapper[(state_mapper["state"] == self.StateAbb) & (state_mapper["coverage"] == coverage)]
            default_map = state_mapper[(state_mapper["state"] == "Default") & (state_mapper["coverage"] == coverage)]
            if not filtered_map.empty:
                mappers[coverage] = filtered_map
            else:
                mappers[coverage] = default_map

        # Auto-populate rate book sheet names based on the input file

        liabTable = None
        otcTable = None
        collTable = None
        pipTable = None
        medTable = None
        ppiTable = None

        # loads the sheets which may not exist. Then makes sure the territory column is first in order.
        def reorder_columns(df):
            first_col = ['Territory']
            last_cols = [col for col in df.columns if ('premium' in col.lower())]
            middle_cols = [col for col in df.columns if (col not in first_col + last_cols)]
            return df.reindex(columns=first_col + middle_cols + last_cols)

        try:
            LIAB_RATE_BOOK_SHEET_NAME = mappers['csl']["sheet"].values[0]
            liabTable = pd.DataFrame(self.rateTables[company][LIAB_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][LIAB_RATE_BOOK_SHEET_NAME][0])
            liabTable = reorder_columns(liabTable).dropna()
        except (KeyError, IndexError) as e:
            # traceback.print_exc()
            message = f"{rate_type} Liability base rates have failed."
            warnings.warn(message, CSLWarning)
            pass

        try:
            OTC_RATE_BOOK_SHEET_NAME = mappers["otc"]['sheet'].values[0]
            otcTable = pd.DataFrame(self.rateTables[company][OTC_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][OTC_RATE_BOOK_SHEET_NAME][0])
            otcTable = reorder_columns(otcTable).dropna()
        except (KeyError, IndexError) as e:
            # traceback.print_exc()
            warnings.warn(f"{rate_type} OTC base rates have failed.", OTCWarning)
            pass

        try:
            COLL_RATE_BOOK_SHEET_NAME = mappers["coll"]['sheet'].values[0]
            collTable = pd.DataFrame(self.rateTables[company][COLL_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][COLL_RATE_BOOK_SHEET_NAME][0])
            collTable = reorder_columns(collTable).dropna()

        except (KeyError, IndexError) as e:
            # traceback.print_exc()
            warnings.warn(f"{rate_type} Collision base rates have failed.", COLLWarning)
            pass

        try:
            PIP_RATE_BOOK_SHEET_NAME = mappers["pip"]['sheet'].values[0]
            pipTable = pd.DataFrame(self.rateTables[company][PIP_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][PIP_RATE_BOOK_SHEET_NAME][0])
            pipTable = reorder_columns(pipTable).dropna()

        except (KeyError, IndexError) as e:
            # traceback.print_exc()
            if self.StateAbb != "OR": # OR will always have this error when it shouldn't Special Case later in method.
                warnings.warn(f"{rate_type} PIP base rates have failed. This is intended if PIP is not in your state.", PIPWarning)
            pass

        try:

            if self.StateAbb in self.no_med_states["states"].values:
                # Michigan has the sheet but not any values.
                pass

            else:
                MED_RATE_BOOK_SHEET_NAME = mappers['med']['sheet'].values[0]
                medTable = pd.DataFrame(self.rateTables[company][MED_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][MED_RATE_BOOK_SHEET_NAME][0])

                # Check if the data is from Utah and select only the first three columns
                # Do not know why or how this is occuring. Reading in a None column.
                medTable = reorder_columns(medTable).dropna()

        except (KeyError, IndexError) as e:
            # traceback.print_exc()
            warnings.warn(f"{rate_type} Medpay base rates have failed. This is intended if Medpay is not in your state.", MEDPAYWarning)
            pass

        try:
            # Below is a Michigan only coverage -- Property Protection Insurance (PPI)
            PPI_RATE_BOOK_SHEET_NAME = mappers['ppi']["sheet"].values[0]
            ppiTable = pd.DataFrame(self.rateTables[company][PPI_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][PPI_RATE_BOOK_SHEET_NAME][0])

        except (KeyError, IndexError) as e:
            # traceback.print_exc()
            if self.StateAbb == "MI":
                warnings.warn(f"{rate_type} Property Insurance base rates have failed.")

        # Special case for DC in which "Not Applicable" appears in two columns.
        public = ["School Buses", "Van Pools", "Other Buses", "Taxis"]
        if (self.StateAbb == "DC") and (rate_type in public):
            pipTable = pipTable[pipTable["PIPMedicalExpenseBenefits"] != "Not Applicable"]

        # If you think the below code is complicated, there was no better way to do it.

        # First it looks at 'groups' they are duplications of the base rate tables generally because of options. Ex: KY liab A,B,C.
        # Then it stores all those in tables_list
        # Next applies filters in the input file to a respective group.
        master_coverage_list = []
        master_table_list = []

        for coverage in coverages:

            tables_list = []
            tables_group_idx = []  # NEW: parallel list tracking which group produced each table

            # Create a table for every group in the input file
            groups = mappers[coverage][["group1", "group2", "group3"]]
            group_empty = pd.isnull(groups).all().all()

            if not group_empty:
                # Make the tables across each group with group-aware tracking
                for _, group_row in groups.iterrows():  # Iterate over mapping rows
                    for col_name, group in group_row.items():  # Iterate over 'group1'/'group2'/'group3' in the row
                        if pd.isnull(group):
                            continue

                        table = None
                        if coverage == "csl" and liabTable is not None:
                            table = liabTable.copy()
                        elif coverage == "med" and medTable is not None:
                            table = medTable.copy()
                        elif coverage == "coll" and collTable is not None:
                            table = collTable.copy()
                        elif coverage == "otc" and otcTable is not None:
                            table = otcTable.copy()
                        elif coverage == "pip" and pipTable is not None:
                            table = pipTable.copy()
                        elif coverage == "ppi" and ppiTable is not None:
                            table = ppiTable.copy()

                        if table is None:
                            continue

                        # Apply the group selection: find the column containing this group's value, filter, drop that column,
                        # and rename the last column to the group label (unchanged behavior).
                        for column in table.columns:
                            if group in table[column].to_list():
                                table = table[table[column] == group]
                                table = table.drop(columns=[column])
                                table.rename(columns={table.columns[-1]: group}, inplace=True)
                                break  # stop after the first match for this group value

                        tables_list.append(table)
                        # Extract the group index from 'group1'/'group2'/'group3' → 1/2/3
                        gi = int(col_name.replace("group", ""))
                        tables_group_idx.append(gi)
            else:
                # If all groups are null, use the original table without filtering on group,
                # but mark group index as 0 to trigger fallback to global filters.
                if coverage == "csl" and liabTable is not None:
                    tables_list.append(liabTable)
                    tables_group_idx.append(0)
                elif coverage == "med" and medTable is not None:
                    tables_list.append(medTable)
                    tables_group_idx.append(0)
                elif coverage == "coll" and collTable is not None:
                    tables_list.append(collTable)
                    tables_group_idx.append(0)
                elif coverage == "otc" and otcTable is not None:
                    tables_list.append(otcTable)
                    tables_group_idx.append(0)
                elif coverage == "pip" and pipTable is not None:
                    tables_list.append(pipTable)
                    tables_group_idx.append(0)
                elif coverage == "ppi" and ppiTable is not None:
                    tables_list.append(ppiTable)
                    tables_group_idx.append(0)

            # ---- GROUP-TARGETED FILTERS ----
            # For each grouped table, apply its own filters: filter_<group_index>_1 .. filter_<group_index>_3.
            # If group_index == 0 (no groups present), fallback to the original global filters: filter1..filter3.
            for i, table in enumerate(tables_list):
                gi = tables_group_idx[i]

                if gi == 0:
                    # Backward-compatible fallback: original indiscriminate filters
                    filters = mappers[coverage][["filter1", "filter2", "filter3"]]
                else:
                    # NEW: group-specific filters for group index gi
                    filters = mappers[coverage][[f"filter_{gi}_1", f"filter_{gi}_2", f"filter_{gi}_3"]]

                filter_empty = pd.isnull(filters).all().all()
                if filter_empty:
                    continue

                # Apply filters sequentially to this table
                for _, filter_row in filters.iterrows():  # iterate rows in the mapping selection
                    for filt in filter_row:  # iterate filter_?_? values in that row
                        if pd.isnull(filt):
                            continue
                        for column in table.columns:  # find the column containing this filter value
                            if filt in table[column].values:
                                table = table[table[column] == filt]
                                table = table.drop(columns=[column])
                                break  # stop after first match for this filter value

                # Save the updated table back to the list
                tables_list[i] = table

            # As before: record coverage + its (now group-filtered) tables
            master_coverage_list.append(coverage)
            master_table_list.append(tables_list)

        # Pip, ppi, specified and comp names are auto set
        column_name_map = {
            "TTT" : {"csl" : "Liability (222.C.2.a.)",
                     "med" : "Medical Payments (222.C.4.a.(2)., 222.C.4.b.(1).)",
                     "coll" : "Collision (222.C.3.c.)",
                     "Specified Causes Of Loss" : "Specified Causes of Loss (222.C.3.c.)",
                     "Comprehensive" : "Comprehensive (222.C.3.c.)"},

            "PPT": {"csl": "Liability (232.B.1.a.)",
                    "med": "Medical Payments (232.B.1.a.)",
                    "coll": "Collision (222.C.3.c.)",
                    "Comprehensive": "Comprehensive (232.B.4.a.)"},

            "School Buses": {"csl": "Liability (239.C.2.a.)",
                             "med": "Medical Payments (239.C.4.a.)",
                             "coll": "Collision (239.C.3.a.)",
                             "Specified Causes Of Loss": "Specified Causes of Loss (239.C.3.a.)",
                             "Comprehensive": "Comprehensive (239.C.3.a.)"},

            "Van Pools": {"csl": "Liability (239.C.2.a.)",
                          "med": "Medical Payments (239.C.4.a.)",
                          "coll": "Collision (239.C.3.a.)",
                          "Specified Causes Of Loss": "Specified Causes of Loss (239.C.3.a.)",
                          "Comprehensive": "Comprehensive (239.C.3.a.)"},

            "Other Buses": {"csl": "Liability (239.C.2.a.)",
                            "med": "Medical Payments (239.C.4.a.)",
                            "coll": "Collision (239.C.3.a.)",
                            "Specified Causes Of Loss": "Specified Causes of Loss (239.C.3.a.)",
                            "Comprehensive": "Comprehensive (239.C.3.a.)"},

            "Taxis" : {"csl": "Liability (239.C.2.a.)",
                       "med": "Medical Payments (239.C.4.a.)",
                       "coll": "Collision (239.C.3.a.)",
                       "Specified Causes Of Loss": "Specified Causes of Loss (239.C.3.a.)",
                       "Comprehensive": "Comprehensive (239.C.3.a.)"},
        }

        column_rename_dict = column_name_map[rate_type]

        for coverage, tables in zip(master_coverage_list, master_table_list):
            for table in tables:
                if coverage in ["csl", "med", "coll"] and table is not None:
                    # Directly apply the column name for csl, med, and coll if BasePremium
                    new_column_name = column_rename_dict[coverage]
                    right_column_name = (table.columns[-1] == "BasePremium") or (table.columns[-1] == "Premium")
                    last_col = table.columns[-1]
                    table.columns = [new_column_name if ((col == last_col) and right_column_name) else col for col in table.columns]
                elif coverage in ["otc"] and table is not None:
                    # Apply the column name only if the column exists in the table
                    if rate_type == "PPT":
                        new_column_name = column_rename_dict["Comprehensive"]
                        table.columns = [new_column_name if col == table.columns[-1] else col for col in table.columns]
                    else:
                        for coverage_sub in ["Specified Causes Of Loss", "Comprehensive"]:
                            if coverage_sub in table.columns:
                                new_column_name = column_rename_dict[coverage_sub]
                                table.columns = [new_column_name if col == table.columns[-1] else col for col in table.columns]

        output_table = None
        for tables in master_table_list:
            for table in tables:
                if output_table is None and table is not None:
                    output_table = table
                elif table is not None:
                    output_table['Territory'] = output_table['Territory'].astype(int).astype(str) # Removing decimals then making it a string.
                    table['Territory'] = table['Territory'].astype(int).astype(str) # Removing decimals then making it a string.
                    output_table = output_table.merge(table, on="Territory", how="inner")

        output_table = output_table.loc[~(output_table == "No Coverage").all(axis=1)]

        if self.StateAbb == "OR" and rate_type == "TTT":
            # Oregon has a special PIP factor that gets applied to TTT CSL for the PIP items.
            factors = pd.DataFrame(self.rateTables[company]['PIPOtherThanPrivatePassengerAndGarageFactor'][1:], index=None, columns=self.rateTables[company]['PIPOtherThanPrivatePassengerAndGarageFactor'][0])
            npobe_factors = factors.loc[factors["TruckPIPRatingBasis"] == "Not Principally Operated by Employees", "Factor"].iloc[0]
            pobe_factors = factors.loc[factors["TruckPIPRatingBasis"] == "Principally Operated by Employees", "Factor"].iloc[0]
            output_table["Principally Operated By Employees"] = output_table["Liability (222.C.2.a.)"] * pobe_factors
            output_table["Not Principally Operated By Employees"] = output_table["Liability (222.C.2.a.)"] * npobe_factors

            liability_col = "Liability (222.C.2.a.)"
            liability_index = output_table.columns.get_loc(liability_col)

            output_table.insert(
                liability_index + 1,
                "Principally Operated By Employees",
                output_table[liability_col] * pobe_factors
            )
            output_table.insert(
                liability_index + 2,
                "Not Principally Operated By Employees",
                output_table[liability_col] * npobe_factors
            )


        for column in output_table.columns:
            if output_table[column].dtype == 'float64' or output_table[column].dtype == 'int64':
                output_table[column] = output_table[column].apply(lambda x: f"{x:.2f}")

        # Replacing 0's with NAs. Mainly for taxi pip.
        output_table = output_table.map(lambda x: "NA" if isinstance(x, (int, float)) and (abs(x) < 1e-6) else x)
        output_table = output_table.map(lambda x: "NA" if x == "0.00" else x)


        return output_table

    # Builds the TTT NAICS table
    # Returns a dataframe
    @log_exceptions
    def buildTTTNAICSFactors(self, company):
        NAICSReference = pd.DataFrame(data=self.NAICSDescriptions).astype({'NAICS Six-Digit Code': 'int64'}) # Converting to int first to get rid of decimal places
        NAICSCoverages = pd.DataFrame(self.rateTables[company]['NAICSFactors_Ext'][1:], index=None, columns=self.rateTables[company]['NAICSFactors_Ext'][0])
        NAICSCoverages = NAICSCoverages.pivot(index='NAICSCode', columns='VehicleAndCoverageType', values='Factor').reset_index('NAICSCode').rename(columns={'NAICSCode' : 'NAICS Six-Digit Code'}).filter(items=['NAICS Six-Digit Code', 'NAICS Category', 'Trucks, Tractors, And Trailers Liability', 'Trucks, Tractors, And Trailers Comprehensive And Specified Causes Of Loss', 'Trucks And Truck-tractors Collision', 'Trailers Collision', 'Private Passenger Types Liability', 'Private Passenger Types Collision', 'Private Passenger Types Comprehensive']).astype({'NAICS Six-Digit Code': 'int64'})
        #NAICSCoverages = self.buildDataFrame("NAICSFactors_Ext").pivot(index='NAICSCode', columns='VehicleAndCoverageType', values='Factor').reset_index('NAICSCode').rename(columns={'NAICSCode' : 'NAICS Six-Digit Code'}).filter(items=['NAICS Six-Digit Code', 'NAICS Category', 'Trucks, Tractors, And Trailers Liability', 'Trucks, Tractors, And Trailers Comprehensive And Specified Causes Of Loss', 'Trucks And Truck-tractors Collision', 'Trailers Collision']).astype({'NAICS Six-Digit Code': 'int64'})
        NAICSCoverages = NAICSCoverages.astype(object)
        NAICSCoverages.iloc[:, 1:] = NAICSCoverages.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")        #NAICSCoverages = self.buildDataFrame("NAICSFactors_Ext").pivot(index='NAICSCode', columns='VehicleAndCoverageType', values='Factor').reset_index('NAICSCode').rename(columns={'NAICSCode' : 'NAICS Six-Digit Code'}).filter(items=['NAICS Six-Digit Code', 'NAICS Category', 'Trucks, Tractors, And Trailers Liability', 'Trucks, Tractors, And Trailers Comprehensive And Specified Causes Of Loss', 'Trucks And Truck-tractors Collision', 'Trailers Collision']).astype({'NAICS Six-Digit Code': 'int64'})
        TTTNAICSFactors = pd.merge(NAICSReference, NAICSCoverages, on = 'NAICS Six-Digit Code', how = 'inner')
        TTTNAICSFactors.drop(columns = "NAICS Definition", inplace = True)
        return TTTNAICSFactors

    # Builds the TTT Liability Fleet Size Factors table
    # Returns a dataframe

    @log_exceptions
    def buildTTTLiabFleetFactors(self, company):
        TTTLiabFleetFactors = pd.DataFrame(self.rateTables[company]['LiabilityFleetSizeFactors_Ext'][1:], index=None, columns=self.rateTables[company]['LiabilityFleetSizeFactors_Ext'][0])
        TTTLiabFleetFactors = TTTLiabFleetFactors.query(f'VehicleType != "Private Passenger"')
        #TTTLiabFleetFactors = self.buildDataFrame("LiabilityFleetSizeFactors_Ext").query(f'VehicleType != "Private Passenger"')
        pivotedTTTLiabFleetFactors = TTTLiabFleetFactors.pivot(index='NumberOfPoweredVehicles', columns='VehicleType', values='Factor').reset_index('NumberOfPoweredVehicles').rename(columns={'NumberOfPoweredVehicles' : 'Number of Powered Vehicles'}).filter(items=['Number of Powered Vehicles', 'Light Truck', 'Medium Truck', 'Heavy Truck', 'Extra-Heavy Truck', 'Heavy Truck-Tractor', 'Extra Heavy Truck-Tractor', 'Semitrailer', 'Trailer', 'Service or Utility Trailer'])
        pivotedTTTLiabFleetFactors["Number of Powered Vehicles"] = self.self_propelled_vehicles

        return pivotedTTTLiabFleetFactors

    # Builds the TTT PhysDam Fleet Size Factors table
    # Returns a dataframe
    @log_exceptions
    def buildTTTPhysDamFleetFactors(self, company):
        TTTCollTruckFleetFactors = pd.DataFrame(self.rateTables[company]['CollisionFleetSizeFactor_Ext'][1:], index=None, columns=self.rateTables[company]['CollisionFleetSizeFactor_Ext'][0])
        TTTCollTruckFleetFactors = TTTCollTruckFleetFactors.query(f'VehicleType == "Trucks And Truck-tractors"').drop(columns='VehicleType').pivot(index='NumberOfPoweredVehicles', columns='TruckBusinessUseClass', values='Factor').reset_index('NumberOfPoweredVehicles')
        TTTCollTrailerFleetFactors = pd.DataFrame(self.rateTables[company]['CollisionFleetSizeFactor_Ext'][1:], index=None, columns=self.rateTables[company]['CollisionFleetSizeFactor_Ext'][0])
        TTTCollTrailerFleetFactors = TTTCollTrailerFleetFactors.query(f'VehicleType == "Trailer Types"').drop(columns='VehicleType').pivot(index='NumberOfPoweredVehicles', columns='TruckBusinessUseClass', values='Factor').reset_index('NumberOfPoweredVehicles')
        TTTSCoLFleetFactors = pd.DataFrame(self.rateTables[company]['ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext'][1:], index=None, columns=self.rateTables[company]['ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext'][0])
        TTTSCoLFleetFactors = TTTSCoLFleetFactors.query(f'VehicleType == "Trucks, Tractors And Trailers"').drop(columns='VehicleType').rename(columns={'Factor' : 'Other Than Collision (All Vehicle Types)'})
        TTTCollFleetFactors = pd.merge(TTTCollTruckFleetFactors, TTTCollTrailerFleetFactors, on = 'NumberOfPoweredVehicles', how = 'inner')
        TTTFleetFactors = pd.merge(TTTCollFleetFactors, TTTSCoLFleetFactors, on = 'NumberOfPoweredVehicles', how = 'inner').rename(columns={'NumberOfPoweredVehicles' : 'Number of Powered Vehicles', 'Commercial' : 'Commercial Use', 'Not Applicable_x' : 'Extra-Heavy Vehicles (All Uses)', 'Retail': 'Retail Use', 'Service' : 'Service Use', 'Not Applicable_y' : 'Trailer Types'}).filter(items=['Number of Powered Vehicles', 'Service Use', 'Retail Use', 'Commercial Use', 'Extra-Heavy Vehicles (All Uses)', 'Trailer Types', 'Other Than Collision (All Vehicle Types)'])


        TTTFleetFactors["Number of Powered Vehicles"] = self.self_propelled_vehicles
        TTTFleetFactors = TTTFleetFactors.drop(columns = 'Other Than Collision (All Vehicle Types)') # This is for ISO Currency where it has been moved to the below table.

        return TTTFleetFactors

    def buildTTTOTCFleetFactors(self, company):
        table = pd.DataFrame(self.rateTables[company]['ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext'][1:], index = None, columns = self.rateTables[company]['ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext'][0])
        table = table[table["VehicleType"] == "Trucks, Tractors And Trailers"].sort_values(by = "NumberOfPoweredVehicles")
        table["NumberOfPoweredVehicles"] = self.self_propelled_vehicles
        table = table.drop(columns = "VehicleType").rename(columns={'NumberOfPoweredVehicles' : 'Number of Powered Vehicles', 'Factor' : 'Trucks, Tractors And Trailers'})

        return table

    # Builds the TTT PhysDam Fleet Size Factors table
    # Returns a dataframe
    @log_exceptions
    def buildPrimaryFactors(self, company):
        TTTPrimaryFactorsLocal = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][0]).query(f'TruckRadiusClass == "Local"')
        #TTTPrimaryFactorsLocal = self.buildDataFrame("TrucksTractorsAndTrailersPrimaryFactors_Ext").query(f'TruckRadiusClass == "Local"')
        TTTPrimaryFactorsLocal['ClassCode'] = TTTPrimaryFactorsLocal['TruckSizeClass'] + "-" + TTTPrimaryFactorsLocal['TruckBusinessUseClass']
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Light Truck-Service", 'Class (Non-Fleet, Fleet)'] = '011, 014'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Light Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '021, 024'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Light Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '031, 034'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Medium Truck-Service", 'Class (Non-Fleet, Fleet)'] = '211, 214'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Medium Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '221, 224'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Medium Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '231, 234'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Heavy Truck-Service", 'Class (Non-Fleet, Fleet)'] = '311, 314'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Heavy Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '321, 324'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Heavy Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '331, 334'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Extra-Heavy Truck-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '401, 404'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Heavy Truck-Tractor-Service", 'Class (Non-Fleet, Fleet)'] = '341, 344'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Heavy Truck-Tractor-Retail", 'Class (Non-Fleet, Fleet)'] = '351, 354'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Heavy Truck-Tractor-Commercial", 'Class (Non-Fleet, Fleet)'] = '361, 364'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Extra Heavy Truck-Tractor-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '501, 504'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Semitrailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '671, 674'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '681, 684'
        TTTPrimaryFactorsLocal.loc[TTTPrimaryFactorsLocal['ClassCode'] == "Service or Utility Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '691, 694'
        TTTPrimaryFactorsLocal = TTTPrimaryFactorsLocal.pivot(index=['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)'], columns='Coverage', values='Factor').reset_index(['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)']). \
            rename(columns={'TruckSizeClass' : 'Size Class', 'TruckBusinessUseClass' : 'Business Use Class', 'Collision' : 'Collision Factor', 'Comprehensive And Specified Causes Of Loss' : 'OTC Factor', 'Liability' : 'Liability Factor'}).sort_values('Class (Non-Fleet, Fleet)')
        TTTPrimaryFactorsLocal = TTTPrimaryFactorsLocal[['Size Class', 'Business Use Class', 'Class (Non-Fleet, Fleet)', 'Liability Factor', 'OTC Factor', 'Collision Factor']]

        TTTPrimaryFactorsInt = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][0]).query(f'TruckRadiusClass == "Intermediate"')
        #TTTPrimaryFactorsInt = self.buildDataFrame("TrucksTractorsAndTrailersPrimaryFactors_Ext").query(f'TruckRadiusClass == "Intermediate"')
        TTTPrimaryFactorsInt['ClassCode'] = TTTPrimaryFactorsInt['TruckSizeClass'] + "-" + TTTPrimaryFactorsInt['TruckBusinessUseClass']
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Light Truck-Service", 'Class (Non-Fleet, Fleet)'] = '012, 015'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Light Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '022, 025'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Light Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '032, 035'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Medium Truck-Service", 'Class (Non-Fleet, Fleet)'] = '212, 215'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Medium Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '222, 225'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Medium Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '232, 235'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Heavy Truck-Service", 'Class (Non-Fleet, Fleet)'] = '312, 315'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Heavy Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '322, 325'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Heavy Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '332, 335'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Extra-Heavy Truck-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '402, 405'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Heavy Truck-Tractor-Service", 'Class (Non-Fleet, Fleet)'] = '342, 345'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Heavy Truck-Tractor-Retail", 'Class (Non-Fleet, Fleet)'] = '352, 355'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Heavy Truck-Tractor-Commercial", 'Class (Non-Fleet, Fleet)'] = '362, 365'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Extra Heavy Truck-Tractor-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '502, 505'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Semitrailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '672, 675'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '682, 685'
        TTTPrimaryFactorsInt.loc[TTTPrimaryFactorsInt['ClassCode'] == "Service or Utility Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '692, 695'
        TTTPrimaryFactorsInt = TTTPrimaryFactorsInt.pivot(index=['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)'], columns='Coverage', values='Factor').reset_index(['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)']). \
            rename(columns={'TruckSizeClass' : 'Size Class', 'TruckBusinessUseClass' : 'Business Use Class', 'Collision' : 'Collision Factor', 'Comprehensive And Specified Causes Of Loss' : 'OTC Factor', 'Liability' : 'Liability Factor'}).sort_values('Class (Non-Fleet, Fleet)')
        TTTPrimaryFactorsInt = TTTPrimaryFactorsInt[['Size Class', 'Business Use Class', 'Class (Non-Fleet, Fleet)', 'Liability Factor', 'OTC Factor', 'Collision Factor']]
        TTTPrimaryFactorsLong = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][0]).query(f'TruckRadiusClass == "Long Distance"')
        #TTTPrimaryFactorsLong = self.buildDataFrame("TrucksTractorsAndTrailersPrimaryFactors_Ext").query(f'TruckRadiusClass == "Long Distance"')
        TTTPrimaryFactorsLong['ClassCode'] = TTTPrimaryFactorsLong['TruckSizeClass'] + "-" + TTTPrimaryFactorsLong['TruckBusinessUseClass']
        TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Light Truck-Service", 'Class (Non-Fleet, Fleet)'] = '013, 016'
        TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Light Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '023, 026'
        TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Light Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '033, 036'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Medium Truck-Service", 'Class (Non-Fleet, Fleet)'] = '213, 216'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Medium Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '223, 226'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Medium Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '233, 236'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Service", 'Class (Non-Fleet, Fleet)'] = '313, 316'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '323, 326'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '333, 336'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Extra-Heavy Truck-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '403, 406'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Tractor-Service", 'Class (Non-Fleet, Fleet)'] = '343, 346'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Tractor-Retail", 'Class (Non-Fleet, Fleet)'] = '353, 356'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Tractor-Commercial", 'Class (Non-Fleet, Fleet)'] = '363, 366'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Extra Heavy Truck-Tractor-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '503, 506'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Semitrailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '673, 676'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '683, 686'
        # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Service or Utility Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '693, 696'

        TTTPrimaryFactorsLong = TTTPrimaryFactorsLong[TTTPrimaryFactorsLong['ClassCode'].str.startswith("Light Truck")]
        TTTPrimaryFactorsLong = TTTPrimaryFactorsLong.pivot(index=['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)'], columns='Coverage', values='Factor').reset_index(['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)']). \
            rename(columns={'TruckSizeClass' : 'Size Class', 'TruckBusinessUseClass' : 'Business Use Class', 'Collision' : 'Collision Factor', 'Comprehensive And Specified Causes Of Loss' : 'OTC Factor', 'Liability' : 'Liability Factor'}).sort_values('Class (Non-Fleet, Fleet)')
        TTTPrimaryFactorsLong = TTTPrimaryFactorsLong[['Size Class', 'Business Use Class', 'Class (Non-Fleet, Fleet)', 'Liability Factor', 'OTC Factor', 'Collision Factor']]

        TTTPrimaryFactors = pd.merge(TTTPrimaryFactorsLocal, TTTPrimaryFactorsInt, on=['Size Class', 'Business Use Class'], how='left')
        TTTPrimaryFactors = pd.merge(TTTPrimaryFactors, TTTPrimaryFactorsLong, on=['Size Class', 'Business Use Class'], how='left')

        return TTTPrimaryFactors

    # Builds the TTT PhysDam Fleet Size Factors table
    # Returns a dataframe
    @log_exceptions
    def buildZonePrimaryFactors(self, company):
        # Which sheet was choosen to be used has been switched around multiple times. Its very annoying, so the function ended up being a mess.
        ZonePrimaryFactors = pd.DataFrame({"Empty":[1]}) # Made so non ISO Currency reviews can be tested.
        try:
            TTTPrimaryFactorsLong = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersPrimaryFactors_Ext'][0]).query(f'TruckRadiusClass == "Long Distance"')
            #TTTPrimaryFactorsLong = self.buildDataFrame("TrucksTractorsAndTrailersPrimaryFactors_Ext").query(f'TruckRadiusClass == "Long Distance"')
            TTTPrimaryFactorsLong['ClassCode'] = TTTPrimaryFactorsLong['TruckSizeClass'] + "-" + TTTPrimaryFactorsLong['TruckBusinessUseClass']
            # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Light Truck-Service", 'Class (Non-Fleet, Fleet)'] = '013, 016'
            # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Light Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '023, 026'
            # TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Light Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '033, 036'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Medium Truck-Service", 'Class (Non-Fleet, Fleet)'] = '213, 216'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Medium Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '223, 226'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Medium Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '233, 236'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Service", 'Class (Non-Fleet, Fleet)'] = '313, 316'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Retail", 'Class (Non-Fleet, Fleet)'] = '323, 326'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Commercial", 'Class (Non-Fleet, Fleet)'] = '333, 336'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Extra-Heavy Truck-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '403, 406'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Tractor-Service", 'Class (Non-Fleet, Fleet)'] = '343, 346'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Tractor-Retail", 'Class (Non-Fleet, Fleet)'] = '353, 356'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Heavy Truck-Tractor-Commercial", 'Class (Non-Fleet, Fleet)'] = '363, 366'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Extra Heavy Truck-Tractor-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '503, 506'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Semitrailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '673, 676'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '683, 686'
            TTTPrimaryFactorsLong.loc[TTTPrimaryFactorsLong['ClassCode'] == "Service or Utility Trailer-Not Applicable", 'Class (Non-Fleet, Fleet)'] = '693, 696'

            TTTPrimaryFactorsLong = TTTPrimaryFactorsLong[~TTTPrimaryFactorsLong['ClassCode'].str.startswith("Light Truck")]
            TTTPrimaryFactorsLong = TTTPrimaryFactorsLong.pivot(index=['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)'], columns='Coverage', values='Factor').reset_index(['TruckSizeClass', 'TruckBusinessUseClass', 'Class (Non-Fleet, Fleet)']). \
                rename(columns={'TruckSizeClass' : 'Size Class', 'TruckBusinessUseClass' : 'Business Use Class', 'Collision' : 'Collision Factor', 'Comprehensive And Specified Causes Of Loss' : 'OTC Factor', 'Liability' : 'Liability Factor'}).sort_values('Class (Non-Fleet, Fleet)')

            ZonePrimaryFactors = TTTPrimaryFactorsLong[['Class (Non-Fleet, Fleet)', 'Liability Factor', 'OTC Factor', 'Collision Factor']]

        except Exception:
            # traceback.print_exc()

            TTTPrimaryFactorsLiab = pd.DataFrame(self.rateTables[company]['ZoneRatedLiabilityPrimaryFactorNumeric'][1:], index=None, columns=self.rateTables[company]['ZoneRatedLiabilityPrimaryFactorNumeric'][0]).query(f'TruckPrimaryClassCodeNumeric == 213 | TruckPrimaryClassCodeNumeric == 223 | TruckPrimaryClassCodeNumeric == 233 | TruckPrimaryClassCodeNumeric == 313 | TruckPrimaryClassCodeNumeric == 323 | TruckPrimaryClassCodeNumeric == 333 | TruckPrimaryClassCodeNumeric == 403 | TruckPrimaryClassCodeNumeric == 343 | TruckPrimaryClassCodeNumeric == 353 | TruckPrimaryClassCodeNumeric == 363 | TruckPrimaryClassCodeNumeric == 503 | TruckPrimaryClassCodeNumeric == 673 | TruckPrimaryClassCodeNumeric == 683 |TruckPrimaryClassCodeNumeric == 693')
            TTTPrimaryFactorsOTC = pd.DataFrame(self.rateTables[company]['ZoneRatedPhysicalDamagePrimaryFactorNumeric'][1:], index=None, columns=self.rateTables[company]['ZoneRatedPhysicalDamagePrimaryFactorNumeric'][0]).query(f'TruckPrimaryClassCodeNumeric == 213 | TruckPrimaryClassCodeNumeric == 223 | TruckPrimaryClassCodeNumeric == 233 | TruckPrimaryClassCodeNumeric == 313 | TruckPrimaryClassCodeNumeric == 323 | TruckPrimaryClassCodeNumeric == 333 | TruckPrimaryClassCodeNumeric == 403 | TruckPrimaryClassCodeNumeric == 343 | TruckPrimaryClassCodeNumeric == 353 | TrucskPrimaryClassCodeNumeric == 363 | TruckPrimaryClassCodeNumeric == 503 | TruckPrimaryClassCodeNumeric == 673 | TruckPrimaryClassCodeNumeric == 683 |TruckPrimaryClassCodeNumeric == 693')


            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 213, 'Class (Non-Fleet, Fleet)'] = '213, 216'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 223, 'Class (Non-Fleet, Fleet)'] = '223, 226'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 233, 'Class (Non-Fleet, Fleet)'] = '233, 236'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 313, 'Class (Non-Fleet, Fleet)'] = '313, 316'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 323, 'Class (Non-Fleet, Fleet)'] = '323, 326'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 333, 'Class (Non-Fleet, Fleet)'] = '333, 336'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 403, 'Class (Non-Fleet, Fleet)'] = '403, 406'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 343, 'Class (Non-Fleet, Fleet)'] = '343, 346'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 353, 'Class (Non-Fleet, Fleet)'] = '353, 356'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 363, 'Class (Non-Fleet, Fleet)'] = '363, 366'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 503, 'Class (Non-Fleet, Fleet)'] = '503, 506'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 673, 'Class (Non-Fleet, Fleet)'] = '673, 676'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 683, 'Class (Non-Fleet, Fleet)'] = '683, 686'
            TTTPrimaryFactorsLiab.loc[
                TTTPrimaryFactorsLiab['TruckPrimaryClassCodeNumeric'] == 693, 'Class (Non-Fleet, Fleet)'] = '693, 696'
            TTTPrimaryFactorsLiab = TTTPrimaryFactorsLiab.sort_values('Class (Non-Fleet, Fleet)').rename(
                columns={'Factor': 'Liability Factor'})
            TTTPrimaryFactorsLiab = TTTPrimaryFactorsLiab[['Class (Non-Fleet, Fleet)', 'Liability Factor']]

            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 213, 'Class (Non-Fleet, Fleet)'] = '213, 216'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 223, 'Class (Non-Fleet, Fleet)'] = '223, 226'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 233, 'Class (Non-Fleet, Fleet)'] = '233, 236'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 313, 'Class (Non-Fleet, Fleet)'] = '313, 316'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 323, 'Class (Non-Fleet, Fleet)'] = '323, 326'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 333, 'Class (Non-Fleet, Fleet)'] = '333, 336'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 403, 'Class (Non-Fleet, Fleet)'] = '403, 406'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 343, 'Class (Non-Fleet, Fleet)'] = '343, 346'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 353, 'Class (Non-Fleet, Fleet)'] = '353, 356'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 363, 'Class (Non-Fleet, Fleet)'] = '363, 366'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 503, 'Class (Non-Fleet, Fleet)'] = '503, 506'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 673, 'Class (Non-Fleet, Fleet)'] = '673, 676'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 683, 'Class (Non-Fleet, Fleet)'] = '683, 686'
            TTTPrimaryFactorsOTC.loc[
                TTTPrimaryFactorsOTC['TruckPrimaryClassCodeNumeric'] == 693, 'Class (Non-Fleet, Fleet)'] = '693, 696'
            TTTPrimaryFactorsOTC = TTTPrimaryFactorsOTC.sort_values('Class (Non-Fleet, Fleet)').rename(
                columns={'Factor': 'OTC Factor'})
            TTTPrimaryFactorsOTC['Collision Factor'] = TTTPrimaryFactorsOTC['OTC Factor']
            TTTPrimaryFactorsOTC = TTTPrimaryFactorsOTC[['Class (Non-Fleet, Fleet)', 'OTC Factor', 'Collision Factor']]

            ZonePrimaryFactors = pd.merge(TTTPrimaryFactorsLiab, TTTPrimaryFactorsOTC, on='Class (Non-Fleet, Fleet)',
                                          how='left')
            ZonePrimaryFactors = ZonePrimaryFactors[['Class (Non-Fleet, Fleet)', 'Liability Factor', 'OTC Factor', 'Collision Factor']]

        finally:
            return ZonePrimaryFactors

    def buildFleetSizeRatingFactorsZone(self,company):
        table = pd.DataFrame(self.rateTables[company]['FleetSizeRatingFactorsZoneRated_Ext'][1:], index=None, columns=self.rateTables[company]['FleetSizeRatingFactorsZoneRated_Ext'][0])
        table = table.pivot(index=  "Number Of Powered Vehicles", columns = "Coverage", values = "Factor").reset_index()
        table.sort_values(by = "Number Of Powered Vehicles", inplace = True)
        table = table.iloc[1:,:]
        table["Number Of Powered Vehicles"] = self.self_propelled_vehicles
        table = table.astype(object)
        table.iloc[:,1:] = table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return table

    # Builds the Showroom Factors table
    # Returns a dataframe
    @log_exceptions
    def buildTTTSecondaryFactors(self, company):
        SecondaryFactorsLiabOTC = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersSecondaryFactorsLiabilityComprehensiveAndSCOL_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersSecondaryFactorsLiabilityComprehensiveAndSCOL_Ext'][0])
        #SecondaryFactorsLiabOTC = self.buildDataFrame("TrucksTractorsAndTrailersSecondaryFactorsLiabilityComprehensiveAndSCOL_Ext")
        SecondaryFactorsLiabOTC = SecondaryFactorsLiabOTC.pivot(index='TruckSecondaryClassification', columns='Coverage', values='Factor').reset_index('TruckSecondaryClassification').rename(columns={'TruckSecondaryClassification' : 'Secondary Class', 'Comprehensive And Specified Causes Of Loss' : 'OTC'})
        SecondaryFactorsLiabOTC = SecondaryFactorsLiabOTC[['Secondary Class', 'Liability', 'OTC']]
        SecondaryFactorsColl = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersSecondaryFactorsCollision_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersSecondaryFactorsCollision_Ext'][0])
        #SecondaryFactorsColl = self.buildDataFrame("TrucksTractorsAndTrailersSecondaryFactorsCollision_Ext")
        SecondaryFactorsColl = SecondaryFactorsColl.pivot(index='TruckSecondaryClassification', columns='Coverage', values='Factor').reset_index('TruckSecondaryClassification').rename(columns={'TruckSecondaryClassification' : 'Secondary Class', 'Trucks And Truck-tractors Collision' : 'Collision'})
        SecondaryFactorsColl = SecondaryFactorsColl[['Secondary Class', 'Collision', 'Trailers Collision']]
        SecondaryFactors = pd.merge(SecondaryFactorsLiabOTC, SecondaryFactorsColl, on='Secondary Class', how='left')

        #Class Code Hardcoded based on Descriptions below:
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Carrier Both Private Carriage & Transport Goods/Materials/Commodities", '4th-5th Digits of Class Code'] = '02'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Tow Trucks For-Hire", '4th-5th Digits of Class Code'] = '03'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Common Carriers", '4th-5th Digits of Class Code'] = '21'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Contract Carriers (Other than Chemical or Iron and Steel Haulers)", '4th-5th Digits of Class Code'] = '22'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Contract Carriers Hauling Chemicals", '4th-5th Digits of Class Code'] = '23'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Contract Carriers Hauling Iron and Steel", '4th-5th Digits of Class Code'] = '24'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Exempt Carriers (Other than Livestock Haulers)", '4th-5th Digits of Class Code'] = '25'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - Exempt Carriers Hauling Livestock", '4th-5th Digits of Class Code'] = '26'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Truckers - All Other", '4th-5th Digits of Class Code'] = '29'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Food Delivery - Canneries and Packing Plants", '4th-5th Digits of Class Code'] = '31'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Food Delivery - Fish and Seafood", '4th-5th Digits of Class Code'] = '32'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Food Delivery - Frozen Food", '4th-5th Digits of Class Code'] = '33'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Food Delivery - Fruit and Vegetable", '4th-5th Digits of Class Code'] = '34'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Food Delivery - Meat or Poultry", '4th-5th Digits of Class Code'] = '35'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Food Delivery - All Other", '4th-5th Digits of Class Code'] = '39'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Specialized Delivery - Armored Cars", '4th-5th Digits of Class Code'] = '41'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Specialized Delivery - Film Delivery", '4th-5th Digits of Class Code'] = '42'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Specialized Delivery - Magazines or Newspapers", '4th-5th Digits of Class Code'] = '43'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Specialized Delivery - Mail and Parcel Post", '4th-5th Digits of Class Code'] = '44'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Specialized Delivery - All Other", '4th-5th Digits of Class Code'] = '49'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Waste Disposal - Auto Dismantlers", '4th-5th Digits of Class Code'] = '51'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Waste Disposal - Building Wrecking Operators", '4th-5th Digits of Class Code'] = '52'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Waste Disposal - Garbage", '4th-5th Digits of Class Code'] = '53'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Waste Disposal - Junk Dealers", '4th-5th Digits of Class Code'] = '54'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Waste Disposal - All Other", '4th-5th Digits of Class Code'] = '59'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Farmers - Individually Owned or Family Corp (Other than Livestock Hauling)", '4th-5th Digits of Class Code'] = '61'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Farmers - Livestock Hauling", '4th-5th Digits of Class Code'] = '62'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Farmers - All Other", '4th-5th Digits of Class Code'] = '69'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Dump and Transit Mix - Excavating", '4th-5th Digits of Class Code'] = '71'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Dump and Transit Mix - Sand and Gravel (Other than Quarrying)", '4th-5th Digits of Class Code'] = '72'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Dump and Transit Mix - Mining", '4th-5th Digits of Class Code'] = '73'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Dump and Transit Mix - Quarrying", '4th-5th Digits of Class Code'] = '74'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Dump and Transit Mix - All Other", '4th-5th Digits of Class Code'] = '79'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Contractors - Building - Commercial", '4th-5th Digits of Class Code'] = '81'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Contractors - Building - Private Dwellings", '4th-5th Digits of Class Code'] = '82'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Contractors - Electrical, Plumbing, Masonry, Plastering, Other Repair/Service", '4th-5th Digits of Class Code'] = '83'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Contractors - Excavating", '4th-5th Digits of Class Code'] = '84'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Contractors - Street and Road", '4th-5th Digits of Class Code'] = '85'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Contractors - All Other", '4th-5th Digits of Class Code'] = '89'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Not Otherwise Specified - Logging and Lumbering", '4th-5th Digits of Class Code'] = '91'
        SecondaryFactors.loc[SecondaryFactors['Secondary Class'] == "Not Otherwise Specified - All Other", '4th-5th Digits of Class Code'] = '99'

        # Need to hard code 05-06. They are not in the book but should be filed.
        SecondaryFactors = SecondaryFactors[['Secondary Class', '4th-5th Digits of Class Code', 'Liability', 'OTC', 'Collision', 'Trailers Collision']]

        code_05 = {"Secondary Class" : """Movers""",
                   "4th-5th Digits of Class Code" : "05",
                   "Liability" : "1.000",
                   "OTC" : "1.000",
                   "Collision" : "1.000",
                   "Trailers Collision" : "1.000"}

        code_06 = {"Secondary Class" : """Owned trucks, tractors and trailers that the insured leases or rents to others to use in their trucking or motor carrier operations when the insured lessor has agreed, in writing, to provide primary coverage and to hold the lessee harmless in a written lease agreement""",
                   "4th-5th Digits of Class Code" : "06",
                   "Liability" : "2.180",
                   "OTC" : "1.900",
                   "Collision" : "2.240",
                   "Trailers Collision" : "1.990"}

        code_05_df = pd.DataFrame([code_05])
        code_06_df = pd.DataFrame([code_06])

        # Concatenate the new rows at the third and fourth positions
        SecondaryFactors = pd.concat(
            [code_05_df, code_06_df, SecondaryFactors.iloc[2:]],
            ignore_index=True
        )

        SecondaryFactors = SecondaryFactors.sort_values('4th-5th Digits of Class Code')

        return SecondaryFactors

    def buildZoneSecondaryFactors(self, company):
        data = pd.DataFrame(self.rateTables[company]['Secondary Classification Factors Zone Rated_Ext'][1:], index=None, columns=self.rateTables[company]['Secondary Classification Factors Zone Rated_Ext'][0])

        #Class Code Hardcoded based on Descriptions below:
        secondary_factors = pd.DataFrame({
            'Secondary Class': [
                "Truckers - Carrier Both Private Carriage & Transport Goods/Materials/Commodities",
                "Truckers - Tow Trucks For-Hire",
                "Movers",
                "Owned trucks, tractors and trailers that the insured leases or rents to others to use in their trucking or motor carrier operations when the insured lessor has agreed, in writing, to provide primary coverage and to hold the lessee harmless in a written lease agreement",
                "Truckers - Common Carriers",
                "Truckers - Contract Carriers (Other than Chemical or Iron and Steel Haulers)",
                "Truckers - Contract Carriers Hauling Chemicals",
                "Truckers - Contract Carriers Hauling Iron and Steel",
                "Truckers - Exempt Carriers (Other than Livestock Haulers)",
                "Truckers - Exempt Carriers Hauling Livestock",
                "Truckers - All Other",
                "Food Delivery - Canneries and Packing Plants",
                "Food Delivery - Fish and Seafood",
                "Food Delivery - Frozen Food",
                "Food Delivery - Fruit and Vegetable",
                "Food Delivery - Meat or Poultry",
                "Food Delivery - All Other",
                "Specialized Delivery - Armored Cars",
                "Specialized Delivery - Film Delivery",
                "Specialized Delivery - Magazines or Newspapers",
                "Specialized Delivery - Mail and Parcel Post",
                "Specialized Delivery - All Other",
                "Waste Disposal - Auto Dismantlers",
                "Waste Disposal - Building Wrecking Operators",
                "Waste Disposal - Garbage",
                "Waste Disposal - Junk Dealers",
                "Waste Disposal - All Other",
                "Farmers - Individually Owned or Family Corp (Other than Livestock Hauling)",
                "Farmers - Livestock Hauling",
                "Farmers - All Other",
                "Dump and Transit Mix - Excavating",
                "Dump and Transit Mix - Sand and Gravel (Other than Quarrying)",
                "Dump and Transit Mix - Mining",
                "Dump and Transit Mix - Quarrying",
                "Dump and Transit Mix - All Other",
                "Contractors - Building - Commercial",
                "Contractors - Building - Private Dwellings",
                "Contractors - Electrical, Plumbing, Masonry, Plastering, Other Repair/Service",
                "Contractors - Excavating",
                "Contractors - Street and Road",
                "Contractors - All Other",
                "Not Otherwise Specified - Logging and Lumbering",
                "Not Otherwise Specified - All Other"
            ],
            '4th-5th Digits of Class Code': [
                '02', '03', '05', '06', '21',
                '22', '23', '24', '25', '26',
                '29', '31', '32', '33', '34',
                '35', '39', '41', '42', '43',
                '44', '49', '51', '52', '53',
                '54', '59', '61', '62', '69',
                '71', '72', '73', '74', '79',
                '81', '82', '83', '84', '85',
                '89', '91', '99'
            ]
        })

        # Step 1: Create a unique identifier by combining 'Coverage' and 'Vehicle Type'
        data['Unique Identifier'] = data['Vehicle Type'] + ' - ' + data['Coverage']
        data = data.rename(columns = {"Code": "4th-5th Digits of Class Code"})

        # Step 2: Create a pivot table with 'Code' as the index, 'Unique Identifier' as the columns, and 'Factor' as the values
        pivot_table = data.pivot_table(index='4th-5th Digits of Class Code', columns='Unique Identifier', values='Factor', aggfunc='first')

        # Step 3: Perform a left join with the secondary_factors DataFrame using the 'Code' column
        SecondaryFactors = secondary_factors.merge(pivot_table, on='4th-5th Digits of Class Code', how='left')
        SecondaryFactors = SecondaryFactors.rename(columns = {
            "Trucks, Tractors And Trailers - Liability" : "Liability",
            "Trucks And Trucktractors - Collision" : "Collision",
            "Trucks, Tractors And Trailers - Other Than Collision" : "OTC",
            "Trailers - Collision" : "Trailers Collision"
        })

        # Need to hard code 05-06. They are not in the book but should be filed.
        SecondaryFactors = SecondaryFactors[['Secondary Class', '4th-5th Digits of Class Code', 'Liability', 'OTC', 'Collision', 'Trailers Collision']]

        SecondaryFactors = SecondaryFactors.sort_values('4th-5th Digits of Class Code')
        SecondaryFactors = SecondaryFactors.astype(object)
        SecondaryFactors.iloc[:, 2:] = SecondaryFactors.iloc[:, 2:].astype(float).map(lambda x: f"{x:.3f}")

        return SecondaryFactors


    # Builds the Zone Base Rate table
    # Returns a dataframe
    @log_exceptions
    def buildZoneBaseRates(self, company):
        """
        Rule 225, similar idea to the buildBaseRates method

        :param company:
        :return: Dictionary {"Zone": zones, "Output Tables" : output_tables, "Med Factor" : med_factor, "Pip Factor" : pip_factor}
        -- Zone tables in a list along with the subtitles list:
                           [Zone A Rates, Zone B Rates,..., Zone Z Rates, Med Factor, Pip Factor],
                           [Zone A Subtitle, Zone B Subtitle, ..., Zone Z Subtitle]
        -- Med Factor: Table containing med factor
        -- Pip Factor: Table containing med factor
        """

        coverages = ["csl", "coll", "otc"]

        sheet_name = "225 Zone Rated"
        state_mapper = pd.read_excel(BA_INPUT_FILE, sheet_name=sheet_name)

        mappers = {}
        for coverage in coverages:
            filtered_map = state_mapper[(state_mapper["state"] == self.StateAbb) & (state_mapper["coverage"] == coverage)]
            default_map = state_mapper[(state_mapper["state"] == "Default") & (state_mapper["coverage"] == coverage)]
            if not filtered_map.empty:
                mappers[coverage] = filtered_map
            else:
                mappers[coverage] = default_map

        # Auto-populate rate book sheet names based on the input file

        liabTable = None
        otcTable = None
        collTable = None

        try:
            LIAB_RATE_BOOK_SHEET_NAME = mappers['csl']["sheet"].values[0]
            liabTable = pd.DataFrame(self.rateTables[company][LIAB_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][LIAB_RATE_BOOK_SHEET_NAME][0])
            liabTable.rename(columns = {"BasePremium" : "Liability (225.D.1.a.)"}, inplace = True)
        except (KeyError, IndexError):
            pass

        try:
            OTC_RATE_BOOK_SHEET_NAME = mappers["otc"]['sheet'].values[0]
            otcTable = pd.DataFrame(self.rateTables[company][OTC_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][OTC_RATE_BOOK_SHEET_NAME][0])
            otcTable.rename(columns={"BasePremium" : "Comp. (225.D.4.a.)"}, inplace=True)
        except (KeyError, IndexError):
            pass

        try:
            COLL_RATE_BOOK_SHEET_NAME = mappers["coll"]['sheet'].values[0]
            collTable = pd.DataFrame(self.rateTables[company][COLL_RATE_BOOK_SHEET_NAME][1:], index=None, columns=self.rateTables[company][COLL_RATE_BOOK_SHEET_NAME][0])
            collTable.rename(columns = {"BasePremium" : "Coll. (225.D.4.a.)"}, inplace=True)
        except (KeyError, IndexError):
            pass


        master_table_list = [] # List of unjoined tables: [[Zone1Liab, Zone1Coll],[Zone2liab,Zone2Coll]]
        zones = liabTable["GaragingZone"].unique()
        for zone in zones:
            # Making tables per zone
            liabTemp = liabTable[liabTable["GaragingZone"] == zone].drop(columns = {"GaragingZone"})
            collTemp = collTable[collTable["GaragingZone"] == zone].drop(columns = {"GaragingZone"})
            otcTemp = otcTable[otcTable["GaragingZone"] == zone].drop(columns = {"GaragingZone"})

            master_table_list.append([liabTemp,collTemp,otcTemp])

        output_tables = []
        for zone, tables in zip(zones, master_table_list):
            output_table = None
            for table in tables:
                if output_table is None:
                    output_table = table
                else:
                    output_table = output_table.merge(table, on="FarthestTerminalZone", how="inner")
            # The below is regex to parse the strings in FarthestTerminalZone grabbing the Zone number and Location.
            output_table["Zone"] = output_table["FarthestTerminalZone"].apply(lambda x: re.match(r"Zone (\d{2}) (.+)", x).group(1))
            output_table["Description"] = output_table["FarthestTerminalZone"].apply(lambda x: re.match(r"Zone (\d{2}) (.+)", x).group(2))
            output_table.drop(columns = ["FarthestTerminalZone"], inplace = True)

            # Reordering columns to [Zone, Description, Premium]
            first_cols = ["Zone", "Description"]
            last_cols = [col for col in output_table.columns if col not in first_cols]
            output_table = output_table[first_cols + last_cols] # Reordering the table columns

            output_tables.append(output_table)

        # Attempting to build med factor
        med_factor = None
        pip_factor = None


        if self.StateAbb in self.no_med_states["states"].values:
            # Michigan has the sheet but not any values.
            pass

        else:
            sheet_names = ['ZoneRatedMedicalPaymentsTextFactor']
            orig_values = ['5,000']
            replace_values = ['Zone-Rated Factor']
            med_factor = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values=orig_values)
            med_factor = med_factor.rename(columns={med_factor.columns[0]: '', med_factor.columns[1]: ' '})

            for column in med_factor.columns:
                if med_factor[column].dtype == 'float64' or med_factor[column].dtype == 'int64':
                    med_factor[column] = med_factor[column].apply(lambda x: f"{x:.3f}")

        if self.StateAbb in self.pip_states:
            sheet_names = ['ZoneRatedPIPFactor']
            orig_values = ['Y']
            replace_values = ['Zone-Rated Factor']

            if self.StateAbb == "MI":
                # Michigan has an annoying pip system.
                sheet_names = ['ZoneRatedPIPFactorMI', 'ZoneRatedPIPFactorMI']
                orig_values = ["Interstate","Intrastate"]
                replace_values = ["Interstate","Intrastate"]

            pip_factor = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values=orig_values)
            pip_factor = pip_factor.rename(columns={pip_factor.columns[0]: '', pip_factor.columns[1]: 'Factor'})

        if pip_factor is not None:
            for column in pip_factor.columns:
                if pip_factor[column].dtype == 'float64' or pip_factor[column].dtype == 'int64':
                    pip_factor[column] = pip_factor[column].apply(lambda x: f"{x:.3f}")


        for output_table in output_tables:
            for column in output_table.columns:
                if output_table[column].dtype == 'float64' or output_table[column].dtype == 'int64':
                    output_table[column] = output_table[column].apply(lambda x: f"{x:.2f}")


        return_dict = {"Zones" : zones,
                       "Output Tables" : output_tables,
                       "Med Factor" : med_factor,
                       "Pip Factor" : pip_factor}

        return return_dict


    # Builds the Showroom Factors table
    # Returns a dataframe
    @log_exceptions
    def build231C(self, company):

        # Check if NC for state specific version of table
        if self.StateAbb == "NC":

            NC_rule231_table = {
                "Class Code": ["7398"],
                "Collision": [1],
                "Comprehensive": [1],
                "Liability and Medical Payments": [1]
            }

            NC_rule231_table = pd.DataFrame(NC_rule231_table)
            NC_rule231_table = NC_rule231_table.astype(object)
            NC_rule231_table.iloc[:, 1:] = NC_rule231_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

            return NC_rule231_table
        # Normal version of table
        else:
            class_codes = pd.DataFrame(self.rateTables[company]['PrivatePassengerClassCode'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerClassCode'][0])
            factors = pd.DataFrame(self.rateTables[company]['PrivatePassengerTypesClassFactors_Ext'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerTypesClassFactors_Ext'][0])

            class_codes.drop(columns=["PrivatePassengerFleet"], inplace=True)

            table = pd.merge(factors, class_codes, on=['PrivatePassengerType', 'PrivatePassengerOperatorExperience', 'PrivatePassengerUse'], how='left')
            table = table.pivot(index=['ClassCode'], columns='Coverage', values='Factor').reset_index()
            table = table.rename(columns={"ClassCode": "Class Code", "Liability": "Liability and Medical Payments"})
            table = table.astype(object)
            table.iloc[:, 1:] = table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

            return table

    # Builds the Showroom Factors table
    # Returns a dataframe
    @log_exceptions
    def buildShowRoomFactors(self, company):
        ShowRoomLiabFactors = pd.DataFrame(self.rateTables[company]['ShowroomLiabilityFactor'][1:], index=None, columns=self.rateTables[company]['ShowroomLiabilityFactor'][0])
        #ShowRoomLiabFactors = self.buildDataFrame("ShowroomLiabilityFactor")
        ShowRoomLiabFactors['Constant'] = "Liability:"
        ShowRoomMedFactors = pd.DataFrame(self.rateTables[company]['ShowroomMedicalPaymentsFactor'][1:], index=None, columns=self.rateTables[company]['ShowroomMedicalPaymentsFactor'][0])
        #ShowRoomMedFactors = self.buildDataFrame("ShowroomMedicalPaymentsFactor")
        ShowRoomMedFactors['Constant'] = "Medical Payments:"
        ShowRoomFactors = pd.concat([ShowRoomLiabFactors, ShowRoomMedFactors]).rename(columns={'Constant' : 'Trailers and Semi-Trailers Used As Showrooms:'})
        return ShowRoomFactors

    # Builds the Other Than Zone Rated Autos Factors table
    # Returns a dataframe
    @log_exceptions
    def buildShowRoom2Factors(self, company):
        ShowRoom2firstFactors = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext'][0]).query(f'VehicleType == "Dump And Transit-mix Vehicles That Are Heavy Or Extra-heavy"')
        #ShowRoom2firstFactors = self.buildDataFrame("TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext").query(f'VehicleType == "Dump And Transit-mix Vehicles That Are Heavy Or Extra-heavy"')
        ShowRoom2firstFactors['VehicleType'] = "Collision:"
        ShowRoom2secondFactors = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext'][0]).query(f'VehicleType == "All Other Vehicles"')
        #ShowRoom2secondFactors = self.buildDataFrame("TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext").query(f'VehicleType == "All Other Vehicles"')
        ShowRoom2secondFactors['VehicleType'] = "All Other Vehicles:"
        ShowRoom2Factors = pd.concat([ShowRoom2firstFactors, ShowRoom2secondFactors])
        return ShowRoom2Factors

    # Builds the Zone Rated Autos Factors table
    # Returns a dataframe
    @log_exceptions
    def buildShowRoom3Factors(self, company):
        ShowRoom3firstFactors = pd.DataFrame(self.rateTables[company]['TruckDumpingRelativity'][1:], index=None, columns=self.rateTables[company]['TruckDumpingRelativity'][0])
        #ShowRoom3firstFactors = self.buildDataFrame("TruckDumpingRelativity")
        ShowRoom3firstFactors['Constant'] = "Collision:"
        ShowRoom3secondFactors = pd.DataFrame(self.rateTables[company]['TruckDumpingRelativity'][1:], index=None, columns=self.rateTables[company]['TruckDumpingRelativity'][0])
        #ShowRoom3secondFactors = self.buildDataFrame("TruckDumpingRelativity")
        ShowRoom3secondFactors['Constant'] = "All Other Vehicles:"
        ShowRoom3secondFactors['Factor'] = 1
        ShowRoom3Factors = pd.concat([ShowRoom3firstFactors, ShowRoom3secondFactors])
        return ShowRoom3Factors

    # Builds the Other Than Zone Rated Autos Factors table
    # Returns a dataframe
    @log_exceptions
    def buildShowRoom4Factors(self, company):
        ShowRoom4firstFactors = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext'][0]).query(f'VehicleType == "Farm Use Vehicles That Are Heavy Or Extra-heavy"')
        #ShowRoom4firstFactors = self.buildDataFrame("TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext").query(f'VehicleType == "Farm Use Vehicles That Are Heavy Or Extra-heavy"')
        ShowRoom4firstFactors['VehicleType'] = "Other than Zone:"
        ShowRoom4secondFactors = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext'][0]).query(f'VehicleType == "All Other Vehicles"')
        #ShowRoom4secondFactors = self.buildDataFrame("TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext").query(f'VehicleType == "All Other Vehicles"')
        ShowRoom4secondFactors['VehicleType'] = "Zone:"
        ShowRoom4Factors = pd.concat([ShowRoom4firstFactors, ShowRoom4secondFactors])
        return ShowRoom4Factors


    # Builds the Auto Layup Factors table
    # Returns a dataframe
    @log_exceptions
    # Decimal edit
    def buildLayupFactors(self, company):
        # Rule 225.D and Rule 239.D.1.c
        LayupFactors01 = pd.DataFrame(self.rateTables[company]['AutoLayUpFactor_Ext'][1:], index=None, columns=self.rateTables[company]['AutoLayUpFactor_Ext'][0]).query(f'MonthsLaidUp == "0-1"')
        #LayupFactors01 = self.buildDataFrame("AutoLayUpFactor_Ext").query(f'MonthsLaidUp == "0-1"')
        LayupFactors26 = pd.DataFrame(self.rateTables[company]['AutoLayUpFactor_Ext'][1:], index=None, columns=self.rateTables[company]['AutoLayUpFactor_Ext'][0]).query(f'MonthsLaidUp != "0-1"')
        #LayupFactors26 = self.buildDataFrame("AutoLayUpFactor_Ext").query(f'MonthsLaidUp != "0-1"')
        LayupFactors26['MonthsLaidUp'].replace('6', '6+', inplace=True)
        LayupFactors = pd.concat([LayupFactors01, LayupFactors26]).rename(columns={'MonthsLaidUp' : 'Months Laid Up'})

        LayupFactors = LayupFactors.astype(object)
        LayupFactors.iloc[:, 1:] = LayupFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return LayupFactors

    # Builds the PPT NAICS table
    # Returns a dataframe
    @log_exceptions
    def buildPPTNAICSFactors(self, company):
        NAICSReference = pd.DataFrame(data=self.NAICSDescriptions).astype({'NAICS Six-Digit Code': 'int64'})
        NAICSCoverages = pd.DataFrame(self.rateTables[company]['NAICSFactors_Ext'][1:], index=None, columns=self.rateTables[company]['NAICSFactors_Ext'][0]).pivot(index='NAICSCode', columns='VehicleAndCoverageType', values='Factor').reset_index('NAICSCode').rename(columns={'NAICSCode' : 'NAICS Six-Digit Code'}).filter(items=['NAICS Six-Digit Code', 'NAICS Category', 'Private Passenger Types Liability', 'Private Passenger Types Collision', 'Private Passenger Types Comprehensive']).astype({'NAICS Six-Digit Code': 'int64'})
        #NAICSCoverages = self.buildDataFrame("NAICSFactors_Ext").pivot(index='NAICSCode', columns='VehicleAndCoverageType', values='Factor').reset_index('NAICSCode').rename(columns={'NAICSCode' : 'NAICS Six-Digit Code'}).filter(items=['NAICS Six-Digit Code', 'NAICS Category', 'Private Passenger Types Liability', 'Private Passenger Types Collision', 'Private Passenger Types Comprehensive']).astype({'NAICS Six-Digit Code': 'int64'})
        PPTNAICSFactors = pd.merge(NAICSReference, NAICSCoverages, on = 'NAICS Six-Digit Code', how = 'inner')
        return PPTNAICSFactors

    # Builds the PPT Liability Fleet Size Factors table
    # Returns a dataframe
    @log_exceptions
    def buildPPTLiabFleetFactors(self, company):
        PPTLiabFleetFactors = pd.DataFrame(self.rateTables[company]['LiabilityFleetSizeFactors_Ext'][1:], index=None, columns=self.rateTables[company]['LiabilityFleetSizeFactors_Ext'][0]).query(f'VehicleType == "Private Passenger"')
        #PPTLiabFleetFactors = self.buildDataFrame("LiabilityFleetSizeFactors_Ext").query(f'VehicleType == "Private Passenger"')
        PPTLiabFleetFactors = PPTLiabFleetFactors.drop(columns='VehicleType').rename(columns={'NumberOfPoweredVehicles' : 'Number of Powered Vehicles', 'Factor' : 'Liability'})
        PPTLiabFleetFactors['Number of Powered Vehicles'] = self.self_propelled_vehicles

        return PPTLiabFleetFactors

    # Builds the PPT PhysDam Fleet Size Factors table
    # Returns a dataframe
    @log_exceptions
    def buildPPTPhysDamFleetFactors(self, company):
        PPTCollFleetFactors = pd.DataFrame(self.rateTables[company]['CollisionFleetSizeFactor_Ext'][1:], index=None, columns=self.rateTables[company]['CollisionFleetSizeFactor_Ext'][0]).query(f'VehicleType == "Private Passenger Types"').drop(columns='VehicleType').rename(columns={'Factor' : 'Collision'})
        #PPTCollFleetFactors = self.buildDataFrame("CollisionFleetSizeFactor_Ext").query(f'VehicleType == "Private Passenger Types"').drop(columns='VehicleType').rename(columns={'Factor' : 'Collision'})
        PPTSCoLFleetFactors = pd.DataFrame(self.rateTables[company]['ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext'][1:], index=None, columns=self.rateTables[company]['ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext'][0]).query(f'VehicleType == "Private Passenger Types"').drop(columns='VehicleType').rename(columns={'Factor' : 'Comprehensive'})
        #PPTSCoLFleetFactors = self.buildDataFrame("ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext").query(f'VehicleType == "Private Passenger Types"').drop(columns='VehicleType').rename(columns={'Factor' : 'Comprehensive'})
        PPTPhysDamFleetFactors = pd.merge(PPTCollFleetFactors, PPTSCoLFleetFactors, on = 'NumberOfPoweredVehicles', how = 'inner').rename(columns={'NumberOfPoweredVehicles' : 'Number of Powered Vehicles'}).filter(items=['Number of Powered Vehicles', 'Collision', 'Comprehensive'])
        PPTPhysDamFleetFactors['Number of Powered Vehicles'] = self.self_propelled_vehicles

        return PPTPhysDamFleetFactors

    # Builds the PPT Types Classifications Farm Factors table
    # Returns a dataframe
    @log_exceptions
    def buildPPTFarmTypes(self, company):
        PPTFarmTypes = pd.DataFrame(self.rateTables[company]['PrivatePassengerFarmFactor2'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerFarmFactor2'][0]).rename(columns={'TypeCoverage' : 'Coverage'})
        #PPTFarmTypes = self.buildDataFrame("PrivatePassengerFarmFactor2").rename(columns={'TypeCoverage' : 'Coverage'})
        med_name = "Med"

        if self.StateAbb in self.pip_states.values:
            med_name = "Med/No-Fault"
        if self.StateAbb in self.no_med_states.values:
            med_name = "No-Fault"

        PPTFarmTypes['Coverage'].replace('MedicalPayments', med_name, inplace=True)
        PPTFarmTypes['Coverage'].replace('PhysDamOTCACV', 'Comprehensive', inplace=True)
        PPTFarmTypes['Coverage'].replace('PhysDamCollSA', 'Collision', inplace=True)
        PPTFarmTypes['Coverage'].replace('UM/UIM', 'Uninsured Motorists', inplace=True)
        PPTFarmTypes = PPTFarmTypes.query(f'Coverage == "Liability" or Coverage == "Med/No-Fault" or Coverage == "Med" or Coverage == "Uninsured Motorists" or Coverage == "Comprehensive" or Coverage == "Collision"')

        return PPTFarmTypes

    # Builds the Towing and Labor rate table
    # Returns a dataframe
    @log_exceptions
    def buildTowingAndLabor(self, company):
        TowingAndLabor = pd.DataFrame(self.rateTables[company]['TowingLaborRate'][1:], index=None, columns=self.rateTables[company]['TowingLaborRate'][0])
        TowingAndLabor = TowingAndLabor.rename(columns = {TowingAndLabor.columns[0] : "Limit", TowingAndLabor.columns[1] : "Rate"})
        TowingAndLabor = TowingAndLabor.astype(object)
        TowingAndLabor.iloc[:, 1:] = TowingAndLabor.iloc[:, 1:].map(lambda x: f"{x:.2f}")

        return TowingAndLabor

    # Iso Currency 239
    def buildPublicAutoLiabFleetSizeFactor(self, company):
        # 239.C.2.d. Fleet Size Factors For Liability And Medical Payments
        self_propelled_vehicles = self.self_propelled_vehicles
        sheet_name = 'PublicTypesFleetSizeFactorsForLiabilityAndMedicalPayments_Ext'
        fleet_factor = pd.DataFrame(self.rateTables[company][sheet_name][1:], index=None, columns=self.rateTables[company][sheet_name][0])
        output_table = fleet_factor.pivot(index='Number Of Powered Vehicles', columns='Group', values='Factor')
        output_table.columns.name = None
        output_table.index = self_propelled_vehicles
        output_table.reset_index(inplace = True)
        output_table.rename(columns={'index': 'Number Of Self-propelled Vehicles'}, inplace=True)
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildPublicAutoCollFleetSizeFactor(self, company):
        # 239.C.3.d. Fleet Size Factors For Collision
        self_propelled_vehicles = self.self_propelled_vehicles
        sheet_name = 'PublicTransportationCollisionFleetSizeFactor_Ext'
        fleet_factor = pd.DataFrame(self.rateTables[company][sheet_name][1:], index=None, columns=self.rateTables[company][sheet_name][0])
        output_table = fleet_factor.pivot(index='Number Of Powered Vehicles', columns='Group', values='Factor')
        output_table.columns.name = None
        output_table.index = self_propelled_vehicles
        output_table.reset_index(inplace = True)
        output_table.rename(columns={'index': 'Number Of Self-propelled Vehicles'}, inplace=True)
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")
        return output_table

    def buildPublicAutoOTCFleetSizeFactor(self, company):
        # 239.C.3.d. Fleet Size Factors For Other Than Collision
        self_propelled_vehicles = self.self_propelled_vehicles
        sheet_name = 'PublicTransportationOtherThanCollisionFleetSizeFactor_Ext'
        fleet_factor = pd.DataFrame(self.rateTables[company][sheet_name][1:], index=None, columns=self.rateTables[company][sheet_name][0])
        output_table = fleet_factor.pivot(index='Number Of Powered Vehicles', columns='Group', values='Factor')
        output_table.columns.name = None
        output_table.index = self_propelled_vehicles
        output_table.reset_index(inplace = True)
        output_table.rename(columns={'index': 'Number Of Self-propelled Vehicles'}, inplace=True)
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildMechanicalLiftFactor(self,company):
        # Rule 241.D.1.d, Rule 239.C.4.d Mechanical Lift Factor
        sheet_names = ['MechanicalLiftFactorOtherThanZoneRated']
        orig_values = ['Y']
        replace_values = ['Factor']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[0, 1] = f"{output_table.iloc[0, 1]:.3f}"

        return output_table

    def buildSpecifiecCausesofLossCoverageFactor(self,company):
        # Rule 241.D.6.b, a rule present in VA
        sheet_names = ['ZoneRatedOtherThanCollisionSpecifiedCauseLossFactor']
        orig_values = ['Y']
        replace_values = ['Factor']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[0, 1] = f"{output_table.iloc[0, 1]:.3f}"

        return output_table

    # Builds the Showroom Factors table
    # Returns a dataframe
    @log_exceptions
    def build240(self, company):
        publicPrimaryLiab = pd.DataFrame(self.rateTables[company]['PublicTransportationLiabilityPrimaryFactor'][1:], index=None, columns=self.rateTables[company]['PublicTransportationLiabilityPrimaryFactor'][0]).rename(columns={'Factor' : 'Liability'})
        publicPrimaryPhys = pd.DataFrame(self.rateTables[company]['PublicTransportationPhysicalDamagePrimaryFactor'][1:], index=None, columns=self.rateTables[company]['PublicTransportationPhysicalDamagePrimaryFactor'][0]).rename(columns={'Factor' : 'Physical Damage'})


        non_fleet = [5718,5719,4118,4119,5178,615,625,635,515,525,535,545,555,565,645,655,4398,585,5728,5729,4128,4129,5278,616,626,636,516,526,536,546,556,566,646,656,4498,586,5738,5739,4138,4139,5378,617,627,637,5279,5379,5479,5579,5679,6479,6579,5879]
        fleet = [5748,5749,4218,4219,5478,618,628,638,518,528,538,548,558,568,648,658,4338,588,5758,5759,4228,4229,5578,619,629,639,519,529,539,549,559,569,649,659,4438,589,5768,5769,4238,4239,5678,610,620,630,5209,5309,5409,5509,5609,6409,6509,5809]

        output_table = pd.DataFrame({"Fleet" : fleet,
                                     "Non-Fleet" : non_fleet})

        output_table = pd.merge(output_table, publicPrimaryLiab, how = "left", left_on = "Fleet", right_on = "PrimaryClassCode")
        output_table = pd.merge(output_table, publicPrimaryPhys, how = "left", left_on = "Fleet", right_on = "PrimaryClassCode")

        output_table = output_table.fillna({'Fleet': 'DELETE'})
        output_table = output_table.query(f'Fleet != "DELETE"')

        output_table = output_table[["Fleet","Non-Fleet","Liability","Physical Damage"]]
        output_table[["Non-Fleet","Fleet"]] = output_table[["Non-Fleet","Fleet"]].astype(int)
        output_table = output_table.astype(object)
        output_table.iloc[:,-1] = output_table.iloc[:,-1].apply(lambda x: f"{x:.3f}")

        return output_table

    def buildVanPrimaryClassFactor(self, company):
        liab_factor = pd.DataFrame(self.rateTables[company]['PublicTransportationLiabilityPrimaryFactor'][1:], index=None, columns=self.rateTables[company]['PublicTransportationLiabilityPrimaryFactor'][0])

        phys_factor = pd.DataFrame(self.rateTables[company]['PublicTransportationPhysicalDamagePrimaryFactor'][1:], index=None,columns=self.rateTables[company]['PublicTransportationPhysicalDamagePrimaryFactor'][0])

        merged_df = pd.merge(liab_factor, phys_factor, on='PrimaryClassCode', how='left')

        categories_data = {
            'Categories': [],
            'Seating Capacity': [],
            'Class Code': [],
            'Liability Factor': [],
            'Physical Damage Factor': []
        }

        for index, row in merged_df.iterrows():
            if row['PrimaryClassCode'] in [4111, 4112, 4113, 4114]:
                categories_data['Categories'].append('Van Pools - Employer Furnished')
                if row['PrimaryClassCode'] in [4111]:
                    categories_data['Seating Capacity'].append('1 to 8')
                elif row['PrimaryClassCode'] in [4112]:
                    categories_data['Seating Capacity'].append('9 to 20')
                elif row['PrimaryClassCode'] in [4113]:
                    categories_data['Seating Capacity'].append('21 to 60')
                elif row['PrimaryClassCode'] in [4114]:
                    categories_data['Seating Capacity'].append('Over 60')
                categories_data['Class Code'].append(int(row['PrimaryClassCode']))
                categories_data['Liability Factor'].append(round(row['Factor_x'], 3))
                categories_data['Physical Damage Factor'].append(round(row['Factor_y'], 3))
            elif row['PrimaryClassCode'] in [4121, 4122, 4123, 4124]:
                categories_data['Categories'].append('Van Pools - All Other')
                if row['PrimaryClassCode'] in [4121]:
                    categories_data['Seating Capacity'].append('1 to 8')
                elif row['PrimaryClassCode'] in [4122]:
                    categories_data['Seating Capacity'].append('9 to 20')
                elif row['PrimaryClassCode'] in [4123]:
                    categories_data['Seating Capacity'].append('21 to 60')
                elif row['PrimaryClassCode'] in [4124]:
                    categories_data['Seating Capacity'].append('Over 60')
                categories_data['Class Code'].append(int(row['PrimaryClassCode']))
                categories_data['Liability Factor'].append(round(row['Factor_x'], 3))
                categories_data['Physical Damage Factor'].append(round(row['Factor_y'], 3))

        categories_df = pd.DataFrame(categories_data)

        pivot_table = categories_df.pivot_table(index=['Categories', 'Seating Capacity', 'Class Code'],values=['Liability Factor', 'Physical Damage Factor'])

        # Sort by Class Code to maintain order
        pivot_table = pivot_table.sort_values(by='Class Code').reset_index()

        return pivot_table

    def buildVanSecondary(self, company):
        liab_factor = pd.DataFrame(self.rateTables[company]['PublicTransportationLiabilitySecondaryFactor'][1:], index=None, columns=self.rateTables[company]['PublicTransportationLiabilitySecondaryFactor'][0])
        phys_factor = pd.DataFrame(self.rateTables[company]['PublicTransportationPhysicalDamageSecondaryRate'][1:], index=None, columns=self.rateTables[company]['PublicTransportationPhysicalDamageSecondaryRate'][0])

        class_code_map = pd.DataFrame({
            'Secondary Class Code': ['1', '2', '3', '4'],
            'Seating Capacity': ['1 to 8', '9 to 20', '21 to 60', 'Over 60']
        })

        # Filter the factors for School and Church Buses
        school_church_liab = liab_factor[liab_factor['PublicTransportationGroup'] == 'School And Church Buses'].drop(columns = ["PublicTransportationGroup"]).rename(columns = {"SeatingCapacity" : "Seating Capacity"})
        school_church_phys = phys_factor[phys_factor['PublicTransportationGroup'] == 'School And Church Buses'].drop(columns = ["PublicTransportationGroup"]).rename(columns = {"SeatingCapacity" : "Seating Capacity"})

        # Filter the factors for Other Buses
        other_buses_liab = liab_factor[liab_factor['PublicTransportationGroup'] == 'Other Buses'].drop(columns = ["PublicTransportationGroup"]).rename(columns = {"SeatingCapacity" : "Seating Capacity"})
        other_buses_phys = phys_factor[phys_factor['PublicTransportationGroup'] == 'Other Buses'].drop(columns = ["PublicTransportationGroup"]).rename(columns = {"SeatingCapacity" : "Seating Capacity"})

        class_code_map = class_code_map.merge(school_church_liab, on='Seating Capacity', how='left').rename(
            columns={"Factor": "School and Church Buses Liability"}
        )
        class_code_map = class_code_map.merge(other_buses_liab, on='Seating Capacity', how='left').rename(
            columns={"Factor": "Other Buses Liability"}
        )

        class_code_map = class_code_map.merge(school_church_phys, on='Seating Capacity', how='left').rename(
            columns={"Factor": "School and Church Buses Physical"}
        )
        class_code_map = class_code_map.merge(other_buses_phys, on='Seating Capacity', how='left').rename(
            columns={"Factor": "Other Buses Physical"}
        )

        # Create multi-index
        class_code_map.columns = pd.MultiIndex.from_tuples([
            ('', 'Secondary Class Code'),
            ('', 'Seating Capacity'),
            ('Liability Factor', 'School and Church Buses'),
            ('Liability Factor', 'Other Buses'),
            ('Physical Damage Factor', 'School and Church Buses'),
            ('Physical Damage Factor', 'Other Buses')
        ])

        return class_code_map



    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildGarageKeepers1(self, company):
        GKComp = pd.DataFrame(self.rateTables[company]['GaragekeepersOtherThanCollisionPreliminaryBasePremium'][1:], index=None, columns=self.rateTables[company]['GaragekeepersOtherThanCollisionPreliminaryBasePremium'][0])
        #GKComp = self.buildDataFrame("GaragekeepersOtherThanCollisionPreliminaryBasePremium")
        GKComp = GKComp.query(f'GaragekeepersOtherThanCollisionCoverageType == "Comprehensive" & GaragekeepersOtherThanCollisionRatingBase == "Direct (Primary)" & GaragekeepersOtherThanCollisionLimit != 999999999').rename(columns={'GaragekeepersOtherThanCollisionLimit' : 'Limit', 'Premium' : '$100 per auto/$500 per loss deductible'}).filter(items=['Limit', '$100 per auto/$500 per loss deductible'])
        GKColl = pd.DataFrame(self.rateTables[company]['GaragekeepersCollisionPreliminaryBasePremium'][1:], index=None, columns=self.rateTables[company]['GaragekeepersCollisionPreliminaryBasePremium'][0])
        #GKColl = self.buildDataFrame("GaragekeepersCollisionPreliminaryBasePremium")
        GKColl = GKColl.query(f'GaragekeepersCollisionRatingBase == "Direct (Primary)" & GaragekeepersCollisionLimit != 999999999 & GaragekeepersCollisionDeductible == "100"').rename(columns={'GaragekeepersCollisionLimit' : 'Limit', 'Premium' : '$100 per auto deductible'}).filter(items=['Limit', '$100 per auto deductible'])
        GKFactors = pd.merge(GKComp, GKColl, on= 'Limit', how = 'inner')
        GKFactors = GKFactors.astype(object)
        GKFactors.iloc[:, 1:] = GKFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.0f}")

        return GKFactors

    # Builds the Driver Based Rating table
    # Returns a dataframe
    # rule 55
    @log_exceptions
    def buildGarageKeepers2(self, company):
        GKComp = pd.DataFrame(self.rateTables[company]['GaragekeepersOtherThanCollisionDeductibleFactor'][1:], index=None, columns=self.rateTables[company]['GaragekeepersOtherThanCollisionDeductibleFactor'][0])
        #GKComp = self.buildDataFrame("GaragekeepersOtherThanCollisionDeductibleFactor")
        GKComp = GKComp.rename(columns={'GaragekeepersOtherThanCollisionDeductible' : 'Comprehensive'}).sort_values('Factor', ascending=False)

        return GKComp

    Factor100 = ''
    Factor250 = ''
    Factor500 = ''
    Factor1000 = ''
    Factor2500 = ''
    Factor5000 = ''
    Factor10000 = ''
    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildGarageKeepers3(self, company):
        global Factor100
        global Factor250
        global Factor500
        global Factor1000
        global Factor2500
        global Factor5000
        global Factor10000
        GKColl = pd.DataFrame(self.rateTables[company]['GaragekeepersCollisionPreliminaryBasePremium'][1:], index=None, columns=self.rateTables[company]['GaragekeepersCollisionPreliminaryBasePremium'][0])
        #GKColl = self.buildDataFrame("GaragekeepersCollisionPreliminaryBasePremium")
        GKColl = GKColl.query(f'GaragekeepersCollisionLimit == 999999999 & GaragekeepersCollisionRatingBase == "Direct (Primary)"').sort_values('GaragekeepersCollisionDeductible').rename(columns={'GaragekeepersCollisionDeductible' : 'Collision'}).filter(items=['Collision', 'Premium']).sort_values('Premium', ascending=False)
        Factor100 = GKColl.iloc[0,1]
        Factor250 = GKColl.iloc[1,1]
        Factor500 = GKColl.iloc[2,1]
        Factor1000 = GKColl.iloc[3,1]
        Factor2500 = GKColl.iloc[4,1]
        Factor5000 = GKColl.iloc[5,1]
        Factor10000 = GKColl.iloc[6,1]

        GKColl.loc[GKColl['Collision'] == "100", 'Factor'] = GKColl.iloc[0,1] / GKColl.iloc[0,1]
        GKColl.loc[GKColl['Collision'] == "250", 'Factor'] = GKColl.iloc[1,1] / GKColl.iloc[0,1]
        GKColl.loc[GKColl['Collision'] == "500", 'Factor'] = GKColl.iloc[2,1] / GKColl.iloc[0,1]
        GKColl.loc[GKColl['Collision'] == "1,000", 'Factor'] = GKColl.iloc[3,1] / GKColl.iloc[0,1]
        GKColl.loc[GKColl['Collision'] == "2,500", 'Factor'] = GKColl.iloc[4,1] / GKColl.iloc[0,1]
        GKColl.loc[GKColl['Collision'] == "5,000", 'Factor'] = GKColl.iloc[5,1] / GKColl.iloc[0,1]
        GKColl.loc[GKColl['Collision'] == "10,000", 'Factor'] = GKColl.iloc[6,1] / GKColl.iloc[0,1]

        GKColl = GKColl[['Collision', 'Factor']]
        GKColl = GKColl.astype(object)
        GKColl.iloc[:, 1:] = GKColl.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return GKColl

    # Rule 264.B.1
    def buildAmbulanceFactors1(self, company):

        sheet_names = ['SpecialTypesAmbulanceFactor'] * 6
        orig_values = ['Liability','MedicalPayments','No-Fault','PhysDamOTCACV','PhysDamCollACV','UM/UIM']
        replace_values = ['Liability','Med','PIP','Other than Collision','Collision','All Other']

        if self.StateAbb not in self.pip_states.values:
            orig_values.remove("No-Fault")
            replace_values.remove("PIP")
        if self.StateAbb in self.no_med_states.values:
            orig_values.remove("MedicalPayments")
            replace_values.remove("Med")

        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values=orig_values)
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Rule 264.B.1
    def buildAmbulanceFactors2(self, company):
        sheet_names = ['SpecialTypesEmergencyVehicleBuybackFactor']
        orig_values = ['Y']
        replace_values = ['']

        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: 'Factor'})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule68table1(self, company):
        # 268.C. Driver Training Owned Auto Factors
        Rule68Table1a = pd.DataFrame(self.rateTables[company]['SpecialTypesDriverTrainingFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesDriverTrainingFactor'][0]).query(f'SpecialTypesClassCode == "7201"').replace({'TypeCoverage' : {'MedicalPayments' : 'Med', 'No-Fault' : 'PIP', 'PhysDamOTCACV' : 'Other than Collision', 'PhysDamCollACV' : 'Collision', 'UM/UIM' : 'All Other' }}).rename(columns={'Factor' : 'Equipped with Dual Controls - Class 7201', 'TypeCoverage' : 'Coverage'}).filter(items=['Coverage', 'Equipped with Dual Controls - Class 7201'])
        Rule68Table1b = pd.DataFrame(self.rateTables[company]['SpecialTypesDriverTrainingFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesDriverTrainingFactor'][0]).query(f'SpecialTypesClassCode == "7202"').replace({'TypeCoverage' : {'MedicalPayments' : 'Med', 'No-Fault' : 'PIP', 'PhysDamOTCACV' : 'Other than Collision', 'PhysDamCollACV' : 'Collision', 'UM/UIM' : 'All Other' }}).rename(columns={'Factor' : 'Not Equipped with Dual Controls - Class 7202', 'TypeCoverage' : 'Coverage'}).filter(items=['Coverage', 'Not Equipped with Dual Controls - Class 7202'])
        Rule68Table1 = pd.merge(Rule68Table1a, Rule68Table1b, on= 'Coverage', how = 'inner').query(f'Coverage != "PhysDamCollSA" & Coverage != "PhysDamOTCSA"')

        Rule68Table1a = pd.DataFrame(self.rateTables[company]['SpecialTypesDriverTrainingFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesDriverTrainingFactor'][0]).query(f'SpecialTypesClassCode == "7223"').replace({'TypeCoverage' : {'MedicalPayments' : 'Med', 'No-Fault' : 'PIP', 'PhysDamOTCACV' : 'Other than Collision', 'PhysDamCollACV' : 'Collision', 'UM/UIM' : 'All Other' }}).rename(columns={'Factor' : 'Equipped with Dual Controls - Class 7223', 'TypeCoverage' : 'Coverage'}).filter(items=['Coverage', 'Equipped with Dual Controls - Class 7223'])
        Rule68Table1b = pd.DataFrame(self.rateTables[company]['SpecialTypesDriverTrainingFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesDriverTrainingFactor'][0]).query(f'SpecialTypesClassCode == "7224"').replace({'TypeCoverage' : {'MedicalPayments' : 'Med', 'No-Fault' : 'PIP', 'PhysDamOTCACV' : 'Other than Collision', 'PhysDamCollACV' : 'Collision', 'UM/UIM' : 'All Other' }}).rename(columns={'Factor' : 'Not Equipped with Dual Controls - Class 7224', 'TypeCoverage' : 'Coverage'}).filter(items=['Coverage', 'Not Equipped with Dual Controls - Class 7224'])

        Rule68TableAB = pd.merge(Rule68Table1a, Rule68Table1b, on= 'Coverage', how = 'inner').query(f'Coverage != "PhysDamCollSA" & Coverage != "PhysDamOTCSA"')
        Rule68Table2 = pd.merge(Rule68Table1, Rule68TableAB, on= 'Coverage', how = 'inner').query(f'Coverage != "PhysDamCollSA" & Coverage != "PhysDamOTCSA"')
        Rule68Table2 = Rule68Table2.astype(object)
        Rule68Table2.iloc[:, 1:] = Rule68Table2.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        no_med = pd.read_excel(BA_INPUT_FILE, sheet_name="No MedPay")

        if self.StateAbb not in self.pip_states.values:
            Rule68Table2 = Rule68Table2[Rule68Table2["Coverage"] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            Rule68Table2 = Rule68Table2[Rule68Table2["Coverage"] != "Med"]

        return Rule68Table2

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule68table3(self, company):
        Rule68Table3 = pd.DataFrame(self.rateTables[company]['DriverTrainingLiabilityAndMedicalPaymentsCoveragesFactor'][1:], index=None, columns=self.rateTables[company]['DriverTrainingLiabilityAndMedicalPaymentsCoveragesFactor'][0]).query(f'Constant == "Y"').filter(items=['Factor'])
        #Rule68Table3 = self.buildDataFrame("DriverTrainingLiabilityAndMedicalPaymentsCoveragesFactor").query(f'Constant == "Y"').filter(items=['Factor'])
        Rule68Table3 = Rule68Table3.astype(object)
        Rule68Table3.iloc[:,:] = Rule68Table3.iloc[:,:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule68Table3

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule68table4(self, company):
        Rule68Table4 = pd.DataFrame(self.rateTables[company]['DriverTrainingMedicalPaymentsCoverageFactors'][1:], index=None, columns=self.rateTables[company]['DriverTrainingMedicalPaymentsCoverageFactors'][0]).query(f'MedicalPaymentsLimitText != "No Coverage"').rename(columns={'MedicalPaymentsLimitText' : 'Limit Per Person'}).sort_values('Factor')
        #Rule68Table4 = self.buildDataFrame("DriverTrainingMedicalPaymentsCoverageFactors").query(f'MedicalPaymentsLimitText != "No Coverage"').rename(columns={'MedicalPaymentsLimitText' : 'Limit Per Person'}).sort_values('Factor')
        Rule68Table4 = Rule68Table4.astype(object)
        Rule68Table4.iloc[:, 1:] = Rule68Table4.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule68Table4

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule79table1(self, company):
        Rule79Table1a = pd.DataFrame(self.rateTables[company]['SpecialTypesRepossessedAutosLiabilityBasePremium'][1:], index=None, columns=self.rateTables[company]['SpecialTypesRepossessedAutosLiabilityBasePremium'][0]).rename(columns={'Factor' : 'Rate'}).filter(items=['Rate'])
        Rule79Table1b = pd.DataFrame(self.rateTables[company]['SpecialTypesRepossessedAutosLiabilityMinimumPremium'][1:], index=None, columns=self.rateTables[company]['SpecialTypesRepossessedAutosLiabilityMinimumPremium'][0]).rename(columns={'Premium' : 'Minimum Premium'}).filter(items=['Minimum Premium'])
        Rule79Table1 = pd.concat([Rule79Table1a, Rule79Table1b], axis=1, join="inner")
        Rule79Table1['Rate'] = Rule79Table1['Rate'].apply(lambda x: "${:,.2f}".format(x))

        return Rule79Table1

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule79table2(self, company):
        Rule79Table2 = pd.DataFrame(self.rateTables[company]['GarageDealersOtherThanCollisionRate'][1:], index=None, columns=self.rateTables[company]['GarageDealersOtherThanCollisionRate'][0])\

        helper_map = pd.read_excel(BA_INPUT_FILE, sheet_name="283 Helper Map")
        # Create a dictionary for mapping Input values to Output values
        mapping_dict = dict(zip(helper_map["Input"], helper_map["Output"]))
        Rule79Table2.replace(mapping_dict, inplace=True)

        Rule79Table2 = Rule79Table2.query(f'SpecialTypesOtherThanCollisionCoverageType == "Stated Amount" & SpecialTypesSupplementaryType == "Miscellaneous Type Vehicles Buildings"').filter(items=['Territory', 'Factor']).rename(columns={'Territory' : 'Territories', 'Factor' : 'Rate per $100'})
        Rule79Table2['Territories'] = "All Territories"
        Rule79Table2 = Rule79Table2.iloc[:1]
        Rule79Table2['Rate per $100'] = Rule79Table2['Rate per $100']
        Rule79Table2['Rate per $100'] = Rule79Table2['Rate per $100'].apply(lambda x: "${:,.2f}".format(x))

        return Rule79Table2

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule279table3(self, company):

        # State exception for MI. Very long--annoying table.
        if self.StateAbb == "MI":
            # Load the rate and deductible tables
            rate_table = pd.DataFrame(
                self.rateTables[company]['GarageDealersCollisionBlanketRateMI'][1:],
                columns=self.rateTables[company]['GarageDealersCollisionBlanketRateMI'][0]
            )
            deduct_table = pd.DataFrame(
                self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][1:],
                columns=self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][0]
            )

            deduct_table = deduct_table[deduct_table["SpecialTypesCollisionDeductible"] != "100"]

            # Convert numeric columns to float
            deduct_table["Factor"] = deduct_table["Factor"].astype(float)
            rate_table["Factor"] = rate_table["Factor"].astype(float)

            # Prepare output list
            combined_tables = []

            # Process each collision type separately
            for collision_type in ["Broadened Collision", "Collision", "Limited Collision"]:
                # Filter rate table for this collision type
                sub_rate = rate_table[rate_table["CollisionTypeGarageDealers"] == collision_type]

                # Extract scaling factors for $250 deductible
                low = float(sub_rate[(sub_rate["SpecialTypesCollisionDeductible"] == "250") & (
                            sub_rate["CoverageBand"] == "Low")]["Factor"].values[0])
                med = float(sub_rate[(sub_rate["SpecialTypesCollisionDeductible"] == "250") & (
                            sub_rate["CoverageBand"] == "Medium")]["Factor"].values[0])
                high = float(sub_rate[(sub_rate["SpecialTypesCollisionDeductible"] == "250") & (
                            sub_rate["CoverageBand"] == "High")]["Factor"].values[0])

                # Copy and scale deductible table
                temp = deduct_table.copy()
                temp["First $50,000"] = temp["Factor"] * low
                temp["$50,000 to $100,000"] = temp["Factor"] * med
                temp["Over $100,000"] = temp["Factor"] * high
                temp["CollisionTypeGarageDealers"] = collision_type

                # Drop original factor column
                temp = temp.drop(columns=["Factor"])
                combined_tables.append(temp)

            # Combine all collision types
            final_df = pd.concat(combined_tables, ignore_index=True)

            # Rename and sort
            final_df = final_df.rename(
                columns={"CollisionTypeGarageDealers": "Coverage", "SpecialTypesCollisionDeductible": "Deductible"})
            final_df = final_df[["Deductible", "Coverage", "First $50,000", "$50,000 to $100,000", "Over $100,000"]]
            final_df["Deductible"] = final_df["Deductible"].str.replace(",","").astype(int)
            final_df = final_df.sort_values(by=["Deductible", "Coverage"])
            final_df["Deductible"] = final_df["Deductible"].apply(lambda x: "{:,.0f}".format(x))

            # Format currency
            for col in ["First $50,000", "$50,000 to $100,000", "Over $100,000"]:
                final_df[col] = final_df[col].apply(lambda x: "${:,.2f}".format(x))

            return final_df

        rate_table = pd.DataFrame(self.rateTables[company]['GarageDealersCollisionBlanketRate'][1:], index=None, columns=self.rateTables[company]['GarageDealersCollisionBlanketRate'][0])
        deduct_table = pd.DataFrame(self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][1:], index=None, columns=self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][0])

        output_table = deduct_table.rename(columns = {"SpecialTypesCollisionDeductible" : "Deductible", "Factor" : "First $50,000"})
        output_table["$50,000 to $100,000"] = output_table["First $50,000"]
        output_table["Over $100,000"] = output_table["First $50,000"]

        # Convert to numeric, sort, then format with commas
        output_table["Deductible"] = (
            output_table["Deductible"]
            .str.replace(r"[^\d.]", "", regex=True)  # Remove non-numeric characters
            .astype(float)  # Convert to float
        )

        # Sort by numeric deductible
        output_table = output_table.sort_values("Deductible")

        # Format back with commas and no decimals
        output_table["Deductible"] = output_table["Deductible"].apply(
            lambda x: f"{x:,.0f}" if pd.notna(x) else ""
        )

        low_100_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "100") & (rate_table["CoverageBand"] == "Low")].iloc[0,2]
        med_100_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "100") & (rate_table["CoverageBand"] == "Medium")].iloc[0,2]
        high_100_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "100") & (rate_table["CoverageBand"] == "High")].iloc[0,2]

        low_250_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "250") & (rate_table["CoverageBand"] == "Low")].iloc[0,2]
        med_250_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "250") & (rate_table["CoverageBand"] == "Medium")].iloc[0,2]
        high_250_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "250") & (rate_table["CoverageBand"] == "High")].iloc[0,2]

        # Apply the 100 factor to the 100$ Deductible
        output_table.iloc[0,1:] = output_table.iloc[0,1:] * [low_100_factor, med_100_factor, high_100_factor]

        # Apply the 250 factor to the 250$+ Deductibles.
        output_table.iloc[1:,1:] = output_table.iloc[1:,1:] * [low_250_factor, med_250_factor, high_250_factor]

        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].map(lambda x: "${:,.2f}".format(x))

        return output_table

    # Rule 283
    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule83table1(self, company):
        Rule83Table1a = pd.DataFrame(self.rateTables[company]['GarageDealersOtherThanCollisionRate'][1:], index=None, columns=self.rateTables[company]['GarageDealersOtherThanCollisionRate'][0])

        helper_map = pd.read_excel(BA_INPUT_FILE, sheet_name="283 Helper Map")

        # Create a dictionary for mapping Input values to Output values
        mapping_dict = dict(zip(helper_map["Input"], helper_map["Output"]))

        # Replace values in Rule83Table1a based on mapping
        Rule83Table1a.replace(mapping_dict, inplace=True)
        Rule83Table1a = Rule83Table1a.rename(columns={'SpecialTypesOtherThanCollisionCoverageType' : 'CoverageType', 'SpecialTypesSupplementaryType' : 'SuppType'}).query(f'CoverageType == "Stated Amount - Specified Causes of Loss"')
        Rule83Table1a = Rule83Table1a.pivot(index=['Territory', 'CoverageType'], columns='SuppType', values='Factor').reset_index(['Territory', 'CoverageType']).filter(items=['Territory', 'Miscellaneous Type Vehicles Buildings', 'Miscellaneous Type Vehicles Open Lots', 'Personal Auto Type Vehicles Buildings', 'Personal Auto Type Vehicles Non-Standard Open Lots', 'Personal Auto Type Vehicles Standard Open Lots'])
        Rule83Table1a = Rule83Table1a.astype(object)
        Rule83Table1a.iloc[:, 1:] = Rule83Table1a.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule83Table1a

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule83table2(self, company):
        Rule83Table2a = pd.DataFrame(self.rateTables[company]['GarageDealersOtherThanCollisionRate'][1:], index=None, columns=self.rateTables[company]['GarageDealersOtherThanCollisionRate'][0])
        #Rule83Table2a = self.buildDataFrame("GarageDealersOtherThanCollisionRate").rename(columns={'SpecialTypesOtherThanCollisionCoverageType' : 'CoverageType', 'SpecialTypesSupplementaryType' : 'SuppType'}).query(f'CoverageType == "Stated Amount - Limited Specified Causes of Loss"')
        helper_map = pd.read_excel(BA_INPUT_FILE, sheet_name="283 Helper Map")

        # Create a dictionary for mapping Input values to Output values
        mapping_dict = dict(zip(helper_map["Input"], helper_map["Output"]))

        # Replace values in Rule83Table1a based on mapping
        Rule83Table2a.replace(mapping_dict, inplace=True)
        Rule83Table2a = Rule83Table2a.rename(columns={'SpecialTypesOtherThanCollisionCoverageType' : 'CoverageType', 'SpecialTypesSupplementaryType' : 'SuppType'}).replace({"Stated Amount - Specified Perils": "Stated Amount - Limited Specified Causes of Loss"}).query(f'CoverageType == "Stated Amount - Limited Specified Causes of Loss"')

        Rule83Table2a = Rule83Table2a.pivot(index=['Territory', 'CoverageType'], columns='SuppType', values='Factor').reset_index(['Territory', 'CoverageType']).filter(items=['Territory', 'Miscellaneous Type Vehicles Buildings', 'Miscellaneous Type Vehicles Open Lots', 'Personal Auto Type Vehicles Buildings', 'Personal Auto Type Vehicles Non-Standard Open Lots', 'Personal Auto Type Vehicles Standard Open Lots'])
        Rule83Table2a = Rule83Table2a.astype(object)
        Rule83Table2a.iloc[:, 1:] = Rule83Table2a.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule83Table2a

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule83table3(self, company):
        Rule83Table3a = pd.DataFrame(self.rateTables[company]['GarageDealersOtherThanCollisionRate'][1:], index=None, columns=self.rateTables[company]['GarageDealersOtherThanCollisionRate'][0])

        helper_map = pd.read_excel(BA_INPUT_FILE, sheet_name="283 Helper Map")

        # Create a dictionary for mapping Input values to Output values
        mapping_dict = dict(zip(helper_map["Input"], helper_map["Output"]))

        # Replace values in Rule83Table1a based on mapping
        Rule83Table3a.replace(mapping_dict, inplace=True)
        Rule83Table3a = Rule83Table3a.rename(columns={'SpecialTypesOtherThanCollisionCoverageType': 'CoverageType',
                         'SpecialTypesSupplementaryType': 'SuppType'}).query(f'CoverageType == "Stated Amount"')
        Rule83Table3a = Rule83Table3a.pivot(index=['Territory', 'CoverageType'], columns='SuppType', values='Factor').reset_index(['Territory', 'CoverageType']).filter(items=['Territory', 'Miscellaneous Type Vehicles Buildings', 'Miscellaneous Type Vehicles Open Lots', 'Personal Auto Type Vehicles Buildings', 'Personal Auto Type Vehicles Non-Standard Open Lots', 'Personal Auto Type Vehicles Standard Open Lots'])
        Rule83Table3a = Rule83Table3a.astype(object)
        Rule83Table3a.iloc[:, 1:] = Rule83Table3a.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule83Table3a

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule83table4(self, company):

        # State exception for MI. Very long--annoying table.
        if self.StateAbb == "MI":
            # Load the rate and deductible tables
            rate_table = pd.DataFrame(
                self.rateTables[company]['GarageDealersCollisionBlanketRateMI'][1:],
                columns=self.rateTables[company]['GarageDealersCollisionBlanketRateMI'][0]
            )
            deduct_table = pd.DataFrame(
                self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][1:],
                columns=self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][0]
            )

            deduct_table = deduct_table[deduct_table["SpecialTypesCollisionDeductible"] != "100"]

            # Convert numeric columns to float
            deduct_table["Factor"] = deduct_table["Factor"].astype(float)
            rate_table["Factor"] = rate_table["Factor"].astype(float)

            # Prepare output list
            combined_tables = []

            # Process each collision type separately
            for collision_type in ["Broadened Collision", "Collision", "Limited Collision"]:
                # Filter rate table for this collision type
                sub_rate = rate_table[rate_table["CollisionTypeGarageDealers"] == collision_type]

                # Extract scaling factors for $250 deductible
                low = float(sub_rate[(sub_rate["SpecialTypesCollisionDeductible"] == "250") & (
                            sub_rate["CoverageBand"] == "Low")]["Factor"].values[0])
                med = float(sub_rate[(sub_rate["SpecialTypesCollisionDeductible"] == "250") & (
                            sub_rate["CoverageBand"] == "Medium")]["Factor"].values[0])
                high = float(sub_rate[(sub_rate["SpecialTypesCollisionDeductible"] == "250") & (
                            sub_rate["CoverageBand"] == "High")]["Factor"].values[0])

                # Copy and scale deductible table
                temp = deduct_table.copy()
                temp["First $50,000"] = temp["Factor"] * low
                temp["$50,000 to $100,000"] = temp["Factor"] * med
                temp["Over $100,000"] = temp["Factor"] * high
                temp["CollisionTypeGarageDealers"] = collision_type

                # Drop original factor column
                temp = temp.drop(columns=["Factor"])
                combined_tables.append(temp)

            # Combine all collision types
            final_df = pd.concat(combined_tables, ignore_index=True)

            # Rename and reorder columns
            final_df = final_df.rename(columns={
                "CollisionTypeGarageDealers": "Coverage",
                "SpecialTypesCollisionDeductible": "Deductible"
            })
            final_df = final_df[["Deductible", "Coverage", "First $50,000", "$50,000 to $100,000", "Over $100,000"]]

            # Clean and convert Deductible
            final_df["Deductible"] = final_df["Deductible"].astype(str).str.replace(",", "")
            final_df["Deductible"] = final_df["Deductible"].astype(int)

            # Replace coverage names
            final_df["Coverage"] = final_df["Coverage"].replace({
                "Broadened Collision": "Broadened",
                "Limited Collision": "Limited",
                "Collision": "Regular"
            })

            # Set custom order for Coverage
            coverage_order = ["Regular", "Limited", "Broadened"]
            final_df["Coverage"] = pd.Categorical(final_df["Coverage"], categories=coverage_order, ordered=True)

            # Sort by Deductible and Coverage
            final_df = final_df.sort_values(by=["Deductible", "Coverage"])

            # Format Deductible as string with commas
            final_df["Deductible"] = final_df["Deductible"].apply(lambda x: "{:,.0f}".format(x))

            # Format currency
            for col in ["First $50,000", "$50,000 to $100,000", "Over $100,000"]:
                final_df[col] = final_df[col].apply(lambda x: "${:,.2f}".format(x))

            return final_df


        # Regular calcualtion for almost all states. Though MI was made a bit better,
        rate_table = pd.DataFrame(self.rateTables[company]['GarageDealersCollisionBlanketRate'][1:], index=None, columns=self.rateTables[company]['GarageDealersCollisionBlanketRate'][0])
        deduct_table = pd.DataFrame(self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][1:], index=None, columns=self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][0])

        output_table = deduct_table.rename(columns = {"SpecialTypesCollisionDeductible" : "Deductible", "Factor" : "First $50,000"})
        output_table["$50,000 to $100,000"] = output_table["First $50,000"]
        output_table["Over $100,000"] = output_table["First $50,000"]

        # Convert to numeric, sort, then format with commas
        output_table["Deductible"] = (
            output_table["Deductible"]
            .str.replace(r"[^\d.]", "", regex=True)  # Remove non-numeric characters
            .astype(float)  # Convert to float
        )

        # Sort by numeric deductible
        output_table = output_table.sort_values("Deductible")

        # Format back with commas and no decimals
        output_table["Deductible"] = output_table["Deductible"].apply(
            lambda x: f"{x:,.0f}" if pd.notna(x) else ""
        )

        low_100_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "100") & (rate_table["CoverageBand"] == "Low")].iloc[0,2]
        med_100_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "100") & (rate_table["CoverageBand"] == "Medium")].iloc[0,2]
        high_100_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "100") & (rate_table["CoverageBand"] == "High")].iloc[0,2]

        low_250_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "250") & (rate_table["CoverageBand"] == "Low")].iloc[0,2]
        med_250_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "250") & (rate_table["CoverageBand"] == "Medium")].iloc[0,2]
        high_250_factor = rate_table[(rate_table["SpecialTypesCollisionDeductible"] == "250") & (rate_table["CoverageBand"] == "High")].iloc[0,2]

        # Apply the 100 factor to the 100$ Deductible
        output_table.iloc[0,1:] = output_table.iloc[0,1:] * [low_100_factor, med_100_factor, high_100_factor]

        # Apply the 250 factor to the 250$+ Deductibles.
        output_table.iloc[1:,1:] = output_table.iloc[1:,1:] * [low_250_factor, med_250_factor, high_250_factor]

        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].map(lambda x: f"{x:.2f}")

        return output_table

    def buildRule283_MI(self,company):
        # 283.B.2.f. Service Operations Collisions Adjustment Factor
        # No ratebook sheet exists for this rule
        data = pd.DataFrame({"Factor" : ["1.000"]})

        return data


    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table1(self, company):
        Rule89Table1 = pd.DataFrame(self.rateTables[company]['NonOwnedBasePremium'][1:], index=None, columns=self.rateTables[company]['NonOwnedBasePremium'][0]).rename(columns={'BasePremium' : 'Liability Limit'})
        #Rule89Table1 = self.buildDataFrame("NonOwnedBasePremium").rename(columns={'BasePremium' : '$100,000 Liability Limit *'})
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 0, 'Total Number of Employees:'] = '0-9'
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 10, 'Total Number of Employees:'] = '10-19'
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 20, 'Total Number of Employees:'] = '20-25'
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 26, 'Total Number of Employees:'] = '26-100'
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 101, 'Total Number of Employees:'] = '101-500'
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 501, 'Total Number of Employees:'] = '501-1,000'
        Rule89Table1.loc[Rule89Table1['NumberOfEmployees'] == 1001, 'Total Number of Employees:'] = 'Over 1,000'
        Rule89Table1 = Rule89Table1.filter(items=['Total Number of Employees:', 'Liability Limit'])
        Rule89Table1 = Rule89Table1[['Total Number of Employees:', 'Liability Limit']]
        Rule89Table1['Liability Limit'] = Rule89Table1['Liability Limit']

        Rule89Table1 = Rule89Table1.astype(object)
        Rule89Table1.iloc[:,1:] = Rule89Table1.iloc[:,1:].map(lambda x: f"{x:.0f}")

        return Rule89Table1

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table2(self, company):
        Rule89Table2 = pd.DataFrame(self.rateTables[company]['NonOwnedPartnershipFactor'][1:], index=None, columns=self.rateTables[company]['NonOwnedPartnershipFactor'][0]).query(f'Constant == "Y"').filter(items=['Factor'])
        #Rule89Table2 = self.buildDataFrame("NonOwnedPartnershipFactor").query(f'Constant == "Y"').filter(items=['Factor'])
        Rule89Table2 = Rule89Table2.astype(object)
        Rule89Table2.iloc[:,:] = Rule89Table2.iloc[:,:].astype(float).map(lambda x: f"{x:.3f}")
        return Rule89Table2

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table3(self, company):
        Rule89Table3 = pd.DataFrame(self.rateTables[company]['GarageServicesLiabilityEmployeesFactor'][1:], index=None, columns=self.rateTables[company]['GarageServicesLiabilityEmployeesFactor'][0]).query(f'Constant == "Y"').rename(columns={'EmployeeLiabilityFactor' : 'Factor'}).filter(items=['Factor'])
        #Rule89Table3 = self.buildDataFrame("GarageServicesLiabilityEmployeesFactor").query(f'Constant == "Y"').rename(columns={'EmployeeLiabilityFactor' : 'Factor'}).filter(items=['Factor'])
        Rule89Table3 = Rule89Table3.astype(object)
        Rule89Table3.iloc[:,:] = Rule89Table3.iloc[:,:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule89Table3

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table4(self, company):
        Rule89Table4 = pd.DataFrame(self.rateTables[company]['NonOwnedAutoFoodorGoodsDeliveryExposureFactor_Ext'][1:], index=None, columns=self.rateTables[company]['NonOwnedAutoFoodorGoodsDeliveryExposureFactor_Ext'][0]).query(f'Constant == "Y"').filter(items=['Factor'])
        #Rule89Table4 = self.buildDataFrame("NonOwnedAutoFoodorGoodsDeliveryExposureFactor_Ext").query(f'Constant == "Y"').filter(items=['Factor'])
        Rule89Table4 = Rule89Table4.astype(object)
        Rule89Table4.iloc[:,:] = Rule89Table4.iloc[:,:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule89Table4

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table5(self, company):
        Rule89Table5 = pd.DataFrame(self.rateTables[company]['NonOwnedAutoFoodorGoodsDeliveryIncreasedLimitFactor_Ext'][1:], index=None, columns=self.rateTables[company]['NonOwnedAutoFoodorGoodsDeliveryIncreasedLimitFactor_Ext'][0])
        #Rule89Table5 = self.buildDataFrame("NonOwnedAutoFoodorGoodsDeliveryIncreasedLimitFactor_Ext")
        Rule89Table5 = Rule89Table5.astype(object)
        Rule89Table5.iloc[:, 0] = Rule89Table5.iloc[:, 0].astype(str).str.replace(",", "", regex=False)
        Rule89Table5 = Rule89Table5.astype(object)
        Rule89Table5.iloc[:, :1] = Rule89Table5.iloc[:, :1].astype(float).map(lambda x: f"{x:,.0f}")
        Rule89Table5 = Rule89Table5.astype(object)
        Rule89Table5.iloc[:,1:] = Rule89Table5.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return Rule89Table5

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table6(self, company):
        Rule89Table6 = pd.DataFrame(self.rateTables[company]['NonOwnedAutoFoodorGoodsDeliveryMinimumPremiumPerPolicy_Ext'][1:], index=None, columns=self.rateTables[company]['NonOwnedAutoFoodorGoodsDeliveryMinimumPremiumPerPolicy_Ext'][0]).query(f'Constant == "Y"').filter(items=['MinimumPremium']).rename(columns={'MinimumPremium' : 'Minimum Premium'})
        #Rule89Table6 = self.buildDataFrame("NonOwnedAutoFoodorGoodsDeliveryMinimumPremiumPerPolicy_Ext").query(f'Constant == "Y"').filter(items=['MinimumPremium']).rename(columns={'MinimumPremium' : 'Minimum Premium'})

        return Rule89Table6

    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildRule89table7(self, company):
        Rule89Table7a = pd.DataFrame(self.rateTables[company]['NonOwnedVolunteersBasePremium'][1:], index=None, columns=self.rateTables[company]['NonOwnedVolunteersBasePremium'][0]).query(f'Constant == "Y"').filter(items=['Factor']).rename(columns={'Factor' : 'Liability Limit'})
        #Rule89Table7a = self.buildDataFrame("NonOwnedVolunteersBasePremium").query(f'Constant == "Y"').filter(items=['Factor']).rename(columns={'Factor' : '$100,000 Liability Limit *'})
        Rule89Table7b = pd.DataFrame(self.rateTables[company]['NonOwnedVolunteersMinimumPremium'][1:], index=None, columns=self.rateTables[company]['NonOwnedVolunteersMinimumPremium'][0]).query(f'Constant == "Y"').filter(items=['BasePremium']).rename(columns={'BasePremium' : 'Minimum Premium'})
        #Rule89Table7b = self.buildDataFrame("NonOwnedVolunteersMinimumPremium").query(f'Constant == "Y"').filter(items=['BasePremium']).rename(columns={'BasePremium' : 'Minimum Premium - $100,000 Limit*'})
        Rule89Table7c = pd.DataFrame(self.rateTables[company]['VolunteersAsInsuredsBasePremium'][1:], index=None, columns=self.rateTables[company]['VolunteersAsInsuredsBasePremium'][0]).query(f'Constant == "Y"').filter(items=['Factor']).rename(columns={'Factor' : 'Liability Limit'})
        Rule89Table7d = pd.DataFrame(self.rateTables[company]['VolunteersAsInsuredsMinimumPremium'][1:], index=None, columns=self.rateTables[company]['VolunteersAsInsuredsMinimumPremium'][0]).query(f'Constant == "Y"').filter(items=['BasePremium']).rename(columns={'BasePremium' : 'Minimum Premium'})
        #Rule89Table7d = self.buildDataFrame("VolunteersAsInsuredsMinimumPremium").query(f'Constant == "Y"').filter(items=['BasePremium']).rename(columns={'BasePremium' : 'Minimum Premium - $100,000 Limit*'})
        Rule89Table7ab = pd.concat([Rule89Table7a, Rule89Table7b], axis=1, join="inner")
        Rule89Table7ab['Volunteer'] = "Blanket Individual"
        Rule89Table7cd = pd.concat([Rule89Table7c, Rule89Table7d], axis=1, join="inner")
        Rule89Table7cd['Volunteer'] = "Volunteers Liability"
        Rule89Table7 = pd.concat([Rule89Table7ab, Rule89Table7cd])
        Rule89Table7 = Rule89Table7[['Volunteer', 'Minimum Premium']] # Edited to only have minimum premium.
        Rule89Table7 = Rule89Table7.rename(columns = {"Volunteer" : "Coverage"})
        Rule89Table7 = Rule89Table7.astype(object)
        Rule89Table7.iloc[:, 1:] = Rule89Table7.iloc[:,1:].map(lambda x: f"{x:.0f}")

        return Rule89Table7

    def buildRule89table8(self, company):
        #289.B.2.b.(3).(b) Extended Non-ownership Liability Employee Coverage Factor
        Rule89Table8 = pd.DataFrame(self.rateTables[company]['EmployeesAsInsuredsFactor'][1:], index=None, columns=self.rateTables[company]['EmployeesAsInsuredsFactor'][0]).query(f'Constant == "Y"').filter(items=['Factor'])
        #Rule89Table3 = self.buildDataFrame("GarageServicesLiabilityEmployeesFactor").query(f'Constant == "Y"').rename(columns={'EmployeeLiabilityFactor' : 'Factor'}).filter(items=['Factor'])

        Rule89Table8 = Rule89Table8.astype(object)
        Rule89Table8.iloc[:, :] = Rule89Table8.iloc[:, :].map(lambda x: f"{x:.3f}")

        return Rule89Table8

    def buildRule89tableB1b(self, company):
        # 289.B.1.b
        output_table = pd.DataFrame(self.rateTables[company]['VolunteersAsInsuredsBasePremium'][1:], index=None, columns=self.rateTables[company]['VolunteersAsInsuredsBasePremium'][0]).query(f'Constant == "Y"').filter(items=['Factor']).rename(columns={'Factor' : 'Liability Limit'})
        output_table = output_table.rename(columns = {"Liability Limit" : "Rate"})
        output_table = output_table.astype(object)
        output_table.iloc[:, :] = output_table.iloc[:, :].map(lambda x: f"{x:.2f}")

        return output_table

    def buildRule89tableB2b2a(self, company):
        # 289.B.2.b.(2).(a)
        output_table = pd.DataFrame(self.rateTables[company]['NonOwnedVolunteersBasePremium'][1:], index=None, columns=self.rateTables[company]['NonOwnedVolunteersBasePremium'][0]).query(f'Constant == "Y"').filter(items=['Factor']).rename(columns={'Factor' : 'Liability Limit'})
        output_table = output_table.rename(columns={"Liability Limit": "Rate"})
        output_table = output_table.astype(object)
        output_table.iloc[:, :] = output_table.iloc[:, :].map(lambda x: f"{x:.2f}")

        return output_table

    # fixing 298 Liability, needed complement taken on factors.
    @log_exceptions
    def buildRule298table1(self, company):
        csl_deductibles = [
            "250 Combined Single Limit", "500 Combined Single Limit", "1,000 Combined Single Limit",
            "2,500 Combined Single Limit", "5,000 Combined Single Limit", "10,000 Combined Single Limit",
            "20,000 Combined Single Limit", "25,000 Combined Single Limit", "50,000 Combined Single Limit",
            "75,000 Combined Single Limit", "100,000 Combined Single Limit"
        ]

        pd_deductibles = [
            "250 Property Damage Per Accident", "500 Property Damage Per Accident", "1,000 Property Damage Per Accident",
            "2,500 Property Damage Per Accident", "5,000 Property Damage Per Accident", "10,000 Property Damage Per Accident",
            "20,000 Property Damage Per Accident", "25,000 Property Damage Per Accident", "50,000 Property Damage Per Accident",
            "75,000 Property Damage Per Accident", "100,000 Property Damage Per Accident"
        ]

        csl_deductible_map = {
            "250 Combined Single Limit": 250,
            "500 Combined Single Limit": 500,
            "1,000 Combined Single Limit": 1000,
            "2,500 Combined Single Limit": 2500,
            "5,000 Combined Single Limit": 5000,
            "10,000 Combined Single Limit": 10000,
            "20,000 Combined Single Limit": 20000,
            "25,000 Combined Single Limit": 25000,
            "50,000 Combined Single Limit": 50000,
            "75,000 Combined Single Limit": 75000,
            "100,000 Combined Single Limit": 100000
        }

        pd_deductible_map = {
            "250 Property Damage Per Accident": 250,
            "500 Property Damage Per Accident": 500,
            "1,000 Property Damage Per Accident": 1000,
            "2,500 Property Damage Per Accident": 2500,
            "5,000 Property Damage Per Accident": 5000,
            "10,000 Property Damage Per Accident": 10000,
            "20,000 Property Damage Per Accident": 20000,
            "25,000 Property Damage Per Accident": 25000,
            "50,000 Property Damage Per Accident": 50000,
            "75,000 Property Damage Per Accident": 75000,
            "100,000 Property Damage Per Accident": 100000
        }

        Rule98Table1a = (
            pd.DataFrame(self.rateTables[company]['LiabilityDeductibleFactor'][1:],
                         columns=self.rateTables[company]['LiabilityDeductibleFactor'][0])
            .query("LiabilityDeductible in @csl_deductibles")
            .replace({'LiabilityDeductible': csl_deductible_map})
            .rename(columns={'LiabilityDeductible': 'Deductible Amount', 'Factor': 'Combined Single Limit'})
            .sort_values('Deductible Amount')
        )

        Rule98Table1b = (
            pd.DataFrame(self.rateTables[company]['LiabilityDeductibleFactor'][1:],
                         columns=self.rateTables[company]['LiabilityDeductibleFactor'][0])
            .query("LiabilityDeductible in @pd_deductibles")
            .replace({'LiabilityDeductible': pd_deductible_map})
            .rename(columns={'LiabilityDeductible': 'Deductible Amount', 'Factor': 'Property Damage Per Accident'})
            .sort_values('Deductible Amount')
        )

        Rule98Table1c = (
            pd.DataFrame(self.rateTables[company]['ZoneRatedLiabilityDeductibleFactor'][1:],
                         columns=self.rateTables[company]['ZoneRatedLiabilityDeductibleFactor'][0])
            .query("LiabilityDeductible in @csl_deductibles")
            .replace({'LiabilityDeductible': csl_deductible_map})
            .rename(columns={'LiabilityDeductible': 'Deductible Amount', 'Factor': 'Combined Single Limit'})
            .sort_values('Deductible Amount')
        )

        Rule98Table1d = (
            pd.DataFrame(self.rateTables[company]['ZoneRatedLiabilityDeductibleFactor'][1:],
                         columns=self.rateTables[company]['ZoneRatedLiabilityDeductibleFactor'][0])
            .query("LiabilityDeductible in @pd_deductibles")
            .replace({'LiabilityDeductible': pd_deductible_map})
            .rename(columns={'LiabilityDeductible': 'Deductible Amount', 'Factor': 'Property Damage Per Accident'})
            .sort_values('Deductible Amount')
        )

        Rule98Table1bd = pd.merge(Rule98Table1b, Rule98Table1d, on='Deductible Amount', how="inner")
        Rule98Table1ac = pd.merge(Rule98Table1a, Rule98Table1c, on='Deductible Amount', how="inner").filter(items=['Combined Single Limit_x', 'Combined Single Limit_y'])
        Rule98Table1 = pd.concat([Rule98Table1bd, Rule98Table1ac], axis=1, join="inner")
        Rule98Table1 = Rule98Table1.rename(columns={'Deductible Amount' : 'Deductible','Property Damage Per Accident_x' : 'Other Than Zone-Rated', 'Property Damage Per Accident_y' : 'Zone-Rated', 'Combined Single Limit_x' : 'Other Than Zone-Rated', 'Combined Single Limit_y' : 'Zone-Rated'})

        Rule98Table1.iloc[:,1:] = 1 - Rule98Table1.iloc[:,1:]
        Rule98Table1 = Rule98Table1.astype(object)
        Rule98Table1.iloc[:,:1] = Rule98Table1.iloc[:,:1].astype(int).map(lambda x: "{:,}".format(x)) # Deductible Commas
        Rule98Table1 = Rule98Table1.astype(object)
        Rule98Table1.iloc[:, 1:] = Rule98Table1.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}") # 3 Place float factors
        Rule98Table1.fillna("NA", inplace = True)
        Rule98Table1.replace({"nan" : "NA"}, inplace=True)

        return Rule98Table1

    @log_exceptions
    def buildRule298PPT(self,company):
        # 298.B.2.a. Private Passenger Types Deductible Discount Factors
        """Deductibles on this rule are very inconsistent state by state. Additionally, full glass states get an extra column."""

        factor_table = pd.DataFrame(self.rateTables[company]['PhysicalDamageDeductibleFactors_Ext'][1:], index=None, columns=self.rateTables[company]['PhysicalDamageDeductibleFactors_Ext'][0])


        # Function to clean the Factor column
        # Super weird bug, Factor column reads in "=Round(0.XXXXXXXXXX,3)" instead of the rounded float. This is a function to fix that.
        def clean_factor(value):
            if isinstance(value, str) and value.startswith('=ROUND'):
                # Extract the number from the string
                match = re.search(r'\d+\.\d+', value)
                if match:
                    return float(match.group(0))
            try:
                return float(value)
            except ValueError:
                return None

            # Apply the cleaning function to the Factor column
        factor_table['Factor'] = factor_table['Factor'].apply(clean_factor)

        coll_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 Coll Map").fillna("Y")
        otc_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 OTC Map").fillna("Y")
        otc_fg_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name="298 OTC FG Map").fillna("Y") # Full glass

        factor_table = factor_table
        deductible_amount = factor_table[factor_table["VehicleAndCoverageType"] == "Private Passenger Types Collision"]["Deductible"]
        collision_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Private Passenger Types Collision"].drop(columns = "VehicleAndCoverageType")
        otc_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Private Passenger Types Comprehensive"].drop(columns = "VehicleAndCoverageType")
        otc_fg_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Private Passenger Types Comprehensive: Full Glass Coverage"].drop(columns = "VehicleAndCoverageType")

        # Filtering down the deductibles in each coverage based on their corresponding map.

        state_lookup_coll = self.StateAbb
        state_lookup_otc = self.StateAbb
        state_lookup_otc_fg = self.StateAbb
        if self.StateAbb not in coll_deductible_map.columns:
            state_lookup_coll = "All Other"
        if self.StateAbb not in otc_deductible_map.columns:
            state_lookup_otc = "All Other"
        if self.StateAbb not in otc_fg_deductible_map.columns:
            state_lookup_otc_fg = "All Other"

        coll_deduc_map = coll_deductible_map[coll_deductible_map[state_lookup_coll] != "N"]
        coll_deduc_map = pd.DataFrame(coll_deduc_map["Limit"])

        otc_deduc_map = otc_deductible_map[otc_deductible_map[state_lookup_otc] != "N"]
        otc_deduc_map = pd.DataFrame(otc_deduc_map["Limit"])

        otc_fg_deduc_map = otc_fg_deductible_map[otc_fg_deductible_map[state_lookup_otc_fg] != "N"]
        otc_fg_deduc_map = pd.DataFrame(otc_fg_deduc_map["Limit"])

        coll_deduc_map['Limit'] = coll_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_deduc_map['Limit'] = otc_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_fg_deduc_map['Limit'] = otc_fg_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.

        # Collision factors get inner joined -> drop unneeded new limits column ->  rename factor column -> repeat for remaining coverages.
        collision_factors = pd.merge(collision_factors, coll_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Collision"})
        otc_factors = pd.merge(otc_factors, otc_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "OTC"})
        otc_fg_factors = pd.merge(otc_fg_factors, otc_fg_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "OTC FG"})


        collision_factors['Collision'] = collision_factors['Collision'].astype(float).round(3) # Very weird things are happening...
        otc_factors['OTC'] = otc_factors['OTC'].astype(float).round(3)
        otc_fg_factors['OTC FG'] = otc_fg_factors['OTC FG'].astype(float).round(3)

        # Now left join those coverages onto the original deductible list
        output_table = pd.DataFrame(deductible_amount)

        if collision_factors.shape[0] > 0:
            output_table = output_table.merge(collision_factors, on='Deductible', how='left')
        if otc_factors.shape[0] > 0:
            output_table = output_table.merge(otc_factors, on='Deductible', how='left')
        if otc_fg_factors.shape[0] > 0:
            output_table = output_table.merge(otc_fg_factors, on='Deductible', how='left')

        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: int(x.replace(',', '')))
        output_table.sort_values(by='Deductible', inplace=True)
        output_table.dropna(subset=output_table.columns[2:], how='all', inplace=True)
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        # Put the commas back in the Deductible column
        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: "{:,}".format(x))
        output_table = output_table.rename(columns = {"OTC" : "Comprehensive All Perils Deductible", "OTC FG" : "Comprehensive All Perils Deductible With Full Glass Coverage"})
        output_table.fillna("NA", inplace = True)
        output_table.replace({"nan" : "NA"}, inplace=True)

        # Removing comprehensive all perils in a mandatory FG state:
        fg_state_list = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 FG States")

        if self.StateAbb in fg_state_list.values:
            output_table = output_table.drop(columns = ["Comprehensive All Perils Deductible"])

        return output_table

    @log_exceptions
    def buildRule298TTT(self,company):
        # 298.B.2.b Trucks, Tractors And Trailers And All Autos Except Zone-rated Risks Deductible Discount Factors
        """Deductibles on this rule are very inconsistent state by state. Additionally, full glass states get an extra column."""

        factor_table = pd.DataFrame(self.rateTables[company]['PhysicalDamageDeductibleFactors_Ext'][1:], index=None, columns=self.rateTables[company]['PhysicalDamageDeductibleFactors_Ext'][0])


        # Function to clean the Factor column
        # Super weird bug, Factor column reads in "=Round(0.XXXXXXXXXX,3)" instead of the rounded float. This is a function to fix that.
        def clean_factor(value):
            if isinstance(value, str) and value.startswith('=ROUND'):
                # Extract the number from the string
                match = re.search(r'\d+\.\d+', value)
                if match:
                    return float(match.group(0))
            try:
                return float(value)
            except ValueError:
                return None

            # Apply the cleaning function to the Factor column
        factor_table['Factor'] = factor_table['Factor'].apply(clean_factor)

        coll_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 Coll Map").fillna("Y")
        otc_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 OTC Map").fillna("Y")
        otc_fg_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name="298 OTC FG Map").fillna("Y") # Full glass

        factor_table = factor_table
        deductible_amount = factor_table[factor_table["VehicleAndCoverageType"] == "Private Passenger Types Collision"]["Deductible"]
        trucks_collision_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Trucks And Truck-tractors Collision"].drop(columns = "VehicleAndCoverageType")
        trailer_collision_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Trailer Types Collision"].drop(columns = "VehicleAndCoverageType")
        otc_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Trucks, Tractors and Trailers Comprehensive"].drop(columns = "VehicleAndCoverageType")
        otc_fg_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Trucks, Tractors and Trailers Comprehensive: Full Glass Coverage"].drop(columns = "VehicleAndCoverageType")
        specified_all_perils_factors = factor_table[factor_table["VehicleAndCoverageType"] == "Trucks, Tractors and Trailers Specified Perils"].drop(columns = "VehicleAndCoverageType")
        # Filtering down the deductibles in each coverage based on their corresponding map.

        state_lookup_coll = self.StateAbb
        state_lookup_otc = self.StateAbb
        state_lookup_otc_fg = self.StateAbb
        if self.StateAbb not in coll_deductible_map.columns:
            state_lookup_coll = "All Other"
        if self.StateAbb not in otc_deductible_map.columns:
            state_lookup_otc = "All Other"
        if self.StateAbb not in otc_fg_deductible_map.columns:
            state_lookup_otc_fg = "All Other"

        coll_deduc_map = coll_deductible_map[coll_deductible_map[state_lookup_coll] != "N"]
        coll_deduc_map = pd.DataFrame(coll_deduc_map["Limit"])

        otc_deduc_map = otc_deductible_map[otc_deductible_map[state_lookup_otc] != "N"]
        otc_deduc_map = pd.DataFrame(otc_deduc_map["Limit"])

        otc_fg_deduc_map = otc_fg_deductible_map[otc_fg_deductible_map[state_lookup_otc_fg] != "N"]
        otc_fg_deduc_map = pd.DataFrame(otc_fg_deduc_map["Limit"])

        coll_deduc_map['Limit'] = coll_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_deduc_map['Limit'] = otc_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_fg_deduc_map['Limit'] = otc_fg_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.

        # Collision factors get inner joined -> drop unneeded new limits column ->  rename factor column -> repeat for remaining coverages.
        trucks_collision_factors = pd.merge(trucks_collision_factors, coll_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Truck-Tractor Collision"})
        trailers_collision_factors = pd.merge(trailer_collision_factors, coll_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Trailer Collision"})
        otc_factors = pd.merge(otc_factors, otc_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Comp All Perils"})
        otc_fg_factors = pd.merge(otc_fg_factors, otc_fg_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Comp All Perils FG"})
        spec_all_peril_factors = pd.merge(specified_all_perils_factors, otc_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Spec All Perils"})


        trucks_collision_factors['Truck-Tractor Collision'] = trucks_collision_factors['Truck-Tractor Collision'].astype(float).round(3)
        trailers_collision_factors['Trailer Collision'] = trailers_collision_factors['Trailer Collision'].astype(float).round(3)
        otc_factors['Comp All Perils'] = otc_factors['Comp All Perils'].astype(float).round(3)
        otc_fg_factors['Comp All Perils FG'] = otc_fg_factors['Comp All Perils FG'].astype(float).round(3)
        spec_all_peril_factors['Spec All Perils'] = spec_all_peril_factors['Spec All Perils'].astype(float).round(3)

        # Now left join those coverages onto the original deductible list
        output_table = pd.DataFrame(deductible_amount)

        if trucks_collision_factors.shape[0] > 0:
            output_table = output_table.merge(trucks_collision_factors, on='Deductible', how='left')
        if trailers_collision_factors.shape[0] > 0:
            output_table = output_table.merge(trailers_collision_factors, on='Deductible', how='left')
        if otc_factors.shape[0] > 0:
            output_table = output_table.merge(otc_factors, on='Deductible', how='left')
        if otc_fg_factors.shape[0] > 0:
            output_table = output_table.merge(otc_fg_factors, on='Deductible', how='left')
        if spec_all_peril_factors.shape[0] > 0:
            output_table = output_table.merge(spec_all_peril_factors, on='Deductible', how='left')

        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: int(x.replace(',', '')))
        output_table.sort_values(by='Deductible', inplace=True)
        output_table.dropna(subset=output_table.columns[2:], how='all', inplace=True)
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        # Put the commas back in the Deductible column
        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: "{:,}".format(x))
        output_table = output_table.rename(columns = {"Truck-Tractor Collision" : "Trucks And Truck-tractors Collision", "Trailer Collision" : "Trailer Types Collision",
                             "Comp All Perils":"Comprehensive All Perils Deductible","Comp All Perils FG":"Comprehensive All Perils Deductible With Full Glass Coverage",
                             "Spec All Perils":"Specified Causes Of Loss All Perils Deductible"})
        output_table.fillna("NA", inplace = True)
        output_table.replace({"nan" : "NA"}, inplace=True)

        # Removing comprehensive all perils in a mandatory FG state:
        fg_state_list = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 FG States")

        if self.StateAbb in fg_state_list.values:
            output_table = output_table.drop(columns = ["Comprehensive All Perils Deductible"])

        return output_table

    @log_exceptions
    def buildRule298Zone(self,company):
        # 298.B.3. Zone-rated Vehicles Deductible Discount Factors
        """Deductibles on this rule are very inconsistent state by state. Additionally, full glass states get an extra column."""

        factor_table = pd.DataFrame(self.rateTables[company]['ZoneRatedVehiclesDeductibleDiscountFactors_Ext'][1:], index=None, columns=self.rateTables[company]['ZoneRatedVehiclesDeductibleDiscountFactors_Ext'][0])

        # Function to clean the Factor column
        # Super weird bug, Factor column reads in "=Round(0.XXXXXXXXXX,3)" instead of the rounded float. This is a function to fix that.
        def clean_factor(value):
            if isinstance(value, str) and value.startswith('=ROUND'):
                # Extract the number from the string
                match = re.search(r'\d+\.\d+', value)
                if match:
                    return float(match.group(0))
            try:
                return float(value)
            except ValueError:
                return None
            # Apply the cleaning function to the Factor column
        factor_table['Factor'] = factor_table['Factor'].apply(clean_factor)

        coll_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 Coll Map").fillna("Y")
        otc_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 OTC Map").fillna("Y")
        otc_fg_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name="298 OTC FG Map").fillna("Y") # Full glass

        factor_table = factor_table
        deductible_amount = factor_table[factor_table["Vehicle And Coverage Type"] == "Trucks And Truck-tractors Collision"]["Deductible Amount"]
        trucks_collision_factors = factor_table[factor_table["Vehicle And Coverage Type"] == "Trucks And Truck-tractors Collision"].drop(columns = "Vehicle And Coverage Type")
        trailer_collision_factors = factor_table[factor_table["Vehicle And Coverage Type"] == "Trailer Types Collision"].drop(columns = "Vehicle And Coverage Type")
        otc_factors = factor_table[factor_table["Vehicle And Coverage Type"] == "Trucks, Tractors and Trailers Comprehensive"].drop(columns = "Vehicle And Coverage Type")
        otc_fg_factors = factor_table[factor_table["Vehicle And Coverage Type"] == "Trucks, Tractors and Trailers Comprehensive: Full Glass Coverage"].drop(columns = "Vehicle And Coverage Type")
        specified_all_perils_factors = factor_table[factor_table["Vehicle And Coverage Type"] == "Trucks, Tractors and Trailers Specified Perils"].drop(columns = "Vehicle And Coverage Type")
        # Filtering down the deductibles in each coverage based on their corresponding map.

        state_lookup_coll = self.StateAbb
        state_lookup_otc = self.StateAbb
        state_lookup_otc_fg = self.StateAbb
        if self.StateAbb not in coll_deductible_map.columns:
            state_lookup_coll = "All Other"
        if self.StateAbb not in otc_deductible_map.columns:
            state_lookup_otc = "All Other"
        if self.StateAbb not in otc_fg_deductible_map.columns:
            state_lookup_otc_fg = "All Other"

        coll_deduc_map = coll_deductible_map[coll_deductible_map[state_lookup_coll] != "N"]
        coll_deduc_map = pd.DataFrame(coll_deduc_map["Limit"])

        otc_deduc_map = otc_deductible_map[otc_deductible_map[state_lookup_otc] != "N"]
        otc_deduc_map = pd.DataFrame(otc_deduc_map["Limit"])

        otc_fg_deduc_map = otc_fg_deductible_map[otc_fg_deductible_map[state_lookup_otc_fg] != "N"]
        otc_fg_deduc_map = pd.DataFrame(otc_fg_deduc_map["Limit"])

        coll_deduc_map['Limit'] = coll_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_deduc_map['Limit'] = otc_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_fg_deduc_map['Limit'] = otc_fg_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.


        # Collision factors get inner joined -> drop unneeded new limits column ->  rename factor column -> repeat for remaining coverages.
        trucks_collision_factors = pd.merge(trucks_collision_factors, coll_deduc_map, left_on='Deductible Amount', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Truck-Tractor Collision"})
        trailers_collision_factors = pd.merge(trailer_collision_factors, coll_deduc_map, left_on='Deductible Amount', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Trailer Collision"})
        otc_factors = pd.merge(otc_factors, otc_deduc_map, left_on='Deductible Amount', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Comp All Perils"})
        otc_fg_factors = pd.merge(otc_fg_factors, otc_fg_deduc_map, left_on='Deductible Amount', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Comp All Perils FG"})
        spec_all_peril_factors = pd.merge(specified_all_perils_factors, otc_deduc_map, left_on='Deductible Amount', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"Factor" : "Spec All Perils"})

        trucks_collision_factors['Truck-Tractor Collision'] = trucks_collision_factors['Truck-Tractor Collision'].astype(float).round(3)
        trailers_collision_factors['Trailer Collision'] = trailers_collision_factors['Trailer Collision'].astype(float).round(3)
        otc_factors['Comp All Perils'] = otc_factors['Comp All Perils'].astype(float).round(3)
        otc_fg_factors['Comp All Perils FG'] = otc_fg_factors['Comp All Perils FG'].astype(float).round(3)
        spec_all_peril_factors['Spec All Perils'] = spec_all_peril_factors['Spec All Perils'].astype(float).round(3)

        # Now left join those coverages onto the original deductible list
        output_table = pd.DataFrame(deductible_amount)

        # Check if each DataFrame has exactly two columns before merging
        if trucks_collision_factors.shape[0] > 0:
            output_table = output_table.merge(trucks_collision_factors, on='Deductible Amount', how='left')
        if trailers_collision_factors.shape[0] > 0:
            output_table = output_table.merge(trailers_collision_factors, on='Deductible Amount', how='left')
        if otc_factors.shape[0] > 0:
            output_table = output_table.merge(otc_factors, on='Deductible Amount', how='left')
        if otc_fg_factors.shape[0] > 0:
            output_table = output_table.merge(otc_fg_factors, on='Deductible Amount', how='left')
        if spec_all_peril_factors.shape[0] > 0:
            output_table = output_table.merge(spec_all_peril_factors, on='Deductible Amount', how='left')

        output_table['Deductible Amount'] = output_table['Deductible Amount'].apply(lambda x: int(x.replace(',', '')))
        output_table.sort_values(by='Deductible Amount', inplace=True)
        output_table.rename(columns = {"Deductible Amount" : "Deductible"}, inplace = True)
        output_table.dropna(subset=output_table.columns[2:], how='all', inplace=True)
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        # Put the commas back in the Deductible column
        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: "{:,}".format(x))
        output_table = output_table.rename(columns = {"Truck-Tractor Collision" : "Trucks And Truck-tractors Collision", "Trailer Collision" : "Trailer Types Collision",
                             "Comp All Perils":"Comprehensive All Perils Deductible","Comp All Perils FG":"Comprehensive All Perils Deductible With Full Glass Coverage",
                             "Spec All Perils":"Specified Causes Of Loss All Perils Deductible"})
        output_table.fillna("NA", inplace = True)
        output_table.replace({"nan" : "NA"}, inplace=True)

        # Removing comprehensive all perils in a mandatory FG state:
        fg_state_list = pd.read_excel(BA_INPUT_FILE, sheet_name = "298 FG States")

        if self.StateAbb in fg_state_list.values:
            output_table = output_table.drop(columns = ["Comprehensive All Perils Deductible"])

        return output_table

    @log_exceptions
    def buildRule298AutoBlanket(self,company):
        # 298.B.4.a. Auto Dealers Blanket Collision Deductible Factors
        """Not a big fan of how this is setup but we are short on time."""

        factor_table = pd.DataFrame(self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][1:], index=None, columns=self.rateTables[company]['GarageDealersCollisionBlanketDeductibleFactor'][0])
        factor_table.rename(columns = {"SpecialTypesCollisionDeductible" : "Deductible"}, inplace = True)

        # Sort

        factor_table["Deductible"] = factor_table["Deductible"].apply(lambda x: int(x.replace(',', '')))
        factor_table.sort_values(by="Deductible", inplace=True)

        # Format back with commas and no decimals
        factor_table["Deductible"] = factor_table["Deductible"].apply(
            lambda x: f"{x:,.0f}" if pd.notna(x) else ""
        )

        factor_table = factor_table.astype(object)
        factor_table.iloc[:, 1:] = factor_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        factor_table = factor_table[factor_table["Deductible"] != "100"]


        return factor_table

    @log_exceptions
    def buildRule298AutoGarageOTCFactors_1(self, company):
        # 298.B.4.b. Auto Dealers and Garagekeepers Other Than Collision Deductible Factors (Theft, Mischief and Vandalism)
        factor_table = pd.DataFrame(self.rateTables[company]['GarageDealersOtherThanCollisionDeductibleFactor'][1:], index=None, columns=self.rateTables[company]['GarageDealersOtherThanCollisionDeductibleFactor'][0])

        data = {
            "Coverage": ["Fire Only", "Fire and Theft Only", "Limited Specified Causes of Loss","Specified Causes of Loss", "Comprehensive"],
        }

        factor_100_500 = factor_table[factor_table["SpecialTypesOtherThanCollisionDeductible"] == "100 per car / 500 per occurrence (Theft, Mischief or Vandalism)"].iloc[0,1]
        factor_250_500 = factor_table[factor_table["SpecialTypesOtherThanCollisionDeductible"] == "250 per car / 1000 per occurrence (Theft, Mischief or Vandalism)"].iloc[0,1]
        factor_500_1000 = factor_table[factor_table["SpecialTypesOtherThanCollisionDeductible"] == "500 per car / 2500 per occurrence (Theft, Mischief or Vandalism)"].iloc[0,1]
        factor_fire = factor_table[factor_table["SpecialTypesOtherThanCollisionDeductible"] == "Full Coverage (Fire Only)"].iloc[0,1]

        output_table = pd.DataFrame(data)
        output_table["$100/500"] = [factor_fire] + [factor_100_500] * 4
        output_table["$250/1,000"] = [factor_fire] + [factor_250_500] * 4
        output_table["$500/2,500"] = [factor_fire] + [factor_500_1000] * 4

        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        output_table.fillna("NA", inplace = True)

        return output_table

    @log_exceptions
    def buildRule298AutoGarageOTCFactors_2(self, company):
        # 298.B.4.b. Auto Dealers and Garagekeepers Other Than Collision Deductible Factors (All Perils)
        """This is just a factor being multiplied to the theft/vandalism table above."""

        output_table = self.buildRule298AutoGarageOTCFactors_1(company)

        factor = pd.DataFrame(self.rateTables[company]['GarageDealersOtherThanCollisionAllPerilsDeductibleFactor'][1:], index=None, columns=self.rateTables[company]['GarageDealersOtherThanCollisionAllPerilsDeductibleFactor'][0]).iloc[0,1]

        for col in output_table.columns[1:]:
            output_table[col] = pd.to_numeric(output_table[col], errors='coerce')

        output_table.iloc[:, 1:] = output_table.iloc[:, 1:] * factor
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        output_table.fillna("NA", inplace = True)

        return output_table


    # Builds the Driver Based Rating table
    # Returns a dataframe
    @log_exceptions
    def buildDBRFactors(self, company):
        DBRLiabFactors = pd.DataFrame(self.rateTables[company]['DriverBasedRatingLiabilityFactor_Ext'][1:], index=None, columns=self.rateTables[company]['DriverBasedRatingLiabilityFactor_Ext'][0]).rename(columns={'Factor' : 'Liability'})
        DBRCollFactors = pd.DataFrame(self.rateTables[company]['DriverBasedRatingCollisionFactor_Ext'][1:], index=None, columns=self.rateTables[company]['DriverBasedRatingCollisionFactor_Ext'][0]).rename(columns={'Factor' : 'Collision'})
        DBRFactors = pd.merge(DBRLiabFactors, DBRCollFactors, on= 'Band', how = 'left')

        return DBRFactors

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildRetentionFactors(self, company):
        RetentionFactors = pd.DataFrame(self.rateTables[company]['RetentionGradeFactor_Ext'][1:], index=None, columns=self.rateTables[company]['RetentionGradeFactor_Ext'][0])
        grades_to_keep = ["A","B","C","D","E","F"]
        RetentionFactors = RetentionFactors.rename(columns={'RetentionGrade' : 'Grade', 'RetentionFactor' : 'Factor'})
        RetentionFactors = RetentionFactors[RetentionFactors["Grade"].isin(grades_to_keep)]
        return RetentionFactors.rename(columns={'RetentionGrade' : 'Grade', 'RetentionFactor' : 'Factor'})

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildFuneralDirectors1(self, company):
        FuneralDirectors1 = pd.DataFrame(self.rateTables[company]['SpecialTypesFuneralDirectorFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesFuneralDirectorFactor'][0])
        #FuneralDirectors1 = self.buildDataFrame("SpecialTypesFuneralDirectorFactor")
        FuneralDirectors1 = FuneralDirectors1.query(f'SpecialTypesClassCode == "7915" & TypeCoverage != "PhysDamOTCACV" & TypeCoverage != "PhysDamCollACV"').replace({'TypeCoverage' : {'MedicalPayments' : 'Med', 'No-Fault' : 'PIP', 'PhysDamOTCSA' : 'Other than Collision', 'PhysDamCollSA' : 'Collision', 'UM/UIM' : 'All Other'}})
        FuneralDirectors1 = FuneralDirectors1.rename(columns={'TypeCoverage': 'Coverage'}).filter(items=['Coverage', 'Factor'])

        if self.StateAbb not in self.pip_states.values:
            FuneralDirectors1 = FuneralDirectors1[FuneralDirectors1["Coverage"] != "PIP"]
        if self.StateAbb in self.no_med_states:
            FuneralDirectors1 = FuneralDirectors1[FuneralDirectors1["Coverage"] != "Med"]

        return FuneralDirectors1

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildFuneralDirectors2(self, company):
        FuneralDirectors2 = pd.DataFrame(self.rateTables[company]['SpecialTypesFuneralDirectorFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesFuneralDirectorFactor'][0])
        #FuneralDirectors2 = self.buildDataFrame("SpecialTypesFuneralDirectorFactor")
        FuneralDirectors2 = FuneralDirectors2.query(f'SpecialTypesClassCode == "7922" & TypeCoverage != "PhysDamOTCACV" & TypeCoverage != "PhysDamCollACV"').replace({'TypeCoverage' : {'MedicalPayments' : 'Med', 'No-Fault' : 'PIP', 'PhysDamOTCSA' : 'Other than Collision', 'PhysDamCollSA' : 'Collision', 'UM/UIM' : 'All Other'}})

        FuneralDirectors2 = FuneralDirectors2.rename(columns={'TypeCoverage' : 'Coverage'}).filter(items=['Coverage', 'Factor'])

        if self.StateAbb not in self.pip_states.values:
            FuneralDirectors2 = FuneralDirectors2[FuneralDirectors2["Coverage"] != "PIP"]
        if self.StateAbb in self.no_med_states:
            FuneralDirectors2 = FuneralDirectors2[FuneralDirectors2["Coverage"] != "Med"]

        return FuneralDirectors2

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildFuneralDirectors3(self, company) :
        FuneralDirectors3 = pd.DataFrame(self.rateTables[company]['FuneralDirectorMedicalPaymentsHiredNonOwnedFactor'][1:], index=None, columns=self.rateTables[company]['FuneralDirectorMedicalPaymentsHiredNonOwnedFactor'][0])
        #FuneralDirectors3 = self.buildDataFrame("FuneralDirectorMedicalPaymentsHiredNonOwnedFactor")
        return FuneralDirectors3.filter(items=['Factor'])

    # Builds the Special Types Golf Carts and Low Speed Vehicles table
    # Returns a dataframe
    @log_exceptions
    def buildSpecialGolfandLow(self, company):
        SpecialGolfandLow = pd.DataFrame(self.rateTables[company]['SpecialTypesGolfCartsAndLowSpeedVehiclesFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesGolfCartsAndLowSpeedVehiclesFactor'][0]).query(f'TypeCoverage == "Liability" | TypeCoverage == "MedicalPayments" | TypeCoverage == "UM/UIM" | TypeCoverage == "PhysDamOTCACV" | TypeCoverage == "PhysDamCollACV"')
        #SpecialGolfandLow = self.buildDataFrame("SpecialTypesGolfCartsAndLowSpeedVehiclesFactor").query(f'TypeCoverage == "Liability" | TypeCoverage == "MedicalPayments" | TypeCoverage == "UM/UIM" | TypeCoverage == "PhysDamOTCACV" | TypeCoverage == "PhysDamCollACV"')
        PivotedSpecialGolfandLow = SpecialGolfandLow.pivot(index='TypeCoverage', columns='SpecialTypesClassCode', values='Factor').reset_index('TypeCoverage')
        return PivotedSpecialGolfandLow.rename(columns={'TypeCoverage' : 'Coverage', '9461' : 'Golf Carts Used On Golf Courses', '9462' : 'Golf Carts Other Commercial Purposes', '9463' : 'Low Speed Vehicle'}). \
            replace({'Coverage': {'MedicalPayments': "Med", 'UM/UIM': "Uninsured Motorist (UM or UIM)", 'PhysDamOTCACV': "Other than Collision", 'PhysDamCollACV': "Collision"}})

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMotorcycles1(self, company):
        Motorcycles1 = pd.DataFrame(self.rateTables[company]['SpecialTypesMotorcycleLiabilityFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesMotorcycleLiabilityFactor'][0])
        #Motorcycles1 = self.buildDataFrame("SpecialTypesMotorcycleLiabilityFactor")
        Motorcycles1 = Motorcycles1.rename(columns={'EngineSize' : 'Engine Size (cc)'})
        Motorcycles1 = Motorcycles1.astype(object)
        Motorcycles1.iloc[:, 1:] = Motorcycles1.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Motorcycles1

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMotorcycles2(self, company):
        Motorcycles2 = pd.DataFrame(self.rateTables[company]['SpecialTypesMotorcycleFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesMotorcycleFactor'][0])
        #Motorcycles2 = self.buildDataFrame("SpecialTypesMotorcycleFactor")
        Motorcycles2 = Motorcycles2.query(f'TypeCoverage == "UM/UIM"').rename(columns={'TypeCoverage' : 'Coverage'}).replace({'Coverage' : {'UM/UIM' : 'UM'}})
        Motorcycles2 = Motorcycles2.astype(object)
        Motorcycles2.iloc[:, 1:] = Motorcycles2.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Motorcycles2

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMotorcycles3(self, company):
        Motorcycles3 = pd.DataFrame(self.rateTables[company]['SpecialTypesMotorcycleFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesMotorcycleFactor'][0])
        #Motorcycles3 = self.buildDataFrame("SpecialTypesMotorcycleFactor")
        Motorcycles3 = Motorcycles3.query(f'TypeCoverage == "MedicalPayments"').rename(columns={'TypeCoverage' : 'Coverage'}).replace({'Coverage' : {'MedicalPayments' : 'Med'}})
        Motorcycles3 = Motorcycles3.astype(object)
        Motorcycles3.iloc[:, 1:] = Motorcycles3.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Motorcycles3

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMotorcycles4(self, company):
        Motorcycles4 = pd.DataFrame(self.rateTables[company]['SpecialTypesMotorcycleFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesMotorcycleFactor'][0])
        #Motorcycles4 = self.buildDataFrame("SpecialTypesMotorcycleFactor")
        Motorcycles4 = Motorcycles4.query(f'TypeCoverage == "PhysDamOTCACV" | TypeCoverage == "PhysDamCollACV"').rename(columns={'TypeCoverage' : 'Coverage'}).replace({'Coverage' : {'PhysDamOTCACV' : 'Comp', 'PhysDamCollACV' : 'Coll'}})
        Motorcycles4 = Motorcycles4.astype(object)
        Motorcycles4.iloc[:, 1:] = Motorcycles4.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Motorcycles4

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMotorcycles5(self, company):
        Motorcycles5 = pd.DataFrame(self.rateTables[company]['SpecialTypesMotorcycleFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesMotorcycleFactor'][0])
        #Motorcycles5 = self.buildDataFrame("SpecialTypesMotorcycleFactor")
        Motorcycles5 = Motorcycles5.query(f'TypeCoverage == "PhysDamOTCSA" | TypeCoverage == "PhysDamCollSA"').rename(columns={'TypeCoverage' : 'Coverage'}).replace({'Coverage' : {'PhysDamOTCSA' : 'Comp', 'PhysDamCollSA' : 'Coll'}})
        Motorcycles5 = Motorcycles5.astype(object)
        Motorcycles5.iloc[:, 1:] = Motorcycles5.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Motorcycles5

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMedPayments92A(self, company):

        # function to get factors
        def process_med_payments(df, territory, group=None):
            # Find the Limits
            med_limit_col = next((col for col in df.columns if 'MedicalPaymentsLimitText' in col), None)
            if med_limit_col is None:
                raise ValueError("No column containing 'MedicalPaymentsLimitText' found.")

            # useful debug statements
            # print(f"\n[DEBUG] Identified medical limit column: {med_limit_col}")
            # print(f"[DEBUG] Filtering for Territory: {territory}, Group: {group}")
            # print("[DEBUG] Available Territories:", df['Territory'].unique())

            query_str = f'Territory == @territory'
            if group:
                query_str += f' & PublicTransportationGroup == "{group}"'

            # Name the Columns
            df = df.query(query_str).rename(
                columns={med_limit_col: 'Limit', 'BasePremium': 'Premium'}).filter(
                items=['Limit', 'Premium'])

            # Get the Factors
            df = df.query('Limit != "No Coverage" & Limit != "500"')
            factor = df[df["Limit"] == "5,000"].iloc[0, 1] if not df.empty else 1
            df['Premium'] = df['Premium'] / factor

            return df

        # Function to format the data frames
        def process_sheet(sheet_data, territory, sheet_label, group=None):
            df = pd.DataFrame(sheet_data[1:], columns=sheet_data[0])
            indicator_col = next((col for col in df.columns if 'NoFault' in col), None)

            # No Fault Detection, if Medpay has a No Fault indicator then make columns for each unique indication value
            if indicator_col:
                unique_indicators = df[indicator_col].dropna().unique()
                processed_dfs = []

                # call process med pay function for each data frame from TTT. PPT, and Public for each No fault indicator
                for indicator in unique_indicators:
                    filtered_df = df[df[indicator_col] == indicator]
                    processed_df = process_med_payments(filtered_df, territory, group)

                    # Only add the indicator to the column name if more than one indicator is present
                    col_name = f"{sheet_label}"
                    if len(unique_indicators) > 1:
                        col_name = col_name + f" - {indicator}"

                    processed_df.rename(columns={'Premium': col_name}, inplace=True)
                    processed_dfs.append(processed_df)

                # combine the different indicator types for the same group
                return reduce(lambda left, right: pd.merge(left, right, on='Limit', how='outer'), processed_dfs)

            # Case where No Fault indicator is not present. Just processes the sheets.
            else:
                df = process_med_payments(df, territory, group)
                df.rename(columns={'Premium': sheet_label}, inplace=True)
                return df

        # Load sheet name mappings
        School_239 = pd.read_excel(BA_INPUT_FILE, sheet_name="239 School Buses")
        Other_239 = pd.read_excel(BA_INPUT_FILE, sheet_name="239 Other Buses")
        Van_239 = pd.read_excel(BA_INPUT_FILE, sheet_name="239 Van Pools")
        Taxi_239 = pd.read_excel(BA_INPUT_FILE, sheet_name="239 Taxis")
        TTT_222 = pd.read_excel(BA_INPUT_FILE, sheet_name="222 TTT")
        PPT_232 = pd.read_excel(BA_INPUT_FILE, sheet_name="232 PPT")

        # Function to fetch sheet name from Input File
        def get_sheetname(df, state, coverage):
            if not df[(df['state'] == state) & (df['coverage'] == coverage)].empty:
                return df.loc[(df['state'] == state) & (df['coverage'] == coverage), 'sheet'].values[0]
            return df.loc[(df['state'] == 'Default') & (df['coverage'] == coverage), 'sheet'].values[0]

        # Use of sheetname function
        school_sheetname = get_sheetname(School_239, self.StateAbb, 'med')
        other_sheetname = get_sheetname(Other_239, self.StateAbb, 'med')
        van_sheetname = get_sheetname(Van_239, self.StateAbb, 'med')
        taxi_sheetname = get_sheetname(Taxi_239, self.StateAbb, 'med')
        ttt_sheetname = get_sheetname(TTT_222, self.StateAbb, 'med')
        ppt_sheetname = get_sheetname(PPT_232, self.StateAbb, 'med')

        Territory92a = pd.DataFrame(self.rateTables[company][ttt_sheetname][1:],
                                    columns=self.rateTables[company][ttt_sheetname][0])
        territory_col = 'Territory'
        Territory92a_value = Territory92a[territory_col].dropna().iloc[0]

        # Process all sheets
        MedPaymentsTTT = process_sheet(self.rateTables[company][ttt_sheetname], Territory92a_value, "Trucks, Tractors, Trailers")
        MedPaymentsSchool = process_sheet(self.rateTables[company][school_sheetname], Territory92a_value,
                                          "School And Church Buses", "School And Church Buses")
        MedPaymentsOther = process_sheet(self.rateTables[company][other_sheetname], Territory92a_value, "Other Buses",
                                         "Other Buses")
        MedPaymentsVan = process_sheet(self.rateTables[company][van_sheetname], Territory92a_value, "Van Pools",
                                       "Van Pools")
        MedPaymentsTaxi = process_sheet(self.rateTables[company][taxi_sheetname], Territory92a_value,
                                        "Taxicabs And Limousines", "Taxicabs And Limousines")
        MedPaymentsPPT = process_sheet(self.rateTables[company][ppt_sheetname], Territory92a_value,
                                       "Private Passenger Types")

        # Merge all DataFrames on 'Limit'
        MedDfs = [MedPaymentsTTT, MedPaymentsSchool, MedPaymentsOther, MedPaymentsVan, MedPaymentsTaxi, MedPaymentsPPT]

        result = MedDfs[0]
        for i, df in enumerate(MedDfs[1:], start=1):
            result = pd.merge(result, df, on='Limit', how='inner', suffixes=(None, f'_df{i}'))

        # Some table formatting
        result['Limit'] = result['Limit'].str.replace(',', '').astype(int)
        result = result.sort_values(by='Limit')
        result['Limit'] = result['Limit'].apply(lambda x: f"{x:,}")
        result = result.astype(object)
        result.iloc[:, 1:] = result.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return result

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildMedPayments92B(self, company):
        MedPaymentsZone = pd.DataFrame(self.rateTables[company]['ZoneRatedMedicalPaymentsTextFactor'][1:], index=None, columns=self.rateTables[company]['ZoneRatedMedicalPaymentsTextFactor'][0])
        #MedPaymentsZone = self.buildDataFrame("ZoneRatedMedicalPaymentsTextFactor")
        MedPaymentsZone = MedPaymentsZone.rename(columns={'MedicalPaymentsLimitText' : 'Limit', 'Factor' : 'Zone-Rated'}).filter(items=['Limit', 'Zone-Rated'])
        MedPaymentsZone = MedPaymentsZone.query(f'Limit != "No Coverage" & Limit != "500"').sort_values('Zone-Rated')
        MedPaymentsZoneFactor = MedPaymentsZone[MedPaymentsZone["Limit"] == "5,000"].iloc[0,1]
        MedPaymentsZone['Zone-Rated'] = MedPaymentsZone['Zone-Rated'] / MedPaymentsZoneFactor
        MedPaymentsZone = MedPaymentsZone.astype(object)
        MedPaymentsZone.iloc[:, 1:] = MedPaymentsZone.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return MedPaymentsZone

    # Builds RULE 101.A.4.a. TRUCKS AND TRUCK TRACTORS TYPE COLLISION VEHICLE AGE AND PRICE BRACKET FACTOR
    # Returns a dataframe
    @log_exceptions
    def build101A11(self, company):
        Table101PublicOTC = pd.DataFrame(
            self.rateTables[company]['PublicTransportationOtherThanCollisionCostNewRelativity'][1:], index=None,
            columns=self.rateTables[company]['PublicTransportationOtherThanCollisionCostNewRelativity'][0])
        # Table101PublicOTC = self.buildDataFrame("PublicTransportationOtherThanCollisionCostNewRelativity")
        Table101PublicOTC = Table101PublicOTC.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Other Than Collision'})
        Table101PublicOTC = Table101PublicOTC.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                             6001: "6,001 to 8,000",
                                                                             8001: "8,001 to 10,000",
                                                                             10001: "10,001 to 15,000", \
                                                                             15001: "15,001 to 20,000",
                                                                             20001: "20,001 to 25,000",
                                                                             25001: "25,001 to 40,000",
                                                                             40001: "40,001 to 65,000",
                                                                             65001: "65,001 to 90,000"}})

        Table101PublicColl = pd.DataFrame(
            self.rateTables[company]['PublicTransportationCollisionCostNewRelativity'][1:], index=None,
            columns=self.rateTables[company]['PublicTransportationCollisionCostNewRelativity'][0])
        # Table101PublicColl = self.buildDataFrame("PublicTransportationCollisionCostNewRelativity")
        Table101PublicColl = Table101PublicColl.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Collision'})
        Table101PublicColl = Table101PublicColl.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                               6001: "6,001 to 8,000",
                                                                               8001: "8,001 to 10,000",
                                                                               10001: "10,001 to 15,000", \
                                                                               15001: "15,001 to 20,000",
                                                                               20001: "20,001 to 25,000",
                                                                               25001: "25,001 to 40,000",
                                                                               40001: "40,001 to 65,000",
                                                                               65001: "65,001 to 90,000"}})

        Table101ZoneOTC = pd.DataFrame(self.rateTables[company]['ZoneRatedOtherThanCollisionCostNewRelativity'][1:],
                                       index=None,
                                       columns=self.rateTables[company]['ZoneRatedOtherThanCollisionCostNewRelativity'][
                                           0])
        Table101ZoneOTC = self.buildDataFrame("ZoneRatedOtherThanCollisionCostNewRelativity")
        Table101ZoneOTC = Table101ZoneOTC.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Other Than Collision'})
        Table101ZoneOTC = Table101ZoneOTC.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                         6001: "6,001 to 8,000",
                                                                         8001: "8,001 to 10,000",
                                                                         10001: "10,001 to 15,000", \
                                                                         15001: "15,001 to 20,000",
                                                                         20001: "20,001 to 25,000",
                                                                         25001: "25,001 to 40,000",
                                                                         40001: "40,001 to 65,000",
                                                                         65001: "65,001 to 90,000"}})

        Table101ZoneColl = pd.DataFrame(self.rateTables[company]['ZoneRatedCollisionCostNewRelativity'][1:], index=None,
                                        columns=self.rateTables[company]['ZoneRatedCollisionCostNewRelativity'][0])
        # Table101ZoneColl = self.buildDataFrame("ZoneRatedCollisionCostNewRelativity")
        Table101ZoneColl = Table101ZoneColl.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Collision'})
        Table101ZoneColl = Table101ZoneColl.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                           6001: "6,001 to 8,000",
                                                                           8001: "8,001 to 10,000",
                                                                           10001: "10,001 to 15,000", \
                                                                           15001: "15,001 to 20,000",
                                                                           20001: "20,001 to 25,000",
                                                                           25001: "25,001 to 40,000",
                                                                           40001: "40,001 to 65,000",
                                                                           65001: "65,001 to 90,000"}})

        Table101TTTOTC = pd.DataFrame(self.rateTables[company]['TruckOtherThanCollisionCostNewRelativity'][1:],
                                      index=None,
                                      columns=self.rateTables[company]['TruckOtherThanCollisionCostNewRelativity'][0])
        # Table101TTTOTC = self.buildDataFrame("TruckOtherThanCollisionCostNewRelativity")
        Table101TTTOTC = Table101TTTOTC.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Other Than Collision'})
        Table101TTTOTC = Table101TTTOTC.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                       6001: "6,001 to 8,000", 8001: "8,001 to 10,000",
                                                                       10001: "10,001 to 15,000", \
                                                                       15001: "15,001 to 20,000",
                                                                       20001: "20,001 to 25,000",
                                                                       25001: "25,001 to 40,000",
                                                                       40001: "40,001 to 65,000",
                                                                       65001: "65,001 to 90,000"}})

        Table101TTTColl = pd.DataFrame(self.rateTables[company]['TruckCollisionCostNewRelativity'][1:], index=None,
                                       columns=self.rateTables[company]['TruckCollisionCostNewRelativity'][0])
        # Table101TTTColl = self.buildDataFrame("TruckCollisionCostNewRelativity")
        Table101TTTColl = Table101TTTColl.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Collision'})
        Table101TTTColl = Table101TTTColl.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                         6001: "6,001 to 8,000",
                                                                         8001: "8,001 to 10,000",
                                                                         10001: "10,001 to 15,000", \
                                                                         15001: "15,001 to 20,000",
                                                                         20001: "20,001 to 25,000",
                                                                         25001: "25,001 to 40,000",
                                                                         40001: "40,001 to 65,000",
                                                                         65001: "65,001 to 90,000"}})

        Table101PPTOTC = pd.DataFrame(
            self.rateTables[company]['PrivatePassengerOtherThanCollisionCostNewRelativity'][1:], index=None,
            columns=self.rateTables[company]['PrivatePassengerOtherThanCollisionCostNewRelativity'][0])
        # Table101PPTOTC = self.buildDataFrame("PrivatePassengerOtherThanCollisionCostNewRelativity")
        Table101PPTOTC = Table101PPTOTC.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Other Than Collision'})
        Table101PPTOTC = Table101PPTOTC.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                       6001: "6,001 to 8,000", 8001: "8,001 to 10,000",
                                                                       10001: "10,001 to 15,000", \
                                                                       15001: "15,001 to 20,000",
                                                                       20001: "20,001 to 25,000",
                                                                       25001: "25,001 to 40,000",
                                                                       40001: "40,001 to 65,000",
                                                                       65001: "65,001 to 90,000"}})

        Table101PPTColl = pd.DataFrame(self.rateTables[company]['PrivatePassengerCollisionCostNewRelativity'][1:],
                                       index=None,
                                       columns=self.rateTables[company]['PrivatePassengerCollisionCostNewRelativity'][
                                           0])
        # Table101PPTColl = self.buildDataFrame("PrivatePassengerCollisionCostNewRelativity")
        Table101PPTColl = Table101PPTColl.rename(
            columns={'OriginalCostNew': 'Original Cost New', 'Factor': 'Collision'})
        Table101PPTColl = Table101PPTColl.replace({'Original Cost New': {0: "0 to 4,500", 4501: "4,501 to 6,000",
                                                                         6001: "6,001 to 8,000",
                                                                         8001: "8,001 to 10,000",
                                                                         10001: "10,001 to 15,000", \
                                                                         15001: "15,001 to 20,000",
                                                                         20001: "20,001 to 25,000",
                                                                         25001: "25,001 to 40,000",
                                                                         40001: "40,001 to 65,000",
                                                                         65001: "65,001 to 90,000"}})

        Table101Public = pd.merge(Table101PublicOTC, Table101PublicColl, on='Original Cost New', how='inner')
        Table101Zone = pd.merge(Table101ZoneOTC, Table101ZoneColl, on='Original Cost New', how='inner')
        Table101TTT = pd.merge(Table101TTTOTC, Table101TTTColl, on='Original Cost New', how='inner')
        Table101PPT = pd.merge(Table101PPTOTC, Table101PPTColl, on='Original Cost New', how='inner')

        Table101PublicZone = pd.merge(Table101Public, Table101Zone, on='Original Cost New', how='inner')
        Table101PublicZoneTTT = pd.merge(Table101PublicZone, Table101TTT, on='Original Cost New', how='inner')
        Table101PublicZoneTTTPPT = pd.merge(Table101PublicZoneTTT, Table101PPT, on='Original Cost New', how='inner', suffixes= (f'x2', f'_y2'))

        return Table101PublicZoneTTTPPT

    # Builds RULE 101.A.4.a. TRUCKS AND TRUCK TRACTORS TYPE COLLISION VEHICLE AGE AND PRICE BRACKET FACTOR
    # Returns a dataframe
    @log_exceptions
    def build101A12(self, company):
        Table101PublicOTC = pd.DataFrame(
            self.rateTables[company]['PublicTransportationOtherThanCollisionAgeGroupRelativity'][1:], index=None,
            columns=self.rateTables[company]['PublicTransportationOtherThanCollisionAgeGroupRelativity'][0])
        # Table101PublicOTC = self.buildDataFrame("PublicTransportationOtherThanCollisionAgeGroupRelativity")
        Table101PublicOTC = Table101PublicOTC.rename(
            columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Other Than Collision'})
        Table101PublicOTC = Table101PublicOTC.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101PublicColl = pd.DataFrame(
            self.rateTables[company]['PublicTransportationCollisionAgeGroupRelativity'][1:], index=None,
            columns=self.rateTables[company]['PublicTransportationCollisionAgeGroupRelativity'][0])
        # Table101PublicColl = self.buildDataFrame("PublicTransportationCollisionAgeGroupRelativity")
        Table101PublicColl = Table101PublicColl.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Collision'})
        Table101PublicColl = Table101PublicColl.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101ZoneOTC = pd.DataFrame(self.rateTables[company]['ZoneRatedOtherThanCollisionAgeGroupRelativity'][1:],
                                       index=None, columns=
                                       self.rateTables[company]['ZoneRatedOtherThanCollisionAgeGroupRelativity'][0])
        # Table101ZoneOTC = self.buildDataFrame("ZoneRatedOtherThanCollisionAgeGroupRelativity")
        Table101ZoneOTC = Table101ZoneOTC.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Other Than Collision'})
        Table101ZoneOTC = Table101ZoneOTC.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101ZoneColl = pd.DataFrame(self.rateTables[company]['ZoneRatedCollisionAgeGroupRelativity'][1:],
                                        index=None,
                                        columns=self.rateTables[company]['ZoneRatedCollisionAgeGroupRelativity'][0])
        # Table101ZoneColl = self.buildDataFrame("ZoneRatedCollisionAgeGroupRelativity")
        Table101ZoneColl = Table101ZoneColl.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Collision'})
        Table101ZoneColl = Table101ZoneColl.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101TTTOTC = pd.DataFrame(self.rateTables[company]['TruckOtherThanCollisionAgeGroupRelativity'][1:],
                                      index=None,
                                      columns=self.rateTables[company]['TruckOtherThanCollisionAgeGroupRelativity'][0])
        # Table101TTTOTC = self.buildDataFrame("TruckOtherThanCollisionAgeGroupRelativity")
        Table101TTTOTC = Table101TTTOTC.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Other Than Collision'})
        Table101TTTOTC = Table101TTTOTC.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101TTTColl = pd.DataFrame(self.rateTables[company]['TruckCollisionAgeGroupRelativity'][1:], index=None,
                                       columns=self.rateTables[company]['TruckCollisionAgeGroupRelativity'][0])
        # Table101TTTColl = self.buildDataFrame("TruckCollisionAgeGroupRelativity")
        Table101TTTColl = Table101TTTColl.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Collision'})
        Table101TTTColl = Table101TTTColl.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101PPTOTC = pd.DataFrame(
            self.rateTables[company]['PrivatePassengerOtherThanCollisionAgeGroupRelativity'][1:], index=None,
            columns=self.rateTables[company]['PrivatePassengerOtherThanCollisionAgeGroupRelativity'][0])
        # Table101PPTOTC = self.buildDataFrame("PrivatePassengerOtherThanCollisionAgeGroupRelativity")
        Table101PPTOTC = Table101PPTOTC.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Other Than Collision'})
        Table101PPTOTC = Table101PPTOTC.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101PPTColl = pd.DataFrame(self.rateTables[company]['PrivatePassengerCollisionAgeGroupRelativity'][1:],
                                       index=None,
                                       columns=self.rateTables[company]['PrivatePassengerCollisionAgeGroupRelativity'][
                                           0])
        # Table101PPTColl = self.buildDataFrame("PrivatePassengerCollisionAgeGroupRelativity")
        Table101PPTColl = Table101PPTColl.rename(columns={'AgeGroup': 'Vehicle Age', 'Factor': 'Collision'})
        Table101PPTColl = Table101PPTColl.replace(
            {'Vehicle Age': {1: "Current Model Year", 2: "First Preceding Year", 3: "2nd", 4: "3rd", 5: "4th", \
                             6: "5th", 7: "6th", 8: "7th", 9: "8th", 10: "9th", 11: "10th", 12: "All Other"}})

        Table101Public = pd.merge(Table101PublicOTC, Table101PublicColl, on='Vehicle Age', how='inner')
        Table101Zone = pd.merge(Table101ZoneOTC, Table101ZoneColl, on='Vehicle Age', how='inner')
        Table101TTT = pd.merge(Table101TTTOTC, Table101TTTColl, on='Vehicle Age', how='inner')
        Table101PPT = pd.merge(Table101PPTOTC, Table101PPTColl, on='Vehicle Age', how='inner')

        Table101PublicZone = pd.merge(Table101Public, Table101Zone, on='Vehicle Age', how='inner')
        Table101PublicZoneTTT = pd.merge(Table101PublicZone, Table101TTT, on='Vehicle Age', how='inner')
        Table101PublicZoneTTTPPT = pd.merge(Table101PublicZoneTTT, Table101PPT, on='Vehicle Age', how='inner', suffixes= (f'x2', f'_y2'))

        return Table101PublicZoneTTTPPT

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def build101b1(self, company):
        table101b1 = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][0]).rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        #table101b1 = self.buildDataFrame("TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext").rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        table101b1 = table101b1.replace({'Price Bracket (OCN Range)' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        #for ind in table101b1.index:
        #    table101b1['Price Bracket (OCN Range)'][ind] = (table101b1['Price Bracket (OCN Range)'][ind]).astype(str)[:-2] + " to " + (table101b1['Price Bracket (OCN Range)'][ind] + 1).astype(str)[:-2]

        #if table101b1['Price Bracket (OCN Range)'][ind] == 0:
        #    table101b1['Price Bracket (OCN Range)'][ind] = (table101b1['Price Bracket (OCN Range)'][ind]).astype(str)[:-2] + " to " + (table101b1['Price Bracket (OCN Range)'][ind] + 1).astype(str)[:-2]
        #else:
        #    table101b1['Number'][ind] = table101b1['NumberOfUnitsMinimum'][ind].astype(str) + " to " + table101b1['NumberofUnitsMaximum'][ind].astype(str)[:-2]
        return table101b1

    # Builds the Retention Grade table
    # Returns a dataframe)
    @log_exceptions
    def build101b2(self, company):
        table101b2 = pd.DataFrame(self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][0]).rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        #table101b2 = self.buildDataFrame("TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext").rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        table101b2 = table101b2.replace({'Price Bracket (OCN Range)' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        return table101b2

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def build101b3(self, company):
        table101b3 = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext'][0]).rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        #table101b3 = self.buildDataFrame("TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext").rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        table101b3 = table101b3.replace({'Price Bracket (OCN Range)' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        return table101b3

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def build101b4(self, company):
        table101b4 = pd.DataFrame(self.rateTables[company]['PrivatePassengerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][0]).rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        #table101b4 = self.buildDataFrame("PrivatePassengerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext").rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        table101b4 = table101b4.replace({'Price Bracket (OCN Range)' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        return table101b4

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def build101b5(self, company):
        table101b5 = pd.DataFrame(self.rateTables[company]['PrivatePassengerTypesComprehensiveVehicleValueFactorsStatedAmountVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerTypesComprehensiveVehicleValueFactorsStatedAmountVehicles_Ext'][0]).rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        #table101b5 = self.buildDataFrame("PrivatePassengerTypesComprehensiveVehicleValueFactorsStatedAmountVehicles_Ext").rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 'Factor' : 'All Stated Amount Vehicles'})
        table101b5 = table101b5.replace({'Price Bracket (OCN Range)' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        return table101b5


    def buildZoneRatedTrailersVVFColl(self, company):
        # RULE 301.C.1.A.(1). ZONE-RATED TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH STATED AMOUNT RATING

        factor_table = pd.DataFrame(
            self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:], index=None,
            columns=self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][0])

        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildZoneRatedNonTrailersVVFColl(self, company):
        # RULE 301.C.1.A.(2). ZONE-RATED NON-TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH STATE AMOUNT RATING

        factor_table = pd.DataFrame(
            self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:],
            index=None,
            columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][0])
        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildZoneRatedVehiclesVVFOTC(self, company):
        # RULE 301.C.1.B.(1). ZONE-RATED VEHICLES VEHICLE VALUE FACTORS - OTHER THAN COLLISION WITH STATE AMOUNT RATING

        factor_table = pd.DataFrame(self.rateTables[company][
                                        'TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext'][
                                    1:], index=None, columns=self.rateTables[company][
            'TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext'][0])
        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]
        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildPPTVVFColl(self, company):
        # RULE 301.C.1.A.(3). PRIVATE PASSENGER TYPES VEHICLE VALUE FACTORS - COLLISIONS WITH STATE AMOUNT RATING

        factor_table = pd.DataFrame(
            self.rateTables[company]['PrivatePassengerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:],
            index=None,
            columns=self.rateTables[company]['PrivatePassengerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][
                0])

        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildNonZoneRatedNonTrailersVVFColl(self, company):
        # RULE 301.C.1.A.(4) NON-ZONE-RATED TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH STATED AMOUNT RATING

        factor_table = pd.DataFrame(
            self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:], index=None,
            columns=self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][0])
        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildAllOtherVehiclesVVFColl(self, company):
        # 301.C.1.A.(5). ALL OTHER VEHICLES VEHICLE VALUE FACTORS - COLLISIONS WITH STATED AMOUNT RATING

        factor_table = pd.DataFrame(
            self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][1:],
            index=None,
            columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext'][
                0])
        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildPPTVehiclesVVFOTC(self, company):
        # RULE 301.C.1.B.(2). PRIVATE PASSENGER TYPES VEHICLE VALUE FACTORS - OTHER THAN COLLISION WITH STATED AMOUNT RATING

        factor_table = pd.DataFrame(
            self.rateTables[company]['PrivatePassengerTypesComprehensiveVehicleValueFactorsStatedAmountVehicles_Ext'][1:],
            index=None, columns=
            self.rateTables[company]['PrivatePassengerTypesComprehensiveVehicleValueFactorsStatedAmountVehicles_Ext'][0])
        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildAllOtherVehiclesVVFOTC(self, company):
        # RULE 301.C.1.B.(3) ALL OTHER VEHICLES VEHICLE VALUE FACTORS - OTHER THAN COLLISION WITH STATED AMOUNT RATING

        factor_table = pd.DataFrame(self.rateTables[company][
                                        'TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext'][
                                    1:], index=None, columns=self.rateTables[company][
            'TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext'][0])
        self.price_brackets = [
            "0 to 999",
            "1,000 to 1,999",
            "2,000 to 2,999",
            "3,000 to 3,999",
            "4,000 to 4,999",
            "5,000 to 5,999",
            "6,000 to 7,999",
            "8,000 to 9,999",
            "10,000 to 11,999",
            "12,000 to 13,999",
            "14,000 to 15,999",
            "16,000 to 17,999",
            "18,000 to 19,999",
            "20,000 to 24,999",
            "25,000 to 29,999",
            "30,000 to 34,999",
            "35,000 to 39,999",
            "40,000 to 44,999",
            "45,000 to 49,999",
            "50,000 to 54,999",
            "55,000 to 64,999",
            "65,000 to 74,999",
            "75,000 to 84,999",
            "85,000 to 99,999",
            "100,000 to 114,999",
            "115,000 to 129,999",
            "130,000 to 149,999",
            "150,000 to 174,999",
            "175,000 to 199,999",
            "200,000 to 229,999",
            "230,000 to 259,999",
            "260,000 to 299,999",
            "300,000 to 349,999",
            "350,000 to 399,999",
            "400,000 to 449,999",
            "450,000 to 499,999",
            "500,000 to 599,999",
            "600,000 to 699,999",
            "700,000 to 799,999",
            "800,000 to 899,999",
            "900,000+"
        ]

        output_table = factor_table.rename(
            columns={"PriceBracket": "Price Bracket", "Factor": "All Stated Amount Vehicles"})
        output_table["Price Bracket"] = self.price_brackets
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildExperienceRating(self, company):
        ExperienceRatingEF = pd.DataFrame(self.rateTables[company]['ExperienceRatingExpectedFrequencyPerPowerUnit_Ext'][1:], index=None, columns=self.rateTables[company]['ExperienceRatingExpectedFrequencyPerPowerUnit_Ext'][0]).rename(columns={'Constant' : 'Variable', 'ExpectedFrequencyPerPowerUnit' : 'Factor'}).replace({'Variable' : {'Y' : 'EF'}})
        ExperienceRatingZ = pd.DataFrame(self.rateTables[company]['ExperienceRatingCredibility_Ext'][1:], index=None, columns=self.rateTables[company]['ExperienceRatingCredibility_Ext'][0]).rename(columns={'Constant' : 'Variable', 'Credibility' : 'Factor'}).replace({'Variable' : {'Y' : 'Z'}})
        ExperienceRatingBZ = pd.DataFrame(self.rateTables[company]['ExperienceRatingBaseCredibility_Ext'][1:], index=None, columns=self.rateTables[company]['ExperienceRatingBaseCredibility_Ext'][0]).rename(columns={'Constant' : 'Variable', 'Credibility' : 'Factor'}).replace({'Variable' : {'Y' : 'BZ'}})
        ExperienceRating = pd.concat([ExperienceRatingEF, ExperienceRatingZ, ExperienceRatingBZ])

        return ExperienceRating

    def buildExperienceRatingMinMax(self, company):
        table = pd.DataFrame(self.rateTables[company]["ExperienceRatingModifierRange_Ext"][1:], index=None,
                             columns=self.rateTables[company]["ExperienceRatingModifierRange_Ext"][0])

        table.rename(columns = {"Constant" : "Modification: ", "Factor" : ""}, inplace = True)
        table = table.astype(object)
        table.iloc[0, 1] = f"{table.iloc[0, 1] * 100:.0f}%"
        table = table.astype(object)
        table.iloc[1, 1] = f"{table.iloc[1, 1] * 100:.0f}%"

        return table
    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildCorporalPunish(self, company):
        global MinPremFactor154
        global MinPrem154
        CorporalPunish1 = pd.DataFrame(self.rateTables[company]['CorporalPunishmentBaseRate_Ext'][1:], index=None, columns=self.rateTables[company]['CorporalPunishmentBaseRate_Ext'][0])
        CorporalPunish2 = pd.DataFrame(self.rateTables[company]['MiscellaneousMinimumMaximumPremium_Ext'][1:], index=None, columns=self.rateTables[company]['MiscellaneousMinimumMaximumPremium_Ext'][0])
        #CorporalPunish2 = self.buildDataFrame("MiscellaneousMinimumMaximumPremium_Ext")
        CorporalPunish2 = CorporalPunish2.query(f'CoverageType == "CorporalPunishment_Ext"').filter(items=['CoverageType', 'Premium']).rename(columns={'CoverageType' : 'Limit', 'Premium' : 'BaseRate'})
        CorporalPunish = pd.concat([CorporalPunish1, CorporalPunish2]).replace({'Limit' : {'CorporalPunishment_Ext' : 'Minimum Premium'}})
        CorporalPunish = CorporalPunish.astype(object)
        CorporalPunish.iloc[:2,1:] = CorporalPunish.iloc[:2,1:].astype(float).map(lambda x: f"{x:.2f}")
        CorporalPunish = CorporalPunish.astype(object)
        CorporalPunish.iloc[2:3,1:2] = CorporalPunish.iloc[2:3,1:2].astype(float).map(lambda x: f"{x:.0f}")

        MinPremFactor154 = pd.DataFrame(self.rateTables[company]['BroadFormSchoolBusOperatorsCoverageFactor_Ext'][1:], index=None, columns=self.rateTables[company]['BroadFormSchoolBusOperatorsCoverageFactor_Ext'][0])
        #MinPremFactor154 = self.buildDataFrame("BroadFormSchoolBusOperatorsCoverageFactor_Ext")
        MinPremFactor154 = MinPremFactor154.iloc[0, 1]
        MinPrem154 = pd.DataFrame(self.rateTables[company]['MiscellaneousMinimumMaximumPremium_Ext'][1:], index=None, columns=self.rateTables[company]['MiscellaneousMinimumMaximumPremium_Ext'][0])
        #MinPrem154 = self.buildDataFrame("MiscellaneousMinimumMaximumPremium_Ext").query(f'CoverageType == "BroadForm_Ext"')
        MinPrem154 = MinPrem154.iloc[0, 3]

        return CorporalPunish.rename(columns={'Limit' : 'Limit each Claim/Each Aggregate', 'BaseRate' : 'Charge'})

    #MinPremFactor = self.buildDataFrame("BroadFormSchoolBusOperatorsCoverageFactor_Ext")
    #MinPremFactor = MinPremFactor.iloc[0, 1]
    #MinPremFactor = " "
    #MinPremLimit = " "

    # Builds the Retention Grade table
    # Returns a dataframe
    @log_exceptions
    def buildRiskTiering(self, company):
        RiskTieringLiab = pd.DataFrame(self.rateTables[company]['TieringLiabilityFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TieringLiabilityFactor_Ext'][0]).rename(columns={'Factor' : 'Liability'}).sort_values('Liability')
        RiskTieringCollision = pd.DataFrame(self.rateTables[company]['TieringCollisionFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TieringCollisionFactor_Ext'][0]).rename(columns={'Factor' : 'Collision'}).sort_values('Collision')
        RiskTieringOTC = pd.DataFrame(self.rateTables[company]['TieringOtherThanCollisionFactor_Ext'][1:], index=None, columns=self.rateTables[company]['TieringOtherThanCollisionFactor_Ext'][0]).rename(columns={'Factor' : 'Other Than Collision'}).sort_values('Other Than Collision')
        RiskTieringLiabColl = pd.merge(RiskTieringLiab, RiskTieringCollision, on='Grade', how='inner')
        RiskTiering = pd.merge(RiskTieringLiabColl, RiskTieringOTC, on='Grade', how='inner')
        return RiskTiering

    # Builds RULE 66. ANTIQUE AUTOS table
    # Returns a dataframe
    @log_exceptions
    # Slightly edited for decimals and PIP, added company use.
    def buildAntiqueAutoLiabFactors(self, company):
        # Rule 266.B. Premium Computation (Liability and Basic No-Fault)
        AntiqueAutoLiabFactors = pd.DataFrame(data=self.rateTables[company]["SpecialTypesAntiqueAutoFactor"][1:], index=None, columns=self.rateTables[company]["SpecialTypesAntiqueAutoFactor"][0])

        AntiqueAutoLiabFactors['TypeCoverage'].replace('UM/UIM', 'Uninsured Motorists', inplace=True)
        AntiqueAutoLiabFactors['TypeCoverage'].replace('No-Fault', 'PIP', inplace=True)
        AntiqueAutoLiabFactors['TypeCoverage'].replace('MedicalPayments', 'Medical Payments', inplace=True)
        AntiqueAutoLiabFactors = AntiqueAutoLiabFactors.rename(columns={'TypeCoverage':'Coverage'})

        if self.StateAbb not in self.pip_states.values:
            AntiqueAutoLiabFactors = AntiqueAutoLiabFactors[AntiqueAutoLiabFactors['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            AntiqueAutoLiabFactors = AntiqueAutoLiabFactors[AntiqueAutoLiabFactors['Coverage'] != "Medical Payments"]

        AntiqueAutoLiabFactors.iloc[:, 1:] = AntiqueAutoLiabFactors.iloc[:, 1:].round(3)

        return AntiqueAutoLiabFactors


    @log_exceptions
    def buildAntiqueAutoPDRates(self, company):
        # Rule 266.B. Premium Computation (Physical Damage)

        factor_table = pd.DataFrame(self.rateTables[company]['SpecialTypesAntiquePhysicalDamageRate_Ext'][1:], index=None, columns=self.rateTables[company]['SpecialTypesAntiquePhysicalDamageRate_Ext'][0])

        # Function to clean the Factor column
        # Super weird bug, Factor column reads in "=Round(0.XXXXXXXXXX,3)" instead of the rounded float. This is a function to fix that.
        def clean_factor(value):
            if isinstance(value, str) and value.startswith('=ROUND'):
                # Extract the number from the string
                match = re.search(r'\d+\.\d+', value)
                if match:
                    return float(match.group(0))
            try:
                return float(value)
            except ValueError:
                return None

            # Apply the cleaning function to the Factor column
        factor_table['BasePremium'] = factor_table['BasePremium'].apply(clean_factor)

        coll_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "266 Coll Map").fillna("Y")
        otc_deductible_map = pd.read_excel(BA_INPUT_FILE, sheet_name = "266 OTC Map").fillna("Y")
        factor_table = factor_table
        deductible_amount = factor_table[factor_table["typeCoverage"] == "PhysDamCollSA"]["Deductible"]
        collision_factors = factor_table[factor_table["typeCoverage"] == "PhysDamCollSA"].drop(columns = "typeCoverage")
        otc_factors = factor_table[factor_table["typeCoverage"] == "PhysDamOTCSA"].drop(columns = "typeCoverage")

        # Filtering down the deductibles in each coverage based on their corresponding map.

        state_lookup_coll = self.StateAbb
        state_lookup_otc = self.StateAbb
        if self.StateAbb not in coll_deductible_map.columns:
            state_lookup_coll = "All Other"
        if self.StateAbb not in otc_deductible_map.columns:
            state_lookup_otc = "All Other"

        coll_deduc_map = coll_deductible_map[coll_deductible_map[state_lookup_coll] != "N"]
        coll_deduc_map = pd.DataFrame(coll_deduc_map["Limit"])

        otc_deduc_map = otc_deductible_map[otc_deductible_map[state_lookup_otc] != "N"]
        otc_deduc_map = pd.DataFrame(otc_deduc_map["Limit"])

        coll_deduc_map['Limit'] = coll_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.
        otc_deduc_map['Limit'] = otc_deduc_map['Limit'].apply(lambda x: "{:,}".format(x)) # Limit column loses it's commas for some reason, need to add back.

        # Collision factors get inner joined -> drop unneeded new limits column ->  rename factor column -> repeat for remaining coverages.
        collision_factors = pd.merge(collision_factors, coll_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"BasePremium" : "Collision"})
        otc_factors = pd.merge(otc_factors, otc_deduc_map, left_on='Deductible', right_on='Limit', how='inner').drop(columns = "Limit").rename(columns = {"BasePremium" : "Comprehensive"})

        collision_factors['Collision'] = collision_factors['Collision'].astype(float).round(3) # Very weird things are happening...
        otc_factors['Comprehensive'] = otc_factors['Comprehensive'].astype(float).round(3)

        # Now left join those coverages onto the original deductible list
        output_table = pd.DataFrame(deductible_amount)

        if collision_factors.shape[0] > 0:
            output_table = output_table.merge(collision_factors, on='Deductible', how='left')
        if otc_factors.shape[0] > 0:
            output_table = output_table.merge(otc_factors, on='Deductible', how='left')

        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: int(x.replace(',', '')))
        output_table.sort_values(by='Deductible', inplace=True)
        output_table.dropna(subset=output_table.columns[1:], how='all', inplace=True)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].round(3)

        # Put the commas back in the Deductible column
        output_table['Deductible'] = output_table['Deductible'].apply(lambda x: "{:,}".format(x))
        output_table.fillna("NA", inplace = True)

        return output_table

    # Iso Currency 271
    def buildFireDepartmentPPTFactors(self,company):
    # 271.B.1. Private Passenger Type Autos
        sheet_names = ['SpecialTypesFireDepartmentFactor']*6
        orig_values = ['7908 Liability', '7908 MedicalPayments','7908 No-Fault','7908 PhysDamOTCACV','7908 PhysDamCollACV','7908 UM/UIM']
        replace_values = ['Liability', 'Med', 'PIP', 'Other than Collision', 'Collision', 'All Other']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values,replace_values, filter_values=orig_values)

        if self.StateAbb not in self.pip_states.values:
            output_table = output_table[output_table['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            output_table = output_table[output_table['Coverage'] != "Med"]
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildFireDepartmentOtherThanPPTFactors(self,company):
        # 271.B.3. Other than Private Passenger Autos or Trailers
        sheet_names = ['SpecialTypesFireDepartmentFactor']*6
        orig_values = ['7909 Liability', '7909 MedicalPayments','7909 No-Fault','7909 PhysDamOTCACV','7909 PhysDamCollACV','7909 UM/UIM']
        replace_values = ['Liability', 'Med', 'PIP', 'Other than Collision', 'Collision', 'All Other']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values,replace_values, filter_values=orig_values)

        if self.StateAbb not in self.pip_states.values:
            output_table = output_table[output_table['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            output_table = output_table[output_table['Coverage'] != "Med"]
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildFireDepartmentBuyBackFactor(self, company):
        # 271.B.3. Other than Private Passenger Autos or Trailers
        sheet_names = ['SpecialTypesEmergencyVehicleBuybackFactor']
        orig_values = ['Y']
        replace_values = ['']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: 'Factor'})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Iso Currency 274
    def buildLawEnforcementPPTFactors(self,company):
        # 274.B.1. Private Passenger Types
        sheet_names = ['SpecialTypesLawEnforcementFactor']*6
        orig_values = ['7911 Liability', '7911 MedicalPayments','7911 No-Fault','7911 PhysDamOTCACV','7911 PhysDamCollACV','7911 UM/UIM']
        replace_values = ['Liability', 'Med', 'PIP', 'Other than Collision', 'Collision', 'All Other']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values,replace_values, filter_values=orig_values)

        if self.StateAbb not in self.pip_states.values:
            output_table = output_table[output_table['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            output_table = output_table[output_table['Coverage'] != "Med"]
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildLawEnforcementOtherThanPPTFactors(self,company):
        # 274.B.1. Private Passenger Types
        sheet_names = ['SpecialTypesLawEnforcementFactor']*6
        orig_values = ['7912 Liability', '7912 MedicalPayments','7912 No-Fault','7912 PhysDamOTCACV','7912 PhysDamCollACV','7912 UM/UIM']
        replace_values = ['Liability', 'Med', 'PIP', 'Other than Collision', 'Collision', 'All Other']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values,replace_values, filter_values=orig_values)

        if self.StateAbb not in self.pip_states.values:
            output_table = output_table[output_table['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            output_table = output_table[output_table['Coverage'] != "Med"]
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildLawEnforcementBuyBackFactor(self, company):
        # 274.B.5. Fellow Volunteer Workers Liability Coverage
        sheet_names = ['SpecialTypesEmergencyVehicleBuybackFactor']
        orig_values = ['Y']
        replace_values = ['']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: 'Factor'})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Builds RULE 75. LeasingOrRentalConcerns table
    # Returns a dataframe
    @log_exceptions
    def buildLeasingOrRentalConcernsFactors(self, company):
        LeasingOrRentalConcernsFactors = pd.DataFrame(self.rateTables[company]['LeasingOrRentalConcernsContingentBasePremium_Ext'][1:], index=None, columns=self.rateTables[company]['LeasingOrRentalConcernsContingentBasePremium_Ext'][0])
        #LeasingOrRentalConcernsFactors = self.buildDataFrame("LeasingOrRentalConcernsContingentBasePremium_Ext")
        LeasingOrRentalConcernsFactors['VehicleType'].replace('Special', 'Special Types', inplace=True)
        LeasingOrRentalConcernsFactors['VehicleType'].replace('Truck', 'Trucks, Tractors and Trailers', inplace=True)
        LeasingOrRentalConcernsFactors.rename(columns={'VehicleType': 'Vehicle Type', 'UM/UIM': 'Base Premium', 'BasePremium': 'Base Premium'}, inplace = True)

        LeasingOrRentalConcernsFactors = LeasingOrRentalConcernsFactors.astype(object)
        LeasingOrRentalConcernsFactors.iloc[:, 1:] = LeasingOrRentalConcernsFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.0f}")

        return LeasingOrRentalConcernsFactors

    # builds VA Rule 275.B.1.(b). Short Term - Autos Leased by the Hour, Day, or Week
    # returns a dataframe
    def buildRule275VA_TTTFactors(self,company):
        # pull data from sheet
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesLeasingConcernFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesLeasingConcernFactor'][0])

        # Rename columns (will be swapped to columns later)
        # Rename values in columns using dictionaries
        table['SpecialTypesClassCode'].replace({
            '7211': 'Trucks',
            '7212': 'Tractors',
            '7213': 'Trailers'
        }, inplace=True)

        # Rename Coverages
        table['TypeCoverage'].replace({
            'MedicalPayments': 'Med',
            'PhysDamCollACV': 'Collision',
            'PhysDamOTCACV': 'Comprehensive',
            'UM/UIM': 'UM'
        }, inplace=True)

        # Remove unnessecary factors
        # Define the values to exclude
        exclude_special_types = ['7203', '7204', '7214']
        exclude_type_coverage = ['No-Fault', 'PhysDamCollSA', 'PhysDamOTCSA']

        # Filter the DataFrame
        table = table[
            ~table['SpecialTypesClassCode'].isin(exclude_special_types) &
            ~table['TypeCoverage'].isin(exclude_type_coverage)
            ]

        # Clean up the dataframe
        output_table = table.pivot(index="TypeCoverage", columns="SpecialTypesClassCode", values="Factor")
        output_table.reset_index(inplace=True)
        output_table.rename(columns={output_table.columns[0]: 'Coverage'}, inplace=True)
        output_table = output_table[[output_table.columns[0], output_table.columns[3], output_table.columns[1], output_table.columns[2]]]
        output_table = output_table.iloc[[2, 3, 0, 1, 4]]

        return output_table

    def buildRule275VA_PPTFactors(self,company):
        # pull data from sheet
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesLeasingConcernFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesLeasingConcernFactor'][0])

        # Rename column (will be swapped to column later)
        table['SpecialTypesClassCode'].replace('7214', 'Factor', inplace=True)

        # Rename coverages
        table['TypeCoverage'].replace({
            'MedicalPayments': 'Med',
            'PhysDamCollACV': 'Collision',
            'PhysDamOTCACV': 'Comprehensive',
            'UM/UIM': 'UM'
        }, inplace=True)

        # Remove unnessecary factors
        # Define the values to exclude
        exclude_special_types = ['7203', '7204', '7211', '7212', '7213']
        exclude_type_coverage = ['No-Fault', 'PhysDamCollSA', 'PhysDamOTCSA']

        # Filter the DataFrame
        table = table[
            ~table['SpecialTypesClassCode'].isin(exclude_special_types) &
            ~table['TypeCoverage'].isin(exclude_type_coverage)
            ]

        # Clean up the dataframe
        output_table = table.pivot(index="TypeCoverage", columns="SpecialTypesClassCode", values="Factor")
        output_table.reset_index(inplace=True)
        output_table.rename(columns={output_table.columns[0]: 'Coverage'}, inplace=True)
        output_table = output_table.iloc[[2, 3, 0, 1, 4]]

        return output_table

    def buildRule275VA_MotorcycleFactors(self,company):
        # pull data from sheet
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesLeasingConcernMiscellaneousFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesLeasingConcernMiscellaneousFactor'][0])

        # Rename column (will be swapped to column later)
        table['SpecialTypesSupplementaryType'].replace('Motorcycle', 'Factor', inplace=True)

        # Rename coverages
        table['TypeCoverage'].replace({
            'MedicalPayments': 'Med',
            'PhysDamCollACV': 'Collision',
            'PhysDamOTCACV': 'Comprehensive',
            'UM/UIM': 'UM'
        }, inplace=True)

        # Remove unnessecary factors
        # Define the values to exclude
        exclude_special_types = ['All Other Special Types', 'Non-Dealer Garage Risks', 'Snowmobile With Passenger Hazard', 'Snowmobile Without Passenger Hazard']
        exclude_type_coverage = ['No-Fault', 'PhysDamCollSA', 'PhysDamOTCSA']

        # Filter the DataFrame
        table = table[
            ~table['SpecialTypesSupplementaryType'].isin(exclude_special_types) &
            ~table['TypeCoverage'].isin(exclude_type_coverage)
            ]

        # Clean up the dataframe
        output_table = table.pivot(index="TypeCoverage", columns="SpecialTypesSupplementaryType",values="Factor")
        output_table.reset_index(inplace=True)
        output_table.rename(columns={output_table.columns[0]: 'Coverage'}, inplace=True)
        output_table = output_table.iloc[[2, 3, 0, 1, 4]]

        return output_table

    def buildRule275VA_SnowmobilesFactors(self,company):
        # pull data from sheet
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesLeasingConcernMiscellaneousFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesLeasingConcernMiscellaneousFactor'][0])

        # Rename column (will be swapped to column later)
        table['SpecialTypesSupplementaryType'].replace('Snowmobile With Passenger Hazard', 'Factor', inplace=True)

        # Rename coverages
        table['TypeCoverage'].replace({
            'MedicalPayments': 'Med',
            'PhysDamCollSA': 'Collision',
            'PhysDamOTCSA': 'Comprehensive',
            'UM/UIM': 'UM'
        }, inplace=True)

        # Remove unnessecary factors
        # Define the values to exclude
        exclude_special_types = ['All Other Special Types', 'Non-Dealer Garage Risks', 'Motorcycle', 'Snowmobile Without Passenger Hazard']
        exclude_type_coverage = ['No-Fault']

        # Filter the DataFrame
        table = table[
            ~table['SpecialTypesSupplementaryType'].isin(exclude_special_types) &
            ~table['TypeCoverage'].isin(exclude_type_coverage)
            ]

        # Clean up the dataframe
        output_table = table.pivot(index="TypeCoverage", columns="SpecialTypesSupplementaryType",values="Factor")
        output_table.reset_index(inplace=True)
        output_table.rename(columns={output_table.columns[0]: 'Coverage'}, inplace=True)
        output_table = output_table.iloc[[2, 3, 0, 1, 4]]

        return output_table

    def buildRule275VA_ExceptMotorHomesFactors(self,company):
        # pull data from sheet
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesLeasingConcernMiscellaneousFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesLeasingConcernMiscellaneousFactor'][0])

        # Rename column (will be swapped to column later)
        table['SpecialTypesSupplementaryType'].replace('All Other Special Types', 'Factor', inplace=True)

        # Rename coverages
        table['TypeCoverage'].replace({
            'MedicalPayments': 'Med',
            'PhysDamCollACV': 'Collision',
            'PhysDamOTCACV': 'Comprehensive',
            'UM/UIM': 'UM'
        }, inplace=True)

        # Remove unnessecary factors
        # Define the values to exclude
        exclude_special_types = ['Motorcycle', 'Non-Dealer Garage Risks', 'Snowmobile With Passenger Hazard', 'Snowmobile Without Passenger Hazard']
        exclude_type_coverage = ['No-Fault', 'PhysDamCollSA', 'PhysDamOTCSA']

        # Filter the DataFrame
        table = table[
            ~table['SpecialTypesSupplementaryType'].isin(exclude_special_types) &
            ~table['TypeCoverage'].isin(exclude_type_coverage)
            ]

        # Clean up the dataframe
        output_table = table.pivot(index="TypeCoverage", columns="SpecialTypesSupplementaryType",values="Factor")
        output_table.reset_index(inplace=True)
        output_table.rename(columns={output_table.columns[0]: 'Coverage'}, inplace=True)
        output_table = output_table.iloc[[2, 3, 0, 1, 4]]

        return output_table

    def buildRule275VA_MotorHomesFactors(self,company):
        # pull data from sheet
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesLeasingConcernFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesLeasingConcernFactor'][0])

        # Rename columns (will be swapped to columns later)
        # Rename values in columns using dictionaries
        table['SpecialTypesClassCode'].replace({
            '7203': "Up to 22'",
            '7204': "Over 22'"
        }, inplace=True)

        # Rename Coverages
        table['TypeCoverage'].replace({
            'MedicalPayments': 'Med',
            'PhysDamCollSA': 'Collision',
            'PhysDamOTCSA': 'Comprehensive',
            'UM/UIM': 'UM'
        }, inplace=True)

        # Remove unnessecary factors
        # Define the values to exclude
        exclude_special_types = ['7211', '7212', '7213', '7214']
        exclude_type_coverage = ['No-Fault']

        # Filter the DataFrame
        table = table[
            ~table['SpecialTypesClassCode'].isin(exclude_special_types) &
            ~table['TypeCoverage'].isin(exclude_type_coverage)
            ]

        # Clean up the dataframe
        output_table = table.pivot(index="TypeCoverage", columns="SpecialTypesClassCode", values="Factor")
        output_table.reset_index(inplace=True)
        output_table.rename(columns={output_table.columns[0]: 'Coverage'}, inplace=True)
        output_table = output_table[[output_table.columns[0], output_table.columns[2], output_table.columns[1]]]
        output_table = output_table.iloc[[2, 3, 0, 1, 4]]

        return output_table

    # Builds RULE 76. MOBILE HOMES table
    # Returns a dataframe
    @log_exceptions
    def buildMobileHomeFactors(self, company):
        MobileHomeFactorsRaw1 = pd.DataFrame(self.rateTables[company]['SpecialTypesMobileHomeFactor'][1:], index=None, columns=self.rateTables[company]['SpecialTypesMobileHomeFactor'][0]).query(f'TypeCoverage == "Liability" | TypeCoverage == "PhysDamCollSA" | TypeCoverage == "UM/UIM" ')
        #MobileHomeFactorsRaw1= self.buildDataFrame("SpecialTypesMobileHomeFactor").query(f'TypeCoverage == "Liability" | TypeCoverage == "PhysDamCollSA" | TypeCoverage == "UM/UIM" ')
        MobileHomeFactorsRaw2 = MobileHomeFactorsRaw1.query(f'SpecialTypesClassCode == "7960" | SpecialTypesClassCode == "7961" | SpecialTypesClassCode == "7962" | SpecialTypesClassCode == "7963" ')
        MobileHomeFactorsRaw2.loc[MobileHomeFactorsRaw2['SpecialTypesClassCode'] == '7960', 'Class Description'] = 'Motor Homes-Self Propelled Up to 22 feet'
        MobileHomeFactorsRaw2.loc[MobileHomeFactorsRaw2['SpecialTypesClassCode'] == '7961', 'Class Description'] = 'Motor Homes-Self Propelled More than 22 feet'
        MobileHomeFactorsRaw2.loc[MobileHomeFactorsRaw2['SpecialTypesClassCode'] == '7962', 'Class Description'] = 'Pickup Trucks with Camper Bodies'
        MobileHomeFactorsRaw2.loc[MobileHomeFactorsRaw2['SpecialTypesClassCode'] == '7963', 'Class Description'] = 'Trailer Equipped as Living Quarters'
        MobileHomeFactorsRaw2 = MobileHomeFactorsRaw2.rename(columns={'SpecialTypesClassCode': 'Class Code'})
        MobileHomeFactors = MobileHomeFactorsRaw2.pivot(index=['Class Code','Class Description'], columns='TypeCoverage', values='Factor').reset_index(['Class Code','Class Description'])
        MobileHomeFactors = MobileHomeFactors[["Class Code", "Class Description","Liability","PhysDamCollSA","UM/UIM"]]
        MobileHomeFactors = MobileHomeFactors.rename(columns={'PhysDamCollSA':'Physical Damage','UM/UIM':'All Other Coverages'})
        MobileHomeFactors.drop(columns = ["Class Description"], inplace = True)
        MobileHomeFactors = MobileHomeFactors.astype(object)
        MobileHomeFactors.iloc[:, 1:] = MobileHomeFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return MobileHomeFactors

    # Rule 276
    def buildMobileHomeAdditionalFactor(self, company):
        # 276.B.6 Limited Other Than Collision Coverage on Contents
        sheet_names = ['MobileHomesAdditionalCoveragesFactor']*6
        orig_values = ['Stated Amount - Fire Only',
                       'Stated Amount - Fire and Theft',
                       'Stated Amount - Limited Specified Causes of Loss',
                       'Stated Amount - Limited Specified Causes of Loss Including Theft',
                       'Stated Amount - Specified Causes of Loss',
                       'Stated Amount - Specified Causes of Loss Including Theft']

        replace_values = ['Fire Only',
                          'Fire And Theft Only',
                          'Limited Specified Causes of Loss',
                          'Limited Specified Causes of Loss Including Theft',
                          'Specified Causes of Loss (Excludes Theft)',
                          'Specified Causes of Loss Including Theft (Includes Theft)']

        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values,replace_values, filter_values=orig_values)
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Builds the RULE 78. REGISTRATION PLATES NOT ISSUED FOR A SPECIFIC AUTO Table
    # Returns a dataframe
    @log_exceptions
    def buildRegistrationPlateFactors(self, company):
        RegistrationPlateFactorsRaw1 = pd.DataFrame(self.rateTables[company]['SpecialTypesRegistrationPlatesFactor_Ext'][1:], index=None, columns=self.rateTables[company]['SpecialTypesRegistrationPlatesFactor_Ext'][0]).query(f'TypeCoverage == "Liability"').query(f'TypeCoverage == "Liability"')
        #RegistrationPlateFactorsRaw1 = self.buildDataFrame("SpecialTypesRegistrationPlatesFactor_Ext").query(f'TypeCoverage == "Liability"')
        RegistrationPlateFactorsRaw2 = RegistrationPlateFactorsRaw1.drop(columns='TypeCoverage')
        RegistrationPlateFactors = RegistrationPlateFactorsRaw2.rename(columns={'RegistrationPlateType': 'Registration Plate Type'})
        return RegistrationPlateFactors

    # Builds RULE 90. HiredAuto table
    # Returns a dataframe
    @log_exceptions
    def buildHiredAutoLiabFactors(self, company):
        HiredAutoLiabFactors1 = pd.DataFrame(self.rateTables[company]['LiabilityCostOfHireRate'][1:], index=None, columns=self.rateTables[company]['LiabilityCostOfHireRate'][0])
        HiredAutoMinFactors = pd.DataFrame(self.rateTables[company]['HiredAutoMinimumLiabilityPremium'][1:], index=None, columns=self.rateTables[company]['HiredAutoMinimumLiabilityPremium'][0]).rename(columns={'Premium': 'Rate'})
        HiredAutoLiabFactors = pd.concat([HiredAutoLiabFactors1, HiredAutoMinFactors]).rename(columns={'Constant':'' , 'Rate': 'Rate'})
        HiredAutoLiabFactors = HiredAutoLiabFactors.astype(object)
        HiredAutoLiabFactors.iloc[0, 0] = 'Per $100 Cost of Hire'
        HiredAutoLiabFactors = HiredAutoLiabFactors.astype(object)
        HiredAutoLiabFactors.iloc[1, 0] = 'Minimum Premium - Non Truckers'
        return HiredAutoLiabFactors


    @log_exceptions
    # Rule 290
    def buildHiredAutoPDFactors(self, company):
        HiredAutoCompFactors = pd.DataFrame(self.rateTables[company]['ComprehensiveVehicleWithDriverCostOfHireRate_Ext'][1:], index=None, columns=self.rateTables[company]['ComprehensiveVehicleWithDriverCostOfHireRate_Ext'][0]).rename(columns={'HiredAutoOtherThanCollisionDeductible': 'Deductible','HiredAutoOtherThanCollisionCoverageType': 'Coverage'})
        #HiredAutoCompFactors = self.buildDataFrame("ComprehensiveVehicleWithDriverCostOfHireRate_Ext").rename(columns={'HiredAutoOtherThanCollisionDeductible': 'Deductible','HiredAutoOtherThanCollisionCoverageType': 'Coverage'})
        HiredAutoCompFactors = HiredAutoCompFactors.query(f'Coverage == "Comprehensive" | Coverage == "Specified Causes of Loss" ')
        HiredAutoCompFactors.iloc[:, 2] = HiredAutoCompFactors.iloc[:, 2]
        HiredAutoCollFactors = pd.DataFrame(self.rateTables[company]['CollisionVehicleWithDriverCostOfHireRate_Ext'][1:], index=None, columns=self.rateTables[company]['CollisionVehicleWithDriverCostOfHireRate_Ext'][0]).rename(columns={'HiredAutoCollisionDeductible': 'Deductible'})
        #HiredAutoCollFactors = self.buildDataFrame("CollisionVehicleWithDriverCostOfHireRate_Ext").rename(columns={'HiredAutoCollisionDeductible': 'Deductible'})
        HiredAutoCollFactors.iloc[:, 1] = HiredAutoCollFactors.iloc[:, 1]
        HiredAutoCollFactors['Coverage'] = "Collision"
        HiredAutoCollFactors.loc[HiredAutoCollFactors['Deductible'] == '100', 'Rate'] = 'N/A'
        HiredAutoPDFactors = pd.concat([HiredAutoCompFactors, HiredAutoCollFactors])
        #add min premium to the table
        HiredAutoCompMinFactors = pd.DataFrame(self.rateTables[company]['Hired Auto MinimumOtherThanCollisionPremium_Ext'][1:], index=None, columns=self.rateTables[company]['Hired Auto MinimumOtherThanCollisionPremium_Ext'][0]).rename(columns={'HiredAutoOtherThanCollisionCoverageType': 'Coverage','MinimumPremium':'Rate'})
        #HiredAutoCompMinFactors = self.buildDataFrame("Hired Auto MinimumOtherThanCollisionPremium_Ext").rename(columns={'HiredAutoOtherThanCollisionCoverageType': 'Coverage','MinimumPremium':'Rate'})
        HiredAutoCompMinFactors['Deductible'] = "Minimum Premium"
        HiredAutoCollMinFactors = pd.DataFrame(self.rateTables[company]['HiredAutoMinimumCollisionPremium_Ext'][1:], index=None, columns=self.rateTables[company]['HiredAutoMinimumCollisionPremium_Ext'][0]).rename(columns={'Constant': 'Coverage','MinimumPremium':'Rate'})
        #HiredAutoCollMinFactors = self.buildDataFrame("HiredAutoMinimumCollisionPremium_Ext").rename(columns={'Constant': 'Coverage','MinimumPremium':'Rate'})
        HiredAutoCollMinFactors['Deductible'] = "Minimum Premium"
        HiredAutoCollMinFactors['Coverage'] = "Collision"
        HiredAutoPDFactors = pd.concat([HiredAutoPDFactors, HiredAutoCompMinFactors,HiredAutoCollMinFactors])
        HiredAutoPDFactors = HiredAutoPDFactors.pivot(index=['Deductible'],columns='Coverage', values='Rate').reset_index(['Deductible'])
        HiredAutoPDFactors = HiredAutoPDFactors[['Deductible', 'Comprehensive', 'Specified Causes of Loss','Collision']]
        HiredAutoPDFactors = HiredAutoPDFactors.sort_values(by=['Comprehensive'], ascending=False)


        # Move "Minimum Premium" row to the bottom
        min_premium_row = HiredAutoPDFactors[HiredAutoPDFactors['Deductible'] == 'Minimum Premium']
        HiredAutoPDFactors = HiredAutoPDFactors[HiredAutoPDFactors['Deductible'] != 'Minimum Premium']
        HiredAutoPDFactors = pd.concat([HiredAutoPDFactors, min_premium_row], ignore_index=True)

        def safe_format(x):
            try:
                return f"{float(x):.2f}"

            except (ValueError, TypeError):
                return x

        HiredAutoPDFactors = HiredAutoPDFactors.astype(object)
        HiredAutoPDFactors.iloc[:-1, 1:] = HiredAutoPDFactors.iloc[:-1, 1:].map(safe_format)

        return HiredAutoPDFactors

    # Builds RULE RULE 115. BUSINESS INTERRUPTION COVERAGE
    # Returns a dataframe
    @log_exceptions
    def buildBusinessInterruptionFactors(self, company):
        BusinessInterruptionCompFactors = pd.DataFrame(self.rateTables[company]['BusinessInterruptionCoverageOtherThanCollisionBaseLossCost'][1:], index=None, columns=self.rateTables[company]['BusinessInterruptionCoverageOtherThanCollisionBaseLossCost'][0]).rename(columns={'BasePremium': 'Comprehensive'})
        #BusinessInterruptionCompFactors = self.buildDataFrame("BusinessInterruptionCoverageOtherThanCollisionBaseLossCost").rename(columns={'BasePremium': 'Comprehensive'})
        BusinessInterruptionSpecifiedCauseFactors = pd.DataFrame(self.rateTables[company]['BusinessInterruptionCoverageSpecifiedCausesOfLossBaseLossCost'][1:], index=None, columns=self.rateTables[company]['BusinessInterruptionCoverageSpecifiedCausesOfLossBaseLossCost'][0]).rename(columns={'BasePremium': 'Specified Causes Of Loss'})
        #BusinessInterruptionSpecifiedCauseFactors = self.buildDataFrame("BusinessInterruptionCoverageSpecifiedCausesOfLossBaseLossCost").rename(columns={'BasePremium': 'Specified Causes Of Loss'})
        BusinessInterruptionCollFactors = pd.DataFrame(self.rateTables[company]['BusinessInterruptionCoverageCollisionBaseLossCost'][1:], index=None, columns=self.rateTables[company]['BusinessInterruptionCoverageCollisionBaseLossCost'][0]).rename(columns={'BasePremium': 'Collision'})
        #BusinessInterruptionCollFactors = self.buildDataFrame("BusinessInterruptionCoverageCollisionBaseLossCost").rename(columns={'BasePremium': 'Collision'})
        BusinessInterruptionFactors = pd.merge(BusinessInterruptionCompFactors, BusinessInterruptionSpecifiedCauseFactors, on='BusinessIncomeCoverageType', how='inner')
        BusinessInterruptionFactors = pd.merge(BusinessInterruptionFactors,BusinessInterruptionCollFactors,on='BusinessIncomeCoverageType', how='inner').rename(columns={'BusinessIncomeCoverageType': 'Covered Causes Of Loss Option'})
        BusinessInterruptionFactors = BusinessInterruptionFactors.astype(object)
        BusinessInterruptionFactors.iloc[:, 1:] = BusinessInterruptionFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.2f}")

        return BusinessInterruptionFactors


    @log_exceptions
    def buildExtendedBusinessFactors(self, company):
        ExtendedBusinessFactors = pd.DataFrame(self.rateTables[company]['ExtendedBusinessIncomeAdditionalCoverageFactor'][1:], index=None, columns=self.rateTables[company]['ExtendedBusinessIncomeAdditionalCoverageFactor'][0]).rename(columns={'ExtendedBusinessIncomeAdditionalCoverageNumberOfDays': 'Number of Days'})
        #ExtendedBusinessFactors = self.buildDataFrame("ExtendedBusinessIncomeAdditionalCoverageFactor").rename(columns={'ExtendedBusinessIncomeAdditionalCoverageNumberOfDays': 'Number of Days'})
        ExtendedBusinessFactors = ExtendedBusinessFactors.astype(object)
        ExtendedBusinessFactors.iloc[:, 1:] = ExtendedBusinessFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return ExtendedBusinessFactors

    @log_exceptions
    def buildWaitingBusinessFactors(self, company):
        WaitingBusinessFactors = pd.DataFrame(self.rateTables[company]['BusinessIncomeCoverageWaitingPeriodFactor'][1:], index=None, columns=self.rateTables[company]['BusinessIncomeCoverageWaitingPeriodFactor'][0]).rename(columns={'BusinessIncomeCoverageWaitingPeriod': 'Duration for Waiting Period'})
        #WaitingBusinessFactors = self.buildDataFrame("BusinessIncomeCoverageWaitingPeriodFactor").rename(columns={'BusinessIncomeCoverageWaitingPeriod': 'Duration for Waiting Period'})
        WaitingBusinessFactors = WaitingBusinessFactors.astype(object)
        WaitingBusinessFactors.iloc[:, 1:] = WaitingBusinessFactors.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return WaitingBusinessFactors

    @log_exceptions
    def buildInsuranceToExposureFactor(self, company):
        InsuranceToExposureFactor = pd.DataFrame(self.rateTables[company]['InsuranceToExposureFactor'][1:], index=None, columns=self.rateTables[company]['InsuranceToExposureFactor'][0]).rename(columns={'PercentageOfInsuranceToExposure': 'Percentage'})
        #InsuranceToExposureFactor = self.buildDataFrame("InsuranceToExposureFactor").rename(columns={'PercentageOfInsuranceToExposure': 'Percentage'})
        InsuranceToExposureFactor['Percentage'].replace(0, 'Under 25%', inplace=True)
        InsuranceToExposureFactor['Percentage'].replace(25, '25 to 49%', inplace=True)
        InsuranceToExposureFactor['Percentage'].replace(50, '50 to 74%', inplace=True)
        InsuranceToExposureFactor['Percentage'].replace(75, '75% or more', inplace=True)
        return InsuranceToExposureFactor

    # Builds RULE 301.C.2.A.1  ZONE-RATED TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    # Returns a dataframe
    @log_exceptions
    def build101A1(self, company):
        Table101A1 = pd.DataFrame(self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A1 = self.buildDataFrame("TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A1['OriginalCostNewVehicles'] = Table101A1['OriginalCostNewVehicles']
        Table101A1 = Table101A1.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A1['Price Bracket'] = Table101A1['PriceBracket']

        Table101A1 = Table101A1.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A1 = Table101A1.sort_values('Price Bracket')
        Table101A1 = Table101A1.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A1 = Table101A1[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A1 = Table101A1.astype(object)
        Table101A1.iloc[:, 1:] = Table101A1.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Table101A1

    # Returns a dataframe RULE 301.C.2.A.2. ZONE-RATED NON-TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    @log_exceptions
    def build101A2(self, company):
        Table101A2 = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A2 = self.buildDataFrame("TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A2['OriginalCostNewVehicles'] = Table101A2['OriginalCostNewVehicles']
        Table101A2 = Table101A2.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A2['Price Bracket'] = Table101A2['PriceBracket']
        Table101A2 = Table101A2.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A2 = Table101A2.sort_values('Price Bracket')
        Table101A2 = Table101A2.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A2 = Table101A2[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A2 = Table101A2.astype(object)
        Table101A2.iloc[:, 1:] = Table101A2.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Table101A2

    # Returns a dataframe RULE 301.C.2.A.3  PRIVATE PASSENGER TYPES VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    @log_exceptions
    def build101A3(self, company):
        Table101A3 = pd.DataFrame(self.rateTables[company]['PrivatePassengerTypesCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerTypesCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A3 = self.buildDataFrame("TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A3['OriginalCostNewVehicles'] = Table101A3['OriginalCostNewVehicles']
        Table101A3 = Table101A3.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A3['Price Bracket'] = Table101A3['PriceBracket']
        Table101A3 = Table101A3.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A3 = Table101A3.sort_values('Price Bracket')
        Table101A3 = Table101A3.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A3 = Table101A3[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A3 = Table101A3.astype(object)
        Table101A3.iloc[:, 1:] = Table101A3.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Table101A3

    # Returns a dataframe RULE 301.C.2.A.4. NON-ZONE-RATED TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    @log_exceptions
    def build101A4(self, company):
        Table101A4 = pd.DataFrame(self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A4 = self.buildDataFrame("PrivatePassengerTypesCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A4['OriginalCostNewVehicles'] = Table101A4['OriginalCostNewVehicles']
        Table101A4 = Table101A4.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A4['Price Bracket'] = Table101A4['PriceBracket']
        Table101A4 = Table101A4.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A4 = Table101A4.sort_values('Price Bracket')
        Table101A4 = Table101A4.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A4 = Table101A4[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A4 = Table101A4.astype(object)
        Table101A4.iloc[:, 1:] = Table101A4.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Table101A4

    # Returns a dataframe RULE 301.C.2.A.5  ALL OTHER VEHICLES VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    @log_exceptions
    def build101A5(self, company):
        Table101A5 = pd.DataFrame(self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A5 = self.buildDataFrame("PrivatePassengerTypesComprehensiveVehicleValueFactorsOCNVehicles_Ext")
        Table101A5['OriginalCostNewVehicles'] = Table101A5['OriginalCostNewVehicles']
        Table101A5 = Table101A5.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A5['Price Bracket'] = Table101A5['PriceBracket']
        Table101A5 = Table101A5.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A5 = Table101A5.sort_values('Price Bracket')
        Table101A5 = Table101A5.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A5 = Table101A5[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A5 = Table101A5.astype(object)
        Table101A5.iloc[:, 1:] = Table101A5.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Table101A5

    # Builds RULE 301.C.2.A.1  ZONE-RATED TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    # Returns a dataframe
    @log_exceptions
    def build101B1(self, company):
        Table101A1 = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A1 = self.buildDataFrame("TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A1['OriginalCostNewVehicles'] = Table101A1['OriginalCostNewVehicles']
        Table101A1 = Table101A1.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A1['Price Bracket'] = Table101A1['PriceBracket']
        Table101A1 = Table101A1.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A1 = Table101A1.sort_values('Price Bracket')
        Table101A1 = Table101A1.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A1 = Table101A1[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A1 = Table101A1.astype(object)
        Table101A1.iloc[:, 1:] = Table101A1.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Table101A1

    # Returns a dataframe RULE 301.C.2.A.2. ZONE-RATED NON-TRAILERS VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    @log_exceptions
    def build101B2(self, company):
        Table101A2 = pd.DataFrame(self.rateTables[company]['PrivatePassengerTypesComprehensiveVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['PrivatePassengerTypesComprehensiveVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A2 = self.buildDataFrame("TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A2['OriginalCostNewVehicles'] = Table101A2['OriginalCostNewVehicles']
        Table101A2 = Table101A2.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A2['Price Bracket'] = Table101A2['PriceBracket']
        Table101A2 = Table101A2.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A2 = Table101A2.sort_values('Price Bracket')
        Table101A2 = Table101A2.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A2 = Table101A2[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A2 = Table101A2.astype(object)
        Table101A2.iloc[:, 1:] = Table101A2.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Table101A2

    # Returns a dataframe RULE 301.C.2.A.3  PRIVATE PASSENGER TYPES VEHICLE VALUE FACTORS - COLLISION WITH ACTUAL CASH VALUE RATING
    @log_exceptions
    def build101B3(self, company):
        Table101A3 = pd.DataFrame(self.rateTables[company]['TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext'][1:], index=None, columns=self.rateTables[company]['TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext'][0])
        #Table101A3 = self.buildDataFrame("TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext")
        Table101A3['OriginalCostNewVehicles'] = Table101A3['OriginalCostNewVehicles']
        Table101A3 = Table101A3.pivot(index=['PriceBracket'], columns='OriginalCostNewVehicles', values='Factor').reset_index(['PriceBracket'])
        Table101A3['Price Bracket'] = Table101A3['PriceBracket']
        Table101A3 = Table101A3.rename(columns={'PriceBracket' : 'Price Bracket (OCN Range)', 1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        Table101A3 = Table101A3.sort_values('Price Bracket')
        Table101A3 = Table101A3.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A3 = Table101A3[['Price Bracket', 'Current Model Year', 'First Preceding Model Year','2nd', '3rd', '4th', '5th', '6th', '7th', '8th', '9th', '10th', '11th', '12th', '13th', '14th', '15th', '16th', '17th', '18th', '19th', '20th', '21st', '22nd', '23rd', '24th', '25th', '26th', '27th and older' ]]
        Table101A3 = Table101A3.astype(object)
        Table101A3.iloc[:, 1:] = Table101A3.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return Table101A3

    def build301D1(self, company):
        # RULE 301.D.1. LIABILITY ORIGINAL COST NEW FACTORS
        # Stole this code from a different function.

        Table101A1 = pd.DataFrame(self.rateTables[company]['LiabilityOriginalCostNewFactor_Ext'][1:], index=None, columns=self.rateTables[company]['LiabilityOriginalCostNewFactor_Ext'][0])
        Table101A1 = Table101A1.pivot(index=['PriceBracket'], columns='VehicleType', values='Factor').reset_index(['PriceBracket'])
        Table101A1['Price Bracket'] = Table101A1['PriceBracket']
        Table101A1 = Table101A1.sort_values('PriceBracket')
        Table101A1 = Table101A1.replace({'Price Bracket' : {0 : "0 to 999", 1000 : "1,000 to 1,999", 2000 : "2,000 to 2,999", 3000 : "3,000 to 3,999", 4000 : "4,000 to 4,999", \
                                                                        5000 : "5,000 to 5,999", 6000 : "6,000 to 7,999", 8000 : "8,000 to 9,999", 10000 : "10,000 to 11,999", 12000 : "12,000 to 13,999", \
                                                                        14000 : "14,000 to 15,999", 16000 : "16,000 to 17,999", 18000 : "18,000 to 19,999", 20000 : "20,000 to 24,999", 25000 : "25,000 to 29,999", \
                                                                        30000 : "30,000 to 34,999", 35000 : "35,000 to 39,999", 40000 : "40,000 to 44,999", 45000 : "45,000 to 49,999", 50000 : "50,000 to 54,999", \
                                                                        55000 : "55,000 to 64,999", 65000 : "65,000 to 74,999", 75000 : "75,000 to 84,999", 85000 : "85,000 to 99,999", 100000 : "100,000 to 114,999", \
                                                                        115000 : "115,000 to 129,999", 130000 : "130,000 to 149,999", 150000 : "150,000 to 174,999", 175000 : "175,000 to 199,999", 200000 : "200,000 to 229,999", \
                                                                        230000 : "230,000 to 259,999", 260000 : "260,000 to 299,999", 300000 : "300,000 to 349,999", 350000 : "350,000 to 399,999", 400000 : "400,000 to 449,999", \
                                                                        450000 : "450,000 to 499,999", 500000 : "500,000 to 599,999", 600000 : "600,000 to 699,999", 700000 : "700,000 to 799,999", 800000 : "800,000 to 899,999", \
                                                                        900000 : "900,000+"}})
        Table101A1 = Table101A1.drop(columns = "PriceBracket")
        Table101A1 = Table101A1[['Price Bracket', 'Extra Heavy Truck-Tractor', 'Extra-Heavy Truck', 'Heavy Truck',
                                   'Heavy Truck-Tractor', 'Light Truck', 'Medium Truck',
                                   'Private Passenger Types', 'Semitrailer', 'Service or Utility Trailer',
                                   'Trailer']]
        Table101A1 = Table101A1.astype(object)
        Table101A1.iloc[:, 1:] = Table101A1.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")
        return Table101A1

    def build301D2(self, company):

        sheet_names = ['LiabilityVehicleAgeFactorsStatedAmountVehicles_Ext']*2
        orig_values = ['Trucks, Tractors, And Trailers','Private Passenger Types']
        replace_values = ['Trucks, Tractors, And Trailers', 'Private Passenger Types']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'All Ages', orig_values,replace_values, filter_values=orig_values)
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def build301D3(self, company):
        # 301.D.2.b Liability Vehicle Age Factors - Original Cost New Vehicles
        # Stole this code from a different function.

        table = pd.DataFrame(self.rateTables[company]['LiabilityVehicleAgeFactorsOriginalCostNewVehicle_Ext'][1:], index=None, columns=self.rateTables[company]['LiabilityVehicleAgeFactorsOriginalCostNewVehicle_Ext'][0])
        table = table.pivot(index=['AgeGroup'], columns='VehicleType', values='Factor').reset_index(['AgeGroup'])
        table = table.sort_values('AgeGroup')
        table["AgeGroup"] = table["AgeGroup"].astype(int)
        table = table.rename(columns={'AgeGroup' : 'Vehicle Age'})
        table["Vehicle Age"] = table["Vehicle Age"].replace({1 : 'Current Model Year', 2 : 'First Preceding Model Year', 3 : '2nd', 4 : '3rd', 5 : '4th', 6 : '5th', 7 : '6th', 8 : '7th', 9 : '8th', 10 : '9th', 11 : '10th', 12 : '11th', 13 : '12th', 14 : '13th', 15 : '14th', 16 : '15th', 17 : '16th', 18 : '17th', 19 : '18th', 20 : '19th', 21 : '20th', 22 : '21st', 23 : '22nd', 24 : '23rd', 25 : '24th', 26 : '25th', 27 : '26th', 28 : '27th and older'})
        table = table.astype(object)
        table.iloc[:, 1:] = table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return table

    # Builds RULE RULE 107. FELLOW EMPLOYEE COVERAGE
    # Returns a dataframe
    @log_exceptions
    def buildFellowEmployeeFactors(self, company):
        FellowEmployeeAllFactors = pd.DataFrame(self.rateTables[company]['FellowEmployeeBaseRate_v2_Ext'][1:], index=None, columns=self.rateTables[company]['FellowEmployeeBaseRate_v2_Ext'][0]).rename(columns={'Premium': 'All Employees'})
        #FellowEmployeeAllFactors = self.buildDataFrame("FellowEmployeeBaseRate_v2_Ext").rename(columns={'Premium': 'All Employees'})
        FellowEmployeeScheduleFactors = pd.DataFrame(self.rateTables[company]['FellowEmployeeCoverageForDesignatedEmployeesPositionsBaseRate_v2_Ext'][1:], index=None, columns=self.rateTables[company]['FellowEmployeeCoverageForDesignatedEmployeesPositionsBaseRate_v2_Ext'][0]).rename(columns={'Premium': 'Scheduled Employees'})
        #FellowEmployeeScheduleFactors = self.buildDataFrame("FellowEmployeeCoverageForDesignatedEmployeesPositionsBaseRate_v2_Ext").rename(columns={'Premium': 'Scheduled Employees'})
        FellowEmployeeAllFactors.loc[FellowEmployeeAllFactors['BusinessAutoProtectionType'] == 'Platinum', 'All Employees'] = 'N/A'
        FellowEmployeeScheduleFactors.loc[FellowEmployeeScheduleFactors['BusinessAutoProtectionType'] == 'Platinum', 'Scheduled Employees'] = 'N/A'
        FellowEmployeeFactors = pd.merge(FellowEmployeeAllFactors, FellowEmployeeScheduleFactors,on='BusinessAutoProtectionType', how='inner').rename(columns={'BusinessAutoProtectionType': 'Option'})
        FellowEmployeeFactors['Option'].replace('Gold', 'Auto Protection Gold', inplace=True)
        FellowEmployeeFactors['Option'].replace('None', 'Without Auto Protection', inplace=True)
        FellowEmployeeFactors['Option'].replace('Plus', 'Auto Protection Plus', inplace=True)

        # Applying Option ranking
        category_order = ['Without Auto Protection', 'Auto Protection Plus', 'Auto Protection Gold', 'Platinum']
        FellowEmployeeFactors['Option'] = pd.Categorical(FellowEmployeeFactors['Option'], categories=category_order, ordered=True)
        FellowEmployeeFactors = FellowEmployeeFactors.sort_values('Option')
        FellowEmployeeFactors = FellowEmployeeFactors.astype(object)
        FellowEmployeeFactors.iloc[:3, 1:] = FellowEmployeeFactors.iloc[:3, 1:].map(lambda x: f"{x:.2f}")

        return FellowEmployeeFactors

    # Builds RULE 310. LOSS OF USE EXPENSES - RENTAL VEHICLES - OPTIONAL LIMITS
    # Returns a dataframe
    @log_exceptions
    def buildLossofUseFactors(self, company):
        LossofUseFactors = pd.DataFrame(self.rateTables[company]['OptionalLimitsLossofUseExpensesBasePremium_Ext'][1:], index=None, columns=self.rateTables[company]['OptionalLimitsLossofUseExpensesBasePremium_Ext'][0]).rename(columns={'Amount Per Day/ Maximum': 'Limit','Base Premium': 'Premium'})
        #LossofUseFactors = self.buildDataFrame("OptionalLimitsLossofUseExpensesBasePremium_Ext").rename(columns={'Amount Per Day/ Maximum': 'Limit','Base Premium': 'Premium'})
        return LossofUseFactors

    # Builds blank dataframe
    # Returns a dataframe
    def buildBlank(self):
        Blank = pd.DataFrame()
        return Blank

    # Builds RULE 125. WAIVER OF TRANSFER OF RIGHTS OF RECOVERY AGAINST OTHERS TO US
    # Returns a dataframe
    @log_exceptions
    def buildWaiver1Factors(self, company):
        Waiver1Factors = pd.DataFrame(self.rateTables[company]['WaiverofSubrogationBlanket_Ext'][1:], index=None, columns=self.rateTables[company]['WaiverofSubrogationBlanket_Ext'][0]).drop(columns='Constant') \
            .rename(columns={'Premium': 'Premium charge per policy term'})
        #Waiver1Factors = self.buildDataFrame("WaiverofSubrogationBlanket_Ext").drop(columns='Constant')\
        return Waiver1Factors

    @log_exceptions
    def buildWaiver2Factors(self, company):
        Waiver2Factors = pd.DataFrame(self.rateTables[company]['Waiver_of_Subrogation_Ext'][1:], index=None, columns=self.rateTables[company]['Waiver_of_Subrogation_Ext'][0]).drop(columns='Constant') \
            .rename(columns={'Premium': 'Premium charge per person or organization per policy term'})
        #Waiver2Factors = self.buildDataFrame("Waiver_of_Subrogation_Ext").drop(columns='Constant')\
        return Waiver2Factors

    # Builds RULE 127. ORIGINAL EQUIPMENT MANUFACTURER PARTS COVERAGE
    # Returns a dataframe
    @log_exceptions
    def buildOriginalFactors(self, company):
        OriginalFactors = pd.DataFrame(self.rateTables[company]['OriginalEquipmentManufacturerPartsCoverageFactor_Ext'][1:], index=None, columns=self.rateTables[company]['OriginalEquipmentManufacturerPartsCoverageFactor_Ext'][0]).drop(columns='Constant') \
            .rename(columns={'Factor': 'Multiply the Physical Damage Coverage Premiums by the Following Factor:'})
        #OriginalFactors = self.buildDataFrame("OriginalEquipmentManufacturerPartsCoverageFactor_Ext").drop(columns='Constant')\
        return OriginalFactors

    # Builds RULE 100. ILF
    # Returns a dataframe
    @log_exceptions
    def buildILF(self, company):
        ILF = pd.DataFrame(self.rateTables[company]['IncreasedLimitFactorText'][1:], index=None, columns=self.rateTables[company]['IncreasedLimitFactorText'][0]).rename(columns={'LiabilityLimitText': 'Combined Single Limit of Liability'})

        limits_list = pd.read_excel(BA_INPUT_FILE, sheet_name="CSL Limits")

        limits_list = limits_list[self.StateAbb].dropna().astype(float).apply(lambda x: f"{x:,.0f}").values

        #ILF = self.buildDataFrame("IncreasedLimitFactorText").rename(columns={'LiabilityLimitText': 'Combined Single Limit of Liability'})
        ILF = ILF.pivot(index=['Combined Single Limit of Liability'],columns='IncreaseLiabilityClass', values='Factor').reset_index(['Combined Single Limit of Liability'])
        ILF = ILF[['Combined Single Limit of Liability','Light And Medium Trucks', 'Heavy Trucks And Truck-Tractors', 'Extra-Heavy Trucks And Truck-Tractors','Trucks, Tractors, And Trailers Zone Rated','All Other Risks']]
        ILF = ILF.sort_values(by=['Light And Medium Trucks'])

        ILF = ILF[ILF['Combined Single Limit of Liability'].isin(limits_list)]

        ILF = ILF.astype(object)
        ILF.iloc[:, 1:] = ILF.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return ILF


    # Builds RULE 288. DRIVE OTHER CAR
    # Returns a dataframe
    @log_exceptions
    def buildDriveOtherFactors(self, company):
        DriveOtherFactorsLiab = pd.DataFrame(self.rateTables[company]['DriveOtherCarLiabilityFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarLiabilityFactor'][0])
        DriveOtherFactorsLiab['Constant'] = "Liability"
        DriveOtherFactorsMed = pd.DataFrame(self.rateTables[company]['DriveOtherCarMedicalPaymentsFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarMedicalPaymentsFactor'][0])
        DriveOtherFactorsMed['Constant'] = "Medical"
        DriveOtherFactorsPIP = pd.DataFrame(self.rateTables[company]['DriveOtherCarLiabilityFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarLiabilityFactor'][0])
        DriveOtherFactorsPIP['Constant'] = "PIP"
        DriveOtherFactorsOTC = pd.DataFrame(self.rateTables[company]['DriveOtherCarOtherThanCollisionFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarOtherThanCollisionFactor'][0])
        DriveOtherFactorsOTC['Constant'] = "Other than Collision"
        DriveOtherFactorsColl = pd.DataFrame(self.rateTables[company]['DriveOtherCarCollisionFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarCollisionFactor'][0])
        DriveOtherFactorsColl['Constant'] = "Collision"
        DriveOtherFactorsUM = pd.DataFrame(self.rateTables[company]['DriveOtherCarUninsuredMotoristFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarUninsuredMotoristFactor'][0])
        DriveOtherFactorsUM['Constant'] = "UM"
        DriveOtherFactorsUIM = pd.DataFrame(self.rateTables[company]['DriveOtherCarUnderinsuredMotoristFactor'][1:], index=None, columns=self.rateTables[company]['DriveOtherCarUnderinsuredMotoristFactor'][0])
        DriveOtherFactorsUIM['Constant'] = "UIM"
        DriveOtherFactors = pd.concat([DriveOtherFactorsLiab, DriveOtherFactorsMed,DriveOtherFactorsPIP,DriveOtherFactorsOTC,DriveOtherFactorsColl,DriveOtherFactorsUM,DriveOtherFactorsUIM])

        DriveOtherFactors = DriveOtherFactors.rename(columns={'Constant':'Coverage','Factor':'Per Named'})

        # Filtering state specific coverages
        um_info = pd.read_excel(BA_INPUT_FILE, sheet_name = "297 Map")
        um_info = um_info[um_info["State"] == self.StateAbb]
        combined_text = um_info.astype(str).apply(lambda x: ' '.join(x), axis=1).str.cat(sep=' ').lower()

        UM_flag = "uninsured" in combined_text
        UIM_flag = "underinsured" in combined_text
        inlcudes_underinsured = "(includes underinsured)" in combined_text

        if not UM_flag:
            DriveOtherFactors = DriveOtherFactors[DriveOtherFactors["Coverage"] != "UM"]
        if not UIM_flag  or inlcudes_underinsured:
            DriveOtherFactors = DriveOtherFactors[DriveOtherFactors["Coverage"] != "UIM"]
        if self.StateAbb not in self.pip_states.values:
            DriveOtherFactors = DriveOtherFactors[DriveOtherFactors["Coverage"] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            DriveOtherFactors = DriveOtherFactors[DriveOtherFactors["Coverage"] != "Medical"]

        DriveOtherFactors = DriveOtherFactors.astype(object)
        DriveOtherFactors.iloc[:,1:] = DriveOtherFactors.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return DriveOtherFactors

    # Builds RULE 294. RENTAL REIMBURSEMENT
    # Returns a dataframe
    @log_exceptions
    def buildRentalFactors(self, company):
        RentalFactors = pd.DataFrame(self.rateTables[company]['RentalReimbursementFactor'][1:], index=None, columns=self.rateTables[company]['RentalReimbursementFactor'][0])
        RentalFactors.iloc[:,1] = RentalFactors.iloc[:,1]
        RentalFactors = RentalFactors.rename(columns={'CoverageIndicator': '','Factor': ''})
        RentalFactors = RentalFactors.astype(object)
        RentalFactors.iloc[:,1:] = RentalFactors.iloc[:,1:].map(lambda x: f"{x:.2f}")

        return RentalFactors


    # Builds RULE 295. AUDIO, VISUAL, AND DATA ELECTRONIC EQUIPMENT
    # Returns a dataframe
    @log_exceptions
    def buildAudioFactors(self, company):
        AudioFactors = pd.DataFrame(self.rateTables[company]['AudioVisualDataEquipmentBasePremium2'][1:], index=None, columns=self.rateTables[company]['AudioVisualDataEquipmentBasePremium2'][0])
        #AudioFactors = self.buildDataFrame("AudioVisualDataEquipmentBasePremium2")
        AudioFactors = AudioFactors.rename(columns={'CostOfNew': 'Cost of Equipment', 'Premium': 'Rate'}) \
            .sort_values(by=['Rate'])
        AudioFactors = AudioFactors.astype(object)
        AudioFactors.iloc[:, 1:] = AudioFactors.iloc[:, 1:].map(lambda x: f"{x:.0f}")

        return AudioFactors

    # Builds RULE 296. TAPES, RECORDS AND DISCS COVERAGE
    # Returns a dataframe
    @log_exceptions
    def buildTapeFactors(self, company):
        TapeFactors = pd.DataFrame(self.rateTables[company]['TapesRecordsAndDiscsBasePremium'][1:], index=None, columns=self.rateTables[company]['TapesRecordsAndDiscsBasePremium'][0])
        #TapeFactors = self.buildDataFrame("TapesRecordsAndDiscsBasePremium")
        TapeFactors = TapeFactors.rename(columns={'BasePremium': 'Premium per auto:'}).drop(columns='Constant')
        TapeFactors = TapeFactors.astype(object)
        TapeFactors.iloc[:, 1:] = TapeFactors.iloc[:, 1:].map(lambda x: f"{x:.2f}")

        return TapeFactors

    # Builds RULE 103. POLLUTION LIABILITY
    # Returns a dataframe
    @log_exceptions
    def buildPollutionFactors(self, company):
        PollutionFactors = pd.DataFrame(self.rateTables[company]['PollutionLiabilityRate_Ext'][1:], index=None, columns=self.rateTables[company]['PollutionLiabilityRate_Ext'][0]).rename(columns={'Rate':'Factor'})
        #PollutionFactors = self.buildDataFrame("PollutionLiabilityRate_Ext").rename(columns={'Rate':'Factor'})
        PollutionMinFactors = pd.DataFrame(self.rateTables[company]['PollutionLiabilityMinimumPremium_Ext'][1:], index=None, columns=self.rateTables[company]['PollutionLiabilityMinimumPremium_Ext'][0]).rename(columns={'MinimumPremium': 'Minimum Premium'})
        PollutionFactors = pd.merge(PollutionFactors, PollutionMinFactors, on='Hazard Grade', how='inner')
        PollutionFactors = PollutionFactors.sort_values(by=['Factor'])
        return PollutionFactors

    # Builds RULE 126. BUSINESS AUTO PROTECTION ENDORSEMENTS
    # Returns a dataframe
    @log_exceptions
    def buildProtectionFactors(self, company):
        ProtectionFactors = pd.DataFrame(self.rateTables[company]['BusinessAutoProtectionFactor_Ext'][1:], index=None, columns=self.rateTables[company]['BusinessAutoProtectionFactor_Ext'][0]).rename(columns={'Business Auto Protection Type': 'SubTypeCode','Factor': 'Premium'})
        ProtectionFactors['PremiumType']= "Multiply the Developed Policy Premium by the Following Factor:"
        ProtectionMinFactors = pd.DataFrame(self.rateTables[company]['MiscellaneousMinimumMaximumPremium_Ext'][1:], index=None, columns=self.rateTables[company]['MiscellaneousMinimumMaximumPremium_Ext'][0]).query(f'CoverageType == "CA7BusinessAutoProtection_Ext"').drop(columns='CoverageType')
        ProtectionMinFactors['PremiumType'].replace('Minimum', 'Minimum Premium Per Policy:', inplace=True)
        ProtectionMinFactors['PremiumType'].replace('Maximum', 'Maximum Premium Per Policy:', inplace=True)
        ProtectionFactors = pd.concat([ProtectionFactors, ProtectionMinFactors])
        ProtectionFactors = ProtectionFactors.pivot(index=['PremiumType'],columns='SubTypeCode', values='Premium').reset_index(['PremiumType'])
        # Florida only offers gold
        if self.StateAbb != "FL":
            ProtectionFactors = ProtectionFactors[['PremiumType', 'Plus', 'Gold','Platinum']]
            ProtectionFactors = ProtectionFactors.sort_values(by=['Plus'])
            ProtectionFactors = ProtectionFactors.rename(columns={'PremiumType': ''})
        elif self.StateAbb == "FL":
            ProtectionFactors = ProtectionFactors[['PremiumType', 'Gold']]
            ProtectionFactors = ProtectionFactors.sort_values(by=['Gold'])
            ProtectionFactors = ProtectionFactors.rename(columns={'PremiumType': ''})
        return ProtectionFactors

    # Builds RULE 80. SNOWMOBILES
    # Returns a dataframe
    @log_exceptions
    def buildSnowMobileFactors(self, company):
        SnowMobileLiabFactors = pd.DataFrame(self.rateTables[company]['SpecialTypesSnowmobileLiabilityBasePremium'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSnowmobileLiabilityBasePremium'][0]).query(f'SpecialTypesSupplementaryType == "Snowmobile With Passenger Hazard"'). \
            rename(columns={'SpecialTypesSupplementaryType': 'Coverage','BasePremium': 'Premium'})
        SnowMobileLiabFactors['Coverage'] = "Liability"
        SnowMobileLiabFactors.iloc[:,1] = SnowMobileLiabFactors.iloc[:,1]
        SnowMobileMedFactors = pd.DataFrame(self.rateTables[company]['SpecialTypesSnowmobileMedicalPaymentsRate'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSnowmobileMedicalPaymentsRate'][0]).rename(columns={'Constant': 'Coverage'})
        SnowMobileMedFactors['Coverage'] = "Medical"
        SnowMobileMedFactors.iloc[:,1] = SnowMobileMedFactors.iloc[:,1]
        SnowMobileOTCFactors = pd.DataFrame(self.rateTables[company]['SpecialTypesSnowmobileOtherThanCollisionRate'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSnowmobileOtherThanCollisionRate'][0]).query(f'SpecialTypesOtherThanCollisionCoverageType == "Stated Amount - With Falling Through Ice"'). \
            rename(columns={'SpecialTypesOtherThanCollisionCoverageType': 'Coverage','Factor': 'Premium'})
        SnowMobileOTCFactors['Coverage'] = "Comprehensive (Per $100 of Ins.)"
        SnowMobileCollFactors = pd.DataFrame(self.rateTables[company]['SpecialTypesSnowmobileCollisionRate'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSnowmobileCollisionRate'][0]).rename(columns={'Constant': 'Coverage','Factor': 'Premium'})
        SnowMobileCollFactors['Coverage'] = "Collision (Per $100 of Ins.)"
        SnowMobileCollFactors.iloc[:,1] = SnowMobileCollFactors.iloc[:,1]
        SnowMobileOtherFactors = pd.DataFrame(self.rateTables[company]['SpecialTypesSnowmobileCollisionRate'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSnowmobileCollisionRate'][0]).rename(columns={'Constant': 'Coverage','Factor': 'Premium'})
        SnowMobileOtherFactors['Coverage'] = "All Other Coverages"
        SnowMobileOtherFactors['Premium'] = "Use PPT Rates"
        SnowMobileFactors = pd.concat([SnowMobileLiabFactors, SnowMobileMedFactors,SnowMobileOTCFactors,SnowMobileCollFactors,SnowMobileOtherFactors])

        SnowMobileFactors = SnowMobileFactors.astype(object)
        SnowMobileFactors.iloc[:-1,1:] = SnowMobileFactors.iloc[:-1,1:].astype(float).map(lambda x: f"{x:.2f}")

        return SnowMobileFactors

    # Iso Currency 281
    def buildMobileandFarmPremiumFactor(self, company):
        # 281.C.2. Premium Computation
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesSpecialEquipmentFactor1'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSpecialEquipmentFactor1'][0])
        mobile_code = "7996"
        farm_code = "7993"

        mobile_table = table[
            (table["SpecialTypesSupplementaryType1"] == "Specified Auto Basis") &
            (table["SpecialTypesClassCode"] == mobile_code) &
            (~table["TypeCoverage"].isin(["PhysDamCollSA", "PhysDamOTCSA"]))
            ].drop(columns = ["SpecialTypesSupplementaryType1","SpecialTypesClassCode"])

        farm_table = table[
            (table["SpecialTypesSupplementaryType1"] == "Specified Auto Basis") &
            (table["SpecialTypesClassCode"] == farm_code) &
            (~table["TypeCoverage"].isin(["PhysDamCollSA", "PhysDamOTCSA"]))
            ].drop(columns = ["SpecialTypesSupplementaryType1","SpecialTypesClassCode"])

        mobile_table.replace({
            "MedicalPayments": "Medical",
            "PhysDamCollACV": "Collision",
            "PhysDamOTCACV": "Other than Collision",
            "No-Fault": "PIP"
        }, inplace = True)

        farm_table.replace({
            "MedicalPayments": "Medical",
            "PhysDamCollACV": "Collision",
            "PhysDamOTCACV": "Other than Collision",
            "No-Fault": "PIP"
        }, inplace = True)

        output_table = pd.merge(mobile_table, farm_table, on="TypeCoverage", how="left")
        output_table.columns = ["Coverage", "Mobile Equipment", "Farm Equipment"]
        if self.StateAbb not in self.pip_states.values:
            output_table = output_table[output_table['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            output_table = output_table[output_table['Coverage'] != "Medical"]
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildMobileandFarmCostOfHireFactor(self, company):
        # 281.D.2.b. Cost of Hire Basis Coverage Factors
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesSpecialEquipmentFactor1'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSpecialEquipmentFactor1'][0])
        mobile_code = "7994"
        farm_code = "7991"

        mobile_table = table[
            (table["SpecialTypesSupplementaryType1"] == "Cost Of Hire Basis") &
            (table["SpecialTypesClassCode"] == mobile_code) &
            (~table["TypeCoverage"].isin(["PhysDamCollSA", "PhysDamOTCSA"]))
            ].drop(columns = ["SpecialTypesSupplementaryType1","SpecialTypesClassCode"])

        farm_table = table[
            (table["SpecialTypesSupplementaryType1"] == "Cost Of Hire Basis") &
            (table["SpecialTypesClassCode"] == farm_code) &
            (~table["TypeCoverage"].isin(["PhysDamCollSA", "PhysDamOTCSA"]))
            ].drop(columns = ["SpecialTypesSupplementaryType1","SpecialTypesClassCode"])

        mobile_table.replace({
            "MedicalPayments": "Medical",
            "PhysDamCollACV": "Collision",
            "PhysDamOTCACV": "Other than Collision",
            "No-Fault": "PIP"
        }, inplace = True)

        farm_table.replace({
            "MedicalPayments": "Medical",
            "PhysDamCollACV": "Collision",
            "PhysDamOTCACV": "Other than Collision",
            "No-Fault": "PIP"
        }, inplace = True)

        output_table = pd.merge(mobile_table, farm_table, on="TypeCoverage", how="left")
        output_table.columns = ["Coverage", "Mobile Equipment", "Farm Equipment"]
        if self.StateAbb not in self.pip_states.values:
            output_table = output_table[output_table['Coverage'] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            output_table = output_table[output_table['Coverage'] != "Medical"]

        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildRentalBasisFactors(self, company):
        # 281.E Rental Period Basis Factors
        table = pd.DataFrame(self.rateTables[company]['SpecialTypesSpecialEquipmentFactor1'][1:], index=None, columns=self.rateTables[company]['SpecialTypesSpecialEquipmentFactor1'][0])
        mobile_code = "7995"
        farm_code = "7992"

        mobile_table = table[
            (table["SpecialTypesSupplementaryType1"] == "Rental Period Basis") &
            (table["SpecialTypesClassCode"] == mobile_code) &
            (table["TypeCoverage"].isin(["Liability"]))
            ].drop(columns = ["SpecialTypesSupplementaryType1","SpecialTypesClassCode"])

        farm_table = table[
            (table["SpecialTypesSupplementaryType1"] == "Rental Period Basis") &
            (table["SpecialTypesClassCode"] == farm_code) &
            (table["TypeCoverage"].isin(["Liability"]))
            ].drop(columns = ["SpecialTypesSupplementaryType1","SpecialTypesClassCode"])

        output_table = pd.merge(mobile_table, farm_table, on="TypeCoverage", how="left")
        output_table.columns = ["Coverage", "Mobile Equipment", "Farm Equipment"]

        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Builds RULE 117. Schedule Rating
    # Returns a dataframe
    @log_exceptions
    def buildRule117(self, company):
        Rule117 = pd.DataFrame(self.rateTables[company]['ScheduleEligibility_Ext'][1:], index=None, columns=self.rateTables[company]['ScheduleEligibility_Ext'][0])
        #Rule117 = self.buildDataFrame("ScheduleEligibility_Ext")
        Rule117 = Rule117.query(f'Constant == "Y"').filter(items=['ScheduleEligibleAmount']).rename(columns={'ScheduleEligibleAmount' : 'Minimum Eligible Premium'})
        return Rule117

    # Builds the Rate Capping table
    # Returns a dataframe
    @log_exceptions
    def buildTelematicsFactors(self, company):
        TelematicsFactors1 = pd.DataFrame(self.rateTables[company]['TelematicsFactor'][1:], index=None, columns=self.rateTables[company]['TelematicsFactor'][0]) \
            .rename(columns={'Participation Policy': 'ParticipationPolicy','Participation Vehicle': 'ParticipationVehicle','Coverage Pattern Code': 'CoveragePatternCode','Policy Term Min': 'PolicyTermMin'})
        #TelematicsFactors1 = self.buildDataFrame("TelematicsFactor")\
        TelematicsFactors1= TelematicsFactors1.query(f'ParticipationPolicy == "Yes" & ParticipationVehicle == "Yes" & CoveragePatternCode == "CA7VehicleLiabTTT" & PolicyTermMin == 1')
        TelematicsFactors = TelematicsFactors1[["Telematics Factor"]]
        TelematicsFactors["Telematics Factor"] = 1 - TelematicsFactors["Telematics Factor"]
        TelematicsFactors = TelematicsFactors.rename(columns={"Telematics Factor": ""})
        return TelematicsFactors

    # Builds RULE T1- Telematics
    @log_exceptions
    def buildRule451(self, company):
        sm_mig_list = ["NICOF","NACO","NAFF","CCMIC","HICNJ"]
        sm_run = ["NGIC"]
        mm_mig_list = ["NICOA","AICOA","NPCIC"]
        mm_run = ["NMIC"]

        capping_data = pd.DataFrame(self.rateTables[company]['RateCappingPremiumRange 2_Ext'][1:], index=None, columns=self.rateTables[company]['RateCappingPremiumRange 2_Ext'][0])

        if company in sm_mig_list:
            capping_data = capping_data[(capping_data['RateCapType'] == 'Migration') & (capping_data['FiledRatesReachedIndicator'] == False)]
        elif company in sm_run:
            capping_data = capping_data[(capping_data['RateCapType'] == 'Run') & (capping_data['FiledRatesReachedIndicator'] == False)]
        elif company in mm_mig_list:
            capping_data = capping_data[(capping_data['RateCapType'] == 'MMMigration') & (capping_data['FiledRatesReachedIndicator'] == False)]
        elif company in mm_run:
            capping_data = capping_data[(capping_data['RateCapType'] == 'MMRun') & (capping_data['FiledRatesReachedIndicator'] == False)]

        capping_data = capping_data.rename(columns={"MaximumRange": "Maximum", "MinimumRange": "Minimum"})
        capping_data = capping_data.loc[capping_data['YrsOnPCMax'].idxmin()] # Grab the minimum YrsonPCMax, not comprehensive but when would we ever use the other functionality...

        # Apply rounding and replacement logic
        capping_data['Minimum'] = round(-1 * capping_data['Maximum'], 2) if round(capping_data['Minimum'], 2) != -999.00 else 'Not Applicable'
        capping_data['Maximum'] = round(capping_data['Maximum'], 2) if round(capping_data['Maximum'], 2) != 999.00 else 'Not Applicable'

        capping_data = pd.DataFrame(capping_data[["Maximum", "Minimum"]]).reset_index()
        capping_data.columns = ["",""]

        return capping_data

    def buildRule451Renewals(self, company):
        sm_mig_list = ["NICOF","NACO","NAFF","CCMIC","HICNJ"]
        sm_run = ["NGIC"]
        mm_mig_list = ["NICOA","AICOA","NPCIC"]
        mm_run = ["NMIC"]

        capping_data = pd.DataFrame(self.rateTables[company]['RateCappingPremiumRange 2_Ext'][1:], index=None, columns=self.rateTables[company]['RateCappingPremiumRange 2_Ext'][0])

        if company in sm_mig_list:
            capping_data = capping_data[(capping_data['RateCapType'] == 'Migration') & (capping_data['FiledRatesReachedIndicator'] == False)]
        elif company in sm_run:
            capping_data = capping_data[(capping_data['RateCapType'] == 'Run') & (capping_data['FiledRatesReachedIndicator'] == False)]
        elif company in mm_mig_list:
            capping_data = capping_data[(capping_data['RateCapType'] == 'MMMigration') & (capping_data['FiledRatesReachedIndicator'] == False)]
        elif company in mm_run:
            capping_data = capping_data[(capping_data['RateCapType'] == 'MMRun') & (capping_data['FiledRatesReachedIndicator'] == False)]

        min_value = capping_data['YrsOnPCMax'].min() # Always 1 less than value
        if min_value == 999:
            min_value = "Not Applicable"
        else:
            min_value = min_value - 1

        # Creating a new DataFrame with empty column names
        capping_data = pd.DataFrame({"Col1" : ["Renewals"], "Col2": [min_value]})

        # Assign empty column names
        capping_data.columns = ["", ""]

        return capping_data

    # Returns a dataframe

    # Builds RULE DP-1. Distribution Plan
    # Returns a dataframe
    @log_exceptions
    def buildDP1(self, company):
        Dp1 = pd.DataFrame(self.rateTables[company]['DistributionFactor_Ext'][1:], index=None, columns=self.rateTables[company]['DistributionFactor_Ext'][0]) \
            .rename(columns={'DistributionGroup': 'Distribution Group','Factor': 'Distribution Factor'})
        #Dp1 = self.buildDataFrame("DistributionFactor_Ext")\
        Dp1= Dp1.query(f'`Distribution Group` != "DG99"').replace({'Distribution Group' : {'DG00' : 'DistributionGroup0', 'DG01' : 'DistributionGroup1', 'DG02' : 'DistributionGroup2', 'DG03' : 'DistributionGroup3', 'DG04' : 'DistributionGroup4', 'DG05' : 'DistributionGroup5', 'DG06' : 'DistributionGroup6', 'DG07' : 'DistributionGroup7', 'DG08' : 'DistributionGroup8', 'DG09' : 'DistributionGroup9', 'DG10' : 'DistributionGroup10', 'DG11' : 'DistributionGroup11', 'DG12' : 'DistributionGroup12', 'DG13' : 'DistributionGroup13', 'DG14' : 'DistributionGroup14', 'DG15' : 'DistributionGroup15'}})
        return Dp1

    @log_exceptions
    def build293Table(self, company, rule_order : int):
        """
        Builds the tables for rule 293.

        Description:
        The build293Table method generates tables for rule 293 based on the state abbreviation and rule order.
        Often uses a builder function simple_long_table_build to help build the tables.
        Was made before the input file was created. Could have been nicer with the input file but it's not worth changing.

        Parameters:
        - company (str): The name of the company for which the rate tables are being processed.
        - rule_order (int): The order number of the rule, which matches the order the table appears in the template file.

        Returns:
        - output_table (pd.DataFrame): A DataFrame containing the table for the specified rule order.

        :param company: The name of the company for which the rate tables are being processed.
        :param rule_order: The order number of the rule, which matches the order the table appears in the review file.
        :return: A DataFrame containing the table for the specified rule order.
        """

        output_table = None # Table that is to be returned at the end of the function.
        sheet_names = None

        if self.StateAbb == "AR":
            # 293.B.1. Accidental Death Benefits

            # Had too much difficulty making it look/load well with the helper
            if rule_order == 1:
                sheet_names = ['SpecialTypesMotorcycleDeathBenefitsBasePremium', 'DeathBenefitsBasePremium']
                table_1 = pd.DataFrame(self.rateTables[company][sheet_names[0]][1:],columns=self.rateTables[company][sheet_names[0]][0])
                table_2 = pd.DataFrame(self.rateTables[company][sheet_names[1]][1:],columns=self.rateTables[company][sheet_names[1]][0])
                output_table = pd.DataFrame({
                    "Principal Sum": ["$5,000"],
                    "Motorcycles": [table_1.iloc[0, 1]],
                    "All Others": [table_2.iloc[0, 1]]
                })

            elif rule_order == 2:
                sheet_names = ['SpecialTypesMotorcycleWorkLossCoverageBasePremium', 'WorkLossCoverageBasePremium']
                table_1 = pd.DataFrame(self.rateTables[company][sheet_names[0]][1:],columns=self.rateTables[company][sheet_names[0]][0])
                table_2 = pd.DataFrame(self.rateTables[company][sheet_names[1]][1:],columns=self.rateTables[company][sheet_names[1]][0])
                output_table = pd.DataFrame({
                    "Income Earner": ["$140/week max"],
                    "Non-Income Earner": ["$70/week max"],
                    "Motorcycles": [table_1.iloc[0, 1]],
                    "All Others": [table_2.iloc[0, 1]]
                })

        elif self.StateAbb == "DE":

            # 293.B.1. Single Limit Personal Injury Protection Factor
            if rule_order == 1:
                sheet_names = ['SingleLimitPIPFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.B.2. No-Fault Factors
            elif rule_order == 2:
                sheet_names = ['NoFaultPIPFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.C.2. Single Limit Added Personal Injury Protection Factors
            elif rule_order == 3:
                sheet_names = ['AddedPIPLimitsFactor'] * 3
                orig_values = ['20,000 Added PIP', '70,000 Added PIP', '270,000 Added PIP']
                replace_values = ['20,000', '70,000', '270,000']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Additional PIP Limits', orig_values, replace_values, filter_values = orig_values)

        elif self.StateAbb == "DC":
            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultPIPMedicalExpenseBenefitsFactor',
                               'NoFaultPIPMedicalExpenseBenefitsFactor',
                               'NoFaultPIPWorkLossBenefitsFactor',
                               'NoFaultPIPWorkLossBenefitsFactor',
                               'NoFaultPIPFuneralExpenseBenefitsFactor']
                orig_values = ['50,000',
                               '100,000',
                               '12,000',
                               '24,000',
                               '4,000']
                replace_values = ['$50,000 Medical Expense',
                               '$100,000 Medical Expense',
                               '$12,000 Wage Loss',
                               '$24,000 Wage Loss',
                               '$4,000 Funeral Expense']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

        elif self.StateAbb == "FL":

            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor']
                orig_values = ['Y', 'Y']
                replace_values = ['$100,000 Liability', 'Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.B.3. Autos Not Provided Personal Injury Protection Factors
            elif rule_order == 2:
                sheet_names = ['NoPIPLiabilityFactorFL', 'NoPIPLiabilityFactorFL']
                orig_values = ['Private Passenger', 'All Other']
                replace_values = ['PPT and Related Classes', 'All Other']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.C.2.a. Exclusion Of Disability Benefits Only Factors
            elif rule_order == 3:
                sheet_names = ['PIPWorkLossExclusionFactor', 'PIPWorkLossExclusionFactor']
                orig_values = ['Exclude Work Loss - Named Insured Only', 'Exclude Work Loss - Named Insured and Resident Relative']
                replace_values = ['Applying Only to the Named Insured', 'Applying to the Named Insured and Resident Relative']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.D.2.a. Extended Personal Injury Protection - Work Loss Included Factor
            elif rule_order == 4:
                sheet_names = ['PIPExtendedPIPFactor']
                orig_values = ['Include Work Loss Extended']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.D.2.b. Extended Personal Injury Protection - Work Loss Excluded Factor
            elif rule_order == 5:
                sheet_names = ['PIPExtendedPIPFactor']
                orig_values = ['Exclude Work Loss - Named Insured and Resident Relative Extended']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.E.2. Added Personal Injury Protection
            elif rule_order == 6:
                output_table = pd.DataFrame(self.rateTables[company]["AddedPIPPremiumFL"][1:], index=None, columns=self.rateTables[company]["AddedPIPPremiumFL"][0])
                output_table = output_table.pivot_table(values = "Premium", index = "AddedPIPLimitText", columns = "PIPWorkLoss")
                output_table.drop(columns = "Exclude Work Loss - Named Insured Only", inplace = True)
                output_table.reset_index(inplace=True)
                output_table.rename(columns = {"Exclude Work Loss - Named Insured and Resident Relative" : "Excluding Disability Benefits",
                                               "Include Work Loss" : "Including Disability",
                                               "AddedPIPLimitText" : "Limit"}, inplace = True)
                output_table.columns = pd.MultiIndex.from_tuples([("Rate Per Auto", col) for col in output_table.columns])


            # 293.F.1. Broadened Personal Injury Protection For Named Individuals
            elif rule_order == 7:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.F.2. Extended Personal Injury Protection For Named Individuals Factor
            elif rule_order == 8:
                sheet_names = ['BroadenedPIPExtendedPIPFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

        elif self.StateAbb == "KS":

            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor']
                orig_values = ['Y', 'Y']
                replace_values = ['$100,000 Liability', 'Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.B.3. Autos Not Provided Personal Injury Protection Factors
            elif rule_order == 2:
                sheet_names = ['NonNoFaultBusesVehicleTypesFactor','NonNoFaultAllOtherVehicleTypesFactor']
                orig_values = ['Y', 'Y']
                replace_values = ['Buses', 'All Other']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Vehicle Type', orig_values, replace_values, filter_values = orig_values)

            # 293.C. Added Personal Injury Protection Options
            elif rule_order == 3:
                sheet_names = ['AdditionalPIPPremiumKS', 'AdditionalPIPPremiumKS']
                orig_values = ['Option 1', 'Option 2']
                replace_values = ['Option 1', 'Option 2']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)
                output_table.rename(columns = {"Premium" : "Rate Per Auto"}, inplace = True)

            # 293.D.1. Broadened Personal Injury Protection
            elif rule_order == 4:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

        elif self.StateAbb == "KY":

            # 293.C.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['ZoneRatedPIPLiabilityFactorKY'] * 3 + ['NoFaultPIPFactor']
                orig_values = ['A','B','C'] + ['Y']
                replace_values = ['$100,000 Liability (A)', '$100,000 Liability (B)', '$100,000 Liability (C)'] + ['Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Additional PIP Limits', orig_values, replace_values, filter_values = orig_values)

            # 293.D.1. Added Personal Injury Protection
            elif rule_order == 2:
                sheet_names = ['AddedPIPPerAutoBasePremium']
                output_table = pd.DataFrame(self.rateTables[company][sheet_names[0]][1:],columns=self.rateTables[company][sheet_names[0]][0])
                output_table = output_table.pivot(index='AdditionalPIPCoverageTypeKY', columns='AddedPIPVehicleCount', values='BasePremium')
                output_table.columns = ['0', '1', '2-4', '5-9', '10-20', '21-40']
                output_table.reset_index(inplace = True)  # Adjusting index
                output_table = output_table[output_table['AdditionalPIPCoverageTypeKY'] != "No Coverage"]
                output_table.rename(columns = {"AdditionalPIPCoverageTypeKY" : ''}, inplace = True) # Hiding the column name

            elif rule_order == 3:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

        elif self.StateAbb == "MD":

            # 293.C.3. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor']
                orig_values = ['Y'] * 2
                replace_values = ['$100,000 Liability', 'Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.C.4. Waive Personal Injury Protection - Full Coverage Factor
            elif rule_order == 2:
                sheet_names = ['PIPWaiverFactor']
                orig_values = ['Basic Personal Injury Protection - with Exclusion of Benefits']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

        elif self.StateAbb == "MI":

            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor', 'NoFaultPropertyProtectionInsuranceFactor']
                orig_values = ['Y'] * 3
                replace_values = ['$100,000 Liability', 'Personal Injury Protection', 'Property Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.B.3. Not Required Personal Injury Protection And Property Protection Factor
            elif rule_order == 2:
                sheet_names = ['NoPIPLiabilityFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.B.5. Medical Expenses Limits Factors
            elif rule_order == 3:
                sheet_names = ['MedicalExpensesLimitsFactors_Ext'] * 4
                orig_values = ['50,000','250,000','500,000']
                replace_values = ['50,000','250,000','500,000']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Limit', orig_values, replace_values, filter_values = orig_values)
                output_table = pd.concat([output_table, pd.DataFrame({"Limit": ["Unlimited"], "Factor": [1]})]) # Adding unlimited, not in ratebook.

            # 293.C.1. Medical Expense Rejection Factor
            elif rule_order == 4:
                sheet_names = ['MedicalExpenseRejectionFactor_Ext']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.C.2. Exclusion of Personal Injury Protection Medical Expenses
            elif rule_order == 5:
                sheet_names = ['QualifiedHealthCoverageMedicalExpensesLimitsFactors_Ext'] * 2
                orig_values = ['All household members are covered by qualified health coverage', 'One or more household member(s), but not all, is (are) covered by qualified health coverage']
                replace_values = ['All household members are covered by qualified health coverage', 'One or more household member(s), but not all, is (are) covered by qualified health coverage']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Qualified', orig_values, replace_values, filter_values = orig_values)

            # 293.C.4.a. Excess Personal Injury Protection Medical Expenses Coverage Factor
            elif rule_order == 6:
                sheet_names = ['CoordinationOfBenefitsExcessPIPFactor']
                orig_values = ['Medical Expenses']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.C.4.b. Excess Personal Injury Protection Work Loss Coverage Factor
            elif rule_order == 7:
                sheet_names = ['CoordinationOfBenefitsExcessPIPFactor']
                orig_values = ['Work Loss Benefits']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.C.4.c. Excess Personal Injury Protection Medical Expenses and Work Loss Coverages Factor
            elif rule_order == 8:
                sheet_names = ['CoordinationOfBenefitsExcessPIPFactor']
                orig_values = ['Medical Expenses and Work Loss Benefits']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.D. Excess Attendant Care Coverage Factors
            elif rule_order == 9:
                sheet_names = ['ExcessAttendantCareCoverageFactors_Ext'] * 3
                orig_values = ['50,000','250,000','500,000']
                replace_values = ['50,000','250,000','500,000']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)

            # 293.E. Broadened Personal Injury Protection Coverage for Named Individuals
            elif rule_order == 10:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names,'', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

        elif self.StateAbb == "MN":

            # 293.B.2. Personal Injury Protection Coverage Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor']
                orig_values = ['Y'] * 2
                replace_values = ['$100,000 Liability', 'Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.B.3. No Personal Injury Protection Factor
            elif rule_order == 2:
                sheet_names = ['NoPIPLiabilityFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.B.5.a. Personal Injury Protection Stacking Factors
            elif rule_order == 3:
                sheet_names = ['PIPStackingFactorMN'] * 8
                orig_values = [1,2,3,4,5,10,21,41] # Excel defaults to int representation, not strings.
                replace_values = ['1','2','3','4','5-9','10-20','21-40','Over 40']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Number of Exposures', orig_values, replace_values, filter_values = orig_values)

            # 293.C. Added Personal Injury Protection
            elif rule_order == 4:
                sheet_names = ['AdditionalPIPIncreasedPremium'] * 6
                orig_values = ['30,000 Medical Expenses and 20,000 Other PIP Benefits',
                               '40,000 Medical Expenses and 20,000 Other PIP Benefits',
                               '50,000 Medical Expenses and 20,000 Other PIP Benefits',
                               '50,000 Medical Expenses and 25,000 Other PIP Benefits',
                               '75,000 Medical Expenses and 25,000 Other PIP Benefits',
                               '100,000 Medical Expenses and 50,000 Other PIP Benefits'
                               ]
                replace_values = ['1','2','3','4','5','6']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Option', orig_values, replace_values, filter_values = orig_values)
                output_table.rename({output_table.columns[1] : "Rate Per Auto"}, inplace = True)

            # 293.D.1. Broadened Personal Injury Protection
            elif rule_order == 5:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names,'', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.E.3. PIP Exclusion of Work Loss Benefits Factors
            elif rule_order == 6:
                sheet_names = ['PIPExclusionOfWorkLossFactorMN'] * 2
                orig_values = ['Named Insured Age 65 or Older Only',
                               'Named Insured/Relative both Age 60 and Older if Retired and Receiving a Pension']
                replace_values = ['Named Insured Age 65 or Older Only',
                                  'Named Insured/Relative both Age 60 and Older if Retired and Receiving a Pension']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Classification', orig_values, replace_values, filter_values = orig_values)

        elif self.StateAbb == "NJ":

            # 293.A.2. Extended Medical Expense Benefits
            if rule_order == 1:
                sheet_names = ['PIPAdditionalExtendedMedicalExpensePremium']
                orig_values = ['Y']
                replace_values = ['$10,000 Limit Rate Per Auto Or Auto Dealer Rating Unit']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.B.2. Van Pools Combined Liability and Personal Injury Protection Coverages Factor
            elif rule_order == 2:
                sheet_names = ['VanPoolsPIPFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

            # 293.B.3. All Other Types Liability and Personal Injury Protection Coverages Factors
            elif rule_order == 3:
                sheet_names = ['NoFaultLiabilityFactorNJ', 'NoFaultPIPFactor']
                table_1 = pd.DataFrame(self.rateTables[company][sheet_names[0]][1:],columns=self.rateTables[company][sheet_names[0]][0])
                table_2 = pd.DataFrame(self.rateTables[company][sheet_names[1]][1:],columns=self.rateTables[company][sheet_names[1]][0])
                no_fault_pip_factor = table_2.iloc[0,1]
                liab_with_tort = table_1[table_1["TruckPIPRatingBasis"] == "With Tort Limitation"].iloc[0,1]
                liab_without_tort = table_1[table_1["TruckPIPRatingBasis"] == "Without Tort Limitation"].iloc[0,1]
                output_table = pd.DataFrame({
                    "Coverage": ["100,000 Liability", "Personal Injury Protection"],
                    "With Tort Limitation": [liab_with_tort, no_fault_pip_factor],
                    "Without Tort Limitation ": [liab_without_tort, no_fault_pip_factor]
                })

            # 293.B.4. Pedestrian Personal Injury Protection Coverages Factor
            elif rule_order == 4:
                sheet_names = ['PedestrianPersonalInjuryProtectionCoveragesFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.C.3.a. Private Passenger Types Liability Tort Limitation Elimination Factor
            elif rule_order == 5:
                sheet_names = ['TortLimitationFactor']
                orig_values = ['No']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.C.3.b. Garages And Van Pools Tort Limitation Elimination Factor
            elif rule_order == 6:
                sheet_names = ['TortLimitationFactor']
                orig_values = ['No']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.C.4.a. Private Passenger Types Combined Options Factors
            elif rule_order == 7:
                output_table = pd.DataFrame(self.rateTables[company]["PIPDeductibleFactorNJ2"][1:], index=None, columns=self.rateTables[company]["PIPDeductibleFactorNJ2"][0])
                output_table = output_table.pivot_table(values = "Factor", index = "PIPDeductible", columns = "CoverageType")
                output_table.reset_index(inplace = True)

                output_table.rename(columns = {"Basic PIP and Pedestrian PIP" : "Basic PIP",
                                               "Medical Expenses and Pedestrian PIP" : "Medical Expenses Only",
                                               "PIPDeductible" : "PIP Deductible"
                                               }, inplace = True)

                output_table.drop(columns = ["Pedestrian PIP"], inplace = True) # Dropping unneeded column

                # Fixing sort issues on the deductible column
                output_table = output_table[output_table["PIP Deductible"] != "Not Applicable"] # removing not applicable
                output_table["PIP Deductible"] = output_table["PIP Deductible"].str.replace(",","")
                output_table["PIP Deductible"] = output_table["PIP Deductible"].astype(int)
                output_table = output_table.sort_values(by = "PIP Deductible") # Sorting


            # 293.C.4.b.(2). Garages And Van Pools Combined Options Factor
            elif rule_order == 8:
                sheet_names = ['GarageDealersCombinedOptionsFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.C.4.b.(3). Garages And Van Pools Subject To Tort Or Limited Tort Liability Factors
            elif rule_order == 9:
                sheet_names = ['GarageDealersTortLimitationFactor'] * 2
                orig_values = ['Yes','No']
                replace_values = ['Limited Tort Liability','Subject to Tort Liability']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)

            # 293.C.5.c. Medical Expense Benefits-as-secondary Option - Personal Injury Protection Coverage Factor
            elif rule_order == 10:
                sheet_names = ['PIPMedicalExpenseSecondaryFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.D.1. Added Personal Injury Protection
            elif rule_order == 11:
                # This rule has some hard coded values on the rate pages template. Those have been copied here.
                # The order of which the values are concatenated is important. Be careful when editing.
                pip_limit_codes = [f"{x:02}" for x in range(2, 21)] # 1:1 with option in terms of ranking
                per_day_ess_serv_ben = [12] + [20]*7 + [12] + [20]*7
                total_agg_ess_serv_ben = ['8760', '14600', '14600', '14600', '14600', '14600', '14600', '14600', '8760', '14600', '14600', '14600', '14600', '14600', '14600', '14600']
                death_benefit = [10000] * 16
                funeral_benefit = [2000] * 16

                # Below this point no info is hardcoded.
                sheet_names = ['AdditionalPIPFirstCarPremium']
                table_1 = pd.DataFrame(self.rateTables[company][sheet_names[0]][1:],columns=self.rateTables[company][sheet_names[0]][0])
                table_2 = pd.DataFrame(self.rateTables[company][sheet_names[0]][1:],columns=self.rateTables[company][sheet_names[0]][0])

                table_1.rename(columns = {"Premium" : "First Auto Or Auto Dealer Rating Unit"}, inplace = True)
                table_2.rename(columns = {"Premium" : "Each Additional Auto Or Auto Dealer Rating Unit"}, inplace = True)

                output_table = pd.merge(table_1,table_2, how = 'left', on = 'AddedPIPCoverageType')

                output_table["Option"] = output_table["AddedPIPCoverageType"].str.extract(r'Option (\d+):').astype(int) # For sorting
                output_table["Aggregate"] = output_table["AddedPIPCoverageType"].str.extract(r'Aggregate ([\w\s$;,]+)')
                output_table["Aggregate"] = output_table["Aggregate"].str.extract(r'(\d[\d,]*)')[0].str.replace(',', '')
                output_table["Weekly"] = output_table["AddedPIPCoverageType"].str.extract(r'Weekly (\$\d+)')

                output_table.drop(columns="AddedPIPCoverageType", inplace=True)
                output_table.sort_values(by="Option", inplace=True)

                output_table["PIP Limit Code"] = pip_limit_codes[:len(output_table)]
                output_table["Weekly Income Benefit"] = per_day_ess_serv_ben[:len(output_table)]
                output_table["Total Aggregate Income Benefit"] = total_agg_ess_serv_ben[:len(output_table)]
                output_table["Total Aggregate Essential Services Benefit"] = total_agg_ess_serv_ben[:len(output_table)]
                output_table["Death Benefit"] = death_benefit[:len(output_table)]
                output_table["Funeral Benefit"] = funeral_benefit[:len(output_table)]

                column_order = [
                    "Option", "PIP Limit Code", "Weekly Income Benefit", "Total Aggregate Income Benefit",
                    "Total Aggregate Essential Services Benefit", "Death Benefit", "Funeral Benefit",
                    "First Auto Or Auto Dealer Rating Unit", "Each Additional Auto Or Auto Dealer Rating Unit"
                ]

                output_table = output_table[column_order]

                # Adding line breaks for the excel formatting later. Column names are too long for regular print.
                output_table.columns = ['\n'.join(col.split()) for col in output_table.columns]


            # 293.D.2. Resident Relative Added Personal Injury Protection
            elif rule_order == 12:
                sheet_names = ['AdditionalPIPResidentRelativePremium']
                orig_values = ['Y']
                replace_values = ['Rate for Each Resident Relative']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.E.1. Broadened Personal Injury Protection For Named Individuals
            elif rule_order == 13:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate for Each Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.F.2. Medical Expense Benefits Coverage - Motor Bus Passengers
            elif rule_order == 14:
                sheet_names = ['PIPMotorBusLiabilityFactor', 'PIPMotorBusMedicalExpensePIPFactor']
                orig_values = ['Y'] * 2
                replace_values = ['$100,000 Liability', "Medical Expense Benefits"]
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            elif rule_order == 15:
            # 293.G.1 Pedestrian PIP Factor
                sheet_names = ['PedestrianPersonalInjuryProtectionCoveragesFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values,
                                                            replace_values, filter_values=orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})  # Removing index that appears in the first column name from the transpose operation

        elif self.StateAbb == "NY":

            # 293.C.1. Work Loss Coordination Factor
            if rule_order == 1:
                sheet_names = ['PIPExclusionOfWorkLossFactorNY']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

            # 293.D.2. Medical Expense Elimination
            elif rule_order == 2:
                sheet_names = ['MedicalExpenseEliminationFactor', 'MedicalExpenseEliminationFactor']
                orig_values = ['Named Insured Only', 'Named Insured and Relative']
                replace_values = ['Named Insured Only', 'Named Insured and Relatives']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.E.2. Additional Personal Injury Protection
            elif rule_order == 3:
                sheet_names = ['AdditionalPIPPremium'] * 4
                orig_values = ['Option A', 'Option B', 'Option C', 'Option D']
                replace_values = ['A', 'B', 'C', 'D']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Option', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {"Premium" : "Rate Per Auto"}) # Removing index that appears in the first column name from the transpose operation

            # 293.E.4. Additional Death Benefit
            elif rule_order == 4:
                sheet_names = ['AdditionalPIPDeathBenefitPremium']
                orig_values = ['Yes']
                replace_values = ['Rate Per Auto']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.F.1. Broadened Personal Injury Protection
            elif rule_order == 5:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate for Each Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '}) # Removing index that appears in the first column name from the transpose operation

            # 293.G.2. Optional Basic Economic Loss Coverage
            elif rule_order == 6:
                table = pd.DataFrame(self.rateTables[company]["AllOtherOptionalBasicEconomicLossPremium"][1:], index=None,columns=self.rateTables[company]["AllOtherOptionalBasicEconomicLossPremium"][0])
                factor_yes = table[table[table.columns[0]] == "Yes"].iloc[0,1]
                exposure_type = ["Private Passenger", "Trucks, Tractors, Trailers", "Auto Dealers", "All Other"]
                class_code = ["9502","9503","9504","9509"]
                limit = ["25,000"] * 4

                output_table = pd.DataFrame({
                    "Exposure Type": exposure_type,
                    "Class Code": class_code,
                    "Limit": limit,
                    "Rate Per Auto" : [factor_yes] * 4
                })

        elif self.StateAbb == "OR":

            # 293.B. Premium Development
            if rule_order == 1:
                sheet_names = ['PIPGarageFactor'] + ['PIPOtherThanPrivatePassengerAndGarageFactor'] * 2
                orig_values = ['Y'] + ['Principally Operated by Employees', 'Not Principally Operated by Employees']
                replace_values = ['Garages'] + ['All Other Types (Principally Operated by Employees)', 'All Other Types (Not Principally Operated by Employees)']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Classification', orig_values, replace_values, filter_values = orig_values)

            # 293.C. Broadened Personal Injury Protection for Named Individuals
            elif rule_order == 2:
                sheet_names = ['BroadenedPIPPremium']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})

        elif self.StateAbb == "PA":

            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor','NoFaultPIPFactor']
                orig_values = ['Y'] * 2
                replace_values = ['$100,000 Liability', 'First Party Benefits']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)

            # 293.C. Added First Party Benefits
            elif rule_order == 2:
                sheet_names = ['NoFaultMedicalExpenseBenefitPremium'] * 4
                orig_values = ['10,000','25,000','50,000','100,000']
                replace_values = ['10,000','25,000','50,000','100,000']
                output_table_1 = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Total Limit', orig_values, replace_values, filter_values = orig_values)
                output_table_1["Coverage"] = ["Medical Expense Benefits"] * 4 # Place holder value for formatting coverage later.

                sheet_names = ['NoFaultWorkLossBenefitPremium'] * 4
                orig_values = ['1,000 Monthly / 5,000 Total','1,000 Monthly / 15,000 Total','1,500 Monthly / 25,000 Total','2,500 Monthly / 50,000 Total']
                replace_values = ['1,000 / 5,000','1,000 / 15,000','1,500 / 25,000','2,500 / 50,000']
                output_table_2 = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Total Limit', orig_values, replace_values, filter_values = orig_values)
                output_table_2["Coverage"] = ["Work Loss Benefits"] * 4 # Place holder value for formatting coverage later.

                sheet_names = ['NoFaultFuneralExpenseBenefitPremium'] * 2
                orig_values = ['1,500','2,500']
                replace_values = ['1,500','2,500']
                output_table_3 = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Total Limit', orig_values, replace_values, filter_values = orig_values)
                output_table_3["Coverage"] = ["Funeral Expense Benefits"] * 2 # Place holder value for formatting coverage later.

                sheet_names = ['NoFaultAccidentalDeathBenefitPremium'] * 3
                orig_values = ['5,000','10,000','25,000']
                replace_values = ['5,000','10,000','25,000']
                output_table_4 = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Total Limit', orig_values, replace_values, filter_values = orig_values)
                output_table_4["Coverage"] = ["Accidental Death Benefits"] * 3 # Place holder value for formatting coverage later.

                output_table = pd.concat([output_table_1, output_table_2, output_table_3, output_table_4])
                output_table.rename(columns = {output_table.columns[1] : "Rate Per Auto"}, inplace = True)
                output_table = output_table[["Coverage", "Total Limit", "Rate Per Auto"]] # Fixing column order

            # 293.C. Public Vehicles Added First Party Benefits Coverage Factor
            elif rule_order == 3:
                sheet_names = ['NoFaultAdditionalFirstPartyBenefitsSeatingCapacityFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

            # 293.D. Combination First Party Benefits
            elif rule_order == 4:
                sheet_names = ['NoFaultCombinationBenefitPremium'] * 4
                orig_values = ['50,000','100,000','177,500','277,500']
                replace_values = ['50,000','100,000','177,500','277,500']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, "Total Benefits Limit", orig_values, replace_values, filter_values = orig_values)
                output_table["Funeral Expense"] = ["2,500"] * 4 # Hardcoded in template
                output_table["Accidental Death"] = ["10,000"] * 2 + ["25,000"] * 2 # Hardcoded in template
                output_table.rename(columns = {"Premium" : "Rate Per Auto", "NoFaultCombinationBenefit" : "Total Benefits Limit"}, inplace = True)
                column_order = ["Total Benefits Limit", "Funeral Expense", "Accidental Death", "Rate Per Auto"]

                output_table = output_table[column_order]

            # 293.D. Public Vehicles Combination First Party Benefits Coverage Factor
            elif rule_order == 5:
                sheet_names = ['NoFaultCombinationFirstPartyBenefitsSeatingCapacityFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

            # 293.E.1. Broadened First Party Benefits for Named Individuals
            elif rule_order == 6:
                sheet_names = ['BroadenedPIPPremiumPA']
                orig_values = ['Y']
                replace_values = ['Rate Per Named Individual']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

            # 293.F.4. Extraordinary Medical Benefits Coverage
            elif rule_order == 7:
                sheet_names = ['NoFaultExtraordinaryMedicalBenefitsPremium'] * 4
                orig_values = ['100,000 Principally Operated By Employees/Covered By Workers\' Compensation', # Function concatenates the two columns by a space.
                               '3000,000 Principally Operated By Employees/Covered By Workers\' Compensation',
                               '500,000 Principally Operated By Employees/Covered By Workers\' Compensation',
                               '1,000,000 Principally Operated By Employees/Covered By Workers\' Compensation']

                replace_values = ['100,000', '300,000', '500,000', '1,000,000']
                output_table_1 = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Limit', orig_values, replace_values, filter_values = orig_values)
                output_table_1.rename(columns = {"Premium" : "WO/WC \nRate Per Auto"}, inplace = True)

                sheet_names = ['NoFaultExtraordinaryMedicalBenefitsPremium'] * 4
                orig_values = ['100,000 Not Principally Operated By Employees/Not Covered By Workers\' Compensation', # Function concatenates the two columns by a space.
                               '3000,000 Not Principally Operated By Employees/Not Covered By Workers\' Compensation',
                               '500,000 Not Principally Operated By Employees/Not Covered By Workers\' Compensation',
                               '1,000,000 Not Principally Operated By Employees/Not Covered By Workers\' Compensation']

                replace_values = ['100,000', '300,000', '500,000', '1,000,000']

                output_table_2 = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Limit', orig_values, replace_values, filter_values = orig_values)
                output_table_2.rename(columns = {"Premium" : "With/WC \nRate Per Auto"}, inplace = True)

                output_table = output_table_1.merge(output_table_2, how = 'left', on = 'Limit')

            # 293.F.5. Extraordinary Medical Benefits Coverage Factors - Public Vehicles\
            elif rule_order == 8:
                sheet_names = ['NoFaultPublicTransportationExtraordinaryMedicalBenefitSeatingCapacityFactor'] * 4
                orig_values = ['1 to 8', '9 to 20', '21 to 60', 'Over 60']
                replace_values = ['1-8', '9-20', '21-60', 'Over 60']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Seating Capacity', orig_values, replace_values, filter_values = orig_values)

        elif self.StateAbb == "TX":

            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor']
                orig_values = ['Y'] * 2
                replace_values = ['$100,000 Liability' , 'Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            elif rule_order == 2:
                sheet_names = ['PIPIncreasedPremium'] * 7
                orig_values = ["2,500","5,000","10,000","25,000","50,000","75,000","100,000"]
                replace_values = ["2,500","5,000","10,000","25,000","50,000","75,000","100,000"]
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Limit', orig_values, replace_values, filter_values = orig_values)

        elif self.StateAbb == "UT":

            # 293.B.2. No-Fault Factors
            if rule_order == 1:
                sheet_names = ['NoFaultLiabilityFactor', 'NoFaultPIPFactor']
                orig_values = ['Y'] * 2
                replace_values = ['$100,000 Liability' , 'Personal Injury Protection']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values = orig_values)

            # 293.B.2. No-Fault Factors
            elif rule_order == 2:
                sheet_names = ['PublicTransportationBusesOnInterstatePIPFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

            # 293.C.1. Additional Personal Injury Protection
            elif rule_order == 3:
                sheet_names = ['AddedPIPFactor'] * 2
                orig_values = ['Basic with Option A', 'Basic with Option B']
                replace_values = ['a.' , 'b.']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Option', orig_values, replace_values, filter_values = orig_values)

            # 293.D.2. Loss of Gross Income and Earning Capacity Benefits Exclusion Factor
            elif rule_order == 4:
                sheet_names = ['PIPExclusionOfLossOfIncomeBenefitsFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

        elif self.StateAbb == "WA":
            # 293.B.1. Medical Payment Coverage Factor - With Basic Personal Injury Protection Benefits
            if rule_order == 1:
                sheet_names = ['BasicPersonalInjuryProtectionFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

            # 293.B.2. Medical Payment Coverage Factor - With Increased Limits
            elif rule_order == 2:
                sheet_names = ['IncreasedLimitsPersonalInjuryProtectionFactor']
                orig_values = ['Y']
                replace_values = ['Factor']
                output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values, replace_values, filter_values = orig_values)
                output_table = output_table.rename(columns = {output_table.columns[0] : '', output_table.columns[1] : ' '})

        # Format numeric values and factors if sheet_names contains "Premium"
        if sheet_names and any("Premium" in name for name in sheet_names):
            for column in output_table.columns:
                if output_table[column].dtype == 'float64' or output_table[column].dtype == 'int64':
                    output_table[column] = output_table[column].apply(lambda x: f"{x:.2f}")
        elif sheet_names and any("Factor" in name for name in sheet_names):
            for column in output_table.columns:
                if output_table[column].dtype == 'float64' or output_table[column].dtype == 'int64':
                    output_table[column] = output_table[column].apply(lambda x: f"{x:.3f}")

        return output_table

    def build297Table_unstacked(self, company, table_num: int):
        """
        Builds the non-stacked tables in rule 297.

        Description:
        The build297Table_unstacked method generates non-stacked tables for rule 297.
        It processes rate tables from an Excel file and applies filters based on the state and coverage type.

        Parameters:
        - company (str): The name of the company for which the rate tables are being processed.
        - table_num (int): The table number corresponding to the input file sheet names and the Rate Pages Excel template.

        Returns:
        - output_table (pd.DataFrame): A DataFrame containing the non-stacked table for the specified table number.

        Method Overview:
        1. Mapping Sheet Names: The method uses a dictionary to map the table_num parameter to corresponding sheet names in the Excel file.
        2. State Mapping: The method reads the state-specific mappings from the Excel file and applies them to filter the rate tables.
        3. Loading Rate Tables: It loads the relevant rate tables for each coverage type from the Excel file.
        4. Filtering: The method applies filters on the rate tables based on the state-specific mappings.
            a. This determines the level of data that is wanted.
        5. Renaming Columns: It renames the columns of the rate tables based on predefined mappings.
        6. Merging Tables: Finally, it merges the filtered and renamed rate tables into a single output table.

        :param company: The name of the company for which the rate tables are being processed.
        :param table_num: The table number corresponding to the input file sheet names and the Rate Pages Excel template.
        :return: A DataFrame containing the non-stacked table for the specified table number.
        """

        sheet_names = {
            1: "297 UI - Non Stacked (T1)",
            2: "297 UI - Non Stacked (T2)",
            3: "297 UI - Non Stacked (T3)",
            4: "297 UI - Non Stacked (T4)",
            5: "297 Hired Auto N-O UI (T5)",
            6: "297 Hired Auto N-O UDI (T6)",
            10: "297 UM PD (T10)",
        }

        column_names = {
            1: ["Limit", "PPT", "Other Than PPT"],
            2: ["Limit", "PPT", "Other Than PPT"],
            3: ["Limit", "PPT", "Other Than PPT"],
            4: ["Limit", "PPT", "Other Than PPT"],
            5: ["Limit", "Non-owned-Autos Per Employee", "Hired Autos per $100 Cost of Hire"],
            6: ["Limit", "Non-owned-Autos Per Employee", "Hired Autos per $100 Cost of Hire"],
            10: ["Limit", "PPT", "Other Than PPT"]
        }

        if self.StateAbb == "OH" and table_num == 10:
            state_mapper = pd.read_excel(BA_INPUT_FILE, sheet_name=sheet_names[table_num])
            state_mapper = state_mapper[state_mapper["state"] == self.StateAbb]

        sheet_name = sheet_names[table_num]
        limits_list = pd.read_excel(BA_INPUT_FILE, sheet_name="UM-UIM-UMPD Limits", header=[0, 1])

        # Fill forward the merged state names (top-level headers)
        limits_list.columns = limits_list.columns.to_frame().fillna(method='ffill').apply(tuple, axis=1)

        # Combine the two header levels into a single string for each column
        limits_list.columns = [f"{state}_{coverage}" for state, coverage in limits_list.columns]

        um_info = pd.read_excel(BA_INPUT_FILE, sheet_name="297 Map")

        subtitle_name = um_info[(um_info["State"] == self.StateAbb)][f"Table {table_num}"]
        subtitle_name = subtitle_name.values[0]

        UM_flag = "uninsured" in str(subtitle_name).lower()
        UIM_flag = "underinsured" in str(subtitle_name).lower()
        UMPD_flag = "property damage" in str(subtitle_name).lower()

        # Looks at the least common flag first because names are not mutually exclusive.
        if UMPD_flag:
            try:
                limits_list = limits_list[f"{self.StateAbb}_UMPD"].dropna()
            except (KeyError) as e:
                limits_list = limits_list[f"{self.StateAbb}_UM"].dropna()

        elif UIM_flag:
            try:
                limits_list = limits_list[f"{self.StateAbb}_UIM"].dropna()
            except (KeyError) as e:
                limits_list = limits_list[f"{self.StateAbb}_UM"].dropna()

        elif UM_flag:
            limits_list = limits_list[f"{self.StateAbb}_UM"].dropna()

        state_mapper = pd.read_excel(BA_INPUT_FILE, sheet_name=sheet_name)

        if self.StateAbb in state_mapper["state"].values:
            state_mapper = state_mapper[state_mapper["state"] == self.StateAbb]
        else:
            state_mapper = state_mapper[state_mapper["state"] == "Default"]

        try:
            sheet_1 = state_mapper['privatepassenger'].iloc[0]
            table_1 = pd.DataFrame(self.rateTables[company][sheet_1][1:], index=None,columns=self.rateTables[company][sheet_1][0]).dropna()

            sheet_2 = state_mapper['nonprivatepassenger'].iloc[0]
            table_2 = pd.DataFrame(self.rateTables[company][sheet_2][1:], index=None,columns=self.rateTables[company][sheet_2][0]).dropna()

            table_1.rename(columns = {"Premium":"BasePremium"}, inplace = True)
            table_2.rename(columns = {"Premium":"BasePremium"}, inplace = True)

            # Dynamically goes through every column and applies the filter if the value exists in said column. Every column should be mutually
            # distinct in its set of values, therefore this approach works. Else a special case is needed.

            for value in state_mapper.iloc[0, 3:state_mapper.shape[1]].values:

                # Skipping null filter values that exist in the filter file.
                if pd.isnull(value):
                    continue

                for column in table_1.columns[:-1]:

                    if value in table_1[column].values:
                        table_1 = table_1[table_1[column] == value]
                        table_1 = table_1.drop(columns=[column])

                        table_2 = table_2[table_2[column] == value]
                        table_2 = table_2.drop(columns=[column])

            # Ghost columns existing in sheets, below is a subset
            # The ghost column will of type None, even the column name.
            col_1 = [col for col in table_1.columns if (col.startswith('UninsuredMotoristCombined') or col.startswith('UnderinsuredMotoristCombined') or col.startswith('UninsuredMotoristPropertyDamage'))]
            col_2 = [col for col in table_2.columns if (col.startswith('UninsuredMotoristCombined') or col.startswith('UnderinsuredMotoristCombined') or col.startswith('UninsuredMotoristPropertyDamage'))]

            table_1 = table_1[[col_1[0], "BasePremium"]]
            table_2 = table_2[[col_2[0], "BasePremium"]]

            # Merging the tables together after we have the correct data level
            output_table = pd.merge(table_1, table_2, on=table_1.columns[0], how="left")
            output_table = output_table[output_table[output_table.columns[0]] != "Not Applicable"]  # Removing Not Applicable.
            output_table = output_table[output_table[output_table.columns[0]] != "No Coverage"]  # Removing No Coverage.
            # Mapping column names
            col_names = column_names[table_num]

            if self.StateAbb == "MT":
                col_names = ["Limit", "UM", "UIM"]

            output_table.columns = col_names

            # Sorting first column by changing values into integers, sorting, then returning them back to strings.
            output_table = output_table.replace(r'^\s*$', np.nan, regex=True).dropna()
            output_table[output_table.columns[0]] = output_table[output_table.columns[0]].str.replace(',', '').astype(float)
            output_table.sort_values(by=output_table.columns[0], inplace=True)
            output_table = output_table[output_table[output_table.columns[0]].astype(float).isin(limits_list)]
            output_table[output_table.columns[0]] = output_table[output_table.columns[0]].apply(lambda x: f"{x:,.0f}")

            for column in output_table.columns:
                if output_table[column].dtype == 'float64' or output_table[column].dtype == 'int64':
                    output_table[column] = output_table[column].apply(lambda x: f"{x:.2f}")

            output_table.set_index(output_table.columns[0], inplace = True)


            # Special case for OR table 10 (PD) having only PPT.
            if sheet_1 == sheet_2:
                output_table = output_table.drop(columns = [output_table.columns[-1]])


        except KeyError as e:
            warnings.warn("297 unstacked has encountered an error! One or more tables may be blank/missing.", Um_Unstacked_Warning)
            traceback.print_exc()

            output_table = None

        return output_table

    def build297Table_stacked(self, company, table_num: int):
        """
        Builds the non-stacked tables in rule 297.

        Description:
        The build297Table_unstacked method generates Stacked tables for rule 297.
        It processes rate tables from an Excel file and applies filters based on the state and coverage type.

        Parameters:
        - company (str): The name of the company for which the rate tables are being processed.
        - table_num (int): The table number corresponding to the input file sheet names and the Rate Pages Excel template.

        Returns:
        - output_table (dict): A dict containing the Stacked tables for the specified table number (PPT and Non PPT)

        Method Overview:
        1. Mapping Sheet Names: The method uses a dictionary to map the table_num parameter to corresponding sheet names in the Excel file.
        2. State Mapping: The method reads the state-specific mappings from the Excel file and applies them to filter the rate tables.
        3. Loading Rate Tables: It loads the relevant rate tables for each coverage type from the Excel file.
        4. Filtering: The method applies filters on the rate tables based on the state-specific mappings.
            a. This determines the level of data that is wanted.
        5. Renaming Columns: It renames the columns of the rate tables based on predefined mappings.
        6. Pivot Tables: Transforms the data from long format to a matrix format.

        :param company: The name of the company for which the rate tables are being processed.
        :param table_num: The table number corresponding to the input file sheet names and the Rate Pages Excel template.
        :return: A dict of DataFrames containing the Stacked tables for the specified table number.

        Note: Rate Pages had these tables side by side, this is not possible in python due to column aliasing being undefined.
        """

        sheet_names = {
            7: "297 UI - Stacked (T7)",
            8: "297 UI - Stacked (T8)",
            9: "297 UI - Stacked (T9)",
        }

        sheet_name = sheet_names[table_num]
        state_mapper = pd.read_excel(BA_INPUT_FILE, sheet_name=sheet_name)
        limits_list = pd.read_excel(BA_INPUT_FILE, sheet_name="UM-UIM-UMPD Limits", header=[0, 1])

        # Fill forward the merged state names (top-level headers)
        limits_list.columns = limits_list.columns.to_frame().fillna(method='ffill').apply(tuple, axis=1)

        # Combine the two header levels into a single string for each column
        limits_list.columns = [f"{state}_{coverage}" for state, coverage in limits_list.columns]

        um_info = pd.read_excel(BA_INPUT_FILE, sheet_name="297 Map")

        subtitle_name = um_info[(um_info["State"] == self.StateAbb)][f"Table {table_num}"]
        subtitle_name = subtitle_name.values[0]

        UM_flag = "uninsured" in str(subtitle_name).lower()
        UIM_flag = "underinsured" in str(subtitle_name).lower()
        UMPD_flag = "property damage" in str(subtitle_name).lower()

        if UM_flag:
            limits_list = limits_list[f"{self.StateAbb}_UM"].dropna()
        elif UIM_flag:
            try:
                limits_list = limits_list[f"{self.StateAbb}_UIM"].dropna()
            except (KeyError) as e:
                limits_list = limits_list[f"{self.StateAbb}_UM"].dropna()

        elif UMPD_flag:
            try:
                limits_list = limits_list[f"{self.StateAbb}_UMPD"].dropna()
            except (KeyError) as e:
                limits_list = limits_list[f"{self.StateAbb}_UM"].dropna()

        if self.StateAbb in state_mapper["state"].values:
            state_mapper = state_mapper[state_mapper["state"] == self.StateAbb]
        else:
            state_mapper = state_mapper[state_mapper["state"] == "Default"]

        try:
            sheet_1 = state_mapper['privatepassenger'].iloc[0]
            table_1 = pd.DataFrame(self.rateTables[company][sheet_1][1:], index=None,
                                   columns=self.rateTables[company][sheet_1][0])

            sheet_2 = state_mapper['nonprivatepassenger'].iloc[0]
            table_2 = pd.DataFrame(self.rateTables[company][sheet_2][1:], index=None,
                                   columns=self.rateTables[company][sheet_2][0])

            table_1.rename(columns={"Premium": "BasePremium"}, inplace=True)
            table_2.rename(columns={"Premium": "BasePremium"}, inplace=True)

            # Dynamically goes through every column and applies the filter if the value exists in said column. Every column should be mutually
            # distinct in its set of values, therefore this approach works.

            for value in state_mapper.iloc[0, 3:state_mapper.shape[1]].values:

                # Skipping null filter values that exist in the filter file.
                if pd.isnull(value):
                    continue

                for column in table_1.columns[:-1]:

                    if value in table_1[column].values:
                        table_1 = table_1[table_1[column] == value]
                        table_1 = table_1.drop(columns=[column])

                        table_2 = table_2[table_2[column] == value]
                        table_2 = table_2.drop(columns=[column])

            index_cols_t1 = [col for col in table_1.columns if (
                        col.startswith('UninsuredMotoristCombinedSingleLimitText') or col.startswith(
                    'UnderinsuredMotoristCombinedSingleLimitText'))]
            index_cols_t2 = [col for col in table_2.columns if (
                        col.startswith('UninsuredMotoristCombinedSingleLimitText') or col.startswith(
                    'UnderinsuredMotoristCombinedSingleLimitText'))]

            col_1 = [col for col in table_1.columns if (
                        col.startswith('UninsuredMotoristVehicleCount') or col.startswith(
                    'UnderinsuredMotoristVehicleCount'))]
            col_2 = [col for col in table_2.columns if (
                        col.startswith('UninsuredMotoristVehicleCount') or col.startswith(
                    'UnderinsuredMotoristVehicleCount'))]

            table_1 = table_1[
                (table_1[index_cols_t1[0]] != 'Not Applicable') & (table_1[index_cols_t1[0]] != 'No Coverage')
                ].rename(
                columns={index_cols_t1[0]: 'Limit', col_1[0]: 'Total Number of Exposures'}
            )  # Removing Not Applicable. Renaming.

            table_2 = table_2[
                (table_2[index_cols_t1[0]] != 'Not Applicable') & (table_2[index_cols_t1[0]] != 'No Coverage')
                ].rename(
                columns={index_cols_t2[0]: 'Limit', col_2[0]: 'Total Number of Exposures'}
            )  # Removing Not Applicable. Renaming.

            exposure_map = {
                0: '1',  # It is important that these are integers to Str, as excel is weird with types.
                1: '2',
                2: '3 - 4',
                4: '5 - 9',
                9: '10 - 30',
                30: '> 30'
            }
            exposure_order = ['1', '2', '3 - 4', '5 - 9', '10 - 30', '> 30']  # To reorder the columns later.

            table_1['Total Number of Exposures'] = table_1['Total Number of Exposures'].map(exposure_map)
            table_2['Total Number of Exposures'] = table_2['Total Number of Exposures'].map(exposure_map)

            pivot_table_1 = table_1.pivot(
                index='Limit',
                columns=['Total Number of Exposures'],
                values='BasePremium'
            )

            pivot_table_2 = table_2.pivot(
                index='Limit',
                columns=['Total Number of Exposures'],
                values='BasePremium'
            )

            # Putting 'Total Number of Exposures' On top instead of on the left. Along with PPT
            # Need to also make sure the column names exist before reordering/editing for exposure_order
            # Filter exposure_order to include only columns that exist in the DataFrame
            valid_columns = [col for col in exposure_order if col in pivot_table_1.columns.get_level_values(0)]

            # Reindex using only the valid columns
            pivot_table_1 = pivot_table_1.reindex(columns=valid_columns, level=1)
            pivot_table_2 = pivot_table_2.reindex(columns=valid_columns, level=1)

            existing_columns = [col for col in exposure_order if col in pivot_table_1.columns.get_level_values(0)]

            columns1 = pd.MultiIndex.from_product([['PPT'], ['Total Number of Exposures'], existing_columns])
            columns2 = pd.MultiIndex.from_product([['Other Than PPT'], ['Total Number of Exposures'], existing_columns])

            pivot_table_1.columns = columns1
            pivot_table_2.columns = columns2

            output_table = pd.concat([pivot_table_1, pivot_table_2], axis=1)

            # Sorting first column by changing values into integers, sorting, then returning them back to strings.
            output_table = output_table.replace(r'^\s*$', np.nan, regex=True).dropna()
            output_table.index = output_table.index.str.replace(',', '').astype(float)
            output_table = output_table[output_table.index.isin(limits_list.values)]
            output_table.sort_index(inplace=True)
            output_table.index = output_table.index.map(lambda x: f"{x:,.0f}")

            for column in output_table.columns:
                if output_table[column].dtype == 'float64' or output_table[column].dtype == 'int64':
                    output_table[column] = output_table[column].apply(lambda x: f"{x:.2f}")


            return output_table

        except KeyError as e:
            warnings.warn("297 stacked has encountered an error! One or more tables may be blank/missing.", Um_Stacked_Warning)
            output_table = None

        return output_table

    # VA Rule VAPCD
    def buildVAPCD(self,company):
        sheet_names = ['AccidentPreventionDiscountFactorVA_Ext']*4
        orig_values = ['Y']*4
        replace_values = ['Liability:','Med:','Other Than Colisision:',"Collision:"]

        table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        table = table.rename(columns={table.columns[0]: 'Coverage', table.columns[1]: 'Factor'})

        table = table.astype(object)
        table.iloc[:,1:] = table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        # UM is hardcoded to be 1 in the old pages, so it is here too
        um_table = pd.DataFrame({'Coverage': ['UM:'],'Factor': ['1']})
        um_table = um_table.rename(columns={um_table.columns[0]: 'Coverage', um_table.columns[1]: 'Factor'})
        um_table = um_table.astype(object)
        um_table.iloc[:,1:] = um_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        output_table = pd.concat([table,um_table], ignore_index=True)

        return output_table

    # CT: Rule A1
    def buildCT_A1(self, company):
        sheet_names = ['LiabilityofMunicipalitiesPremiumCT_Ext']
        orig_values = ['Y']
        replace_values = ['Premium:']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.0f}")

        return output_table

    def buildFL_PublicsSeasonalMigrantFarm(self, company):
        """FL Rule 243"""
        sheet_names = ['FarmLaborContractorPassengerHazardFactor']
        orig_values = ['Included']
        replace_values = ['Factor']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildFL_AutoBody(self, company):
        """FL Rule 267"""
        sheet_names = ['SpecialTypesAutoBodyFactor']
        orig_values = ['Liability']
        replace_values = ['Factor']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table


    def buildFL_DriveAwayContractors(self, company):
        """FL Rule 269"""

        sheet_names = ['SpecialTypesDriveAwayContractorFactor']
        orig_values = ['Liability']
        replace_values = ['Factor']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    def buildFL_AutoLeaseGapCoverage(self, company):
        """FL Rule 309"""

        sheet_names = ['AutoLoanLeaseGapCoverageFactor']
        orig_values = ['Y']
        replace_values = ['Multiply the Physical Damage Coverage Premiums by the Following Factor:']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:, 1:] = output_table.iloc[:, 1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # KS: Rule A1
    def buildKS_A1(self, company):
        sheet_names = ['AccidentPreventionDiscountFactor']
        orig_values = ['Y']
        replace_values = ['Apply the following factor to applicable premiums:']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # ND: Rule A1
    def buildND_A1(self, company):
        sheet_names = ['RentalVehicleCoverageFactor_Ext']
        orig_values = ['Y']
        replace_values = ['Apply the following factor to applicable premiums:']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # ND: Rule A3
    def buildND_A3(self, company):
        sheet_names = ['AccidentPreventionDiscountFactorND_Ext']
        orig_values = ['Y']
        replace_values = ['Apply the following factor to applicable premiums:']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # MI: Rule A2
    def buildMI_A2(self, company):
        sheet_names = ['PropertyDamageLiabilityCoverageBuybackVehiclePremium']
        orig_values = ['Y']
        replace_values = ['Rate']

        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ''})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.2f}")

        return output_table

    # MI: Rule A4
    def buildMI_A4A(self, company):
        # Load the table and assign column names
        output_table = pd.DataFrame(
            self.rateTables[company]['MIMCCAFee_v2_Ext'][1:],
            columns=self.rateTables[company]['MIMCCAFee_v2_Ext'][0]
        )

        # Filter out historic vehicles
        output_table = output_table[output_table['RegisteredHistoricVehicle'] == "No"]

        # Filter for Unlimited and NotApplicable pipMedicalExpensesLimit
        unlimited_df = output_table[output_table['pipMedicalExpensesLimit'] == "Unlimited"]
        not_applicable_df = output_table[output_table['pipMedicalExpensesLimit'] == "NotApplicable"]

        # Further filter based on rejectionOfPIPMedicalExpensesCoverage
        unlimited_df = unlimited_df[unlimited_df['rejectionOfPIPMedicalExpensesCoverage'] == "No"]
        not_applicable_df = not_applicable_df[not_applicable_df['rejectionOfPIPMedicalExpensesCoverage'] == "Yes"]

        # Modify the pipMedicalExpensesLimit values
        unlimited_df['pipMedicalExpensesLimit'] = "Unlimited PIP"
        not_applicable_df['pipMedicalExpensesLimit'] = "Other-than-Unlimited"

        # Concatenate the two filtered DataFrames
        final_df = pd.concat([unlimited_df, not_applicable_df], ignore_index=True)

        # Rename columns and drop unnecessary ones
        final_df = final_df.drop(columns=['RegisteredHistoricVehicle', 'rejectionOfPIPMedicalExpensesCoverage'])
        final_df = final_df.rename(columns={
            'pipMedicalExpensesLimit': '',
            'Fee': ''
        })

        final_df = final_df.astype(object)
        final_df.iloc[:, 1:] = final_df.iloc[:, 1:].astype(float).map(lambda x: f"{x:.2f}")

        return final_df

    # MI: Rule A4
    def buildMI_A4B(self, company):
        # Load the table and assign column names
        output_table = pd.DataFrame(
            self.rateTables[company]['MIMCCAFee_v2_Ext'][1:],
            columns=self.rateTables[company]['MIMCCAFee_v2_Ext'][0]
        )

        # Filter out historic vehicles
        output_table = output_table[output_table['RegisteredHistoricVehicle'] == "Yes"]

        # Filter for Unlimited and NotApplicable pipMedicalExpensesLimit
        unlimited_df = output_table[output_table['pipMedicalExpensesLimit'] == "Unlimited"]
        not_applicable_df = output_table[output_table['pipMedicalExpensesLimit'] == "NotApplicable"]

        # Further filter based on rejectionOfPIPMedicalExpensesCoverage
        unlimited_df = unlimited_df[unlimited_df['rejectionOfPIPMedicalExpensesCoverage'] == "No"]
        not_applicable_df = not_applicable_df[not_applicable_df['rejectionOfPIPMedicalExpensesCoverage'] == "Yes"]

        # Modify the pipMedicalExpensesLimit values
        unlimited_df['pipMedicalExpensesLimit'] = "Unlimited PIP"
        not_applicable_df['pipMedicalExpensesLimit'] = "Other-than-Unlimited"

        # Concatenate the two filtered DataFrames
        final_df = pd.concat([unlimited_df, not_applicable_df], ignore_index=True)

        # Rename columns and drop unnecessary ones
        final_df = final_df.drop(columns=['RegisteredHistoricVehicle', 'rejectionOfPIPMedicalExpensesCoverage'])
        final_df = final_df.rename(columns={
            'pipMedicalExpensesLimit': '',
            'Fee': ''
        })

        final_df = final_df.astype(object)
        final_df.iloc[:, 1:] = final_df.iloc[:, 1:].astype(float).map(lambda x: f"{x:.2f}")

        return final_df

    def buildNJ_A5(self, company):
        sheet_names = ['JitneysLiabilityBasePremium']
        orig_values = ['Y']
        replace_values = ['Liability Base Rate:']

        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, 'Coverage', orig_values, replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ''})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.2f}")

        return output_table

    # NV: Rule A2
    def buildNV_A2(self, company):
        sheet_names = ['PassiveRestraintDiscountFactor']*8
        orig_values = ['Passive seat belts/dual front airbags',
                       'Passive seat belts/driver side airbag',
                       'Passive seat belts/no airbags',
                       'Manual seat belts/dual front, side and rear side-impact airbags',
                       'Manual seat belts/dual front airbags',
                       'Manual seat belts/driver side airbags',
                       'Manual seat belts/no airbags',
                       'Not Applicable']
        replace_values = ['Passive seat belts/dual front airbags',
                       'Passive seat belts/driver side airbag',
                       'Passive seat belts/no airbags',
                       'Manual seat belts/dual front, side and rear side-impact airbags',
                       'Manual seat belts/dual front airbags',
                       'Manual seat belts/driver side airbags',
                       'Manual seat belts/no airbags',
                       'Not Applicable']

        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: 'Passive Restraint System', output_table.columns[1]: 'Factor'})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # RI: Rule A1
    def buildRI_A1(self, company):
        sheet_names = ['AccidentPreventionDiscountFactorRI']
        orig_values = ['Y']
        replace_values = ['Factor:']
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # RI: Rule A2
    def buildRI_A2(self, company):
        sheet_names = ['AntiTheftDeviceDiscountFactor1']*6
        orig_values = ['Alarm Only','Active Disabling Device Only','Passive Disabling Device Only','Vehicle Recovery System Only','Vehicle Recovery System and Active Disabling Device','Vehicle Recovery System and Passive Disabling Device']
        replace_values = ['A2.B.2. Category 1 Alarm Only Anti-Theft Devices Discount Factor',
                        'A2.C.2. Category 2 Active Disabling Anti-Theft Devices Discount Factor',
                        'A2.D.2. Category 3 Passive Disabling Anti-Theft Devices Discount Factor',
                        'A2.E.2. Category 4 Vehicle Recovery System Anti-Theft Devices Discount Factor',
                        'A2.F.1. Categories 4/1 Or 4/2 Multiple Anti-Theft Devices Discount Factor',
                        'A2.F.2. Categories 4/3 Multiple Anti-Theft Devices Discount Factor'
                        ]
        output_table = self.simple_long_table_build(self.rateTables, company, sheet_names, '', orig_values,
                                                    replace_values, filter_values=orig_values)
        output_table = output_table.rename(columns={output_table.columns[0]: '', output_table.columns[1]: ' '})
        output_table = output_table.astype(object)
        output_table.iloc[:,1:] = output_table.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        return output_table

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Fix to weird column line breaks
    def insert_newline_before_dollar(self, value):
        if isinstance(value, str) and '$' in value:
            return value.replace('$', '\n$')
        return value

    def insert_newline_before_first_open_parenthesis(self, value):
        if isinstance(value, str) and '(' in value:
            return value.replace('(', '\n(', 1)
        return value

    # Same as above, but used specifically for VA
    def insert_newline_before_301Arules(self, value):
        if isinstance(value, str) and '301.A.' in value:
            return value.replace('301.A.', '\n301.A.', 1)
        return value

    def insert_newline_before_301Brules(self, value):
        if isinstance(value, str) and '301.B.' in value:
            return value.replace('301.B.', '\n301.B.', 1)
        return value

    # Applies manual formatting to the base rates worksheet
    def formatBaseRates(self, ws, rule : str):

        # --- Small local helpers -------------------------------------------------
        insert_newline_before_first_open_parenthesis = lambda s: (
            s if s is None else (s if "(" not in str(s) else str(s).replace("(", "\n(", 1))
        )
        insert_newline_before_dollar = lambda s: (
            s if s is None else (s if "$" not in str(s) else str(s).replace("$", "\n$", 1))
        )

        def right_coord(coord: str) -> str:
            """Return the cell coordinate immediately to the right of coord (supports AA, AB...)."""
            col_letters = "".join([c for c in coord if c.isalpha()])
            row_digits = "".join([c for c in coord if c.isdigit()])
            col_idx = openpyxl.utils.column_index_from_string(col_letters)
            next_col_letter = get_column_letter(col_idx + 1)
            return f"{next_col_letter}{row_digits}"

        def merge_with_right(start_coord: str, span: int = 1):
            """Merge start_coord with 'span' cells to its right."""
            col_letters = "".join([c for c in start_coord if c.isalpha()])
            row_digits = "".join([c for c in start_coord if c.isdigit()])
            start_col_idx = openpyxl.utils.column_index_from_string(col_letters)
            row_idx = int(row_digits)
            end_col_letter = get_column_letter(start_col_idx + span)
            ws.merge_cells(f"{start_coord}:{end_col_letter}{row_idx}")

        center_bottom_wrap = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        center_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # --- Column widths -------------------------------------------------------
        for col_letter, px in {
            "A": 82, "B": 220, "C": 220, "D": 220, "E": 220,
            "F": 220, "G": 220, "H": 220, "I": 220, "J": 220, "K": 220
        }.items():
            ws.column_dimensions[col_letter].width = self.pixelsToInches(px)

        # --- Header line-breaks and moving lines (Row 3, Cols B..max) -----------
        start_idx = openpyxl.utils.column_index_from_string("B")
        for col_idx in range(start_idx, ws.max_column + 1):
            cell = ws.cell(row=3, column=col_idx)
            if cell.value is None:
                continue
            # Apply line-breaks
            cell.value = insert_newline_before_first_open_parenthesis(cell.value)
            cell.value = insert_newline_before_dollar(cell.value)
            cell.alignment = center_bottom_wrap

            # Move FIRST TWO lines to the cell above, if present
            lines = str(cell.value).split("\n")
            if len(lines) > 1:
                above_cell = ws.cell(row=2, column=col_idx)
                num_to_move = min(2, len(lines))  # change to 1 if you want previous behavior
                moved = "\n".join(lines[:num_to_move])
                above_cell.value = (str(above_cell.value) + "\n" + moved) if above_cell.value else moved
                above_cell.alignment = center_bottom_wrap
                cell.value = "\n".join(lines[num_to_move:])

        CSL_RULE_LOOKUP = {
            "222" : "Liability (222.C.2.a.)",
            "232": "Liability (232.C.2.a.)",
            "239": "Liability (239.C.2.a.)",

        }

        # --- PIP header names and formatting ------------------------------------
        PIP_RULE_LOOKUP = {
            "222": "Personal Injury Protection (222.C.4.a.(2)., 222.C.4.b.(1).)",
            "232": "Personal Injury Protection (232.B.1.a)",
            "239": "Personal Injury Protection (239.C.2.a)"
        }
        PIP_HEADER_NAME = PIP_RULE_LOOKUP.get(rule)
        pip_cell = "C2"
        ws_title = ws["A1"].value

        # Determine if state is PIP state (keep existing I/O)
        if self.StateAbb in self.pip_states.values and PIP_HEADER_NAME:
            # Match original branching rules
            if "Medical" in str(ws["C2"].value):
                pip_cell = "D2"
            if self.StateAbb == "KY":
                pip_cell = "F2"
            if self.StateAbb == "KS" and rule == "232":
                pip_cell = "F2"
            if self.StateAbb == "OR" and rule == "232":
                pip_cell = "E2"
            if self.StateAbb == "MI" and rule == "222":
                pip_cell = "D2"
            if self.StateAbb == "MI" and any(keyword in ws_title for keyword in ["TRUCKS", "BUSES"]):
                pip_cell = "D2"

            ws[pip_cell] = PIP_HEADER_NAME
            ws[pip_cell].value = insert_newline_before_first_open_parenthesis(ws[pip_cell].value)

            # Decide if PIP is one column by checking the cell to the right
            NON_PIP_RIGHT_TOKENS = ("Liability", "Medical", "Specified", "Comprehensive", "Collision")
            right_cell_value = str(ws[right_coord(pip_cell)].value)
            pip_1_col = any(tok in right_cell_value for tok in NON_PIP_RIGHT_TOKENS)

            # Merge decisions
            if self.StateAbb == "KY" and any(keyword in ws_title for keyword in ["BUSES"]):
                merge_with_right(pip_cell, span=1)

            elif self.StateAbb == "KY" and any(keyword in ws_title for keyword in ["VAN", "TAXICABS", "PRIVATE","TRUCKS"]):
                merge_with_right(pip_cell, span=1)

            elif self.StateAbb == "MI" and any(keyword in ws_title for keyword in ["TRUCKS", "VAN"]):
                merge_with_right(pip_cell, span=2)

            elif self.StateAbb == "MI":
                merge_with_right(pip_cell, span=1)

            elif not pip_1_col:
                merge_with_right(pip_cell, span=1)

            ws[pip_cell].alignment = center_bottom_wrap
            ws.row_dimensions[3].height = 30

        # --- KS / OR special cases
        if self.StateAbb == "KS" and rule == "232":
            ws["B2"] = "Liability \n(232.B.1.a.)"
            ws["B3"] = "Subject to No Fault"
            ws["C3"] = "Not Subject to No Fault"
            ws.merge_cells('B2:C2')

            ws['D2'] = "Medical Payments \n(232.B.2.a.)"
            ws["D3"] = "Subject to No Fault"
            ws["E3"] = "Not Subject to No Fault"
            ws.merge_cells('D2:E2')

            for addr in ("B2", "C2", "D2", "E2"):
                ws[addr].alignment = center_center_wrap

        if self.StateAbb == "OR" and rule == "232":
            ws['C2'] = "Medical Payments \n(232.B.1.a.)"
            ws["C3"] = "Subject to No Fault"
            ws["D3"] = "Not Subject to No Fault"
            ws.merge_cells('C2:D2')
            for addr in ("C2", "D2"):
                ws[addr].alignment = center_center_wrap

        # KY Liability formatting header.
        if self.StateAbb in ["KY"]:
            ws['B2'].value = CSL_RULE_LOOKUP[rule]
            ws.merge_cells('B2:D2')

        # --- Number formatting for inner table values ---------------------------
        # Single pass over all cells from row 4 to max, columns A..max
        for row in range(4, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).number_format = '#,##0.00'

    #Format Towing and Labor Rate Table
    def formatNaics(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(120)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)
        ws.column_dimensions['C'].width = self.pixelsToInches(150)
        ws.column_dimensions['D'].width = self.pixelsToInches(150)
        ws.column_dimensions['E'].width = self.pixelsToInches(150)
        ws.column_dimensions['F'].width = self.pixelsToInches(150)
        ws.column_dimensions['G'].width = self.pixelsToInches(150)
        ws.column_dimensions['H'].width = self.pixelsToInches(150)


        rules = {"Trucks, Tractors, And Trailers Liability" : "Liability \n(223.C.2.e.)",
                 "Trucks, Tractors, And Trailers Comprehensive And Specified Causes Of Loss": "Comprehensive and Specified Causes of Loss\n(223.C.3.h.)",
                 "Trucks And Truck-tractors Collision": "Collision\n(223.C.3.h.)",
                 "Trailers Collision" : "Collision\n(223.C.3.h.)",
                 "Private Passenger Types Liability": "Liability\n(232.B.1.e.)",
                 "Private Passenger Types Collision": "Collision\n(232.B.2.e.)",
                 "Private Passenger Types Comprehensive": "Comprehensive\n(232.B.2.e.)"}

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1:
                    cell.number_format = self.NAICS # Applying currency formatting to columns A-B
                cell.font = Font(size=9, name = "Arial")
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(3, 5):
                cell = ws[char + str(row)]
                if row == 3 and str(cell.value) in rules:
                    cell.value = rules[str(cell.value)]
                cell.font = Font(size=9, name="Arial", bold = True)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4,5):
                cell = ws[char + str(row)]
                cell.font = Font(size=9, name="Arial", bold = False)

        ws.insert_rows(2)
        ws.merge_cells('B3:C3')
        ws['B3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        ws['B3'] = 'Trucks, Tractors, and Trailers'
        ws['D3'] = 'Trucks and Truck-Tractors'
        ws['D3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        ws['E3'] = 'Trailers'
        ws['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        ws.merge_cells('F3:H3')
        ws['F3'] = 'Private Passenger Types'
        ws['F3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in range(3, 5):  # Rows 3 and 4
            for col in range(1, 9):  # Columns A to H
                cell = ws.cell(row=row, column=col)
                if cell.coordinate != "A3":  # Exclude A3
                    cell.border = border

    def format222B(self, ws):
        bold_font = Font(name='Arial', size=10, bold=True)
        italic_font = Font(name='Arial', size=10, italic=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_alignment = Alignment(horizontal='center', vertical='center')

        ws['A2'].font = italic_font
        ws['A3'].border = None  # Assuming you want to apply a thin border

        ws.insert_rows(25, 2)
        ws.insert_rows(49, 2)

        ws['A26'] = "222.B.1.b. Collision Fleet Size Factors"
        ws['A26'].font = italic_font
        ws['A27'].border = None
        ws['A50'] = "222.B.1.c. Other Than Collision Fleet Size Factors"
        ws['A50'].font = italic_font
        ws['A51'].border = None

    def format222E(self, ws):
        ws.column_dimensions['B'].width = self.pixelsToInches(85)
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 13

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("222"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Months', 'Factor']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        ws.delete_rows(3)

    #Format School Bus Operations Rate Table
    def format23B(self, ws):
        """Formats 223.B.5, has important note at bottom."""

        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['B'].width = self.pixelsToInches(125)

        for cell in ws['8:8']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))

            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws['K9'] = ws['A24'].value
        ws['L9'] = ws['B24'].value
        ws['M9'] = ws['C24'].value
        ws['N9'] = ws['D24'].value

        ws['K10'] = ws['A25'].value
        ws['L10'] = ws['B25'].value
        ws['M10'] = ws['C25'].value
        ws['N10'] = ws['D25'].value

        ws['K11'] = ws['A26'].value
        ws['L11'] = ws['B26'].value
        ws['M11'] = ws['C26'].value
        ws['N11'] = ws['D26'].value

        ws['K12'] = ws['A27'].value
        ws['L12'] = ws['B27'].value
        ws['M12'] = ws['C27'].value
        ws['N12'] = ws['D27'].value

        ws['K13'] = ws['A28'].value
        ws['L13'] = ws['B28'].value
        ws['M13'] = ws['C28'].value
        ws['N13'] = ws['D28'].value

        ws['K14'] = ws['A29'].value
        ws['L14'] = ws['B29'].value
        ws['M14'] = ws['C29'].value
        ws['N14'] = ws['D29'].value

        ws['K15'] = ws['A30'].value
        ws['L15'] = ws['B30'].value
        ws['M15'] = ws['C30'].value
        ws['N15'] = ws['D30'].value

        ws['K16'] = ws['A31'].value
        ws['L16'] = ws['B31'].value
        ws['M16'] = ws['C31'].value
        ws['N16'] = ws['D31'].value

        ws['K17'] = ws['A32'].value
        ws['L17'] = ws['B32'].value
        ws['M17'] = ws['C32'].value
        ws['N17'] = ws['D32'].value

        ws['K18'] = ws['A33'].value
        ws['L18'] = ws['B33'].value
        ws['M18'] = ws['C33'].value
        ws['N18'] = ws['D33'].value

        ws['K19'] = ws['A34'].value
        ws['L19'] = ws['B34'].value
        ws['M19'] = ws['C34'].value
        ws['N19'] = ws['D34'].value

        ws['K20'] = ws['A35'].value
        ws['L20'] = ws['B35'].value
        ws['M20'] = ws['C35'].value
        ws['N20'] = ws['D35'].value

        ws['K21'] = ws['A36'].value
        ws['L21'] = ws['B36'].value
        ws['M21'] = ws['C36'].value
        ws['N21'] = ws['D36'].value

        ws['K22'] = ws['A37'].value
        ws['L22'] = ws['B37'].value
        ws['M22'] = ws['C37'].value
        ws['N22'] = ws['D37'].value

        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)
        ws.delete_rows(23)

        ws.insert_rows(3)
        ws.insert_rows(3)
        ws.insert_rows(3)
        ws['A4'] = 'Size Class'
        ws['B4'] = 'Business Use Class'
        ws['C4'] = 'Radius Class'
        ws.merge_cells('C4:N4')
        ws['C5'] = 'Local'
        ws.merge_cells('C5:F5')
        ws['G5'] = 'Intermediate'
        ws.merge_cells('G5:J5')
        ws['K5'] = 'Long Distance'
        ws.merge_cells('K5:N5')
        ws['C6'] = 'Up to 50 Miles'
        ws.merge_cells('C6:F6')
        ws['G6'] = '51 to 200 Miles'
        ws.merge_cells('G6:J6')
        ws['K6'] = 'Over 200 Miles'
        ws.merge_cells('K6:N6')
        ws['C7'] = 'Class (Non-Fleet, Fleet)'
        ws['D7'] = 'Liability Factor'
        ws['E7'] = 'OTC Factor'
        ws['F7'] = 'Collision Factor'
        ws['G7'] = 'Class (Non-Fleet, Fleet)'
        ws['H7'] = 'Liability Factor'
        ws['I7'] = 'OTC Factor'
        ws['J7'] = 'Collision Factor'

        ws.merge_cells('K11:N11')
        ws.merge_cells('A11:J11')
        ws.merge_cells('A4:A7')
        ws.merge_cells('B4:B7')

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)

        for cell in ws['5:5']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)

        for cell in ws['6:6']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)

        for cell in ws['7:7']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)

        """The following is a deletion of part of the table. Since ISO Updates moved the zone rated portion to another sheet,
            it is easier to delete it from this position than to edit the code earlier."""


        start_cell = "K11"
        end_cell = "N25"
        # Unmerge cells in the specified range
        merged_ranges_to_unmerge = []
        for merged_range in ws.merged_cells.ranges:
            if start_cell in merged_range:
                merged_ranges_to_unmerge.append(merged_range)

        for merged_range in merged_ranges_to_unmerge:
            ws.unmerge_cells(str(merged_range))

        # Clear content and formatting
        for row in ws[start_cell:end_cell]:
            for cell in row:
                cell.value = None  # Clear content
                cell.font = Font()  # Reset font
                cell.border = Border()  # Reset border
                cell.alignment = Alignment()  # Reset alignment

        ws.insert_rows(21)
        ws.insert_rows(18)
        ws.insert_rows(15)

    #Format Towing and Labor Rate Table
    def format23C(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(380)

        for cell in ws[2]:
            cell.font = Font(bold=False, italic = True, size = 10, name = 'Arial')

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col == 1: # Applying unique column width and no decimal formatting to column A
                for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.font = Font(name='Arial', size=9)

        ws['A4'].font = Font(name = "Arial", size = 10, bold = True)
        ws['A4'].alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    # Format 222C
    def format222C(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(110)
        ws.column_dimensions['B'].width = self.pixelsToInches(85)

        ws["A4"].value = ""
        ws["A10"].value = ""
        ws["A16"].value = ""
        ws["A22"].value = ""
        ws["A28"].value = ""

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        # for i in [31,27,15,9,3]:
        #     ws.delete_rows(i)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith('222.'):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in ['Factor','Months']):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
                        cell.number_format = "0.000"
                        cell.border = border


    def format225D(self,ws):

        ws.insert_rows(27)

        ws.column_dimensions['A'].width = self.pixelsToInches(100)
        for col_idx in range(2, 4):  # B=2, D=4
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith('225.D'):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in
                             ['Number Of Powered Vehicles', 'Months Laid Up', 'Liability', 'Collision',
                              'Other Than Collision', 'Factor']):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border


    #Format Towing and Labor Rate Table
    def format31C(self, ws):
        ws["A4"] = "Class Code"
        ws.column_dimensions['B'].width = self.pixelsToInches(120)
        ws.column_dimensions['C'].width = self.pixelsToInches(120)
        ws.column_dimensions['D'].width = self.pixelsToInches(120)
        ws.column_dimensions['G'].width = self.pixelsToInches(120)

    # Format 32B
    def format32B(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['C'].width = self.pixelsToInches(110)
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col == 1: # Applying unique column width and no decimal formatting to column A
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.number_format = '#0'
        # Insert a horizontal page break at row 25
        ws.row_breaks.append(Break(id=25))
        ws['A26'].alignment = Alignment(wrap_text=False)


        # Format PPP Classifications - Farm
    def format33(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        # for col in range(2, 3):
        #     char = get_column_letter(col) # Letter representing the current column
        #     for row in range(5, ws.max_row + 1):
        #         cell = ws[char + str(row)]
        #         cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #         cell.number_format = cell.number_format = '#0'
        # Format Public Auto Classifications

    def format239C(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)


        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = self.pixelsToInches(100)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        ws.column_dimensions['C'].width = self.pixelsToInches(100)
        ws.column_dimensions['D'].width = self.pixelsToInches(100)
        ws.column_dimensions['E'].width = self.pixelsToInches(100)


        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '239' in cell.value:
                        cell.font = italic
                    elif any(word in cell.value for word in ['Number', 'Other', 'School', 'Taxicabs','Van']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif cell.value == "Factor":
                        cell.font = bold
                        cell.alignment = align_center
                        row_index = cell.row
                        ws.delete_rows(row_index - 1)
                    elif any(char.isdigit() or char.isalpha() for char in cell.value) and '239' not in cell.value:  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

    def format240(self, ws):
        """This sheet was done in a quick fix, its formatting is messy. Feel free to redo."""

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 13

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        ws.column_dimensions['A'].width = self.pixelsToInches(90)
        ws.column_dimensions['B'].width = self.pixelsToInches(90)
        ws.column_dimensions['C'].width = self.pixelsToInches(90)
        ws.column_dimensions['D'].width = self.pixelsToInches(90)
        ws.column_dimensions['E'].width = self.pixelsToInches(90)
        ws.column_dimensions['F'].width = self.pixelsToInches(90)

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, 56 + 1):
                cell = ws[char + str(row)]
                if col in [1,2]:
                    cell.number_format = '0'
                if col == 3:
                    cell.number_format = '0.000'

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("240"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Factor', 'Non-Fleet','Fleet','Secondary','Class','Categories','Seating','School','Other','Liability','Physical']):
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        # Merge specified cells and apply formatting
        merge_ranges = ['A61:A64', 'A65:A68', 'C72:D72', 'E72:F72']
        for merge_range in merge_ranges:
            # Clear border formatting before merging
            for row in ws[merge_range]:
                for cell in row:
                    cell.border = Border()  # Clear the border

            ws.merge_cells(merge_range)
            for row in ws[merge_range]:
                for cell in row:
                    cell.font = Font(name='Arial', bold=True, size=10)

        for row in ws.iter_rows(min_row=61, max_row=68, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = '0'
                cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                     right=Side(border_style='thin', color='C1C1C1'),
                                     top=Side(border_style='thin', color='C1C1C1'),
                                     bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows(min_row=74, max_row=77, min_col=3, max_col=6):
            for cell in row:
                cell.number_format = '0.000'
                cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                     right=Side(border_style='thin', color='C1C1C1'),
                                     top=Side(border_style='thin', color='C1C1C1'),
                                     bottom=Side(border_style='thin', color='C1C1C1'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for row in ws.iter_rows(min_row=61, max_row=68, min_col=4, max_col=5):
            for cell in row:
                cell.number_format = '0.000'
                cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                     right=Side(border_style='thin', color='C1C1C1'),
                                     top=Side(border_style='thin', color='C1C1C1'),
                                     bottom=Side(border_style='thin', color='C1C1C1'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws[60]:
            cell.font = Font(name='Arial', bold=True, size=10)

        for row in ws['A61:A68']:
            for cell in row:
                cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                     right=Side(border_style='thin', color='C1C1C1'),
                                     top=Side(border_style='thin', color='C1C1C1'),
                                     bottom=Side(border_style='thin', color='C1C1C1'))
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws[72]:
            cell.font = Font(name='Arial', bold=True, size=10)
            cell.alignment = align_center

        for row in ws.iter_rows(min_row=73, max_row=73, min_col=1, max_col=6):
            for cell in row:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = align_center
                cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                     right=Side(border_style='thin', color='C1C1C1'),
                                     top=Side(border_style='thin', color='C1C1C1'),
                                     bottom=Side(border_style='thin', color='C1C1C1'))

    def format241(self, ws):

        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 13

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("241"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Months', 'Factor']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        ws.delete_rows(3)

    def format243(self, ws):
        """FL 243"""
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("243"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Liability']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        ws.delete_rows(3)

    #Format Towing and Labor Rate Table
    def format55(self, ws, boldFont):
        ws.insert_rows(3)
        ws.insert_rows(3)
        ws['A4'] = "Limit"
        ws['B4'] = "Direct Primary Option"
        ws['B5'] = "Comprehensive"
        ws['C5'] = "Collision"

        ws.merge_cells('A4:A6')
        ws.merge_cells('B4:C4')

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, italic=False, name='Arial', size=10)
        ws['A2'].font = italic

        for cell in ws['4:4']:
            cell.font = boldFont
        for cell in ws['5:5']:
            cell.font = boldFont
        for cell in ws['6:6']:
            cell.font = boldFont

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws['5:5']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for cell in ws['6:6']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        ws.column_dimensions['B'].width = self.pixelsToInches(125)
        ws.column_dimensions['C'].width = self.pixelsToInches(100)

        for cell in ws['35:35']:
            cell.font = bold

        for cell in ws['45:45']:
            cell.font = bold

        for cell in ws['34:34']:
            cell.border = None

        for cell in ws['44:44']:
            cell.border = None

        ws.insert_rows(34)
        ws.insert_rows(34)
        ws.insert_rows(34)
        ws.insert_rows(34)
        ws.insert_rows(34)
        ws.insert_rows(34)


        ws['A35'] = "Apply a factor of 0.74 for Legal Liability Option."
        ws['A36'] = "Apply a factor of 1.15 to the Legal Liability premium for Direct Excess Option."
        ws['A38'] = "255.D. Deductibles"
        ws['A39'] = "Apply the following factors for additional deductibles:"

        for cell in ws['38:38']:
            cell.font = italic

        ws['A60'] = "255.E.2. Garagekeepers Enhancement Endorsement"

        for cell in ws['60:60']:
            cell.font = italic

        ws['A62'] = "Apply the following factor to the Garagekeepers premium:"

        ws['A64'] = "Factor"
        ws['A65'] = '1.100'

        for cell in ws['64:64']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = bold

        for cell in ws['65:65']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'),
                                 right=Side(border_style='thin', color='C1C1C1'),
                                 top=Side(border_style='thin', color='C1C1C1'),
                                 bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.merge_cells('A64:C64')
        ws.merge_cells('A65:C65')

        for col in range(1, 4):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(7, 34):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#,0'

        # Remove all existing row and column breaks

    #Format Towing and Labor Rate Table
    def format68(self, ws, boldFont):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 25

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("268"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Coverage', 'Equipped', 'Person','Factor']):
                        cell.font = bold
                        cell.alignment = align_down
                        cell.border = border
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        # Adding headers
        # Merge and format
        ws.insert_rows(3)
        ws.merge_cells('B4:C4')
        cell_bc = ws['B4']
        cell_bc.value = "Educational Institutions"
        cell_bc.font = bold
        cell_bc.alignment = align_center
        cell_bc.border = border

        # Merge and format
        ws.merge_cells('D4:E4')
        cell_de = ws['D4']
        cell_de.value = "Commercial Driving Schools"
        cell_de.font = bold
        cell_de.alignment = align_center
        cell_de.border = border


    def format267(self, ws):
        """FL 267"""
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("267"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Factor']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        ws.delete_rows(3)

    def format269(self, ws):
        """FL 269"""
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("269"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Factor']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        ws.delete_rows(3)

    def format309(self, ws):
        """FL 309"""
        ws.column_dimensions["A"].width = self.pixelsToInches(460)
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("309"):
                        cell.font = italic
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        ws.delete_rows(3)

    # Format Rule 72

    def format72(self, ws):
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        # Define styles
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("272"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Factor', 'Coverage']):
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("272"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border
                        cell.number_format = "0.000"

    # Format Golf Carts and Low Speed Vehicles
    def format73(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        ws.column_dimensions['B'].width = self.pixelsToInches(125)
        ws.column_dimensions['C'].width = self.pixelsToInches(150)
        ws.column_dimensions['D'].width = self.pixelsToInches(125)

    # Format Law Enforcement agencies
    def format274(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = self.pixelsToInches(100)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("274"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Coverage']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif cell.value == "Factor":
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)) and not str(cell.value).startswith("274"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        for col in range(2, 3):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(5, 19):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '#,##0.00'

        # Format Mobile Homes

    def format276(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 26
        for col_idx in range(2, 5):  # B=2, E=5
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("276"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Coverage', 'Class', 'Liability', 'Physical', 'Other']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif cell.value == "Factor":
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("276"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        # for col in range(3, 6):
        #     char = get_column_letter(col)  # Letter representing the current column
        #     for row in range(5, 9):
        #         cell = ws[char + str(row)]
        #         cell.number_format = cell.number_format = '#,##0.00'

    # Format Motorcycles -Add table headers and descriptions... State Specific?

    def format77(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(135)
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("277"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Engine', 'Factor', 'Coverage']):
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("277"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

    # Format rule 78
    def format78(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)

    #Format Towing and Labor Rate Table
    def format79(self, ws, boldFont):
        """
        The below formatting function is very poorly made and VERY annoying to edit. Sadly don't have time to remake this one.
        """
        ws.insert_rows(2)
        ws.insert_rows(8)
        ws.insert_rows(8)
        ws.insert_rows(8)
        ws.insert_rows(8)
        ws.insert_rows(8)
        ws.insert_rows(14)


        ws['A4'] = "Per Auto"
        ws['B4'] = "Min. Premium"
        ws['A9'] = "279.B.2. Physical Damage"
        ws['A11'] = "Coverage is provided on a non-reporting basis"
        ws['A12'] = "       Comprehensive - $100/$500 deductible:"

        ws.merge_cells('A4:B4')

        for cell in ws['4:4']:
            cell.font = boldFont
        for cell in ws['5:5']:
            cell.font = boldFont
        for cell in ws['13:13']:
            cell.font = boldFont
        for cell in ws['17:17']:
            cell.font = boldFont
        for cell in ws['9:9']:
            cell.font = Font(italic=True)
        for cell in ws['3:3']:
            cell.font = Font(italic=True, bold=False)

        for cell in ws['7:7']:
            cell.border = None
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        for cell in ws['12:12']:
            cell.border = None
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        ws.column_dimensions['A'].width = self.pixelsToInches(130)
        ws.column_dimensions['B'].width = self.pixelsToInches(130)
        ws.column_dimensions['C'].width = self.pixelsToInches(130)
        ws.column_dimensions['D'].width = self.pixelsToInches(100)
        ws.column_dimensions['E'].width = self.pixelsToInches(100)

        cell = ws["B6"]
        cell.number_format = cell.number_format = '$#0'
        cell = ws["A6"]
        cell.number_format = cell.number_format = '$#,##0.00'
        cell = ws["B14"]
        cell.number_format = cell.number_format = '$#,##0.00'

        for col in range(2, 5):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(18, 23):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#,##0.00'

        ws.insert_rows(16)
        ws.delete_rows(14)

        ws['A16'] = "Collision (Rate per $100):"
        ws["A16"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws["A16"].border = None
    # Format Snowmobiles

    def format80(self, ws):

        ws.column_dimensions['A'].width = self.pixelsToInches(205)

        for col in range(2, 3):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, 9):
                cell = ws[char + str(row)]
                cell.number_format = '0.00'

    # Format Mobile and Farm Equipment

    def format81(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, 4):  # B=2, D=4
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("281"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Coverage', 'Equipment']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("281"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border

        for col in range(2, 4):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(5, 24):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '#,##0.00'

    def format283(self,ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 18

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("283"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Deductible', '$50,000', '$100,000','Territory','Personal','Miscellaneous','Vehicle', 'Factor', 'Coverage']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border


    def format284(self,ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_down = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.column_dimensions['A'].width = 15
        for col_idx in range(2, 4):  # B=2, D=4
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("284"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in
                             ['Vehicles', 'Coverage']):
                        cell.font = bold
                        cell.alignment = align_down
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("284"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border


    # Format Towing and Labor Rate Table
    def format88(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(130)

    #Format Towing and Labor Rate Table
    def format89(self, ws):
        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))
        ws.column_dimensions['A'].width = self.pixelsToInches(110)
        ws.column_dimensions['B'].width = self.pixelsToInches(110)
        ws.column_dimensions['C'].width = self.pixelsToInches(110)

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith('289'):
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in
                             ['Factor', 'Limit', 'Premium','Total','Coverage','Rate']) or 'Volunteer' == str(
                            cell.value):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border

        # ws['A62'].border = None
        # ws['A62'].font = Font(name='Arial', size=10, italic = False)
        # ws['A62'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        #
        # ws['A63'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        # ws['A63'].border = None
        # ws['A63'].font = Font(name='Arial', size=10, italic = False)

    #Format Towing and Labor Rate Table
    def format92(self, ws):

        ws['B6'].alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)
        ws['B7'].alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)

        ws['B4'].value = "Trucks, Tractors, Trailers"
        ws['C4'].value = "School and Church Buses"
        ws['D4'].value = "All Other Buses"
        ws['E4'].value = "Van Pools"
        ws['F4'].value = "Taxicabs and Limousines"
        ws['G4'].value = "Private Passenger Types"


        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['B'].width = self.pixelsToInches(140)
        ws.column_dimensions['C'].width = self.pixelsToInches(100)
        ws.column_dimensions['D'].width = self.pixelsToInches(100)
        ws.column_dimensions['E'].width = self.pixelsToInches(100)
        ws.column_dimensions['F'].width = self.pixelsToInches(100)
        ws.column_dimensions['G'].width = self.pixelsToInches(140)

    # Format Rental Reimbursement
    def format293(self, ws):
        # Define styles
        italic_font = Font(italic=True, name = "Arial", size = 10)
        bold_font = Font(bold=True, name = "Arial", size = 10)
        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))



        # Set column A width to 175 pixels
        ws.column_dimensions['A'].width = 175 / 7  # openpyxl uses width in characters, not pixels
        for col_idx in range(2, 8):  # B=2, H=8
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        # Iterate through all cells in the worksheet

        # Check if the cell value is "Factor" or "Rate Per Named Individual"
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in ["Factor", "Rate Per Named Individual"] and cell.column == 1:
                    # Get the row index of the current cell
                    row_idx = cell.row
                    # Delete the above two rows
                    ws.delete_rows(row_idx - 2, 2)


        for row in ws.iter_rows():
            for cell in row:
                # Apply italics if the cell contains the word '293'
                if '293' in str(cell.value):
                    cell.font = italic_font

                # The re search exists cause f*** excel
                if cell.value is not None and re.search(r'[a-zA-Z0-9]', str(cell.value)) and not ('293' in str(cell.value)):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text = True)

        # Apply Bold font to A1
        ws['A1'].font = bold_font

    def format94(self, ws):

        ws.column_dimensions['A'].width = self.pixelsToInches(160)

    # Format Audio, Visual, and Data Electronic Equipment

    def format95(self, ws):

        ws.column_dimensions['A'].width = self.pixelsToInches(140)

        for col in range(2, 3):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, 25):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#,0'

    # Format Tapes, Records, and Discs coverage

    def format96(self, ws):

        ws.column_dimensions['A'].width = self.pixelsToInches(130)
        cell = ws["A5"]
        cell.number_format = cell.number_format = '$#,0'

    def format297(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(85)
        # Define styles
        italic_font = Font(italic=True, name = "Arial", size = 10)
        bold_font = Font(bold=True, name = "Arial", size = 10)
        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(100)

        # Apply Bold font to A1
        ws['A1'].font = bold_font
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)

                    elif (str(cell.value).startswith("Single Limits") or str(cell.value).startswith("Uninsured Motorists")): # Rule name
                        cell.font = Font(italic=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

                    elif str(cell.value) in ['Limit', 'PPT','Total Number of Exposures', 'Other Than PPT',"Non-owned-Autos Per Employee", "Hired Autos per $100 Cost of Hire"]:
                        if str(cell.value) == 'Limit':
                            start_row = cell.row
                            col = cell.column
                            end_row = start_row
                            while ws.cell(row=end_row + 1, column=col).value in (None, ''):
                                end_row += 1
                                if end_row > 30:
                                    break
                            if end_row > start_row:
                                ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

                        if (str(cell.value) == 'PPT') and (str(ws.cell(row=cell.row-1, column=cell.column).value == 'Total Number of Exposures')):
                            start_col = cell.column
                            row_ = cell.row
                            end_col = start_col
                            while 'Other Than PPT' != str(ws.cell(row=row_, column=end_col + 1).value):
                                end_col += 1

                                if end_col > 30: # Search has gone too far, this item doesn't have Other Than PPT.
                                    end_col = start_col
                                    break

                            ws.merge_cells(start_row=row_, start_column=start_col, end_row=row_, end_column=end_col)
                            non_ppt_col = end_col + 1
                            ppt_span = end_col - start_col + 1
                            ws.merge_cells(start_row=row_, start_column=non_ppt_col, end_row=row_,
                                           end_column=non_ppt_col + ppt_span - 1)

                        if (str(cell.value) == 'Total Number of Exposures') and (cell.column < 3):
                            start_col = cell.column
                            row_ = cell.row
                            end_col = start_col
                            while str(ws.cell(row=row_, column=end_col + 1).value) != 'Total Number of Exposures':
                                end_col += 1

                                if end_col > 30: # Search has gone too far, this item doesn't have Other Than PPT.
                                    end_col = start_col
                                    break

                            ws.merge_cells(start_row=row_, start_column=start_col, end_row=row_, end_column=end_col)

                            second_start_col = end_col + 1
                            span = end_col - start_col + 1
                            ws.merge_cells(start_row=row_, start_column=second_start_col, end_row=row_,
                                           end_column=second_start_col + span - 1)

                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border

                        # Last case of Other Than PPT not fitting in the row height for single limits
                        if str(cell.value) == "Other Than PPT" and cell.column == 3:
                            ws.row_dimensions[cell.row].height = 40

                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border

    #Format Towing and Labor Rate Table
    def format98(self, ws):
        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        ws.column_dimensions['B'].width = self.pixelsToInches(125)
        ws.column_dimensions['C'].width = self.pixelsToInches(125)
        ws.column_dimensions['D'].width = self.pixelsToInches(125)
        ws.column_dimensions['E'].width = self.pixelsToInches(125)
        ws.column_dimensions['F'].width = self.pixelsToInches(125)


        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("298") and any(
                            char.isdigit() for char in str(cell.value)):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif cell.column == "A" and any(word in str(cell.value) for word in
                             ['Fire', 'Specified', 'Comprehensive']) and not str(cell.value).startswith("298"):
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    elif any(word in str(cell.value) for word in
                             ['Coverage','Deductible', 'Zone-Rated', 'Collision', 'All Perils',"Factor",'$100/500','$250/1,000','$500/2,500']):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        cell.border = border


        ws["A17"] = "For Physical Damage premium computations, when the Deductible Discount Factor is subtracted from the Vehicle Value Factor,"
        ws["A18"] = "the resulting value (Vehicle Value Factor – Deductible Discount Factor) is subject to a minimum value of 0.10"
        ws["A17"].font = Font(name='Arial', size=10)
        ws["A18"].font = Font(name='Arial', size=10)
        # Making some space between the next rule and text above.
        ws.insert_rows(19)

        ws.insert_rows(3)
        # Adding headers to first table.
        # Merge and format B3:C3
        ws.merge_cells('B4:C4')
        cell = ws['B4']
        cell.value = "Property Damage Per Accident"
        cell.font = Font(bold=True, name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        # Merge and format D3:E3
        cell = ws['D4']
        ws.merge_cells('D4:E4')
        cell.value = "Combined Single Limit"
        cell.font = Font(bold=True, name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        cell = ws['A4']
        cell.value = str(ws['A5'].value) # Moving coverage name up so it doesn't get deleted via merging.
        ws.merge_cells('A4:A5')
        cell.font = Font(bold=True, name='Arial', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

        for row in ws.iter_rows(min_row = 5, max_row = 5, min_col=1, max_col=5):
            for cell in row:
                cell.border = border

    #Format Zone Base Rates Table
    def formatZoneRates(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(60)
        ws.column_dimensions['B'].width = self.pixelsToInches(175)

        for col_idx in range(3, 6):  # C=3, E=5
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(110)


        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif 'Zone' in str(cell.value) and any(char.isdigit() for char in str(cell.value)) or 'Medical' in str(cell.value)\
                            or 'Personal' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in ['Zone', 'Description', 'Liability', 'Coll.','Comp.', 'Factor']) and 'Medical' not in str(cell.value) \
                            and 'Personal' not in str(cell.value):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)) and '225.D.' not in str(cell.value):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border

    def format225C2(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(160)
        ws.column_dimensions['B'].width = self.pixelsToInches(85)
        ws.column_dimensions['C'].width = self.pixelsToInches(85)
        ws.column_dimensions['D'].width = self.pixelsToInches(85)


        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '225.C.2.' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in ['Class','Factor']):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)) and '225.C.2.' not in str(cell.value):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border


        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(5, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '#,###0.000'
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def format225C3(self, ws):
        ws.column_dimensions['A'].width = 64

        for col_idx in range(2, 7):  # B=2, F=6
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))


        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                if row == 4:
                    cell.font = Font(name='Arial', size=9, bold=True)
                elif col in range(1,3):
                    cell.border = border
                elif col in range(2,ws.max_column + 1) :
                    cell.border = border
                    cell.number_format = cell.number_format = '#,###0.000'

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col == 1: # Applying unique column width and no decimal formatting to column A
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
                    cell.font = Font(name='Arial', size=9)

        ws['A2'].font = Font(name='Arial', size=10, italic=True)
        ws['A1'].font = Font(name='Arial', size=10, bold=True)

    #Format Towing and Labor Rate Table
    def formatTowingAndLabor(self, ws):

        ws.column_dimensions['A'].width = 15

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '317.B' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in ["Limit","Rate"]):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border
                        cell.number_format = '0.00'

        ws['A3'].value = 'Premium per Auto Per-disablement limit'
        ws['A3'].alignment = Alignment(wrap_text=False)

    #Format School Bus Operations Rate Table
    def formatMedPayments(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        #ws['A9'] = Territory92a

    # Format ILFs
    def format100(self, ws):
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20


    def format301(self, ws):
        ws.column_dimensions['A'].width = 21
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        for col_idx in range(4, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        bold_cells = [
        "Price Bracket", "Current Model Year", "First Preceding Model Year", "2nd","3rd", "4th", "5th", "6th", "7th",
        "8th", "9th", "10th", "11th", "12th", "13th", "14th", "15th", "16th", "17th", "18th", "19th", "20th",
        "21st", "22nd", "23rd", "24th", "25th", "26th", "27th and older", "Vehicles","Trucks","Private",
        ]


        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '301.C' in str(cell.value) or '301.D' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in bold_cells):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border

    def format301VA(self,ws):
        ws.column_dimensions['A'].width = 21
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        for col_idx in range(4, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        bold_cells = [
        "Price Bracket", "Current Model Year", "First Preceding Model Year", "2nd","3rd", "4th", "5th", "6th", "7th",
        "8th", "9th", "10th", "11th", "12th", "13th", "14th", "15th", "16th", "17th", "18th", "19th", "20th",
        "21st", "22nd", "23rd", "24th", "25th", "26th", "27th and older", "Vehicles","Trucks","Private",
        ]


        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline logic
                    self.insert_newline_before_301Arules(cell.value)
                    self.insert_newline_before_301Brules(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '301.A' in str(cell.value) or '301.B' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in bold_cells):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = border



    def format301D1(self, ws):
        ws.column_dimensions['A'].width = 21
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 13
        ws.column_dimensions['D'].width = 13
        ws.column_dimensions['E'].width = 13
        ws.column_dimensions['F'].width = 13
        ws.column_dimensions['G'].width = 13
        ws.column_dimensions['H'].width = 13
        ws.column_dimensions['I'].width = 13
        ws.column_dimensions['J'].width = 13
        ws.column_dimensions['k'].width = 13
        ws.column_dimensions['L'].width = 13


        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        bold_cells = [
        "Vehicle","Truck","Private","All Ages","Factor","Trailer","Semitrailer","Price"
        ]


        # this is pretty messy, but was in a rush.
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '301.C' in str(cell.value) or '301.D' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in bold_cells) or (str(cell.value) == "Trucks, Tractors, And Trailers") or (str(cell.value) == "Private Passenger Types"):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                        cell.border = border



    def format301D2(self, ws):
        ws.column_dimensions['A'].width = 21
        ws.column_dimensions['B'].width = 13
        ws.column_dimensions['C'].width = 13


        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        bold_cells = [
        "Vehicle","Truck","Private","All Ages","Factor","Trailer","Semitrailer","Price"
        ]


        # this is pretty messy, but was in a rush.
        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    # Apply newline functions
                    self.insert_newline_before_first_open_parenthesis(cell.value)
                    self.insert_newline_before_dollar(cell.value)

                    # Apply specific formatting
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif '301.C' in str(cell.value) or '301.D' in str(cell.value):  # Detect if it contains Zone + digits
                        cell.font = Font(italic=True, name='Arial', size=10)
                    elif any(word in str(cell.value) for word in bold_cells) or (str(cell.value) == "Trucks, Tractors, And Trailers") or (str(cell.value) == "Private Passenger Types"):
                        cell.font = Font(bold=True, name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
                        cell.border = border
        ws['A5'].font = Font(name='Arial', size=10, bold = False)
        ws['A5'].border = border
        ws['A5'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)
        ws['A6'].font = Font(name='Arial', size=10, bold = False)
        ws['A6'].border = border
        ws['A6'].alignment = Alignment(horizontal='center', vertical='center',wrap_text=True)

    #Format School Bus Operations Rate Table
    # 416
    def formatExperienceRating(self, ws):
        # This format function was initially done weird. Not worth redoing.
        ws['A2'].font = Font(italic=True, name='Arial', size=10)
        for col in range(2, 3):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, 8):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '#,####0.0000'
        ws['A3'].border = None
        ws['A8'].border = None
        ws['B9'].border = None

    #Format School Bus Operations Rate Table
    def formatSchoolBusOps(self, ws, fontItalic):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        ws['A9'] = "454.B. Commercial Broad Form Endorsement School Bus Operators"
        ws['A10'] = ""
        ws['A11'] = "Charge is " + str(MinPremFactor154) + " of the scheduled auto liability and physical damage premium under Business Auto Coverage form, subject to minimum premiums of " + str(MinPrem154) + ". Minimum premium is not subject to increased limit factors."
        for cell in ws['9:9']:
            cell.font = fontItalic
        for cell in ws['11:11']:
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws.merge_cells('A11:F11')
        ws.row_dimensions[11].height = self.pixelsToInches(500)

    # format ambulance services sheet
    def formatAmbulance(self, ws):
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        # Define styles
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("264"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Factor', 'Coverage']):
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("264"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border
                        cell.number_format = "0.000"

    # Format ANTIQUEAUTOS Table
    def formatANTIQUEAUTOS(self, ws):
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15

        ws['A2'] = '266.B. Premium Computation'
        ws['A2'].font = Font(italic = True, name='Arial', size=10)

    def formatFireDepartments(self,ws):
        ws.column_dimensions['A'].width = 20
        for col_idx in range(2, 3):  # B=2, C=3
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = self.pixelsToInches(85)

        # Define styles
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, name='Arial', size=10)

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("271"):
                        cell.font = italic
                    elif any(word in str(cell.value) for word in ['Factor', 'Coverage']):
                        cell.font = bold
                        cell.alignment = align_center
                    elif any(char.isdigit() or char.isalpha() for char in
                             str(cell.value)) and not str(cell.value).startswith("271"):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border
                        cell.number_format = "0.000"

    # Format LEASINGORRENTALCONCERNS Table
    def formatLEASINGORRENTALCONCERNS(self, ws, italicFont):
        # Normal rule 275 formatting
        if self.StateAbb != "VA":
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 14
            ws.insert_rows(3)
            ws['A11'] = '275.B.4.b. Exclusion of Certain Leased Autos'
            ws['A12'] = 'Coverage is provided at no charge.'
            ws['A14'] = '275.B.5.b. Schedule of Limits for Owned Autos'
            ws['A15'] = 'Coverage is provided at no charge.'
            for cell in ws['11:11']:
                cell.font = italicFont
            for cell in ws['14:14']:
                cell.font = italicFont

            for col in range(2, 3):
                char = get_column_letter(col)  # Letter representing the current column
                for row in range(5, 10):
                    cell = ws[char + str(row)]
                    cell.number_format = cell.number_format = '$#0'
        # VA Specific formatting
        elif self.StateAbb == "VA":
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 14
            ws.column_dimensions['D'].width = 14

            # insert missing rule text
            ws.insert_rows(10, amount=2)
            ws['A10'] = "275.B.1.(b). Short Term - Autos Leased by the Hour, Day, or Week"

            # Define styles
            italic = Font(bold=False, italic=True, name='Arial', size=10)
            bold = Font(bold=True, name='Arial', size=10)

            align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

            border = Border(left=Side(border_style='thin', color='C1C1C1'),
                            right=Side(border_style='thin', color='C1C1C1'),
                            top=Side(border_style='thin', color='C1C1C1'),
                            bottom=Side(border_style='thin', color='C1C1C1'))

            # Apply styles
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("275"):
                            cell.font = italic
                        elif str(cell.value).startswith("("):
                            cell.font = italic
                        elif any(word in str(cell.value) for word in ['Factor', 'Coverage', 'Vehicle Type', 'Base Premium',
                                                                      'Trucks', 'Tractors', 'Trailers', "Up to 22'", "Over 22'"]) \
                                                                      and "Trucks, Tractors and Trailers" not in str(cell.value):
                            cell.font = bold
                            cell.alignment = align_center
                        elif any(char.isdigit() or char.isalpha() for char in
                                 str(cell.value)) and not str(cell.value).startswith(
                            "275"):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
                            if cell.row <= 9: # Formats the base premiums seperately from factors
                                cell.number_format = "$#0"
                            if cell.row > 9:
                                cell.number_format = "0.000"


        for col in range(2, 3):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(5, 10):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#0'

    # Format rule 90.hired auto Table
    def formathiredauto(self, ws):
        ws.column_dimensions['B'].width = 14.5
        ws['A3'] = 'The following rates apply to Liability Limits:'
        ws['A9'] = 'Rate per $100 Cost of Hire:'
        for cell in ws['3:3']:
            cell.alignment = Alignment(wrap_text=False)
        for cell in ws['7:7']:
            cell.alignment = Alignment(wrap_text=False)
        for cell in ws['9:9']:
            cell.alignment = Alignment(wrap_text=False)
        ws.insert_rows(8)
        ws.insert_rows(2)

        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):  # Check if the cell contains a number
                    cell.number_format = openpyxl.styles.numbers.FORMAT_CURRENCY_USD_SIMPLE  # Apply simple USD currency format


        for row in ws.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value == "Minimum Premium":
                    row_num = cell.row
                    for col_cells in ws.iter_rows(min_row=row_num, max_row=row_num):
                        for num_cell in col_cells:
                            num_cell.number_format = '"$"#,##0'
                    break  # Exit loop once formatting is applied

        ws["B7"].number_format = '"$"#,##0'

    # Format RULE 315. BUSINESS INTERRUPTION COVERAGE
    def formatBusinessInterruption(self, ws):
        ws.insert_rows(8)
        ws.insert_rows(23)
        ws.insert_rows(29)
        ws['A3'] = 'Base Rates Per $100 Of Insurance'
        ws['A8'] = '315.B.3. Extended Business Income Additional Coverage factors:'
        ws['A23'] = '315.B.4. Waiting Period factors:'
        ws['A29'] = '315.B.5. Percentage of Insurance to Exposure factors:'
        ws.column_dimensions['A'].width = 44
        ws.column_dimensions['B'].width = 23
        ws.column_dimensions['C'].width = 23
        ws.column_dimensions['C'].width = 23

        ws.insert_rows(3)
        ws['A2'].font = Font(italic=True, name='Arial', size=10)
        ws['A9'].font = Font(italic = True, name='Arial', size=10)
        ws['A24'].font = Font(italic = True, name='Arial', size=10)
        ws['A30'].font = Font(italic = True, name='Arial', size=10)
        ws['A8'].border = None
        ws['A23'].border = None
        ws['A29'].border = None

    # Format RULE 107. FELLOW EMPLOYEE COVERAGE
    def formatFellowEmployee(self, ws):
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['A'].width = 17
        ws.column_dimensions['A'].width = 20

        for cell in ws['3:3']:
            cell.alignment = Alignment(wrap_text=False)

    # Format RULE 310. LOSS OF USE EXPENSES - RENTAL VEHICLES - OPTIONAL LIMITS
    def formatLossofUse(self, ws):
        ws.column_dimensions['A'].width = 40
        for col in range(2, 3):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, 8):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#,0'

    # Format RULE 113. SILICA OR SILICA-RELATED DUST LIABILITY
    def format113(self, ws):
        ws['A4'] = 'Coverage is provided at no cost.'
        ws['A4'].font = Font(name='Arial', size=10)


    # Format RULE 125
    def format125(self, ws):
        ws.column_dimensions['A'].width = 53
        ws['A5'].number_format = '$#,0'
        ws['A10'].number_format = '$#,0'

    # Format RULE 127
    def format127(self, ws):
        ws.column_dimensions['A'].width = 66
        ws['A4'] = 'Physical Damage Factor:'

    # Format RULE 103
    def format103(self, ws):
        ws.column_dimensions['C'].width = self.pixelsToInches(150)
        ws['A9'].font = Font(name='Arial', size=10)
        for col in range(3, 4):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(5, 9):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#,0'



    # Format RULE 126
    def format126(self, ws):
        ws.column_dimensions['A'].width = 52

        for cell in ws['A:A']:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        for col in range(2, 5):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(6, 8):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '$#,0'

    #Format Schedule Rating Table
    def format117(self, ws, fontItalic):
        if self.StateAbb == "NE":
            ws['A1'].value = "RULE 417. FlEX RATING"

        ws['A3'] = "417.1.b. Minimum Eligible Premium"
        ws['A6'] = "417.2. Maximum Modification: +/- " + str(self.SchedRatingMod) + "%"
        ws['A4'].number_format = '$#,0'
        ws.column_dimensions['A'].width = 33
        ws['A3'].font = Font(italic = True, bold = False, name = "Arial",size = 10)
        ws['A6'].font = Font(italic = True, bold = False, name = "Arial",size = 10)

    # Format Schedule Rating Table
    def format150(self, ws):

        ws['A29'] = 'Use a 1.000 factor for large fleets (20+ Vehicles)'
        for col in range(1, 2):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, 28):
                cell = ws[char + str(row)]
                cell.number_format = cell.number_format = '#0'

    #Format Rate Capping Table
    def format151(self, ws, fontItalic):
        ws.column_dimensions['B'].width = self.pixelsToInches(100)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("451"):
                        cell.font = italic
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center
                        cell.border = border
        ws.delete_rows(10)
        ws.delete_rows(9)
        ws.delete_rows(4)
        ws.delete_rows(3)

        ws["B3"].number_format = "0.00"
        ws["B4"].number_format = "0.00"
        ws["B7"].number_format = "0"

        if str(ws["B3"].value) != "Not Applicable":
            ws.column_dimensions['B'].width = self.pixelsToInches(55)

    #Format School Bus Operations Rate Table
    def formatDp1(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)

    # Format RULE T1
    def formatT1(self, ws):
        ws['A3'] = 'Apply the following discount to appropriate coverage premiums:'
        ws['A3'].alignment = Alignment(wrap_text=False)
        ws['A5'].number_format = '0%'

    #Format Rate Capping Table
    def formatR1(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(60)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)
        ws.column_dimensions['C'].width = self.pixelsToInches(85)
        ws.column_dimensions['D'].width = self.pixelsToInches(150)
        ws.column_dimensions['E'].width = self.pixelsToInches(200)
        ws.column_dimensions['F'].width = self.pixelsToInches(160)
        ws.column_dimensions['I'].width = self.pixelsToInches(110)

        italic = Font(bold=False, italic=True, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        for row in ws.iter_rows():
            for cell in row:
                if cell.value:
                    if cell.coordinate == 'A1':
                        cell.font = Font(bold=True, name='Arial', size=10)
                    elif str(cell.value).startswith("R1"):
                        cell.font = italic
                    elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = align_center

    def formatA1(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        if self.StateAbb == "CT":

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A1"):
                            cell.font = italic
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(4)
            ws.delete_rows(3)

        if self.StateAbb == "KS":
            ws.column_dimensions['A'].width = self.pixelsToInches(300)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A1"):
                            cell.font = italic
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(4)
            ws.delete_rows(3)

        if self.StateAbb == "ND":
            ws.column_dimensions['A'].width = self.pixelsToInches(300)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A1"):
                            cell.font = italic
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(4)
            ws.delete_rows(3)

        if self.StateAbb == "RI":
            ws.column_dimensions['A'].width = self.pixelsToInches(85)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A1"):
                            cell.font = italic
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(4)
            ws.delete_rows(3)

    def formatA2(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        bold = Font(bold=True, italic=False, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        if self.StateAbb == "MI":

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A2"):
                            cell.font = italic
                        elif any(word in str(cell.value) for word in ['Rate']):
                            cell.font = bold
                            cell.alignment = align_center
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(3)

        if self.StateAbb == "NV":
            ws.column_dimensions['A'].width = self.pixelsToInches(400)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A2"):
                            cell.font = italic
                        elif any(word in str(cell.value) for word in ['Factor', 'Passive Restraint System']):
                            cell.font = bold
                            cell.alignment = align_center
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(3)

        if self.StateAbb == "RI":
            ws.column_dimensions['A'].width = self.pixelsToInches(500)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A2"):
                            cell.font = italic
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(4)
            ws.delete_rows(3)

    def formatA3(self, ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        if self.StateAbb == "ND":
            ws.column_dimensions['A'].width = self.pixelsToInches(300)
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A3"):
                            cell.font = italic
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(4)
            ws.delete_rows(3)

    def formatA4(self,ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold = Font(bold=True, italic=False, name='Arial', size=10)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        if self.StateAbb == "MI":
            ws.column_dimensions['A'].width = self.pixelsToInches(180)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A4"):
                            cell.font = italic
                        elif any(word in str(cell.value) for word in ['Rate']):
                            cell.font = bold
                            cell.alignment = align_center
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border
            ws.delete_rows(10)
            ws.delete_rows(4)

    def formatA5(self,ws):
        italic = Font(bold=False, italic=True, name='Arial', size=10)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold = Font(bold=True, italic=False, name='Arial', size=10)

        border = Border(left=Side(border_style='thin', color='C1C1C1'),
                        right=Side(border_style='thin', color='C1C1C1'),
                        top=Side(border_style='thin', color='C1C1C1'),
                        bottom=Side(border_style='thin', color='C1C1C1'))

        if self.StateAbb == "NJ":
            ws.column_dimensions['A'].width = self.pixelsToInches(180)

            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        if cell.coordinate == 'A1':
                            cell.font = Font(bold=True, name='Arial', size=10)
                        elif str(cell.value).startswith("A4"):
                            cell.font = italic
                        elif any(word in str(cell.value) for word in ['Rate']):
                            cell.font = bold
                            cell.alignment = align_center
                        elif any(char.isdigit() or char.isalpha() for char in str(cell.value)):  # Apply formatting if the cell contains a digit or character
                            cell.font = Font(name='Arial', size=10)
                            cell.alignment = align_center
                            cell.border = border

    def overideFooter(self, ws, companies):
        # Dictionary to map company abbreviations to full names
        company_names = {
            'NAFF': 'Nationwide Affinity Insurance Company of America',
            'NACO': 'Nationwide Assurance Company',
            'NGIC': 'Nationwide General Insurance Company',
            'CCMIC' : 'Colonial County Mutual Insurance Company',
            'HICNJ' : 'Harleysville Insurance Company of New Jersey',
            'NICOF': 'Nationwide Insurance Company of Florida',
            'NMIC' : 'Nationwide Mutual Insurance Company',
            "AICOA" : "Allied Insurance Company of America",
            "NICOA" : "Nationwide Insurance Company of America",
            "NPCIC" : "Nationwide Property Casualty Insurance Company",
            "NWAG" : "Nationwide Agribusiness Insurance Company"
        }

        company_numbers = {
            'NAFF': '21',
            'NACO': '27',
            'NGIC': '25',
            'CCMIC' : '12',
            'HICNJ' : 'G',
            'NICOF': '40',
            'NMIC' : '01',
            "AICOA" : "32",
            "NICOA" : "07",
            "NPCIC" : "28",
            "NWAG" : "AG"
        }

        # Clear existing footers
        ws.oddFooter.left.text = ""
        ws.oddFooter.right.text = ""
        # Check if any of the abbreviated companies are in the split string

        # Special Case for DP-1, over-rides regular behavior. NGIC only page.
        if "RULE DP-1" in str(ws["A1"].value):
            name = company_names["NGIC"]
            ws.oddFooter.left.text = f"{name}"
            ws.oddFooter.right.text = ws.oddFooter.center.text
            ws.oddFooter.center.text = ""
            ws.oddFooter.left.text = ws.oddFooter.left.text.strip()  # Removing annoying spaces.

            return None

        # Check if any of the abbreviated companies are in the split string
        if isinstance(companies, str):
            companies = [company.strip() for company in companies.replace(' ', '').split(',')]
        elif isinstance(companies, list):
            companies = [company.strip() for company in companies]
        # Check if any of the abbreviated companies are in the split string
        included_companies = [company for company in companies if company in company_names.keys()]

        # Move all formatting and values from ws.oddFooter.center to ws.oddFooter.right
        # Altering the tab name such that it does not get cut off in the footer.
        tab_name = copy.deepcopy(ws.title)

        if len(ws.title) > 31:
            for key, value in company_numbers.items():
                ws.title = ws.title.replace(key, value)

        ws.oddFooter.center.text = f"{self.StateAbb} - {tab_name} - &P"

        ws.oddFooter.right.text = ws.oddFooter.center.text
        ws.oddFooter.center.text = ""

        # If none are found, assume all companies are included
        if len(self.CompanyListDif) == 1:
            included_companies = self.existing_companies

        # Apply footers based on companies
        if len(included_companies) == 5:
            ws.oddFooter.left.text = (
                f"{company_names[included_companies[0]]} \n "
                f"{company_names[included_companies[1]]} \n "
                f"{company_names[included_companies[2]]} \n "
                f"{company_names[included_companies[3]]} \n"
                f"{company_names[included_companies[4]]}"
            )

        elif len(included_companies) == 4:
            ws.oddFooter.left.text = (
                f"{company_names[included_companies[0]]} \n "
                f"{company_names[included_companies[1]]} \n "
                f"{company_names[included_companies[2]]} \n "
                f"{company_names[included_companies[3]]}"
            )
        elif len(included_companies) == 3:
            ws.oddFooter.left.text = (
                f"{company_names[included_companies[0]]} \n "
                f"{company_names[included_companies[1]]} \n "
                f"{company_names[included_companies[2]]}"
            )
        elif len(included_companies) == 2:
            ws.oddFooter.left.text = (
                f"{company_names[included_companies[0]]} \n "
                f"{company_names[included_companies[1]]}"
            )
        elif len(included_companies) == 1:
            ws.oddFooter.left.text = f"{company_names[included_companies[0]]}"
        else:
            ws.oddFooter.left.text = ""


    def overideHeaderFL(self, AutoPages):
        # In FL space is needed within the right sided header for a stamp. This code moves the right header to the left header.
        # Couldn't think of a better solution that took 30 seconds to implement.

        """Does not add the spaces correctly as of 6-19-2025"""
        for sheet_name in AutoPages.sheetnames:
            ws = AutoPages[sheet_name]

            # Safely get the right header text
            right_text = ws.oddHeader.right.text or ""
            right_text = right_text.replace("\n","\n ") # Fixing that odd header formatting

            # Move the modified right header text to the left header
            left_text = ws.oddHeader.left.text or ""

            ws.oddHeader.left.text = f"{left_text}\n {right_text}"

            # Clear the right header
            ws.oddHeader.right.text = ""

    def extract_company_name(self, company_test):
        # Find the position of the first comma
        comma_index = company_test.find(',')

        # Extract the substring before the comma, or the whole string if no comma is found
        base_name = company_test[:comma_index] if comma_index != -1 else company_test

        return base_name


    # Sets up the Auto Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildBAPages(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW' or company == "MM": # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        RatePages = ExcelSettingsBA.Excel(StateAbb=self.StateAbb, State=self.State, nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = RatePages.getFontName()
        fontSize = RatePages.getFontSize()

        ################################################################################################################################################################################
        #Generates a tab for each rate page
        #Pages with multiple tables on the same page need to be added and customized via the table specific functions in the ExcelSetting.py program
        #Creates additional tables for the Migration Companies if the table values differ from NGIC
        #Comment out any of the below Rules if they are not needed for your state or delete if they are no longer offered
        ################################################################################################################################################################################

        # Performing Nesting Procedure
        self.nesting()
        # Grabbing input for file for sheets that vary
        state_sheet_exceptions = pd.read_excel(BA_INPUT_FILE, sheet_name=None, engine='openpyxl')


        if self.StateAbb == "MT":
            print("Warning: Rule 297 will not be correct.")

        if self.StateAbb == "DC":
            print("Warning: Additional PIP Base Rate Tables for 222, 232, 239 will be absent.")

        if (self.StateAbb == "MI"):
            print("Warning: Base Rate Formatting incomplete. 298 has special exceptions not yet built.")

        if  (self.StateAbb == "VA"):
            print("Warning: Due to large shifts in manual presentation, this manual is incomplete.")

        if (self.StateAbb == "NY") or (self.StateAbb == "CA"):
            print("Warning: Rule 297 for this state was not built out.")

        def sheet_fetch(rule_sheet):
            """
            Fetches the appropriate sheet names for a given state from the rule_sheet.
            Falls back to 'Default' if no match is found for a coverage or if 'SPECIAL' is the only entry.

            returns a list of table codes that require state specific items from the BA Input file.
            """

            sheet_df = state_sheet_exceptions[rule_sheet]

            # --- State and Default dataframes ---
            state_rows_df = sheet_df[sheet_df["state"] == self.StateAbb]
            default_rows_df = sheet_df[sheet_df["state"] == "Default"]

            # --- Determine coverage already included by state ---
            state_coverages = set(
                state_rows_df["coverage"].dropna().astype(str).str.strip().tolist()
            )

            # --- Default rows to supplement (coverage NOT already in state) ---
            supplemental_default_df = default_rows_df[
                ~default_rows_df["coverage"].astype(str).str.strip().isin(state_coverages)
            ]

            # --- Combine state + supplemental default ---
            combined = pd.concat([state_rows_df, supplemental_default_df], ignore_index=True)

            # --- Convert to a cleaned list of sheet names ---
            sheet_list = (
                combined["sheet"]
                .dropna()
                .astype(str)
                .str
                .strip()
            )

            sheet_list = sheet_list[~sheet_list.eq("") & ~sheet_list.str.casefold().eq("nan")].tolist()

            # --- SPECIAL logic ---
            if "SPECIAL" in sheet_list:
                non_specials = [s for s in sheet_list if s != "SPECIAL"]

                if not non_specials:
                    # SPECIAL is the only item → fallback to Default
                    return (
                        default_rows_df["sheet"]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .replace("nan", "")
                        .tolist()
                    )
                else:
                    # SPECIAL + others → drop SPECIAL
                    return non_specials

            # --- Normal case ---
            return sheet_list

        # VA has many different rule numberings. This dictionary is to prevent lots of potential ugliness in the code later.
        # This can quite easily be modified, or other dictionaries added, to support other states as well
        VARules = {
            # "Rule VAPCD": ["VAPCD.E. Premium Computation"],
            # "Rule 222.C": ["222.C.2.b. Liability Fleet Size Factors", "222.C.3.d. Collision Fleet Size Factors", "222.C.3.d. Other Than Collision Fleet Size Factors"],
            # "Rule 223.B.5": ["223.B.5. Primary Classification Factors and Statistical Codes - Truck, Tractors, and Trailers"],
            # "Rule 223.C.": ["223.C.2. Secondary Classification Factors "],
            # 222.C skipped for now
            # 222.D skipped for now
            # 224 skipped for now
            # 240 skipped for now
            "Rule 241": ["241.C.2.d. Mechanical Lift Factor", "241.C.3.c. Specified Causes of Loss Coverage Factor",
                         "241.D.(c). Commercial Lay-up Credit"],
            "Rule 267": ["267.B. Auto Body Manufacturers And Installers Factor:"],
            "Rule 268": ["268. Driver Training Owned Auto Factors",
                         "268.D.2.c. Instructors in Excess of Owned Autos Used For Driver Training Liability and Medical Payments Coverages",
                         "268.D.2.(d).(2). Medical Expense Coverage Factors"],
            "Rule 275": ["275.B.1.(a).(2).Charge the following premiums for the Applicable Vehicle Type:",
                                         "(1). Trucks, Tractors, and Trailers Factors",
                                         "(2). Private Passenger Types Factor",
                                         "(3).(a). Motorcycles Factors",
                                         "(3).(b). Snowmobiles Factors",
                                         "(3).(c). All Other Special Types Except Motor Homes Factors",
                                         "(5). Motor Homes Factors"],
            "Rule 281": ["281.C.2. Premium Computation", "281.D. Cost of Hire Basis Coverage Factors",
                         "281.E Rental Period Basis Factors"],
            # 283 skipped for now
            # "Rule 288": ["288.B.7. Individual Named Insured "],
            # "Rule 289": ["289.C.1.a.(2). Other Than Social Service Agency Risks", "289.C.1.b. Extended Non-Ownership Liability Employee Coverage Factor", "289.C.1.c.(2). Partnership and LLC Non-ownership Liability Coverage Factor", "289.C.1.d.(2). Non-ownership Liability Coverage Factor", "289.C.1.e.(3).(c) Food or Goods Delivery Risks", "289 C.1.e.(3).(e). Apply the following factors for higher limits:", "289.C.1.e.(3).(f)", "289.C.2. Social Service Agency Risks"],
            # 290 skipped for now
            # 292 skipped for now
            # "Rule 294": ["294.B.3. Rate per $100 limit for the selected coverage:"],
            # "Rule 295": ["295.B.2.a. Audio, Visual and Data Electronic Equipment"],
            # "Rule 300": ["300.C. Increased Limit Factors"]
            "Rule 301.A": ["301.A. Zone-Rated Trailers Vehicle Value Factors - Collision with Actual Cash Value Rating",
                        "301.A. Zone-Rated Non-Trailers Vehicle Value Factors - Collision with Actual Cash Value Rating",
                        "301.A. Zone-Rated Private Passenger Types Vehicle Value Factors - Collision with Actual Cash Value Rating",
                        "301.A. Non-Zone-Rated Trailers Vehicle Value Factors - Collision with Actual Cash Value Rating",
                        "301.A. All Other Vehicle Value Factors - Collision with Actual Cash Value Rating",
                        "301.A. Zone-Rated Vehicles Vehicle Value Factors - Other Than Collision With Actual Cash Value Rating",
                        "301.A. Private Passenger Types Vehicle Value Factors - Other Than Collision With Actual Cash Value Rating",
                        "301.A.  All Other Vehicles Vehicle Value Factors - Other Than Collision With Actual Cash Value Rating"],
            "Rule 301.B": ["301.B. Zone-Rated Trailers Vehicle Value Factors - Collision with Stated Amount Rating",
                        "301.B. Zone-Rated Non-Trailers Vehicle Value Factors - Collision with Stated Amount Rating",
                        "301.B. Zone-Rated Vehicles Vehicle Value Factors - Other Than Collision with Stated Amount Rating",
                        "301.B. Private Passenger Types Vehicle Value Factors - Collision with Stated Amount Rating",
                        "301.B. Non-Zone-Rated Trailers Vehicle Value Factors - Collision with Stated Amount Rating",
                        "301.B. All Other Vehicles Vehicle Value Factors - Collision with Stated Amount Rating",
                        "301.B. Private Passenger Types Vehicle Value Factors - Other Than Collision with Stated Amount Rating",
                        "301.B. All Other Vehicles Vehicle Value Factors - Other Than Collision with Stated Amount Rating"],
            "Rule 301.C.1": ["301.C.1. Liability Original Cost New Factors"],
            "Rule 301.C.2/3": ["301.C.2 Liability Vehicle Age Factors - Original Cost New Vehicles",
                        "301.C.3 Liability Vehicle Age Factors - Stated Amount Vehicles"]
        }

        # Rule VAPCD
        if self.StateAbb == "VA":
            self.compareCompanies('AccidentPreventionDiscountFactorVA_Ext')
            for CompanyTest in self.CompanyListDif: # List of company clusters "XXXX,XXXXX,...."
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = "" # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheet('Rule VAPCD '+ self.title_company_name, 'RULE VAPCD. VIRGINIA ACCIDENT PREVENTION COURSE DISCOUNT ' + self.title_company_name, 'VAPCD.E. Premium Computation', self.buildVAPCD(comp_name), False, True)
                self.overideFooter(RatePages.getWB()['Rule VAPCD '+ self.title_company_name],CompanyTest)

        #Rule 208
        self.compareCompanies('ExpenseConstant_Ext')
        for CompanyTest in self.CompanyListDif: # List of company clusters "XXXX,XXXXX,...."
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = "" # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 208 '+ self.title_company_name, 'RULE 208. EXPENSE CONSTANT ' + self.title_company_name, '208.B. Rate and Premium Computation', self.buildExpenseConstant(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 208 '+ self.title_company_name],CompanyTest)

        #Rule 222 Premium
        sheet_to_compare = sheet_fetch("222 TTT")

        self.compareCompanies(sheet_to_compare)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = "" # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 222 TTT BR '+ self.title_company_name, 'RULE 222. TRUCKS, TRACTORS, AND TRAILERS BASE RATES '+ self.title_company_name, ' ', self.buildBaseRates(comp_name, "TTT"), False, True )
            self.overideFooter(RatePages.getWB()['Rule 222 TTT BR '+ self.title_company_name],CompanyTest)

        #Rule 222 B
        self.compareCompanies(["LiabilityFleetSizeFactors_Ext","CollisionFleetSizeFactor_Ext","ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = "" # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet3tables('Rule 222 B '+ self.title_company_name, 'RULE 222. PREMIUM DEVELOPMENT - TRUCK, TRACTOR, TRAILER TYPES ' + self.title_company_name, '222.B.1.a. Liability Fleet Size Factors', self.buildTTTLiabFleetFactors(comp_name), self.buildTTTPhysDamFleetFactors(comp_name), self.buildTTTOTCFleetFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 222 B '+ self.title_company_name],CompanyTest)

        #Rule 222 C
        self.compareCompanies(["ShowroomLiabilityFactor","TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext","TruckDumpingRelativity", "TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            tables = [self.buildShowRoomFactors(comp_name),
                      self.buildShowRoom2Factors(comp_name),
                      self.buildShowRoom3Factors(comp_name),
                      self.buildShowRoom4Factors(comp_name)]

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ['222.C.2.g - 222.C.4.b.(4)  Trailers and Semi-Trailers Used As Showrooms:',
                         '222.C.3.e Heavy Dumping Factor Other Than Zone-Rated Autos:',
                         '222.C.3.e Heavy Dumping Factor Zone-Rated Autos:',
                         '222.C.3.f Heavy Farming Factor']

            RatePages.generateWorksheetTablesX('Rule 222 C'+ self.title_company_name, 'RULE 222.C  TRUCKS, TRACTORS, TRAILERS CLASSIFICATION - Special Provisions for Certain Risks ' + self.title_company_name, subtitles, tables,False, True)
            self.overideFooter(RatePages.getWB()['Rule 222 C' + self.title_company_name],CompanyTest)

        #Rule 222 E
        self.compareCompanies("AutoLayUpFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            tables = [self.buildLayupFactors(comp_name)]

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ['222.E.1.c Commercial Lay-up Credit']

            RatePages.generateWorksheetTablesX('Rule 222 E'+ self.title_company_name, 'RULE 222.E PREMIUM DEVELOPMENT - TRUCK, TRACTOR, TRAILER TYPES ' + self.title_company_name, subtitles, tables,False, True)
            self.overideFooter(RatePages.getWB()['Rule 222 E' + self.title_company_name],CompanyTest)


        #Rule 223 B.5
        # Below has been adjusted such that the zone is deleted in the format function. It's easier than rewriting the function.
        self.compareCompanies("TrucksTractorsAndTrailersPrimaryFactors_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet23B('Rule 223 B.5 '+ self.title_company_name, 'RULE 223. TRUCKS, TRACTORS, TRAILERS CLASSIFICATION ' + self.title_company_name, '223.B.5. Primary Classification Factors and Statistical Codes - Truck, Tractors, and Trailers', self.buildPrimaryFactors(comp_name), self.buildZonePrimaryFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 223 B.5 '+ self.title_company_name],CompanyTest)


        #Rule 223 C
        self.compareCompanies("TrucksTractorsAndTrailersSecondaryFactorsLiabilityComprehensiveAndSCOL_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 223 C '+ self.title_company_name, 'RULE 223. TRUCKS, TRACTORS, TRAILERS CLASSIFICATION ' + self.title_company_name, '223.C.4. Secondary Classification Factors ', self.buildTTTSecondaryFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 223 C '+ self.title_company_name],CompanyTest)


        #Rule 225 C.2
        self.compareCompanies("TrucksTractorsAndTrailersPrimaryFactors_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            tables = [self.buildZonePrimaryFactors(comp_name)]
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            subtitles = ['225.C.2. Primary Classification Factors and Statistical Codes - Zone Rated Autos']

            # You may think this looks strange. 200B is a no length invisible unicode character to add uniqueness to the titles.
            # This uniqueness is needed for the marked up macro to run correctly.
            title_start = 'RULE 225.\u200B PREMIUM DEVELOPMENT - ZONE-RATED AUTOS '

            RatePages.generateWorksheetTablesX('Rule 225.C.2 ' + self.title_company_name,title_start + self.title_company_name, subtitles, tables,False, True)
            self.overideFooter(RatePages.getWB()['Rule 225.C.2 ' + self.title_company_name],CompanyTest)


        # Rule 225 C.3
        self.compareCompanies("Secondary Classification Factors Zone Rated_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            tables = [self.buildZoneSecondaryFactors(comp_name)]
            subtitles = ['225.C.3. Secondary Classification Factors - Zone Rated Autos']

            # You may think this looks strange. 200B is a no length invisible unicode character to add 'uniqueness' to the titles.
            # This uniqueness is needed for the marked up macro to run correctly. The user will never tell the difference.
            title_start = 'RULE 225.\u200B\u200B PREMIUM DEVELOPMENT - ZONE-RATED AUTOS '

            RatePages.generateWorksheetTablesX('Rule 225.C.3 ' + self.title_company_name,title_start + self.title_company_name, subtitles, tables,False, True)
            self.overideFooter(RatePages.getWB()['Rule 225.C.3 ' + self.title_company_name],CompanyTest)

        # Rule 225
        self.compareCompanies(["ZoneRatedLiabilityBasePremium","ZoneRatedCollisionBasePremium","ZoneRatedOtherThanCollisionBasePremium"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            zone_info = self.buildZoneBaseRates(comp_name)
            zones = zone_info["Zones"].tolist()
            rate_tables = zone_info["Output Tables"]
            med_factor = zone_info["Med Factor"]
            pip_factor = zone_info["Pip Factor"]

            med_rule_name = "225.D.2.a.(2) Medical Payments Liability Factor"
            pip_rule_name = "225.D.3 Personal Injury Protection Factor"

            subtitles = zones.copy()
            tables = rate_tables.copy()

            # Add med_factor and pip_factor only if they are not None
            if med_factor is not None:
                tables.append(med_factor)
                subtitles.append(med_rule_name)
            if pip_factor is not None:
                tables.append(pip_factor)
                subtitles.append(pip_rule_name)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheetTablesX('Rule 225 Zone BR' + self.title_company_name,
                                               'RULE 225. ZONE-RATED AUTOS BASE RATES' + self.title_company_name,
                                               subtitles, tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 225 Zone BR' + self.title_company_name], CompanyTest)

        # #Rule 225 D
        self.compareCompanies("AutoLayUpFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ['225.D Fleet Size Rating Factors - Zone Rated',
                         '225.D.5.c. Commercial Lay-up Credit']

            tables = [self.buildFleetSizeRatingFactorsZone(comp_name),
                      self.buildLayupFactors(comp_name)]

            # You may think this looks strange. 200B is a no length invisible unicode character to add 'uniqueness' to the titles.
            # This uniqueness is needed for the marked up macro to run correctly. The user will never tell the difference.
            title_start = 'RULE 225.\u200B\u200B\u200B PREMIUM DEVELOPMENT - ZONE-RATED AUTOS '
            RatePages.generateWorksheetTablesX('Rule 225.D ' + self.title_company_name,
                                               title_start + self.title_company_name,
                                               subtitles, tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 225.D ' + self.title_company_name], CompanyTest)

        #Rule 231 C
        self.compareCompanies(["PrivatePassengerClassCode","PrivatePassengerTypesClassFactors_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 231 C ' + self.title_company_name,'RULE 231. PRIVATE PASSENGER TYPES ' + self.title_company_name,'231.C.2.d. Use and Operator Experience Factors', self.build231C(comp_name),False, True)
            self.overideFooter(RatePages.getWB()['Rule 231 C ' + self.title_company_name], CompanyTest)

        #Rule 232
        sheet_to_compare = sheet_fetch("232 PPT")

        self.compareCompanies(sheet_to_compare)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 232 PPT BR '+ self.title_company_name, 'RULE 232. PRIVATE PASSENGER BASE RATES ' + self.title_company_name, ' ', self.buildBaseRates(comp_name, "PPT"), False, True )
            self.overideFooter(RatePages.getWB()['Rule 232 PPT BR ' + self.title_company_name], CompanyTest)

        #Rule 232 B

        self.compareCompanies(["LiabilityFleetSizeFactors_Ext","CollisionFleetSizeFactor_Ext","ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet2tables('Rule 232 B '+ self.title_company_name, 'RULE 232. PREMIUM DEVELOPMENT - PRIVATE PASSENGER TYPES ' + self.title_company_name, '232.B.1.b. Liability Fleet Size Factors', self.buildPPTLiabFleetFactors(comp_name), '232.B.4.d. Physical Damage Fleet Size Factors', self.buildPPTPhysDamFleetFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 232 B ' + self.title_company_name], CompanyTest)


        #Rule 233
        self.compareCompanies("PrivatePassengerFarmFactor2")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 233 '+ self.title_company_name, 'RULE 233. PRIVATE PASSENGER TYPES CLASSIFICATIONS - FARM ' + self.title_company_name, '233.B.2. Farm Use - Fleet Vehicle Factors (class code 7399)', self.buildPPTFarmTypes(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 233 ' + self.title_company_name], CompanyTest)


        #Rule 239 School Bus Premium
        sheet_to_compare = sheet_fetch("239 School Buses")

        self.compareCompanies(sheet_to_compare)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 239 SB BR '+ self.title_company_name, 'RULE 239. SCHOOL AND CHURCH BUSES BASE RATES ' + self.title_company_name, ' ', self.buildBaseRates(comp_name, "School Buses"), False, True )
            self.overideFooter(RatePages.getWB()['Rule 239 SB BR ' + self.title_company_name], CompanyTest)

        #Rule 239 Other Bus Premium
        sheet_to_compare = sheet_fetch("239 Other Buses")
        self.compareCompanies(sheet_to_compare)

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 239 OB BR '+ self.title_company_name, 'RULE 239. ALL OTHER BUSES BASE RATES ' + self.title_company_name, ' ', self.buildBaseRates(comp_name, "Other Buses"), False, True )
            self.overideFooter(RatePages.getWB()['Rule 239 OB BR '+ self.title_company_name], CompanyTest)


        #Rule 239 Van Bus Premium
        sheet_to_compare = sheet_fetch("239 Van Pools")

        self.compareCompanies(sheet_to_compare)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 239 VP BR '+ self.title_company_name, 'RULE 239. VAN POOLS BASE RATES ' + self.title_company_name, ' ', self.buildBaseRates(comp_name, "Van Pools"), False, True )
            self.overideFooter(RatePages.getWB()['Rule 239 VP BR ' + self.title_company_name], CompanyTest)


        #Rule 239 Taxi Premium
        sheet_to_compare = sheet_fetch("239 Taxis")
        self.compareCompanies(sheet_to_compare)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 239 T-L BR '+ self.title_company_name, 'RULE 239. TAXICABS AND LIMOUSINES BASE RATES ' + self.title_company_name, ' ', self.buildBaseRates(comp_name, "Taxis"), False, True )
            self.overideFooter(RatePages.getWB()['Rule 239 T-L BR ' + self.title_company_name], CompanyTest)

        #Rule 239 C
        self.compareCompanies(["PublicTypesFleetSizeFactorsForLiabilityAndMedicalPayments_Ext",
                               "PublicTransportationCollisionFleetSizeFactor_Ext",
                               "PublicTransportationOtherThanCollisionFleetSizeFactor_Ext",
                               "MechanicalLiftFactorOtherThanZoneRated"])

        subtitles = ["239.C.2.d. Fleet Size Factors For Liability And Medical Payments",
                     "239.C.3.d. Fleet Size Factors For Collision",
                     "239.C.3.d. Fleet Size Factors For Other Than Collision",
                     "239.C.4.d Mechanical Lift Factor"]

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            tables = [self.buildPublicAutoLiabFleetSizeFactor(comp_name),
                      self.buildPublicAutoCollFleetSizeFactor(comp_name),
                      self.buildPublicAutoOTCFleetSizeFactor(comp_name),
                      self.buildMechanicalLiftFactor(comp_name)]
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheetTablesX('Rule 239 C '+ self.title_company_name, 'RULE 239. PUBLIC AUTO PREMIUM DEVELOPMENT - OTHER THAN ZONE-RATED AUTOS ' + self.title_company_name, subtitles, tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 239 C ' + self.title_company_name], CompanyTest)


        #Rule 239 D

        self.compareCompanies("AutoLayUpFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            # You may think this looks strange. 200B is a no length invisible unicode character to add 'uniqueness' to the titles.
            # This uniqueness is needed for the marked up macro to run correctly. The user will never tell the difference.
            title_start = 'RULE 239.\u200B PUBLIC AUTO PREMIUM DEVELOPMENT - OTHER THAN ZONE-RATED AUTOS '

            RatePages.generateWorksheet('Rule 239 D '+ self.title_company_name, title_start + self.title_company_name, '239.D.1.c. Commercial Lay-up Credit', self.buildLayupFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 239 D '+ self.title_company_name], CompanyTest)

        #Rule 240
        self.compareCompanies(["PublicTransportationLiabilityPrimaryFactor","PublicTransportationPhysicalDamagePrimaryFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ['240.C.3.a Primary Classifications - Public Auto Use Classes Except Van Pools',
                      '240.C.3.b Primary Classifications - Van Pools',
                      '240.D. Secondary Classifications']
            tables = [self.build240(comp_name),
                      self.buildVanPrimaryClassFactor(comp_name),
                      self.buildVanSecondary(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 240 '+ self.title_company_name, 'RULE 240. PUBLIC AUTO CLASSIFICATIONS' + self.title_company_name, subtitles, tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 240 '+ self.title_company_name], CompanyTest)


        #Rule 241
        self.compareCompanies(["AutoLayUpFactor_Ext","MechanicalLiftFactorOtherThanZoneRated"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            if self.StateAbb == "VA":
                subtitles = VARules["Rule 241"]
            else:
                subtitles = ['241.D.1.d Mechanical Lift Factor',
                             '241.E.1.c. Commercial Lay-up Credit']

            if self.StateAbb == "VA":
                tables = [self.buildMechanicalLiftFactor(comp_name),
                          self.buildSpecifiecCausesofLossCoverageFactor(comp_name),
                          self.buildLayupFactors(comp_name)]
            else:
                tables = [self.buildMechanicalLiftFactor(comp_name),
                          self.buildLayupFactors(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 241 '+ self.title_company_name, 'RULE 241. PUBLIC AUTO PREMIUM DEVELOPMENT - ZONE-RATED AUTOS ' + self.title_company_name, subtitles, tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 241 '+ self.title_company_name], CompanyTest)

        # Rule 243
        self.compareCompanies("FarmLaborContractorPassengerHazardFactor")
        if self.StateAbb == "FL" or self.StateAbb == "VA":
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.

                subtitles = ['243.B Passenger Hazard Included Liability Coverage Factor']

                tables = [self.buildFL_PublicsSeasonalMigrantFarm(comp_name)]

                RatePages.generateWorksheetTablesX('Rule 243 ' + self.title_company_name,
                                                   'RULE 243. PUBLIC AUTO TRANSPORTATION OF SEASONAL OR MIGRANT FARM WORKERS ' + self.title_company_name,
                                                   subtitles, tables, False, True)
                self.overideFooter(RatePages.getWB()['Rule 243 ' + self.title_company_name], CompanyTest)

        #Rule 255
        self.compareCompanies(["GaragekeepersOtherThanCollisionPreliminaryBasePremium",
                               "GaragekeepersOtherThanCollisionDeductibleFactor",
                               "GaragekeepersCollisionPreliminaryBasePremium"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet3tables('Rule 255 '+ self.title_company_name, 'RULE 255. GARAGEKEEPERS INSURANCE ' + self.title_company_name, '255.C. Rates', self.buildGarageKeepers1(comp_name), self.buildGarageKeepers2(comp_name), self.buildGarageKeepers3(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 255 '+ self.title_company_name], CompanyTest)

        #Rule 264
        self.compareCompanies("SpecialTypesAntiqueAutoFactor")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            subtitles = ["264.B.1. Coverage Excluding Volunteer Workers",
                         "264.B.2. Coverage for Volunteer Firefighters and Volunteer Workers"]
            tables = [self.buildAmbulanceFactors1(comp_name),self.buildAmbulanceFactors2(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 264 '+ self.title_company_name, 'RULE 264. AMBULANCE SERVICES ' + self.title_company_name, subtitles, tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 264 '+ self.title_company_name], CompanyTest)

        #Rule 266
        self.compareCompanies(["SpecialTypesAntiqueAutoFactor","SpecialTypesAntiquePhysicalDamageRate_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet2tables('Rule 266 '+ self.title_company_name, 'RULE 266. ANTIQUE AUTOS ' + self.title_company_name, 'Liability and Basic No-Fault', self.buildAntiqueAutoLiabFactors(comp_name), 'Physical Damage', self.buildAntiqueAutoPDRates(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 266 '+ self.title_company_name], CompanyTest)

        # Rule 267
        self.compareCompanies("SpecialTypesAutoBodyFactor")
        if self.StateAbb == "FL" or self.StateAbb == "VA":
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.

                if self.StateAbb == "VA":
                    subtitles = VARules["Rule 267"]
                else:
                    subtitles = ['267.B.2. Auto Body Manufacturers And Installers Factor:']

                tables = [self.buildFL_AutoBody(comp_name)]

                RatePages.generateWorksheetTablesX('Rule 267 ' + self.title_company_name,
                                                   'RULE 267. AUTO BODY MANUFACTURERS AND INSTALLERS ' + self.title_company_name,
                                                   subtitles, tables, False, True)
                self.overideFooter(RatePages.getWB()['Rule 267 ' + self.title_company_name], CompanyTest)

        #Rule 268
        self.compareCompanies(["SpecialTypesDriverTrainingFactor",
                               "DriverTrainingLiabilityAndMedicalPaymentsCoveragesFactor",
                               "DriverTrainingMedicalPaymentsCoverageFactors"])

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            if self.StateAbb == "VA":
                subtitles = VARules["Rule 268"]
            else:
                subtitles = ["268.C. Driver Training Owned Auto Factors",
                             "268.E.2.c.(3). Instructors in Excess of Owned Autos Used For Driver Training Liability and Medical Payments Coverages",
                             "268.E.2.d. Medical Payments Coverage Factors"]
            tables = [self.buildRule68table1(comp_name),
                      self.buildRule68table3(comp_name),
                      self.buildRule68table4(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 268 '+ self.title_company_name, 'RULE 268. DRIVER TRAINING PROGRAMS ' + self.title_company_name, subtitles, tables, False, True)

            self.overideFooter(RatePages.getWB()['Rule 268 '+ self.title_company_name], CompanyTest)

        # Rule 269
        self.compareCompanies("SpecialTypesDriveAwayContractorFactor")
        if self.StateAbb == "FL" or self.StateAbb == "VA":
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.

                subtitles = ['269.B.1.a. Drive-away Contractors Liability Coverage Factor']

                tables = [self.buildFL_DriveAwayContractors(comp_name)]

                RatePages.generateWorksheetTablesX('Rule 269 ' + self.title_company_name,
                                                   'RULE 269. DRIVE-AWAY CONTRACTORS ' + self.title_company_name,
                                                   subtitles, tables, False, True)
                self.overideFooter(RatePages.getWB()['Rule 269 ' + self.title_company_name], CompanyTest)
        #Rule 271
        self.compareCompanies(["SpecialTypesFireDepartmentFactor","SpecialTypesEmergencyVehicleBuybackFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ["271.B.1. Private Passenger Type Autos",
                         "271.B.3. Other than Private Passenger Autos or Trailers",
                         "271.B.4. Coverage for Volunteer Firefighters and Volunteer Workers"]
            tables = [self.buildFireDepartmentPPTFactors(comp_name),
                      self.buildFireDepartmentOtherThanPPTFactors(comp_name),
                      self.buildFireDepartmentBuyBackFactor(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 271 '+ self.title_company_name, 'RULE 271. FIRE DEPARTMENTS ' + self.title_company_name, subtitles, tables, False, True)

            self.overideFooter(RatePages.getWB()['Rule 271 '+ self.title_company_name], CompanyTest)

        #Rule 272
        self.compareCompanies(["SpecialTypesFuneralDirectorFactor","FuneralDirectorMedicalPaymentsHiredNonOwnedFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ["272.B.1. Limousines",
                         "272.B.2. Hearses and Flower Cars",
                         "272.C.2. Medical Payments Coverage for Hired and Non-Owned Autos"]
            tables = [self.buildFuneralDirectors1(comp_name),
                      self.buildFuneralDirectors2(comp_name),
                      self.buildFuneralDirectors3(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 272 '+ self.title_company_name, 'RULE 272. FUNERAL DIRECTORS ' + self.title_company_name, subtitles, tables, False, True)

            self.overideFooter(RatePages.getWB()['Rule 272 '+ self.title_company_name], CompanyTest)


        #Rule 273
        self.compareCompanies("SpecialTypesGolfCartsAndLowSpeedVehiclesFactor")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 273 '+ self.title_company_name, 'RULE 273. GOLF CARTS AND LOW SPEED VEHICLES ' + self.title_company_name, '273.C. Premium Computation',  self.buildSpecialGolfandLow(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 273 '+ self.title_company_name], CompanyTest)

        # Rule 274

        self.compareCompanies(["SpecialTypesLawEnforcementFactor","SpecialTypesEmergencyVehicleBuybackFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            subtitles = ["274.B.1. Private Passenger Types",
                         "274.B.4. All Other Types",
                         "274.B.5. Fellow Volunteer Workers Liability Coverage"]
            tables = [self.buildLawEnforcementPPTFactors(comp_name),
                      self.buildLawEnforcementOtherThanPPTFactors(comp_name),
                      self.buildLawEnforcementBuyBackFactor(comp_name)]
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheetTablesX('Rule 274 ' + self.title_company_name,'RULE 274. LAW ENFORCEMENT AGENCIES ' + self.title_company_name, subtitles, tables,False, True)
            self.overideFooter(RatePages.getWB()['Rule 274 '+ self.title_company_name], CompanyTest)

        #Rule 275
        if self.StateAbb == "VA":
            self.compareCompanies("LeasingOrRentalConcernsContingentBasePremium_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                subtitles = VARules["Rule 275"]
                tables = [self.buildLeasingOrRentalConcernsFactors(comp_name),
                          self.buildRule275VA_TTTFactors(comp_name),
                          self.buildRule275VA_PPTFactors(comp_name),
                          self.buildRule275VA_MotorcycleFactors(comp_name),
                          self.buildRule275VA_SnowmobilesFactors(comp_name),
                          self.buildRule275VA_ExceptMotorHomesFactors(comp_name),
                          self.buildRule275VA_MotorHomesFactors(comp_name)]
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 275 '+ self.title_company_name, 'RULE 275. LEASING OR RENTAL CONCERNS ' + self.title_company_name, subtitles, tables, False, True)
                self.overideFooter(RatePages.getWB()['Rule 275 '+ self.title_company_name], CompanyTest)
        else:
            self.compareCompanies("LeasingOrRentalConcernsContingentBasePremium_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheet('Rule 275 ' + self.title_company_name,
                                            'RULE 275. LEASING OR RENTAL CONCERNS ' + self.title_company_name,
                                            '275.B. Premium Computation',
                                            self.buildLeasingOrRentalConcernsFactors(comp_name), False, True)
                self.overideFooter(RatePages.getWB()['Rule 275 ' + self.title_company_name], CompanyTest)


        #Rule 276

        self.compareCompanies(["SpecialTypesMobileHomeFactor","MobileHomesAdditionalCoveragesFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ["276.B. Premium Computation - Trailers, Pickup Trucks, and Motor Homes",
                         "276.B.6 Limited Other Than Collision Coverage on Contents"]
            tables = [self.buildMobileHomeFactors(comp_name),
                      self.buildMobileHomeAdditionalFactor(comp_name)]
            RatePages.generateWorksheetTablesX('Rule 276 ' + self.title_company_name,
                                               'RULE 276. MOBILE HOMES - ' + self.title_company_name, subtitles,
                                               tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 276 ' + self.title_company_name], CompanyTest)

        #Rule 277
        self.compareCompanies(["SpecialTypesMotorcycleLiabilityFactor",
                               "SpecialTypesMotorcycleFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ["277.B.1.d For Liability, apply the following factors based on engine size CCs:",
                         "277.B.2. Multiply the Private Passenger Base Rate by the following factor:",
                         "277.B.4.d. Physical Damage Coverage Factor - ACV",
                         "277.B.4.d. Physical Damage Coverage Factor - SA",
                         "277.B.5. Multiply the Private Passenger Base Rate by the following factor:"]
            tables = [self.buildMotorcycles1(comp_name), self.buildMotorcycles3(comp_name),
                      self.buildMotorcycles4(comp_name), self.buildMotorcycles5(comp_name),
                      self.buildMotorcycles2(comp_name)]
            RatePages.generateWorksheetTablesX('Rule 277 ' + self.title_company_name,
                                               'RULE 277. MOTORCYCLES - ' + self.title_company_name, subtitles,
                                               tables, False, True)

            self.overideFooter(RatePages.getWB()['Rule 277 '+ self.title_company_name], CompanyTest)


        #Rule 278
        self.compareCompanies("SpecialTypesRegistrationPlatesFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 278 ' + self.title_company_name,
                                        'RULE 278. REGISTRATION PLATES NOT ISSUED FOR A SPECIFIC AUTO ' + self.title_company_name,
                                        '278.B. Premium Computation',
                                        self.buildRegistrationPlateFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 278 '+ self.title_company_name], CompanyTest)

        #Rule 279
        table_codes = ["SpecialTypesRepossessedAutosLiabilityBasePremium",
                      "GarageDealersOtherThanCollisionRate",
                      "GarageDealersCollisionBlanketRate",
                      "GarageDealersCollisionBlanketDeductibleFactor"]

        if self.StateAbb == "MI":
            table_codes[2] = "GarageDealersCollisionBlanketRateMI"

        self.compareCompanies(table_codes)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet3tables('Rule 279 ' + self.title_company_name, 'RULE 279. REPOSSESSED AUTOS ' + self.title_company_name,
                                               '279.B.1 Liability', self.buildRule79table1(comp_name),
                                               self.buildRule79table2(comp_name),
                                               self.buildRule279table3(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 279 '+ self.title_company_name], CompanyTest)

        #Rule 280

        self.compareCompanies(["SpecialTypesSnowmobileLiabilityBasePremium",
                              "SpecialTypesSnowmobileMedicalPaymentsRate",
                              "SpecialTypesSnowmobileOtherThanCollisionRate",
                              "SpecialTypesSnowmobileCollisionRate"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 280 ' + self.title_company_name, 'RULE 280. SNOWMOBILES ' + self.title_company_name,
                                        '280.B. Premium Computation', self.buildSnowMobileFactors(comp_name), False,
                                        True)
            self.overideFooter(RatePages.getWB()['Rule 280 ' + self.title_company_name], CompanyTest)

        # Rule 281
        self.compareCompanies("SpecialTypesSpecialEquipmentFactor1")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            if self.StateAbb == "VA":
                subtitles = VARules["Rule 281"]
            else:
                subtitles = ["281.C.2. Premium Computation",
                             "281.D.2.b. Cost of Hire Basis Coverage Factors",
                             "281.E Rental Period Basis Factors"]
            tables = [self.buildMobileandFarmPremiumFactor(comp_name),
                      self.buildMobileandFarmCostOfHireFactor(comp_name),
                      self.buildRentalBasisFactors(comp_name)]
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheetTablesX('Rule 281 ' + self.title_company_name,
                                               'RULE 281. MOBILE AND FARM EQUIPMENT ' + self.title_company_name, subtitles,
                                               tables, False, True)
            self.overideFooter(RatePages.getWB()['Rule 281 '+ self.title_company_name], CompanyTest)


        #Rule 283
        table_codes = ["GarageDealersOtherThanCollisionRate",
                      "GarageDealersCollisionBlanketRate",
                      "GarageDealersCollisionBlanketDeductibleFactor"]

        if self.StateAbb == "MI":
            table_codes[1] = "GarageDealersCollisionBlanketRateMI"

        self.compareCompanies(table_codes)
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            subtitles = ["283.B Specified Causes of Loss",
                         "283.B Limited Specified Causes of Loss",
                         "283.B Comprehensive",
                         "283.B Blanket Collision"]
            tables = [self.buildRule83table1(comp_name),
                      self.buildRule83table2(comp_name),
                      self.buildRule83table3(comp_name),
                      self.buildRule83table4(comp_name)]

            if self.StateAbb == "MI":
                tables = tables + [self.buildRule283_MI(comp_name)]
                subtitles = subtitles + ['283.B.2.f. Service Operations Collisions Adjustment Factor']

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheetTablesX('Rule 283 ' + self.title_company_name,
                                               'RULE 283. AUTOS HELD FOR SALE BY NON-DEALERS ' + self.title_company_name,
                                               subtitles,tables, False, True)

            self.overideFooter(RatePages.getWB()['Rule 283 ' + self.title_company_name], CompanyTest)

        #Rule 284
        # This needed to be hardcoded.
        #
        subtitles = ["284.C. Premium Computation"]
        data = {
            "Coverage": [
                "Liability", "Medical", "PIP", "Other than Collision", "Collision",
                "UM", "UIM"
            ],
            "All-terrain Vehicles": [0.300, 2.000, 2.000, 0.500, 0.850, 1.000, 1.000],
            "Utility Task Vehicles": [0.600, 2.000, 2.000, 0.500, 0.850, 1.000, 1.000]
        }

        um_info = pd.read_excel(BA_INPUT_FILE, sheet_name = "297 Map")
        um_info = um_info[um_info["State"] == self.StateAbb]

        combined_text = um_info.astype(str).apply(lambda x: ' '.join(x), axis=1).str.cat(sep=' ').lower()

        UM_flag = "uninsured" in combined_text
        UIM_flag = "underinsured" in combined_text
        UMPD_flag = "property damage" in combined_text
        inlcudes_underinsured = "(includes underinsured)" in combined_text


        data = pd.DataFrame(data)
        if not UM_flag:
            data = data[data["Coverage"] != "UM"]
        if not UIM_flag or inlcudes_underinsured:
            data = data[data["Coverage"] != "UIM"]
        if not UMPD_flag:
            data = data[data["Coverage"] != "UMPD"]
        if self.StateAbb not in self.pip_states.values:
            data = data[data["Coverage"] != "PIP"]
        if self.StateAbb in self.no_med_states.values:
            data = data[data["Coverage"] != "Medical"]

        data = data.astype(object)
        data.iloc[:,1:] = data.iloc[:,1:].astype(float).map(lambda x: f"{x:.3f}")

        tables = [data]

        RatePages.generateWorksheetTablesX('Rule 284',
                                           'RULE 284. ALL-TERRAIN VEHICLES AND UTILITY TASK VEHICLES ',
                                           subtitles,tables, False, True)

        self.overideFooter(RatePages.getWB()['Rule 284'], self.default_company)

        #Rule 288
        self.compareCompanies(["DriveOtherCarLiabilityFactor",
                               "DriveOtherCarMedicalPaymentsFactor",
                               "DriveOtherCarOtherThanCollisionFactor",
                               "DriveOtherCarCollisionFactor",
                               "DriveOtherCarUninsuredMotoristFactor",
                               "DriveOtherCarUnderinsuredMotoristFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 288 ' + self.title_company_name, 'RULE 288. DRIVE OTHER CAR ' + self.title_company_name,
                                        '288.B Individual Named Insured', self.buildDriveOtherFactors(comp_name),
                                        False, True)

            self.overideFooter(RatePages.getWB()['Rule 288 ' + self.title_company_name], CompanyTest)

        #Rule 289
        self.compareCompanies(["NonOwnedBasePremium",
                              "VolunteersAsInsuredsBasePremium",
                              "NonOwnedVolunteersBasePremium",
                              "EmployeesAsInsuredsFactor",
                              "NonOwnedVolunteersMinimumPremium",
                              "NonOwnedAutoFoodorGoodsDeliveryMinimumPremiumPerPolicy_Ext",
                              "NonOwnedAutoFoodorGoodsDeliveryIncreasedLimitFactor_Ext",
                              "NonOwnedAutoFoodorGoodsDeliveryExposureFactor_Ext",
                              "GarageServicesLiabilityEmployeesFactor",
                              "NonOwnedPartnershipFactor"])

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            subtitles = ["289.B.1.b Blanket Individual Liability Rate",
                         "289.B.2.a.(1).(a). Other Than Auto Dealers Risks",
                         "289.B.2.a.(2).(d). Non-ownership Liability Coverage Factor",
                         "289.B.2.b.(1).(d). Partnership and LLC Non-ownership Liability Coverage Factor",
                         "289.B.2.b.(2).(a) Volunteers Liability Rate",
                         "289.B.2.b.(3).(b) Extended Non-ownership Liability Employee Coverage Factor",
                         "289.B.2.b.(4).(c).(iii). Food or Goods Delivery Risks",
                         "289 B.2.b.(4).(c).(v). Apply the following factors for higher limits:",
                         "289.B.2.(b).(4).(c).(vi). Minimum premium per policy:",
                         "289.B.2.(5) Minimum Premium"
                         ]

            # These tables are horrifically out of order. Just reference the rule names above. Adjusted last minute, CW manual had errors.
            tables = [self.buildRule89tableB1b(comp_name),
                      self.buildRule89table1(comp_name),
                      self.buildRule89table3(comp_name),
                      self.buildRule89table2(comp_name),
                      self.buildRule89tableB2b2a(comp_name),
                      self.buildRule89table8(comp_name),
                      self.buildRule89table4(comp_name),
                      self.buildRule89table5(comp_name),
                      self.buildRule89table6(comp_name),
                      self.buildRule89table7(comp_name)]

            RatePages.generateWorksheetTablesX('Rule 289 ' + self.title_company_name,
                                               'RULE 289. NON-OWNERSHIP LIABILITY ' + self.title_company_name,
                                               subtitles, tables, False, True)

            self.overideFooter(RatePages.getWB()['Rule 289 ' + self.title_company_name], CompanyTest)

        #Rule 290

        self.compareCompanies(["LiabilityCostOfHireRate",
                              "ComprehensiveVehicleWithDriverCostOfHireRate_Ext",
                              "CollisionVehicleWithDriverCostOfHireRate_Ext",
                              "Hired Auto MinimumOtherThanCollisionPremium_Ext",
                              "HiredAutoMinimumCollisionPremium_Ext"])

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet2tables('Rule 290 ' + self.title_company_name, 'RULE 290. HIRED AUTOS ' + self.title_company_name,
                                               '290.B.3. Liability', self.buildHiredAutoLiabFactors(comp_name),
                                               '290.C.3. Physical Damage', self.buildHiredAutoPDFactors(comp_name),
                                               False, True)

            self.overideFooter(RatePages.getWB()['Rule 290 ' + self.title_company_name], CompanyTest)

        # Rule 292
        try:
            self.compareCompanies("ZoneRatedMedicalPaymentsTextFactor") # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheet2tables('Rule 292 '+ self.title_company_name, 'RULE 292. MEDICAL PAYMENTS ' + self.title_company_name, '292.A. Increased Limit Factors - Trucks, Tractors, Trailers, Public Autos and Private Passenger Types', self.buildMedPayments92A(comp_name), '292.B. Increased Limit Factors - Zone-Rates Autos', self.buildMedPayments92B(comp_name), False, True)
                self.overideFooter(RatePages.getWB()['Rule 292 ' + self.title_company_name], CompanyTest)
        except (KeyError) as e:
            warnings.warn(f"292 sheet has failed to build. This is intended if Medpay is not in your state.", MEDPAYWarning)

        # Needs an alternative approach for testing whether the companies are equal as it's unclear what speicific tables may differ
        # state to state. This was made before the current comparison system. It is much more powerful than even the current but also more rule specific.
        # This approach is probably overkill, but it's what I could make work with the current system.
        #Rule 293 (Varies by state)
        ratebook_names = ['NGIC', 'NAFF', 'NACO','CCMIC','HICNJ', 'NICOF','NMIC','NICOA','NPCIC','AICOA']
        ratebooks = [self.NGICRatebook, self.NAFFRatebook, self.NACORatebook, self.rateTables['CCMIC'], self.rateTables['HICNJ'], self.NICOFRatebook, self.rateTables["NMIC"],self.rateTables["NICOA"],self.rateTables["NPCIC"],self.rateTables["AICOA"]]  # Replace with actual ratebooks

        available_companies = []
        available_books = []

        for company, book in zip(ratebook_names, ratebooks):
            if book != "Not found" and book is not None:
                available_books.append(book)
                available_companies.append(company)

        map_293 = pd.read_excel(BA_INPUT_FILE, sheet_name="293 Map", engine='openpyxl')

        # Create a tables_list for each company
        company_tables_list = {name: [] for name in available_companies}

        last_non_na_col = map_293.loc[map_293['State'] == self.StateAbb].notna().sum(axis=1).values[0] - 1  # Replace 'StateAbb' with actual state abbreviation

        for company in available_companies:
            tables = []
            for table_num in range(1, last_non_na_col + 1):
                table = self.build293Table(company, table_num)  # Replace with actual function call
                if table is not None:
                    tables.append(table)
            company_tables_list[company] = tables  # Store tables directly

        def compare_tables_list(tables1, tables2):
            if len(tables1) != len(tables2):
                return False
            for df1, df2 in zip(tables1, tables2):
                if not df1.equals(df2):
                    return False
            return True

        def cluster_companies(company_tables_list):
            clusters = []
            visited = set()

            for company1, tables1 in company_tables_list.items():
                if company1 not in visited:
                    cluster = [company1]
                    visited.add(company1)
                    for company2, tables2 in company_tables_list.items():
                        if company2 not in visited and compare_tables_list(tables1, tables2):
                            cluster.append(company2)
                            visited.add(company2)
                    clusters.append(", ".join(cluster))

            return clusters

        def all_tables_empty(company_tables_list):
            for tables in company_tables_list.values():
                if any(not table.empty for table in tables):  # Check if there's any non-empty DataFrame
                    return False
            return True

        # Get the clusters of companies with equal tables_list
        company_clusters = cluster_companies(company_tables_list)

        # Check if all values are NaN for the given state
        rule_names_check = map_293.loc[map_293['State'] == self.StateAbb].isna().all().all()
        # Filter out NaN values and convert to list
        rule_names = map_293.loc[map_293['State'] == self.StateAbb].dropna(axis=1).values.tolist()[0][1:] # Not grabbing state name

        # Check if rule_names does not contain only NaN values
        if not all_tables_empty(company_tables_list):
            if not rule_names_check:
                self.CompanyListDif = company_clusters
                for company_group in company_clusters:
                    company_name = company_group.split(',')[0]
                    self.title_company_name = company_group
                    if len(company_clusters) == 1:
                        self.title_company_name = ""
                    RatePages.generateWorksheetTablesX('Rule 293 ' + self.title_company_name, 'RULE 293. NO-FAULT COVERAGES ' + self.title_company_name, rule_names, company_tables_list[company_name], False, True)
                    self.overideFooter(RatePages.getWB()['Rule 293 ' + self.title_company_name], company_group)

        #Rule 294
        self.compareCompanies("RentalReimbursementFactor")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 294 ' + self.title_company_name, 'RULE 294. RENTAL REIMBURSEMENT ' + self.title_company_name,
                                        '294.B. Rate per $100 limit for the selected coverage:',
                                        self.buildRentalFactors(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 294 ' + self.title_company_name], CompanyTest)


        #Rule 295

        self.compareCompanies("AudioVisualDataEquipmentBasePremium2")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 295 ' + self.title_company_name,
                                        'RULE 295. AUDIO, VISUAL, AND DATA ELECTRONIC EQUIPMENT ' + self.title_company_name,
                                        '295.B.2. Audio, Visual and Data Electronic Equipment',
                                        self.buildAudioFactors(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 295 ' + self.title_company_name], CompanyTest)

        #Rule 296
        self.compareCompanies("TapesRecordsAndDiscsBasePremium") # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 296 ' + self.title_company_name,
                                        'RULE 296. TAPES, RECORDS AND DISCS COVERAGE - ' + self.title_company_name,
                                        '296.B. Premium Development', self.buildTapeFactors(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 296 ' + self.title_company_name], CompanyTest)

        ### An alteration of the 293 code.
        #Rule 297 (Varies by state)
        available_companies = []
        available_books = []

        for company, book in zip(ratebook_names, ratebooks):
            if book != "Not found" and book is not None:
                available_books.append(book)
                available_companies.append(company)

        map_297 = pd.read_excel(BA_INPUT_FILE, sheet_name="297 Map", engine='openpyxl')

        # Create a tables_list for each company
        company_tables_list = {name: [] for name in available_companies}

        non_blank_cols = map_297.loc[map_297['State'] == self.StateAbb].dropna(axis=1).columns[1:]  # Exclude the 'State' column

        # The below has a try except at a request from support. This function breaks more than intended and causes a headache.
        # This makes it such that if it fails, it doesn't break the process.
        for company in available_companies:
            tables = []
            for col in non_blank_cols:
                try:
                    table_num = int(col.split(' ')[-1])  # Assuming column names are like 'Table 1', 'Table 2', etc.
                    table = None
                    if table_num in [1, 2, 3, 4, 5, 6, 10]:
                        table = self.build297Table_unstacked(company, table_num)
                    elif table_num in [7, 8, 9]:
                        table = self.build297Table_stacked(company, table_num)
                    if table is not None:
                        tables.append(table)
                except:
                    pass

            # Only create the list if exists.
            if len(company_tables_list) != 0:
                company_tables_list[company] = tables  # Store tables directly

        # Prevents code from breaking by trying to compare/insert empty dataframes.
        if company_tables_list:
            # Get the clusters of companies with equal tables_list
            company_clusters = cluster_companies(company_tables_list)

            # Check if all values are NaN for the given state
            rule_names_check = map_297.loc[map_297['State'] == self.StateAbb].isna().all().all()

            # Filter out NaN values and convert to list
            rule_names = map_297.loc[map_297['State'] == self.StateAbb].dropna(axis=1).values.tolist()[0][1:]  # Not grabbing state name

            # Check if rule_names does not contain only NaN values
            if not all_tables_empty(company_tables_list):
                if not rule_names_check:
                    self.CompanyListDif = company_clusters
                    for company_group in company_clusters:
                        company_name = company_group.split(',')[0]
                        self.title_company_name = company_group
                        if len(company_clusters) == 1:
                            self.title_company_name = ""
                        RatePages.generateWorksheetTablesX('Rule 297 ' + self.title_company_name,'RULE 297. UNINSURED MOTORISTS INSURANCE ' + self.title_company_name, rule_names, company_tables_list[company_name], True, True)
                        self.overideFooter(RatePages.getWB()['Rule 297 ' + self.title_company_name], company_group)

        #Rule 298

        self.compareCompanies(["LiabilityDeductibleFactor",
                               "ZoneRatedLiabilityDeductibleFactor",
                               "PhysicalDamageDeductibleFactors_Ext",
                               "ZoneRatedVehiclesDeductibleDiscountFactors_Ext",
                               "GarageDealersCollisionBlanketDeductibleFactor",
                               "GarageDealersOtherThanCollisionDeductibleFactor",
                               "GarageDealersOtherThanCollisionAllPerilsDeductibleFactor"])

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            tables = [self.buildRule298table1(comp_name),
                      pd.DataFrame(),  # Empty because text will go here.
                      self.buildRule298PPT(comp_name),
                      self.buildRule298TTT(comp_name),
                      self.buildRule298Zone(comp_name),
                      self.buildRule298AutoBlanket(comp_name),
                      self.buildRule298AutoGarageOTCFactors_1(comp_name),
                      self.buildRule298AutoGarageOTCFactors_2(comp_name)]

            subtitles = ["298.A.2 Liability Deductible Discount Factors",
                         "298.B.1.c. Special Provisions",
                         "298.B.2.a. Private Passenger Types Deductible Discount Factors",
                         "298.B.2.b. Trucks, Tractors And Trailers And All Autos Except Zone-rated Risks Deductible Discount Factors",
                         "298.B.3. Zone-rated Vehicles Deductible Discount Factors",
                         "298.B.4.a. Auto Dealers Blanket Collision Deductible Factors",
                         "298.B.4.b. Auto Dealers and Garagekeepers Other Than Collision Deductible Factors (Applicable to Theft, Mischief, and Vandalism)",
                         "298.B.4.b. Auto Dealers and Garagekeepers Other Than Collision Deductible Factors (All Perils)"]


            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheetTablesX('Rule 298 ' + self.title_company_name,
                                               'RULE 298. DEDUCTIBLE INSURANCE ' + self.title_company_name, subtitles, tables,
                                               False, True)

            self.overideFooter(RatePages.getWB()['Rule 298 ' + self.title_company_name], CompanyTest)

        #Rule 300
        self.compareCompanies("IncreasedLimitFactorText")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 300 ' + self.title_company_name,
                                        'RULE 300. LIABILITY INCREASED LIMIT FACTORS ' + self.title_company_name,
                                        '300.B. Increased Limit Factors', self.buildILF(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 300 ' + self.title_company_name], CompanyTest)


        # 301.C
        # VA uses different numbers, but this is the contents of most other states rule 301.C
        if self.StateAbb == "VA":
            self.compareCompanies(
                "TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext")  # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)

                tables_301A = [
                    self.build101A1(comp_name),
                    self.build101A2(comp_name),
                    self.build101A3(comp_name),
                    self.build101A4(comp_name),
                    self.build101A5(comp_name),
                    self.build101B1(comp_name),
                    self.build101B2(comp_name),
                    self.build101B3(comp_name)
                ]

                tables_301B = [
                    self.buildZoneRatedTrailersVVFColl(comp_name),
                    self.buildZoneRatedNonTrailersVVFColl(comp_name),
                    self.buildPPTVVFColl(comp_name),
                    self.buildNonZoneRatedNonTrailersVVFColl(comp_name),
                    self.buildAllOtherVehiclesVVFColl(comp_name),
                    self.buildZoneRatedVehiclesVVFOTC(comp_name),
                    self.buildPPTVehiclesVVFOTC(comp_name),
                    self.buildAllOtherVehiclesVVFOTC(comp_name)
                ]

                subtitles_301A = VARules["Rule 301.A"]
                subtitles_301B = VARules["Rule 301.B"]

                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.A ' + self.title_company_name,
                                                   'RULE 301.A Vehicle Age and Price Bracket - Actual Cash Value' + self.title_company_name, subtitles_301A,
                                                   tables_301A,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.A ' + self.title_company_name], CompanyTest)

                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.B ' + self.title_company_name,
                                                   'RULE 301.B Vehicle Age and Price Bracket - Stated Amount Rating' + self.title_company_name, subtitles_301B,
                                                   tables_301B,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.B ' + self.title_company_name], CompanyTest)


        else:
            self.compareCompanies(["TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext",
                                   "TrucksAndTruckTractorsCollisionVehicleValueFactorsStatedAmountVehicles_Ext",
                                   "TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorStatedAmountVehicle_Ext",
                                   "PrivatePassengerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext",
                                   "PrivatePassengerTypesComprehensiveVehicleValueFactorsStatedAmountVehicles_Ext",
                                   "TrailerTypesCollisionVehicleValueFactorsOCNVehicles_Ext",
                                   "PrivatePassengerTypesCollisionVehicleValueFactorsOCNVehicles_Ext",
                                   "TrucksAndTruckTractorsCollisionVehicleValueFactorsOCNVehicles_Ext",
                                   "TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext",
                                   "PrivatePassengerTypesComprehensiveVehicleValueFactorsOCNVehicles_Ext",
                                   "TrucksTractorsAndTrailersOtherThanCollisionVehicleValueFactorsOCNVehicles_Ext",])
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)

                tables = [self.buildZoneRatedTrailersVVFColl(comp_name),
                          self.buildZoneRatedNonTrailersVVFColl(comp_name),
                          self.buildPPTVVFColl(comp_name),
                          self.buildNonZoneRatedNonTrailersVVFColl(comp_name),
                          self.buildAllOtherVehiclesVVFColl(comp_name),
                          self.buildZoneRatedVehiclesVVFOTC(comp_name),
                          self.buildPPTVehiclesVVFOTC(comp_name),
                          self.buildAllOtherVehiclesVVFOTC(comp_name),
                          self.build101A1(comp_name),
                          self.build101A2(comp_name),
                          self.build101A3(comp_name),
                          self.build101A4(comp_name),
                          self.build101A5(comp_name),
                          self.build101B1(comp_name),
                          self.build101B2(comp_name),
                          self.build101B3(comp_name)
                          ]


                subtitles = [
                    "301.C.1.A.(1) Zone-Rated Trailers Vehicle Value Factors - Collision with Stated Amount Rating",
                    "301.C.1.A.(2) Zone-Rated Non-Trailers Vehicle Value Factors - Collision with State Amount Rating",
                    "301.C.1.A.(3) Private Passenger Types Vehicle Value Factors - Collision with State Amount Rating",
                    "301.C.1.A.(4) Non-Zone-Rated Trailers Vehicle Value Factors - Collision with Stated Amount Rating",
                    "301.C.1.A.(5) All Other Vehicles Vehicle Value Factors - Collision with Stated Amount Rating",
                    "301.C.1.B.(1) Zone-Rated Vehicles Vehicle Value Factors - Other Than Collision with State Amount Rating",
                    "301.C.1.B.(2) Private Passenger Types Vehicle Value Factors - Other Than Collision with Stated Amount Rating",
                    "301.C.1.B.(3) All Other Vehicles Vehicle Value Factors - Other Than Collision with Stated Amount Rating",
                    "301.C.2.a.1 Zone-Rated Trailers Vehicle Value Factors - Collision with Actual Cash Value Rating",
                    "301.C.2.a.2 Zone-Rated Non-Trailers Vehicle Value Factors - Collision with Actual Cash Value Rating",
                    "301.C.2.a.3 Zone-Rated Private Passenger Types Vehicle Value Factors - Collision with Actual Cash Value Rating",
                    "301.C.2.a.4 Non-Zone-Rated Trailers Vehicle Value Factors - Collision with Actual Cash Value Rating",
                    "301.C.2.a.5 All Other Vehicle Value Factors - Collision with Actual Cash Value Rating",
                    "301.C.2.B.1 Zone-Rated Vehicles Vehicle Value Factors - Other Than Collision With Actual Cash Value Rating",
                    "301.C.2.B.2. Private Passenger Types Vehicle Value Factors - Other Than Collision WIth Actual Cash Value Rating",
                    "301.C.2.B.3  All Other Vehicles Vehicle Value Factors - Other Than Collision With Actual Cash Value Rating"
                ]


                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.C ' + self.title_company_name,
                                                   '301.C Vehicle Age and Price Bracket ' + self.title_company_name, subtitles,
                                                   tables,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.C ' + self.title_company_name], CompanyTest)



        # 301.D.1
        if self.StateAbb != "VA":
            self.compareCompanies("LiabilityOriginalCostNewFactor_Ext")  # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)

                tables = [self.build301D1(comp_name)]

                subtitles = ["301.D.1. Liability Original Cost New Factors"]

                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.D.1 ' + self.title_company_name, '301.D.1. LIABILITY ORIGINAL COST NEW FACTORS' + self.title_company_name, subtitles, tables,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.D.1 ' + self.title_company_name], CompanyTest)
        else:
            self.compareCompanies("LiabilityOriginalCostNewFactor_Ext")  # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)

                tables = [self.build301D1(comp_name)]

                subtitles = VARules["Rule 301.C.1"]

                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.C.1 ' + self.title_company_name, 'RULE 301.C.1. LIABILITY ORIGINAL COST NEW FACTORS' + self.title_company_name, subtitles, tables,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.C.1 ' + self.title_company_name], CompanyTest)

        # 301.D.2
        if self.StateAbb != "VA":
            self.compareCompanies(["LiabilityVehicleAgeFactorsStatedAmountVehicles_Ext","LiabilityVehicleAgeFactorsOriginalCostNewVehicle_Ext"])  # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)

                tables = [self.build301D2(comp_name),
                          self.build301D3(comp_name)]

                subtitles = ["301.D.2.a Liability Vehicle Age Factors - Stated Amount Vehicles",
                             "301.D.2.b Liability Vehicle Age Factors - Original Cost New Vehicles"]

                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.D.2 ' + self.title_company_name, '301.D.2 LIABILITY VEHICLE AGE FACTORS ' + self.title_company_name, subtitles, tables,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.D.2 ' + self.title_company_name], CompanyTest)
        else:
            self.compareCompanies(
                "LiabilityOriginalCostNewFactor_Ext")  # Just a temp replacement, don't know what to do about this sheet, doesn't fit well.
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)

                tables = [self.build301D3(comp_name),
                          self.build301D2(comp_name)]

                subtitles = VARules["Rule 301.C.2/3"]

                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheetTablesX('Rule 301.C.2 and 301.C.3' + self.title_company_name,
                                                   'RULE 301.C.2. & 301.C.3.' + self.title_company_name,
                                                   subtitles, tables,
                                                   False, True)

                self.overideFooter(RatePages.getWB()['Rule 301.C.2 and 301.C.3' + self.title_company_name], CompanyTest)


        #Rule 303

        self.compareCompanies(["PollutionLiabilityRate_Ext",
                              "PollutionLiabilityMinimumPremium_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 303 ' + self.title_company_name, 'RULE 303. POLLUTION LIABILITY ' + self.title_company_name,
                                        '303.B.1.b. Premium Computation ', self.buildPollutionFactors(comp_name),
                                        False, True)

            self.overideFooter(RatePages.getWB()['Rule 303 ' + self.title_company_name], CompanyTest)

        #Rule 305
        RatePages.generateWorksheet('Rule 305', 'RULE 305. LIMITED MEXICO COVERAGE', '305.B. Premium Computation', self.buildBlank(), False, True )
        self.overideFooter(RatePages.getWB()['Rule 305'], self.default_company)

        #Rule 306 NAICS

        self.compareCompanies("NAICSFactors_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.
            RatePages.generateWorksheet('Rule 306 NAICS ' + self.title_company_name, 'RULE 306. NAICS FACTORS - ' + self.title_company_name, ' ',
                                        self.buildTTTNAICSFactors(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 306 NAICS ' + self.title_company_name], CompanyTest)

        #Rule 307
        # Rule excluded from Florida
        if self.StateAbb != "FL":
            self.compareCompanies(["FellowEmployeeBaseRate_v2_Ext","FellowEmployeeCoverageForDesignatedEmployeesPositionsBaseRate_v2_Ext"])
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.
                RatePages.generateWorksheet('Rule 307 ' + self.title_company_name, 'RULE 307. FELLOW EMPLOYEE COVERAGE ' + self.title_company_name,
                                            '307.C. Premium Computation', self.buildFellowEmployeeFactors(comp_name),
                                            False, True)

                self.overideFooter(RatePages.getWB()['Rule 307 ' + self.title_company_name], CompanyTest)

        # Rule 309
        if self.StateAbb == "FL" or self.StateAbb == "VA":
            self.compareCompanies(["AutoLoanLeaseGapCoverageFactor"])
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""  # Every Company, no point in putting in sheet name.

                subtitles = ['309.B.1.a. Drive-away Contractors Liability Coverage Factor']

                tables = [self.buildFL_AutoLeaseGapCoverage(comp_name)]

                RatePages.generateWorksheetTablesX('Rule 309 ' + self.title_company_name,
                                                   'RULE 309.  AUTO LOAN/LEASE GAP COVERAGE ' + self.title_company_name,
                                                   subtitles, tables, False, True)
                self.overideFooter(RatePages.getWB()['Rule 309 ' + self.title_company_name], CompanyTest)

        #Rule 310
        self.compareCompanies("OptionalLimitsLossofUseExpensesBasePremium_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 310 ' + self.title_company_name,
                                        'RULE 310. LOSS OF USE EXPENSES - RENTAL VEHICLES - OPTIONAL LIMITS ' + self.title_company_name, "",
                                        self.buildLossofUseFactors(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 310 ' + self.title_company_name], CompanyTest)

        #Rule 313
        RatePages.generateWorksheet('Rule 313', 'RULE 313. SILICA OR SILICA-RELATED DUST LIABILITY', '313.B. Premium Computation',self.buildBlank(), False, True )
        self.overideFooter(RatePages.getWB()['Rule 313'], self.default_company)

        #Rule 315

        self.compareCompanies(["BusinessInterruptionCoverageOtherThanCollisionBaseLossCost",
                               "BusinessInterruptionCoverageSpecifiedCausesOfLossBaseLossCost",
                               "BusinessInterruptionCoverageCollisionBaseLossCost",
                               "ExtendedBusinessIncomeAdditionalCoverageFactor",
                               "BusinessIncomeCoverageWaitingPeriodFactor",
                               "InsuranceToExposureFactor"])

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet4tables('Rule 315 ' + self.title_company_name,
                                               'RULE 315. BUSINESS INTERRUPTION COVERAGE - ' + self.title_company_name,
                                               '315.B.1. Business Interruption Coverage Base Premium:',
                                               self.buildBusinessInterruptionFactors(comp_name),
                                               self.buildExtendedBusinessFactors(comp_name),
                                               self.buildWaitingBusinessFactors(comp_name),
                                               self.buildInsuranceToExposureFactor(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 315 ' + self.title_company_name], CompanyTest)

        #Rule 317
        self.compareCompanies("TowingLaborRate")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)

            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""  # Every Company, no point in putting in sheet name.

            RatePages.generateWorksheet('Rule 317 '+ self.title_company_name, 'RULE 317. TOWING AND LABOR ' + self.title_company_name, '317.B. Premium Computation', self.buildTowingAndLabor(comp_name), False, True)
            self.overideFooter(RatePages.getWB()['Rule 317 ' + self.title_company_name], CompanyTest)

        #Rule 416

        self.compareCompanies(["ExperienceRatingExpectedFrequencyPerPowerUnit_Ext",
                               "ExperienceRatingModifierRange_Ext",
                               "ExperienceRatingCredibility_Ext",
                               "ExperienceRatingBaseCredibility_Ext"])

        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet2tbls('Rule 416 ' + self.title_company_name, 'RULE 416. EXPERIENCE RATING ' + self.title_company_name,
                                             '416.3.d. Rating Variables', self.buildExperienceRating(comp_name),
                                             self.buildExperienceRatingMinMax(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 416 ' + self.title_company_name], CompanyTest)

        #Rule 417

        self.compareCompanies("ScheduleEligibility_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 417 ' + self.title_company_name, 'RULE 417. SCHEDULE RATING ' + self.title_company_name, ' ',
                                        self.buildRule117(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 417 ' + self.title_company_name], CompanyTest)

        #Rule 425

        self.compareCompanies(["WaiverofSubrogationBlanket_Ext","Waiver_of_Subrogation_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet2tables('Rule 425 ' + self.title_company_name,
                                               'RULE 425. WAIVER OF TRANSFER OF RIGHTS OF RECOVERY AGAINST OTHERS TO US ' + self.title_company_name,
                                               '425.A.3. Premium Computation', self.buildWaiver1Factors(comp_name),
                                               '425.B.3. Premium Computation', self.buildWaiver2Factors(comp_name),
                                               False, True)

            self.overideFooter(RatePages.getWB()['Rule 425 ' + self.title_company_name], CompanyTest)

        #Rule 426

        self.compareCompanies(["BusinessAutoProtectionFactor_Ext","MiscellaneousMinimumMaximumPremium_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 426 ' + self.title_company_name,
                                        'RULE 426. BUSINESS AUTO PROTECTION ENDORSEMENTS ' + self.title_company_name,
                                        '426.B. Premium Computation', self.buildProtectionFactors(comp_name), False,
                                        True)

            self.overideFooter(RatePages.getWB()['Rule 426 ' + self.title_company_name], CompanyTest)

        #Rule 427

        self.compareCompanies("OriginalEquipmentManufacturerPartsCoverageFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 427 ' + self.title_company_name,
                                        'RULE 427. ORIGINAL EQUIPMENT MANUFACTURER PARTS COVERAGE ' + self.title_company_name,
                                        '427.C. Premium Computation', self.buildOriginalFactors(comp_name), False,
                                        True)

            self.overideFooter(RatePages.getWB()['Rule 427 ' + self.title_company_name], CompanyTest)

        #Rule 450

        self.compareCompanies(["DriverBasedRatingLiabilityFactor_Ext","DriverBasedRatingCollisionFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 450 ' + self.title_company_name, 'RULE 450. DRIVER BASED RATING PLAN ' + self.title_company_name,
                                        '450.B.5. Premium Computation', self.buildDBRFactors(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 450 ' + self.title_company_name], CompanyTest)

        #Rule 451
        # CO and MT had objections over Rule 451 being filed after transition ended.
        if self.StateAbb in ["MT","CO"]:
            pass
        else:
            for company in available_companies:
                tables = []
                table = self.buildRule451(company)
                if table is not None:
                    tables.append(table)
                company_tables_list[company] = tables  # Store tables directly
            company_clusters = cluster_companies(company_tables_list)

            self.CompanyListDif = company_clusters
            # Check if rule_names contain only NaN values
            if not rule_names_check:
                for company_group in company_clusters:
                    company_name = company_group.split(',')[0]
                    self.title_company_name = company_group
                    if len(company_clusters) == 1:
                        self.title_company_name = ""

                    tables = [self.buildRule451(company_name),
                              self.buildRule451Renewals(company_name)]

                    subtitles = ["451.B",
                                 "451.E Number Of Subsequent Renewals"]

                    RatePages.generateWorksheetTablesX('Rule 451 ' + self.title_company_name, 'RULE 451. TRANSITION CAPPING PROGRAM ' + self.title_company_name, subtitles, tables, False, True )
                    self.overideFooter(RatePages.getWB()['Rule 451 ' + self.title_company_name], company_group)


        #Rule 452

        leaf_list = pd.read_excel(BA_INPUT_FILE, sheet_name="452 Leaf")
        show_leaf = np.bool(leaf_list[leaf_list["State"] == self.State].iloc[0,1]) # Lsit from product has full state names.

        if show_leaf:
            self.compareCompanies("RetentionGradeFactor_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""
                RatePages.generateWorksheet('Rule 452 ' + self.title_company_name,
                                            'RULE 452. LIFETIME EXPENSE ALLOCATION FACTOR ' + self.title_company_name,
                                            '452.C.1. Lifetime Expense Allocation Factor',
                                            self.buildRetentionFactors(comp_name), False, True)

                self.overideFooter(RatePages.getWB()['Rule 452 ' + self.title_company_name], CompanyTest)

        #Rule 453


        self.compareCompanies(["TieringLiabilityFactor_Ext","TieringCollisionFactor_Ext","TieringOtherThanCollisionFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 453 ' + self.title_company_name, 'RULE 453. RISK TIER RATING PLAN ' + self.title_company_name,
                                        '453.C.1. Tiering Factor', self.buildRiskTiering(comp_name), False, True)

            self.overideFooter(RatePages.getWB()['Rule 453 ' + self.title_company_name], CompanyTest)

        #Rule 454

        self.compareCompanies(["CorporalPunishmentBaseRate_Ext","MiscellaneousMinimumMaximumPremium_Ext","BroadFormSchoolBusOperatorsCoverageFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            RatePages.generateWorksheet('Rule 454 ' + self.title_company_name, 'RULE 454. SCHOOL BUS OPERATIONS ' + self.title_company_name,
                                        '454.A. Corporal Punishment', self.buildCorporalPunish(comp_name), False,
                                        True)

            self.overideFooter(RatePages.getWB()['Rule 454 ' + self.title_company_name], CompanyTest)

        #Rule DP1
        if self.StateAbb in ['AR', 'AZ', 'CT', 'DC', 'DE', 'IA', 'IL', 'KY', 'MD', 'ME', 'MI', 'MN', 'MO', 'MS', 'NC', 'NE', 'NH', 'NM', 'NV', 'OR', 'PA', 'RI', 'SD', 'TN', 'TX', 'UT', 'VT', 'WI', 'WV', 'WY']:

            if self.rateTables["NGIC"] is not None and self.rateTables["NGIC"] != "Not Found":
                    RatePages.generateWorksheet('Rule DP-1', 'RULE DP-1. DISTRIBUTION PLAN',
                                                ' ', self.buildDP1("NGIC"), False, True)

                    self.overideFooter(RatePages.getWB()['Rule DP-1'], None)

        # #Rule T1 THIS HAS BEEN DISCONTINUED AS OF 1/28/2026.
        # # Try except because state list is missing.
        # t1_exclude_list = pd.read_excel(BA_INPUT_FILE, sheet_name = "T1 Exclude")
        # has_t1 = True if self.StateAbb not in t1_exclude_list["State"].values else False
        #
        # if has_t1:
        #     self.compareCompanies("TelematicsFactor")
        #     for CompanyTest in self.CompanyListDif:
        #         comp_name = self.extract_company_name(CompanyTest)
        #         self.title_company_name = CompanyTest
        #         if len(self.CompanyListDif) == 1:
        #             self.title_company_name = ""
        #
        #         RatePages.generateWorksheet('Rule T1 ' + self.title_company_name, 'RULE T-1. BUSINESS AUTO TELEMATICS - ' + self.title_company_name,
        #                                     '', self.buildTelematicsFactors(comp_name), False, True)
        #
        #         self.overideFooter(RatePages.getWB()['Rule T1 ' + self.title_company_name], CompanyTest)
        # else:
        #     pass


        # Below begins the list of non standard rules starting with a letter.
        if self.StateAbb == "CT":
            # CT: Rule A1
            self.compareCompanies("LiabilityofMunicipalitiesPremiumCT_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A1', 'RULE A1. LIABILITY OF MUNICIPALITIES' + self.title_company_name,['A1.B'], [self.buildCT_A1(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A1'], self.default_company)

            # CT: Rule R1
            df_raw = pd.read_excel(BA_INPUT_FILE, sheet_name="CT R1", header=None)
            R1 = df_raw.fillna('')
            R1.columns = ['']*9 # Making columns blank.

            RatePages.generateWorksheetTablesX('Rule R1', 'RULE R1. Rate Order of Calculation' + self.title_company_name,[''], [R1], False, True)

            self.overideFooter(RatePages.getWB()['Rule R1'], self.default_company)

        if self.StateAbb == "KS":
            # KS: Rule A1
            self.compareCompanies("AccidentPreventionDiscountFactor")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A1', 'RULE A1.  ACCIDENT PREVENTION COURSE PREMIUM REDUCTION' + self.title_company_name,['A1.D'], [self.buildKS_A1(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A1'], self.default_company)

        if self.StateAbb == "MI":
            # MI: Rule A2
            self.compareCompanies("PropertyDamageLiabilityCoverageBuybackVehiclePremium")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A2', 'RULE A2. PROPERTY DAMAGE LIABILITY COVERAGE BUYBACK ' + self.title_company_name,
                                                   ['A2.C. Charge the following per auto:'],
                                                   [self.buildMI_A2(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A2'], self.default_company)

            # MI: Rule A4
            self.compareCompanies("MIMCCAFee_v2_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A4', 'RULE A4. MICHIGAN CATASTROPHIC CLAIMS ASSOCIATION SURCHARGE ' + self.title_company_name,
                                                   ['A4.A. Charge the following per auto:',
                                                    'A4.B. Charge the following per historic vehicle:'],
                                                   [self.buildMI_A4A(comp_name),
                                                    self.buildMI_A4B(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A2'], self.default_company)

        if self.StateAbb == "ND":
            # ND: Rule A1
            self.compareCompanies("RentalVehicleCoverageFactor_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A1', 'RULE A1. RENTAL VEHICLE COVERAGE' + self.title_company_name,['A1.B.2 Premium Development'], [self.buildND_A1(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A1'], self.default_company)

            # ND: Rule A3
            self.compareCompanies("AccidentPreventionDiscountFactorND_Ext")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A3', 'RULE A3. ACCIDENT PREVENTION COURSE DISCOUNT' + self.title_company_name,['A3.E.'], [self.buildND_A3(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A3'], self.default_company)

        if self.StateAbb == "NJ":
            self.compareCompanies("JitneysLiabilityBasePremium")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A5', 'RULE A5. JITNEYS' + self.title_company_name,[''], [self.buildNJ_A5(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A5'], self.default_company)

        if self.StateAbb == "NV":
            # NV: Rule A2
            self.compareCompanies("PassiveRestraintDiscountFactor")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A2', 'RULE A2. PASSIVE RESTRAINT DISCOUNT ' + self.title_company_name,
                                                   ['A2. Multiply the medical payments charge and the uninsured motorists charge by the following factors:'],
                                                   [self.buildNV_A2(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A2'], self.default_company)

        if self.StateAbb == "RI":
            # RI: Rule A1
            self.compareCompanies("AccidentPreventionDiscountFactorRI")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A1', 'RULE A1. Accident Prevention Course Discount' + self.title_company_name,['A1.D Accident Prevention Course Discount Factor'], [self.buildRI_A1(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A1'], self.default_company)

            # RI: Rule A2
            self.compareCompanies("AntiTheftDeviceDiscountFactor1")
            for CompanyTest in self.CompanyListDif:
                comp_name = self.extract_company_name(CompanyTest)
                self.title_company_name = CompanyTest
                if len(self.CompanyListDif) == 1:
                    self.title_company_name = ""

                RatePages.generateWorksheetTablesX('Rule A2', 'RULE A2. Anti-Theft Device Discount' + self.title_company_name,[''], [self.buildRI_A2(comp_name)], False, True)

                self.overideFooter(RatePages.getWB()['Rule A2'], self.default_company)


        ################################################################################################################################################################################
        #End of Rate Page Creation
        #Formatting of specific rate pages below:
        ################################################################################################################################################################################


        RatePages.createIndex()
        AutoPages = RatePages.getWB()
        excel_Sheet_names = AutoPages.sheetnames



        Rule22_TTTBaseRates = [name for name in excel_Sheet_names if name.startswith('Rule 222 TTT')]
        for Rule in Rule22_TTTBaseRates:
            self.formatBaseRates(AutoPages[Rule], "222")

        Rule222_B_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 222 B')]
        for Rule in Rule222_B_Sheets:
            self.format222B(AutoPages[Rule])

        Rule222C_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 222 C')]
        for Rule in Rule222C_Sheets:
            self.format222C(AutoPages[Rule])

        Rule222E_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 222 E')]
        for Rule in Rule222E_Sheets:
            self.format222E(AutoPages[Rule])

        Rule23_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 223 B.5')]
        for Rule in Rule23_Sheets:
            self.format23B(AutoPages[Rule])

        Rule23C_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 223 C')]
        for Rule in Rule23C_Sheets:
            self.format23C(AutoPages[Rule])

        Rule225C2_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 225.C.2')]
        for Rule in Rule225C2_Sheets:
            self.format225C2(AutoPages[Rule])

        Rule225C3_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 225.C.3')]
        for Rule in Rule225C3_Sheets:
            self.format225C3(AutoPages[Rule])

        Rule225Z_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 225 Zone BR')]
        for Rule in Rule225Z_Sheets:
            self.formatZoneRates(AutoPages[Rule])

        Rule225D_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 225.D')]
        for Rule in Rule225D_Sheets:
            self.format225D(AutoPages[Rule])

        Rule31C_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 231 C')]
        for Rule in Rule31C_Sheets:
            self.format31C(AutoPages[Rule])

        Rule32_PPTBaseRates = [name for name in excel_Sheet_names if name.startswith('Rule 232 PPT')]
        for Rule in Rule32_PPTBaseRates:
            self.formatBaseRates(AutoPages[Rule], "232")

        Rule32B_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 232 B')]
        for Rule in Rule32B_Sheets:
            self.format32B(AutoPages[Rule])

        Rule33_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 233')]
        for Rule in Rule33_Sheets:
            self.format33(AutoPages[Rule])

        Rule34_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 234')]
        for Rule in Rule34_Sheets:
            self.formatTowingAndLabor(AutoPages[Rule])

        Rule39_SchoolBaseRates = [name for name in excel_Sheet_names if name.startswith('Rule 239 SB')]
        for Rule in Rule39_SchoolBaseRates:
            self.formatBaseRates(AutoPages[Rule], "239")

        Rule39_OtherBusBaseRates = [name for name in excel_Sheet_names if name.startswith('Rule 239 OB')]
        for Rule in Rule39_OtherBusBaseRates:
            self.formatBaseRates(AutoPages[Rule], "239")

        Rule39_VanBaseRates = [name for name in excel_Sheet_names if name.startswith('Rule 239 VP')]
        for Rule in Rule39_VanBaseRates:
            self.formatBaseRates(AutoPages[Rule],"239")

        Rule39_TaxiBaseRates = [name for name in excel_Sheet_names if name.startswith('Rule 239 T-L')]
        for Rule in Rule39_TaxiBaseRates:
            self.formatBaseRates(AutoPages[Rule], "239")

        Rule_239C = [name for name in excel_Sheet_names if name.startswith('Rule 239 C')]
        for Rule in Rule_239C:
            self.format239C(AutoPages[Rule])

        Rule240_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 240')]
        for Rule in Rule240_Sheets:
            self.format240(AutoPages[Rule])

        Rule241_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 241')]
        for Rule in Rule241_Sheets:
            self.format241(AutoPages[Rule])

        # Currently FL and VA Only
        Rule243_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 243')]
        for Rule in Rule243_Sheets:
            self.format243(AutoPages[Rule])

        Rule55_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 255')]
        for Rule in Rule55_Sheets:
            self.format55(AutoPages[Rule], Font(name=fontName, size=fontSize, bold=True))

        Rule264_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 264')]
        for Rule in Rule264_Sheets:
            self.formatAmbulance(AutoPages[Rule])

        Rule266_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 266')]
        for Rule in Rule266_Sheets:
            self.formatANTIQUEAUTOS(AutoPages[Rule])

        # Currently FL and VA Only
        Rule267_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 267')]
        for Rule in Rule267_Sheets:
            self.format267(AutoPages[Rule])

        Rule68_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 268')]
        for Rule in Rule68_Sheets:
            self.format68(AutoPages[Rule], Font(name=fontName, size=fontSize, bold=True))

        # Currently FL and VA Only
        Rule269_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 269')]
        for Rule in Rule269_Sheets:
            self.format269(AutoPages[Rule])

        Rule271_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 271')]
        for Rule in Rule271_Sheets:
            self.formatFireDepartments(AutoPages[Rule])

        Rule72_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 272')]
        for Rule in Rule72_Sheets:
            self.format72(AutoPages[Rule])

        Rule73_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 273')]
        for Rule in Rule73_Sheets:
            self.format73(AutoPages[Rule])
        Rule274_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 274')]
        for Rule in Rule274_Sheets:
            self.format274(AutoPages[Rule])
        Rule75_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 275')]
        for Rule in Rule75_Sheets:
            self.formatLEASINGORRENTALCONCERNS(AutoPages[Rule], Font(name=fontName, size=fontSize, italic=True))
        Rule276_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 276')]
        for Rule in Rule276_Sheets:
            self.format276(AutoPages[Rule])
        Rule77_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 277')]
        for Rule in Rule77_Sheets:
            self.format77(AutoPages[Rule])

        Rule78_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 278')]
        for Rule in Rule78_Sheets:
            self.format78(AutoPages[Rule])

        Rule79_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 279')]
        for Rule in Rule79_Sheets:
            self.format79(AutoPages[Rule], Font(name=fontName, size=fontSize, bold=True))

        Rule80_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 280')]
        for Rule in Rule80_Sheets:
            self.format80(AutoPages[Rule])

        Rule81_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 281')]
        for Rule in Rule81_Sheets:
            self.format81(AutoPages[Rule])

        Rule283_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 283')]
        for Rule in Rule283_Sheets:
            self.format283(AutoPages[Rule])

        Rule284_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 284')]
        for Rule in Rule284_Sheets:
            self.format284(AutoPages[Rule])

        Rule88_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 288')]
        for Rule in Rule88_Sheets:
            self.format88(AutoPages[Rule])

        Rule89_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 289')]
        for Rule in Rule89_Sheets:
            self.format89(AutoPages[Rule])

        Rule90_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 290')]
        for Rule in Rule90_Sheets:
            self.formathiredauto(AutoPages[Rule])

        Rule92_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 292')]
        for Rule in Rule92_Sheets:
            self.format92(AutoPages[Rule])

        Rule293_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 293')]
        for Rule in Rule293_Sheets:
            self.format293(AutoPages[Rule])

        Rule94_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 294')]
        for Rule in Rule94_Sheets:
            self.format94(AutoPages[Rule])

        Rule95_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 295')]
        for Rule in Rule95_Sheets:
            self.format95(AutoPages[Rule])

        Rule96_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 296')]
        for Rule in Rule96_Sheets:
            self.format96(AutoPages[Rule])

        Rule297_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 297')]
        for Rule in Rule297_Sheets:
            self.format297(AutoPages[Rule])

        Rule98_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 298')]
        for Rule in Rule98_Sheets:
            self.format98(AutoPages[Rule])

        Rule100_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 300')]
        for Rule in Rule100_Sheets:
            self.format100(AutoPages[Rule])

        # VA rule 301 has an exception
        if self.StateAbb != "VA":
            Rule301_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.C')]
            for Rule in Rule301_Sheets:
                self.format301(AutoPages[Rule])
        elif self.StateAbb == "VA":
            Rule301A_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.A')]
            for Rule in Rule301A_Sheets:
                self.format301VA(AutoPages[Rule])
            Rule301B_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.B')]
            for Rule in Rule301B_Sheets:
                self.format301VA(AutoPages[Rule])

        # VA rule 301 has an exception
        if self.StateAbb != "VA":
            Rule301D1_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.D.1')]
            for Rule in Rule301D1_Sheets:
                self.format301D1(AutoPages[Rule])
        elif self.StateAbb == "VA":
            Rule301C1_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.C.1')]
            for Rule in Rule301C1_Sheets:
                self.format301D1(AutoPages[Rule])

        # VA rule 301 has an exception
        if self.StateAbb != "VA":
            Rule301D2_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.D.2')]
            for Rule in Rule301D2_Sheets:
                self.format301D2(AutoPages[Rule])
        elif self.StateAbb == "VA":
            Rule301C2_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 301.C.2')]
            for Rule in Rule301C2_Sheets:
                self.format301D2(AutoPages[Rule])

        Rule103_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 303')]
        for Rule in Rule103_Sheets:
            self.format103(AutoPages[Rule])

        Rule105_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 305')]
        for Rule in Rule105_Sheets:
            self.format113(AutoPages[Rule])

        Rule306_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 306 NAICS')]
        for Rule in Rule306_Sheets:
            self.formatNaics(AutoPages[Rule])

        Rule107_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 307')]
        for Rule in Rule107_Sheets:
            self.formatFellowEmployee(AutoPages[Rule])

        # Currently FL and VA Only
        Rule309_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 309')]
        for Rule in Rule309_Sheets:
            self.format309(AutoPages[Rule])

        Rule110_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 310')]
        for Rule in Rule110_Sheets:
            self.formatLossofUse(AutoPages[Rule])

        Rule113_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 313')]
        for Rule in Rule113_Sheets:
            self.format113(AutoPages[Rule])

        Rule115_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 315')]
        for Rule in Rule115_Sheets:
            self.formatBusinessInterruption(AutoPages[Rule])

        Rule317_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 317')]
        for Rule in Rule317_Sheets:
            self.formatTowingAndLabor(AutoPages[Rule])

        Rule116_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 416')]
        for Rule in Rule116_Sheets:
            self.formatExperienceRating(AutoPages[Rule])

        Rule117_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 417')]
        for Rule in Rule117_Sheets:
            self.format117(AutoPages[Rule], Font(name=fontName, size=fontSize, italic=True))

        Rule125_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 425')]
        for Rule in Rule125_Sheets:
            self.format125(AutoPages[Rule])

        Rule126_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 426')]
        for Rule in Rule126_Sheets:
            self.format126(AutoPages[Rule])

        Rule127_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 427')]
        for Rule in Rule127_Sheets:
            self.format127(AutoPages[Rule])

        Rule150_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 450')]
        for Rule in Rule150_Sheets:
            self.format150(AutoPages[Rule])

        Rule151_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 451')]
        for Rule in Rule151_Sheets:
            self.format151(AutoPages[Rule], Font(name=fontName, size=fontSize, italic=True))

        Rule154_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule 454')]
        for Rule in Rule154_Sheets:
            self.formatSchoolBusOps(AutoPages[Rule], Font(name=fontName, size=fontSize, italic=True))

        RuleDP1_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule DP-1')]
        for Rule in RuleDP1_Sheets:
            self.formatDp1(AutoPages[Rule])

        RuleT1_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule T1')]
        for Rule in RuleT1_Sheets:
            self.formatT1(AutoPages[Rule])

        RuleA1_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule A1')]
        for Rule in RuleA1_Sheets:
            self.formatA1(AutoPages[Rule])

        RuleA2_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule A2')]
        for Rule in RuleA2_Sheets:
            self.formatA2(AutoPages[Rule])

        RuleA3_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule A3')]
        for Rule in RuleA3_Sheets:
            self.formatA3(AutoPages[Rule])

        RuleA4_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule A4')]
        for Rule in RuleA4_Sheets:
            self.formatA4(AutoPages[Rule])

        RuleA5_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule A5')]
        for Rule in RuleA5_Sheets:
            self.formatA5(AutoPages[Rule])

        RuleR1_Sheets = [name for name in excel_Sheet_names if name.startswith('Rule R1')]
        for Rule in RuleR1_Sheets:
            self.formatR1(AutoPages[Rule])

        if self.StateAbb == "FL":
            # FL needs room to stamp
            self.overideHeaderFL(AutoPages)

        return AutoPages