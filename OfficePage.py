# This module formats the Office State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class Office:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.officeProgramCode = 60000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building", 
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')
    
    # Builds the territory multiplier table for the given coverage (either building, bpp, or liability)
    # Returns a dataframe
    #def buildTerritoryMultiplier(self, coverage):
    #    territorialFactor = self.buildDataFrame("BP7_Peril_TerritorialFactor")
    #    filteredTerritorialFactor = territorialFactor.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'TerritoryCode': 'Territory'})
    #    if coverage.casefold() == 'building': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BldgTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'bpp': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BPPTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'liability': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='LiabilityTerritoryFactor').reset_index('Territory')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'] == 'L-Products'].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'}) # Converting to int first to get rid of decimal places
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0', 
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe    
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.officeProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.officeProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_BLDG_BPP_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCodeMin == {self.officeProgramCode}').rename(columns={'BldgBPPLimit_Min': 'Limit Min', 'BldgBPPLimit_Max': 'Limit Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Limit Min', 'Limit Max'], columns='Peril TypeCode', values='Factor').reset_index(['Limit Min', 'Limit Max']).fillna({'Limit Max': 'and over'})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.officeProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance' : 'int32'})

    # Builds the general liability occupancy modifiers
    # Returns a dataframe
    def buildGeneralOccupancyMod(self):
        generalOccupancyModifier = self.buildDataFrame("BP7_Peril_Occupant_Factor")
        filteredGeneralOccupancyMod = generalOccupancyModifier.query(f'ClassCode_Min == {self.officeProgramCode} & `Peril TypeCode` == "liability1"'). \
                rename(columns={'OccupancyType': 'Occupancy', 'BLDGOccupantFactor': 'Building', 'BPPOccupantFactor': 'Business Personal Property'})
        return filteredGeneralOccupancyMod.replace({'Occupancy': {'Condominium': 'Condo Unit-owner', 'buildingOwnerLessorsrisk': "Lessor's Risk", 'buildingOwnerOccupant': 'Owner Occupant', 'tenant': 'Tenant'}}). \
                filter(items=['Occupancy', 'Building', 'Business Personal Property'])

    # Builds the directors and officers liability insurance worksheet
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Office"').copy() # Getting a new copy of the data here
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 or More'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])

    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Office"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Office"').copy() # Getting a new copy of the data here
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index = filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])
    
    # Builds the optometrists professional liability table
    # Returns a dataframe
    def buildOptometristsProfessionalLiab(self):
        optometristsProfessionalLiab = self.buildDataFrame("BP7_ProfLiabOptical_Rate")
        optometristsProfessionalLiab['LiabilityAmount'] = optometristsProfessionalLiab['LiabilityAmount'].apply(lambda x: "${0:,.0f}".format(x))
        optometristsProfessionalLiab['AggregateLimit'] = optometristsProfessionalLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        optometristsProfessionalLiab['Occurence / Aggregate'] = optometristsProfessionalLiab['LiabilityAmount'] + ' / ' + optometristsProfessionalLiab['AggregateLimit']
        return optometristsProfessionalLiab.rename(columns={'OptometristRate': 'Each Optometrist', 'OpticianRate': 'Each Optician'}).filter(items=['Occurence / Aggregate', 'Each Optometrist', 'Each Optician'])

    def buildVetSpecializedLiab(self):
        data = [
            ("$25,000", "$32", "$17"),
            ("$50,000", "$49", "$32"),
            ("$100,000", "$65", "$49")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the veterinarians professional liability table
    # Returns a dataframe
    def buildVetProfessionalLiab(self):
        vetProfessionalLiab = self.buildDataFrame("BP7_Pol_VeterinariansExtraCov_Rate")
        vetProfessionalLiab['LiabLimitPerOcc'] = vetProfessionalLiab['LiabLimitPerOcc'].apply(lambda x: "${0:,.0f}".format(x))
        vetProfessionalLiab['LiabLimitPerAgg'] = vetProfessionalLiab['LiabLimitPerAgg'].apply(lambda x: "${0:,.0f}".format(x))
        vetProfessionalLiab['Occurence / Aggregate'] = vetProfessionalLiab['LiabLimitPerOcc'] + ' / ' + vetProfessionalLiab['LiabLimitPerAgg']
        return vetProfessionalLiab.rename(columns={'Otherthanhouseholdanimals': 'Other than household animals', 'HouseholdAnimalsOnly': 'Household animals only'}). \
                filter(items=['Occurence / Aggregate', 'Other than household animals', 'Household animals only'])

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.officeProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Office premises'})

    # Builds the franchise upgrage endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.officeProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(), 
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(), 
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    def buildArchitectsEngineersEndorsement(self):

        data = {"Base Premium for each Office Premises" : ["$200.00"]}
        data = pd.DataFrame(data)

        return data

    def buildConsultantSpecializedEndorsement(self):

        data = {"Base Premium for each Office Premises" : ["$200.00"]}
        data = pd.DataFrame(data)

        return data

    def buildProfessionaServicesEnforsement(self):

        data = {"Base Premium for each Office Premises" : ["$125.00"]}
        data = pd.DataFrame(data)

        return data

    def buildAccountantsSpecialziedEndorsement(self):

        data = {"Base Premium for each Office Premises" : ["$200.00"]}
        data = pd.DataFrame(data)

        return data

    def buildAttorneySpecializedEndorsement(self):

        data = {"Base Premium for each Office Premises" : ["$200.00"]}
        data = pd.DataFrame(data)

        return data

    def buildHealthCareSpecializedEndorsement(self):

        data = {"Base Premium for each Office Premises" : ["$300.00", "Plus $10 for each additional employee above 5\nfor employee dishonesty coverage"]}
        data = pd.DataFrame(data)

        return data

    def buildVeterinarianSpecializedEndorsement(self):
        data = {"Base Premium for each Office Premises" : ["$269.00"]}
        data = pd.DataFrame(data)

        return data

    def buildVeterinarianSpecializedEndorsementIncome(self):
        data = [
            ("$25,000", "$32", "$17"),
            ("$50,000", "$49", "$32"),
            ("$100,000", "$65", "$49")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])


    def buildVetProfLiabHousehold(self):
        # Data for household pet veterinarian rates
        data = [
            ("300,000/600,000", "$60"),
            ("500,000/1,000,000", "$69"),
            ("1,000,000/2,000,000", "$86"),
            ("2,000,000/4,000,000", "$175")
        ]
        data = pd.DataFrame(data, columns=["Limits", "Rate"])

        return data

    def buildVetProfLiabNonHousehold(self):

        # Data for non-household pet veterinarian rates
        data = [
            ("300,000/600,000", "$105"),
            ("500,000/1,000,000", "$118"),
            ("1,000,000/2,000,000", "$135"),
            ("2,000,000/4,000,000", "$225")
        ]

        data = pd.DataFrame(data, columns=["Limits", "Rate"])

        return data

    import pandas as pd

    def buildPetServicesMobileEquip(self):
        data = [
            ("$15,000", "$49"),
            ("$25,000", "$85"),
            ("$50,000", "$166"),
            ("$100,000", "$220")
        ]
        return pd.DataFrame(data, columns=["Limits", "Mobile Equipment"])

    def buildPetServicesCustomizedVehicle(self):
        data = [
            ("$25,000", "$91", "$46"),
            ("$50,000", "$104", "$59"),
            ("$100,000", "$117", "$71")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Vehicle", "Each Additional Vehicle"])

    def buildPetServicesBusinessIncome(self):
        data = [
            ("$25,000", "$13", "$7"),
            ("$50,000", "$20", "$13"),
            ("$100,000", "$26", "$20")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    def buildVetMobileEquip(self):
        data = [
            ("$15,000", "$122"),
            ("$25,000", "$211"),
            ("$50,000", "$414"),
            ("$100,000", "$549")
        ]
        return pd.DataFrame(data, columns=["Limits", "Mobile Equipment"])

    def buildVetCustomizedVehicle(self):
        data = [
            ("$25,000", "$227", "$113"),
            ("$50,000", "$260", "$146"),
            ("$100,000", "$293", "$179")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Vehicle", "Each Additional Vehicle"])

    def buildVetBusinessIncome(self):
        data = [
            ("$25,000", "$32", "$17"),
            ("$50,000", "$49", "$32"),
            ("$100,000", "$65", "$49")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    def buildPetServicesSpecializedEndorsement(self):
        data = {"Base Premium for each Office Premises" : ["$212.00"]}
        data = pd.DataFrame(data)

        return data

    def buildPetServicesSpecializedEndorsementIncome(self):
        data = [
            ("$25,000", "$13", "$7"),
            ("$50,000", "$20", "$13"),
            ("$100,000", "$26", "$20")
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    def buildPetServicesProfLiab(self):
        data = [
            ("300,000/600,000", "$43"),
            ("500,000/1,000,000", "$56"),
            ("1,000,000/2,000,000", "$68"),
            ("2,000,000/4,000,000", "$83")
        ]
        return pd.DataFrame(data, columns=["Limits", "Rate"])

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the base rates worksheet
    def formatBaseRates(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(159)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = '#,##0.0000' # 4 values after the decimal point for the base rates

    # Applies manual formatting to the territory multiplier worksheet
    #def formatTerritoryMultiplier(self, ws):
    #    ws.column_dimensions['A'].width = self.pixelsToInches(70)
    #    for col in range(2, ws.max_column + 1):
    #        char = get_column_letter(col) # Letter representing the current column
    #        ws.column_dimensions[char].width = self.pixelsToInches(54)  

    # Applies manual formatting to the construction factor worksheet
    def formatConstructionFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    # Applies manual formatting to the theft options worksheet
    def formatTheftOptions(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(80)

    # Applies manual formatting to the year built modifier worksheet
    def formatYearBuiltModifier(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(131)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '###0'

    # Applies manual formatting to the equipment breakdown base rate worksheet
    def formatEBBaseRate(self, ws):
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the property damage deductible worksheet
    def formatPropertyDamageDeductible(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(187)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the liability size of risk modifier worksheet
    def formatLiabilitySizeRisk(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Building plus Business Personal Property'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col < 3:
                ws.column_dimensions[char].width = self.pixelsToInches(140)
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(100)

    # Applies manual formatting to the liability limit factor worksheet
    def formatLiabilityLimitFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat
    
    # Applies manual formatting to the general liability occupancy modifiers worksheet
    def formatGeneralOccupancyMod(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(70)
        ws.column_dimensions['C'].width = self.pixelsToInches(215)

    # Applies manual formatting to the directors and officers liability insurance worksheet
    def formatDirsOfficersLiabIns(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(130)
        ws.column_dimensions['D'].width = self.pixelsToInches(140)
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['B' + str(row)].number_format = self.noDecimalFormat
            ws['C' + str(row)].number_format = '$#,##0.00'
            ws['D' + str(row)].number_format = '$#,##0.00'

    # Applies manual formatting to the directors and officers liability insurance - non-monetary relief worksheet
    def formatDirsOfficersNonMonetaryRelief(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(225)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat
            ws['B' + str(row)].number_format = '$#,##0.00'

    # Applies manual formatting to the directors and officers liability insurance - extended reporting periods worksheet
    def formatDirsOfficersReportingPeriods(self, ws):
        ws.column_dimensions['B'].width = self.pixelsToInches(215)
    
    # Applies manual formatting to the optometrists professional liability worksheet
    def formatOptometristsProfessionalLiab(self, ws):
        for col in range(1, ws.max_column + 1): 
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(170)
            if col > 1: # Applying specific currency formatting to every column except A
                for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = '$#,##0.00'

    # Applies manual formatting to the veterinarians specialized liability worksheet
    def formatVetSpecializedLiab(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(120)
            if col > 1: # Applying specific currency formatting to every column except A
                for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = '$#,##0.00'

        ws.insert_rows(3)
        ws["A3"] = "Veterinarian Services - Business Income"
        ws["A3"].alignment = Alignment(horizontal='left',vertical='center',wrap_text=False)
        ws["A3"].font = Font(bold = True, name = "Arial", size = 10)

        # Bolding fourth row
        for col in range(1, ws.max_column + 1):
            cell = ws[f"{get_column_letter(col)}4"]
            cell.font = Font(bold=True, name="Arial", size=10)

    # Applies manual formatting to the veterinarians professional liability worksheet
    def formatVetProfessionalLiab(self, ws):
        for col in range(1, ws.max_column + 1): 
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(205)
            if col > 1: # Applying specific currency formatting to every column except A
                for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = '$#,##0.00'

    # Applies manual formatting to the endorsement charge worksheet
    def formatEndorsementCharge(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(350)
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the franchise endorsement worksheet
    def formatFranchiseEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            cell = ws['B' + str(row)]
            cell.number_format = '$#,##0.00'

    def formatArchitectsEngineersEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(300)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatConsultantSpecializedEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(300)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatProfessionaServicesEnforsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(300)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatAccountantsSpecialziedEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(300)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatAttorneySpecializedEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(300)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatHealthCareSpecializedEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(300)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatVeterinarianSpecializedEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(120)
        ws.column_dimensions['B'].width = self.pixelsToInches(120)
        ws.column_dimensions['C'].width = self.pixelsToInches(120)


        ws.insert_rows(6)
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:C4")
        ws["A5"].border = Border()
        ws["A6"] = "Veterinarian Services - Business Income"
        bold_font = Font(bold=True, name = 'Arial', size = 10)

        rows_to_bold = [6]

        for row in rows_to_bold:
            for r in [row, row + 1]:
                for cell in ws[r]:
                    cell.font = bold_font


    def formatVetProfLiab(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'

        ws.insert_rows(2)
        ws.insert_rows(9)

        ws["A3"] = "Rate per Veterinarian - Household Pet"
        ws["A3"].alignment = Alignment(horizontal='left', vertical='center',wrap_text = False)

        ws["A10"] = "Rate per Veterinarian - Non Household Pet"
        ws["A10"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)


        bold_font = Font(bold=True, name = 'Arial', size = 10)
        rows_to_bold = [3, 10]

        for row in rows_to_bold:
            for r in [row, row + 1]:
                for cell in ws[r]:
                    cell.font = bold_font

        for row in rows_to_bold:
            ws[f"A{row}"].border = Border()


    def formatMobilePetVetServices(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)
        ws.column_dimensions['C'].width = self.pixelsToInches(175)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'
            ws['C' + str(row)].number_format = '$#,##0.00'

        ws.insert_rows(2)
        ws.insert_rows(9)
        ws.insert_rows(15)
        ws.insert_rows(21)
        ws.insert_rows(28)
        ws.insert_rows(34)

        ws["A3"] = "Pet Services"
        ws["A10"] = "Pet Services Customized Vehicle"
        ws["A16"] = "Pet Services - Business Income"
        ws["A22"] = "Veterinarian"
        ws["A29"] = "Veterinarian Services per Customized Vehicle"
        ws["A35"] = "Veterinarian Services - Business Income"

        ws["A3"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws["A10"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws["A16"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws["A22"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws["A29"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
        ws["A35"].alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        bold_font = Font(bold=True, name = 'Arial', size = 10)
        rows_to_bold = [3, 10, 16, 22, 29, 35]

        for row in rows_to_bold:
            for r in [row, row + 1]:
                for cell in ws[r]:
                    cell.font = bold_font

        for row in rows_to_bold:
            ws[f"A{row}"].border = Border()

    def formatPetServicesSpecializedEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(120)
        ws.column_dimensions['B'].width = self.pixelsToInches(120)
        ws.column_dimensions['C'].width = self.pixelsToInches(120)


        ws.insert_rows(6)
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:C4")
        ws["A5"].border = Border()
        ws["A6"] = "Pet Services - Business Income"
        bold_font = Font(bold=True, name = 'Arial', size = 10)

        rows_to_bold = [6]

        for row in rows_to_bold:
            for r in [row, row + 1]:
                for cell in ws[r]:
                    cell.font = bold_font


    def formatPetServicesProfLiab(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '$#,##0.00'


    # Sets up the Office Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildOfficePage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        Office = ExcelSettings.Excel(state=self.state, programName='Office', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = Office.getFontName()
        fontSize = Office.getFontSize()

        if 'NACO' in self.rateTables.keys():
            Office.generateWorksheet('BRNACO', 'O Table 3.B.1. NW Assurance State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NAFF' in self.rateTables.keys():
            Office.generateWorksheet('BRNAFF', 'O Table 3.B.1. NW Affinity State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NGIC' in self.rateTables.keys():
            Office.generateWorksheet('BRNGIC', 'O Table 3.B.1. NW General Insurance Company', self.buildBaseRates('NGIC'), False, True)
        if 'NICOF' in self.rateTables.keys():
            Office.generateWorksheet('BRNICOF', 'O Table 3.B.1. NICOF State Base Rates', self.buildBaseRates('NGIC'), False, True)
        #Office.generateWorksheet('TRBG', 'O Table 3.C.1.a. State Territory Multiplier - Building', self.buildTerritoryMultiplier('Building'), False, True)
        #Office.generateWorksheet('TRPP', 'O Table 3.C.1.a. State Territory Multiplier - BPP', self.buildTerritoryMultiplier('BPP'), False, True)
        #Office.generateWorksheet('TRLB', 'O Table 3.C.1.a. State Territory Multiplier - Liability', self.buildTerritoryMultiplier('Liability'), False, True)
        Office.generateWorksheet('CBG', 'O Table 3.C.2.c. Construction Factor - Building', self.buildConstructionType('Building'), False, True)
        Office.generateWorksheet('CPP', 'O Table 3.C.2.c. Construction Factor - BPP', self.buildConstructionType('BPP'), False, True)
        Office.generateWorksheet('ET', 'O Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions(), False, True)
        Office.generateWorksheet('YBBG', 'O Table 3.C.2.p. Year Built Modifier - Building', self.buildYearBuiltModifier('Building'), False, True)
        Office.generateWorksheet('YBPP', 'O Table 3.C.2.p. Year Built Modifier - BPP', self.buildYearBuiltModifier('BPP'), False, True)
        Office.generateWorksheet('EBB', 'O Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate(), False, True)
        Office.generateWorksheet('PDLD', 'O Table 3.C.4.b. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount(), False, True)
        Office.generateWorksheet('LS', 'O Table 3.C.4.d. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk(), False, True)
        Office.generateWorksheet('LL', 'O Table 3.C.4.e. Liability Limit Factor', self.buildLiabilityLimitFactor(), False, True)
        Office.generateWorksheet('GLO', 'O Table 3.D.1.c. General Liability Occupancy Modifiers', self.buildGeneralOccupancyMod(), False, True)
        Office.generateWorksheet('DO', 'O Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns(), False, True)
        Office.generateWorksheet('DONM', 'O Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief(), False, True)
        Office.generateWorksheet('ERP', 'O Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods(), False, True)
        Office.generateWorksheet('OPTO', 'O Table 4.B.5.a. Optometrists Professional Liability', self.buildOptometristsProfessionalLiab(), False, True)
        Office.generateWorksheet('VSPL', 'O Table 4.C.4.A. Veterinarian Specialized Endorsement With Professional Liability', self.buildVetSpecializedLiab(), False, True)
        Office.generateWorksheet('VET', 'O Table 4.C.5.a. Veterinarians Professional Liability', self.buildVetProfessionalLiab(), False, True)
        Office.generateWorksheet('PLUS', 'O Table 4.D. Office PLUS Endorsement', self.buildEndorsementCharge(), False, True)
        Office.generateWorksheet('FR', 'O Table 4.E. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement(), False, True)
        Office.generateWorksheet('AES', 'O Table 4.F. Architects and Engineers Specialized Endorsement', self.buildArchitectsEngineersEndorsement(), False, True)
        Office.generateWorksheet('CS', 'O Table 4.G. Consultants Specialized Endorsement', self.buildConsultantSpecializedEndorsement(), False, True)
        Office.generateWorksheet('PFSS', 'O Table 4.H. Professional Services Specialized Endorsement', self.buildProfessionaServicesEnforsement(), False, True)
        Office.generateWorksheet('ACS', 'O Table 4.I. Accountants Specialized Endorsement', self.buildAccountantsSpecialziedEndorsement(), False, True)
        Office.generateWorksheet('ATS', 'O Table 4.J. Attorneys Specialized Endorsement', self.buildAttorneySpecializedEndorsement(), False, True)
        Office.generateWorksheet('HCS', 'O Table 4.K. Health Care Specialized Endorsement', self.buildHealthCareSpecializedEndorsement(), False, True)
        Office.generateWorksheet2tables('VS', 'O Table 4.L.3. Veterinarian Specialized Endorsement', self.buildVeterinarianSpecializedEndorsement(), self.buildVeterinarianSpecializedEndorsementIncome(), False, True)
        Office.generateWorksheet2tables('VPL', 'O Table 4.M. Veterinarian Professional Liability', self.buildVetProfLiabHousehold(), self.buildVetProfLiabNonHousehold(), False, True)
        Office.generateWorksheet6tables('MPVS', 'O Table 4.N. Mobile Pet and Veterinarian Services Endorsement', self.buildPetServicesMobileEquip(), self.buildPetServicesCustomizedVehicle(), self.buildPetServicesBusinessIncome(),self.buildVetMobileEquip(), self.buildVetCustomizedVehicle(), self.buildVetBusinessIncome(),False, True)
        Office.generateWorksheet2tables('PSS', 'O Table 4.O. Pet Services Specialized Endorsement', self.buildPetServicesSpecializedEndorsement(), self.buildPetServicesSpecializedEndorsementIncome(), False, True)
        Office.generateWorksheet('PSPL', 'O Table 4.P. Pet Services Professional Liability', self.buildPetServicesProfLiab(), False, True)

        Office.createIndex()
        OfficePages = Office.getWB()

        if 'NACO' in self.rateTables.keys():
            self.formatBaseRates(OfficePages['BRNACO'])
        if 'NAFF' in self.rateTables.keys():
            self.formatBaseRates(OfficePages['BRNAFF'])
        if 'NGIC' in self.rateTables.keys():
            self.formatBaseRates(OfficePages['BRNGIC'])
        if 'NICOF' in self.rateTables.keys():
            self.formatBaseRates(OfficePages['BRNICOF'])
        #self.formatTerritoryMultiplier(OfficePages['TRBG'])
        #self.formatTerritoryMultiplier(OfficePages['TRPP']) 
        #self.formatTerritoryMultiplier(OfficePages['TRLB'])
        self.formatConstructionFactor(OfficePages['CBG'])  
        self.formatConstructionFactor(OfficePages['CPP'])  
        self.formatTheftOptions(OfficePages['ET'])
        self.formatYearBuiltModifier(OfficePages['YBBG'])
        self.formatYearBuiltModifier(OfficePages['YBPP'])  
        self.formatEBBaseRate(OfficePages['EBB'])
        self.formatPropertyDamageDeductible(OfficePages['PDLD'])
        self.formatLiabilitySizeRisk(OfficePages['LS'], Font(name=fontName, size=fontSize, bold=True))
        self.formatLiabilityLimitFactor(OfficePages['LL'])
        self.formatGeneralOccupancyMod(OfficePages['GLO'])
        self.formatDirsOfficersLiabIns(OfficePages['DO'])
        self.formatDirsOfficersNonMonetaryRelief(OfficePages['DONM'])
        self.formatDirsOfficersReportingPeriods(OfficePages['ERP'])
        self.formatOptometristsProfessionalLiab(OfficePages['OPTO'])
        self.formatVetSpecializedLiab(OfficePages['VSPL'])
        self.formatVetProfessionalLiab(OfficePages['VET'])
        self.formatEndorsementCharge(OfficePages['PLUS'])
        self.formatFranchiseEndorsement(OfficePages['FR'])
        self.formatArchitectsEngineersEndorsement(OfficePages['AES'])
        self.formatConsultantSpecializedEndorsement(OfficePages['CS'])
        self.formatProfessionaServicesEnforsement(OfficePages['PFSS'])
        self.formatAccountantsSpecialziedEndorsement(OfficePages['ACS'])
        self.formatAttorneySpecializedEndorsement(OfficePages['ATS'])
        self.formatHealthCareSpecializedEndorsement(OfficePages['HCS'])
        self.formatVeterinarianSpecializedEndorsement(OfficePages['VS'])
        self.formatVetProfLiab(OfficePages['VPL'])
        self.formatMobilePetVetServices(OfficePages['MPVS'])
        self.formatPetServicesSpecializedEndorsement(OfficePages['PSS'])
        self.formatPetServicesProfLiab(OfficePages['PSPL'])

        return OfficePages