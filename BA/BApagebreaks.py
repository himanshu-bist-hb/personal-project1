"""
BApagebreaks.py
================
Post-processing for the BA Rate Pages workbook.

PIPELINE
  1. openpyxl: truncate sheet names > 31 chars, move Index to position 0,
     apply per-rule page breaks and print settings (in-memory; no COM).
  2. XML sanitize pass on the saved .xlsx zip: removes the few openpyxl
     output quirks that make Excel show "Open and Repair" on file open.

Pure Python, no Excel.exe. Typical run: ~1-2 seconds end-to-end.

----------------------------------------------------------------------------
HOW TO ADD A PAGE-BREAK RULE FOR A NEW SHEET
----------------------------------------------------------------------------

Step 1 - Write a handler function. Signature: (ws, dest_filename) -> None
         where `ws` is an openpyxl Worksheet and `dest_filename` is the
         absolute file path (use it for state/file-specific branching).

    def _handle_rule_999(ws, dest_filename):
        ws.print_area = f"A1:H{ws.max_row}"
        disable_fit_to_page(ws)
        add_break_after(ws, 37)

Step 2 - Register it in SHEET_RULES below (one line):

    ("Rule 999", _handle_rule_999),

Order matters: more-specific prefixes BEFORE less-specific ones
(e.g., "Rule 239 C" must come before "Rule 239 ").

----------------------------------------------------------------------------
HELPERS YOU CAN USE INSIDE A HANDLER
----------------------------------------------------------------------------

  fit_single_page(ws)         entire sheet onto one printed page
  fit_width_only(ws)          fit width to 1 page; height grows with content
  disable_fit_to_page(ws)     turn off fit-to-page; manual breaks rule
  add_break_after(ws, row)    add a horizontal page break AFTER the given row

OPENPYXL API CHEATSHEET
  ws.page_setup.orientation     = "portrait" or "landscape"
  ws.page_setup.fitToWidth      = 1
  ws.page_setup.fitToHeight     = 1   (or 0 for unlimited tall)
  ws.print_area                 = "A1:H50"
  ws.print_title_rows           = "1:3"   (or None to clear)
  ws.print_options.horizontalCentered = False
  ws.print_options.verticalCentered   = False
  ws.page_margins.top           = 1.00     (inches)
  ws.row_breaks.append(Break(id=N))         break ABOVE row N
"""

import os
import re
import shutil
import subprocess
import time
import zipfile
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break, ColBreak, RowBreak


# ============================================================================
#  HELPERS  -  use these inside rule handlers
# ============================================================================

def fit_single_page(ws):
    """Fit the entire content of the sheet onto a single printed page."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1


def fit_width_only(ws):
    """Fit width to 1 page; height grows as needed (manual breaks honored)."""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def disable_fit_to_page(ws):
    """Turn off fit-to-page entirely; manual breaks are honored as-is."""
    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.fitToWidth = 0
    ws.page_setup.fitToHeight = 0


def add_break_after(ws, row):
    """Add a horizontal page break AFTER the given row (1-indexed)."""
    ws.row_breaks.append(Break(id=row))


# ============================================================================
#  RULE HANDLERS
#  Signature: (ws, dest_filename) -> None
# ============================================================================

def _handle_index(ws, dest_filename):
    # NOTE: print_title_rows = None (NOT "0:0") so we don't write an invalid
    # definedName like Index!$0:$0 - that is what triggers Open-and-Repair.
    ws.print_title_rows = None
    ws.print_area = f"A1:J{ws.max_row}"
    fit_width_only(ws)


def _handle_rule_222b(ws, dest_filename):
    ws.print_area = f"A1:J{ws.max_row}"
    fit_width_only(ws)
    add_break_after(ws, 25)
    add_break_after(ws, 49)


def _handle_rule_222ttt(ws, dest_filename):
    fit_single_page(ws)
    ws.print_title_rows = "1:3"


def _handle_rule_223b5(ws, dest_filename):
    ws.page_setup.orientation = "landscape"


def _handle_rule_223c(ws, dest_filename):
    fit_single_page(ws)


def _handle_rule_225_zone(ws, dest_filename):
    ws.print_area = f"A1:G{ws.max_row}"
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    fit_width_only(ws)
    add_break_after(ws, 51)
    add_break_after(ws, 103)


def _handle_rule_225c3(ws, dest_filename):
    fit_single_page(ws)


def _handle_rule_232ppt(ws, dest_filename):
    fit_single_page(ws)
    ws.print_title_rows = "1:3"


def _handle_rule_239_general(ws, dest_filename):
    fit_single_page(ws)
    ws.print_title_rows = "1:3"


def _handle_rule_239c(ws, dest_filename):
    fit_single_page(ws)
    ws.page_margins.top = 1.00


def _handle_rule_240(ws, dest_filename):
    fit_single_page(ws)
    ws.print_options.verticalCentered = True
    ws.print_title_rows = "1:3"
    ws.print_area = f"A1:G{ws.max_row}"
    ws.page_margins.top = 1.00


def _handle_rule_255(ws, dest_filename):
    ws.print_area = f"A1:H{ws.max_row}"
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    disable_fit_to_page(ws)
    add_break_after(ws, 37)


def _handle_rule_275(ws, dest_filename):
    if ws["A10"].value == "275.B.1.(b). Short Term - Autos Leased by the Hour, Day, or Week":
        ws.print_title_rows = "1:1"
        fit_single_page(ws)


def _handle_rule_283(ws, dest_filename):
    ws.print_area = f"A1:P{ws.max_row}"
    targets = {
        "283.B Limited Specified Causes of Loss",
        "283.B Comprehensive",
        "283.B Blanket Collision",
    }
    for row in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=1).value)
        if cell_value in targets and row > 3:
            # break BEFORE this row (= after row-1)
            ws.row_breaks.append(Break(id=row - 1))
    fit_width_only(ws)


def _handle_rule_289(ws, dest_filename):
    ws.print_area = f"A1:H{ws.max_row}"
    disable_fit_to_page(ws)
    add_break_after(ws, 37)


def _handle_rule_297(ws, dest_filename):
    ws.print_area = f"A1:G{ws.max_row}"
    ws.col_breaks = ColBreak()
    ws.row_breaks = RowBreak()
    fit_width_only(ws)
    occurrence_count = 0
    for row in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=1).value)
        if cell_value.startswith("Single") or cell_value.startswith("Uninsured"):
            occurrence_count += 1
        if (occurrence_count % 3 == 0) and (occurrence_count != 0):
            occurrence_count += 1
            add_break_after(ws, row - 1)


def _handle_rule_298(ws, dest_filename):
    ws.print_area = f"A1:G{ws.max_row}"
    ws.col_breaks = ColBreak()
    ws.row_breaks = RowBreak()
    fit_width_only(ws)
    add_break_after(ws, 33)
    add_break_after(ws, 57)
    add_break_after(ws, 88)


_VA_VEHICLE_TYPES = {
    "Extra Heavy Truck-Tractor", "Extra-Heavy Truck", "Heavy Truck",
    "Heavy Truck-Tractor", "Light Truck", "Medium Truck",
    "Private Passenger Types", "Semitrailer",
    "Service or Utility Trailer", "Trailer",
}


def _handle_rule_301ab(ws, dest_filename):
    if ws["B4"].value in _VA_VEHICLE_TYPES:
        return
    fit_width_only(ws)
    if ws.title.startswith("Rule 301.B"):
        ws.print_area = f"A1:T{ws.max_row}"
    for row in range(46, ws.max_row, 45):
        add_break_after(ws, row)
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.page_setup.orientation = "landscape"
    ws.page_margins.top = 1.00


def _handle_rule_301c(ws, dest_filename):
    if ws["B4"].value in _VA_VEHICLE_TYPES:
        if "FL" not in dest_filename:
            ws.page_margins.top = 1.00
        fit_single_page(ws)
        return

    # Factor columns (D onwards) default to ~12 char-units wide (0.94" each).
    # With 26 such columns, Set 2 is ~28" wide — forcing fit_width_only to
    # scale down to ~33%.  At 33%, Set 1 (7 cols) fills only ~26% of the page.
    # Width 6 reduces Set 2 to ~16.6", giving a ~57% scale.  At 57% scale,
    # the automatic page break falls at row ~46, coinciding with the manual
    # break — no spurious dashed lines appear within any section.  Width 5
    # gave 65% scale but that put the auto break at row ~41 (before the
    # manual break at 46), creating unwanted intermediate pages.
    for col_idx in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 6

    ws.col_breaks = ColBreak()
    ws.row_breaks = RowBreak()
    ws.page_setup.orientation = "landscape"
    ws.print_area = f"A1:AC{ws.max_row}"
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    fit_width_only(ws)

    for row in [46, 91, 136, 181, 226, 271, 316, 361]:
        add_break_after(ws, row)
    for row in [406, 451, 496, 541, 586, 631, 676, 720]:
        add_break_after(ws, row)


def _handle_rule_301cd(ws, dest_filename):
    if ws["B4"].value not in _VA_VEHICLE_TYPES:
        return
    if "FL" not in dest_filename:
        ws.page_margins.top = 1.00
    fit_single_page(ws)


def _handle_rule_306(ws, dest_filename):
    fit_width_only(ws)
    ws.print_title_rows = "1:4"


def _handle_rule_315(ws, dest_filename):
    fit_width_only(ws)
    add_break_after(ws, 23)


def _handle_rule_r1(ws, dest_filename):
    ws.print_area = f"A1:M{ws.max_row}"
    disable_fit_to_page(ws)
    occurrence_count = 0
    for row in range(1, ws.max_row + 1):
        cell_value = str(ws.cell(row=row, column=1).value)
        if cell_value.startswith("R1"):
            occurrence_count += 1
        if occurrence_count in (3, 6):
            occurrence_count += 1
            ws.row_breaks.append(Break(id=row - 1))


# ============================================================================
#  RULE REGISTRY
#  Order matters: more-specific prefixes BEFORE less-specific ones.
#  To add a new rule: write the handler above, then add one line here.
# ============================================================================

SHEET_RULES = [
    ("Index",        _handle_index),

    ("Rule 222 B",   _handle_rule_222b),
    ("Rule 222 TTT", _handle_rule_222ttt),

    ("Rule 223 B.5", _handle_rule_223b5),
    ("Rule 223 C",   _handle_rule_223c),

    ("Rule 225 Zone", _handle_rule_225_zone),
    ("Rule 225.C.3", _handle_rule_225c3),

    ("Rule 232 PPT", _handle_rule_232ppt),

    ("Rule 239 C",   _handle_rule_239c),       # specific BEFORE generic
    ("Rule 239 ",    _handle_rule_239_general),

    ("Rule 240 ",    _handle_rule_240),

    ("Rule 255",     _handle_rule_255),

    ("Rule 275",     _handle_rule_275),

    ("Rule 283",     _handle_rule_283),
    ("Rule 289",     _handle_rule_289),
    ("Rule 297",     _handle_rule_297),
    ("Rule 298",     _handle_rule_298),

    ("Rule 301.C",   _handle_rule_301c),
    ("Rule 301.D",   _handle_rule_301cd),
    ("Rule 301.A",   _handle_rule_301ab),
    ("Rule 301.B",   _handle_rule_301ab),

    ("Rule 306",     _handle_rule_306),
    ("Rule 315",     _handle_rule_315),

    ("Rule R1",      _handle_rule_r1),
]


def _apply_matching_rule(sheet_name, ws, dest_filename):
    """Walk SHEET_RULES and run the first handler whose prefix matches."""
    for prefix, handler in SHEET_RULES:
        if sheet_name.startswith(prefix):
            handler(ws, dest_filename)
            return True
    return False


# ============================================================================
#  PUBLIC ENTRY POINT
# ============================================================================

def process_pagebreaks(dest_filename1, dest_filename2):
    """
    Apply page breaks / print settings to dest_filename1.

    dest_filename2 is accepted for backward compatibility (was a PDF path).
    """
    print(f"[BApagebreaks] Processing: {dest_filename1}")
    dest_filename1 = os.path.normpath(os.path.abspath(dest_filename1))

    workbook = openpyxl.load_workbook(dest_filename1)

    # Truncate sheet names exceeding Excel's 31-character limit
    for original_name in list(workbook.sheetnames):
        if len(original_name) > 31:
            workbook[original_name].title = original_name[:31]

    # Apply defaults + rules to every sheet
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.print_title_rows = "1:1"
        fit_single_page(ws)
        _apply_matching_rule(sheet_name, ws, dest_filename1)

    # Index goes to position 0, visible, active on open
    if "Index" in workbook.sheetnames:
        ws_index = workbook["Index"]
        ws_index.sheet_state = "visible"
        if workbook.sheetnames.index("Index") != 0:
            workbook._sheets.remove(ws_index)
            workbook._sheets.insert(0, ws_index)
        workbook.active = 0

    workbook.save(dest_filename1)
    workbook.close()

    # Sanitize the saved zip to prevent Excel's "Open and Repair" popup
    _sanitize_xlsx(dest_filename1)
    print("[BApagebreaks] Done.")


# ============================================================================
#  XML SANITIZE PASS
# ============================================================================
#
#  Why: openpyxl's xlsx output sometimes contains tiny XML quirks that make
#  Excel show "We found a problem with some content" on open. The safest
#  fixes Excel's repair would apply, applied here directly:
#
#    * Drop <definedName> entries whose ref contains "$0" - rows are 1-indexed
#      in Excel; "$0" never matches anything and Excel flags it as invalid.
#
#    * Drop <definedName> entries whose ref is empty.
#
#    * Drop empty <definedNames/> wrappers if they end up with no children.
#
#  Operates directly on the .xlsx zip. ~0.5 seconds even for large workbooks.
#  No Excel process needed.
# ============================================================================

# Match an entire <definedName ...>...</definedName> element whose body
# contains "$0" anywhere - the only practical way openpyxl produces an
# invalid range like Sheet1!$0:$0.
_BAD_DEFINED_NAME_RE = re.compile(
    rb"<definedName\b[^>]*>[^<]*\$0[^<]*</definedName>"
)
# Also catch self-closing <definedName ... ref="..."/> variants and
# definedName elements with empty body.
_EMPTY_DEFINED_NAME_RE = re.compile(
    rb"<definedName\b[^>]*>\s*</definedName>"
)
# If the wrapper ends up empty, drop the whole <definedNames/> element.
_EMPTY_DEFINED_NAMES_WRAPPER_RE = re.compile(
    rb"<definedNames\s*/?>\s*</definedNames>|<definedNames\s*/>"
)


def _sanitize_xlsx(filename):
    """
    Open the xlsx as a zip, rewrite xl/workbook.xml in-place to drop any
    <definedName> elements with invalid "$0" or empty refs, and save the
    zip back. Pure Python, no Excel.
    """
    tmp = filename + ".sanitize.tmp"
    changed = False

    with zipfile.ZipFile(filename, "r") as zin:
        infos = zin.infolist()
        contents = {info.filename: zin.read(info.filename) for info in infos}

    if "xl/workbook.xml" in contents:
        original = contents["xl/workbook.xml"]
        cleaned = _BAD_DEFINED_NAME_RE.sub(b"", original)
        cleaned = _EMPTY_DEFINED_NAME_RE.sub(b"", cleaned)
        cleaned = _EMPTY_DEFINED_NAMES_WRAPPER_RE.sub(b"", cleaned)
        if cleaned != original:
            contents["xl/workbook.xml"] = cleaned
            changed = True

    if not changed:
        # Nothing to fix - keep original file as-is
        return

    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            zout.writestr(info, contents[info.filename])

    os.replace(tmp, filename)


# ============================================================================
#  PDF EXPORT  -  drives Excel via COM to preserve every print setting
#  (page breaks, fit-to-page, print areas, orientation, margins, headers).
# ============================================================================

_XL_TYPE_PDF        = 0   # xlTypePDF
_XL_QUALITY_STD     = 0   # xlQualityStandard


def _strip_pdf_tags(pdf):
    """Drop the accessibility tag tree + metadata from an open pikepdf.Pdf.
    Excel writes one structure element per table cell, which dominates the
    file size of dense rate pages (a 10 MB workbook can export as 170 MB)."""
    root = pdf.Root
    for key in ("/StructTreeRoot", "/MarkInfo", "/Metadata",
                "/PieceInfo", "/Lang"):
        if key in root:
            del root[key]
    for page in pdf.pages:
        for key in ("/StructParents", "/Tabs", "/PieceInfo"):
            if key in page:
                del page[key]
    pdf.remove_unreferenced_resources()


def _compress_pdf(src_path, dest_path):
    """
    Write a compressed copy of src_path to dest_path.

    Excel's ExportAsFixedFormat produces a "tagged" PDF: every table cell
    gets an accessibility structure element, which for rate pages (hundreds
    of pages of dense tables) is most of the file — a 10 MB workbook can
    export as a 170 MB PDF. The DOI filing copy doesn't need the tag tree,
    so drop it and recompress every stream (measured ~5x smaller).

    Best-effort: if pikepdf is not installed or anything fails, the raw
    PDF is moved to dest_path unchanged so export never breaks.
    """
    try:
        import pikepdf
    except ImportError:
        print("[pagebreaks] pikepdf not installed — PDF left uncompressed "
              "(pip install pikepdf to shrink output ~5x)")
        shutil.move(src_path, dest_path)
        return

    try:
        with pikepdf.open(src_path) as pdf:
            _strip_pdf_tags(pdf)
            pdf.save(dest_path,
                     compress_streams=True,
                     recompress_flate=True,
                     object_stream_mode=pikepdf.ObjectStreamMode.generate)
        raw_mb  = os.path.getsize(src_path) / 1e6
        out_mb  = os.path.getsize(dest_path) / 1e6
        print(f"[pagebreaks] PDF compressed {raw_mb:.1f} MB -> {out_mb:.1f} MB")
        os.remove(src_path)
    except Exception as exc:
        print(f"[pagebreaks] PDF compression failed ({exc}) — keeping raw PDF")
        shutil.move(src_path, dest_path)


def _kill_excel_instances():
    """Best-effort: terminate any orphan EXCEL.EXE so COM gets a clean slate."""
    try:
        subprocess.run(
            ["taskkill", "/f", "/im", "excel.exe"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False,
        )
    except Exception:
        pass


# ============================================================================
#  PARALLEL PDF EXPORT
#
#  Excel's PDF renderer is single-threaded at ~8k cells/second, so a sheet
#  like BOP's TRDEF (82k rows x 21 cols = 1.7M cells) dominates the export.
#  When a workbook contains a sheet that big, several Excel instances each
#  export a different slice of the workbook — a huge sheet is sliced by ROW
#  RANGE via an in-memory print area, so the .xlsx keeps its single tab and
#  the printed footer/tab name is unchanged. Page numbers stay continuous:
#  every worker first counts its own pages (phase 1), then the cumulative
#  start is stamped via FirstPageNumber before exporting (phase 2).
#  Measured on TRDEF-scale data: 224s single-instance -> 86s with 4 workers.
#
#  Any failure falls back to the normal single-instance export.
# ============================================================================

_HUGE_SHEET_ROWS  = 10_000   # a sheet taller than this triggers the parallel path
_PARALLEL_WORKERS = 4


def _plan_parallel_export(xlsx_path):
    """
    Return a list of worker groups, or None when a plain single-instance
    export is the right choice. Each group is a list of
    (sheet_name, first_row, last_row, max_col) units in workbook order;
    first_row/last_row are None for "whole sheet".
    """
    try:
        import pikepdf  # required to merge the per-worker PDFs
    except ImportError:
        return None

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            if ws.title == "Index" or ws.sheet_state != "visible":
                continue
            sheets.append((ws.title, ws.max_row or 1, ws.max_column or 1))
    finally:
        wb.close()

    if not any(rows > _HUGE_SHEET_ROWS for _, rows, _ in sheets):
        return None

    # Units: whole small sheets; huge sheets sliced into row ranges of
    # roughly equal weight so the groups balance.
    total_weight = sum(rows * cols for _, rows, cols in sheets)
    target = total_weight / _PARALLEL_WORKERS
    units = []
    for name, rows, cols in sheets:
        weight = rows * cols
        if rows <= _HUGE_SHEET_ROWS:
            units.append((name, None, None, cols, weight))
            continue
        n_slices = max(2, min(_PARALLEL_WORKERS * 2, round(weight / target) or 2))
        step = -(-rows // n_slices)  # ceil
        r = 1
        while r <= rows:
            r2 = min(r + step - 1, rows)
            units.append((name, r, r2, cols, (r2 - r + 1) * cols))
            r = r2 + 1

    # Contiguous partition into up to _PARALLEL_WORKERS groups, coalescing
    # adjacent slices of the same sheet that land in the same group (a
    # worker can only print one area per sheet).
    groups, cur, cur_weight = [], [], 0
    for unit in units:
        name, r1, r2, cols, weight = unit
        if cur and cur_weight >= target and len(groups) < _PARALLEL_WORKERS - 1:
            groups.append(cur)
            cur, cur_weight = [], 0
        if cur and cur[-1][0] == name and cur[-1][2] is not None and r1 is not None:
            prev = cur[-1]
            cur[-1] = (name, prev[1], r2, cols)
        else:
            cur.append((name, r1, r2, cols))
        cur_weight += weight
    if cur:
        groups.append(cur)

    return groups if len(groups) >= 2 else None


def _parallel_worker(xlsx_path, group, out_pdf, idx, counts, counted,
                     starts, go, errors):
    """Phase 1: isolate this group's slice and count its pages. Phase 2 (after
    the coordinator computes cumulative starts): stamp FirstPageNumber and
    export. Runs in its own thread with its own Excel instance."""
    import win32com.client
    import pythoncom

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False

        workbook = excel.Workbooks.Open(xlsx_path, ReadOnly=True, UpdateLinks=0)

        by_name = {u[0]: u for u in group}
        first_sheet = None
        n_pages = 0
        for sheet in workbook.Sheets:
            unit = by_name.get(sheet.Name)
            if unit is None:
                sheet.Visible = 0  # xlSheetHidden — in-memory only (ReadOnly)
                continue
            _, r1, r2, cols = unit
            if r1 is not None:
                area = f"A{r1}:{get_column_letter(cols)}{r2}"
                sheet.PageSetup.PrintArea = area
            if first_sheet is None:
                first_sheet = sheet
            n_pages += sheet.PageSetup.Pages.Count

        counts[idx] = n_pages
        counted[idx].set()

        go.wait()
        if errors:          # another worker failed during phase 1 — abort
            return
        first_sheet.PageSetup.FirstPageNumber = starts[idx]
        first_sheet.Activate()
        workbook.ExportAsFixedFormat(
            Type=_XL_TYPE_PDF,
            Filename=out_pdf,
            Quality=_XL_QUALITY_STD,
            IncludeDocProperties=False,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )
    except Exception as exc:
        errors.append(exc)
        counted[idx].set()
    finally:
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        pythoncom.CoUninitialize()


def _export_pdf_parallel(xlsx_path, groups, pdf_path, progress):
    """Run the worker pool, then merge + strip + compress into pdf_path.
    Raises on any failure (caller falls back to the sequential export)."""
    import tempfile
    import threading
    import pikepdf

    k = len(groups)
    part_pdfs = [
        os.path.join(tempfile.gettempdir(),
                     f"~rate_pages_part{i}_{os.getpid()}.pdf")
        for i in range(k)
    ]
    counts  = [None] * k
    counted = [threading.Event() for _ in range(k)]
    starts  = [1] * k
    go      = threading.Event()
    errors  = []

    progress(f"PDF — rendering with {k} Excel instances in parallel...")
    threads = [
        threading.Thread(
            target=_parallel_worker,
            args=(xlsx_path, groups[i], part_pdfs[i], i,
                  counts, counted, starts, go, errors),
            daemon=True,
        )
        for i in range(k)
    ]
    try:
        for t in threads:
            t.start()
        for evt in counted:
            evt.wait()
        if errors:
            raise errors[0]
        for i in range(1, k):
            starts[i] = starts[i - 1] + counts[i - 1]
        go.set()
        for t in threads:
            t.join()
        if errors:
            raise errors[0]
    finally:
        go.set()  # never leave workers blocked

    for p in part_pdfs:
        if not (os.path.exists(p) and os.path.getsize(p) > 0):
            raise RuntimeError(f"parallel export part missing: {p}")

    progress("PDF — merging and compressing...")
    try:
        merged = pikepdf.new()
        parts = [pikepdf.open(p) for p in part_pdfs]
        try:
            for part in parts:
                merged.pages.extend(part.pages)
            _strip_pdf_tags(merged)
            merged.save(pdf_path,
                        compress_streams=True,
                        recompress_flate=True,
                        object_stream_mode=pikepdf.ObjectStreamMode.generate)
        finally:
            for part in parts:
                part.close()
        raw_mb = sum(os.path.getsize(p) for p in part_pdfs) / 1e6
        print(f"[pagebreaks] PDF compressed {raw_mb:.1f} MB -> "
              f"{os.path.getsize(pdf_path)/1e6:.1f} MB")
    finally:
        for p in part_pdfs:
            try: os.remove(p)
            except OSError: pass


# ============================================================================
#  DIRECT GRID RENDER
#
#  Even split across 4 Excel instances, Excel's renderer needs ~85s for a
#  TRDEF-scale sheet (82k rows x 21 cols = 1.7M cells). A sheet that big in
#  this codebase is always the simple generateWorksheet layout — title row,
#  blank row, one column-header row, then a uniform bordered grid — which
#  reportlab can draw straight from the cell values in a fraction of the
#  time. When every huge sheet sits at the END of the workbook, Excel
#  exports only the small sheets (in a worker thread, concurrently) while
#  the huge ones are drawn here, and the two PDFs are stitched with pikepdf.
#  Page numbering already restarts at 1 on every sheet (useFirstPageNumber),
#  so the stitched footer numbers stay correct.
#
#  Fidelity approximations (invisible at the 10% fit-to-page scale these
#  sheets print at): data rows are assumed single-line, and auto-fit row
#  heights come from font metrics rather than Excel's. Merged cells are not
#  supported — any such sheet (and any failure here) falls back to the
#  parallel / single-instance Excel export.
# ============================================================================

# Excel XlPaperSize -> (width, height) in points; unknown codes -> Letter.
_PAPER_SIZES_PT = {
    1: (612.0, 792.0),      # Letter
    3: (792.0, 1224.0),     # Tabloid
    5: (612.0, 1008.0),     # Legal
    8: (841.89, 1190.55),   # A3
    9: (595.276, 841.89),   # A4
    11: (419.53, 595.276),  # A5
}


def _plan_direct_render(xlsx_path):
    """Return {"sheets": [...], "has_small": bool} when the workbook's huge
    sheets can be drawn directly (all of them trailing), else None."""
    try:
        import reportlab  # noqa: F401
        import pikepdf    # noqa: F401
    except ImportError:
        return None

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        vis = [(ws.title, ws.max_row or 1) for ws in wb.worksheets
               if ws.title != "Index" and ws.sheet_state == "visible"]
    finally:
        wb.close()

    huge = [n for n, r in vis if r > _HUGE_SHEET_ROWS]
    if not huge:
        return None
    if [n for n, _ in vis[-len(huge):]] != huge:
        return None  # a huge sheet sits mid-workbook — parallel path instead
    return {"sheets": huge, "has_small": len(vis) > len(huge)}


def _xml_attrs(tag, xml_text):
    m = re.search(rf"<{tag}\b([^>]*?)/?>", xml_text)
    if not m:
        return {}
    return dict(re.findall(r'(\w+)="([^"]*)"', m.group(1)))


def _parse_grid_sheet_settings(xlsx_path, sheet_name):
    """Print-relevant settings for one sheet, parsed straight from the xlsx
    zip (openpyxl's read-only mode exposes cell values but no page setup).
    Only the head and tail of the sheet XML are kept in memory."""
    from xml.etree import ElementTree as ET
    from xml.sax.saxutils import unescape

    ns = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(xlsx_path) as z:
        wb_root = ET.fromstring(z.read("xl/workbook.xml"))
        rel_root = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        rid_target = {
            rel.get("Id"): rel.get("Target")
            for rel in rel_root
        }
        rid_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
        target = None
        sheet_pos = None
        for pos, sheet in enumerate(wb_root.iter(f"{ns}sheet")):
            if sheet.get("name") == sheet_name:
                target = rid_target.get(sheet.get(rid_attr))
                sheet_pos = pos
                break
        if target is None:
            raise RuntimeError(f"sheet {sheet_name!r} not found in workbook.xml")
        target = target.lstrip("/")
        if not target.startswith("xl/"):
            target = "xl/" + target

        title_rows = 0
        for dn in wb_root.iter(f"{ns}definedName"):
            if (dn.get("name") == "_xlnm.Print_Titles"
                    and dn.get("localSheetId") == str(sheet_pos)):
                m = re.search(r"\$\d+:\$(\d+)", dn.text or "")
                if m:
                    title_rows = int(m.group(1))

        has_merges = False
        with z.open(target) as fh:
            head = fh.read(262144)
            tail = head
            has_merges = b"<mergeCell" in head
            while True:
                chunk = fh.read(1 << 20)
                if not chunk:
                    break
                if not has_merges and b"<mergeCell" in tail[-12:] + chunk:
                    has_merges = True
                tail = (tail + chunk)[-65536:]

    head_text = head.decode("utf-8", "replace")
    tail_text = tail.decode("utf-8", "replace")

    margins = {k: float(v) for k, v in _xml_attrs("pageMargins", tail_text).items()}
    setup = _xml_attrs("pageSetup", tail_text)
    fmt_pr = _xml_attrs("sheetFormatPr", head_text)
    fit_to_page = _xml_attrs("pageSetUpPr", head_text).get("fitToPage") == "1"

    col_widths = {}
    cols_m = re.search(r"<cols>(.*?)</cols>", head_text, re.S)
    if cols_m:
        for col_attrs in re.findall(r"<col\b([^>]*?)/>", cols_m.group(1)):
            a = dict(re.findall(r'(\w+)="([^"]*)"', col_attrs))
            if "width" in a and "min" in a and "max" in a:
                for ci in range(int(a["min"]), int(a["max"]) + 1):
                    col_widths[ci] = float(a["width"])

    hf = {}
    for tag in ("oddHeader", "oddFooter"):
        m = re.search(rf"<{tag}>(.*?)</{tag}>", tail_text, re.S)
        raw = unescape(m.group(1), {"&quot;": '"', "&apos;": "'"}) if m else ""
        hf[tag] = raw.replace("_x000a_", "\n").replace("_x000d_", "")

    return {
        "has_merges": has_merges,
        "title_rows": title_rows,
        "margins": {
            "left": margins.get("left", 0.7), "right": margins.get("right", 0.7),
            "top": margins.get("top", 0.75), "bottom": margins.get("bottom", 0.75),
            "header": margins.get("header", 0.3), "footer": margins.get("footer", 0.3),
        },
        "fit_to_page": fit_to_page,
        "fit_w": int(setup.get("fitToWidth", 1)),
        "fit_h": int(setup.get("fitToHeight", 1)),
        "scale_pct": int(setup.get("scale", 100) or 100),
        "paper_size": int(setup["paperSize"]) if setup.get("paperSize") else None,
        "first_page_number": int(setup.get("firstPageNumber", 1) or 1),
        "default_row_height": float(fmt_pr.get("defaultRowHeight", 15.0)),
        "default_col_width": float(fmt_pr.get("defaultColWidth", 8.43)),
        "col_widths": col_widths,
        "odd_header": hf["oddHeader"],
        "odd_footer": hf["oddFooter"],
    }


_HF_CODE_RE = re.compile(r'&"(?P<font>[^"]*)"|&K[0-9A-Fa-f]{6}|&(?P<size>\d+)')


def _parse_hf_sections(raw, default_size=10.0):
    """Split an Excel header/footer string into its L/C/R sections, stripping
    font/size/color codes but keeping &P/&N/&A tokens. Returns
    {"L"|"C"|"R": (lines, font_name, bold, size)} for non-empty sections."""
    chunks = re.split(r"&([LCR])", raw)
    section_text = {"L": "", "C": "", "R": ""}
    if chunks[0].strip():
        section_text["C"] = chunks[0]  # text before any code is centered
    for marker, text in zip(chunks[1::2], chunks[2::2]):
        section_text[marker] += text

    out = {}
    for key, text in section_text.items():
        if not text:
            continue
        found = {"font": None, "size": None}

        def _grab(m, found=found):
            if m.group("font") is not None:
                found["font"] = m.group("font")
            elif m.group("size") is not None:
                found["size"] = float(m.group("size"))
            return ""

        cleaned = _HF_CODE_RE.sub(_grab, text)
        if not cleaned.strip():
            continue
        name, _, style = (found["font"] or "Arial").partition(",")
        out[key] = (cleaned.split("\n"), name.strip() or "Arial",
                    "bold" in style.lower(), found["size"] or default_size)
    return out


_GRID_FONT_CACHE = {}


def _grid_pdf_font(name, bold):
    """Register (once) and return a reportlab font for the given family,
    preferring the real Windows TTF so metrics match Excel's output."""
    key = ((name or "Arial").lower(), bool(bold))
    if key in _GRID_FONT_CACHE:
        return _GRID_FONT_CACHE[key]

    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    base = key[0].replace(" ", "")
    fonts_dir = os.path.join(os.environ.get("WINDIR", r"C:\Windows"), "Fonts")
    candidates = [f"{base}bd.ttf", f"{base}b.ttf"] if bold else [f"{base}.ttf"]
    reg = None
    for cand in candidates:
        path = os.path.join(fonts_dir, cand)
        if os.path.exists(path):
            try:
                reg_name = f"XLGrid-{base}-B" if bold else f"XLGrid-{base}"
                pdfmetrics.registerFont(TTFont(reg_name, path))
                reg = reg_name
                break
            except Exception:
                pass
    if reg is None:
        reg = "Helvetica-Bold" if bold else "Helvetica"
    _GRID_FONT_CACHE[key] = reg
    return reg


def _render_grid_sheet_pdf(xlsx_path, sheet_name, out_pdf, page_size, progress):
    """Draw a simple grid sheet (title / blank / header / bordered data rows)
    to out_pdf with reportlab, replicating Excel's fit-to-page print layout."""
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.pdfbase import pdfmetrics

    s = _parse_grid_sheet_settings(xlsx_path, sheet_name)
    if s["has_merges"]:
        raise RuntimeError(f"{sheet_name}: merged cells are not direct-renderable")
    title_rows = s["title_rows"] or 3

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    try:
        ws = wb[sheet_name]
        probe = list(ws.iter_rows(min_row=1, max_row=title_rows + 1))

        def _cellfont(cell, fallback=("Arial", 10.0, False)):
            try:
                f = cell.font
                return (f.name or fallback[0], float(f.size or fallback[1]),
                        bool(f.bold))
            except Exception:
                return fallback

        title_font = _cellfont(probe[0][0], ("Arial", 10.0, True)) if probe else ("Arial", 10.0, True)
        header_font = (_cellfont(probe[title_rows - 1][0], ("Arial", 10.0, True))
                       if len(probe) >= title_rows else ("Arial", 10.0, True))
        data_font = ("Arial", 10.0, False)
        border_rgb = "C1C1C1"
        col_fmts = {}
        if len(probe) > title_rows:
            first_data = probe[title_rows]
            data_font = _cellfont(first_data[0])
            for ci, cell in enumerate(first_data):
                try:
                    side = cell.border.left
                    if side is not None and side.style and side.color and side.color.rgb:
                        border_rgb = str(side.color.rgb)[-6:]
                except Exception:
                    pass
                try:
                    if cell.number_format and cell.number_format != "General":
                        col_fmts[ci] = cell.number_format
                except Exception:
                    pass

        values = [tuple(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(values) <= title_rows:
        raise RuntimeError(f"{sheet_name}: no data rows to render")
    n_cols = max(len(r) for r in values)
    values = [r if len(r) == n_cols else r + (None,) * (n_cols - len(r))
              for r in values]

    # ── Geometry (points) ───────────────────────────────────────────────────
    page_w, page_h = page_size
    mg = s["margins"]
    ml, mr = mg["left"] * 72.0, mg["right"] * 72.0
    mt, mb = mg["top"] * 72.0, mg["bottom"] * 72.0
    printable_w = page_w - ml - mr
    printable_h = page_h - mt - mb

    def _chars_to_pt(chars):
        return int(chars * 7 + 5) * 0.75  # Excel char width -> px -> pt

    default_w = _chars_to_pt(s["default_col_width"])
    col_w = [_chars_to_pt(s["col_widths"][ci + 1]) if (ci + 1) in s["col_widths"]
             else default_w for ci in range(n_cols)]
    row_h = s["default_row_height"]

    h_font = _grid_pdf_font(header_font[0], True)
    t_fontname = _grid_pdf_font(title_font[0], title_font[2])
    d_fontname = _grid_pdf_font(data_font[0], data_font[2])
    d_size = data_font[1]
    h_size = header_font[1]
    line_h = h_size * 1.275

    string_width = pdfmetrics.stringWidth

    def _wrap(text, font, size, avail):
        if string_width(text, font, size) <= avail or " " not in text:
            return [text]
        lines, cur = [], None
        for word in text.split(" "):
            trial = word if cur is None else f"{cur} {word}"
            if cur is not None and string_width(trial, font, size) > avail:
                lines.append(cur)
                cur = word
            else:
                cur = trial
        lines.append(cur)
        return lines

    header_vals = values[title_rows - 1]
    header_lines = [
        _wrap(str(v), h_font, h_size, col_w[ci] - 3.0) if v is not None else []
        for ci, v in enumerate(header_vals)
    ]
    max_hdr_lines = max((len(l) for l in header_lines), default=1) or 1
    title_heights = [row_h] * title_rows
    title_heights[-1] = max(row_h, max_hdr_lines * line_h + 2.0)
    title_block_h = sum(title_heights)

    # ── Scale + pagination, mimicking Excel's fit-to-page (10% floor) ───────
    n_data = len(values) - title_rows
    content_w = sum(col_w)
    content_h = title_block_h + n_data * row_h
    if s["fit_to_page"]:
        scales = [1.0]
        if s["fit_w"]:
            scales.append(printable_w * s["fit_w"] / content_w)
        if s["fit_h"]:
            scales.append(printable_h * s["fit_h"] / content_h)
        scale = max(0.10, int(min(scales) * 100) / 100.0)
    else:
        scale = max(0.10, min(4.0, s["scale_pct"] / 100.0))

    limit_w = printable_w / scale
    col_pages, cur, cur_w = [], [], 0.0
    for ci in range(n_cols):
        if cur and cur_w + col_w[ci] > limit_w + 0.01:
            col_pages.append(cur)
            cur, cur_w = [], 0.0
        cur.append(ci)
        cur_w += col_w[ci]
    col_pages.append(cur)

    avail_h = printable_h / scale - title_block_h
    rows_per_page = max(1, int(avail_h // row_h))
    row_chunks = [(r, min(r + rows_per_page, len(values)))
                  for r in range(title_rows, len(values), rows_per_page)]
    total_pages = len(col_pages) * len(row_chunks)
    progress(f"PDF — drawing {sheet_name} ({n_data:,} rows, {total_pages} pages)...")

    # ── Value display cache (General + the few config number formats) ───────
    def _make_fmt(excel_fmt):
        if excel_fmt and excel_fmt.startswith("$"):
            return lambda v: f"${v:,.0f}" if isinstance(v, (int, float)) else str(v)
        if excel_fmt and "#,##0" in excel_fmt:
            return lambda v: f"{v:,.0f}" if isinstance(v, (int, float)) else str(v)
        return None

    col_fmt_fns = {ci: fn for ci, fn in
                   ((ci, _make_fmt(f)) for ci, f in col_fmts.items()) if fn}

    def _general(v):
        if isinstance(v, float):
            return str(int(v)) if v.is_integer() else f"{v:.10g}"
        return str(v)

    hdr_secs = _parse_hf_sections(s["odd_header"])
    ftr_secs = _parse_hf_sections(s["odd_footer"])
    mh_pt = mg["header"] * 72.0
    mf_pt = mg["footer"] * 72.0

    c = rl_canvas.Canvas(out_pdf, pagesize=page_size)
    c.setPageCompression(1)
    br = int(border_rgb[0:2], 16) / 255.0
    bg = int(border_rgb[2:4], 16) / 255.0
    bb = int(border_rgb[4:6], 16) / 255.0

    def _hf_expand(text, page_no):
        return (text.replace("&&", "\x00").replace("&P", str(page_no))
                .replace("&N", str(total_pages)).replace("&A", sheet_name)
                .replace("\x00", "&"))

    def _draw_hf(page_no):
        for secs, top in ((hdr_secs, True), (ftr_secs, False)):
            for key, (lines, fname, bold, fsize) in secs.items():
                font = _grid_pdf_font(fname, bold)
                c.setFont(font, fsize)
                leading = fsize * 1.2
                if top:
                    y = page_h - mh_pt - fsize
                else:
                    y = mf_pt + (len(lines) - 1) * leading + 2.0
                for ln in lines:
                    txt = _hf_expand(ln, page_no)
                    if txt:
                        if key == "L":
                            c.drawString(ml, y, txt)
                        elif key == "C":
                            c.drawCentredString(page_w / 2.0, y, txt)
                        else:
                            c.drawRightString(page_w - mr, y, txt)
                    y -= leading

    d_descent = pdfmetrics.getDescent(d_fontname) / 1000.0 * d_size
    d_base_off = -d_descent + 0.8       # baseline above cell bottom
    h_descent = pdfmetrics.getDescent(h_font) / 1000.0 * h_size
    h_base_off = -h_descent + 0.8
    width_cache = {}
    disp_cache = {}

    page_no = s["first_page_number"]
    for colset in col_pages:
        xs = [0.0]
        for ci in colset:
            xs.append(xs[-1] + col_w[ci])
        centers = [(xs[i] + xs[i + 1]) / 2.0 for i in range(len(colset))]

        for r0, r1 in row_chunks:
            _draw_hf(page_no)
            c.saveState()
            c.translate(ml, page_h - mt)
            c.scale(scale, scale)
            c.setStrokeColorRGB(br, bg, bb)
            c.setLineWidth(0.75)

            # Title block (repeated on every page, like print_title_rows).
            y = 0.0
            c.setFont(t_fontname, title_font[1])
            for ti in range(title_rows - 1):
                y -= title_heights[ti]
                v = values[ti][0] if values[ti] else None
                if v is not None:
                    c.drawString(xs[0] + 1.5, y + d_base_off, str(v))
            y -= title_heights[-1]
            c.setFont(h_font, h_size)
            for i, ci in enumerate(colset):
                lines = header_lines[ci]
                for li, ln in enumerate(reversed(lines)):
                    c.drawCentredString(centers[i], y + h_base_off + li * line_h, ln)
            grid_top = y

            # Borders: one grid of shared lines when the block is fully
            # populated (the normal case), else per-row runs of filled cells.
            n_rows = r1 - r0
            dense = all(values[r][ci] is not None
                        for r in range(r0, r1) for ci in colset)
            segs = []
            if dense:
                y_bot = grid_top - n_rows * row_h
                for i in range(n_rows + 1):
                    yy = grid_top - i * row_h
                    segs.append((xs[0], yy, xs[-1], yy))
                for xx in xs:
                    segs.append((xx, grid_top, xx, y_bot))
            else:
                for r in range(r0, r1):
                    y_top = grid_top - (r - r0) * row_h
                    y_bot = y_top - row_h
                    row = values[r]
                    run = None
                    for i, ci in enumerate(colset + [None]):
                        filled = ci is not None and row[ci] is not None
                        if filled and run is None:
                            run = i
                        elif not filled and run is not None:
                            segs.append((xs[run], y_top, xs[i], y_top))
                            segs.append((xs[run], y_bot, xs[i], y_bot))
                            for k in range(run, i + 1):
                                segs.append((xs[k], y_top, xs[k], y_bot))
                            run = None
            c.lines(segs)

            # Cell text.
            tx = c.beginText()
            tx.setFont(d_fontname, d_size)
            set_origin = tx.setTextOrigin
            text_out = tx.textOut
            for r in range(r0, r1):
                row = values[r]
                y_base = grid_top - (r - r0 + 1) * row_h + d_base_off
                for i, ci in enumerate(colset):
                    v = row[ci]
                    if v is None:
                        continue
                    txt = disp_cache.get(v)
                    if txt is None:
                        fmt = col_fmt_fns.get(ci)
                        txt = fmt(v) if fmt else _general(v)
                        disp_cache[v] = txt
                    w = width_cache.get(txt)
                    if w is None:
                        w = string_width(txt, d_fontname, d_size)
                        width_cache[txt] = w
                    set_origin(centers[i] - w * 0.5, y_base)
                    text_out(txt)
            c.drawText(tx)
            c.restoreState()
            c.showPage()
            page_no += 1

    c.save()
    return out_pdf


def _export_pdf_direct(xlsx_path, direct_sheets, has_small, pdf_path, progress):
    """Excel exports the small sheets in a worker thread while the huge grid
    sheets are drawn with reportlab; the parts are stitched with pikepdf.
    Raises on any failure (caller falls back to the Excel-only exports)."""
    import tempfile
    import threading
    import pikepdf
    import win32com.client
    import pythoncom

    tmpdir = tempfile.gettempdir()
    raw_excel = os.path.join(tmpdir, f"~rate_pages_small_{os.getpid()}.pdf")
    try:
        os.remove(raw_excel)
    except OSError:
        pass

    hide = set(direct_sheets) | {"Index"}
    paper = {}
    ready = threading.Event()
    errors = []

    def _excel_worker():
        pythoncom.CoInitialize()
        excel = None
        workbook = None
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            excel.EnableEvents = False
            workbook = excel.Workbooks.Open(xlsx_path, ReadOnly=True, UpdateLinks=0)
            try:
                paper["size"] = int(workbook.Sheets(1).PageSetup.PaperSize)
            except Exception:
                pass
            ready.set()
            first_visible = None
            for sheet in workbook.Sheets:
                if sheet.Name in hide:
                    sheet.Visible = 0  # xlSheetHidden — in-memory only (ReadOnly)
                elif first_visible is None:
                    first_visible = sheet
            first_visible.Activate()
            workbook.ExportAsFixedFormat(
                Type=_XL_TYPE_PDF,
                Filename=raw_excel,
                Quality=_XL_QUALITY_STD,
                IncludeDocProperties=False,
                IgnorePrintAreas=False,
                OpenAfterPublish=False,
            )
        except Exception as exc:
            errors.append(exc)
        finally:
            try:
                if workbook is not None:
                    workbook.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel is not None:
                    excel.Quit()
            except Exception:
                pass
            pythoncom.CoUninitialize()
            ready.set()

    thread = None
    if has_small:
        progress("PDF — Excel exporting small sheets in parallel...")
        thread = threading.Thread(target=_excel_worker, daemon=True)
        thread.start()
        ready.wait(timeout=60)  # only blocks until PaperSize is known

    page_size = _PAPER_SIZES_PT.get(paper.get("size"), _PAPER_SIZES_PT[1])

    parts = []
    try:
        for i, name in enumerate(direct_sheets):
            part = os.path.join(tmpdir, f"~rate_pages_grid{i}_{os.getpid()}.pdf")
            _render_grid_sheet_pdf(xlsx_path, name, part, page_size, progress)
            parts.append(part)

        if thread is not None:
            progress("PDF — waiting for Excel to finish the small sheets...")
            thread.join()
            if errors:
                raise errors[0]
            if not (os.path.exists(raw_excel) and os.path.getsize(raw_excel) > 0):
                raise RuntimeError("Excel produced no PDF for the small sheets")

        progress("PDF — merging and compressing...")
        merged = pikepdf.open(raw_excel) if has_small else pikepdf.new()
        with merged:
            if has_small:
                _strip_pdf_tags(merged)
            part_pdfs = [pikepdf.open(p) for p in parts]
            try:
                for pp in part_pdfs:
                    merged.pages.extend(pp.pages)
                merged.save(pdf_path,
                            compress_streams=True,
                            recompress_flate=True,
                            object_stream_mode=pikepdf.ObjectStreamMode.generate)
            finally:
                for pp in part_pdfs:
                    pp.close()
    finally:
        for p in parts + [raw_excel]:
            try:
                os.remove(p)
            except OSError:
                pass


def export_to_pdf(xlsx_path, pdf_path, progress_callback=None):
    """
    Convert an .xlsx file to PDF using Excel via COM.

    All sheets except 'Index' are included. The xlsx is opened read-only
    and never modified. Raises RuntimeError if the PDF is not produced.
    """
    import win32com.client
    import pythoncom
    import tempfile

    def _progress(msg):
        print(f"[pagebreaks] {msg}")
        if progress_callback:
            progress_callback(msg)

    xlsx_path = os.path.normpath(os.path.abspath(xlsx_path))
    pdf_path  = os.path.normpath(os.path.abspath(pdf_path))

    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"Excel file not found: {xlsx_path}")

    os.makedirs(os.path.dirname(pdf_path), exist_ok=True)
    if os.path.exists(pdf_path):
        try: os.remove(pdf_path)
        except PermissionError:
            raise RuntimeError(f"PDF is open in another program: {pdf_path}")

    # Excel writes the raw (tagged, ~5x larger) PDF to the local temp dir so
    # the oversized intermediate never lands in a OneDrive-synced folder;
    # only the compressed copy is written to pdf_path.
    raw_pdf = os.path.join(
        tempfile.gettempdir(),
        f"~rate_pages_export_{os.getpid()}.pdf",
    )
    if os.path.exists(raw_pdf):
        try: os.remove(raw_pdf)
        except OSError: pass

    # Opening the workbook from a OneDrive-synced folder stalls on hydration
    # and sync locks; give Excel a local temp copy to read instead.
    src_for_excel = xlsx_path
    local_xlsx = None
    onedrive = os.environ.get("OneDrive", "")
    if onedrive and xlsx_path.lower().startswith(os.path.normpath(onedrive).lower()):
        local_xlsx = os.path.join(
            tempfile.gettempdir(),
            f"~rate_pages_src_{os.getpid()}.xlsx",
        )
        t0 = time.perf_counter()
        shutil.copy2(xlsx_path, local_xlsx)
        src_for_excel = local_xlsx
        print(f"[pagebreaks] copied source out of OneDrive in {time.perf_counter()-t0:0.1f}s")

    _kill_excel_instances()

    # Fastest path for workbooks with a huge trailing grid sheet (BOP's
    # 82k-row TRDEF): draw it directly with reportlab while Excel exports
    # only the small sheets, then stitch. Anything going wrong falls through
    # to the parallel / single-instance Excel exports below.
    try:
        direct = _plan_direct_render(src_for_excel)
    except Exception as exc:
        print(f"[pagebreaks] direct-render planning skipped ({exc})")
        direct = None
    if direct:
        t_export = time.perf_counter()
        try:
            _export_pdf_direct(src_for_excel, direct["sheets"],
                               direct["has_small"], pdf_path, _progress)
            if not (os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0):
                raise RuntimeError("direct-render PDF missing or empty")
            print(f"[pagebreaks] direct-render export took "
                  f"{time.perf_counter()-t_export:0.1f}s")
            if local_xlsx is not None:
                try: os.remove(local_xlsx)
                except OSError: pass
            return pdf_path
        except Exception as exc:
            print(f"[pagebreaks] direct render failed ({exc}) — "
                  f"falling back to Excel export")
            try: os.remove(pdf_path)
            except OSError: pass

    # Workbooks with a huge sheet render ~2.6x faster split across several
    # Excel instances; anything going wrong here falls through to the normal
    # single-instance export below.
    try:
        plan = _plan_parallel_export(src_for_excel)
    except Exception as exc:
        print(f"[pagebreaks] parallel planning skipped ({exc})")
        plan = None
    if plan:
        t_export = time.perf_counter()
        try:
            _export_pdf_parallel(src_for_excel, plan, pdf_path, _progress)
            print(f"[pagebreaks] parallel export took "
                  f"{time.perf_counter()-t_export:0.1f}s")
            if local_xlsx is not None:
                try: os.remove(local_xlsx)
                except OSError: pass
            if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
                return pdf_path
            raise RuntimeError("merged PDF missing or empty")
        except Exception as exc:
            print(f"[pagebreaks] parallel export failed ({exc}) — "
                  f"falling back to single-instance export")
            try: os.remove(pdf_path)
            except OSError: pass

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    t_export = time.perf_counter()
    try:
        _progress("PDF — launching Excel...")
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False

        _progress("PDF — opening workbook...")
        workbook = excel.Workbooks.Open(src_for_excel, ReadOnly=True, UpdateLinks=0)

        # Hide Index from the PDF without touching the source file.
        for sheet in workbook.Sheets:
            if sheet.Name == "Index":
                sheet.Visible = 0  # xlSheetHidden
            else:
                # Make sure the first non-Index sheet is active so Excel
                # picks a sensible page-1.
                try: sheet.Activate()
                except Exception: pass
                break

        _progress("PDF — Excel is rendering pages (the long step)...")
        workbook.ExportAsFixedFormat(
            Type=_XL_TYPE_PDF,
            Filename=raw_pdf,
            Quality=_XL_QUALITY_STD,
            IncludeDocProperties=False,
            IgnorePrintAreas=False,
            OpenAfterPublish=False,
        )
    finally:
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        del workbook, excel
        pythoncom.CoUninitialize()
        if local_xlsx is not None:
            try: os.remove(local_xlsx)
            except OSError: pass
    print(f"[pagebreaks] Excel export took {time.perf_counter()-t_export:0.1f}s")

    # Verify Excel produced the raw PDF, then shrink it into pdf_path.
    for _ in range(10):
        if os.path.exists(raw_pdf) and os.path.getsize(raw_pdf) > 0:
            break
        time.sleep(0.2)
    else:
        raise RuntimeError(f"PDF was not created: {raw_pdf}")

    _progress("PDF — compressing...")
    t_comp = time.perf_counter()
    _compress_pdf(raw_pdf, pdf_path)
    print(f"[pagebreaks] compression took {time.perf_counter()-t_comp:0.1f}s")

    if os.path.exists(pdf_path) and os.path.getsize(pdf_path) > 0:
        return pdf_path
    raise RuntimeError(f"PDF was not created: {pdf_path}")


# Example usage:
# process_pagebreaks(r"C:\path\to\workbook.xlsx", "ignored.pdf")
# export_to_pdf(r"C:\path\to\workbook.xlsx", r"C:\path\to\workbook.pdf")
