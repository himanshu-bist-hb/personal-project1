"""
ExcelSettingsBOP.py
====================
Excel workbook / worksheet factory for BOP Rate Pages, mirroring the shape
of BA/ExcelSettingsBA.py.

WHY THIS FILE LOOKS DIFFERENT FROM BA'S VERSION
------------------------------------------------
BA has ~15 hand-written formatWorksheet* methods because its sheets have
genuinely different layouts (stacked tables, two-part Zone tables, the fixed
Rule 222 grid, ...). BOP's All Programs page has one layout used ~30 times
(title, optional merged sub-header row, one data table) with only the
column widths / number formats / sub-header text differing per table. So
instead of ~30 near-duplicate formatXxx() methods, there is ONE generic
format_table(ws, table_code) that looks up those per-table differences in
"BOP/BOP Input File.xlsx" (loaded via bop_config.load_bop_config()).

This is also what makes header/sub-header/title/table/footer sizing
consistent across every page: every sheet reads the exact same
Formatting Defaults values instead of each format method picking its own.
"""

from copy import copy

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows

from config.constants import COMPANY_NAMES
from .bop_config import load_bop_config


class Excel:
    """
    Factory class that creates an openpyxl Workbook and populates it with
    formatted worksheets, one per rate table.

    Typical usage (from AllProgramsPage.AllPrograms.buildAllProgramsPage):
        pages = ExcelSettingsBOP.Excel(state, "All Programs", nEff, rEff, companies)
        pages.generateWorksheet("SPR", title, df, False, True)
        ...
        wb = pages.getWB()
    """

    def __init__(self, state, programName, nEffective, rEffective, companyList) -> None:
        self.wb = Workbook()
        self.wb.active.title = "Index"

        self.state = state
        self.programName = programName
        self.nEffective = nEffective
        self.rEffective = rEffective
        self.companyList = list(companyList) if companyList else []

        cfg = load_bop_config()
        self._cfg = cfg
        fmt = cfg.formatting

        self.fontName = fmt.get("FontName", "Arial")
        self.fontSize = fmt.get("FontSize", 10)
        self.headerFontName = fmt.get("HeaderFontName", self.fontName)
        self.headerFontSize = fmt.get("HeaderFontSize", self.fontSize)
        self.footerFontName = fmt.get("FooterFontName", self.fontName)
        self.footerFontSize = fmt.get("FooterFontSize", self.fontSize)
        self.leftMargin = fmt.get("LeftMargin", 0.25)
        self.rightMargin = fmt.get("RightMargin", 0.25)
        self.topMargin = fmt.get("TopMargin", 1.25)
        self.bottomMargin = fmt.get("BottomMargin", 0.95)
        self.headerMargin = fmt.get("HeaderMargin", 0.5)
        self.footerMargin = fmt.get("FooterMargin", 0.25)
        self.borderColor = fmt.get("BorderColor", "C1C1C1")
        self.currencyFormat = fmt.get("CurrencyFormat", "$#,##0")
        self.noDecimalFormat = fmt.get("NoDecimalFormat", "#,##0")
        self.zipCodeFormat = fmt.get("ZipCodeFormat", "####0")
        self._defaultPrintTitleRows = fmt.get("PrintTitleRows", "1:3")

        self.font = Font(name=self.fontName, size=self.fontSize)
        self.fontBold = Font(name=self.fontName, size=self.fontSize, bold=True)
        self.fontItalic = Font(name=self.fontName, size=self.fontSize, italic=True)
        # Shared style objects — constructing a new Alignment per cell makes
        # openpyxl re-hash the style for every cell, which turns large sheets
        # (TRDEF has tens of thousands of rows) into a multi-minute format pass.
        self._centerAlign = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
        self.headerFont = f"{self.headerFontName},Bold"
        self.footerFont = f"{self.footerFontName},Bold"

        self._thinBorder = Border(
            left=Side(border_style="thin", color=self.borderColor),
            right=Side(border_style="thin", color=self.borderColor),
            top=Side(border_style="thin", color=self.borderColor),
            bottom=Side(border_style="thin", color=self.borderColor),
        )
        self._formatMap = {
            "Currency": self.currencyFormat,
            "NoDecimal": self.noDecimalFormat,
            "ZipCode": self.zipCodeFormat,
        }

    # ==========================================================================
    #  PRIVATE HELPERS
    # ==========================================================================

    def _apply_page_setup(self, ws) -> None:
        ws.page_setup.orientation = "portrait"
        ws.page_setup.blackAndWhite = False
        ws.page_setup.firstPageNumber = 1
        ws.page_setup.useFirstPageNumber = True
        ws.page_setup.fitToHeight = False
        ws.sheet_view.showGridLines = False
        ws.print_title_rows = self._defaultPrintTitleRows
        ws.page_margins.left = self.leftMargin
        ws.page_margins.right = self.rightMargin
        ws.page_margins.top = self.topMargin
        ws.page_margins.bottom = self.bottomMargin
        ws.page_margins.header = self.headerMargin
        ws.page_margins.footer = self.footerMargin

    def _token_sub(self, template: str) -> str:
        return template.format(
            state=self.state,
            state_abb=self.state,
            n_effective=self.nEffective,
            r_effective=self.rEffective,
            companies=self._footer_companies_text(),
        )

    def _footer_companies_text(self) -> str:
        names = [COMPANY_NAMES.get(c, c) for c in self.companyList if c != "CW"]
        return "\n".join(f"{n} " for n in names)

    def _apply_standard_header(self, ws) -> None:
        hf = self._cfg.header_footer
        ws.oddHeader.left.text = hf.get("HeaderLeftText", "")
        ws.oddHeader.left.size = self.headerFontSize
        ws.oddHeader.left.font = self.headerFont
        ws.oddHeader.center.text = self._token_sub(hf.get("HeaderCenterTemplate", ""))
        ws.oddHeader.center.size = self.headerFontSize
        ws.oddHeader.center.font = self.headerFont
        ws.oddHeader.right.text = self._token_sub(hf.get("HeaderRightTemplate", ""))
        ws.oddHeader.right.size = self.headerFontSize
        ws.oddHeader.right.font = self.headerFont

    def _apply_default_footer(self, ws) -> None:
        hf = self._cfg.header_footer
        ws.oddFooter.left.text = self._token_sub(hf.get("FooterLeftTemplate", ""))
        ws.oddFooter.left.size = self.footerFontSize
        ws.oddFooter.left.font = self.footerFont
        ws.oddFooter.center.text = self._token_sub(hf.get("FooterCenterTemplate", ""))
        ws.oddFooter.center.size = self.footerFontSize
        ws.oddFooter.center.font = self.footerFont
        ws.oddFooter.right.text = self._token_sub(hf.get("FooterRightTemplate", ""))
        ws.oddFooter.right.size = self.footerFontSize
        ws.oddFooter.right.font = self.footerFont

    def _apply_page_header_footer(self, ws) -> None:
        self._apply_page_setup(ws)
        self._apply_standard_header(ws)
        self._apply_default_footer(ws)

    def _write_df_block(self, ws, df, use_index: bool, use_header: bool, index_cell: str) -> None:
        for r in dataframe_to_rows(df, use_index, use_header):
            if use_index and len(list(r)) == 1:
                ws[index_cell] = list(r)[0]
                continue
            ws.append(r)

    def _resolve_col_range(self, range_str: str, max_col: int):
        if not range_str:
            return None
        start, _, end = range_str.partition(":")
        start_idx = column_index_from_string(start)
        end_idx = max_col if end.strip().upper() == "REST" else column_index_from_string(end)
        return start_idx, end_idx

    def _apply_sub_header(self, ws, table_code: str) -> int:
        """
        Inserts the merged sub-header row for tables that need one (per the
        "Sub Headers" tab), e.g. a row reading "Amount of Insurance" spanning
        several columns above the real column headers.

        Returns the inserted row number, or 0 if this table has no sub-header.
        """
        sh = self._cfg.sub_headers.get(table_code)
        if not sh:
            return 0

        row_idx = sh["insert_at_row"]
        ws.insert_rows(row_idx)
        max_col = ws.max_column

        for label_range, label_text in (
            (sh["label1_range"], sh["label1_text"]),
            (sh["label2_range"], sh["label2_text"]),
        ):
            rng = self._resolve_col_range(label_range, max_col)
            if not rng:
                continue
            start_idx, end_idx = rng
            if label_text:
                ws.cell(row=row_idx, column=start_idx).value = label_text
            if end_idx > start_idx:
                ws.merge_cells(start_row=row_idx, start_column=start_idx,
                                end_row=row_idx, end_column=end_idx)

        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = self._thinBorder
            cell.font = self.fontBold
            cell.alignment = self._centerAlign

        if sh["print_title_rows"]:
            ws.print_title_rows = sh["print_title_rows"]

        return row_idx

    # ==========================================================================
    #  WORKSHEET GENERATORS
    # ==========================================================================

    def generateWorksheet(self, table_code, title, df, useIndex, useHeader, layout_key=None):
        """
        Single-table sheet: A1 = title, A2 = blank, row 3+ = data (headers + rows).

        layout_key: which key to look up in the Table Layout / Number Formats /
        Sub Headers config tabs, if different from table_code (the actual tab
        name). Used e.g. by BCEG, whose tab is always named "BCEG" but whose
        column-width profile depends on whether the state has multiple
        building-code groups (BCEG_MULTI vs BCEG_SINGLE).
        """
        ws = self.wb.create_sheet(title=table_code)
        ws["A1"] = title
        ws["A2"] = ""
        self._write_df_block(ws, df, useIndex, useHeader, index_cell="A3")
        self.format_table(ws, layout_key or table_code)
        return ws

    def generateTerrWorksheet(self, table_code, title, df, useIndex, useHeader):
        """Territory-definitions sheet — same single-table layout as generateWorksheet."""
        return self.generateWorksheet(table_code, title, df, useIndex, useHeader)

    # ==========================================================================
    #  GENERIC FORMATTER — replaces the ~30 near-duplicate format*() methods
    # ==========================================================================

    def format_table(self, ws, table_code: str) -> None:
        self._apply_page_header_footer(ws)

        sub_header_row = self._apply_sub_header(ws, table_code)
        header_row = (sub_header_row + 1) if sub_header_row else 3

        max_row, max_col = ws.max_row, ws.max_column

        # bestFit is a column-level property — set it once per column, not once
        # per cell (per-cell it dominated the build time of large sheets).
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].bestFit = True

        # Title, spacer and column-header rows: a handful of cells, style
        # directly. (A sub-header row, if any, was styled by _apply_sub_header.)
        for col in range(1, max_col + 1):
            ws.cell(row=1, column=col).font = self.fontBold
            ws.cell(row=2, column=col).font = self.fontItalic
            header_cell = ws.cell(row=header_row, column=col)
            header_cell.font = self.fontBold
            header_cell.alignment = self._centerAlign

        self._format_data_rows(ws, table_code, header_row + 1, max_row, max_col)

        for col_start, col_end, width_px in self._cfg.table_layout.get(table_code, []):
            end = ws.max_column if col_end == "REST" else col_end
            for col in range(col_start, end + 1):
                ws.column_dimensions[get_column_letter(col)].width = width_px / 7.0

        for cell_ref, text in self._cfg.footnotes.get(table_code, []):
            c = ws[cell_ref]
            if c._style is not None:
                # Data cells share StyleArrays (see _format_data_rows); give
                # this cell a private copy before restyling it.
                c._style = copy(c._style)
            c.value = text
            c.font = self.font
            c.alignment = Alignment(horizontal="left", vertical="bottom", wrap_text=False)
            c.border = Border()

    def _format_data_rows(self, ws, table_code: str, first_row: int, max_row: int, max_col: int) -> None:
        """
        Font / alignment / border / number-format for the data region.

        openpyxl rebuilds and re-hashes a cell's style on every .font /
        .alignment / .border / .number_format assignment, which made
        grid-level sheets (TRDEF: tens of thousands of rows) take minutes.
        Instead, build the handful of distinct styles once — through a scratch
        cell, so the ids land in the same workbook style registries openpyxl
        would use — then stamp the shared StyleArray straight onto each cell.

        The stamped arrays are SHARED between cells, so restyling a data cell
        afterwards through the normal descriptors would corrupt every cell
        sharing its array — replace cell._style with a copy first (as the
        footnote loop in format_table does).
        """
        if first_row > max_row:
            return

        # Per-column number format from config; for overlapping ranges the
        # last config row wins, matching the sequential application this
        # replaces.
        col_fmt = {}
        for col_start, col_end, row_start, fmt_name in self._cfg.number_formats.get(table_code, []):
            end = max_col if col_end == "REST" else col_end
            excel_fmt = self._formatMap.get(fmt_name, fmt_name)
            for col in range(col_start, min(end, max_col) + 1):
                col_fmt[col] = (row_start, excel_fmt)

        # Distinct styles, each as a (no-border, with-border) pair. The
        # scratch cell is inside the data region, so the stamping loop below
        # overwrites its style like any other cell — no cleanup needed.
        scratch = ws.cell(row=first_row, column=1)
        scratch.font = self.font
        scratch.alignment = self._centerAlign
        border_id = self.wb._borders.add(self._thinBorder)

        def _pair(style_array):
            plain = copy(style_array)
            bordered = copy(style_array)
            bordered.borderId = border_id
            return plain, bordered

        base_pair = _pair(scratch._style)
        fmt_pairs = {}
        for _, excel_fmt in col_fmt.values():
            if excel_fmt not in fmt_pairs:
                scratch.number_format = excel_fmt
                fmt_pairs[excel_fmt] = _pair(scratch._style)

        col_styles = []
        for col in range(1, max_col + 1):
            row_start, excel_fmt = col_fmt.get(col, (None, None))
            col_styles.append((row_start, fmt_pairs.get(excel_fmt), base_pair))

        for row_cells in ws.iter_rows(min_row=first_row, max_row=max_row, max_col=max_col):
            row = row_cells[0].row
            for cell, (row_start, fmt_pair, base) in zip(row_cells, col_styles):
                plain, bordered = fmt_pair if (row_start is not None and row >= row_start) else base
                cell._style = bordered if cell.value is not None else plain

    # ==========================================================================
    #  INDEX + GETTERS
    # ==========================================================================

    def createIndex(self) -> None:
        ws_index = self.wb["Index"]
        ws_index["A1"] = f"{self.state} - INDEX"
        ws_index["A1"].font = self.fontBold

        sheet_names = self.wb.sheetnames
        for i in range(1, len(sheet_names)):
            name = sheet_names[i]
            row = i + 2
            cell = ws_index[f"A{row}"]
            target_title = self.wb[name]["A1"].value
            cell.value = target_title if target_title else name
            cell.hyperlink = f"#'{name}'!A1"
            cell.font = self.font
            cell.style = "Hyperlink"

        ws_index.column_dimensions["A"].width = 100

    def getWB(self):
        return self.wb

    def getFontName(self):
        return self.fontName

    def getFontSize(self):
        return self.fontSize
