# This module builds and formats the Service State Page workbook (pre-2.0).
#
# Same idea as ServicePage.py (BP-2.0) — see that file's module docstring —
# but transcribed from the root-level ServicePageCurrent.py, which predates
# the Territory Definitions workbook (State Territory Multiplier tables are
# built straight from the ratebook's own BP7_Peril_TerritorialFactor table,
# same TRBG/TRPP/TRLB codes + PROGRAM_TR layout key already used by Hab/Auto
# Service/Retail's pre-2.0 pages) and does not have the cat4/L-Products
# exclusions BP-2.0 added, nor the Repair Services (RSS), Pet Services
# (PSS/PSPL) or Mobile Pet and Veterinarian Services (MPVS) endorsements —
# those are BP-2.0-only additions with no pre-2.0 equivalent, same as
# Retail's RTS/PSS/PSPL. Property Damage Liability Deductible and Liability
# Size of Risk are also unpivoted here (single "liability1" peril, not broken
# out by peril like the BP-2.0 version).

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment

from . import ExcelSettingsBOP


class Service:
    # Tables buildBaseRates() needs present for a given company before it can
    # be built without a KeyError.
    _BASE_RATE_TABLES = (
        'BP7_Peril_Building_Base_Rates',
        'BP7_Peril_BPP_Base_Rates',
        'BP7_Peril_Liability_Base_Rates',
    )

    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        # Individual program pages (unlike All Programs) also show the
        # "AllPeril" row/column — see [[bop_auto_service_allperil_row_fix]].
        self.perils = list(perils) + ['allperil']
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective
        self.rEffective = rEffective

        self.serviceProgramCode = 70000

    # Builds a dataframe for the given table code
    # The hierarchy matches Business Auto: lower-level company (NACO/NAFF/NICOF)
    # first, then NGIC (state-level default), then CW as the country-wide
    # fallback. See [[bop-nesting-order]] — the root-level ServicePageCurrent.py
    # checked NGIC first, which is backwards.
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if 'NACO' in self.rateTables.keys():
            if tableCode in self.rateTables['NACO'].keys():
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys():
            if tableCode in self.rateTables['NAFF'].keys():
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys():
            if tableCode in self.rateTables['NICOF'].keys():
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if tableCode in self.rateTables['NGIC'].keys():
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0])

    # Builds the dry cleaners EXTRA factor table
    # Returns a dataframe
    def buildDryCleanersFactor(self):
        dryCleanersFactor = self.buildDataFrame("BP7_Dry_Cleaners_Extra_Factor")
        filteredDryCleanersFactor = dryCleanersFactor.query(f'Class_Code_Min == {self.serviceProgramCode}')
        data = {'TierLimit': ['BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits',
                    'Tier2', 'Tier2', 'Tier2', 'Tier2', 'Tier2',
                    'Tier3', 'Tier3', 'Tier3', 'Tier3', 'Tier3',
                    'Tier4', 'Tier4', 'Tier4', 'Tier4', 'Tier4'],
                'Coverage': ['Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets'],
                'Limits (Per Item / Per Occurrence)': ['$1,000/ALS', '$1,000/$10,000', '$1,000/$5,000', '$1,000/$5,000', '$1,000/$5,000',
                    '$2,000/ALS', '$2,000/$15,000', '$2,000/$6,000', '$2,000/$6,000', '$2,000/$6,000',
                    '$3,000/ALS', '$3,000/$20,000', '$3,000/$9,000', '$3,000/$9,000', '$3,000/$9,000',
                    '$5,000/ALS', '$5,000/$25,000', '$5,000/$10,000', '$5,000/$10,000', '$5,000/$10,000']}
        dryCleaners = pd.DataFrame(data)
        finalDryCleaners = pd.merge(dryCleaners, filteredDryCleanersFactor, how='left', on=['TierLimit'])
        return finalDryCleaners.rename(columns={'TierLimit': 'Tier', 'DryCleanerExtraFactor': 'Factor'}).filter(items=['Tier', 'Coverage', 'Limits (Per Item / Per Occurrence)', 'Factor']). \
                replace({'Tier': {'BaseLimits': 'Base', 'Tier2': 'Tier 2', 'Tier3': 'Tier 3', 'Tier4': 'Tier 4'}})

    # Builds the dry cleaners EXTRA earthquake factor table
    # Returns a dataframe
    def buildDryCleanersEQFactor(self):
        miscFactors = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        dryCleanersEQFactor = miscFactors.query(f'FactorName == "DryCleanersExtraEQ"')
        return dryCleanersEQFactor.filter(items=['Factor']).rename(columns={'Factor': 'Dry Cleaners EXTRA EQ Factor'})

    # Builds the funeral home EXTRA endorsement table
    # Returns a dataframe
    def buildFuneralHomeEndorsement(self):
        funeralHomeEndorsement = self.buildDataFrame("BP7_FuneralExtraBaseRate")
        funeralHomeEndorsement['Occurence'] = funeralHomeEndorsement['LiabilityLimitOccurence'].apply(lambda x: "${0:,.0f}".format(x))
        funeralHomeEndorsement['Aggregate'] = funeralHomeEndorsement['LiabilityLimtAggregate'].apply(lambda x: "${0:,.0f}".format(x))
        funeralHomeEndorsement['Occurence / Aggregate'] = funeralHomeEndorsement['Occurence'] + ' / ' + funeralHomeEndorsement['Aggregate']
        pivotedFuneralHome = funeralHomeEndorsement.pivot(index=['LiabilityLimitOccurence', 'Occurence / Aggregate'], columns='IncrementalDescedents', values='FuneralExtraBaseRate'). \
                rename(columns={100: 'First 100 decedents', 200: 'Next 200 decedents', 300: 'Next 300 decedents', 400: 'Next 400 decedents', 1000: 'Over 1,000 decedents'}). \
                reset_index(['LiabilityLimitOccurence', 'Occurence / Aggregate']).sort_values(by=['LiabilityLimitOccurence'])
        del pivotedFuneralHome['LiabilityLimitOccurence']
        return pivotedFuneralHome

    # Builds the funeral home EXTRA endorsement minimum premium
    # Returns a dataframe
    def buildFuneralHomeMinPrem(self):
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        funeralHomeMinPrem = miscMinMaxPrem.query(f'CoverageType == "BP7FuneralDirectorsProflLiab"')
        return funeralHomeMinPrem.filter(items=['Premium'])

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building",
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the territory multiplier table for the given coverage (either building, bpp, or liability)
    # Returns a dataframe
    def buildTerritoryMultiplier(self, coverage):
        territorialFactor = self.buildDataFrame("BP7_Peril_TerritorialFactor")
        filteredTerritorialFactor = territorialFactor.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'TerritoryCode': 'Territory'})
        if coverage.casefold() == 'building':
            return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BldgTerritoryFactor').reset_index('Territory')
        elif coverage.casefold() == 'bpp':
            return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BPPTerritoryFactor').reset_index('Territory')
        elif coverage.casefold() == 'liability':
            return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='LiabilityTerritoryFactor').reset_index('Territory')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction')
        elif coverage.casefold() == 'bpp':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction')

    # Builds the exclude theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'})
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range')
        elif coverage.casefold() == 'bpp':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range')

    # Builds the equipment breakdown base rate table
    # Returns a dataframe
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.serviceProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']). \
                replace({'P.D. Deductible Amount': {0: 'No Deductible'}}).filter(items=['P.D. Deductible Amount', 'Factor'])

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance': 'int32'})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_BLDG_BPP_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCodeMin == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"').rename(columns={'BldgBPPLimit_Min': 'Limit Min', 'BldgBPPLimit_Max': 'Limit Max', 'LiabilityFactor': 'Factor'})
        return filteredLiabilitySizeRisk.filter(items=['Limit Min', 'Limit Max', 'Factor']).fillna({'Limit Max': 'and over'})

    # Builds the general liability occupancy modifiers table
    # Returns a dataframe
    def buildGeneralOccupancyMod(self):
        generalOccupancyModifier = self.buildDataFrame("BP7_Peril_Occupant_Factor")
        filteredGeneralOccupancyMod = generalOccupancyModifier.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"'). \
                rename(columns={'OccupancyType': 'Occupancy', 'BLDGOccupantFactor': 'Building', 'BPPOccupantFactor': 'Business Personal Property'})
        return filteredGeneralOccupancyMod.replace({'Occupancy': {'Condominium': 'Condo Unit-owner', 'buildingOwnerLessorsrisk': "Lessor's Risk", 'buildingOwnerOccupant': 'Owner Occupant', 'tenant': 'Tenant'}}). \
                filter(items=['Occupancy', 'Building', 'Business Personal Property'])

    # Builds the directors and officers liability insurance table
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Service"').copy()
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 or More'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])

    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Service"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Service"').copy()
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index=filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])

    # Builds the barber or beauty shops professional liability table
    # Returns a dataframe
    def buildBarberProfLiab(self):
        barberProfLiab = self.buildDataFrame("BP7_ProfLiabarbersBeauticians_Rate")
        barberProfLiab['Occurence'] = barberProfLiab['LiabilityLimit'].apply(lambda x: "${0:,.0f}".format(x))
        barberProfLiab['Aggregate'] = barberProfLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        barberProfLiab['Occurence / Aggregate'] = barberProfLiab['Occurence'] + ' / ' + barberProfLiab['Aggregate']
        pivotedBarberProf = barberProfLiab.pivot(index=['LiabilityLimit', 'Occurence / Aggregate'], columns='ProfessionType', values='BaseRate').reset_index(['LiabilityLimit', 'Occurence / Aggregate']). \
                rename(columns={'Barber': 'Each Barber', 'Beautician': 'Each Beautician', 'Manicurist': 'Each Manicurist'}).sort_values(by=['LiabilityLimit'])
        del pivotedBarberProf['LiabilityLimit']
        return pivotedBarberProf

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.serviceProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Service premises'})

    # Builds the franchise upgrade endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.serviceProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(),
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(),
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Merges the "Number of Units" column of the D&O table into its 2 bands
    # ("Under 51" / "51 or More" — Service only has 2, unlike Hab's 5).
    # Handled here directly (rather than through the generic Sub Headers
    # config) because it merges DATA rows, not a header row.
    def _formatDirsOfficersLiabIns(self, ws):
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Groups the Dry Cleaners EXTRA Factor table's Tier (col A) and Factor
    # (col D) columns into 5-row blocks, one per tier — the data is always
    # exactly 20 rows (4 tiers x 5 coverage types) in this fixed order, so a
    # static merge is safe.
    def _formatDryCleanersFactor(self, ws):
        for start in (4, 9, 14, 19):
            end = start + 4
            ws.merge_cells(f'A{start}:A{end}')
            ws.merge_cells(f'D{start}:D{end}')
            ws[f'A{start}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws[f'D{start}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Sets up the Service Excel file and creates a separate worksheet for
    # each of the given dataframes. progress_callback (optional) is called
    # with a short message before each sheet is built.
    # Returns the Excel workbook
    def buildServicePage(self, progress_callback=None):
        companies = [c for c in self.rateTables.keys() if c != 'CW']

        Service = ExcelSettingsBOP.Excel(state=self.state, programName='Service', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        sheetSpecs = [
            ('DC', 'S Table 1.A.5.c. Dry Cleaners EXTRA Factor', self.buildDryCleanersFactor, False, True, None, self._formatDryCleanersFactor),
            ('DCEQ', 'S Table 1.A.5.d. Dry Cleaners EXTRA Earthquake Factor', self.buildDryCleanersEQFactor, False, True, None, None),
            ('FU', 'S Table 1.C.5.a. Funeral Home EXTRA Endorsement', self.buildFuneralHomeEndorsement, False, True, None, None),
            ('FUMP', 'S Table 1.C.5.c. Funeral Home EXTRA Endorsement Minimum Premium', self.buildFuneralHomeMinPrem, False, True, None, None),
        ]
        # A company can be present in rateTables (its ratebook was uploaded)
        # without having filed its own base-rate tables — a deviation
        # ratebook may only override a handful of tables. Check for the
        # specific tables buildBaseRates() needs, not just company
        # membership, or it KeyErrors on that company's missing table.
        for company, tab, label in (('NACO', 'BRNACO', 'NW Assurance'), ('NAFF', 'BRNAFF', 'NW Affinity'),
                                     ('NGIC', 'BRNGIC', 'NW General Insurance Company'), ('NICOF', 'BRNICOF', 'NICOF')):
            if company in self.rateTables and all(t in self.rateTables[company] for t in self._BASE_RATE_TABLES):
                sheetSpecs.append((tab, f'S Table 3.B.1. {label} State Base Rates', lambda c=company: self.buildBaseRates(c), False, True, 'AS_BR', None))

        sheetSpecs += [
            ('TRBG', 'S Table 3.C.1.a. State Territory Multiplier - Building', lambda: self.buildTerritoryMultiplier('Building'), False, True, 'PROGRAM_TR', None),
            ('TRPP', 'S Table 3.C.1.a. State Territory Multiplier - BPP', lambda: self.buildTerritoryMultiplier('BPP'), False, True, 'PROGRAM_TR', None),
            ('TRLB', 'S Table 3.C.1.a. State Territory Multiplier - Liability', lambda: self.buildTerritoryMultiplier('Liability'), False, True, 'PROGRAM_TR', None),
            ('CBG', 'S Table 3.C.2.c. Construction Factor - Building', lambda: self.buildConstructionType('Building'), False, True, None, None),
            ('CPP', 'S Table 3.C.2.c. Construction Factor - BPP', lambda: self.buildConstructionType('BPP'), False, True, None, None),
            ('ET', 'S Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions, False, True, None, None),
            ('YBBG', 'S Table 3.C.2.p. Year Built Modifier - Building', lambda: self.buildYearBuiltModifier('Building'), False, True, None, None),
            ('YBPP', 'S Table 3.C.2.p. Year Built Modifier - BPP', lambda: self.buildYearBuiltModifier('BPP'), False, True, None, None),
            ('EBB', 'S Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate, False, True, None, None),
            ('PDLD', 'S Table 3.C.4.b. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount, False, True, None, None),
            ('LL', 'S Table 3.C.4.d. Liability Limit Factor', self.buildLiabilityLimitFactor, False, True, None, None),
            ('LS', 'S Table 3.C.4.e. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk, False, True, 'LS_RETAIL_CURRENT', None),
            ('GLO', 'S Table 3.D.1.c. General Liability Occupancy Modifiers', self.buildGeneralOccupancyMod, False, True, None, None),
            ('DO', 'S Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns, False, True, None, self._formatDirsOfficersLiabIns),
            ('DONM', 'S Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief, False, True, None, None),
            ('ERP', 'S Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods, False, True, None, None),
            ('BB', 'S Table 4.B.1.e.(1). Barber or Beauty Shops Professional Liability', self.buildBarberProfLiab, False, True, None, None),
            ('PLUS', 'S Table 4.C. Service PLUS Endorsement', self.buildEndorsementCharge, False, True, None, None),
            ('FR', 'S Table 4.D. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement, False, True, None, None),
        ]

        total = len(sheetSpecs)
        for i, (tableCode, title, build, useIndex, useHeader, layoutKey, postFormat) in enumerate(sheetSpecs, start=1):
            if progress_callback:
                progress_callback(f"Building sheet {i}/{total}: {tableCode}...")
            print(f"  [{i}/{total}] Building sheet: {tableCode}")
            ws = Service.generateWorksheet(tableCode, title, build(), useIndex, useHeader, layout_key=layoutKey)
            if postFormat:
                postFormat(ws)

        if progress_callback:
            progress_callback("Building Index sheet...")
        print(f"  [{total}/{total}] Building sheet: Index")
        Service.createIndex()
        return Service.getWB()
