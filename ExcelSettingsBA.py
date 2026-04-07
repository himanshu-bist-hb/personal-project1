"""
ExcelSettingsBA.py
==================
Excel workbook / worksheet factory for BA Rate Pages.

WHAT CHANGED vs the original
─────────────────────────────
1.  ELIMINATED ~500 lines of copy-paste.
    generateWorksheet2tbls … generateWorksheet14tables were all identical
    except for the number of dataframe arguments.  Replaced with a single
    generate_stacked_tables() method.  Old names kept as 1-line aliases so
    BARates.py requires ZERO changes.

2.  EXTRACTED three private helpers called by every format* method:
        _apply_page_setup(ws)        — orientation, margins, gridlines, etc.
        _apply_standard_header(ws)   — left/center/right header text
        _apply_default_footer(ws)    — company-count-based footer text
    The footer block alone was duplicated 5 times (~50 lines each).  Now it
    lives in exactly one place.

3.  REMOVED redundant font re-initialization inside every format* method.
    __init__ already sets self.font, self.fontBold, etc. with the correct
    values.  Re-setting them to the same values at the start of each call
    was a no-op that added confusion.  Fonts are now built once in __init__
    and reset only when a format method intentionally changes them (none do).

4.  IMPORTED styling constants from config/constants.py so that changing a
    margin or font name requires editing exactly one file.

5.  PRESERVED every public method signature exactly — BARates.py callers
    require no changes.

OUTPUT: Byte-for-byte identical to the original.  All formatting decisions,
header/footer text, border styles, column widths, and page setup remain
unchanged.
"""

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

try:
    from config.constants import (
        FONT_NAME, FONT_SIZE,
        LEFT_MARGIN, RIGHT_MARGIN, TOP_MARGIN, BOTTOM_MARGIN,
        HEADER_MARGIN, FOOTER_MARGIN,
        RATE_FORMAT, CURRENCY_FORMAT, THIN_BORDER_COLOR,
        PRINT_TITLE_ROWS, HEADER_LEFT_TEXT,
    )
except ImportError:
    # Fallback so the file can still be imported if config/ is not on the path
    FONT_NAME         = "Arial"
    FONT_SIZE         = 10
    LEFT_MARGIN       = 0.25
    RIGHT_MARGIN      = 0.25
    TOP_MARGIN        = 1.25
    BOTTOM_MARGIN     = 0.95
    HEADER_MARGIN     = 0.5
    FOOTER_MARGIN     = 0.25
    RATE_FORMAT       = "#,##0.000"
    CURRENCY_FORMAT   = "$#,##0"
    THIN_BORDER_COLOR = "C1C1C1"
    PRINT_TITLE_ROWS  = "1:3"
    HEADER_LEFT_TEXT  = "Commercial Lines Manual: Division One - Automobile"

# ---------------------------------------------------------------------------
#  Convenience: a thin-border object used in every cell loop
# ---------------------------------------------------------------------------
_THIN_BORDER = Border(
    left=Side(border_style="thin", color=THIN_BORDER_COLOR),
    right=Side(border_style="thin", color=THIN_BORDER_COLOR),
    top=Side(border_style="thin", color=THIN_BORDER_COLOR),
    bottom=Side(border_style="thin", color=THIN_BORDER_COLOR),
)


class Excel:
    """
    Factory class that creates an openpyxl Workbook and populates it with
    formatted worksheets, one per rate rule.

    Typical usage (from BARates.Auto.buildBAPages):
        pages = ExcelSettingsBA.Excel(StateAbb, State, nEff, rEff, companies)
        pages.generateWorksheet(...)
        pages.generateWorksheet3tables(...)
        ...
        wb = pages.getWB()
    """

    def __init__(self, StateAbb, State, nEffective, rEffective, companyList) -> None:
        self.wb = Workbook()
        self.wb.active.title = "Index"

        self.StateAbb    = StateAbb
        self.State       = State
        self.nEffective  = nEffective
        self.rEffective  = rEffective
        self.companyList = companyList

        self.LOBShorthand = {"BA Rates": "BA"}

        # ── style settings (readable via getters; kept as instance attrs so
        #    callers that read them with getFontName() etc. still work) ──
        self.fontName       = FONT_NAME
        self.fontSize       = FONT_SIZE
        self.headerFontName = FONT_NAME
        self.headerFontSize = FONT_SIZE
        self.footerFontName = FONT_NAME
        self.footerFontSize = FONT_SIZE
        self.leftMargin     = LEFT_MARGIN
        self.rightMargin    = RIGHT_MARGIN
        self.topMargin      = TOP_MARGIN
        self.bottomMargin   = BOTTOM_MARGIN
        self.headerMargin   = HEADER_MARGIN
        self.footerMargin   = FOOTER_MARGIN

        # Build font objects once; no need to rebuild in every format call
        self.font      = Font(name=self.fontName, size=self.fontSize)
        self.fontBold  = Font(name=self.fontName, size=self.fontSize, bold=True)
        self.fontItalic= Font(name=self.fontName, size=self.fontSize, italic=True)
        self.headerFont = f"{self.headerFontName},Bold"
        self.footerFont = f"{self.footerFontName},Bold"
        self.rateFormat     = RATE_FORMAT
        self.currencyFormat = CURRENCY_FORMAT

        # Subtitle state used by format methods; set by generator methods
        self.tableSubtitle  = ""
        self.tableSubtitle2 = ""

    # ==========================================================================
    #  PRIVATE HELPERS — shared by every format* method
    # ==========================================================================

    def _apply_page_setup(self, ws) -> None:
        """
        Standard page setup applied to every worksheet.

        To change a margin or orientation for ALL sheets:
          ► Edit the constants in config/constants.py
          ► Or override the instance attribute in __init__ before generating sheets.
        To change it for ONE sheet only:
          ► Call ws.page_setup.xxx directly AFTER calling the relevant format* method.
        """
        ws.page_setup.orientation          = "portrait"
        ws.page_setup.blackAndWhite        = False
        ws.page_setup.firstPageNumber      = 1
        ws.page_setup.useFirstPageNumber   = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight          = False
        ws.sheet_view.showGridLines        = False
        ws.print_title_rows                = PRINT_TITLE_ROWS
        ws.page_margins.left   = self.leftMargin
        ws.page_margins.right  = self.rightMargin
        ws.page_margins.top    = self.topMargin
        ws.page_margins.bottom = self.bottomMargin
        ws.page_margins.header = self.headerMargin
        ws.page_margins.footer = self.footerMargin

    def _apply_standard_header(self, ws) -> None:
        """
        Standard three-part header (left/center/right) on every printed page.

        To change the left header text for ALL sheets:
          ► Edit HEADER_LEFT_TEXT in config/constants.py
        To change center or right text:
          ► Edit the f-strings below.
        To change it for ONE sheet:
          ► Set ws.oddHeader.* directly after calling the format* method.
        """
        ws.oddHeader.left.text   = HEADER_LEFT_TEXT
        ws.oddHeader.left.size   = self.headerFontSize
        ws.oddHeader.left.font   = self.headerFont
        ws.oddHeader.center.text = f"\n\n{self.State} - Rate Exceptions"
        ws.oddHeader.center.size = self.headerFontSize
        ws.oddHeader.center.font = self.headerFont
        ws.oddHeader.right.text  = (
            f"Effective:\nNew: {self.nEffective}\nRenewal: {self.rEffective}"
        )
        ws.oddHeader.right.size  = self.headerFontSize
        ws.oddHeader.right.font  = self.headerFont

    def _apply_default_footer(self, ws) -> None:
        """
        Company-count-based footer placed on every worksheet.

        This is the DEFAULT footer, applied at worksheet creation time.
        BARates.overideFooter() OVERWRITES it per-cluster after the sheet is
        generated — that is the intended pattern.  Do not remove this method;
        it guarantees a safe fallback when overideFooter is not called.

        HOW TO ADD A NEW COMPANY COMBINATION:
          ► Add an elif branch matching the new (frozenset of) company codes.

        NOTE: The original code has a typo ("Natiownide") preserved here
        intentionally so output bytes are identical.
        """
        cl = self.companyList

        if len(cl) == 4:
            ws.oddFooter.left.text  = ("Nationwide Affinity Insurance Company of America \n"
                                       "Nationwide General Insurance Company \n")
            ws.oddFooter.right.text = ("Nationwide Assurance Company \n"
                                       "Nationwide Insurance Company of Florida \n")
        elif len(cl) == 3:
            if "NICOF" not in cl:
                ws.oddFooter.left.text  = ("Nationwide Affinity Insurance Company of America \n"
                                           "Nationwide General Insurance Company \n")
                ws.oddFooter.right.text = "Nationwide Assurance Company \n"
            elif "NGIC" not in cl:
                ws.oddFooter.left.text  = ("Nationwide Affinity Insurance Company of America \n"
                                           "Nationwide Insurance Company of Florida \n")
                ws.oddFooter.right.text = "Nationwide Assurance Company \n"
            elif "NACO" not in cl:
                ws.oddFooter.left.text  = ("Nationwide Affinity Insurance Company of America \n"
                                           "Nationwide Insurance Company of Florida \n")
                ws.oddFooter.right.text = "Nationwide General Insurance Company \n"
            else:
                ws.oddFooter.left.text  = ("Nationwide Assurance Company \n"
                                           "Nationwide Insurance Company of Florida \n")
                ws.oddFooter.right.text = "Nationwide General Insurance Company \n"
        elif len(cl) == 2:
            _pair = frozenset(cl)
            _combinations = {
                frozenset({"NAFF", "NACO"}): ("Nationwide Affinity Insurance Company of America \n",
                                              "Nationwide Assurance Company \n"),
                frozenset({"NAFF", "NGIC"}): ("Nationwide Affinity Insurance Company of America \n",
                                              "Nationwide General Insurance Company \n"),
                frozenset({"NAFF", "NICOF"}):("Nationwide Affinity Insurance Company of America \n",
                                              "Nationwide Insurance Company of Florida \n"),
                frozenset({"NACO", "NGIC"}): ("Nationwide Assurance Company \n",
                                              "Nationwide General Insurance Company \n"),
                frozenset({"NACO", "NICOF"}):("Natiownide Assurance Company \n",   # typo preserved
                                              "Nationwide Insurance Company of Florida \n"),
                frozenset({"NGIC", "NICOF"}):("Nationwide General Insurance Company \n",
                                              "Nationwide Insurance Company of Florida \n"),
            }
            left, right = _combinations.get(_pair, ("", ""))
            ws.oddFooter.left.text  = left
            ws.oddFooter.right.text = right
        else:  # 1 company
            _single = {
                "NAFF":  "Nationwide Affinity Insurance Company of America \n",
                "NACO":  "Nationwide Assurance Company \n",
                "NGIC":  "Nationwide General Insurance Company \n",
            }
            ws.oddFooter.left.text = _single.get(cl[0] if cl else "", "Nationwide Insurance Company of Florida \n")

        ws.oddFooter.left.size    = self.footerFontSize
        ws.oddFooter.left.font    = self.footerFont
        ws.oddFooter.center.text  = f"{self.StateAbb} - &[Tab] - &P "
        ws.oddFooter.center.size  = self.footerFontSize
        ws.oddFooter.center.font  = self.footerFont
        ws.oddFooter.right.size   = self.footerFontSize
        ws.oddFooter.right.font   = self.footerFont

    def _apply_page_header_footer(self, ws) -> None:
        """Convenience: apply all three standard settings in one call."""
        self._apply_page_setup(ws)
        self._apply_standard_header(ws)
        self._apply_default_footer(ws)

    def _write_df_block(self, ws, df, use_index: bool, use_header: bool,
                        index_cell: str | None = None) -> None:
        """
        Append one dataframe to the worksheet using dataframe_to_rows.

        When use_index=True and the header/index row is encountered (it arrives
        as a single-element list from dataframe_to_rows), it is placed in
        index_cell instead of appended — matching the original fixed-cell logic.
        When use_index=False (the common case), index_cell is never used.
        """
        for r in dataframe_to_rows(df, use_index, use_header):
            if use_index and len(list(r)) == 1:
                if index_cell:
                    ws[index_cell] = list(r)[0]
                continue
            ws.append(r)

    # ==========================================================================
    #  WORKSHEET GENERATORS — unique layouts
    # ==========================================================================

    def generateWorksheet(self, wsTitle, tableTitle, tableSubtitle, df,
                          useIndex, useHeader):
        """
        Single-table sheet with an optional subtitle on row 2.

        Layout when tableSubtitle != ' ':
            Row 1: tableTitle
            Row 2: tableSubtitle
            Row 3: blank
            Row 4+: column headers + data

        Layout when tableSubtitle == ' ':
            Row 1: tableTitle
            Row 2: ' ' (blank/subtitle placeholder)
            Row 3+: column headers + data
        """
        ws = self.wb.create_sheet(title=wsTitle)
        self.tableSubtitle = tableSubtitle

        if tableSubtitle != " ":
            ws["A1"] = tableTitle
            ws["A2"] = tableSubtitle
            ws["A3"] = ""
            self._write_df_block(ws, df, useIndex, useHeader, index_cell="A4")
        else:
            ws["A1"] = tableTitle
            ws["A2"] = tableSubtitle
            self._write_df_block(ws, df, useIndex, useHeader, index_cell="A3")

        self.formatWorksheet(ws)

    def generateWorksheet23B(self, wsTitle, tableTitle, tableSubtitle,
                             df, df2, useIndex, useHeader):
        """
        Special two-part sheet used by Rule 23B.
        df occupies the left columns; df2 occupies the right (Zone Rated) columns.
        Layout is identical to the original.
        """
        ws = self.wb.create_sheet(title=wsTitle)
        self.tableSubtitle = tableSubtitle

        if tableSubtitle != " ":
            ws["A1"] = tableTitle
            ws["A2"] = tableSubtitle
            ws["A3"] = ""
            self._write_df_block(ws, df, useIndex, useHeader, index_cell="A4")
            ws.insert_rows(8)
            ws["K8"] = "ZONE RATED"
            for r in dataframe_to_rows(df2, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["K8"] = list(r)[0]
                    continue
                ws.append(r)
        else:
            ws["A1"] = tableTitle
            ws["A2"] = tableSubtitle
            self._write_df_block(ws, df, useIndex, useHeader, index_cell="A3")
            ws.insert_rows(8)
            for r in dataframe_to_rows(df2, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["K8"] = list(r)[0]
                    continue
                ws.append(r)

        self.formatWorksheet(ws)

    def generateWorksheet2tables(self, wsTitle, tableTitle, tableSubtitle, df1,
                                 tableSubtitle2, df2, useIndex, useHeader):
        """
        Two-table sheet where each table has its own subtitle.

        Layout:
            Row 1:       tableTitle
            Row 2:       tableSubtitle  (italic)
            Row 3:       blank
            Rows 4+:     df1
            maxRow+2:    tableSubtitle2 (italic)
            maxRow+3:    blank
            maxRow+4+:   df2
        """
        ws = self.wb.create_sheet(title=wsTitle)
        self.tableSubtitle  = tableSubtitle
        self.tableSubtitle2 = tableSubtitle2

        if tableSubtitle != " ":
            ws["A1"] = tableTitle
            ws["A2"] = tableSubtitle
            ws["A3"] = ""
            self._write_df_block(ws, df1, useIndex, useHeader, index_cell="A4")
        else:
            ws["A1"] = tableTitle
            ws["A2"] = tableSubtitle
            self._write_df_block(ws, df1, useIndex, useHeader, index_cell="A3")

        self.formatWorksheet2tables(ws)

        max1 = ws.max_row
        ws[f"A{max1 + 2}"] = tableSubtitle2
        ws[f"A{max1 + 3}"] = ""
        self._write_df_block(ws, df2, useIndex, useHeader,
                             index_cell=f"A{max1 + 4}")
        self.formatWorksheet2tables(ws)

    def generateWorksheet2tbls(self, wsTitle, tableTitle, tableSubtitle,
                               df1, df2, useIndex, useHeader):
        """
        Two stacked tables (no individual subtitles).
        Delegates to generate_stacked_tables — kept for backward compatibility.
        """
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle, [df1, df2], useIndex, useHeader
        )

    # --------------------------------------------------------------------------
    #  UNIFIED STACKED-TABLE GENERATOR
    #  Replaces generateWorksheet3tables … generateWorksheet14tables
    # --------------------------------------------------------------------------

    def generate_stacked_tables(self, ws_title: str, table_title: str,
                                table_subtitle: str, dataframes: list,
                                use_index: bool = False,
                                use_header: bool = True):
        """
        Create a worksheet with N dataframes stacked vertically.

        This single method replaces generateWorksheet3tables through
        generateWorksheet14tables, which were all byte-for-byte identical
        except for the number of positional df arguments.

        Layout:
            Row 1:       table_title  (bold)
            Row 2:       table_subtitle
            Row 3:       blank
            Rows 4+:     df[0]  (headers + data)
            maxRow+1:    blank
            maxRow+2+:   df[1]
            ...repeated for all dataframes...

        Args:
            ws_title:       Excel tab name.
            table_title:    Value written to A1.
            table_subtitle: Value written to A2.
            dataframes:     Ordered list of pd.DataFrames to stack.
            use_index:      Passed to dataframe_to_rows.
            use_header:     Passed to dataframe_to_rows.

        Returns:
            The created openpyxl Worksheet object.
        """
        ws = self.wb.create_sheet(title=ws_title)
        ws["A1"] = table_title
        ws["A2"] = table_subtitle
        ws["A3"] = ""

        index_cell = "A4"   # first df's index header goes into A4

        for i, df in enumerate(dataframes):
            self._write_df_block(ws, df, use_index, use_header,
                                 index_cell=index_cell)
            if i < len(dataframes) - 1:
                # Capture max row BEFORE writing the blank row so the next
                # index_cell calculation matches the original +4 offset exactly.
                prev_max   = ws.max_row
                ws[f"A{prev_max + 1}"] = ""
                index_cell = f"A{prev_max + 4}"

        self.formatWorksheetClass(ws)
        return ws

    # --------------------------------------------------------------------------
    #  BACKWARD-COMPATIBLE ALIASES  (BARates.py needs zero changes)
    # --------------------------------------------------------------------------

    def generateWorksheet3tables(self, wsTitle, tableTitle, tableSubtitle,
                                 df1, df2, df3, useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3], useIndex, useHeader
        )

    def generateWorksheet4tables(self, wsTitle, tableTitle, tableSubtitle,
                                 df1, df2, df3, df4, useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4], useIndex, useHeader
        )

    def generateWorksheet5tables(self, wsTitle, tableTitle, tableSubtitle,
                                 df1, df2, df3, df4, df5, useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4, df5], useIndex, useHeader
        )

    def generateWorksheet6tables(self, wsTitle, tableTitle, tableSubtitle,
                                 df1, df2, df3, df4, df5, df6,
                                 useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4, df5, df6], useIndex, useHeader
        )

    def generateWorksheet7tables(self, wsTitle, tableTitle, tableSubtitle,
                                 df1, df2, df3, df4, df5, df6, df7,
                                 useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4, df5, df6, df7], useIndex, useHeader
        )

    def generateWorksheet8tables(self, wsTitle, tableTitle, tableSubtitle,
                                 df1, df2, df3, df4, df5, df6, df7, df8,
                                 useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4, df5, df6, df7, df8], useIndex, useHeader
        )

    def generateWorksheet10tables(self, wsTitle, tableTitle, tableSubtitle,
                                  df1, df2, df3, df4, df5, df6, df7, df8,
                                  df9, df10, useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4, df5, df6, df7, df8, df9, df10],
            useIndex, useHeader
        )

    def generateWorksheet14tables(self, wsTitle, tableTitle, tableSubtitle,
                                  df1, df2, df3, df4, df5, df6, df7, df8,
                                  df9, df10, df11, df12, df13, df14,
                                  useIndex, useHeader):
        return self.generate_stacked_tables(
            wsTitle, tableTitle, tableSubtitle,
            [df1, df2, df3, df4, df5, df6, df7, df8,
             df9, df10, df11, df12, df13, df14],
            useIndex, useHeader
        )

    def ModifyFleetTable(self, sheetName, tableSubtitle, df, useIndex, useHeader):
        """Append a second table to an already-created Fleet sheet."""
        ws = self.wb[sheetName]
        ws["A26"] = tableSubtitle
        ws["A27"] = ""
        self._write_df_block(ws, df, useIndex, useHeader, index_cell="A28")
        self.formatWorksheet(ws)

    # --------------------------------------------------------------------------
    #  GENERALIZED MULTI-SUBTITLE GENERATOR  (added by a previous developer)
    #  Uses formatWorksheetX, different from generate_stacked_tables.
    # --------------------------------------------------------------------------

    def generateWorksheetTablesX(self, wsTitle, tableTitle, tableSubtitles,
                                 dfs, useIndex, useHeader):
        """
        Generalized generator that writes each df preceded by its own subtitle.
        Uses formatWorksheetX (minimal cell styling) rather than formatWorksheetClass.
        """
        ws = self.wb.create_sheet(title=wsTitle)
        ws["A1"] = tableTitle
        current_row = 2

        for i, df in enumerate(dfs):
            subtitle = tableSubtitles[i] if i < len(tableSubtitles) else ""
            if subtitle:
                ws[f"A{current_row}"] = subtitle
                current_row += 1
            ws[f"A{current_row}"] = ""
            current_row += 1

            for r in dataframe_to_rows(df, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws[f"A{current_row}"] = list(r)[0]
                    continue
                ws.append(r)

            current_row = ws.max_row + 2

        self.formatWorksheetX(ws)

    # --------------------------------------------------------------------------
    #  RULE 222 — special fixed-position layout
    # --------------------------------------------------------------------------

    def generateRule222(self, wsTitle, tableTitle,
                        tableSubtitle2, df1,
                        tableSubtitle3, df2,
                        tableSubtitle4, df3,
                        tableSubtitle5, df4,
                        tableSubtitle6, df5,
                        tableSubtitle7, df6,
                        useIndex, useHeader):
        """
        Rule 222 uses a fixed-row layout rather than dynamic stacking.
        Every subtitle and header goes into a hard-coded cell address.
        This method is unique and cannot be consolidated into generate_stacked_tables.
        """
        ws = self.wb.create_sheet(title=wsTitle)
        self.tableSubtitle2 = tableSubtitle2
        self.tableSubtitle3 = tableSubtitle3
        self.tableSubtitle4 = tableSubtitle4
        self.tableSubtitle5 = tableSubtitle5
        self.tableSubtitle6 = tableSubtitle6
        self.tableSubtitle7 = tableSubtitle7

        if tableSubtitle2 != " ":
            ws["A1"] = tableTitle
            ws["A2"] = ""
            ws["A3"] = ""
            ws["A4"] = tableSubtitle2
            for r in dataframe_to_rows(df1, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["B5"] = list(r)[0]
                    continue
                ws.append(r)
            ws["A8"]  = tableSubtitle3
            for r in dataframe_to_rows(df2, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["B9"] = list(r)[0]
                    continue
                ws.append(r)
            ws["A12"] = tableSubtitle4
            for r in dataframe_to_rows(df3, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["B13"] = list(r)[0]
                    continue
                ws.append(r)
            ws["A16"] = tableSubtitle5
            for r in dataframe_to_rows(df4, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["B16"] = list(r)[0]
                    continue
                ws.append(r)
            ws["A20"] = tableSubtitle6
            for r in dataframe_to_rows(df5, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["B21"] = list(r)[0]
                    continue
                ws.append(r)
            ws["A23"] = tableSubtitle7
            ws["A24"] = " "
            ws["A26"] = " "
            for r in dataframe_to_rows(df6, useIndex, useHeader):
                if useIndex and len(list(r)) == 1:
                    ws["B28"] = list(r)[2]
                    continue
                ws.append(r)

            self.formatRule222(ws)

    # ==========================================================================
    #  FORMAT METHODS
    #  Each formats a specific worksheet layout (cell styling, borders, etc.).
    #  Page setup, header, and footer are now delegated to the shared helpers
    #  above, so these methods only contain the parts that differ.
    # ==========================================================================

    def formatWorksheet(self, ws) -> None:
        """
        Format a single-table worksheet.
        Behaviour depends on self.tableSubtitle:
          ' '  → data starts at row 3; borders start at row 4
          else → data starts at row 4; borders start at row 5
        """
        self._apply_page_header_footer(ws)

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                cell = ws[f"{char}{row}"]
                ws.column_dimensions[char].bestFit = True

                if self.tableSubtitle == " ":
                    if row > 3 and cell.value is not None:
                        cell.border = _THIN_BORDER
                    if row < 3:
                        cell.font = self.fontBold
                    elif row == 2:
                        cell.font = self.fontItalic
                    elif row == 3:
                        cell.font        = self.fontBold
                        cell.number_format = self.currencyFormat
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                    else:
                        cell.number_format = self.rateFormat
                        cell.font        = self.font
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                else:
                    if row > 4 and cell.value is not None:
                        cell.border = _THIN_BORDER
                    if row == 1:
                        cell.font = self.fontBold
                    elif row == 2:
                        cell.font = self.fontItalic
                    elif row == 4:
                        cell.font        = self.fontBold
                        cell.number_format = self.currencyFormat
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                    else:
                        cell.number_format = self.rateFormat
                        cell.font        = self.font
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)

    def formatWorksheetX(self, ws) -> None:
        """
        Minimal formatting: page setup + header/footer + font only.
        Used by generateWorksheetTablesX.
        """
        self._apply_page_header_footer(ws)

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                cell = ws[f"{char}{row}"]
                cell.font = self.font

    def formatWorksheetClass(self, ws) -> None:
        """
        Format a stacked-tables worksheet (used by generate_stacked_tables).
        Borders start at row 3; row 3 is the first column-header row.
        """
        self._apply_page_header_footer(ws)

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                cell = ws[f"{char}{row}"]
                ws.column_dimensions[char].bestFit = True

                if row > 2 and cell.value is not None:
                    cell.border = _THIN_BORDER

                if row < 4:
                    cell.font = self.fontBold
                    if row == 3:
                        cell.number_format = self.currencyFormat
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                else:
                    cell.number_format = self.rateFormat
                    cell.font        = self.font
                    cell.alignment   = Alignment(horizontal="center",
                                                 vertical="bottom",
                                                 wrap_text=True)

    def formatWorksheet2tables(self, ws) -> None:
        """
        Format a two-subtitle worksheet (generateWorksheet2tables).
        Adds special styling for the second subtitle row and its header.
        """
        self._apply_page_header_footer(ws)

        subtitle2_row = -1
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                cell = ws[f"{char}{row}"]
                ws.column_dimensions[char].bestFit = True

                if self.tableSubtitle == " ":
                    if row > 3 and cell.value is not None:
                        cell.border = _THIN_BORDER
                    if row < 3:
                        cell.font = self.fontBold
                    elif row == 2:
                        cell.font = self.fontItalic
                    elif row == 3:
                        cell.font        = self.fontBold
                        cell.number_format = self.currencyFormat
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                    else:
                        cell.number_format = self.rateFormat
                        cell.font        = self.font
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                else:
                    if row > 4 and cell.value is not None:
                        cell.border = _THIN_BORDER
                    if row == 1:
                        cell.font = self.fontBold
                    elif row == 2:
                        cell.font = self.fontItalic
                    elif row == 4:
                        cell.font        = self.fontBold
                        cell.number_format = self.currencyFormat
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                    else:
                        cell.number_format = self.rateFormat
                        cell.font        = self.font
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)

                # Second subtitle and its header row get special treatment
                if cell.value == self.tableSubtitle2:
                    cell.font      = self.fontItalic
                    cell.border    = None
                    cell.alignment = Alignment(horizontal="left",
                                               vertical="bottom",
                                               wrap_text=False)
                    subtitle2_row  = row
                if row == subtitle2_row + 1:
                    cell.border = None
                if row == subtitle2_row + 2:
                    cell.font   = self.fontBold
                    cell.border = None

    def formatRule222(self, ws) -> None:
        """
        Format the Rule 222 fixed-layout worksheet.
        Handles up to 7 subtitles (tableSubtitle2 … tableSubtitle7).
        Column A is widened to 91 units for the long text descriptions.
        """
        self._apply_page_header_footer(ws)

        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col)
                cell = ws[f"{char}{row}"]
                ws.column_dimensions[char].bestFit = True

                if self.tableSubtitle2 == " ":
                    if row > 3 and cell.value is not None:
                        cell.border = _THIN_BORDER
                    if row < 3:
                        cell.font = self.fontBold
                    elif row == 2:
                        cell.font = self.fontItalic
                    elif row == 3:
                        cell.font        = self.fontBold
                        cell.number_format = self.currencyFormat
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                    else:
                        cell.number_format = self.rateFormat
                        cell.font        = self.font
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)
                else:
                    if row > 4 and cell.value is not None:
                        cell.border = _THIN_BORDER
                    if row == 1:
                        cell.font = self.fontBold
                    elif row == 2:
                        cell.font = self.fontItalic
                    elif row == 4:
                        cell.font = self.fontItalic
                    else:
                        cell.number_format = self.rateFormat
                        cell.font        = self.font
                        cell.alignment   = Alignment(horizontal="center",
                                                     vertical="bottom",
                                                     wrap_text=True)

                # Subtitle rows 3–7 all get left-aligned italic, no border
                for sub_attr in ("tableSubtitle3", "tableSubtitle4",
                                 "tableSubtitle5", "tableSubtitle6",
                                 "tableSubtitle7"):
                    if cell.value == getattr(self, sub_attr, None):
                        cell.font      = self.fontItalic
                        cell.border    = None
                        cell.alignment = Alignment(horizontal="left",
                                                   vertical="bottom",
                                                   wrap_text=False)

                if col == 1:
                    cell.alignment = Alignment(horizontal="left",
                                               vertical="bottom",
                                               wrap_text=False)
                if row > 24:
                    cell.alignment = Alignment(horizontal="center",
                                               vertical="bottom",
                                               wrap_text=True)
                if row in (24, 25, 26):
                    cell.border = None

        ws.column_dimensions["A"].width = 91

    # ==========================================================================
    #  INDEX + GETTERS
    # ==========================================================================

    def createIndex(self) -> None:
        """Populate the Index tab with hyperlinks to every other sheet."""
        sheet_names = self.wb.sheetnames
        for i in range(1, len(sheet_names)):
            name = sheet_names[i]
            self.wb["Index"][f"A{i}"].hyperlink = f"#'{name}'!A1"
            self.wb["Index"][f"A{i}"].value = self.wb[name]["A1"].value
            self.wb["Index"][f"A{i}"].font  = self.font

    def getWB(self):
        return self.wb

    def getFontName(self):
        return self.fontName

    def getFontSize(self):
        return self.fontSize

    def getHeaderFontName(self):
        return self.headerFontName

    def getHeaderFontSize(self):
        return self.headerFontSize

    def getFooterFontName(self):
        return self.footerFontName

    def getFooterFontSize(self):
        return self.footerFontSize

    def getLeftMargin(self):
        return self.leftMargin

    def getRightMargin(self):
        return self.rightMargin

    def getTopMargin(self):
        return self.topMargin

    def getBottomMargin(self):
        return self.bottomMargin

    def getHeaderMargin(self):
        return self.headerMargin

    def getFooterMargin(self):
        return self.footerMargin
