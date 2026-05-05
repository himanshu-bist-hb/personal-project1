"""
BA Rate Pages - Comprehensive Code Documentation Generator
===========================================================
Produces BA_Code_Documentation.xlsx with one row per rule covering:
  - Build function(s) + full source code
  - generateWorksheet* variant used
  - Custom format function(s) + full source code  (BARates.py post-processing)
  - Page-break handler + full source code          (BApagebreaks.py)
  - Ratebook sheet codes used for clustering
  - State restrictions / special notes

Run:  python generate_ba_docs.py
Output: BA_Code_Documentation.xlsx (overwrite-safe)
"""

import inspect
import os
import re
import textwrap

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

# ── Source file paths ──────────────────────────────────────────────────────
BA_RATES      = "BA/BARates.py"
BA_SETTINGS   = "BA/ExcelSettingsBA.py"
BA_BREAKS     = "BA/BApagebreaks.py"
OUTPUT_FILE   = "BA_Code_Documentation.xlsx"

# ── Colour palette ─────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark navy
SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # medium blue
ODD_FILL      = PatternFill("solid", fgColor="EBF3FB")   # very light blue
EVEN_FILL     = PatternFill("solid", fgColor="FFFFFF")   # white
ACCENT_FILL   = PatternFill("solid", fgColor="D6E4F0")   # soft blue accent

HEADER_FONT   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEAD_FONT  = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BODY_FONT     = Font(name="Consolas", size=9)
LABEL_FONT    = Font(name="Calibri", size=10, bold=True)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="B0C4DE")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ═══════════════════════════════════════════════════════════════════════════
#  FUNCTION EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════

def _read_file(path):
    if not os.path.exists(path):
        return ""
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def _extract_function(source_text, func_name):
    """
    Return the full source of `func_name` from source_text.
    Handles both top-level defs and class method defs (any indent).
    Returns a note string if the function isn't found.
    """
    lines = source_text.splitlines(keepends=True)
    pattern = re.compile(r"^(\s*)def\s+" + re.escape(func_name) + r"\s*\(")

    start = -1
    base_indent = -1
    for i, line in enumerate(lines):
        m = pattern.match(line)
        if m:
            start = i
            base_indent = len(m.group(1))
            break

    if start == -1:
        return f"# ⚠ Function '{func_name}' not found in source file."

    collected = [lines[start]]
    for line in lines[start + 1:]:
        stripped = line.lstrip()
        indent   = len(line) - len(stripped)
        # stop when we hit a non-empty line at same or lesser indent
        if stripped and indent <= base_indent:
            break
        collected.append(line)

    return "".join(collected).rstrip()


def get_func(file_path, func_name, _cache={}):
    """Cached function extractor."""
    key = (file_path, func_name)
    if key not in _cache:
        src = _read_file(file_path)
        _cache[key] = _extract_function(src, func_name)
    return _cache[key]


def get_funcs(file_path, func_names):
    """Return concatenated source of multiple functions, separated by a divider."""
    parts = []
    for name in func_names:
        code = get_func(file_path, name)
        parts.append(f"# {'─'*60}\n# FUNCTION: {name}\n# {'─'*60}\n{code}")
    return "\n\n".join(parts)


# ═══════════════════════════════════════════════════════════════════════════
#  RULE REGISTRY
#  Each dict: rule_no, rule_name, sheet_prefix, category,
#             build_funcs (list), generate_func,
#             format_funcs (list, BARates.py),
#             pagebreak_handler (str or None),
#             compare_sheets (str – ratebook keys),
#             state_notes
# ═══════════════════════════════════════════════════════════════════════════

RULES = [
    # ── VA-only special rule (before 208) ──────────────────────────────────
    dict(rule_no="VAPCD",   rule_name="Virginia Accident Prevention Course Discount",
         sheet_prefix="Rule VAPCD",  category="Virginia Special",
         build_funcs=["buildVAPCD"],
         generate_func="generateWorksheet",
         format_funcs=[],
         pagebreak_handler=None,
         compare_sheets="AccidentPreventionDiscountFactorVA_Ext",
         state_notes="VA only"),

    # ── Core numbered rules ────────────────────────────────────────────────
    dict(rule_no="208",   rule_name="Expense Constant",
         sheet_prefix="Rule 208",  category="Premium Development",
         build_funcs=["buildExpenseConstant"],
         generate_func="generateWorksheet",
         format_funcs=[],
         pagebreak_handler=None,
         compare_sheets="ExpenseConstant_Ext",
         state_notes="All states"),

    dict(rule_no="222 TTT BR", rule_name="Trucks Tractors Trailers Base Rates",
         sheet_prefix="Rule 222 TTT BR", category="Base Rates",
         build_funcs=["buildBaseRates"],
         generate_func="generateWorksheet",
         format_funcs=["formatBaseRates"],
         pagebreak_handler="_handle_rule_222ttt",
         compare_sheets="sheet_fetch('222 TTT')  [dynamic from BA Input File]",
         state_notes="buildBaseRates called with rate_type='TTT'"),

    dict(rule_no="222 B", rule_name="TTT Fleet Size Factors",
         sheet_prefix="Rule 222 B", category="Fleet Factors",
         build_funcs=["buildTTTLiabFleetFactors","buildTTTPhysDamFleetFactors","buildTTTOTCFleetFactors"],
         generate_func="generateWorksheet3tables",
         format_funcs=["format222B"],
         pagebreak_handler="_handle_rule_222b",
         compare_sheets="LiabilityFleetSizeFactors_Ext, CollisionFleetSizeFactor_Ext, ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext",
         state_notes="3 tables: Liability / Phys Dam / OTC fleet factors"),

    dict(rule_no="222 C", rule_name="TTT Special Provisions (Showroom/Dump/Farm)",
         sheet_prefix="Rule 222 C", category="Classification",
         build_funcs=["buildShowRoomFactors","buildShowRoom2Factors","buildShowRoom3Factors","buildShowRoom4Factors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format222C"],
         pagebreak_handler=None,
         compare_sheets="ShowroomLiabilityFactor, TrucksAndTruckTractorsCollisionHeavyDumpingFactor_Ext, TruckDumpingRelativity, TrucksAndTruckTractorsCollisionHeavyFarmingFactor_Ext",
         state_notes="4 subtables: Showroom Trailers, Heavy Dumping (non-zone), Heavy Dumping (zone), Farming"),

    dict(rule_no="222 E", rule_name="TTT Commercial Lay-up Credit",
         sheet_prefix="Rule 222 E", category="Premium Development",
         build_funcs=["buildLayupFactors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format222E"],
         pagebreak_handler=None,
         compare_sheets="AutoLayUpFactor_Ext",
         state_notes="Single table"),

    dict(rule_no="223 B.5", rule_name="TTT Primary Classification Factors",
         sheet_prefix="Rule 223 B.5", category="Classification",
         build_funcs=["buildPrimaryFactors","buildZonePrimaryFactors"],
         generate_func="generateWorksheet23B",
         format_funcs=["format23B"],
         pagebreak_handler="_handle_rule_223b5",
         compare_sheets="TrucksTractorsAndTrailersPrimaryFactors_Ext",
         state_notes="Special two-part layout: non-zone left, zone-rated right starting at col K"),

    dict(rule_no="223 C", rule_name="TTT Secondary Classification Factors",
         sheet_prefix="Rule 223 C", category="Classification",
         build_funcs=["buildTTTSecondaryFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format23C"],
         pagebreak_handler="_handle_rule_223c",
         compare_sheets="TrucksTractorsAndTrailersSecondaryFactorsLiabilityComprehensiveAndSCOL_Ext",
         state_notes="Single table"),

    dict(rule_no="225 C.2", rule_name="Zone-Rated Primary Classification Factors",
         sheet_prefix="Rule 225.C.2", category="Zone-Rated",
         build_funcs=["buildZonePrimaryFactors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format225C2"],
         pagebreak_handler=None,
         compare_sheets="TrucksTractorsAndTrailersPrimaryFactors_Ext",
         state_notes="Uses invisible unicode 200B in title for macro uniqueness"),

    dict(rule_no="225 C.3", rule_name="Zone-Rated Secondary Classification Factors",
         sheet_prefix="Rule 225.C.3", category="Zone-Rated",
         build_funcs=["buildZoneSecondaryFactors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format225C3"],
         pagebreak_handler="_handle_rule_225c3",
         compare_sheets="Secondary Classification Factors Zone Rated_Ext",
         state_notes="Uses invisible unicode 200B x2 in title for macro uniqueness"),

    dict(rule_no="225 Zone BR", rule_name="Zone-Rated Autos Base Rates",
         sheet_prefix="Rule 225 Zone BR", category="Base Rates",
         build_funcs=["buildZoneBaseRates"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatZoneRates"],
         pagebreak_handler="_handle_rule_225_zone",
         compare_sheets="ZoneRatedLiabilityBasePremium, ZoneRatedCollisionBasePremium, ZoneRatedOtherThanCollisionBasePremium",
         state_notes="Returns dict: 'Output Tables', 'Zones', 'Med Factor', 'Pip Factor'. Med/PIP appended if present."),

    dict(rule_no="225 D", rule_name="Zone-Rated Fleet Size & Lay-up",
         sheet_prefix="Rule 225.D", category="Zone-Rated",
         build_funcs=["buildFleetSizeRatingFactorsZone","buildLayupFactors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format225D"],
         pagebreak_handler=None,
         compare_sheets="AutoLayUpFactor_Ext",
         state_notes="2 tables: Fleet Size Rating + Lay-up. Uses 200B x3 in title."),

    dict(rule_no="231 C", rule_name="Private Passenger Classification Factors",
         sheet_prefix="Rule 231 C", category="Classification",
         build_funcs=["build231C"],
         generate_func="generateWorksheet",
         format_funcs=["format31C"],
         pagebreak_handler=None,
         compare_sheets="PrivatePassengerClassCode, PrivatePassengerTypesClassFactors_Ext",
         state_notes="Use and Operator Experience Factors"),

    dict(rule_no="232 PPT BR", rule_name="Private Passenger Types Base Rates",
         sheet_prefix="Rule 232 PPT BR", category="Base Rates",
         build_funcs=["buildBaseRates"],
         generate_func="generateWorksheet",
         format_funcs=["formatBaseRates"],
         pagebreak_handler="_handle_rule_232ppt",
         compare_sheets="sheet_fetch('232 PPT')  [dynamic from BA Input File]",
         state_notes="buildBaseRates called with rate_type='PPT'"),

    dict(rule_no="232 B", rule_name="PPT Fleet Size Factors",
         sheet_prefix="Rule 232 B", category="Fleet Factors",
         build_funcs=["buildPPTLiabFleetFactors","buildPPTPhysDamFleetFactors"],
         generate_func="generateWorksheet2tables",
         format_funcs=["format32B"],
         pagebreak_handler=None,
         compare_sheets="LiabilityFleetSizeFactors_Ext, CollisionFleetSizeFactor_Ext, ComprehensiveAndSpecifiedCausesOfLossFleetSizeFactor_Ext",
         state_notes="2 tables: Liability Fleet + Physical Damage Fleet"),

    dict(rule_no="233", rule_name="Private Passenger Farm Use Fleet Factors",
         sheet_prefix="Rule 233", category="Classification",
         build_funcs=["buildPPTFarmTypes"],
         generate_func="generateWorksheet",
         format_funcs=["format33"],
         pagebreak_handler=None,
         compare_sheets="PrivatePassengerFarmFactor2",
         state_notes="Farm use classification for PPT"),

    dict(rule_no="239 SB BR", rule_name="School & Church Buses Base Rates",
         sheet_prefix="Rule 239 SB BR", category="Base Rates",
         build_funcs=["buildBaseRates"],
         generate_func="generateWorksheet",
         format_funcs=["formatBaseRates"],
         pagebreak_handler="_handle_rule_239_general",
         compare_sheets="sheet_fetch('239 School Buses')  [dynamic]",
         state_notes="buildBaseRates called with rate_type='School Buses'"),

    dict(rule_no="239 OB BR", rule_name="All Other Buses Base Rates",
         sheet_prefix="Rule 239 OB BR", category="Base Rates",
         build_funcs=["buildBaseRates"],
         generate_func="generateWorksheet",
         format_funcs=["formatBaseRates"],
         pagebreak_handler="_handle_rule_239_general",
         compare_sheets="sheet_fetch('239 Other Buses')  [dynamic]",
         state_notes="buildBaseRates called with rate_type='Other Buses'"),

    dict(rule_no="239 VP BR", rule_name="Van Pools Base Rates",
         sheet_prefix="Rule 239 VP BR", category="Base Rates",
         build_funcs=["buildBaseRates"],
         generate_func="generateWorksheet",
         format_funcs=["formatBaseRates"],
         pagebreak_handler="_handle_rule_239_general",
         compare_sheets="sheet_fetch('239 Van Pools')  [dynamic]",
         state_notes="buildBaseRates called with rate_type='Van Pools'"),

    dict(rule_no="239 T-L BR", rule_name="Taxicabs & Limousines Base Rates",
         sheet_prefix="Rule 239 T-L BR", category="Base Rates",
         build_funcs=["buildBaseRates"],
         generate_func="generateWorksheet",
         format_funcs=["formatBaseRates"],
         pagebreak_handler="_handle_rule_239_general",
         compare_sheets="sheet_fetch('239 Taxis')  [dynamic]",
         state_notes="buildBaseRates called with rate_type='Taxis'"),

    dict(rule_no="239 C", rule_name="Public Auto Fleet & Mechanical Lift Factors",
         sheet_prefix="Rule 239 C", category="Premium Development",
         build_funcs=["buildPublicAutoLiabFleetSizeFactor","buildPublicAutoCollFleetSizeFactor",
                      "buildPublicAutoOTCFleetSizeFactor","buildMechanicalLiftFactor"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format239C"],
         pagebreak_handler="_handle_rule_239c",
         compare_sheets="PublicTypesFleetSizeFactorsForLiabilityAndMedicalPayments_Ext, PublicTransportationCollisionFleetSizeFactor_Ext, PublicTransportationOtherThanCollisionFleetSizeFactor_Ext, MechanicalLiftFactorOtherThanZoneRated",
         state_notes="4 tables: Liab Fleet, Coll Fleet, OTC Fleet, Mechanical Lift"),

    dict(rule_no="239 D", rule_name="Public Auto Commercial Lay-up Credit",
         sheet_prefix="Rule 239 D", category="Premium Development",
         build_funcs=["buildLayupFactors"],
         generate_func="generateWorksheet",
         format_funcs=[],
         pagebreak_handler=None,
         compare_sheets="AutoLayUpFactor_Ext",
         state_notes="Uses invisible 200B unicode in title"),

    dict(rule_no="240", rule_name="Public Auto Classifications",
         sheet_prefix="Rule 240", category="Classification",
         build_funcs=["build240","buildVanPrimaryClassFactor","buildVanSecondary"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format240"],
         pagebreak_handler="_handle_rule_240",
         compare_sheets="PublicTransportationLiabilityPrimaryFactor, PublicTransportationPhysicalDamagePrimaryFactor",
         state_notes="3 tables: Public Auto Primary (except Vans), Van Pools Primary, Secondary"),

    dict(rule_no="241", rule_name="Zone-Rated Public Auto - Mechanical Lift & Lay-up",
         sheet_prefix="Rule 241", category="Zone-Rated",
         build_funcs=["buildMechanicalLiftFactor","buildLayupFactors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format241"],
         pagebreak_handler=None,
         compare_sheets="AutoLayUpFactor_Ext, MechanicalLiftFactorOtherThanZoneRated",
         state_notes="VA: adds buildSpecifiecCausesofLossCoverageFactor as 2nd table"),

    dict(rule_no="243", rule_name="Public Auto - Seasonal/Migrant Farm Workers",
         sheet_prefix="Rule 243", category="Special Types",
         build_funcs=["buildFL_PublicsSeasonalMigrantFarm"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format243"],
         pagebreak_handler=None,
         compare_sheets="FarmLaborContractorPassengerHazardFactor",
         state_notes="FL and VA only"),

    dict(rule_no="255", rule_name="Garagekeepers Insurance",
         sheet_prefix="Rule 255", category="Garage",
         build_funcs=["buildGarageKeepers1","buildGarageKeepers2","buildGarageKeepers3"],
         generate_func="generateWorksheet3tables",
         format_funcs=["format55"],
         pagebreak_handler="_handle_rule_255",
         compare_sheets="GaragekeepersOtherThanCollisionPreliminaryBasePremium, GaragekeepersOtherThanCollisionDeductibleFactor, GaragekeepersCollisionPreliminaryBasePremium",
         state_notes="3 tables: OTC Base Premium, OTC Ded Factor, Collision Base Premium"),

    dict(rule_no="264", rule_name="Ambulance Services",
         sheet_prefix="Rule 264", category="Special Types",
         build_funcs=["buildAmbulanceFactors1","buildAmbulanceFactors2"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatAmbulance"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesAntiqueAutoFactor",
         state_notes="2 tables: Excl Volunteers / For Volunteers"),

    dict(rule_no="266", rule_name="Antique Autos",
         sheet_prefix="Rule 266", category="Special Types",
         build_funcs=["buildAntiqueAutoLiabFactors","buildAntiqueAutoPDRates"],
         generate_func="generateWorksheet2tables",
         format_funcs=["formatANTIQUEAUTOS"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesAntiqueAutoFactor, SpecialTypesAntiquePhysicalDamageRate_Ext",
         state_notes="2 tables: Liability & No-Fault / Physical Damage"),

    dict(rule_no="267", rule_name="Auto Body Manufacturers and Installers",
         sheet_prefix="Rule 267", category="Special Types",
         build_funcs=["buildFL_AutoBody"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format267"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesAutoBodyFactor",
         state_notes="FL and VA only"),

    dict(rule_no="268", rule_name="Driver Training Programs",
         sheet_prefix="Rule 268", category="Special Types",
         build_funcs=["buildRule68table1","buildRule68table3","buildRule68table4"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format68"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesDriverTrainingFactor, DriverTrainingLiabilityAndMedicalPaymentsCoveragesFactor, DriverTrainingMedicalPaymentsCoverageFactors",
         state_notes="3 tables: Owned Auto Factors, Instructors Factors, Medical Payments Factors"),

    dict(rule_no="269", rule_name="Drive-Away Contractors",
         sheet_prefix="Rule 269", category="Special Types",
         build_funcs=["buildFL_DriveAwayContractors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format269"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesDriveAwayContractorFactor",
         state_notes="FL and VA only"),

    dict(rule_no="271", rule_name="Fire Departments",
         sheet_prefix="Rule 271", category="Special Types",
         build_funcs=["buildFireDepartmentPPTFactors","buildFireDepartmentOtherThanPPTFactors","buildFireDepartmentBuyBackFactor"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatFireDepartments"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesFireDepartmentFactor, SpecialTypesEmergencyVehicleBuybackFactor",
         state_notes="3 tables: PPT Autos, Other Autos, Volunteer Buyback"),

    dict(rule_no="272", rule_name="Funeral Directors",
         sheet_prefix="Rule 272", category="Special Types",
         build_funcs=["buildFuneralDirectors1","buildFuneralDirectors2","buildFuneralDirectors3"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format72"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesFuneralDirectorFactor, FuneralDirectorMedicalPaymentsHiredNonOwnedFactor",
         state_notes="3 tables: Limousines, Hearses & Flower Cars, Medical Payments"),

    dict(rule_no="273", rule_name="Golf Carts and Low Speed Vehicles",
         sheet_prefix="Rule 273", category="Special Types",
         build_funcs=["buildSpecialGolfandLow"],
         generate_func="generateWorksheet",
         format_funcs=["format73"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesGolfCartsAndLowSpeedVehiclesFactor",
         state_notes="Single table"),

    dict(rule_no="274", rule_name="Law Enforcement Agencies",
         sheet_prefix="Rule 274", category="Special Types",
         build_funcs=["buildLawEnforcementPPTFactors","buildLawEnforcementOtherThanPPTFactors","buildLawEnforcementBuyBackFactor"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format274"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesLawEnforcementFactor, SpecialTypesEmergencyVehicleBuybackFactor",
         state_notes="3 tables: PPT, All Other Types, Fellow Volunteer Workers"),

    dict(rule_no="275", rule_name="Leasing or Rental Concerns",
         sheet_prefix="Rule 275", category="Special Types",
         build_funcs=["buildLeasingOrRentalConcernsFactors"],
         generate_func="generateWorksheet  [generateWorksheetTablesX for VA (7 tables)]",
         format_funcs=["formatLEASINGORRENTALCONCERNS"],
         pagebreak_handler="_handle_rule_275",
         compare_sheets="LeasingOrRentalConcernsContingentBasePremium_Ext",
         state_notes="VA: 7-table variant adds buildRule275VA_TTTFactors, VA_PPTFactors, VA_MotorcycleFactors, VA_SnowmobilesFactors, VA_ExceptMotorHomesFactors, VA_MotorHomesFactors"),

    dict(rule_no="276", rule_name="Mobile Homes",
         sheet_prefix="Rule 276", category="Special Types",
         build_funcs=["buildMobileHomeFactors","buildMobileHomeAdditionalFactor"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format276"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesMobileHomeFactor, MobileHomesAdditionalCoveragesFactor",
         state_notes="2 tables: Premium Computation + Limited OTC on Contents"),

    dict(rule_no="277", rule_name="Motorcycles",
         sheet_prefix="Rule 277", category="Special Types",
         build_funcs=["buildMotorcycles1","buildMotorcycles3","buildMotorcycles4","buildMotorcycles5","buildMotorcycles2"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format77"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesMotorcycleLiabilityFactor, SpecialTypesMotorcycleFactor",
         state_notes="5 tables (note: order is 1,3,4,5,2 matching manual order): Liab/Engine CC, Med Payments Factor, ACV PD Factor, SA PD Factor, Modify Rate Factor"),

    dict(rule_no="278", rule_name="Registration Plates Not Issued",
         sheet_prefix="Rule 278", category="Special Types",
         build_funcs=["buildRegistrationPlateFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format78"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesRegistrationPlatesFactor_Ext",
         state_notes="Single table"),

    dict(rule_no="279", rule_name="Repossessed Autos",
         sheet_prefix="Rule 279", category="Garage",
         build_funcs=["buildRule79table1","buildRule79table2","buildRule279table3"],
         generate_func="generateWorksheet3tables",
         format_funcs=["format79"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesRepossessedAutosLiabilityBasePremium, GarageDealersOtherThanCollisionRate, GarageDealersCollisionBlanketRate [MI: GarageDealersCollisionBlanketRateMI]",
         state_notes="3 tables: Liability Base Premium, OTC Rate, Blanket Collision Rate"),

    dict(rule_no="280", rule_name="Snowmobiles",
         sheet_prefix="Rule 280", category="Special Types",
         build_funcs=["buildSnowMobileFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format80"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesSnowmobileLiabilityBasePremium, SpecialTypesSnowmobileMedicalPaymentsRate, SpecialTypesSnowmobileOtherThanCollisionRate, SpecialTypesSnowmobileCollisionRate",
         state_notes="Single merged table with all coverages"),

    dict(rule_no="281", rule_name="Mobile and Farm Equipment",
         sheet_prefix="Rule 281", category="Special Types",
         build_funcs=["buildMobileandFarmPremiumFactor","buildMobileandFarmCostOfHireFactor","buildRentalBasisFactors"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format81"],
         pagebreak_handler=None,
         compare_sheets="SpecialTypesSpecialEquipmentFactor1",
         state_notes="3 tables: Premium Computation, Cost of Hire Basis, Rental Period Basis. VA: different subtitles."),

    dict(rule_no="283", rule_name="Autos Held for Sale by Non-Dealers",
         sheet_prefix="Rule 283", category="Garage",
         build_funcs=["buildRule83table1","buildRule83table2","buildRule83table3","buildRule83table4"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format283"],
         pagebreak_handler="_handle_rule_283",
         compare_sheets="GarageDealersOtherThanCollisionRate, GarageDealersCollisionBlanketRate [MI: GarageDealersCollisionBlanketRateMI], GarageDealersCollisionBlanketDeductibleFactor",
         state_notes="4 tables: Specified Causes, Limited Specified, Comprehensive, Blanket Collision. MI: adds 5th table via buildRule283_MI"),

    dict(rule_no="284", rule_name="All-Terrain & Utility Task Vehicles",
         sheet_prefix="Rule 284", category="Special Types",
         build_funcs=["(hardcoded DataFrame in buildBAPages)"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format284"],
         pagebreak_handler=None,
         compare_sheets="n/a – hardcoded; UM/UIM presence checked via '297 Map' input sheet",
         state_notes="Fixed rates table. UM/UIM rows conditionally dropped based on state's 297 Map"),

    dict(rule_no="288", rule_name="Drive Other Car",
         sheet_prefix="Rule 288", category="Premium Development",
         build_funcs=["buildDriveOtherFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format88"],
         pagebreak_handler=None,
         compare_sheets="DriveOtherCarLiabilityFactor, DriveOtherCarMedicalPaymentsFactor, DriveOtherCarOtherThanCollisionFactor, DriveOtherCarCollisionFactor, DriveOtherCarUninsuredMotoristFactor, DriveOtherCarUnderinsuredMotoristFactor",
         state_notes="Single merged table with all coverages"),

    dict(rule_no="289", rule_name="Non-Ownership Liability",
         sheet_prefix="Rule 289", category="Premium Development",
         build_funcs=["buildRule89tableB1b","buildRule89table1","buildRule89table3","buildRule89table2",
                      "buildRule89tableB2b2a","buildRule89table8","buildRule89table4","buildRule89table5",
                      "buildRule89table6","buildRule89table7"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format89"],
         pagebreak_handler="_handle_rule_289",
         compare_sheets="NonOwnedBasePremium, VolunteersAsInsuredsBasePremium, NonOwnedVolunteersBasePremium, EmployeesAsInsuredsFactor, NonOwnedVolunteersMinimumPremium, NonOwnedAutoFoodorGoodsDeliveryMinimumPremiumPerPolicy_Ext, NonOwnedAutoFoodorGoodsDeliveryIncreasedLimitFactor_Ext, NonOwnedAutoFoodorGoodsDeliveryExposureFactor_Ext, GarageServicesLiabilityEmployeesFactor, NonOwnedPartnershipFactor",
         state_notes="10 tables. Note: build order intentionally differs from subtitle order (CW manual had errors)"),

    dict(rule_no="290", rule_name="Hired Autos",
         sheet_prefix="Rule 290", category="Premium Development",
         build_funcs=["buildHiredAutoLiabFactors","buildHiredAutoPDFactors"],
         generate_func="generateWorksheet2tables",
         format_funcs=["formathiredauto"],
         pagebreak_handler=None,
         compare_sheets="LiabilityCostOfHireRate, ComprehensiveVehicleWithDriverCostOfHireRate_Ext, CollisionVehicleWithDriverCostOfHireRate_Ext, Hired Auto MinimumOtherThanCollisionPremium_Ext, HiredAutoMinimumCollisionPremium_Ext",
         state_notes="2 tables: Liability / Physical Damage"),

    dict(rule_no="292", rule_name="Medical Payments Increased Limit Factors",
         sheet_prefix="Rule 292", category="Medical / PIP",
         build_funcs=["buildMedPayments92A","buildMedPayments92B"],
         generate_func="generateWorksheet2tables",
         format_funcs=["format92"],
         pagebreak_handler=None,
         compare_sheets="ZoneRatedMedicalPaymentsTextFactor",
         state_notes="Wrapped in try/except with MEDPAYWarning. Skipped if no MedPay in state"),

    dict(rule_no="293", rule_name="No-Fault Coverages",
         sheet_prefix="Rule 293", category="Medical / PIP",
         build_funcs=["build293Table"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format293"],
         pagebreak_handler=None,
         compare_sheets="'293 Map' sheet in BA Input File drives dynamic table count; tables compared per-company for clustering",
         state_notes="Variable number of tables (1–N) from 293 Map. Custom clustering logic identical to 297 approach"),

    dict(rule_no="294", rule_name="Rental Reimbursement",
         sheet_prefix="Rule 294", category="Premium Development",
         build_funcs=["buildRentalFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format94"],
         pagebreak_handler=None,
         compare_sheets="RentalReimbursementFactor",
         state_notes="Single table"),

    dict(rule_no="295", rule_name="Audio Visual and Data Electronic Equipment",
         sheet_prefix="Rule 295", category="Premium Development",
         build_funcs=["buildAudioFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format95"],
         pagebreak_handler=None,
         compare_sheets="AudioVisualDataEquipmentBasePremium2",
         state_notes="Single table"),

    dict(rule_no="296", rule_name="Tapes Records and Discs Coverage",
         sheet_prefix="Rule 296", category="Premium Development",
         build_funcs=["buildTapeFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format96"],
         pagebreak_handler=None,
         compare_sheets="TapesRecordsAndDiscsBasePremium",
         state_notes="Single table"),

    dict(rule_no="297", rule_name="Uninsured Motorists Insurance",
         sheet_prefix="Rule 297", category="UM/UIM",
         build_funcs=["build297Table_unstacked","build297Table_stacked"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format297"],
         pagebreak_handler="_handle_rule_297",
         compare_sheets="'297 Map' sheet in BA Input File drives which tables and naming. Tables 1,2,3,4,5,6,10 = unstacked; 7,8,9 = stacked",
         state_notes="Skipped for NY and CA. MT: warning issued. Variable table count per state."),

    dict(rule_no="298", rule_name="Deductible Insurance",
         sheet_prefix="Rule 298", category="Premium Development",
         build_funcs=["buildRule298table1","buildRule298PPT","buildRule298TTT","buildRule298Zone",
                      "buildRule298AutoBlanket","buildRule298AutoGarageOTCFactors_1","buildRule298AutoGarageOTCFactors_2"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format98"],
         pagebreak_handler="_handle_rule_298",
         compare_sheets="LiabilityDeductibleFactor, ZoneRatedLiabilityDeductibleFactor, PhysicalDamageDeductibleFactors_Ext, ZoneRatedVehiclesDeductibleDiscountFactors_Ext, GarageDealersCollisionBlanketDeductibleFactor, GarageDealersOtherThanCollisionDeductibleFactor, GarageDealersOtherThanCollisionAllPerilsDeductibleFactor",
         state_notes="8 tables (one is always an empty DataFrame for text placeholder at subtitle 2)"),

    dict(rule_no="300", rule_name="Liability Increased Limit Factors",
         sheet_prefix="Rule 300", category="Premium Development",
         build_funcs=["buildILF"],
         generate_func="generateWorksheet",
         format_funcs=["format100"],
         pagebreak_handler=None,
         compare_sheets="IncreasedLimitFactorText",
         state_notes="Single table"),

    dict(rule_no="301.C", rule_name="Vehicle Age and Price Bracket (Non-VA)",
         sheet_prefix="Rule 301.C", category="Vehicle Value Factors",
         build_funcs=["buildZoneRatedTrailersVVFColl","buildZoneRatedNonTrailersVVFColl","buildPPTVVFColl",
                      "buildNonZoneRatedNonTrailersVVFColl","buildAllOtherVehiclesVVFColl",
                      "buildZoneRatedVehiclesVVFOTC","buildPPTVehiclesVVFOTC","buildAllOtherVehiclesVVFOTC",
                      "build101A1","build101A2","build101A3","build101A4","build101A5",
                      "build101B1","build101B2","build101B3"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format301"],
         pagebreak_handler="_handle_rule_301cd",
         compare_sheets="TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext (+ 10 more _Ext sheets)",
         state_notes="Non-VA only. 16 tables: 8 SAM + 8 ACV combinations across Zone/Non-Zone, Trailer/Non-Trailer, PPT/Other for Coll/OTC"),

    dict(rule_no="301.A (VA)", rule_name="Vehicle Value Factors ACV – VA Only",
         sheet_prefix="Rule 301.A", category="Vehicle Value Factors",
         build_funcs=["build101A1","build101A2","build101A3","build101A4","build101A5",
                      "build101B1","build101B2","build101B3"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format301VA"],
         pagebreak_handler="_handle_rule_301ab",
         compare_sheets="TrailerTypesCollisionVehicleValueFactorsStatedAmountVehicles_Ext",
         state_notes="VA only. 8 ACV tables with VARules['Rule 301.A'] subtitles"),

    dict(rule_no="301.B (VA)", rule_name="Vehicle Value Factors SAM – VA Only",
         sheet_prefix="Rule 301.B", category="Vehicle Value Factors",
         build_funcs=["buildZoneRatedTrailersVVFColl","buildZoneRatedNonTrailersVVFColl","buildPPTVVFColl",
                      "buildNonZoneRatedNonTrailersVVFColl","buildAllOtherVehiclesVVFColl",
                      "buildZoneRatedVehiclesVVFOTC","buildPPTVehiclesVVFOTC","buildAllOtherVehiclesVVFOTC"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format301VA"],
         pagebreak_handler="_handle_rule_301ab",
         compare_sheets="(same as 301.A VA)",
         state_notes="VA only. 8 SAM tables with VARules['Rule 301.B'] subtitles"),

    dict(rule_no="301.D.1", rule_name="Liability Original Cost New Factors",
         sheet_prefix="Rule 301.D.1", category="Vehicle Value Factors",
         build_funcs=["build301D1"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format301D1"],
         pagebreak_handler="_handle_rule_301cd",
         compare_sheets="LiabilityOriginalCostNewFactor_Ext",
         state_notes="Non-VA. VA uses same function but different sheet name (Rule 301.C.1)"),

    dict(rule_no="301.D.2", rule_name="Liability Vehicle Age Factors",
         sheet_prefix="Rule 301.D.2", category="Vehicle Value Factors",
         build_funcs=["build301D2","build301D3"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format301D2"],
         pagebreak_handler="_handle_rule_301cd",
         compare_sheets="LiabilityVehicleAgeFactorsStatedAmountVehicles_Ext, LiabilityVehicleAgeFactorsOriginalCostNewVehicle_Ext",
         state_notes="Non-VA: 2 tables (SAM, OCN). VA: same functions, order reversed, different sheet name (Rule 301.C.2)"),

    dict(rule_no="303", rule_name="Pollution Liability",
         sheet_prefix="Rule 303", category="Premium Development",
         build_funcs=["buildPollutionFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format103"],
         pagebreak_handler=None,
         compare_sheets="PollutionLiabilityRate_Ext, PollutionLiabilityMinimumPremium_Ext",
         state_notes="Single table"),

    dict(rule_no="305", rule_name="Limited Mexico Coverage",
         sheet_prefix="Rule 305", category="Premium Development",
         build_funcs=["buildBlank"],
         generate_func="generateWorksheet",
         format_funcs=["format113"],
         pagebreak_handler=None,
         compare_sheets="n/a – always uses default company",
         state_notes="Blank rate table; all states"),

    dict(rule_no="306 NAICS", rule_name="NAICS Industry Code Factors",
         sheet_prefix="Rule 306 NAICS", category="Classification",
         build_funcs=["buildTTTNAICSFactors"],
         generate_func="generateWorksheet",
         format_funcs=["formatNaics"],
         pagebreak_handler="_handle_rule_306",
         compare_sheets="NAICSFactors_Ext",
         state_notes="Large table of NAICS codes with factors"),

    dict(rule_no="307", rule_name="Fellow Employee Coverage",
         sheet_prefix="Rule 307", category="Premium Development",
         build_funcs=["buildFellowEmployeeFactors"],
         generate_func="generateWorksheet",
         format_funcs=["formatFellowEmployee"],
         pagebreak_handler=None,
         compare_sheets="FellowEmployeeBaseRate_v2_Ext, FellowEmployeeCoverageForDesignatedEmployeesPositionsBaseRate_v2_Ext",
         state_notes="Excluded from FL"),

    dict(rule_no="309", rule_name="Auto Loan/Lease Gap Coverage",
         sheet_prefix="Rule 309", category="Special Types",
         build_funcs=["buildFL_AutoLeaseGapCoverage"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format309"],
         pagebreak_handler=None,
         compare_sheets="AutoLoanLeaseGapCoverageFactor",
         state_notes="FL and VA only"),

    dict(rule_no="310", rule_name="Loss of Use Expenses - Optional Limits",
         sheet_prefix="Rule 310", category="Premium Development",
         build_funcs=["buildLossofUseFactors"],
         generate_func="generateWorksheet",
         format_funcs=["formatLossofUse"],
         pagebreak_handler=None,
         compare_sheets="OptionalLimitsLossofUseExpensesBasePremium_Ext",
         state_notes="All states"),

    dict(rule_no="313", rule_name="Silica or Silica-Related Dust Liability",
         sheet_prefix="Rule 313", category="Premium Development",
         build_funcs=["buildBlank"],
         generate_func="generateWorksheet",
         format_funcs=["format113"],
         pagebreak_handler=None,
         compare_sheets="n/a – always uses default company",
         state_notes="Blank rate table; all states"),

    dict(rule_no="315", rule_name="Business Interruption Coverage",
         sheet_prefix="Rule 315", category="Premium Development",
         build_funcs=["buildBusinessInterruptionFactors","buildExtendedBusinessFactors",
                      "buildWaitingBusinessFactors","buildInsuranceToExposureFactor"],
         generate_func="generateWorksheet4tables",
         format_funcs=["formatBusinessInterruption"],
         pagebreak_handler="_handle_rule_315",
         compare_sheets="BusinessInterruptionCoverageOtherThanCollisionBaseLossCost, BusinessInterruptionCoverageSpecifiedCausesOfLossBaseLossCost, BusinessInterruptionCoverageCollisionBaseLossCost, ExtendedBusinessIncomeAdditionalCoverageFactor, BusinessIncomeCoverageWaitingPeriodFactor, InsuranceToExposureFactor",
         state_notes="4 tables: Base Premium, Extended Business, Waiting Period, Insurance-to-Exposure"),

    dict(rule_no="317", rule_name="Towing and Labor",
         sheet_prefix="Rule 317", category="Premium Development",
         build_funcs=["buildTowingAndLabor"],
         generate_func="generateWorksheet",
         format_funcs=["formatTowingAndLabor"],
         pagebreak_handler=None,
         compare_sheets="TowingLaborRate",
         state_notes="Single table"),

    dict(rule_no="416", rule_name="Experience Rating",
         sheet_prefix="Rule 416", category="Rating Plans",
         build_funcs=["buildExperienceRating","buildExperienceRatingMinMax"],
         generate_func="generateWorksheet2tbls",
         format_funcs=["formatExperienceRating"],
         pagebreak_handler=None,
         compare_sheets="ExperienceRatingExpectedFrequencyPerPowerUnit_Ext, ExperienceRatingModifierRange_Ext, ExperienceRatingCredibility_Ext, ExperienceRatingBaseCredibility_Ext",
         state_notes="2 tables: Rating Variables + Min/Max"),

    dict(rule_no="417", rule_name="Schedule Rating",
         sheet_prefix="Rule 417", category="Rating Plans",
         build_funcs=["buildRule117"],
         generate_func="generateWorksheet",
         format_funcs=["format117"],
         pagebreak_handler=None,
         compare_sheets="ScheduleEligibility_Ext",
         state_notes="Single table"),

    dict(rule_no="425", rule_name="Waiver of Transfer of Rights of Recovery",
         sheet_prefix="Rule 425", category="Premium Development",
         build_funcs=["buildWaiver1Factors","buildWaiver2Factors"],
         generate_func="generateWorksheet2tables",
         format_funcs=["format125"],
         pagebreak_handler=None,
         compare_sheets="WaiverofSubrogationBlanket_Ext, Waiver_of_Subrogation_Ext",
         state_notes="2 tables: Blanket / Scheduled"),

    dict(rule_no="426", rule_name="Business Auto Protection Endorsements",
         sheet_prefix="Rule 426", category="Premium Development",
         build_funcs=["buildProtectionFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format126"],
         pagebreak_handler=None,
         compare_sheets="BusinessAutoProtectionFactor_Ext, MiscellaneousMinimumMaximumPremium_Ext",
         state_notes="Single table"),

    dict(rule_no="427", rule_name="Original Equipment Manufacturer Parts Coverage",
         sheet_prefix="Rule 427", category="Premium Development",
         build_funcs=["buildOriginalFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format127"],
         pagebreak_handler=None,
         compare_sheets="OriginalEquipmentManufacturerPartsCoverageFactor_Ext",
         state_notes="Single table"),

    dict(rule_no="450", rule_name="Driver Based Rating Plan",
         sheet_prefix="Rule 450", category="Rating Plans",
         build_funcs=["buildDBRFactors"],
         generate_func="generateWorksheet",
         format_funcs=["format150"],
         pagebreak_handler=None,
         compare_sheets="DriverBasedRatingLiabilityFactor_Ext, DriverBasedRatingCollisionFactor_Ext",
         state_notes="Single table"),

    dict(rule_no="451", rule_name="Transition Capping Program",
         sheet_prefix="Rule 451", category="Rating Plans",
         build_funcs=["buildRule451","buildRule451Renewals"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["format151"],
         pagebreak_handler=None,
         compare_sheets="Custom clustering from 297 map; same cluster_companies() helper",
         state_notes="Excluded for MT and CO. 2 tables: 451.B + 451.E Renewals"),

    dict(rule_no="452", rule_name="Lifetime Expense Allocation Factor (LEAF)",
         sheet_prefix="Rule 452", category="Rating Plans",
         build_funcs=["buildRetentionFactors"],
         generate_func="generateWorksheet",
         format_funcs=[],
         pagebreak_handler=None,
         compare_sheets="RetentionGradeFactor_Ext",
         state_notes="Conditional on '452 Leaf' input sheet. Show_leaf flag per state."),

    dict(rule_no="453", rule_name="Risk Tier Rating Plan",
         sheet_prefix="Rule 453", category="Rating Plans",
         build_funcs=["buildRiskTiering"],
         generate_func="generateWorksheet",
         format_funcs=[],
         pagebreak_handler=None,
         compare_sheets="TieringLiabilityFactor_Ext, TieringCollisionFactor_Ext, TieringOtherThanCollisionFactor_Ext",
         state_notes="Single table"),

    dict(rule_no="454", rule_name="School Bus Operations – Corporal Punishment",
         sheet_prefix="Rule 454", category="Special Types",
         build_funcs=["buildCorporalPunish"],
         generate_func="generateWorksheet",
         format_funcs=["formatSchoolBusOps"],
         pagebreak_handler=None,
         compare_sheets="CorporalPunishmentBaseRate_Ext, MiscellaneousMinimumMaximumPremium_Ext, BroadFormSchoolBusOperatorsCoverageFactor_Ext",
         state_notes="Single table"),

    dict(rule_no="DP-1", rule_name="Distribution Plan",
         sheet_prefix="Rule DP-1", category="Distribution",
         build_funcs=["buildDP1"],
         generate_func="generateWorksheet",
         format_funcs=["formatDp1"],
         pagebreak_handler=None,
         compare_sheets="NGIC ratebook only",
         state_notes="Specific state list only: AR AZ CT DC DE IA IL KY MD ME MI MN MO MS NC NE NH NM NV OR PA RI SD TN TX UT VT WI WV WY"),

    # ── State-specific lettered rules ─────────────────────────────────────
    dict(rule_no="A1 (CT)", rule_name="Liability of Municipalities – CT",
         sheet_prefix="Rule A1", category="State Special",
         build_funcs=["buildCT_A1"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA1"],
         pagebreak_handler=None,
         compare_sheets="LiabilityofMunicipalitiesPremiumCT_Ext",
         state_notes="CT only"),

    dict(rule_no="R1 (CT)", rule_name="Rate Order of Calculation – CT",
         sheet_prefix="Rule R1", category="State Special",
         build_funcs=["(hardcoded DataFrame from 'CT R1' sheet in BA Input File)"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatR1"],
         pagebreak_handler="_handle_rule_r1",
         compare_sheets="CT R1 sheet in BA Input File",
         state_notes="CT only. DataFrame read directly from BA Input File, no ratebook data"),

    dict(rule_no="A1 (KS)", rule_name="Accident Prevention Course Discount – KS",
         sheet_prefix="Rule A1", category="State Special",
         build_funcs=["buildKS_A1"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA1"],
         pagebreak_handler=None,
         compare_sheets="AccidentPreventionDiscountFactor",
         state_notes="KS only"),

    dict(rule_no="A2 (MI)", rule_name="Property Damage Liability Buyback – MI",
         sheet_prefix="Rule A2", category="State Special",
         build_funcs=["buildMI_A2"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA2"],
         pagebreak_handler=None,
         compare_sheets="PropertyDamageLiabilityCoverageBuybackVehiclePremium",
         state_notes="MI only"),

    dict(rule_no="A4 (MI)", rule_name="Michigan Catastrophic Claims Assn Surcharge – MI",
         sheet_prefix="Rule A4", category="State Special",
         build_funcs=["buildMI_A4A","buildMI_A4B"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA4"],
         pagebreak_handler=None,
         compare_sheets="MIMCCAFee_v2_Ext",
         state_notes="MI only. 2 tables: standard autos + historic vehicles"),

    dict(rule_no="A1 (ND)", rule_name="Rental Vehicle Coverage – ND",
         sheet_prefix="Rule A1", category="State Special",
         build_funcs=["buildND_A1"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA1"],
         pagebreak_handler=None,
         compare_sheets="RentalVehicleCoverageFactor_Ext",
         state_notes="ND only"),

    dict(rule_no="A3 (ND)", rule_name="Accident Prevention Course Discount – ND",
         sheet_prefix="Rule A3", category="State Special",
         build_funcs=["buildND_A3"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA3"],
         pagebreak_handler=None,
         compare_sheets="AccidentPreventionDiscountFactorND_Ext",
         state_notes="ND only"),

    dict(rule_no="A5 (NJ)", rule_name="Jitneys – NJ",
         sheet_prefix="Rule A5", category="State Special",
         build_funcs=["buildNJ_A5"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA5"],
         pagebreak_handler=None,
         compare_sheets="JitneysLiabilityBasePremium",
         state_notes="NJ only"),

    dict(rule_no="A2 (NV)", rule_name="Passive Restraint Discount – NV",
         sheet_prefix="Rule A2", category="State Special",
         build_funcs=["buildNV_A2"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA2"],
         pagebreak_handler=None,
         compare_sheets="PassiveRestraintDiscountFactor",
         state_notes="NV only"),

    dict(rule_no="A1 (RI)", rule_name="Accident Prevention Course Discount – RI",
         sheet_prefix="Rule A1", category="State Special",
         build_funcs=["buildRI_A1"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA1"],
         pagebreak_handler=None,
         compare_sheets="AccidentPreventionDiscountFactorRI",
         state_notes="RI only"),

    dict(rule_no="A2 (RI)", rule_name="Anti-Theft Device Discount – RI",
         sheet_prefix="Rule A2", category="State Special",
         build_funcs=["buildRI_A2"],
         generate_func="generateWorksheetTablesX",
         format_funcs=["formatA2"],
         pagebreak_handler=None,
         compare_sheets="AntiTheftDeviceDiscountFactor1",
         state_notes="RI only"),
]


# ═══════════════════════════════════════════════════════════════════════════
#  PAGEBREAK DESCRIPTIONS  (used in the "Page Break Logic" column)
# ═══════════════════════════════════════════════════════════════════════════

PAGEBREAK_DESCRIPTIONS = {
    "_handle_rule_222b":
        "disable_fit_to_page → add breaks after rows 25 and 49",
    "_handle_rule_222ttt":
        "fit_single_page → print_title_rows = '1:3'",
    "_handle_rule_223b5":
        "landscape orientation only",
    "_handle_rule_223c":
        "fit_single_page",
    "_handle_rule_225_zone":
        "print_area A1:M{max_row} → disable_fit_to_page → break every 51 rows (starting row 52)",
    "_handle_rule_225c3":
        "fit_single_page",
    "_handle_rule_232ppt":
        "fit_single_page → print_title_rows = '1:3'",
    "_handle_rule_239_general":
        "fit_single_page → print_title_rows = '1:3'",
    "_handle_rule_239c":
        "fit_single_page → top margin = 1.00\"",
    "_handle_rule_240":
        "fit_single_page → verticalCentered → print_title_rows '1:3' → print_area A1:M{max_row} → top margin 1.00\"",
    "_handle_rule_255":
        "print_area A1:H{max_row} → disable_fit_to_page → break after row 37",
    "_handle_rule_275":
        "conditional: if A10 = '275.B.1.(b). Short Term…' → print_title_rows '1:1' + fit_single_page",
    "_handle_rule_283":
        "print_area A1:P{max_row} → search column A for subtitle strings → break before matching rows → fit_width_only",
    "_handle_rule_289":
        "print_area A1:H{max_row} → disable_fit_to_page → break after row 37",
    "_handle_rule_297":
        "print_area A1:P{max_row} → disable_fit_to_page → count rows starting 'Single'/'Uninsured'; every 3rd occurrence triggers a break",
    "_handle_rule_298":
        "print_area A1:K{max_row} → disable_fit_to_page → count rows starting '298'; break at occurrences 4 and 8",
    "_handle_rule_301ab":
        "VA vehicle types → skip; otherwise fit_width_only + landscape + breaks every 45 rows + top margin 1.00\"",
    "_handle_rule_301cd":
        "Non-VA vehicle types → skip; otherwise fit_single_page + top margin 1.00\" (except FL)",
    "_handle_rule_306":
        "fit_width_only → print_title_rows '1:4'",
    "_handle_rule_315":
        "fit_width_only → break after row 23",
    "_handle_rule_r1":
        "print_area A1:M{max_row} → disable_fit_to_page → break at 3rd and 6th 'R1'-starting rows",
}


# ═══════════════════════════════════════════════════════════════════════════
#  BUILD ROWS
# ═══════════════════════════════════════════════════════════════════════════

def build_row(rule):
    # Build function(s) names
    build_names = rule["build_funcs"]
    build_names_str = "\n".join(build_names)

    # Build function code — skip hardcoded/placeholder entries
    real_builds = [n for n in build_names if not n.startswith("(")]
    build_code = get_funcs(BA_RATES, real_builds) if real_builds else "(see buildBAPages inline code)"

    # Generate function
    generate_func = rule["generate_func"]

    # ExcelSettingsBA internal format method (inferred from generate_func)
    _gen_map = {
        "generateWorksheet":         "formatWorksheet (ExcelSettingsBA.py)",
        "generateWorksheet23B":      "formatWorksheet (ExcelSettingsBA.py)",
        "generateWorksheet2tables":  "formatWorksheet2tables (ExcelSettingsBA.py)",
        "generateWorksheet2tbls":    "formatWorksheetClass (ExcelSettingsBA.py)",
        "generateWorksheet3tables":  "formatWorksheetClass (ExcelSettingsBA.py)",
        "generateWorksheet4tables":  "formatWorksheetClass (ExcelSettingsBA.py)",
        "generateWorksheetTablesX":  "formatWorksheetX (ExcelSettingsBA.py)",
        "generateRule222":           "formatRule222 (ExcelSettingsBA.py)",
    }
    internal_fmt = next(
        (v for k, v in _gen_map.items() if generate_func.startswith(k)),
        "formatWorksheet (ExcelSettingsBA.py)"
    )

    # Custom format function(s) code from BARates.py
    fmt_names = rule.get("format_funcs", [])
    fmt_names_str = "\n".join(fmt_names) if fmt_names else "(none – uses internal ExcelSettings format only)"
    fmt_code = get_funcs(BA_RATES, fmt_names) if fmt_names else "(no post-generation custom format)"

    # Page break handler
    pb_handler = rule.get("pagebreak_handler")
    if pb_handler:
        pb_desc  = PAGEBREAK_DESCRIPTIONS.get(pb_handler, "")
        pb_code  = get_func(BA_BREAKS, pb_handler)
    else:
        pb_desc  = "Default: fit_single_page (applied to all sheets as baseline in process_pagebreaks)"
        pb_code  = "(no rule-specific handler; baseline fit_single_page + print_title_rows='1:1' applied)"

    return {
        "Rule No."              : rule["rule_no"],
        "Rule Name"             : rule["rule_name"],
        "Category"              : rule["category"],
        "Sheet Name Prefix"     : rule["sheet_prefix"],
        "State Restrictions"    : rule["state_notes"],
        "Build Function(s)"     : build_names_str,
        "Build Code\n(DataFrame Creation)"   : build_code,
        "generateWorksheet Variant"          : generate_func,
        "Internal Format\n(ExcelSettingsBA)" : internal_fmt,
        "Custom Format Function(s)\n(BARates.py post-processing)" : fmt_names_str,
        "Custom Format Code"    : fmt_code,
        "Page Break Handler"    : pb_handler or "(none)",
        "Page Break Logic"      : pb_desc,
        "Page Break Code"       : pb_code,
        "Ratebook Sheets / compareCompanies Keys" : rule["compare_sheets"],
    }


# ═══════════════════════════════════════════════════════════════════════════
#  EXCEL WRITER
# ═══════════════════════════════════════════════════════════════════════════

COLUMNS = [
    "Rule No.",
    "Rule Name",
    "Category",
    "Sheet Name Prefix",
    "State Restrictions",
    "Build Function(s)",
    "Build Code\n(DataFrame Creation)",
    "generateWorksheet Variant",
    "Internal Format\n(ExcelSettingsBA)",
    "Custom Format Function(s)\n(BARates.py post-processing)",
    "Custom Format Code",
    "Page Break Handler",
    "Page Break Logic",
    "Page Break Code",
    "Ratebook Sheets / compareCompanies Keys",
]

COL_WIDTHS = {
    "Rule No."              : 14,
    "Rule Name"             : 38,
    "Category"              : 22,
    "Sheet Name Prefix"     : 22,
    "State Restrictions"    : 38,
    "Build Function(s)"     : 34,
    "Build Code\n(DataFrame Creation)"                           : 80,
    "generateWorksheet Variant"                                  : 34,
    "Internal Format\n(ExcelSettingsBA)"                         : 38,
    "Custom Format Function(s)\n(BARates.py post-processing)"    : 34,
    "Custom Format Code"                                         : 80,
    "Page Break Handler"    : 26,
    "Page Break Logic"      : 50,
    "Page Break Code"       : 70,
    "Ratebook Sheets / compareCompanies Keys"                    : 70,
}

# Columns that carry full source code → need taller rows and code font
CODE_COLS = {
    "Build Code\n(DataFrame Creation)",
    "Custom Format Code",
    "Page Break Code",
}

CATEGORY_COLOURS = {
    "Base Rates"           : "FFF2CC",
    "Fleet Factors"        : "E2EFDA",
    "Classification"       : "FCE4D6",
    "Zone-Rated"           : "DDEBF7",
    "Medical / PIP"        : "F4CCCC",
    "UM/UIM"               : "EAD1DC",
    "Garage"               : "D9D2E9",
    "Special Types"        : "CFE2F3",
    "Vehicle Value Factors": "D9EAD3",
    "Premium Development"  : "FFF2CC",
    "Rating Plans"         : "FCE5CD",
    "Distribution"         : "F3F3F3",
    "State Special"        : "F4CCCC",
    "Virginia Special"     : "EAD1DC",
}


def write_excel(rows):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    sum_headers = ["#", "Rule No.", "Rule Name", "Category",
                   "Build Function(s)", "generateWorksheet Variant",
                   "Custom Format Function(s)", "Page Break Handler",
                   "State Restrictions"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BOX

    ws_sum.row_dimensions[1].height = 28
    ws_sum.freeze_panes = "A2"

    for ri, row in enumerate(rows, 1):
        fill = PatternFill("solid",
                           fgColor=CATEGORY_COLOURS.get(row["Category"], "FFFFFF"))
        vals = [
            ri,
            row["Rule No."],
            row["Rule Name"],
            row["Category"],
            row["Build Function(s)"],
            row["generateWorksheet Variant"],
            row["Custom Format Function(s)\n(BARates.py post-processing)"],
            row["Page Break Handler"],
            row["State Restrictions"],
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws_sum.cell(row=ri+1, column=ci, value=v)
            cell.fill  = fill
            cell.font  = Font(name="Calibri", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BOX

    # auto-width summary
    SUM_WIDTHS = [5, 14, 40, 20, 34, 30, 34, 26, 40]
    for ci, w in enumerate(SUM_WIDTHS, 1):
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Full Details ────────────────────────────────────────────────
    ws = wb.create_sheet("Full Details")

    # Header row (row 1)
    for ci, col in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border    = BOX

    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "A2"

    # Data rows
    for ri, row in enumerate(rows, 2):
        cat   = row["Category"]
        color = CATEGORY_COLOURS.get(cat, "FFFFFF")
        base_fill = PatternFill("solid", fgColor=color)

        # Compute row height based on longest code cell
        max_lines = 1
        for col in COLUMNS:
            val = str(row.get(col, ""))
            n = val.count("\n") + 1
            if n > max_lines:
                max_lines = n
        row_height = max(18, min(max_lines * 13, 409))
        ws.row_dimensions[ri].height = row_height

        for ci, col in enumerate(COLUMNS, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = BOX
            cell.alignment = WRAP_ALIGN

            if col in CODE_COLS:
                cell.font = Font(name="Consolas", size=9)
                cell.fill = PatternFill("solid", fgColor="F8F8F8")
            else:
                cell.font = Font(name="Calibri", size=10,
                                 bold=(col in ("Rule No.", "Rule Name")))
                cell.fill = base_fill

    # Column widths
    for ci, col in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 20)

    # ── Sheet 3: Page Break Reference ───────────────────────────────────────
    ws_pb = wb.create_sheet("Page Break Reference")

    pb_headers = ["Handler Function", "Sheet Prefix Trigger",
                  "Logic Summary", "Full Code"]
    for ci, h in enumerate(pb_headers, 1):
        cell = ws_pb.cell(row=1, column=ci, value=h)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BOX
    ws_pb.row_dimensions[1].height = 28
    ws_pb.freeze_panes = "A2"

    # All known handlers with sheet prefixes from SHEET_RULES
    pb_registry = [
        ("_handle_index",           "Index"),
        ("_handle_rule_222b",        "Rule 222 B"),
        ("_handle_rule_222ttt",      "Rule 222 TTT"),
        ("_handle_rule_223b5",       "Rule 223 B.5"),
        ("_handle_rule_223c",        "Rule 223 C"),
        ("_handle_rule_225_zone",    "Rule 225 Zone"),
        ("_handle_rule_225c3",       "Rule 225.C.3"),
        ("_handle_rule_232ppt",      "Rule 232 PPT"),
        ("_handle_rule_239c",        "Rule 239 C  [specific – before generic]"),
        ("_handle_rule_239_general", "Rule 239    [generic – catches all 239 prefixes]"),
        ("_handle_rule_240",         "Rule 240 "),
        ("_handle_rule_255",         "Rule 255"),
        ("_handle_rule_275",         "Rule 275"),
        ("_handle_rule_283",         "Rule 283"),
        ("_handle_rule_289",         "Rule 289"),
        ("_handle_rule_297",         "Rule 297"),
        ("_handle_rule_298",         "Rule 298"),
        ("_handle_rule_301cd",       "Rule 301.C / Rule 301.D  [specific, checked first]"),
        ("_handle_rule_301ab",       "Rule 301.A / Rule 301.B"),
        ("_handle_rule_306",         "Rule 306"),
        ("_handle_rule_315",         "Rule 315"),
        ("_handle_rule_r1",          "Rule R1"),
    ]

    helper_names = ["fit_single_page","fit_width_only","disable_fit_to_page","add_break_after"]

    for ri, (fn, prefix) in enumerate(pb_registry, 2):
        code  = get_func(BA_BREAKS, fn)
        logic = PAGEBREAK_DESCRIPTIONS.get(fn, "")
        vals  = [fn, prefix, logic, code]
        fill  = ODD_FILL if ri % 2 == 0 else EVEN_FILL
        for ci, v in enumerate(vals, 1):
            cell = ws_pb.cell(row=ri, column=ci, value=v)
            cell.border = BOX
            cell.alignment = WRAP_ALIGN
            cell.fill = fill
            cell.font = Font(name="Consolas" if ci == 4 else "Calibri", size=9 if ci == 4 else 10)
        n = code.count("\n") + 1
        ws_pb.row_dimensions[ri].height = max(18, min(n * 13, 409))

    # helpers section
    last_data_row = len(pb_registry) + 1
    sep_row = last_data_row + 2
    c = ws_pb.cell(row=sep_row, column=1, value="── HELPER FUNCTIONS (used inside handlers) ──")
    c.font = Font(name="Calibri", size=11, bold=True)
    c.fill = SUBHEAD_FILL
    c.font = SUBHEAD_FONT
    ws_pb.merge_cells(start_row=sep_row, start_column=1, end_row=sep_row, end_column=4)

    for ri2, hname in enumerate(helper_names, sep_row + 1):
        code = get_func(BA_BREAKS, hname)
        vals = [hname, "(helper)", "(used by handlers above)", code]
        for ci, v in enumerate(vals, 1):
            cell = ws_pb.cell(row=ri2, column=ci, value=v)
            cell.border = BOX
            cell.alignment = WRAP_ALIGN
            cell.fill = ACCENT_FILL
            cell.font = Font(name="Consolas" if ci == 4 else "Calibri", size=9 if ci == 4 else 10)
        n = code.count("\n") + 1
        ws_pb.row_dimensions[ri2].height = max(18, min(n * 13, 409))

    ws_pb.column_dimensions["A"].width = 34
    ws_pb.column_dimensions["B"].width = 46
    ws_pb.column_dimensions["C"].width = 64
    ws_pb.column_dimensions["D"].width = 90

    # ── Sheet 4: generateWorksheet Reference ───────────────────────────────
    ws_gen = wb.create_sheet("generateWorksheet Reference")
    gen_headers = ["Method Name", "# Tables", "Layout Description",
                   "Internal Format Method", "Full Code (ExcelSettingsBA.py)"]
    for ci, h in enumerate(gen_headers, 1):
        cell = ws_gen.cell(row=1, column=ci, value=h)
        cell.fill  = HEADER_FILL
        cell.font  = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = BOX
    ws_gen.row_dimensions[1].height = 28
    ws_gen.freeze_panes = "A2"

    gen_registry = [
        ("generateWorksheet",         "1",   "Title → Subtitle → blank → df[0]",                                      "formatWorksheet"),
        ("generateWorksheet23B",       "2",   "df1 left cols + df2 ZONE RATED starting col K row 8",                   "formatWorksheet"),
        ("generateWorksheet2tables",   "2",   "Title → Sub1 → df[0] → blank → Sub2 → df[1]",                           "formatWorksheet2tables"),
        ("generateWorksheet2tbls",     "2",   "Delegates to generate_stacked_tables",                                  "formatWorksheetClass"),
        ("generate_stacked_tables",    "N",   "Title → Subtitle → blank → df[0] → blanks → df[1] → … (N tables)",     "formatWorksheetClass"),
        ("generateWorksheet3tables",   "3",   "Alias → generate_stacked_tables with [df1,df2,df3]",                    "formatWorksheetClass"),
        ("generateWorksheet4tables",   "4",   "Alias → generate_stacked_tables with [df1..df4]",                       "formatWorksheetClass"),
        ("generateWorksheet5tables",   "5",   "Alias → generate_stacked_tables",                                       "formatWorksheetClass"),
        ("generateWorksheet6tables",   "6",   "Alias → generate_stacked_tables",                                       "formatWorksheetClass"),
        ("generateWorksheet7tables",   "7",   "Alias → generate_stacked_tables",                                       "formatWorksheetClass"),
        ("generateWorksheet8tables",   "8",   "Alias → generate_stacked_tables",                                       "formatWorksheetClass"),
        ("generateWorksheet10tables",  "10",  "Alias → generate_stacked_tables",                                       "formatWorksheetClass"),
        ("generateWorksheet14tables",  "14",  "Alias → generate_stacked_tables",                                       "formatWorksheetClass"),
        ("generateWorksheetTablesX",   "N",   "Per-DF subtitles list; uses formatWorksheetX (minimal: page setup+footer+font only, no borders/number-formats)", "formatWorksheetX"),
    ]

    for ri, (name, ntables, layout, fmt) in enumerate(gen_registry, 2):
        code = get_func(BA_SETTINGS, name)
        vals = [name, ntables, layout, fmt, code]
        fill = ODD_FILL if ri % 2 == 0 else EVEN_FILL
        for ci, v in enumerate(vals, 1):
            cell = ws_gen.cell(row=ri, column=ci, value=v)
            cell.border = BOX
            cell.alignment = WRAP_ALIGN
            cell.fill = fill
            cell.font = Font(name="Consolas" if ci == 5 else "Calibri", size=9 if ci == 5 else 10)
        n = code.count("\n") + 1
        ws_gen.row_dimensions[ri].height = max(18, min(n * 13, 409))

    ws_gen.column_dimensions["A"].width = 32
    ws_gen.column_dimensions["B"].width = 8
    ws_gen.column_dimensions["C"].width = 64
    ws_gen.column_dimensions["D"].width = 36
    ws_gen.column_dimensions["E"].width = 90

    # ── Save ────────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"\nSaved: {OUTPUT_FILE}")
    print(f"  Sheets: Summary | Full Details | Page Break Reference | generateWorksheet Reference")
    print(f"  Rules documented: {len(rows)}")


# ═══════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    print("Scanning BA source files ...")
    rows = []
    for rule in RULES:
        try:
            rows.append(build_row(rule))
        except Exception as e:
            print(f"  WARNING: Error processing rule {rule['rule_no']}: {e}")

    print(f"  {len(rows)} rules processed. Writing Excel …")
    write_excel(rows)
