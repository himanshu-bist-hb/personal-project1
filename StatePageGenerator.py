# This program generates the state pages for the given inputs
# Required packages:
#   Pandas
#   Numpy
#   Openpyxl

import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
import StatePageValidationTest as VT
import ClassModifierPage as C
import AllProgramsPage as PG
import AllPerilPage as AP
import AutoServicePage as AS
import FoodServicePage as FS
import HabPage as H
import OfficePage as O
import ServicePage as S
import RetailPage as R
import WholesalePage as W
import RatingPlansPage as RP
import CommonRulesPage as CR
import AdditionalRulesPage as AR
import OptionalCoveragesPage as OP
#from BOPRatePageUserInterface import *

import pandas as pd
 
initialTime = time.perf_counter() # starting the clock

# **********************************************************************************************************************************************************************
#                                                           START OF INPUTS
# **********************************************************************************************************************************************************************

#                                       **LABELS EACH PROPOSED RATEBOOK INPUT BY THE USER AND IF NO RATEBOOK IS FOUND, IT LABELS IT AS "NOT FOUND"**
try: NGICRatebook
except NameError: NGICRatebook = "Not found"
if NGICRatebook != "Not found":
    NGICRatebook = pd.ExcelFile(NGICRatebook)
try: MMRatebook
except NameError: MMRatebook = "Not found"
if MMRatebook != "Not found":
    MMRatebook = pd.ExcelFile(MMRatebook)
try: NACORatebook
except NameError: NACORatebook = "Not found"
if NACORatebook != "Not found":
    NACORatebook = pd.ExcelFile(NACORatebook)
try: NAFFRatebook
except NameError: NAFFRatebook = "Not found"
if NAFFRatebook != "Not found":
    NAFFRatebook = pd.ExcelFile(NAFFRatebook)
try: NICOFRatebook
except NameError: NICOFRatebook = "Not found"
if NICOFRatebook != "Not found":
    NICOFRatebook = pd.ExcelFile(NICOFRatebook)
#                                       **LABELS EACH CURRENT RATEBOOK INPUT BY THE USER AND IF NO RATEBOOK IS FOUND, IT LABELS IT AS "NOT FOUND"**
try: CurrentNGICRatebook
except NameError: CurrentNGICRatebook = "Not found"
if CurrentNGICRatebook != "Not found":
    CurrentNGICRatebook = pd.ExcelFile(CurrentNGICRatebook)
try: CurrentMMRatebook
except NameError: CurrentMMRatebook = "Not found"
if CurrentMMRatebook != "Not found":
    CurrentMMRatebook = pd.ExcelFile(CurrentMMRatebook)
try: CurrentNACORatebook
except NameError: CurrentNACORatebook = "Not found"
if CurrentNACORatebook != "Not found":
    CurrentNACORatebook = pd.ExcelFile(CurrentNACORatebook)
try: CurrentNAFFRatebook
except NameError: CurrentNAFFRatebook = "Not found"
if CurrentNAFFRatebook != "Not found":
    CurrentNAFFRatebook = pd.ExcelFile(CurrentNAFFRatebook)
try: CurrentNICOFRatebook
except NameError: CurrentNICOFRatebook = "Not found"
if CurrentNICOFRatebook != "Not found":
    CurrentNICOFRatebook = pd.ExcelFile(CurrentNICOFRatebook)
try: CWRatebook
except NameError: CWRatebook = "Not found"
if CWRatebook != "Not found":
    CWRatebook = pd.ExcelFile(CWRatebook)
try: folder_selected
except NameError: folder_selected = "Not found"
if folder_selected != "Not found":
    folder_selected = folder_selected
try: RatingPlansApplies
except NameError: RatingPlansApplies = "Not found"
if RatingPlansApplies != "Not found":
    RatingPlansApplies = RatingPlansApplies
try: IRPMCredit
except NameError: IRPMCredit = "Not found"
if IRPMCredit != "Not found":
    IRPMCredit = IRPMCredit
try: IRPMDebit
except NameError: IRPMDebit = "Not found"
if IRPMDebit != "Not found":
    IRPMDebit = IRPMDebit
try: CommonRulesApplies
except NameError: CommonRulesApplies = "Not found"
if CommonRulesApplies != "Not found":
    CommonRulesApplies = CommonRulesApplies
try: AdditionalRulesApplies
except NameError: AdditionalRulesApplies = "Not found"
if AdditionalRulesApplies != "Not found":
    AdditionalRulesApplies = AdditionalRulesApplies
try: OptionalCoveragesApplies
except NameError: OptionalCoveragesApplies = "Not found"
if OptionalCoveragesApplies != "Not found":
    OptionalCoveragesApplies = OptionalCoveragesApplies
try: ClassApplies
except NameError: ClassApplies = "Not found"
if ClassApplies != "Not found":
    ClassApplies = ClassApplies
try: IndProgramsApplies
except NameError: IndProgramsApplies = "Not found"
if IndProgramsApplies != "Not found":
    IndProgramsApplies = IndProgramsApplies
try: AllPerilApplies
except NameError: AllPerilApplies = "Not found"
if AllPerilApplies != "Not found":
    AllPerilApplies = AllPerilApplies
try: AllProgramApplies
except NameError: AllProgramApplies = "Not found"
if AllProgramApplies != "Not found":
    AllProgramApplies = AllProgramApplies



if NGICRatebook != "Not found":
    RateBookDetails = pd.read_excel(NGICRatebook, sheet_name='Rate Book Details')
else:
    RateBookDetails = pd.read_excel(MMRatebook, sheet_name='Rate Book Details')
StateName = RateBookDetails.iloc[3,4]
nEffective = RateBookDetails.iloc[7,4]
nEffective = datetime.date.strftime(nEffective, "%m-%d-%y")
rEffective = RateBookDetails.iloc[7,4]
rEffective = datetime.date.strftime(rEffective, "%m-%d-%y")

#                                **CREATES A STATE ABBREVIATION VARIABLE THAT CAN BE USED THROUGHOUT THE CODE FOR SPECIFIC STATE EXCEPTIONS**
if StateName == "Alabama":
    state = "AL"
if StateName == "Alaska":
    state = "AK"
if StateName == "Arizona":
    state = "AZ"
if StateName == "Arkansas":
    state = "AR"
if StateName == "California":
    state = "CA"
if StateName == "Colorado":
    state = "CO"
if StateName == "Connecticut":
    state = "CT"
if StateName == "Delaware":
    state = "DE"
if StateName == "District of Columbia":
    state = "DC"
if StateName == "Florida":
    state = "FL"
if StateName == "Georgia":
    state = "GA"
if StateName == "Hawaii":
    state = "HI"
if StateName == "Idaho":
    state = "ID"
if StateName == "Illinois":
    state = "IL"
if StateName == "Indiana":
    state = "IN"
if StateName == "Iowa":
    state = "IA"
if StateName == "Kansas":
    state = "KS"
if StateName == "Kentucky":
    state = "KY"
if StateName == "Louisiana":
    state = "LA"
if StateName == "Maine":
    state = "ME"
if StateName == "Maryland":
    state = "MD"
if StateName == "Massachusetts":
    state = "MA"
if StateName == "Michigan":
    state = "MI"
if StateName == "Minnesota":
    state = "MN"
if StateName == "Mississippi":
    state = "MS"
if StateName == "Missouri":
    state = "MO"
if StateName == "Montana":
    state = "MT"
if StateName == "Nebraska":
    state = "NE"
if StateName == "Nevada":
    state = "NV"
if StateName == "New Hampshire":
    state = "NH"
if StateName == "New Jersey":
    state = "NJ"
if StateName == "New Mexico":
    state = "NM"
if StateName == "New York":
    state = "NY"
if StateName == "North Carolina":
    state = "NC"
if StateName == "North Dakota":
    state = "ND"
if StateName == "Ohio":
    state = "OH"
if StateName == "Oklahoma":
    state = "OK"
if StateName == "Oregon":
    state = "OR"
if StateName == "Pennsylvania":
    state = "PA"
if StateName == "Rhode Island":
    state = "RI"
if StateName == "South Carolina":
    state = "SC"
if StateName == "South Dakota":
    state = "SD"
if StateName == "Tennessee":
    state = "TN"
if StateName == "Texas":
    state = "TX"
if StateName == "Utah":
    state = "UT"
if StateName == "Vermont":
    state = "VT"
if StateName == "Virginia":
    state = "VA"
if StateName == "Washington":
    state = "WA"
if StateName == "West Virginia":
    state = "WV"
if StateName == "Wisconsin":
    state = "WI"
if StateName == "Wyoming":
    state = "WY"

#                                                            **DEFINES WHICH RATEBOOKS ARE BEING USED IN THE REVIEW BASED ON USER INPUTS**

if (NGICRatebook != "Not found" and NACORatebook != "Not found" and NAFFRatebook != "Not found" and NICOFRatebook != "Not found"):
    rateBooks = {'CW' : CWRatebook, # Country-wide ratebook (must be provided)
             'NACO': NACORatebook, # NACO ratebook (optional)
             'NAFF': NAFFRatebook, # NAFF ratebook (optional)
             'NGIC': NGICRatebook, # NGIC ratebook (must be provided)
             'NICOF': NICOFRatebook # NICOF ratebook (optional)
}
    
if (NGICRatebook != "Not found" and NACORatebook != "Not found" and NAFFRatebook != "Not found" and NICOFRatebook == "Not found"):
    rateBooks = {'CW': CWRatebook, # Country-wide ratebook (must be provided)
             'NACO': NACORatebook, # NACO ratebook (optional)
             'NAFF': NAFFRatebook, # NAFF ratebook (optional)
             'NGIC': NGICRatebook # NGIC ratebook (must be provided)
             #'NICOF': NICOFRatebook # NICOF ratebook (optional)
}
    
if (NGICRatebook != "Not found" and NACORatebook != "Not found" and NAFFRatebook == "Not found" and NICOFRatebook == "Not found"):
    rateBooks = {'CW': CWRatebook, # Country-wide ratebook (must be provided)
             'NACO': NACORatebook, # NACO ratebook (optional)
             #'NAFF': NAFFRatebook, # NAFF ratebook (optional)
             'NGIC': NGICRatebook # NGIC ratebook (must be provided)
             #'NICOF': NICOFRatebook # NICOF ratebook (optional)
}
    
if (NGICRatebook != "Not found" and NACORatebook == "Not found" and NAFFRatebook == "Not found" and NICOFRatebook == "Not found"):
    rateBooks = {'CW': CWRatebook, # Country-wide ratebook (must be provided)
             #'NACO': NACORatebook, # NACO ratebook (optional)
             #'NAFF': NAFFRatebook, # NAFF ratebook (optional)
             'NGIC': NGICRatebook # NGIC ratebook (must be provided)
             #'NICOF': NICOFRatebook # NICOF ratebook (optional)
}

if state == "TX":
    perils = ['allother1', 'cat1', 'cat2', 'cat3', 'cat4', 'fire1', 'fire2', 'fire3', 'fire4', 'liability1', 'liability2', 'liability3', 'liability4', 'theft1', 'water1', 'water2', 'weather1', 'weather2']
elif state in ('AZ', 'CA', 'CO', 'ID', 'MT', 'NM', 'NV', 'OR', 'TX', 'UT', 'WA', 'WY'):
    perils = ['allother1', 'cat1', 'cat2', 'cat4', 'fire1', 'fire2', 'fire3', 'fire4', 'liability1', 'liability2', 'liability3', 'liability4', 'theft1', 'water1', 'water2', 'weather1', 'weather2']
elif state in ('AL', 'AR', 'CT', 'DC', 'DE', 'FL', 'GA', 'IL', 'IN', 'KY', 'MA', 'MD', 'ME', 'MO', 'MS', 'NC', 'NH', 'NJ', 'NY', 'OH', 'PA', 'RI', 'SC', 'TN', 'VA', 'VT', 'WV'):
    perils = ['allother1', 'cat1', 'cat2', 'cat3', 'cat4', 'fire1', 'fire3', 'fire4', 'liability1', 'liability2', 'liability3', 'liability4', 'theft1', 'water1', 'water2', 'weather1', 'weather2']
elif state in ('IA', 'KS', 'MI', 'MN', 'ND', 'NE', 'SD', 'WI'):
    perils = ['allother1', 'cat1', 'cat2', 'cat4', 'fire1', 'fire3', 'fire4', 'liability1', 'liability2', 'liability3', 'liability4', 'theft1', 'water1', 'water2', 'weather1', 'weather2']

#                                                            **ASSIGNS THE TERRITORY DEFINTIONS EXCEL FILEPATH**    
if AllProgramApplies.get() == 1:
    print("Loading Territory Definitions")
    TerritoryDefs = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\BOP\\Territory Defs\\4 Territory Defs (By Peril Grid)\\AllStatesDefinitionswLatLon.xlsx')
    TerritoryDefsByST = pd.read_excel(TerritoryDefs, sheet_name=state)


#                                                            **ASSIGNS THE EQ TERRITORY DEFINTIONS TXT FILEPATH**
if OptionalCoveragesApplies.get() == 1:
    print("Loading EQ Territory Definitions")
    EQTerritoryDefs = pd.read_csv('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\BOP\\Territory Defs\\4 Territory Defs (By Peril Grid)\\NWCE_' + state + '_ZIP_1223.txt', sep= '\t', header=0)

#if RatingPlansApplies.get() == 1:
#    perils = ['allother1', 'cat1', 'cat2', 'cat3', 'fire1', 'fire3', 'liability1', 'theft1', 'water1', 'water2', 'weather1', 'weather2']
#else: perils = ['allother1', 'cat1', 'cat2', 'fire1', 'fire3', 'liability1', 'theft1', 'water1', 'water2', 'weather1', 'weather2']

    

# **********************************************************************************************************************************************************************
#                                                           END OF INPUTS
# **********************************************************************************************************************************************************************

#state = 'RI' # 2 letter abbreviation for the state
#rateBooks = {'CW': r"M:\Actshare\Com\Annual_Rate_Reviews\2023\BOP\RI - CLT\Ratebooks\Proposed (Upload)\BP7+CW+Rate+Book+-+Ext_v65_1.xlsx", # Country-wide ratebook (must be provided)
#             'NACO': r"M:\Actshare\Com\Annual_Rate_Reviews\2023\BOP\RI - CLT\Ratebooks\Current (Download)\BP7+RI+NACO+-+v2_Ext (1).xlsx", # NACO ratebook (optional)
#             'NAFF': r"M:\Actshare\Com\Annual_Rate_Reviews\2023\BOP\RI - CLT\Ratebooks\Current (Download)\BP7+RI+NAFF+-+v2_Ext.xlsx", # NAFF ratebook (optional)
#             'NGIC': r"M:\Actshare\Com\Annual_Rate_Reviews\2023\BOP\RI - CLT\Ratebooks\Current (Download)\BP7+RI+Rate+Book+-+v8_Ext.xlsx", # NGIC ratebook (must be provided)
#             'NICOF': r"M:\Actshare\Com\Annual_Rate_Reviews\2023\BOP\RI - CLT\Ratebooks\Current (Download)\BP7+RI+NICOF+-+v2_Ext.xlsx" # NICOF ratebook (optional)
#}

#nEffective = '12-01-2023' # New business effective date
#rEffective = '12-01-2023' # Renewal business effective dated

# Full list of perils for reference: ['allother1', 'allperil', 'cat1', 'cat2', 'cat3', 'cat4', 'fire1', 'fire2', 'fire3', 'fire4', 'liability1', 'liability2', 'liability3', 'liability4', 'theft1', 'water1', 'water2', 'weather1', 'weather2']
# Perils to be used in the All Programs page
#perils = ['allother1', 'cat1', 'cat2', 'cat3', 'fire1', 'fire3', 'liability1', 'theft1', 'water1', 'water2', 'weather1', 'weather2']
# **********************************************************************************************************************************************************************
#                                                           END OF INPUTS
# **********************************************************************************************************************************************************************

# Perils to be used in the individual program pages
programPerils = perils + ['allperil']

# Array of perils to be used in validation testing
perilsValidation = ['allother1', 'cat1', 'cat2', 'cat3', 'cat4', 'fire1', 'fire2', 'fire3', 'fire4', 'liability1', 'liability2', 'liability3', 'liability4', 'theft1', 'water1', 'water2', 'weather1', 'weather2']

# A dictionary to convert between various peril names
#perilsConversions = {'allother1': 'NW-Other', 'allperil': 'All Peril', 'cat1': 'ST', 'cat2': 'WS', 'cat3': 'HU', 'cat4': 'CAT4', 'fire1': 'NW-Fire', 'fire2': 'CAWF', 'fire3': 'FFEQ', 'fire4': 'NC-FFL', \
#                     'liability1': 'LIAB-Other', 'liability2': 'LIAB-Prem', 'liability3': 'LIAB-Prod', 'liability4': 'LIAB-Med', 'theft1': 'NW-Theft', 'water1': 'NW-Water', 'water2': 'NC-Water', 'weather1': 'NC-Other', 'weather2': 'NC-Wind'}

perilsConversions = {'allother1': 'NW-Other', 'allperil': 'AllPeril', 'cat1': 'ST', 'cat2': 'WS', 'cat3': 'HU', 'cat4': 'L-Products', 'fire1': 'NW-Fire', 'fire2': 'WF', 'fire3': 'FFEQ', 'fire4': 'NC-BINC', \
                     'liability1': 'L-SlipFall', 'liability2': 'L-Violence', 'liability3': 'L-OtherMed', 'liability4': 'L-OtherPrem', 'theft1': 'NW-Theft', 'water1': 'NW-Water', 'water2': 'NC-Water', 'weather1': 'NC-Other', 'weather2': 'NC-Wind'}

programCodes = {'Auto': 20000, 'Food': 40000, 'Hab': 10000, 'Office': 60000, 'Service': 70000, 'Retail': 50000, 'Wholesale': 80000}
classCodes = {10000: 'Hab', 20000: 'Auto', 40000: 'Food', 50000: 'Retail', 60000: 'Office', 70000: 'Service', 80000: 'Wholesale'}

# A dictionary to remove excess leading zeros in the protection classes
protectionClassConversions = {'000001': '1', '000002': '2', '000003': '3', '000004': '4', '000005': '5', '000006': '6', '000007': '7', '000008': '8', '000009': '9', '000010': '10', \
                              '00001X': '1X', '00002X': '2X', '00003X': '3X', '00004X': '4X', '00005X': '5X', '00006X': '6X', '00007X': '7X', '00008X': '8X', \
                              '00001Y': '1Y', '00002Y': '2Y', '00003Y': '3Y', '00004Y': '4Y', '00005Y': '5Y', '00006Y': '6Y', '00007Y': '7Y', '00008Y': '8Y', \
                              '00001W': '1W', '00002W': '2W', '00003W': '3W', '00004W': '4W', '00005W': '5W', '00006W': '6W', '00007W': '7W', '00008W': '8W', \
                              '00008B': '8B', '00009E': '9E', '00009S': '9S', '00010W': '10W'}

# A dictionary that contains states that have more than 1 building code
# The outer dictionary maps the states to their BCEG groups, and  the inner dictionary maps the building codes to the group
buildingCodes = {'AL': {'A': ['001'], 'B': ['004'], 'C': ['005'], 'D': ['006']},
                 'FL': {'A': ['011', '012'], 'B': ['010', '015'], 'C': ['002', '007', '008', '014', '016', '017'], 'D': ['009', '013']},
                 'GA': {'A': ['002'], 'B': ['004'], 'C': ['005'], 'D': ['006']},
                 #'MD': {'A': ['701'], 'B': ['702'], 'C': ['704'], 'D': ['705']},
                 'MS': {'A': ['002'], 'B': ['003'], 'C': ['004']},
                 'NC': {'A': ['003'], 'B': ['004'], 'C': ['005'], 'D': ['006']},
                 'NE': {'A': ['701'], 'B': ['703'], 'C': ['704']},
                 #'OR': {'A': ['701'], 'B': ['702']},
                 #'RI': {'A': ['701'], 'B': ['702'], 'C': ['703']},
                 'SC': {'A': ['002'], 'B': ['003'], 'C': ['004']},
                 'TX': {'A': ['004', '005', '006', '007', '008', '009', '015', '016'], 'B': ['010', '011', '012', '013', '014']},
                 'VA': {'A': ['001', '005', '006', '007', '008', '009', '012', '013'], 'B': ['010', '011']},
                 'WY': {'A': ['702'], 'B': ['703']}}

# Stores all the rate tables used in the pages
# Outer dictionary:
#   Keys: company names
#   Values: inner dictionary
# Inner dictionary:
#   Keys: rate table codes (cell B6 in Excel)
#   Values: tables associated with the codes
rateTables = {}

for company, companyFile in rateBooks.items():
    rateBook = load_workbook(companyFile)
    rateTables[company] = {} # Initializing the inner dictionary here for each company
    for sheet in rateBook:
        if sheet.title == 'Rate Book Details' or sheet["A1"].value[-2:] == 'RR': # Skipping the index sheet of the ratebook and any Rate Routine (RR) tabs
            continue
        char = get_column_letter(sheet.max_column) # Letter representing the last column
        cell_range = 'A12:' + char + str(sheet.max_row)
        cr = CellRange(cell_range)
        width = cr.max_col - cr.min_col
        if not width:
            cells = [cell.value for cell in [row[0] for row in sheet[cell_range]]]
        else:
            cells = []
            for row in sheet[cell_range]:
                cells.append([c.value for c in row])
        rateTables[company][sheet["B6"].value] = cells # Using the value in cell B6 as the unique identifier for each table

#ValidationTest = VT.ValidationTest(rateTables, perilsValidation, programCodes)
#ValidationTest.runValidationTest()

# All of the files below will be saved to the current directory by default
# If a different directory is desired, updated the appropiate dest_filenameX variables

if ClassApplies.get() == 1:
    print("Creating Class Pages")
    ClassModifier = C.ClassModifier(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    ClassModifierPages = ClassModifier.buildClassModifierPage()
    dest_filename1 = folder_selected + '/Class State Pages.xlsx'
    ClassModifierPages.save(filename=dest_filename1)

if AllProgramApplies.get() == 1:
    print("Creating All Programs Pages")
    AllPrograms = PG.AllPrograms(state, rateTables, perils, perilsConversions, protectionClassConversions, buildingCodes, nEffective, rEffective, TerritoryDefsByST)
    AllProgramsPages = AllPrograms.buildAllProgramsPage()
    dest_filename2 = folder_selected + '/All Programs State Pages.xlsx'
    AllProgramsPages.save(filename=dest_filename2)

if AllPerilApplies.get() == 1:
    print("Creating All Peril Pages")
    AllPeril = AP.AllPeril(state, rateTables, classCodes, protectionClassConversions, buildingCodes, nEffective, rEffective)
    AllPerilPages = AllPeril.buildAllPerilPage()
    dest_filename3 = folder_selected + '/All Peril State Pages.xlsx'
    AllPerilPages.save(filename=dest_filename3)

if IndProgramsApplies.get() == 1:
    print("Creating Auto Pages")
    AutoService = AS.Auto(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    AutoServicePages = AutoService.buildAutoPage()
    dest_filename4 = folder_selected + '/Auto Service State Pages.xlsx'
    AutoServicePages.save(filename=dest_filename4)

    print("Creating Food Pages")
    FoodService = FS.Food(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    FoodServicePages = FoodService.buildFoodPage()
    dest_filename5 = folder_selected + '/Food Service State Pages.xlsx'
    FoodServicePages.save(filename=dest_filename5)

    print("Creating Hab Pages")
    Habitational = H.Hab(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    HabitationalPages = Habitational.buildHabPage()
    dest_filename6 = folder_selected + '/Hab State Pages.xlsx'
    HabitationalPages.save(filename=dest_filename6)

    print("Creating Office Pages")
    Office = O.Office(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    OfficePages = Office.buildOfficePage()
    dest_filename7 = folder_selected + '/Office State Pages.xlsx'
    OfficePages.save(filename=dest_filename7)

    print("Creating Service Pages")
    Service = S.Service(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    ServicePages = Service.buildServicePage()
    dest_filename8 = folder_selected + '/Service State Pages.xlsx'
    ServicePages.save(filename=dest_filename8)

    print("Creating Retail Pages")
    Retail = R.Retail(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    RetailPages = Retail.buildRetailPage()
    dest_filename9 = folder_selected + '/Retail State Pages.xlsx'
    RetailPages.save(filename=dest_filename9)

    print("Creating Wholesale Pages")
    Wholesale = W.Wholesale(state, rateTables, programPerils, perilsConversions, nEffective, rEffective)
    WholesalePages = Wholesale.buildWholesalePage()
    dest_filename10 = folder_selected + '/Wholesale State Pages.xlsx'
    WholesalePages.save(filename=dest_filename10)

if RatingPlansApplies.get() == 1:
    print("Creating Rating Plan Pages")
    RatingPlans = RP.RatingPlans(state, rateTables, programPerils, perilsConversions, nEffective, rEffective, RatingPlansApplies, IRPMCredit, IRPMDebit)
    RatingPages = RatingPlans.buildRatingPlansPage()
    dest_filename10 = folder_selected + '/Rating Plan State Pages.xlsx'
    RatingPages.save(filename=dest_filename10)

if CommonRulesApplies.get() == 1:
    print("Creating Common Rule Pages")
    CommonRules = CR.CommonRules(state, rateTables, programPerils, perilsConversions, nEffective, rEffective, CommonRulesApplies)
    CommonRulesPages = CommonRules.buildCommonRulesPage()
    dest_filename10 = folder_selected + '/Common Rules State Pages.xlsx'
    CommonRulesPages.save(filename=dest_filename10)

if AdditionalRulesApplies.get() == 1:
    print("Creating Additional Rule Pages")
    AdditionalRules = AR.AdditionalRules(state, rateTables, programPerils, perilsConversions, classCodes, nEffective, rEffective, AdditionalRulesApplies)
    AdditionalRulesPages = AdditionalRules.buildAdditionalRulesPage()
    dest_filename10 = folder_selected + '/Additional Rules State Pages.xlsx'
    AdditionalRulesPages.save(filename=dest_filename10)

if OptionalCoveragesApplies.get() == 1:
    print("Creating Optional Coverages Pages")
    OptionalCoverages = OP.OptionalCoverages(state, rateTables, perils, classCodes, perilsConversions, nEffective, rEffective, OptionalCoveragesApplies, EQTerritoryDefs, NGICRatebook, NACORatebook, NAFFRatebook, NICOFRatebook, MMRatebook)
    OptionalCoveragesPages = OptionalCoverages.buildOptionalCoveragesPage()
    dest_filename10 = folder_selected + '/Optional Coverages State Pages.xlsx'
    OptionalCoveragesPages.save(filename=dest_filename10)

endTime = time.perf_counter() # stopping the clock
print(f'This program ran in {endTime - initialTime:0.4f} seconds')