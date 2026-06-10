# This module formats the Rating Plans State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class CommonRules:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective, CommonRulesApplies) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date
        self.CommonRulesApplies = CommonRulesApplies

        #self.autoProgramCode = 20000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company


    # Builds the Fixed Expense table
    # Returns a dataframe
    def buildExpenseConstant(self):
        ExpenseConstant = self.buildDataFrame("BP7PerilCWSpecificCovRateFactors")
        filteredExpenseConstant = ExpenseConstant.query(f'CoverageSpecific == "Expense Constant"').filter(items=['Value']).rename(columns={'Value': 'Rate'})
        return filteredExpenseConstant


    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)


    # Applies manual formatting to the RPMET worksheet
    def formatFixedExpense(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(70)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '$#,##0'
            ws.column_dimensions['A'].width = self.pixelsToInches(100)

    
    # Sets up the Auto Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildCommonRulesPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        CommonRules = ExcelSettings.Excel(state=self.state, programName='Common Rules', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = CommonRules.getFontName()
        fontSize = CommonRules.getFontSize()

        CommonRules.generateWorksheet('EC', 'CR Table 8. Expense Constant', self.buildExpenseConstant(), False, True)

        CommonRules.createIndex()
        CommonRulesPages = CommonRules.getWB()


        self.formatFixedExpense(CommonRulesPages['EC'])  


        return CommonRulesPages