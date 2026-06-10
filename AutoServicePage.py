# This module formats the Auto Service State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class Auto:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.autoProgramCode = 20000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building", 
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the territory multiplier table for the given coverage (either building, bpp, or liability)
    # Returns a dataframe
    #def buildTerritoryMultiplier(self, coverage):
    #    territorialFactor = self.buildDataFrame("BP7_Peril_TerritorialFactor")
    #    filteredTerritorialFactor = territorialFactor.filter(items=['Peril TypeCode', 'TerritoryCode', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor'])
    #    if coverage.casefold() == 'FFEQ': # Case-insensitive comparison
    #        return filteredTerritorialFactor.query(f'`Peril TypeCode` == "fire3"').rename(columns={'TerritoryCode': 'Territory', 'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'HU': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "HU"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'L-OtherMed': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "L-OtherMed"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'L-OtherPrem': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "L-OtherPrem"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'L-Products': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "L-Products"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'L-SlipFall': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "L-SlipFall"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'L-Violence': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "L-Violence"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NC-BINC': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NC-BINC"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NC-Other': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NC-Other"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NC-Water': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NC-Water"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NC-Wind': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NC-Wind"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NW-Fire': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NW-Fire"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NW-Other': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NW-Other"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NW-Theft': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NW-Theft"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'NW-Water': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "NW-Water"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'ST': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "ST"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'WF': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "WF"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})
        #elif peril.casefold() == 'WS': # Case-insensitive comparison
        #    return filteredTerritorialFactor.query(f'`Peril TypeCode` == "WS"').filter(items=['Territory', 'BldgTerritoryFactor', 'BPPTerritoryFactor', 'LiabilityTerritoryFactor', 'BITerritoryFactor']).rename(columns={'BldgTerritoryFactor': 'Building', 'BPPTerritoryFactor': 'BPP', 'LiabilityTerritoryFactor': 'Liability', 'BITerritoryFactor': 'BI'})

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'] == 'L-Products'].index)
        filteredTheftOptions = filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Full Theft': 'Full', 'Limited Theft': 'Limited'})
        return filteredTheftOptions
        

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'}) # Converting to int first to get rid of decimal places
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0', 
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe    
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.autoProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the car wash factor table
    # Returns a dataframe 
    def buildCarWashFactor(self):
        carWashFactor = self.buildDataFrame("BP7_Peril_Car_Wash_Factor")
        return carWashFactor.dropna().query(f'`Class Code Min` == {self.autoProgramCode} & `Peril TypeCode` == "liability1"').rename(columns={'No Of Bays': 'Number of Bays', 'CarWashFactor': 'Factor'}). \
                replace({'Number of Bays': {1: 'One', 2: 'Two', 3: 'Three', 4: '4 or more'}}).filter(items=['Number of Bays', 'Factor'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.autoProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_Receipts_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCode_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & AutoServType == "N/A" & FoodServType == "N/A"').rename(columns={'ReceiptMin': 'Min', 'ReceiptMax': 'Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Min', 'Max'], columns='Peril TypeCode', values='Factor').reset_index(['Min', 'Max']).fillna({'Max': 'and over'})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.autoProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance' : 'int32'})

    # Builds the liquified petroleum gas (LPG) exposures table
    # Returns a dataframe
    def buildLPGExposure(self):
        lpgExposure = self.buildDataFrame("BP7_LPG_Premium")
        return lpgExposure.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'LPGPremium': 'Premium (each premises)'})

    # Builds the additional insured - garage operations table
    # Returns a dataframe
    def buildGarageOperations(self):
        garageOperations = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates").query(f'CoverageName == "AdditionalInsrdGarageOperations"').astype({'BaseRate': 'str'})
        garageOperations['Rate'] = '$' + garageOperations['BaseRate'] + ' per additional insured subject to the Liability Limit factor'
        return garageOperations.filter(items=['Rate'])

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.autoProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Auto Service premises'})

    # Builds the optional increased limits - employee portable tools and equipment tables
    # The equalLimits parameter determines which table is being built (i.e. the table where the employee limit is equal to the occurence limit or not)
    # Returns a dataframe
    def buildEmpPortToolsAndEquipment(self, equalLimits):
        empPortToolsAndEquipment = self.buildDataFrame("BP7_EmployeePortableToolsAndEquipmentCov_Rate")
        empPortToolsAndEquipment['OptionalLimitsPerEmployee'] = empPortToolsAndEquipment['Optional_IncreasedLimits'].str.replace(',', '').str.split('/').str[0]
        empPortToolsAndEquipment['OptionalLimitsPerOccurence'] = empPortToolsAndEquipment['Optional_IncreasedLimits'].str.replace(',', '').str.split('/').str[1]
        empPortToolsAndEquipment['TotalLimitsPerEmployee'] = empPortToolsAndEquipment['TotalLimits'].str.replace(',', '').str.split('/').str[0]
        empPortToolsAndEquipment['TotalLimitsPerOccurence'] = empPortToolsAndEquipment['TotalLimits'].str.replace(',', '').str.split('/').str[1]
        filteredEmpPortToolsAndEquipment = empPortToolsAndEquipment.filter(items=['OptionalLimitsPerEmployee', 'OptionalLimitsPerOccurence', 'TotalLimitsPerEmployee', 'TotalLimitsPerOccurence', 'Additional_Premium']). \
                rename(columns={'Additional_Premium': 'Additional Premium'})
        if equalLimits:
            return filteredEmpPortToolsAndEquipment.query(f'TotalLimitsPerEmployee == TotalLimitsPerOccurence'). \
                    astype({'OptionalLimitsPerEmployee': 'int32', 'OptionalLimitsPerOccurence': 'int32', 'TotalLimitsPerEmployee': 'int32', 'TotalLimitsPerOccurence': 'int32'}).sort_values(by=['TotalLimitsPerEmployee'])
        return filteredEmpPortToolsAndEquipment.query(f'TotalLimitsPerEmployee != TotalLimitsPerOccurence'). \
                astype({'OptionalLimitsPerEmployee': 'int32', 'OptionalLimitsPerOccurence': 'int32', 'TotalLimitsPerEmployee': 'int32', 'TotalLimitsPerOccurence': 'int32'}).sort_values(by=['TotalLimitsPerEmployee'])

    # Builds the groadened garage liability – defective products and faulty work coverage endorsement base rate table
    # Returns a dataframe
    def buildBroadenedGarageLiab(self):
        miscBaseRates = self.buildDataFrame("BP7_Miscellaneous_Base_Rates")
        return miscBaseRates.query(f'BaseRateName == "BroadenedGarageLiability"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the special property damage deductible table
    # Returns a dataframe
    def buildSpecialPropertyDeductible(self):
        specialPropertyDeductible = self.buildDataFrame("BP7_Special_Property_Damage_Deductible")
        return specialPropertyDeductible.astype({'Deductible': 'int32'})

    # Builds the franchise upgrage endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.autoProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(), 
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(), 
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the base rates worksheet
    def formatBaseRates(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(159)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = '#,##0.0000' # 4 values after the decimal point for the base rates

    # Applies manual formatting to the territory multiplier worksheet
    #def formatTerritoryMultiplier(self, ws):
    #    ws.column_dimensions['A'].width = self.pixelsToInches(70)
    #    for col in range(2, ws.max_column + 1):
    #        char = get_column_letter(col) # Letter representing the current column
    #        ws.column_dimensions[char].width = self.pixelsToInches(54)  

    # Applies manual formatting to the construction factor worksheet
    def formatConstructionFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    # Applies manual formatting to the theft options worksheet
    def formatTheftOptions(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(80)

    # Applies manual formatting to the year built modifier worksheet
    def formatYearBuiltModifier(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(131)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '###0'

    # Applies manual formatting to the equipment breakdown base rate worksheet
    def formatEBBaseRate(self, ws):
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the car wash factor worksheet
    def formatCarWashFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(124)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)

    # Applies manual formatting to the property damage deductible worksheet
    def formatPropertyDamageDeductible(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(187)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the liability size of risk modifier worksheet
    def formatLiabilitySizeRisk(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Receipts Range'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col < 3:
                ws.column_dimensions[char].width = self.pixelsToInches(95)
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(120)

    # Applies manual formatting to the liability limit factor worksheet
    def formatLiabilityLimitFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the Liquified Petroleum Gas (LPG) exposures worksheet
    def formatLPGExposures(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(200)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat
            ws['B' + str(row)].number_format = self.currencyFormat

    # Applies manual formatting to the additional insured - garage operations worksheet
    def formatGarageOperations(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(480)

    # Applies manual formatting to the endorsement charge worksheet
    def formatEndorsementCharge(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(350)
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the employee portable tools and equipment worksheet
    def formatEmpPortToolsAndEquipment(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Optional Increased Limits'
        ws['A4'] = 'Per Employee / Occurrence'
        ws['C3'] = 'Total Limits'
        ws['C4'] = 'Per Employee / Occurrence'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('A4:B4')
        ws.merge_cells('C3:D3')
        ws.merge_cells('C4:D4')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1): 
            char = get_column_letter(col) # Letter representing the current column
            for row in range(5, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 5: # Applies currency formatting to columns A-D
                    cell.number_format = self.currencyFormat
                else:
                    cell.number_format = '$#,##0.00'

    # Applies manual formatting to the broadened garage liability worksheet
    def formatBroadenedGarageLiabiltiy(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(100)

    # Applies manual formatting to the special property damage deductible worksheet
    def formatSpecialPropertyDeductible(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(90)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = self.currencyFormat

    # Applies manual formatting to the franchise endorsement worksheet
    def formatFranchiseEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            cell = ws['B' + str(row)]
            cell.number_format = '$#,##0.00'          

    # Sets up the Auto Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildAutoPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        AutoService = ExcelSettings.Excel(state=self.state, programName='Auto Service', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = AutoService.getFontName()
        fontSize = AutoService.getFontSize()

        if 'NACO' in self.rateTables.keys():
            AutoService.generateWorksheet('BRNACO', 'AS Table 3.B.1. NW Assurance State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NAFF' in self.rateTables.keys():
            AutoService.generateWorksheet('BRNAFF', 'AS Table 3.B.1. NW Affinity State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NGIC' in self.rateTables.keys():
            AutoService.generateWorksheet('BRNGIC', 'AS Table 3.B.1. NW General Insurance Company', self.buildBaseRates('NGIC'), False, True)
        if 'NICOF' in self.rateTables.keys():
            AutoService.generateWorksheet('BRNICOF', 'AS Table 3.B.1. NICOF State Base Rates', self.buildBaseRates('NGIC'), False, True)
        #AutoService.generateWorksheet('TRFFEQ', 'AS Table 3.C.1.a. State Territory Multiplier - FFEQ', self.buildTerritoryMultiplier('FFEQ'), False, True)
        #AutoService.generateWorksheet('TRHU', 'AS Table 3.C.1.a. State Territory Multiplier - HU', self.buildTerritoryMultiplier('HU'), False, True)
        #AutoService.generateWorksheet('TRLOM', 'AS Table 3.C.1.a. State Territory Multiplier - L-OtherMed', self.buildTerritoryMultiplier('L-OtherMed'), False, True)
        AutoService.generateWorksheet('CBG', 'AS Table 3.C.2.c. Construction Factor - Building', self.buildConstructionType('Building'), False, True)
        AutoService.generateWorksheet('CPP', 'AS Table 3.C.2.c. Construction Factor - BPP', self.buildConstructionType('BPP'), False, True)
        AutoService.generateWorksheet('TO', 'AS Table 3.C.2.m. Theft Options', self.buildTheftOptions(), False, True)
        AutoService.generateWorksheet('YBBG', 'AS Table 3.C.2.p. Year Built Modifier - Building', self.buildYearBuiltModifier('Building'), False, True)
        AutoService.generateWorksheet('YBPP', 'AS Table 3.C.2.p. Year Built Modifier - BPP', self.buildYearBuiltModifier('BPP'), False, True)
        AutoService.generateWorksheet('EBB', 'AS Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate(), False, True)
        AutoService.generateWorksheet('CW', 'AS Table 3.C.4.b. Car Wash Factor', self.buildCarWashFactor(), False, True)
        AutoService.generateWorksheet('PDLD', 'AS Table 3.C.4.c. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount(), False, True)
        AutoService.generateWorksheet('LS', 'AS Table 3.C.4.f. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk(), False, True)
        AutoService.generateWorksheet('LL', 'AS Table 3.C.4.g. Liability Limit Factor', self.buildLiabilityLimitFactor(), False, True)
        AutoService.generateWorksheet('LPGE', 'AS Table 3.C.4.h. Liquefied Petroleum Gas (LPG) Exposures', self.buildLPGExposure(), False, True)
        AutoService.generateWorksheet('AIGO', 'AS Table 4.A. Additional Insured - Garage Operations', self.buildGarageOperations(), False, True)
        AutoService.generateWorksheet('PLUS', 'AS Table 4.B.1. Auto Service PLUS Endorsement', self.buildEndorsementCharge(), False, True)
        AutoService.generateWorksheet('OILN', 'AS Table 4.B.2. Optional Increased Limits - Employee Portable Tools and Equipment - Employee Limit is not equal to the Occurence Limit', self.buildEmpPortToolsAndEquipment(False), False, True) 
        AutoService.generateWorksheet('OILE', 'AS Table 4.B.2. Optional Increased Limits - Employee Portable Tools and Equipment - Employee Limit is equal to the Occurence Limit', self.buildEmpPortToolsAndEquipment(True), False, True)
        AutoService.generateWorksheet('BGL', 'AS Table 4.C.1. Broadened Garage Liability – Defective Products and Faulty Work Coverage Endorsement Base Rate', self.buildBroadenedGarageLiab(), False, True)
        AutoService.generateWorksheet('SPD', 'AS Table 4.C.2. Special Property Damage Deductible', self.buildSpecialPropertyDeductible(), False, True)
        AutoService.generateWorksheet('FR', 'AS Table 4.E. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement(), False, True)

        AutoService.createIndex()
        AutoPages = AutoService.getWB()

        if 'NACO' in self.rateTables.keys():
            self.formatBaseRates(AutoPages['BRNACO'])
        if 'NAFF' in self.rateTables.keys():
            self.formatBaseRates(AutoPages['BRNAFF'])
        if 'NGIC' in self.rateTables.keys():
            self.formatBaseRates(AutoPages['BRNGIC'])
        if 'NICOF' in self.rateTables.keys():
            self.formatBaseRates(AutoPages['BRNICOF'])
        #self.formatTerritoryMultiplier(AutoPages['TRFFEQ'])
        #self.formatTerritoryMultiplier(AutoPages['TRHU']) 
        #self.formatTerritoryMultiplier(AutoPages['TRLOM'])
        self.formatConstructionFactor(AutoPages['CBG'])  
        self.formatConstructionFactor(AutoPages['CPP'])  
        self.formatTheftOptions(AutoPages['TO'])
        self.formatYearBuiltModifier(AutoPages['YBBG'])
        self.formatYearBuiltModifier(AutoPages['YBPP'])  
        self.formatEBBaseRate(AutoPages['EBB'])
        self.formatCarWashFactor(AutoPages['CW'])
        self.formatPropertyDamageDeductible(AutoPages['PDLD'])
        self.formatLiabilitySizeRisk(AutoPages['LS'], Font(name=fontName, size=fontSize, bold=True))
        self.formatLiabilityLimitFactor(AutoPages['LL'])
        self.formatLPGExposures(AutoPages['LPGE'])
        self.formatGarageOperations(AutoPages['AIGO'])
        self.formatEndorsementCharge(AutoPages['PLUS'])
        self.formatEmpPortToolsAndEquipment(AutoPages['OILN'], Font(name=fontName, size=fontSize, bold=True))
        self.formatEmpPortToolsAndEquipment(AutoPages['OILE'], Font(name=fontName, size=fontSize, bold=True))
        self.formatBroadenedGarageLiabiltiy(AutoPages['BGL'])
        self.formatSpecialPropertyDeductible(AutoPages['SPD'])
        self.formatFranchiseEndorsement(AutoPages['FR'])

        return AutoPages