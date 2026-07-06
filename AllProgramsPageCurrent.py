# This module formats the All Programs State Page workbook in Excel

import pandas as pd
import math
import ExcelSettingsCurrent
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class AllPrograms:
    def __init__(self, state, rateTables, perils, perilsConversions, protectionClassConversions, buildingCodes, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.protectionClassConversions = protectionClassConversions
        self.buildingCodes = buildingCodes
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the sprinkler factor table
    # Returns a dataframe
    def buildSprinklerFactor(self):
        sprinklerPeril = self.buildDataFrame("BP7_Peril_Sprinkler_Discount")
        filteredSprinklerPeril = sprinklerPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils}').filter(items=['Peril TypeCode', 'Bldg Sprinkler Factor', 'BPP Sprinkler Factor'])
        return filteredSprinklerPeril.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Peril TypeCode': 'Peril', 'Bldg Sprinkler Factor': 'Building', 'BPP Sprinkler Factor': 'BPP'}).sort_values(by=['Peril'])

    # Builds the protection class table for the given coverage (either Building or BPP)
    # Returns a dataframe
    def buildProtectionClass(self, coverage):
        protectionPeril = self.buildDataFrame("BP7_Peril_Protection_Class")
        filteredProtectionPeril = protectionPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions, 'ProtectionClass': self.protectionClassConversions}). \
                rename(columns={'ProtectionClass': 'Protection Class'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            pivotedProtectionPeril = filteredProtectionPeril.pivot(index='Protection Class', columns='Peril TypeCode', values='BldgProtectionClassFactor').reset_index('Protection Class')
            pivotedProtectionPeril['Protection Class Number'] = pivotedProtectionPeril['Protection Class'].apply(self.getProtectionClassValue)
            return pivotedProtectionPeril.sort_values(by=['Protection Class Number']).loc[:, pivotedProtectionPeril.columns != 'Protection Class Number']
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            pivotedProtectionPeril = filteredProtectionPeril.pivot(index='Protection Class', columns='Peril TypeCode', values='BPPProtectionClassFactor').reset_index('Protection Class')
            pivotedProtectionPeril['Protection Class Number'] = pivotedProtectionPeril['Protection Class'].apply(self.getProtectionClassValue)
            return pivotedProtectionPeril.sort_values(by=['Protection Class Number']).loc[:, pivotedProtectionPeril.columns != 'Protection Class Number']

    # Builds the masonry veneer factor table for the given coverage (either Building or BPP)
    # Returns a dataframe
    def buildMasonryVeneer(self, coverage):
        masonryVeneerPeril = self.buildDataFrame("BP7_Peril_Masonry_Veneer")
        updatedMasonryVeneerPeril = masonryVeneerPeril.astype({'Masonry_Veneer_Min_Percent': 'int64', 'Masonry_Veneer_Max_Percent': 'int64'}). \
                astype({'Masonry_Veneer_Min_Percent': 'string', 'Masonry_Veneer_Max_Percent': 'string'}) # Converting to int first to get rid of decimal places
        updatedMasonryVeneerPeril["Masonry Veneer Percentage"] = updatedMasonryVeneerPeril["Masonry_Veneer_Min_Percent"] + ' - ' + updatedMasonryVeneerPeril["Masonry_Veneer_Max_Percent"] + '%' # Creating a single column for the percentage
        filteredMasonryVeneer = updatedMasonryVeneerPeril.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredMasonryVeneer.pivot(index='Masonry Veneer Percentage', columns='Peril TypeCode', values='Bldg_Masonry_Veneer_Factor').reset_index('Masonry Veneer Percentage')
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredMasonryVeneer.pivot(index='Masonry Veneer Percentage', columns='Peril TypeCode', values='BPP_Masonry_Veneer_Factor').reset_index('Masonry Veneer Percentage')

    # Builds the building valuation options factor table
    # Returns a dataframe
    def buildValuationBasis(self):
        valuationBasisPeril = self.buildDataFrame("BP7_Peril_ValuationBasis")
        return  valuationBasisPeril.query(f'`Peril TypeCode` == "allperil" & ValuationTypeCode != "ReplacementCost"').filter(items=['ValuationTypeCode', 'ValuationFactor']). \
                rename(columns={'ValuationTypeCode': 'Valuation Type', 'ValuationFactor': 'Factor'}). \
                replace({'Valuation Type': {'ActualCashValue': 'Actual Cash Value – Building', 'FunctionalValuation': 'Functional Building Valuation', 'ReplacementCostExtension': 'Replacement Cost – Extension'}})

    # Builds the automatic increase in building insurance (A.I.I) table
    # Returns a dataframe
    def buildAnnualIncrease(self):
        annualIncreasePeril = self.buildDataFrame("BP7_Peril_AmountOfAnnualIncrease_Factor")
        updatedAnnualIncrease = annualIncreasePeril.astype({'AmountOfAnnualIncrease': 'int32'}).query(f'`Peril TypeCode` == "allperil" & AmountOfAnnualIncrease <= 10').rename(columns={'AmountAnnualIncreaseFactor': 'Factor'})
        filteredAnnualIncrease = updatedAnnualIncrease.astype({'AmountOfAnnualIncrease': 'string'})
        filteredAnnualIncrease['Amount of Annual Increase'] = filteredAnnualIncrease['AmountOfAnnualIncrease'] + '%'
        return filteredAnnualIncrease.filter(items=['Amount of Annual Increase', 'Factor']).set_index('Amount of Annual Increase')

    # Builds the property deductible table
    # Returns a dataframe
    def buildPropertyDeductible(self):
        propertyDedPeril = self.buildDataFrame("BP7_Peril_PropertyDeductible")
        filteredPropertyDeductible = propertyDedPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'DeductibleAmount': 'Property Deductible', 'BPPTIB_AmtofInsurance_Min': 'BPP Min', 'BPPTIB_AmtOfInsurance_Max': 'BPP Max', 'BLDG_AmtOfInsurance': 'Building'})
        pivotedPropertyDeductible = filteredPropertyDeductible.filter(items=['Peril TypeCode', 'Property Deductible', 'BPP Min', 'BPP Max', 'Building', 'PropertyDeductibleFactor']). \
                pivot(index=['Property Deductible', 'BPP Min', 'BPP Max', 'Building'], columns='Peril TypeCode', values='PropertyDeductibleFactor').reset_index(['Property Deductible', 'BPP Min', 'BPP Max', 'Building'])
        sortedPropertyDeductible = pivotedPropertyDeductible.astype({'Property Deductible': 'int32'}).fillna({'BPP Max': '+'}).sort_values(by=['Property Deductible', 'BPP Min', 'BPP Max', 'Building'])
        propertyDeductibleDimensions = sortedPropertyDeductible.shape # A list where the first element represents the number of rows in the df and the second element represents the number of columns
        for row in range(propertyDeductibleDimensions[0]):
            for column in range(4, propertyDeductibleDimensions[1]): # Starting with column 4 because that is the first column with a peril
                if math.isnan(sortedPropertyDeductible.iloc[row, column]):
                    lowFactor = sortedPropertyDeductible.iloc[row - 1, column]
                    highFactor = sortedPropertyDeductible.iloc[row + 1, column]
                    lowBuildingAmount = sortedPropertyDeductible.iloc[row - 1, 3]
                    currentBuildingAmount = sortedPropertyDeductible.iloc[row, 3]
                    highBuildingAmount = sortedPropertyDeductible.iloc[row + 1, 3]
                    if math.isnan(lowFactor): # In case the low factor is missing too
                        lowFactor = sortedPropertyDeductible.iloc[row - 2, column]
                        lowBuildingAmount = sortedPropertyDeductible.iloc[row - 2, 3]
                    if math.isnan(highFactor): # In case the high factor is missing too
                        highFactor = sortedPropertyDeductible.iloc[row + 2, column]
                        highBuildingAmount = sortedPropertyDeductible.iloc[row + 2, 3]
                    sortedPropertyDeductible.iloc[row, column] = lowFactor + (highFactor - lowFactor) * ((currentBuildingAmount - lowBuildingAmount) / (highBuildingAmount - lowBuildingAmount))
        return sortedPropertyDeductible

    # Builds the wind/hail deductible factor table for the given coverage (either Building or BPP)
    # Returns a dataframe
    def buildWHDeductibleFactor(self, coverage):
        whDedPeril = self.buildDataFrame("BP7_Peril_WH_Deductible_Factor")
        if coverage.casefold() == 'building': # Case-insensitive comparison
            filteredWHDeductibleFactor = whDedPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "Building"'). \
                replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Peril TypeCode': 'Peril', 'BPPTIB_AmtofInsurance_Min': 'BPP Min', 'BPPTIB_AmtofInsurance_Max': 'BPP Max', 'BLDG_AmtofInsurance': 'Building'})
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            filteredWHDeductibleFactor = whDedPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "BPP"'). \
                replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Peril TypeCode': 'Peril', 'BPPTIB_AmtofInsurance_Min': 'BPP Min', 'BPPTIB_AmtofInsurance_Max': 'BPP Max', 'BLDG_AmtofInsurance': 'Building'})
        updatedWHDeductibleFactor = filteredWHDeductibleFactor[filteredWHDeductibleFactor['WH_PercentOrAmount'].notnull()] # Filtering out the missing values from the WH_PercentOrAmount column
        sortedWHDeductibleFactor = updatedWHDeductibleFactor.fillna({'BPP Max': '+'}).sort_values(by=['Peril', 'BPP Min', 'BPP Max', 'Building']) # Filling in missing values and sorting
        return sortedWHDeductibleFactor.astype({'WH_PercentOrAmount': 'int32'}).pivot(index=['Peril', 'BPP Min', 'BPP Max', 'Building'], columns='WH_PercentOrAmount', values='WH Factor').reset_index(['Peril', 'BPP Min', 'BPP Max', 'Building'])

    # Builds the wind/hail deductible per building factor table for the given coverage (either Building or BPP)
    # Returns a dataframe
    def buildWHDeductiblePerBuilding(self, coverage):
        whDedBldgPeril = self.buildDataFrame("BP7 Peril_WH_Deductible_Per_Building")
        if coverage.casefold() == 'building': # Case-insensitive comparison
            filteredWHDedBldg = whDedBldgPeril.query(f'Class_Code_Min == 40000 & {"`Peril TypeCode`".strip()} in {self.perils} & Coverage == "Building" & WHDeductibleAmt != "1" & WHDeductibleAmt != "2" & WHDeductibleAmt != "5"'). \
                    replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'WHDeductibleAmt': 'Wind / Hail Deductible Amount', 'AmtOfInsurance': 'Amount of Insurance'})
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            filteredWHDedBldg = whDedBldgPeril.query(f'Class_Code_Min == 40000 & {"`Peril TypeCode`".strip()} in {self.perils} & Coverage == "BPP" & WHDeductibleAmt != "1" & WHDeductibleAmt != "2" & WHDeductibleAmt != "5"'). \
                    replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'WHDeductibleAmt': 'Wind / Hail Deductible Amount', 'AmtOfInsurance': 'Amount of Insurance'})
        return filteredWHDedBldg.pivot(index=['Wind / Hail Deductible Amount', 'Amount of Insurance'], columns='Peril TypeCode', values='WH_Deductible_Building_Factor').reset_index(['Wind / Hail Deductible Amount', 'Amount of Insurance']). \
                astype({'Wind / Hail Deductible Amount': 'int32'}).sort_values(by=['Wind / Hail Deductible Amount', 'Amount of Insurance'])

    # Builds the wind/hail percentage deductible factor table for the given coverage (either Building or BPP)
    # Returns a dataframe
    def buildWHDeductiblePercentage(self, coverage):
        whDedPeril = self.buildDataFrame("BP7 Peril_WH_Deductible_Per_Building")
        whDedPeril['Wind / Hail Deductible'] = whDedPeril['WHDeductibleAmt'] + '%'
        filteredWHDeductiblePercentage = None
        if coverage.casefold() == 'building': # Case-insensitive comparison
            filteredWHDeductiblePercentage = whDedPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "Building" & (WHDeductibleAmt == "1" | WHDeductibleAmt == "2" | WHDeductibleAmt == "5")')

        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            filteredWHDeductiblePercentage = whDedPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "BPP" & (WHDeductibleAmt == "1" | WHDeductibleAmt == "2" | WHDeductibleAmt == "5")')


        # Temporary fix for the new AOI Column in the ratebook, causing their to be a lack of uniqueness for this subset.
        filteredWHDeductiblePercentage = filteredWHDeductiblePercentage.drop_duplicates(subset=['Wind / Hail Deductible', 'Peril TypeCode'])

        return filteredWHDeductiblePercentage.replace({'Peril TypeCode': self.perilsConversions}).pivot(index='Wind / Hail Deductible', columns='Peril TypeCode', values='WH_Deductible_Building_Factor').reset_index('Wind / Hail Deductible')

    # Builds the burglar alarm factor table
    # Returns a dataframe
    def buildBurglarAlarmFactor(self):
        burglarAlarmPeril = self.buildDataFrame("BP7_Peril_Burglar_Alarm_Factor")
        filteredBurglarAlarmPeril = burglarAlarmPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & (`Burglar Alarm Type` == "local" | `Burglar Alarm Type` == "central")'). \
                replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Peril TypeCode': 'Peril'})
        return filteredBurglarAlarmPeril.pivot(index='Peril', columns='Burglar Alarm Type', values='Burglar Alarm Factor').rename(columns={'central': 'Central Station Alarm', 'local': 'Local Alarm'}).reset_index('Peril')

    # Builds the fire alarm factor table
    # Returns a dataframe
    def buildFireAlarmFactor(self):
        fireAlarmPeril = self.buildDataFrame("BP7_Peril_Fire_Alarm_Factor")
        return fireAlarmPeril.query(f'`Peril TypeCode` in {self.perils} & FireAlarmType == "Central Station Fire Alarm"').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Peril TypeCode': 'Peril'}). \
            pivot(index='Peril', columns='Coverage', values='FireAlarmFactor').rename(columns={'Business Income': 'Bus Inc'})

    # Builds the building age modifier table for the given coverage (either Building, BPP, or Business Income)
    # Returns a dataframe
    def buildBuildingAgeModifier(self, coverage):
        bldgAgePeril = self.buildDataFrame("BP7 Peril Building_Age_Modifier")
        if coverage.casefold() == 'building': # Case-insensitive comparison
            filteredBldgAgePeril = bldgAgePeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "Building"')
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            filteredBldgAgePeril = bldgAgePeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "BPP"')
        elif coverage.casefold() == 'business income': # Case-insensitive comparison
            filteredBldgAgePeril = bldgAgePeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Coverage == "Business Income"')
        return filteredBldgAgePeril.replace({'Peril TypeCode': self.perilsConversions, 'Building_Age_Min': {101: '101-1000'}}).rename(columns={'Building_Age_Min': 'Building Age Years'}). \
            pivot(index='Building Age Years', columns='Peril TypeCode', values='BuildingAge_Factor').reset_index('Building Age Years')

    # Builds the building AOI table
    # Returns a dataframe
    def buildBuildingAOI(self):
        aoiBldgPeril = self.buildDataFrame("BP7_Peril_Building_Amt_Insurance")
        filteredAOIBldgPeril = aoiBldgPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & Building_Limit < 10000000').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Building_Limit': 'Lower Bound'}). \
            pivot(index='Lower Bound', columns='Peril TypeCode', values='AmountOfInsuranceFactor').reset_index('Lower Bound')
        filteredAOIBldgPeril['Upper Bound'] = filteredAOIBldgPeril['Lower Bound'] - 1 # Creating the upper bound column
        filteredAOIBldgPeril['Upper Bound'] = filteredAOIBldgPeril['Upper Bound'].shift(periods=-1, fill_value='and over') # Shifting the upper bound column to align it with lower bound
        cols = filteredAOIBldgPeril.columns.tolist()
        cols = cols[0:1] + cols[-1:] + cols[1:-1] # Rearranging the order of columns
        return filteredAOIBldgPeril[cols]

    # Builds the BPP AOI table
    # Returns a dataframe
    def buildBPPAOI(self):
        aoiBPPPeril = self.buildDataFrame("BP7_Peril_BPP_Amt_Insurance")
        filteredAOIBPPPeril = aoiBPPPeril.query(f'Class_Code_Min == 40000 & `Peril TypeCode` in {self.perils} & BPP_Limit < 10000000').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'BPP_Limit': 'Lower Bound'}). \
            pivot(index='Lower Bound', columns='Peril TypeCode', values='AmountOfInsuranceFactor').reset_index('Lower Bound')
        filteredAOIBPPPeril['Upper Bound'] = filteredAOIBPPPeril['Lower Bound'] - 1 # Creating the upper bound column
        filteredAOIBPPPeril['Upper Bound'] = filteredAOIBPPPeril['Upper Bound'].shift(periods=-1, fill_value='and over') # Shifting the upper bound column to align it with lower bound
        cols = filteredAOIBPPPeril.columns.tolist()
        cols = cols[0:1] + cols[-1:] + cols[1:-1] # Rearranging the order of columns
        return filteredAOIBPPPeril[cols]

    # Builds the blanket insurance factor table
    # Returns a dataframe
    def buildBlanketInsuranceFactor(self):
        blanketInsuranceFactor = self.buildDataFrame("BP7_Peril_Blanket_Insurance_Ind")
        return blanketInsuranceFactor.query(f'`Peril TypeCode` == "allperil" & BlanketInsuranceIndicator == "Y"').rename(columns={'BlanketInsuranceFactor': 'Factor'}).filter(items=['Factor'])

    # Builds the building code effectiveness grade (BCEG) table
    # Returns a dataframe
    def buildBCEG(self):
        bceg = self.buildDataFrame("BP7_Peril_BCEG_Factor")
        filteredBCEG = bceg.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions})
        filteredBCEG['Grade'] = filteredBCEG['Tier Grade'] # Creating a new column for the grade and using the old one for sorting
        updatedBCEG = filteredBCEG.astype({'Tier Grade': 'int32'}).replace({'Grade': {'98': 'Non-Particip', '99': 'Ungraded'}})
        if self.state in self.buildingCodes: # Multiple building codes, so different table format
            updatedBCEG['Territory Group'] = updatedBCEG['Territory Code'].apply(self.getBuildingCode)
            finalBCEG = updatedBCEG.pivot(index=['Territory Group', 'Territory Code', 'Grade', 'Tier Grade'], columns='Peril TypeCode', values='BCEGFactor'). \
                    reset_index(['Territory Group', 'Territory Code', 'Grade', 'Tier Grade']).groupby(by=['Territory Group', 'Grade', 'Tier Grade']).mean(numeric_only=True). \
                    reset_index(['Territory Group', 'Grade', 'Tier Grade']).sort_values(by=['Territory Group', 'Tier Grade'])
            del finalBCEG['Tier Grade'] # Deleting the extra column used for sorting
            finalBCEG['ISO Territory'] = finalBCEG['Territory Group'].apply(self.getTerritoryCodes) # Creating a column for the territory
            cols = finalBCEG.columns.tolist()
            cols = cols[0:1] + cols[-1:] + cols[1:-1] # Rearranging the order of columns
            return finalBCEG[cols]
        finalBCEG = updatedBCEG.pivot(index=['Grade', 'Tier Grade'], columns='Peril TypeCode', values='BCEGFactor').reset_index(['Grade', 'Tier Grade']).sort_values(by=['Tier Grade'])
        del finalBCEG['Tier Grade'] # Deleting the extra column used for sorting
        return finalBCEG

    # Builds the tenants improvements and betterments factor table
    # Returns a dataframe
    def buildTenantsImprovements(self):
        tenantsImprovementsPeril = self.buildDataFrame("BP7_Peril_Tenants_Improvements_and_Betterments_Factor")
        return tenantsImprovementsPeril.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'Peril TypeCode': 'Peril', 'TIBFactor': 'Factor'}).filter(items=['Peril', 'Factor']).sort_values(by='Peril')

    # Builds the equipment breakdown limit relativity table
    # Returns a dataframe
    def buildEBLimitRelativity(self):
        ebLimitRelativity = self.buildDataFrame("BP7_EBLimitsRelativityModifier")
        return ebLimitRelativity.query(f'Class_Code_Min == 40000').filter(items=['TotalPropertyLimitMin', 'TotalPropertyLimitMax', 'LimitRelativityModifier']). \
                rename(columns={'TotalPropertyLimitMin': 'Total Property Limit Min', 'TotalPropertyLimitMax': 'Total Property Limit Max', 'LimitRelativityModifier': 'Limit Relativity Modifier'}).fillna(value={'Total Property Limit Max': 'and over'})

    # Builds the equipment breakdown deductible factor table
    # Returns a dataframe
    def buildEBDeductibleFactor(self):
        ebDeductibleFactor = self.buildDataFrame("BP7_EBDeductibleFactor")
        return ebDeductibleFactor.query(f'Class_Code_Min == 40000').filter(items=['DeductibleAmt', 'Factor']).rename(columns={'DeductibleAmt': 'Deductible Amount'})

    # Builds the medical payments factor table
    # Returns a dataframe
    def buildMedicalPaymentsFactor(self):
        medicalPaymentsPeril = self.buildDataFrame("BP7_Peril_Medical_Payments_Decreased_Limit")
        return medicalPaymentsPeril.query(f'`Peril TypeCode` == "liability1"').filter(items=['MedPayLimit', 'MedPayLimitFactor']).replace({'MedPayLimit': {0: 'Excluded'}}). \
                rename(columns={'MedPayLimit': 'Medical Payments Limit', 'MedPayLimitFactor': 'Factor'})

    # Calculates the "value" of each protection class by multiplying the class number by 200 and adds the Unicode code of the letter if there is one
    # Used for sorting the protection class table
    # Returns the value associated with the given protection class
    def getProtectionClassValue(self, protectionClass):
        if len(protectionClass) == 3:
            return int(protectionClass[:-1]) * 200 + ord(protectionClass[-1])
        elif len(protectionClass) == 2:
            if protectionClass == '10':
                return int(protectionClass) * 200
            return int(protectionClass[:-1]) * 200 + ord(protectionClass[-1])
        return int(protectionClass) * 200

    # Returns the building code (A, B, C or D) from the given territory code
    # Varies by state
    def getBuildingCode(self, territoryCode):
        for buildingCode in self.buildingCodes[self.state]:
            if territoryCode in self.buildingCodes[self.state][buildingCode]:
                return buildingCode
    
    # Returns the list of territories associated with the given building code separated by a comma
    def getTerritoryCodes(self, buildingCode):
        return ', '.join(self.buildingCodes[self.state][buildingCode])

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    def formatSprinklerFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        ws.column_dimensions['B'].width = self.pixelsToInches(68)
        ws.column_dimensions['C'].width = self.pixelsToInches(47)

    def formatProtectionClass(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(131)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    def formatMasonryVeneer(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    def formatBuildingValuation(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(229)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)

    def formatAnnualIncrease(self, ws, font):
        ws.column_dimensions['A'].width = self.pixelsToInches(208)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        # Adding a footnote for the A.I.I. table
        ws['A16'] = 'For each additional 1%, add 0.005'
        ws['A16'].font = font
        ws['A16'].alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False) # Aligns the text at the bottom of the table
        ws['A16'].border = Border(left=None, right=None, top=None, bottom=None) # Removing the border for the footnote of the table

    def formatPropertyDeductible(self, ws, boldFont):
        ws.insert_rows(3)
        ws['B3'] = 'Amount of Insurance'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('B3:D3')
        ws.merge_cells('E3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col <= 3:
                ws.column_dimensions[char].width = self.pixelsToInches(74)
            elif col == 4:
                ws.column_dimensions[char].width = self.pixelsToInches(105) # Unique column width for column D
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(53)
            for row in range(5, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 5: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-D

    def formatWHDeductibleFactor(self, ws, boldFont):
        ws.insert_rows(3)
        ws['B3'] = 'Amount of Insurance'
        ws['E3'] = 'Wind-Hail Deductible'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('B3:D3')
        ws.merge_cells('E3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col <= 3:
                ws.column_dimensions[char].width = self.pixelsToInches(74)
            elif col == 4:
                ws.column_dimensions[char].width = self.pixelsToInches(105) # Unique column width for column D
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(62)
            for row in range(5, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 5: # Applying currency formatting to columns A-D
                    cell.number_format = self.currencyFormat 

    def formatWHDeductiblePerBuilding(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col == 1:
                ws.column_dimensions[char].width = self.pixelsToInches(222)
            elif col == 2:
                ws.column_dimensions[char].width = self.pixelsToInches(159)
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(70)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 3: # Applying currency formatting to columns A and B
                    cell.number_format = self.currencyFormat

    def formatWHDeductiblePercentage(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col == 1:
                ws.column_dimensions[char].width = self.pixelsToInches(159)
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(70)

    def formatBurglarAlarm(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        ws.column_dimensions['B'].width = self.pixelsToInches(166)
        ws.column_dimensions['C'].width = self.pixelsToInches(96)

    def formatFireAlarm(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(91)

    def formatBuildingAge(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col == 1: # Applying unique column width and no decimal formatting to column A
                ws.column_dimensions[char].width = self.pixelsToInches(145)
                for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(53)

    def formatAOI(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Building Limit'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col <= 2:
                ws.column_dimensions[char].width = self.pixelsToInches(82)
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A and B
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(53)

    # TODO: This needs to be updated to reflect states with multiple building codes
    def formatBCEG(self, ws, boldFont):
        if self.state in self.buildingCodes: # Format for multiple building codes in the state
            ws.column_dimensions['A'].width = self.pixelsToInches(73) # Territory Group column
            ws.column_dimensions['B'].width = self.pixelsToInches(66) # Territory column
            ws.column_dimensions['C'].width = self.pixelsToInches(82) # Grade column
            for col in range(4, ws.max_column + 1):
                char = get_column_letter(col) # Letter representing the current column
                ws.column_dimensions[char].width = self.pixelsToInches(53)
        else: # Format for a single building code in the state
            ws.insert_rows(3)
            ws['B3'] = 'Entire State'
            for cell in ws['3:3']:
                cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                    right=Side(border_style='thin', color='C1C1C1'), 
                                    top=Side(border_style='thin', color='C1C1C1'), 
                                    bottom=Side(border_style='thin', color='C1C1C1'))
                cell.font = boldFont
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            ws.merge_cells('B3:' + get_column_letter(ws.max_column) + '3')
            ws.print_title_rows = '1:4'
            for col in range(1, ws.max_column + 1):
                char = get_column_letter(col) # Letter representing the current column
                if col == 1:
                    ws.column_dimensions[char].width = self.pixelsToInches(82)
                else:
                    ws.column_dimensions[char].width = self.pixelsToInches(53) 

    def formatTIB(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        ws.column_dimensions['B'].width = self.pixelsToInches(68)

    def formatEBLimitRelativity(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Total Property Limit'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(180)
            if col < 3:
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A and B                

    def formatEBDeductibleFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(68)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = self.currencyFormat # Applying currency format to column A

    def formatMedicalPayments(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(180)
        ws.column_dimensions['B'].width = self.pixelsToInches(68)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = self.noDecimalFormat # Applying no decimal formatting for column A

    # Sets up the All Programs Excel file and creates a separate worksheet for each of the given dataframes
    # Returns the Excel file
    def buildAllProgramsPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        AllPrograms = ExcelSettingsCurrent.Excel(state=self.state, programName='All Programs', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = AllPrograms.getFontName()
        fontSize = AllPrograms.getFontSize()

        AllPrograms.generateWorksheet('SPR', 'AS, FS, H, O, R, S, W Table 3.C.2.a. Sprinkler Factor', self.buildSprinklerFactor(), False, True)
        AllPrograms.generateWorksheet('PCBG', 'AS, FS, H, O, R, S, W Table 3.C.2.b. Protection Class Factor - Building', self.buildProtectionClass('Building'), False, True)
        AllPrograms.generateWorksheet('PCPP', 'AS, FS, H, O, R, S, W Table 3.C.2.b. Protection Class Factor - BPP', self.buildProtectionClass('BPP'), False, True)
        AllPrograms.generateWorksheet('MVBG', 'AS, FS, H, O, R, S, W Table 3.C.2.c.1. Masonry Veneer Factor - Building', self.buildMasonryVeneer('Building'), False, True)
        AllPrograms.generateWorksheet('MVPP', 'AS, FS, H, O, R, S, W Table 3.C.2.c.1. Masonry Veneer Factor - BPP', self.buildMasonryVeneer('BPP'), False, True)
        AllPrograms.generateWorksheet('BV', 'AS, FS, H, O, R, S, W Table 3.C.2.d. Building Valuation Options Factor', self.buildValuationBasis(), False, True)
        AllPrograms.generateWorksheet('AIBI', 'AS, FS, H, O, R, S, W Table 3.C.2.e. Automatic Increase In Building Insurance (A.I.I.) Factor', self.buildAnnualIncrease(), True, True)
        AllPrograms.generateWorksheet('PD', 'AS, FS, H, O, R, S, W Table 3.C.2.f. Property Deductible Factor', self.buildPropertyDeductible(), False, True)
        AllPrograms.generateWorksheet('WHOBG', 'AS, FS, H, O, R, S, W Table 3.C.2.g.(1). Windstorm or Hail Deductible Factor - Per Occurrence Fixed Deductible Amount - Building', self.buildWHDeductibleFactor('Building'), False, True)
        AllPrograms.generateWorksheet('WHOPP', 'AS, FS, H, O, R, S, W Table 3.C.2.g.(1). Windstorm or Hail Deductible Factor - Per Occurrence Fixed Deductible Amount - BPP', self.buildWHDeductibleFactor('BPP'), False, True)
        AllPrograms.generateWorksheet('WHBBG', 'AS, FS, H, O, R, S, W Table 3.C.2.g.(2). Windstorm or Hail Deductible Factor - Per Building Fixed Deductible Amount - Building', self.buildWHDeductiblePerBuilding('Building'), False, True)
        AllPrograms.generateWorksheet('WHBPP', 'AS, FS, H, O, R, S, W Table 3.C.2.g.(2). Windstorm or Hail Deductible Factor - Per Building Fixed Deductible Amount - BPP', self.buildWHDeductiblePerBuilding('BPP'), False, True)
        AllPrograms.generateWorksheet('WHPBG', 'AS, FS, H, O, R, S, W Table 3.C.2.g.(3). Windstorm or Hail Deductible Factor - Percentage Deductible - Building', self.buildWHDeductiblePercentage('Building'), False, True)
        AllPrograms.generateWorksheet('WHPPP', 'AS, FS, H, O, R, S, W Table 3.C.2.g.(3). Windstorm or Hail Deductible Factor - Percentage Deductible - BPP', self.buildWHDeductiblePercentage('BPP'), False, True)
        AllPrograms.generateWorksheet('BA', 'AS, FS, H, O, R, S, W Table 3.C.2.h. Burglar Alarm Modifier', self.buildBurglarAlarmFactor(), False, True) 
        AllPrograms.generateWorksheet('CSFA', 'AS, FS, H, O, R, S, W Table 3.C.2.i. Central Station Fire Alarm Modifier', self.buildFireAlarmFactor(), True, True)
        AllPrograms.generateWorksheet('BABG', 'AS, FS, H, O, R, S, W Table 3.C.2.j. Building Age Modifier - Building', self.buildBuildingAgeModifier('Building'), False, True)
        AllPrograms.generateWorksheet('BAPP', 'AS, FS, H, O, R, S, W Table 3.C.2.j. Building Age Modifier - BPP', self.buildBuildingAgeModifier('BPP'), False, True)
        AllPrograms.generateWorksheet('BABI', 'FS Table 3.C.2.j. Building Age Modifier - Bus Inc', self.buildBuildingAgeModifier('Business Income'), False, True) # Business Income is only applicable in the food program
        AllPrograms.generateWorksheet('AIBG', 'AS, FS, H, O, R, S, W Table 3.C.2.l.(1). AOI (Amount of Insurance) Relativity Factor - Building', self.buildBuildingAOI(), False, True)
        AllPrograms.generateWorksheet('AIPP', 'AS, FS, H, O, R, S, W Table 3.C.2.l.(1). AOI (Amount of Insurance) Relativity Factor - BPP', self.buildBPPAOI(), False, True)
        AllPrograms.generateWorksheet('BKT', 'H Table 3.C.2.m., AS, FS, O, R, S, W Table 3.C.2.n. Blanket Insurance Factor', self.buildBlanketInsuranceFactor(), False, True)
        AllPrograms.generateWorksheet('BCEG', 'H Table 3.C.2.n., AS, FS, O, R, S, W Table 3.C.2.o. Building Code Effectiveness Grading', self.buildBCEG(), False, True)
        AllPrograms.generateWorksheet('TIB', 'H Table 3.C.2.p., AS, FS, O, R, S, W Table 3.C.2.q. Tenants Improvements and Betterments Factor', self.buildTenantsImprovements(), False, True)
        AllPrograms.generateWorksheet('EBL', 'AS, FS, H, O, R, S, W Table 3.C.3.c. Equipment Breakdown Limits Relativity Modifier', self.buildEBLimitRelativity(), False, True)
        AllPrograms.generateWorksheet('EBD', 'AS, FS, H, O, R, S, W Table 3.C.3.d. Equipment Breakdown Deductible Factor', self.buildEBDeductibleFactor(), False, True)
        AllPrograms.generateWorksheet('MD', 'FS, O, S, W Table 3.C.4.c., R Table 3.C.4.d., AS, H Table 3.C.4.e. Medical Payments Factor', self.buildMedicalPaymentsFactor(), False, True)

        AllPrograms.createIndex()
        AllProgramsPages = AllPrograms.getWB()

        self.formatSprinklerFactor(AllProgramsPages['SPR'])
        self.formatProtectionClass(AllProgramsPages['PCBG'])
        self.formatProtectionClass(AllProgramsPages['PCPP'])
        self.formatMasonryVeneer(AllProgramsPages['MVBG'])
        self.formatMasonryVeneer(AllProgramsPages['MVPP'])
        self.formatBuildingValuation(AllProgramsPages['BV'])
        self.formatAnnualIncrease(AllProgramsPages['AIBI'], Font(name=fontName, size=fontSize))
        self.formatPropertyDeductible(AllProgramsPages['PD'], Font(name=fontName, size=fontSize, bold=True))
        self.formatWHDeductibleFactor(AllProgramsPages['WHOBG'], Font(name=fontName, size=fontSize, bold=True))
        self.formatWHDeductibleFactor(AllProgramsPages['WHOPP'], Font(name=fontName, size=fontSize, bold=True))
        self.formatWHDeductiblePerBuilding(AllProgramsPages['WHBBG'])
        self.formatWHDeductiblePerBuilding(AllProgramsPages['WHBPP'])
        self.formatWHDeductiblePercentage(AllProgramsPages['WHPBG'])
        self.formatWHDeductiblePercentage(AllProgramsPages['WHPPP'])
        self.formatBurglarAlarm(AllProgramsPages['BA'])
        self.formatFireAlarm(AllProgramsPages['CSFA'])
        self.formatBuildingAge(AllProgramsPages['BABG'])
        self.formatBuildingAge(AllProgramsPages['BAPP'])
        self.formatBuildingAge(AllProgramsPages['BABI'])
        self.formatAOI(AllProgramsPages['AIBG'], Font(name=fontName, size=fontSize, bold=True))
        self.formatAOI(AllProgramsPages['AIPP'], Font(name=fontName, size=fontSize, bold=True))
        self.formatBCEG(AllProgramsPages['BCEG'], Font(name=fontName, size=fontSize, bold=True))
        self.formatTIB(AllProgramsPages['TIB'])
        self.formatEBLimitRelativity(AllProgramsPages['EBL'], Font(name=fontName, size=fontSize, bold=True))
        self.formatEBDeductibleFactor(AllProgramsPages['EBD'])
        self.formatMedicalPayments(AllProgramsPages['MD'])

        return AllProgramsPages