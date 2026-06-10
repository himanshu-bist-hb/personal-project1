# This module formats the Hab State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

class Hab:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.habProgramCode = 10000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        propertyBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7PerilBldgHabitationalSwimmingPoolsPropertyBaseRate'][1:], index=None, columns=self.rateTables[company]['BP7PerilBldgHabitationalSwimmingPoolsPropertyBaseRate'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        filteredPropertyBaseRates = propertyBaseRates.query(f'`Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"'). \
                    pivot(index='Peril TypeCode', columns='CoverageCode', values='BaseRate').reset_index().rename_axis(None, axis=1). \
                    rename(columns={'POOL': 'Swimming Pool', 'SPA': 'Spa', 'FENCE': 'Fence'})
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        mergedBaseRates = pd.merge(baseRates, filteredPropertyBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(mergedBaseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building", 
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')
    
    # Builds the liability charges for related additional exposures table
    # Returns a dataframe
    def buildRelatedAddtExposures(self, company):
        relatedAddtExposures = pd.DataFrame(data=self.rateTables[company]['BP7LiabilityChargesForRelatedAddnlExposures'][1:], index=None, columns=self.rateTables[company]['BP7LiabilityChargesForRelatedAddnlExposures'][0])
        return relatedAddtExposures.rename(columns={'ExposureType': 'Exposure', 'BaseRate': 'Rate'}). \
                replace({'Exposure': {'Clubhouse': 'a. Per Clubhouse',
                                      'ExerciseRoom': 'b. Per Exercise Room',
                                      'Playground': 'c. Per Playground',
                                      'Spa': 'd. Per Spa',
                                      'SwimmingPool': 'e. Per Swimming Pool'}})
    
    # Builds the territory multiplier table for the given coverage (either building, bpp, or liability)
    # Returns a dataframe
    #def buildTerritoryMultiplier(self, coverage):
    #    territorialFactor = self.buildDataFrame("BP7_Peril_TerritorialFactor")
    #    filteredTerritorialFactor = territorialFactor.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'TerritoryCode': 'Territory'})
    #    if coverage.casefold() == 'building': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BldgTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'bpp': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BPPTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'liability': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='LiabilityTerritoryFactor').reset_index('Territory')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'}) # Converting to int first to get rid of decimal places
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0', 
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe    
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.habProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the number of units factor table
    # Returns a dataframe
    def buildNumOfUnitsFactor(self):
        numberOfUnits = self.buildDataFrame("BP7_Peril_Liability_No_Of_Units_Factor")
        filteredNumberOfUnits = numberOfUnits.fillna({'NoOfUnits_Max' : 0}). \
                astype({'NoOfUnits_Min': 'int64', 'NoOfUnits_Max' : 'int64'}).astype({'NoOfUnits_Min': 'string', 'NoOfUnits_Max': 'string'}) # Converting to int first to get rid of decimal places
        filteredNumberOfUnits['Units'] = np.where(filteredNumberOfUnits['NoOfUnits_Max'] == '0', # Creating a new column for units
                                                  'Over 75',
                                                  filteredNumberOfUnits['NoOfUnits_Min'] + ' - ' + filteredNumberOfUnits['NoOfUnits_Max'])
        return filteredNumberOfUnits.replace({'Peril TypeCode': self.perilsConversions}).replace({'Units': {'1 - 1': '01 - 01'}}).replace({'Units': {'2 - 3': '02 - 03'}}).replace({'Units': {'4 - 10': '04 - 10'}}).rename(columns={'LiabilityNoOfUnitsFactor': 'Factor'}).pivot(index='Units', columns='Peril TypeCode', values='Factor').reset_index('Units')
    
    # Builds the number of stories factor table
    # Returns a dataframe
    def buildNumOfStoriesFactor(self):
        numberOfStories = self.buildDataFrame("BP7_Peril_Liability_No_Of_Stories_Factor")
        filteredNumberOfStories = numberOfStories.query(f'NoOfStories == NoOfStories').replace({'NoOfStories': {'4': '4 or More'}})
        return filteredNumberOfStories.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'NoOfStories': 'Number of Stories', 'LiabilityNoOfStoriesFactor': 'Factor'}).pivot(index='Number of Stories', columns='Peril TypeCode', values='Factor').reset_index('Number of Stories')

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.habProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.habProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance' : 'int32'})
    
    # Builds the directors and officers liability insurance table
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Habitational"').copy() # Getting a new copy of the data here
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 to 100'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 101, 'Number of Units'] = '101 to 250'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 251, 'Number of Units'] = '251 to 500'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 501, 'Number of Units'] = 'over 500'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])
    
    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Habitational"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])
    
    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Habitational"').copy() # Getting a new copy of the data here
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index = filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.habProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Habitational premises'})

    def buildHabExclusion(self):
        data = pd.DataFrame({
            "Factor" : ["0.98"]
        })

        return data

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the base rates worksheet
    def formatBaseRates(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(120)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = '#,##0.0000' # 4 values after the decimal point for the base rates

    # Applies manual formatting to the liability charges for related additional exposures worksheet
    def formatRelatedAddtExposures(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        for row in range(4, ws.max_row + 1):
            ws['B' + str(row)].number_format = '#,##0.000' # 2 values after decimal point

    # Applies manual formatting to the territory multiplier worksheet
    #def formatTerritoryMultiplier(self, ws):
    #    ws.column_dimensions['A'].width = self.pixelsToInches(70)
    #    for col in range(2, ws.max_column + 1):
    #        char = get_column_letter(col) # Letter representing the current column
    #        ws.column_dimensions[char].width = self.pixelsToInches(54)  

    # Applies manual formatting to the construction factor worksheet
    def formatConstructionFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    # Applies manual formatting to the year built modifier worksheet
    def formatYearBuiltModifier(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(131)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '###0'

    # Applies manual formatting to the equipment breakdown base rate worksheet
    def formatEBBaseRate(self, ws):
        ws['A4'].number_format = '$#,##0.00'
    
    # Applies manumal formatting to the number of stories factor worksheet
    def formatNumOfStoriesFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)

    # Applies manual formatting to the property damage deductible worksheet
    def formatPropertyDamageDeductible(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(187)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the liability limit factor worksheet
    def formatLiabilityLimitFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat
    
    def formatDirsOfficersLiabIns(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(130)
        ws.column_dimensions['D'].width = self.pixelsToInches(140)
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        if self.state == 'WA':
            ws.merge_cells('A10:A11')
            ws.merge_cells('A12:A13')
            ws.merge_cells('A14:A15')
        else:
            ws.merge_cells('A10:A12')
            ws.merge_cells('A13:A15')
            ws.merge_cells('A16:A18')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['B' + str(row)].number_format = self.noDecimalFormat
            ws['C' + str(row)].number_format = '$#,##0.00'
            ws['D' + str(row)].number_format = '$#,##0.00'

    # Applies manual formatting to the directors and officers liability insurance - non-monetary relief
    def formatDirsOfficersNonMonetaryRelief(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(225)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat
            ws['B' + str(row)].number_format = '$#,##0.00'

    def formatDirsOfficersReportingPeriods(self, ws):
        ws.column_dimensions['B'].width = self.pixelsToInches(215)

    # Applies manual formatting to the endorsement charge worksheet
    def formatEndorsementCharge(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(350)
        ws['A4'].number_format = '$#,##0.00'

    def formatHabExclusion(self, ws):
        ws.insert_rows(2,2)
        ws["A3"] = "Multiply the factor below to adjust for the exclusion"

    # Sets up the Hab Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildHabPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        Hab = ExcelSettings.Excel(state=self.state, programName='Habitational', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        if 'NACO' in self.rateTables.keys():
            Hab.generateWorksheet('BRNACO', 'H Table 3.B.1. NW Assurance State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NAFF' in self.rateTables.keys():
            Hab.generateWorksheet('BRNAFF', 'H Table 3.B.1. NW Affinity State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NGIC' in self.rateTables.keys():
            Hab.generateWorksheet('BRNGIC', 'H Table 3.B.1. NW General Insurance Company', self.buildBaseRates('NGIC'), False, True)
        if 'NICOF' in self.rateTables.keys():
            Hab.generateWorksheet('BRNICOF', 'H Table 3.B.1. NICOF State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NACO' in self.rateTables.keys():
            Hab.generateWorksheet('LANACO', 'H Table 3.C.5. Liability Charges for Related Additional Exposures', self.buildRelatedAddtExposures('NGIC'), False, True)
        if 'NAFF' in self.rateTables.keys():
            Hab.generateWorksheet('LANAFF', 'H Table 3.C.5. Liability Charges for Related Additional Exposures', self.buildRelatedAddtExposures('NGIC'), False, True)
        if 'NGIC' in self.rateTables.keys():
            Hab.generateWorksheet('LANGIC', 'H Table 3.C.5. Liability Charges for Related Additional Exposures', self.buildRelatedAddtExposures('NGIC'), False, True)
        if 'NICOF' in self.rateTables.keys():
            Hab.generateWorksheet('LANICOF', 'H Table 3.C.5. Liability Charges for Related Additional Exposures', self.buildRelatedAddtExposures('NGIC'), False, True)
        #Hab.generateWorksheet('TRBG', 'H Table 3.C.1.a. State Territory Multiplier - Building', self.buildTerritoryMultiplier('Building'), False, True)
        #Hab.generateWorksheet('TRPP', 'H Table 3.C.1.a. State Territory Multiplier - BPP', self.buildTerritoryMultiplier('BPP'), False, True)
        #Hab.generateWorksheet('TRLB', 'H Table 3.C.1.a. State Territory Multiplier - Liability', self.buildTerritoryMultiplier('Liability'), False, True)
        Hab.generateWorksheet('CBG', 'H Table 3.C.2.c. Construction Factor - Building', self.buildConstructionType('Building'), False, True)
        Hab.generateWorksheet('CPP', 'H Table 3.C.2.c. Construction Factor - BPP', self.buildConstructionType('BPP'), False, True)
        Hab.generateWorksheet('YBBG', 'H Table 3.C.2.o. Year Built Modifier - Building', self.buildYearBuiltModifier('Building'), False, True)
        Hab.generateWorksheet('YBPP', 'H Table 3.C.2.o. Year Built Modifier - BPP', self.buildYearBuiltModifier('BPP'), False, True)
        Hab.generateWorksheet('EBB', 'H Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate(), False, True)
        Hab.generateWorksheet('NU', 'H Table 3.C.4.a. Number of Units Factor', self.buildNumOfUnitsFactor(), False, True)
        Hab.generateWorksheet('NS', 'H Table 3.C.4.b. Number of Stories Factor', self.buildNumOfStoriesFactor(), False, True)
        Hab.generateWorksheet('PDLD', 'H Table 3.C.4.d. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount(), False, True)
        Hab.generateWorksheet('LL', 'H Table 3.C.4.f. Liability Limit Factor', self.buildLiabilityLimitFactor(), False, True)
        Hab.generateWorksheet('DO', 'H Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns(), False, True)
        Hab.generateWorksheet('DONM', 'H Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief(), False, True)
        Hab.generateWorksheet('ERP', 'H Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods(), False, True)
        Hab.generateWorksheet('PLUS', 'H Table 4.B. Habitational PLUS Endorsement', self.buildEndorsementCharge(), False, True)

        if self.state == "CA":
            Hab.generateWorksheet('HABEX', 'H Table 4.C Habitability Exclusion', self.buildHabExclusion(), False, True)

        Hab.createIndex()
        HabPages = Hab.getWB()

        if 'NACO' in self.rateTables.keys():
            self.formatBaseRates(HabPages['BRNACO'])
        if 'NAFF' in self.rateTables.keys():
            self.formatBaseRates(HabPages['BRNAFF'])
        if 'NGIC' in self.rateTables.keys():
            self.formatBaseRates(HabPages['BRNGIC'])
        if 'NICOF' in self.rateTables.keys():
            self.formatBaseRates(HabPages['BRNICOF'])
        if 'NACO' in self.rateTables.keys():
            self.formatRelatedAddtExposures(HabPages['LANACO'])
        if 'NAFF' in self.rateTables.keys():
            self.formatRelatedAddtExposures(HabPages['LANAFF'])
        if 'NGIC' in self.rateTables.keys():
            self.formatRelatedAddtExposures(HabPages['LANGIC'])
        if 'NICOF' in self.rateTables.keys():
            self.formatRelatedAddtExposures(HabPages['LANICOF'])
        #self.formatTerritoryMultiplier(HabPages['TRBG'])
        #self.formatTerritoryMultiplier(HabPages['TRPP']) 
        #self.formatTerritoryMultiplier(HabPages['TRLB'])
        self.formatConstructionFactor(HabPages['CBG'])  
        self.formatConstructionFactor(HabPages['CPP'])  
        self.formatYearBuiltModifier(HabPages['YBBG'])
        self.formatYearBuiltModifier(HabPages['YBPP'])  
        self.formatEBBaseRate(HabPages['EBB'])
        self.formatNumOfStoriesFactor(HabPages['NS'])
        self.formatPropertyDamageDeductible(HabPages['PDLD'])
        self.formatLiabilityLimitFactor(HabPages['LL'])
        self.formatDirsOfficersLiabIns(HabPages['DO'])
        self.formatDirsOfficersNonMonetaryRelief(HabPages['DONM'])
        self.formatDirsOfficersReportingPeriods(HabPages['ERP'])
        self.formatEndorsementCharge(HabPages['PLUS'])

        if self.state == "CA":
            self.formatHabExclusion(HabPages['HABEX'])

        return HabPages

