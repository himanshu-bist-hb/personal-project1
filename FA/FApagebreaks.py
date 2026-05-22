"""
FA/FApagebreaks.py
==================
Page-break / print-settings post-processing for Farm Auto rate pages.

By default FA uses the IDENTICAL page-break rules as BA.  FA_SHEET_RULES starts
as a copy of BA's rules so any future FA-only rules can be added here without
touching BA code.

HOW TO ADD AN FA-SPECIFIC PAGE-BREAK RULE
------------------------------------------
Step 1 – Write a handler function in this file:

    def _handle_fa_rule_xxx(ws, dest_filename):
        ws.print_area = f"A1:H{ws.max_row}"
        disable_fit_to_page(ws)
        add_break_after(ws, 30)

Step 2 – Append it to FA_SHEET_RULES below:

    FA_SHEET_RULES.append(("Rule FA XXX", _handle_fa_rule_xxx))

    (Add more-specific prefixes BEFORE less-specific ones if needed.)

Step 3 – Done.  FARatePages.py already calls this module's process_pagebreaks,
         which walks FA_SHEET_RULES automatically.

HELPERS (imported from BA for use in any handlers you add here)
-------------------------------------------------------------------
  fit_single_page(ws)       – fit entire sheet onto one printed page
  fit_width_only(ws)        – fit width to 1 page; height grows with content
  disable_fit_to_page(ws)   – turn off fit-to-page; manual breaks rule
  add_break_after(ws, row)  – add a horizontal page break AFTER the given row

OPENPYXL API CHEATSHEET (same as BApagebreaks.py)
  ws.page_setup.orientation     = "portrait" or "landscape"
  ws.print_area                 = "A1:H50"
  ws.print_title_rows           = "1:3"
  ws.page_margins.top           = 1.00   (inches)
"""

import os
import openpyxl

from BA.BApagebreaks import (
    SHEET_RULES as _BA_SHEET_RULES,
    fit_single_page,
    fit_width_only,
    disable_fit_to_page,
    add_break_after,
    _sanitize_xlsx,
    export_to_pdf,        # re-exported: FARatePages.py imports this
)

# ---------------------------------------------------------------------------
# FA_SHEET_RULES: start with all BA rules, then append FA-only rules below.
# ---------------------------------------------------------------------------
FA_SHEET_RULES = list(_BA_SHEET_RULES)

# ── Add FA-only rules here ──────────────────────────────────────────────────

def _handle_fa_rule_223c2(ws, dest_filename):
    # Rule 223 C2 needs its own handler so it doesn't match "Rule 223 C"
    # (which would inherit BA's fit_single_page with default margins).
    # Tighter margins eliminate the large white gap and give ~65-70% scale,
    # making the 11pt font appear as ~7-8pt — far more readable than the
    # previous ~35% scale with 10pt font.
    ws.page_setup.orientation = "portrait"
    ws.page_margins.header = 0.5    # standard; right header is 3 lines tall at 10pt
    ws.page_margins.top    = 1.0    # must be > header + header_text_height (~0.75")
    ws.page_margins.bottom = 0.35   # tighter than standard to reclaim space at bottom
    ws.page_margins.footer = 0.25
    ws.page_margins.left   = 0.25   # restore standard (was 0.5, wider than needed)
    ws.page_margins.right  = 0.25
    fit_single_page(ws)

# Insert before "Rule 223 C" so the more-specific prefix wins.
_223c_idx = next(i for i, (p, _) in enumerate(FA_SHEET_RULES) if p == "Rule 223 C")
FA_SHEET_RULES.insert(_223c_idx, ("Rule 223 C2", _handle_fa_rule_223c2))


def _handle_fa_rule_450(ws, dest_filename):
    # Rule 450: Driver Based Rating Plan — exactly 2 printed pages.
    #   Page 1: Male Liability + Male Collision (complete, no mid-table cuts).
    #   Page 2: Female Liability + Female Collision + Violation table (complete).
    #
    # Strategy: find the female section heading (second "450.B.1.a." in col A),
    # insert a manual break just before it, then compute an explicit scale % so
    # the taller of the two halves fits on exactly one portrait page.
    # Using fitToHeight=2 was unreliable — Excel sometimes ignores manual breaks
    # when auto-fit is active, so we use an explicit scale instead.
    ws.print_area = f"A1:E{ws.max_row}"

    # Locate the female "450.B.1.a." heading (second occurrence; first is male).
    female_heading_row = None
    count = 0
    for row in range(1, ws.max_row + 1):
        if str(ws.cell(row=row, column=1).value or "").strip().startswith("450.B.1.a."):
            count += 1
            if count == 2:
                female_heading_row = row
                break

    if female_heading_row is None:
        fit_single_page(ws)
        return

    # Break after the blank separator row so the female heading starts page 2.
    break_row = female_heading_row - 1
    add_break_after(ws, break_row)

    # Compute scale so the larger section fits on one portrait page.
    # Portrait letter: printable height ≈ 9.5" (11" − 0.75" top − 0.75" bottom).
    # Default Excel row height = 15 pt = 15/72".
    PAGE_H     = 9.5
    ROW_H      = 15 / 72
    page1_rows = break_row
    page2_rows = ws.max_row - break_row
    scale_f    = PAGE_H / (max(page1_rows, page2_rows) * ROW_H)
    scale_pct  = max(50, min(100, int(scale_f * 100)))

    ws.sheet_properties.pageSetUpPr.fitToPage = False
    ws.page_setup.scale = scale_pct


FA_SHEET_RULES.append(("Rule 450", _handle_fa_rule_450))


# ---------------------------------------------------------------------------
# process_pagebreaks: FA version — uses FA_SHEET_RULES instead of BA default
# ---------------------------------------------------------------------------

def process_pagebreaks(dest_filename1, dest_filename2):
    """
    Apply page breaks / print settings to dest_filename1 using FA_SHEET_RULES.

    dest_filename2 is accepted for backward compatibility (was a PDF path).
    FA_SHEET_RULES contains all BA rules plus any FA-only additions above,
    so adding a FA rule here automatically takes effect — no other file changes.
    """
    print(f"[FApagebreaks] Processing: {dest_filename1}")
    dest_filename1 = os.path.normpath(os.path.abspath(dest_filename1))

    workbook = openpyxl.load_workbook(dest_filename1)

    # Truncate sheet names exceeding Excel's 31-character limit
    for original_name in list(workbook.sheetnames):
        if len(original_name) > 31:
            workbook[original_name].title = original_name[:31]

    # Apply defaults + FA rules to every sheet
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.print_title_rows = "1:1"
        fit_single_page(ws)
        for prefix, handler in FA_SHEET_RULES:
            if sheet_name.startswith(prefix):
                handler(ws, dest_filename1)
                break

    # Index goes to position 0, visible, active on open
    if "Index" in workbook.sheetnames:
        ws_index = workbook["Index"]
        ws_index.sheet_state = "visible"
        if workbook.sheetnames.index("Index") != 0:
            workbook._sheets.remove(ws_index)
            workbook._sheets.insert(0, ws_index)
        workbook.active = 0

    workbook.save(dest_filename1)
    workbook.close()

    # Sanitize the saved zip to prevent Excel's "Open and Repair" popup
    _sanitize_xlsx(dest_filename1)
    print("[FApagebreaks] Done.")