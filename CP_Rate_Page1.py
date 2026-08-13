# <editor-fold desc="Libraries">
import datetime
import warnings
warnings.simplefilter(action='ignore')
import time
import sqlite3
import numpy as np
import pandas as pd
import os.path
import xlwings as xw
import xlsxwriter as xlw
import datetime
import tabulate
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, DEFAULT_FONT
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter
# </editor-fold>

#<editor-fold desc="Functions and Preset Formatting">
warnings.simplefilter(action='ignore')

def round_half_up(n, decimals=0):
    multiplier = 10 ** decimals
    if n > 0:
        return np.floor(((n+0.000000001) * multiplier) + 0.5) / multiplier
    elif n == 0:
        return 0
    else:
        return np.floor(((n-0.000000001) * multiplier) + 0.5) / multiplier

def pixelsToInches(px):
    return px / float(7)

def splitdf(inputdata,splitnum):
    import math
    splitindex = math.ceil(inputdata.shape[0]/splitnum)
    splitlist = list()
    for i in range(1,splitnum+1):
        splitlist.append(inputdata.iloc[(i-1)*splitindex:i*splitindex].reset_index(drop=True))
    return pd.concat(splitlist,axis=1)

def formatwkstSM(wkstname,titlerows,A1title,A2title,dfname,statename,stabb, effdate):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.styles.borders import Border, Side

    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    headerFont = headerFontName + ',Bold'
    footerFont = footerFontName + ',Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'

    wkstname['A1'] = A1title
    wkstname['A2'] = A2title
    cella2 = wkstname['A2']
    cella2.alignment = Alignment(horizontal='left',vertical='center')
    for r in dataframe_to_rows(dfname, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            wkstname['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
            continue
        wkstname.append(r)
    wkstname.page_setup.orientation = 'portrait'  # Landscape orientation for printing
    wkstname.page_setup.firstPageNumber = 1  # Resetting the page counter for the footer on each worksheet
    wkstname.page_setup.useFirstPageNumber = True
    wkstname.sheet_view.showGridLines = False  # Turning off gridlines
    wkstname.print_title_rows = '1:'+ titlerows # Number of title rows to repeat through printed pages
    wkstname.page_margins.left = leftMargin
    wkstname.page_margins.right = rightMargin
    wkstname.page_margins.top = topMargin
    wkstname.page_margins.bottom = bottomMargin
    wkstname.page_margins.header = headerMargin
    wkstname.page_margins.footer = footerMargin
    wkstname.print_options.horizontalCentered = True
    # Left Header
    if statename == 'Florida':
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    wkstname.oddHeader.left.font = headerFont
    wkstname.oddHeader.left.size = headerFontSize
    # Center header
    wkstname.oddHeader.center.text = "\n\n" + statename + " - Rate Pages"
    wkstname.oddHeader.center.size = headerFontSize
    wkstname.oddHeader.center.font = headerFont
    # Right header
    wkstname.oddHeader.right.size = headerFontSize
    wkstname.oddHeader.right.text = "Effective Date: " + effdate
    wkstname.oddHeader.right.font = headerFont
    # Left footer
    wkstname.oddFooter.left.size = 7
    wkstname.oddFooter.left.font = footerFont
    # Center footer
    wkstname.oddFooter.center.text = stabb + " - SRP - &[Tab] - &P"
    wkstname.oddFooter.center.size = footerFontSize
    wkstname.oddFooter.center.font = footerFont

    if NICOFCondition == True:
        wkstname.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFCondition == False and NAFFCondition == True:
        wkstname.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFCondition == False and NAFFCondition == False and NACOCondition == True:
        wkstname.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        wkstname.oddFooter.left.text = "\nNationwide General Insurance Company"

def formatwkstMM(wkstname,titlerows,A1title,A2title,dfname,statename,stabb, effdate):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.styles.borders import Border, Side
    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    headerFont = headerFontName + ',Bold'
    footerFont = footerFontName + ',Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'

    wkstname['A1'] = A1title
    wkstname['A2'] = A2title
    cella2 = wkstname['A2']
    cella2.alignment = Alignment(horizontal='left',vertical='center')
    for r in dataframe_to_rows(dfname, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            wkstname['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
            continue
        wkstname.append(r)

    wkstname.page_setup.orientation = 'portrait'  # Landscape orientation for printing
    wkstname.page_setup.firstPageNumber = 1  # Resetting the page counter for the footer on each worksheet
    wkstname.page_setup.useFirstPageNumber = True
    wkstname.sheet_view.showGridLines = False  # Turning off gridlines
    wkstname.print_title_rows = '1:'+ titlerows
    wkstname.page_margins.left = leftMargin
    wkstname.page_margins.right = rightMargin
    wkstname.page_margins.top = topMargin
    wkstname.page_margins.bottom = bottomMargin
    wkstname.page_margins.header = headerMargin
    wkstname.page_margins.footer = footerMargin
    wkstname.print_options.horizontalCentered = True
    # Left Header
    if statename == 'Florida':
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    wkstname.oddHeader.left.size = headerFontSize
    wkstname.oddHeader.left.font = headerFont
    # Center header
    wkstname.oddHeader.center.text = "\n\n" + statename + " - Rate Pages"
    wkstname.oddHeader.center.size = headerFontSize
    wkstname.oddHeader.center.font = headerFont
    # Right header
    wkstname.oddHeader.right.text = "Effective Date: " + effdate
    wkstname.oddHeader.right.size = headerFontSize
    wkstname.oddHeader.right.font = headerFont
    if stabb != "WI":
        wkstname.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company \nNationwide Property & Casualty Insurance Company \nAllied Insurance Company of America"
    else:
        wkstname.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company"
    # Left footer
    wkstname.oddFooter.left.size = footerFontSize
    wkstname.oddFooter.left.font = footerFont
    # Center footer
    wkstname.oddFooter.center.text = stabb + " - SRP - &[Tab] - &P"
    wkstname.oddFooter.center.size = footerFontSize
    wkstname.oddFooter.center.font = footerFont

def borderfnct(wkstname):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.borders import Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.pagebreak import Break
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.styles.borders import Border, Side
    import math
    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    headerFont = headerFontName + ',Bold'
    footerFont = footerFontName + ',Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'
    for row in range(1, wkstname.max_row + 1):
        for col in range(1, wkstname.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            cell = wkstname[char + str(row)]
            wkstname.column_dimensions[char].bestFit = True  # Using bestfit as the default option for column widths
            if row > 2 and cell.value != '':  # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                      right=Side(border_style='thin', color='00000000'),
                                      top=Side(border_style='thin', color='00000000'),
                                      bottom=Side(border_style='thin', color='00000000'))
            if row > 2 and cell.value == '' or cell.value == ' ':
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                      right=Side(border_style='thin', color='00000000'),
                                      bottom=Side(border_style='thin', color='00000000'))
            if row < 2:  # Applies bold font on row 1, which is a header row
                cell.font = fontBoldUnderline
            elif col >= 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            if col == 1 and row == 2:
                cell.alignment = Alignment(horizontal='left',vertical = 'bottom',wrap_text=False)
    for col in wkstname.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells 
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        wkstname.column_dimensions[column].width = adjusted_width
    wkstname.print_area = 'A1:'+ get_column_letter(wkstname.max_column) + str(wkstname.max_row+1)
    wkstname.page_setup.fitToPage = True
    if wkstname.max_row > 30:
        tablefac = math.ceil((wkstname.max_row + 1)/50)

        for rnum in range(1, tablefac):
            row_num = rnum*50
            page_break = Break(id=row_num)
            wkstname.row_breaks.append(page_break)
        wkstname.page_setup.fitToPage = True
        wkstname.page_setup.fitToWidth = tablefac
        wkstname.page_setup.fitToHeight = tablefac

def addlformat(wkstname):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    wkstname['A1'].alignment = Alignment(horizontal='center')
    wkstname['A2'].alignment = Alignment(horizontal='center')
    wkstname.insert_cols(0,4)
    wkstname.column_dimensions['A'].width = 9.1
    wkstname.column_dimensions['B'].width = 9.1
    wkstname.column_dimensions['C'].width = 9.1
    wkstname.column_dimensions['D'].width = 9.1
    wkstname['A3'] = '*'
    wkstname['A3'].font = Font(color="FFFFFF")
    wkstname.print_area = 'A1:' + get_column_letter(wkstname.max_column + 4) + str(wkstname.max_row + 1)
    wkstname.page_setup.fitToPage = True

print('Creating Rate Pages')

# </editor-fold>

#<editor-fold desc="Company Info">
start_time = time.time()

pd.set_option('display.max_columns', None)
pd.options.display.width = None
pd.options.mode.chained_assignment = None

global NGICCondition
global MMCondition
global NACOCondition
global NAFFCondition
global NICOFCondition

try: NGICRatebook
except:
    NGICRatebook = "Not found"
    NGICCondition = False
else:
    if NGICRatebook == '':
        NGICRatebook = "Not found"
        NGICCondition = False
    else:
        NGICRatebook = pd.ExcelFile(NGICRatebook)
        NGICCondition = True

try: MMRatebook
except:
    MMRatebook = "Not found"
    MMCondition = False
else:
    if MMRatebook == '':
        MMRatebook = "Not found"
        MMCondition = False
    else:
        MMRatebook = pd.ExcelFile(MMRatebook)
        MMCondition = True

try: NACORatebook
except:
    NACORatebook = "Not found"
    NACOCondition = False
else:
    if NACORatebook == '':
        NACORatebook = "Not found"
        NACOCondition = False
    else:
        NACORatebook = pd.ExcelFile(NACORatebook)
        NACOCondition = True

try: NAFFRatebook
except:
    NAFFRatebook = "Not found"
    NAFFCondition = False
else:
    if NAFFRatebook == '':
        NAFFRatebook = "Not found"
        NAFFCondition = False
    else:
        NAFFRatebook = pd.ExcelFile(NAFFRatebook)
        NAFFCondition = True

try: NICOFRatebook
except:
    NICOFRatebook = "Not found"
    NICOFCondition = False
else:
    if NICOFRatebook == '':
        NICOFRatebook = "Not found"
        NICOFCondition = False
    else:
        NICOFRatebook = pd.ExcelFile(NICOFRatebook)
        NICOFCondition = True


if NGICRatebook != "Not found":
    NGICwb = load_workbook(NGICRatebook, read_only=True)
    RateBookDetails = pd.read_excel(NGICRatebook, sheet_name='Rate Book Details')
else:
    RateBookDetails = pd.read_excel(MMRatebook, sheet_name='Rate Book Details')
State = RateBookDetails.iloc[3,4]
EffectiveDate = RateBookDetails.iloc[7,4]
EffectiveDate = datetime.date.strftime(EffectiveDate, "%m-%d-%y")

try: TerrAdjRatebook
except:
    TerrAdjRatebook = "Not found"
else:
    if TerrAdjRatebook == '':
        TerrAdjRatebook = "Not found"
    else:
        TerrAdjRatebook = pd.ExcelFile(TerrAdjRatebook)

if State == "Alabama":
    StateAbb = "AL"
if State == "Alaska":
    StateAbb = "AK"
if State == "Arizona":
    StateAbb = "AZ"
if State == "Arkansas":
    StateAbb = "AR"
if State == "California":
    StateAbb = "CA"
if State == "Colorado":
    StateAbb = "CO"
if State == "Connecticut":
    StateAbb = "CT"
if State == "Delaware":
    StateAbb = "DE"
if State == "District of Columbia":
    StateAbb = "DC"
if State == "Florida":
    StateAbb = "FL"
if State == "Georgia":
    StateAbb = "GA"
if State == "Hawaii":
    StateAbb = "HI"
if State == "Idaho":
    StateAbb = "ID"
if State == "Illinois":
    StateAbb = "IL"
if State == "Indiana":
    StateAbb = "IN"
if State == "Iowa":
    StateAbb = "IA"
if State == "Kansas":
    StateAbb = "KS"
if State == "Kentucky":
    StateAbb = "KY"
if State == "Louisiana":
    StateAbb = "LA"
if State == "Maine":
    StateAbb = "ME"
if State == "Maryland":
    StateAbb = "MD"
if State == "Massachusetts":
    StateAbb = "MA"
if State == "Michigan":
    StateAbb = "MI"
if State == "Minnesota":
    StateAbb = "MN"
if State == "Mississippi":
    StateAbb = "MS"
if State == "Missouri":
    StateAbb = "MO"
if State == "Montana":
    StateAbb = "MT"
if State == "Nebraska":
    StateAbb = "NE"
if State == "Nevada":
    StateAbb = "NV"
if State == "New Hampshire":
    StateAbb = "NH"
if State == "New Jersey":
    StateAbb = "NJ"
if State == "New Mexico":
    StateAbb = "NM"
if State == "New York":
    StateAbb = "NY"
if State == "North Carolina":
    StateAbb = "NC"
if State == "North Dakota":
    StateAbb = "ND"
if State == "Ohio":
    StateAbb = "OH"
if State == "Oklahoma":
    StateAbb = "OK"
if State == "Oregon":
    StateAbb = "OR"
if State == "Pennsylvania":
    StateAbb = "PA"
if State == "Rhode Island":
    StateAbb = "RI"
if State == "South Carolina":
    StateAbb = "SC"
if State == "South Dakota":
    StateAbb = "SD"
if State == "Tennessee":
    StateAbb = "TN"
if State == "Texas":
    StateAbb = "TX"
if State == "Utah":
    StateAbb = "UT"
if State == "Vermont":
    StateAbb = "VT"
if State == "Virginia":
    StateAbb = "VA"
if State == "Washington":
    StateAbb = "WA"
if State == "West Virginia":
    StateAbb = "WV"
if State == "Wisconsin":
    StateAbb = "WI"
if State == "Wyoming":
    StateAbb = "WY"

# try: CWRatebook
# except:
#     CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025.xlsx')
# else:
#     if CWRatebook == '':
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025.xlsx')
#     elif StateAbb in ['VT', 'KY']:
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025 (VT, KY).xlsx')
#     elif StateAbb == NY:
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025 (NY).xlsx')
#     elif StateAbb in ['UT','MI','DE','TN','AL']:
#         CWRatebook = pd.ExcelFile(
#             '\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025 (UT, MI, DE, TN, AL).xlsx')
#     else:
#         CWRatebook = pd.ExcelFile(CWRatebook)

CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025.xlsx')
#</editor-fold>

# <editor-fold desc="Create LCM Dataframes">
if NAFFRatebook != "Not found":
    NAFFLCM = pd.read_excel(NAFFRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NAFFLCM = NAFFLCM.iloc[0, 1]
    NAFFCRIMELCM = pd.read_excel(NAFFRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NAFFCRIMELCM = NAFFCRIMELCM.iloc[0, 1]
if NACORatebook != "Not found":
    NACOLCM = pd.read_excel(NACORatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NACOLCM = NACOLCM.iloc[0, 1]
    NACOCRIMELCM = pd.read_excel(NACORatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NACOCRIMELCM = NACOCRIMELCM.iloc[0, 1]
if NICOFRatebook != "Not found":
    NICOFLCM = pd.read_excel(NICOFRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NICOFLCM = NICOFLCM.iloc[0, 1]
    NICOFCRIMELCM = pd.read_excel(NICOFRatebook, sheet_name='CrimeLossCostMultiplier_Ext',skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NICOFCRIMELCM = NICOFCRIMELCM.iloc[0, 1]
if NGICRatebook != "Not found":
    NGICLCM = pd.read_excel(NGICRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NGICLCM = NGICLCM.iloc[0, 1]
    NGICCRIMELCM = pd.read_excel(NGICRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NGICCRIMELCM = NGICCRIMELCM.iloc[0, 1]
if MMRatebook != "Not found":
    if StateAbb != "WI":
        MMLCM = pd.read_excel(MMRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCRIMELCM = pd.read_excel(MMRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCompDev = pd.read_excel(MMRatebook, sheet_name='Company Deviation Factor_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        NWMLCM = MMLCM.iloc[0, 1]
        NWMCRIMELCM = MMCRIMELCM.iloc[0, 1]
        AICOA = ['aicoa_ext']
        AICOA = pd.DataFrame(AICOA, columns=['UnderwritingCompanyCode'])
        AICOACompDev = pd.merge(MMCompDev, AICOA, on = 'UnderwritingCompanyCode', how = 'inner')
        AICOACompDev = AICOACompDev.iloc[0, 2]
        NICOA = ['nicoa_ext']
        NICOA = pd.DataFrame(NICOA, columns=['UnderwritingCompanyCode'])
        NICOACompDev = pd.merge(MMCompDev, NICOA, on='UnderwritingCompanyCode', how='inner')
        NICOACompDev = NICOACompDev.iloc[0, 2]
        NWPC = ['npcic_ext']
        NWPC = pd.DataFrame(NWPC, columns=['UnderwritingCompanyCode'])
        NWPCCompDev = pd.merge(MMCompDev, NWPC, on='UnderwritingCompanyCode', how='inner')
        NWPCCompDev = NWPCCompDev.iloc[0, 2]
        AICOALCM = round_half_up(NWMLCM*AICOACompDev,3)
        NICOALCM = round_half_up(NWMLCM * NICOACompDev, 3)
        NWPCLCM = round_half_up(NWMLCM * NWPCCompDev, 3)
        AICOACRIMELCM = round_half_up(NWMCRIMELCM*AICOACompDev,3)
        NICOACRIMELCM = round_half_up(NWMCRIMELCM * NICOACompDev, 3)
        NWPCCRIMELCM = round_half_up(NWMCRIMELCM * NWPCCompDev, 3)
    else:
        MMLCM = pd.read_excel(MMRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCRIMELCM = pd.read_excel(MMRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCompDev = pd.read_excel(MMRatebook, sheet_name='Company Deviation Factor_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        NWMLCM = MMLCM.iloc[0, 1]
        NWMCRIMELCM = MMCRIMELCM.iloc[0, 1]
        NICOA = ['nicoa_ext']
        NICOA = pd.DataFrame(NICOA, columns=['UnderwritingCompanyCode'])
        NICOACompDev = pd.merge(MMCompDev, NICOA, on='UnderwritingCompanyCode', how='inner')
        NICOACompDev = NICOACompDev.iloc[0, 2]
        NICOALCM = round_half_up(NWMLCM * NICOACompDev, 3)
        NICOACRIMELCM = round_half_up(NWMCRIMELCM * NICOACompDev, 3)
if NICOFRatebook != "Not found":
    SMLCMs = [['Nationwide Affinity Insurance Company of America', NAFFLCM], ['Nationwide Assurance Company', NACOLCM],
              ['Nationwide Insurance Company of Florida', NICOFLCM], ['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    SMCRIMELCMs = [['Nationwide Affinity Insurance Company of America', NAFFCRIMELCM], ['Nationwide Assurance Company', NACOCRIMELCM],
              ['Nationwide Insurance Company of Florida', NICOFCRIMELCM], ['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])
if NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
    SMLCMs = [['Nationwide Affinity Insurance Company of America', NAFFLCM], ['Nationwide Assurance Company', NACOLCM],
               ['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    SMCRIMELCMs = [['Nationwide Affinity Insurance Company of America', NAFFCRIMELCM], ['Nationwide Assurance Company', NACOCRIMELCM],
                ['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])
if NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
    SMLCMs = [['Nationwide Assurance Company', NACOLCM], ['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    SMCRIMELCMs = [['Nationwide Assurance Company', NACOCRIMELCM], ['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])
if NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook == "Not found" and NGICRatebook != "Not found":
    SMLCMs = [['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    SMCRIMELCMs = [['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])
if MMRatebook != "Not found":
    if StateAbb != "WI":
        MMLCMs = [['Nationwide Insurance Company of America', NICOALCM], ['Nationwide Mutual Insurance Company', NWMLCM],
                  ['Nationwide Property & Casualty Insurance Company', NWPCLCM], ['Allied Insurance Company of America', AICOALCM]]
        MMLCMTable = pd.DataFrame(MMLCMs, columns=['Company', 'LCM'])
        MMCRIMELCMs = [['Nationwide Insurance Company of America', NICOACRIMELCM], ['Nationwide Mutual Insurance Company', NWMCRIMELCM],
                  ['Nationwide Property & Casualty Insurance Company', NWPCCRIMELCM], ['Allied Insurance Company of America', AICOACRIMELCM]]
        MMCRIMELCMTable = pd.DataFrame(MMCRIMELCMs, columns=['Company', 'LCM'])
    else:
        MMLCMs = [['Nationwide Insurance Company of America', NICOALCM], ['Nationwide Mutual Insurance Company', NWMLCM]]
        MMLCMTable = pd.DataFrame(MMLCMs, columns=['Company', 'LCM'])
        MMCRIMELCMs = [['Nationwide Insurance Company of America', NICOACRIMELCM], ['Nationwide Mutual Insurance Company', NWMCRIMELCM]]
        MMCRIMELCMTable = pd.DataFrame(MMCRIMELCMs, columns=['Company', 'LCM'])
# </editor-fold>

# <editor-fold desc="Create Territory Adjustment Dataframes">
if TerrAdjRatebook != "Not found":

    # Load the workbook
    Terrwb = pd.ExcelFile(TerrAdjRatebook)

    # Read + sort each sheet by column A (Grid ID), then keep the factor column
    bgi = (
        pd.read_excel(
            Terrwb, sheet_name='BGI',
            skiprows=[0, 1, 2, 3, 4, 5, 6],
            usecols=[0, 1],  # A=Grid ID, B=BGI Factor
            names=['Grid ID', 'BGI Factor'],
            engine='openpyxl'
        )
        .sort_values('Grid ID', kind='mergesort')
        .reset_index(drop=True)
    )

    bgii = (
        pd.read_excel(
            Terrwb, sheet_name='BGII',
            skiprows=[0, 1, 2, 3, 4, 5, 6],
            usecols=[0, 1],  # A=Grid ID, B=BGII Factor
            names=['Grid ID', 'BGII Factor'],
            engine='openpyxl'
        )
        .sort_values('Grid ID', kind='mergesort')
        .reset_index(drop=True)
    )

    scol = (
        pd.read_excel(
            Terrwb, sheet_name='SCOL',
            skiprows=[0, 1, 2, 3, 4, 5, 6],
            usecols=[0, 1],  # A=Grid ID, B=SCOL Factor
            names=['Grid ID', 'SCOL Factor'],
            engine='openpyxl'
        )
        .sort_values('Grid ID', kind='mergesort')
        .reset_index(drop=True)
    )

    GridFac = pd.concat(
         [
             bgi[['Grid ID']],
             bgi[['BGI Factor']],
             bgii[['BGII Factor']],
             scol[['SCOL Factor']]
         ],
         axis=1
        )

    TerrCoord = pd.ExcelFile(
        '\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Property\\Territory Defs (Independent)\\3.0\\Development\\R Shiny Interface\\RShiny Smoothed Relativities\\' + StateAbb + ' Smoothed Relativities.xlsx')
    Coords = pd.read_excel(TerrCoord,sheet_name='Sheet1',usecols=[1,2,3], names=['Grid ID', 'Longitude', 'Latitude']).round(7)
    GridFac = pd.merge(GridFac, Coords, how='inner', on='Grid ID')
    GridFac = GridFac[['Grid ID','Longitude','Latitude','BGI Factor','BGII Factor','SCOL Factor']]
    GridFac = GridFac.rename(columns={"Grid ID":"Territory"})
    GridFac = GridFac.drop_duplicates().reset_index(drop=True)
    GridLen = math.ceil(GridFac.shape[0]/3)
    GridFac1 = GridFac.iloc[:GridLen].reset_index(drop=True)
    GridFac2 = GridFac.iloc[GridLen:2*GridLen].reset_index(drop=True)
    GridFac3 = GridFac.iloc[2*GridLen:].reset_index(drop=True)
    GridFacXL = pd.concat([GridFac1,GridFac2,GridFac3],axis=1).fillna("   ")
    GridFacXL.insert(6,'','')
    GridFacXL.insert(13,' ','')
    print("territory set up")
else:
    pass
# </editor-fold>

# <editor-fold desc="Create PMF Dataframes">
if MMRatebook != "Not found":
    MMwb = load_workbook(MMRatebook, read_only=False)
    if 'PackageModifierFactorTable_Ext' in MMwb.sheetnames:
        MMPMF = pd.read_excel(MMRatebook, sheet_name='PackageModifierFactorTable_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    else:
        MMPMF = pd.read_excel(NGICRatebook, sheet_name='PackageModifierFactorTable_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    MMPMF = MMPMF.rename(columns={'PackageModifierCode': 'Package Code'})
    MMPMF = MMPMF.drop(labels=[0, 9], axis=0)
else:
    MMwb = Workbook()

if NGICRatebook != "Not found":
    SMPMF = pd.read_excel(NGICRatebook, sheet_name='PackageModifierFactorTable_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    SMPMF = SMPMF.rename(columns={'PackageModifierCode': 'Package Code'})
    SMPMF = SMPMF.drop(labels=[0, 9], axis=0)
# </editor-fold>

# <editor-fold desc="Create Capping Dataframe">
if v5.get() == 1 and v4.get() == 1:
    SMMinCap = SMMinCapRange.get()
    SMMaxCap = SMMaxCapRange.get()
    if NICOFRatebook != "Not found":
        SMCapping = [['Nationwide Affinity Insurance Company of America', SMMinCap, SMMaxCap], ['Nationwide Assurance Company', SMMinCap, SMMaxCap],
                ['Nationwide Insurance Company of Florida', SMMinCap, SMMaxCap]]
        SMCapTable = pd.DataFrame(SMCapping, columns=['Company', 'Minimum', 'Maximum'])
    if NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        SMCapping = [['Nationwide Affinity Insurance Company of America', SMMinCap, SMMaxCap], ['Nationwide Assurance Company', SMMinCap, SMMaxCap]]
        SMCapTable = pd.DataFrame(SMCapping, columns=['Company', 'Minimum', 'Maximum'])
    if NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        SMCapping = [['Nationwide Assurance Company', SMMinCap, SMMaxCap]]
        SMCapTable = pd.DataFrame(SMCapping, columns=['Company', 'Minimum', 'Maximum'])
if v10.get() == 1 and v11.get() == 1:
    MMMinCap = MMMinCapRange.get()
    MMMaxCap = MMMaxCapRange.get()
    MMCapping = [['Nationwide Insurance Company of America', MMMinCap, MMMaxCap], ['Nationwide Mutual Insurance Company', MMMinCap, MMMaxCap],
                ['Nationwide Property & Casualty Insurance Company', MMMinCap, MMMaxCap], ['Allied Insurance Company of America', MMMinCap, MMMaxCap]]
    MMCapTable = pd.DataFrame(MMCapping, columns=['Company', 'Minimum', 'Maximum'])
# </editor-fold>

# <editor-fold desc="Create Crime Deductible Dataframe">
CrimeDed = pd.read_excel(CWRatebook, sheet_name='CrimeDeductibleFactor_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
CrimeDed = CrimeDed.rename(columns={'Deductible': 'Deductible Amount'})
#CrimeDed = CrimeDed.drop(labels=[0, 14], axis=0)

# </editor-fold>

# <editor-fold desc="Create AOI Factor Dataframe">

NGICwb = load_workbook(NGICRatebook, read_only=True)

if 'BasicGroupILOIFactorBldg' in NGICwb.sheetnames:
    BGIBLOI = pd.read_excel(MMRatebook, sheet_name='BasicGroupILOIFactorBldg',skiprows=11)

elif 'BasicGroupILOIFactorBldg' in NGICwb.sheetnames:
    BGIBLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupILOIFactorBldg',skiprows=11)

else:
    BGIBLOI = pd.read_excel(CWRatebook,sheet_name='BasicGroupILOIFactorBldg',skiprows=11)

if 'BasicGroupIILOIFactorBldg' in NGICwb.sheetnames:
    BGIIBLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupIILOIFactorBldg',skiprows=11)

else:
    BGIIBLOI = pd.read_excel(CWRatebook,sheet_name='BasicGroupIILOIFactorBldg',skiprows=11)

if 'BroadSpecialLOIFactorBldg' in NGICwb.sheetnames:
    SCOLBLOI = pd.read_excel(NGICRatebook, sheet_name='BroadSpecialLOIFactorBldg',skiprows=11)

else:
    SCOLBLOI = pd.read_excel(CWRatebook, sheet_name='BroadSpecialLOIFactorBldg',skiprows=11)

if 'BasicGroupILOIFactorPersProp' in NGICwb.sheetnames:
    BGIPPLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupILOIFactorPersProp',skiprows=11)

else:
    BGIPPLOI = pd.read_excel(CWRatebook,sheet_name='BasicGroupILOIFactorPersProp',skiprows=11)

if 'BasicGroupIILOIFactorPersProp' in NGICwb.sheetnames:
    BGIIPPLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupIILOIFactorPersProp',skiprows=11)

else:
    BGIIPPLOI = pd.read_excel(CWRatebook,sheet_name='BasicGroupIILOIFactorPersProp',skiprows=11)

if 'BroadSpecialLOIFactorPrsnlProp' in NGICwb.sheetnames:
    SCOLPPLOI = pd.read_excel(NGICRatebook, sheet_name='BroadSpecialLOIFactorPrsnlProp',skiprows=11)

else:
    SCOLPPLOI = pd.read_excel(CWRatebook, sheet_name='BroadSpecialLOIFactorPrsnlProp',skiprows=11)

BGIBLOI = BGIBLOI.pivot(index='Limit',columns='ConstructionCode',values='Factor').reset_index(names=['Limit',1,2,3,4,5,6])
BGIBLOI = BGIBLOI.drop(columns=[2,3,5,6])
BGIBLOI = BGIBLOI.drop(BGIBLOI.tail(1).index)
BGIBLOI['Limit'] = BGIBLOI['Limit'].apply('{:,}'.format)
BGIBLOI.iloc[-1, BGIBLOI.columns.get_loc('Limit')] = BGIBLOI.iloc[-1,BGIBLOI.columns.get_loc('Limit')] + "+"
BGIBLOI = BGIBLOI.rename(columns={1:"Construction Group 1-3",4:"Construction Group 4-6"})
BGIBLOI = splitdf(BGIBLOI, 2).fillna('')

BGIIBLOI['Limit'] = BGIIBLOI['Limit'].apply('{:,}'.format)
BGIIBLOI = BGIIBLOI.drop(BGIIBLOI.tail(1).index)
BGIIBLOI.iloc[-1, BGIIBLOI.columns.get_loc('Limit')] = BGIIBLOI.iloc[-1, BGIIBLOI.columns.get_loc('Limit')] + "+"
BGIIBLOI = splitdf(BGIIBLOI,2).fillna('')

SCOLBLOI['Limit'] = SCOLBLOI['Limit'].apply('{:,}'.format)
SCOLBLOI = SCOLBLOI.drop(SCOLBLOI.tail(1).index)
SCOLBLOI.iloc[-1, SCOLBLOI.columns.get_loc('Limit')] = SCOLBLOI.iloc[-1, SCOLBLOI.columns.get_loc('Limit')] + "+"
SCOLBLOI = splitdf(SCOLBLOI,2).fillna('')

BGIPPLOI = BGIPPLOI.pivot(index='Limit',columns='ConstructionCode',values='Factor').reset_index(names=['Limit',1,2,3,4,5,6])
BGIPPLOI = BGIPPLOI.drop(columns=[2,3,5,6])
BGIPPLOI = BGIPPLOI.drop(BGIPPLOI.tail(1).index)
BGIPPLOI['Limit'] = BGIPPLOI['Limit'].apply('{:,}'.format)
BGIPPLOI.iloc[-1, BGIPPLOI.columns.get_loc('Limit')] = BGIPPLOI.iloc[-1, BGIPPLOI.columns.get_loc('Limit')] + "+"
BGIPPLOI = BGIPPLOI.rename(columns={1:"Construction Group 1-3",4:"Construction Group 4-6"})
BGIPPLOI = splitdf(BGIPPLOI, 2).fillna('')

BGIIPPLOI['Limit'] = BGIIPPLOI['Limit'].apply('{:,}'.format)
BGIIPPLOI = BGIIPPLOI.drop(BGIIPPLOI.tail(1).index)
BGIIPPLOI.iloc[-1, BGIIPPLOI.columns.get_loc('Limit')] = BGIIPPLOI.iloc[-1, BGIIPPLOI.columns.get_loc('Limit')] + "+"
BGIIPPLOI = splitdf(BGIIPPLOI,2).fillna('')

SCOLPPLOI['Limit'] = SCOLPPLOI['Limit'].apply('{:,}'.format)
SCOLPPLOI = SCOLPPLOI.drop(SCOLPPLOI.tail(1).index)
SCOLPPLOI.iloc[-1, SCOLPPLOI.columns.get_loc('Limit')] = SCOLPPLOI.iloc[-1, SCOLPPLOI.columns.get_loc('Limit')] + "+"
SCOLPPLOI = splitdf(SCOLPPLOI,2).fillna('')

# </editor-fold>

# <editor-fold desc="Create Tiering Grade Dataframes">
if 'GroupITierFactor_Ext' in NGICwb.sheetnames:
    BGITier = pd.read_excel(NGICRatebook, sheet_name='GroupITierFactor_Ext', skiprows=11)

else:
    BGITier = pd.read_excel(CWRatebook,sheet_name='GroupITierFactor_Ext', skiprows=11)

if 'GroupIITierFactor_Ext' in NGICwb.sheetnames:
    BGIITier = pd.read_excel(NGICRatebook, sheet_name='GroupIITierFactor_Ext',skiprows=11)

else:
    BGIITier = pd.read_excel(CWRatebook, sheet_name='GroupIITierFactor_Ext', skiprows=11)

if 'SpecialCauseofLossTierFactor_Ex' in NGICwb.sheetnames:
    SCOLTier = pd.read_excel(NGICRatebook, sheet_name='SpecialCauseofLossTierFactor_Ex',skiprows=11)

else:
    SCOLTier = pd.read_excel(CWRatebook, sheet_name='SpecialCauseofLossTierFactor_Ex',skiprows=11)

BGITier = BGITier.pivot(index='TierGradeGroup I', columns='TieringNAICSGroup', values='Factor')
BGITier = BGITier[[col for col in BGITier.columns if col != 'All Other'] + ['All Other']]
BGITier['Tier Grade'] = BGITier.index
BGITier = BGITier[['Tier Grade'] + [col for col in BGITier.columns if col != 'Tier Grade']]
BGITier.index.name = None
BGITier = BGITier.rename_axis(None, axis=1)
BGITier = BGITier.reset_index(drop=True)

BGIITier = BGIITier.pivot(index='TierGradeGroup II', columns='TieringNAICSGroup', values='Factor')
BGIITier = BGIITier[[col for col in BGIITier.columns if col != 'All Other'] + ['All Other']]
BGIITier['Tier Grade'] = BGIITier.index
BGIITier = BGIITier[['Tier Grade'] + [col for col in BGIITier.columns if col != 'Tier Grade']]
BGIITier.index.name = None
BGIITier = BGIITier.rename_axis(None, axis=1)
BGIITier = BGIITier.reset_index(drop=True)

SCOLTier = SCOLTier.pivot(index='TierGradeSpecialCOL', columns='TieringNAICSGroup',values='Factor')
SCOLTier = SCOLTier[[col for col in SCOLTier.columns if col != 'All Other'] + ['All Other']]
SCOLTier['Tier Grade'] = SCOLTier.index
SCOLTier = SCOLTier[['Tier Grade'] + [col for col in SCOLTier.columns if col != 'Tier Grade']]
SCOLTier.index.name = None
SCOLTier = SCOLTier.rename_axis(None, axis=1)
SCOLTier = SCOLTier.reset_index(drop=True)

# </editor-fold>

# <editor-fold desc="Create AOB Dataframes">

if 'AgeOfBuildingFactor_Ext' in NGICwb.sheetnames:
    AOB = pd.read_excel(NGICRatebook, sheet_name='AgeOfBuildingFactor_Ext',skiprows=11)
else:
    AOB = pd.read_excel(CWRatebook, sheet_name='AgeOfBuildingFactor_Ext', skiprows=11)

AOB = AOB.pivot_table(index='AgeOfBuildingFrom', columns=['CoveredObject', 'CauseOfLossGroup'], values='Factor').reset_index(names='Age')
AOB.columns = ["Age", "BGITime", "BGIITime", "BroadTime", "EQ1", "SCOLTime", "BGIContents", "BGIIContents",
                   "PPContents", "EQ2", "SCOLContents", "BGIBldg", "BGIIBldg", "STRBldg", "EQ3", "SCOLBldg"]

BGIAOB = AOB[["Age", "BGIBldg", "BGIContents", "BGITime"]].rename(columns={"BGIBldg":"Building","BGIContents":"Contents","BGITime":"Time"})
BGIIAOB = AOB[["Age", "BGIIBldg", "BGIIContents", "BGIITime"]].rename(columns={"BGIIBldg":"Building","BGIIContents":"Contents","BGIITime":"Time"})
SCOLAOB = AOB[["Age", "SCOLBldg", "SCOLContents", "SCOLTime"]].rename(columns={"SCOLBldg":"Building","SCOLContents":"Contents","SCOLTime":"Time"})

BGIAOB = splitdf(BGIAOB,3)
BGIIAOB = splitdf(BGIIAOB,3)
SCOLAOB = splitdf(SCOLAOB,3)

# </editor-fold>

# <editor-fold desc="Create Deductible Factor Dataframe">
if 'DeductibleFactor' in MMwb.sheetnames:
    DIP1 = pd.read_excel(MMRatebook, sheet_name='DeductibleFactor',skiprows=11)
elif 'DeductibleFactor' in NGICwb.sheetnames:
    DIP1 = pd.read_excel(NGICRatebook, sheet_name='DeductibleFactor', skiprows=11)
else:
    DIP1 = pd.read_excel(CWRatebook, sheet_name='DeductibleFactor', skiprows=11)

if 'Deductible250Factor' in MMwb.sheetnames:
    DIP2 = pd.read_excel(MMRatebook, sheet_name='Deductible250Factor',skiprows=11)
elif 'Deductible250Factor' in NGICwb.sheetnames:
    DIP2 = pd.read_excel(NGICRatebook, sheet_name='Deductible250Factor', skiprows=11)
else:
    DIP2 = pd.read_excel(CWRatebook, sheet_name='Deductible250Factor', skiprows=11)

DIP = DIP1.pivot_table(index=['Deductible','Limit'],columns = 'CauseOfLossDeductible',values='Factor').reset_index(names=['Deductible','Limit'])
# remove pivot_table commas from formatting
DIP[['Deductible']] = DIP[['Deductible']].replace({',': ''}, regex=True).astype(float)
# find unique limit values to use for deductible 250 page, transpose to turn into a column value
DedLim = pd.DataFrame(DIP['Limit'].unique().T, columns=['Limit'])
# create dataframe with deductible limit values, repeated by the length of the deductible 250 page
DedLim = pd.concat([DedLim]*len(DIP2['CauseOfLoss']))
# lowercase column names
DIP.columns = DIP.columns.str.lower()

# convert data to array
Ded250 = np.array(DIP2['CauseOfLoss'])
Ded250F = np.array(DIP2['Factor'])
# define new columns in DedLim, repeat Ded250 limits repeatedly until DedLim is full
DedLim['CauseOfLoss'] = Ded250[DedLim.groupby('Limit').cumcount()]
DedLim['Factor'] = Ded250F[DedLim.groupby('Limit').cumcount()]
# define new deductible
DedLim = DedLim.assign(Deductible='250')
DedLim = DedLim[['Deductible','CauseOfLoss','Limit','Factor']]
# pivot
DedLim2 = DedLim.pivot_table(index=['Deductible','Limit'],columns = 'CauseOfLoss',values='Factor').reset_index(names=['Deductible','Limit'])
# drop broad
DedLim2 = DedLim2.drop(['Broad'],axis=1).rename(columns={"Basic":"Basic Group I","All Other":"Basic Group II","Special":"Other Cause Of Loss"})
DedLim2.columns = DedLim2.columns.str.lower()
DIP = DIP.sort_values(by=['deductible','limit'])

DIP['deductible'] = DIP['deductible'].astype(int).apply('{:,}'.format)
DIP = DIP.reset_index(drop=True)
DIP = pd.concat([DedLim2,DIP],ignore_index=True)

DIPI = DIP[['deductible','limit','basic group i']].rename(columns={"basic group i":"Factor"})
DIPII = DIP[['deductible','limit','basic group ii']].rename(columns={"basic group ii":"Factor"})
DIPSCOL = DIP[['deductible','limit','other cause of loss']].rename(columns={"other cause of loss":"Factor"})
DIPI.columns = DIPI.columns.map(str.title)
DIPII.columns = DIPII.columns.map(str.title)
DIPSCOL.columns = DIPSCOL.columns.map(str.title)

DIPILim = DIPI['Limit'].astype(str)
DIPIILim = DIPII['Limit'].astype(str)
DIPSCOLLim = DIPSCOL['Limit'].astype(str)

##Formatting Limit Column
for i in range(len(DIPI['Limit'])):
    if int(DIPI['Limit'][i]) not in [min(DIPI['Limit'].astype(int)),max(DIPI['Limit'].astype(int))]:
        DIPILim[i] = str(f"{DIPI['Limit'][i-1]:,}") + ' - ' + str(f"{DIPI['Limit'][i]:,}")
    elif int(DIPI['Limit'][i]) == min(DIPI['Limit'].astype(int)):
        DIPILim[i] = str(f"{DIPI['Limit'][i]:,}") + ' or less'
    else:
        DIPILim[i] = 'More than ' + str(f"{DIPI['Limit'][i-1]:,}")

for i in range(len(DIPII['Limit'])):
    if int(DIPII['Limit'][i]) not in [min(DIPII['Limit'].astype(int)),max(DIPII['Limit'].astype(int))]:
        DIPIILim[i] = str(f"{DIPII['Limit'][i-1]:,}") + ' - ' + str(f"{DIPII['Limit'][i]:,}")
    elif int(DIPII['Limit'][i]) == min(DIPII['Limit'].astype(int)):
        DIPIILim[i] = str(f"{DIPII['Limit'][i]:,}") + ' or less'
    else:
        DIPIILim[i] = 'More than ' + str(f"{DIPII['Limit'][i-1]:,}")

for i in range(len(DIPSCOL['Limit'])):
    if int(DIPSCOL['Limit'][i]) not in [min(DIPSCOL['Limit'].astype(int)),max(DIPSCOL['Limit'].astype(int))]:
        DIPSCOLLim[i] = str(f"{DIPSCOL['Limit'][i-1]:,}") + ' - ' + str(f"{DIPSCOL['Limit'][i]:,}")
    elif int(DIPSCOL['Limit'][i]) == min(DIPSCOL['Limit'].astype(int)):
        DIPSCOLLim[i] = str(f"{DIPSCOL['Limit'][i]:,}") + ' or less'
    else:
        DIPSCOLLim[i] = 'More than ' + str(f"{DIPSCOL['Limit'][i-1]:,}")

DIPI['Limit'] = DIPILim
DIPII['Limit'] = DIPIILim
DIPSCOL['Limit'] = DIPSCOLLim

DIPI = splitdf(DIPI,3)
DIPII = splitdf(DIPII,3)
DIPSCOL = splitdf(DIPSCOL,3)
# </editor-fold>

#<editor-fold desc="Create M&S Dataframe">
if 'CrimeTerritoryBaseRate_Ext' in NGICwb.sheetnames:
    CrimeMSDF = pd.read_excel(NGICRatebook, sheet_name='CrimeTerritoryBaseRate_Ext', skiprows=11).rename(columns={"CrimeTerritory":"Territory","Base Premium":"Premium"})
else:
    CrimeMSDF = pd.read_excel(CWRatebook, sheet_name='CrimeTerritoryBaseRate_Ext', skiprows=11).rename(columns={"CrimeTerritory":"Territory","Base Premium":"Premium"})

if 'MoneyandSecuritiesOccupancyFact' in NGICwb.sheetnames:
    MSDF = pd.read_excel(NGICRatebook, sheet_name='MoneyandSecuritiesOccupancyFact', skiprows=11)
else:
    MSDF = pd.read_excel(CWRatebook, sheet_name='MoneyandSecuritiesOccupancyFact', skiprows=11)

MSDF = MSDF.pivot_table(index=['Inside','Outside'],columns='Occupancy',values='Factor').reset_index()
MSDF = MSDF.rename(columns={'Inside':'Inside Limit','Outside':'Outside Limit'})
MSDF['Inside Limit'] = MSDF['Inside Limit'].apply('{:,}'.format)
MSDF['Outside Limit'] = MSDF['Outside Limit'].apply('{:,}'.format)

#</editor-fold>

#<editor-fold desc="Create Employee Dishonesty Dataframe">
if 'EmployeeDishonestyLimitsBaseRat' in NGICwb.sheetnames:
    EmployeeDF = pd.read_excel(NGICRatebook, sheet_name='EmployeeDishonestyLimitsBaseRat', skiprows=11)
else:
    EmployeeDF = pd.read_excel(CWRatebook, sheet_name='EmployeeDishonestyLimitsBaseRat', skiprows=11)

if 'EmployeeDishonestyLimitsRateabl' in NGICwb.sheetnames:
    EDLR = pd.read_excel(NGICRatebook, sheet_name='EmployeeDishonestyLimitsRateabl', skiprows=11)
else:
    EDLR = pd.read_excel(CWRatebook, sheet_name='EmployeeDishonestyLimitsRateabl', skiprows=11)

EmployeeDF["Each Add'l Rateable Employee"] = EDLR['Premium Per Rateable Employee >5']
EmployeeDF.rename(columns={'Premium < 5 Rateable Employees':'1-5 Rateable Employees'},inplace=True)
EmployeeDF['Limit'] = EmployeeDF['Limit'].apply('{:,}'.format)
#</editor-fold>

#<editor-fold desc="Create Fraudulent Impersonation DF">
if 'FraudulentImpersonationEmployee' in NGICwb.sheetnames:
    Fraudone = pd.read_excel(NGICRatebook, sheet_name='FraudulentImpersonationEmployee', skiprows=11)
else:
    Fraudone = pd.read_excel(CWRatebook, sheet_name='FraudulentImpersonationEmployee', skiprows=11)
if 'FraudulentImpersonationEmpl (1)' in NGICwb.sheetnames:
    Fraudtwo = pd.read_excel(NGICRatebook, sheet_name='FraudulentImpersonationEmpl (1)', skiprows=11)
else:
    Fraudtwo = pd.read_excel(CWRatebook, sheet_name='FraudulentImpersonationEmpl (1)', skiprows=11)

FraudDF = Fraudtwo.merge(Fraudone, how='inner',on='Limit').rename(columns={"Premium < 5 Rateable Employees":"1-5 Rateable Employees","Premium Per Rateable Employee >5":"Each Add'l Rateable Employee"})
FraudDF['Limit'] = FraudDF['Limit'].apply('{:,}'.format)

if 'VerificationFactorEmployeesFact' in NGICwb.sheetnames:
    FraudDF2 = pd.read_excel(NGICRatebook, sheet_name='VerificationFactorEmployeesFact', skiprows=11)
else:
    FraudDF2 = pd.read_excel(CWRatebook, sheet_name='VerificationFactorEmployeesFact', skiprows=11)

FraudDF2['Verification Type'] = FraudDF2['Verification Type'].map({'OptionAVerifReqdAllTransInstr':"Verification required for all transfer instructions",
                                                               'OptionBVerifReqdAllTransInstrExcessSpecifiedAmt':"Verification required for all transfer instructions in excess of an amount",
                                                               'OptionCVerifTransferInstrNotReqd':"Verification of transfer instructions not required"})

if 'VerificationFactorEmployeesFact' in NGICwb.sheetnames:
    FraudDF3 = pd.read_excel(NGICRatebook, sheet_name='VerificationFactorCustomerorVen', skiprows=11)
else:
    FraudDF3 = pd.read_excel(CWRatebook, sheet_name='VerificationFactorCustomerorVen', skiprows=11)

FraudDF3['Verification Type'] = FraudDF3['Verification Type'].map({'OptionAVerifReqdAllTransInstr':"Verification required for all transfer instructions",
                                                               'OptionBVerifReqdAllTransInstrExcessSpecifiedAmt':"Verification required for all transfer instructions in excess of an amount",
                                                               'OptionCVerifTransferInstrNotReqd':"Verification of transfer instructions not required"})

#</editor-fold

#<editor-fold desc="IRPM DF">

if 'Schedule Rating Threshold_Ext' in NGICwb.sheetnames:
    IRPM1 = pd.read_excel(NGICRatebook, sheet_name='Schedule Rating Threshold_Ext', skiprows=11).drop('AbsoluteThreshold',axis=1).rename(columns={'ScheduleEligibilityIndicator':'Constant'})
else:
    IRPM1 = pd.read_excel(CWRatebook, sheet_name='Schedule Rating Threshold_Ext', skiprows=11).drop('AbsoluteThreshold',axis=1).rename(columns={'ScheduleEligibilityIndicator':'Constant'})
if 'IRPMMaximumCredit' in NGICwb.sheetnames:
    IRPM2 = pd.read_excel(NGICRatebook, sheet_name='IRPMMaximumCredit', skiprows=11)
else:
    IRPM2 = pd.read_excel(CWRatebook, sheet_name='IRPMMaximumCredit', skiprows=11)
if 'IRPMMaximumDebit' in NGICwb.sheetnames:
    IRPM3 = pd.read_excel(NGICRatebook, sheet_name='IRPMMaximumDebit', skiprows=11)
else:
    IRPM3 = pd.read_excel(CWRatebook, sheet_name='IRPMMaximumDebit', skiprows=11)

IRPM = pd.DataFrame(index=range(3),columns=range(2))
IRPM.iloc[0,0]="Minimum Eligible Premium"
IRPM.iloc[0,1]=IRPM1.iloc[0,1]
IRPM.iloc[1,0]="Maximum Credit"
IRPM.iloc[1,1]=IRPM2.iloc[0,1].astype(str) + "%"
IRPM.iloc[2,0]="Maximum Debit"
IRPM.iloc[2,1]=IRPM3.iloc[0,1].astype(str) + "%"

#</editor-fold>

#<editor-fold desc="Emergency Evacuation DF">

if 'EmergencyEvacuationIncludin (1)' in NGICwb.sheetnames:
    EmergencyDF = pd.read_excel(NGICRatebook, sheet_name='EmergencyEvacuationIncludin (1)', skiprows=11)
else:
    EmergencyDF = pd.read_excel(CWRatebook, sheet_name='EmergencyEvacuationIncludin (1)', skiprows=11)

EmergencyDF['Limit'] = EmergencyDF['Limit'].apply('{:,}'.format)

#</editor-fold>

#<editor-fold desc="Civil Authority DF">
if 'CivilAuthorityIncreasedRadiusCo' in NGICwb.sheetnames:
    CivilDF = pd.read_excel(NGICRatebook, sheet_name='CivilAuthorityIncreasedRadiusCo', skiprows=11).rename(columns={"Radius":"Radius In Miles"})
else:
    CivilDF = pd.read_excel(CWRatebook, sheet_name='CivilAuthorityIncreasedRadiusCo', skiprows=11).rename(columns={"Radius":"Radius in Miles"})

#</editor-fold>

#<editor-fold desc="Computer DFs">

if 'ComputerandFundsTransferFraudBa' in NGICwb.sheetnames:
    ComputerDF = pd.read_excel(NGICRatebook, sheet_name='ComputerandFundsTransferFraudBa', skiprows=11)
else:
    ComputerDF = pd.read_excel(CWRatebook, sheet_name='ComputerandFundsTransferFraudBa', skiprows=11)

if 'ComputerandFundsTransferFraudAn' in NGICwb.sheetnames:
    ComputerAnnualDF = pd.read_excel(NGICRatebook,sheet_name='ComputerandFundsTransferFraudAn', skiprows=11).fillna('0').astype(str)
else:
    ComputerAnnualDF = pd.read_excel(CWRatebook,sheet_name='ComputerandFundsTransferFraudAn', skiprows=11).fillna('0').astype(str)

ComputerAnnualDF['Total Sales Min'] = ComputerAnnualDF['Total Sales Min'].apply(lambda x:"{:,}".format(int(float(x))))
ComputerAnnualDF['Total Sales Max'] = ComputerAnnualDF['Total Sales Max'].apply(lambda x:"{:,}".format(int(float(x))))
ComputerAnnualDF.iloc[0,0] = 0
ComputerAnnualDF['Sales'] = ComputerAnnualDF['Total Sales Min'].astype(str) + '-' + ComputerAnnualDF['Total Sales Max'].astype(str)
ComputerAnnualDF = ComputerAnnualDF.drop(columns=['Total Sales Min','Total Sales Max'])
ComputerAnnualDF = ComputerAnnualDF[['Sales']+[col for col in ComputerAnnualDF.columns if col!='Sales']]
ComputerAnnualDF = ComputerAnnualDF.iloc[0:5]
ComputerAnnualDF.iloc[4,0] = "over 25,000,001"
ComputerAnnualDF.iloc[4,1] = "2.00 + .10 for each addn'l 10,000,000"

#</editor-fold>

#<editor-fold desc="Forgery or Alteration DF">
if 'ForgeryOrAlterationLimitsRateab' in NGICwb.sheetnames:
    ForgeryDF = pd.read_excel(NGICRatebook, sheet_name='ForgeryOrAlterationLimitsRateab', skiprows=11)
else:
    ForgeryDF = pd.read_excel(CWRatebook, sheet_name='ForgeryOrAlterationLimitsRateab', skiprows=11)

if 'ForgeryorAlterationLimitsBaseRa' in NGICwb.sheetnames:
    ForgeryDF2 = pd.read_excel(NGICRatebook, sheet_name='ForgeryorAlterationLimitsBaseRa', skiprows=11)
else:
    ForgeryDF2 = pd.read_excel(CWRatebook, sheet_name='ForgeryorAlterationLimitsBaseRa', skiprows=11)

ForgeryDF2["Each Add'l Rateable Employee"] = ForgeryDF2['Premium Per Rateable Employee >5']
ForgeryDF2.rename(columns={"Premium < 5 Rateable Employees":"1-5 Rateable Employees"},inplace=True)
ForgeryDF2['Limit'] = ForgeryDF2['Limit'].apply('{:,}'.format)

#</editor-fold>

#<editor-fold desc="LPDP Dataframe DF">
#Pull in dataframe
if 'Large Premium Discount Factor_E' in NGICwb.sheetnames:
    LPDFDF = pd.read_excel(NGICRatebook, sheet_name='Large Premium Discount Factor_E', skiprows=11).rename(columns={"PremiumRangeMinimum":"Premium Range Minimum", "PremiumRangeMaximum":"Premium Range Maximum"})
else:
    LPDFDF = pd.read_excel(CWRatebook, sheet_name='Large Premium Discount Factor_E', skiprows=11).rename(columns={"PremiumRangeMinimum":"Premium Range Minimum","PremiumRangeMaximum":"Premium Range Maximum"})
#Changing the column label for this dataframe
LPDFDF['Premium Range Maximum'] = LPDFDF['Premium Range Maximum'].astype(str)
#Adjusting deductible maximum limit value (iloc)
LPDFDF.iloc[0,1] = "999,999,999"
#</editor-fold>

#<editor-fold desc="Counterfeit Currency 313 DF">
if 'MoneyOrdersAndCounterfeitLimits' in NGICwb.sheetnames:
    CFCurrDF = pd.read_excel(NGICRatebook, sheet_name='MoneyOrdersAndCounterfeitLimits', skiprows=11)
else:
    CFCurrDF = pd.read_excel(CWRatebook, sheet_name='MoneyOrdersAndCounterfeitLimits', skiprows=11)
#</editor-fold>

#<editor-fold desc="Money and Securities DF">
if 'MoneyandSecuritiesOccupancyFact' in NGICwb.sheetnames:
    MSDF = pd.read_excel(NGICRatebook, sheet_name='MoneyandSecuritiesOccupancyFact', skiprows=11)
else:
    MSDF = pd.read_excel(CWRatebook, sheet_name='MoneyandSecuritiesOccupancyFact', skiprows=11)
MSDF = MSDF.pivot_table(index=['Inside','Outside'],columns='Occupancy',values='Factor').reset_index()
MSDF = MSDF.rename(columns={'Inside':'Inside Limit','Outside':'Outside Limit'})
MSDF['Inside Limit'] = MSDF['Inside Limit'].apply('{:,}'.format)
MSDF['Outside Limit'] = MSDF['Outside Limit'].apply('{:,}'.format)
#</editor-fold>

#<editor-fold desc="BI ALS DF">
if 'BusinessIncomeActualLossSus (1)' in NGICwb.sheetnames:
    BusinessDF = pd.read_excel(NGICRatebook, sheet_name='BusinessIncomeActualLossSus (1)', skiprows=11).rename(columns={"TypeOfRisk":"Type Of Risk","PeriodOfRestoration":"Period of Restoration"})
else:
    BusinessDF = pd.read_excel(CWRatebook, sheet_name='BusinessIncomeActualLossSus (1)', skiprows=11).rename(columns={"TypeOfRisk":"Type Of Risk","PeriodOfRestoration":"Period of Restoration"})
#</editor-fold>

#<editor-fold desc="All Other DFs">
if 'ExpenseConstant_Ext' in NGICwb.sheetnames:
    ExpConstDF = pd.read_excel(NGICRatebook, sheet_name='ExpenseConstant_Ext', skiprows=11)
else:
    ExpConstDF = pd.read_excel(CWRatebook, sheet_name='ExpenseConstant_Ext', skiprows=11)

ExpConstDF.iloc[0, 0] = "All Policies"
ExpConstDF['Premium'] = '$' + ExpConstDF['Premium'].astype(str)


if 'PolicyMinimumPremium_Ext' in NGICwb.sheetnames:
    PolicyMinDF = pd.read_excel(NGICRatebook, sheet_name='PolicyMinimumPremium_Ext', skiprows=11)
else:
    PolicyMinDF = pd.read_excel(CWRatebook, sheet_name='PolicyMinimumPremium_Ext', skiprows=11)

PolicyMinDF.iloc[0,0] = "All Policies"
PolicyMinDF['Minimum'] = '$' + PolicyMinDF['Minimum'].astype(str)

if 'Nut Hullers And Processors Cove' in NGICwb.sheetnames:
    NutHullProcDF = pd.read_excel(NGICRatebook, sheet_name='Nut Hullers And Processors Cove', skiprows=11).drop(columns = 'Constant')
else:
    NutHullProcDF = pd.read_excel(CWRatebook, sheet_name='Nut Hullers And Processors Cove', skiprows=11).drop(columns = 'Constant')

if 'NutHullersAndProcessorsCoverage' in NGICwb.sheetnames:
    NutHullILF = pd.read_excel(NGICRatebook, sheet_name = 'NutHullersAndProcessorsCoverage', skiprows=11).rename(columns={"StockPile Occurrence Limit":"Limit Per Stockpile Per Occurrence"})
else:
    NutHullILF = pd.read_excel(CWRatebook, sheet_name = 'NutHullersAndProcessorsCoverage', skiprows=11).rename(columns={"StockPile Occurrence Limit":"Limit Per Stockpile Per Occurrence"})

if 'CosmeticExclusiononCoverageforS' in NGICwb.sheetnames:
    CosmeticDF = pd.read_excel(NGICRatebook, sheet_name='CosmeticExclusiononCoverageforS', skiprows=11).drop(columns='Constant')
else:
    CosmeticDF = pd.read_excel(CWRatebook, sheet_name='CosmeticExclusiononCoverageforS', skiprows=11).drop(columns='Constant')

if 'TerrsmPrsnlPropBaseRate' in NGICwb.sheetnames:
    TerrorismDF = pd.read_excel(NGICRatebook, sheet_name='TerrorismLoadFactor_Ext', skiprows=11).rename(columns={"Rate":"Factor"}).drop(columns='Terrorism Territory Tier')
else:
    TerrorismDF = pd.read_excel(CWRatebook, sheet_name='TerrorismLoadFactor_Ext', skiprows=11).rename(columns={"Rate":"Factor"}).drop(columns='Terrorism Territory Tier')

if 'Controlled Atmosphere and S (1)' in NGICwb.sheetnames:
    ControlAtmDF = pd.read_excel(NGICRatebook, sheet_name='Controlled Atmosphere and S (1)',skiprows=11).rename(columns={"Premium": "Base Rate"}).drop(columns='Constant')
else:
    ControlAtmDF = pd.read_excel(CWRatebook, sheet_name='Controlled Atmosphere and S (1)',skiprows=11).rename(columns={"Premium": "Base Rate"}).drop(columns='Constant')

if 'ControlledAtmosphereStorage (1)' in NGICwb.sheetnames:
    ControlAtChInjDF = pd.read_excel(NGICRatebook, sheet_name='ControlledAtmosphereStorage (1)',skiprows=11).rename(columns={"Rate": "Base Rate"})
else:
    ControlAtChInjDF = pd.read_excel(CWRatebook, sheet_name='ControlledAtmosphereStorage (1)',skiprows=11).rename(columns={"Rate": "Base Rate"})

if 'Hops Growers Coverage_Ext' in NGICwb.sheetnames:
    HopsDF = pd.read_excel(NGICRatebook, sheet_name='Hops Growers Coverage_Ext', skiprows=11).rename(columns={"Premium":"Rate"}).drop(columns='Constant')
else:
    HopsDF = pd.read_excel(CWRatebook, sheet_name='Hops Growers Coverage_Ext', skiprows=11).rename(columns={"Premium":"Rate"}).drop(columns='Constant')

if 'Fruit Trees, Trellises, Sta (1)' in NGICwb.sheetnames:
    FruitDF = pd.read_excel(NGICRatebook, sheet_name='Fruit Trees, Trellises, Sta (1)', skiprows=11).rename(columns={"Premium":"Rate"}).drop(columns='Constant')
else:
    FruitDF = pd.read_excel(CWRatebook, sheet_name='Fruit Trees, Trellises, Sta (1)', skiprows=11).rename(columns={"Premium":"Rate"}).drop(columns='Constant')

if 'WineryEndorsementPremium_Ext' in NGICwb.sheetnames:
    WineEndDF = pd.read_excel(NGICRatebook, sheet_name='WineryEndorsementPremium_Ext', skiprows=11).drop(columns='Constant')
else:
    WineEndDF = pd.read_excel(CWRatebook, sheet_name='WineryEndorsementPremium_Ext', skiprows=11).drop(columns='Constant')

if 'WineContaminationIncreasedLimit' in NGICwb.sheetnames:
    WineConDF = pd.read_excel(NGICRatebook, sheet_name='WineContaminationIncreasedLimit', skiprows=11).rename(columns={'Premium':'Rate'})
else:
    WineConDF = pd.read_excel(CWRatebook, sheet_name='WineContaminationIncreasedLimit', skiprows=11).rename(columns={'Premium':'Rate'})
if 'WineLeakageIncreasedLimits_Ext' in NGICwb.sheetnames:
    WineLeakDF = pd.read_excel(NGICRatebook, sheet_name='WineLeakageIncreasedLimits_Ext', skiprows=11).rename(columns={'Premium':'Rate'})
else:
    WineLeakDF = pd.read_excel(CWRatebook, sheet_name='WineLeakageIncreasedLimits_Ext', skiprows=11).rename(columns={'Premium':'Rate'})
if 'WineProcessingErrorsIncreasedLi' in NGICwb.sheetnames:
    WineProDF = pd.read_excel(NGICRatebook, sheet_name='WineProcessingErrorsIncreasedLi', skiprows=11).rename(columns={'Premium':'Rate'})
else:
    WineProDF = pd.read_excel(CWRatebook, sheet_name='WineProcessingErrorsIncreasedLi', skiprows=11).rename(columns={'Premium':'Rate'})

WineConDF.insert(0, 'Coverage', 'Wine Contamination')
WineLeakDF.insert(0, 'Coverage', 'Wine Leakage')
WineProDF.insert(0, 'Coverage', 'Wine Processing Errors')
WineConDF = WineConDF.tail(3)
WineLeakDF = WineLeakDF.tail(3)
WineProDF = WineProDF.tail(3)
WineryDF = pd.concat([WineConDF,WineLeakDF,WineProDF],ignore_index=True)

if 'WineLeakageEndorsementMinimum_E' in NGICwb.sheetnames:
    WineLkgDF1 = pd.read_excel(NGICRatebook, sheet_name='WineLeakageEndorsementMinimum_E', skiprows=11)
else:
    WineLkgDF1 = pd.read_excel(CWRatebook, sheet_name='WineLeakageEndorsementMinimum_E', skiprows=11)

WineLkgDF1.iloc[0,0] = 'Minimum Charge'
if 'WineLeakageEndorsementRate_Ext' in NGICwb.sheetnames:
    WineLkgDF2 = pd.read_excel(NGICRatebook, sheet_name='WineLeakageEndorsementRate_Ext', skiprows=11)
else:
    WineLkgDF2 = pd.read_excel(CWRatebook, sheet_name='WineLeakageEndorsementRate_Ext', skiprows=11)


if 'HumanSvcWrkplaceViolncLossofInc' in NGICwb.sheetnames:
    HumSvcsDF = pd.read_excel(NGICRatebook, sheet_name='HumanSvcWrkplaceViolncLossofInc', skiprows=11).drop(columns='Constant')
else:
    HumSvcsDF = pd.read_excel(CWRatebook, sheet_name='HumanSvcWrkplaceViolncLossofInc', skiprows=11).drop(columns='Constant')

if 'ClientsPropertyCoverageFactor_E' in NGICwb.sheetnames:
    ClientsDF = pd.read_excel(NGICRatebook, sheet_name='ClientsPropertyCoverageFactor_E', skiprows=11).drop(columns='Constant')
else:
    ClientsDF = pd.read_excel(CWRatebook, sheet_name='ClientsPropertyCoverageFactor_E', skiprows=11).drop(columns='Constant')

if 'HumanServicesEnhancementCoverag' in NGICwb.sheetnames:
    HumSvcsEnhDF = pd.read_excel(NGICRatebook, sheet_name='HumanServicesEnhancementCoverag', skiprows=11).drop(columns='Constant')
else:
    HumSvcsEnhDF = pd.read_excel(CWRatebook, sheet_name='HumanServicesEnhancementCoverag', skiprows=11).drop(columns='Constant')

if 'HumanServicesPropertyEndors (1)' in NGICwb.sheetnames:
    HumSvcsPetDF = pd.read_excel(NGICRatebook, sheet_name='HumanServicesPropertyEndors (1)', skiprows=11).drop(columns='Constant')
else:
    HumSvcsPetDF = pd.read_excel(CWRatebook, sheet_name='HumanServicesPropertyEndors (1)', skiprows=11).drop(columns='Constant')

if 'WaterDamageDeductibleWeight_Ext' in NGICwb.sheetnames:
    WtrDmgDF = pd.read_excel(NGICRatebook, sheet_name='WaterDamageDeductibleWeight_Ext', skiprows=11).drop(columns='Constant')
else:
    WtrDmgDF = pd.read_excel(CWRatebook, sheet_name='WaterDamageDeductibleWeight_Ext', skiprows=11).drop(columns='Constant')

if 'SeniorLivingCommunitiesProp (1)' in NGICwb.sheetnames:
    SeniorDF1 = pd.read_excel(NGICRatebook, sheet_name='SeniorLivingCommunitiesProp (1)', skiprows=11)
else:
    SeniorDF1 = pd.read_excel(CWRatebook, sheet_name='SeniorLivingCommunitiesProp (1)', skiprows=11)
if 'SeniorLivingCommunitiesProp (2)' in NGICwb.sheetnames:
    SeniorDF2 = pd.read_excel(NGICRatebook, sheet_name='SeniorLivingCommunitiesProp (2)', skiprows=11)
else:
    SeniorDF2 = pd.read_excel(CWRatebook, sheet_name='SeniorLivingCommunitiesProp (2)', skiprows=11)
if 'SeniorLivingCommunitiesProp (3)' in NGICwb.sheetnames:
    SeniorDF3 = pd.read_excel(NGICRatebook, sheet_name='SeniorLivingCommunitiesProp (3)', skiprows=11).rename(columns={"Premium":"Minimum"})
else:
    SeniorDF3 = pd.read_excel(CWRatebook, sheet_name='SeniorLivingCommunitiesProp (3)', skiprows=11).rename(columns={"Premium":"Minimum"})

if 'Premium' in SeniorDF1.columns:
    SeniorDFmin = SeniorDF1
    if 'Maximum' in SeniorDF2.columns:
        SeniorDFmax = SeniorDF2
        SeniorDFfac = SeniorDF3
    else:
        SeniorDFmax = SeniorDF3
        SeniorDFfac = SeniorDF2
elif 'Premium' in SeniorDF2.columns:
    SeniorDFmin = SeniorDF2
    if 'Maximum' in SeniorDF1.columns:
        SeniorDFmax = SeniorDF1
        SeniorDFfac = SeniorDF3
    else:
        SeniorDFmax = SeniorDF3
        SeniorDFfac = SeniorDF1
else:
    SeniorDFmin = SeniorDF3
    if 'Maximum' in SeniorDF1.columns:
        SeniorDFmax = SeniorDF1
        SeniorDFfac = SeniorDF2
    else:
        SeniorDFmax = SeniorDF2
        SeniorDFfac = SeniorDF1

SeniorDF = SeniorDFmin.merge(SeniorDFmax,on='Constant')
SeniorDF = SeniorDF.merge(SeniorDFfac,on='Constant').drop(columns='Constant')

if 'PropertyProtectionPlusEndor (3)' in NGICwb.sheetnames:
    PropPPDF1 = pd.read_excel(NGICRatebook, sheet_name='PropertyProtectionPlusEndor (3)', skiprows=5, header=None)
else:
    PropPPDF1 = pd.read_excel(CWRatebook, sheet_name='PropertyProtectionPlusEndor (3)', skiprows=5, header=None)

if 'PropertyProtectionPlusEndor (2)' in NGICwb.sheetnames:
    PropPPDF2 = pd.read_excel(NGICRatebook, sheet_name='PropertyProtectionPlusEndor (2)', skiprows=5, header=None)
else:
    PropPPDF2 = pd.read_excel(CWRatebook, sheet_name='PropertyProtectionPlusEndor (2)', skiprows=5, header=None)

if 'PropertyProtectionPlusEndor (1)' in NGICwb.sheetnames:
    PropPPDF3 = pd.read_excel(NGICRatebook, sheet_name='PropertyProtectionPlusEndor (1)', skiprows=5, header=None)
else:
    PropPPDF3 = pd.read_excel(CWRatebook, sheet_name='PropertyProtectionPlusEndor (1)',skiprows=5, header=None)

if 'PropertyProtectionPlusEndorsementMinimum_Ext' in PropPPDF1.iloc[0,1]:
    PropPPDFmin = PropPPDF1
    if 'PropertyProtectionPlusEndorsementMaximum_Ext' in PropPPDF2.iloc[0,1]:
        PropPPDFmax = PropPPDF2
        PropPPDFfac = PropPPDF3
    else:
        PropPPDFmax = PropPPDF3
        PropPPDFfac = PropPPDF2
elif 'PropertyProtectionPlusEndorsementMaximum_Ext' in PropPPDF1.iloc[0,1]:
    PropPPDFmax = PropPPDF1
    if 'PropertyProtectionPlusEndorsementMinimum_Ext' in PropPPDF2.iloc[0,1]:
        PropPPDFmin = PropPPDF2
        PropPPDFfac = PropPPDF3
    else:
        PropPPDFmin = PropPPDF3
        PropPPDFfac = PropPPDF2
else:
    PropPPDFfac = PropPPDF1
    if 'PropertyProtectionPlusEndorsementMinimum_ext' in PropPPDF2.iloc[0,1]:
        PropPPDFmin = PropPPDF2
        PropPPDFmax = PropPPDF3
    else:
        PropPPDFmin = PropPPDF3
        PropPPDFmax = PropPPDF2

PropPPDFmin = PropPPDFmin[7:].set_axis(PropPPDFmin.iloc[6],axis='columns').reset_index(drop=True)
PropPPDFmax = PropPPDFmax[7:].set_axis(PropPPDFmax.iloc[6],axis='columns').reset_index(drop=True)
PropPPDFfac = PropPPDFfac[7:].set_axis(PropPPDFfac.iloc[6],axis='columns').reset_index(drop=True)
PropPPDFmin = PropPPDFmin.rename(columns={"Constant":"Option", "Limit":"Minimum Charge"})
PropPPDFmax = PropPPDFmax.rename(columns={"Constant":"Option", "Limit":"Maximum Charge"})
PropPPDFfac = PropPPDFfac.rename(columns={"Constant":"Option", "Limit":"Factor"})
PropPPDF = PropPPDFmin.merge(PropPPDFmax, on='Option')
PropPPDF = PropPPDF.merge(PropPPDFfac, on='Option')
PropPPDF['Option'] = pd.Categorical(PropPPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
PropPPDF = PropPPDF.sort_values('Option').reset_index(drop=True)

if 'BreweryProtectionPlusEndors (3)' in NGICwb.sheetnames:
    BrewPPDF1 = pd.read_excel(NGICRatebook, sheet_name='BreweryProtectionPlusEndors (3)', skiprows=5, header=None)
else:
    BrewPPDF1 = pd.read_excel(CWRatebook, sheet_name='BreweryProtectionPlusEndors (3)', skiprows=5, header=None)

if 'BreweryProtectionPlusEndors (2)' in NGICwb.sheetnames:
    BrewPPDF2 = pd.read_excel(NGICRatebook, sheet_name='BreweryProtectionPlusEndors (2)', skiprows=5, header=None)
else:
    BrewPPDF2 = pd.read_excel(CWRatebook, sheet_name='BreweryProtectionPlusEndors (2)', skiprows=5, header=None)

if 'BreweryProtectionPlusEndors (1)' in NGICwb.sheetnames:
    BrewPPDF3 = pd.read_excel(NGICRatebook, sheet_name='BreweryProtectionPlusEndors (1)', skiprows=5, header=None)
else:
    BrewPPDF3 = pd.read_excel(CWRatebook, sheet_name='BreweryProtectionPlusEndors (1)', skiprows=5, header=None)

if 'BreweryProtectionPlusEndorsementMinimum_Ext' in BrewPPDF1.iloc[0,1]:
    BrewPPDFmin = BrewPPDF1
    if 'BreweryProtectionPlusEndorsementMaximum_Ext' in BrewPPDF2.iloc[0,1]:
        BrewPPDFmax = BrewPPDF2
        BrewPPDFfac = BrewPPDF3
    else:
        BrewPPDFmax = BrewPPDF3
        BrewPPDFfac = BrewPPDF2
elif 'BreweryProtectionPlusEndorsementMaximum_Ext' in BrewPPDF1.iloc[0,1]:
    BrewPPDFmax = BrewPPDF1
    if 'BreweryProtectionPlusEndorsementMinimum_Ext' in BrewPPDF2.iloc[0,1]:
        BrewPPDFmin = BrewPPDF2
        BrewPPDFfac = BrewPPDF3
    else:
        BrewPPDFmin = BrewPPDF3
        BrewPPDFfac = BrewPPDF2
else:
    BrewPPDFfac = BrewPPDF1
    if 'BreweryProtectionPlusEndorsementMinimum_ext' in BrewPPDF2.iloc[0,1]:
        BrewPPDFmin = BrewPPDF2
        BrewPPDFmax = BrewPPDF3
    else:
        BrewPPDFmin = BrewPPDF3
        BrewPPDFmax = BrewPPDF2

BrewPPDFmin = BrewPPDFmin[7:].set_axis(BrewPPDFmin.iloc[6],axis='columns').reset_index(drop=True)
BrewPPDFmax = BrewPPDFmax[7:].set_axis(BrewPPDFmax.iloc[6],axis='columns').reset_index(drop=True)
BrewPPDFfac = BrewPPDFfac[7:].set_axis(BrewPPDFfac.iloc[6],axis='columns').reset_index(drop=True)
BrewPPDFmin = BrewPPDFmin.rename(columns={"Constant":"Option", "Minimum":"Minimum Charge"})
BrewPPDFmax = BrewPPDFmax.rename(columns={"Constant":"Option", "Maximum":"Maximum Charge"})
BrewPPDFfac = BrewPPDFfac.rename(columns={"Constant":"Option", "Limit":"Factor"})
BrewPPDF = BrewPPDFmin.merge(BrewPPDFmax, on='Option')
BrewPPDF = BrewPPDF.merge(BrewPPDFfac, on='Option')
BrewPPDF['Option'] = pd.Categorical(BrewPPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
BrewPPDF = BrewPPDF.sort_values('Option').reset_index(drop=True)

if 'GolfProtectionPlusEndorsementMi' in NGICwb.sheetnames:
    GolfPPDF1 = pd.read_excel(NGICRatebook, sheet_name='GolfProtectionPlusEndorsementMi', skiprows=11).rename(columns={"Minimum":"Minimum Charge"})
else:
    GolfPPDF1 = pd.read_excel(CWRatebook, sheet_name='GolfProtectionPlusEndorsementMi', skiprows=11).rename(columns={"Minimum":"Minimum Charge"})
if 'GolfProtectionPlusEndorsementMa' in NGICwb.sheetnames:
    GolfPPDF2 = pd.read_excel(NGICRatebook, sheet_name='GolfProtectionPlusEndorsementMa', skiprows=11).rename(columns={"Maximum":"Maximum Charge"})
else:
    GolfPPDF2 = pd.read_excel(CWRatebook, sheet_name='GolfProtectionPlusEndorsementMa', skiprows=11).rename(columns={"Maximum":"Maximum Charge"})
if 'GolfProtectionPlusEndorsementFa' in NGICwb.sheetnames:
    GolfPPDF3 = pd.read_excel(NGICRatebook, sheet_name='GolfProtectionPlusEndorsementFa', skiprows=11)
else:
    GolfPPDF3 = pd.read_excel(CWRatebook, sheet_name='GolfProtectionPlusEndorsementFa', skiprows=11)

GolfPPDF = GolfPPDF1.merge(GolfPPDF2,on="Option")
GolfPPDF = GolfPPDF.merge(GolfPPDF3,on="Option")
GolfPPDF['Option'] = pd.Categorical(GolfPPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
GolfPPDF = GolfPPDF.sort_values('Option').reset_index(drop=True)

if 'ManufacturerProtectionPlusE (3)' in NGICwb.sheetnames:
    ManuPPDF1 = pd.read_excel(NGICRatebook, sheet_name='ManufacturerProtectionPlusE (3)', skiprows=5,header=None)
else:
    ManuPPDF1 = pd.read_excel(CWRatebook, sheet_name='ManufacturerProtectionPlusE (3)', skiprows=5, header=None)
if 'ManufacturerProtectionPlusE (2)' in NGICwb.sheetnames:
    ManuPPDF2 = pd.read_excel(NGICRatebook, sheet_name='ManufacturerProtectionPlusE (2)', skiprows=5,header=None)
else:
    ManuPPDF2 = pd.read_excel(CWRatebook, sheet_name='ManufacturerProtectionPlusE (2)', skiprows=5, header=None)
if 'ManufacturerProtectionPlusE (1)' in NGICwb.sheetnames:
    ManuPPDF3 = pd.read_excel(NGICRatebook, sheet_name='ManufacturerProtectionPlusE (1)', skiprows=5,header=None)
else:
    ManuPPDF3 = pd.read_excel(CWRatebook, sheet_name='ManufacturerProtectionPlusE (1)', skiprows=5, header=None)

if 'ManufacturerProtectionPlusEndorsementMinimum_Ext' in ManuPPDF1.iloc[0,1]:
    ManuPPDFmin = ManuPPDF1
    if 'ManufacturerProtectionPlusEndorsementMaximum_Ext' in ManuPPDF2.iloc[0,1]:
        ManuPPDFmax = ManuPPDF2
        ManuPPDFfac = ManuPPDF3
    else:
        ManuPPDFmax = ManuPPDF3
        ManuPPDFfac = ManuPPDF2
elif 'ManufacturerProtectionPlusEndorsementMaximum_Ext' in ManuPPDF1.iloc[0,1]:
    ManuPPDFmax = ManuPPDF1
    if 'ManufacturerProtectionPlusEndorsementMinimum_Ext' in ManuPPDF2.iloc[0,1]:
        ManuPPDFmin = ManuPPDF2
        ManuPPDFfac = ManuPPDF3
    else:
        ManuPPDFmin = ManuPPDF3
        ManuPPDFfac = ManuPPDF2
else:
    ManuPPDFfac = ManuPPDF1
    if 'ManufacturerProtectionPlusEndorsementMinimum_ext' in ManuPPDF2.iloc[0,1]:
        ManuPPDFmin = ManuPPDF2
        ManuPPDFmax = ManuPPDF3
    else:
        ManuPPDFmin = ManuPPDF3
        ManuPPDFmax = ManuPPDF2

ManuPPDFmin = ManuPPDFmin[7:].set_axis(ManuPPDFmin.iloc[6],axis='columns').reset_index(drop=True)
ManuPPDFmax = ManuPPDFmax[7:].set_axis(ManuPPDFmax.iloc[6],axis='columns').reset_index(drop=True)
ManuPPDFfac = ManuPPDFfac[7:].set_axis(ManuPPDFfac.iloc[6],axis='columns').reset_index(drop=True)
ManuPPDFmin = ManuPPDFmin.rename(columns={"Constant":"Option", "Minimum":"Minimum Charge"})
ManuPPDFmax = ManuPPDFmax.rename(columns={"Constant":"Option", "Maximum":"Maximum Charge"})
ManuPPDFfac = ManuPPDFfac.rename(columns={"Constant":"Option", "Limit":"Factor"})
ManuPPDF = ManuPPDFmin.merge(ManuPPDFmax, on='Option')
ManuPPDF = ManuPPDF.merge(ManuPPDFfac, on='Option')
ManuPPDF['Option'] = pd.Categorical(ManuPPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
ManuPPDF = ManuPPDF.sort_values('Option').reset_index(drop=True)

if 'RestaurantProtectionPlusEnd (3)' in NGICwb.sheetnames:
    RestPPDF1 = pd.read_excel(NGICRatebook, sheet_name='RestaurantProtectionPlusEnd (3)', skiprows=5,header=None)
else:
    RestPPDF1 = pd.read_excel(CWRatebook, sheet_name='RestaurantProtectionPlusEnd (3)', skiprows=5,header=None)
if 'RestaurantProtectionPlusEnd (2)' in NGICwb.sheetnames:
    RestPPDF2 = pd.read_excel(NGICRatebook, sheet_name='RestaurantProtectionPlusEnd (2)', skiprows=5,header=None)
else:
    RestPPDF2 = pd.read_excel(CWRatebook, sheet_name='RestaurantProtectionPlusEnd (2)', skiprows=5, header=None)
if 'RestaurantProtectionPlusEnd (1)' in NGICwb.sheetnames:
    RestPPDF3 = pd.read_excel(NGICRatebook, sheet_name='RestaurantProtectionPlusEnd (1)', skiprows=5, header=None)
else:
    RestPPDF3 = pd.read_excel(CWRatebook, sheet_name='RestaurantProtectionPlusEnd (1)',skiprows=5, header=None)

if 'RestaurantProtectionPlusEndorsementMinimum_Ext' in RestPPDF1.iloc[0,1]:
    RestPPDFmin = RestPPDF1
    if 'RestaurantProtectionPlusEndorsementMaximum_Ext' in RestPPDF2.iloc[0,1]:
        RestPPDFmax = RestPPDF2
        RestPPDFfac = RestPPDF3
    else:
        RestPPDFmax = RestPPDF3
        RestPPDFfac = RestPPDF2
elif 'RestaurantProtectionPlusEndorsementMaximum_Ext' in RestPPDF1.iloc[0,1]:
    RestPPDFmax = RestPPDF1
    if 'RestaurantProtectionPlusEndorsementMinimum_Ext' in RestPPDF2.iloc[0,1]:
        RestPPDFmin = RestPPDF2
        RestPPDFfac = RestPPDF3
    else:
        RestPPDFmin = RestPPDF3
        RestPPDFfac = RestPPDF2
else:
    RestPPDFfac = RestPPDF1
    if 'RestaurantProtectionPlusEndorsementMinimum_ext' in RestPPDF2.iloc[0,1]:
        RestPPDFmin = RestPPDF2
        RestPPDFmax = RestPPDF3
    else:
        RestPPDFmin = RestPPDF3
        RestPPDFmax = RestPPDF2

RestPPDFmin = RestPPDFmin[7:].set_axis(RestPPDFmin.iloc[6],axis='columns').reset_index(drop=True)
RestPPDFmax = RestPPDFmax[7:].set_axis(RestPPDFmax.iloc[6],axis='columns').reset_index(drop=True)
RestPPDFfac = RestPPDFfac[7:].set_axis(RestPPDFfac.iloc[6],axis='columns').reset_index(drop=True)
RestPPDFmin = RestPPDFmin.rename(columns={"Constant":"Option", "Minimum":"Minimum Charge"})
RestPPDFmax = RestPPDFmax.rename(columns={"Constant":"Option", "Maximum":"Maximum Charge"})
RestPPDFfac = RestPPDFfac.rename(columns={"Constant":"Option", "Limit":"Factor"})
RestPPDF = RestPPDFmin.merge(RestPPDFmax, on='Option')
RestPPDF = RestPPDF.merge(RestPPDFfac, on='Option')
RestPPDF['Option'] = pd.Categorical(RestPPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
RestPPDF = RestPPDF.sort_values('Option').reset_index(drop=True)

if 'WholesalerProtectionPlusEnd (3)' in NGICwb.sheetnames:
    WholePPDF1 = pd.read_excel(NGICRatebook, sheet_name='WholesalerProtectionPlusEnd (3)', skiprows=5,header=None)
else:
    WholePPDF1 = pd.read_excel(CWRatebook, sheet_name='WholesalerProtectionPlusEnd (3)', skiprows=5,header=None)
if 'WholesalerProtectionPlusEnd (2)' in NGICwb.sheetnames:
    WholePPDF2 = pd.read_excel(NGICRatebook, sheet_name='WholesalerProtectionPlusEnd (2)', skiprows=5,header=None)
else:
    WholePPDF2 = pd.read_excel(CWRatebook, sheet_name='WholesalerProtectionPlusEnd (2)', skiprows=5,header=None)
if 'WholesalerProtectionPlusEnd (1)' in NGICwb.sheetnames:
    WholePPDF3 = pd.read_excel(NGICRatebook, sheet_name='WholesalerProtectionPlusEnd (1)', skiprows=5,header=None)
else:
    WholePPDF3 = pd.read_excel(CWRatebook, sheet_name='WholesalerProtectionPlusEnd (1)', skiprows=5,header=None)

if 'WholesalerProtectionPlusEndorsementMinimum_Ext' in WholePPDF1.iloc[0,1]:
    WholePPDFmin = WholePPDF1
    if 'WholesalerProtectionPlusEndorsementMaximum_Ext' in WholePPDF2.iloc[0,1]:
        WholePPDFmax = WholePPDF2
        WholePPDFfac = WholePPDF3
    else:
        WholePPDFmax = WholePPDF3
        WholePPDFfac = WholePPDF2
elif 'WholesalerProtectionPlusEndorsementMaximum_Ext' in WholePPDF1.iloc[0,1]:
    WholePPDFmax = WholePPDF1
    if 'WholesalerProtectionPlusEndorsementMinimum_Ext' in WholePPDF2.iloc[0,1]:
        WholePPDFmin = WholePPDF2
        WholePPDFfac = WholePPDF3
    else:
        WholePPDFmin = WholePPDF3
        WholePPDFfac = WholePPDF2
else:
    WholePPDFfac = WholePPDF1
    if 'WholesalerProtectionPlusEndorsementMinimum_ext' in WholePPDF2.iloc[0,1]:
        WholePPDFmin = WholePPDF2
        WholePPDFmax = WholePPDF3
    else:
        WholePPDFmin = WholePPDF3
        WholePPDFmax = WholePPDF2

WholePPDFmin = WholePPDFmin[7:].set_axis(WholePPDFmin.iloc[6],axis='columns').reset_index(drop=True)
WholePPDFmax = WholePPDFmax[7:].set_axis(WholePPDFmax.iloc[6],axis='columns').reset_index(drop=True)
WholePPDFfac = WholePPDFfac[7:].set_axis(WholePPDFfac.iloc[6],axis='columns').reset_index(drop=True)
WholePPDFmin = WholePPDFmin.rename(columns={"Constant":"Option", "Minimum":"Minimum Charge"})
WholePPDFmax = WholePPDFmax.rename(columns={"Constant":"Option", "Maximum":"Maximum Charge"})
WholePPDFfac = WholePPDFfac.rename(columns={"Constant":"Option", "Limit":"Factor"})
WholePPDF = WholePPDFmin.merge(WholePPDFmax, on='Option')
WholePPDF = WholePPDF.merge(WholePPDFfac, on='Option')
WholePPDF['Option'] = pd.Categorical(WholePPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
WholePPDF = WholePPDF.sort_values('Option').reset_index(drop=True)

if 'HotelProtectionPlusEndorsem (1)' in NGICwb.sheetnames:
    HotelPPDF1 = pd.read_excel(NGICRatebook, sheet_name='HotelProtectionPlusEndorsem (1)', skiprows=5,header=None)
else:
    HotelPPDF1 = pd.read_excel(CWRatebook, sheet_name='HotelProtectionPlusEndorsem (1)', skiprows=5,header=None)
if 'HotelProtectionPlusEndorsementM' in NGICwb.sheetnames:
    HotelPPDF2 = pd.read_excel(NGICRatebook, sheet_name='HotelProtectionPlusEndorsementM', skiprows=5,header=None)
else:
    HotelPPDF2 = pd.read_excel(CWRatebook, sheet_name='HotelProtectionPlusEndorsementM', skiprows=5,header=None)
if 'HotelProtectionPlusEndorsementF' in NGICwb.sheetnames:
    HotelPPDF3 = pd.read_excel(NGICRatebook, sheet_name='HotelProtectionPlusEndorsementF', skiprows=5,header=None)
else:
    HotelPPDF3 = pd.read_excel(CWRatebook, sheet_name='HotelProtectionPlusEndorsementF',skiprows=5, header=None)

if 'HotelProtectionPlusEndorsementMinimum_Ext' in HotelPPDF1.iloc[0,1]:
    HotelPPDFmin = HotelPPDF1
    if 'HotelProtectionPlusEndorsementMaximum_Ext' in HotelPPDF2.iloc[0,1]:
        HotelPPDFmax = HotelPPDF2
        HotelPPDFfac = HotelPPDF3
    else:
        HotelPPDFmax = HotelPPDF3
        HotelPPDFfac = HotelPPDF2
elif 'HotelProtectionPlusEndorsementMaximum_Ext' in HotelPPDF1.iloc[0,1]:
    HotelPPDFmax = HotelPPDF1
    if 'HotelProtectionPlusEndorsementMinimum_Ext' in HotelPPDF2.iloc[0,1]:
        HotelPPDFmin = HotelPPDF2
        HotelPPDFfac = HotelPPDF3
    else:
        HotelPPDFmin = HotelPPDF3
        HotelPPDFfac = HotelPPDF2
else:
    HotelPPDFfac = HotelPPDF1
    if 'HotelProtectionPlusEndorsementMinimum_ext' in HotelPPDF2.iloc[0,1]:
        HotelPPDFmin = HotelPPDF2
        HotelPPDFmax = HotelPPDF3
    else:
        HotelPPDFmin = HotelPPDF3
        HotelPPDFmax = HotelPPDF2

HotelPPDFmin = HotelPPDFmin[7:].set_axis(HotelPPDFmin.iloc[6],axis='columns').reset_index(drop=True)
HotelPPDFmax = HotelPPDFmax[7:].set_axis(HotelPPDFmax.iloc[6],axis='columns').reset_index(drop=True)
HotelPPDFfac = HotelPPDFfac[7:].set_axis(HotelPPDFfac.iloc[6],axis='columns').reset_index(drop=True)
HotelPPDFmin = HotelPPDFmin.rename(columns={"Constant":"Option", "Minimum":"Minimum Charge"})
HotelPPDFmax = HotelPPDFmax.rename(columns={"Constant":"Option", "Maximum":"Maximum Charge"})
HotelPPDFfac = HotelPPDFfac.rename(columns={"Constant":"Option", "Limit":"Factor"})
HotelPPDF = HotelPPDFmin.merge(HotelPPDFmax, on='Option')
HotelPPDF = HotelPPDF.merge(HotelPPDFfac, on='Option')
HotelPPDF['Option'] = pd.Categorical(HotelPPDF['Option'], categories=['Gold','Platinum','Diamond'], ordered=True)
HotelPPDF = HotelPPDF.sort_values('Option').reset_index(drop=True)

if 'FranchiseUpgradeEndorsement_Ext' in NGICwb.sheetnames:
    FranUpgDF = pd.read_excel(NGICRatebook, sheet_name='FranchiseUpgradeEndorsement_Ext', skiprows=11).rename(columns={"Premium":"Base Rate"}).drop(columns='Constant')
else:
    FranUpgDF = pd.read_excel(CWRatebook, sheet_name='FranchiseUpgradeEndorsement_Ext', skiprows=11).rename(columns={"Premium":"Base Rate"}).drop(columns='Constant')
#remember header=False

#</editor-fold>

# <editor-fold desc="Create Segmentation Removal Tiering, LEAF, Age of Building Dataframes">
if v10.get() == 1 and v12.get() == 1:
    SegRem = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Segmentation Removal Pages\\Segmentation Removal Rate Pages.xlsx')
    SegTier = pd.read_excel(SegRem, sheet_name='Tiering', usecols="E:J,M:R,U:Z",skiprows=range(0,2))
    SegTier.columns = SegTier.iloc[0]
    SegTier = SegTier[1:]
    SegLeaf = pd.read_excel(SegRem, sheet_name='LEAF', usecols="E,H",skiprows=range(0,4))
    SegLeaf.columns = SegLeaf.iloc[0]
    SegLeaf = SegLeaf[1:]
    SegAOB = pd.read_excel(SegRem,sheet_name="Age of Building", usecols="F:AO",skiprows=range(0,1))
    SegAOB.columns = SegAOB.iloc[0]
    SegAOB = SegAOB[1:]

    BGIBLOIMM = pd.read_excel(MMRatebook, sheet_name='BasicGroupILOIFactorBldg', skiprows=11)
    BGIIBLOIMM = pd.read_excel(MMRatebook, sheet_name='BasicGroupIILOIFactorBldg', skiprows=11)
    SCOLBLOIMM = pd.read_excel(MMRatebook, sheet_name='BroadSpecialLOIFactorBldg', skiprows=11)
    BGIPPLOIMM = pd.read_excel(MMRatebook, sheet_name='BasicGroupILOIFactorPersProp', skiprows=11)
    BGIIPPLOIMM = pd.read_excel(MMRatebook, sheet_name='BasicGroupIILOIFactorPersProp', skiprows=11)
    SCOLPPLOIMM = pd.read_excel(MMRatebook, sheet_name='BroadSpecialLOIFactorPrsnlProp', skiprows=11)

    BGIBLOIMM = BGIBLOIMM.pivot(index='Limit', columns='ConstructionCode', values='Factor').reset_index(names=['Limit', 1, 2, 3, 4, 5, 6])
    BGIBLOIMM = BGIBLOIMM.drop(columns=[2, 3, 5, 6])
    BGIBLOIMM = BGIBLOIMM.drop(BGIBLOIMM.tail(1).index)
    BGIBLOIMM['Limit'] = BGIBLOIMM['Limit'].apply('{:,}'.format)
    BGIBLOIMM.iloc[-1, BGIBLOIMM.columns.get_loc('Limit')] = "10,000,000+"
    BGIBLOIMM = BGIBLOIMM.rename(columns={1: "Construction Group 1-3", 4: "Construction Group 4-6"})
    BGIBLOIMM = splitdf(BGIBLOIMM, 2).fillna('')

    BGIIBLOIMM['Limit'] = BGIIBLOIMM['Limit'].apply('{:,}'.format)
    BGIIBLOIMM = BGIIBLOIMM.drop(BGIIBLOIMM.tail(1).index)
    BGIIBLOIMM.iloc[-1, BGIIBLOIMM.columns.get_loc('Limit')] = BGIIBLOIMM.iloc[-1, BGIIBLOIMM.columns.get_loc('Limit')] + "+"
    BGIIBLOIMM = splitdf(BGIIBLOIMM, 2).fillna('')

    SCOLBLOIMM['Limit'] = SCOLBLOIMM['Limit'].apply('{:,}'.format)
    SCOLBLOIMM = SCOLBLOIMM.drop(SCOLBLOIMM.tail(1).index)
    SCOLBLOIMM.iloc[-1, SCOLBLOIMM.columns.get_loc('Limit')] = SCOLBLOIMM.iloc[-1, SCOLBLOIMM.columns.get_loc('Limit')] + "+"
    SCOLBLOIMM = splitdf(SCOLBLOIMM, 2).fillna('')

    BGIPPLOIMM = BGIPPLOIMM.pivot(index='Limit', columns='ConstructionCode', values='Factor').reset_index(names=['Limit', 1, 2, 3, 4, 5, 6])
    BGIPPLOIMM = BGIPPLOIMM.drop(columns=[2, 3, 5, 6])
    BGIPPLOIMM = BGIPPLOIMM.drop(BGIPPLOIMM.tail(1).index)
    BGIPPLOIMM['Limit'] = BGIPPLOIMM['Limit'].apply('{:,}'.format)
    BGIPPLOIMM.iloc[-1, BGIPPLOIMM.columns.get_loc('Limit')] = BGIPPLOIMM.iloc[-1, BGIPPLOIMM.columns.get_loc('Limit')] + "+"
    BGIPPLOIMM = BGIPPLOIMM.rename(columns={1: "Construction Group 1-3", 4: "Construction Group 4-6"})
    BGIPPLOIMM = splitdf(BGIPPLOIMM, 2).fillna('')

    BGIIPPLOIMM['Limit'] = BGIIPPLOIMM['Limit'].apply('{:,}'.format)
    BGIIPPLOIMM = BGIIPPLOIMM.drop(BGIIPPLOIMM.tail(1).index)
    BGIIPPLOIMM.iloc[-1, BGIIPPLOIMM.columns.get_loc('Limit')] = BGIIPPLOIMM.iloc[-1, BGIIPPLOIMM.columns.get_loc('Limit')] + "+"
    BGIIPPLOIMM = splitdf(BGIIPPLOIMM, 2).fillna('')

    SCOLPPLOIMM['Limit'] = SCOLPPLOIMM['Limit'].apply('{:,}'.format)
    SCOLPPLOIMM = SCOLPPLOIMM.drop(SCOLPPLOIMM.tail(1).index)
    SCOLPPLOIMM.iloc[-1, SCOLPPLOIMM.columns.get_loc('Limit')] = SCOLPPLOIMM.iloc[-1, SCOLPPLOIMM.columns.get_loc('Limit')] + "+"
    SCOLPPLOIMM = splitdf(SCOLPPLOIMM, 2).fillna('')

else:
    pass

# </editor-fold>

#<editor-fold desc="SM Index Sheet">
if v5.get() == 1:
    wb = Workbook()
    wb.active.title = "Index"
#Initializing Default Settings
fontName = 'Arial'
fontSize = 10
headerFontName = 'Arial'
headerFontSize = 10
footerFontName = 'Arial'
footerFontSize = 10
leftMargin = 0.5
rightMargin = 0.5
topMargin = 2.25
bottomMargin = 1.5
headerMargin = 1.5
footerMargin = 0.5

#Building Fonts and Formats
font = Font(name=fontName, size=fontSize)
fontBold = Font(name=fontName, size=fontSize, bold=True)
fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
fontBoldItalics = Font(name=fontName, size = fontSize, italic=True, bold=True)
fontBlue = Font(name=fontName, size=fontSize, color='0000FF')
headerFont = headerFontName + ',Bold'
footerFont = footerFontName + ',Bold'
rateFormat = '#,##0.000'
rateFormat2 = '#,##0.00'
codeFormat = '#,##0'
currencyFormat = '$#,##0'
percentFormat = '0.0%'
# </editor-fold>

# <editor-fold desc="SM LCM Excel Sheet">
if v5.get() == 1 and v1.get() == 1:
    ws = wb.create_sheet(title='LCM')
    formatwkstSM(wkstname=ws,titlerows='3',A1title='BASE RATE CALCULATION',A2title='Loss Cost Multiplier',dfname=SMLCMTable,statename=State,stabb=StateAbb, effdate=EffectiveDate)

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            cell = ws[char + str(row)]
            ws.column_dimensions[char].bestFit = True  # Using bestfit as the default option for column widths
            if row > 3 and cell.value is not None:  # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
                if row < 2:  # Applies bold font to rows 1-3, which are header rows
                    cell.font = fontBoldUnderline
                elif row == 2:
                    cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
                elif row == 4:  # Additional formatting for row 4 (table header row)
                    cell.number_format = rateFormat
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
                else:
                    cell.number_format = rateFormat  # Default format for rates
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    if NICOFRatebook != "Not found":
        ws.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws.oddFooter.left.text = "\nNationwide General Insurance Company"
    # Left footer
    ws.oddFooter.left.size = footerFontSize
    ws.oddFooter.left.font = footerFont
    # Center footer
    ws.oddFooter.center.text = StateAbb + " - SRP - &[Tab] - &P"
    ws.oddFooter.center.size = footerFontSize
    ws.oddFooter.center.font = footerFont
    # </editor-fold>

# <editor-fold desc="SM PMF Excel Sheet">
if v5.get() == 1 and v2.get() == 1:
    ws2 = wb.create_sheet(title='PMF')
    formatwkstSM(wkstname = ws2, titlerows = '3', A1title = 'PACKAGE MODIFICATION FACTOR', A2title = 'Blank', dfname = SMPMF, statename = State, stabb = StateAbb, effdate = EffectiveDate)

    for row in range(1, ws2.max_row + 1):
        for col in range(1, ws2.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            cell = ws2[char + str(row)]
            ws2.column_dimensions[char].bestFit = True  # Using bestfit as the default option for column widths
            if row > 2 and cell.value is not None:  # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                        right=Side(border_style='thin', color='00000000'),
                                        top=Side(border_style='thin', color='00000000'),
                                        bottom=Side(border_style='thin', color='00000000'))
                if row < 2:  # Applies bold font to rows 1-3, which are header rows
                    cell.font = fontBoldUnderline
                elif row == 2:
                    cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
                elif col == 1:
                    cell.number_format = codeFormat
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
                else:
                    cell.number_format = rateFormat2
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
        for col in ws2.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws2.column_dimensions[column].width = adjusted_width
    if NICOFRatebook != "Not found":
        ws2.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws2.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws2.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws2.oddFooter.left.text = "\nNationwide General Insurance Company"
    # Left footer
    ws2.oddFooter.left.size = footerFontSize
    ws2.oddFooter.left.font = footerFont
    # Center footer
    ws2.oddFooter.center.text = StateAbb + " - SRP - &[Tab] - &P"
    ws2.oddFooter.center.size = footerFontSize
    ws2.oddFooter.center.font = footerFont
# </editor-fold>

# <editor-fold desc="SM Crime Endorsement Excel Sheet">
if v5.get() == 1 and v1.get() == 1:
    ws3 = wb.create_sheet(title='Crime')
    ws3['A1'] = 'CRIME ENDORSEMENT'
    ws3['A2'] = 'Blank'
    for r in dataframe_to_rows(CrimeDed, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & len(list(r)) == 1:  # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            ws3['B4'] = list(r)[0]  # Manually adding the 1 index value to the header row in Excel
            continue
        ws3.append(r)
    ws3['A19'] = 'Blank'
    ws3['A20'] = 'Blank'
    ws3['A21'] = 'Blank'
    for r in dataframe_to_rows(SMCRIMELCMTable, False, True):
        ws3.append(r)

    ws3.page_setup.orientation = 'portrait'  # Landscape orientation for printing
    ws3.page_setup.blackAndWhite = False
    ws3.page_setup.firstPageNumber = 1  # Resetting the page counter for the footer on each worksheet
    ws3.page_setup.useFirstPageNumber = True
    ws3.sheet_view.showGridLines = False  # Turning off gridlines
    ws3.print_title_rows = '1:3'
    ws3.page_margins.left = leftMargin
    ws3.page_margins.right = rightMargin
    ws3.page_margins.top = topMargin
    ws3.page_margins.bottom = bottomMargin
    ws3.page_margins.header = headerMargin
    ws3.page_margins.footer = footerMargin
    ws3.print_options.horizontalCentered = True
    # Left Header
    if State == 'Florida':
        ws3.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        ws3.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    ws3.oddHeader.left.size = headerFontSize
    ws3.oddHeader.left.font = headerFont
    # Center header
    ws3.oddHeader.center.text = "\n\n" + State + " - Rate Pages"
    ws3.oddHeader.center.size = headerFontSize
    ws3.oddHeader.center.font = headerFont
    # Right header
    ws3.oddHeader.right.text = "Effective Date: " + EffectiveDate
    ws3.oddHeader.right.size = headerFontSize
    ws3.oddHeader.right.font = headerFont
    if NICOFRatebook != "Not found":
        ws3.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws3.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws3.oddFooter.left.text = "Nationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws3.oddFooter.left.text = "\nNationwide General Insurance Company"
    # Left footer
    ws3.oddFooter.left.size = footerFontSize
    ws3.oddFooter.left.font = footerFont
    # Center footer
    ws3.oddFooter.center.text = StateAbb + " - SRP - &[Tab] - &P"
    ws3.oddFooter.center.size = footerFontSize
    ws3.oddFooter.center.font = footerFont
    for row in range(1, ws3.max_row + 1):
        for col in range(1, ws3.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            cell = ws3[char + str(row)]
            ws3.column_dimensions[char].bestFit = True  # Using bestfit as the default option for column widths
            if 3 < row < 19 and cell.value is not None:  # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                        right=Side(border_style='thin', color='00000000'),
                                        top=Side(border_style='thin', color='00000000'),
                                        bottom=Side(border_style='thin', color='00000000'))
            if row < 2:  # Applies bold font to rows 1-3, which are header rows
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif col == 1:
                cell.number_format = codeFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            elif col == 2 and row < 17:
                cell.number_format = rateFormat2
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            else:
                cell.number_format = rateFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            if 18 < row < 22:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif row > 22 and cell.value is not None:
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                        right=Side(border_style='thin', color='00000000'),
                                        top=Side(border_style='thin', color='00000000'),
                                        bottom=Side(border_style='thin', color='00000000'))
        for col in ws3.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws3.column_dimensions[column].width = adjusted_width
    # </editor-fold>

# <editor-fold desc="SM Capping Page">
if v5.get() == 1 and v4.get() == 1:
    ws4 = wb.create_sheet(title='Capping')
    formatwkstSM(wkstname = ws4, titlerows = '3', A1title = 'RATE CAPPING RANGE', A2title =
    'Blank', dfname = SMCapTable, statename = State, stabb = StateAbb, effdate = EffectiveDate)

    for row in range(1, ws4.max_row + 1):
        for col in range(1, ws4.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            cell = ws4[char + str(row)]
            ws4.column_dimensions[char].bestFit = True  # Using bestfit as the default option for
            column widths
            if row > 2 and cell.value is not None:  # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                        right=Side(border_style='thin', color='00000000'),
                                        top=Side(border_style='thin', color='00000000'),
                                        bottom=Side(border_style='thin', color='00000000'))
            if row < 2:  # Applies bold font to rows 1-3, which are header rows
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif col == 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            elif col > 1:
                cell.number_format = percentFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
        for col in ws4.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws4.column_dimensions[column].width = adjusted_width
    if NICOFRatebook != "Not found":
        ws4.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws4.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws4.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws4.oddFooter.left.text = "\nNationwide General Insurance Company"
    # Left footer
    ws4.oddFooter.left.size = footerFontSize
    ws4.oddFooter.left.font = footerFont
    # Center footer
    ws4.oddFooter.center.text = StateAbb + " - SRP - &[Tab] - &P"
    ws4.oddFooter.center.size = footerFontSize
    ws4.oddFooter.center.font = footerFont
# </editor-fold>

    # <editor-fold desc="SM Territory Sheets">
    if v3.get() == 1 and v5.get() == 1:
        ws10 = wb.create_sheet(title='Territory')
        formatwkstSM(wkstname=ws10, titlerows = '3', A1title = 'TERRITORY ADJUSTMENT FACTORS',
        A2title = 'Blank', dfname = GridFacXL, statename = State, stabb = StateAbb, effdate =
        EffectiveDate)
        print("worksheet set up")
        # Define reusable styles
        thin_border = Border(left=Side(border_style='thin', color='00000000'),
                              right=Side(border_style='thin', color='00000000'),
                              top=Side(border_style='thin', color='00000000'),
                              bottom=Side(border_style='thin', color='00000000'))

        side_border = Border(left=Side(border_style='thin', color='00000000'),
                              right=Side(border_style='thin', color='00000000'))

        header_font = fontBoldUnderline
        subheader_font = Font(name=fontName, size=fontSize, color='FFFFFFFF', bold=True)

        # Apply styles and borders
        for row in ws10.iter_rows(min_row=1, max_row=ws10.max_row, max_col=ws10.max_column):
            for cell in row:
                r, c = cell.row, cell.column

                # Font and alignment
                if r == 1:
                    cell.font = header_font
                elif r == 2:
                    cell.font = subheader_font
                else:
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

                # Borders
                if r > 2:
                    if cell.value and str(cell.value).strip():
                        cell.border = thin_border
                    else:
                        cell.border = side_border

        # Bold font for row 3
        for cell in ws10["3:3"]:
            cell.font = fontBold

        # Apply number format for data rows
        for col in ws10.iter_cols(min_row=3):
            for cell in col:
                cell.number_format = "0.000"

        # Adjust column widths
        for col in ws10.iter_cols(min_row=3):
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
            ws10.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2

        # Set print area and page setup
        ws10.print_area = f"A1:{get_column_letter(ws10.max_column)}{ws10.max_row}"
        ws10.page_setup.fitToPage = True

        # Add page breaks every 120 rows
        pagefacnum = math.ceil(ws10.max_row / 120)
        for rnum in range(1, pagefacnum):
            ws10.row_breaks.append(Break(id=rnum * 120))

        ws10.page_setup.fitToWidth = 1
        ws10.page_setup.fitToHeight = pagefacnum
        print("formatting done")
    # </editor-fold>

    #<editor-fold desc="MM Index Sheet">
    if v10.get() == 1:
        wb2 = Workbook()
        wb2.active.title = "Index"
    #</editor-fold>

    # <editor-fold desc="MM LCM Excel Sheet">
    if v10.get() == 1 and v7.get() == 1:
        ws5 = wb2.create_sheet(title='LCM')
        formatwkstMM(wkstname = ws5, titlerows = '3', A1title = 'BASE RATE CALCULATION', A2title ='Loss Cost Multiplier', dfname = MMLCMTable, statename = State, stabb = StateAbb, effdate = EffectiveDate)

        for row in range(1, ws5.max_row + 1):
            for col in range(1, ws5.max_column + 1):
                char = get_column_letter(col)  # Letter representing the current column
                cell = ws5[char + str(row)]
                ws5.column_dimensions[char].bestFit = True  # Using bestfit as the default option for column widths
                if row > 3 and cell.value is not None:  # Adding a border to the table data
                    cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                          right=Side(border_style='thin', color='00000000'),
                                          top=Side(border_style='thin', color='00000000'),
                                          bottom=Side(border_style='thin', color='00000000'))
                if row < 2:  # Applies bold font to rows 1-3, which are header rows
                    cell.font = fontBoldUnderline
                elif row == 2:
                    cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
                elif row == 4:  # Additional formatting for row 4 (table header row)
                    cell.number_format = rateFormat
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
                else:
                    cell.number_format = rateFormat  # Default format for rates
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            for col in ws5.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws5.column_dimensions[column].width = adjusted_width
    # </editor-fold>

    # <editor-fold desc="MM PMF Excel Sheet">
    if v10.get() == 1 and v8.get() == 1:
        ws6 = wb2.create_sheet(title='PMF')
        formatwkstMM(wkstname = ws6, titlerows = '3', A1title = 'PACKAGE MODIFICATION FACTOR',
        A2title = 'Blank', dfname = MMPMF, statename = State, stabb = StateAbb, effdate = EffectiveDate)

        for row in range(1, ws6.max_row + 1):
            for col in range(1, ws6.max_column + 1):
                char = get_column_letter(col)  # Letter representing the current column
                cell = ws6[char + str(row)]
                ws6.column_dimensions[char].bestFit = True  # Using bestfit as the default option for
                column widths
                if row > 2 and cell.value is not None:  # Adding a border to the table data
                    cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                          right=Side(border_style='thin', color='00000000'),
                                          top=Side(border_style='thin', color='00000000'),
                                          bottom=Side(border_style='thin', color='00000000'))
                if row < 2:  # Applies bold font to rows 1-3, which are header rows
                    cell.font = fontBoldUnderline
                elif row == 2:
                    cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
                elif col == 1:
                    cell.number_format = codeFormat
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
                else:
                    cell.number_format = rateFormat2
                    cell.font = font
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            for col in ws6.columns:
                max_length = 0
                column = col[0].column_letter  # Get the column name
                for cell in col:
                    try:  # Necessary to avoid error on empty cells
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws6.column_dimensions[column].width = adjusted_width
    # </editor-fold>

    # <editor-fold desc="MM Crime Endorsement Excel Sheet">
    if v10.get() == 1 and v7.get() == 1:
        ws7 = wb2.create_sheet(title='Crime')
        ws7['A1'] = 'CRIME ENDORSEMENT'
        ws7['A2'] = 'Blank'
        for r in dataframe_to_rows(CrimeDed, False, True):
            # The header is the first row and the index is the second row, but they need to be on the same row in Excel
            if False & (len(list(r)) == 1): # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
                ws7['B4'] = list(r)[0] # Manually adding the 1 index value to the header row in Excel
                continue
            ws7.append(r)

        ws7['A19'] = 'Blank'
        ws7['A20'] = 'Blank'
        ws7['A21'] = 'Blank'
        for r in dataframe_to_rows(MMCRIMELCMTable, False, True):
            ws7.append(r)

        ws7.page_setup.orientation = 'portrait' # Landscape orientation for printing
        ws7.page_setup.blackAndWhite = False
        ws7.page_setup.firstPageNumber = 1 # Resetting the page counter for the footer on each worksheet
        ws7.page_setup.useFirstPageNumber = True
        ws7.sheet_view.showGridLines = False # Turning off gridlines
        ws7.print_title_rows = '1:3'
        ws7.page_margins.left = leftMargin
        ws7.page_margins.right = rightMargin
        ws7.page_margins.top = topMargin
        ws7.page_margins.bottom = bottomMargin
        ws7.page_margins.header = headerMargin
        ws7.page_margins.footer = footerMargin
        ws7.print_options.horizontalCentered = True

    # Left Header
    if State == 'Florida':
        ws7.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        ws7.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    ws7.oddHeader.left.size = headerFontSize
    ws7.oddHeader.left.font = headerFont
    
    # Center header
    ws7.oddHeader.center.text = "\n\n" + State + " Rate Pages"
    ws7.oddHeader.center.size = headerFontSize
    ws7.oddHeader.center.font = headerFont
    
    # Right header
    ws7.oddHeader.right.text = "Effective Date: " + EffectiveDate
    ws7.oddHeader.right.size = headerFontSize
    ws7.oddHeader.right.font = headerFont

    if StateAbb != "WI":
        ws7.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company \nNationwide Property & Casualty Insurance Company \nAllied Insurance Company of America"
    else:
        ws7.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company"

    # Left footer
    ws7.oddFooter.left.size = footerFontSize
    ws7.oddFooter.left.font = footerFont
    
    # Center footer
    ws7.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws7.oddFooter.center.size = footerFontSize
    ws7.oddFooter.center.font = footerFont

    for row in range(1, ws7.max_row + 1):
        for col in range(1, ws7.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws7[char + str(row)]
            ws7.column_dimensions[char].bestFit = True # Using bestfit as the default option for column widths
            if 3 < row < 19 and cell.value is not None: # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            if row < 2: # Applies bold font to rows 1-3, which are header rows
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif col == 1:
                cell.number_format = rateFormat2
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            elif col == 2 and row < 17:
                cell.number_format = rateFormat2
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            else:
                cell.number_format = rateFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

            if 18 < row < 22:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif row > 22 and cell.value is not None:
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))

    for col in ws7.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws7.column_dimensions[column].width = adjusted_width
# </editor-fold>

# <editor-fold desc="MM Capping Page">
if v10.get() == 1 and v11.get() == 1:
    ws8 = wb2.create_sheet(title='Capping')
    formatwkstMM(wkstname=ws8, titlerows='3', A1title='RATE CAPPING RANGE', A2title='Blank', dfname=MMCapTable, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    
    for row in range(1, ws8.max_row + 1):
        for col in range(1, ws8.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws8[char + str(row)]
            ws8.column_dimensions[char].bestFit = True # Using bestfit as the default option for column widths
            if row > 2 and cell.value is not None: # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            if row < 2: # Applies bold font to rows 1-3, which are header rows
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif col == 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            elif col == 2:
                cell.number_format = percentFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
                
    for col in ws8.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws8.column_dimensions[column].width = adjusted_width
# </editor-fold>

# <editor-fold desc="MM Territory Sheets">
if v10.get() == 1 and v7.get() == 1:
    ws11 = wb2.create_sheet(title='Territory')
    formatwkstMM(wkstname=ws11, titlerows='3', A1title='TERRITORY ADJUSTMENT FACTORS',
                 A2title='Blank', dfname=GridFacXL, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    print("worksheet set up")

    # Define reusable styles
    thin_border = Border(left=Side(border_style='thin', color='00000000'),
                         right=Side(border_style='thin', color='00000000'),
                         top=Side(border_style='thin', color='00000000'),
                         bottom=Side(border_style='thin', color='00000000'))

    side_border = Border(left=Side(border_style='thin', color='00000000'),
                         right=Side(border_style='thin', color='00000000'))

    header_font = fontBoldUnderline
    subheader_font = Font(name=fontName, size=fontSize, color='FFFFFFFF', bold=True)

    # Apply styles and borders
    for row in ws11.iter_rows(min_row=1, max_row=ws11.max_row, max_col=ws11.max_column):
        for cell in row:
            r, c = cell.row, cell.column
            # Font and alignment
            if r == 1:
                cell.font = header_font
            elif r == 2:
                cell.font = subheader_font
            else:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            
            # Borders
            if r > 2:
                if cell.value and str(cell.value).strip():
                    cell.border = thin_border
                else:
                    cell.border = side_border

    # Bold font for row 3
    for cell in ws11["3:3"]:
        cell.font = fontBold

    # Apply number format for data rows
    for col in ws11.iter_cols(min_row=3):
        for cell in col:
            cell.number_format = "0.000"

    # Adjust column widths
    for col in ws11.iter_cols(min_row=3):
        max_length = max((len(str(cell.value)) for cell in col if cell.value), default=0)
        ws11.column_dimensions[col[0].column_letter].width = (max_length + 2) * 1.2

    # Set print area and page setup
    ws11.print_area = f"A1:{get_column_letter(ws11.max_column)}{ws11.max_row}"
    ws11.page_setup.fitToPage = True

    # Add page breaks every 120 rows
    tablefac2 = math.ceil(ws11.max_row / 120)
    for rnum in range(1, tablefac2):
        ws11.row_breaks.append(Break(id=rnum * 120))

    ws11.page_setup.fitToWidth = 1
    ws11.page_setup.fitToHeight = tablefac2
    print("formatting done")
# </editor-fold>

# <editor-fold desc="MM Segmentation Removal Sheets">
if v10.get() == 1 and v12.get() == 1:
    ws12 = wb2.create_sheet(title='Tiering')
    ws12['A1'] = 'TIERING GRADE FACTORS'
    ws12['A2'] = 'Basic Group I Tiering Factors'
    ws12['G1'] = 'TIERING GRADE FACTORS'
    ws12['G2'] = 'Basic Group II Tiering Factors'
    ws12['M1'] = 'TIERING GRADE FACTORS'
    ws12['M2'] = 'Special Cause of Loss Tiering Factors'
    for r in dataframe_to_rows(SegTier, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & (len(list(r)) == 1): # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            ws12['B4'] = list(r)[0] # Manually adding the 1 index value to the header row in Excel
            continue
        ws12.append(r)

    ws12.page_setup.orientation = 'portrait' # Landscape orientation for printing
    ws12.page_setup.blackAndWhite = False
    ws12.page_setup.firstPageNumber = 1 # Resetting the page counter for the footer on each worksheet
    ws12.page_setup.useFirstPageNumber = True
    ws12.sheet_view.showGridLines = False # Turning off gridlines
    ws12.print_title_rows = '1:3'
    ws12.page_margins.left = leftMargin
    ws12.page_margins.right = rightMargin
    ws12.page_margins.top = topMargin
    ws12.page_margins.bottom = bottomMargin
    ws12.page_margins.header = headerMargin
    ws12.page_margins.footer = footerMargin
    ws12.print_options.horizontalCentered = True

    # Left Header
    if State == 'Florida':
        ws12.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        ws12.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    ws12.oddHeader.left.size = headerFontSize
    ws12.oddHeader.left.font = headerFont
    
    # Center header
    ws12.oddHeader.center.text = "\n\n" + State + " Rate Pages"
    ws12.oddHeader.center.size = headerFontSize
    ws12.oddHeader.center.font = headerFont
    
    # Right header
    ws12.oddHeader.right.text = "Effective Date: " + EffectiveDate
    ws12.oddHeader.right.size = headerFontSize
    ws12.oddHeader.right.font = headerFont

    if StateAbb != "WI":
        ws12.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company \nNationwide Property & Casualty Insurance Company \nAllied Insurance Company of America"
    else:
        ws12.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company"

    # Left footer
    ws12.oddFooter.left.size = footerFontSize
    ws12.oddFooter.left.font = footerFont
    
    # Center footer
    ws12.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws12.oddFooter.center.size = footerFontSize
    ws12.oddFooter.center.font = footerFont

    for row in range(1, ws12.max_row + 1):
        for col in range(1, ws12.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws12[char + str(row)]
            ws12.column_dimensions[char].bestFit = True # Using bestfit as the default option for column widths
            if row > 2 and cell.value != '': # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            if row < 2: # Applies bold font to row 1, which is a header row
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF', bold=True)
            elif col >= 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

    ws12.merge_cells('A2:B2')
    currCell = ws12['A2']
    currCell.alignment = Alignment(horizontal='left')
    
    ws12.merge_cells('G2:H2')
    currCell = ws12['G2']
    currCell.alignment = Alignment(horizontal='left')
    
    ws12.merge_cells('M2:O2')
    currCell = ws12['M2']
    currCell.alignment = Alignment(horizontal='left')

    for col in ws12.iter_cols(min_row=3):
        for cell in col:
            cell.number_format = "0.000"

    for row in range(3, ws12.max_row + 1):
        for col in range(1, ws12.max_column + 1):
            if row == 3:
                char = get_column_letter(col)
                cell = ws12[char + str(row)]
                cell.font = fontBold
            else:
                pass
        
        cell = ws12["A" + str(row)]
        cell.number_format = "0"
        cell = ws12["G" + str(row)]
        cell.number_format = "0"
        cell = ws12["M" + str(row)]
        cell.number_format = "0"

    for col in ws12.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws12.column_dimensions[column].width = adjusted_width

    ws12.print_area = 'A1:' + get_column_letter(ws12.max_column+1) + str(ws12.max_row+1)
    ws12.page_setup.fitToPage = True
    ws12.page_setup.fitToWidth = 3
    ws12.page_setup.fitToHeight = False

    ws13 = wb2.create_sheet(title='LEAF')
    ws13['A1'] = 'LIFETIME EXPENSE ALLOCATION FACTOR'
    ws13['A2'] = 'Blank'
    for r in dataframe_to_rows(SegLeaf, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & (len(list(r)) == 1): # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            ws13['B4'] = list(r)[0] # Manually adding the 1 index value to the header row in Excel
            continue
        ws13.append(r)

    ws13.page_setup.orientation = 'portrait' # Landscape orientation for printing
    ws13.page_setup.blackAndWhite = False
    ws13.page_setup.firstPageNumber = 1 # Resetting the page counter for the footer on each worksheet
    ws13.page_setup.useFirstPageNumber = True
    ws13.sheet_view.showGridLines = False # Turning off gridlines
    ws13.print_title_rows = '1:3'
    ws13.page_margins.left = leftMargin
    ws13.page_margins.right = rightMargin
    ws13.page_margins.top = topMargin
    ws13.page_margins.bottom = bottomMargin
    ws13.page_margins.header = headerMargin
    ws13.page_margins.footer = footerMargin
    ws13.print_options.horizontalCentered = True

    # Left Header
    if State == 'Florida':
        ws13.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        ws13.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    ws13.oddHeader.left.size = headerFontSize
    ws13.oddHeader.left.font = headerFont
    
    # Center header
    ws13.oddHeader.center.text = "\n\n" + State + " Rate Pages"
    ws13.oddHeader.center.size = headerFontSize
    ws13.oddHeader.center.font = headerFont
    
    # Right header
    ws13.oddHeader.right.text = "Effective Date: " + EffectiveDate
    ws13.oddHeader.right.size = headerFontSize
    ws13.oddHeader.right.font = headerFont

    if StateAbb != "WI":
        ws13.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company \nNationwide Property & Casualty Insurance Company \nAllied Insurance Company of America"
    else:
        ws13.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company"

    # Left footer
    ws13.oddFooter.left.size = footerFontSize
    ws13.oddFooter.left.font = footerFont
    
    # Center footer
    ws13.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws13.oddFooter.center.size = footerFontSize
    ws13.oddFooter.center.font = footerFont

    for row in range(1, ws13.max_row + 1):
        for col in range(1, ws13.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws13[char + str(row)]
            ws13.column_dimensions[char].bestFit = True # Using bestfit as the default option for column widths
            if row > 2 and cell.value != '': # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            if row < 2: # Applies bold font to row 1, which is a header row
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF', bold=True)
            elif col >= 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

    for col in ws13.iter_cols(min_row=3):
        for cell in col:
            cell.number_format = "0.000"

    for cell in ws13["3:3"]:
        cell.font = fontBold

    for row in range(1, ws13.max_row + 1):
        cell = ws13["A" + str(row)]
        cell.number_format = "0"

    for col in ws13.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws13.column_dimensions[column].width = adjusted_width

    ws13.print_area = 'A1:' + get_column_letter(ws13.max_column+1) + str(ws13.max_row+1)
    ws13.page_setup.fitToPage = True

    ws14 = wb2.create_sheet(title='Age of Building')
    ws14['A1'] = 'AGE OF BUILDING FACTORS'
    ws14['A2'] = 'Basic Group I Factors'
    ws14['M1'] = 'AGE OF BUILDING FACTORS'
    ws14['M2'] = 'Basic Group II Factors'
    ws14['Y1'] = 'AGE OF BUILDING FACTORS'
    ws14['Y2'] = 'Special Cause of Loss Factors'

    for r in dataframe_to_rows(SegAOB, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & (len(list(r)) == 1): # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            ws14['B4'] = list(r)[0] # Manually adding the 1 index value to the header row in Excel
            continue
        ws14.append(r)

    ws14.page_setup.orientation = 'portrait' # Landscape orientation for printing
    ws14.page_setup.blackAndWhite = False
    ws14.page_setup.firstPageNumber = 1 # Resetting the page counter for the footer on each worksheet
    ws14.page_setup.useFirstPageNumber = True
    ws14.sheet_view.showGridLines = False # Turning off gridlines
    ws14.print_title_rows = '1:3'
    ws14.page_margins.left = leftMargin
    ws14.page_margins.right = rightMargin
    ws14.page_margins.top = topMargin
    ws14.page_margins.bottom = bottomMargin
    ws14.page_margins.header = headerMargin
    ws14.page_margins.footer = footerMargin
    ws14.print_options.horizontalCentered = True

    # Left Header
    if State == 'Florida':
        ws14.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        ws14.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    ws14.oddHeader.left.size = headerFontSize
    ws14.oddHeader.left.font = headerFont
    
    # Center header
    ws14.oddHeader.center.text = "\n\n" + State + " Rate Pages"
    ws14.oddHeader.center.size = headerFontSize
    ws14.oddHeader.center.font = headerFont
    
    # Right header
    ws14.oddHeader.right.text = "Effective Date: " + EffectiveDate
    ws14.oddHeader.right.size = headerFontSize
    ws14.oddHeader.right.font = headerFont

    if StateAbb != "WI":
        ws14.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company \nNationwide Property & Casualty Insurance Company \nAllied Insurance Company of America"
    else:
        ws14.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company"

    # Left footer
    ws14.oddFooter.left.size = footerFontSize
    ws14.oddFooter.left.font = footerFont
    
    # Center footer
    ws14.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws14.oddFooter.center.size = footerFontSize
    ws14.oddFooter.center.font = footerFont

    for row in range(1, ws14.max_row + 1):
        for col in range(1, ws14.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws14[char + str(row)]
            ws14.column_dimensions[char].bestFit = True # Using bestfit as the default option for column widths
            if row > 2: # Adding a border to the table data
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            if row < 2: # Applies bold font to row 1, which is a header row
                cell.font = fontBoldUnderline
            elif col >= 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

    ws14.merge_cells('A2:C2')
    currCell = ws14['A2']
    currCell.alignment = Alignment(horizontal='left')
    
    ws14.merge_cells('M2:O2')
    currCell = ws14['M2']
    currCell.alignment = Alignment(horizontal='left')
    
    ws14.merge_cells('Y2:AB2')
    currCell = ws14['Y2']
    currCell.alignment = Alignment(horizontal='left')

    for col in ws14.iter_cols(min_row=3):
        for cell in col:
            cell.number_format = "0.000"

    for row in range(3, ws14.max_row + 1):
        for col in range(1, ws14.max_column + 1):
            if row == 3:
                char = get_column_letter(col)
                cell = ws14[char + str(row)]
                cell.font = fontBold
            else:
                pass
            
        cell = ws14["A" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["D" + str(row)]
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thick', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
                             
        cell = ws14["E" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["H" + str(row)]
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thick', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))

        cell = ws14["I" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["M" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["P" + str(row)]
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thick', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
                             
        cell = ws14["Q" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["T" + str(row)]
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thick', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
                             
        cell = ws14["U" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["Y" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["AB" + str(row)]
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thick', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
                             
        cell = ws14["AC" + str(row)]
        cell.number_format = "0"
        
        cell = ws14["AF" + str(row)]
        cell.border = Border(left=Side(border_style='thin', color='00000000'),
                             right=Side(border_style='thick', color='00000000'),
                             top=Side(border_style='thin', color='00000000'),
                             bottom=Side(border_style='thin', color='00000000'))
                             
        cell = ws14["AG" + str(row)]
        cell.number_format = "0"

    for col in ws14.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws14.column_dimensions[column].width = adjusted_width
        
    ws14.print_area = 'A1:' + get_column_letter(ws14.max_column+1) + str(ws14.max_row+1)
    ws14.page_setup.fitToPage = True
    ws14.page_setup.fitToWidth = 3
    ws14.page_setup.fitToHeight = False

    ws15 = wb2.create_sheet(title='AOI Bldg Grp I')
    formatwkstMM(wkstname=ws15, titlerows='3', A1title='AOI CURVES', A2title='Limit of Insurance Relativity Factors Building Basic Group I', dfname=BGIBLOIMM, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ws15)
    for row in range(4, ws15.max_row+1):
        for col in range(1, ws15.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws15[char + str(row)]
            cell.number_format = rateFormat

    ws16 = wb2.create_sheet(title='AOI Bldg Grp II')
    formatwkstMM(wkstname=ws16, titlerows='3', A1title='AOI CURVES', A2title='Basic Group II Limit Factors Building', dfname=BGIIBLOIMM, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ws16)
    for row in range(4, ws16.max_row+1):
        for col in range(1, ws16.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws16[char + str(row)]
            cell.number_format = rateFormat

    ws17 = wb2.create_sheet(title='AOI Bldg SCOL')
    formatwkstMM(wkstname=ws17, titlerows='3', A1title='AOI CURVES', A2title='Special Cause of Loss Limit Factors Building', dfname=SCOLBLOIMM, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ws17)
    for row in range(4, ws17.max_row+1):
        for col in range(1, ws17.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws17[char + str(row)]
            cell.number_format = rateFormat

    ws18 = wb2.create_sheet(title='AOI PP Grp I')
    formatwkstMM(wkstname=ws18, titlerows='3', A1title='AOI CURVES', A2title='Basic Group I Limit Factors Contents', dfname=BGIPPLOIMM, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ws18)
    for row in range(4, ws18.max_row+1):
        for col in range(1, ws18.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws18[char + str(row)]
            cell.number_format = rateFormat

    ws19 = wb2.create_sheet(title='AOI PP Grp II')
    formatwkstMM(wkstname=ws19, titlerows='3', A1title='AOI CURVES', A2title='Basic Group II Limit Factors Contents', dfname=BGIIPPLOIMM, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ws19)
    for row in range(4, ws19.max_row+1):
        for col in range(1, ws19.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws19[char + str(row)]
            cell.number_format = rateFormat

    ws20 = wb2.create_sheet(title='AOI PP SCOL')
    formatwkstMM(wkstname=ws20, titlerows='3', A1title='AOI CURVES', A2title='Special Cause of Loss Limit Factors - Contents', dfname=SCOLPPLOIMM, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ws20)
    for row in range(4, ws20.max_row+1):
        for col in range(1, ws20.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = ws20[char + str(row)]
            cell.number_format = rateFormat

# <editor-fold desc="SM Deductible tables">
if v5.get() == 1 and v14.get() == 1:
    DedbyAOIBG1SM = wb.create_sheet(title='Deductible by AOI Grp I')
    formatwkstSM(wkstname=DedbyAOIBG1SM, titlerows='3', A1title='RULE 81. DEDUCTIBLE INSURANCE PLAN', A2title='Basic Group I Factors', dfname=DIPI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(DedbyAOIBG1SM)
    for row in range(4, DedbyAOIBG1SM.max_row + 1):
        for col in range(1, DedbyAOIBG1SM.max_column + 1):
            cell = DedbyAOIBG1SM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = DedbyAOIBG1SM["A" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBG1SM["D" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBG1SM["G" + str(row)]
        cell.number_format = "0"
        
    DedbyAOIBG1SM.column_dimensions['I'].width = 9.5
    DedbyAOIBG1SM.insert_cols(4, 1)
    DedbyAOIBG1SM.insert_cols(8, 1)
    
    for col in DedbyAOIBG1SM.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                max_length = 1
                pass
        adjusted_width = (max_length + 2) * 1.2
        DedbyAOIBG1SM.column_dimensions[column].width = adjusted_width
        
    DedbyAOIBG1SM.column_dimensions['G'].width = 9.5
    DedbyAOIBG1SM.column_dimensions['K'].width = 9.5
    DedbyAOIBG1SM.print_area = 'A1:' + get_column_letter(DedbyAOIBG1SM.max_column) + str(DedbyAOIBG1SM.max_row + 1)
    DedbyAOIBG1SM.page_setup.fitToPage = True
    DedbyAOIBG1SM.page_setup.fitToHeight = 1
    DedbyAOIBG1SM.page_setup.fitToWidth = 1

    DedbyAOIBG2SM = wb.create_sheet(title='Deductible by AOI Grp II')
    formatwkstSM(wkstname=DedbyAOIBG2SM, titlerows='3', A1title='RULE 81. DEDUCTIBLE INSURANCE PLAN', A2title='Basic Group II Factors', dfname=DIPII, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(DedbyAOIBG2SM)
    
    for row in range(4, DedbyAOIBG2SM.max_row + 1):
        for col in range(1, DedbyAOIBG2SM.max_column + 1):
            cell = DedbyAOIBG2SM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = DedbyAOIBG2SM["A" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBG2SM["D" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBG2SM["G" + str(row)]
        cell.number_format = "0"
        
    DedbyAOIBG2SM.insert_cols(4, 1)
    DedbyAOIBG2SM.insert_cols(8, 1)
    
    for col in DedbyAOIBG2SM.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                max_length = 1
                pass
        adjusted_width = (max_length + 2) * 1.2
        DedbyAOIBG2SM.column_dimensions[column].width = adjusted_width
        
    DedbyAOIBG2SM.column_dimensions['G'].width = 9.5
    DedbyAOIBG2SM.column_dimensions['K'].width = 9.5
    DedbyAOIBG2SM.print_area = 'A1:' + get_column_letter(DedbyAOIBG2SM.max_column) + str(DedbyAOIBG2SM.max_row + 1)
    DedbyAOIBG2SM.page_setup.fitToPage = True
    DedbyAOIBG2SM.page_setup.fitToHeight = 1
    DedbyAOIBG2SM.page_setup.fitToWidth = 1

    DedbyAOISCOLSM = wb.create_sheet(title='Deductible by AOI SCOL')
    formatwkstSM(wkstname=DedbyAOISCOLSM, titlerows='3', A1title='RULE 81. DEDUCTIBLE INSURANCE PLAN', A2title='Special Cause of Loss Factors', dfname=DIPSCOL, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(DedbyAOISCOLSM)
    
    for row in range(4, DedbyAOISCOLSM.max_row + 1):
        for col in range(1, DedbyAOISCOLSM.max_column + 1):
            cell = DedbyAOISCOLSM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = DedbyAOISCOLSM["A" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOISCOLSM["D" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOISCOLSM["G" + str(row)]
        cell.number_format = "0"
        
    DedbyAOISCOLSM.insert_cols(4, 1)
    DedbyAOISCOLSM.insert_cols(8, 1)
    
    for col in DedbyAOISCOLSM.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                max_length = 1
                pass
        adjusted_width = (max_length + 2) * 1.2
        DedbyAOISCOLSM.column_dimensions[column].width = adjusted_width
        
    DedbyAOISCOLSM.column_dimensions['G'].width = 9.5
    DedbyAOISCOLSM.column_dimensions['K'].width = 9.5
    DedbyAOISCOLSM.print_area = 'A1:' + get_column_letter(DedbyAOISCOLSM.max_column) + str(DedbyAOISCOLSM.max_row + 1)
    DedbyAOISCOLSM.page_setup.fitToPage = True
    DedbyAOISCOLSM.page_setup.fitToHeight = 1
    DedbyAOISCOLSM.page_setup.fitToWidth = 1

# <editor-fold desc="SM Full Manual">
if v5.get() == 1 and v6.get() == 1:
    ExpSM = wb.create_sheet(title='Expense Constant')
    formatwkstSM(wkstname=ExpSM, titlerows='3', A1title='RULE 8 EXPENSE CONSTANT', A2title='8. Rate and Premium Computation', dfname=ExpConstDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ExpSM)

    MinPSM = wb.create_sheet(title='Minimum Premium')
    formatwkstSM(wkstname=MinPSM, titlerows='3', A1title='RULE 8 EXPENSE CONSTANT', A2title='8. Rate and Premium Computation', dfname=PolicyMinDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(MinPSM)

    AOIBG1SM = wb.create_sheet(title='AOI Bldg Grp I')
    formatwkstSM(wkstname=AOIBG1SM, titlerows='3', A1title='AOI CURVES', A2title='Limit of Insurance Relativity Factors Building Basic Group I', dfname=BGIBLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIBG1SM)
    for row in range(4, AOIBG1SM.max_row+1):
        for col in range(1, AOIBG1SM.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIBG1SM[char + str(row)]
            cell.number_format = rateFormat

    AOIBG2SM = wb.create_sheet(title='AOI Bldg Grp II')
    formatwkstSM(wkstname=AOIBG2SM, titlerows='3', A1title='AOI CURVES', A2title='Basic Group II Limit Factors Building', dfname=BGIIBLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIBG2SM)
    for row in range(4, AOIBG2SM.max_row+1):
        for col in range(1, AOIBG2SM.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIBG2SM[char + str(row)]
            cell.number_format = rateFormat

    AOIBSCOLSM = wb.create_sheet(title='AOI Bldg SCOL')
    formatwkstSM(wkstname=AOIBSCOLSM, titlerows='3', A1title='AOI CURVES', A2title='Special Cause of Loss Limit Factors Building', dfname=SCOLBLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIBSCOLSM)
    for row in range(4, AOIBSCOLSM.max_row+1):
        for col in range(1, AOIBSCOLSM.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIBSCOLSM[char + str(row)]
            cell.number_format = rateFormat

    AOIPPG1SM = wb.create_sheet(title='AOI PP Grp I')
    formatwkstSM(wkstname=AOIPPG1SM, titlerows='3', A1title='AOI CURVES', A2title='Basic Group I Limit Factors Contents', dfname=BGIPPLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIPPG1SM)
    for row in range(4, AOIPPG1SM.max_row+1):
        for col in range(1, AOIPPG1SM.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIPPG1SM[char + str(row)]
            cell.number_format = rateFormat

    AOIPPG2SM = wb.create_sheet(title='AOI PP Grp II')
    formatwkstSM(wkstname=AOIPPG2SM, titlerows='3', A1title='AOI CURVES', A2title='Basic Group II Limit Factors Contents', dfname=BGIIPPLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIPPG2SM)
    for row in range(4, AOIPPG2SM.max_row+1):
        for col in range(1, AOIPPG2SM.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIPPG2SM[char + str(row)]
            cell.number_format = rateFormat

    AOIPPSCOLSM = wb.create_sheet(title='AOI PP SCOL')
    formatwkstSM(wkstname=AOIPPSCOLSM, titlerows='3', A1title='AOI CURVES', A2title='Special Cause of Loss Limit Factors - Contents', dfname=SCOLPPLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIPPSCOLSM)
    for row in range(4, AOIPPSCOLSM.max_row+1):
        for col in range(1, AOIPPSCOLSM.max_column+1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIPPSCOLSM[char + str(row)]
            cell.number_format = rateFormat

    TieringG1SM = wb.create_sheet(title='Tiering Grp I')
    formatwkstSM(wkstname=TieringG1SM, titlerows='3', A1title='TIERING GRADE FACTORS', A2title='Basic Group I Tiering Factors', dfname=BGITier, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TieringG1SM)
    for row in range(4, TieringG1SM.max_row + 1):
        for col in range(1, TieringG1SM.max_column + 1):
            cell = TieringG1SM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    TieringG2SM = wb.create_sheet(title='Tiering Grp II')
    formatwkstSM(wkstname=TieringG2SM, titlerows='3', A1title='TIERING GRADE FACTORS', A2title='Basic Group II Tiering Factors', dfname=BGIITier, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TieringG2SM)
    for row in range(4, TieringG2SM.max_row + 1):
        for col in range(1, TieringG2SM.max_column + 1):
            cell = TieringG2SM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    TieringSCOLSM = wb.create_sheet(title='Tiering SCOL')
    formatwkstSM(wkstname=TieringSCOLSM, titlerows='3', A1title='TIERING GRADE FACTORS', A2title='Special Cause of Loss Tiering Factors', dfname=SCOLTier, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TieringSCOLSM)
    for row in range(4, TieringSCOLSM.max_row + 1):
        for col in range(1, TieringSCOLSM.max_column + 1):
            cell = TieringSCOLSM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    EmergencySM = wb.create_sheet(title='Emergency Evacuation')
    formatwkstSM(wkstname=EmergencySM, titlerows='3', A1title='EMERGENCY EVACUATION INCLUDING CIVIL AUTHORITY COVERAGE', A2title='', dfname=EmergencyDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(EmergencySM)
    addlformat(EmergencySM)

    HumanSvEnhSM = wb.create_sheet(title='Human Serv. Enhancement Cover.')
    formatwkstSM(wkstname=HumanSvEnhSM, titlerows='3', A1title='HUMAN SERVICES ENHANCEMENT COVERAGE', A2title='', dfname=HumSvcsEnhDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(HumanSvEnhSM)
    addlformat(HumanSvEnhSM)

    ClientsSM = wb.create_sheet(title='Clients Property Coverage')
    formatwkstSM(wkstname=ClientsSM, titlerows='3', A1title='CLIENTS PROPERTY COVERAGE FACTOR', A2title='', dfname=ClientsDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ClientsSM)
    addlformat(ClientsSM)

    HumanSvcWorkSM = wb.create_sheet(title='Human Serv. Workplace Violence')
    formatwkstSM(wkstname=HumanSvcWorkSM, titlerows='3', A1title='HUMAN SERVICES WORKPLACE VIOLENCE LOSS OF INCOME', A2title='', dfname=HumSvcsDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(HumanSvcWorkSM)
    addlformat(HumanSvcWorkSM)

    HumanSvcPrpSM = wb.create_sheet(title='Human Serv. Prop Endorsement')
    formatwkstSM(wkstname=HumanSvcPrpSM, titlerows='3', A1title='HUMAN SERVICES PROPERTY ENDORSEMENT FACTOR', A2title='', dfname=HumSvcsFctDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(HumanSvcPrpSM)
    addlformat(HumanSvcPrpSM)

    SeniorPropSM = wb.create_sheet(title='Senior Living Property')
    formatwkstSM(wkstname=SeniorPropSM, titlerows='3', A1title='SENIOR LIVING COMMUNITIES PROPERTY ENDORSEMENT', A2title='', dfname=SeniorDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(SeniorPropSM)
    addlformat(SeniorPropSM)

    WaterDmgSM = wb.create_sheet(title='Water Damage Deductible Weight')
    formatwkstSM(wkstname=WaterDmgSM, titlerows='3', A1title='WATER DAMAGE DEDUCTIBLE WEIGHT', A2title='', dfname=WtrDmgDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(WaterDmgSM)
    addlformat(WaterDmgSM)

    AOBGrp1SM = wb.create_sheet(title='Age of Building Grp I')
    formatwkstSM(wkstname=AOBGrp1SM, titlerows='3', A1title='AGE OF BUILDING FACTORS', A2title='Basic Group I Factors', dfname=BGIAOB, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOBGrp1SM)
    
    for row in range(4, AOBGrp1SM.max_row + 1):
        for col in range(1, AOBGrp1SM.max_column + 1):
            cell = AOBGrp1SM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = AOBGrp1SM["A" + str(row)]
        cell.number_format = "0"
        cell = AOBGrp1SM["E" + str(row)]
        cell.number_format = "0"
        cell = AOBGrp1SM["I" + str(row)]
        cell.number_format = "0"

    AOBGrp1SM.column_dimensions['B'].width = 12
    AOBGrp1SM.column_dimensions['C'].width = 12
    AOBGrp1SM.column_dimensions['D'].width = 12
    AOBGrp1SM.insert_cols(5, 1)
    AOBGrp1SM.insert_cols(10, 1)
    AOBGrp1SM.column_dimensions['H'].width = 12
    AOBGrp1SM.column_dimensions['J'].width = 5.8
    AOBGrp1SM.column_dimensions['I'].width = 12
    AOBGrp1SM.column_dimensions['M'].width = 12
    AOBGrp1SM.column_dimensions['N'].width = 12
    AOBGrp1SM.print_area = 'A1:' + get_column_letter(AOBGrp1SM.max_column + 1) + str(AOBGrp1SM.max_row + 1)

    AOBGrp2SM = wb.create_sheet(title='Age of Building Grp II')
    formatwkstSM(wkstname=AOBGrp2SM, titlerows='3', A1title='AGE OF BUILDING FACTORS', A2title='Basic Group II Factors', dfname=BGIIAOB, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOBGrp2SM)
    
    for row in range(4, AOBGrp2SM.max_row + 1):
        for col in range(1, AOBGrp2SM.max_column + 1):
            cell = AOBGrp2SM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = AOBGrp2SM["A" + str(row)]
        cell.number_format = "0"
        cell = AOBGrp2SM["E" + str(row)]
        cell.number_format = "0"
        cell = AOBGrp2SM["I" + str(row)]
        cell.number_format = "0"

    AOBGrp2SM.column_dimensions['B'].width = 12
    AOBGrp2SM.column_dimensions['C'].width = 12
    AOBGrp2SM.column_dimensions['D'].width = 12
    AOBGrp2SM.insert_cols(5, 1)
    AOBGrp2SM.insert_cols(10, 1)
    AOBGrp2SM.column_dimensions['H'].width = 12
    AOBGrp2SM.column_dimensions['J'].width = 5.8
    AOBGrp2SM.column_dimensions['I'].width = 12
    AOBGrp2SM.column_dimensions['M'].width = 12
    AOBGrp2SM.column_dimensions['N'].width = 12
    AOBGrp2SM.print_area = 'A1:' + get_column_letter(AOBGrp2SM.max_column + 1) + str(AOBGrp2SM.max_row + 1)

    AOBSCOLSM = wb.create_sheet(title='Age of Building SCOL')
    formatwkstSM(wkstname=AOBSCOLSM, titlerows='3', A1title='AGE OF BUILDING FACTORS', A2title='Special Cause of Loss Factors', dfname=SCOLAOB, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOBSCOLSM)
    
    for row in range(4, AOBSCOLSM.max_row + 1):
        for col in range(1, AOBSCOLSM.max_column + 1):
            cell = AOBSCOLSM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = AOBSCOLSM["A" + str(row)]
        cell.number_format = "0"
        cell = AOBSCOLSM["E" + str(row)]
        cell.number_format = "0"
        cell = AOBSCOLSM["I" + str(row)]
        cell.number_format = "0"

    AOBSCOLSM.column_dimensions['B'].width = 12
    AOBSCOLSM.column_dimensions['C'].width = 12
    AOBSCOLSM.column_dimensions['D'].width = 12
    AOBSCOLSM.insert_cols(5, 1)
    AOBSCOLSM.insert_cols(10, 1)
    AOBSCOLSM.column_dimensions['H'].width = 12
    AOBSCOLSM.column_dimensions['J'].width = 5.8
    AOBSCOLSM.column_dimensions['I'].width = 12
    AOBSCOLSM.column_dimensions['M'].width = 12
    AOBSCOLSM.column_dimensions['N'].width = 12
    AOBSCOLSM.print_area = 'A1:' + get_column_letter(AOBSCOLSM.max_column + 1) + str(AOBSCOLSM.max_row + 1)

    CivilTimeSM = wb.create_sheet(title='Civil Authority Time Element')
    formatwkstSM(wkstname=CivilTimeSM, titlerows='3', A1title='RULE 54. E. CIVIL AUTHORITY CHANGES', A2title='Table 54.E.3.b: Civil Authority Increased Radius Coverage Factor', dfname=CivilDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(CivilTimeSM)
    addlformat(CivilTimeSM)
    CivilTimeSM['E3'].alignment = Alignment(horizontal='center')
    CivilTimeSM.column_dimensions['E'].width = 14

    CosmeticSM = wb.create_sheet(title='Cosmetic Exclusion')
    formatwkstSM(wkstname=CosmeticSM, titlerows='3', A1title='COSMETIC EXCLUSION ON COVERAGE FOR SIDING', A2title='Multiply the Basic Group II building rate by the below factor', dfname=CosmeticDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(CosmeticSM)
    addlformat(CosmeticSM)
    CosmeticSM['D4'].alignment = Alignment(horizontal='center')

    TerrorismSM = wb.create_sheet(title='Terrorism')
    formatwkstMM(wkstname=TerrorismSM, titlerows='3', A1title='Coverage for Certified Acts of Terrorism', A2title='', dfname=TerrorismDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TerrorismSM)
    addlformat(TerrorismSM)

    Money304SM = wb.create_sheet(title='Money & Securities 304')
    formatwkstSM(wkstname=Money304SM, titlerows='3', A1title='304 Money and Securities', A2title='Base Premium', dfname=CrimeMSDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(Money304SM)

    Money304RFSM = wb.create_sheet(title='Money & Securities 304 RF')
    formatwkstSM(wkstname=Money304RFSM, titlerows='3', A1title='Rating Factor', A2title='Base Premium', dfname=MSDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(Money304RFSM)
    for row in range(4, Money304RFSM.max_row + 1):
        for col in range(1, Money304RFSM.max_column + 1):
            cell = Money304RFSM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = Money304RFSM["A" + str(row)]
        cell.number_format = "#,##0"
        cell = Money304RFSM["B" + str(row)]
        cell.number_format = "#,##0"

    LPDPSM = wb.create_sheet(title='LPDP')
    formatwkstSM(wkstname=LPDPSM, titlerows='3', A1title='Large Premium Discount Factor', A2title='', dfname=LPDFDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(LPDPSM)

    Employee308SM = wb.create_sheet(title='Employee Dishonesty 308')
    formatwkstSM(wkstname=Employee308SM, titlerows='3', A1title='308 Employee Dishonesty', A2title='Base Premium', dfname=EmployeeDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(Employee308SM)

    CompFraudSM = wb.create_sheet(title='Computer Fraud 311 BP')
    formatwkstSM(wkstname=CompFraudSM, titlerows='3', A1title='311 Computer and Funds Transfer Fraud', A2title='Base Premium', dfname=ComputerDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(CompFraudSM)
    for row in range(4, CompFraudSM.max_row + 1):
        for col in range(1, CompFraudSM.max_column + 1):
            cell = CompFraudSM[get_column_letter(col) + str(row)]
            cell.number_format = "#,##0"

    CompFraudGSFSM = wb.create_sheet(title='Computer Fraud 311 GSF')
    formatwkstSM(wkstname=CompFraudGSFSM, titlerows='3', A1title='Annual Gross Sales Factor', A2title='', dfname=ComputerAnnualDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(CompFraudGSFSM)

    CounterfeitSM = wb.create_sheet(title='Counterfeit Currency 313')
    formatwkstSM(wkstname=CounterfeitSM, titlerows='3', A1title='313 Counterfeit Currency', A2title='', dfname=CFCurrDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(CounterfeitSM)
    for row in range(4, CounterfeitSM.max_row + 1):
        for col in range(1, CounterfeitSM.max_column + 1):
            cell = CounterfeitSM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = CounterfeitSM["A" + str(row)]
        cell.number_format = "#,##0"

    ForgerySM = wb.create_sheet(title='Forgery or Alteration 314')
    formatwkstSM(wkstname=ForgerySM, titlerows='3', A1title='314 Forgery or Alteration', A2title='Base Premium', dfname=ForgeryDF2, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ForgerySM)

    Fraud516SM = wb.create_sheet(title='Fraudulent Impersonation 516')
    formatwkstSM(wkstname=Fraud516SM, titlerows='3', A1title='516 Fraudulent Impersonation', A2title='Base Premium', dfname=FraudDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(Fraud516SM)

    ##
    WineLeakSM = wb.create_sheet(title='Wine Leakage Endorsement')
    formatwkstSM(wkstname=WineLeakSM, titlerows='3', A1title='WINE LEAKAGE ENDORSEMENT', A2title='', dfname=WineLkgDF1, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    for r in dataframe_to_rows(WineLkgDF2, False, True):
        # The header is the first row and the index is the second row, but they need to be on the same row in Excel
        if False & (len(list(r)) == 1): # Checking to see when the index row is reached since it will contain only the number of the indices that are in the dataframe (currently only works for 1 index)
            WineLeakSM['B5'] = list(r)[0] # Manually adding the 1 index value to the header row in Excel
            continue
        WineLeakSM.append(r)
    
    borderfnct(WineLeakSM)
    WineLeakSM.delete_rows(3)
    WineLeakSM.insert_rows(4)
    WineLeakSM.delete_rows(2)
    WineLeakSM.insert_rows(2)
    addlformat(WineLeakSM)
    WineLeakSM['E1'].alignment = Alignment(horizontal='left')
    WineLeakSM.column_dimensions['E'].width = 14.8

    BIALSSM = wb.create_sheet(title='BI ALS')
    formatwkstSM(wkstname=BIALSSM, titlerows='3', A1title='RULE 51. BUSINESS INCOME COVERAGE OPTIONS', A2title='Table 51.C.3: Business Income Actual Loss Sustained Factors', dfname=BusinessDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(BIALSSM)
    for row in range(4, BIALSSM.max_row + 1):
        for col in range(1, BIALSSM.max_column + 1):
            cell = BIALSSM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    NutHullerSM = wb.create_sheet(title='Nut Hullers')
    formatwkstSM(wkstname=NutHullerSM, titlerows='3', A1title='NUT HULLERS AND PROCESSORS COVERAGE', A2title='Charge the below premium', dfname=NutHullProcDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(NutHullerSM)
    addlformat(NutHullerSM)

    ConAtmoSM = wb.create_sheet(title='Con Atmo and Storage')
    formatwkstSM(wkstname=ConAtmoSM, titlerows='3', A1title='CONTROLLED ATMOSPHERE AND STORAGE COVERAGE', A2title='a. The base rate for all types of occupancies is below (per $100).', dfname=ControlAtmDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ConAtmoSM)
    addlformat(ConAtmoSM)
    ConAtmoSM['E5'] = 'b. Modify the rate by the Property Deductible factor, if applicable.'
    ConAtmoSM['E5'].alignment = Alignment(horizontal='center')
    ConAtmoSM['E5'].font = Font(name='Arial', size=10)
    ConAtmoSM['E6'] = 'c. Multiply the rate thus developed by the applicable total Limit of Insurance'
    ConAtmoSM['E6'].alignment = Alignment(horizontal='center')
    ConAtmoSM['E6'].font = Font(name='Arial', size=10)
    ConAtmoSM['E7'] = ' (per $100) to determine the Manual Premium.'
    ConAtmoSM['E7'].alignment = Alignment(horizontal='center')
    ConAtmoSM['E7'].font = Font(name='Arial', size=10)
    ConAtmoSM.print_area = 'A1:' + get_column_letter(ConAtmoSM.max_column + 3) + str(ConAtmoSM.max_row + 1)
    ConAtmoSM.page_setup.fitToPage = True

    ConAtmoStorSM = wb.create_sheet(title='Con Atmo Stor and Chem Inj')
    formatwkstSM(wkstname=ConAtmoStorSM, titlerows='3', A1title='CONTROLLED ATMOSPHERE, STORAGE AND CHEMICAL INJURY COVERAGE', A2title='a. The base rate for all types of occupancies is below (per $100).', dfname=ControlAtChInjDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ConAtmoStorSM)
    addlformat(ConAtmoStorSM)
    ConAtmoStorSM['E6'] = 'b. Modify the rate by the Property Deductible factor, if applicable.'
    ConAtmoStorSM['E6'].alignment = Alignment(horizontal='center')
    ConAtmoStorSM['E6'].font = Font(name='Arial', size=10)
    ConAtmoStorSM['E7'] = 'c. Multiply the rate thus developed by the applicable total Limit of Insurance'
    ConAtmoStorSM['E7'].alignment = Alignment(horizontal='center')
    ConAtmoStorSM['E7'].font = Font(name='Arial', size=10)
    ConAtmoStorSM['E8'] = ' (per $100) to determine the Manual Premium.'
    ConAtmoStorSM['E8'].alignment = Alignment(horizontal='center')
    ConAtmoStorSM['E8'].font = Font(name='Arial', size=10)
    ConAtmoStorSM.column_dimensions['E'].width = 11.5
    ConAtmoStorSM.print_area = 'A1:' + get_column_letter(ConAtmoStorSM.max_column + 3) + str(ConAtmoStorSM.max_row + 1)
    ConAtmoStorSM.page_setup.fitToPage = True

    HopsGrowerSM = wb.create_sheet(title='Hops Grower')
    formatwkstSM(wkstname=HopsGrowerSM, titlerows='3', A1title='HOPS GROWER', A2title='', dfname=HopsDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(HopsGrowerSM)
    addlformat(HopsGrowerSM)

    FruitTreeSM = wb.create_sheet(title='Fruit Trees and Trelisses')
    formatwkstSM(wkstname=FruitTreeSM, titlerows='3', A1title='FRUIT TREES, TRELLISES, STAKES & VINES', A2title='', dfname=FruitDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(FruitTreeSM)
    addlformat(FruitTreeSM)

    WinerySM = wb.create_sheet(title='Winery Endorsement')
    formatwkstSM(wkstname=WinerySM, titlerows='3', A1title='WINERY ENDORSEMENT', A2title='', dfname=WineEndDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(WinerySM)
    addlformat(WinerySM)

    WineryIncSM = wb.create_sheet(title='Winery Increased Limits Schedul')
    formatwkstSM(wkstname=WineryIncSM, titlerows='3', A1title='WINERY INCREASED LIMITS SCHEDULE', A2title='', dfname=WineryDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(WineryIncSM)
    for row in range(4, WineryIncSM.max_row + 1):
        for col in range(1, WineryIncSM.max_column + 1):
            cell = WineryIncSM[get_column_letter(col) + str(row)]
            cell.number_format = "#,##0"

    ##
    CPPPlusSM = wb.create_sheet(title='PLUS Endorsements CPP')
    formatwkstSM(wkstname=CPPPlusSM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='COMMERCIAL PROPERTY PROTECTION PLUS ENDORSEMENTS', dfname=PropPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(CPPPlusSM)

    BMGPlusSM = wb.create_sheet(title='PLUS Endorsements BMG')
    formatwkstSM(wkstname=BMGPlusSM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='BEVERAGE MAKERS GOLD & PLATINUM PROTECTION PLUS ENDORSEMENT', dfname=BrewPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(BMGPlusSM)

    ##
    GBGPlusSM = wb.create_sheet(title='PLUS Endorsements GCG')
    formatwkstSM(wkstname=GBGPlusSM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='GOLF COURSE GOLD & PLATINUM PROTECTION PLUS ENDORSEMENT', dfname=GolfPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(GBGPlusSM)

    HGPlusSM = wb.create_sheet(title='PLUS Endorsements HG')
formatwkstSM(wkstname=HGPlusSM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS',
             A2title='HOTEL GOLD & PLATINUM PROTECTION PLUS ENDORSEMENTS', dfname=HotelPPDF,
             statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(HGPlusSM)

FranchiseSM = wb.create_sheet(title='Franchise Upgrade Endorsement')
formatwkstSM(wkstname=FranchiseSM, titlerows='3', A1title='FRANCHISE UPGRADE ENDORSEMENT',
             A2title='', dfname=FranUpgDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(FranchiseSM)
addlformat(FranchiseSM)

NutHullsSM = wb.create_sheet(title='Nuts, Hulls, or Shells')
formatwkstSM(wkstname=NutHullsSM, titlerows='3', A1title='NUTS, HULLS OR SHELLS IN THE OPEN',
             A2title='', dfname=NutHullILF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(NutHullsSM)
for row in range(4, NutHullsSM.max_row + 1):
    for col in range(1, NutHullsSM.max_column + 1):
        cell = NutHullsSM[get_column_letter(col) + str(row)]
        cell.number_format = rateFormat
    cell = NutHullsSM["A" + str(row)]
    cell.number_format = "#,##0"
    cell = NutHullsSM["B" + str(row)]
    cell.number_format = "#,##0"

#<editor-fold desc="MM Deductible">
if v10.get() == 1 and v15.get() == 1:
    DedbyAOIBGIMM = wb2.create_sheet(title='Deductible by AOI BGI')
    formatwkstMM(wkstname=DedbyAOIBGIMM, titlerows='3', A1title='RULE 81. DEDUCTIBLE INSURANCE PLAN',
                 A2title='Basic Group I Factors', dfname=DIPI, statename=State, stabb=StateAbb,
                 effdate=EffectiveDate)
    borderfnct(DedbyAOIBGIMM)
    for row in range(4, DedbyAOIBGIMM.max_row + 1):
        for col in range(1, DedbyAOIBGIMM.max_column + 1):
            cell = DedbyAOIBGIMM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = DedbyAOIBGIMM["A" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBGIMM["D" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBGIMM["G" + str(row)]
        cell.number_format = "0"
    DedbyAOIBGIMM.insert_cols(4,1)
    DedbyAOIBGIMM.insert_cols(8,1)
    for col in DedbyAOIBGIMM.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                max_length = 1
                pass
        adjusted_width = (max_length + 2) * 1.2
        DedbyAOIBGIMM.column_dimensions[column].width = adjusted_width
    DedbyAOIBGIMM.column_dimensions['G'].width = 9.5
    DedbyAOIBGIMM.column_dimensions['K'].width = 9.5
    DedbyAOIBGIMM.print_area = 'A1:' + get_column_letter(DedbyAOIBGIMM.max_column) + str(DedbyAOIBGIMM.max_row+1)
    DedbyAOIBGIMM.page_setup.fitToPage = True
    DedbyAOIBGIMM.page_setup.fitToHeight = 1
    DedbyAOIBGIMM.page_setup.fitToWidth = 1

    DedbyAOIBG2MM = wb2.create_sheet(title='Deductible by AOI BGII')
    formatwkstMM(wkstname=DedbyAOIBG2MM, titlerows='3', A1title='RULE 81. DEDUCTIBLE INSURANCE PLAN',
                 A2title='Basic Group II Factors', dfname=DIPII, statename=State, stabb=StateAbb,
                 effdate=EffectiveDate)
    borderfnct(DedbyAOIBG2MM)
    for row in range(4, DedbyAOIBG2MM.max_row + 1):
        for col in range(1, DedbyAOIBG2MM.max_column + 1):
            cell = DedbyAOIBG2MM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = DedbyAOIBG2MM["A" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBG2MM["D" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOIBG2MM["G" + str(row)]
        cell.number_format = "0"
    DedbyAOIBG2MM.insert_cols(4,1)
    DedbyAOIBG2MM.insert_cols(8,1)
    for col in DedbyAOIBG2MM.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                max_length = 1
                pass
        adjusted_width = (max_length + 2) * 1.2
        DedbyAOIBG2MM.column_dimensions[column].width = adjusted_width
    DedbyAOIBG2MM.column_dimensions['G'].width = 9.5
    DedbyAOIBG2MM.column_dimensions['K'].width = 9.5
    DedbyAOIBG2MM.print_area = 'A1:' + get_column_letter(DedbyAOIBG2MM.max_column) + str(DedbyAOIBG2MM.max_row+1)
    DedbyAOIBG2MM.page_setup.fitToPage = True
    DedbyAOIBG2MM.page_setup.fitToHeight = 1
    DedbyAOIBG2MM.page_setup.fitToWidth = 1

    DedbyAOISCOLMM = wb2.create_sheet(title='Deductible by AOI SCOL')
    formatwkstMM(wkstname=DedbyAOISCOLMM, titlerows='3', A1title='RULE 81. DEDUCTIBLE INSURANCE PLAN',
                 A2title='Special Cause of Loss Factors', dfname=DIPSCOL, statename=State,
                 stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(DedbyAOISCOLMM)
    for row in range(4, DedbyAOISCOLMM.max_row + 1):
        for col in range(1, DedbyAOISCOLMM.max_column + 1):
            cell = DedbyAOISCOLMM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = DedbyAOISCOLMM["A" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOISCOLMM["D" + str(row)]
        cell.number_format = "0"
        cell = DedbyAOISCOLMM["G" + str(row)]
        cell.number_format = "0"
    DedbyAOISCOLMM.insert_cols(4,1)
    DedbyAOISCOLMM.insert_cols(8,1)
    for col in DedbyAOISCOLMM.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try: # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                max_length = 1
                pass
        adjusted_width = (max_length + 2) * 1.2
        DedbyAOISCOLMM.column_dimensions[column].width = adjusted_width
    DedbyAOISCOLMM.column_dimensions['G'].width = 9.5
    DedbyAOISCOLMM.column_dimensions['K'].width = 9.5
    DedbyAOISCOLMM.print_area = 'A1:' + get_column_letter(DedbyAOISCOLMM.max_column) + str(DedbyAOISCOLMM.max_row+1)
    DedbyAOISCOLMM.page_setup.fitToPage = True
    DedbyAOISCOLMM.page_setup.fitToHeight = 1
    DedbyAOISCOLMM.page_setup.fitToWidth = 1

#<editor-fold desc="MM Full Manual">
if v10.get() == 1 and v13.get() == 1:
    ExpMM = wb2.create_sheet(title='Expense Constant')
    formatwkstMM(wkstname=ExpMM, titlerows='3', A1title='RULE 8 EXPENSE CONSTANT', A2title='8. Rate and Premium Computation', dfname=ExpConstDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(ExpMM)

    MinPMM = wb2.create_sheet(title='Minimum Premium')
    formatwkstMM(wkstname=MinPMM, titlerows='3', A1title='RULE 8 EXPENSE CONSTANT', A2title='8. Rate and Premium Computation', dfname=PolicyMinDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(MinPMM)

if v12.get() == 1:
    AOIBG1MM = wb2.create_sheet(title='AOI Bldg Grp I')
    formatwkstMM(wkstname=AOIBG1MM, titlerows='3', A1title='AOI CURVES', A2title='Limit of Insurance Relativity Factors Building Basic Group I', dfname=BGIBLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIBG1MM)
    for row in range(4, AOIBG1MM.max_row + 1):
        for col in range(1, AOIBG1MM.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIBG1MM[char + str(row)]
            cell.number_format = rateFormat

    AOIBG2MM = wb2.create_sheet(title='AOI Bldg Grp II')
    formatwkstMM(wkstname=AOIBG2MM, titlerows='3', A1title='AOI CURVES', A2title='Basic Group II Limit Factors Building', dfname=BGIIBLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIBG2MM)
    for row in range(4, AOIBG2MM.max_row + 1):
        for col in range(1, AOIBG2MM.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIBG2MM[char + str(row)]
            cell.number_format = rateFormat

    AOIBSCOLMM = wb2.create_sheet(title='AOI Bldg SCOL')
    formatwkstMM(wkstname=AOIBSCOLMM, titlerows='3', A1title='AOI CURVES', A2title='Special Cause of Loss Limit Factors Building', dfname=SCOLBLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIBSCOLMM)
    for row in range(4, AOIBSCOLMM.max_row + 1):
        for col in range(1, AOIBSCOLMM.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIBSCOLMM[char + str(row)]
            cell.number_format = rateFormat

    AOIPPG1MM = wb2.create_sheet(title='AOI PP Grp I')
    formatwkstMM(wkstname=AOIPPG1MM, titlerows='3', A1title='AOI CURVES', A2title='Basic Group I Limit Factors Contents', dfname=BGIPPLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIPPG1MM)
    for row in range(4, AOIPPG1MM.max_row + 1):
        for col in range(1, AOIPPG1MM.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIPPG1MM[char + str(row)]
            cell.number_format = rateFormat

    AOIPPG2MM = wb2.create_sheet(title='AOI PP Grp II')
    formatwkstMM(wkstname=AOIPPG2MM, titlerows='3', A1title='AOI CURVES', A2title='Basic Group II Limit Factors Contents', dfname=BGIIPPLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIPPG2MM)
    for row in range(4, AOIPPG2MM.max_row + 1):
        for col in range(1, AOIPPG2MM.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIPPG2MM[char + str(row)]
            cell.number_format = rateFormat

    AOIPPSCOLMM = wb2.create_sheet(title='AOI PP SCOL')
    formatwkstMM(wkstname=AOIPPSCOLMM, titlerows='3', A1title='AOI CURVES', A2title='Special Cause of Loss Limit Factors Contents', dfname=SCOLPPLOI, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOIPPSCOLMM)
    for row in range(4, AOIPPSCOLMM.max_row + 1):
        for col in range(1, AOIPPSCOLMM.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            cell = AOIPPSCOLMM[char + str(row)]
            cell.number_format = rateFormat

    TierG1MM = wb2.create_sheet(title='Tiering Grp I')
    formatwkstMM(wkstname=TierG1MM, titlerows='3', A1title='TIERING GRADE FACTORS', A2title='Basic Group I Tiering Factors', dfname=BGITier, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TierG1MM)
    for row in range(4, TierG1MM.max_row + 1):
        for col in range(1, TierG1MM.max_column + 1):
            cell = TierG1MM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    TierG2MM = wb2.create_sheet(title='Tiering Grp II')
    formatwkstMM(wkstname=TierG2MM, titlerows='3', A1title='TIERING GRADE FACTORS', A2title='Basic Group II Tiering Factors', dfname=BGIITier, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TierG2MM)
    for row in range(4, TierG2MM.max_row + 1):
        for col in range(1, TierG2MM.max_column + 1):
            cell = TierG2MM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    TierSCOLMM = wb2.create_sheet(title='Tiering SCOL')
    formatwkstMM(wkstname=TierSCOLMM, titlerows='3', A1title='TIERING GRADE FACTORS', A2title='Special Cause of Loss Tiering Factors', dfname=SCOLTier, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(TierSCOLMM)
    for row in range(4, TierSCOLMM.max_row + 1):
        for col in range(1, TierSCOLMM.max_column + 1):
            cell = TierSCOLMM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat

    AOBG1MM = wb2.create_sheet(title='Age of Building Grp I')
    formatwkstMM(wkstname=AOBG1MM, titlerows='3', A1title='AGE OF BUILDING FACTORS', A2title='Basic Group I Factors', dfname=BGIAOB, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOBG1MM)
    for row in range(4, AOBG1MM.max_row + 1):
        for col in range(1, AOBG1MM.max_column + 1):
            cell = AOBG1MM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = AOBG1MM["A" + str(row)]
        cell.number_format = "0"
        cell = AOBG1MM["E" + str(row)]
        cell.number_format = "0"
        cell = AOBG1MM["I" + str(row)]
        cell.number_format = "0"
    AOBG1MM.column_dimensions['B'].width = 12
    AOBG1MM.column_dimensions['C'].width = 12
    AOBG1MM.column_dimensions['D'].width = 12
    AOBG1MM.insert_cols(5, 1)
    AOBG1MM.insert_cols(10, 1)
    AOBG1MM.column_dimensions['H'].width = 12
    AOBG1MM.column_dimensions['J'].width = 5.8
    AOBG1MM.column_dimensions['I'].width = 12
    AOBG1MM.column_dimensions['M'].width = 12
    AOBG1MM.column_dimensions['N'].width = 12
    AOBG1MM.print_area = 'A1:' + get_column_letter(AOBG1MM.max_column+1) + str(AOBG1MM.max_row + 1)

    AOBG2MM = wb2.create_sheet(title='Age of Building Grp II')
    formatwkstMM(wkstname=AOBG2MM, titlerows='3', A1title='AGE OF BUILDING FACTORS', A2title='Basic Group II Factors', dfname=BGIIAOB, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOBG2MM)
    for row in range(4, AOBG2MM.max_row + 1):
        for col in range(1, AOBG2MM.max_column + 1):
            cell = AOBG2MM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = AOBG2MM["A" + str(row)]
        cell.number_format = "0"
        cell = AOBG2MM["E" + str(row)]
        cell.number_format = "0"
        cell = AOBG2MM["I" + str(row)]
        cell.number_format = "0"
    AOBG2MM.column_dimensions['B'].width = 12
    AOBG2MM.column_dimensions['C'].width = 12
    AOBG2MM.column_dimensions['D'].width = 12
    AOBG2MM.insert_cols(5, 1)
    AOBG2MM.insert_cols(10, 1)
    AOBG2MM.column_dimensions['H'].width = 12
    AOBG2MM.column_dimensions['J'].width = 5.8
    AOBG2MM.column_dimensions['I'].width = 12
    AOBG2MM.column_dimensions['M'].width = 12
    AOBG2MM.column_dimensions['N'].width = 12
    AOBG2MM.print_area = 'A1:' + get_column_letter(AOBG2MM.max_column+1) + str(AOBG2MM.max_row + 1)

    AOBSCOLMM = wb2.create_sheet(title='Age of Building SCOL')
    formatwkstMM(wkstname=AOBSCOLMM, titlerows='3', A1title='AGE OF BUILDING FACTORS', A2title='Special Cause of Loss Factors', dfname=SCOLAOB, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    borderfnct(AOBSCOLMM)
    for row in range(4, AOBSCOLMM.max_row + 1):
        for col in range(1, AOBSCOLMM.max_column + 1):
            cell = AOBSCOLMM[get_column_letter(col) + str(row)]
            cell.number_format = rateFormat
        cell = AOBSCOLMM["A" + str(row)]
        cell.number_format = "0"
        cell = AOBSCOLMM["E" + str(row)]
        cell.number_format = "0"
        cell = AOBSCOLMM["I" + str(row)]
        cell.number_format = "0"
    AOBSCOLMM.column_dimensions['B'].width = 12
    AOBSCOLMM.column_dimensions['C'].width = 12
    AOBSCOLMM.column_dimensions['D'].width = 12
    AOBSCOLMM.insert_cols(5, 1)
    AOBSCOLMM.insert_cols(10, 1)
    AOBSCOLMM.column_dimensions['H'].width = 12
    AOBSCOLMM.column_dimensions['J'].width = 5.8
    AOBSCOLMM.column_dimensions['I'].width = 12
    AOBSCOLMM.column_dimensions['M'].width = 12
    AOBSCOLMM.column_dimensions['N'].width = 12
    AOBSCOLMM.print_area = 'A1:' + get_column_letter(AOBSCOLMM.max_column+1) + str(AOBSCOLMM.max_row + 1)

else:
    pass

EmergencyMM = wb2.create_sheet(title='Emergency Evacuation')
formatwkstMM(wkstname=EmergencyMM, titlerows='3', A1title='EMERGENCY EVACUATION INCLUDING CIVIL AUTHORITY COVERAGE', A2title='', dfname=EmergencyDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(EmergencyMM)
addlformat(EmergencyMM)

HumanSvEnhMM = wb2.create_sheet(title='Human Serv. Enhancement Cover.')
formatwkstMM(wkstname=HumanSvEnhMM, titlerows='3', A1title='HUMAN SERVICES PROPERTY ENDORSEMENT FACTOR', A2title='', dfname=HumSvcsEnhDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(HumanSvEnhMM)
addlformat(HumanSvEnhMM)

ClientsMM = wb2.create_sheet(title='Clients Property Coverage')
formatwkstMM(wkstname=ClientsMM, titlerows='3', A1title='CLIENTS PROPERTY COVERAGE FACTOR', A2title='', dfname=ClientsDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(ClientsMM)
addlformat(ClientsMM)

HumanSvcWorkMM = wb2.create_sheet(title='Human Serv. Workplace Violence')
formatwkstMM(wkstname=HumanSvcWorkMM, titlerows='3', A1title='HUMAN SERVICES WORKPLACE VIOLENCE LOSS OF INCOME', A2title='', dfname=HumSvcsDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(HumanSvcWorkMM)
addlformat(HumanSvcWorkMM)

HumanSvcPrpMM = wb2.create_sheet(title='Human Serv. Prop Endorsement')
formatwkstSM(wkstname=HumanSvcPrpMM, titlerows='3', A1title='HUMAN SERVICES PROPERTY ENDORSEMENT FACTOR', A2title='', dfname=HumSvcsFctDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(HumanSvcPrpMM)
addlformat(HumanSvcPrpMM)

SeniorPropMM = wb2.create_sheet(title='Senior Living Property Factor')
formatwkstMM(wkstname=SeniorPropMM, titlerows='3', A1title='SENIOR LIVING COMMUNITIES PROPERTY ENDORSEMENT', A2title='', dfname=SeniorDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(SeniorPropMM)
addlformat(SeniorPropMM)

WaterDmgMM = wb2.create_sheet(title='Water Damage Deductible Weight')
formatwkstMM(wkstname=WaterDmgMM, titlerows='3', A1title='WATER DAMAGE DEDUCTIBLE WEIGHT', A2title='', dfname=WtrDmgDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(WaterDmgMM)
addlformat(WaterDmgMM)

CivilTimeMM = wb2.create_sheet(title='Civil Authority Time Element')
formatwkstMM(wkstname=CivilTimeMM, titlerows='3', A1title='RULE 54. E. CIVIL AUTHORITY CHANGES', A2title='Table 54.E.3.b: Civil Authority Increased Radius Coverage Factor', dfname=CivilDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(CivilTimeMM)
addlformat(CivilTimeMM)
CivilTimeMM.column_dimensions['E'].width = 14

Cosmetic = wb2.create_sheet(title='Cosmetic Exclusion')
formatwkst(wkstname=Cosmetic, titlerows='3', A1title='COSMETIC EXCLUSION ON COVERAGE FOR SIDING', A2title='Multiply the Basic Group II building rate by the below factor', dfname=CosmeticDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(Cosmetic)
addlformat(Cosmetic)

Terrorism = wb2.create_sheet(title='Terrorism')
formatwkst(wkstname=Terrorism, titlerows='3', A1title='Coverage for Certified Acts of Terrorism', A2title='', dfname=TerrorismDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(Terrorism)
addlformat(Terrorism)

LPDPMM = wb2.create_sheet(title='LPDP')
formatwkstMM(wkstname=LPDPMM, titlerows='3', A1title='Large Premium Discount Factor', A2title='', dfname=LPDPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(LPDPMM)

Money304MM = wb2.create_sheet(title='Money and Securities 304')
formatwkstMM(wkstname=Money304MM, titlerows='3', A1title='304 Money and Securities', A2title='', dfname=CrimeMSDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(Money304MM)

Money304RFMM = wb2.create_sheet(title='Money and Securities 304 RF')
formatwkstMM(wkstname=Money304RFMM, titlerows='3', A1title='Rating Factor', A2title='Base Premium', dfname=MSDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(Money304RFMM)
for row in range(4, Money304RFMM.max_row + 1):
    for col in range(1, Money304RFMM.max_column + 1):
        cell = Money304RFMM[get_column_letter(col) + str(row)]
        cell.number_format = rateFormat
    cell = Money304RFMM["A" + str(row)]
    cell.number_format = "0"
    cell = Money304RFMM["B" + str(row)]
    cell.number_format = "#,##0"

Employee306MM = wb2.create_sheet(title='Employee Dishonesty 306')
formatwkstMM(wkstname=Employee306MM, titlerows='3', A1title='306 Employee Dishonesty', A2title='', dfname=EmployeeDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(Employee306MM)

CompFraudMM = wb2.create_sheet(title='Computer Fraud 311')
formatwkstMM(wkstname=CompFraudMM, titlerows='3', A1title='311 Computer and Funds Transfer Fraud', A2title='Base Premium', dfname=ComputerDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(CompFraudMM)

CompFraudGSFMM = wb2.create_sheet(title='Computer Fraud 311 GSF')
formatwkstMM(wkstname=CompFraudGSFMM, titlerows='3', A1title='Annual Gross Sales Factor', A2title='', dfname=ComputerAnnualDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(CompFraudGSFMM)

CounterfeitMM = wb2.create_sheet(title='Counterfeit Currency 313')
formatwkstMM(wkstname=CounterfeitMM, titlerows='3', A1title='313 Counterfeit Currency', A2title='', dfname=CFCurrDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(CounterfeitMM)
for row in range(4, CounterfeitMM.max_row + 1):
    for col in range(1, CounterfeitMM.max_column + 1):
        cell = CounterfeitMM[get_column_letter(col) + str(row)]
        cell.number_format = rateFormat
    cell = CounterfeitMM["A" + str(row)]
    cell.number_format = "#,##0"

ForgeryMM = wb2.create_sheet(title='Forgery or Alteration 314')
formatwkstMM(wkstname=ForgeryMM, titlerows='3', A1title='314 Forgery or Alteration', A2title='Base Premium', dfname=ForgeryDF2, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(ForgeryMM)

Fraud516MM = wb2.create_sheet(title='Fraudulent Impersonation 516')
formatwkstMM(wkstname=Fraud516MM, titlerows='3', A1title='516 Fraudulent Impersonation', A2title='Base Premium', dfname=FraudDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(Fraud516MM)

WineLeakMM = wb2.create_sheet(title='Wine Leakage Endorsement')
formatwkstMM(wkstname=WineLeakMM, titlerows='3', A1title='WINE LEAKAGE ENDORSEMENT', A2title='', dfname=WineLkgDF1, statename=State, stabb=StateAbb, effdate=EffectiveDate)
for r in dataframe_to_rows(WineLkgDF2, False, True):
    #The header is the first row and the index is the second row, but they need to be on the
    #same row in Excel
    if False & len(list(r))==1: # Checking to see when the index row is reached since it
        #will contain only the number of the indices that are in the dataframe (currently only
        #works for 1 index)
        WineLeakMM['B5'] = list(r)[0] # Manually adding the 1 index value to the header row in Excel
        continue
    WineLeakMM.append(r)
borderfnct(WineLeakMM)
WineLeakMM.delete_rows(3)
WineLeakMM.insert_rows(4)
WineLeakMM.delete_rows(2)
WineLeakMM.insert_rows(2)
addlformat(WineLeakMM)
WineLeakMM['E1'].alignment = Alignment(horizontal='left')
WineLeakMM.column_dimensions['E'].width = 14.8
for row in range(4, WineLeakMM.max_row + 1):
    for col in range(1, WineLeakMM.max_column + 1):
        cell = WineLeakMM[get_column_letter(col) + str(row)]
        cell.number_format = rateFormat
    cell = WineLeakMM["E" + str(row)]
    cell.number_format = "#,##0"

BIALSMM = wb2.create_sheet(title='BI ALS')
formatwkstMM(wkstname=BIALSMM, titlerows='3', A1title='RULE 51. BUSINESS INCOME COVERAGE OPTIONS', A2title='Table 51.C.3: Business Income Actual Loss Sustained Factors', dfname=BusinessDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(BIALSMM)
for row in range(4, BIALSMM.max_row + 1):
    for col in range(1, BIALSMM.max_column + 1):
        cell = BIALSMM[get_column_letter(col) + str(row)]
        cell.number_format = rateFormat

NutHullerMM = wb2.create_sheet(title='Nut Hullers')
formatwkstMM(wkstname=NutHullerMM, titlerows='3', A1title='NUT HULLERS AND PROCESSORS COVERAGE', A2title='Charge the below premiums', dfname=NutHullProcDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(NutHullerMM)
addlformat(NutHullerMM)

ConAtmoMM = wb2.create_sheet(title='Con Atmo and Storage')
formatwkstMM(wkstname=ConAtmoMM, titlerows='3', A1title='CONTROLLED ATMOSPHERE AND STORAGE COVERAGE', A2title='a. The base rate for all types of occupancies is below (per $100)', dfname=ControlAtmDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(ConAtmoMM)
addlformat(ConAtmoMM)
ConAtmoMM['E5'] = 'b. Modify the rate by the Property Deductible factor, if applicable.'
ConAtmoMM['E5'].alignment = Alignment(horizontal='center')
ConAtmoMM['E5'].font = Font(name='Arial',size=10)
ConAtmoMM['E6'] = 'c. Multiply the rate thus developed by the applicable total Limit of Insurance (per $100)'
ConAtmoMM['E6'].alignment = Alignment(horizontal='center')
ConAtmoMM['E6'].font = Font(name='Arial',size=10)
ConAtmoMM['E7'] = ' to determine the Manual Premium.'
ConAtmoMM['E7'].alignment = Alignment(horizontal='center')
ConAtmoMM['E7'].font = Font(name='Arial',size=10)
ConAtmoMM.print_area = 'A1:' + get_column_letter(ConAtmoMM.max_column + 3) + str(ConAtmoMM.max_row + 1)
ConAtmoMM.page_setup.fitToPage = True

ConAtmoStorMM = wb2.create_sheet(title='Con Atmo Stor and Chem Inj')
formatwkstMM(wkstname=ConAtmoStorMM, titlerows='3', A1title='CONTROLLED ATMOSPHERE, STORAGE AND CHEMICAL INJURY COVERAGE', A2title='a. The base rate for all types of occupancies is below (per $100)', dfname=ControlAtChInjDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(ConAtmoStorMM)
addlformat(ConAtmoStorMM)
ConAtmoStorMM['E6'] = 'b. Modify the rate by the Property Deductible factor, if applicable.'
ConAtmoStorMM['E6'].alignment = Alignment(horizontal='center')
ConAtmoStorMM['E6'].font = Font(name='Arial',size=10)
ConAtmoStorMM['E7'] = 'c. Multiply the rate thus developed by the applicable total Limit of Insurance (per $100)'
ConAtmoStorMM['E7'].alignment = Alignment(horizontal='center')
ConAtmoStorMM['E7'].font = Font(name='Arial',size=10)
ConAtmoStorMM['E8'] = ' to determine the Manual Premium.'
ConAtmoStorMM['E8'].alignment = Alignment(horizontal='center')
ConAtmoStorMM['E8'].font = Font(name='Arial',size=10)
ConAtmoStorMM.column_dimensions['E'].width = 11.5
ConAtmoStorMM.print_area = 'A1:' + get_column_letter(ConAtmoStorMM.max_column + 3) + str(ConAtmoStorMM.max_row + 1)
ConAtmoStorMM.page_setup.fitToPage = True

HopsGrowerMM = wb2.create_sheet(title='Hops Grower')
formatwkstMM(wkstname=HopsGrowerMM, titlerows='3', A1title='HOPS GROWER', A2title='', dfname=HopsDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(HopsGrowerMM)
addlformat(HopsGrowerMM)

FruitTreeMM = wb2.create_sheet(title='Fruit Trees and Trelisses')
formatwkstMM(wkstname=FruitTreeMM, titlerows='3', A1title='FRUIT TREES, TRELLISES, STAKES & VINES', A2title='', dfname=FruitDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(FruitTreeMM)
addlformat(FruitTreeMM)

WineryMM = wb2.create_sheet(title='Winery Endorsement')
formatwkstMM(wkstname=WineryMM, titlerows='3', A1title='WINERY ENDORSEMENT', A2title='', dfname=WineEndDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(WineryMM)
addlformat(WineryMM)

WineryIncMM = wb2.create_sheet(title='Winery Increased Limits Schedul')
formatwkstMM(wkstname=WineryIncMM, titlerows='3', A1title='WINERY INCREASED LIMITS SCHEDULE', dfname=WineryDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(WineryIncMM)
for row in range(4, WineryIncMM.max_row + 1):
    for col in range(1, WineryIncMM.max_column + 1):
        cell = WineryIncMM[get_column_letter(col) + str(row)]
        cell.number_format = "#,##0"

CPPPlusMM = wb2.create_sheet(title='PLUS Endorsements CPP')
formatwkstMM(wkstname=CPPPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='COMMERCIAL PROPERTY PROTECTION PLUS ENDORSEMENTS', dfname=PropPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(CPPPlusMM)

BMGPlusMM = wb2.create_sheet(title='PLUS Endorsements BMG')
formatwkstMM(wkstname=BMGPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='BEVERAGE MAKERS GOLD & PLATINUM PROTECTION PLUS ENDORSEMENT', dfname=BrewPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(BMGPlusMM)

GCGPlusMM = wb2.create_sheet(title='PLUS Endorsements GCG')
formatwkstMM(wkstname=GCGPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='GOLF COURSE GOLD & PLATINUM PROTECTION PLUS ENDORSEMENT', dfname=GolfPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(GCGPlusMM)

MGPlusMM = wb2.create_sheet(title='PLUS Endorsements MG')
formatwkstMM(wkstname=MGPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='MANUFACTURER GOLD & PLATINUM PROTECTION PLUS ENDORSEMENT', dfname=ManuPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(MGPlusMM)

RGPlusMM = wb2.create_sheet(title='PLUS Endorsements RG')
formatwkstMM(wkstname=RGPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='RESTAURANT GOLD & PLATINUM PROTECTION PLUS ENDORSEMENTS', dfname=RestPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(RGPlusMM)

WGPlusMM = wb2.create_sheet(title='PLUS Endorsements WG')
formatwkstMM(wkstname=WGPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='WHOLESALER GOLD & PLATINUM PROTECTION PLUS ENDORSEMENTS', dfname=WholePPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(WGPlusMM)

HGPlusMM = wb2.create_sheet(title='PLUS Endorsements HG')
formatwkstMM(wkstname=HGPlusMM, titlerows='3', A1title='PROTECTION PLUS ENDORSEMENTS', A2title='HOTEL GOLD & PLATINUM PROTECTION PLUS ENDORSEMENTS', dfname=HotelPPDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(HGPlusMM)

FranchiseMM = wb2.create_sheet(title='Franchise Upgrade Endorsement')
formatwkstMM(wkstname=FranchiseMM, titlerows='3', A1title='Franchise Upgrade Endorsement', A2title='', dfname=FranUpgDF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(FranchiseMM)
addlformat(FranchiseMM)

NutsHullsMM = wb2.create_sheet(title='Nuts, Hulls, or Shells')
formatwkstMM(wkstname=NutsHullsMM, titlerows='3', A1title='NUTS, HULLS, OR SHELLS IN THE OPEN', A2title='', dfname=NutHullILF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
borderfnct(NutsHullsMM)
for row in range(4, NutsHullsMM.max_row + 1):
    for col in range(1, NutsHullsMM.max_column + 1):
        cell = NutsHullsMM[get_column_letter(col) + str(row)]
        cell.number_format = rateFormat
    cell = NutsHullsMM["A" + str(row)]
    cell.number_format = "#,##0"
    cell = NutsHullsMM["B" + str(row)]
    cell.number_format = "#,##0"

#<editor-fold desc="Save SM Workbook">
if v5.get() == 1:
    sheetNames = wb.sheetnames # A list containing the name of all worksheets in the workbook
    for i in range(1, len(sheetNames)):
        if wb[sheetNames[i]]['A3'].value == '*':
            wb["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames[i]}'!E1")
            wb["Index"]['A' + str(i)].value = wb[sheetNames[i]]['E1'].value
            wb["Index"]['A' + str(i)].font = fontBlue
        else:
            wb["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames[i]}'!A1")
            wb["Index"]['A' + str(i)].value = wb[sheetNames[i]]['A1'].value
            wb["Index"]['A' + str(i)].font = fontBlue
    wb.save(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Small Market Rate Pages.xlsx'))
    wb = xw.Book(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Small Market Rate Pages.xlsx'))
    current_label = wb.api.SensitivityLabel.GetLabel()
    if current_label.LabelName == "":
        labelinfo = wb.api.SensitivityLabel.CreateLabelInfo()
        labelinfo.AssignmentMethod = 2
        labelinfo.Justification = 'init'
        labelinfo.LabelId = 'fbefcacb-1e54-47c6-a321-a2f3e970fe0d'
        labelinfo.LabelName = 'Restricted'
        wb.api.SensitivityLabel.SetLabel(labelinfo, labelinfo)
    wb.save()
#</editor-fold>

#<editor-fold desc="Save MM Workbook">
if v10.get() == 1:
    sheetNames2 = wb2.sheetnames # A list containing the name of all worksheets in the workbook
    for i in range(1, len(sheetNames2)):
        if wb2[sheetNames2[i]]['A3'].value == '*':
            wb2["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames2[i]}'!E1")
            wb2["Index"]['A' + str(i)].value = wb2[sheetNames2[i]]['E1'].value
            wb2["Index"]['A' + str(i)].font = fontBlue
        else:
            wb2["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames2[i]}'!A1")
            wb2["Index"]['A' + str(i)].value = wb2[sheetNames2[i]]['A1'].value
            wb2["Index"]['A' + str(i)].font = fontBlue
    wb2.save(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Middle Market Rate Pages.xlsx'))
    wb2 = xw.Book(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Middle Market Rate Pages.xlsx'))
    current_label = wb2.api.SensitivityLabel.GetLabel()
    if current_label.LabelName == "":
        labelinfo = wb2.api.SensitivityLabel.CreateLabelInfo()
        labelinfo.AssignmentMethod = 2
        labelinfo.Justification = 'init'
        labelinfo.LabelId = 'fbefcacb-1e54-47c6-a321-a2f3e970fe0d'
        labelinfo.LabelName = 'Restricted'
        wb2.api.SensitivityLabel.SetLabel(labelinfo, labelinfo)
    wb2.save()
#</editor-fold>

#<editor-fold desc="Time taken">
print(time.time() - start_time)
#</editor-fold>