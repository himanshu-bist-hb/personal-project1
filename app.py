# ba_rate_pages_app.py  —  streamlit run app.py

import copy
import io
import os
import re
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import numpy as np
import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.properties import PageSetupProperties

import subprocess, sys

st.set_page_config(
    page_title="RatePage Builder · Nationwide",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Session state ────────────────────────────────────────────────────────────
REQUIRED = ["NGIC", "MM", "NACO", "NAFF", "NICOF", "HICNJ", "CCMIC", "NWAG"]
OPTIONAL = ["CW"]
ALL_KEYS = REQUIRED + OPTIONAL

for k in ALL_KEYS:
    st.session_state.setdefault(f"file_{k}", None)
st.session_state.setdefault("save_dir",    "")
st.session_state.setdefault("sched_mod",   0)
st.session_state.setdefault("run_status",  "idle")
st.session_state.setdefault("run_msg",     "")
st.session_state.setdefault("xlsx_path",   "")
st.session_state.setdefault("pdf_path",    "")
st.session_state.setdefault("pdf_status",  "idle")
st.session_state.setdefault("lob",          "Business Auto")
st.session_state.setdefault("confirm_step", "idle")
st.session_state.setdefault("upload_reset", 0)
st.session_state.setdefault("mode",          "individual")
st.session_state.setdefault("multi_src_dir", "")
st.session_state.setdefault("multi_save_dir","")
st.session_state.setdefault("multi_gen_pdf", False)
st.session_state.setdefault("multi_step",    "idle")
st.session_state.setdefault("multi_results", [])
st.session_state.setdefault("multi_sched",        0)    # kept for compat
st.session_state.setdefault("multi_sched_map",    {})   # {state_name: int} per-state mods
st.session_state.setdefault("multi_sched_mode",   "upload")  # "upload" | "manual"
st.session_state.setdefault("multi_sched_excel",  {})   # raw parsed excel {state: rating}

st.session_state.setdefault("active_tool",             "rate_pages")
st.session_state.setdefault("cmp_comp_data",           None)
st.session_state.setdefault("cmp_file_ids",            (None, None))
st.session_state.setdefault("cmp_tracked_bytes",       None)
st.session_state.setdefault("cmp_tracked_filename",    None)
st.session_state.setdefault("cmp_tracked_fmt",         None)
st.session_state.setdefault("cmp_mass_results",        None)
st.session_state.setdefault("cmp_mass_zip_bytes",      None)
st.session_state.setdefault("cmp_mass_file_ids",       (None, None))
st.session_state.setdefault("cmp_mass_fmt",            None)
st.session_state.setdefault("cmp_mass_save_dir",       None)
st.session_state.setdefault("cmp_mass_output_path",    "")

LOB_NAV = [("Business Auto","🚗"),
    ("Farm Auto",         "🚜"),
    ("General Liability", "⚖️"),
    ("Property",          "🏠"),
]
LOB_NAMES   = [l for l, _ in LOB_NAV]
LOB_OPTIONS = [f"{ic}  {nm}" for nm, ic in LOB_NAV]

# ─── Helpers ──────────────────────────────────────────────────────────────────
def _valid(k):
    v = st.session_state.get(f"file_{k}")
    return v is not None and "error" not in v

def n_req():    return sum(1 for k in REQUIRED if _valid(k))
def all_req():  return n_req() == len(REQUIRED)
def any_req():  return n_req() > 0
def has_ngic(): return _valid("NGIC")

def chip(f):
    if f:
        return f'<span class="chip-ok">&#10003; {Path(f["name"]).stem[:17]}</span>'
    return '<span class="chip-none">&#8212; none</span>'

def browse_folder():
    # Run tkinter in a separate process — calling tk.Tk() from a Streamlit
    # background thread on Windows crashes the server process entirely.
    try:
        result = subprocess.run(
            [sys.executable, "-c",
             "import tkinter as tk; from tkinter import filedialog; "
             "root=tk.Tk(); root.withdraw(); root.wm_attributes('-topmost',True); "
             "print(filedialog.askdirectory(title='Select Save Location') or '', end='')"],
            capture_output=True, text=True, timeout=120,
        )
        folder = result.stdout.strip()
        return folder or None
    except Exception:
        return None

def spacer(px=20):
    st.markdown(f"<div style='height:{px}px'></div>", unsafe_allow_html=True)


# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=Libre+Baskerville:wght@700&display=swap');

/* ── tokens ── */
:root {
  --nw-blue:     #1A5DAB;
  --nw-deep:     #0D3F7A;
  --nw-lt:       #EBF2FB;
  --gold:        #C8A951;
  --gold-lt:     #EDD97A;
  --off:         #F4F7FB;
  --surface:     #FFFFFF;
  --border:      #D4DFEF;
  --text:        #0C1A35;
  --muted:       #6B7A9E;
  --ok-bg:       #EAF5EE;
  --ok-fg:       #196B38;
  --radius:      10px;
  --shadow:      0 2px 12px rgba(13,63,122,0.08);
  --content-pad: 72px;   /* ← single knob: push main content right/left */
}

/* ── kill white top gap ── */
html, body { margin:0 !important; padding:0 !important; background:var(--off) !important; }
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
section.main { background:var(--off) !important; padding:0 !important; margin:0 !important; }
.block-container {
  background:var(--off) !important;
  padding: 0 var(--content-pad) 64px var(--content-pad) !important;
  margin:0 !important; max-width:100% !important;
}
.stMainBlockContainer { padding-top:0 !important; }
#MainMenu, footer,
[data-testid="stToolbar"],
[data-testid="stDecoration"],
[data-testid="stHeader"],
header { display:none !important; height:0 !important; min-height:0 !important; }
[data-testid="stVerticalBlock"] > div:first-child { padding-top:0 !important; }

/* ── sidebar shell ── */
[data-testid="stSidebar"] {
  background: #F7F9FC !important;
  border-right: 1px solid #DDE5F0 !important;
  padding: 0 !important;
}
[data-testid="stSidebar"] > div:first-child {
  padding: 0 !important;
  background: #F7F9FC !important;
}


/* ══════════════════════════════════════════════════════════════════
   SIDEBAR NAV  —  st.radio disguised as a nav menu
   Strategy: hide every Streamlit chrome element, style only the
   <label> elements.  The CHECKED state is read via aria-checked="true"
   on the <label> itself — this is set by the browser's native radio
   behaviour and is 100% reliable across Chrome / Edge / Firefox.
══════════════════════════════════════════════════════════════════ */

/* Hide the radio group label */
[data-testid="stSidebar"] [data-testid="stRadio"] > label { display:none !important; }

/* Zero out EVERY wrapper between the sidebar edge and the label */
[data-testid="stSidebar"] [data-testid="stRadio"] > div { gap:0 !important; padding:0 !important; margin:0 !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] [data-baseweb="radio"] { margin:0 !important; padding:0 !important; width:100% !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] [data-baseweb="radio"] > div { margin:0 !important; padding:0 !important; width:100% !important; }
[data-testid="stSidebar"] [data-testid="stRadio"] > div > div { margin:0 !important; padding:0 !important; width:100% !important; }

/* Hide the actual circle dot */
[data-testid="stSidebar"] [data-testid="stRadio"] label > div:first-child { display:none !important; }

/* ── base style for every nav label ── */
[data-testid="stSidebar"] [data-testid="stRadio"] label {
  display: flex !important;
  align-items: center !important;
  gap: 10px !important;
  width: 100% !important;
  min-width: 100% !important;
  margin: 0 !important;
  padding: 13px 20px !important;
  border-radius: 0 !important;
  cursor: pointer !important;
  font-size: 13px !important;
  font-family: 'Inter', sans-serif !important;
  font-weight: 500 !important;
  color: #566278 !important;
  background: transparent !important;
  border: none !important;
  border-left: 3px solid transparent !important;
  transition: background 0.14s ease, color 0.14s ease !important;
  user-select: none !important;
  box-sizing: border-box !important;
}

/* ── hover ── */
[data-testid="stSidebar"] [data-testid="stRadio"] label:hover {
  background: #EBF2FB !important;
  color: #0D3F7A !important;
  border-left: 3px solid #A8C4E8 !important;
}

/* ── SELECTED item — full-width blue row, every selector variant ── */

/* 1. aria-checked on label (Streamlit >= 1.28) */
[data-testid="stSidebar"] [data-testid="stRadio"] label[aria-checked="true"] {
  background: linear-gradient(90deg, #1A5DAB 0%, #0D3F7A 100%) !important;
  color: #FFFFFF !important;
  font-weight: 700 !important;
  border-left: 3px solid #EDD97A !important;
  box-shadow: none !important;
}
[data-testid="stSidebar"] [data-testid="stRadio"] label[aria-checked="true"]:hover {
  background: linear-gradient(90deg, #1A5DAB 0%, #0D3F7A 100%) !important;
  color: #FFFFFF !important;
}
            
/* Container div — zero all spacing */
[data-testid="stSidebar"] [data-testid="stImage"]        { margin:0; padding:0; line-height:0; }
[data-testid="stSidebar"] [data-testid="stImage"] > div  { margin:0; padding:0; }

/* The img itself gets the breathing room via padding instead */
[data-testid="stSidebar"] [data-testid="stImage"] img    { display:block; padding:16px 20px 12px; margin:0; }

/* 2. checked input sibling */
[data-testid="stSidebar"] [data-testid="stRadio"] input[type="radio"]:checked ~ div,
[data-testid="stSidebar"] [data-testid="stRadio"] input[type="radio"]:checked + div {
  background: linear-gradient(90deg, #1A5DAB 0%, #0D3F7A 100%) !important;
  color: #FFFFFF !important;
  font-weight: 700 !important;
}

/* 3. data-checked attribute (BaseWeb / older Streamlit) */
[data-testid="stSidebar"] [data-testid="stRadio"] label[data-checked="true"] {
  background: linear-gradient(90deg, #1A5DAB 0%, #0D3F7A 100%) !important;
  color: #FFFFFF !important;
  font-weight: 700 !important;
  border-left: 3px solid #EDD97A !important;
}

/* 4. BaseWeb checked state on the radio wrapper */
[data-testid="stSidebar"] [data-testid="stRadio"] [data-baseweb="radio"][aria-checked="true"] label,
[data-testid="stSidebar"] [data-testid="stRadio"] [data-checked="true"] label {
  background: linear-gradient(90deg, #1A5DAB 0%, #0D3F7A 100%) !important;
  color: #FFFFFF !important;
  font-weight: 700 !important;
  border-left: 3px solid #EDD97A !important;
}

/* ── collapse button in sidebar ── */
[data-testid="stSidebar"] [data-testid="stButton"] button {
  background: #ECF0F7 !important;
  border: 1px solid #D0D9E8 !important;
  color: #3A4D6B !important;
  border-radius: 7px !important;
  font-size: 11px !important;
  font-weight: 600 !important;
  letter-spacing: 0.3px !important;
  padding: 7px 14px !important;
  transition: background 0.13s, color 0.13s !important;
}
[data-testid="stSidebar"] [data-testid="stButton"] button:hover {
  background: #DAE2EF !important;
  color: #0D3F7A !important;
}

/* ── HEADER ── */
.nw-header {
  background: var(--nw-blue);
  margin: 0 calc(-1 * var(--content-pad));
  padding: 0 var(--content-pad);
  display: flex; align-items: center; justify-content: space-between;
  height: 64px;
  box-shadow: 0 2px 16px rgba(13,63,122,0.22);
  position: relative; z-index: 100;
}
.nw-header-left { display:flex; align-items:center; gap:14px; }
.nw-eagle {
  width:32px; height:32px;
  background:rgba(255,255,255,0.15);
  border:1px solid rgba(255,255,255,0.22);
  border-radius:7px;
  display:flex; align-items:center; justify-content:center; font-size:17px;
}
.nw-brand      { font-family:'Libre Baskerville',serif; font-size:16px; color:#fff; }
.nw-brand span { color:var(--gold-lt); }
.nw-sep        { width:1px; height:18px; background:rgba(255,255,255,0.18); }
.nw-pgname     { font-size:12px; font-weight:500; color:rgba(255,255,255,0.68); }
.nw-right      { font-size:10px; color:rgba(255,255,255,0.32); letter-spacing:1.2px; text-transform:uppercase; }
.gold-line {
  height:3px;
  background:linear-gradient(90deg,var(--gold) 0%,var(--gold-lt) 45%,transparent 100%);
  margin:0 calc(-1 * var(--content-pad)) 32px;
}

/* ── section label ── */
.sec-label {
  font-size:10px; font-weight:700; letter-spacing:2.2px; text-transform:uppercase;
  color:var(--nw-blue); display:flex; align-items:center; gap:10px; margin:0 0 16px;
}
.sec-label::after { content:''; flex:1; height:1px; background:var(--border); }

/* ── field helpers ── */
.f-label { font-size:10px; font-weight:700; letter-spacing:1.8px; text-transform:uppercase; color:var(--nw-blue); margin:0 0 8px; }
.f-hint  { font-size:10px; color:var(--muted); margin:5px 0 0; }
.f-ok    { font-size:10px; color:var(--ok-fg); margin:5px 0 0; }

/* ── chips ── */
.chip-ok   { display:inline-flex; align-items:center; gap:4px; background:var(--ok-bg); border:1px solid #9ECDB0; border-radius:5px; padding:2px 8px; font-size:10px; color:var(--ok-fg); font-weight:600; max-width:100%; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; }
.chip-none { display:inline-flex; align-items:center; gap:4px; border:1px dashed var(--border); border-radius:5px; padding:2px 8px; font-size:10px; color:var(--muted); }

/* ── file uploader (multi-file drop zone) ── */
[data-testid="stFileUploader"] { background:var(--surface) !important; border:2px dashed var(--border) !important; border-radius:10px !important; transition:border-color 0.18s !important; }
[data-testid="stFileUploader"]:hover { border-color:var(--nw-blue) !important; }
[data-testid="stFileUploader"] section { padding:20px 16px !important; }
[data-testid="stFileUploaderDropzoneInstructions"] > div > span { font-size:12px !important; color:var(--muted) !important; }
[data-testid="stFileUploaderDropzoneInstructions"] > div > small { font-size:10px !important; color:var(--border) !important; }
[data-testid="stFileUploaderDropzone"] { padding:0 !important; }
[data-testid="stFileUploaderDropzone"] > div { flex-direction:column !important; align-items:center !important; gap:8px !important; }
[data-testid="stFileUploaderDropzone"] button { font-size:11px !important; font-weight:600 !important; padding:7px 18px !important; border-radius:6px !important; min-height:unset !important; background:var(--nw-blue) !important; color:#fff !important; border:none !important; }
[data-testid="stFileUploaderDropzone"] button:hover { background:var(--nw-deep) !important; }
[data-testid="stFileUploadedFile"] { padding:3px 8px !important; font-size:10px !important; }

/* ── assignment table ── */
.assign-wrap { border:1px solid var(--border); border-radius:var(--radius); background:var(--surface); margin-top:14px; overflow:hidden; box-shadow:var(--shadow); }
.assign-hdr  { display:flex; justify-content:space-between; align-items:center; padding:9px 16px; background:var(--off); border-bottom:1px solid var(--border); font-size:9px; font-weight:700; letter-spacing:2px; text-transform:uppercase; color:var(--muted); }
.arow        { display:grid; grid-template-columns:120px 1fr 28px; align-items:center; padding:8px 16px; border-bottom:1px solid var(--border); gap:12px; }
.arow:last-child { border-bottom:none; }
.arow-ok    { background:var(--surface); }
.arow-empty { background:var(--surface); }
.arow-error { background:#FFF5F5; }
.aco        { font-size:11px; font-weight:700; color:var(--text); display:flex; align-items:center; gap:5px; flex-wrap:wrap; }
.afile      { font-size:11px; color:var(--muted); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.afile-assigned { color:var(--text) !important; }
.afile-err  { color:#C8102E !important; font-size:10px; }
.astat      { font-size:12px; font-weight:700; text-align:right; }
.astat-ok   { color:var(--ok-fg); }
.astat-empty{ color:var(--border); font-weight:400; }
.astat-err  { color:#C8102E; }
.ab-req     { display:inline-block; background:#FEE8E8; color:#C8102E; border:1px solid #F5C0C0; border-radius:3px; font-size:8px; font-weight:700; letter-spacing:0.5px; text-transform:uppercase; padding:1px 5px; line-height:1.5; }
.ab-opt     { display:inline-block; background:var(--nw-lt); color:var(--nw-blue); border:1px solid var(--border); border-radius:3px; font-size:8px; font-weight:600; letter-spacing:0.5px; text-transform:uppercase; padding:1px 5px; line-height:1.5; }

/* ── widget labels ── */
label[data-testid="stWidgetLabel"] p { font-size:10px !important; font-weight:700 !important; letter-spacing:1.5px !important; text-transform:uppercase !important; color:var(--nw-blue) !important; margin-bottom:4px !important; }

/* ── inputs ── */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input { border:1.5px solid var(--border) !important; border-radius:7px !important; background:var(--surface) !important; font-family:'Inter',sans-serif !important; font-size:13px !important; color:var(--text) !important; padding:9px 12px !important; }
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus { border-color:var(--nw-blue) !important; box-shadow:0 0 0 3px rgba(26,93,171,0.10) !important; outline:none !important; }

/* ── slider ── */
[data-testid="stSlider"] { padding:4px 0 !important; }
[data-baseweb="slider"] div[role="slider"] { background:var(--nw-blue) !important; border:2px solid #fff !important; box-shadow:0 0 0 2px var(--nw-blue) !important; width:18px !important; height:18px !important; }
[data-baseweb="slider"] div[role="progressbar"] { background:var(--nw-blue) !important; }

/* ── expander ── */
[data-testid="stExpander"] { border:1px solid var(--border) !important; border-radius:var(--radius) !important; background:var(--surface) !important; box-shadow:var(--shadow) !important; margin-bottom:0 !important; }
details summary { font-family:'Inter',sans-serif !important; font-size:10px !important; font-weight:700 !important; letter-spacing:2px !important; text-transform:uppercase !important; color:var(--nw-blue) !important; padding:13px 20px !important; }
details[open] summary { border-bottom:1px solid var(--border); }

/* ── readiness card ── */
.rdy-card  { border:1px solid var(--border); border-radius:var(--radius); background:var(--surface); padding:8px 16px 4px; margin-bottom:22px; }
.rdy-row   { display:flex; align-items:flex-start; gap:11px; padding:11px 0; border-bottom:1px solid var(--border); }
.rdy-row:last-child { border-bottom:none; padding-bottom:6px; }
.rdy-dot   { width:22px; height:22px; border-radius:50%; display:flex; align-items:center; justify-content:center; font-size:11px; font-weight:700; flex-shrink:0; margin-top:1px; }
.dot-ok    { background:var(--ok-bg); color:var(--ok-fg); }
.dot-wait  { background:var(--nw-lt); color:var(--nw-blue); }
.rdy-title { font-size:12px; font-weight:600; color:var(--text); line-height:1.4; }
.rdy-sub   { font-size:10px; color:var(--muted); margin-top:2px; }

/* ── run button ── */
div.btn-ready > div > button { background:linear-gradient(135deg,var(--nw-deep) 0%,var(--nw-blue) 100%) !important; color:#fff !important; border:none !important; border-radius:9px !important; font-weight:700 !important; font-size:13px !important; letter-spacing:0.5px !important; padding:13px 28px !important; width:100% !important; box-shadow:0 4px 18px rgba(13,63,122,0.28) !important; transition:all 0.18s !important; }
div.btn-ready > div > button:hover { background:linear-gradient(135deg,#0a2f5e 0%,var(--nw-deep) 100%) !important; box-shadow:0 7px 26px rgba(13,63,122,0.36) !important; transform:translateY(-1px) !important; }
div.btn-wait  > div > button { background:var(--border) !important; color:var(--muted) !important; border:none !important; border-radius:9px !important; font-weight:600 !important; font-size:13px !important; padding:13px 28px !important; width:100% !important; box-shadow:none !important; }

/* ── secondary button ── */
.stButton > button[kind="secondary"] { background:transparent !important; color:var(--nw-blue) !important; border:1.5px solid var(--border) !important; border-radius:7px !important; font-size:11px !important; font-weight:500 !important; }
.stButton > button[kind="secondary"]:hover { border-color:var(--nw-blue) !important; background:var(--nw-lt) !important; }

[data-testid="stAlert"] { border-radius:8px !important; font-size:12px !important; }
[data-testid="column"]  { padding:0 6px !important; }

/* ── inline warning box ── */
.warn-box {
  border:1.5px solid #EDD97A; border-radius:10px; background:#FFFDF5;
  padding:16px 20px; margin-top:12px;
  box-shadow:0 2px 10px rgba(200,169,81,0.10);
}
.warn-box .wb-head { display:flex; align-items:center; gap:8px; margin-bottom:8px; }
.warn-box .wb-icon { font-size:18px; }
.warn-box .wb-title { font-size:13px; font-weight:700; color:#0C1A35; }
.warn-box .wb-body { font-size:11px; color:#6B7A9E; line-height:1.6; margin:0 0 4px; }

/* ── inline spinner ── */
.inline-loader { display:flex; align-items:center; gap:14px; padding:18px 20px; margin-top:12px;
  border:1.5px solid var(--border); border-radius:10px; background:var(--surface); }
.spin-ring {
  width:28px; height:28px; border:3px solid rgba(26,93,171,0.15);
  border-top:3px solid #1A5DAB; border-radius:50%; flex-shrink:0;
  animation: spin 0.8s linear infinite;
}
@keyframes spin { to { transform: rotate(360deg); } }
.loader-label { font-family:'Inter',sans-serif; font-size:12px; font-weight:600; color:var(--nw-blue); }
.loader-sub   { font-family:'Inter',sans-serif; font-size:10px; color:var(--muted); margin-top:2px; }

/* ── coming soon ── */
.coming-soon { display:flex; flex-direction:column; align-items:center; justify-content:center; padding:80px 32px; background:var(--surface); border:1px solid var(--border); border-radius:var(--radius); text-align:center; margin-top:24px; }
.coming-soon .cs-icon  { font-size:52px; margin-bottom:16px; }
.coming-soon .cs-title { font-family:'Libre Baskerville',serif; font-size:22px; font-weight:700; color:var(--nw-deep); margin-bottom:8px; }
.coming-soon .cs-sub   { font-size:13px; color:var(--muted); max-width:380px; line-height:1.6; }
.coming-soon .cs-tag   { margin-top:20px; display:inline-block; background:var(--nw-lt); color:var(--nw-blue); border:1px solid var(--border); border-radius:20px; font-size:10px; font-weight:700; letter-spacing:1.5px; text-transform:uppercase; padding:5px 14px; }

/* ── inline radio (sched mode toggle) ── */
[data-testid="stRadio"][aria-label="sched_mode_pick"] > div,
div:has(> [data-testid="stRadio"] > div > [data-baseweb="radio"]:first-child) {
  gap:0 !important;
}
/* Per-state schedule rating grid container */
.sched-state-grid { border:1px solid var(--border); border-top:none; border-radius:0 0 8px 8px; overflow:hidden; }

/* ── multi-state progress card ── */
.multi-progress-card { display:flex; align-items:center; gap:14px; padding:14px 18px; border:1.5px solid var(--border); border-radius:10px; background:var(--surface); margin-top:10px; }
.mp-label { font-size:9px; font-weight:700; letter-spacing:1.8px; text-transform:uppercase; color:var(--muted); flex-shrink:0; }
.mp-state { font-size:15px; font-weight:700; color:var(--nw-deep); flex:1; }
.mp-count { font-size:11px; color:var(--muted); flex-shrink:0; }

/* ══════════════════════════════════════════════════════════════════
   SIDEBAR SERVICES NAV BUTTON  (Tracked Pages)
══════════════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] .svc-btn-wrap > div > button {
  display: flex !important; align-items: center !important; gap: 10px !important;
  width: 100% !important; padding: 13px 20px !important; border-radius: 0 !important;
  font-size: 13px !important; font-family: 'Inter', sans-serif !important;
  font-weight: 500 !important; color: #566278 !important;
  background: transparent !important; border: none !important;
  border-left: 3px solid transparent !important; box-shadow: none !important;
  text-align: left !important; justify-content: flex-start !important;
  transition: background 0.14s ease, color 0.14s ease !important;
}
[data-testid="stSidebar"] .svc-btn-wrap > div > button:hover {
  background: #EBF2FB !important; color: #0D3F7A !important;
  border-left: 3px solid #A8C4E8 !important;
}
[data-testid="stSidebar"] .svc-active > div > button {
  background: linear-gradient(90deg,#1A5DAB 0%,#0D3F7A 100%) !important;
  color: #FFFFFF !important; font-weight: 700 !important;
  border-left: 3px solid #EDD97A !important;
}
[data-testid="stSidebar"] .svc-active > div > button:hover {
  background: linear-gradient(90deg,#1A5DAB 0%,#0D3F7A 100%) !important;
  color: #FFFFFF !important;
}

/* ══════════════════════════════════════════════════════════════════
   TRACKED PAGES COMPARATOR  — styles use app.py design tokens
══════════════════════════════════════════════════════════════════ */

/* ── Service badge ── */
.svc-badge {
  display:inline-flex; align-items:center; gap:6px;
  background:var(--nw-lt); color:var(--nw-blue);
  border:1px solid var(--border); border-radius:20px;
  font-size:9px; font-weight:700; letter-spacing:1.5px; text-transform:uppercase;
  padding:4px 12px; margin-top:6px;
}

/* ── Metric cards grid ── */
.cmp-metric-grid { display:flex; gap:12px; flex-wrap:wrap; margin:14px 0; }
.cmp-metric-card {
  flex:1 1 110px; background:var(--surface); border-radius:var(--radius);
  padding:14px 12px 12px; border:1px solid var(--border);
  text-align:center; box-shadow:var(--shadow);
}
.cmp-metric-val   { font-family:'Libre Baskerville',serif; font-size:1.8rem; font-weight:700; line-height:1.1; }
.cmp-metric-label { font-size:10px; color:var(--muted); text-transform:uppercase; letter-spacing:0.8px; margin-top:4px; }

/* ── Sheet pill tags ── */
.cmp-pills {
  display:flex; flex-wrap:wrap; gap:8px;
  padding:12px 14px; background:var(--off);
  border:1px solid var(--border); border-radius:var(--radius); margin:10px 0 14px;
}
.cmp-pill {
  padding:4px 14px; border-radius:20px; font-size:12px;
  font-weight:600; border:1.5px solid; cursor:default;
  display:inline-flex; align-items:center; gap:5px; font-family:'Inter',sans-serif;
}
.cmp-pill-unchanged { background:#EBF2FB; color:#0D3F7A; border-color:#A8C4E8; }
.cmp-pill-new       { background:#D1FAE5; color:#065f46; border-color:#10b981; }
.cmp-pill-deleted   { background:#fee2e2; color:#991b1b; border-color:#ef4444; }
.cmp-pill-modified  { background:#FFF8E7; color:#78350f; border-color:var(--gold); }

/* ── Legend ── */
.cmp-legend { display:flex; flex-wrap:wrap; gap:16px; margin:8px 0 12px; }
.cmp-legend-item { display:flex; align-items:center; gap:7px; font-size:12px; color:var(--text); }
.cmp-legend-dot  { width:14px; height:14px; border-radius:4px; flex-shrink:0; }

/* ── Diff table ── */
.cmp-diff-wrap {
  overflow-x:auto; border-radius:var(--radius);
  border:1px solid var(--border); box-shadow:var(--shadow);
  max-height:500px; overflow-y:auto;
}
table.cmp-diff { border-collapse:collapse; width:100%; font-size:12px; white-space:nowrap; font-family:'Inter',sans-serif; }
table.cmp-diff thead th {
  background:var(--nw-deep); color:white;
  padding:8px 12px; font-weight:600; text-align:left;
  position:sticky; top:0; z-index:2;
  border-right:1px solid rgba(255,255,255,0.1);
}
table.cmp-diff thead th:first-child { width:40px; text-align:center; }
table.cmp-diff tbody td {
  padding:6px 12px; border-bottom:1px solid #f0f2f5; border-right:1px solid #f0f2f5;
  vertical-align:top; max-width:240px; overflow:hidden; text-overflow:ellipsis;
}
table.cmp-diff tbody tr:hover td { filter:brightness(0.97); }
.cmp-r-added   td { background:#ecfdf5 !important; }
.cmp-r-deleted td { background:#fef2f2 !important; }
.cmp-c-changed { background:#FFFBEB !important; }
.cmp-c-added   { background:#ecfdf5 !important; }
.cmp-c-deleted { background:#fef2f2 !important; }
.cmp-rn { color:#9ca3af; font-size:10px; text-align:center; user-select:none; }
.cmp-val-old  { text-decoration:line-through; color:#dc2626; font-size:10px; display:block; line-height:1.3; }
.cmp-val-new  { color:#16a34a; font-size:11px; display:block; font-weight:600; line-height:1.4; }
.cmp-val-only { font-size:12px; }

/* ── Info / instruction box  — uses app.py gold accent ── */
.cmp-info-box {
  background:var(--nw-lt); border-left:4px solid var(--nw-blue);
  padding:14px 16px; border-radius:0 var(--radius) var(--radius) 0; margin:8px 0;
  font-size:13px; line-height:1.7; color:var(--text);
}
.cmp-info-box ol, .cmp-info-box ul { margin:6px 0 0; padding-left:20px; }
.cmp-info-box strong { color:var(--nw-deep); }

/* ── upload label inside comparator ── */
.cmp-upload-label { font-size:11px; font-weight:700; letter-spacing:1.5px; text-transform:uppercase; color:var(--nw-blue); margin-bottom:6px; display:block; }

/* ── Truncation note ── */
.cmp-truncation-note { text-align:center; padding:8px 12px; color:var(--muted); font-style:italic; font-size:12px; background:var(--off); }
</style>
""", unsafe_allow_html=True)


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:

    st.image("Nationwide-logo.png",  width=200)

    # # Logo
    st.markdown("""
    <div style="padding:16px 20px 6px; font-size:8px; font-weight:700;
                letter-spacing:2.8px; text-transform:uppercase; color:#9BAABF;">
      Line of Business
    </div>
    """, unsafe_allow_html=True)

    # ── LOB navigation ──
    # st.radio is the correct widget: it has persistent checked state that the
    # browser maintains natively via aria-checked on the label element.
    # CSS selector  label[aria-checked="true"]  always works — no class names,
    # no JS, no Streamlit internals.  This is the only reliable approach.
    current_idx = LOB_NAMES.index(st.session_state.lob) if st.session_state.lob in LOB_NAMES else 0

    # When in comparator mode, deselect the radio (index=None) so that clicking
    # the already-active LOB still registers as a new selection and triggers a rerun.
    _radio_idx = None if st.session_state.active_tool == "comparator" else current_idx

    chosen = st.radio(
        "lob_selector",
        options=LOB_OPTIONS,
        index=_radio_idx,
        key="lob_radio",
        label_visibility="collapsed",
    )

    # Derive the plain LOB name from the chosen option label
    if chosen is not None:
        chosen_lob = LOB_NAMES[LOB_OPTIONS.index(chosen)]
        if chosen_lob != st.session_state.lob or st.session_state.active_tool == "comparator":
            st.session_state.lob         = chosen_lob
            st.session_state.run_status  = "idle"
            st.session_state.active_tool = "rate_pages"
            st.rerun()

    # JS: style active label AND fix all parent wrapper widths/margins
    _active = st.session_state.lob
    st.markdown(f"""
    <script>
    (function highlight() {{
      var active = {repr(_active)};
      var sidebar = window.parent.document.querySelector('[data-testid="stSidebar"]');
      if (!sidebar) return;

      // Zero out every wrapper div inside the radio group so nothing clips the label
      var radio = sidebar.querySelector('[data-testid="stRadio"]');
      if (radio) {{
        radio.querySelectorAll('div').forEach(function(d) {{
          d.style.setProperty('padding', '0', 'important');
          d.style.setProperty('margin', '0', 'important');
          d.style.setProperty('width', '100%', 'important');
          d.style.setProperty('max-width', '100%', 'important');
          d.style.setProperty('box-sizing', 'border-box', 'important');
        }});
      }}

      var labels = sidebar.querySelectorAll('[data-testid="stRadio"] label');
      labels.forEach(function(lbl) {{
        // Also fix the label's own parent chain
        var el = lbl.parentElement;
        while (el && el !== radio) {{
          el.style.setProperty('padding', '0', 'important');
          el.style.setProperty('margin', '0', 'important');
          el.style.setProperty('width', '100%', 'important');
          el = el.parentElement;
        }}

        var txt = lbl.innerText || lbl.textContent || '';
        if (txt.indexOf(active) !== -1) {{
          lbl.style.setProperty('background', 'linear-gradient(90deg,#1A5DAB 0%,#0D3F7A 100%)', 'important');
          lbl.style.setProperty('color', '#FFFFFF', 'important');
          lbl.style.setProperty('font-weight', '700', 'important');
          lbl.style.setProperty('border-left', '3px solid #EDD97A', 'important');
          lbl.style.setProperty('width', '100%', 'important');
          lbl.style.setProperty('min-width', '100%', 'important');
          lbl.style.setProperty('margin', '0', 'important');
          lbl.style.setProperty('border-radius', '0', 'important');
          lbl.style.setProperty('box-sizing', 'border-box', 'important');
          lbl.style.setProperty('padding', '13px 20px', 'important');
        }} else {{
          lbl.style.removeProperty('background');
          lbl.style.removeProperty('color');
          lbl.style.removeProperty('font-weight');
          lbl.style.removeProperty('border-left');
        }}
      }});
    }})();
    setTimeout(highlight, 150);
    setTimeout(highlight, 400);

    function highlight() {{
      var active = {repr(_active)};
      var sidebar = window.parent.document.querySelector('[data-testid="stSidebar"]');
      if (!sidebar) return;
      var radio = sidebar.querySelector('[data-testid="stRadio"]');
      if (radio) {{
        radio.querySelectorAll('div').forEach(function(d) {{
          d.style.setProperty('padding', '0', 'important');
          d.style.setProperty('margin', '0', 'important');
          d.style.setProperty('width', '100%', 'important');
          d.style.setProperty('max-width', '100%', 'important');
          d.style.setProperty('box-sizing', 'border-box', 'important');
        }});
      }}
      var labels = sidebar ? sidebar.querySelectorAll('[data-testid="stRadio"] label') : [];
      labels.forEach(function(lbl) {{
        var el = lbl.parentElement;
        while (el && el !== radio) {{
          el.style.setProperty('padding', '0', 'important');
          el.style.setProperty('margin', '0', 'important');
          el.style.setProperty('width', '100%', 'important');
          el = el.parentElement;
        }}
        var txt = lbl.innerText || lbl.textContent || '';
        if (txt.indexOf(active) !== -1) {{
          lbl.style.setProperty('background', 'linear-gradient(90deg,#1A5DAB 0%,#0D3F7A 100%)', 'important');
          lbl.style.setProperty('color', '#FFFFFF', 'important');
          lbl.style.setProperty('font-weight', '700', 'important');
          lbl.style.setProperty('border-left', '3px solid #EDD97A', 'important');
          lbl.style.setProperty('width', '100%', 'important');
          lbl.style.setProperty('min-width', '100%', 'important');
          lbl.style.setProperty('margin', '0', 'important');
          lbl.style.setProperty('border-radius', '0', 'important');
          lbl.style.setProperty('box-sizing', 'border-box', 'important');
          lbl.style.setProperty('padding', '13px 20px', 'important');
        }} else {{
          lbl.style.removeProperty('background');
          lbl.style.removeProperty('color');
          lbl.style.removeProperty('font-weight');
          lbl.style.removeProperty('border-left');
        }}
      }});
    }}
    </script>
    """, unsafe_allow_html=True)

    # ── Analytics Services section ────────────────────────────────────────────
    st.markdown("""
    <div style="margin:20px 20px 0; border-top:1px solid #DDE5F0; padding-top:12px;">
      <div style="font-size:8px;font-weight:700;letter-spacing:2.8px;text-transform:uppercase;color:#9BAABF;margin-bottom:4px;">
        Analytics Services
      </div>
    </div>
    """, unsafe_allow_html=True)

    _is_cmp = st.session_state.active_tool == "comparator"
    st.markdown(f'<div class="svc-btn-wrap {"svc-active" if _is_cmp else "svc-inactive"}">', unsafe_allow_html=True)
    if st.button("📊  Tracked Pages", key="btn_tracked_pages", use_container_width=True):
        st.session_state.active_tool = "comparator"
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)


# ─── HEADER ───────────────────────────────────────────────────────────────────
active_lob  = st.session_state.lob
active_tool = st.session_state.active_tool

LOB_ICONS = {"Business Auto":"🚗","General Liability":"⚖️","Farm Auto":"🚜","Property":"🏠"}
LOB_SUBS  = {
    "Business Auto":     "Upload proposed ratebooks &nbsp;&middot;&nbsp; Configure options &nbsp;&middot;&nbsp; Generate output",
    "General Liability": "General Liability rate page configuration",
    "Farm Auto":         "Farm Auto rate page configuration",
    "Property":          "Property rate page configuration",
}

if active_tool == "comparator":
    st.markdown("""
    <div class="nw-header">
      <div class="nw-header-left">
        <div class="nw-eagle">📊</div>
        <div class="nw-brand">Nationwide <span>Insurance</span></div>
        <div class="nw-sep"></div>
        <div class="nw-pgname">📊 &nbsp;Tracked Pages &middot; Rate Change Analysis</div>
      </div>
      <div class="nw-right">BA &nbsp;&middot;&nbsp; Analytics &nbsp;&middot;&nbsp; Internal Tool</div>
    </div>
    <div class="gold-line"></div>
    """, unsafe_allow_html=True)
    h1, _ = st.columns([3, 1])
    with h1:
        st.markdown("""
        <p style="font-family:'Libre Baskerville',serif;font-size:25px;font-weight:700;
                  color:#0D3F7A;margin:0 0 5px;">Compare Rate Pages</p>
        <p style="font-size:13px;color:#6B7A9E;margin:0;">
          Upload current &amp; proposed rate pages &nbsp;&middot;&nbsp; Detect cell-level changes &nbsp;&middot;&nbsp; Generate tracked Excel reports
        </p>
        """, unsafe_allow_html=True)
else:
    st.markdown(f"""
    <div class="nw-header">
      <div class="nw-header-left">
        <div class="nw-eagle">📋</div>
        <div class="nw-brand">Nationwide <span>Insurance</span></div>
        <div class="nw-sep"></div>
        <div class="nw-pgname">{LOB_ICONS[active_lob]} &nbsp;{active_lob} &middot; Rate Page Builder</div>
      </div>
      <div class="nw-right">BA &nbsp;&middot;&nbsp; Analytics &nbsp;&middot;&nbsp; Internal Tool</div>
    </div>
    <div class="gold-line"></div>
    """, unsafe_allow_html=True)

    # Page heading
    h1, h2 = st.columns([3, 1])
    with h1:
        st.markdown(f"""
        <p style="font-family:'Libre Baskerville',serif;font-size:25px;font-weight:700;
                  color:#0D3F7A;margin:0 0 5px;">Build {active_lob} Rate Pages</p>
        <p style="font-size:13px;color:#6B7A9E;margin:0;">{LOB_SUBS[active_lob]}</p>
        """, unsafe_allow_html=True)
    with h2:
        if active_lob == "Business Auto":
            nr = n_req(); tot = len(REQUIRED); pct = int(nr/tot*100)
            fg = "#196B38" if nr == tot else "#1A5DAB"
            st.markdown(f"""
            <div style="text-align:right;padding-top:4px;">
              <div style="font-size:10px;color:#6B7A9E;letter-spacing:1.5px;text-transform:uppercase;margin-bottom:5px;">Uploaded Files</div>
              <div style="font-size:28px;font-weight:700;color:{fg};line-height:1;font-family:'Libre Baskerville',serif;">
                {nr}<span style="font-size:13px;font-weight:400;color:#6B7A9E;">/{tot}</span>
              </div>
              <div style="background:#D4DFEF;border-radius:3px;height:3px;margin-top:8px;overflow:hidden;">
                <div style="width:{pct}%;height:100%;background:linear-gradient(90deg,#0D3F7A,#1A5DAB);border-radius:3px;"></div>
              </div>
            </div>""", unsafe_allow_html=True)

spacer(28)


# ─── BUSINESS AUTO ────────────────────────────────────────────────────────────
if active_tool == "comparator":
    pass  # handled below after all helper functions are defined

elif active_lob == "Business Auto":

    DETECT_ORDER = ["HICNJ", "CCMIC", "NICOF", "NWAG", "NACO", "NAFF", "NGIC", "MM", "CW"]

    def scan_states(src_path):
        states = []
        try:
            for item in sorted(Path(src_path).iterdir()):
                if item.is_dir() and not item.name.startswith(('.', '_')):
                    books = {}
                    for f in item.iterdir():
                        if f.is_file() and f.suffix.lower() in ('.xlsx', '.xlsm', '.xls'):
                            name_up = f.name.upper()
                            matched = next((k for k in DETECT_ORDER if k in name_up), None)
                            if matched:
                                books.setdefault(matched, []).append(f)
                    has_ngic  = "NGIC" in books and len(books["NGIC"]) == 1
                    conflicts = [k for k, v in books.items() if len(v) > 1]
                    states.append({"name": item.name, "path": item, "books": books,
                                   "ready": has_ngic and not conflicts,
                                   "conflicts": conflicts, "has_ngic": has_ngic})
        except Exception:
            pass
        return states

    # ── Mode toggle ────────────────────────────────────────────────────────────
    mode = st.session_state.mode
    tc1, tc2, _ = st.columns([2, 2, 8])
    with tc1:
        if st.button("Individual State", key="btn_ind", use_container_width=True,
                     type="primary" if mode == "individual" else "secondary"):
            if mode != "individual":
                st.session_state.mode = "individual"
                st.session_state.confirm_step = "idle"
                st.rerun()
    with tc2:
        if st.button("Multiple States", key="btn_mul", use_container_width=True,
                     type="primary" if mode == "multiple" else "secondary"):
            if mode != "multiple":
                st.session_state.mode = "multiple"
                st.session_state.multi_step = "idle"
                st.rerun()

    spacer(16)

    # ══════════════════════════════════════════════════════════════════════════
    # INDIVIDUAL STATE MODE
    # ══════════════════════════════════════════════════════════════════════════
    if mode == "individual":
        L, R = st.columns([13, 7], gap="large")

        with L:
            st.markdown('<div class="sec-label">&#128194; &nbsp;Proposed Ratebooks</div>', unsafe_allow_html=True)

            uploaded = st.file_uploader(
            "Select all ratebook files at once — filenames must contain the company code (NGIC, MM, NACO, …)",
            type=["xlsx", "xlsm", "xls"],
            accept_multiple_files=True,
            key=f"multi_up_{st.session_state.upload_reset}",
        )

            # ── Auto-detect & assign ───────────────────────────────────────
            if uploaded:
                grouped = {}
                for f in uploaded:
                    name_up = f.name.upper()
                    matched = next((k for k in DETECT_ORDER if k in name_up), None)
                    grouped.setdefault(matched or "__unknown__", []).append(f)
                for key in ALL_KEYS:
                    files = grouped.get(key, [])
                    if len(files) == 1:
                        st.session_state[f"file_{key}"] = {"name": files[0].name, "bytes": files[0].read()}
                    elif len(files) > 1:
                        st.session_state[f"file_{key}"] = {"error": "multiple", "names": [f.name for f in files]}

            # ── Assignment table ───────────────────────────────────────────
            LABELS = {"NGIC": "Required", "CW": "Optional"}
            rows_html = ""; n_ok = n_err = 0
            for key in ALL_KEYS:
                val = st.session_state.get(f"file_{key}")
                bh = '<span class="ab-req">Required</span>' if LABELS.get(key) == "Required" else ('<span class="ab-opt">Optional</span>' if LABELS.get(key) == "Optional" else "")
                if val is None:
                    rows_html += f'<div class="arow arow-empty"><span class="aco">{key} {bh}</span><span class="afile">Not uploaded</span><span class="astat astat-empty">—</span></div>'
                elif "error" in val:
                    n_err += 1
                    rows_html += f'<div class="arow arow-error"><span class="aco">{key} {bh}</span><span class="afile afile-err">&#9888;&nbsp; Multiple files: {", ".join(val["names"])}</span><span class="astat astat-err">&#10005;</span></div>'
                else:
                    n_ok += 1
                    rows_html += f'<div class="arow arow-ok"><span class="aco">{key} {bh}</span><span class="afile afile-assigned">{val["name"]}</span><span class="astat astat-ok">&#10003;</span></div>'
            summary = f'{n_ok} assigned' + (f' &nbsp;&middot;&nbsp; <span style="color:#C8102E;font-weight:700;">{n_err} conflict{"s" if n_err>1 else ""}</span>' if n_err else '')
            st.markdown(f'<div class="assign-wrap"><div class="assign-hdr"><span>File Assignment</span><span>{summary}</span></div>{rows_html}</div>', unsafe_allow_html=True)

            if any(st.session_state[f"file_{k}"] for k in ALL_KEYS):
                spacer(8)
                _, clr = st.columns([5, 1])
                with clr:
                    if st.button("Clear all", type="secondary"):
                        for k in ALL_KEYS: st.session_state[f"file_{k}"] = None
                        st.session_state.upload_reset += 1
                        st.session_state.run_status = "idle"
                        st.rerun()

        with R:
            st.markdown('<div class="sec-label">&#9881; &nbsp;Configuration</div>', unsafe_allow_html=True)
            st.markdown('<p class="f-label">&#128193; &nbsp;Save Location</p>', unsafe_allow_html=True)
            typed = st.text_input("save_path", value=st.session_state.save_dir, placeholder="Paste path or click Browse", label_visibility="collapsed")
            if typed != st.session_state.save_dir: st.session_state.save_dir = typed
            if st.button("Browse", key="browse_btn"):
                folder = browse_folder()
                if folder: st.session_state.save_dir = folder; st.rerun()
            if st.session_state.save_dir:
                p = st.session_state.save_dir
                st.markdown(f'<p class="f-ok">&#10003; &nbsp;{("…"+p[-38:]) if len(p)>40 else p}</p>', unsafe_allow_html=True)
            else:
                st.markdown('<p class="f-hint">Browse your device or paste the full folder path</p>', unsafe_allow_html=True)

            spacer(6)
            st.markdown('<p class="f-label">&#128202; &nbsp;Schedule Rating Mod</p>', unsafe_allow_html=True)
            nc, pc = st.columns([3, 1])
            with nc:
                tm = st.number_input("mod_num", min_value=0, max_value=100, value=st.session_state.sched_mod, step=1, label_visibility="collapsed")
                if tm != st.session_state.sched_mod: st.session_state.sched_mod = int(tm)
            with pc:
                st.markdown(f'<div style="display:flex;align-items:center;height:42px;padding-left:4px;"><span style="font-size:22px;font-weight:700;color:#1A5DAB;line-height:1;">{st.session_state.sched_mod}<span style="font-size:13px;font-weight:400;color:#6B7A9E;">%</span></span></div>', unsafe_allow_html=True)
            spacer(6)
            sv = st.slider("mod_slider", 0, 100, value=st.session_state.sched_mod, step=1, format="%d%%", label_visibility="collapsed")
            if sv != st.session_state.sched_mod: st.session_state.sched_mod = sv; st.rerun()
            st.markdown('<p class="f-hint">Rule 417 &middot; State Schedule Rating Maximum Modification Threshold</p>', unsafe_allow_html=True)

            spacer(6)
            st.markdown('<div class="sec-label">&#128203; &nbsp;Readiness</div>', unsafe_allow_html=True)
            has_files = any_req(); save_ok = bool(st.session_state.save_dir)
            nr_now = n_req(); sdv = st.session_state.save_dir; mv = st.session_state.sched_mod
            req_sub  = f"All {len(REQUIRED)} ratebooks uploaded" if all_req() else f"{nr_now} of {len(REQUIRED)} ratebooks uploaded"
            save_sub = (("…"+sdv[-36:]) if len(sdv)>38 else sdv) if save_ok else "Not yet selected"

            def rdy(ok, title, sub):
                d = "dot-ok" if ok else "dot-wait"; i = "&#10003;" if ok else "&#9675;"
                return f'<div class="rdy-row"><div class="rdy-dot {d}">{i}</div><div><div class="rdy-title">{title}</div><div class="rdy-sub">{sub}</div></div></div>'

            ngic_uploaded = has_ngic(); ngic_sub = "Uploaded" if ngic_uploaded else "Required — please upload NGIC"
            st.markdown('<div class="rdy-card">'
                + rdy(ngic_uploaded, 'NGIC Ratebook <span style="font-size:10px;color:#C8102E;font-weight:600;">REQUIRED</span>', ngic_sub)
                + rdy(has_files, f'Other Ratebooks &nbsp;<span style="font-size:10px;color:#6B7A9E;font-weight:400;">{nr_now}/{len(REQUIRED)}</span>', req_sub)
                + rdy(save_ok, "Save location", save_sub)
                + rdy(True, f'Schedule Mod &nbsp;<span style="font-size:10px;color:#6B7A9E;font-weight:400;">{mv}%</span>', "Rule 417 threshold")
                + '</div>', unsafe_allow_html=True)

            ngic_ok = has_ngic(); ready = ngic_ok and save_ok

            if st.session_state.confirm_step == "idle":
                if ready:
                    st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                    if st.button("&#129413;  Create Rate Pages", key="run_btn", use_container_width=True):
                        st.session_state.confirm_step = "confirm"; st.session_state.run_status = "idle"; st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                else:
                    missing = (["NGIC ratebook"] if not ngic_ok else []) + (["save location"] if not save_ok else [])
                    st.markdown('<div class="btn-wait">', unsafe_allow_html=True)
                    st.button(f"Waiting \u2014 {', '.join(missing)}", key="run_btn_dis", use_container_width=True, disabled=True)
                    st.markdown('</div>', unsafe_allow_html=True)

            elif st.session_state.confirm_step == "confirm":
                st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                st.button("&#129413;  Create Rate Pages", key="run_btn_cfm", use_container_width=True, disabled=True)
                st.markdown('</div>', unsafe_allow_html=True)
                st.markdown('<div class="warn-box"><div class="wb-head"><span class="wb-icon">⚠️</span><span class="wb-title">Close &amp; save all open Excel files</span></div><p class="wb-body">The builder needs exclusive access to the workbooks. Please save and close any open <code>.xlsx</code> / <code>.xlsm</code> files before proceeding.</p></div>', unsafe_allow_html=True)
                spacer(8)
                bc1, bc2 = st.columns(2)
                with bc1:
                    if st.button("Cancel", key="cancel_btn", use_container_width=True, type="secondary"):
                        st.session_state.confirm_step = "idle"; st.rerun()
                with bc2:
                    if st.button("Proceed", key="proceed_btn", use_container_width=True, type="primary"):
                        st.session_state.confirm_step = "processing"; st.rerun()

            elif st.session_state.confirm_step == "processing":
                st.markdown('<div class="btn-wait">', unsafe_allow_html=True)
                st.button("Processing Excel...", key="run_btn_proc", use_container_width=True, disabled=True)
                st.markdown('</div>', unsafe_allow_html=True)
                loader_ph = st.empty()
                def update_progress(msg):
                    loader_ph.markdown(f'<div class="inline-loader"><div class="spin-ring"></div><div><div class="loader-label">Creating Excel rate pages…</div><div class="loader-sub">{msg}</div></div></div>', unsafe_allow_html=True)
                update_progress("Please wait while the workbooks are processed.")
                from BARatePages import run as run_rate_pages
                try:
                    def _rb(k):
                        f = st.session_state.get(f"file_{k}")
                        return io.BytesIO(f["bytes"]) if f and "error" not in f else None
                    xlsx_out, pdf_out = run_rate_pages(
                        NGICRatebook=_rb("NGIC"), MMRatebook=_rb("MM"), NACORatebook=_rb("NACO"),
                        NAFFRatebook=_rb("NAFF"), NICOFRatebook=_rb("NICOF"), HICNJRatebook=_rb("HICNJ"),
                        CCMICRatebook=_rb("CCMIC"), NWAGRatebook=_rb("NWAG"), CWRatebook=_rb("CW"),
                        folder_selected=st.session_state.save_dir,
                        SchedRatingMod=int(st.session_state.sched_mod) or None,
                        progress_callback=update_progress, skip_pdf=True)
                    st.session_state.xlsx_path = xlsx_out; st.session_state.pdf_path = pdf_out
                    st.session_state.run_status = "success"; st.session_state.pdf_status = "idle"
                except Exception as e:
                    import traceback; traceback.print_exc()
                    st.session_state.run_status = "error"; st.session_state.run_msg = str(e)
                st.session_state.confirm_step = "idle"; st.rerun()

            elif st.session_state.confirm_step == "pdf_processing":
                st.markdown('<div class="btn-wait">', unsafe_allow_html=True)
                st.button("Generating PDF...", key="pdf_btn_proc", use_container_width=True, disabled=True)
                st.markdown('</div>', unsafe_allow_html=True)
                loader_ph2 = st.empty()
                def update_pdf_progress(msg):
                    loader_ph2.markdown(f'<div class="inline-loader"><div class="spin-ring"></div><div><div class="loader-label">Converting to PDF…</div><div class="loader-sub">{msg}</div></div></div>', unsafe_allow_html=True)
                from BARatePages import generate_pdf_only
                try:
                    generate_pdf_only(st.session_state.xlsx_path, st.session_state.pdf_path, progress_callback=update_pdf_progress)
                    st.session_state.pdf_status = "success"
                except Exception as e:
                    import traceback; traceback.print_exc()
                    st.session_state.pdf_status = "error"; st.session_state.run_msg = str(e)
                st.session_state.confirm_step = "idle"; st.rerun()

            if st.session_state.run_status == "success":
                spacer(10)
                st.success(f"&#10003;  Excel created: {Path(st.session_state.xlsx_path).name}")
                if st.session_state.pdf_status != "success":
                    st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                    if st.button("Generate PDF Document", key="gen_pdf_btn", use_container_width=True):
                        st.session_state.confirm_step = "pdf_processing"; st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                    if st.session_state.pdf_status == "error":
                        st.error(f"PDF Error: {st.session_state.run_msg}")
                else:
                    st.success(f"&#10003;  PDF created: {Path(st.session_state.pdf_path).name}")
            elif st.session_state.run_status == "error":
                spacer(10); st.error(st.session_state.run_msg)

            spacer(24)
            st.markdown('<div style="padding-top:14px;border-top:1px solid var(--border);"><p style="font-size:10px;color:#8892A4;letter-spacing:0.8px;text-transform:uppercase;text-align:center;margin:0;line-height:1.9;">Nationwide Insurance &nbsp;&middot;&nbsp; BA Analytics Division<br>Internal Use Only</p></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # MULTIPLE STATES MODE
    # ══════════════════════════════════════════════════════════════════════════
    else:
        L, R = st.columns([13, 7], gap="large")

        with L:
            st.markdown('<div class="sec-label">&#128194; &nbsp;Source Folder</div>', unsafe_allow_html=True)
            st.markdown('<p class="f-label">&#128193; &nbsp;Main Ratebooks Folder</p>', unsafe_allow_html=True)
            st.markdown('<p class="f-hint">Select the folder that contains one subfolder per state. Each subfolder must contain that state\'s ratebook files.</p>', unsafe_allow_html=True)
            spacer(4)
            typed_src = st.text_input("multi_src_path", value=st.session_state.multi_src_dir, placeholder="Paste path or click Browse", label_visibility="collapsed", key="multi_src_input")
            if typed_src != st.session_state.multi_src_dir:
                st.session_state.multi_src_dir = typed_src; st.session_state.multi_step = "idle"; st.session_state.multi_results = []; st.session_state.multi_sched_map = {}
            if st.button("Browse", key="browse_src_btn"):
                folder = browse_folder()
                if folder: st.session_state.multi_src_dir = folder; st.session_state.multi_step = "idle"; st.session_state.multi_results = []; st.session_state.multi_sched_map = {}; st.rerun()

            if st.session_state.multi_src_dir:
                src_path = Path(st.session_state.multi_src_dir)
                if src_path.exists() and src_path.is_dir():
                    p = str(src_path)
                    st.markdown(f'<p class="f-ok">&#10003; &nbsp;{("…"+p[-38:]) if len(p)>40 else p}</p>', unsafe_allow_html=True)
                    spacer(14)
                    states = scan_states(src_path)
                    if not states:
                        st.markdown('<p class="f-hint">No state subfolders found in this directory.</p>', unsafe_allow_html=True)
                    else:
                        n_ready  = sum(1 for s in states if s["ready"])
                        n_issues = len(states) - n_ready
                        issue_txt = f' &nbsp;&middot;&nbsp; <span style="color:#C8102E;font-weight:700;">{n_issues} issue{"s" if n_issues>1 else ""}</span>' if n_issues else ''
                        state_rows = ""
                        for s in states:
                            found_str = ", ".join(sorted(s["books"].keys())) or "—"
                            if s["ready"]:
                                state_rows += f'<div class="arow arow-ok" style="grid-template-columns:110px 1fr 28px;"><span class="aco">{s["name"]}</span><span class="afile afile-assigned">{found_str}</span><span class="astat astat-ok">&#10003;</span></div>'
                            elif s["conflicts"]:
                                state_rows += f'<div class="arow arow-error" style="grid-template-columns:110px 1fr 28px;"><span class="aco">{s["name"]}</span><span class="afile afile-err">&#9888;&nbsp; Conflicts: {", ".join(s["conflicts"])}</span><span class="astat astat-err">&#9888;</span></div>'
                            elif not s["has_ngic"]:
                                state_rows += f'<div class="arow arow-error" style="grid-template-columns:110px 1fr 28px;"><span class="aco">{s["name"]}</span><span class="afile afile-err">Missing NGIC &nbsp;({found_str})</span><span class="astat astat-err">&#10005;</span></div>'
                            else:
                                state_rows += f'<div class="arow arow-empty" style="grid-template-columns:110px 1fr 28px;"><span class="aco">{s["name"]}</span><span class="afile">{found_str}</span><span class="astat astat-empty">—</span></div>'
                        st.markdown(f'<div class="assign-wrap"><div class="assign-hdr"><span>States Detected &nbsp;({len(states)})</span><span>{n_ready} ready{issue_txt}</span></div>{state_rows}</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<p style="font-size:11px;color:#C8102E;margin:4px 0;">Path not found or not a directory.</p>', unsafe_allow_html=True)

        with R:
            st.markdown('<div class="sec-label">&#9881; &nbsp;Configuration</div>', unsafe_allow_html=True)
            st.markdown('<p class="f-label">&#128193; &nbsp;Save Location</p>', unsafe_allow_html=True)
            st.markdown('<p class="f-hint">Each state\'s Excel will be saved in <code>Save Location / State /</code></p>', unsafe_allow_html=True)
            spacer(4)
            typed_save = st.text_input("multi_save_path", value=st.session_state.multi_save_dir, placeholder="Paste path or click Browse", label_visibility="collapsed", key="multi_save_input")
            if typed_save != st.session_state.multi_save_dir: st.session_state.multi_save_dir = typed_save
            if st.button("Browse", key="browse_save_btn"):
                folder = browse_folder()
                if folder: st.session_state.multi_save_dir = folder; st.rerun()
            if st.session_state.multi_save_dir:
                p = st.session_state.multi_save_dir
                st.markdown(f'<p class="f-ok">&#10003; &nbsp;{("…"+p[-38:]) if len(p)>40 else p}</p>', unsafe_allow_html=True)
            else:
                st.markdown('<p class="f-hint">Choose where all output files will be saved.</p>', unsafe_allow_html=True)

            spacer(6)
            st.markdown('<p class="f-label">&#128202; &nbsp;Schedule Rating Mod &mdash; Per State</p>', unsafe_allow_html=True)
            st.markdown('<p class="f-hint">Rule 417 &middot; Each state can have its own threshold</p>', unsafe_allow_html=True)
            spacer(4)

            # ── Mode toggle ─────────────────────────────────────────────────
            _sched_opts = ["Upload Excel", "Enter Manually"]
            _sched_idx  = 0 if st.session_state.multi_sched_mode == "upload" else 1
            _sched_pick = st.radio(
                "sched_mode_pick",
                options=_sched_opts,
                index=_sched_idx,
                horizontal=True,
                label_visibility="collapsed",
                key="multi_sched_mode_radio",
            )
            _new_mode = "upload" if _sched_pick == "Upload Excel" else "manual"
            if _new_mode != st.session_state.multi_sched_mode:
                st.session_state.multi_sched_mode = _new_mode
                st.rerun()

            # ── Detect states early for this section ────────────────────────
            _src_ok_sched = bool(st.session_state.multi_src_dir) and Path(st.session_state.multi_src_dir).is_dir()
            _states_sched = scan_states(st.session_state.multi_src_dir) if _src_ok_sched else []

            # ── Upload Excel branch ─────────────────────────────────────────
            if st.session_state.multi_sched_mode == "upload":
                spacer(4)
                st.markdown(
                    '<p class="f-hint"><b>Col A</b> = State name &nbsp;&middot;&nbsp; '
                    '<b>Col B</b> = Schedule Rating % &nbsp;(no header row required).<br>'
                    'Only rows whose state name matches a detected folder are applied.</p>',
                    unsafe_allow_html=True,
                )
                _sched_file = st.file_uploader(
                    "Upload schedule rating Excel",
                    type=["xlsx", "xls"],
                    key="multi_sched_excel_uploader",
                    label_visibility="collapsed",
                )
                if _sched_file is not None:
                    import pandas as _pd
                    try:
                        _df = _pd.read_excel(io.BytesIO(_sched_file.read()), header=None)
                        _new_excel = {}
                        for _, _row in _df.iterrows():
                            try:
                                _sn = str(_row.iloc[0]).strip()
                                _sv = int(float(str(_row.iloc[1]).strip()))
                                if _sn and _sn.lower() not in ("nan", "none", ""):
                                    _new_excel[_sn] = max(0, min(100, _sv))
                            except Exception:
                                pass
                        if _new_excel != st.session_state.multi_sched_excel:
                            st.session_state.multi_sched_excel = _new_excel
                            # Pre-populate sched_map for matched detected states
                            for _s in _states_sched:
                                _nm = _s["name"]
                                if _nm in _new_excel:
                                    st.session_state.multi_sched_map[_nm] = _new_excel[_nm]
                            st.rerun()
                    except Exception as _exc:
                        st.markdown(f'<p style="font-size:11px;color:#C8102E;margin:4px 0;">&#9888; Could not parse file: {_exc}</p>', unsafe_allow_html=True)

            # ── Per-state input grid ─────────────────────────────────────────
            if _states_sched:
                spacer(6)
                _n_set = sum(
                    1 for _s in _states_sched
                    if st.session_state.multi_sched_map.get(_s["name"], 0) > 0
                )
                # Apply any excel defaults not yet in sched_map
                if st.session_state.multi_sched_excel and st.session_state.multi_sched_mode == "upload":
                    for _s in _states_sched:
                        _nm = _s["name"]
                        if _nm not in st.session_state.multi_sched_map and _nm in st.session_state.multi_sched_excel:
                            st.session_state.multi_sched_map[_nm] = st.session_state.multi_sched_excel[_nm]

                st.markdown(
                    f'<div class="assign-hdr" style="border:1px solid var(--border);border-radius:8px 8px 0 0;'
                    f'padding:9px 14px;margin-bottom:0;">'
                    f'<span>Per-State Schedule Rating</span>'
                    f'<span style="color:{"var(--ok-fg)" if _n_set==len(_states_sched) else "var(--muted)"};">'
                    f'{_n_set}&thinsp;/&thinsp;{len(_states_sched)} configured</span></div>',
                    unsafe_allow_html=True,
                )
                st.markdown('<div style="border:1px solid var(--border);border-top:none;border-radius:0 0 8px 8px;overflow:hidden;">', unsafe_allow_html=True)
                for _i, _s in enumerate(_states_sched):
                    _nm  = _s["name"]
                    _exc = st.session_state.multi_sched_excel.get(_nm)
                    _cur = int(st.session_state.multi_sched_map.get(_nm, _exc if _exc is not None else 0))

                    _bg   = "#F9FBF9" if _i % 2 == 0 else "#FFFFFF"
                    _src_badge = ""
                    if _exc is not None and st.session_state.multi_sched_mode == "upload":
                        _src_badge = ' &nbsp;<span style="font-size:8px;font-weight:700;color:var(--ok-fg);letter-spacing:0.5px;text-transform:uppercase;">excel</span>'

                    _lc, _ic, _pc = st.columns([5, 3, 1])
                    with _lc:
                        st.markdown(
                            f'<div style="display:flex;align-items:center;height:38px;padding:0 8px;'
                            f'font-size:12px;font-weight:600;color:var(--text);background:{_bg};">'
                            f'{_nm}{_src_badge}</div>',
                            unsafe_allow_html=True,
                        )
                    with _ic:
                        _nv = st.number_input(
                            f"sr_{_nm}",
                            min_value=0, max_value=100,
                            value=_cur,
                            step=1,
                            key=f"msm_{_nm}",
                            label_visibility="collapsed",
                        )
                        if int(_nv) != st.session_state.multi_sched_map.get(_nm):
                            st.session_state.multi_sched_map[_nm] = int(_nv)
                    with _pc:
                        st.markdown(
                            f'<div style="display:flex;align-items:center;height:38px;font-size:12px;'
                            f'color:var(--muted);">%</div>',
                            unsafe_allow_html=True,
                        )
                st.markdown('</div>', unsafe_allow_html=True)
            elif not _src_ok_sched:
                st.markdown('<p class="f-hint">Select a source folder above to configure per-state values.</p>', unsafe_allow_html=True)
            else:
                st.markdown('<p class="f-hint">No state folders detected yet.</p>', unsafe_allow_html=True)

            spacer(6)
            gen_pdf = st.checkbox("Generate PDF for each state", value=st.session_state.multi_gen_pdf, key="multi_pdf_chk")
            if gen_pdf != st.session_state.multi_gen_pdf: st.session_state.multi_gen_pdf = gen_pdf
            if gen_pdf:
                st.markdown('<p class="f-hint">PDFs saved to <code>Save Location / PDF /</code></p>', unsafe_allow_html=True)

            spacer(6)
            st.markdown('<div class="sec-label">&#128203; &nbsp;Readiness</div>', unsafe_allow_html=True)
            src_ok      = bool(st.session_state.multi_src_dir) and Path(st.session_state.multi_src_dir).is_dir()
            save_ok_m   = bool(st.session_state.multi_save_dir)
            states_list = scan_states(st.session_state.multi_src_dir) if src_ok else []
            n_ready_m   = sum(1 for s in states_list if s["ready"])
            ready_m     = src_ok and save_ok_m and n_ready_m > 0

            def rdy(ok, title, sub):
                d = "dot-ok" if ok else "dot-wait"; i = "&#10003;" if ok else "&#9675;"
                return f'<div class="rdy-row"><div class="rdy-dot {d}">{i}</div><div><div class="rdy-title">{title}</div><div class="rdy-sub">{sub}</div></div></div>'

            src_sub  = (f"{len(states_list)} folders, {n_ready_m} ready" if states_list else "No states found") if src_ok else "Not selected"
            save_sub = (("…"+st.session_state.multi_save_dir[-36:]) if len(st.session_state.multi_save_dir)>38 else st.session_state.multi_save_dir) if save_ok_m else "Not yet selected"
            st.markdown('<div class="rdy-card">'
                + rdy(src_ok and bool(states_list), "Source Folder", src_sub)
                + rdy(n_ready_m > 0, f'States Ready &nbsp;<span style="font-size:10px;color:#6B7A9E;">{n_ready_m}/{len(states_list)}</span>', f"{n_ready_m} state{'s' if n_ready_m!=1 else ''} with valid NGIC ratebook")
                + rdy(save_ok_m, "Save Location", save_sub)
                + rdy(True, f'Schedule Mod &nbsp;<span style="font-size:10px;color:#6B7A9E;">Per State</span>', f'{sum(1 for s in states_list if st.session_state.multi_sched_map.get(s["name"],0)>0)} of {len(states_list)} states configured' if states_list else "No states detected")
                + '</div>', unsafe_allow_html=True)

            if st.session_state.multi_step == "idle":
                if ready_m:
                    st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                    if st.button(f"&#129413;  Create Rate Pages for {n_ready_m} State{'s' if n_ready_m!=1 else ''}", key="multi_run_btn", use_container_width=True):
                        st.session_state.multi_step = "confirm"; st.rerun()
                    st.markdown('</div>', unsafe_allow_html=True)
                else:
                    missing_m = (["source folder"] if not src_ok else (["no ready states"] if not n_ready_m else [])) + (["save location"] if not save_ok_m else [])
                    st.markdown('<div class="btn-wait">', unsafe_allow_html=True)
                    st.button(f"Waiting \u2014 {', '.join(missing_m)}", key="multi_run_dis", use_container_width=True, disabled=True)
                    st.markdown('</div>', unsafe_allow_html=True)

            elif st.session_state.multi_step == "confirm":
                st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                st.button(f"&#129413;  Create Rate Pages for {n_ready_m} State{'s' if n_ready_m!=1 else ''}", key="multi_run_cfm", use_container_width=True, disabled=True)
                st.markdown('</div>', unsafe_allow_html=True)
                st.markdown('<div class="warn-box"><div class="wb-head"><span class="wb-icon">⚠️</span><span class="wb-title">Close &amp; save all open Excel files</span></div><p class="wb-body">The builder needs exclusive access to all workbooks. Please save and close any open <code>.xlsx</code> / <code>.xlsm</code> files before proceeding.</p></div>', unsafe_allow_html=True)
                spacer(8)
                mc1, mc2 = st.columns(2)
                with mc1:
                    if st.button("Cancel", key="multi_cancel_btn", use_container_width=True, type="secondary"):
                        st.session_state.multi_step = "idle"; st.rerun()
                with mc2:
                    if st.button("Proceed", key="multi_proceed_btn", use_container_width=True, type="primary"):
                        st.session_state.multi_step = "processing"; st.rerun()

            elif st.session_state.multi_step == "processing":
                st.markdown('<div class="btn-wait">', unsafe_allow_html=True)
                st.button("Processing states...", key="multi_proc_btn", use_container_width=True, disabled=True)
                st.markdown('</div>', unsafe_allow_html=True)

                from BARatePages import run as run_rate_pages, generate_pdf_only
                src_path_m   = Path(st.session_state.multi_src_dir)
                save_path_m  = Path(st.session_state.multi_save_dir)
                states_to_run = [s for s in scan_states(src_path_m) if s["ready"]]
                pdf_dir = save_path_m / "PDF"
                if st.session_state.multi_gen_pdf: pdf_dir.mkdir(parents=True, exist_ok=True)

                prog_bar    = st.progress(0.0)
                state_ph    = st.empty()
                results = []

                for idx, s in enumerate(states_to_run):
                    sname = s["name"]
                    state_ph.markdown(f'<div class="multi-progress-card"><div class="mp-label">Processing</div><div class="mp-state">{sname}</div><div class="mp-count">{idx+1} of {len(states_to_run)}</div></div>', unsafe_allow_html=True)
                    out_dir = save_path_m / sname
                    out_dir.mkdir(parents=True, exist_ok=True)
                    def _p(k): files = s["books"].get(k); return files[0] if files and len(files)==1 else None
                    try:
                        xlsx_out, _ = run_rate_pages(
                            NGICRatebook=_p("NGIC"), MMRatebook=_p("MM"), NACORatebook=_p("NACO"),
                            NAFFRatebook=_p("NAFF"), NICOFRatebook=_p("NICOF"), HICNJRatebook=_p("HICNJ"),
                            CCMICRatebook=_p("CCMIC"), NWAGRatebook=_p("NWAG"), CWRatebook=None,
                            folder_selected=str(out_dir), SchedRatingMod=int(st.session_state.multi_sched_map.get(sname, 0)) or None, skip_pdf=True)
                        pdf_out = None
                        if st.session_state.multi_gen_pdf:
                            pdf_file = pdf_dir / (Path(xlsx_out).stem + ".pdf")
                            generate_pdf_only(xlsx_out, str(pdf_file))
                            pdf_out = str(pdf_file)
                        results.append({"state": sname, "xlsx": xlsx_out, "pdf": pdf_out, "error": None})
                    except Exception as e:
                        import traceback; traceback.print_exc()
                        results.append({"state": sname, "xlsx": None, "pdf": None, "error": str(e)})
                    prog_bar.progress((idx + 1) / len(states_to_run))

                state_ph.empty()
                st.session_state.multi_results = results
                st.session_state.multi_step = "done"
                st.rerun()

            elif st.session_state.multi_step == "done":
                results   = st.session_state.multi_results
                n_success = sum(1 for r in results if not r["error"])
                n_failed  = len(results) - n_success
                if n_failed == 0:
                    st.success(f"&#10003;  All {n_success} state{'s' if n_success!=1 else ''} completed successfully")
                else:
                    st.warning(f"{n_success} completed &nbsp;&middot;&nbsp; {n_failed} failed")
                result_rows = ""
                for r in results:
                    if r["error"]:
                        result_rows += f'<div class="arow arow-error"><span class="aco">{r["state"]}</span><span class="afile afile-err">{str(r["error"])[:70]}</span><span class="astat astat-err">&#10005;</span></div>'
                    else:
                        pdf_tag = ' &nbsp;<span style="color:var(--ok-fg);font-size:10px;">+ PDF</span>' if r["pdf"] else ''
                        result_rows += f'<div class="arow arow-ok"><span class="aco">{r["state"]}</span><span class="afile afile-assigned">{Path(r["xlsx"]).name}{pdf_tag}</span><span class="astat astat-ok">&#10003;</span></div>'
                st.markdown(f'<div class="assign-wrap"><div class="assign-hdr"><span>Results</span><span>{n_success}/{len(results)} completed</span></div>{result_rows}</div>', unsafe_allow_html=True)
                spacer(8)
                if st.button("Start Over", key="multi_reset_btn", type="secondary"):
                    st.session_state.multi_step = "idle"; st.session_state.multi_results = []; st.rerun()

            spacer(24)
            st.markdown('<div style="padding-top:14px;border-top:1px solid var(--border);"><p style="font-size:10px;color:#8892A4;letter-spacing:0.8px;text-transform:uppercase;text-align:center;margin:0;line-height:1.9;">Nationwide Insurance &nbsp;&middot;&nbsp; BA Analytics Division<br>Internal Use Only</p></div>', unsafe_allow_html=True)


# ─── OTHER LOBs ───────────────────────────────────────────────────────────────
elif active_tool != "comparator":
    CS = {
        "General Liability": ("&#9878;",  "General Liability", 'Wire up your GL backend in the <code>elif active_lob == "General Liability"</code> block.'),
        "Farm Auto":         ("&#128668;", "Farm Auto",         'Wire up your FA backend in the <code>elif active_lob == "Farm Auto"</code> block.'),
        "Property":          ("&#127968;", "Property",          'Wire up your Property backend in the <code>elif active_lob == "Property"</code> block.'),
    }
    ic, ti, de = CS[active_lob]
    st.markdown(f"""
    <div class="coming-soon">
      <div class="cs-icon">{ic}</div>
      <div class="cs-title">{ti} Rate Pages</div>
      <div class="cs-sub">{de}</div>
      <div class="cs-tag">Coming Soon</div>
    </div>""", unsafe_allow_html=True)


# ─── TRACKED PAGES COMPARATOR — helper functions ──────────────────────────────

def _cmp_file_bytes(uploaded_file) -> bytes:
    if hasattr(uploaded_file, "seek"):
        uploaded_file.seek(0)
    return uploaded_file.read()

def _cmp_detect_engine(name: str) -> str:
    return "xlrd" if name.lower().endswith(".xls") else "openpyxl"

def _cmp_read_excel_sheets(raw: bytes, filename: str) -> Dict[str, pd.DataFrame]:
    engine = _cmp_detect_engine(filename)
    try:
        buf = io.BytesIO(raw)
        xl  = pd.ExcelFile(buf, engine=engine)
        result: Dict[str, pd.DataFrame] = {}
        for sheet in xl.sheet_names:
            buf.seek(0)
            df = pd.read_excel(buf, sheet_name=sheet, header=None, dtype=str, engine=engine)
            result[sheet] = df.fillna("")
        return result
    except Exception as exc:
        st.error(f"Could not read **{filename}**: {exc}")
        return {}

def _cmp_cell_str(val) -> str:
    s = str(val).strip() if val is not None else ""
    if re.fullmatch(r"-?\d+\.0+", s):
        s = s[: s.index(".")]
    return s

def _cmp_esc(text: str) -> str:
    return str(text).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

def _cmp_compare_dataframes(old_df: pd.DataFrame, new_df: pd.DataFrame):
    nr = max(len(old_df), len(new_df))
    nc = max(
        len(old_df.columns) if len(old_df) else 0,
        len(new_df.columns) if len(new_df) else 0,
    )
    if nc == 0:
        empty = pd.DataFrame()
        return empty, empty, {}, {}, {"total_rows":0,"total_cols":0,"added_rows":0,"deleted_rows":0,"changed_rows":0,"changed_cells":0}

    old_a = old_df.reindex(range(nr)).reindex(columns=range(nc)).fillna("")
    new_a = new_df.reindex(range(nr)).reindex(columns=range(nc)).fillna("")
    old_row_range = set(range(len(old_df))); new_row_range = set(range(len(new_df)))
    cell_status: dict = {}; row_status: dict = {}
    added_rows = deleted_rows = changed_rows = changed_cells = 0
    for i in range(nr):
        in_old = i in old_row_range; in_new = i in new_row_range
        if in_new and not in_old:
            row_status[i] = "added"; added_rows += 1
            for j in range(nc): cell_status[(i,j)] = "added"
        elif in_old and not in_new:
            row_status[i] = "deleted"; deleted_rows += 1
            for j in range(nc): cell_status[(i,j)] = "deleted"
        else:
            row_changed = False
            for j in range(nc):
                ov = _cmp_cell_str(old_a.iat[i,j]); nv = _cmp_cell_str(new_a.iat[i,j])
                if ov != nv:
                    cell_status[(i,j)] = "changed"; row_changed = True; changed_cells += 1
                else:
                    cell_status[(i,j)] = "same"
            if row_changed: row_status[i] = "changed"; changed_rows += 1
            else: row_status[i] = "same"
    return old_a, new_a, cell_status, row_status, {
        "total_rows":nr,"total_cols":nc,"added_rows":added_rows,
        "deleted_rows":deleted_rows,"changed_rows":changed_rows,"changed_cells":changed_cells,
    }

_CMP_DATE_RE = re.compile(r"\b\d{2}-\d{2}-\d{4}\b")

def _cmp_base_name(filename: str) -> str:
    stem    = os.path.splitext(filename)[0]
    cleaned = _CMP_DATE_RE.sub("", stem)
    return re.sub(r"[\s_\-]+"," ", cleaned).strip().lower()

def _cmp_match_pairs(current_files, proposed_files):
    cur_map:  Dict[str, object] = {_cmp_base_name(f.name): f for f in current_files}
    prop_map: Dict[str, object] = {_cmp_base_name(f.name): f for f in proposed_files}
    matched            = [(cur_map[k], prop_map[k]) for k in cur_map if k in prop_map]
    unmatched_current  = [cur_map[k]  for k in cur_map  if k not in prop_map]
    unmatched_proposed = [prop_map[k] for k in prop_map if k not in cur_map]
    return matched, unmatched_current, unmatched_proposed

_MAX_DISP = 1000

def _cmp_render_diff_table(old_a, new_a, cell_status, row_status, max_rows=_MAX_DISP) -> str:
    nr = len(old_a); nc = len(old_a.columns)
    col_headers = [get_column_letter(j+1) for j in range(nc)]
    parts = ['<div class="cmp-diff-wrap"><table class="cmp-diff"><thead><tr>']
    parts.append('<th class="cmp-rn">#</th>')
    for ch in col_headers: parts.append(f"<th>{ch}</th>")
    parts.append("</tr></thead><tbody>")
    show = min(nr, max_rows)
    for i in range(show):
        rs = row_status.get(i,"same")
        row_cls = {"added":"cmp-r-added","deleted":"cmp-r-deleted"}.get(rs,"")
        parts.append(f'<tr class="{row_cls}">')
        parts.append(f'<td class="cmp-rn">{i+1}</td>')
        for j in range(nc):
            cs = cell_status.get((i,j),"same")
            ov = _cmp_esc(_cmp_cell_str(old_a.iat[i,j])); nv = _cmp_esc(_cmp_cell_str(new_a.iat[i,j]))
            if rs == "added":
                cell_cls = "cmp-c-added"; inner = f'<span class="cmp-val-only">{nv}</span>'
            elif rs == "deleted":
                cell_cls = "cmp-c-deleted"; inner = f'<span class="cmp-val-only">{ov}</span>'
            elif cs == "changed":
                cell_cls = "cmp-c-changed"
                if ov and nv: inner = f'<span class="cmp-val-old">{ov}</span><span class="cmp-val-new">{nv}</span>'
                else: inner = f'<span class="cmp-val-only">{nv or ov}</span>'
            else:
                cell_cls = ""; inner = f'<span class="cmp-val-only">{nv}</span>'
            td_cls = f' class="{cell_cls}"' if cell_cls else ""
            parts.append(f"<td{td_cls}>{inner}</td>")
        parts.append("</tr>")
    if nr > max_rows:
        parts.append(f'<tr><td colspan="{nc+1}" class="cmp-truncation-note">&#9888; Showing first {max_rows:,} of {nr:,} rows. Download the Excel report to view all rows.</td></tr>')
    parts.append("</tbody></table></div>")
    return "".join(parts)

# ── Excel export fills / fonts / borders ──────────────────────────────────────
_CMP_FILL_CHG     = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_CMP_FILL_ADDED   = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
_CMP_FILL_DELETED = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_CMP_FILL_HDR     = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
_CMP_FILL_SUM_HDR = PatternFill(start_color="1A4A7A", end_color="1A4A7A", fill_type="solid")
_CMP_FONT_HDR     = Font(color="FFFFFF", bold=True, name="Segoe UI", size=10)
_CMP_FONT_NORM    = Font(name="Segoe UI", size=10)
_CMP_FONT_BOLD    = Font(name="Segoe UI", size=10, bold=True)
_CMP_THIN_BDR = Border(
    left=Side(style="thin",color="D1D5DB"), right=Side(style="thin",color="D1D5DB"),
    top=Side(style="thin",color="D1D5DB"),  bottom=Side(style="thin",color="D1D5DB"),
)
_CMP_TAB_COLOR = {"new":"10B981","deleted":"EF4444","modified":"F59E0B","unchanged":"6B7280"}
_CMP_FILL_SBS_CHG_NEW  = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_CMP_FILL_SBS_NEW_SHT  = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
_CMP_FILL_SBS_DEL_SHT  = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
_CMP_FILL_SBS_SEP      = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_CMP_FILL_SBS_HDR_OLD  = PatternFill(start_color="0F2942", end_color="0F2942", fill_type="solid")
_CMP_FILL_SBS_HDR_NEW  = PatternFill(start_color="1E6091", end_color="1E6091", fill_type="solid")
_CMP_FILL_SBS_HDR_SEP  = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
_CMP_FONT_STRIKE       = Font(name="Segoe UI", size=10, strike=True, color="C00000")
_CMP_FONT_SBS_HDR      = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
_CMP_BDR_SEP = Border(
    left=Side(style="medium",color="595959"), right=Side(style="medium",color="595959"),
    top=Side(style="thin",color="D0D0D0"),    bottom=Side(style="thin",color="D0D0D0"),
)

def _cmp_auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0; col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try: max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except: pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

def _cmp_apply_border(ws, max_row, max_col):
    for r in range(1, max_row+1):
        for c in range(1, max_col+1):
            ws.cell(r,c).border = _CMP_THIN_BDR

def _cmp_load_src_wb(raw, filename):
    try:
        if filename.lower().endswith(".xls"): return None
        return openpyxl.load_workbook(io.BytesIO(raw))
    except: return None

def _cmp_copy_print_settings(target_ws, src_wb, sheet_name):
    if src_wb is None or sheet_name not in src_wb.sheetnames: return
    try:
        src_ws = src_wb[sheet_name]
        src_ps = src_ws.page_setup; tgt_ps = target_ws.page_setup
        for attr in ('orientation','paperSize','scale','firstPageNumber','pageOrder',
                     'usePrinterDefaults','blackAndWhite','draft','cellComments',
                     'useFirstPageNumber','horizontalDpi','verticalDpi','copies','errors'):
            try:
                val = getattr(src_ps, attr, None)
                if val is not None: setattr(tgt_ps, attr, val)
            except: pass
        src_pm = src_ws.page_margins; tgt_pm = target_ws.page_margins
        for attr in ('left','right','top','bottom','header','footer'):
            try:
                val = getattr(src_pm, attr, None)
                if val is not None: setattr(tgt_pm, attr, val)
            except: pass
        target_ws.HeaderFooter = copy.deepcopy(src_ws.HeaderFooter)
    except: pass

def _cmp_copy_cell_style(src_cell, dst_cell):
    if src_cell is None: return
    for attr, copy_fn in [
        ('font', copy.copy), ('alignment', copy.copy),
        ('border', copy.copy), ('number_format', None),
    ]:
        try:
            val = getattr(src_cell, attr, None)
            if val is None: continue
            if attr == 'number_format':
                if val and val != 'General': dst_cell.number_format = val
            else:
                setattr(dst_cell, attr, copy_fn(val))
        except: pass
    try:
        if src_cell.fill is not None and src_cell.fill.fill_type not in (None, "none"):
            dst_cell.fill = copy.copy(src_cell.fill)
    except: pass

def _cmp_font_with_strike(src_font):
    if src_font is None: return _CMP_FONT_STRIKE
    try:
        return Font(name=src_font.name, size=src_font.size, bold=src_font.bold,
                    italic=src_font.italic, underline=src_font.underline, strike=True,
                    color="C00000", vertAlign=src_font.vertAlign,
                    charset=src_font.charset, scheme=src_font.scheme)
    except: return _CMP_FONT_STRIKE

def _cmp_copy_row_col_dims(src_ws, dst_ws):
    if src_ws is None: return
    try:
        for row_idx, rd in src_ws.row_dimensions.items():
            if rd.height is not None: dst_ws.row_dimensions[row_idx].height = rd.height
        for col_ltr, cd in src_ws.column_dimensions.items():
            if cd.width is not None: dst_ws.column_dimensions[col_ltr].width = cd.width
    except: pass

def _cmp_build_summary_sheet(wb, ordered, new_only, deleted_only, sheet_stats, old_fn, new_fn):
    ws = wb.create_sheet("📋 Summary", 0)
    ws.sheet_properties.tabColor = "0F2942"
    ws.merge_cells("A1:G1")
    c = ws["A1"]; c.value = "Excel Comparison Report"
    c.font = Font(name="Segoe UI",size=16,bold=True,color="0F2942")
    c.alignment = Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height = 32
    for idx,(label,value) in enumerate([("Current Pages",old_fn),("Proposed Pages",new_fn),("Generated",datetime.now().strftime("%Y-%m-%d  %H:%M:%S"))], start=2):
        ws.cell(idx,1,label).font = _CMP_FONT_BOLD; ws.cell(idx,2,value).font = _CMP_FONT_NORM
    headers = ["Sheet Name","Status","Changed Cells","Added Rows","Deleted Rows","Changed Rows","Notes"]
    hr = 6
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(hr,c_idx,h); cell.fill = _CMP_FILL_SUM_HDR; cell.font = _CMP_FONT_HDR
        cell.alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[hr].height = 22
    STATUS_FILL = {
        "Added":     PatternFill(start_color="D1FAE5",end_color="D1FAE5",fill_type="solid"),
        "Deleted":   PatternFill(start_color="FEE2E2",end_color="FEE2E2",fill_type="solid"),
        "Modified":  PatternFill(start_color="FFF3CD",end_color="FFF3CD",fill_type="solid"),
        "Unchanged": PatternFill(start_color="F3F4F6",end_color="F3F4F6",fill_type="solid"),
    }
    for row_off, sname in enumerate(ordered, start=1):
        r = hr + row_off
        if sname in new_only:       status, sv, note = "Added", {}, "New sheet in revised file"
        elif sname in deleted_only: status, sv, note = "Deleted", {}, "Removed from revised file"
        else:
            sv = sheet_stats.get(sname, {})
            has_chg = sv.get("changed_cells",0)+sv.get("added_rows",0)+sv.get("deleted_rows",0) > 0
            status = "Modified" if has_chg else "Unchanged"; note = ""
        rf = STATUS_FILL.get(status)
        data = [sname, status, sv.get("changed_cells","—"), sv.get("added_rows","—"), sv.get("deleted_rows","—"), sv.get("changed_rows","—"), note]
        for ci, val in enumerate(data, start=1):
            cell = ws.cell(r,ci,val); cell.font = _CMP_FONT_NORM; cell.alignment = Alignment(vertical="center")
            if rf: cell.fill = rf
    _cmp_apply_border(ws, hr + len(ordered), len(headers)); _cmp_auto_width(ws)

def _cmp_build_highlighted_excel(old_sheets, new_sheets, ordered, new_only, deleted_only, sheet_stats, old_fn, new_fn) -> bytes:
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for name in ordered:
        ws = wb.create_sheet(title=name[:31])
        if name in new_only:
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if not df.empty:
                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(i,j, None if val=="" else val); c.fill = _CMP_FILL_ADDED; c.font = _CMP_FONT_NORM
                _cmp_apply_border(ws,len(df),len(df.columns)); _cmp_auto_width(ws)
        elif name in deleted_only:
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if not df.empty:
                for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
                    for j, val in enumerate(row, 1):
                        c = ws.cell(i,j, None if val=="" else val); c.fill = _CMP_FILL_DELETED; c.font = _CMP_FONT_NORM
                _cmp_apply_border(ws,len(df),len(df.columns)); _cmp_auto_width(ws)
        else:
            old_df = old_sheets.get(name, pd.DataFrame()); new_df = new_sheets.get(name, pd.DataFrame())
            old_a, new_a, cell_status, row_status, stats = _cmp_compare_dataframes(old_df, new_df)
            has_chg = stats["changed_cells"]+stats["added_rows"]+stats["deleted_rows"] > 0
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["modified"] if has_chg else _CMP_TAB_COLOR["unchanged"]
            nr, nc = len(new_a), len(new_a.columns)
            for i in range(nr):
                rs = row_status.get(i,"same")
                for j in range(nc):
                    cs = cell_status.get((i,j),"same")
                    if rs=="deleted": raw=old_a.iat[i,j]; fill=_CMP_FILL_DELETED
                    else: raw=new_a.iat[i,j]; fill=(_CMP_FILL_ADDED if rs=="added" else (_CMP_FILL_CHG if cs=="changed" else None))
                    cell = ws.cell(i+1,j+1, None if raw=="" else raw); cell.font = _CMP_FONT_NORM
                    if fill: cell.fill = fill
            if nr and nc: _cmp_apply_border(ws,nr,nc); _cmp_auto_width(ws)
    _cmp_build_summary_sheet(wb, ordered, new_only, deleted_only, sheet_stats, old_fn, new_fn)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def _cmp_build_sidebyside_excel(old_sheets, new_sheets, new_only, deleted_only, sheet_stats, sheet_data, old_fn, new_fn, old_raw, new_raw) -> bytes:
    old_src_wb = _cmp_load_src_wb(old_raw, old_fn); new_src_wb = _cmp_load_src_wb(new_raw, new_fn)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    all_names = sorted(set(list(old_sheets.keys()) + list(new_sheets.keys())))

    def _finalise(ws, sep_col, src_wb, src_name):
        _cmp_auto_width(ws); ws.column_dimensions[get_column_letter(sep_col)].width = 3
        ws.sheet_view.showGridLines = False; _cmp_copy_print_settings(ws, src_wb, src_name)
        ws.page_setup.orientation = 'landscape'
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.fitToWidth = 1; ws.page_setup.fitToHeight = 0

    for name in all_names:
        ws = wb.create_sheet(title=name[:31])
        if name in new_only:
            # New sheet: proposed (green) on LEFT, blank on RIGHT
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if df.empty: continue
            nc_new = len(df.columns); sep_col = nc_new+1; old_start = sep_col+1
            for i, row_vals in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
                for jj, val in enumerate(row_vals, start=1):                        # LEFT: green (proposed)
                    c = ws.cell(i,jj, None if val=="" else val); c.fill = _CMP_FILL_SBS_NEW_SHT; c.font = _CMP_FONT_NORM
                ws.cell(i, sep_col).fill = _CMP_FILL_SBS_SEP
                for j in range(old_start, old_start+nc_new): ws.cell(i,j)           # RIGHT: blank
            _finalise(ws, sep_col, new_src_wb, name)
        elif name in deleted_only:
            # Deleted sheet: blank on LEFT, current (red) on RIGHT
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if df.empty: continue
            nc_old = len(df.columns); sep_col = nc_old+1; old_start = sep_col+1
            for i, row_vals in enumerate(dataframe_to_rows(df, index=False, header=False), start=1):
                for j in range(1, nc_old+1): ws.cell(i,j)                           # LEFT: blank
                ws.cell(i, sep_col).fill = _CMP_FILL_SBS_SEP
                for jj, val in enumerate(row_vals, start=old_start):                # RIGHT: red (current)
                    c = ws.cell(i,jj, None if val=="" else val); c.fill = _CMP_FILL_SBS_DEL_SHT; c.font = _CMP_FONT_NORM
            _finalise(ws, sep_col, old_src_wb, name)
        else:
            old_df = old_sheets.get(name, pd.DataFrame()); new_df = new_sheets.get(name, pd.DataFrame())
            if name in sheet_data: old_a, new_a, cell_status, row_status = sheet_data[name]
            else: old_a, new_a, cell_status, row_status, _ = _cmp_compare_dataframes(old_df, new_df)
            if old_a.empty and new_a.empty: continue
            nr = max(len(old_a) if not old_a.empty else 0, len(new_a) if not new_a.empty else 0)
            nc_old = len(old_a.columns) if not old_a.empty else 0
            nc_new = len(new_a.columns) if not new_a.empty else 0
            sv = sheet_stats.get(name, {})
            has_chg = (sv.get("changed_cells",0)+sv.get("added_rows",0)+sv.get("deleted_rows",0)) > 0
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["modified"] if has_chg else _CMP_TAB_COLOR["unchanged"]
            # Proposed (NEW) on LEFT, separator, Current (OLD) on RIGHT
            sep_col = nc_new+1; old_start = sep_col+1
            for i in range(nr):
                rs = row_status.get(i,"same")
                # LEFT: NEW (proposed) — yellow fill for changed cells
                for j in range(nc_new):
                    val = _cmp_cell_str(new_a.iat[i,j]) if i < len(new_a) else ""
                    cs = cell_status.get((i,j),"same")
                    c = ws.cell(i+1, j+1, None if val=="" else val)
                    if rs=="added":   c.fill = _CMP_FILL_ADDED;      c.font = _CMP_FONT_NORM
                    elif rs=="deleted": c.value = None;              c.font = _CMP_FONT_NORM
                    elif cs=="changed": c.fill = _CMP_FILL_SBS_CHG_NEW; c.font = _CMP_FONT_NORM
                    else: c.font = _CMP_FONT_NORM
                ws.cell(i+1, sep_col).fill = _CMP_FILL_SBS_SEP
                # RIGHT: OLD (current) — strikethrough red for changed cells
                for j in range(nc_old):
                    val = _cmp_cell_str(old_a.iat[i,j]) if i < len(old_a) else ""
                    cs = cell_status.get((i,j),"same")
                    c = ws.cell(i+1, old_start+j, None if val=="" else val)
                    if rs=="deleted": c.fill = _CMP_FILL_DELETED;  c.font = _CMP_FONT_NORM
                    elif rs=="added": c.value = None;              c.font = _CMP_FONT_NORM
                    elif cs=="changed": c.font = _CMP_FONT_STRIKE
                    else: c.font = _CMP_FONT_NORM
            _finalise(ws, sep_col, new_src_wb, name)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def _cmp_xml_esc(s: str) -> str:
    return s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")

def _cmp_replace_cell_rich(xml_text, ref, old_val, new_val):
    if old_val and new_val:
        is_xml = (f'<r><rPr><strike/><color rgb="FFC00000"/></rPr><t>{_cmp_xml_esc(old_val)}</t></r>'
                  f'<r><t xml:space="preserve">  {_cmp_xml_esc(new_val)}</t></r>')
    elif old_val:
        is_xml = f'<r><rPr><strike/><color rgb="FFC00000"/></rPr><t>{_cmp_xml_esc(old_val)}</t></r>'
    else:
        is_xml = f'<t>{_cmp_xml_esc(new_val or "")}</t>'
    pat = r'(<c r="' + re.escape(ref) + r'")((?:[^>]*))>(.*?)</c>'
    def _sub(m):
        attrs = re.sub(r'\s+t="[^"]*"','',m.group(2))
        return f'{m.group(1)}{attrs} t="inlineStr"><is>{is_xml}</is></c>'
    return re.sub(pat, _sub, xml_text, flags=re.DOTALL)

def _cmp_patch_inline_rich_cells(wb_bytes, wb_obj, rich_map):
    sheet_zip_path = {ws.title: f"xl/worksheets/sheet{i}.xml" for i, ws in enumerate(wb_obj.worksheets, 1)}
    in_buf = io.BytesIO(wb_bytes); out_buf = io.BytesIO()
    with zipfile.ZipFile(in_buf,"r") as zin:
        with zipfile.ZipFile(out_buf,"w",compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                for sheet_name, cells in rich_map.items():
                    if sheet_zip_path.get(sheet_name) == item.filename and cells:
                        xml_text = data.decode("utf-8")
                        for ref, (ov, nv) in cells.items():
                            xml_text = _cmp_replace_cell_rich(xml_text, ref, ov, nv)
                        data = xml_text.encode("utf-8"); break
                zout.writestr(item, data)
    return out_buf.getvalue()

def _cmp_build_inline_excel(old_sheets, new_sheets, new_only, deleted_only, sheet_stats, sheet_data, old_fn, new_fn, old_raw, new_raw) -> bytes:
    old_src_wb = _cmp_load_src_wb(old_raw, old_fn); new_src_wb = _cmp_load_src_wb(new_raw, new_fn)
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    all_names = sorted(set(list(old_sheets.keys()) + list(new_sheets.keys())))
    rich_cells: Dict[str, Dict[str,tuple]] = {}

    def _get_src_ws(wb_obj, sn):
        if wb_obj is None: return None
        try: return wb_obj[sn] if sn in wb_obj.sheetnames else None
        except: return None

    def _finalise_inline(ws, src_wb, src_name, src_ws=None):
        if src_ws is not None: _cmp_copy_row_col_dims(src_ws, ws)
        else: _cmp_auto_width(ws)
        ws.sheet_view.showGridLines = False; _cmp_copy_print_settings(ws, src_wb, src_name)

    for name in all_names:
        ws = wb.create_sheet(title=name[:31])
        if name in new_only:
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["new"]
            df = new_sheets.get(name, pd.DataFrame())
            if df.empty: continue
            src_ws = _get_src_ws(new_src_wb, name)
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    val = _cmp_cell_str(df.iat[i,j]); c = ws.cell(i+1,j+1, None if val=="" else val)
                    _cmp_copy_cell_style(src_ws.cell(i+1,j+1) if src_ws else None, c); c.fill = _CMP_FILL_ADDED
            _finalise_inline(ws, new_src_wb, name, src_ws)
        elif name in deleted_only:
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["deleted"]
            df = old_sheets.get(name, pd.DataFrame())
            if df.empty: continue
            src_ws = _get_src_ws(old_src_wb, name)
            for i in range(len(df)):
                for j in range(len(df.columns)):
                    val = _cmp_cell_str(df.iat[i,j]); c = ws.cell(i+1,j+1, None if val=="" else val)
                    src_c = src_ws.cell(i+1,j+1) if src_ws else None
                    _cmp_copy_cell_style(src_c, c); c.font = _cmp_font_with_strike(src_c.font if src_c else None)
            _finalise_inline(ws, old_src_wb, name, src_ws)
        else:
            old_df = old_sheets.get(name, pd.DataFrame()); new_df = new_sheets.get(name, pd.DataFrame())
            if name in sheet_data: old_a, new_a, cell_status, row_status = sheet_data[name]
            else: old_a, new_a, cell_status, row_status, _ = _cmp_compare_dataframes(old_df, new_df)
            if old_a.empty and new_a.empty: continue
            nr = max(len(old_a) if not old_a.empty else 0, len(new_a) if not new_a.empty else 0)
            nc = max(len(old_a.columns) if not old_a.empty else 0, len(new_a.columns) if not new_a.empty else 0)
            sv = sheet_stats.get(name, {})
            has_chg = (sv.get("changed_cells",0)+sv.get("added_rows",0)+sv.get("deleted_rows",0)) > 0
            ws.sheet_properties.tabColor = _CMP_TAB_COLOR["modified"] if has_chg else _CMP_TAB_COLOR["unchanged"]
            new_src_ws = _get_src_ws(new_src_wb, name); old_src_ws = _get_src_ws(old_src_wb, name)
            for i in range(nr):
                rs = row_status.get(i,"same")
                for j in range(nc):
                    cs = cell_status.get((i,j),"same")
                    old_val = _cmp_cell_str(old_a.iat[i,j]) if i < len(old_a) else ""
                    new_val = _cmp_cell_str(new_a.iat[i,j]) if i < len(new_a) else ""
                    c = ws.cell(i+1, j+1)
                    if rs=="deleted":
                        c.value = None if old_val=="" else old_val
                        src_c = old_src_ws.cell(i+1,j+1) if old_src_ws else None
                        _cmp_copy_cell_style(src_c, c); c.font = _cmp_font_with_strike(src_c.font if src_c else None)
                    elif rs=="added":
                        c.value = None if new_val=="" else new_val
                        src_c = new_src_ws.cell(i+1,j+1) if new_src_ws else None
                        _cmp_copy_cell_style(src_c, c); c.fill = _CMP_FILL_ADDED
                    elif cs=="changed":
                        src_c = new_src_ws.cell(i+1,j+1) if new_src_ws else None
                        _cmp_copy_cell_style(src_c, c)
                        placeholder = new_val if new_val else old_val
                        c.value = placeholder if placeholder else None
                        if placeholder:
                            ref = f"{get_column_letter(j+1)}{i+1}"
                            rich_cells.setdefault(name, {})[ref] = (old_val, new_val)
                    else:
                        c.value = None if new_val=="" else new_val
                        src_c = new_src_ws.cell(i+1,j+1) if new_src_ws else None
                        _cmp_copy_cell_style(src_c, c)
            _finalise_inline(ws, new_src_wb, name, new_src_ws)
    buf = io.BytesIO(); wb.save(buf); raw_out = buf.getvalue()
    if any(rich_cells.values()):
        raw_out = _cmp_patch_inline_rich_cells(raw_out, wb, rich_cells)
    return raw_out

def _cmp_process_file_pair(old_f, new_f, fmt):
    try:
        old_raw = _cmp_file_bytes(old_f); new_raw = _cmp_file_bytes(new_f)
        old_sheets = _cmp_read_excel_sheets(old_raw, old_f.name)
        new_sheets = _cmp_read_excel_sheets(new_raw, new_f.name)
        if not old_sheets or not new_sheets: return None, {"error":"Could not read one or both files"}
        old_names_s: Set[str] = set(old_sheets); new_names_s: Set[str] = set(new_sheets)
        new_only_s = new_names_s - old_names_s; deleted_only_s = old_names_s - new_names_s
        common_s = old_names_s & new_names_s
        ordered_s = list(old_sheets.keys()) + [s for s in new_sheets.keys() if s not in old_sheets]
        sheet_stats_s: Dict[str,dict] = {}; sheet_data_s: Dict[str,tuple] = {}
        for sname in common_s:
            oa, na, cs, rs, stats = _cmp_compare_dataframes(old_sheets[sname], new_sheets[sname])
            sheet_stats_s[sname] = stats; sheet_data_s[sname] = (oa, na, cs, rs)
        if fmt == "Side-by-Side":
            tracked = _cmp_build_sidebyside_excel(old_sheets, new_sheets, new_only_s, deleted_only_s, sheet_stats_s, sheet_data_s, old_f.name, new_f.name, old_raw, new_raw)
        else:
            tracked = _cmp_build_inline_excel(old_sheets, new_sheets, new_only_s, deleted_only_s, sheet_stats_s, sheet_data_s, old_f.name, new_f.name, old_raw, new_raw)
        total_chg = sum(v["changed_cells"]+v["added_rows"]+v["deleted_rows"] for v in sheet_stats_s.values())
        modified_s = sum(1 for v in sheet_stats_s.values() if v["changed_cells"]+v["added_rows"]+v["deleted_rows"] > 0)
        return tracked, {
            "error": None, "current_file": old_f.name, "proposed_file": new_f.name,
            "total_sheets": len(old_names_s | new_names_s), "new_sheets": len(new_only_s),
            "deleted_sheets": len(deleted_only_s), "modified_sheets": modified_s,
            "unchanged_sheets": len(common_s) - modified_s,
            "changed_cells": sum(v["changed_cells"] for v in sheet_stats_s.values()),
            "added_rows":    sum(v["added_rows"]    for v in sheet_stats_s.values()),
            "deleted_rows":  sum(v["deleted_rows"]  for v in sheet_stats_s.values()),
            "total_changes": total_chg,
        }
    except Exception as exc:
        return None, {"error": str(exc), "current_file": old_f.name, "proposed_file": new_f.name}


# ─── TRACKED PAGES COMPARATOR — UI ────────────────────────────────────────────
if active_tool == "comparator":
    _tab_ind, _tab_mass = st.tabs(["📄  Individual Pages", "📁  Mass Processing"])

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 1 — INDIVIDUAL PAGES
    # ══════════════════════════════════════════════════════════════════════════
    with _tab_ind:
        spacer(8)
        st.markdown('<div class="sec-label">&#128196; &nbsp;Upload Rate Page Files</div>', unsafe_allow_html=True)
        up_c1, up_c2 = st.columns(2, gap="large")
        with up_c1:
            st.markdown('<span class="cmp-upload-label">&#128194; Current Pages</span>', unsafe_allow_html=True)
            old_file = st.file_uploader("old_up", type=["xlsx","xls"], key="cmp_old_file",
                                        label_visibility="collapsed",
                                        help="The existing / current version of the rate pages")
        with up_c2:
            st.markdown('<span class="cmp-upload-label">&#128194; Proposed Pages</span>', unsafe_allow_html=True)
            new_file = st.file_uploader("new_up", type=["xlsx","xls"], key="cmp_new_file",
                                        label_visibility="collapsed",
                                        help="The proposed / updated version of the rate pages")

        spacer(8)
        st.markdown('<div class="sec-label">&#9881; &nbsp;Export Format &amp; Generate</div>', unsafe_allow_html=True)
        fmt_col, btn_col = st.columns([2, 1])
        with fmt_col:
            st.markdown('<p class="f-label">&#128202; &nbsp;Tracked Pages Format</p>', unsafe_allow_html=True)
            st.radio("cmp_export_fmt",
                     options=["Side-by-Side","Inline Diff"], horizontal=True,
                     label_visibility="collapsed", key="cmp_ind_fmt",
                     help="Side-by-Side: OLD on left, NEW on right.\nInline Diff: single table — changed cells show ~~old~~ new.")
        with btn_col:
            spacer(20)
            _ind_gen = st.button("⚙️  Generate Tracked Pages", type="primary",
                                 use_container_width=True, key="cmp_ind_generate_btn",
                                 disabled=not (old_file and new_file))

        if not old_file or not new_file:
            spacer(8)
            st.markdown("""
            <div class="cmp-info-box">
              <strong>How to use — Individual mode</strong>
              <ol>
                <li>Upload your <strong>Current Pages</strong> Excel on the left.</li>
                <li>Upload your <strong>Proposed Pages</strong> Excel on the right.</li>
                <li>Choose <strong>Side-by-Side</strong> or <strong>Inline Diff</strong> format.</li>
                <li>Click <strong>⚙️ Generate Tracked Pages</strong>.</li>
              </ol>
              <strong>Detected changes:</strong>
              <ul>
                <li>🟢 <strong>New sheets</strong> — added in Proposed Pages</li>
                <li>🔴 <strong>Deleted sheets</strong> — removed from Current Pages</li>
                <li>🟡 <strong>Changed cells</strong> — old ↦ new value with strikethrough</li>
                <li>🟢 <strong>Added rows</strong> &nbsp;·&nbsp; 🔴 <strong>Deleted rows</strong></li>
              </ul>
            </div>""", unsafe_allow_html=True)
        else:
            _old_fid  = f"{old_file.name}:{old_file.size}"
            _new_fid  = f"{new_file.name}:{new_file.size}"
            _cur_fids = (_old_fid, _new_fid)
            if st.session_state.cmp_file_ids != _cur_fids:
                st.session_state.cmp_comp_data        = None
                st.session_state.cmp_tracked_bytes    = None
                st.session_state.cmp_tracked_filename = None
                st.session_state.cmp_tracked_fmt      = None

            if _ind_gen:
                _fmt = st.session_state.get("cmp_ind_fmt","Side-by-Side")
                ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
                with st.spinner("Reading workbooks…"):
                    old_raw    = _cmp_file_bytes(old_file)
                    new_raw    = _cmp_file_bytes(new_file)
                    old_sheets = _cmp_read_excel_sheets(old_raw, old_file.name)
                    new_sheets = _cmp_read_excel_sheets(new_raw, new_file.name)
                if old_sheets and new_sheets:
                    old_names_g: Set[str] = set(old_sheets); new_names_g: Set[str] = set(new_sheets)
                    new_only_g     = new_names_g - old_names_g
                    deleted_only_g = old_names_g - new_names_g
                    common_g       = old_names_g & new_names_g
                    ordered_g: List[str] = list(old_sheets.keys()) + [s for s in new_sheets.keys() if s not in old_sheets]
                    with st.spinner("Computing differences…"):
                        sheet_stats_g: Dict[str,dict] = {}; sheet_data_g: Dict[str,tuple] = {}
                        for sname in common_g:
                            oa, na, cs, rs, stats = _cmp_compare_dataframes(old_sheets[sname], new_sheets[sname])
                            sheet_stats_g[sname] = stats; sheet_data_g[sname] = (oa, na, cs, rs)
                    if _fmt == "Side-by-Side":
                        with st.spinner("Generating Tracked Pages (Side-by-Side)…"):
                            tracked_bytes = _cmp_build_sidebyside_excel(old_sheets, new_sheets, new_only_g, deleted_only_g, sheet_stats_g, sheet_data_g, old_file.name, new_file.name, old_raw, new_raw)
                            tracked_fname = f"tracked_pages_sidebyside_{ts}.xlsx"
                    else:
                        with st.spinner("Generating Tracked Pages (Inline Diff)…"):
                            tracked_bytes = _cmp_build_inline_excel(old_sheets, new_sheets, new_only_g, deleted_only_g, sheet_stats_g, sheet_data_g, old_file.name, new_file.name, old_raw, new_raw)
                            tracked_fname = f"tracked_pages_inline_{ts}.xlsx"
                    with st.spinner("Building highlighted summary report…"):
                        report_bytes = _cmp_build_highlighted_excel(old_sheets, new_sheets, ordered_g, new_only_g, deleted_only_g, sheet_stats_g, old_file.name, new_file.name)
                        report_fname = f"excel_diff_{ts}.xlsx"
                    st.session_state.cmp_comp_data = {
                        "old_sheets":old_sheets,"new_sheets":new_sheets,"ordered":ordered_g,
                        "new_only":new_only_g,"deleted_only":deleted_only_g,"common":common_g,
                        "sheet_stats":sheet_stats_g,"sheet_data":sheet_data_g,
                        "old_name":old_file.name,"new_name":new_file.name,
                        "report_bytes":report_bytes,"report_fname":report_fname,
                    }
                    st.session_state.cmp_tracked_bytes    = tracked_bytes
                    st.session_state.cmp_tracked_filename = tracked_fname
                    st.session_state.cmp_tracked_fmt      = _fmt
                    st.session_state.cmp_file_ids         = _cur_fids
                else:
                    st.error("Could not read one or both files. Please check they are valid Excel workbooks.")

            _cd = st.session_state.cmp_comp_data
            if _cd is None:
                spacer(8)
                st.info("Both files are ready. Choose a format and click **⚙️ Generate Tracked Pages** to run the comparison.", icon="ℹ️")
            else:
                old_sheets   = _cd["old_sheets"];  new_sheets   = _cd["new_sheets"]
                ordered      = _cd["ordered"];     new_only     = _cd["new_only"]
                deleted_only = _cd["deleted_only"]; common       = _cd["common"]
                sheet_stats  = _cd["sheet_stats"]; sheet_data   = _cd["sheet_data"]
                old_names    = set(old_sheets);    new_names    = set(new_sheets)

                total_changes  = sum(v["changed_cells"]+v["added_rows"]+v["deleted_rows"] for v in sheet_stats.values())
                modified_count = sum(1 for v in sheet_stats.values() if v["changed_cells"]+v["added_rows"]+v["deleted_rows"] > 0)
                total_sheets_seen = len(old_names | new_names)

                spacer(12)
                st.markdown('<div class="sec-label">&#128200; &nbsp;Comparison Summary</div>', unsafe_allow_html=True)
                st.markdown(f"""
                <div class="cmp-metric-grid">
                  <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:var(--nw-deep)">{total_sheets_seen}</div><div class="cmp-metric-label">Total Sheets</div></div>
                  <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#10b981">{len(new_only)}</div><div class="cmp-metric-label">Added Sheets</div></div>
                  <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#ef4444">{len(deleted_only)}</div><div class="cmp-metric-label">Deleted Sheets</div></div>
                  <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:var(--gold)">{modified_count}</div><div class="cmp-metric-label">Modified Sheets</div></div>
                  <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:var(--muted)">{len(common) - modified_count}</div><div class="cmp-metric-label">Unchanged Sheets</div></div>
                  <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#dc2626">{total_changes:,}</div><div class="cmp-metric-label">Total Changes</div></div>
                </div>""", unsafe_allow_html=True)

                spacer(6)
                st.markdown('<div class="sec-label">&#128204; &nbsp;Sheet Overview</div>', unsafe_allow_html=True)
                st.markdown("""<div class="cmp-legend">
                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#10b981"></div>New sheet</div>
                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#ef4444"></div>Deleted sheet</div>
                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#f59e0b"></div>Modified sheet</div>
                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#A8C4E8"></div>Unchanged sheet</div>
                </div>""", unsafe_allow_html=True)
                pills_html = ['<div class="cmp-pills">']
                for sname in ordered:
                    if sname in new_only:       cls, icon = "cmp-pill cmp-pill-new", "＋"
                    elif sname in deleted_only: cls, icon = "cmp-pill cmp-pill-deleted", "−"
                    else:
                        sv = sheet_stats.get(sname, {})
                        has_chg = sv.get("changed_cells",0)+sv.get("added_rows",0)+sv.get("deleted_rows",0) > 0
                        cls = "cmp-pill cmp-pill-modified" if has_chg else "cmp-pill cmp-pill-unchanged"
                        icon = "~" if has_chg else "✓"
                    pills_html.append(f'<span class="{cls}">{_cmp_esc(f"{icon} {sname}")}</span>')
                pills_html.append("</div>")
                st.markdown("".join(pills_html), unsafe_allow_html=True)

                spacer(4)
                st.markdown('<div class="sec-label">&#128269; &nbsp;Sheet-by-Sheet Analysis</div>', unsafe_allow_html=True)
                tab_labels: List[str] = []
                for sname in ordered:
                    if sname in new_only:       tab_labels.append(f"🟢 {sname}")
                    elif sname in deleted_only: tab_labels.append(f"🔴 {sname}")
                    else:
                        sv = sheet_stats.get(sname, {})
                        has_chg = sv.get("changed_cells",0)+sv.get("added_rows",0)+sv.get("deleted_rows",0) > 0
                        tab_labels.append(f"🟡 {sname}" if has_chg else f"⚪ {sname}")
                sheet_tabs = st.tabs(tab_labels)
                for sname, stab in zip(ordered, sheet_tabs):
                    with stab:
                        if sname in new_only:
                            st.success(f"**'{sname}'** is a **new sheet** — exists only in Proposed Pages.")
                            df = new_sheets[sname]; st.caption(f"{len(df):,} rows × {len(df.columns):,} columns")
                            st.dataframe(df, use_container_width=True, height=360, hide_index=True)
                        elif sname in deleted_only:
                            st.error(f"**'{sname}'** was **deleted** — exists only in Current Pages.")
                            df = old_sheets[sname]; st.caption(f"{len(df):,} rows × {len(df.columns):,} columns")
                            st.dataframe(df, use_container_width=True, height=360, hide_index=True)
                        else:
                            old_a, new_a, cell_status, row_status = sheet_data[sname]
                            sv = sheet_stats[sname]
                            has_chg = sv["changed_cells"]+sv["added_rows"]+sv["deleted_rows"] > 0
                            if not has_chg:
                                st.info(f"✅ No changes detected in **'{sname}'**.")
                                st.dataframe(new_sheets[sname], use_container_width=True, height=300, hide_index=True)
                            else:
                                m1,m2,m3,m4 = st.columns(4)
                                m1.metric("Rows compared", f"{sv['total_rows']:,}")
                                m2.metric("Added rows",    sv["added_rows"],   delta=f"+{sv['added_rows']}"   if sv["added_rows"]   else None)
                                m3.metric("Deleted rows",  sv["deleted_rows"], delta=f"-{sv['deleted_rows']}" if sv["deleted_rows"] else None, delta_color="inverse")
                                m4.metric("Changed cells", f"{sv['changed_cells']:,}")
                                st.markdown("""<div class="cmp-legend" style="margin-top:10px">
                                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#FFFBEB;border:1.5px solid #f59e0b"></div>Changed cell</div>
                                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#ecfdf5;border:1.5px solid #10b981"></div>Added row</div>
                                  <div class="cmp-legend-item"><div class="cmp-legend-dot" style="background:#fef2f2;border:1.5px solid #ef4444"></div>Deleted row</div>
                                </div>""", unsafe_allow_html=True)
                                st.markdown(_cmp_render_diff_table(old_a, new_a, cell_status, row_status), unsafe_allow_html=True)

                spacer(12)
                st.markdown('<div class="sec-label">&#128190; &nbsp;Export</div>', unsafe_allow_html=True)
                exp_c1, exp_c2 = st.columns(2)
                with exp_c1:
                    if st.session_state.cmp_tracked_bytes:
                        _tfmt = st.session_state.cmp_tracked_fmt or "Side-by-Side"
                        st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                        st.download_button(
                            label="📋 Download Tracked Pages" + (" — Side-by-Side" if _tfmt=="Side-by-Side" else " — Inline Diff"),
                            data=st.session_state.cmp_tracked_bytes,
                            file_name=st.session_state.cmp_tracked_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True, type="primary",
                        )
                        st.markdown('</div>', unsafe_allow_html=True)
                with exp_c2:
                    st.download_button(
                        label="📥 Download Highlighted Summary",
                        data=_cd["report_bytes"], file_name=_cd["report_fname"],
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Colour-coded sheet tabs and highlighted cells",
                    )

        spacer(24)
        st.markdown('<div style="padding-top:14px;border-top:1px solid var(--border);"><p style="font-size:10px;color:#8892A4;letter-spacing:0.8px;text-transform:uppercase;text-align:center;margin:0;line-height:1.9;">Nationwide Insurance &nbsp;&middot;&nbsp; BA Analytics Division<br>Internal Use Only</p></div>', unsafe_allow_html=True)

    # ══════════════════════════════════════════════════════════════════════════
    # TAB 2 — MASS PROCESSING
    # ══════════════════════════════════════════════════════════════════════════
    with _tab_mass:
        spacer(8)
        st.markdown('<div class="sec-label">&#128196; &nbsp;Upload Rate Page Files — Batch</div>', unsafe_allow_html=True)
        mc1, mc2 = st.columns(2, gap="large")
        with mc1:
            st.markdown('<span class="cmp-upload-label">&#128194; Current Pages</span>', unsafe_allow_html=True)
            mass_cur = st.file_uploader("mass_cur", type=["xlsx","xls"], accept_multiple_files=True,
                                        key="cmp_mass_cur", label_visibility="collapsed",
                                        help="Select multiple files — hold Ctrl and click, or Ctrl+A to select all")
            if mass_cur: st.caption(f"✅ {len(mass_cur)} file(s) selected")
        with mc2:
            st.markdown('<span class="cmp-upload-label">&#128194; Proposed Pages</span>', unsafe_allow_html=True)
            mass_prop = st.file_uploader("mass_prop", type=["xlsx","xls"], accept_multiple_files=True,
                                         key="cmp_mass_prop", label_visibility="collapsed",
                                         help="Select multiple files — hold Ctrl and click, or Ctrl+A to select all")
            if mass_prop: st.caption(f"✅ {len(mass_prop)} file(s) selected")

        spacer(8)
        st.markdown('<div class="sec-label">&#9881; &nbsp;Output &amp; Format</div>', unsafe_allow_html=True)
        st.markdown('<p class="f-label">&#128193; &nbsp;Output Save Location</p>', unsafe_allow_html=True)
        st.markdown('<p class="f-hint">Files will be saved to a <code>Tracked Pages</code> subfolder here.</p>', unsafe_allow_html=True)

        if "_cmp_mass_path_pending" in st.session_state:
            st.session_state.cmp_mass_output_path = st.session_state.pop("_cmp_mass_path_pending")

        path_col, browse_col = st.columns([4,1])
        with path_col:
            st.text_input("cmp_mass_path_input", label_visibility="collapsed",
                          placeholder="Paste a folder path or click Browse",
                          key="cmp_mass_output_path")
        with browse_col:
            if st.button("Browse", key="cmp_mass_browse_btn", use_container_width=True):
                _picked = browse_folder()
                if _picked:
                    st.session_state["_cmp_mass_path_pending"] = _picked
                    st.rerun()

        _raw_mass_path = st.session_state.cmp_mass_output_path.strip().strip('"').strip("'")
        if _raw_mass_path:
            _mass_resolved = os.path.join(_raw_mass_path, "Tracked Pages")
            if os.path.isdir(_raw_mass_path):
                st.markdown(f'<p class="f-ok">&#10003; &nbsp;Files will be saved to: <code>{_mass_resolved}</code></p>', unsafe_allow_html=True)
            else:
                st.markdown(f'<p class="f-hint">&#9888; Folder not found — will be created on Generate.</p>', unsafe_allow_html=True)
        else:
            _mass_resolved = None
            st.markdown('<p class="f-hint">No location set — files available as ZIP download only.</p>', unsafe_allow_html=True)

        spacer(4)
        st.markdown('<p class="f-label">&#128202; &nbsp;Tracked Pages Format</p>', unsafe_allow_html=True)
        mfmt_col, mbtn_col = st.columns([2,1])
        with mfmt_col:
            st.radio("cmp_mass_fmt_sel", options=["Side-by-Side","Inline Diff"], horizontal=True,
                     label_visibility="collapsed", key="cmp_mass_fmt_radio")
        with mbtn_col:
            _mass_gen = st.button("⚙️  Generate All Tracked Pages", type="primary",
                                  use_container_width=True, key="cmp_mass_generate_btn",
                                  disabled=not (mass_cur and mass_prop))

        if not mass_cur or not mass_prop:
            spacer(8)
            st.markdown("""
            <div class="cmp-info-box">
              <strong>How to use — Mass Processing mode</strong>
              <ol>
                <li>Click <em>Current Pages</em> and select all current Excel files at once (hold <strong>Ctrl</strong> or <strong>Ctrl+A</strong>).</li>
                <li>Do the same for <em>Proposed Pages</em>.</li>
                <li>Files are matched automatically — only the <strong>date portion</strong> (DD-MM-YYYY) may differ in the filename.</li>
                <li>Set an <strong>Output Save Location</strong> (optional) — files are also available as a ZIP download.</li>
                <li>Click <strong>⚙️ Generate All Tracked Pages</strong>.</li>
              </ol>
              <strong>Output naming:</strong> each file is named after the Proposed Pages file with <em>" - TRACKED PAGES"</em> appended.
            </div>""", unsafe_allow_html=True)
        else:
            _mass_cur_fid  = "|".join(sorted(f"{f.name}:{f.size}" for f in mass_cur))
            _mass_prop_fid = "|".join(sorted(f"{f.name}:{f.size}" for f in mass_prop))
            _mass_fids     = (_mass_cur_fid, _mass_prop_fid)
            if st.session_state.cmp_mass_file_ids != _mass_fids:
                st.session_state.cmp_mass_results   = None
                st.session_state.cmp_mass_zip_bytes = None
                st.session_state.cmp_mass_fmt       = None

            matched_pairs, unmatched_cur, unmatched_prop = _cmp_match_pairs(mass_cur, mass_prop)
            spacer(8)
            mc_a, mc_b, mc_c = st.columns(3)
            mc_a.metric("Matched pairs",        len(matched_pairs))
            mc_b.metric("Unmatched — Current",  len(unmatched_cur))
            mc_c.metric("Unmatched — Proposed", len(unmatched_prop))

            if unmatched_cur or unmatched_prop:
                with st.expander("&#9888;&nbsp; Unmatched files (will be skipped)"):
                    if unmatched_cur:
                        st.markdown("**No match in Proposed Pages:**")
                        for f in unmatched_cur: st.markdown(f"&nbsp;&nbsp;• `{f.name}`")
                    if unmatched_prop:
                        st.markdown("**No match in Current Pages:**")
                        for f in unmatched_prop: st.markdown(f"&nbsp;&nbsp;• `{f.name}`")

            if not matched_pairs:
                st.warning("No matching file pairs found. Check that filenames match (aside from the date portion).")
            else:
                if _mass_gen:
                    _mfmt = st.session_state.get("cmp_mass_fmt_radio","Side-by-Side")
                    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
                    _disk_dir: Optional[str] = None
                    if _raw_mass_path:
                        _disk_dir = os.path.join(_raw_mass_path, "Tracked Pages")
                        try: os.makedirs(_disk_dir, exist_ok=True)
                        except Exception as _me: st.error(f"Could not create output folder: {_me}"); _disk_dir = None

                    mass_results_list = []; zip_buf = io.BytesIO(); total_pairs = len(matched_pairs)
                    with st.status(f"&#9881; Generating {total_pairs} Tracked Page(s)…", expanded=True) as _status:
                        prog = st.progress(0, text="Initialising…")
                        with zipfile.ZipFile(zip_buf,"w",compression=zipfile.ZIP_DEFLATED) as zf:
                            for idx, (cur_f, prop_f) in enumerate(matched_pairs):
                                st.write(f"**[{idx+1}/{total_pairs}]** Processing `{prop_f.name}`…")
                                prog.progress(idx/total_pairs, text=f"Processing {idx+1} of {total_pairs}: {prop_f.name}")
                                tracked_bytes, stats = _cmp_process_file_pair(cur_f, prop_f, _mfmt)
                                mass_results_list.append(stats)
                                if tracked_bytes:
                                    stem = os.path.splitext(prop_f.name)[0]
                                    out_fname = f"{stem} - TRACKED PAGES.xlsx"
                                    zf.writestr(out_fname, tracked_bytes)
                                    if _disk_dir:
                                        try:
                                            with open(os.path.join(_disk_dir, out_fname),"wb") as fh:
                                                fh.write(tracked_bytes)
                                        except Exception as _we: stats["save_error"] = str(_we)
                        prog.progress(1.0, text=f"✅ Done — {total_pairs} file(s) processed.")
                        _status.update(label=f"✅ Complete — {total_pairs} file(s) processed.", state="complete", expanded=False)

                    st.session_state.cmp_mass_results   = mass_results_list
                    st.session_state.cmp_mass_zip_bytes = zip_buf.getvalue()
                    st.session_state.cmp_mass_fmt       = _mfmt
                    st.session_state.cmp_mass_file_ids  = _mass_fids
                    st.session_state.cmp_mass_save_dir  = _disk_dir

                _mr = st.session_state.cmp_mass_results
                if _mr is None:
                    spacer(8)
                    st.info(f"**{len(matched_pairs)} pair(s)** ready. Choose a format and click **⚙️ Generate All Tracked Pages** to start.", icon="ℹ️")
                else:
                    ok_results  = [r for r in _mr if not r.get("error")]
                    err_results = [r for r in _mr if r.get("error")]
                    total_files_proc = len(ok_results)
                    total_chg_across = sum(r["total_changes"]   for r in ok_results)
                    total_cell_across= sum(r["changed_cells"]   for r in ok_results)
                    total_add_across = sum(r["added_rows"]      for r in ok_results)
                    total_del_across = sum(r["deleted_rows"]    for r in ok_results)
                    files_with_chg   = sum(1 for r in ok_results if r["total_changes"]>0)
                    files_no_chg     = total_files_proc - files_with_chg

                    spacer(12)
                    st.markdown('<div class="sec-label">&#128200; &nbsp;Batch Summary</div>', unsafe_allow_html=True)
                    st.markdown(f"""
                    <div class="cmp-metric-grid">
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:var(--nw-deep)">{total_files_proc}</div><div class="cmp-metric-label">Files Processed</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:var(--gold)">{files_with_chg}</div><div class="cmp-metric-label">Files With Changes</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#10b981">{files_no_chg}</div><div class="cmp-metric-label">Files Unchanged</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#dc2626">{total_chg_across:,}</div><div class="cmp-metric-label">Total Changes</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:var(--nw-blue)">{total_cell_across:,}</div><div class="cmp-metric-label">Changed Cells</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#10b981">{total_add_across:,}</div><div class="cmp-metric-label">Added Rows</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#ef4444">{total_del_across:,}</div><div class="cmp-metric-label">Deleted Rows</div></div>
                      <div class="cmp-metric-card"><div class="cmp-metric-val" style="color:#ef4444">{len(err_results)}</div><div class="cmp-metric-label">Errors</div></div>
                    </div>""", unsafe_allow_html=True)

                    spacer(8)
                    st.markdown('<div class="sec-label">&#128203; &nbsp;Per-File Results</div>', unsafe_allow_html=True)
                    tbl_rows = []
                    for r in _mr:
                        if r.get("error"):
                            tbl_rows.append({"Current File":r.get("current_file","—"),"Proposed File":r.get("proposed_file","—"),"Status":f"❌ Error: {r['error']}","Sheets":"—","New":"—","Deleted":"—","Modified":"—","Changed Cells":"—","Added Rows":"—","Deleted Rows":"—"})
                        else:
                            status = "✅ No changes" if r["total_changes"]==0 else f"🟡 {r['total_changes']:,} changes"
                            tbl_rows.append({"Current File":r["current_file"],"Proposed File":r["proposed_file"],"Status":status,"Sheets":r["total_sheets"],"New":r["new_sheets"],"Deleted":r["deleted_sheets"],"Modified":r["modified_sheets"],"Changed Cells":r["changed_cells"],"Added Rows":r["added_rows"],"Deleted Rows":r["deleted_rows"]})
                    st.dataframe(pd.DataFrame(tbl_rows), use_container_width=True, hide_index=True, height=min(400, 60+38*len(tbl_rows)))
                    if err_results:
                        st.error(f"{len(err_results)} file(s) failed — see Status column above.")

                    spacer(8)
                    st.markdown('<div class="sec-label">&#128190; &nbsp;Output</div>', unsafe_allow_html=True)
                    _saved_dir = st.session_state.cmp_mass_save_dir
                    save_errors = [r for r in ok_results if r.get("save_error")]
                    if _saved_dir:
                        saved_count = len(ok_results) - len(save_errors)
                        st.success(f"**{saved_count} file(s) saved** to `{_saved_dir}`", icon="✅")
                        if save_errors:
                            with st.expander(f"⚠️ {len(save_errors)} file(s) could not be saved to disk"):
                                for r in save_errors:
                                    st.markdown(f"• `{r.get('proposed_file','?')}` — {r['save_error']}")
                    else:
                        st.info("No save location was set — files available as ZIP download only.", icon="ℹ️")

                    _mfmt_label = st.session_state.get("cmp_mass_fmt","Side-by-Side")
                    st.markdown('<div class="btn-ready">', unsafe_allow_html=True)
                    st.download_button(
                        label=f"📦 Download All Tracked Pages (ZIP) — {_mfmt_label}",
                        data=st.session_state.cmp_mass_zip_bytes,
                        file_name=f"tracked_pages_batch_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip",
                        mime="application/zip", use_container_width=True, type="primary",
                    )
                    st.markdown('</div>', unsafe_allow_html=True)

        spacer(24)
        st.markdown('<div style="padding-top:14px;border-top:1px solid var(--border);"><p style="font-size:10px;color:#8892A4;letter-spacing:0.8px;text-transform:uppercase;text-align:center;margin:0;line-height:1.9;">Nationwide Insurance &nbsp;&middot;&nbsp; BA Analytics Division<br>Internal Use Only</p></div>', unsafe_allow_html=True)