"""
bop_config.py
=============
Loads "BOP/BOP Input File.xlsx" once and exposes its tabs as plain Python
dicts/lists. This is the ONLY place that reads that workbook — every other
BOP module (ExcelSettingsBOP.py, AllProgramsPage.py, BOPRatePages.py) calls
load_bop_config() and reads the returned object's attributes.

To let a non-technical person change a font, a column width, a page-break
rule, or a rating lookup table: edit "BOP Input File.xlsx" directly. No
Python file needs to change.
"""

from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from config.constants import BOP_INPUT_FILE


@dataclass
class BOPConfig:
    formatting: Dict[str, object] = field(default_factory=dict)
    header_footer: Dict[str, str] = field(default_factory=dict)
    table_layout: Dict[str, List[Tuple[int, object, int]]] = field(default_factory=dict)
    number_formats: Dict[str, List[Tuple[int, int, int, str]]] = field(default_factory=dict)
    sub_headers: Dict[str, dict] = field(default_factory=dict)
    footnotes: Dict[str, List[Tuple[str, str]]] = field(default_factory=dict)
    page_break_rules: List[Tuple[str, str]] = field(default_factory=list)
    perils_by_state: Dict[str, List[str]] = field(default_factory=dict)
    peril_conversions: Dict[str, str] = field(default_factory=dict)
    protection_class_conversions: Dict[str, str] = field(default_factory=dict)
    building_codes_by_state: Dict[str, Dict[str, List[str]]] = field(default_factory=dict)


def _rows(ws):
    """Yield each data row (skipping the header row) as a tuple of values."""
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        yield row


def _col(value):
    """Parse a Table Layout / Number Formats column boundary cell: an int, or 'REST'."""
    if isinstance(value, str) and value.strip().upper() == "REST":
        return "REST"
    return int(value)


@lru_cache(maxsize=1)
def load_bop_config(path: str = None) -> BOPConfig:
    wb = load_workbook(str(path or BOP_INPUT_FILE), read_only=True, data_only=True)
    cfg = BOPConfig()

    for setting, value in _rows(wb["Formatting Defaults"]):
        cfg.formatting[setting] = value

    for field_name, value in _rows(wb["Header Footer Text"]):
        cfg.header_footer[field_name] = value or ""

    for table_code, col_start, col_end, width_px in _rows(wb["Table Layout"]):
        cfg.table_layout.setdefault(table_code, []).append(
            (int(col_start), _col(col_end), float(width_px))
        )

    for table_code, col_start, col_end, row_start, fmt in _rows(wb["Number Formats"]):
        cfg.number_formats.setdefault(table_code, []).append(
            (int(col_start), _col(col_end), int(row_start), fmt)
        )

    for row in _rows(wb["Sub Headers"]):
        table_code, insert_at_row, print_title_rows, l1_range, l1_text, l2_range, l2_text = row
        cfg.sub_headers[table_code] = {
            "insert_at_row": int(insert_at_row),
            "print_title_rows": print_title_rows,
            "label1_range": l1_range or "",
            "label1_text": l1_text or "",
            "label2_range": l2_range or "",
            "label2_text": l2_text or "",
        }

    for table_code, cell, text in _rows(wb["Footnotes"]):
        cfg.footnotes.setdefault(table_code, []).append((cell, text or ""))

    for prefix, rule in _rows(wb["Page Break Rules"]):
        cfg.page_break_rules.append((prefix, rule))

    for state, perils_csv in _rows(wb["Perils By State"]):
        cfg.perils_by_state[state] = [p.strip() for p in str(perils_csv).split(",") if p.strip()]

    for peril_code, display_name in _rows(wb["Peril Conversions"]):
        cfg.peril_conversions[peril_code] = display_name

    for code, display_value in _rows(wb["Protection Class Conversions"]):
        cfg.protection_class_conversions[str(code)] = str(display_value)

    for state, group, codes_csv in _rows(wb["Building Codes By State"]):
        codes = [c.strip() for c in str(codes_csv).split(",") if c.strip()]
        cfg.building_codes_by_state.setdefault(state, {})[group] = codes

    wb.close()
    return cfg
