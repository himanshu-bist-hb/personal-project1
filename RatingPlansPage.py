# This module formats the Rating Plans State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class RatingPlans:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective, RatingPlansApplies, IRPMCredit, IRPMDebit) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date
        self.RatingPlansApplies = RatingPlansApplies
        self.IRPMCredit = IRPMCredit
        self.IRPMDebit = IRPMDebit

        #self.autoProgramCode = 20000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the Multi Building Credit Factor table
    # Returns a dataframe
    def buildMultiBuildingCredit(self):
        MultiBuildingAllOther = self.buildDataFrame("BP7_Peril_Multi_Building_Credit").query(f'Class_Code_Min == 20000 & `Peril TypeCode` == "allperil"').fillna({'Building_No_Max': 0}). \
                astype({'Building_No_Min': 'int64', 'Building_No_Max': 'int64'}).astype({'Building_No_Min': 'string', 'Building_No_Max': 'string'})
        MultiBuildingAllOther['Total # of Buildings'] = np.where(MultiBuildingAllOther['Building_No_Max'] == '0', 
                                                                 MultiBuildingAllOther['Building_No_Min'] + '+',
                                                                 MultiBuildingAllOther['Building_No_Min'] + ' - ' + MultiBuildingAllOther['Building_No_Max'])
        FilteredMultiBuildingAllOther = MultiBuildingAllOther.filter(items=['Total # of Buildings', 'Multi_Building_Credit_Factor']).rename(columns={'Multi_Building_Credit_Factor': 'All Other'})
        MultiBuildingHab = self.buildDataFrame("BP7_Peril_Multi_Building_Credit").query(f'Class_Code_Min == 10000 & `Peril TypeCode` == "allperil"').fillna({'Building_No_Max': 0}). \
                astype({'Building_No_Min': 'int64', 'Building_No_Max': 'int64'}).astype({'Building_No_Min': 'string', 'Building_No_Max': 'string'})
        MultiBuildingHab['Total # of Buildings'] = np.where(MultiBuildingHab['Building_No_Max'] == '0', 
                                                                 MultiBuildingHab['Building_No_Min'] + '+',
                                                                 MultiBuildingHab['Building_No_Min'] + ' - ' + MultiBuildingHab['Building_No_Max'])
        FilteredMultiBuildingHab = MultiBuildingHab.filter(items=['Total # of Buildings', 'Multi_Building_Credit_Factor']).rename(columns={'Multi_Building_Credit_Factor': 'Habitational'})
        MultiBuildingNRPAllOther = self.buildDataFrame("BP7 Peril Multi Building Credit NRP").query(f'Class_Code_Min == 20000 & `Peril TypeCode` == "allperil"').fillna({'Building_No_Max': 0}). \
                astype({'Building_No_Min': 'int64', 'Building_No_Max': 'int64'}).astype({'Building_No_Min': 'string', 'Building_No_Max': 'string'})
        MultiBuildingNRPAllOther['Total # of Buildings'] = np.where(MultiBuildingNRPAllOther['Building_No_Max'] == '0', 
                                                                 MultiBuildingNRPAllOther['Building_No_Min'] + '+',
                                                                 MultiBuildingNRPAllOther['Building_No_Min'] + ' - ' + MultiBuildingNRPAllOther['Building_No_Max'])
        FilteredMultiBuildingNRPAllOther = MultiBuildingNRPAllOther.filter(items=['Total # of Buildings', 'Multi_Building_Credit_Factor']).rename(columns={'Multi_Building_Credit_Factor': 'All Other (NRP)'})
        MultiBuildingNRPHab = self.buildDataFrame("BP7 Peril Multi Building Credit NRP").query(f'Class_Code_Min == 10000 & `Peril TypeCode` == "allperil"').fillna({'Building_No_Max': 0}). \
                astype({'Building_No_Min': 'int64', 'Building_No_Max': 'int64'}).astype({'Building_No_Min': 'string', 'Building_No_Max': 'string'})
        MultiBuildingNRPHab['Total # of Buildings'] = np.where(MultiBuildingNRPHab['Building_No_Max'] == '0', 
                                                                 MultiBuildingNRPHab['Building_No_Min'] + '+',
                                                                 MultiBuildingNRPHab['Building_No_Min'] + ' - ' + MultiBuildingNRPHab['Building_No_Max'])
        FilteredMultiBuildingNRPHab = MultiBuildingNRPHab.filter(items=['Total # of Buildings', 'Multi_Building_Credit_Factor']).rename(columns={'Multi_Building_Credit_Factor': 'Habitational (NRP)'})
        MultiBuildingNonNRP = pd.merge(FilteredMultiBuildingAllOther, FilteredMultiBuildingHab, on = 'Total # of Buildings', how = 'inner')
        MultiBuildingNRP = pd.merge(FilteredMultiBuildingNRPAllOther, FilteredMultiBuildingNRPHab, on = 'Total # of Buildings', how = 'inner')
        MultiBuilding = pd.merge(MultiBuildingNonNRP, MultiBuildingNRP, on = 'Total # of Buildings', how = 'inner')
        return MultiBuilding
    
    # Builds the Tiering Factor table
    # Returns a dataframe
    def buildTieringFactor(self):
        TieringFactorHab = self.buildDataFrame("BP7_Peril_Tiering_Factor")
        filteredTieringFactorHab = TieringFactorHab.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'TierFactor': 'Factor'}).query(f'Grade in {"1", "2", "3", "4", "5", "6", "7", "8", "9"}')
        filteredTieringFactorHab['Program'] = 'H'
        TieringFactorASFSW = self.buildDataFrame("BP7_Peril_Tiering_Factor")
        filteredTieringFactorASFSW = TieringFactorASFSW.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'TierFactor': 'Factor'}).query(f'Grade in {"A", "B", "C", "D", "E", "F", "G", "H", "I", "J"}')
        filteredTieringFactorASFSW['Program'] = 'AS/FS/W'
        TieringFactorORS = self.buildDataFrame("BP7_Peril_Tiering_Factor")
        filteredTieringFactorORS = TieringFactorORS.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'TierFactor': 'Factor'}).query(f'Grade in {"K", "L", "M", "N", "O", "P", "Q", "R", "S", "T"}')
        filteredTieringFactorORS['Program'] = 'O/R/S'
        TieringFactorHASFSW = pd.concat([filteredTieringFactorHab, filteredTieringFactorASFSW])
        TieringFactor = pd.concat([TieringFactorHASFSW, filteredTieringFactorORS])
        return TieringFactor.sort_values(by= 'Program', ascending=False).sort_values(by=['Peril TypeCode', 'Grade']).rename(columns={'Peril TypeCode': 'Peril'}).filter(items=['Peril', 'Program', 'Grade', 'Factor'])

    # Builds the IRPM Threshold table
    # Returns a dataframe
    def buildIRPMThreshold(self):
        IRPMThreshold = self.buildDataFrame("BP7_IRPM_Eligibility_Threshold")
        filteredIRPMThreshold = IRPMThreshold.query(f'ProgramCode == "Auto Service"').filter(items=['IRPMEligibleAmount']).rename(columns={'IRPMEligibleAmount': 'Amount'})
        return filteredIRPMThreshold

    # Builds the IRPM Modification Plan table
    # Returns a dataframe
    def buildIRPMModPlan(self):
        if self.RatingPlansApplies.get() == 1:
            IRPMCredit = self.IRPMCredit.get()
            IRPMDebit = self.IRPMDebit.get()
            IRPMData = [[IRPMCredit, IRPMDebit]]
            IRPMModPlan = pd.DataFrame(IRPMData, columns=['Credit', 'Debit'])
            return IRPMModPlan

    # Builds the Lifetime Expense Allocation Factor tables
    # Returns a dataframe
    def buildLEAFfactors(self):
        LEAFfactors = self.buildDataFrame("BP7_Peril_Retention_Factor")
        filteredLEAFfactors = LEAFfactors.query(f'RetentionGrade in {"A", "B", "C", "D", "E", "F"}').filter(items=['Peril TypeCode', 'RetentionGrade', 'RetentionFactor']).replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'Peril TypeCode': 'Peril', 'RetentionGrade': 'Grade', 'RetentionFactor': 'Factor'}). \
                sort_values(by=['Peril', 'Grade'])
        return filteredLEAFfactors

    # Builds the Large Premium Discount Plan factor table
    # Returns a dataframe
    def buildLPDPFactors(self):
        LPDPNRPFactors = self.buildDataFrame("BP7 LPDP NRP Factor_v1_Ext").query(f'ProgramCode == "Auto Service"').replace({'PremiumRange': {1: "$0 to $5,000", 5001: "$5,001 to $6,000", 6001: "$6,001 to $8,000", 8001: "$8,001 to $10,000", 10001: "$10,001 to $15,000", 15001: "$15,001 to $20,000", 20001: "$20,001 to $25,000", 25001: "$25,001 to $30,000", 30001: "$30,001 to $35,000", 35001: "$35,001 to $40,000", 40001: "$40,001 to $45,000", 45001: "$45,001 to $50,000", 50001: "$50,001 to $75,000", 75001: "$75,001 to $100,000", 100001: "$100,001 to $150,000", 150001: "$150,001 to $200,000", 200001: "$200,001 to $250,000", 250001: "$250,001 and greater"}})
        FilteredLPDPNRPFactors = LPDPNRPFactors.filter(items=['PremiumRange', 'LPDPFactor']).rename(columns={'PremiumRange': 'Annual Premium','LPDPFactor': 'National Retail Program Factor Range'})
        LPDPNonNRPFactors = self.buildDataFrame("BP7 LPDP Factor_v2_Ext").query(f'ProgramCode == "Auto Service"').replace({'PremiumRange': {0: "$0 to $5,000", 5001: "$5,001 to $6,000", 6001: "$6,001 to $8,000", 8001: "$8,001 to $10,000", 10001: "$10,001 to $15,000", 15001: "$15,001 to $20,000", 20001: "$20,001 to $25,000", 25001: "$25,001 to $30,000", 30001: "$30,001 to $35,000", 35001: "$35,001 to $40,000", 40001: "$40,001 to $45,000", 45001: "$45,001 to $50,000", 50001: "$50,001 to $75,000", 75001: "$75,001 to $100,000", 100001: "$100,001 to $150,000", 150001: "$150,001 to $200,000", 200001: "$200,001 to $250,000", 250001: "$250,001 and greater"}})
        FilteredLPDPNonNRPFactors = LPDPNonNRPFactors.filter(items=['PremiumRange', 'LPDPFactor']).rename(columns={'PremiumRange': 'Annual Premium','LPDPFactor': 'All Other Factor Range'})
        LPDPFactors = pd.merge(FilteredLPDPNonNRPFactors, FilteredLPDPNRPFactors, on = 'Annual Premium', how = 'inner')
        return LPDPFactors

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the MBCP worksheet
    def formatMultiBuildingCredit(self, ws):
        ws.insert_rows(3)
        ws['B2'] = 'MBCP Factor'
        ws['B3'] = 'All Other'
        ws['D3'] = 'National Retail Program'
        for cell in ws['2:2']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('B2:E2')
        ws.merge_cells('B3:C3')
        ws.merge_cells('D3:E3')
        ws.print_title_rows = '1:4'
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(80)
        ws.merge_cells('A2:A4')
        ws['A2'] = 'Total # of Buildings'
        #for cell in ws['A2']:
        #        cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    # Applies manual formatting to the RTRP worksheet
    def formatTieringFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(80)
        for row in range(4, ws.max_row + 1):
            cell = ws['D' + str(row)]
            cell.number_format = '#,##0.0000'

    # Applies manual formatting to the RPMET worksheet
    def formatIRPMThreshold(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(70)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '$#,##0'

    # Applies manual formatting to the RPMP worksheet
    def formatIRPMModPlan(self, ws):
        ws.insert_rows(3)
        ws.insert_rows(4)
        ws['A3'] = 'Total Modification:'
        for cell in ws['3:3']:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
        ws.column_dimensions['A'].width = self.pixelsToInches(80)
        ws.column_dimensions['B'].width = self.pixelsToInches(80)
        for row in range(6, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '#,##0%'
        for row in range(6, ws.max_row + 1):
            cell = ws['B' + str(row)]
            cell.number_format = '#,##0%'

    # Applies manual formatting to the LEAF worksheet
    #def formatLEAFFactors(self, ws):

    # Applies manual formatting to the LPDP worksheet
    def formatLPDPFactors(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(200)
        ws.column_dimensions['B'].width = self.pixelsToInches(95)
        ws.column_dimensions['C'].width = self.pixelsToInches(150)
    
    # Sets up the Auto Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildRatingPlansPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        RatingPlans = ExcelSettings.Excel(state=self.state, programName='Rating Plans', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = RatingPlans.getFontName()
        fontSize = RatingPlans.getFontSize()

        RatingPlans.generateWorksheet('MBCP', 'RP Table 90.B. Multiple Building Credit Plan', self.buildMultiBuildingCredit(), False, True)
        RatingPlans.generateWorksheet('RTRP', 'RP Table 91.C. Risk Tier Rating Plan', self.buildTieringFactor(), False, True)
        RatingPlans.generateWorksheet('RPMET', 'RP Table 92.A. State Individual Risk Premium Modification Eligibility Threshold', self.buildIRPMThreshold(), False, True)
        RatingPlans.generateWorksheet('RPMP', 'RP Table 92.C. State Individual Risk Premium Modification Plan', self.buildIRPMModPlan(), False, True)
        RatingPlans.generateWorksheet('LEAF', 'RP Table 94.C.1 Lifetime Expense Allocation Factor', self.buildLEAFfactors(), False, True)
        RatingPlans.generateWorksheet('LPDP', 'RP Table 95.C. Large Premium Discount Plan', self.buildLPDPFactors(), False, True)

        RatingPlans.createIndex()
        RatingPages = RatingPlans.getWB()

        self.formatMultiBuildingCredit(RatingPages['MBCP'])  
        self.formatTieringFactor(RatingPages['RTRP'])  
        self.formatIRPMThreshold(RatingPages['RPMET'])
        self.formatIRPMModPlan(RatingPages['RPMP'])
        #self.formatYearBuiltModifier(RatingPages['LEAF'])  
        self.formatLPDPFactors(RatingPages['LPDP'])

        return RatingPages