"""
BOPpagebreaks.py
=================
Post-processing for the BOP Rate Pages workbook — mirrors BA/BApagebreaks.py.

Unlike BA (which has ~25 sheet-specific rule handlers accumulated from years
of real print testing), BOP's page-break rules are driven entirely by the
"Page Break Rules" tab in "BOP/BOP Input File.xlsx": a list of
(sheet-name-prefix -> rule name) pairs, most-specific-first, with "*" as the
catch-all default. Add a row there to change how a sheet paginates — no
Python required — unless the sheet needs genuinely custom per-cell logic,
in which case add a handler to _RULE_HANDLERS below and reference its name
from the config tab.

The generic openpyxl helpers (fit_single_page, etc.) and the COM/zip
utilities (_sanitize_xlsx, _kill_excel_instances, export_to_pdf) are not
BA-specific, so they are imported from BA.BApagebreaks rather than
duplicated here.
"""

import os

import openpyxl

from BA.BApagebreaks import (
    fit_single_page,
    fit_width_only,
    disable_fit_to_page,
    _sanitize_xlsx,
    export_to_pdf,
)
from .bop_config import load_bop_config


def _handle_index(ws, dest_filename):
    # print_title_rows = None (not "0:0") so we don't write an invalid
    # definedName like Index!$0:$0 - that is what triggers Open-and-Repair.
    ws.print_title_rows = None
    ws.print_area = f"A1:J{ws.max_row}"
    fit_width_only(ws)


def _handle_fit_single_page(ws, dest_filename):
    fit_single_page(ws)


def _handle_fit_width_only(ws, dest_filename):
    fit_width_only(ws)


def _handle_disable_fit_to_page(ws, dest_filename):
    disable_fit_to_page(ws)


# Rule name (as written in the "Page Break Rules" tab) -> handler function.
_RULE_HANDLERS = {
    "index": _handle_index,
    "fit_single_page": _handle_fit_single_page,
    "fit_width_only": _handle_fit_width_only,
    "disable_fit_to_page": _handle_disable_fit_to_page,
}


def _apply_matching_rule(sheet_name, ws, dest_filename, page_break_rules):
    for prefix, rule_name in page_break_rules:
        if prefix == "*" or sheet_name.startswith(prefix):
            handler = _RULE_HANDLERS.get(rule_name)
            if handler:
                handler(ws, dest_filename)
            return True
    return False


def process_pagebreaks(dest_filename1, dest_filename2=None):
    """
    Apply page breaks / print settings to dest_filename1.

    dest_filename2 is accepted for backward compatibility (was a PDF path)
    and is unused, matching BA.BApagebreaks.process_pagebreaks.
    """
    print(f"[BOPpagebreaks] Processing: {dest_filename1}")
    dest_filename1 = os.path.normpath(os.path.abspath(dest_filename1))

    page_break_rules = load_bop_config().page_break_rules

    workbook = openpyxl.load_workbook(dest_filename1)

    for original_name in list(workbook.sheetnames):
        if len(original_name) > 31:
            workbook[original_name].title = original_name[:31]

    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        _apply_matching_rule(sheet_name, ws, dest_filename1, page_break_rules)

    if "Index" in workbook.sheetnames:
        ws_index = workbook["Index"]
        ws_index.sheet_state = "visible"
        if workbook.sheetnames.index("Index") != 0:
            workbook._sheets.remove(ws_index)
            workbook._sheets.insert(0, ws_index)
        workbook.active = 0

    workbook.save(dest_filename1)
    workbook.close()

    _sanitize_xlsx(dest_filename1)
    print("[BOPpagebreaks] Done.")
