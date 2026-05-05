"""
SCRIPT 2 — JSON to Excel
Usage: python json_to_excel.py

Reads ba_input_data.json from the current directory and reconstructs
BA Input File.xlsx with all original sheets, column headers, and data.
"""

import json
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE  = "ba_input_data.json"
OUTPUT_FILE = "BA Input File.xlsx"


def style_header_row(ws, num_cols):
    """Apply a clean header style to row 1."""
    header_fill   = PatternFill("solid", fgColor="1F3864")
    header_font   = Font(bold=True, color="FFFFFF", size=10)
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side     = Side(style="thin", color="AAAAAA")
    thin_border   = Border(left=thin_side, right=thin_side,
                           top=thin_side, bottom=thin_side)

    for col in range(1, num_cols + 1):
        cell             = ws.cell(row=1, column=col)
        cell.fill        = header_fill
        cell.font        = header_font
        cell.alignment   = header_align
        cell.border      = thin_border

    ws.row_dimensions[1].height = 22


def auto_col_widths(ws, num_cols, max_width=50):
    """Set column widths based on content length."""
    for col in range(1, num_cols + 1):
        col_letter = get_column_letter(col)
        max_len    = 0
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len  = max(max_len, cell_len)
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def json_to_excel(input_path, output_path):
    if not os.path.exists(input_path):
        print(f"ERROR: '{input_path}' not found in current directory.")
        return

    print(f"Reading: {input_path}")
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, rows in data.items():
            if not rows:
                print(f"  ✗ {sheet_name:35s}  SKIPPED — empty")
                continue

            try:
                headers   = rows[0]
                data_rows = rows[1:]
                df        = pd.DataFrame(data_rows, columns=headers)

                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Apply styling
                ws         = writer.sheets[sheet_name]
                num_cols   = len(headers)
                style_header_row(ws, num_cols)
                auto_col_widths(ws, num_cols)
                ws.freeze_panes = "A2"

                print(f"  ✓ {sheet_name:35s}  ({len(data_rows)} data rows)")

            except Exception as e:
                print(f"  ✗ {sheet_name:35s}  ERROR — {e}")

    size_kb = os.path.getsize(output_path) / 1024
    print(f"\nSaved: {output_path}  ({size_kb:.1f} KB)")
    print(f"Sheets created: {len(data)}")
    print("\nBA Input File.xlsx has been fully reconstructed.")


if __name__ == "__main__":
    json_to_excel(INPUT_FILE, OUTPUT_FILE)
