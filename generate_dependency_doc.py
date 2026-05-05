"""
Generates: BA_Input_File_Dependency_Map.xlsx
A full reference document of every BA Input File.xlsx sheet dependency.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Color Palette ─────────────────────────────────────────────────────────────
C_HEADER_DARK   = "1F3864"   # dark navy  – sheet headers
C_HEADER_MED    = "2E75B6"   # mid blue   – section headers
C_HEADER_LIGHT  = "D6E4F0"   # pale blue  – column headers
C_HIGH          = "E2EFDA"   # green      – High reduction potential
C_MEDIUM_HIGH   = "FFF2CC"   # yellow     – Medium-High
C_MEDIUM        = "FCE4D6"   # orange     – Medium
C_LOW           = "FFDCE1"   # red-pink   – Low
C_CAT_COLORS = {
    "State Config":         "DEEAF1",
    "LCM / Company Dev":    "E2EFDA",
    "Rate Sheet Routing":   "FFF2CC",
    "Rule-Specific Maps":   "FCE4D6",
    "Classification Tables":"EAD1DC",
    "Limits Tables":        "D9D2E9",
    "Specialized / State":  "F4CCCC",
}
POTENTIAL_COLOR = {
    "HIGH":        "70AD47",
    "MEDIUM-HIGH": "FFC000",
    "MEDIUM":      "ED7D31",
    "LOW":         "FF0000",
}

# ── Data ──────────────────────────────────────────────────────────────────────
MASTER_DATA = [
    # (Sheet Name, Category, Rule/Function, What it provides,
    #  Key Columns, Read Frequency, Reduction Potential, Action / Notes)
    (
        "No MedPay", "State Config", "Initialization / __init__",
        "List of states that do NOT have Medical Payments coverage",
        "states",
        "Once at startup", "HIGH",
        "Replace with a Python constant list — rarely changes across filings"
    ),
    (
        "PIP States", "State Config", "Initialization / __init__",
        "List of states that DO have PIP (Personal Injury Protection) coverage",
        "states",
        "Once at startup", "HIGH",
        "Replace with a Python constant list — rarely changes across filings"
    ),
    (
        "LCM-Comp Dev Mapping", "LCM / Company Dev", "Initialization / __init__",
        "Maps every ratebook sheet to flags for LCM and Company Deviation application",
        "sheetnames, coverage, lcm, company_dev",
        "Once at startup", "MEDIUM",
        "Changes per filing; consider moving to YAML/JSON config instead of Excel sheet"
    ),
    (
        "LCM-Map", "LCM / Company Dev", "Initialization / __init__",
        "Maps coverage types to their LCM multiplier type",
        "coverage, lcm_type",
        "Once at startup", "MEDIUM",
        "Small lookup; could be a Python dict if coverage-to-LCM relationship is stable"
    ),
    (
        "222 TTT", "Rate Sheet Routing", "buildBaseRates()",
        "Maps state + coverage to the correct ratebook sheet for Trucks/Tractors/Trailers",
        "state, coverage, sheet, group1, group2, group3, filter1, filter2, filter3",
        "Per state per run", "LOW",
        "Core dynamic routing — changes per state/filing; keep in Excel"
    ),
    (
        "232 PPT", "Rate Sheet Routing", "buildBaseRates()",
        "Maps state + coverage to the correct ratebook sheet for Private Passenger Types",
        "state, coverage, sheet, group1, group2, group3, filter1, filter2, filter3",
        "Per state per run", "LOW",
        "Core dynamic routing — changes per state/filing; keep in Excel"
    ),
    (
        "225 Zone Rated", "Rate Sheet Routing", "buildZoneBaseRates()",
        "Maps state + coverage to the correct ratebook sheet for Zone-Rated risks",
        "state, coverage, sheet, group1, group2, group3, filter1, filter2, filter3",
        "Per state per run", "LOW",
        "Core dynamic routing — changes per state/filing; keep in Excel"
    ),
    (
        "283 Helper Map", "Rule-Specific Maps", "Rule 283 — Autos Held for Sale (Non-Dealers)",
        "Configuration table that drives table structure for Rule 283 pages",
        "Varies (table config columns)",
        "4x per run (called from multiple table-building methods)", "MEDIUM",
        "Read 4 times; consider caching the read result to avoid repeated I/O"
    ),
    (
        "298 Coll Map", "Rule-Specific Maps", "Rule 298 — Deductible Factors",
        "Collision deductible factors by state and coverage",
        "state, coverage, factor columns",
        "3x per run", "MEDIUM",
        "Read 3 times; cache the DataFrame; .fillna('Y') is applied on read"
    ),
    (
        "298 OTC Map", "Rule-Specific Maps", "Rule 298 — Deductible Factors",
        "Other Than Collision deductible factors by state and coverage",
        "state, coverage, factor columns",
        "3x per run", "MEDIUM",
        "Read 3 times; cache the DataFrame; .fillna('Y') is applied on read"
    ),
    (
        "298 OTC FG Map", "Rule-Specific Maps", "Rule 298 — Deductible Factors",
        "Full Glass OTC deductible factors by state and coverage",
        "state, coverage, factor columns",
        "3x per run", "MEDIUM",
        "Read 3 times; cache the DataFrame; .fillna('Y') is applied on read"
    ),
    (
        "298 FG States", "Rule-Specific Maps", "Rule 298 — Full Glass",
        "List of states where Full Glass coverage is available",
        "states",
        "2x per run", "MEDIUM-HIGH",
        "Small list; could be a Python constant if FG state list is stable"
    ),
    (
        "266 Coll Map", "Rule-Specific Maps", "Rule 266 — Antique Autos",
        "Collision deductible factor mapping for Antique Auto coverage",
        "state, coverage, factor columns",
        "Once per run", "MEDIUM",
        "Filing-specific; acceptable to keep in Excel"
    ),
    (
        "266 OTC Map", "Rule-Specific Maps", "Rule 266 — Antique Autos",
        "OTC deductible factor mapping for Antique Auto coverage",
        "state, coverage, factor columns",
        "Once per run", "MEDIUM",
        "Filing-specific; acceptable to keep in Excel"
    ),
    (
        "297 Map", "Rule-Specific Maps", "Rule 297 — UM/UIM Coverage",
        "Drives UM/UIM coverage table structure, clustering, and naming by state",
        "state, coverage structure columns",
        "4x per run (4 different rule variants)", "MEDIUM",
        "Read 4 times — hardcoded literal on line ~10327; should use BA_INPUT_FILE constant"
    ),
    (
        "239 School Buses", "Classification Tables", "Rule 239 — Bus/Vehicle Classification",
        "School bus classification and rate data",
        "classification, rate columns (accessed via .iloc[])",
        "Once per run", "MEDIUM-HIGH",
        "Embed as Python dict/DataFrame constant if classification doesn't change between filings"
    ),
    (
        "239 Other Buses", "Classification Tables", "Rule 239 — Bus/Vehicle Classification",
        "Other bus types classification and rate data",
        "classification, rate columns (accessed via .iloc[])",
        "Once per run", "MEDIUM-HIGH",
        "Embed as Python dict/DataFrame constant if classification doesn't change between filings"
    ),
    (
        "239 Van Pools", "Classification Tables", "Rule 239 — Bus/Vehicle Classification",
        "Van pool vehicle classification and rate data",
        "classification, rate columns (accessed via .iloc[])",
        "Once per run", "MEDIUM-HIGH",
        "Embed as Python dict/DataFrame constant if classification doesn't change between filings"
    ),
    (
        "239 Taxis", "Classification Tables", "Rule 239 — Bus/Vehicle Classification",
        "Taxi and limousine classification and rate data",
        "classification, rate columns (accessed via .iloc[])",
        "Once per run", "MEDIUM-HIGH",
        "Embed as Python dict/DataFrame constant if classification doesn't change between filings"
    ),
    (
        "CSL Limits", "Limits Tables", "Rule 295 — CSL Liability Limits",
        "Commercial Single Limit liability limit options by state",
        "limit values, codes",
        "Once per run", "LOW",
        "State-mandated limits change frequently; keep in Excel"
    ),
    (
        "UM-UIM-UMPD Limits", "Limits Tables", "Rules 297 / UIM",
        "UM, UIM, and UMPD limit options — multi-level header structure",
        "Multi-level columns: UM, UIM, UMPD limit tiers",
        "2x per run", "LOW",
        "Read with header=[0,1] — complex structure; state-mandated; keep in Excel"
    ),
    (
        "293 Map", "Specialized / State", "Rule 293 — No-Fault Coverage",
        "Drives dynamic table generation for No-Fault coverage per state",
        "state-specific columns",
        "Once per run", "LOW",
        "Read with engine='openpyxl'; complex per-state logic; keep in Excel"
    ),
    (
        "452 Leaf", "Specialized / State", "Rule 452 — Lifetime Expense Factor",
        "Lifetime Expense Allocation Factor data — only loaded if show_leaf flag is True",
        "factor columns",
        "Conditional — only if show_leaf=True", "MEDIUM",
        "Gate-flagged; rarely used; could be embedded or skipped entirely if flag is always False"
    ),
    (
        "CT R1", "Specialized / State", "Connecticut-specific Rule R1",
        "CT-only rate order calculation data — raw read with no headers",
        "Raw columns (header=None)",
        "Once per run (CT only)", "LOW",
        "State-specific and complex; keep in Excel. Only runs for Connecticut"
    ),
]

CATEGORY_SUMMARY = [
    # (Category, Sheet Count, Reduction Potential, Strategy, Priority)
    ("State Config",         2,  "HIGH",        "Replace with Python constant lists",                                          1),
    ("Classification Tables",4,  "MEDIUM-HIGH", "Embed as Python dicts if stable across filings",                              2),
    ("LCM / Company Dev",    2,  "MEDIUM",      "Move to YAML/JSON config; cache reads at startup",                            3),
    ("Rule-Specific Maps",   7,  "MEDIUM",      "Cache DataFrames to avoid repeated reads; keep in Excel for filing flexibility",4),
    ("Rate Sheet Routing",   3,  "LOW",         "Core routing logic — keep in Excel, changes per state/filing",                 5),
    ("Limits Tables",        2,  "LOW",         "State-mandated values — keep in Excel",                                       5),
    ("Specialized / State",  3,  "LOW",         "Complex state-specific logic — keep in Excel",                                5),
]

BUG_DATA = [
    (
        "Hardcoded filename", "BARates.py", "~10327",
        "pd.read_excel('BA Input File.xlsx', sheet_name=\"297 Map\", engine='openpyxl')",
        "Replace 'BA Input File.xlsx' literal with BA_INPUT_FILE constant",
        "MEDIUM"
    ),
    (
        "Repeated reads (no cache)", "BARates.py", "Multiple",
        "Sheets 283 Helper Map, 298 Coll/OTC/OTC-FG Map, 297 Map each read 3-4x",
        "Read once into a dict at startup, e.g. self._maps = {}; reuse cached DataFrame",
        "HIGH"
    ),
]

# ── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def bold_font(size=11, color="000000", italic=False):
    return Font(bold=True, size=size, color=color, italic=italic)

def normal_font(size=10, color="000000"):
    return Font(size=size, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    side = Side(style="thin", color="AAAAAA")
    return Border(left=side, right=side, top=side, bottom=side)

def thick_border():
    side = Side(style="medium", color="2E75B6")
    return Border(left=side, right=side, top=side, bottom=side)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def write_cell(ws, row, col, value, font=None, fill_=None, align=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:   cell.font   = font
    if fill_:  cell.fill   = fill_
    if align:  cell.alignment = align
    if border: cell.border = border
    return cell

def merge_header(ws, row, col_start, col_end, value, bg, fg="FFFFFF", font_size=14):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    cell = ws.cell(row=row, column=col_start, value=value)
    cell.fill      = fill(bg)
    cell.font      = bold_font(size=font_size, color=fg)
    cell.alignment = center()
    cell.border    = thick_border()

# ── Sheet 1 — Cover ───────────────────────────────────────────────────────────
def build_cover(wb):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height  = 20
    ws.row_dimensions[2].height  = 60
    ws.row_dimensions[3].height  = 30
    ws.row_dimensions[4].height  = 20
    ws.row_dimensions[5].height  = 25
    ws.row_dimensions[6].height  = 25
    ws.row_dimensions[7].height  = 25
    ws.row_dimensions[8].height  = 25
    ws.row_dimensions[9].height  = 25
    ws.row_dimensions[10].height = 25
    ws.row_dimensions[11].height = 25
    ws.row_dimensions[12].height = 40
    ws.row_dimensions[13].height = 20

    for c in ["A","B","C","D"]:
        ws.column_dimensions[c].width = 30

    merge_header(ws, 2, 1, 4,
                 "BA Input File.xlsx — Dependency Reference Document",
                 C_HEADER_DARK, font_size=18)

    merge_header(ws, 3, 1, 4,
                 "Business Auto (BA) Rate Pages System  |  Codebase: BA/BARates.py",
                 C_HEADER_MED, font_size=12)

    # Stats table
    stats = [
        ("Total Sheets in BA Input File",   "25"),
        ("Categories",                       "7"),
        ("HIGH reduction potential sheets",  "2"),
        ("MEDIUM-HIGH potential sheets",     "5"),
        ("MEDIUM potential sheets",          "9"),
        ("LOW potential sheets",             "8"),
        ("Known bugs / issues found",        "2"),
    ]
    ws.row_dimensions[4].height = 12

    for i, (label, val) in enumerate(stats, start=5):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font      = bold_font(size=11)
        lc.fill      = fill("D6E4F0")
        lc.alignment = left()
        lc.border    = thin_border()

        vc = ws.cell(row=i, column=2, value=val)
        vc.font      = bold_font(size=11, color=C_HEADER_DARK)
        vc.fill      = fill("EBF5FB")
        vc.alignment = center()
        vc.border    = thin_border()

        for col in [3, 4]:
            ec = ws.cell(row=i, column=col)
            ec.fill   = fill("F8FBFE")
            ec.border = thin_border()

    # Navigation guide
    nav_row = 13
    merge_header(ws, nav_row, 1, 4, "Navigation Guide", C_HEADER_MED, font_size=12)

    nav_items = [
        ("Master Overview",    "Full list of all 25 sheets with details"),
        ("Category Summary",   "Grouped view + reduction strategy per category"),
        ("Code References",    "Which Python function reads each sheet"),
        ("Reduction Roadmap",  "Prioritised action plan to reduce dependencies"),
        ("Bugs & Issues",      "Identified problems and quick-win fixes"),
    ]
    for j, (sheet, desc) in enumerate(nav_items, start=14):
        ws.row_dimensions[j].height = 22
        sc = ws.cell(row=j, column=1, value=sheet)
        sc.font      = bold_font(size=10, color=C_HEADER_DARK)
        sc.fill      = fill(C_HEADER_LIGHT)
        sc.alignment = center()
        sc.border    = thin_border()

        dc = ws.cell(row=j, column=2, value=desc)
        dc.font      = normal_font(size=10)
        dc.fill      = fill("FAFAFA")
        dc.alignment = left()
        dc.border    = thin_border()

        for col in [3, 4]:
            ec = ws.cell(row=j, column=col)
            ec.fill   = fill("FAFAFA")
            ec.border = thin_border()

# ── Sheet 2 — Master Overview ─────────────────────────────────────────────────
def build_master(wb):
    ws = wb.create_sheet("Master Overview")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Column widths
    widths = [26, 20, 32, 44, 36, 18, 16, 48]
    headers = [
        "Sheet Name", "Category", "Rule / Function",
        "What It Provides", "Key Columns Used",
        "Read Frequency", "Reduction Potential", "Action / Notes"
    ]
    col_letters = [get_column_letter(i+1) for i in range(len(widths))]
    for letter, w in zip(col_letters, widths):
        ws.column_dimensions[letter].width = w

    # Title row
    ws.row_dimensions[1].height = 36
    merge_header(ws, 1, 1, len(headers),
                 "Master Overview — All BA Input File.xlsx Sheet Dependencies",
                 C_HEADER_DARK, font_size=14)

    # Column headers
    ws.row_dimensions[2].height = 30
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font      = bold_font(size=10, color="FFFFFF")
        c.fill      = fill(C_HEADER_MED)
        c.alignment = center()
        c.border    = thin_border()

    # Data rows
    prev_cat = None
    for row_idx, row_data in enumerate(MASTER_DATA, start=3):
        ws.row_dimensions[row_idx].height = 42
        (sheet_name, category, rule_func, what_provides,
         key_cols, read_freq, potential, notes) = row_data

        # Alternate category row shading
        cat_color = C_CAT_COLORS.get(category, "FFFFFF")
        pot_color = POTENTIAL_COLOR.get(potential, "AAAAAA")

        values = [sheet_name, category, rule_func, what_provides,
                  key_cols, read_freq, potential, notes]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.alignment = left() if col_idx not in [6, 7] else center()
            c.border    = thin_border()

            if col_idx == 7:   # Reduction Potential — colored badge
                c.fill = fill(pot_color)
                c.font = bold_font(size=10, color="FFFFFF")
                c.alignment = center()
            elif col_idx == 1:
                c.fill = fill(cat_color)
                c.font = bold_font(size=10)
            elif col_idx == 2:
                c.fill = fill(cat_color)
                c.font = normal_font(size=10)
            else:
                # Alternate row shading
                bg = "F7FBFF" if row_idx % 2 == 0 else "FFFFFF"
                c.fill = fill(bg)
                c.font = normal_font(size=10)

    # Legend below table
    legend_row = len(MASTER_DATA) + 4
    ws.row_dimensions[legend_row - 1].height = 20
    ws.merge_cells(start_row=legend_row-1, start_column=1,
                   end_row=legend_row-1, end_column=len(headers))
    lh = ws.cell(row=legend_row-1, column=1, value="Reduction Potential Legend")
    lh.font      = bold_font(size=11, color=C_HEADER_DARK)
    lh.fill      = fill(C_HEADER_LIGHT)
    lh.alignment = center()

    legends = [
        ("HIGH",        POTENTIAL_COLOR["HIGH"],        "Safe to remove Excel dependency — embed as Python code"),
        ("MEDIUM-HIGH", POTENTIAL_COLOR["MEDIUM-HIGH"], "Likely removable — verify stability across filings first"),
        ("MEDIUM",      POTENTIAL_COLOR["MEDIUM"],      "Keep in Excel but optimise (cache reads, reduce duplication)"),
        ("LOW",         POTENTIAL_COLOR["LOW"],         "Keep in Excel — changes frequently or is state-mandated"),
    ]
    for i, (label, color, desc) in enumerate(legends, start=legend_row):
        ws.row_dimensions[i].height = 20
        lc = ws.cell(row=i, column=1, value=label)
        lc.fill = fill(color); lc.font = bold_font(10, "FFFFFF"); lc.alignment = center()
        lc.border = thin_border()
        dc = ws.cell(row=i, column=2, value=desc)
        dc.font = normal_font(10); dc.alignment = left(); dc.border = thin_border()
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=len(headers))

# ── Sheet 3 — Category Summary ────────────────────────────────────────────────
def build_category(wb):
    ws = wb.create_sheet("Category Summary")
    ws.sheet_view.showGridLines = False

    col_widths = [22, 14, 18, 60, 12]
    col_letters = [get_column_letter(i+1) for i in range(5)]
    for letter, w in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = w

    ws.row_dimensions[1].height = 36
    merge_header(ws, 1, 1, 5,
                 "Category Summary — Grouped View & Reduction Strategy",
                 C_HEADER_DARK, font_size=14)

    headers = ["Category", "Sheet Count", "Overall Potential", "Reduction Strategy", "Priority"]
    ws.row_dimensions[2].height = 30
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = bold_font(10, "FFFFFF"); c.fill = fill(C_HEADER_MED)
        c.alignment = center(); c.border = thin_border()

    for row_idx, (cat, count, potential, strategy, priority) in enumerate(CATEGORY_SUMMARY, start=3):
        ws.row_dimensions[row_idx].height = 40
        cat_color = C_CAT_COLORS.get(cat, "FFFFFF")
        pot_color = POTENTIAL_COLOR.get(potential, "AAAAAA")

        vals = [cat, count, potential, strategy, priority]
        for col_idx, val in enumerate(vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border()
            c.alignment = center() if col_idx in [2, 3, 5] else left()

            if col_idx == 1:
                c.fill = fill(cat_color); c.font = bold_font(10)
            elif col_idx == 3:
                c.fill = fill(pot_color); c.font = bold_font(10, "FFFFFF"); c.alignment = center()
            else:
                bg = "F7FBFF" if row_idx % 2 == 0 else "FFFFFF"
                c.fill = fill(bg); c.font = normal_font(10)

    # Sheet breakdown per category
    br_row = len(CATEGORY_SUMMARY) + 5
    merge_header(ws, br_row - 1, 1, 5,
                 "Sheet Breakdown by Category", C_HEADER_MED, font_size=12)

    cat_groups = {}
    for row in MASTER_DATA:
        cat_groups.setdefault(row[1], []).append(row[0])

    current_row = br_row
    for cat, sheets in cat_groups.items():
        cat_color = C_CAT_COLORS.get(cat, "FFFFFF")
        ws.row_dimensions[current_row].height = 22
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=5)
        hc = ws.cell(row=current_row, column=1, value=f"  {cat}")
        hc.fill = fill(cat_color); hc.font = bold_font(11); hc.border = thin_border()
        current_row += 1

        for sheet in sheets:
            ws.row_dimensions[current_row].height = 20
            sc = ws.cell(row=current_row, column=1, value=f"     • {sheet}")
            sc.font = normal_font(10); sc.fill = fill("FAFAFA")
            sc.border = thin_border()
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=5)
            current_row += 1

# ── Sheet 4 — Code References ─────────────────────────────────────────────────
def build_code_refs(wb):
    ws = wb.create_sheet("Code References")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    col_widths = [26, 36, 16, 24, 20, 42]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    merge_header(ws, 1, 1, 6,
                 "Code References — Which Function Reads Each Sheet",
                 C_HEADER_DARK, font_size=14)

    headers = ["Sheet Name", "Python Function / Method",
               "Approx. Line(s)", "Read Engine",
               "Times Read per Run", "Special Handling / Notes"]
    ws.row_dimensions[2].height = 30
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = bold_font(10, "FFFFFF"); c.fill = fill(C_HEADER_MED)
        c.alignment = center(); c.border = thin_border()

    code_data = [
        ("No MedPay",          "__init__",                      "111",       "pandas (default)", "1",  "Stored in self.no_med_states; filtered with .loc[]"),
        ("PIP States",         "__init__",                      "112",       "pandas (default)", "1",  "Stored in self.pip_states"),
        ("LCM-Comp Dev Mapping","__init__",                     "187",       "pandas (default)", "1",  "Iterates rows 0 to shape[0]; drives LCM application loop"),
        ("LCM-Map",            "__init__",                      "188",       "pandas (default)", "1",  "Lookup via iloc[0, 2] — reads 3rd column for multiplier type"),
        ("222 TTT",            "buildBaseRates()",              "3025",      "pandas (default)", "Per state", "Filtered by self.StateAbb; columns: state, coverage, sheet, group1-3, filter1-3"),
        ("232 PPT",            "buildBaseRates()",              "3026",      "pandas (default)", "Per state", "Filtered by self.StateAbb; same column pattern as 222 TTT"),
        ("225 Zone Rated",     "buildZoneBaseRates()",          "1262",      "pandas (default)", "Per state", "Same routing pattern; zone-rated risks only"),
        ("283 Helper Map",     "Multiple table-build methods",  "1901, 2029, 2049, 2070", "pandas (default)", "4", "Called 4 separate times — same sheet read 4x (cache opportunity)"),
        ("298 Coll Map",       "Rule 298 deductible methods",   "2470, 2567, 2674", "pandas (default)", "3", ".fillna('Y') applied; read 3x (cache opportunity)"),
        ("298 OTC Map",        "Rule 298 deductible methods",   "2471, 2568, 2675", "pandas (default)", "3", ".fillna('Y') applied; read 3x (cache opportunity)"),
        ("298 OTC FG Map",     "Rule 298 deductible methods",   "2472, 2569, 2676", "pandas (default)", "3", ".fillna('Y') applied; read 3x (cache opportunity)"),
        ("298 FG States",      "Rule 298 Full Glass check",     "2537, 2645, 2754", "pandas (default)", "2", "State list — small; read 2x"),
        ("266 Coll Map",       "Rule 266 Antique Auto",         "3671",      "pandas (default)", "1",  "Filing-specific deductible map"),
        ("266 OTC Map",        "Rule 266 Antique Auto",         "3672",      "pandas (default)", "1",  "Filing-specific deductible map"),
        ("297 Map",            "Rule 297 UM/UIM (4 variants)",  "4564, 5758, 5910, 10327", "openpyxl (line 10327)", "4", "BUG: line 10327 uses hardcoded 'BA Input File.xlsx' — should use BA_INPUT_FILE constant"),
        ("239 School Buses",   "Rule 239 classification",       "3021",      "pandas (default)", "1",  "Accessed via .iloc[] — raw positional access"),
        ("239 Other Buses",    "Rule 239 classification",       "3022",      "pandas (default)", "1",  "Accessed via .iloc[]"),
        ("239 Van Pools",      "Rule 239 classification",       "3023",      "pandas (default)", "1",  "Accessed via .iloc[]"),
        ("239 Taxis",          "Rule 239 classification",       "3024",      "pandas (default)", "1",  "Accessed via .iloc[]"),
        ("CSL Limits",         "Rule 295 liability limits",     "4525",      "pandas (default)", "1",  "Limit options and codes"),
        ("UM-UIM-UMPD Limits", "Rules 297 / UIM processing",   "5750, 5902","pandas (default)", "2",  "header=[0,1] — multi-level column index; complex structure"),
        ("293 Map",            "Rule 293 No-Fault",             "10212",     "openpyxl",         "1",  "engine='openpyxl' explicitly; drives dynamic per-state table generation"),
        ("452 Leaf",           "Rule 452 Lifetime Expense",     "10856",     "pandas (default)", "Conditional", "Only read if show_leaf=True for the state"),
        ("CT R1",              "Connecticut Rule R1",           "10944",     "pandas (default)", "1 (CT only)", "header=None — raw read; Connecticut-specific only"),
    ]

    for row_idx, row_vals in enumerate(code_data, start=3):
        ws.row_dimensions[row_idx].height = 38
        cat = next((r[1] for r in MASTER_DATA if r[0] == row_vals[0]), None)
        cat_color = C_CAT_COLORS.get(cat, "FFFFFF")

        for col_idx, val in enumerate(row_vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border()
            if col_idx == 1:
                c.fill = fill(cat_color); c.font = bold_font(10)
                c.alignment = left()
            else:
                bg = "F7FBFF" if row_idx % 2 == 0 else "FFFFFF"
                c.fill = fill(bg)
                c.font = normal_font(10)
                c.alignment = left() if col_idx in [2, 6] else center()

            # Highlight the bug row
            if row_vals[0] == "297 Map" and col_idx == 6:
                c.fill = fill("FFDCE1")
                c.font = bold_font(10, "C00000")

# ── Sheet 5 — Reduction Roadmap ───────────────────────────────────────────────
def build_roadmap(wb):
    ws = wb.create_sheet("Reduction Roadmap")
    ws.sheet_view.showGridLines = False

    col_widths = [8, 20, 36, 36, 18, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    merge_header(ws, 1, 1, 6,
                 "Reduction Roadmap — Prioritised Action Plan",
                 C_HEADER_DARK, font_size=14)

    headers = ["Priority", "Sheets Affected", "Current Behaviour", "Recommended Action", "Effort", "Impact"]
    ws.row_dimensions[2].height = 30
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = bold_font(10, "FFFFFF"); c.fill = fill(C_HEADER_MED)
        c.alignment = center(); c.border = thin_border()

    roadmap = [
        (
            "P1", "No MedPay\nPIP States",
            "Read from Excel at startup — 2 sheets just for state lists",
            "Define as Python constant sets in config/constants.py:\n"
            "NO_MED_STATES = {'AK', 'AL', ...}\n"
            "PIP_STATES = {'FL', 'MI', ...}",
            "Low (1–2 hrs)", "HIGH — eliminates 2 sheet reads"
        ),
        (
            "P2", "283 Helper Map\n298 Coll Map\n298 OTC Map\n298 OTC FG Map\n298 FG States\n297 Map",
            "Same sheet read multiple times per run (3–4 reads each)\n"
            "Total wasted reads: ~18 extra pd.read_excel() calls",
            "Cache all maps at startup into a dict:\n"
            "self._maps = {\n"
            "  '283 Helper Map': pd.read_excel(...),\n"
            "  '298 Coll Map':   pd.read_excel(...),\n"
            "  ...}\n"
            "Then reference self._maps['sheet'] everywhere",
            "Low (2–3 hrs)", "HIGH — reduces I/O from ~40 reads to ~25"
        ),
        (
            "P3", "239 School Buses\n239 Other Buses\n239 Van Pools\n239 Taxis",
            "4 sheets read from Excel for Rule 239 vehicle classification\n"
            "Data accessed with .iloc[] (positional — fragile)",
            "Verify if classification changes between filings.\n"
            "If stable: embed as Python DataFrames or dicts in code.\n"
            "If changes: keep in Excel but switch to named column access",
            "Medium (3–5 hrs)", "MEDIUM — removes 4 Excel sheets"
        ),
        (
            "P4", "LCM-Comp Dev Mapping\nLCM-Map",
            "Two config sheets drive LCM multiplier routing\n"
            "Changes per filing but rarely within a filing",
            "Move to a structured YAML or JSON config file.\n"
            "Keep version-controlled with the code.\n"
            "Eliminates 2 Excel sheets; adds auditability via git diff",
            "Medium (4–6 hrs)", "MEDIUM — better audit trail"
        ),
        (
            "P5", "452 Leaf",
            "Sheet read conditionally (show_leaf flag)\n"
            "May never be used in most state runs",
            "Confirm if show_leaf is ever True in production.\n"
            "If not: remove the read entirely.\n"
            "If yes: cache at startup like other maps",
            "Low (1 hr)", "MEDIUM — removes conditional dead read"
        ),
        (
            "P6", "222 TTT\n232 PPT\n225 Zone Rated\nCSL Limits\nUM-UIM-UMPD Limits\n293 Map\nCT R1",
            "Core routing and state-mandated limits — these MUST\n"
            "be in Excel as they change per filing / state",
            "No change to storage.\n"
            "Action: ensure all reads go through the BA_INPUT_FILE\n"
            "constant (no hardcoded strings); add read-once caching\n"
            "where called more than once",
            "Very Low (30 min)", "LOW — maintenance improvement only"
        ),
    ]

    priority_colors = {
        "P1": "70AD47", "P2": "70AD47",
        "P3": "FFC000", "P4": "FFC000", "P5": "FFC000",
        "P6": "ED7D31",
    }

    for row_idx, (pri, sheets, current, action, effort, impact) in enumerate(roadmap, start=3):
        ws.row_dimensions[row_idx].height = 90
        bg_color = "F7FBFF" if row_idx % 2 == 0 else "FFFFFF"
        pc = priority_colors.get(pri, "AAAAAA")

        vals = [pri, sheets, current, action, effort, impact]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if col_idx == 1:
                c.fill = fill(pc); c.font = bold_font(12, "FFFFFF"); c.alignment = center()
            elif col_idx in [5, 6]:
                c.fill = fill("EBF5FB"); c.font = normal_font(10)
            else:
                c.fill = fill(bg_color); c.font = normal_font(10)

# ── Sheet 6 — Bugs & Issues ───────────────────────────────────────────────────
def build_bugs(wb):
    ws = wb.create_sheet("Bugs & Issues")
    ws.sheet_view.showGridLines = False

    col_widths = [22, 16, 12, 46, 46, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 36
    merge_header(ws, 1, 1, 6,
                 "Bugs & Issues Found During Dependency Audit",
                 C_HEADER_DARK, font_size=14)

    headers = ["Issue", "File", "Line(s)", "Current Code / Behaviour", "Fix", "Severity"]
    ws.row_dimensions[2].height = 30
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col_idx, value=h)
        c.font = bold_font(10, "FFFFFF"); c.fill = fill(C_HEADER_MED)
        c.alignment = center(); c.border = thin_border()

    severity_colors = {"HIGH": "FF0000", "MEDIUM": "FFC000", "LOW": "70AD47"}

    for row_idx, (issue, file_, lines, current, fix, severity) in enumerate(BUG_DATA, start=3):
        ws.row_dimensions[row_idx].height = 70
        sc = severity_colors.get(severity, "AAAAAA")
        bg = "FFF8F8" if row_idx % 2 == 0 else "FFFFFF"

        vals = [issue, file_, lines, current, fix, severity]
        for col_idx, val in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = thin_border()
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if col_idx == 6:
                c.fill = fill(sc); c.font = bold_font(10, "FFFFFF"); c.alignment = center()
            elif col_idx in [2, 3]:
                c.fill = fill("EBF5FB"); c.font = Font(name="Courier New", size=9)
            elif col_idx in [4, 5]:
                c.fill = fill(bg); c.font = Font(name="Courier New", size=9)
            else:
                c.fill = fill(bg); c.font = bold_font(10)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()

    build_cover(wb)
    build_master(wb)
    build_category(wb)
    build_code_refs(wb)
    build_roadmap(wb)
    build_bugs(wb)

    out = "BA_Input_File_Dependency_Map.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == "__main__":
    main()
