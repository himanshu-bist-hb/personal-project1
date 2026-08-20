# This module builds and formats the Office State Page workbook (BP-2.0).
#
# The build*() methods below are unchanged business logic, transcribed from
# the root-level OfficePage.py: each pulls a table out of the ratebook data
# (via buildDataFrame's nesting waterfall) and shapes it into the DataFrame
# the rate page needs, filtered to the Office program (Class_Code_Min ==
# 60000). Same porting pattern as [[bop_hab_autoservice_port]],
# [[bop_retail_port]] and [[bop_service_port]]: the [[bop-nesting-order]] fix
# (lower-level company -> NGIC -> CW, not NGIC-first), the per-company
# base-rate fix (root hardcoded buildBaseRates('NGIC') for every company's
# BRNACO/BRNAFF/BRNICOF tab — same bug already found and fixed in Hab/Auto
# Service/Retail/Service), the "AllPeril" row/column addition for individual
# program pages, and the broader liability-peril exclusion list (drop
# L-Products/L-Violence/L-OtherMed/L-OtherPrem/WF/NC-BINC, rename L-SlipFall
# -> LIAB-Other) already applied to Retail/Service's Construction
# Factor/Theft Options/Year Built Modifier tables — the root-level OfficePage.py
# still only dropped 'L-Products' by itself, an older/narrower version of the
# same fix.
#
# All Excel formatting (fonts, column widths, sub-header labels, page setup)
# lives in ExcelSettingsBOP.py, driven by "BOP/BOP Input File.xlsx". Most of
# Office's tables reuse table codes/layout profiles already shared with Hab,
# Auto Service, Retail and Service (CBG/CPP/YBBG/YBPP/EBB/PDLD/LL/DO/DONM/ERP/
# PLUS/FR/AS_BR/LS_RETAIL/PSS). Office-only tables (OPTO/VET/AES/VPL/
# PSPL_OFFICE) get new layout keys; several single-value "Base Premium for
# each Office Premises" endorsements (CS/PFSS/ACS/ATS/HCS) reuse the AES key
# since they share its exact 300px single-column shape (same convention as
# Retail/Service's PSPL reusing the PED key). The Directors & Officers 2-band
# merge reuses the exact same post-processing as Retail/Service's (Office
# only has "Under 51"/"51 or More", unlike Hab's 5 bands).
#
# Office adds several BP-2.0-only endorsements with no Retail equivalent but
# a close Service one: Pet Services Specialized Endorsement (PSS, a two-table
# sheet, identical shape to Retail/Service's PSS) and Mobile Pet and
# Veterinarian Services Endorsement (MPVS, a SIX-table sheet, content-identical
# to Service's MPVS — same Pet Services / Veterinarian blocks). Office also
# has veterinary content Service does NOT: a standalone Veterinarian
# Specialized Endorsement summary (VSPL), a Veterinarian Professional
# Liability table (VET), a two-table Veterinarian Specialized Endorsement
# (VS, same "base premium + business income" shape as PSS), and a two-peer-
# table Veterinarian Professional Liability - Household/Non-Household sheet
# (VPL). Like Retail's PSS, none of these sheets' underlying root-level
# generateWorksheet2tables/6tables (ExcelSettings.py) module is present in
# this repo, so _appendLabeledBlocks reconstructs the same semantic layout
# (label, header, data, repeated per block with a blank separator row
# between) from scratch. Unlike Service's own _appendLabeledBlocks, the
# section label here is NOT bordered — a feedback fix from Retail's PSS
# table (R Table 4.E): the label should read as a plain bold caption, not a
# tightly-boxed cell (see [[bop_retail_pss_border_width_fix]]).

from copy import copy

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import ExcelSettingsBOP

_THIN_BORDER = Border(
    left=Side(border_style='thin', color='C1C1C1'),
    right=Side(border_style='thin', color='C1C1C1'),
    top=Side(border_style='thin', color='C1C1C1'),
    bottom=Side(border_style='thin', color='C1C1C1'),
)


class Office:
    # Tables buildBaseRates() needs present for a given company before it can
    # be built without a KeyError.
    _BASE_RATE_TABLES = (
        'BP7_Peril_Building_Base_Rates',
        'BP7_Peril_BPP_Base_Rates',
        'BP7_Peril_Liability_Base_Rates',
    )

    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        # Individual program pages (unlike All Programs) also show the
        # "AllPeril" row/column — see [[bop_auto_service_allperil_row_fix]].
        self.perils = list(perils) + ['allperil']
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective
        self.rEffective = rEffective

        self.officeProgramCode = 60000

    # Builds a dataframe for the given table code
    # The hierarchy matches Business Auto: lower-level company (NACO/NAFF/NICOF)
    # first, then NGIC (state-level default), then CW as the country-wide
    # fallback. See [[bop-nesting-order]] — the root-level OfficePage.py
    # checked NGIC first, which is backwards.
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if 'NACO' in self.rateTables.keys():
            if tableCode in self.rateTables['NACO'].keys():
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys():
            if tableCode in self.rateTables['NAFF'].keys():
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys():
            if tableCode in self.rateTables['NICOF'].keys():
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if tableCode in self.rateTables['NGIC'].keys():
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0])

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building",
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction'). \
                    drop(columns=['L-Products', 'L-Violence', 'L-OtherMed', 'L-OtherPrem', 'WF', 'NC-BINC'], errors='ignore').rename(columns={'L-SlipFall': 'LIAB-Other'})
        elif coverage.casefold() == 'bpp':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction'). \
                    drop(columns=['L-Products', 'L-Violence', 'L-OtherMed', 'L-OtherPrem', 'WF', 'NC-BINC'], errors='ignore').rename(columns={'L-SlipFall': 'LIAB-Other'})

    # Builds the exclude theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'].isin(['L-Products', 'NC-BINC', 'WF'])].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.officeProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'})
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop(columns=['L-Products', 'WF', 'NC-BINC'], errors='ignore')
        elif coverage.casefold() == 'bpp':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop(columns=['L-Products', 'WF', 'NC-BINC'], errors='ignore')

    # Builds the equipment breakdown base rate table
    # Returns a dataframe
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.officeProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.officeProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_BLDG_BPP_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCodeMin == {self.officeProgramCode}').rename(columns={'BldgBPPLimit_Min': 'Limit Min', 'BldgBPPLimit_Max': 'Limit Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Limit Min', 'Limit Max'], columns='Peril TypeCode', values='Factor').reset_index(['Limit Min', 'Limit Max']).fillna({'Limit Max': 'and over'})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.officeProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance': 'int32'})

    # Builds the general liability occupancy modifiers table
    # Returns a dataframe
    def buildGeneralOccupancyMod(self):
        generalOccupancyModifier = self.buildDataFrame("BP7_Peril_Occupant_Factor")
        filteredGeneralOccupancyMod = generalOccupancyModifier.query(f'ClassCode_Min == {self.officeProgramCode} & `Peril TypeCode` == "liability1"'). \
                rename(columns={'OccupancyType': 'Occupancy', 'BLDGOccupantFactor': 'Building', 'BPPOccupantFactor': 'Business Personal Property'})
        return filteredGeneralOccupancyMod.replace({'Occupancy': {'Condominium': 'Condo Unit-owner', 'buildingOwnerLessorsrisk': "Lessor's Risk", 'buildingOwnerOccupant': 'Owner Occupant', 'tenant': 'Tenant'}}). \
                filter(items=['Occupancy', 'Building', 'Business Personal Property'])

    # Builds the directors and officers liability insurance table
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Office"').copy()
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 or More'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])

    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Office"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Office"').copy()
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index=filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])

    # Builds the optometrists professional liability table
    # Returns a dataframe
    def buildOptometristsProfessionalLiab(self):
        optometristsProfessionalLiab = self.buildDataFrame("BP7_ProfLiabOptical_Rate")
        optometristsProfessionalLiab['LiabilityAmount'] = optometristsProfessionalLiab['LiabilityAmount'].apply(lambda x: "${0:,.0f}".format(x))
        optometristsProfessionalLiab['AggregateLimit'] = optometristsProfessionalLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        optometristsProfessionalLiab['Occurrence / Aggregate'] = optometristsProfessionalLiab['LiabilityAmount'] + ' / ' + optometristsProfessionalLiab['AggregateLimit']
        return optometristsProfessionalLiab.rename(columns={'OptometristRate': 'Each Optometrist', 'OpticianRate': 'Each Optician'}).filter(items=['Occurrence / Aggregate', 'Each Optometrist', 'Each Optician'])

    # Builds the veterinarian specialized endorsement with professional
    # liability summary table (VSPL) — flat premium, hardcoded in the root
    # tool rather than pulled from the ratebook
    # Returns a dataframe
    def buildVetSpecializedLiab(self):
        data = [
            ("$25,000", "$32", "$17"),
            ("$50,000", "$49", "$32"),
            ("$100,000", "$65", "$49"),
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the veterinarians professional liability table
    # Returns a dataframe
    def buildVetProfessionalLiab(self):
        vetProfessionalLiab = self.buildDataFrame("BP7_Pol_VeterinariansExtraCov_Rate")
        vetProfessionalLiab['LiabLimitPerOcc'] = vetProfessionalLiab['LiabLimitPerOcc'].apply(lambda x: "${0:,.0f}".format(x))
        vetProfessionalLiab['LiabLimitPerAgg'] = vetProfessionalLiab['LiabLimitPerAgg'].apply(lambda x: "${0:,.0f}".format(x))
        vetProfessionalLiab['Occurrence / Aggregate'] = vetProfessionalLiab['LiabLimitPerOcc'] + ' / ' + vetProfessionalLiab['LiabLimitPerAgg']
        return vetProfessionalLiab.rename(columns={'Otherthanhouseholdanimals': 'Other than household animals', 'HouseholdAnimalsOnly': 'Household animals only'}). \
                filter(items=['Occurrence / Aggregate', 'Other than household animals', 'Household animals only'])

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.officeProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Office premises'})

    # Builds the franchise upgrade endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.officeProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(),
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(),
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Builds the Architects and Engineers Specialized Endorsement table (flat
    # premium, hardcoded in the root tool rather than pulled from the ratebook)
    # Returns a dataframe
    def buildArchitectsEngineersEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$200.00"]})

    # Builds the Consultants Specialized Endorsement table (flat premium)
    # Returns a dataframe
    def buildConsultantSpecializedEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$200.00"]})

    # Builds the Professional Services Specialized Endorsement table (flat premium)
    # Returns a dataframe
    def buildProfessionalServicesEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$125.00"]})

    # Builds the Accountants Specialized Endorsement table (flat premium)
    # Returns a dataframe
    def buildAccountantsSpecializedEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$200.00"]})

    # Builds the Attorneys Specialized Endorsement table (flat premium)
    # Returns a dataframe
    def buildAttorneySpecializedEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$200.00"]})

    # Builds the Health Care Specialized Endorsement table (flat premium plus
    # a per-employee dishonesty-coverage surcharge line)
    # Returns a dataframe
    def buildHealthCareSpecializedEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": [
            "$300.00", "Plus $10 for each additional employee above 5\nfor employee dishonesty coverage",
        ]})

    # Builds the veterinarian specialized endorsement base premium table (the
    # first, single-value block of the VS sheet — see _formatVS)
    # Returns a dataframe
    def buildVeterinarianSpecializedEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$269.00"]})

    # Builds the veterinarian specialized endorsement business income table
    # (the second, appended block of the VS sheet — see _formatVS)
    # Returns a dataframe
    def buildVeterinarianSpecializedEndorsementIncome(self):
        data = [
            ("$25,000", "$32", "$17"),
            ("$50,000", "$49", "$32"),
            ("$100,000", "$65", "$49"),
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the household-pet veterinarian rate table (the first block of
    # the VPL sheet — see _formatVPL)
    # Returns a dataframe
    def buildVetProfLiabHousehold(self):
        data = [
            ("300,000/600,000", "$60"),
            ("500,000/1,000,000", "$69"),
            ("1,000,000/2,000,000", "$86"),
            ("2,000,000/4,000,000", "$175"),
        ]
        return pd.DataFrame(data, columns=["Limits", "Rate"])

    # Builds the non-household-pet veterinarian rate table (the second block
    # of the VPL sheet — see _formatVPL)
    # Returns a dataframe
    def buildVetProfLiabNonHousehold(self):
        data = [
            ("300,000/600,000", "$105"),
            ("500,000/1,000,000", "$118"),
            ("1,000,000/2,000,000", "$135"),
            ("2,000,000/4,000,000", "$225"),
        ]
        return pd.DataFrame(data, columns=["Limits", "Rate"])

    # Builds the Pet Services block of the Mobile Pet and Veterinarian
    # Services Endorsement sheet (MPVS) — Returns a dataframe
    def buildPetServicesMobileEquip(self):
        data = [("$15,000", "$49"), ("$25,000", "$85"), ("$50,000", "$166"), ("$100,000", "$220")]
        return pd.DataFrame(data, columns=["Limits", "Mobile Equipment"])

    # Builds the Pet Services per Customized Vehicle block of MPVS
    # Returns a dataframe
    def buildPetServicesCustomizedVehicle(self):
        data = [("$25,000", "$91", "$46"), ("$50,000", "$104", "$59"), ("$100,000", "$117", "$71")]
        return pd.DataFrame(data, columns=["Limits", "1st Vehicle", "Each Additional Vehicle"])

    # Builds the Pet Services - Business Income block of MPVS
    # Returns a dataframe
    def buildPetServicesBusinessIncome(self):
        data = [("$25,000", "$13", "$7"), ("$50,000", "$20", "$13"), ("$100,000", "$26", "$20")]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the Veterinarian block of MPVS
    # Returns a dataframe
    def buildVetMobileEquip(self):
        data = [("$15,000", "$122"), ("$25,000", "$211"), ("$50,000", "$414"), ("$100,000", "$549")]
        return pd.DataFrame(data, columns=["Limits", "Mobile Equipment"])

    # Builds the Veterinarian Services per Customized Vehicle block of MPVS
    # Returns a dataframe
    def buildVetCustomizedVehicle(self):
        data = [("$25,000", "$227", "$113"), ("$50,000", "$260", "$146"), ("$100,000", "$293", "$179")]
        return pd.DataFrame(data, columns=["Limits", "1st Vehicle", "Each Additional Vehicle"])

    # Builds the Veterinarian Services - Business Income block of MPVS
    # Returns a dataframe
    def buildVetBusinessIncome(self):
        data = [("$25,000", "$32", "$17"), ("$50,000", "$49", "$32"), ("$100,000", "$65", "$49")]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the Pet Services Specialized Endorsement base premium table (the
    # first block of the PSS sheet — see _formatPSSplzdEndo)
    # Returns a dataframe
    def buildPetServicesSpecializedEndorsement(self):
        return pd.DataFrame({"Base Premium for each Office Premises": ["$212.00"]})

    # Builds the Pet Services - Business Income table (the second, appended
    # block of the PSS sheet — see _formatPSSplzdEndo)
    # Returns a dataframe
    def buildPetServicesSpecializedEndorsementIncome(self):
        data = [("$25,000", "$13", "$7"), ("$50,000", "$20", "$13"), ("$100,000", "$26", "$20")]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the table for Pet Services Professional Liability
    # Returns a dataframe
    def buildPetServicesProfLiab(self):
        data = [
            ("300,000/600,000", "$43"),
            ("500,000/1,000,000", "$56"),
            ("1,000,000/2,000,000", "$68"),
            ("2,000,000/4,000,000", "$83"),
        ]
        return pd.DataFrame(data, columns=["Limits", "Rate"])

    # Merges the "Number of Units" column of the D&O table into its 2 bands
    # ("Under 51" / "51 or More" — Office only has 2, unlike Hab's 5).
    # Handled here directly (rather than through the generic Sub Headers
    # config) because it merges DATA rows, not a header row.
    def _formatDirsOfficersLiabIns(self, ws):
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Appends a series of (label, dataframe) blocks below ws's current
    # content, each as a bolded (unboxed) section label row + bolded,
    # bordered column-header row + bordered plain data rows, with one blank
    # separator row between blocks (and, if blank_before_first, before the
    # first block too). Used to reconstruct the root tool's
    # generateWorksheet2tables/6tables (not present in this repo) by hand for
    # VS/VPL/MPVS/PSS.
    #
    # Each row is only bordered/aligned across its OWN block's column count
    # (n_cols), not the sheet's running max column — otherwise a narrower
    # block following a wider one would pick up a stray bordered empty cell
    # on its right, and bestFit below would size that column from blank
    # cells instead of its real content (see [[bop_service_port]]'s
    # _appendLabeledBlocks, which this mirrors).
    #
    # UNLIKE Service's _appendLabeledBlocks, the label row is intentionally
    # left unboxed here — a feedback fix confirmed against Retail's PSS table
    # (R Table 4.E): the section label should read as a plain bold caption
    # above the table, not a tightly-boxed cell (see
    # [[bop_retail_pss_border_width_fix]]).
    def _appendLabeledBlocks(self, ws, boldFont, font, blocks, blank_before_first=False):
        row = ws.max_row + 1
        max_col = 1
        for i, (label, df) in enumerate(blocks):
            if i > 0 or blank_before_first:
                row += 1  # blank separator row, left untouched (no border)
            label_row = row
            header_row = label_row + 1
            n_cols = len(df.columns)
            label_cell = ws.cell(row=label_row, column=1, value=label)
            label_cell.font = boldFont
            for col, name in enumerate(df.columns, start=1):
                ws.cell(row=header_row, column=col, value=name)
            for r_off, (_, data_row) in enumerate(df.iterrows()):
                for col, val in enumerate(data_row, start=1):
                    ws.cell(row=header_row + 1 + r_off, column=col, value=val)
            for col in range(1, n_cols + 1):
                cell = ws.cell(row=header_row, column=col)
                cell.font = boldFont
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            for r_off in range(len(df)):
                for col in range(1, n_cols + 1):
                    cell = ws.cell(row=header_row + 1 + r_off, column=col)
                    cell.font = font
                    cell.border = _THIN_BORDER
                    cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            max_col = max(max_col, n_cols)
            row = header_row + len(df)
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].bestFit = True

    # Appends the Veterinarian Services - Business Income table below the
    # base premium table on the same sheet (widened to span columns A-C via
    # the merge below, matching the ["PSS", 1, 3, 120] width config reused
    # for VS's layout_key).
    def _formatVS(self, ws, boldFont, font):
        ws.merge_cells('A3:C3')
        ws.merge_cells('A4:C4')
        blocks = [("Veterinarian Services - Business Income", self.buildVeterinarianSpecializedEndorsementIncome())]
        self._appendLabeledBlocks(ws, boldFont, font, blocks, blank_before_first=True)

    # Builds the two-table Veterinarian Professional Liability - Household /
    # Non-Household sheet (VPL) from scratch — generateWorksheet is called
    # with an EMPTY dataframe for this table code (just the title in A1),
    # same pattern as MPVS below.
    def _formatVPL(self, ws, boldFont, font):
        blocks = [
            ("Rate per Veterinarian - Household Pet", self.buildVetProfLiabHousehold()),
            ("Rate per Veterinarian - Non Household Pet", self.buildVetProfLiabNonHousehold()),
        ]
        self._appendLabeledBlocks(ws, boldFont, font, blocks)

    # Mobile Pet and Veterinarian Services Endorsement (MPVS) — a SIX-table
    # sheet, content-identical to Service's own MPVS (same Pet Services /
    # Veterinarian blocks — Office's Class Code just happens to also carry
    # this endorsement). generateWorksheet is called with an EMPTY dataframe
    # for this table code (just the title in A1), and this method builds the
    # entire body from row 3 down via _appendLabeledBlocks.
    #
    # Like Service's _formatMPVS, this is unverified against a real
    # reference (no real Office ratebook to test against yet) — flag to the
    # user to check the actual PDF output for O Table 4.N. the first time
    # real data is available. See [[bop_service_port]]'s _formatMPVS for the
    # same caveat.
    def _formatMPVS(self, ws, boldFont, font):
        blocks = [
            ("Pet Services", self.buildPetServicesMobileEquip()),
            ("Pet Services per Customized Vehicle", self.buildPetServicesCustomizedVehicle()),
            ("Pet Services - Business Income", self.buildPetServicesBusinessIncome()),
            ("Veterinarian", self.buildVetMobileEquip()),
            ("Veterinarian Services per Customized Vehicle", self.buildVetCustomizedVehicle()),
            ("Veterinarian Services - Business Income", self.buildVetBusinessIncome()),
        ]
        self._appendLabeledBlocks(ws, boldFont, font, blocks)

    # Appends the Pet Services - Business Income table below the base
    # premium table on the same sheet (widened to span columns A-C, same
    # merge/shape as VS's above and Retail/Service's PSS).
    def _formatPSSplzdEndo(self, ws, boldFont, font):
        ws.merge_cells('A3:C3')
        ws.merge_cells('A4:C4')
        blocks = [("Pet Services - Business Income", self.buildPetServicesSpecializedEndorsementIncome())]
        self._appendLabeledBlocks(ws, boldFont, font, blocks, blank_before_first=True)

    # Sets up the Office Excel file and creates a separate worksheet for
    # each of the given dataframes. progress_callback (optional) is called
    # with a short message before each sheet is built.
    # Returns the Excel workbook
    def buildOfficePage(self, progress_callback=None):
        companies = [c for c in self.rateTables.keys() if c != 'CW']

        Office = ExcelSettingsBOP.Excel(state=self.state, programName='Office', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        sheetSpecs = []
        # A company can be present in rateTables (its ratebook was uploaded)
        # without having filed its own base-rate tables — a deviation
        # ratebook may only override a handful of tables. Check for the
        # specific tables buildBaseRates() needs, not just company
        # membership, or it KeyErrors on that company's missing table.
        for company, tab, label in (('NACO', 'BRNACO', 'NW Assurance'), ('NAFF', 'BRNAFF', 'NW Affinity'),
                                     ('NGIC', 'BRNGIC', 'NW General Insurance Company'), ('NICOF', 'BRNICOF', 'NICOF')):
            if company in self.rateTables and all(t in self.rateTables[company] for t in self._BASE_RATE_TABLES):
                sheetSpecs.append((tab, f'O Table 3.B.1. {label} State Base Rates', lambda c=company: self.buildBaseRates(c), False, True, 'AS_BR', None))

        sheetSpecs += [
            ('CBG', 'O Table 3.C.2.c. Construction Factor - Building', lambda: self.buildConstructionType('Building'), False, True, None, None),
            ('CPP', 'O Table 3.C.2.c. Construction Factor - BPP', lambda: self.buildConstructionType('BPP'), False, True, None, None),
            ('ET', 'O Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions, False, True, None, None),
            ('YBBG', 'O Table 3.C.2.p. Year Built Modifier - Building', lambda: self.buildYearBuiltModifier('Building'), False, True, None, None),
            ('YBPP', 'O Table 3.C.2.p. Year Built Modifier - BPP', lambda: self.buildYearBuiltModifier('BPP'), False, True, None, None),
            ('EBB', 'O Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate, False, True, None, None),
            ('PDLD', 'O Table 3.C.4.b. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount, False, True, None, None),
            ('LS', 'O Table 3.C.4.d. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk, False, True, 'LS_RETAIL', None),
            ('LL', 'O Table 3.C.4.e. Liability Limit Factor', self.buildLiabilityLimitFactor, False, True, None, None),
            ('GLO', 'O Table 3.D.1.c. General Liability Occupancy Modifiers', self.buildGeneralOccupancyMod, False, True, None, None),
            ('DO', 'O Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns, False, True, None, self._formatDirsOfficersLiabIns),
            ('DONM', 'O Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief, False, True, None, None),
            ('ERP', 'O Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods, False, True, None, None),
            ('OPTO', 'O Table 4.B.5.a. Optometrists Professional Liability', self.buildOptometristsProfessionalLiab, False, True, None, None),
            ('VSPL', 'O Table 4.C.4.A. Veterinarian Specialized Endorsement With Professional Liability', self.buildVetSpecializedLiab, False, True, 'PSS', None),
            ('VET', 'O Table 4.C.5.a. Veterinarians Professional Liability', self.buildVetProfessionalLiab, False, True, None, None),
            ('PLUS', 'O Table 4.D. Office PLUS Endorsement', self.buildEndorsementCharge, False, True, None, None),
            ('FR', 'O Table 4.E. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement, False, True, None, None),
            ('AES', 'O Table 4.F. Architects and Engineers Specialized Endorsement', self.buildArchitectsEngineersEndorsement, False, True, None, None),
            ('CS', 'O Table 4.G. Consultants Specialized Endorsement', self.buildConsultantSpecializedEndorsement, False, True, 'AES', None),
            ('PFSS', 'O Table 4.H. Professional Services Specialized Endorsement', self.buildProfessionalServicesEndorsement, False, True, 'AES', None),
            ('ACS', 'O Table 4.I. Accountants Specialized Endorsement', self.buildAccountantsSpecializedEndorsement, False, True, 'AES', None),
            ('ATS', 'O Table 4.J. Attorneys Specialized Endorsement', self.buildAttorneySpecializedEndorsement, False, True, 'AES', None),
            ('HCS', 'O Table 4.K. Health Care Specialized Endorsement', self.buildHealthCareSpecializedEndorsement, False, True, 'AES', None),
            ('VS', 'O Table 4.L.3. Veterinarian Specialized Endorsement', self.buildVeterinarianSpecializedEndorsement, False, True, 'PSS',
             lambda ws: self._formatVS(ws, Office.fontBold, Office.font)),
            ('VPL', 'O Table 4.M. Veterinarian Professional Liability', lambda: pd.DataFrame(), False, False, None,
             lambda ws: self._formatVPL(ws, Office.fontBold, Office.font)),
            ('MPVS', 'O Table 4.N. Mobile Pet and Veterinarian Services Endorsement', lambda: pd.DataFrame(), False, False, None,
             lambda ws: self._formatMPVS(ws, Office.fontBold, Office.font)),
            ('PSS', 'O Table 4.O. Pet Services Specialized Endorsement', self.buildPetServicesSpecializedEndorsement, False, True, None,
             lambda ws: self._formatPSSplzdEndo(ws, Office.fontBold, Office.font)),
            ('PSPL', 'O Table 4.P. Pet Services Professional Liability', self.buildPetServicesProfLiab, False, True, 'PSPL_OFFICE', None),
        ]

        total = len(sheetSpecs)
        for i, (tableCode, title, build, useIndex, useHeader, layoutKey, postFormat) in enumerate(sheetSpecs, start=1):
            if progress_callback:
                progress_callback(f"Building sheet {i}/{total}: {tableCode}...")
            print(f"  [{i}/{total}] Building sheet: {tableCode}")
            ws = Office.generateWorksheet(tableCode, title, build(), useIndex, useHeader, layout_key=layoutKey)
            if postFormat:
                postFormat(ws)

        if progress_callback:
            progress_callback("Building Index sheet...")
        print(f"  [{total}/{total}] Building sheet: Index")
        Office.createIndex()
        return Office.getWB()
