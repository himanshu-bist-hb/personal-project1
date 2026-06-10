# This module formats the Optional Coverages State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class OptionalCoverages:
    def __init__(self, state, rateTables, perils, classCodes, perilsConversions, nEffective, rEffective, OptionalCoveragesApplies, EQTerritoryDefs, NGICRatebook, NACORatebook, NAFFRatebook, NICOFRatebook, MMRatebook) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.classCodes = classCodes
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date
        self.OptionalCoveragesApplies = OptionalCoveragesApplies
        self.EQTerritoryDefs = EQTerritoryDefs
        self.NGICRatebook = NGICRatebook
        self.NACORatebook = NACORatebook
        self.NAFFRatebook = NAFFRatebook
        self.NICOFRatebook = NICOFRatebook
        self.MMRatebook = MMRatebook
        #self.TerritoryDefsByST = TerritoryDefsByST

        self.currencyFormat = '$#,##0'
        self.currencywdecFormat = '$#,##0.00'
        self.twodecimal = '#,##0.00'
        self.noDecimalFormat = '#,##0'
        self.ZipCodeFormat = '####0'


    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

# Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAccountsReceivableILF(self):
        AccountsReceivableILF = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        AccountsReceivableILF = AccountsReceivableILF.query(f'`CoverageName` == "AccountsReceivable"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
        AccountsReceivableILF = AccountsReceivableILF.iloc[0:]
        AccountsReceivableILF = AccountsReceivableILF.reset_index(drop=True)
        return AccountsReceivableILF

# Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildForgeryILF(self):
        ForgeryILF = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        return ForgeryILF.query(f'`FactorName` == "ForgeryAndAlterationIncreasedLimits"').rename(columns={'Factor': 'Factor'}).filter(items=['Factor'])
    
# Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildMoneyInside(self):
        MoneyInside = self.buildDataFrame("BP7_MoneyandSecuritiesInsideBaseRate")
        MoneyInside = MoneyInside.filter(items=['AdditionalLimit_Inside', 'TotalLimit_Inside', 'Class_Code_Min', 'MoneyandSecuritiesInsideBaseRate']).rename(columns={'AdditionalLimit_Inside': 'Additional Limit INSIDE', 'TotalLimit_Inside': 'Total Limit INSIDE', 'MoneyandSecuritiesInsideBaseRate' : 'Base Rate'}).replace({'Class_Code_Min' : {10000 : 'Habitational', 60000 : 'Office', 20000 : 'All Other'}})
        MoneyInside = MoneyInside.query(f'Class_Code_Min == "Habitational" | Class_Code_Min == "Office" | Class_Code_Min == "All Other"')
        pivotedMoneyInside = MoneyInside.pivot(index=['Additional Limit INSIDE', 'Total Limit INSIDE'], columns='Class_Code_Min', values='Base Rate').reset_index(['Additional Limit INSIDE', 'Total Limit INSIDE'])
        pivotedMoneyInside = pivotedMoneyInside[['Additional Limit INSIDE', 'Total Limit INSIDE', 'Habitational', 'Office', 'All Other']]
        return pivotedMoneyInside
    
# Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildMoneyOutside(self):
        MoneyOutside = self.buildDataFrame("BP7 MoneyandSecuritiesOutside BaseRate")
        MoneyOutside = MoneyOutside.filter(items=['AdditionalLimit_Outside', 'TotalLimit_Outside', 'Class_Code_Min', 'MoneyandSecuritiesOutsideBaseRate']).rename(columns={'AdditionalLimit_Outside': 'Additional Limit OUTSIDE', 'TotalLimit_Outside': 'Total Limit OUTSIDE', 'MoneyandSecuritiesOutsideBaseRate' : 'Base Rate'}).replace({'Class_Code_Min' : {10000 : 'Habitational', 60000 : 'Office', 20000 : 'All Other'}})
        MoneyOutside = MoneyOutside.query(f'Class_Code_Min == "Habitational" | Class_Code_Min == "Office" | Class_Code_Min == "All Other"')
        pivotedMoneyOutside = MoneyOutside.pivot(index=['Additional Limit OUTSIDE', 'Total Limit OUTSIDE'], columns='Class_Code_Min', values='Base Rate').reset_index(['Additional Limit OUTSIDE', 'Total Limit OUTSIDE'])
        pivotedMoneyOutside = pivotedMoneyOutside[['Additional Limit OUTSIDE', 'Total Limit OUTSIDE', 'Habitational', 'Office', 'All Other']]
        return pivotedMoneyOutside

# Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildOutdoorSignsILF(self):
        OutdoorSignsILF = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return OutdoorSignsILF.query(f'`CoverageName` == "OutdoorSigns"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildOutdoorTreesILF(self):
        OutdoorTreesILF = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return OutdoorTreesILF.query(f'`CoverageName` == "OutdoorTreesShrubs"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildValuablePapersILF(self):
        ValuablePapersILF = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        return ValuablePapersILF.query(f'`FactorName` == "ValuablePapers"').rename(columns={'Factor': 'Factor'}).filter(items=['Factor'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildBackupSewerILF(self):
        BackupSewerILF = self.buildDataFrame("BP7_BackupSewerDrain_IncreasedLimitFactor")
        filteredBackupSewerILF = BackupSewerILF.query(f'ClassCode_Min == 40000').filter(items=['PerBuilding_IncreasedLimit', 'PolicyAggr_IncreasedLimit', 'PerBuilding_TotalLimit', 'PolicyAggr_TotalLimit', 'BackupSewerDrainFactor' ])
        return filteredBackupSewerILF.rename(columns={'PerBuilding_IncreasedLimit': 'Per Building', 'PolicyAggr_IncreasedLimit': 'Policy Aggregate', 'PerBuilding_TotalLimit': 'Per Building', 'PolicyAggr_TotalLimit': 'Policy Aggregate', 'BackupSewerDrainFactor': 'Rate per $100'})

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildAwayFromPremisesILF(self):
        AwayFromPremisesILF = self.buildDataFrame("BP7_PersonalProp_InTransit")
        return AwayFromPremisesILF.rename(columns={'CoverageApplies': 'Coverage', 'IntransitPremisesFactor': 'Factor'}).replace({'Coverage' : {'OnlyOnInsuredVehicles' : ' While in Transit - Primary on Vehicles Owned/Operated by Insured', 'WhileOnAnyVehicle' : 'While in Transit - Otherwise Shipped'}}).fillna({'Coverage' : 'Temporarily Away from Premises'}).sort_values('Factor', ascending=False)
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildElectronicDataILF(self):
        ElectronicDataILF = self.buildDataFrame("BP7_ElectronicDataBaseRates")
        return ElectronicDataILF.rename(columns={'ElectronicDataBaseRate': 'Additional Premium'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildInterruptionILF(self):
        InterruptionILF = self.buildDataFrame("BP7 InterrOfCompOper Coverage Base Rate")
        return InterruptionILF.rename(columns={'InterrOfCompOperBaseRate': 'Additional Premium'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildBLDGPropertyILF(self):
        BLDGPropertyILF = self.buildDataFrame("BP7_Miscellaneous_Base_Rates")
        return BLDGPropertyILF.query(f'`BaseRateName` == "BuildingPropertyOfOthers"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildComputerFraudILF(self):
        ComputerFraudILF = self.buildDataFrame("BP7_Computer_Fraud_And_Funds_Transfer_Fraud_Base_Rate")
        return ComputerFraudILF.rename(columns={'LimitOfInsurance': 'Limit Of Insurance', 'EachAdditionalEmployee': 'Each Additional Employee Over 5'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildCondoLossAsses(self):
        CondoLossAsses = self.buildDataFrame("BP7 CondoOwners_LossAssessmentLimit")
        CondoLossAsses = CondoLossAsses.astype({'LimitOfInsurance': 'int64'})
        return CondoLossAsses.rename(columns={'LimitOfInsurance': 'Limit Of Insurance', 'CondoLossAssessmentFactor': 'Additional Premium'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildCondo(self):
        Condo = self.buildDataFrame("BP7_Miscellaneous_Base_Rates")
        return Condo.query(f'`BaseRateName` == "CommlCondoUnitOwnersMiscProp"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the EQ Deductible Options Factor table
    # Returns a dataframe
    def buildEQPropertyDed(self):
        EQPropertyDed = self.buildDataFrame("BP7EarthquakeDeductibleOptions")
        updatedEQPropertyDed = EQPropertyDed.astype({'SubLimit_MinPct': 'int64', 'SubLimit_MaxPct': 'int64'}). \
                astype({'SubLimit_MinPct': 'string', 'SubLimit_MaxPct': 'string'}) # Converting to int first to get rid of decimal places
        updatedEQPropertyDed["Sub-Limit Percentage (%)"] = updatedEQPropertyDed["SubLimit_MinPct"] + ' - ' + updatedEQPropertyDed["SubLimit_MaxPct"] + '%' # Creating a single column for the percentage
        filteredEQPropertyDed = updatedEQPropertyDed.pivot(index=['DeductibleTier', 'BldgClass', 'Sub-Limit Percentage (%)'], columns='Deductible', values='DeductibleFactor' ).reset_index(['DeductibleTier', 'BldgClass', 'Sub-Limit Percentage (%)']). \
            replace({'Sub-Limit Percentage (%)': {'0 - 1%': "00 - 01%", '2 - 2%': "02 - 02%", '3 - 3%': "03 - 03%", '4 - 4%': "04 - 04%", '5 - 5%': "05 - 05%", '6 - 10%': "06 - 10%"}})
        filteredEQPropertyDed = filteredEQPropertyDed.sort_values(['DeductibleTier', 'BldgClass', 'Sub-Limit Percentage (%)']).rename(columns={10 : '10%', 15 : '15%', 20 : '20%', 25 : '25%'})
        return filteredEQPropertyDed

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildComputerFraudILF(self):
        ComputerFraudILF = self.buildDataFrame("BP7_Computer_Fraud_And_Funds_Transfer_Fraud_Base_Rate")
        return ComputerFraudILF.rename(columns={'LimitOfInsurance': 'Limit Of Insurance', 'EachAdditionalEmployee': 'Each Additional Employee Over 5'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildBPPStorage(self):
        BPPStorage = self.buildDataFrame("BP7_BPP_Temporarily_in_Storage_Base_Rate")
        return BPPStorage.filter(items=['Factor'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQSprinkler(self):
        EQSprinkler = self.buildDataFrame("BP7_Factors_Deductible_Coverage")
        return EQSprinkler.query(f'`PropertyDeductible` == 500 | PropertyDeductible == 1000 | PropertyDeductible == 2500 | PropertyDeductible == 5000 | PropertyDeductible == 10000 | PropertyDeductible == 25000 | PropertyDeductible == 50000').rename(columns={'PropertyDeductible': 'Deductible Amount'})

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQotherthanfunctional(self):
        EQotherthanfunctional = self.buildDataFrame("BP7PerilCWSpecificCovRateFactors")
        return EQotherthanfunctional.query(f'`CoverageSpecific` == "EarthquakeLossOfIncome" | CoverageSpecific == "EarthquakePackgageDisc"').filter(items=['CoverageSpecific', 'Value' ]).rename(columns={'CoverageSpecific': 'Coverage', 'Value': 'Factor'}).replace({'Coverage' : {'EarthquakeLossOfIncome' : 'Loss of Income', 'EarthquakePackgageDisc' : 'Package Discount Factor'}})

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQFunctional(self):
        EQFunctional = self.buildDataFrame("BP7EarthquakeFuncBldgValuation")
        return EQFunctional.query(f'`BuildingValuation` == "FunctionalValuation"').rename(columns={'FunctionalValuation': 'Factor'}).filter(items=['Factor'])


    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQUndamagedLoss(self):
        EQUndamagedLoss = self.buildDataFrame("BP7EarthquakeOrdLawFactor")
        return EQUndamagedLoss.query(f'`CovLossUndamagedPortion` == 1').filter(items=['BuildingValuation', 'Factor' ]).rename(columns={'BuildingValuation': 'Building Valuation', 'Factor': 'Factor'}).replace({'Building Valuation' : {'ActualCashValue' : 'Actual Cash Value - Building', 'Functional' : 'Functional Buidling Valuation', 'ReplacementCost' : 'Replacement Cost', 'ReplacementCostExtension' : 'Replacement Cost - Extension'}})

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQSprinklerLeakage(self):
        EQSprinklerLeakage = self.buildDataFrame("BP7PerilCWSpecificCovRateFactors")
        return EQSprinklerLeakage.query(f'`CoverageSpecific` == "EarthquakeLossOfIncome" | CoverageSpecific == "EarthquakePackgageDisc"').filter(items=['CoverageSpecific', 'Value' ]).rename(columns={'CoverageSpecific': 'Coverage', 'Value': 'Factor'}).replace({'Coverage' : {'EarthquakeLossOfIncome' : 'Loss of Income', 'EarthquakePackgageDisc' : 'Package Discount Factor'}})
    
    # Builds the Earthquake and Volcanic Eruption factor table
    # Returns a dataframe
    def buildEQDeductibleOptions(self):
        EQDeductibleOptions = self.buildDataFrame("BP7EarthquakeDeductibleOptions")
        filteredEQDeductibleOptions = EQDeductibleOptions.query(f'`SubLimit_MinPct` == 76 & (BldgClass == "1C" | BldgClass == "2A" | BldgClass == "3C" | BldgClass == "4B")')
        pivotedEQDeductibleOptions = filteredEQDeductibleOptions.pivot(index=['DeductibleTier', 'BldgClass'], columns='Deductible', values='DeductibleFactor').reset_index(['DeductibleTier', 'BldgClass']).replace({'BldgClass' : {'1C' : "1C, 1D", '2A' : "2A, 2B, 3A, 3B, 4A", '3C' : "3C, 4C, 4D, 5B, 5C, 5AA", '4B' : "4B, 5A"}}). \
            rename(columns={'DeductibleTier' : "Deductible Tier", 'BldgClass' : "Building Classes"}).rename(columns={10 : '10%', 15 : '15%', 20 : '20%', 25 : '25%'})
        return pivotedEQDeductibleOptions
    
    # Builds the Territory Multiplier factor table
    # Returns a dataframe
    def buildEQTerritoryDefinitions(self):
        EQTerritoryDefsByST = pd.DataFrame(data=self.EQTerritoryDefs)
        EQTerritoryDefsByST = EQTerritoryDefsByST.filter(items=['ZIP', 'COMM_EQ_CODE']).rename(columns={'ZIP' : 'Zip Code', 'COMM_EQ_CODE' : 'EQ Territory'}).astype({'EQ Territory': 'string'})
        if self.state =="CA":
            EQDedTier = self.buildDataFrame("BP7 Earthquake DeductibleTier_Ext2")
        else:
            EQDedTier = self.buildDataFrame("BP7EarthquakeDeductibleTier")
        EQDedTier = EQDedTier.rename(columns={'EarthquakeTerritory' : 'EQ Territory', 'DeductibleTier' : 'Deductible Tier'}).astype({'EQ Territory': 'string'})
        EQTerritoryDefs = pd.merge(EQTerritoryDefsByST, EQDedTier, on= 'EQ Territory', how = 'left')
        return EQTerritoryDefs

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQMasonryVeneer(self):
        EQMasonryVeneer = self.buildDataFrame("BP7EarthquakeMasonryVennerFactor")
        updatedEQMasonryVeneer = EQMasonryVeneer.astype({'VeneerBldg_MinPct': 'int64', 'VeneerBldg_MaxPct': 'int64'}). \
                astype({'VeneerBldg_MinPct': 'string', 'VeneerBldg_MaxPct': 'string'}) # Converting to int first to get rid of decimal places
        updatedEQMasonryVeneer["Veneer Building Percent"] = updatedEQMasonryVeneer["VeneerBldg_MinPct"] + ' - ' + updatedEQMasonryVeneer["VeneerBldg_MaxPct"] + '%' # Creating a single column for the percentage
        return updatedEQMasonryVeneer.query(f'`VeneerBldg_MinPct` == "10" | VeneerBldg_MinPct == "26" | VeneerBldg_MinPct == "51" ').filter(items=['Veneer Building Percent', 'VeenerFactor' ]).rename(columns={'Veneer Building Percent': 'Percentage Of Total Exterior Wall Areas Faced With Masonry Veneer', 'VeenerFactor': 'Factor'})
    

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQCoinsurance(self):
        EQCoinsurance = self.buildDataFrame("BP7EarthquakeCoinsuranceFactor")
        EQCoinsurance = EQCoinsurance.astype({'CoinsurancePct': 'int64'}).astype({'CoinsurancePct': 'string'})
        EQCoinsurance['CoinsurancePct'] = EQCoinsurance['CoinsurancePct'] + '%'
        return EQCoinsurance.rename(columns={'CoinsurancePct': 'Percent Of Coinsurance', 'CoinsuranceFactor' : 'Factor'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQSprinkleredRisk(self):
        EQSprinkleredRisk = self.buildDataFrame("BP7EarthquakeSprinkleredRiskFactor")
        return EQSprinkleredRisk.query(f'`SprinklerIndicator` == 1').filter(items=['SprinklerFactor' ]).rename(columns={'SprinklerFactor': 'Factor'})
    
    # Builds the Earthquake and Volcanic Eruption factor table
    # Returns a dataframe
    def buildEQBuildingHeight(self):
        EQBuildingHeight4 = self.buildDataFrame("BP7EarthquakeBuildingHeightModificationFactor").query(f'`Stories_Min` == 4').rename(columns={'Bldg Class' : 'Building Class'})
        pivotedEQBuildingHeight4 = EQBuildingHeight4.pivot(index= 'Building Class', columns= 'Tier', values= 'Factor').reset_index('Building Class')
        EQBuildingHeight8 = self.buildDataFrame("BP7EarthquakeBuildingHeightModificationFactor").query(f'`Stories_Min` == 8').rename(columns={'Bldg Class' : 'Building Class'})
        pivotedEQBuildingHeight8 = EQBuildingHeight8.pivot(index= 'Building Class', columns= 'Tier', values= 'Factor').reset_index('Building Class')
        EQBuildingHeight = pd.merge(pivotedEQBuildingHeight4, pivotedEQBuildingHeight8, on= 'Building Class', how = 'left')
        return EQBuildingHeight

    def buildAddInsuredOwnerLeaseContractor(self):
        # OC Table D.3.J.3. Additional Insured – Owners, Lessees or Contractors
        #  – Scheduled Person or Organization

        data = pd.DataFrame({"Rate":["$25"]})


        # Below is an example of multiple columns in two different formats.
        # VERSION 1
        # data = [
        #     {"Limit": "$100,000", "Factor": "0.95"},
        #     {"Limit": "$150,000", "Factor": "0.96"},
        #     {"Limit": "$250,000", "Factor": "0.97"},
        #     {"Limit": "$500,000", "Factor": "0.98"},
        #     {"Limit": "$1,000,000", "Factor": "0.99"}
        # ]
        #
        # # Create the DataFrame
        # formatted_df = pd.DataFrame(data)
        # VERSION 2
        # data = pd.DataFrame({"Limit": ["$100,000","$150,000","$250,000","$500,000","$1,000,000"],
        #                      "Factor" : ["0.95","0.96","0.97","0.98","0.99"]})

        return data

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    Territory = " "
    Territory1 = " "
    Territory2 = " "
    Territory3 = " "
    Territory4 = " "
    Territory5 = " "
    Territory6 = " "
    Territory7 = " "
    Territory8 = " "
    Territory9 = " "
    Territory10 = " "
    Territory11 = " "
    Territory12 = " "
    Territory13 = " "
    Territory14 = " "
    Territory15 = " "
    Territory16 = " "
    Territory17 = " "


    def buildEQClassRated(self):
        global Territory
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="MN" or self.state == "CT" or self.state == "NE" or self.state == "SD" or self.state == "DC" or self.state == "ND" or self.state == "RI" or self.state == "WI":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 1').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        elif self.state =="TX":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 5').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        elif self.state =="MD" or self.state == "IA" or self.state == "MI" or self.state == "FL" or self.state == "KS" or self.state == "WV":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 11').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else: 
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="MN" or self.state == "CT" or self.state == "NE" or self.state == "SD" or self.state == "DC" or self.state == "ND" or self.state == "RI" or self.state == "WI":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 1')
            Territory = "1"
        elif self.state =="TX":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 5')
            Territory = "5"
        elif self.state =="MD" or self.state == "IA" or self.state == "MI" or self.state == "FL" or self.state == "KS" or self.state == "WV":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 11')
            Territory = "11"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
            Territory = "21"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedMS1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 11').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
            Territory1 = "21"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 11')
            Territory1 = "11"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedMS2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 12').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
            Territory2 = "22"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 12')
            Territory2 = "12"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedMS3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 23').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 13').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 23')
            Territory3 = "23"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 13')
            Territory3 = "13"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedMS4(self):
        global Territory4
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 24').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 14').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="OH" or self.state =="IN" or self.state =="NM" or self.state =="TN":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 24')
            Territory4 = "24"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 14')
            Territory4 = "14"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedKY1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="KY" or self.state =="OR":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 1').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="KY" or self.state =="OR":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
            Territory1 = "21"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 1')
            Territory1 = "1"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedKY2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="KY" or self.state =="OR":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 2').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="KY" or self.state =="OR":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
            Territory2 = "22"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 2')
            Territory2 = "2"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedKY3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="KY" or self.state =="OR":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 23').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 3').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="KY" or self.state =="OR":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 23')
            Territory3 = "23"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 3')
            Territory3 = "3"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedKY4(self):
        global Territory4
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="KY" or self.state =="OR":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 24').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 4').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="KY" or self.state =="OR":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 24')
            Territory4 = "24"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 4')
            Territory4 = "4"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedKY5(self):
        global Territory5
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="KY" or self.state =="OR":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 25').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 5').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="KY" or self.state =="OR":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 25')
            Territory5 = "25"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 5')
            Territory5 = "5"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated


    def buildEQClassRatedIL1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="IL" or self.state=="MT":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 1').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})        
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="IL" or self.state=="MT":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
            Territory1 = "21"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 1')
            Territory1 = "1"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedIL2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="IL" or self.state=="MT":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 2').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="IL" or self.state=="MT":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
            Territory2 = "22"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 2')
            Territory2 = "2"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedIL3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="IL" or self.state=="MT":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 23').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 3').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="IL" or self.state=="MT":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 23')
            Territory3 = "23"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 3')
            Territory3 = "3"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedIL4(self):
        global Territory4
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="IL" or self.state=="MT":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 24').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 4').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="IL" or self.state=="MT":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 24')
            Territory4 = "24"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 4')
            Territory4 = "4"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedIL5(self):
        global Territory5
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="IL":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 25').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        elif self.state =="MT":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 91').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 5').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="IL":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 25')
            Territory5 = "25"
        elif self.state =="MT":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 91')
            Territory5 = "21A"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 5')
            Territory5 = "5"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedIL6(self):
        global Territory6
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        if self.state =="IL":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 26').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        elif self.state =="MT":
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 92').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        else:
            filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 6').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        if self.state =="IL":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 26')
            Territory6 = "26"
        elif self.state =="MT":
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 92')
            Territory6 = "22A"
        else:
            filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 6')
            Territory6 = "6"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
        Territory1 = "21"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
        Territory2 = "22"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 23').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 23')
        Territory3 = "23"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR4(self):
        global Territory4
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 24').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 24')
        Territory4 = "24"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR5(self):
        global Territory5
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 25').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 25')
        Territory5 = "25"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR6(self):
        global Territory6
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 26').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 26')
        Territory6 = "26"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedAR7(self):
        global Territory7
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 27').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 27')
        Territory7 = "27"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedNH1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
        Territory1 = "21"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedNH2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
        Territory2 = "22"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedNH3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 23').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 23')
        Territory3 = "23"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedID1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
        Territory1 = "21"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedID2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
        Territory2 = "22"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 21').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 21')
        Territory1 = "21"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 22').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 22')
        Territory2 = "22"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 92').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 92')
        Territory3 = "22A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT4(self):
        global Territory4
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 23').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 23')
        Territory4 = "23"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT5(self):
        global Territory5
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 93').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 93')
        Territory5 = "23A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT6(self):
        global Territory6
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 24').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 24')
        Territory6 = "24"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT7(self):
        global Territory7
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 95').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 95')
        Territory7 = "24A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedUT8(self):
        global Territory8
        BldgEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBLDG")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == 25').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7EarthquakeDvisionFiveLossCostBPP")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == 25')
        Territory8 = "25"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA1(self):
        global Territory1
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "21"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "21"')
        Territory1 = "21"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA2(self):
        global Territory2
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "91"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "91"')
        Territory2 = "21A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA3(self):
        global Territory3
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "22"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "22"')
        Territory3 = "22"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA4(self):
        global Territory4
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "92"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "92"')
        Territory4 = "22A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA5(self):
        global Territory5
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "23"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "23"')
        Territory5 = "23"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA6(self):
        global Territory6
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "93"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "93"')
        Territory6 = "23A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA7(self):
        global Territory7
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "94"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "94"')
        Territory7 = "23B"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA8(self):
        global Territory8
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "24"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "24"')
        Territory8 = "24"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA9(self):
        global Territory9
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "95"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "95"')
        Territory9 = "24A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA10(self):
        global Territory10
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "25"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "25"')
        Territory10 = "25"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA11(self):
        global Territory11
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "96"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "96"')
        Territory11 = "25A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA12(self):
        global Territory12
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "26"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "26"')
        Territory12 = "26"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA13(self):
        global Territory13
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "97"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "97"')
        Territory13 = "26A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA14(self):
        global Territory14
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "98"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "98"')
        Territory14 = "26B"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA15(self):
        global Territory15
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "27"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "27"')
        Territory15 = "27"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA16(self):
        global Territory16
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "99"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "99"')
        Territory16 = "27A"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    
    def buildEQClassRatedCA17(self):
        global Territory17
        BldgEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBLDG_Ext2")
        filteredBldgEQClassRated = BldgEQClassRated.query(f'`EarthquakeTerritory` == "28"').filter(items=['BldgClass', 'MandatoryDeductible', 'EarthquakeBldgLossCost']).astype({'MandatoryDeductible' : 'int64'}).astype({'MandatoryDeductible' : 'string'})
        filteredBldgEQClassRated['MandatoryDeductible'] = filteredBldgEQClassRated['MandatoryDeductible'] + '%'
        filteredBldgEQClassRated = filteredBldgEQClassRated.rename(columns={'BldgClass' : 'Bldg Class', 'MandatoryDeductible' : 'Mand. Deduct.', 'EarthquakeBldgLossCost' : 'Bldg.'})
        BPPEQClassRated = self.buildDataFrame("BP7 Earthquake DvisionFiveLossCostBPP_Ext2")
        filteredBPPEQClassRated = BPPEQClassRated.query(f'`EarthquakeTerritory` == "28"')
        Territory17 = "28"
        pivotedBPPEQClassRated = filteredBPPEQClassRated.pivot(index= 'BldgClass', columns= 'ContentsGrade', values= 'EarthquakeBPPLossCost').reset_index('BldgClass').rename(columns={'BldgClass' : 'Bldg Class'})
        EQClassRated = pd.merge(filteredBldgEQClassRated, pivotedBPPEQClassRated, on= 'Bldg Class', how = 'left').rename(columns={1 : '1', 2 : '2', 3 : '3', 4 : '4'})
        return EQClassRated
    


    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQLCM(self):
        EQNGICLcm = self.buildDataFrame("BP7EarthquakeDivisionFiveLossCostMultiplier")
        EQNGICLcm = EQNGICLcm.filter(items=['Underwriting CompanyDisplay Name', 'EarthquakeLossCostMultiplier' ]).rename(columns={'Underwriting CompanyDisplay Name': 'Company Name', 'EarthquakeLossCostMultiplier': 'Factor'})
        EQLcm = EQNGICLcm
        if self.NACORatebook != "Not found":
            EQNACOLcm = pd.DataFrame(self.rateTables['NACO']['BP7EarthquakeDivisionFiveLossCostMultiplier'][1:], index=None, columns=self.rateTables['NACO']['BP7EarthquakeDivisionFiveLossCostMultiplier'][0])
            EQNACOLcm = EQNACOLcm.filter(items=['Underwriting CompanyDisplay Name', 'EarthquakeLossCostMultiplier' ]).rename(columns={'Underwriting CompanyDisplay Name': 'Company Name', 'EarthquakeLossCostMultiplier': 'Factor'})
            EQLcm = pd.concat([EQLcm, EQNACOLcm])
        if self.NAFFRatebook != "Not found":
            EQNAFFLcm = pd.DataFrame(self.rateTables['NAFF']['BP7EarthquakeDivisionFiveLossCostMultiplier'][1:], index=None, columns=self.rateTables['NAFF']['BP7EarthquakeDivisionFiveLossCostMultiplier'][0])
            EQNAFFLcm = EQNAFFLcm.filter(items=['Underwriting CompanyDisplay Name', 'EarthquakeLossCostMultiplier' ]).rename(columns={'Underwriting CompanyDisplay Name': 'Company Name', 'EarthquakeLossCostMultiplier': 'Factor'})
            EQLcm = pd.concat([EQLcm, EQNAFFLcm])
        if self.NICOFRatebook != "Not found":
            EQNICOFLcm = pd.DataFrame(self.rateTables['NICOF']['BP7EarthquakeDivisionFiveLossCostMultiplier'][1:], index=None, columns=self.rateTables['NICOF']['BP7EarthquakeDivisionFiveLossCostMultiplier'][0])
            EQNICOFLcm = EQNICOFLcm.filter(items=['Underwriting CompanyDisplay Name', 'EarthquakeLossCostMultiplier' ]).rename(columns={'Underwriting CompanyDisplay Name': 'Company Name', 'EarthquakeLossCostMultiplier': 'Factor'})
            EQLcm = pd.concat([EQLcm, EQNICOFLcm])
        return EQLcm

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQSprinklerLeakCoinsurance(self):
        EQSprinklerLeakCoinsurance = self.buildDataFrame("BP7EarthquakeCoinsuranceFactor")
        EQSprinklerLeakCoinsurance = EQSprinklerLeakCoinsurance.astype({'CoinsurancePct': 'int64'}).astype({'CoinsurancePct': 'string'})
        EQSprinklerLeakCoinsurance['CoinsurancePct'] = EQSprinklerLeakCoinsurance['CoinsurancePct'] + '%'
        return EQSprinklerLeakCoinsurance.rename(columns={'CoinsurancePct': 'Percent Of Coinsurance', 'CoinsuranceFactor' : 'Factor'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEQSprinklerLeakEQExt(self):
        EQSprinklerLeakEQExt = self.buildDataFrame("BP7_EarthquakeSprinklerLeakageExtnFactor")
        return EQSprinklerLeakEQExt.rename(columns={'SusceptibilityGrade': 'Susceptibility Grade', 'SprinklerLeakageFactor' : 'Factor'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEmployeeDishonesty(self):
        EmployeeDishonesty = self.buildDataFrame("BP7_Pol_EmployeeDishonesty_BaseRate")
        return EmployeeDishonesty.rename(columns={'1to5EmployeesFactor': 'Premium 1-5 Employees', 'EachAddlEmpFactor' : 'Each Additional Employee'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEmployeeDishonestyERISA(self):
        EmployeeDishonestyERISA = self.buildDataFrame("BP7_Pol_EmployeeDishonesty_ERISAIndicator")
        return EmployeeDishonestyERISA.query(f'`ERISA_ComplianceIndicator` == 1').rename(columns={'ERISA_ComplianceIndicatorFactor' : 'Factor'}).filter(items=['Factor'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildExtendedBusinessIncome(self):
        ExtendedBusinessIncomeFood = self.buildDataFrame("BP7_Peril_ExtendedBusinessIncPeriodOfIdemnityFactor")
        filteredExtendedBusinessIncomeFood = ExtendedBusinessIncomeFood.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 40000 & IndemnityDays != 60').filter(items=['IndemnityDays', 'ExtendedBusinessIncPeriodOfIndemnityFactor'])
        ExtendedBusinessIncomeOther = self.buildDataFrame("BP7_Peril_ExtendedBusinessIncPeriodOfIdemnityFactor")
        filteredExtendedBusinessIncomeOther = ExtendedBusinessIncomeOther.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 20000 & IndemnityDays != 60').filter(items=['IndemnityDays', 'ExtendedBusinessIncPeriodOfIndemnityFactor'])
        ExtendedBusinessIncome = pd.merge(filteredExtendedBusinessIncomeFood, filteredExtendedBusinessIncomeOther, on= 'IndemnityDays', how = 'left')
        ExtendedBusinessIncome = ExtendedBusinessIncome.rename(columns={'IndemnityDays' : "Number of Days", 'ExtendedBusinessIncPeriodOfIndemnityFactor_x' : "Food Service Business Income", 'ExtendedBusinessIncPeriodOfIndemnityFactor_y' : "All Other Programs Building and BPP"})
        return ExtendedBusinessIncome

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildBusinessIncomeOrdinary(self):
        BusinessIncomeOrdFood = self.buildDataFrame("BP7_Peril_BusinessIncOrdinaryPayrollFactor")
        filteredBusinessIncomeOrdFood = BusinessIncomeOrdFood.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 40000 & NoOfDays != 60').filter(items=['NoOfDays', 'BusinessIncOrdinaryPayrollFactor'])
        BusinessIncomeOrdOther = self.buildDataFrame("BP7_Peril_BusinessIncOrdinaryPayrollFactor")
        filteredBusinessIncomeOrdOther = BusinessIncomeOrdOther.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 10000 & NoOfDays != 60').filter(items=['NoOfDays', 'BusinessIncOrdinaryPayrollFactor'])
        BusinessIncomeOrd = pd.merge(filteredBusinessIncomeOrdFood, filteredBusinessIncomeOrdOther, on= 'NoOfDays', how = 'left')
        BusinessIncomeOrd = BusinessIncomeOrd.rename(columns={'NoOfDays' : "Number of Days", 'BusinessIncOrdinaryPayrollFactor_x' : "Food Service Business Income", 'BusinessIncOrdinaryPayrollFactor_y' : "All Other Programs Building and BPP"})
        return BusinessIncomeOrd

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildBusinessIncomeActual(self):
        BusinessIncomeActFood = self.buildDataFrame("BP7_Peril_ActualLossSustainedFactor")
        filteredBusinessIncomeActFood = BusinessIncomeActFood.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 40000').filter(items=['NoOfMonths', 'BusinessIncomeActualLossSustained'])
        BusinessIncomeActOther = self.buildDataFrame("BP7_Peril_ActualLossSustainedFactor")
        filteredBusinessIncomeActOther = BusinessIncomeActOther.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 20000').filter(items=['NoOfMonths', 'BusinessIncomeActualLossSustained'])
        BusinessIncomeAct = pd.merge(filteredBusinessIncomeActFood, filteredBusinessIncomeActOther, on= 'NoOfMonths', how = 'left')
        BusinessIncomeAct = BusinessIncomeAct.rename(columns={'NoOfMonths' : "Number of Months", 'BusinessIncomeActualLossSustained_x' : "Food Service", 'BusinessIncomeActualLossSustained_y' : "All Other Programs"})
        BusinessIncomeAct['Food Service'] = BusinessIncomeAct['Food Service'] + 1
        BusinessIncomeAct['All Other Programs'] = BusinessIncomeAct['All Other Programs'] + 1
        return BusinessIncomeAct
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildBusinessIncomeWaiting(self):
        BusinessIncomeWaitFood = self.buildDataFrame("BP7_Peril_WaitingPeriodFactor")
        filteredBusinessIncomeWaitFood = BusinessIncomeWaitFood.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 40000').filter(items=['NoOfHours', 'BusinessIncomeWaitingPeriod'])
        BusinessIncomeWaitOther = self.buildDataFrame("BP7_Peril_WaitingPeriodFactor")
        filteredBusinessIncomeWaitOther = BusinessIncomeWaitOther.query(f'`Peril TypeCode` == "allperil" & Class_Code_Min == 20000').filter(items=['NoOfHours', 'BusinessIncomeWaitingPeriod'])
        BusinessIncomeWait = pd.merge(filteredBusinessIncomeWaitFood, filteredBusinessIncomeWaitOther, on= 'NoOfHours', how = 'left')
        BusinessIncomeWait = BusinessIncomeWait.rename(columns={'NoOfHours' : "Number of Hours", 'BusinessIncomeWaitingPeriod_x' : "Food Service", 'BusinessIncomeWaitingPeriod_y' : "All Other Programs"})
        BusinessIncomeWait['Food Service'] = BusinessIncomeWait['Food Service'] + 1
        BusinessIncomeWait['All Other Programs'] = BusinessIncomeWait['All Other Programs'] + 1
        return BusinessIncomeWait

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildOrdinanceLoss(self):
        OrdinanceLoss = self.buildDataFrame("BP7_Ordinance_Or_Law_Factor")
        return OrdinanceLoss.filter(items=['Class_Code_Min', 'Ordinance Or Law Factor' ]).rename(columns={'Class_Code_Min': 'Program', 'Ordinance Or Law Factor': 'Factor'}).replace({'Program': self.classCodes}).replace({'Program' : {'Hab' : 'Habitational', 'Food' : 'Food Service', 'Auto' : 'Auto Service'}}).sort_values('Program')
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildOrdinanceEndorsement(self):
        OrdinanceEndorsement = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return OrdinanceEndorsement.query(f'`CoverageName` == "OrdLawBroadened"').rename(columns={'BaseRate' : 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildOutdoorSignsBusInc(self):
        OutdoorSignsBusInc = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return OutdoorSignsBusInc.query(f'`CoverageName` == "BusinessIncomeOutsideSigns"').rename(columns={'BaseRate' : 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildSpoilage(self):
        Spoilage = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return Spoilage.query(f'`CoverageName` == "SpoilagePowerOutage"').rename(columns={'BaseRate' : 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildUtilityServices1(self):
        UtilityServicesComm = self.buildDataFrame("BP7_UtilitySrv_CommunicationSupply")
        UtilityServicesComm = UtilityServicesComm.pivot(index='CoverageCode', columns='CommunicationSupplyOption', values='CommunicationSupplyFactor').reset_index('CoverageCode')
        UtilityServicesComm = UtilityServicesComm.rename(columns={'CoverageCode' : 'Coverage', 'ExcludingOverheadTransmissionLines' : 'Communication', 'IncludingOverheadTransmissionLines' : 'CommIncluding'}).filter(items=['Coverage', 'Communication'])
        UtilityServicesPower = self.buildDataFrame("BP7_UtilitySrv_PowerSupply")
        UtilityServicesPower = UtilityServicesPower.pivot(index='CoverageCode', columns='PowerSupplyOption', values='PowerSupplyFactor').reset_index('CoverageCode')
        UtilityServicesPower = UtilityServicesPower.rename(columns={'CoverageCode' : 'Coverage', 'ExcludingOverheadTransmissionLines' : 'Power', 'IncludingOverheadTransmissionLines' : 'PowerIncluding'}).filter(items=['Coverage', 'Power'])
        #UtilityServicesPower = UtilityServicesPower.pivot(index='CoverageCode', columns='PowerSupplyOption', values='PowerSupplyFactor').reset_index('CoverageCode')
        #UtilityServicesWater = self.buildDataFrame("BP7_UtilitySrv_WaterSupply")
        UtilityServices1 = pd.merge(UtilityServicesComm, UtilityServicesPower, on='Coverage', how='left').replace({'Coverage': {'BppOnly': 'Personal Property', 'BuildingOnly': 'Building'}}).sort_values('Coverage')
        return UtilityServices1
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildUtilityServices2(self):
        UtilityServicesComm = self.buildDataFrame("BP7_UtilitySrv_CommunicationSupply")
        UtilityServicesComm = UtilityServicesComm.pivot(index='CoverageCode', columns='CommunicationSupplyOption', values='CommunicationSupplyFactor').reset_index('CoverageCode')
        UtilityServicesComm = UtilityServicesComm.rename(columns={'CoverageCode' : 'Coverage', 'ExcludingOverheadTransmissionLines' : 'CommExcluding', 'IncludingOverheadTransmissionLines' : 'Communication'}).filter(items=['Coverage', 'Communication'])
        UtilityServicesPower = self.buildDataFrame("BP7_UtilitySrv_PowerSupply")
        UtilityServicesPower = UtilityServicesPower.pivot(index='CoverageCode', columns='PowerSupplyOption', values='PowerSupplyFactor').reset_index('CoverageCode')
        UtilityServicesPower = UtilityServicesPower.rename(columns={'CoverageCode' : 'Coverage', 'ExcludingOverheadTransmissionLines' : 'PowerExcluding', 'IncludingOverheadTransmissionLines' : 'Power'}).filter(items=['Coverage', 'Power'])
        #UtilityServicesPower = UtilityServicesPower.pivot(index='CoverageCode', columns='PowerSupplyOption', values='PowerSupplyFactor').reset_index('CoverageCode')
        #UtilityServicesWater = self.buildDataFrame("BP7_UtilitySrv_WaterSupply")
        UtilityServices2 = pd.merge(UtilityServicesComm, UtilityServicesPower, on='Coverage', how='left').replace({'Coverage': {'BppOnly': 'Personal Property', 'BuildingOnly': 'Building'}}).sort_values('Coverage')
        return UtilityServices2
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildUtilityServices3(self):
        UtilityServicesWater = self.buildDataFrame("BP7_UtilitySrv_WaterSupply")
        UtilityServices3 = UtilityServicesWater.rename(columns={'CoverageCode': 'Coverage', 'WaterSupplyFactor': 'Factor'}).replace({'Coverage': {'BppOnly': 'Personal Property', 'BuildingOnly': 'Building'}}).sort_values('Coverage')
        return UtilityServices3

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildScheduledPropFloater(self):
        ScheduledPropFloater = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return ScheduledPropFloater.query(f'`CoverageName` == "ScheduledPropertyFloater"').rename(columns={'BaseRate' : 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildVehicleDamBase(self):
        VehicleDamBase = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return VehicleDamBase.query(f'`CoverageName` == "VehicleDamageToProp"').rename(columns={'BaseRate' : 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildVehicleDamAdditional(self):
        VehicleDamAdditional = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return VehicleDamAdditional.query(f'`CoverageName` == "VehicleDamageToPropAddLimit"').rename(columns={'BaseRate' : 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildRebuildingExpense(self):
        RebuildingExpense = self.buildDataFrame("BP7PerilCWSpecificCovRateFactors")
        return RebuildingExpense.query(f'`CoverageSpecific` == "Increase In Rebuilding Expense"').filter(items=['Value' ]).rename(columns={'Value': 'Factor'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildRentalLoss(self):
        RentalLoss = self.buildDataFrame("BP7_Miscellaneous_Base_Rates")
        return RentalLoss.query(f'`BaseRateName` == "LossOFRentalDesignLandlord"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildExclusionRoofSiding(self):
        roof_factor = self.buildDataFrame("BP7_Cosmetic_Loss_Exclusion_Factor")
        roof_factor = roof_factor[roof_factor["Peril TypeCode"] == "allperil"].drop(columns = ["Peril TypeCode"])
        roof_factor = roof_factor[roof_factor["Peril TypeDisplay Name"] == "All Peril"].drop(columns=["Peril TypeDisplay Name"])
        roof_factor = roof_factor[roof_factor["Exclusion Option"] == "Roof Covering And Siding"].drop(columns=["Exclusion Option"])
        return roof_factor
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildExclusionRoof(self):
        roof_factor = self.buildDataFrame("BP7_Cosmetic_Loss_Exclusion_Factor")
        roof_factor = roof_factor[roof_factor["Peril TypeCode"] == "allperil"].drop(columns = ["Peril TypeCode"])
        roof_factor = roof_factor[roof_factor["Peril TypeDisplay Name"] == "All Peril"].drop(columns=["Peril TypeDisplay Name"])
        roof_factor = roof_factor[roof_factor["Exclusion Option"] == "Roof Covering"].drop(columns=["Exclusion Option"])
        return roof_factor
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildWindhailACVSettlement(self):
        data = {'Factor': [1]}
        WindhailACVSettlement  = pd.DataFrame(data)
        return WindhailACVSettlement
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildACVRoof(self):
        data = {'Factor': [1]}
        ACVRoof  = pd.DataFrame(data)
        return ACVRoof

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildFalsePretense(self):
        FalsePretense = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return FalsePretense.query(f'`CoverageName` == "False Pretense"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodBaseRates(self):
        FloodBaseRates = self.buildDataFrame("BP7_FloodBaseRateMinimums_Ext")

        filteredFloodBaseRatesLow = FloodBaseRates.query(f'`Flood Zone` == "C"')
        filteredFloodBaseRatesLow = filteredFloodBaseRatesLow.replace({'Flood Zone' : {'C' : "C, X"}})
        filteredFloodBaseRatesLow['Hazard Area'] = ['Low']
        filteredFloodBaseRatesLow = filteredFloodBaseRatesLow.filter(items={'Hazard Area', 'Flood Zone', 'Base Rate'})

        filteredFloodBaseRatesMod = FloodBaseRates.query(f'`Flood Zone` == "X500"')
        filteredFloodBaseRatesMod = filteredFloodBaseRatesMod.replace({'Flood Zone' : {'X500' : "X500, B, X500L, BL"}})
        filteredFloodBaseRatesMod['Hazard Area'] = ['Moderate']
        filteredFloodBaseRatesMod = filteredFloodBaseRatesMod.filter(items={'Hazard Area', 'Flood Zone', 'Base Rate'})

        filteredFloodBaseRatesHigh = FloodBaseRates.query(f'`Flood Zone` == "A"')
        filteredFloodBaseRatesHigh = filteredFloodBaseRatesHigh.replace({'Flood Zone' : {'A' : "A, AH, AO, A1-A30, A99, AE, AR"}})
        filteredFloodBaseRatesHigh['Hazard Area'] = ['High']
        filteredFloodBaseRatesHigh = filteredFloodBaseRatesHigh.filter(items={'Hazard Area', 'Flood Zone', 'Base Rate'})

        FloodBaseRatesLowMod = pd.concat([filteredFloodBaseRatesLow, filteredFloodBaseRatesMod])
        FloodBaseRates = pd.concat([FloodBaseRatesLowMod, filteredFloodBaseRatesHigh])
        FloodBaseRates = FloodBaseRates[['Hazard Area', 'Flood Zone', 'Base Rate']]

        return FloodBaseRates

    # Builds the masonry veneer factor table for the given coverage (either Building or BPP)
    # Returns a dataframe
    def buildFloodSubLimit(self):
        FloodSubLimit = self.buildDataFrame("BP7_FloodSubLimitFactor_Ext")
        updatedFloodSubLimit = FloodSubLimit.astype({'Min Sub-Limit Percentage': 'int64', 'Max Sub-Limit Percentage': 'int64'}). \
                astype({'Min Sub-Limit Percentage': 'string', 'Max Sub-Limit Percentage': 'string'}) # Converting to int first to get rid of decimal places
        updatedFloodSubLimit["Sub-Limit Percentage (%)"] = updatedFloodSubLimit["Min Sub-Limit Percentage"] + ' - ' + updatedFloodSubLimit["Max Sub-Limit Percentage"] + '%' # Creating a single column for the percentage
        filteredFloodSubLimit = updatedFloodSubLimit.filter(items=['Sub-Limit Percentage (%)', 'Sub-Limit Factor'])
        return filteredFloodSubLimit

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodBLDGConstruction(self):
        FloodBLDGConstruction = self.buildDataFrame("BP7_FloodConstructionTypeFactor_Ext")
        FloodBLDGConstruction.loc[FloodBLDGConstruction['Construction TypeDisplay Name'] == 'Frame', 'Code'] = '1'
        FloodBLDGConstruction.loc[FloodBLDGConstruction['Construction TypeDisplay Name'] == 'Joisted Masonry', 'Code'] = '2'
        FloodBLDGConstruction.loc[FloodBLDGConstruction['Construction TypeDisplay Name'] == 'Non-Combustible', 'Code'] = '3'
        FloodBLDGConstruction.loc[FloodBLDGConstruction['Construction TypeDisplay Name'] == 'Masonry Non-Combustible', 'Code'] = '4'
        FloodBLDGConstruction.loc[FloodBLDGConstruction['Construction TypeDisplay Name'] == 'Modified Fire-Resistive', 'Code'] = '5'
        FloodBLDGConstruction.loc[FloodBLDGConstruction['Construction TypeDisplay Name'] == 'Fire-Resistive', 'Code'] = '6'
        filteredFloodBLDGConstruction = FloodBLDGConstruction.filter(items=['Construction TypeDisplay Name', 'Code', 'Construction Type Factor']).rename(columns={'Construction TypeDisplay Name': 'Construction', 'Construction Type Factor': 'Factor'})
        filteredFloodBLDGConstruction = filteredFloodBLDGConstruction[['Construction', 'Code', 'Factor']].sort_values('Code')
        return filteredFloodBLDGConstruction

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodBPP(self):
        FloodBPP = self.buildDataFrame("BP7_FloodConcentrationValuesFactor_Ext")
        FloodBPP = FloodBPP.replace({'Concentration Risk' : {'Second Floor': 'Second Floor or Higher'}}). \
            rename(columns={'Concentration Risk': 'Description', 'Concentration Factor': 'Factor'}).sort_values('Factor')
        FloodBPP.loc[FloodBPP['Description'] == 'Second Floor or Higher', 'Concentration of Values'] = 'Low'
        FloodBPP.loc[FloodBPP['Description'] == 'First Floor', 'Concentration of Values'] = 'Moderate'
        FloodBPP.loc[FloodBPP['Description'] == 'Basement', 'Concentration of Values'] = 'High'
        FloodBPP = FloodBPP[['Concentration of Values', 'Description', 'Factor']].sort_values('Factor')
        return FloodBPP


    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodDedFactors(self):
        FloodDedFactors = self.buildDataFrame("BP7_FloodDeductibleFactor_Ext")
        return FloodDedFactors.rename(columns={'Deductible Factor' : 'Factor'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodMinAdjRate(self):
        FloodMinPrem = self.buildDataFrame("BP7_FloodBaseRateMinimums_Ext")
        filteredFloodMinPrem = FloodMinPrem.query(f'`Flood Zone` == "C" | `Flood Zone` == "X500" | `Flood Zone` == "A"')
        filteredFloodMinPrem.loc[filteredFloodMinPrem['Flood Zone'] == "C", 'Hazard Area'] = 'Low'
        filteredFloodMinPrem.loc[filteredFloodMinPrem['Flood Zone'] == "X500", 'Hazard Area'] = 'Moderate'
        filteredFloodMinPrem.loc[filteredFloodMinPrem['Flood Zone'] == "A", 'Hazard Area'] = 'High'
        filteredFloodMinPrem = filteredFloodMinPrem.replace({'Flood Zone' : {"C" : "C, X", "X500" : "X500, B, X500L, BL", "A" : "A, AH, AO, A1-A30, A99, AE, AR"}})
        filteredFloodMinPrem = filteredFloodMinPrem.filter(items=['Hazard Area', 'Flood Zone', 'Minimum Rate']).sort_values('Minimum Rate')
        return filteredFloodMinPrem[['Hazard Area', 'Flood Zone', 'Minimum Rate']]
        

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodAggLimit(self):
        FloodAggLimit = self.buildDataFrame("BP7_FloodAggregateLimitMultiplier_Ext")
        return FloodAggLimit.rename(columns={'Aggregate Limit Multiplier' : 'Multiplier'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildFloodMinPrem(self):
        FloodMinPrem = self.buildDataFrame("BP7_FloodBaseRateMinimums_Ext")

        filteredFloodMinPremLow = FloodMinPrem.query(f'`Flood Zone` == "C"')
        filteredFloodMinPremLow = filteredFloodMinPremLow.replace({'Flood Zone' : {'C' : "C, X"}})
        filteredFloodMinPremLow['Hazard Area'] = ['Low']
        filteredFloodMinPremLow = filteredFloodMinPremLow.filter(items={'Hazard Area', 'Flood Zone', 'Minimum Premium'})

        filteredFloodMinPremMod = FloodMinPrem.query(f'`Flood Zone` == "X500"')
        filteredFloodMinPremMod = filteredFloodMinPremMod.replace({'Flood Zone' : {'X500' : "X500, B, X500L, BL"}})
        filteredFloodMinPremMod['Hazard Area'] = ['Moderate']
        filteredFloodMinPremMod = filteredFloodMinPremMod.filter(items={'Hazard Area', 'Flood Zone', 'Minimum Premium'})

        filteredFloodMinPremHigh = FloodMinPrem.query(f'`Flood Zone` == "A"')
        filteredFloodMinPremHigh = filteredFloodMinPremHigh.replace({'Flood Zone' : {'A' : "A, AH, AO, A1-A30, A99, AE, AR"}})
        filteredFloodMinPremHigh['Hazard Area'] = ['High']
        filteredFloodMinPremHigh = filteredFloodMinPremHigh.filter(items={'Hazard Area', 'Flood Zone', 'Minimum Premium'})

        FloodMinPremCombined = pd.concat([filteredFloodMinPremLow, filteredFloodMinPremMod, filteredFloodMinPremHigh])
        FloodMinPremium = FloodMinPremCombined[['Hazard Area', 'Flood Zone', 'Minimum Premium']]

        return FloodMinPremium

    def buildBusinessIncomeExtraExpense(self):

        data = pd.DataFrame({
            "Limit": ["$100,000", "$150,000", "$250,000", "$500,000", "$1,000,000"],
            "Factor": ['0.95', '0.96', '0.97', '0.98', '0.99']
        })

        return data

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildAmendmentAdditional(self):
        AmendmentAdditional = self.buildDataFrame("BP7PerilCWSpecificCovRateFactors")
        return AmendmentAdditional.query(f'`CoverageSpecific` == "AmendmentInsClauseAddinsrd"').filter(items=['Value' ]).rename(columns={'Value': 'Rate'})
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAdditionalFairs(self):
        AdditionalFairs = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return AdditionalFairs.query(f'`CoverageName` == "AdditionalInsuredFarisCarnivalBaseRatesForHabOffice" | CoverageName == "AdditionalInsuredFarisCarnivalBaseRatesForOthers"').\
            rename(columns={'CoverageName' : 'Coverage','BaseRate': 'Rate'}).filter(items=['Coverage', 'Rate']).replace({'Coverage' : {'AdditionalInsuredFarisCarnivalBaseRatesForHabOffice' : 'Habitational and Office Programs', 'AdditionalInsuredFarisCarnivalBaseRatesForOthers' : 'All Other Programs'}})
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAdditionalServices(self):
        AdditionalFairs = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return AdditionalFairs.query(f'`CoverageName` == "AdditionalInsuredServicesPerformedOnPremisesBaseRatesForHabOffice" | CoverageName == "AdditionalInsuredServicesPerformedOnPremisesBaseRatesForOthers"'). \
            rename(columns={'CoverageName' : 'Coverage','BaseRate': 'Rate'}).filter(items=['Coverage', 'Rate']).replace({'Coverage' : {'AdditionalInsuredServicesPerformedOnPremisesBaseRatesForHabOffice' : 'Habitational and Office Programs', 'AdditionalInsuredServicesPerformedOnPremisesBaseRatesForOthers' : 'All Other Programs'}})

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAdditionalVendors(self):
        AdditionalVendors = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return AdditionalVendors.query(f'`CoverageName` == "AdditionalInsrdVendor"').rename(columns={'CoverageName' : 'Coverage','BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAdditionalDesignated(self):
        AdditionalDesignated = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return AdditionalDesignated.query(f'`CoverageName` == "AdditionalInsuredDesignatedPersonOrOrganizationRates"').rename(columns={'CoverageName' : 'Coverage','BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEmployeeBodilyInj(self):
        EmployeeBodilyInj = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        return EmployeeBodilyInj.query(f'`FactorName` == "EmployeeBodilyInjuryToAnotherEmployee"').rename(columns={'Factor': 'Rate'}).filter(items=['Rate'])
    #NEED TO INCLUDE MINIMUM PREMIUM 
        # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildEmployeeBodilyInjMinPrem(self):
        EmployeeBodilyInjMinPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        EmployeeBodilyInjMinPrem = EmployeeBodilyInjMinPrem.query(f'CoverageType == "BP7Pol_EmployeeBodilyInjuryDesignatedPositionsCov_Ext"').filter(items=['Premium'])
        return EmployeeBodilyInjMinPrem.rename(columns={'Premium': 'Minimum Premium'})

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeeBenefitsLiab(self):
        EmployeeBenefitsLiab = self.buildDataFrame("BP7_EmpBenefitsLiabBaseRate")
        updatedEmployeeBenefitsLiab = EmployeeBenefitsLiab.fillna({'MaxNumofEmployees': 0}).astype({'MinNumofEmployees': 'int64', 'MaxNumofEmployees': 'int64'}). \
                astype({'MinNumofEmployees': 'string', 'MaxNumofEmployees': 'string'}) # Converting to int first to get rid of decimal places
        updatedEmployeeBenefitsLiab["Number of Employees"] = updatedEmployeeBenefitsLiab["MinNumofEmployees"] + ' - ' + updatedEmployeeBenefitsLiab["MaxNumofEmployees"] # Creating a single column for the percentage
        updatedEmployeeBenefitsLiab['Liability Limit Occurrence'] = updatedEmployeeBenefitsLiab['Liability Limit Occurrence'] / 1000
        updatedEmployeeBenefitsLiab['LiabilityLimitAggregate'] = updatedEmployeeBenefitsLiab['LiabilityLimitAggregate'] / 1000
        updatedEmployeeBenefitsLiab = updatedEmployeeBenefitsLiab.astype({'Liability Limit Occurrence': 'int64', 'LiabilityLimitAggregate': 'int64'}). \
                astype({'Liability Limit Occurrence': 'string', 'LiabilityLimitAggregate': 'string'}) # Converting to int first to get rid of decimal places
        updatedEmployeeBenefitsLiab['Liability'] = '$' + updatedEmployeeBenefitsLiab['Liability Limit Occurrence'] + ' / $' +  updatedEmployeeBenefitsLiab['LiabilityLimitAggregate']
        updatedEmployeeBenefitsLiab = updatedEmployeeBenefitsLiab.replace({'Number of Employees' : {'1001 - 0' : 'Over 1000', '51 - 100' : '051 - 100', '0 - 50': '000 - 050'}})
        updatedEmployeeBenefitsLiab = updatedEmployeeBenefitsLiab.filter(items=['Liability', 'Number of Employees', 'BaseRate'])
        updatedEmployeeBenefitsLiab = updatedEmployeeBenefitsLiab.pivot(index='Number of Employees', columns='Liability', values='BaseRate').reset_index('Number of Employees')
        updatedEmployeeBenefitsLiab = updatedEmployeeBenefitsLiab[['Number of Employees', '$300 / $600', '$500 / $1000', '$1000 / $2000', '$2000 / $4000']]
        updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == 'Over 1000', '$300 / $600'] = str(updatedEmployeeBenefitsLiab.iloc[4,1].astype('int64')) + ' + $' +  str(updatedEmployeeBenefitsLiab.iloc[5,1]) + ' per employee'
        updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == 'Over 1000', '$500 / $1000'] = str(updatedEmployeeBenefitsLiab.iloc[4,2].astype('int64')) + ' + $' +  str(updatedEmployeeBenefitsLiab.iloc[5,2]) + ' per employee'
        updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == 'Over 1000', '$1000 / $2000'] = str(updatedEmployeeBenefitsLiab.iloc[4,3].astype('int64')) + ' + $' +  str(updatedEmployeeBenefitsLiab.iloc[5,3]) + ' per employee'
        updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == 'Over 1000', '$2000 / $4000'] = str(updatedEmployeeBenefitsLiab.iloc[4,4].astype('int64')) + ' + $' +  str(updatedEmployeeBenefitsLiab.iloc[5,4]) + ' per employee'
        #updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == 'Over 1000', '$300 / $600'] = updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == '501 - 1000', '$300 / $600'] #+ ' + ' + updatedEmployeeBenefitsLiab.loc[updatedEmployeeBenefitsLiab['Number of Employees'] == 'Over 1000', '$300 / $600'] + ' per employee'
        return updatedEmployeeBenefitsLiab

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeeBenefitsLiabExt(self):
        EmployeeBenefitsLiabExt = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        EmployeeBenefitsLiabExt = EmployeeBenefitsLiabExt.query(f'`FactorName` == "EmpBenefitsLiabERPAnnualPremium"')
        EmployeeBenefitsLiabExt['Factor'] = (EmployeeBenefitsLiabExt['Factor']*100)
        EmployeeBenefitsLiabExt = EmployeeBenefitsLiabExt.astype({'Factor' : 'int64'}).astype({'Factor' : 'string'})
        EmployeeBenefitsLiabExt['Factor'] = EmployeeBenefitsLiabExt['Factor'] + "%"
        EmployeeBenefitsLiabExt = EmployeeBenefitsLiabExt.rename(columns={'Factor': '% of Annual Premium'}).filter(items=['% of Annual Premium'])
        return EmployeeBenefitsLiabExt
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildGarageKeepersBase(self):
        GarageKeepersBase = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        return GarageKeepersBase.query(f'`CoverageName` == "GarageKeepers"').rename(columns={'CoverageName' : 'Coverage','BaseRate': 'Rate'}).filter(items=['Rate'])
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildGarageKeepersDed(self):
        GarageKeepersDed = self.buildDataFrame("BP7_Bldg_GarageKeepers_DeducitbleFactors")
        GarageKeepersDed = GarageKeepersDed.pivot(index='Deductible', columns='CausesType', values='DeductableFactor').reset_index('Deductible').rename(columns={'AllCauses': 'All Causes', 'LimitedCauses': 'Limited Causes', 'Deductible': 'Deductible Amount'}).sort_values('All Causes', ascending=False )
        GarageKeepersDed = GarageKeepersDed[['Deductible Amount', 'Limited Causes', 'All Causes']]
        GarageKeepersDed = GarageKeepersDed.astype({'Deductible Amount' : 'int64'})
        return GarageKeepersDed
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildGarageKeepersDedLOI(self):
        GarageKeepersDedLOI = self.buildDataFrame("BP7_Bldg_GarageKeepers_LimitOfInsuranceFactor")
        GarageKeepersDedLOI = GarageKeepersDedLOI.filter(items=['LimitOfInsurance_Min', 'LimitOfInsuranceFactor']).rename(columns={'LimitOfInsurance_Min': 'Limit of Insurance', 'LimitOfInsuranceFactor': 'Factor'})
        return GarageKeepersDedLOI
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildGarageKeepersBlanket(self):
        GarageKeepersBlanket = self.buildDataFrame("BP7_Bldg_GarageKeepers_Blanket Insurance")
        return GarageKeepersBlanket.query(f'BlanketInsuranceIndicator == 1').rename(columns={'BlanketInsuranceFactor': 'Factor'}).filter(items=['Factor'])
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildGarageKeepersRateMod(self):
        GarageKeepersBasisMod = self.buildDataFrame("BP7_Bldg_GarageKeepers_BasisModifier")
        GarageKeepersBasisMod = GarageKeepersBasisMod.query(f'LiabilityCoverageBasis == "DirectPrimary"').rename(columns={'LiabilityCoverageBasis' : 'Modifier', 'BMParamFactor' : 'Factor'}).replace({'Modifier' : {'DirectPrimary' : 'Basis Modifier'}})

        GarageKeepersProtected = self.buildDataFrame("BP7_Bldg_GarageKeepers_ProtectedLotModifier")
        GarageKeepersProtected = GarageKeepersProtected.query(f'PlMParamName == "true"').rename(columns={'PlMParamName' : 'Modifier', 'PLMParamFactor' : 'Factor'}).replace({'Modifier' : {'true' : 'Protected Lot Modifier'}})

        GarageKeepersBurglar = self.buildDataFrame("BP7_Bldg_GarageKeepers_BurglarAlarmIndicator")
        GarageKeepersBurglar = GarageKeepersBurglar.query(f'AlarmTypeParamName == "central" | AlarmTypeParamName == "local"').rename(columns={'AlarmTypeParamName' : 'Type', 'AlarmTypeFactor' : 'Factor'}).replace({'Type' : {'local' : 'Local', 'central' : 'Central Station'}})

        GarageKeepersRateMod = pd.concat([GarageKeepersBasisMod, GarageKeepersProtected, GarageKeepersBurglar])
        GarageKeepersRateMod.loc[GarageKeepersRateMod['Type'] == "Central Station", 'Modifier'] = 'Burglar Alarm Modifier'
        GarageKeepersRateMod = GarageKeepersRateMod[['Modifier', 'Type', 'Factor']]
        
        return GarageKeepersRateMod
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildHiredAuto(self):
        HiredAuto = self.buildDataFrame("BP7_HiredAutoLiability")
        return HiredAuto.rename(columns={'Hired Auto Premium' : 'Premium'})
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildNonOwnedAuto(self):
        NonOwnedAuto = self.buildDataFrame("BP7_Non_Owned_Auto_Liability")
        updatedNonOwnedAuto = NonOwnedAuto.fillna({'NoofEmployeesMax': 0}).astype({'NoofEmployeesMin': 'int64', 'NoofEmployeesMax': 'int64'}). \
                astype({'NoofEmployeesMin': 'string', 'NoofEmployeesMax': 'string'}) # Converting to int first to get rid of decimal places
        updatedNonOwnedAuto["Number of Employees"] = updatedNonOwnedAuto["NoofEmployeesMin"] + ' - ' + updatedNonOwnedAuto["NoofEmployeesMax"] # Creating a single column for the percentage
        updatedNonOwnedAuto = updatedNonOwnedAuto.replace({'Number of Employees' : {'1001 - 0' : 'Over 1,000'}}).filter(items=['Number of Employees', 'Limit', 'BaseRate'])
        updatedNonOwnedAuto = updatedNonOwnedAuto.pivot(index='Number of Employees', columns='Limit', values='BaseRate').reset_index('Number of Employees').sort_values(300000)
        return updatedNonOwnedAuto
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildLiquorLiabFood1(self):
        LiquorLiabFood1 = self.buildDataFrame("BP7_LiquorBaseRateForFood")
        LiquorLiabFood1 = LiquorLiabFood1.fillna({'FoodServiceType': 'Other'}).replace({'FoodServiceType' : {'EXQ': 'Exquisite Fine Dining', 'FIN' : 'Fine Dining'}}).rename(columns={'FoodServiceType': 'Coverage', 'LiquorBaseRateForFood' : 'Rate'})
        return LiquorLiabFood1
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildLiquorLiabFood2(self):
        LiquorLiabFood2 = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        LiquorLiabFood2 = LiquorLiabFood2.query(f'CoverageType == "BP7LiquorLiabCov"').filter(items=['Premium']).rename(columns={'Premium': 'Rate'})
        return LiquorLiabFood2
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildLiquorLiabFood3(self):
        LiquorLiabFood3 = self.buildDataFrame("BP7_LiquorLiabilityLimitFactor")
        return LiquorLiabFood3.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'LiquorLiabilityLimitFactor' : 'Liquor Liability Limit Factor'})
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildLiquorLiabAllOther(self):
        LiquorLiabAllOther = self.buildDataFrame("BP7_LiquorBaseRateForOthers")
        return LiquorLiabAllOther.rename(columns={'LiabilityLimitOfInsurance' : 'Liability Limit of Insurance', 'LiquorBaseRateForOthers' : 'Rate (Each Premises)'})
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildLiquorLiabAmendment(self):
        LiquorLiabAmendment = self.buildDataFrame("BP7_Liquor_Base_Rates")
        return LiquorLiabAmendment.rename(columns={'LiabilityLimitOfInsurance' : 'Liability Limit of Insurance', 'Premium' : 'Premium (Each Event)'})
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildTenantsPropDam(self):
        TenantsPropDam = self.buildDataFrame("BP7PerilCWSpecificCovRateFactors")
        return TenantsPropDam.query(f'`CoverageSpecific` == "TenantsPropDamLegalLiability"').filter(items=['Value' ]).rename(columns={'Value': 'Factor'})

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeePracticesOutside(self):
        EmployeePracticesOutside = self.buildDataFrame("BP7_EmployementPracticesLiability_BaseRate")
        updatedEmployeePracticesOutside = EmployeePracticesOutside.fillna({'PrimaryPolicyState' : 'Other'})
        updatedEmployeePracticesOutside = updatedEmployeePracticesOutside.query(f'PrimaryPolicyState == "AR"')
        updatedEmployeePracticesOutside = updatedEmployeePracticesOutside.pivot(index='LimitOfLiability', columns='DeductibleAmt', values='EmploymentPracticesLiabilityBaseRateFactor').reset_index('LimitOfLiability').rename(columns={'LimitOfLiability' : 'Limit of Liability'})
        updatedEmployeePracticesOutside = updatedEmployeePracticesOutside.fillna({2500 : 'N/A', 5000 : 'N/A', 10000 : 'N/A'})
        return updatedEmployeePracticesOutside

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeePracticesInside(self):
        EmployeePracticesInside = self.buildDataFrame("BP7_EmployementPracticesLiability_BaseRate")
        updatedEmployeePracticesInside = EmployeePracticesInside.fillna({'PrimaryPolicyState' : 'Other'})
        updatedEmployeePracticesInside = updatedEmployeePracticesInside.query(f'PrimaryPolicyState == "Other"')
        updatedEmployeePracticesInside = updatedEmployeePracticesInside.pivot(index='LimitOfLiability', columns='DeductibleAmt', values='EmploymentPracticesLiabilityBaseRateFactor').reset_index('LimitOfLiability').rename(columns={'LimitOfLiability' : 'Limit of Liability'})
        updatedEmployeePracticesInside = updatedEmployeePracticesInside.fillna({2500 : 'N/A', 5000 : 'N/A', 10000 : 'N/A'})
        return updatedEmployeePracticesInside

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeeRelatedState(self):
        EmployeeRelatedState = self.buildDataFrame("BP7_EmployementPracticesLiabilityFactor")
        return EmployeeRelatedState.rename(columns={'EmployementPracticesLiabilityFactor' : 'Factor'}).filter(items=['Factor'])
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeeRelatedNAICS(self):
        EmployeeRelatedNAICS = self.buildDataFrame("BP7_EmployementPracticesLiability_NAICS_Factor")
        return EmployeeRelatedNAICS
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildEmployeeRelatedSuppERP(self):
        EmployeeRelatedSuppERP = self.buildDataFrame("BP7 Employment Practices Liability ERP Factor")
        EmployeeRelatedSuppERP = EmployeeRelatedSuppERP.fillna({'State' : 'Other'})
        if self.state != "CT" and self.state != "MD" and self.state != "VA" and self.state != "VT" and self.state != "WY":
            EmployeeRelatedSuppERP['Form'] = 'PB4311'
            EmployeeRelatedSuppERP = EmployeeRelatedSuppERP.query(f'State == "Other"').filter(items=['Form', 'Years In Program < =', 'ERP Period', 'ERP Factor']).rename(columns={'Years In Program < =' : 'Years in Program'})
            updatedEmployeeRelatedSuppERP = EmployeeRelatedSuppERP.pivot(index=['Form', 'Years in Program'], columns='ERP Period', values='ERP Factor').reset_index(['Form', 'Years in Program']).rename(columns={'12': '12 Months', '24' : '24 Months', '36' : '36 Months', '0' : 'Unlimited'}).replace({'Years in Program' : {999 : "3+"}})
            updatedEmployeeRelatedSuppERP = updatedEmployeeRelatedSuppERP[['Form', 'Years in Program', '12 Months', '36 Months']]
        elif self.state == "CT" or self.state == "VT":
            EmployeeRelatedSuppERP['Form'] = 'PB4312'
            EmployeeRelatedSuppERP = EmployeeRelatedSuppERP.query(f'State == "CT"').filter(items=['Form', 'Years In Program < =', 'ERP Period', 'ERP Factor']).rename(columns={'Years In Program < =' : 'Years in Program'})
            updatedEmployeeRelatedSuppERP = EmployeeRelatedSuppERP.pivot(index=['Form', 'Years in Program'], columns='ERP Period', values='ERP Factor').reset_index(['Form', 'Years in Program']).rename(columns={'12': '12 Months', '24' : '24 Months', '36' : '36 Months', '0' : 'Unlimited'}).replace({'Years in Program' : {999 : "3+"}})
            updatedEmployeeRelatedSuppERP = updatedEmployeeRelatedSuppERP[['Form', 'Years in Program', '12 Months', '36 Months']]
        elif self.state == "VA":
            EmployeeRelatedSuppERP['Form'] = 'PB4313'
            EmployeeRelatedSuppERP = EmployeeRelatedSuppERP.query(f'State == "VA"').filter(items=['Form', 'Years In Program < =', 'ERP Period', 'ERP Factor']).rename(columns={'Years In Program < =' : 'Years in Program'})
            updatedEmployeeRelatedSuppERP = EmployeeRelatedSuppERP.pivot(index=['Form', 'Years in Program'], columns='ERP Period', values='ERP Factor').reset_index(['Form', 'Years in Program']).rename(columns={'12': '12 Months', '24' : '24 Months', '36' : '36 Months', '0' : 'Unlimited'}).replace({'Years in Program' : {999 : "3+"}})
            updatedEmployeeRelatedSuppERP = updatedEmployeeRelatedSuppERP[['Form', 'Years in Program', '12 Months', '24 Months', '36 Months']]
        elif self.state == "MD" or self.state == "WY":
            EmployeeRelatedSuppERP['Form'] = 'PB4314'
            EmployeeRelatedSuppERP = EmployeeRelatedSuppERP.query(f'State == "MD"').filter(items=['Form', 'Years In Program < =', 'ERP Period', 'ERP Factor']).rename(columns={'Years In Program < =' : 'Years in Program'})
            updatedEmployeeRelatedSuppERP = EmployeeRelatedSuppERP.pivot(index=['Form', 'Years in Program'], columns='ERP Period', values='ERP Factor').reset_index(['Form', 'Years in Program']).rename(columns={'12': '12 Months', '24' : '24 Months', '36' : '36 Months', '0' : 'Unlimited'}).replace({'Years in Program' : {999 : "3+"}})
            updatedEmployeeRelatedSuppERP = updatedEmployeeRelatedSuppERP[['Form', 'Years in Program', '12 Months', '36 Months', 'Unlimited']]
        return updatedEmployeeRelatedSuppERP
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildCarWashLiab(self):
        CarWashLiab = self.buildDataFrame("BP7_Car_Wash_Base_Rate")
        CarWashLiab = CarWashLiab.query(f'NoOfBays != 6').rename(columns={'NoOfBays' : 'No. of Bays', 'CarWashFactor': 'Rate per Bay'}).replace({'No. of Bays' : {1 : 'One', 2 : 'Two', 3 : 'Three', 4 : 'Four', 5 : 'Five', 6 : 'Six'}}).fillna({'No. of Bays' : 'Six or More'}).sort_values('Rate per Bay', ascending=False)
        return CarWashLiab
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildCarWashDed(self):
        CarWashDed = self.buildDataFrame("BP7_Car_Wash_Deductible_Factor")
        return CarWashDed.rename(columns={'DeductibleFactor' : 'Factor'})

    def buildEmploymentPracticesLiability(self):
        data = pd.DataFrame({
            "Limit": ["$25,000", "$50,000", "$100,000"],
            "Factor ": ["0.07", "0.10", "0.10"]
        })

        return data

    def buildEmploymentTypeFactor(self):

        df_employee_type_factor = pd.DataFrame(
            {
                "Employee Type": [
                    "Full-time (32 or more hours)",
                    "Part-time (less than 32 hours)",
                    "Temporary workers/leased workers",
                    "Endorsed Independent contrators",  # spelled as provided
                    "Independent Contractors - unendorsed",
                ],
                "Factor": ['1.00', '0.75', '0.75', '0.75', '0.10'],
            }
        )

        return df_employee_type_factor

    def buildDefenseInside(self):

        df_employee_band_rates = pd.DataFrame(
            {
                "Band Label": ["1 - 100", "The next 150"],
                "Band Range": ["1 - 100", "101 - 250"],
                "Rate": ['54.62', '32.77'],
            }
        )

        return df_employee_band_rates

    def buildDefesneWithin(self):

        data = pd.DataFrame(
            {
                "Limit": ['50,000', '100,000', '250,000', '500,000', '1,000,000'],
                "2,500": ['45.90', '54.62', '83.38', 'NA', 'NA'],
                "5,000": ['43.97', '52.70', '81.01', '102.69', '124.83'],
                "10,000": ['41.08', '49.60', '77.05', '98.44', '120.41'],
                "25,000": ['35.07', '42.72', '67.86', '88.42', '109.91'],
            }
        )

        return data

    def buildMatchingDefense(self):

        data = pd.DataFrame(
            {
                "Indemnity Limit": ['50,000', '100,000', '250,000', '500,000', '1,000,000'],
                "ALAE Limit": ['50,000', '100,000', '250,000', '500,000', '1,000,000'],
                "2,500": ['62.38', '69.86',' 98.66', 'NA', 'NA'],
                "5,000": ['60.07', '67.69', '96.19', '118.19', '138.29'],
                "10,000": ['56.46', '64.09', '92.02', '113.82', '133.80'],
                "25,000": ['48.54', '55.89', '82.25', '103.46', '123.11'],
            }
        )

        return data

    def buildEmploymentPractFactor(self):
        df_factor = pd.DataFrame({"Factor": ['0.600']})

        return df_factor

    def buildEmploymentPracticesLiabilityState(self):

        _state_pairs_in_order = [
            ("AL", 0.847), ("KS", 0.858), ("NJ", 1.507), ("VT", 0.627),
            ("AR", 0.759), ("KY", 0.561), ("NM", 1.023), ("WA", 0.726),
            ("AZ", 0.913), ("MA", 'NA'), ("NV", 'NA'), ("WI", 0.715),
            ("CA", 2.717), ("MD", 0.858), ("NY", 1.199), ("WV", 0.616),
            ("CO", 0.814), ("ME", 0.660), ("OH", 0.968), ("WY", 0.715),
            ("CT", 1.188), ("MI", 1.000), ("OR", 0.814),
            ("DC", 1.903), ("MN", 0.616), ("PA", 0.979),
            ("DE", 0.880), ("MO", 1.067), ("RI", 0.451),
            ("FL", 0.847), ("MS", 0.814), ("SC", 0.781),
            ("GA", 1.023), ("MT", 0.605), ("SD", 0.286),
            ("IA", 0.704), ("NC", 0.858), ("TN", 0.869),
            ("ID", 0.572), ("ND", 0.363), ("TX", 1.000),
            ("IL", 1.012), ("NE", 0.880), ("UT", 0.561),
            ("IN", 0.715), ("NH", 0.473), ("VA", 1.045),
        ]

        # Convert to strings; blank for missing numeric values
        pairs_as_strings = [
            (state, str(val))
            for state, val in _state_pairs_in_order
        ]

        # Chunk into rows of 4 (State, Relativity) pairs
        pairs_per_row = 4
        rows: list[list[str]] = []
        for i in range(0, len(pairs_as_strings), pairs_per_row):
            chunk = pairs_as_strings[i:i + pairs_per_row]
            # Pad the last row with empty pairs if needed
            while len(chunk) < pairs_per_row:
                chunk.append(("", ""))
            # Flatten: [State1, Rel1, State2, Rel2, State3, Rel3, State4, Rel4]
            flat = []
            for s, r in chunk:
                flat.extend([s, r])
            rows.append(flat)

        # Wide DataFrame: all string values, four state/relativity pairs per row
        columns = ["State", "Factor", "State", "Factor", "State", "Factor", "State", "Factor"]
        df_state_relativities_wide = pd.DataFrame(rows, columns=columns)

        # Ensure dtype is string (object-strings)
        df_state_relativities_wide = df_state_relativities_wide.astype(str)

        return df_state_relativities_wide

    def buildEmploymentPracticesLiabilityThirdParty(self):
        df_factor = pd.DataFrame({"Factor": ['0.15']})

        return df_factor

    def buildLimitedCoverageUnmannedAircraft(self):

        data = pd.DataFrame({
            "Coverage": ["BI/PD", "P/AI"],
            "Rates": ["$150", "$100"]
        })

        return data

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAdvantageRate(self):
        AdvantageRate = self.buildDataFrame("BP7_Businessowners_Advantage_Rate")
        AdvantageRate = AdvantageRate.rename(columns={'EmployeeDishonestyChoice' : 'Employee Dishonesty Choice', 'Class_Code_Min' : 'Program', 'BusinessownerAdvRate' : 'Rate'}).replace({'Program': self.classCodes}).replace({'Program' : {'Hab' : 'Habitational', 'Food' : 'Food Service', 'Auto' : 'Auto Service'}}).sort_values('Program').filter(items=['Employee Dishonesty Choice', 'Program', 'Rate']).replace({'Employee Dishonesty Choice' : {'ExcludesEmployeeDishonesty' : 'Without Employee Dishonesty Rate', 'IncludesEmployeeDishonesty' : 'With Limited Employee Dishonesty Rate'}})
        pivotedAdvantageRate = AdvantageRate.pivot(index='Program', columns='Employee Dishonesty Choice', values='Rate').reset_index('Program')
        pivotedAdvantageRate = pivotedAdvantageRate[['Program', 'Without Employee Dishonesty Rate', 'With Limited Employee Dishonesty Rate']]
        return pivotedAdvantageRate
    
# Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildAdvantageILF(self):
        AdvantageILF = self.buildDataFrame("BP7_Businessowners_Advantage_Increased_Limit")
        AdvantageILF = AdvantageILF.rename(columns={'LimitofInsurance' : 'Limit of Insurance', 'Class_Code_Min' : 'Program', 'IncreasedLimitFactor' : 'Increased Limit Factor'}).filter(items=['Limit of Insurance', 'Program', 'Increased Limit Factor']).replace({'Program': self.classCodes}).replace({'Program' : {'Hab' : 'Habitational', 'Food' : 'Food Service', 'Auto' : 'Auto Service'}})
        pivotedAdvantageILF = AdvantageILF.pivot(index='Limit of Insurance', columns='Program', values='Increased Limit Factor').reset_index('Limit of Insurance')
        return pivotedAdvantageILF

    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildCyberSuite(self):
        CyberSuite = self.buildDataFrame("BP7_Cyber_Suite_Premium")
        filteredCyberSuite = CyberSuite.rename(columns={'ProgramCodeDisplay Name' : 'Program', 'DeductibleAnnualAggrLimit' : 'Aggregate Limit / Deductible', 'CyberSuiteCovPremium' : 'Premium'}).filter(items=['Program', 'Aggregate Limit / Deductible', 'Premium'])
        pivotedCyberSuite = filteredCyberSuite.pivot(index='Aggregate Limit / Deductible', columns='Program', values='Premium').reset_index('Aggregate Limit / Deductible').replace({'Aggregate Limit / Deductible' : {50000 : '$50,000 / $1,000', 100000 : '100,000 / 1,000', 250000 : '250,000 / 1,000', 500000 : '500,000 / 5,000', 1000000 : '1,000,000 / 10,000'}})
        return pivotedCyberSuite.rename(columns={'Auto Service' : 'Auto', 'Food Service' : 'Food'})
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildCyberSuiteThird(self):
        CyberSuiteThird = self.buildDataFrame("BP7_Cyber_Suite_3Party_Premium")
        filteredCyberSuiteThird = CyberSuiteThird.rename(columns={'ProgramCodeDisplay Name' : 'Program', 'DeductibleAnnualAggrLimit' : 'Aggregate Limit / Deductible', 'CyberSuiteCovPremium' : 'Premium'}).filter(items=['Program', 'Aggregate Limit / Deductible', 'Premium'])
        pivotedCyberSuiteThird = filteredCyberSuiteThird.pivot(index='Aggregate Limit / Deductible', columns='Program', values='Premium').reset_index('Aggregate Limit / Deductible').replace({'Aggregate Limit / Deductible' : {50000 : '$50,000 / $1,000', 100000 : '100,000 / 1,000', 250000 : '250,000 / 1,000', 500000 : '500,000 / 5,000', 1000000 : '1,000,000 / 10,000'}})
        return pivotedCyberSuiteThird.rename(columns={'Auto Service' : 'Auto', 'Food Service' : 'Food'})
    

    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildCyberSuiteSubLimit1(self):
        data = {'Aggregate Limit/Deductible' : ['$50,000 / $1,000', '100,000 / 1,000', '250,000 / 1,000', '500,000 / 5,000', '1,000,000 / 10,000'], 'Forensic IT Review, Legal Review, Regulatory Fines & Penalties, PCI Fines & Penalties' : ['$25,000', '$50,000', '$125,000', '$250,000', '$500,000'], 'DC RE' : ['$5,000', '$5,000', '$5,000', '$5,000', '$5,000'], 'CA' : ['$5,000', '$5,000', '$5,000', '$5,000', '$5,000'], 'Cyber Extortion' : ['$10,000', '$10,000', '$25,000', '$25,000', '$25,000'], 'Misdirected Payment Fraud' : ['$10,000', '$10,000', '$25,000', '$25,000', '$25,000'], 'Computer Fraud' : ['$10,000', '$10,000', '$25,000', '$25,000', '$25,000']}
        CyberSuiteSubLimit1  = pd.DataFrame(data)
        return CyberSuiteSubLimit1
    
    # Builds the Forgery and Alteration factor table
    # Returns a dataframe
    def buildCyberSuiteSubLimit2(self):
        data = [['$25,000', '$5,000', '$1,000', '$1,000']]
        CyberSuiteSubLimit2  = pd.DataFrame(data, columns=['Annual Aggregate Limit', 'Lost Wages and Child or Elder Care', 'Mental Health Counseling', 'Miscellaneous Expense'])
        return CyberSuiteSubLimit2
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildGarageKeepersTerritoryMult(self):
        data = [['A000A000A000A000', 1]]
        GKMultipliers  = pd.DataFrame(data, columns=['Territory', 'Building'])
        return GKMultipliers
    
    # Converts the given pixels to inches
    # Returns a decimal 
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)
    
    def formatAccountsReceivableILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatMoneyILF(self, ws, boldFont):
        for cell in ws['22:22']:
            cell.border = None
            #cell.font = boldFont
            #cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col <= 3:
                ws.column_dimensions[char].width = self.pixelsToInches(110)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 3: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col >= 3: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatOutdoorSignsILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatOutdoorTreesILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatValuablePapersILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatBackupSewerILF(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Increased Limit Increments'
        ws.merge_cells('A3:B3')
        ws['C3'] = 'Total Limit'
        ws.merge_cells('C3:D3')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(5, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 5: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
    
    def formatAwayFromPremisesILF(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(225)
        #for row in range(4, ws.max_row + 1):
        #    cell = ws['B' + str(row)]
        #    cell.number_format = self.currencywdecFormat # Applying no decimal formatting for column B

    def formatElectronicDataILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 3: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col >= 3: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatInterruptionILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 3: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col >= 3: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatBLDGPropertyILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatComputerFraudILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 2: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col >= 2: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatCondoLossAsses(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 2: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col >= 2: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatCondo(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatEQPropertyDed(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A-B
                elif col >= 4: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(140)

    def formatEQSprinkler(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B

    def formatEQDeductibleOptions(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A-B
                elif col >= 3: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B
    
    def formatEQMasonryVeneer(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatEQCoinsurance(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatEQSprinkleredRisk(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatEQBuildingHeight(self, ws, boldFont):
        ws.insert_rows(3)
        ws['B3'] = '4-7 Stories'
        ws.merge_cells('B3:D3')
        ws['E3'] = '8 Or More Stories'
        ws.merge_cells('E3:G3')
        ws['B4'] = 'Tier 1'
        ws['C4'] = 'Tier 2'
        ws['D4'] = 'Tier 3'
        ws['E4'] = 'Tier 1'
        ws['F4'] = 'Tier 2'
        ws['G4'] = 'Tier 3'

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    def formatEQClassRated(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)


    def formatEQClassRatedID(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)


    def formatEQClassRatedNH(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    
    def formatEQClassRatedMS(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(59)
        ws.insert_rows(59)
        ws['C60'] = 'Territory ' + Territory4 + ' Loss Costs'
        ws.merge_cells('C60:G60')
        ws['D61'] = 'Contents Grade'
        ws.merge_cells('D61:G61')

        for cell in ws['60:60']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['61:61']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['62:62']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    def formatEQClassRatedKY(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(59)
        ws.insert_rows(59)
        ws['C60'] = 'Territory ' + Territory4 + ' Loss Costs'
        ws.merge_cells('C60:G60')
        ws['D61'] = 'Contents Grade'
        ws.merge_cells('D61:G61')

        for cell in ws['60:60']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['61:61']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['62:62']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(78)
        ws.insert_rows(78)
        ws['C79'] = 'Territory ' + Territory5 + ' Loss Costs'
        ws.merge_cells('C79:G79')
        ws['D80'] = 'Contents Grade'
        ws.merge_cells('D80:G80')

        for cell in ws['79:79']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['80:80']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['81:81']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            
    def formatEQClassRatedIL(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(59)
        ws.insert_rows(59)
        ws['C60'] = 'Territory ' + Territory4 + ' Loss Costs'
        ws.merge_cells('C60:G60')
        ws['D61'] = 'Contents Grade'
        ws.merge_cells('D61:G61')

        for cell in ws['60:60']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['61:61']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['62:62']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(78)
        ws.insert_rows(78)
        ws['C79'] = 'Territory ' + Territory5 + ' Loss Costs'
        ws.merge_cells('C79:G79')
        ws['D80'] = 'Contents Grade'
        ws.merge_cells('D80:G80')

        for cell in ws['79:79']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['80:80']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['81:81']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(97)
        ws.insert_rows(97)
        ws['C98'] = 'Territory ' + Territory6 + ' Loss Costs'
        ws.merge_cells('C98:G98')
        ws['D99'] = 'Contents Grade'
        ws.merge_cells('D99:G99')

        for cell in ws['98:98']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['99:99']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['100:100']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        

    def formatEQClassRatedAR(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(59)
        ws.insert_rows(59)
        ws['C60'] = 'Territory ' + Territory4 + ' Loss Costs'
        ws.merge_cells('C60:G60')
        ws['D61'] = 'Contents Grade'
        ws.merge_cells('D61:G61')

        for cell in ws['60:60']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['61:61']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['62:62']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(78)
        ws.insert_rows(78)
        ws['C79'] = 'Territory ' + Territory5 + ' Loss Costs'
        ws.merge_cells('C79:G79')
        ws['D80'] = 'Contents Grade'
        ws.merge_cells('D80:G80')

        for cell in ws['79:79']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['80:80']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['81:81']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(97)
        ws.insert_rows(97)
        ws['C98'] = 'Territory ' + Territory6 + ' Loss Costs'
        ws.merge_cells('C98:G98')
        ws['D99'] = 'Contents Grade'
        ws.merge_cells('D99:G99')

        for cell in ws['98:98']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['99:99']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['100:100']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(116)
        ws.insert_rows(116)
        ws['C117'] = 'Territory ' + Territory7 + ' Loss Costs'
        ws.merge_cells('C117:G117')
        ws['D118'] = 'Contents Grade'
        ws.merge_cells('D118:G118')

        for cell in ws['117:117']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['118:118']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['119:119']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
            
    def formatEQClassRatedUT(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(59)
        ws.insert_rows(59)
        ws['C60'] = 'Territory ' + Territory4 + ' Loss Costs'
        ws.merge_cells('C60:G60')
        ws['D61'] = 'Contents Grade'
        ws.merge_cells('D61:G61')

        for cell in ws['60:60']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['61:61']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['62:62']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(78)
        ws.insert_rows(78)
        ws['C79'] = 'Territory ' + Territory5 + ' Loss Costs'
        ws.merge_cells('C79:G79')
        ws['D80'] = 'Contents Grade'
        ws.merge_cells('D80:G80')

        for cell in ws['79:79']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['80:80']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['81:81']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(97)
        ws.insert_rows(97)
        ws['C98'] = 'Territory ' + Territory6 + ' Loss Costs'
        ws.merge_cells('C98:G98')
        ws['D99'] = 'Contents Grade'
        ws.merge_cells('D99:G99')

        for cell in ws['98:98']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['99:99']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['100:100']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(116)
        ws.insert_rows(116)
        ws['C117'] = 'Territory ' + Territory7 + ' Loss Costs'
        ws.merge_cells('C117:G117')
        ws['D118'] = 'Contents Grade'
        ws.merge_cells('D118:G118')

        for cell in ws['117:117']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['118:118']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['119:119']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(135)
        ws.insert_rows(135)
        ws['C136'] = 'Territory ' + Territory8 + ' Loss Costs'
        ws.merge_cells('C136:G136')
        ws['D137'] = 'Contents Grade'
        ws.merge_cells('D137:G137')

        for cell in ws['136:136']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['137:137']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['138:138']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    def formatEQClassRatedCA(self, ws, boldFont):
        ws.insert_rows(3)
        ws['C3'] = 'Territory ' + Territory1 + ' Loss Costs'
        ws.merge_cells('C3:G3')
        ws.insert_rows(4)
        ws['D4'] = 'Contents Grade'
        ws.merge_cells('D4:G4')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(21)
        ws.insert_rows(21)
        ws['C22'] = 'Territory ' + Territory2 + ' Loss Costs'
        ws.merge_cells('C22:G22')
        ws['D23'] = 'Contents Grade'
        ws.merge_cells('D23:G23')

        for cell in ws['22:22']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['23:23']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['24:24']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(40)
        ws.insert_rows(40)
        ws['C41'] = 'Territory ' + Territory3 + ' Loss Costs'
        ws.merge_cells('C41:G41')
        ws['D42'] = 'Contents Grade'
        ws.merge_cells('D42:G42')

        for cell in ws['41:41']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['42:42']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['43:43']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(59)
        ws.insert_rows(59)
        ws['C60'] = 'Territory ' + Territory4 + ' Loss Costs'
        ws.merge_cells('C60:G60')
        ws['D61'] = 'Contents Grade'
        ws.merge_cells('D61:G61')

        for cell in ws['60:60']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['61:61']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['62:62']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(78)
        ws.insert_rows(78)
        ws['C79'] = 'Territory ' + Territory5 + ' Loss Costs'
        ws.merge_cells('C79:G79')
        ws['D80'] = 'Contents Grade'
        ws.merge_cells('D80:G80')

        for cell in ws['79:79']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['80:80']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['81:81']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(97)
        ws.insert_rows(97)
        ws['C98'] = 'Territory ' + Territory6 + ' Loss Costs'
        ws.merge_cells('C98:G98')
        ws['D99'] = 'Contents Grade'
        ws.merge_cells('D99:G99')

        for cell in ws['98:98']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['99:99']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['100:100']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(116)
        ws.insert_rows(116)
        ws['C117'] = 'Territory ' + Territory7 + ' Loss Costs'
        ws.merge_cells('C117:G117')
        ws['D118'] = 'Contents Grade'
        ws.merge_cells('D118:G118')

        for cell in ws['117:117']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['118:118']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['119:119']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(135)
        ws.insert_rows(135)
        ws['C136'] = 'Territory ' + Territory8 + ' Loss Costs'
        ws.merge_cells('C136:G136')
        ws['D137'] = 'Contents Grade'
        ws.merge_cells('D137:G137')

        for cell in ws['136:136']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['137:137']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['138:138']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(154)
        ws.insert_rows(154)
        ws['C155'] = 'Territory ' + Territory9 + ' Loss Costs'
        ws.merge_cells('C155:G155')
        ws['D156'] = 'Contents Grade'
        ws.merge_cells('D156:G156')

        for cell in ws['155:155']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['156:156']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['157:157']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(173)
        ws.insert_rows(173)
        ws['C174'] = 'Territory ' + Territory10 + ' Loss Costs'
        ws.merge_cells('C174:G174')
        ws['D175'] = 'Contents Grade'
        ws.merge_cells('D175:G175')

        for cell in ws['174:174']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['175:175']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['176:176']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(192)
        ws.insert_rows(192)
        ws['C193'] = 'Territory ' + Territory11 + ' Loss Costs'
        ws.merge_cells('C193:G193')
        ws['D194'] = 'Contents Grade'
        ws.merge_cells('D194:G194')

        for cell in ws['193:193']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['194:194']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['195:195']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(211)
        ws.insert_rows(211)
        ws['C212'] = 'Territory ' + Territory12 + ' Loss Costs'
        ws.merge_cells('C212:G212')
        ws['D213'] = 'Contents Grade'
        ws.merge_cells('D213:G213')

        for cell in ws['212:212']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['213:213']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['214:214']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(230)
        ws.insert_rows(230)
        ws['C231'] = 'Territory ' + Territory13 + ' Loss Costs'
        ws.merge_cells('C231:G231')
        ws['D232'] = 'Contents Grade'
        ws.merge_cells('D232:G232')

        for cell in ws['231:231']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['232:232']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['233:233']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(249)
        ws.insert_rows(249)
        ws['C250'] = 'Territory ' + Territory14 + ' Loss Costs'
        ws.merge_cells('C250:G250')
        ws['D251'] = 'Contents Grade'
        ws.merge_cells('D251:G251')

        for cell in ws['250:250']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['251:251']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['252:252']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(268)
        ws.insert_rows(268)
        ws['C269'] = 'Territory ' + Territory15 + ' Loss Costs'
        ws.merge_cells('C269:G269')
        ws['D270'] = 'Contents Grade'
        ws.merge_cells('D270:G270')

        for cell in ws['269:269']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['270:270']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['271:271']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(287)
        ws.insert_rows(287)
        ws['C288'] = 'Territory ' + Territory16 + ' Loss Costs'
        ws.merge_cells('C288:G288')
        ws['D289'] = 'Contents Grade'
        ws.merge_cells('D289:G289')

        for cell in ws['288:288']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['289:289']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['290:290']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(306)
        ws.insert_rows(306)
        ws['C307'] = 'Territory ' + Territory17 + ' Loss Costs'
        ws.merge_cells('C307:G307')
        ws['D308'] = 'Contents Grade'
        ws.merge_cells('D308:G308')

        for cell in ws['307:307']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['308:308']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['309:309']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

            
    def formatEQLCM(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(225)

    def formatEQSprinklerLeakCoinsurance(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatEmployeeDishonesty(self, ws, boldFont):
        ws.insert_rows(3)
        ws['B3'] = 'Premium'
        ws.merge_cells('B3:C3')

        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col < 2: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col >= 2: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatExtendedBusinessIncome(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A-B

    def formatOrdinanceEndorsement(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatEQTerritoryDefs(self, ws):
        ws.column_dimensions['C'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = self.ZipCodeFormat # Applying no decimal formatting for column A
        for row in range(4, ws.max_row + 1):
            cell = ws['B' + str(row)]
            cell.number_format = self.ZipCodeFormat # Applying no decimal formatting for column A
        for row in range(4, ws.max_row + 1):
            cell = ws['C' + str(row)]
            cell.number_format = self.ZipCodeFormat # Applying no decimal formatting for column A

    def formatSpoilage(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatUtilityServices(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Communication Supply and Power Supply Property'
        ws.insert_rows(4)
        ws['B4'] = 'Not including overhead transmission lines'
        ws.insert_rows(9)
        ws['B9'] = 'Including overhead transmission lines'
        ws.insert_rows(14)
        ws['A14'] = 'Water Supply Property'

        ws.merge_cells('A3:C3')
        ws.merge_cells('B4:C4')
        ws.merge_cells('B9:C9')
        ws.merge_cells('A14:B14')

        for cell in ws['3:3']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['4:4']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['9:9']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['10:10']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['14:14']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['15:15']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col <= 1:
                ws.column_dimensions[char].width = self.pixelsToInches(110)
            elif col > 1: 
                ws.column_dimensions[char].width = self.pixelsToInches(150)

        for cell in ws['8:8']:
            cell.border = None
        for cell in ws['13:13']:
            cell.border = None
        for row in range(15, ws.max_row + 1):
            cell = ws['C' + str(row)]
            cell.border = None # Applying no decimal formatting for column A

    def formatScheduledPropFloater(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B
        ws['B4'] = '(per $100)'

    def formatVehicleDamBase(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatExclusionRoofSiding(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatFloodSubLimit(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatFloodBLDGConstruction(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 3: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatFloodBPP(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 3: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatFloodDedFactors(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatFloodMinAdjRate(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 3: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-B

    def formatFloodAggLimit(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.twodecimal # Applying currency formatting to columns A-
                    
    def formatFloodMinPrem(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 3: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatAmendmentAdditional(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatAdditionalFairs(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatAdditionalServices(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatAdditionalVendors(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatAdditionalDesignated(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatEmployeeBodilyInj(self, ws, boldFont):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(7, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

        for cell in ws['5:5']:
            cell.border = None

        for cell in ws['6:6']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    def formatEmployeeBenefitsLiab(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A-B

    def formatGarageKeepersBase(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatGarageKeepersDed(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B

    def formatGarageKeepersLOI(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B

    def formatGarageKeepersRateMod(self, ws):
        ws.merge_cells('A6:A7')
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    #cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def formatHiredAuto(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col ==2:
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatNonOwnedAuto(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatLiquorLiabFood(self, ws, boldFont):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(13, 14):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

            for row in range(18, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B

        for cell in ws['3:3']:
            cell.border = None
        for cell in ws['4:4']:
            cell.border = None
        for cell in ws['10:10']:
            cell.border = None
        for cell in ws['11:11']:
            cell.border = None
        for cell in ws['15:15']:
            cell.border = None
        for cell in ws['16:16']:
            cell.border = None

        for cell in ws['3:3']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        for cell in ws['5:5']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['10:10']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        for cell in ws['12:12']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        
        for cell in ws['15:15']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

        for cell in ws['17:17']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.column_dimensions['A'].width = self.pixelsToInches(175)

    def formatLiquorLiabAllOther(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col == 2:
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatLiquorLiabAmendment(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col == 2:
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatEmployeePracticesOutside(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col > 1:
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatEmployeePracticesInside(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
                elif col > 1:
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatEmploymentPracticesLiability(self, ws):
        ws.insert_rows(2,3)
        ws["A3"] = "Apply the following factor to the fully developed EPLI premium calculated from OC Table D.22.A.4, to arrive at the"
        ws["A4"] = "endorsement premium"
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1:
                    cell.number_format = self.currencyFormat # Applying currency formatting

    def formatEmploymentPracticesLiabilityNamedIndCont(self, ws):
        t = Side(style='thin', color='C1C1C1');
        bd = Border(left=t, right=t, top=t, bottom=t)

        def box(rng):
            for row in ws[rng]:
                for c in row:
                    if c.value is not None and str(c.value) != "":
                        c.border = bd

        def bold_row(rng):
            for c in ws[rng][0]:
                if c.value is not None and str(c.value) != "":
                    c.font = Font(bold=True)

        # ---------- Insert missing rule text (hard-coded positions, bottom-safe ordering) ----------
        ws.insert_rows(10, 2)  # before band section
        ws['A10'] = 'Standard Rule - Defense Inside the Limit (DIL) Coverage'
        ws['A11'] = 'Base limit of $100,000, deductible of $2,500'
        ws['A10'].font = Font(bold=True)

        ws.insert_rows(16, 1)  # before DIL rate table
        ws['A16'] = 'Rates for Defense Within Limits - First 100 Employees'
        ws['A16'].font = Font(bold=True)

        ws.insert_rows(24, 1)  # before DOL table
        ws['A24'] = 'Exception Rule - Matching Defense Limit (DOL) Coverage (as required by statute) - Rates First 100 Employees'
        ws['A24'].font = Font(bold=True)

        ws.insert_rows(32, 1)  # before 101-250 factor
        ws['A32'] = 'For 101 - 250 Employees, apply the following factor to arrive at the corresponding rate from tables above:'

        # state header was originally at row 31 -> after inserts it is at row 35; place section label above it
        ws.insert_rows(35, 1)
        ws['A35'] = 'State Relativities'
        ws['A35'].font = Font(bold=True)

        # ---------- Header merges (existing) ----------
        ws.merge_cells('A1:H1');
        ws['A1'].font = Font(bold=True)
        ws.merge_cells('A3:B3');
        bold_row('A3:B3')  # "Employee Type  Factor"



        # ---------- Borders for each table (new row indices) ----------
        for rng in ['A3:B8', 'A12:C14', 'A17:E22', 'A25:F30', 'A33:B34', 'A37:H49']:
            box(rng)

        # ---------- Bold header rows ----------
        bold_row('A12:C12')  # Band header
        bold_row('A17:E17')  # DIL header
        bold_row('A25:F25')  # DOL header
        ws['A33'].font = Font(bold=True)  # "Factor"
        bold_row('A36:H36')  # State table header

        # ---------- Number formats ----------
        CUR, F3, R2 = '$#,##0', '0.000', '$0.00'

        # Employee Type factors
        for r in range(4, 9):
            v = ws[f'B{r}'].value
            ws[f'B{r}'].number_format = F3

        # DIL table
        for c in 'BCDE':
            v = ws[f'{c}17'].value
            ws[f'{c}17'].number_format = CUR  # deductibles
        for r in range(18, 23):
            if isinstance(ws[f'A{r}'].value, (int, float)): ws[f'A{r}'].number_format = CUR  # limits
            for c in 'BCDE':
                v = ws[f'{c}{r}'].value
                ws[f'{c}{r}'].number_format = R2  # rates

        # DOL table
        for c in 'CDEF':
            v = ws[f'{c}25'].value
            ws[f'{c}25'].number_format = CUR  # deductibles row
        for r in range(26, 31):
            for c in 'AB':  # indemnity/ALAE limits
                v = ws[f'{c}{r}'].value
                ws[f'{c}{r}'].number_format = CUR
            for c in 'CDEF':  # rates
                v = ws[f'{c}{r}'].value
                ws[f'{c}{r}'].number_format = R2

        # 101–250 factor
        if isinstance(ws['A34'].value, (int, float)): ws['A34'].number_format = F3

        # State relativities (B,D,F,H numeric)
        for r in range(37, 50):
            for c in 'BDFH':
                v = ws[f'{c}{r}'].value
                if isinstance(v, (int, float)): ws[f'{c}{r}'].number_format = F3

        # Final edits
        ws.insert_rows(17,1)
        ws["B17"] = "Deductible"
        ws.merge_cells('B17:E17')
        bold_row('B17:E17')

        ws.insert_rows(26,1)
        ws["C26"] = "Deductible"
        ws.merge_cells('C26:F26')
        bold_row('C26:F26')

        bold_row('A34:B34')
        for rng in ['B17:E17', 'C26:F26']:
            box(rng)

        ws.insert_rows(37,1)
        bold_row('A40:H40')

    def formatEmploymentPracticesLiabilityThirdParty(self, ws):
        ws.insert_rows(2,3)
        ws["A3"] = "Apply the following factor to the fully developed EPLI premium calculated from OC Table D.22.A.4, to arrive at the"
        ws["A4"] = "endorsement premium"

    def formatEmployeeRelatedSuppERP(self, ws):
        ws.merge_cells('A4:A6')
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col == 2: 
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A-B

    def formatCarWashLiab(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 2: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A-B

    def formatCarWashDed(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.noDecimalFormat # Applying currency formatting to columns A

    def formatAdvantageRate(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A

    def formatAdvantageILF(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A

    def formatCyberSuite(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A

    def formatCyberSuiteThird(self, ws, boldFont):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencywdecFormat # Applying currency formatting to columns A

        for cell in ws['9:9']:
            cell.border = None
        for cell in ws['16:16']:
            cell.border = None

        for cell in ws['10:10']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['17:17']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        ws.insert_rows(9)
        ws.insert_rows(10)
        ws['A10'] = 'Cyber Suite Sublimit Table'
        ws.insert_rows(18)
        ws.insert_rows(19)
        ws['A19'] = '- Identity Recovery Sublimit Table'

        for cell in ws['10:10']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            #cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['19:19']:
            #cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
            #                    right=Side(border_style='thin', color='C1C1C1'), 
            #                    top=Side(border_style='thin', color='C1C1C1'), 
            #                    bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            #cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

    def formatGarageKeepersTerritoryMult(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(175)

    def formatBusinessIncomeExtraExpenses(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        ws.column_dimensions['B'].width = self.pixelsToInches(125)
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1:
                    cell.number_format = self.currencywdecFormat  # Applying currency formatting to columns A-B

    def formatLimitedCoverageUnmannedAircraft(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)
        ws.column_dimensions['B'].width = self.pixelsToInches(125)
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)  # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col == 1:
                    cell.number_format = self.currencywdecFormat  # Applying currency formatting to columns A-B


    # Sets up the Auto Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildOptionalCoveragesPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        OptionalCoverages = ExcelSettings.Excel(state=self.state, programName='Optional Coverages', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = OptionalCoverages.getFontName()
        fontSize = OptionalCoverages.getFontSize()

        OptionalCoverages.generateWorksheet('ARIL', 'OC Table B.1.D.1. Accounts Receivable - Increased Limit', self.buildAccountsReceivableILF(), False, True)
        OptionalCoverages.generateWorksheet('FAIL', 'OC Table B.2.D.1. Forgery and Alteration - Increased Limit', self.buildForgeryILF(), False, True)
        OptionalCoverages.generateWorksheet2tables('MSIL', 'OC Table B.3.D.1. Money and Securities - Increased Limit', self.buildMoneyInside(), self.buildMoneyOutside(), False, True)
        OptionalCoverages.generateWorksheet('OSIL', 'OC Table B.4.D.1. Outdoor Signs - Increased Limit', self.buildOutdoorSignsILF(), False, True)
        OptionalCoverages.generateWorksheet('OTIL', 'OC Table B.5.D.1. Outdoor Trees, Shrubs, Plants, and Lawns - Increased Limit', self.buildOutdoorTreesILF(), False, True)
        OptionalCoverages.generateWorksheet('VPRIL', 'OC Table B.6.D.1. Valuable Papers and Records - Increased Limit', self.buildValuablePapersILF(), False, True)
        OptionalCoverages.generateWorksheet('BSWIL', 'OC Table B.7.D.1. Back Up of Sewer and Drain Water Damage - Increased Limit', self.buildBackupSewerILF(), False, True)
        OptionalCoverages.generateWorksheet('PPAIL', 'OC Table B.8.D. Business Personal Property Away from Premises - Increased Limit', self.buildAwayFromPremisesILF(), False, True)
        OptionalCoverages.generateWorksheet('EDIL', 'OC Table B.9.D.1. Electronic Data - Increased Limit', self.buildElectronicDataILF(), False, True)
        OptionalCoverages.generateWorksheet('ICOIL', 'OC Table B.10.C.1 Interruption of Computer Operations - Increased Limit', self.buildInterruptionILF(), False, True)
        OptionalCoverages.generateWorksheet('BPOIL', 'OC Table B.11.D.1 Building Property of Others - Increased Limit', self.buildBLDGPropertyILF(), False, True)
        OptionalCoverages.generateWorksheet('FTFIL', 'OC Table B.12.C.1 Computer Fraud and Funds Transfer Fraud - Increased Limit', self.buildComputerFraudILF(), False, True)
        OptionalCoverages.generateWorksheet('PPSIL', 'OC Table B.13.D.1 Business Personal Property Temporarily in Portable Storage Units - Increased Limit', self.buildBPPStorage(), False, True)
        OptionalCoverages.generateWorksheet('CCULA', 'OC Table C.3.C.1.a. Condominium Commercial Unit-Owners Optional Coverages - Loss Assessment', self.buildCondoLossAsses(), False, True)
        OptionalCoverages.generateWorksheet('CCUOC', 'OC Table C.3.C.2.a. Condominium Commercial Unit-Owners Optional Coverages', self.buildCondo(), False, True)
        OptionalCoverages.generateWorksheet('EPDF', 'OC Table C.4.B.1. Earthquake and Volcanic Eruption - Property Deductible Factor', self.buildEQPropertyDed(), False, True)
        OptionalCoverages.generateWorksheet('ESLPD', 'OC Table C.4.B.2. Earthquake and Volcanic Eruption - Earthquake Sprinkler Leakage Property Deductible Factor', self.buildEQSprinkler(), False, True)
        OptionalCoverages.generateWorksheet('EXFBV', 'OC Table C.4.D.1. Earthquake and Volcanic Eruption - Earthquake, other than Functional Building Valuation ', self.buildEQotherthanfunctional(), False, True)
        OptionalCoverages.generateWorksheet('EFBV', 'OC Table C.4.D.2. Earthquake and Volcanic Eruption - Functional Building Valuation', self.buildEQFunctional(), False, True) 
        OptionalCoverages.generateWorksheet('ELU', 'OC Table C.4.D.3.a. Earthquake and Volcanic Eruption - Coverage for Loss to the undamaged portion of the Building', self.buildEQUndamagedLoss(), False, True)
        OptionalCoverages.generateWorksheet('ESL', 'OC Table C.4.D.5. Earthquake and Volcanic Eruption - Earthquake Sprinkler Leakage Only', self.buildEQSprinklerLeakage(), False, True)
        OptionalCoverages.generateWorksheet('EDO', 'OC Table C.4.E.2.c.3 Earthquake and Volcanic Eruption - Earthquake Deductible Options', self.buildEQDeductibleOptions(), False, True)
        OptionalCoverages.generateWorksheet('ETD', 'OC Table C.4.E.3 Earthquake Territory Definitions', self.buildEQTerritoryDefinitions(), False, True)
        OptionalCoverages.generateWorksheet('EMVL', 'OC Table C.4.E.4.c. Earthquake and Volcanic Eruption - Masonry Veneer Limitation', self.buildEQMasonryVeneer(), False, True)
        OptionalCoverages.generateWorksheet('EC', 'OC Table C.4.E.6.a Earthquake and Volcanic Eruption - Coinsurance', self.buildEQCoinsurance(), False, True)
        OptionalCoverages.generateWorksheet('ESR', 'OC Table C.4.E.7 Earthquake and Volcanic Eruption - Sprinklered Risk', self.buildEQSprinkleredRisk(), False, True)
        OptionalCoverages.generateWorksheet('EBH', 'OC Table C.4.E.8 Earthquake and Volcanic Eruption - Building Height', self.buildEQBuildingHeight(), False, True)

        #EQ Territory Class Rated pages - format differs per state

        if self.state == "ID" or self.state == "GA" or self.state == "NC" or self.state == "VT" or self.state == "WY": # 2 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet2tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedID1(), self.buildEQClassRatedID2(), False, True)
        elif self.state == "NH" or self.state == "CO" or self.state == "NV" or self.state == "AZ" or self.state == "ME": # 3 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet3tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedNH1(), self.buildEQClassRatedNH2(), self.buildEQClassRatedNH3(), False, True)
        elif self.state == "MS" or self.state == "SC" or self.state == "OH" or self.state == "IN" or self.state == "NM" or self.state == "TN": # 4 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet4tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedMS1(), self.buildEQClassRatedMS2(), self.buildEQClassRatedMS3(), self.buildEQClassRatedMS4(), False, True)
        elif self.state == "KY" or self.state == "OR" or self.state == "WA": # 5 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet5tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedKY1(), self.buildEQClassRatedKY2(), self.buildEQClassRatedKY3(), self.buildEQClassRatedKY4(), self.buildEQClassRatedKY5(), False, True)
        elif self.state == "IL" or self.state == "MT" or self.state == "NY": # 6 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet6tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedIL1(), self.buildEQClassRatedIL2(), self.buildEQClassRatedIL3(), self.buildEQClassRatedIL4(), self.buildEQClassRatedIL5(), self.buildEQClassRatedIL6(), False, True)
        elif self.state == "AR" or self.state == "MO": # 7 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet7tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedAR1(), self.buildEQClassRatedAR2(), self.buildEQClassRatedAR3(), self.buildEQClassRatedAR4(), self.buildEQClassRatedAR5(), self.buildEQClassRatedAR6(), self.buildEQClassRatedAR7(), False, True)
        elif self.state == "UT": # 8 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet8tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedUT1(), self.buildEQClassRatedUT2(), self.buildEQClassRatedUT3(), self.buildEQClassRatedUT4(), self.buildEQClassRatedUT5(), self.buildEQClassRatedUT6(), self.buildEQClassRatedUT7(), self.buildEQClassRatedUT8(), False, True)
        elif self.state == "CA": # 17 Territories displayed on rate pages
            OptionalCoverages.generateWorksheet17tables('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRatedCA1(), self.buildEQClassRatedCA2(), self.buildEQClassRatedCA3(), self.buildEQClassRatedCA4(), self.buildEQClassRatedCA5(), self.buildEQClassRatedCA6(), self.buildEQClassRatedCA7(), self.buildEQClassRatedCA8(), self.buildEQClassRatedCA9(), self.buildEQClassRatedCA10(), self.buildEQClassRatedCA11(), self.buildEQClassRatedCA12(), self.buildEQClassRatedCA13(), self.buildEQClassRatedCA14(), self.buildEQClassRatedCA15(), self.buildEQClassRatedCA16(), self.buildEQClassRatedCA17(), False, True)
        else: # 1 Territory displayed on rate pages
            OptionalCoverages.generateWorksheet('ECR', 'OC Table C.4.F.1.a. Earthquake and Volcanic Eruption - Class Rated', self.buildEQClassRated(), False, True)
        #End of EQ Territory Class Rated pages
        
        OptionalCoverages.generateWorksheet('ELCM', 'OC Table C.4.F.1.b. Earthquake and Volcanic Eruption - Loss Cost Multiplier', self.buildEQLCM(), False, True)
        OptionalCoverages.generateWorksheet('ESLC', 'OC Table C.4.H.2.c. Earthquake and Volcanic Eruption - Earthquake Sprinkler Leakage Only - Coinsurance', self.buildEQSprinklerLeakCoinsurance(), False, True)
        OptionalCoverages.generateWorksheet('ESLEE', 'OC Table C.4.H.3.d. Earthquake and Volcanic Eruption - Earthquake Sprinkler Leakage only - Earthquake Extension', self.buildEQSprinklerLeakEQExt(), False, True)
        OptionalCoverages.generateWorksheet('ED', 'OC Table C.5.D.2. Employee Dishonesty', self.buildEmployeeDishonesty(), False, True)
        OptionalCoverages.generateWorksheet('EDECE', 'OC Table C.5.E. Employee Dishonesty - ERISA Compliance Endorsement', self.buildEmployeeDishonestyERISA(), False, True)
        OptionalCoverages.generateWorksheet('BIIPI', 'OC Table C.7.C. Extended Business Income - Increased Period of Indemnity', self.buildExtendedBusinessIncome(), False, True)
        OptionalCoverages.generateWorksheet('BIOP', 'OC Table C.8.C. Business Income - Ordinary Payroll Increased Coverage Period', self.buildBusinessIncomeOrdinary(), False, True)
        OptionalCoverages.generateWorksheet('BIALS', 'OC Table C.8-A.C. Business Income - Actual Loss Sustained - Alternative Months', self.buildBusinessIncomeActual(), False, True)
        OptionalCoverages.generateWorksheet('BIWP', 'OC Table C.8-B.C. Business Income - Waiting Period', self.buildBusinessIncomeWaiting(), False, True)
        OptionalCoverages.generateWorksheet('OLCLU', 'OC Table C.12.E.1.a.(1). Ordinance or Law Coverage for Loss to the Undamaged Portion of the Building', self.buildOrdinanceLoss(), False, True)
        OptionalCoverages.generateWorksheet('OLB', 'OC Table C.12-A.D. Ordinance or Law Broadened Endorsement', self.buildOrdinanceEndorsement(), False, True)
        OptionalCoverages.generateWorksheet('OSBI', 'OC Table C.13.D.1. Outdoor Signs - Business Income', self.buildOutdoorSignsBusInc(), False, True)
        OptionalCoverages.generateWorksheet('SFPO', 'OC Table C.14.E.1. Spoilage From Power Outage', self.buildSpoilage(), False, True)
        OptionalCoverages.generateWorksheet3tables('USIBI', 'OC Table C.15.D.1. Utility Services Additional Coverage - Including Business Income', self.buildUtilityServices1(), self.buildUtilityServices2(), self.buildUtilityServices3(), False, True)
        OptionalCoverages.generateWorksheet('SPF', 'OC Table C.16.D.2. Scheduled Property Floater', self.buildScheduledPropFloater(), False, True)
        OptionalCoverages.generateWorksheet('VDBP', 'OC Table C.17.E.1. Vehicle Damage to Leased Property Base Premium', self.buildVehicleDamBase(), False, True)
        OptionalCoverages.generateWorksheet('VDALP', 'OC Table C.17.E.2. Vehicle Damage to Leased Property Additional Limit Premium', self.buildVehicleDamAdditional(), False, True)
        OptionalCoverages.generateWorksheet('IREFD', 'OC Table C.18.C.1. Increase in Rebuilding Expenses Following Disaster (Additional Expense Coverage on Annual Aggregate Basis)', self.buildRebuildingExpense(), False, True)
        OptionalCoverages.generateWorksheet('LRVDP', 'OC Table C.19.C. Loss of Rental Value - Landlord as Designated Payee', self.buildRentalLoss(), False, True)
        OptionalCoverages.generateWorksheet('ECRCS', 'OC Table C.20.A.3. Exclusion - Cosmetic Loss to Roof Coverings and Siding', self.buildExclusionRoofSiding(), False, True)
        OptionalCoverages.generateWorksheet('ECRC', 'OC Table C.20.B.3. Exclusion - Cosmetic Loss to Roof Covering', self.buildExclusionRoof(), False, True)
        OptionalCoverages.generateWorksheet('WHRSA', 'OC Table C.21.A.3. Windstorm or Hail Losses to Roof Surfacing - Actual Cash Value Loss Settlement', self.buildWindhailACVSettlement(), False, True)
        OptionalCoverages.generateWorksheet('ACVRS', 'OC Table C.21.B.3. Actual Cash Value for Roof Surfacing', self.buildACVRoof(), False, True)
        OptionalCoverages.generateWorksheet('FP', 'OC Table C.22.C. False Pretense', self.buildFalsePretense(), False, True)
        OptionalCoverages.generateWorksheet('FBR', 'OC Table C.23.G.1. Flood Base Rates', self.buildFloodBaseRates(), False, True)
        OptionalCoverages.generateWorksheet('FSA', 'OC Table C.23.G.2.c. Flood Sub-Limit Adjustment', self.buildFloodSubLimit(), False, True)
        OptionalCoverages.generateWorksheet('FBCF', 'OC Table C.23.G.3.a. Flood Building Construction Factor', self.buildFloodBLDGConstruction(), False, True)
        OptionalCoverages.generateWorksheet('FPPCV', 'OC Table C.23.G.3.b. Flood Business Personal Property of Concentration of Values', self.buildFloodBPP(), False, True)
        OptionalCoverages.generateWorksheet('FDF', 'OC Table C.23.G.4. Flood Deductible Factors', self.buildFloodDedFactors(), False, True)
        OptionalCoverages.generateWorksheet('FMAR', 'OC Table C.23.G.5. Flood Minimum Adjusted Rate', self.buildFloodMinAdjRate(), False, True)
        OptionalCoverages.generateWorksheet('FALM', 'OC Table C.23.G.6. Flood Aggregate Limit Multipliers', self.buildFloodAggLimit(), False, True)
        OptionalCoverages.generateWorksheet('FMP', 'OC Table C.23.G.7. Flood Minimum Premium', self.buildFloodMinPrem(), False, True)
        OptionalCoverages.generateWorksheet('BIEELI', 'OC Table C.25.A.4. Business Income Extra Expense', self.buildBusinessIncomeExtraExpense(),False, True)
        OptionalCoverages.generateWorksheet('AIPN', 'OC Table D.3.C.3. Amendment to the Other Insurance Clause for Additional Insureds - Primary and Non-Contributory', self.buildAmendmentAdditional(), False, True)
        OptionalCoverages.generateWorksheet('AIFCE', 'OC Table D.3.D.3.a-b. Additional Insured – Fairs, Carnivals or Expositions', self.buildAdditionalFairs(), False, True)
        OptionalCoverages.generateWorksheet('AISPP', 'OC Table D.3.E.3.a-b. Additional Insured – Services Performed On Premises of Additional Insured', self.buildAdditionalServices(), False, True)
        OptionalCoverages.generateWorksheet('AIV', 'OC Table D.3.F.3. Additional Insured – Vendors', self.buildAdditionalVendors(), False, True)
        OptionalCoverages.generateWorksheet('AIDPO', 'OC Table D.3.I.3. Additional Insured – Designated Person or Organization', self.buildAdditionalDesignated(), False, True)
        OptionalCoverages.generateWorksheet('AIOLCS', 'OC Table D.3.J.3. Additional Insured – Owners, Lessees or Contractors – Scheduled Person or Organization', self.buildAddInsuredOwnerLeaseContractor(), False, True)
        OptionalCoverages.generateWorksheet2tables('EBIAE', 'OC Table D.5.C. Employee Bodily Injury to Another Employee', self.buildEmployeeBodilyInj(), self.buildEmployeeBodilyInjMinPrem(), False, True)
        OptionalCoverages.generateWorksheet('EBL', 'OC Table D.6.D.1. Employee Benefits Liability', self.buildEmployeeBenefitsLiab(), False, True)
        OptionalCoverages.generateWorksheet('EBERP', 'OC Table D.6.E.3. Employee Benefits Liability - Extended Reporting Period Option  % of Annual Premium', self.buildEmployeeBenefitsLiabExt(), False, True)
        OptionalCoverages.generateWorksheet('GKBR', 'OC Table D.8.C.1. Garage Keepers Coverage - Base Rate', self.buildGarageKeepersBase(), False, True)
        OptionalCoverages.generateWorksheet('GKTM', 'OC Table D.8.C.2.a. Garage Keepers Coverage - Territory Multipliers', self.buildGarageKeepersTerritoryMult(), False, True)
        OptionalCoverages.generateWorksheet('GKDF', 'OC Table D.8.C.2.b. Garage Keepers Coverage - Deductible Factors', self.buildGarageKeepersDed(), False, True)
        OptionalCoverages.generateWorksheet('GKLIF', 'OC Table D.8.C.2.c. Garage Keepers Coverage - Limit of Insurance Factors', self.buildGarageKeepersDedLOI(), False, True)
        OptionalCoverages.generateWorksheet('GKBIF', 'OC Table D.8.C.2.d.(2). Garage Keepers Coverage - Blanket Insurance Factor', self.buildGarageKeepersBlanket(), False, True)
        OptionalCoverages.generateWorksheet('GKRM', 'OC Table D.8.C.2.e. Garage Keepers Coverage - Rate Modifiers', self.buildGarageKeepersRateMod(), False, True)
        OptionalCoverages.generateWorksheet('HNAHA', 'OC Table D.9.E.1. Hired and Non-Owned Auto Liability - Hired Auto', self.buildHiredAuto(), False, True)
        OptionalCoverages.generateWorksheet('HNANA', 'OC Table D.9.E.2. Hired and Non-Owned Auto Liability - Non-Owned Auto', self.buildNonOwnedAuto(), False, True)
        OptionalCoverages.generateLiquorLiability('LLFSR', 'OC Table D.10.A.4.a. Liquor Liability Coverage - Food Service Program Risks Only', 'State Base Rates', self.buildLiquorLiabFood1(), 'Minimum Premium', self.buildLiquorLiabFood2(), 'Liquor Liability Limit Factor', self.buildLiquorLiabFood3(), False, True)
        OptionalCoverages.generateWorksheet('LLAOP', 'OC Table D.10.A.4.b. Liquor Liability Coverage - All Other Programs', self.buildLiquorLiabAllOther(), False, True)
        OptionalCoverages.generateWorksheet('LLE', 'OC Table D.10.B.3.b. Liquor Liability Coverage - Amendment – Liquor Liability Exclusion', self.buildLiquorLiabAmendment(), False, True)
        OptionalCoverages.generateWorksheet('TPDR', 'OC Table D.13.D.1. Tenants Property Damage Legal Liability Rate', self.buildTenantsPropDam(), False, True)
        OptionalCoverages.generateWorksheet('ERPBO', 'OC Table D.14.D.1.b. Employee Related Practices Liability - Base Rate Outside (Only Applicable for AR, MO, MT, NM, and VT)', self.buildEmployeePracticesOutside(), False, True)
        OptionalCoverages.generateWorksheet('ERPBI', 'OC Table D.14.D.2.b. Employee Related Practices Liability - Base Rate Inside (Not Applicable for AR, MO, MT, NM, and VT)', self.buildEmployeePracticesInside(), False, True)
        OptionalCoverages.generateWorksheet('ERPSF', 'OC Table D.14.E. Employee Related Practices Liability - State Factor', self.buildEmployeeRelatedState(), False, True)
        OptionalCoverages.generateWorksheet('ERPNF', 'OC Table D.14.F. Employee Related Practices Liability - NAICS Factor', self.buildEmployeeRelatedNAICS(), False, True)
        OptionalCoverages.generateWorksheet('ERPSE', 'OC Table D.14.G.2. Employee Related Practices Liability - Supplemental ERP Premium', self.buildEmployeeRelatedSuppERP(), False, True)
        OptionalCoverages.generateWorksheet('CWLR', 'OC Table D.15.C. Car Wash Damage to Customers Autos - Liability Rate', self.buildCarWashLiab(), False, True)
        OptionalCoverages.generateWorksheet('CWDF', 'OC Table D.15.D. Car Wash Damage to Customers Autos - Deductible Factor', self.buildCarWashDed(), False, True)
        OptionalCoverages.generateWorksheet('EPLWHC', 'OC Table D.21.A.4 Wage and Hour Claims Expenses - Employment Practices Liability', self.buildEmploymentPracticesLiability(), False, True)
        OptionalCoverages.generateWorksheet6tables('EPLNIC', 'OC Table D.22.A.4 Employment Practices Liability Coverage For Injury To Named Independent Contractors', self.buildEmploymentTypeFactor(), self.buildDefenseInside(), self.buildDefesneWithin(), self.buildMatchingDefense(), self.buildEmploymentPractFactor(), self.buildEmploymentPracticesLiabilityState(), False, True)
        OptionalCoverages.generateWorksheet('EPLTPP', 'OC Table D.23.A.4 Employment Practices Liability Coverage for Third Party Practices', self.buildEmploymentPracticesLiabilityThirdParty(), False, True)
        OptionalCoverages.generateWorksheet('UAVL', 'OC Table D.25.B.4. Limited Coverage for Designated Unmanned Aircraft', self.buildLimitedCoverageUnmannedAircraft(), False, True)
        OptionalCoverages.generateWorksheet('BAR', 'OC Table E.1.D.1. Businessowners ADVANTAGE Rate', self.buildAdvantageRate(), False, True)
        OptionalCoverages.generateWorksheet('BAILF', 'OC Table E.1.D.2. Businessowners ADVANTAGE - Increased Limit Factor', self.buildAdvantageILF(), False, True)
        OptionalCoverages.generateWorksheet('CSC', 'OC Table E.3.A.5. Cyber Suite Coverage', self.buildCyberSuite(), False, True)
        OptionalCoverages.generateWorksheet3tables('CSCTP', 'OC Table E.3.B.5. Cyber Suite Coverage Third Party Only Premiums', self.buildCyberSuiteThird(), self.buildCyberSuiteSubLimit1(), self.buildCyberSuiteSubLimit2(), False, True)

        OptionalCoverages.createIndex()
        OptionalCoveragesPages = OptionalCoverages.getWB()

        self.formatAccountsReceivableILF(OptionalCoveragesPages['ARIL'])
        #self.formatForgeryILF(OptionalCoveragesPages['FAIL'])
        self.formatMoneyILF(OptionalCoveragesPages['MSIL'], Font(name=fontName, size=fontSize, bold=True)) 
        self.formatOutdoorSignsILF(OptionalCoveragesPages['OSIL'])
        self.formatOutdoorTreesILF(OptionalCoveragesPages['OTIL'])  
        #self.formatValuablePapersILF(OptionalCoveragesPages['VPRIL'])  
        self.formatBackupSewerILF(OptionalCoveragesPages['BSWIL'], Font(name=fontName, size=fontSize, bold=True))
        self.formatAwayFromPremisesILF(OptionalCoveragesPages['PPAIL'])
        self.formatElectronicDataILF(OptionalCoveragesPages['EDIL'])  
        self.formatInterruptionILF(OptionalCoveragesPages['ICOIL'])
        self.formatBLDGPropertyILF(OptionalCoveragesPages['BPOIL'])
        self.formatComputerFraudILF(OptionalCoveragesPages['FTFIL'])
        #self.formatBPPStorage(OptionalCoveragesPages['PPSIL'])
        self.formatCondoLossAsses(OptionalCoveragesPages['CCULA'])
        self.formatCondo(OptionalCoveragesPages['CCUOC'])
        self.formatEQPropertyDed(OptionalCoveragesPages['EPDF'])
        self.formatEQSprinkler(OptionalCoveragesPages['ESLPD'])
        #self.formatEQotherthanfunctional(OptionalCoveragesPages['EXFBV'])
        #self.formatEQFunctional(OptionalCoveragesPages['EFBV'], Font(name=fontName, size=fontSize, bold=True))
        #self.formatEQUndamagedLoss(OptionalCoveragesPages['ELU'], Font(name=fontName, size=fontSize, bold=True))
        #self.formatEQSprinklerLeakage(OptionalCoveragesPages['ESL'])
        self.formatEQDeductibleOptions(OptionalCoveragesPages['EDO'])
        self.formatEQTerritoryDefs(OptionalCoveragesPages['ETD'])
        self.formatEQMasonryVeneer(OptionalCoveragesPages['EMVL'])
        self.formatEQCoinsurance(OptionalCoveragesPages['EC'])
        self.formatEQSprinkleredRisk(OptionalCoveragesPages['ESR'])
        self.formatEQBuildingHeight(OptionalCoveragesPages['EBH'], Font(name=fontName, size=fontSize, bold=True))
        if self.state == "ID" or self.state == "GA" or self.state == "NC" or self.state == "VT" or self.state == "WY": #2
            self.formatEQClassRatedID(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "NH" or self.state == "CO" or self.state == "NV" or self.state == "AZ" or self.state == "ME": #3
            self.formatEQClassRatedNH(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "MS" or self.state == "SC" or self.state == "OH" or self.state == "IN" or self.state == "NM" or self.state == "TN": #4
            self.formatEQClassRatedMS(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "KY" or self.state == "OR" or self.state == "WA": #5
            self.formatEQClassRatedKY(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "IL" or self.state == "MT" or self.state == "NY": #6
            self.formatEQClassRatedIL(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "AR" or self.state == "MO": #7
            self.formatEQClassRatedAR(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "UT": #8
            self.formatEQClassRatedUT(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        elif self.state == "CA":
            self.formatEQClassRatedCA(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        else: #1
            self.formatEQClassRated(OptionalCoveragesPages['ECR'], Font(name=fontName, size=fontSize, bold=True))
        self.formatEQLCM(OptionalCoveragesPages['ELCM'])
        self.formatEQSprinklerLeakCoinsurance(OptionalCoveragesPages['ESLC'])
        #self.formatEQSprinklerLeakEQExt(OptionalCoveragesPages['ESLEE'])
        self.formatEmployeeDishonesty(OptionalCoveragesPages['ED'], Font(name=fontName, size=fontSize, bold=True))
        #self.formatEmployeeDishonestyERISA(OptionalCoveragesPages['EDECE'])
        self.formatExtendedBusinessIncome(OptionalCoveragesPages['BIIPI'])
        self.formatExtendedBusinessIncome(OptionalCoveragesPages['BIOP'])
        self.formatExtendedBusinessIncome(OptionalCoveragesPages['BIALS'])
        self.formatExtendedBusinessIncome(OptionalCoveragesPages['BIWP'])
        #self.formatOrdinanceLoss(OptionalCoveragesPages['OLCLU'])
        self.formatOrdinanceEndorsement(OptionalCoveragesPages['OLB'])
        self.formatOrdinanceEndorsement(OptionalCoveragesPages['OSBI'])
        self.formatSpoilage(OptionalCoveragesPages['SFPO'])
        self.formatUtilityServices(OptionalCoveragesPages['USIBI'], Font(name=fontName, size=fontSize, bold=True))
        self.formatScheduledPropFloater(OptionalCoveragesPages['SPF'])
        self.formatVehicleDamBase(OptionalCoveragesPages['VDBP'])
        self.formatVehicleDamBase(OptionalCoveragesPages['VDALP'])
        #self.formatRebuildingExpense(OptionalCoveragesPages['IREFD']
        self.formatVehicleDamBase(OptionalCoveragesPages['LRVDP'])
        self.formatExclusionRoofSiding(OptionalCoveragesPages['ECRCS'])
        self.formatExclusionRoofSiding(OptionalCoveragesPages['ECRC'])
        self.formatExclusionRoofSiding(OptionalCoveragesPages['WHRSA'])
        self.formatExclusionRoofSiding(OptionalCoveragesPages['ACVRS'])
        self.formatVehicleDamBase(OptionalCoveragesPages['FP'])
        #self.formatFloodBaseRates(OptionalCoveragesPages['FBR'])
        self.formatFloodSubLimit(OptionalCoveragesPages['FSA'])
        self.formatFloodBLDGConstruction(OptionalCoveragesPages['FBCF'])
        self.formatFloodBPP(OptionalCoveragesPages['FPPCV'])
        self.formatFloodDedFactors(OptionalCoveragesPages['FDF'])
        self.formatFloodMinAdjRate(OptionalCoveragesPages['FMAR'])
        self.formatFloodAggLimit(OptionalCoveragesPages['FALM'])
        self.formatFloodMinPrem(OptionalCoveragesPages['FMP'])
        self.formatBusinessIncomeExtraExpenses(OptionalCoveragesPages["BIEELI"])
        self.formatAmendmentAdditional(OptionalCoveragesPages['AIPN'])
        self.formatAdditionalFairs(OptionalCoveragesPages['AIFCE'])
        self.formatAdditionalServices(OptionalCoveragesPages['AISPP'])
        self.formatAdditionalVendors(OptionalCoveragesPages['AIV'])
        self.formatAdditionalDesignated(OptionalCoveragesPages['AIDPO'])
        self.formatEmployeeBodilyInj(OptionalCoveragesPages['EBIAE'], Font(name=fontName, size=fontSize, bold=True))
        self.formatEmployeeBenefitsLiab(OptionalCoveragesPages['EBL'])
        #self.formatEmployeeBenefitsLiabExt(OptionalCoveragesPages['EBERP'])
        self.formatGarageKeepersBase(OptionalCoveragesPages['GKBR'])
        self.formatGarageKeepersTerritoryMult(OptionalCoveragesPages['GKTM'])
        self.formatGarageKeepersDed(OptionalCoveragesPages['GKDF'])
        self.formatGarageKeepersLOI(OptionalCoveragesPages['GKLIF'])
        #self.formatGarageKeepersBlanket(OptionalCoveragesPages['GKBIF'])
        self.formatGarageKeepersRateMod(OptionalCoveragesPages['GKRM'])
        self.formatHiredAuto(OptionalCoveragesPages['HNAHA'])
        self.formatNonOwnedAuto(OptionalCoveragesPages['HNANA'])
        self.formatLiquorLiabFood(OptionalCoveragesPages['LLFSR'], Font(name=fontName, size=fontSize, bold=True))
        self.formatLiquorLiabAllOther(OptionalCoveragesPages['LLAOP'])
        self.formatLiquorLiabAmendment(OptionalCoveragesPages['LLE'])
        #self.formatTenantsPropDam(OptionalCoveragesPages['TPDR'])
        self.formatEmployeePracticesOutside(OptionalCoveragesPages['ERPBO'])
        self.formatEmployeePracticesInside(OptionalCoveragesPages['ERPBI'])
        self.formatEmploymentPracticesLiability(OptionalCoveragesPages['EPLWHC'])
        self.formatEmploymentPracticesLiabilityNamedIndCont(OptionalCoveragesPages['EPLNIC'])
        self.formatEmploymentPracticesLiabilityThirdParty(OptionalCoveragesPages['EPLTPP'])
        #self.formatEmployeeRelatedState(OptionalCoveragesPages['ERPSF'])
        #self.formatEmployeeRelatedNAICS(OptionalCoveragesPages['ERPNF'])
        self.formatEmployeeRelatedSuppERP(OptionalCoveragesPages['ERPSE'])
        self.formatCarWashLiab(OptionalCoveragesPages['CWLR'])
        self.formatCarWashDed(OptionalCoveragesPages['CWDF'])
        self.formatLimitedCoverageUnmannedAircraft(OptionalCoveragesPages["UAVL"])
        self.formatAdvantageRate(OptionalCoveragesPages['BAR'])
        self.formatAdvantageILF(OptionalCoveragesPages['BAILF'])
        self.formatCyberSuite(OptionalCoveragesPages['CSC'])
        self.formatCyberSuiteThird(OptionalCoveragesPages['CSCTP'], Font(name=fontName, size=fontSize, bold=True))

        return OptionalCoveragesPages