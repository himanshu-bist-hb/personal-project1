# This module builds and formats the Retail State Page workbook (BP-2.0).
#
# The build*() methods below are unchanged business logic, transcribed from
# the root-level RetailPage.py: each pulls a table out of the ratebook data
# (via buildDataFrame's nesting waterfall) and shapes it into the DataFrame
# the rate page needs, filtered to the Retail program (Class_Code_Min ==
# 50000). Same porting pattern as [[bop_hab_autoservice_port]]: the
# [[bop-nesting-order]] fix (lower-level company -> NGIC -> CW, not
# NGIC-first) and the per-company base-rate fix (root hardcoded
# buildBaseRates('NGIC') for every company's BRNACO/BRNAFF/BRNICOF tab — same
# bug already found and fixed in Hab/Auto Service).
#
# All Excel formatting (fonts, column widths, sub-header labels, page setup)
# lives in ExcelSettingsBOP.py, driven by "BOP/BOP Input File.xlsx". Most of
# Retail's tables reuse table codes/layout profiles already shared with Hab
# and Auto Service (CBG/CPP/YBBG/YBPP/EBB/CW/PDLD/LL/DO/DONM/ERP/LPGE/FR/
# AS_BR/PROGRAM_TR) since the column widths and number formats transcribed
# from the root file's format*() methods matched those exactly. Retail-only
# tables (DC/DCEQ/ET/GLO/FL/HE/OPTI/PED/RTS) and Retail's own Liability Size
# of Risk shape (LS_RETAIL — different sub-header text/widths than Auto
# Service's LS) get new layout keys. The Dry Cleaners EXTRA Factor (DC) row
# grouping, the Directors & Officers 2-band merge (Retail only has "Under
# 51"/"51 or More", unlike Hab's 5 bands), and the Pet Services Specialized
# Endorsement (PSS) two-table sheet are handled with direct worksheet
# post-processing below, same pattern as Hab's DO/HABEX and Auto Service's
# OILN/OILE.

from copy import copy

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import ExcelSettingsBOP

_THIN_BORDER = Border(
    left=Side(border_style='thin', color='C1C1C1'),
    right=Side(border_style='thin', color='C1C1C1'),
    top=Side(border_style='thin', color='C1C1C1'),
    bottom=Side(border_style='thin', color='C1C1C1'),
)


class Retail:
    # Tables buildBaseRates() needs present for a given company before it can
    # be built without a KeyError.
    _BASE_RATE_TABLES = (
        'BP7_Peril_Building_Base_Rates',
        'BP7_Peril_BPP_Base_Rates',
        'BP7_Peril_Liability_Base_Rates',
    )

    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        # Individual program pages (unlike All Programs) also show the
        # "AllPeril" row/column — see [[bop_auto_service_allperil_row_fix]].
        self.perils = list(perils) + ['allperil']
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective
        self.rEffective = rEffective

        self.retailProgramCode = 50000

    # Builds a dataframe for the given table code
    # The hierarchy matches Business Auto: lower-level company (NACO/NAFF/NICOF)
    # first, then NGIC (state-level default), then CW as the country-wide
    # fallback. See [[bop-nesting-order]] — the root-level RetailPage.py
    # checked NGIC first, which is backwards.
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if 'NACO' in self.rateTables.keys():
            if tableCode in self.rateTables['NACO'].keys():
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys():
            if tableCode in self.rateTables['NAFF'].keys():
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys():
            if tableCode in self.rateTables['NICOF'].keys():
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if tableCode in self.rateTables['NGIC'].keys():
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0])

    # Builds the dry cleaners EXTRA factor table
    # Returns a dataframe
    def buildDryCleanersFactor(self):
        dryCleanersFactor = self.buildDataFrame("BP7_Dry_Cleaners_Extra_Factor")
        filteredDryCleanersFactor = dryCleanersFactor.query(f'Class_Code_Min == {self.retailProgramCode}')
        data = {'TierLimit': ['BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits',
                    'Tier2', 'Tier2', 'Tier2', 'Tier2', 'Tier2',
                    'Tier3', 'Tier3', 'Tier3', 'Tier3', 'Tier3',
                    'Tier4', 'Tier4', 'Tier4', 'Tier4', 'Tier4'],
                'Coverage': ['Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets'],
                'Limits (Per Item / Per Occurrence)': ['$1,000/ALS', '$1,000/$10,000', '$1,000/$5,000', '$1,000/$5,000', '$1,000/$5,000',
                    '$2,000/ALS', '$2,000/$15,000', '$2,000/$6,000', '$2,000/$6,000', '$2,000/$6,000',
                    '$3,000/ALS', '$3,000/$20,000', '$3,000/$9,000', '$3,000/$9,000', '$3,000/$9,000',
                    '$5,000/ALS', '$5,000/$25,000', '$5,000/$10,000', '$5,000/$10,000', '$5,000/$10,000']}
        dryCleaners = pd.DataFrame(data)
        finalDryCleaners = pd.merge(dryCleaners, filteredDryCleanersFactor, how='left', on=['TierLimit'])
        return finalDryCleaners.rename(columns={'TierLimit': 'Tier', 'DryCleanerExtraFactor': 'Factor'}).filter(items=['Tier', 'Coverage', 'Limits (Per Item / Per Occurrence)', 'Factor']). \
                replace({'Tier': {'BaseLimits': 'Base', 'Tier2': 'Tier 2', 'Tier3': 'Tier 3', 'Tier4': 'Tier 4'}})

    # Builds the dry cleaners EXTRA earthquake factor table
    # Returns a dataframe
    def buildDryCleanersEQFactor(self):
        miscFactors = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        dryCleanersEQFactor = miscFactors.query(f'FactorName == "DryCleanersExtraEQ"')
        return dryCleanersEQFactor.filter(items=['Factor']).rename(columns={'Factor': 'Dry Cleaners EXTRA EQ Factor'})

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.retailProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.retailProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.retailProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building",
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.retailProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the exclude theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.retailProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'] == 'L-Products'].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.retailProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'})
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.retailProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the car wash factor table
    # Returns a dataframe
    def buildCarWashFactor(self):
        carWashFactor = self.buildDataFrame("BP7_Peril_Car_Wash_Factor")
        return carWashFactor.dropna().query(f'`Class Code Min` == {self.retailProgramCode} & `Peril TypeCode` == "liability1"').rename(columns={'No Of Bays': 'Number of Bays', 'CarWashFactor': 'Factor'}). \
                replace({'Number of Bays': {1: 'One', 2: 'Two', 3: 'Three', 4: '4 or more'}}).filter(items=['Number of Bays', 'Factor'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.retailProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.retailProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance': 'int32'})

    # Builds the liquified petroleum gas (LPG) exposures table
    # Returns a dataframe
    def buildLPGExposure(self):
        lpgExposure = self.buildDataFrame("BP7_LPG_Premium")
        return lpgExposure.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'LPGPremium': 'Premium (each premises)'})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_BLDG_BPP_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCodeMin == {self.retailProgramCode}').rename(columns={'BldgBPPLimit_Min': 'Limit Min', 'BldgBPPLimit_Max': 'Limit Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Limit Min', 'Limit Max'], columns='Peril TypeCode', values='Factor').reset_index(['Limit Min', 'Limit Max']).fillna({'Limit Max': 'and over'})

    # Builds the general liability occupancy modifiers table
    # Returns a dataframe
    def buildGeneralOccupancyMod(self):
        generalOccupancyModifier = self.buildDataFrame("BP7_Peril_Occupant_Factor")
        filteredGeneralOccupancyMod = generalOccupancyModifier.query(f'ClassCode_Min == {self.retailProgramCode} & `Peril TypeCode` == "liability1"'). \
                rename(columns={'OccupancyType': 'Occupancy', 'BLDGOccupantFactor': 'Building', 'BPPOccupantFactor': 'Business Personal Property'})
        return filteredGeneralOccupancyMod.replace({'Occupancy': {'Condominium': 'Condo Unit-owner', 'buildingOwnerLessorsrisk': "Lessor's Risk", 'buildingOwnerOccupant': 'Owner Occupant', 'tenant': 'Tenant'}}). \
                filter(items=['Occupancy', 'Building', 'Business Personal Property'])

    # Builds the directors and officers liability insurance table
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Retail"').copy()
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 or More'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])

    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Retail"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Retail"').copy()
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index=filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])

    # Builds the florists misdelivery table
    # Returns a dataframe
    def buildFloristsMisdelivery(self):
        floristsMisdelivery = self.buildDataFrame("BP7_Florists_Misdelivery")
        return floristsMisdelivery.astype({'LiabilityLimit': 'int32'}).rename(columns={'LiabilityLimit': 'Limit', 'RatePerPremises': 'Each Premises'})

    # Builds the hearing aid stores professional liability table
    # Returns a dataframe
    def buildHearingAidStoresLiab(self):
        hearingAidStoresLiab = self.buildDataFrame("BP7_ProfLiabHearingAidEstablishments")
        hearingAidStoresLiab['Limit'] = hearingAidStoresLiab['Limit'].apply(lambda x: "${0:,.0f}".format(x))
        hearingAidStoresLiab['AggregateLimit'] = hearingAidStoresLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        hearingAidStoresLiab['Occurence / Aggregate'] = hearingAidStoresLiab['Limit'] + ' / ' + hearingAidStoresLiab['AggregateLimit']
        return hearingAidStoresLiab.filter(items=['Occurence / Aggregate', 'Hearing Aid Prof Liab Charge']).rename(columns={'Hearing Aid Prof Liab Charge': 'Each Hearing Aid Specialist'})

    # Builds the Optical Goods Stores Professional Liability table
    # Returns a dataframe
    def buildOpticalGoodsStoresLiab(self):
        opticalGoodsStoresLiab = self.buildDataFrame("BP7_ProfLiabOptical_Rate")
        opticalGoodsStoresLiab['LiabilityAmount'] = opticalGoodsStoresLiab['LiabilityAmount'].apply(lambda x: "${0:,.0f}".format(x))
        opticalGoodsStoresLiab['AggregateLimit'] = opticalGoodsStoresLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        opticalGoodsStoresLiab['Occurence / Aggregate'] = opticalGoodsStoresLiab['LiabilityAmount'] + ' / ' + opticalGoodsStoresLiab['AggregateLimit']
        return opticalGoodsStoresLiab.filter(items=['Occurence / Aggregate', 'OpticianRate']).rename(columns={'OpticianRate': 'Each Optician'})

    # Builds the pedorthists professional liability table
    # Returns a dataframe
    def buildPedorthistsLiab(self):
        pedorthistsLiab = self.buildDataFrame("BP7_ProfLiabPedorthistsBaseRate")
        pedorthistsLiab['Liability Limit Occurrence'] = pedorthistsLiab['Liability Limit Occurrence'].apply(lambda x: "${0:,.0f}".format(x))
        pedorthistsLiab['LiabilityLimitAggregate'] = pedorthistsLiab['LiabilityLimitAggregate'].apply(lambda x: "${0:,.0f}".format(x))
        pedorthistsLiab['Occurence / Aggregate'] = pedorthistsLiab['Liability Limit Occurrence'] + ' / ' + pedorthistsLiab['LiabilityLimitAggregate']
        return pedorthistsLiab.filter(items=['Occurence / Aggregate', 'BaseRate']).rename(columns={'BaseRate': 'Each Certified Pedorthist'})

    # Builds the table for Retail Trade Specialized Endorsement
    # Returns a dataframe
    def buildRTSplzdEndo(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.retailProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Retail Premises'})

    # Builds the franchise upgrade endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.retailProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(),
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(),
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Builds the table for Pet Services Specialized Endorsement (base premium)
    # Returns a dataframe
    def buildPSSplzdEndo(self):
        return pd.DataFrame({"Base premium for each Retail Premises": ["$212.00"]})

    # Builds the Pet Services - Business Income table (the second, appended
    # block of the PSS sheet — see _formatPSSplzdEndo)
    # Returns a dataframe
    def buildPSSBIncome(self):
        data = [
            ("$25,000", "$13", "$7"),
            ("$50,000", "$20", "$13"),
            ("$100,000", "$26", "$20"),
        ]
        return pd.DataFrame(data, columns=["Limits", "1st Worker", "Each Additional Worker"])

    # Builds the table for Pet Services Professional Liability
    # Returns a dataframe
    def buildPSProfLiab(self):
        return pd.DataFrame({
            "Limits": ["$300,000/$600,000", "$500,000/$1,000,000", "$1,000,000/$2,000,000", "$2,000,000/$4,000,000"],
            "Rate": ["$43", "$56", "$68", "$83"],
        })

    # Merges the "Number of Units" column of the D&O table into its 2 bands
    # ("Under 51" / "51 or More" — Retail only has 2, unlike Hab's 5).
    # Handled here directly (rather than through the generic Sub Headers
    # config) because it merges DATA rows, not a header row.
    def _formatDirsOfficersLiabIns(self, ws):
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Groups the Dry Cleaners EXTRA Factor table's Tier (col A) and Factor
    # (col D) columns into 5-row blocks, one per tier — the data is always
    # exactly 20 rows (4 tiers x 5 coverage types) in this fixed order, so a
    # static merge is safe.
    def _formatDryCleanersFactor(self, ws):
        for start in (4, 9, 14, 19):
            end = start + 4
            ws.merge_cells(f'A{start}:A{end}')
            ws.merge_cells(f'D{start}:D{end}')
            ws[f'A{start}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws[f'D{start}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Appends the Pet Services - Business Income table below the base
    # premium table on the same sheet, with its own bolded section label,
    # and widens the base premium table's header/data rows to span columns
    # A-C (matching the ["PSS", 1, 3, 120] width config and RSS's identical
    # merge in ServicePage.py). There's no direct BOP equivalent of the root
    # tool's generateWorksheet2tables (that module isn't present in this
    # repo), so this reconstructs the same two-block layout by hand: the
    # first table (buildPSSplzdEndo, written normally by generateWorksheet)
    # occupies rows 3-4; this appends a bolded "Pet Services - Business
    # Income" label plus the second table's header/data rows below it.
    def _formatPSSplzdEndo(self, ws, boldFont, biDf):
        ws.merge_cells('A3:C3')
        ws.merge_cells('A4:C4')
        start_row = ws.max_row + 2
        ws.cell(row=start_row, column=1, value="Pet Services - Business Income")
        for col in range(1, len(biDf.columns) + 1):
            ws.cell(row=start_row + 1, column=col, value=biDf.columns[col - 1])
        for r, (_, row) in enumerate(biDf.iterrows(), start=start_row + 2):
            for col, val in enumerate(row, start=1):
                ws.cell(row=r, column=col, value=val)
        for row in (start_row, start_row + 1):
            for cell in ws[row]:
                cell.font = boldFont
                cell.border = _THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for col in range(1, len(biDf.columns) + 1):
            char = get_column_letter(col)
            ws.column_dimensions[char].width = 120 / 7.0

    # Sets up the Retail Excel file and creates a separate worksheet for
    # each of the given dataframes. progress_callback (optional) is called
    # with a short message before each sheet is built.
    # Returns the Excel workbook
    def buildRetailPage(self, progress_callback=None):
        companies = [c for c in self.rateTables.keys() if c != 'CW']

        Retail = ExcelSettingsBOP.Excel(state=self.state, programName='Retail', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        sheetSpecs = [
            ('DC', 'R Table 1.A.5.c. Dry Cleaners EXTRA Factor', self.buildDryCleanersFactor, False, True, None, self._formatDryCleanersFactor),
            ('DCEQ', 'R Table 1.A.5.d. Dry Cleaners EXTRA Earthquake Factor', self.buildDryCleanersEQFactor, False, True, None, None),
        ]
        # A company can be present in rateTables (its ratebook was uploaded)
        # without having filed its own base-rate tables — a deviation
        # ratebook may only override a handful of tables. Check for the
        # specific tables buildBaseRates() needs, not just company
        # membership, or it KeyErrors on that company's missing table.
        for company, tab, label in (('NACO', 'BRNACO', 'NW Assurance'), ('NAFF', 'BRNAFF', 'NW Affinity'),
                                     ('NGIC', 'BRNGIC', 'NW General Insurance Company'), ('NICOF', 'BRNICOF', 'NICOF')):
            if company in self.rateTables and all(t in self.rateTables[company] for t in self._BASE_RATE_TABLES):
                sheetSpecs.append((tab, f'R Table 3.B.1. {label} State Base Rates', lambda c=company: self.buildBaseRates(c), False, True, 'AS_BR', None))

        sheetSpecs += [
            ('CBG', 'R Table 3.C.2.c. Construction Factor - Building', lambda: self.buildConstructionType('Building'), False, True, None, None),
            ('CPP', 'R Table 3.C.2.c. Construction Factor - BPP', lambda: self.buildConstructionType('BPP'), False, True, None, None),
            ('ET', 'R Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions, False, True, None, None),
            ('YBBG', 'R Table 3.C.2.p. Year Built Modifier - Building', lambda: self.buildYearBuiltModifier('Building'), False, True, None, None),
            ('YBPP', 'R Table 3.C.2.p. Year Built Modifier - BPP', lambda: self.buildYearBuiltModifier('BPP'), False, True, None, None),
            ('EBB', 'R Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate, False, True, None, None),
            ('CW', 'R Table 3.C.4.b. Car Wash Factor', self.buildCarWashFactor, False, True, None, None),
            ('PDLD', 'R Table 3.C.4.c. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount, False, True, None, None),
            ('LL', 'R Table 3.C.4.e. Liability Limit Factor', self.buildLiabilityLimitFactor, False, True, None, None),
            ('LPGE', 'R Table 3.C.4.f. Liquefied Petroleum Gas (LPG) Exposures', self.buildLPGExposure, False, True, 'LPGE_RETAIL', None),
            ('LS', 'R Table 3.C.4.g. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk, False, True, 'LS_RETAIL', None),
            ('GLO', 'R Table 3.D.1.c. General Liability Occupancy Modifiers', self.buildGeneralOccupancyMod, False, True, None, None),
            ('DO', 'R Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns, False, True, None, self._formatDirsOfficersLiabIns),
            ('DONM', 'R Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief, False, True, None, None),
            ('ERP', 'R Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods, False, True, None, None),
            ('FL', 'R Table 4.B.1.e. Florists Misdelivery', self.buildFloristsMisdelivery, False, True, None, None),
            ('HE', 'R Table 4.B.2.e.(1). Hearing Aid Stores Professional Liability', self.buildHearingAidStoresLiab, False, True, None, None),
            ('OPTI', 'R Table 4.B.3.e.(1). Optical Goods Stores Professional Liability', self.buildOpticalGoodsStoresLiab, False, True, None, None),
            ('PED', 'R Table 4.B.4.e.(1). Pedorthists Professional Liability', self.buildPedorthistsLiab, False, True, None, None),
            ('RTS', 'R Table 4.C. Retail Trade Specialized Endorsement', self.buildRTSplzdEndo, False, True, None, None),
            ('FR', 'R Table 4.D. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement, False, True, None, None),
            ('PSS', 'R Table 4.E. Pet Services Specialized Endorsement', self.buildPSSplzdEndo, False, True, None,
             lambda ws: self._formatPSSplzdEndo(ws, Retail.fontBold, self.buildPSSBIncome())),
            ('PSPL', 'R Table 4.F. Pet Services Professional Liability', self.buildPSProfLiab, False, True, 'PED', None),
        ]

        total = len(sheetSpecs)
        for i, (tableCode, title, build, useIndex, useHeader, layoutKey, postFormat) in enumerate(sheetSpecs, start=1):
            if progress_callback:
                progress_callback(f"Building sheet {i}/{total}: {tableCode}...")
            print(f"  [{i}/{total}] Building sheet: {tableCode}")
            ws = Retail.generateWorksheet(tableCode, title, build(), useIndex, useHeader, layout_key=layoutKey)
            if postFormat:
                postFormat(ws)

        if progress_callback:
            progress_callback("Building Index sheet...")
        print(f"  [{total}/{total}] Building sheet: Index")
        Retail.createIndex()
        return Retail.getWB()
