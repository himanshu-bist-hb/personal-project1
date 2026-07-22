# This module builds and formats the Auto Service State Page workbook (BP-2.0).
#
# Same idea as HabPage.py — see that file's module docstring. The two
# sheets with a two-row sub-header (Optional Increased Limits / Total
# Limits, each with its own second label row) are handled with a small
# amount of direct worksheet post-processing below, since that shape doesn't
# fit the generic single-row Sub Headers config used everywhere else.

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import get_column_letter

from . import ExcelSettingsBOP

_THIN_BORDER = Border(
    left=Side(border_style='thin', color='C1C1C1'),
    right=Side(border_style='thin', color='C1C1C1'),
    top=Side(border_style='thin', color='C1C1C1'),
    bottom=Side(border_style='thin', color='C1C1C1'),
)


class Auto:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective
        self.rEffective = rEffective

        self.autoProgramCode = 20000

    # Builds a dataframe for the given table code
    # The hierarchy matches Business Auto: lower-level company (NACO/NAFF/NICOF)
    # first, then NGIC (state-level default), then CW as the country-wide
    # fallback. See [[bop-nesting-order]] — the root-level AutoServicePage.py
    # checked NGIC first, which is backwards.
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if 'NACO' in self.rateTables.keys():
            if tableCode in self.rateTables['NACO'].keys():
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys():
            if tableCode in self.rateTables['NAFF'].keys():
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys():
            if tableCode in self.rateTables['NICOF'].keys():
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if tableCode in self.rateTables['NGIC'].keys():
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0])

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building",
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'] == 'L-Products'].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Full Theft': 'Full', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'})
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.autoProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the car wash factor table
    # Returns a dataframe
    def buildCarWashFactor(self):
        carWashFactor = self.buildDataFrame("BP7_Peril_Car_Wash_Factor")
        return carWashFactor.dropna().query(f'`Class Code Min` == {self.autoProgramCode} & `Peril TypeCode` == "liability1"').rename(columns={'No Of Bays': 'Number of Bays', 'CarWashFactor': 'Factor'}). \
                replace({'Number of Bays': {1: 'One', 2: 'Two', 3: 'Three', 4: '4 or more'}}).filter(items=['Number of Bays', 'Factor'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.autoProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_Receipts_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCode_Min == {self.autoProgramCode} & `Peril TypeCode` in {self.perils} & AutoServType == "N/A" & FoodServType == "N/A"').rename(columns={'ReceiptMin': 'Min', 'ReceiptMax': 'Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Min', 'Max'], columns='Peril TypeCode', values='Factor').reset_index(['Min', 'Max']).fillna({'Max': 'and over'})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.autoProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance': 'int32'})

    # Builds the liquified petroleum gas (LPG) exposures table
    # Returns a dataframe
    def buildLPGExposure(self):
        lpgExposure = self.buildDataFrame("BP7_LPG_Premium")
        return lpgExposure.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'LPGPremium': 'Premium (each premises)'})

    # Builds the additional insured - garage operations table
    # Returns a dataframe
    def buildGarageOperations(self):
        garageOperations = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates").query(f'CoverageName == "AdditionalInsrdGarageOperations"').astype({'BaseRate': 'str'})
        garageOperations['Rate'] = '$' + garageOperations['BaseRate'] + ' per additional insured subject to the Liability Limit factor'
        return garageOperations.filter(items=['Rate'])

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.autoProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Auto Service premises'})

    # Builds the optional increased limits - employee portable tools and equipment tables
    # The equalLimits parameter determines which table is being built (i.e. the table where the employee limit is equal to the occurence limit or not)
    # Returns a dataframe
    def buildEmpPortToolsAndEquipment(self, equalLimits):
        empPortToolsAndEquipment = self.buildDataFrame("BP7_EmployeePortableToolsAndEquipmentCov_Rate")
        empPortToolsAndEquipment['OptionalLimitsPerEmployee'] = empPortToolsAndEquipment['Optional_IncreasedLimits'].str.replace(',', '').str.split('/').str[0]
        empPortToolsAndEquipment['OptionalLimitsPerOccurence'] = empPortToolsAndEquipment['Optional_IncreasedLimits'].str.replace(',', '').str.split('/').str[1]
        empPortToolsAndEquipment['TotalLimitsPerEmployee'] = empPortToolsAndEquipment['TotalLimits'].str.replace(',', '').str.split('/').str[0]
        empPortToolsAndEquipment['TotalLimitsPerOccurence'] = empPortToolsAndEquipment['TotalLimits'].str.replace(',', '').str.split('/').str[1]
        filteredEmpPortToolsAndEquipment = empPortToolsAndEquipment.filter(items=['OptionalLimitsPerEmployee', 'OptionalLimitsPerOccurence', 'TotalLimitsPerEmployee', 'TotalLimitsPerOccurence', 'Additional_Premium']). \
                rename(columns={'Additional_Premium': 'Additional Premium'})
        if equalLimits:
            return filteredEmpPortToolsAndEquipment.query(f'TotalLimitsPerEmployee == TotalLimitsPerOccurence'). \
                    astype({'OptionalLimitsPerEmployee': 'int32', 'OptionalLimitsPerOccurence': 'int32', 'TotalLimitsPerEmployee': 'int32', 'TotalLimitsPerOccurence': 'int32'}).sort_values(by=['TotalLimitsPerEmployee'])
        return filteredEmpPortToolsAndEquipment.query(f'TotalLimitsPerEmployee != TotalLimitsPerOccurence'). \
                astype({'OptionalLimitsPerEmployee': 'int32', 'OptionalLimitsPerOccurence': 'int32', 'TotalLimitsPerEmployee': 'int32', 'TotalLimitsPerOccurence': 'int32'}).sort_values(by=['TotalLimitsPerEmployee'])

    # Builds the broadened garage liability – defective products and faulty work coverage endorsement base rate table
    # Returns a dataframe
    def buildBroadenedGarageLiab(self):
        miscBaseRates = self.buildDataFrame("BP7_Miscellaneous_Base_Rates")
        return miscBaseRates.query(f'BaseRateName == "BroadenedGarageLiability"').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the special property damage deductible table
    # Returns a dataframe
    def buildSpecialPropertyDeductible(self):
        specialPropertyDeductible = self.buildDataFrame("BP7_Special_Property_Damage_Deductible")
        return specialPropertyDeductible.astype({'Deductible': 'int32'})

    # Builds the franchise upgrade endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.autoProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(),
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(),
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Inserts the 2-row "Optional Increased Limits / Total Limits" sub-header
    # above the employee portable tools & equipment table. Handled directly
    # (rather than through the generic single-row Sub Headers config) because
    # it's two stacked label rows, not one.
    def _formatEmpPortToolsAndEquipment(self, ws, boldFont):
        ws.insert_rows(3, 2)
        ws['A3'] = 'Optional Increased Limits'
        ws['A4'] = 'Per Employee / Occurrence'
        ws['C3'] = 'Total Limits'
        ws['C4'] = 'Per Employee / Occurrence'
        for cell in ws['3:3']:
            cell.border = _THIN_BORDER
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('A4:B4')
        ws.merge_cells('C3:D3')
        ws.merge_cells('C4:D4')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)
            for row in range(5, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = '$#,##0' if col < 5 else '$#,##0.00'

    # Sets up the Auto Service Excel file and creates a separate worksheet
    # for each of the given dataframes. progress_callback (optional) is
    # called with a short message before each sheet is built.
    # Returns the Excel workbook
    def buildAutoPage(self, progress_callback=None):
        companies = [c for c in self.rateTables.keys() if c != 'CW']

        AutoService = ExcelSettingsBOP.Excel(state=self.state, programName='Auto Service', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        sheetSpecs = []
        for company, tab, label in (('NACO', 'BRNACO', 'NW Assurance'), ('NAFF', 'BRNAFF', 'NW Affinity'),
                                     ('NGIC', 'BRNGIC', 'NW General Insurance Company'), ('NICOF', 'BRNICOF', 'NICOF')):
            if company in self.rateTables:
                sheetSpecs.append((tab, f'AS Table 3.B.1. {label} State Base Rates', lambda c=company: self.buildBaseRates(c), False, True, 'AS_BR', None))

        sheetSpecs += [
            ('CBG', 'AS Table 3.C.2.c. Construction Factor - Building', lambda: self.buildConstructionType('Building'), False, True, None, None),
            ('CPP', 'AS Table 3.C.2.c. Construction Factor - BPP', lambda: self.buildConstructionType('BPP'), False, True, None, None),
            ('TO', 'AS Table 3.C.2.m. Theft Options', self.buildTheftOptions, False, True, None, None),
            ('YBBG', 'AS Table 3.C.2.p. Year Built Modifier - Building', lambda: self.buildYearBuiltModifier('Building'), False, True, None, None),
            ('YBPP', 'AS Table 3.C.2.p. Year Built Modifier - BPP', lambda: self.buildYearBuiltModifier('BPP'), False, True, None, None),
            ('EBB', 'AS Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate, False, True, None, None),
            ('CW', 'AS Table 3.C.4.b. Car Wash Factor', self.buildCarWashFactor, False, True, None, None),
            ('PDLD', 'AS Table 3.C.4.c. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount, False, True, None, None),
            ('LS', 'AS Table 3.C.4.f. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk, False, True, None, None),
            ('LL', 'AS Table 3.C.4.g. Liability Limit Factor', self.buildLiabilityLimitFactor, False, True, None, None),
            ('LPGE', 'AS Table 3.C.4.h. Liquefied Petroleum Gas (LPG) Exposures', self.buildLPGExposure, False, True, None, None),
            ('AIGO', 'AS Table 4.A. Additional Insured - Garage Operations', self.buildGarageOperations, False, True, None, None),
            ('PLUS', 'AS Table 4.B.1. Auto Service PLUS Endorsement', self.buildEndorsementCharge, False, True, None, None),
            ('OILN', 'AS Table 4.B.2. Optional Increased Limits - Employee Portable Tools and Equipment - Employee Limit is not equal to the Occurence Limit', lambda: self.buildEmpPortToolsAndEquipment(False), False, True, None,
             lambda ws: self._formatEmpPortToolsAndEquipment(ws, AutoService.fontBold)),
            ('OILE', 'AS Table 4.B.2. Optional Increased Limits - Employee Portable Tools and Equipment - Employee Limit is equal to the Occurence Limit', lambda: self.buildEmpPortToolsAndEquipment(True), False, True, None,
             lambda ws: self._formatEmpPortToolsAndEquipment(ws, AutoService.fontBold)),
            ('BGL', 'AS Table 4.C.1. Broadened Garage Liability – Defective Products and Faulty Work Coverage Endorsement Base Rate', self.buildBroadenedGarageLiab, False, True, None, None),
            ('SPD', 'AS Table 4.C.2. Special Property Damage Deductible', self.buildSpecialPropertyDeductible, False, True, None, None),
            ('FR', 'AS Table 4.E. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement, False, True, None, None),
        ]

        total = len(sheetSpecs)
        for i, (tableCode, title, build, useIndex, useHeader, layoutKey, postFormat) in enumerate(sheetSpecs, start=1):
            if progress_callback:
                progress_callback(f"Building sheet {i}/{total}: {tableCode}...")
            print(f"  [{i}/{total}] Building sheet: {tableCode}")
            ws = AutoService.generateWorksheet(tableCode, title, build(), useIndex, useHeader, layout_key=layoutKey)
            if postFormat:
                postFormat(ws)

        if progress_callback:
            progress_callback("Building Index sheet...")
        print(f"  [{total}/{total}] Building sheet: Index")
        AutoService.createIndex()
        return AutoService.getWB()
