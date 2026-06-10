# This module formats the Class Modifier State Page workbook in Excel

import pandas as pd
import ExcelSettings
from openpyxl.utils import get_column_letter

class ClassModifier:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the property & liability class modifiers table for the given coverage (building, BPP, business income, liability or EB)
    # Returns a dataframe
    def buildPropertyLiabClassModifiers(self, coverage):
        propertyLiabClassModifiers = self.buildDataFrame("BP7_Peril_Class_Codes")
        ebClassModifier = self.buildDataFrame("BP7_EBClassModifier")
        filteredPropertyLiabClass = propertyLiabClassModifiers.query(f'`Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'BuildingClassCode': 'Code'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            filteredPropertyLiabClass = filteredPropertyLiabClass.query(f'`Peril TypeCode` != "L-OtherMed" & `Peril TypeCode` != "L-OtherPrem" & `Peril TypeCode` != "L-Products" & `Peril TypeCode` != "L-SlipFall" & `Peril TypeCode` != "L-Violence"')
            return filteredPropertyLiabClass.pivot(index='Code', columns='Peril TypeCode', values='BuildingClassFactor').reset_index('Code')
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            filteredPropertyLiabClass = filteredPropertyLiabClass.query(f'`Peril TypeCode` != "L-OtherMed" & `Peril TypeCode` != "L-OtherPrem" & `Peril TypeCode` != "L-Products" & `Peril TypeCode` != "L-SlipFall" & `Peril TypeCode` != "L-Violence"')
            return filteredPropertyLiabClass.pivot(index='Code', columns='Peril TypeCode', values='BPPClassFactor').reset_index('Code')
        elif coverage.casefold() == 'business income':
            filteredPropertyLiabClass = filteredPropertyLiabClass.query(f'`Peril TypeCode` != "L-OtherMed" & `Peril TypeCode` != "L-OtherPrem" & `Peril TypeCode` != "L-Products" & `Peril TypeCode` != "L-SlipFall" & `Peril TypeCode` != "L-Violence"')
            return filteredPropertyLiabClass.pivot(index='Code', columns='Peril TypeCode', values='BIClassFactor').reset_index('Code')
        elif coverage.casefold() == 'liability':
            filteredPropertyLiabClass = filteredPropertyLiabClass.query(f'`Peril TypeCode` == "AllPeril" | `Peril TypeCode` == "L-OtherMed" | `Peril TypeCode` == "L-OtherPrem" | `Peril TypeCode` == "L-Products" | `Peril TypeCode` == "L-SlipFall" | `Peril TypeCode` == "L-Violence"')
            return filteredPropertyLiabClass.pivot(index='Code', columns='Peril TypeCode', values='GLClassFactor').reset_index('Code')
        elif coverage.casefold() == 'eb':
            return ebClassModifier.rename(columns={'Classcode': 'Code', 'EBClassModifier': 'Modifier'})

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    def formatPropertyLiabClass(self, ws):
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = '####0' # Applying unique formatting for the code column
        if ws.title == 'CLEB': # The EB worksheet does not have perils, so formatting it slightly differently
            ws.column_dimensions['B'].width = self.pixelsToInches(80)
        else:
            for col in range(2, ws.max_column + 1):
                char = get_column_letter(col) # Letter representing the current column
                ws.column_dimensions[char].width = self.pixelsToInches(53)

    # Sets up the Class Modifier Excel file and creates a separate worksheet for each of the given dataframes
    # Returns the Excel file
    def buildClassModifierPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        ClassModifier = ExcelSettings.Excel(state=self.state, programName='Class', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        ClassModifier.generateWorksheet('CLBG', 'Table 3.C. Property & Liability Class Modifiers - Building', self.buildPropertyLiabClassModifiers('Building'), False, True)
        ClassModifier.generateWorksheet('CLPP', 'Table 3.C. Property & Liability Class Modifiers - BPP', self.buildPropertyLiabClassModifiers('BPP'), False, True)
        ClassModifier.generateWorksheet('CLBI', 'Table 3.C. Property & Liability Class Modifiers - Bus Inc', self.buildPropertyLiabClassModifiers('Business Income'), False, True)
        ClassModifier.generateWorksheet('CLGL', 'Table 3.C. Property & Liability Class Modifiers - Liability', self.buildPropertyLiabClassModifiers('Liability'), False, True)
        ClassModifier.generateWorksheet('CLEB', 'Table 3.C. Property & Liability Class Modifiers - EB', self.buildPropertyLiabClassModifiers('EB'), False, True)

        ClassModifier.createIndex()
        ClassModifierPages = ClassModifier.getWB()

        self.formatPropertyLiabClass(ClassModifierPages['CLBG'])
        self.formatPropertyLiabClass(ClassModifierPages['CLPP'])
        self.formatPropertyLiabClass(ClassModifierPages['CLBI'])
        self.formatPropertyLiabClass(ClassModifierPages['CLGL'])
        self.formatPropertyLiabClass(ClassModifierPages['CLEB'])

        return ClassModifierPages