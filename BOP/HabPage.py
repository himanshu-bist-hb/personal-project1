# This module builds and formats the Habitational (Hab) State Page workbook (BP-2.0).
#
# The build*() methods below are unchanged business logic, transcribed from
# the root-level HabPage.py: each pulls a table out of the ratebook data (via
# buildDataFrame's nesting waterfall) and shapes it into the DataFrame the
# rate page needs, filtered to the Hab program (Class_Code_Min == 10000).
#
# All Excel formatting (fonts, column widths, sub-header labels, page setup)
# lives in ExcelSettingsBOP.py, driven by "BOP/BOP Input File.xlsx" — the
# Hab-specific column profiles use the HAB_* layout keys there. The two
# tables with formatting too specific for that generic system (Directors &
# Officers row-merges, the CA-only Habitability Exclusion note) are handled
# with a small amount of direct worksheet post-processing below, same as the
# root-level file did.

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment

from . import ExcelSettingsBOP


class Hab:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective  # New business effective date
        self.rEffective = rEffective  # Renewal business effective date

        self.habProgramCode = 10000

    # Builds a dataframe for the given table code
    # The hierarchy matches Business Auto: lower-level company (NACO/NAFF/NICOF)
    # first, since a company-specific filing should override the default;
    # then NGIC (the state-level default company); then CW as the final
    # country-wide fallback. See [[bop-nesting-order]] — the root-level
    # HabPage.py checked NGIC first, which is backwards.
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if 'NACO' in self.rateTables.keys():
            if tableCode in self.rateTables['NACO'].keys():
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys():
            if tableCode in self.rateTables['NAFF'].keys():
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys():
            if tableCode in self.rateTables['NICOF'].keys():
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        if tableCode in self.rateTables['NGIC'].keys():
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0])

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        propertyBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7PerilBldgHabitationalSwimmingPoolsPropertyBaseRate'][1:], index=None, columns=self.rateTables[company]['BP7PerilBldgHabitationalSwimmingPoolsPropertyBaseRate'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        filteredPropertyBaseRates = propertyBaseRates.query(f'`Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"'). \
                    pivot(index='Peril TypeCode', columns='CoverageCode', values='BaseRate').reset_index().rename_axis(None, axis=1). \
                    rename(columns={'POOL': 'Swimming Pool', 'SPA': 'Spa', 'FENCE': 'Fence'})
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        mergedBaseRates = pd.merge(baseRates, filteredPropertyBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(mergedBaseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building",
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the liability charges for related additional exposures table
    # Returns a dataframe
    def buildRelatedAddtExposures(self, company):
        relatedAddtExposures = pd.DataFrame(data=self.rateTables[company]['BP7LiabilityChargesForRelatedAddnlExposures'][1:], index=None, columns=self.rateTables[company]['BP7LiabilityChargesForRelatedAddnlExposures'][0])
        return relatedAddtExposures.rename(columns={'ExposureType': 'Exposure', 'BaseRate': 'Rate'}). \
                replace({'Exposure': {'Clubhouse': 'a. Per Clubhouse',
                                      'ExerciseRoom': 'b. Per Exercise Room',
                                      'Playground': 'c. Per Playground',
                                      'Spa': 'd. Per Spa',
                                      'SwimmingPool': 'e. Per Swimming Pool'}})

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp':
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.habProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'})
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.habProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the number of units factor table
    # Returns a dataframe
    def buildNumOfUnitsFactor(self):
        numberOfUnits = self.buildDataFrame("BP7_Peril_Liability_No_Of_Units_Factor")
        filteredNumberOfUnits = numberOfUnits.fillna({'NoOfUnits_Max': 0}). \
                astype({'NoOfUnits_Min': 'int64', 'NoOfUnits_Max': 'int64'}).astype({'NoOfUnits_Min': 'string', 'NoOfUnits_Max': 'string'})
        filteredNumberOfUnits['Units'] = np.where(filteredNumberOfUnits['NoOfUnits_Max'] == '0',
                                                  'Over 75',
                                                  filteredNumberOfUnits['NoOfUnits_Min'] + ' - ' + filteredNumberOfUnits['NoOfUnits_Max'])
        return filteredNumberOfUnits.replace({'Peril TypeCode': self.perilsConversions}).replace({'Units': {'1 - 1': '01 - 01'}}).replace({'Units': {'2 - 3': '02 - 03'}}).replace({'Units': {'4 - 10': '04 - 10'}}).rename(columns={'LiabilityNoOfUnitsFactor': 'Factor'}).pivot(index='Units', columns='Peril TypeCode', values='Factor').reset_index('Units')

    # Builds the number of stories factor table
    # Returns a dataframe
    def buildNumOfStoriesFactor(self):
        numberOfStories = self.buildDataFrame("BP7_Peril_Liability_No_Of_Stories_Factor")
        filteredNumberOfStories = numberOfStories.query(f'NoOfStories == NoOfStories').replace({'NoOfStories': {'4': '4 or More'}})
        return filteredNumberOfStories.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'NoOfStories': 'Number of Stories', 'LiabilityNoOfStoriesFactor': 'Factor'}).pivot(index='Number of Stories', columns='Peril TypeCode', values='Factor').reset_index('Number of Stories')

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.habProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.habProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance': 'int32'})

    # Builds the directors and officers liability insurance table
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Habitational"').copy()
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 to 100'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 101, 'Number of Units'] = '101 to 250'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 251, 'Number of Units'] = '251 to 500'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 501, 'Number of Units'] = 'over 500'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])

    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Habitational"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Habitational"').copy()
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index=filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.habProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Habitational premises'})

    # Builds the CA-only habitability exclusion table
    # Returns a dataframe
    def buildHabExclusion(self):
        return pd.DataFrame({"Factor": ["0.98"]})

    # Merges the "Number of Units" column of the D&O table into 3-row (or,
    # in WA, 2-row) groups, matching the root-level formatDirsOfficersLiabIns.
    # Handled here directly (rather than through the generic Sub Headers
    # config) because it merges DATA rows, not a header row.
    def _formatDirsOfficersLiabIns(self, ws):
        if self.state == 'WA':
            ws.merge_cells('A4:A6')
            ws.merge_cells('A7:A9')
            ws.merge_cells('A10:A11')
            ws.merge_cells('A12:A13')
            ws.merge_cells('A14:A15')
        else:
            ws.merge_cells('A4:A6')
            ws.merge_cells('A7:A9')
            ws.merge_cells('A10:A12')
            ws.merge_cells('A13:A15')
            ws.merge_cells('A16:A18')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Inserts the CA-only habitability exclusion explanatory note above the
    # factor. Handled directly (rather than through Sub Headers) because it's
    # a one-off note, not a repeatable column-label pattern.
    def _formatHabExclusion(self, ws):
        ws.insert_rows(2, 2)
        ws['A3'] = 'Multiply the factor below to adjust for the exclusion'

    # Sets up the Hab Excel file and creates a separate worksheet for each of
    # the given dataframes. progress_callback (optional) is called with a
    # short message before each sheet is built.
    # Returns the Excel workbook
    def buildHabPage(self, progress_callback=None):
        companies = [c for c in self.rateTables.keys() if c != 'CW']

        Hab = ExcelSettingsBOP.Excel(state=self.state, programName='Habitational', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        # (tab name, page title, builder callable, useIndex, useHeader, layout_key, post-format hook)
        sheetSpecs = []
        for company, tab, label in (('NACO', 'BRNACO', 'NW Assurance'), ('NAFF', 'BRNAFF', 'NW Affinity'),
                                     ('NGIC', 'BRNGIC', 'NW General Insurance Company'), ('NICOF', 'BRNICOF', 'NICOF')):
            if company in self.rateTables:
                sheetSpecs.append((tab, f'H Table 3.B.1. {label} State Base Rates', lambda c=company: self.buildBaseRates(c), False, True, 'HAB_BR', None))
        for company, tab in (('NACO', 'LANACO'), ('NAFF', 'LANAFF'), ('NGIC', 'LANGIC'), ('NICOF', 'LANICOF')):
            if company in self.rateTables:
                sheetSpecs.append((tab, 'H Table 3.C.5. Liability Charges for Related Additional Exposures', lambda c=company: self.buildRelatedAddtExposures(c), False, True, 'HAB_LA', None))

        sheetSpecs += [
            ('CBG', 'H Table 3.C.2.c. Construction Factor - Building', lambda: self.buildConstructionType('Building'), False, True, None, None),
            ('CPP', 'H Table 3.C.2.c. Construction Factor - BPP', lambda: self.buildConstructionType('BPP'), False, True, None, None),
            ('YBBG', 'H Table 3.C.2.o. Year Built Modifier - Building', lambda: self.buildYearBuiltModifier('Building'), False, True, None, None),
            ('YBPP', 'H Table 3.C.2.o. Year Built Modifier - BPP', lambda: self.buildYearBuiltModifier('BPP'), False, True, None, None),
            ('EBB', 'H Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate, False, True, None, None),
            ('NU', 'H Table 3.C.4.a. Number of Units Factor', self.buildNumOfUnitsFactor, False, True, None, None),
            ('NS', 'H Table 3.C.4.b. Number of Stories Factor', self.buildNumOfStoriesFactor, False, True, None, None),
            ('PDLD', 'H Table 3.C.4.d. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount, False, True, None, None),
            ('LL', 'H Table 3.C.4.f. Liability Limit Factor', self.buildLiabilityLimitFactor, False, True, None, None),
            ('DO', 'H Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns, False, True, None, self._formatDirsOfficersLiabIns),
            ('DONM', 'H Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief, False, True, None, None),
            ('ERP', 'H Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods, False, True, None, None),
            ('PLUS', 'H Table 4.B. Habitational PLUS Endorsement', self.buildEndorsementCharge, False, True, None, None),
        ]
        if self.state == 'CA':
            sheetSpecs.append(('HABEX', 'H Table 4.C Habitability Exclusion', self.buildHabExclusion, False, True, None, self._formatHabExclusion))

        total = len(sheetSpecs)
        for i, (tableCode, title, build, useIndex, useHeader, layoutKey, postFormat) in enumerate(sheetSpecs, start=1):
            if progress_callback:
                progress_callback(f"Building sheet {i}/{total}: {tableCode}...")
            print(f"  [{i}/{total}] Building sheet: {tableCode}")
            ws = Hab.generateWorksheet(tableCode, title, build(), useIndex, useHeader, layout_key=layoutKey)
            if postFormat:
                postFormat(ws)

        if progress_callback:
            progress_callback("Building Index sheet...")
        print(f"  [{total}/{total}] Building sheet: Index")
        Hab.createIndex()
        return Hab.getWB()
