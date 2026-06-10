# This module formats the Service State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class Service:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.serviceProgramCode = 70000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company
    
    # Builds the dry cleaners EXTRA factor table
    # Returns a dataframe
    def buildDryCleanersFactor(self):
        dryCleanersFactor = self.buildDataFrame("BP7_Dry_Cleaners_Extra_Factor")
        filteredDryCleanersFactor = dryCleanersFactor.query(f'Class_Code_Min == {self.serviceProgramCode}')
        data = {'TierLimit': ['BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits', 'BaseLimits', 
                    'Tier2', 'Tier2', 'Tier2', 'Tier2', 'Tier2', 
                    'Tier3', 'Tier3', 'Tier3', 'Tier3', 'Tier3', 
                    'Tier4', 'Tier4', 'Tier4', 'Tier4', 'Tier4'],
                'Coverage': ['Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets',
                    'Customers Pers Prop - Not In Process', 'Customers Pers Prop - In Process', 'Fur', 'Garments with Jewels', 'Rugs/Carpets'],
                'Limits (Per Item / Per Occurrence)': ['$1,000/ALS', '$1,000/$10,000', '$1,000/$5,000', '$1,000/$5,000', '$1,000/$5,000',
                    '$2,000/ALS', '$2,000/$15,000', '$2,000/$6,000', '$2,000/$6,000', '$2,000/$6,000',
                    '$3,000/ALS', '$3,000/$20,000', '$3,000/$9,000', '$3,000/$9,000', '$3,000/$9,000',
                    '$5,000/ALS', '$5,000/$25,000', '$5,000/$10,000', '$5,000/$10,000', '$5,000/$10,000']}
        dryCleaners = pd.DataFrame(data)
        finalDryCleaners = pd.merge(dryCleaners, filteredDryCleanersFactor, how='left', on=['TierLimit'])
        return finalDryCleaners.rename(columns={'TierLimit': 'Tier', 'DryCleanerExtraFactor': 'Factor'}).filter(items=['Tier', 'Coverage', 'Limits (Per Item / Per Occurrence)', 'Factor']). \
                replace({'Tier': {'BaseLimits': 'Base', 'Tier2': 'Tier 2', 'Tier3': 'Tier 3', 'Tier4': 'Tier 4'}})

    # Builds the dry cleaners EXTRA earthquake factor table
    # Returns a dataframe
    def buildDryCleanersEQFactor(self):
        miscFactors = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        dryCleanersEQFactor = miscFactors.query(f'FactorName == "DryCleanersExtraEQ"')
        return dryCleanersEQFactor.filter(items=['Factor']).rename(columns={'Factor': 'Dry Cleaners EXTRA EQ Factor'})
    
    # Builds the funeral home EXTRA endorsement table
    # Returns a dataframe
    def buildFuneralHomeEndorsement(self):
        funeralHomeEndorsement = self.buildDataFrame("BP7_FuneralExtraBaseRate")
        funeralHomeEndorsement['Occurence'] = funeralHomeEndorsement['LiabilityLimitOccurence'].apply(lambda x: "${0:,.0f}".format(x))
        funeralHomeEndorsement['Aggregate'] = funeralHomeEndorsement['LiabilityLimtAggregate'].apply(lambda x: "${0:,.0f}".format(x))
        funeralHomeEndorsement['Occurence / Aggregate'] = funeralHomeEndorsement['Occurence'] + ' / ' + funeralHomeEndorsement['Aggregate']
        pivotedFuneralHome = funeralHomeEndorsement.pivot(index=['LiabilityLimitOccurence', 'Occurence / Aggregate'], columns='IncrementalDescedents', values='FuneralExtraBaseRate'). \
                rename(columns={100: 'First 100 decedents', 200: 'Next 200 decedents', 300: 'Next 300 decedents', 400: 'Next 400 decedents', 1000: 'Over 1,000 decedents'}). \
                reset_index(['LiabilityLimitOccurence', 'Occurence / Aggregate']).sort_values(by=['LiabilityLimitOccurence'])
        del pivotedFuneralHome['LiabilityLimitOccurence'] # Remvoing the extra column used for sorting
        return pivotedFuneralHome
    
    # Builds the funeral home EXTRA endorsement minimum premium
    # Returns a dataframe
    def buildFuneralHomeMinPrem(self):
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        funeralHomeMinPrem = miscMinMaxPrem.query(f'CoverageType == "BP7FuneralDirectorsProflLiab"')
        return funeralHomeMinPrem.filter(items=['Premium'])

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        filteredBuilingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        baseRates = pd.merge(filteredBuilingBaseRates, filteredBPPBaseRates, how='inner', on='Peril TypeCode')
        finalBaseRates = pd.merge(baseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building", 
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the territory multiplier table for the given coverage (either building, bpp, or liability)
    # Returns a dataframe
    #def buildTerritoryMultiplier(self, coverage):
    #    territorialFactor = self.buildDataFrame("BP7_Peril_TerritorialFactor")
    #    filteredTerritorialFactor = territorialFactor.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'TerritoryCode': 'Territory'})
    #    if coverage.casefold() == 'building': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BldgTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'bpp': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BPPTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'liability': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='LiabilityTerritoryFactor').reset_index('Territory')

    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'] == 'L-Products'].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.serviceProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'}) # Converting to int first to get rid of decimal places
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0', 
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe    
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.serviceProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.serviceProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance' : 'int32'})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_BLDG_BPP_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCodeMin == {self.serviceProgramCode}').rename(columns={'BldgBPPLimit_Min': 'Limit Min', 'BldgBPPLimit_Max': 'Limit Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Limit Min', 'Limit Max'], columns='Peril TypeCode', values='Factor').reset_index(['Limit Min', 'Limit Max']).fillna({'Limit Max': 'and over'})

    # Builds the general liability occupancy modifiers
    # Returns a dataframe
    def buildGeneralOccupancyMod(self):
        generalOccupancyModifier = self.buildDataFrame("BP7_Peril_Occupant_Factor")
        filteredGeneralOccupancyMod = generalOccupancyModifier.query(f'ClassCode_Min == {self.serviceProgramCode} & `Peril TypeCode` == "liability1"'). \
                rename(columns={'OccupancyType': 'Occupancy', 'BLDGOccupantFactor': 'Building', 'BPPOccupantFactor': 'Business Personal Property'})
        return filteredGeneralOccupancyMod.replace({'Occupancy': {'Condominium': 'Condo Unit-owner', 'buildingOwnerLessorsrisk': "Lessor's Risk", 'buildingOwnerOccupant': 'Owner Occupant', 'tenant': 'Tenant'}}). \
                filter(items=['Occupancy', 'Building', 'Business Personal Property'])

    # Builds the directors and officers liability insurance worksheet
    # Returns a dataframe
    def buildDirsOfficersLiabIns(self):
        dirsOfficersLiabIns = self.buildDataFrame("BP7_DirectorsAndOfficersLiability")
        filteredDirsOfficersLiabIns = dirsOfficersLiabIns.query(f'`Class Code` == "Service"').copy() # Getting a new copy of the data here
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 1, 'Number of Units'] = 'Under 51'
        filteredDirsOfficersLiabIns.loc[filteredDirsOfficersLiabIns['NoofUnitsMin'] == 51, 'Number of Units'] = '51 or More'
        return filteredDirsOfficersLiabIns.rename(columns={'Rate': 'Rate per Unit', 'MinimumPremium': 'Minimum Premium'}).filter(items=['Number of Units', 'Limit', 'Rate per Unit', 'Minimum Premium'])
    
    # Builds the directors and officers liability insurance - non-monetary relief table
    # Returns a dataframe
    def buildDirsOfficersNonMonetaryRelief(self):
        dirsOfficersNonMonetaryRelief = self.buildDataFrame("BP7 Directors And Officers Non Monetary Reliefs")
        filteredNonMonetaryRelief = dirsOfficersNonMonetaryRelief.query(f'`Class Code` == "Service"')
        return filteredNonMonetaryRelief.rename(columns={'LiabilityLimitOfInsurance': 'Liability Limit of Insurance', 'FlatFee': 'Flat Fee'}).filter(items=['Liability Limit of Insurance', 'Flat Fee'])

    # Builds the directors and officers liability insurance - extended reporting periods table
    # Returns a dataframe
    def buildDirsOfficersReportingPeriods(self):
        dirsOfficersReportingPeriods = self.buildDataFrame("BP7_DirectorsAndOfficersLiab_ERP_Pct")
        filteredReportingPeriods = dirsOfficersReportingPeriods.query(f'`Class Code` == "Service"').copy() # Getting a new copy of the data here
        filteredReportingPeriods['PremiumCharge'] = pd.Series(["{0:.0f}%".format(val * 100) for val in filteredReportingPeriods['PremiumCharge']], index = filteredReportingPeriods.index)
        filteredReportingPeriods['Premium Charge'] = filteredReportingPeriods['PremiumCharge'] + ' of annual D&O premium'
        return filteredReportingPeriods.replace({'Years': {'1year': 'One', '2years': 'Two', '3years': 'Three'}}).filter(items=['Years', 'Premium Charge'])
    
    # Builds the barber or beauty shops professional liability table
    # Returns a dataframe
    def buildBarberProfLiab(self):
        barberProfLiab = self.buildDataFrame("BP7_ProfLiabarbersBeauticians_Rate")
        barberProfLiab['Occurence'] = barberProfLiab['LiabilityLimit'].apply(lambda x: "${0:,.0f}".format(x))
        barberProfLiab['Aggregate'] = barberProfLiab['AggregateLimit'].apply(lambda x: "${0:,.0f}".format(x))
        barberProfLiab['Occurence / Aggregate'] = barberProfLiab['Occurence'] + ' / ' + barberProfLiab['Aggregate']
        pivotedBarberProf = barberProfLiab.pivot(index=['LiabilityLimit', 'Occurence / Aggregate'], columns='ProfessionType', values='BaseRate').reset_index(['LiabilityLimit', 'Occurence / Aggregate']). \
                rename(columns={'Barber': 'Each Barber', 'Beautician': 'Each Beautician', 'Manicurist': 'Each Manicurist'}).sort_values(by=['LiabilityLimit'])
        del pivotedBarberProf['LiabilityLimit'] # Removing the extra column used for sorting
        return pivotedBarberProf

    # Builds the endorsement charge table
    # Returns a dataframe
    def buildEndorsementCharge(self):
        endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
        return endorsementCharge.query(f'ClassCodeMIn == {self.serviceProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Service premises'})

    # Builds the franchise upgrage endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.serviceProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(), 
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(), 
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])
    
    def buildRepairSpecializedEndorsement(self):
        data = {"Base premium for each Service premises": ["$300.00"]} 
        data = pd.DataFrame(data)
        return data
    def buildPetSpecializedEndorsement(self):
        data = {"Base premium for each Service premises": ["$212.00"]} 
        data = pd.DataFrame(data)
        return data
    def buildPetServicePL(self):
       data = {"Limits": ["$300,000", "$500,000", "$1,000,000", "$2,000,000"], "Rate" : [ "$43", "$56", "$68", "$83"]}
       return pd.DataFrame(data)
    def buildPetServices(self):
       data = pd.DataFrame({ "Limit": ["$15,000", "$25,000", "$50,000","$100,000"], "Mobile Equipment": ["$49", "$85", "$166", "$220"]})
       return data
    def buildPetServicesCustomized(self):
        data = pd.DataFrame({"Limit": ["$25,000", "$50,000", "$100,000"], "1st Vehicle": ["$91", "$104", "$117"],"Each Addl": ["$46", "$59", "$71"]})
        return data
    def buildservicebusinessincome(self):   
        data = pd.DataFrame({"Limit": ["$25,000", "$50,000", "$100,000"],"1st Worker": ["$13", "$20", "$26"],"Each Addl": ["$7", "$13", "$20"]})
        return data
    def buildvet(self):    
        data = pd.DataFrame({ "Limits": ["$15,000", "$25,000", "$50,000", "$100,000"], "Mobile Equipment": ["$122", "$211", "$414", "$549"]})
        return data
    def buildvetcustom(self):    
        data = pd.DataFrame({ "Limits": ["$25,000", "$50,000", "$100,000"], "1st Vehicle": ["$227", "$260", "$293"], "Each Addl": ["$113", "$146", "$179"]})
        return data
    def buildvetbusinessincome(self):    
        data = pd.DataFrame({ "Limit": ["$25,000", "$50,000", "$100,000"], "1st Worker": ["$32", "$49", "$65"], "Each Addl": ["$17","$32", "$49"]})
        return data
        

       

    
    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the dry cleaners EXTRA factor worksheet
    def formatDryCleanersFactor(self, ws):
        ws.column_dimensions['B'].width = self.pixelsToInches(250)
        ws.column_dimensions['C'].width = self.pixelsToInches(245)
        ws.merge_cells('A4:A8')
        ws.merge_cells('A9:A13')
        ws.merge_cells('A14:A18')
        ws.merge_cells('A19:A23')
        ws.merge_cells('D4:D8')
        ws.merge_cells('D9:D13')
        ws.merge_cells('D14:D18')
        ws.merge_cells('D19:D23')
        ws['A4'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['A9'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['A14'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['A19'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['D4'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['D9'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['D14'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws['D19'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # Applies manual formatting to the dry cleaners EXTRA earthquake factor worksheet
    def formatDryCleanersEQFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(245)
        ws['A4'].number_format = '#,##0.000'
    
    # Applies manual formatting to the funeral home EXTRA endorsement worksheet
    def formatFuneralHomeEndorsement(self, ws):
        for col in range(1, ws.max_column + 1): 
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(150)

    # Applies manual formatting to the funeral home EXTRA endorsement minimum premium worksheet
    def formatFuneralHomeMinPrem(self, ws):
        ws['A4'].number_format = '$#,##0'

    # Applies manual formatting to the base rates worksheet
    def formatBaseRates(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(159)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = '#,##0.0000' # 4 values after the decimal point for the base rates

    # Applies manual formatting to the territory multiplier worksheet
    #def formatTerritoryMultiplier(self, ws):
    #    ws.column_dimensions['A'].width = self.pixelsToInches(70)
    #    for col in range(2, ws.max_column + 1):
    #        char = get_column_letter(col) # Letter representing the current column
    #        ws.column_dimensions[char].width = self.pixelsToInches(54)  

    # Applies manual formatting to the construction factor worksheet
    def formatConstructionFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    # Applies manual formatting to the theft options worksheet
    def formatTheftOptions(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(80)

    # Applies manual formatting to the year built modifier worksheet
    def formatYearBuiltModifier(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(131)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = '###0'

    # Applies manual formatting to the equipment breakdown base rate worksheet
    def formatEBBaseRate(self, ws):
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the property damage deductible worksheet
    def formatPropertyDamageDeductible(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(187)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the liability limit factor worksheet
    def formatLiabilityLimitFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the liability size of risk modifier worksheet
    def formatLiabilitySizeRisk(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Building plus Business Personal Property'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col < 3:
                ws.column_dimensions[char].width = self.pixelsToInches(140)
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(100)
    
    # Applies manual formatting to the general liability occupancy modifiers worksheet
    def formatGeneralOccupancyMod(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(70)
        ws.column_dimensions['C'].width = self.pixelsToInches(215)

    # Applies manual formatting to the directors and officers liability insurance worksheet
    def formatDirsOfficersLiabIns(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(130)
        ws.column_dimensions['D'].width = self.pixelsToInches(140)
        ws.merge_cells('A4:A6')
        ws.merge_cells('A7:A9')
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws['B' + str(row)].number_format = self.noDecimalFormat
            ws['C' + str(row)].number_format = '$#,##0.00'
            ws['D' + str(row)].number_format = '$#,##0.00'

    # Applies manual formatting to the directors and officers liability insurance - non-monetary relief worksheet
    def formatDirsOfficersNonMonetaryRelief(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(225)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat
            ws['B' + str(row)].number_format = '$#,##0.00'

    # Applies manual formatting to the directors and officers liability insurance - extended reporting periods worksheet
    def formatDirsOfficersReportingPeriods(self, ws):
        ws.column_dimensions['B'].width = self.pixelsToInches(215)
    
    # Applies manual formatting to the barber or beauty shops professional liability worksheet
    def formatBarberProfLiab(self, ws):
        for col in range(1, ws.max_column + 1): 
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(170)
            if col > 1: # Applying specific currency formatting to every column except A
                for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = '$#,##0.00'

    # Applies manual formatting to the endorsement charge worksheet
    def formatEndorsementCharge(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(350)
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the franchise endorsement worksheet
    def formatFranchiseEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            cell = ws['B' + str(row)]
            cell.number_format = '$#,##0.00' 

    def formatRepairSpecializedEndorsement(self,ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(120)
        ws.column_dimensions['B'].width = self.pixelsToInches(120)
        ws.column_dimensions['C'].width = self.pixelsToInches(120)
        ws['A4'].number_format = '$#,##0.00'
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:C4")

    def formatPetSpecializedEndorsement(self,ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(120)
        ws.column_dimensions['B'].width = self.pixelsToInches(120)
        ws.column_dimensions['C'].width = self.pixelsToInches(120)

        ws.insert_rows(6)
        bold_font = Font(bold=True, name = 'Arial', size = 10)

        ws["A6"] = "Pet Services - Business Income"
        ws["A5"].border = Border()
        ws["A6"] = "Pet Services - Business Income"
        ws.merge_cells("A3:C3")
        ws.merge_cells("A4:C4")


        rows = [6, 7]
        for row in rows:
            for cell in ws[row]:
                cell.font = bold_font
        ws['A4'].number_format = '$#,##0.00'

    def formatPetServicePL(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        ws.insert_rows(3)
    def formatMPVS(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(150)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)
        ws.insert_rows(3)
        ws.insert_rows(10)
        ws.insert_rows(16)
        ws.insert_rows(22)
        ws.insert_rows(29)
        ws.insert_rows(35)
        ws["A3"] = "Pet Services"
        ws["A10"] = "Pet Services per Customized Vehicle"
        ws["A16"] = "Pet Services - Business Income"
        ws["A22"] = "Veterinarian"
        ws["A29"] = "Veterinarian Services per Customized Vehicle"
        ws["A35"] = "Veterinarian Services - Business Income"
        ws["A9"].border = Border()
        ws["A15"].border = Border()
        ws["A21"].border = Border()
        ws["A28"].border = Border()
        ws["A34"].border = Border()

        
        rows = [3,4,10,11,16,17,22,23,29,30,35,36]
        for row in rows:
            for cell in ws[row]:
                cell.font = Font(bold = True, name = "Arial", size = 10)
    

        
    # Sets up the Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildServicePage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW':  # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        Service = ExcelSettings.Excel(
            state=self.state,
            programName='Service',
            nEffective=self.nEffective,
            rEffective=self.rEffective,
            companyList=companies)
        

        fontName = Service.getFontName()
        fontSize = Service.getFontSize()

        Service.generateWorksheet('DC', 'S Table 1.A.5.c. Dry Cleaners EXTRA Factor', self.buildDryCleanersFactor(), False, True)
        Service.generateWorksheet('DCEQ', 'S Table 1.A.5.d. Dry Cleaners EXTRA Earthquake Factor', self.buildDryCleanersEQFactor(), False, True)
        Service.generateWorksheet('FU', 'S Table 1.C.5.a. Funeral Home EXTRA Endorsement', self.buildFuneralHomeEndorsement(), False, True)
        Service.generateWorksheet('FUMP', 'S Table 1.C.5.c. Funeral Home EXTRA Endorsement Minimum Premium', self.buildFuneralHomeMinPrem(), False, True)
        if 'NACO' in self.rateTables.keys():
            Service.generateWorksheet('BRNACO', 'S Table 3.B.1. NW Assurance State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NAFF' in self.rateTables.keys():
            Service.generateWorksheet('BRNAFF', 'S Table 3.B.1. NW Affinity State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NGIC' in self.rateTables.keys():
            Service.generateWorksheet('BRNGIC', 'S Table 3.B.1. NW General Insurance Company', self.buildBaseRates('NGIC'), False, True)
        if 'NICOF' in self.rateTables.keys():
            Service.generateWorksheet('BRNICOF', 'S Table 3.B.1. NICOF State Base Rates', self.buildBaseRates('NGIC'), False, True)
        #Service.generateWorksheet('TRBG', 'S Table 3.C.1.a. State Territory Multiplier - Building', self.buildTerritoryMultiplier('Building'), False, True)
        #Service.generateWorksheet('TRPP', 'S Table 3.C.1.a. State Territory Multiplier - BPP', self.buildTerritoryMultiplier('BPP'), False, True)
        #Service.generateWorksheet('TRLB', 'S Table 3.C.1.a. State Territory Multiplier - Liability', self.buildTerritoryMultiplier('Liability'), False, True)
        Service.generateWorksheet('CBG', 'S Table 3.C.2.c. Construction Factor - Building', self.buildConstructionType('Building'), False, True)
        Service.generateWorksheet('CPP', 'S Table 3.C.2.c. Construction Factor - BPP', self.buildConstructionType('BPP'), False, True)
        Service.generateWorksheet('ET', 'S Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions(), False, True)
        Service.generateWorksheet('YBBG', 'S Table 3.C.2.p. Year Built Modifier - Building', self.buildYearBuiltModifier('Building'), False, True)
        Service.generateWorksheet('YBPP', 'S Table 3.C.2.p. Year Built Modifier - BPP', self.buildYearBuiltModifier('BPP'), False, True)
        Service.generateWorksheet('EBB', 'S Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate(), False, True)
        Service.generateWorksheet('PDLD', 'S Table 3.C.4.b. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount(), False, True)
        Service.generateWorksheet('LL', 'S Table 3.C.4.d. Liability Limit Factor', self.buildLiabilityLimitFactor(), False, True)
        Service.generateWorksheet('LS', 'S Table 3.C.4.e. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk(), False, True)
        Service.generateWorksheet('GLO', 'S Table 3.D.1.c. General Liability Occupancy Modifiers', self.buildGeneralOccupancyMod(), False, True)
        Service.generateWorksheet('DO', 'S Table 4.A.1. Directors and Officers Liability Insurance', self.buildDirsOfficersLiabIns(), False, True)
        Service.generateWorksheet('DONM', 'S Table 4.A.2. Directors and Officers Liability Insurance - Non-Monetary Relief', self.buildDirsOfficersNonMonetaryRelief(), False, True)
        Service.generateWorksheet('ERP', 'S Table 4.A.3. Directors and Officers Liability Insurance - Extended Reporting Periods', self.buildDirsOfficersReportingPeriods(), False, True)
        Service.generateWorksheet('BB', 'S Table 4.B.1.e.(1). Barber or Beauty Shops Professional Liability', self.buildBarberProfLiab(), False, True)
        Service.generateWorksheet('PLUS', 'S Table 4.C. Service PLUS Endorsement', self.buildEndorsementCharge(), False, True)
        Service.generateWorksheet('FR', 'S Table 4.D. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement(), False, True)
        Service.generateWorksheet('RSS', 'S Table 4.E. Repair Services Specialized Endorsement', self.buildRepairSpecializedEndorsement(), False, True)
        Service.generateWorksheet2tables('PSS', 'S Table 4.F. Pet Services Specialized Endorsement', self.buildPetSpecializedEndorsement(), self.buildservicebusinessincome(), False, True)
        Service.generateWorksheet('PSPL', 'S Table 4.G. Pet Services Professional Liability', self.buildPetServicePL(), False, True)
        Service.generateWorksheet6tables('MPVS', 'S Table 4.H. Mobile Pet and Veterinarian Services Endorsement', self.buildPetServices(),self.buildPetServicesCustomized(), self.buildservicebusinessincome(), self.buildvet(), self.buildvetcustom(), self.buildvetbusinessincome(), False, True)


        Service.createIndex()
        ServicePages = Service.getWB()

        self.formatDryCleanersFactor(ServicePages['DC'])
        self.formatDryCleanersEQFactor(ServicePages['DCEQ'])
        self.formatFuneralHomeEndorsement(ServicePages['FU'])
        self.formatFuneralHomeMinPrem(ServicePages['FUMP'])
        if 'NACO' in self.rateTables.keys():
            self.formatBaseRates(ServicePages['BRNACO'])
        if 'NAFF' in self.rateTables.keys():
            self.formatBaseRates(ServicePages['BRNAFF'])
        if 'NGIC' in self.rateTables.keys():
            self.formatBaseRates(ServicePages['BRNGIC'])
        if 'NICOF' in self.rateTables.keys():
            self.formatBaseRates(ServicePages['BRNICOF'])
        #self.formatTerritoryMultiplier(ServicePages['TRBG'])
        #self.formatTerritoryMultiplier(ServicePages['TRPP']) 
        #self.formatTerritoryMultiplier(ServicePages['TRLB'])
        self.formatConstructionFactor(ServicePages['CBG'])  
        self.formatConstructionFactor(ServicePages['CPP'])  
        self.formatTheftOptions(ServicePages['ET'])
        self.formatYearBuiltModifier(ServicePages['YBBG'])
        self.formatYearBuiltModifier(ServicePages['YBPP'])  
        self.formatEBBaseRate(ServicePages['EBB'])
        self.formatPropertyDamageDeductible(ServicePages['PDLD'])
        self.formatLiabilityLimitFactor(ServicePages['LL'])
        self.formatLiabilitySizeRisk(ServicePages['LS'], Font(name=fontName, size=fontSize, bold=True))
        self.formatGeneralOccupancyMod(ServicePages['GLO'])
        self.formatDirsOfficersLiabIns(ServicePages['DO'])
        self.formatDirsOfficersNonMonetaryRelief(ServicePages['DONM'])
        self.formatDirsOfficersReportingPeriods(ServicePages['ERP'])
        self.formatBarberProfLiab(ServicePages['BB'])
        self.formatEndorsementCharge(ServicePages['PLUS'])
        self.formatFranchiseEndorsement(ServicePages['FR'])
        self.formatRepairSpecializedEndorsement(ServicePages['RSS'])
        self.formatPetSpecializedEndorsement(ServicePages['PSS'])
        self.formatPetServicePL(ServicePages['PSPL'])
        self.formatMPVS(ServicePages['MPVS'])
        

        return ServicePages