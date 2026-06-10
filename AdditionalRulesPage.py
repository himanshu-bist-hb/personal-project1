# This module formats the Rating Plans State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class AdditionalRules:
    def __init__(self, state, rateTables, perils, perilsConversions, classCodes, nEffective, rEffective, AdditionalRulesApplies) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.classCodes = classCodes
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date
        self.AdditionalRulesApplies = AdditionalRulesApplies

        #self.autoProgramCode = 20000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'
        self.percentageFormat = '#,##0%'

    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildRateCappingFactors(self):
        RateCapping = self.buildDataFrame("BP7_RateCappingPremiumRange2")
        filteredRateCapping = RateCapping.query(f'RateCapType == "Migration" & FiledRatesReachedIndicator == 0 & YrsOnPCMin == 1').filter(items=['MinimumRange', 'MaximumRange']).rename(columns={'MinimumRange': 'Lower Bound', 'MaximumRange': 'Upper Bound'})
        return filteredRateCapping
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildDistributionFactors(self):
        DistributionFactors = self.buildDataFrame("BP7Distribution Factor_Ext")
        DistributionFactors = DistributionFactors.query(f'DistributionGroup != "DG99"').rename(columns={'DistributionGroup': 'Distribution Group'}).replace({'Distribution Group' : {'DG00' : '00', 'DG01' : '01', 'DG02' : '02', 'DG03' : '03', 'DG04' : '04', 'DG05' : '05', 'DG06' : '06', 'DG07' : '07', 'DG08' : '08', 'DG09' : '09', 'DG10' : '10', 'DG11' : '11', 'DG12' : '12', 'DG13' : '13', 'DG14' : '14', 'DG15' : '15'}})
        return DistributionFactors
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildWHExclusion(self):
        WHExclusion = self.buildDataFrame("BP7_WHExclusionFactor")
        WHExclusion = WHExclusion.rename(columns={'BuildingWHExclusionFactor' : 'Building', 'BPPWHExclusionFactor' : 'BPP', 'ClassCode_Min' : 'Program'}).replace({'Program': self.classCodes}).filter(items=['Program', 'Building', 'BPP']).replace({'Program': self.classCodes}).replace({'Program' : {'Hab' : 'Habitational', 'Food' : 'Food Service', 'Auto' : 'Auto Service', 'Service' : 'Process/Service'}}).sort_values('Program')
        return WHExclusion
    
    # Builds the Liability for Hazards of Lead Factor tables
    # Returns a dataframe
    def buildLiabilityHazards(self):
        LiabilityHazards = self.buildDataFrame("BP7_LiabForHazdOfLead_Factor")
        LiabilityHazards = LiabilityHazards.fillna({'YearBuilt': 'Buildings built 1979 and later:'}).replace({'YearBuilt' : {1979 : 'For buildings built prior to 1979:'}}).replace({'HazardOfLeadFactor' : {1 : 'No Charge'}})
        LiabilityHazards = LiabilityHazards.rename(columns={'YearBuilt' : 'Year Built', 'HazardOfLeadFactor' : 'General Liability'}).sort_values('Year Built', ascending=False)
        return LiabilityHazards
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildIBHSzones(self):
        IBHSzones = self.buildDataFrame("BP7_Wind_Mitigation_Zone")
        IBHSzones = IBHSzones.rename(columns={'County Name': 'Counties'}).replace({'IBHS Zone': {'Central' : 'Central Zone', 'Coastal' : 'Coastal Zone', 'Northern' : 'Northern Zone'}}).filter(items=['Counties', 'IBHS Zone'])
        IBHSzones = IBHSzones.groupby('IBHS Zone')['Counties'].apply(', '.join).reset_index()
        IBHSzones.loc[IBHSzones['IBHS Zone'] == 'Northern Zone', 'Counties'] = 'All Other Counties'
        return IBHSzones
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildIBHSHurricane(self):
        IBHSHurricane = self.buildDataFrame("BP7_Wind_Mitigation")
        IBHSHurricane = IBHSHurricane.query(f'`Sub-Decking` != "No" & `Sub-Decking` != "Yes" & `Roof Age Max` != 999 & `Certificate Type` == "IBHS Hurricane"')
        IBHSHurricane = IBHSHurricane.pivot(index='IBHS Zone', columns='Certificate Level', values='Factor').reset_index('IBHS Zone').rename(columns={'Fortified Bronze' : 'Bronze', 'Fortified Silver' : 'Silver', 'Fortified Gold' : 'Gold'})
        IBHSHurricane = IBHSHurricane[['IBHS Zone', '2006+ IBC', 'Bronze', 'Silver', 'Gold']]
        IBHSHurricane['2006+ IBC'] = 1.00 - IBHSHurricane['2006+ IBC']
        IBHSHurricane['Bronze'] = 1.00 - IBHSHurricane['Bronze']
        IBHSHurricane['Silver'] = 1.00 - IBHSHurricane['Silver']
        IBHSHurricane['Gold'] = 1.00 - IBHSHurricane['Gold']
        return IBHSHurricane
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildIBHSWindHail(self):
        IBHSWindHail = self.buildDataFrame("BP7_Wind_Mitigation")
        IBHSWindHail = IBHSWindHail.query(f'`Sub-Decking` != "No" & `Sub-Decking` != "Yes" & `Roof Age Max` != 999 & `Certificate Type` == "IBHS High Wind and Hail"')
        IBHSWindHail = IBHSWindHail.pivot(index='IBHS Zone', columns='Certificate Level', values='Factor').reset_index('IBHS Zone').rename(columns={'Fortified Bronze' : 'Bronze', 'Fortified Silver' : 'Silver', 'Fortified Gold' : 'Gold'})
        IBHSWindHail = IBHSWindHail[['IBHS Zone', '2006+ IBC', 'Bronze', 'Silver', 'Gold']]
        IBHSWindHail['2006+ IBC'] = 1 - IBHSWindHail['2006+ IBC']
        IBHSWindHail['Bronze'] = 1 - IBHSWindHail['Bronze']
        IBHSWindHail['Silver'] = 1 - IBHSWindHail['Silver']
        IBHSWindHail['Gold'] = 1 - IBHSWindHail['Gold']
        return IBHSWindHail
    
    # Builds the Accounts Receivable base rate table
    # Returns a dataframe
    def buildMineSubsidence(self):
        MineDwelling = self.buildDataFrame("BP7_Bldg_Mine_Subsidence_Charge_Dwelling")
        MineDwelling = MineDwelling.astype({'Amt_Insurance_Min': 'int64', 'Amt_Insurance_Max': 'int64'})
                #astype({'Amt_Insurance_Min': 'string', 'Amt_Insurance_Max': 'string'}) # Converting to int first to get rid of decimal places
        MineDwelling['Amt_Insurance_Min'] = "$" + MineDwelling['Amt_Insurance_Min'].apply('{:,.0f}'.format)
        MineDwelling['Amt_Insurance_Max'] = "$" + MineDwelling['Amt_Insurance_Max'].apply('{:,.0f}'.format)
        MineDwelling["Amount of Insurance D"] = MineDwelling["Amt_Insurance_Min"] + ' - ' + MineDwelling["Amt_Insurance_Max"] # Creating a single column for the percentage
        MineDwelling = MineDwelling.replace({'Amount of Insurance D' : {'$1 - $25,000' : 'Up to $25,000'}}).filter(items=['Amount of Insurance D', 'MineSubsidenceChargeDwelling'])
        MineDwelling = MineDwelling.rename(columns={'MineSubsidenceChargeDwelling' : 'Dwelling Structure'}).sort_values('Dwelling Structure')
        MineNonDwelling = self.buildDataFrame("BP7_Bldg_Mine_Subsidence_Charge_NonDwelling")
        MineNonDwelling = MineNonDwelling.astype({'Amt_Insurance_Min': 'int64', 'Amt_Insurance_Max': 'int64'})
        MineNonDwelling['Amt_Insurance_Min'] = "$" + MineNonDwelling['Amt_Insurance_Min'].apply('{:,.0f}'.format)
        MineNonDwelling['Amt_Insurance_Max'] = "$" + MineNonDwelling['Amt_Insurance_Max'].apply('{:,.0f}'.format)
                #astype({'Amt_Insurance_Min': 'string', 'Amt_Insurance_Max': 'string'}) # Converting to int first to get rid of decimal places
        MineNonDwelling["Amount of Insurance ND"] = MineNonDwelling["Amt_Insurance_Min"] + ' - ' + MineNonDwelling["Amt_Insurance_Max"] # Creating a single column for the percentage
        MineNonDwelling = MineNonDwelling.replace({'Amount of Insurance ND' : {'$1 - $25,000' : 'Up to $25,000'}}).filter(items=['Amount of Insurance ND', 'MineSubsidenceChargeNonDwelling'])
        MineNonDwelling = MineNonDwelling.rename(columns={'MineSubsidenceChargeNonDwelling' : 'Non-Dwelling Structure'}).sort_values('Non-Dwelling Structure')
        MineSubsidence = MineDwelling.join(MineNonDwelling)
        MineSubsidence = MineSubsidence.fillna({'Amount of Insurance D': 'N/A', 'Dwelling Structure' : 'N/A', 'Amount of Insurance ND' : 'N/A', 'Non-Dwelling Structure' : 'N/A'})
        MineSubsidence = MineSubsidence.replace({'Amount of Insurance ND' : {'N/A' : ' '}}).replace({'Non-Dwelling Structure' : {'N/A' : ' '}}).rename(columns={'Amount of Insurance D' : 'Amount of Insurance', 'Amount of Insurance ND' : 'Amount of Insurance'})
        return MineSubsidence
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildMSB(self):
        Msb = self.buildDataFrame("BP7_Bldg_MineSubsidenceCov_Rate")
        Msb = Msb.query(f'County == "ATHENS"').replace({'County' : {'ATHENS' : 'Charge per property location:'}})
        return Msb
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildMSC(self):
        Msc = self.buildDataFrame("BP7_Bldg_MineSubsidenceCov_Rate")
        Msc = Msc.query(f'County == "ERIE"').replace({'County' : {'ERIE' : 'Charge per property location:'}})
        return Msc
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildSGLIL(self):
        SGlil = self.buildDataFrame("BP7_Stop_Gap_Base_Rate")
        SGlil = SGlil.query(f'Limit == "500000/500000/500000" | Limit == "1000000/1000000/1000000"')
        SGlil = SGlil.rename(columns={'Limit' : 'Increased Limits', 'MinimumPremium' : 'Minimum Premium'}).replace({'Increased Limits' : {'100000/100000/500000' : '$100,000/100,000/500,000', '500000/500000/500000' : '$500,000/500,000/500,000', '1000000/1000000/1000000' : '$1,000,000/1,000,000/1,000,000'}})
        return SGlil
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildSGLPD(self):
        SGlpd = self.buildDataFrame("BP7_Stop_Gap_Base_Rate")
        SGlpd = SGlpd.query(f'Limit == "100000/100000/500000"')
        SGlpd = SGlpd.rename(columns={'Limit' : 'Increased Limits', 'MinimumPremium' : 'Minimum Premium'}).replace({'Increased Limits' : {'100000/100000/500000' : '$100,000/100,000/500,000', '500000/500000/500000' : '$500,000/500,000/500,000', '1000000/1000000/1000000' : '$1,000,000/1,000,000/1,000,000'}}). \
            filter(items=['Rate', 'Minimum Premium']).rename(columns={'Rate' : 'Rate (per $100 of payroll)'})
        return SGlpd
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildSGEIL(self):
        SGeil = self.buildDataFrame("BP7_Extended_Stop_Gap_Base_Rate")
        SGeil = SGeil.query(f'Limit == "500000/500000/500000" | Limit == "1000000/1000000/1000000"')
        SGeil = SGeil.rename(columns={'Limit' : 'Increased Limits', 'MinimumPremium' : 'Minimum Premium'}).replace({'Increased Limits' : {'100000/100000/500000' : '$100,000/100,000/500,000', '500000/500000/500000' : '$500,000/500,000/500,000', '1000000/1000000/1000000' : '$1,000,000/1,000,000/1,000,000'}})
        return SGeil
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildSGEPD(self):
        SGepd = self.buildDataFrame("BP7_Extended_Stop_Gap_Base_Rate")
        SGepd = SGepd.query(f'Limit == "100000/100000/500000"')
        SGepd = SGepd.rename(columns={'Limit' : 'Increased Limits', 'MinimumPremium' : 'Minimum Premium'}).replace({'Increased Limits' : {'100000/100000/500000' : '$100,000/100,000/500,000', '500000/500000/500000' : '$500,000/500,000/500,000', '1000000/1000000/1000000' : '$1,000,000/1,000,000/1,000,000'}}). \
            filter(items=['Rate', 'Minimum Premium']).rename(columns={'Rate' : 'Rate (per $100 of payroll)'})
        return SGepd
    
    # Builds the Rate Capping Factor tables
    # Returns a dataframe
    def buildUPS(self):
        Ups = self.buildDataFrame("BP7_Pol_UGPetroleumStorageTank_BaseRate")
        Ups = Ups.rename(columns={'NoOfTanks' : 'Number of Tanks', 'LimitPerTank' : 'Limit Per Tank', 'AggregateLimit' : 'Aggregate Limit', 'UGPetroleumStorageTankBaseRate' : 'Premium'})
        return Ups



    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the Rate Capping worksheet
    def formatRateCappingFactors(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(80)
        ws.column_dimensions['B'].width = self.pixelsToInches(80)
        for cell in ws['4:4']:
            cell.number_format = '#,##0%'

    # Applies manual formatting to the Rate Capping worksheet
    def formatDistributionFactors(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)

    # Applies manual formatting to the Rate Capping worksheet
    def formatWHExclusion(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(125)

    # Applies manual formatting to the Rate Capping worksheet
    def formatLiabilityHazards(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(190)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)

    # Applies manual formatting to the Rate Capping worksheet
    def formatIBHSZones(self, ws, boldFont):
        ws.row_dimensions[4].height = self.pixelsToInches(450)
        ws.row_dimensions[5].height = self.pixelsToInches(250)
        for cell in ws['4:4']:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws['5:5']:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for cell in ws['6:6']:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.merge_cells('B3:E3')
        ws.merge_cells('B4:E4')
        ws.merge_cells('B5:E5')
        ws.merge_cells('B6:E6')
        
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    if col >= 2: 
                        cell.number_format = '#,##0%' # Applying currency formatting to columns A-
        for cell in ws['8:8']:
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        for cell in ws['13:13']:
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)

        for cell in ws['7:7']:
            cell.border = None
        for cell in ws['12:12']:
            cell.border = None

        ws.insert_rows(7)
        ws['B8'] = 'Commercial Hurricane Premium Discounts*'
        ws.insert_rows(13)
        ws.insert_rows(14)
        ws.insert_rows(15)
        ws['A13'] = '* - Adjustments: Metal Roof > 10 years old or metal roof with no sub-decking, or both; all non-metal roofs > 5 years old:'
        ws['A14'] = '10 point reduction from above discounts all zones'
        ws['B16'] = 'Commercial Other Wind & Hail Premium Discounts**'
        ws['A21'] = '** - Adjustments: Metal Roof > 10 years old, All other roofs > 5 years old:'
        ws['A22'] = '10 point reduction from above discounts all zones'

        for cell in ws['8:8']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('B8:E8')

        for cell in ws['16:16']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('B16:E16')

        for cell in ws['14:14']:
            cell.font = boldFont

        for cell in ws['22:22']:
            cell.font = boldFont

    def formatMineSubsidence(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col >= 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(175)
        ws.column_dimensions['C'].width = self.pixelsToInches(175)

    def formatMSB(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(175)
        ws.column_dimensions['C'].width = self.pixelsToInches(175)
        ws.delete_rows(3)
        for cell in ws['3:3']:
            cell.border = None

    def formatMSC(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(175)
        ws.column_dimensions['C'].width = self.pixelsToInches(175)
        ws.delete_rows(3)
        for cell in ws['3:3']:
            cell.border = None

    def formatSGlil(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 2: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(200)
        ws.column_dimensions['C'].width = self.pixelsToInches(150)

    def formatSGlpd(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(175)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)

    def formatSGeil(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 2: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(200)
        ws.column_dimensions['C'].width = self.pixelsToInches(150)

    def formatSGepd(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(175)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)

    def formatUPS(self, ws):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                if col > 1: 
                    cell.number_format = self.currencyFormat # Applying currency formatting to columns A-B
        ws.column_dimensions['A'].width = self.pixelsToInches(175)
        ws.column_dimensions['B'].width = self.pixelsToInches(150)
        ws.column_dimensions['C'].width = self.pixelsToInches(150)
        ws.column_dimensions['D'].width = self.pixelsToInches(150)
        ws['A10'] = "7 or more"
        ws['B10'] = "Ineligible for coverage"
        ws.merge_cells('B10:D10')
        for cell in ws['10:10']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        

    
    # Sets up the Auto Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildAdditionalRulesPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        AdditionalRules = ExcelSettings.Excel(state=self.state, programName='Additional Rules', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = AdditionalRules.getFontName()
        fontSize = AdditionalRules.getFontSize()


        if self.state == "MD":
            AdditionalRules.generateWorksheet('LHL', 'AR Table 1.C.1-2. Liability for Hazards of Lead', self.buildLiabilityHazards(), False, True)
            #AdditionalRules.generateWorksheet('DF', 'AR - 2 Distribution Factor', self.buildDistributionFactors(), False, True)
            AdditionalRules.generateWorksheet('WHE', 'AR Table A.3. Windstorm or Hail Exclusion (Coverages Not Included in By-peril Rating)', self.buildWHExclusion(), False, True)
            AdditionalRules.generateWorksheet('RC', 'AR-4 Table A.4. Transition Capping Program', self.buildRateCappingFactors(), False, True)
        elif self.state == "OH":
            AdditionalRules.generateWorksheet('MSB', 'AR-1 Table B.2. Mine Subsidence', self.buildMSB(), False, True)
            AdditionalRules.generateWorksheet('MSC', 'AR-1 Table C.2. Mine Subsidence', self.buildMSC(), False, True)
            AdditionalRules.generateWorksheet('SGLIL', 'AR-2 Table C.2 Stop Gap - Employers Liability Coverage Increased Limits', self.buildSGLIL(), False, True)
            AdditionalRules.generateWorksheet('SGLPD', 'AR-2 Table D. Stop Gap - Employers Liability Coverage Premium Determination', self.buildSGLPD(), False, True)
            AdditionalRules.generateWorksheet('SGEIL', 'AR-2 Table E.4.C Stop Gap - Extended Coverage Endorsement Increased Limits', self.buildSGEIL(), False, True)
            AdditionalRules.generateWorksheet('SGEPD', 'AR-2 Table E.4 Stop Gap - Extended Coverage Endorsement Premium Determination', self.buildSGEPD(), False, True)
            AdditionalRules.generateWorksheet('UPS', 'AR-3 Table F. Underground Petroleum Storage Tank Deductible Coverage', self.buildUPS(), False, True)
            AdditionalRules.generateWorksheet('RC', 'AR-4 Table B. Transition Capping Program', self.buildRateCappingFactors(), False, True)
        elif self.state == "AL":
            AdditionalRules.generateWorksheet('WHE', 'AR Table 1.3 Windstorm or Hail Exclusion (Coverages Not Included in By-peril Rating)', self.buildWHExclusion(), False, True)
            AdditionalRules.generateWorksheet3tables('IBHS', 'AR Table 2. IBHS Certificate Discounts', self.buildIBHSzones(), self.buildIBHSHurricane(), self.buildIBHSWindHail(), False, True)
        elif self.state == "IN" or self.state == "KY":
            AdditionalRules.generateWorksheet('MS', 'AR Table 1. Mine Subsidence Insurance', self.buildMineSubsidence(), False, True)
        elif self.state == "WA":
            AdditionalRules.generateWorksheet('SGLIL', 'AR-1 Table C.2 Stop Gap - Employers Liability Coverage Increased Limits', self.buildSGLIL(), False, True)
            AdditionalRules.generateWorksheet('SGLPD', 'AR-1 Table D. Stop Gap - Employers Liability Coverage Premium Determination', self.buildSGLPD(), False, True)
        else:
            AdditionalRules.generateWorksheet('RC', 'AR Table 98.C. Rate Capping', self.buildRateCappingFactors(), False, True)
            AdditionalRules.generateWorksheet('DF', 'AR - 2 Distribution Factor', self.buildDistributionFactors(), False, True)



        AdditionalRules.createIndex()
        AdditionalRulesPages = AdditionalRules.getWB()

        if self.state == "MD":
            self.formatLiabilityHazards(AdditionalRulesPages['LHL']) 
            #self.formatDistributionFactors(AdditionalRulesPages['DF']) 
            self.formatWHExclusion(AdditionalRulesPages['WHE'])
            self.formatRateCappingFactors(AdditionalRulesPages['RC'])
        elif self.state == "OH":
             self.formatMSB(AdditionalRulesPages['MSB']) 
             self.formatMSC(AdditionalRulesPages['MSC']) 
             self.formatSGlil(AdditionalRulesPages['SGLIL']) 
             self.formatSGlpd(AdditionalRulesPages['SGLPD']) 
             self.formatSGeil(AdditionalRulesPages['SGEIL']) 
             self.formatSGepd(AdditionalRulesPages['SGEPD']) 
             self.formatUPS(AdditionalRulesPages['UPS']) 
             self.formatRateCappingFactors(AdditionalRulesPages['RC'])
        elif self.state == "AL":
            self.formatWHExclusion(AdditionalRulesPages['WHE'])
            self.formatIBHSZones(AdditionalRulesPages['IBHS'], Font(name=fontName, size=fontSize, bold=True)) 
        elif self.state == "IN" or self.state == "KY":
            self.formatMineSubsidence(AdditionalRulesPages['MS'])
        elif self.state == "WA":
            self.formatSGlil(AdditionalRulesPages['SGLIL']) 
            self.formatSGlpd(AdditionalRulesPages['SGLPD']) 
        else:
            self.formatRateCappingFactors(AdditionalRulesPages['RC']) 
            self.formatDistributionFactors(AdditionalRulesPages['DF']) 

        return AdditionalRulesPages