# This module formats the Food Service State Page workbook in Excel

import pandas as pd
import numpy as np
import ExcelSettings
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side

class Food:
    def __init__(self, state, rateTables, perils, perilsConversions, nEffective, rEffective) -> None:
        self.state = state
        self.rateTables = rateTables
        self.perils = perils
        self.perilsConversions = perilsConversions
        self.nEffective = nEffective # New business effective date
        self.rEffective = rEffective # Renewal business effective date

        self.foodProgramCode = 40000
        self.currencyFormat = '$#,##0'
        self.noDecimalFormat = '#,##0'
    
    # Builds a dataframe for the given table code
    # The hierarchy is as follows: NGIC > Migration > CW
    # Returns the dataframe that was built
    def buildDataFrame(self, tableCode):
        if tableCode in self.rateTables['NGIC'].keys(): # Checking if table exists in the NGIC ratebook
            return pd.DataFrame(data=self.rateTables['NGIC'][tableCode][1:], index=None, columns=self.rateTables['NGIC'][tableCode][0])
        if 'NACO' in self.rateTables.keys(): # Checking if NACO file was given
            if tableCode in self.rateTables['NACO'].keys(): # Checking if table exists in NACO
                return pd.DataFrame(data=self.rateTables['NACO'][tableCode][1:], index=None, columns=self.rateTables['NACO'][tableCode][0])
        if 'NAFF' in self.rateTables.keys(): # Checking if NAFF file was given
            if tableCode in self.rateTables['NAFF'].keys(): # Checking if table exists in NAFF
                return pd.DataFrame(data=self.rateTables['NAFF'][tableCode][1:], index=None, columns=self.rateTables['NAFF'][tableCode][0])
        if 'NICOF' in self.rateTables.keys(): # Checking if NICOF file was given
            if tableCode in self.rateTables['NICOF'].keys(): # Checking if tabl exists in NICOF
                return pd.DataFrame(data=self.rateTables['NICOF'][tableCode][1:], index=None, columns=self.rateTables['NICOF'][tableCode][0])
        return pd.DataFrame(data=self.rateTables['CW'][tableCode][1:], index=None, columns=self.rateTables['CW'][tableCode][0]) # Returning the country-wide table if it wasn't found in any other company

    # Builds the base rates table for the given company
    # Returns a dataframe
    def buildBaseRates(self, company):
        buildingBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Building_Base_Rates'][0])
        bppBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_BPP_Base_Rates'][0])
        liabilityBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Liability_Base_Rates'][0])
        businessIncomeBaseRates = pd.DataFrame(data=self.rateTables[company]['BP7_Peril_Business_Income_Base_Rate'][1:], index=None, columns=self.rateTables[company]['BP7_Peril_Business_Income_Base_Rate'][0])
        filteredBuildingBaseRates = buildingBaseRates.query(f'Class_Code_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4" & `Peril TypeCode` != "fire4"').filter(items=['Peril TypeCode', 'BuildingBaseRate'])
        filteredBPPBaseRates = bppBaseRates.query(f'Class_Code_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4" & `Peril TypeCode` != "fire4"').filter(items=['Peril TypeCode', 'BPPBaseRate'])
        filteredLiabilityBaseRates = liabilityBaseRates.query(f'ClassCode_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils} & OccupanyType != "tenant"'). \
                pivot(index='Peril TypeCode', columns='OccupanyType', values='LiabilityFactor').reset_index().rename_axis(None, axis=1)
        filteredBusinessIncomeBaseRates = businessIncomeBaseRates.query(f'`Peril TypeCode` in {self.perils} & `Peril TypeCode` != "cat4" & `Peril TypeCode` != "water2" & `Peril TypeCode` != "fire1" & `Peril TypeCode` != "allother1" & `Peril TypeCode` != "theft1" & `Peril TypeCode` != "water1" & OccupanyType != "tenant"'). \
                    pivot(index='Peril TypeCode', columns='OccupanyType', values='BusinessIncomeBaseRate').reset_index().rename_axis(None, axis=1). \
                    rename(columns={"buildingOwnerLessorsrisk": "Bus Inc Lessor's Risk", "buildingOwnerOccupant": "Bus Inc Occupant"})
        baseRates = pd.merge(filteredBuildingBaseRates, filteredBPPBaseRates, how='outer', on='Peril TypeCode')
        mergedBaseRates = pd.merge(baseRates, filteredBusinessIncomeBaseRates, how='outer', on='Peril TypeCode')
        finalBaseRates = pd.merge(mergedBaseRates, filteredLiabilityBaseRates, how='outer', on='Peril TypeCode')
        return finalBaseRates.replace({'Peril TypeCode': self.perilsConversions}).rename(columns={"Peril TypeCode": "Peril", "BuildingBaseRate": "Building", 
                "BPPBaseRate": "BPP", "buildingOwnerLessorsrisk": "Liability Lessor's Risk", "buildingOwnerOccupant": "Liability Occupant"}).sort_values(by='Peril')

    # Builds the territory multiplier table for the given coverage (either building, bpp, liability or business income)
    # Returns a dataframe
    #def buildTerritoryMultiplier(self, coverage):
    #    territorialFactor = self.buildDataFrame("BP7_Peril_TerritorialFactor")
    #    filteredTerritorialFactor = territorialFactor.query(f'Class_Code_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}).rename(columns={'TerritoryCode': 'Territory'})
    #    if coverage.casefold() == 'building': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BldgTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'bpp': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BPPTerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'business income': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='BITerritoryFactor').reset_index('Territory')
    #    elif coverage.casefold() == 'liability': # Case-insensitive comparison
    #        return filteredTerritorialFactor.pivot(index='Territory', columns='Peril TypeCode', values='LiabilityTerritoryFactor').reset_index('Territory')
    
    # Builds the construction type table for the given coverage (either building or bpp)
    # Returns a dataframe
    def buildConstructionType(self, coverage):
        constructionType = self.buildDataFrame("BP7 Peril Construction_Type")
        filteredConstructionType = constructionType.query(f'Class_Code_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                rename(columns={'ConstructionClassDisplay Name': 'Construction'})
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BldgConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredConstructionType.pivot(index='Construction', columns='Peril TypeCode', values='BPPConstructionClassFactor').reset_index('Construction').drop('L-Products', axis=1)

    # Builds the theft options table
    # Returns a dataframe
    def buildTheftOptions(self):
        theftOptions = self.buildDataFrame("BP7_Peril_BPP_Theft_Options_Factor")
        filteredTheftOptions = theftOptions.query(f'Class_Code_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils} & `Theft Option` != "Full Theft"'). \
                replace({'Peril TypeCode': self.perilsConversions})
        filteredTheftOptions = filteredTheftOptions.drop(filteredTheftOptions[filteredTheftOptions['Peril TypeCode'] == 'L-Products'].index)
        return filteredTheftOptions.pivot(index='Peril TypeCode', columns='Theft Option', values='BPP Theft Options Factor').reset_index('Peril TypeCode'). \
                rename(columns={'Peril TypeCode': 'Peril', 'Excluded Theft': 'Excluded', 'Limited Theft': 'Limited'})

    # Builds the year built modifier table for the given coverage (either building, bpp or business income)
    # Returns a dataframe
    def buildYearBuiltModifier(self, coverage):
        yearBuiltModifier = pd.DataFrame()
        if coverage.casefold() == 'building': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_Building_Year_Built_Modifier")
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BPP_Year_Built_Modifier")
        elif coverage.casefold() == 'business income':
            yearBuiltModifier = self.buildDataFrame("BP7 Peril_BI_Year_Built_Modifier")
        filteredYearBuiltModifier = yearBuiltModifier.query(f'Class_Code_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils}').replace({'Peril TypeCode': self.perilsConversions}). \
                fillna({'Year_Built_Max': 0}).astype({'Year_Built_Min': 'int64', 'Year_Built_Max': 'int64'}).astype({'Year_Built_Min': 'string', 'Year_Built_Max': 'string'}) # Converting to int first to get rid of decimal places
        filteredYearBuiltModifier['Year Built Range'] = np.where(filteredYearBuiltModifier['Year_Built_Max'] == '0', 
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + '+',
                                                                 filteredYearBuiltModifier['Year_Built_Min'] + ' - ' + filteredYearBuiltModifier['Year_Built_Max'])
        if coverage.casefold() == 'building': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='Bldg_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'bpp': # Case-insensitive comparison
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BPP_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)
        elif coverage.casefold() == 'business income':
            return filteredYearBuiltModifier.pivot(index='Year Built Range', columns='Peril TypeCode', values='BI_Year_Built_Factor').reset_index('Year Built Range').drop('L-Products', axis=1)

    # Builds the equipment breakdown base rate table
    # Returns a dataframe    
    def buildEBBaseRate(self):
        ebBaseRate = self.buildDataFrame("BP7_EBBaseRate")
        return ebBaseRate.query(f'Class_Code_Min == {self.foodProgramCode}').rename(columns={'BaseRate': 'Rate'}).filter(items=['Rate'])

    # Builds the property damage liability deductible factor table
    # Returns a dataframe
    def buildPDDeductibleAmount(self):
        pdDeductibleAmount = self.buildDataFrame("BP7_Peril_Property_Damage_Liability_Factor")
        return pdDeductibleAmount.query(f'ClassCode_Min == {self.foodProgramCode}').rename(columns={'PDDeductibleAmount': 'P.D. Deductible Amount', 'PDDeductibleFactor': 'Factor'}). \
                replace({'P.D. Deductible Amount': {'NoDeductible': '0'}}).astype({'P.D. Deductible Amount': 'int64'}).sort_values(by=['P.D. Deductible Amount']).replace({'Peril TypeCode': self.perilsConversions}). \
                pivot(index='P.D. Deductible Amount', columns='Peril TypeCode', values='Factor').reset_index('P.D. Deductible Amount').replace({'P.D. Deductible Amount': {0: 'No Deductible'}})

    # Builds the liability size of risk modifier table
    # Returns a dataframe
    def buildLiabilitySizeRisk(self):
        liabilitySizeRisk = self.buildDataFrame("BP7_Peril_Liability_Factor_Receipts_Limit")
        filteredLiabilitySizeRisk = liabilitySizeRisk.query(f'ClassCode_Min == {self.foodProgramCode} & `Peril TypeCode` in {self.perils} & AutoServType == "N/A" & FoodServType == "N/A"').rename(columns={'ReceiptMin': 'Min', 'ReceiptMax': 'Max', 'LiabilityFactor': 'Factor'}).replace({'Peril TypeCode': self.perilsConversions})
        return filteredLiabilitySizeRisk.pivot(index=['Min', 'Max'], columns='Peril TypeCode', values='Factor').reset_index(['Min', 'Max']).fillna({'Max': 'and over'})

    # Builds the liability limit factor table
    # Returns a dataframe
    def buildLiabilityLimitFactor(self):
        liabilityLimitFactor = self.buildDataFrame("BP7_Peril_ILF_Factor")
        return liabilityLimitFactor.query(f'ClassCode_Min == {self.foodProgramCode} & `Peril TypeCode` == "liability1"').filter(items=['LiabilityLimit', 'LiabilityFactor']). \
                rename(columns={'LiabilityLimit': 'Liability Limit of Insurance', 'LiabilityFactor': 'Factor'}).astype({'Liability Limit of Insurance' : 'int32'})

    # Builds the endorsement charge table
    # Returns a dataframe
    # def buildEndorsementCharge(self):
    #    endorsementCharge = self.buildDataFrame("BP7_PlusEndorsementCharge")
    #    return endorsementCharge.query(f'ClassCodeMIn == {self.foodProgramCode}').filter(items=['PlusEndorsementCharge']).rename(columns={'PlusEndorsementCharge': 'Base premium for each Food Service premises'})

    # Builds the table for Food Services Specialized Endorsement
    def buildFSSplzdEndo(self):
        data = {"Base premium for each Food Service Premises":["$400.00"]}
        data = pd.DataFrame(data)
        return data

    # Builds the optional increased limits – spoilage from power outage
    # Returns a dataframe
    def buildSpoilagePowerOutage(self):
        optionalCoverageRates = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        spoilagePowerOutage = optionalCoverageRates.query(f'CoverageName == "SpoilagePowerOutageIncreasedLimits"').rename(columns={'CoverageName': 'Coverage', 'BaseRate': 'Optional Increased Limit'})
        return spoilagePowerOutage.replace({'Coverage': {'SpoilagePowerOutageIncreasedLimits': 'Spoilage Power Outage'}})

    # Builds the optional increased limits - food contamination - loss of income table
    # Returns a dataframe
    def buildFoodContamination(self):
        miscBaseRates = self.buildDataFrame("BP7_Miscellaneous_Base_Rates")
        miscFactors = self.buildDataFrame("BP7_Miscellaneous_Factors_Table")
        foodContamEstablishment = self.buildDataFrame("BP7_Food_Contamination_Establishment")
        foodContamFranchise = self.buildDataFrame("BP7_Food_Contamination_Franchise")
        filteredMiscBaseRates = miscBaseRates.query(f'BaseRateName == "FoodContaminationLossOfIncome"')
        filteredMiscFactors = miscFactors.query(f'FactorName == "Receipts"')
        filteredFoodContamEst = foodContamEstablishment.query(f'FoodServiceType == "FIN"')
        filteredFoodContamFranch = foodContamFranchise.query(f'FranchiseInd == "Yes"')
        foodContamination = pd.concat([filteredMiscBaseRates, filteredMiscFactors, filteredFoodContamEst, filteredFoodContamFranch], ignore_index=True)
        foodContamination['Service'] = np.where(foodContamination['BaseRateName'].notnull(),
                                                'Base',
                                                np.where(foodContamination['FoodServiceType'].notnull(),
                                                         'Fine Dining',
                                                         np.where(foodContamination['FranchiseInd'].notnull(),
                                                                  'Franchise',
                                                                  'Receipts')))
        foodContamination['Rate or Factor'] = np.where(foodContamination['BaseRateName'].notnull(),
                                                       foodContamination['BaseRate'],
                                                       np.where(foodContamination['FoodServiceType'].notnull(),
                                                                foodContamination['FoodContaminationEstablishmentFactor'],
                                                                np.where(foodContamination['FranchiseInd'].notnull(),
                                                                         foodContamination['FoodContainmentFranchiseFactor'],
                                                                         foodContamination['Factor'])))
        return foodContamination.filter(items=['Service', 'Rate or Factor'])

    # Builds the off premises valet parking table
    # Returns a dataframe
    def buildValetParking(self):
        optionalCoverageRates = self.buildDataFrame("BP7_Optional_Coverage_Base_Rates")
        offPremValetPark = optionalCoverageRates.query(f'CoverageName == "OffPremisesValetParking"').rename(columns={'CoverageName': 'Coverage', 'BaseRate': 'Rate'})
        return offPremValetPark.replace({'Coverage': {'OffPremisesValetParking': 'Off Premises Valet Parking'}})

    # Builds the franchise upgrage endorsement table for the given program
    # Returns a dataframe
    def buildFranchiseUpgradeEndorsement(self):
        franchiseUpgradeBase = self.buildDataFrame("BP7_Franchise_Upgrade_Base")
        miscMinMaxPrem = self.buildDataFrame("BP7_Miscellaneous_Minimum/Maximum_Premium")
        filteredFranchiseUpgrade = franchiseUpgradeBase.query(f'MinClassCode == {self.foodProgramCode}')
        filteredMiscMinMaxPrem = miscMinMaxPrem.query(f'CoverageType == "BP7Pol_FranchiseUpgradeEndorsementCov_Ext"')
        franchiseUpgradeEndorsement = pd.concat([filteredFranchiseUpgrade, filteredMiscMinMaxPrem], ignore_index=True)
        franchiseUpgradeEndorsement['Rate or Premium'] = np.where(franchiseUpgradeEndorsement['RateType'].isnull(), 
                                                                  'Minimum Premium',
                                                                  'Base Rate')
        franchiseUpgradeEndorsement['Per Building'] = np.where(franchiseUpgradeEndorsement['FranchiseUpgradeBase'].isnull(), 
                                                               franchiseUpgradeEndorsement['Premium'],
                                                               franchiseUpgradeEndorsement['FranchiseUpgradeBase'])
        return franchiseUpgradeEndorsement.filter(items=['Rate or Premium', 'Per Building'])

    # Converts the given pixels to inches
    # Returns a decimal
    # NOTE: This ratio will vary based on font settings
    def pixelsToInches(self, px):
        return px / float(7)

    # Applies manual formatting to the base rates worksheet
    def formatBaseRates(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(130)
            for row in range(4, ws.max_row + 1):
                cell = ws[char + str(row)]
                cell.number_format = '#,##0.0000' # 4 values after the decimal point for the base rates

    # Applies manual formatting to the territory multiplier worksheet
    #def formatTerritoryMultiplier(self, ws):
    #    ws.column_dimensions['A'].width = self.pixelsToInches(70)
    #    for col in range(2, ws.max_column + 1):
    #        char = get_column_letter(col) # Letter representing the current column
    #        ws.column_dimensions[char].width = self.pixelsToInches(54)  

    # Applies manual formatting to the construction factor worksheet
    def formatConstructionFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(138)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)

    # Applies manual formatting to the theft options worksheet
    def formatTheftOptions(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(82)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(80)

    # Applies manual formatting to the year built modifier worksheet
    def formatYearBuiltModifier(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(131)
        for col in range(2, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            ws.column_dimensions[char].width = self.pixelsToInches(53)
        for row in range(4, ws.max_row + 1):
            cell = ws['A' + str(row)]
            cell.number_format = self.noDecimalFormat

    # Applies manual formatting to the equipment breakdown base rate worksheet
    def formatEBBaseRate(self, ws):
        ws['A4'].number_format = '$#,##0.00'

    # Applies manual formatting to the property damage deductible worksheet
    def formatPropertyDamageDeductible(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(187)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the liability size of risk modifier worksheet
    def formatLiabilitySizeRisk(self, ws, boldFont):
        ws.insert_rows(3)
        ws['A3'] = 'Receipts Range'
        for cell in ws['3:3']:
            cell.border = Border(left=Side(border_style='thin', color='C1C1C1'), 
                                right=Side(border_style='thin', color='C1C1C1'), 
                                top=Side(border_style='thin', color='C1C1C1'), 
                                bottom=Side(border_style='thin', color='C1C1C1'))
            cell.font = boldFont
            cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=True)
        ws.merge_cells('A3:B3')
        ws.merge_cells('C3:' + get_column_letter(ws.max_column) + '3')
        ws.print_title_rows = '1:4'
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col) # Letter representing the current column
            if col < 3:
                ws.column_dimensions[char].width = self.pixelsToInches(95)
                for row in range(5, ws.max_row + 1):
                    cell = ws[char + str(row)]
                    cell.number_format = self.noDecimalFormat
            else:
                ws.column_dimensions[char].width = self.pixelsToInches(120)

    # Applies manual formatting to the liability limit factor worksheet
    def formatLiabilityLimitFactor(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(205)
        ws.column_dimensions['B'].width = self.pixelsToInches(54)
        for row in range(4, ws.max_row + 1):
            ws['A' + str(row)].number_format = self.noDecimalFormat

    # Applies manual formatting to the endorsement charge worksheet
    def formatEndorsementCharge(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(350)
        ws['A4'].number_format = '$#,##0.00'
    
    # Applies manual formatting to the optional increased limits – spoilage from power outage worksheet
    def formatSpoilagePower(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(160)
        ws.column_dimensions['B'].width = self.pixelsToInches(175)
    
    # Applies manual formatting to the optional increased limits – food contamination – loss of income worksheet
    def formatFoodContamination(self, ws):
        ws.column_dimensions['B'].width = self.pixelsToInches(135)
    
    def formatValetParking(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(185)
        ws['B4'].number_format = '$#,##0.00'

    # Applies manual formatting to the franchise endorsement worksheet
    def formatFranchiseEndorsement(self, ws):
        ws.column_dimensions['A'].width = self.pixelsToInches(145)
        ws.column_dimensions['B'].width = self.pixelsToInches(100)
        for row in range(4, ws.max_row + 1):
            cell = ws['B' + str(row)]
            cell.number_format = '$#,##0.00'   

    # Sets up the Food Service Excel file using the Excel class
    # A separate worksheet is generated for each table, and most worksheets are manually formatted afterwards
    # Returns the Excel file
    def buildFoodPage(self):
        companies = []
        for company in self.rateTables.keys():
            if company == 'CW': # country-wide is not a company, so ignoring it
                continue
            companies.append(company)

        FoodService = ExcelSettings.Excel(state=self.state, programName='Food Service', nEffective=self.nEffective, rEffective=self.rEffective, companyList=companies)

        fontName = FoodService.getFontName()
        fontSize = FoodService.getFontSize()

        if 'NACO' in self.rateTables.keys():
            FoodService.generateWorksheet('BRNACO', 'FS Table 3.B.1. NW Assurance State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NAFF' in self.rateTables.keys():
            FoodService.generateWorksheet('BRNAFF', 'FS Table 3.B.1. NW Affinity State Base Rates', self.buildBaseRates('NGIC'), False, True)
        if 'NGIC' in self.rateTables.keys():
            FoodService.generateWorksheet('BRNGIC', 'FS Table 3.B.1. NW General Insurance Company', self.buildBaseRates('NGIC'), False, True)
        if 'NICOF' in self.rateTables.keys():
            FoodService.generateWorksheet('BRNICOF', 'FS Table 3.B.1. NICOF State Base Rates', self.buildBaseRates('NGIC'), False, True)
        #FoodService.generateWorksheet('TRBG', 'FS Table 3.C.1.a. State Territory Multiplier - Building', self.buildTerritoryMultiplier('Building'), False, True)
        #FoodService.generateWorksheet('TRPP', 'FS Table 3.C.1.a. State Territory Multiplier - BPP', self.buildTerritoryMultiplier('BPP'), False, True)
        #FoodService.generateWorksheet('TRLB', 'FS Table 3.C.1.a. State Territory Multiplier - Liability', self.buildTerritoryMultiplier('Liability'), False, True)
        FoodService.generateWorksheet('CBG', 'FS Table 3.C.2.c. Construction Factor - Building', self.buildConstructionType('Building'), False, True)
        FoodService.generateWorksheet('CPP', 'FS Table 3.C.2.c. Construction Factor - BPP and Bus Inc', self.buildConstructionType('BPP'), False, True)
        FoodService.generateWorksheet('ET', 'FS Table 3.C.2.m. Exclude Theft Factor', self.buildTheftOptions(), False, True)
        FoodService.generateWorksheet('YBBG', 'FS Table 3.C.2.p. Year Built Modifier - Building', self.buildYearBuiltModifier('Building'), False, True)
        FoodService.generateWorksheet('YBPP', 'FS Table 3.C.2.p. Year Built Modifier - BPP', self.buildYearBuiltModifier('BPP'), False, True)
        FoodService.generateWorksheet('YBBI', 'FS Table 3.C.2.p. Year Built Modifier - Bus Inc', self.buildYearBuiltModifier('Business Income'), False, True)
        FoodService.generateWorksheet('EBB', 'FS Table 3.C.3.a. EB Base Rate', self.buildEBBaseRate(), False, True)
        FoodService.generateWorksheet('PDLD', 'FS Table 3.C.4.b. Property Damage Liability Deductible Factor', self.buildPDDeductibleAmount(), False, True)
        FoodService.generateWorksheet('LS', 'FS Table 3.C.4.d. Liability Size of Risk Modifier', self.buildLiabilitySizeRisk(), False, True)
        FoodService.generateWorksheet('LL', 'FS Table 3.C.4.e. Liability Limit Factor', self.buildLiabilityLimitFactor(), False, True)
        # FoodService.generateWorksheet('PLUS', 'FS Table 4.A.1. Food Service PLUS Endorsement', self.buildEndorsementCharge(), False, True)
        FoodService.generateWorksheet('FSS', 'FS Table 4.A.1. Food Service Specialized Endorsement', self.buildFSSplzdEndo(), False, True)
        FoodService.generateWorksheet('SPO', 'FS Table 4.A.2. Optional Increased Limits – Spoilage From Power Outage', self.buildSpoilagePowerOutage(), False, True)
        FoodService.generateWorksheet('C', 'FS Table 4.A.3. Optional Increased Limits – Food Contamination – Loss of Income', self.buildFoodContamination(), False, True)
        FoodService.generateWorksheet('VAL', 'FS Table 4.B. Off Premises Valet Parking', self.buildValetParking(), False, True)
        FoodService.generateWorksheet('FR', 'FS Table 4.C. Franchise Upgrade Endorsement', self.buildFranchiseUpgradeEndorsement(), False, True)

        FoodService.createIndex()
        FoodPages = FoodService.getWB()

        if 'NACO' in self.rateTables.keys():
            self.formatBaseRates(FoodPages['BRNACO'])
        if 'NAFF' in self.rateTables.keys():
            self.formatBaseRates(FoodPages['BRNAFF'])
        if 'NGIC' in self.rateTables.keys():
            self.formatBaseRates(FoodPages['BRNGIC'])
        if 'NICOF' in self.rateTables.keys():
            self.formatBaseRates(FoodPages['BRNICOF'])
        #self.formatTerritoryMultiplier(FoodPages['TRBG'])
        #self.formatTerritoryMultiplier(FoodPages['TRPP']) 
        #self.formatTerritoryMultiplier(FoodPages['TRLB'])
        self.formatConstructionFactor(FoodPages['CBG'])  
        self.formatConstructionFactor(FoodPages['CPP'])  
        self.formatTheftOptions(FoodPages['ET'])
        self.formatYearBuiltModifier(FoodPages['YBBG'])
        self.formatYearBuiltModifier(FoodPages['YBPP'])
        self.formatYearBuiltModifier(FoodPages['YBBI'])  
        self.formatEBBaseRate(FoodPages['EBB'])
        self.formatPropertyDamageDeductible(FoodPages['PDLD'])
        self.formatLiabilitySizeRisk(FoodPages['LS'], Font(name=fontName, size=fontSize, bold=True))
        self.formatLiabilityLimitFactor(FoodPages['LL'])
        self.formatEndorsementCharge(FoodPages['FSS'])
        self.formatSpoilagePower(FoodPages['SPO'])
        self.formatFoodContamination(FoodPages['C'])
        self.formatValetParking(FoodPages['VAL'])
        self.formatFranchiseEndorsement(FoodPages['FR'])

        return FoodPages
    