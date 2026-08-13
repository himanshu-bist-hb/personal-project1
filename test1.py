# <editor-fold desc="Libraries">
import datetime
import warnings
import time
warnings.simplefilter(action='ignore')
import sqlite3
import numpy as np
import pandas as pd
import os.path
import xlwings as xw
import xlsxwriter as xlw
import tabulate
import math
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, DEFAULT_FONT
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.pagebreak import Break

# </editor-fold>

# <editor-fold desc="Functions and Preset Formatting">
warnings.simplefilter(action='ignore')

def round_half_up(n, decimals=0):
    multiplier = 10 ** decimals
    if n > 0:
        return np.floor(((n+0.000000001) * multiplier) + 0.5) / multiplier
    elif n == 0:
        return 0
    else:
        return np.floor(((n-0.000000001) * multiplier) + 0.5) / multiplier

def pixelsToInches(px):
    return px / float(72)

def splitdf(inputdata, splitnum):
    import math
    splitindex = math.ceil(inputdata.shape[0]/splitnum)
    splitlist = list()
    for i in range(1, splitnum+1):
        splitlist.append(inputdata.iloc[(i-1)*splitindex:i*splitindex].reset_index(drop=True))
    return pd.concat(splitlist, axis=1)

def formatwkstSM(wkstname, titlerows, A1title, A2title, dfname, statename, stabb, effdate):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.styles.borders import Border, Side

    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    headerFont = headerFontName + ', Bold'
    footerFont = footerFontName + ', Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'

    wkstname['A1'] = A1title
    wkstname['A2'] = A2title
    cella2 = wkstname['A2']
    cella2.alignment = Alignment(horizontal='left', vertical='center')

    for r in dataframe_to_rows(dfname, False, True):
        if False & len(list(r))==1: 
            wkstname['B4'] = list(r)[0] 
            continue
        wkstname.append(r)

    wkstname.page_setup.orientation = 'portrait' 
    wkstname.page_setup.firstPageNumber = 1 
    wkstname.page_setup.useFirstPageNumber = True
    wkstname.sheet_view.showGridLines = False 
    wkstname.print_title_rows = '1:' + titlerows 
    wkstname.page_margins.left = leftMargin
    wkstname.page_margins.right = rightMargin
    wkstname.page_margins.top = topMargin
    wkstname.page_margins.bottom = bottomMargin
    wkstname.page_margins.header = headerMargin
    wkstname.page_margins.footer = footerMargin
    wkstname.print_options.horizontalCentered = True

    # Left Header
    if statename == 'Florida':
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    wkstname.oddHeader.left.font = headerFont
    wkstname.oddHeader.left.size = headerFontSize

    # Center header
    wkstname.oddHeader.center.text = "\n\n" + statename + " Rate Pages"
    wkstname.oddHeader.center.size = headerFontSize
    wkstname.oddHeader.center.font = headerFont

    # Right header
    wkstname.oddHeader.right.text = "Effective Date: " + effdate
    wkstname.oddHeader.right.size = headerFontSize
    wkstname.oddHeader.right.font = headerFont

    # Left footer
    wkstname.oddFooter.left.size = footerFontSize
    wkstname.oddFooter.left.font = footerFont

    # Center footer
    wkstname.oddFooter.center.text = stabb + " - SRP &[Tab] - &P"
    wkstname.oddFooter.center.size = footerFontSize
    wkstname.oddFooter.center.font = footerFont

    if NICOFCondition == True:
        wkstname.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFCondition == False and NAFFCondition == True:
        wkstname.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFCondition == False and NAFFCondition == False and NACOCondition == True:
        wkstname.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        wkstname.oddFooter.left.text = "\nNationwide General Insurance Company"

def formatwkstMM(wkstname, titlerows, A1title, A2title, dfname, statename, stabb, effdate):
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.styles.borders import Border, Side

    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    headerFont = headerFontName + ', Bold'
    footerFont = footerFontName + ', Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'

    wkstname['A1'] = A1title
    wkstname['A2'] = A2title
    cella2 = wkstname['A2']
    cella2.alignment = Alignment(horizontal='left', vertical='center')

    for r in dataframe_to_rows(dfname, False, True):
        if False & len(list(r))==1: 
            wkstname['B4'] = list(r)[0]
            continue
        wkstname.append(r)

    wkstname.page_setup.orientation = 'portrait' 
    wkstname.page_setup.firstPageNumber = 1
    wkstname.page_setup.useFirstPageNumber = True
    wkstname.sheet_view.showGridLines = False
    wkstname.print_title_rows = '1:' + titlerows
    wkstname.page_margins.left = leftMargin
    wkstname.page_margins.right = rightMargin
    wkstname.page_margins.top = topMargin
    wkstname.page_margins.bottom = bottomMargin
    wkstname.page_margins.header = headerMargin
    wkstname.page_margins.footer = footerMargin
    wkstname.print_options.horizontalCentered = True

    # Left Header
    if statename == 'Florida':
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        wkstname.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"
    wkstname.oddHeader.left.size = headerFontSize
    wkstname.oddHeader.left.font = headerFont

    # Center header
    wkstname.oddHeader.center.text = "\n\n" + statename + " Rate Pages"
    wkstname.oddHeader.center.size = headerFontSize
    wkstname.oddHeader.center.font = headerFont

    # Right header
    wkstname.oddHeader.right.text = "Effective Date: " + effdate
    wkstname.oddHeader.right.size = headerFontSize
    wkstname.oddHeader.right.font = headerFont

    if stabb != "WI":
        wkstname.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company \nNationwide Property & Casualty Insurance Company \nAllied Insurance Company of America"
    else:
        wkstname.oddFooter.left.text = "\nNationwide Insurance Company of America \nNationwide Mutual Insurance Company"

    # Left footer
    wkstname.oddFooter.left.size = footerFontSize
    wkstname.oddFooter.left.font = footerFont

    # Center footer
    wkstname.oddFooter.center.text = stabb + " - SRP &[Tab] - &P"
    wkstname.oddFooter.center.size = footerFontSize
    wkstname.oddFooter.center.font = footerFont

def borderfnct(wkstname):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.borders import Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.pagebreak import Break
    import math

    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    headerFont = headerFontName + ', Bold'
    footerFont = footerFontName + ', Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'

    for row in range(1, wkstname.max_row + 1):
        for col in range(1, wkstname.max_column + 1):
            char = get_column_letter(col) 
            cell = wkstname[char + str(row)]
            wkstname.column_dimensions[char].bestFit = True 

            if row > 2 and cell.value != '': 
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))

            if row < 2: 
                cell.font = fontBoldUnderline
            elif col >= 1:
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

            if col == 1 and row == 2:
                cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)

    for col in wkstname.iter_cols(min_row=3):
        max_length = 0
        column = col[0].column_letter 
        for cell in col:
            try: 
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        wkstname.column_dimensions[column].width = adjusted_width

    wkstname.print_area = 'A1:' + get_column_letter(wkstname.max_column) + str(wkstname.max_row+1)
    wkstname.page_setup.fitToPage = True

    if wkstname.max_row > 30:
        tablefac = math.ceil((wkstname.max_row + 1) / 50)
        for rnum in range(1, tablefac):
            row_num = rnum * 50
            page_break = Break(id=row_num)
            wkstname.row_breaks.append(page_break)
        wkstname.page_setup.fitToPage = True
        wkstname.page_setup.fitToWidth = tablefac
        wkstname.page_setup.fitToHeight = tablefac

def addlformat(wkstname):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, Font, DEFAULT_FONT
    wkstname['A1'].alignment = Alignment(horizontal='center')
    wkstname['A2'].alignment = Alignment(horizontal='center')
    wkstname.insert_cols(0,4)
    wkstname.column_dimensions['A'].width = 9.1
    wkstname.column_dimensions['B'].width = 9.1
    wkstname.column_dimensions['C'].width = 9.1
    wkstname.column_dimensions['D'].width = 9.1
    wkstname['A3'] = ' * '
    wkstname['A3'].font = Font(color="FFFFFF")
    wkstname.print_area = 'A1:' + get_column_letter(wkstname.max_column + 4) + str(wkstname.max_row + 1)
    wkstname.page_setup.fitToPage = True

print('Creating Rate Pages')

# </editor-fold>

# <editor-fold desc="Company Info">
start_time = time.time()
pd.set_option('display.max_columns', None)
pd.options.display.width = None
pd.options.mode.chained_assignment = None

global NGICCondition
global MMCondition
global NACOCondition
global NAFFCondition
global NICOFCondition

try: NGICRatebook
except:
    NGICRatebook = "Not found"
    NGICCondition = False
else:
    if NGICRatebook == '':
        NGICRatebook = "Not found"
        NGICCondition = False
    else:
        NGICRatebook = pd.ExcelFile(NGICRatebook)
        NGICCondition = True

try: MMRatebook
except:
    MMRatebook = "Not found"
    MMCondition = False
else:
    if MMRatebook == '':
        MMRatebook = "Not found"
        MMCondition = False
    else:
        MMRatebook = pd.ExcelFile(MMRatebook)
        MMCondition = True

try: NACORatebook
except:
    NACORatebook = "Not found"
    NACOCondition = False
else:
    if NACORatebook == '':
        NACORatebook = "Not found"
        NACOCondition = False
    else:
        NACORatebook = pd.ExcelFile(NACORatebook)
        NACOCondition = True

try: NAFFRatebook
except:
    NAFFRatebook = "Not found"
    NAFFCondition = False
else:
    if NAFFRatebook == '':
        NAFFRatebook = "Not found"
        NAFFCondition = False
    else:
        NAFFRatebook = pd.ExcelFile(NAFFRatebook)
        NAFFCondition = True

try: NICOFRatebook
except:
    NICOFRatebook = "Not found"
    NICOFCondition = False
else:
    if NICOFRatebook == '':
        NICOFRatebook = "Not found"
        NICOFCondition = False
    else:
        NICOFRatebook = pd.ExcelFile(NICOFRatebook)
        NICOFCondition = True

if NGICRatebook != "Not found":
    NGICwb = load_workbook(NGICRatebook, read_only=True)
    RateBookDetails = pd.read_excel(NGICRatebook, sheet_name='Rate Book Details')
else:
    RateBookDetails = pd.read_excel(MMRatebook, sheet_name='Rate Book Details')

State = RateBookDetails.iloc[3,4]
EffectiveDate = RateBookDetails.iloc[7,4]
EffectiveDate = datetime.date.strftime(EffectiveDate, "%m-%d-%y")

try: TerrAdjRatebook
except:
    TerrAdjRatebook = "Not found"
else:
    if TerrAdjRatebook == '':
        TerrAdjRatebook = "Not found"
    else:
        TerrAdjRatebook = pd.ExcelFile(TerrAdjRatebook)

if State == "Alabama":
    StateAbb = "AL"
if State == "Alaska":
    StateAbb = "AK"
if State == "Arizona":
    StateAbb = "AZ"
if State == "Arkansas":
    StateAbb = "AR"
if State == "California":
    StateAbb = "CA"
if State == "Colorado":
    StateAbb = "CO"
if State == "Connecticut":
    StateAbb = "CT"
if State == "Delaware":
    StateAbb = "DE"
if State == "District of Columbia":
    StateAbb = "DC"
if State == "Florida":
    StateAbb = "FL"
if State == "Georgia":
    StateAbb = "GA"
if State == "Hawaii":
    StateAbb = "HI"
if State == "Idaho":
    StateAbb = "ID"
if State == "Illinois":
    StateAbb = "IL"
if State == "Indiana":
    StateAbb = "IN"
if State == "Iowa":
    StateAbb = "IA"
if State == "Kansas":
    StateAbb = "KS"
if State == "Kentucky":
    StateAbb = "KY"
if State == "Louisiana":
    StateAbb = "LA"
if State == "Maine":
    StateAbb = "ME"
if State == "Maryland":
    StateAbb = "MD"
if State == "Massachusetts":
    StateAbb = "MA"
if State == "Michigan":
    StateAbb = "MI"
if State == "Minnesota":
    StateAbb = "MN"
if State == "Mississippi":
    StateAbb = "MS"
if State == "Missouri":
    StateAbb = "MO"
if State == "Montana":
    StateAbb = "MT"
if State == "Nebraska":
    StateAbb = "NE"
if State == "Nevada":
    StateAbb = "NV"
if State == "New Hampshire":
    StateAbb = "NH"
if State == "New Jersey":
    StateAbb = "NJ"
if State == "New Mexico":
    StateAbb = "NM"
if State == "New York":
    StateAbb = "NY"
if State == "North Carolina":
    StateAbb = "NC"
if State == "North Dakota":
    StateAbb = "ND"
if State == "Ohio":
    StateAbb = "OH"
if State == "Oklahoma":
    StateAbb = "OK"
if State == "Oregon":
    StateAbb = "OR"
if State == "Pennsylvania":
    StateAbb = "PA"
if State == "Rhode Island":
    StateAbb = "RI"
if State == "South Carolina":
    StateAbb = "SC"
if State == "South Dakota":
    StateAbb = "SD"
if State == "Tennessee":
    StateAbb = "TN"
if State == "Texas":
    StateAbb = "TX"
if State == "Utah":
    StateAbb = "UT"
if State == "Vermont":
    StateAbb = "VT"
if State == "Virginia":
    StateAbb = "VA"
if State == "Washington":
    StateAbb = "WA"
if State == "West Virginia":
    StateAbb = "WV"
if State == "Wisconsin":
    StateAbb = "WI"
if State == "Wyoming":
    StateAbb = "WY"

# try: CWRatebook
# except:
#     CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025.xlsx')
# else:
#     if CWRatebook == '':
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025.xlsx')
#     elif StateAbb in ['VT', 'KY']:
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025 (VT, KY).xlsx')
#     elif StateAbb == 'NY':
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025 (NY).xlsx')
#     elif StateAbb in ['UT', 'MI', 'DE', 'TN', 'AL']:
#         CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025 (UT, MI, DE, TN, AL).xlsx')
#     else:
#         CWRatebook = pd.ExcelFile(CWRatebook)
# CWRatebook = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Jiang\\Rate Page Template\\CP CW RB 2025.xlsx')
# </editor-fold>

# <editor-fold desc="Create LCM Dataframes">
if NAFFRatebook != "Not found":
    NAFFLCM = pd.read_excel(NAFFRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NAFFLCM = NAFFLCM.iloc[0, 1]
    NAFFCRIMELCM = pd.read_excel(NAFFRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NAFFCRIMELCM = NAFFCRIMELCM.iloc[0, 1]

if NACORatebook != "Not found":
    NACOLCM = pd.read_excel(NACORatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NACOLCM = NACOLCM.iloc[0, 1]
    NACOCRIMELCM = pd.read_excel(NACORatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NACOCRIMELCM = NACOCRIMELCM.iloc[0, 1]

if NICOFRatebook != "Not found":
    NICOFLCM = pd.read_excel(NICOFRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NICOFLCM = NICOFLCM.iloc[0, 1]
    NICOFCRIMELCM = pd.read_excel(NICOFRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NICOFCRIMELCM = NICOFCRIMELCM.iloc[0, 1]

if NGICRatebook != "Not found":
    NGICLCM = pd.read_excel(NGICRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NGICLCM = NGICLCM.iloc[0, 1]
    NGICCRIMELCM = pd.read_excel(NGICRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    NGICCRIMELCM = NGICCRIMELCM.iloc[0, 1]

if MMRatebook != "Not found":
    if StateAbb != "WI":
        MMLCM = pd.read_excel(MMRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCRIMELCM = pd.read_excel(MMRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCompDev = pd.read_excel(MMRatebook, sheet_name='Company Deviation Factor_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        
        NWMLCM = MMLCM.iloc[0, 1]
        NWMCRIMELCM = MMCRIMELCM.iloc[0, 1]
        
        AICOA = ['aicoa_ext']
        AICOA = pd.DataFrame(AICOA, columns=['UnderwritingCompanyCode'])
        AICOACompDev = pd.merge(MMCompDev, AICOA, on='UnderwritingCompanyCode', how='inner')
        AICOACompDev = AICOACompDev.iloc[0, 2]
        
        NICOA = ['nicoa_ext']
        NICOA = pd.DataFrame(NICOA, columns=['UnderwritingCompanyCode'])
        NICOACompDev = pd.merge(MMCompDev, NICOA, on='UnderwritingCompanyCode', how='inner')
        NICOACompDev = NICOACompDev.iloc[0, 2]
        
        NWPC = ['npcic_ext']
        NWPC = pd.DataFrame(NWPC, columns=['UnderwritingCompanyCode'])
        NWPCCompDev = pd.merge(MMCompDev, NWPC, on='UnderwritingCompanyCode', how='inner')
        NWPCCompDev = NWPCCompDev.iloc[0, 2]
        
        AICOALCM = round_half_up(NWMLCM * AICOACompDev, 3)
        NICOALCM = round_half_up(NWMLCM * NICOACompDev, 3)
        NWPCLCM = round_half_up(NWMLCM * NWPCCompDev, 3)
        
        AICOACRIMELCM = round_half_up(NWMCRIMELCM * AICOACompDev, 3)
        NICOACRIMELCM = round_half_up(NWMCRIMELCM * NICOACompDev, 3)
        NWPCCRIMELCM = round_half_up(NWMCRIMELCM * NWPCCompDev, 3)
    else:
        MMLCM = pd.read_excel(MMRatebook, sheet_name='LossCostMultiplier', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCRIMELCM = pd.read_excel(MMRatebook, sheet_name='CrimeLossCostMultiplier_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        MMCompDev = pd.read_excel(MMRatebook, sheet_name='Company Deviation Factor_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
        NWMLCM = MMLCM.iloc[0, 1]
        NWMCRIMELCM = MMCRIMELCM.iloc[0, 1]
        NICOA = ['nicoa_ext']
        NICOA = pd.DataFrame(NICOA, columns=['UnderwritingCompanyCode'])
        NICOACompDev = pd.merge(MMCompDev, NICOA, on='UnderwritingCompanyCode', how='inner')
        NICOACompDev = NICOACompDev.iloc[0, 2]
        NICOALCM = round_half_up(NWMLCM * NICOACompDev, 3)
        NICOACRIMELCM = round_half_up(NWMCRIMELCM * NICOACompDev, 3)

if NICOFRatebook != "Not found":
    SMLCMs = [['Nationwide Affinity Insurance Company of America', NAFFLCM],
              ['Nationwide Assurance Company', NACOLCM],
              ['Nationwide Insurance Company of Florida', NICOFLCM],
              ['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    
    SMCRIMELCMs = [['Nationwide Affinity Insurance Company of America', NAFFCRIMELCM],
                   ['Nationwide Assurance Company', NACOCRIMELCM],
                   ['Nationwide Insurance Company of Florida', NICOFCRIMELCM],
                   ['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])

if NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
    SMLCMs = [['Nationwide Affinity Insurance Company of America', NAFFLCM],
              ['Nationwide Assurance Company', NACOLCM],
              ['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    
    SMCRIMELCMs = [['Nationwide Affinity Insurance Company of America', NAFFCRIMELCM],
                   ['Nationwide Assurance Company', NACOCRIMELCM],
                   ['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])

if NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
    SMLCMs = [['Nationwide Assurance Company', NACOLCM],
              ['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    
    SMCRIMELCMs = [['Nationwide Assurance Company', NACOCRIMELCM],
                   ['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])

if NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook == "Not found" and NGICRatebook != "Not found":
    SMLCMs = [['Nationwide General Insurance Company', NGICLCM]]
    SMLCMTable = pd.DataFrame(SMLCMs, columns=['Company', 'LCM'])
    SMCRIMELCMs = [['Nationwide General Insurance Company', NGICCRIMELCM]]
    SMCRIMELCMTable = pd.DataFrame(SMCRIMELCMs, columns=['Company', 'LCM'])

if MMRatebook != "Not found":
    if StateAbb != "WI":
        MMLCMs = [['Nationwide Insurance Company of America', NICOALCM],
                  ['Nationwide Mutual Insurance Company', NWMLCM],
                  ['Nationwide Property & Casualty Insurance Company', NWPCLCM],
                  ['Allied Insurance Company of America', AICOALCM]]
        MMLCMTable = pd.DataFrame(MMLCMs, columns=['Company', 'LCM'])
        
        MMCRIMELCMs = [['Nationwide Insurance Company of America', NICOACRIMELCM],
                       ['Nationwide Mutual Insurance Company', NWMCRIMELCM],
                       ['Nationwide Property & Casualty Insurance Company', NWPCCRIMELCM],
                       ['Allied Insurance Company of America', AICOACRIMELCM]]
        MMCRIMELCMTable = pd.DataFrame(MMCRIMELCMs, columns=['Company', 'LCM'])
    else:
        MMLCMs = [['Nationwide Insurance Company of America', NICOALCM],
                  ['Nationwide Mutual Insurance Company', NWMLCM]]
        MMLCMTable = pd.DataFrame(MMLCMs, columns=['Company', 'LCM'])
        
        MMCRIMELCMs = [['Nationwide Insurance Company of America', NICOACRIMELCM],
                       ['Nationwide Mutual Insurance Company', NWMCRIMELCM]]
        MMCRIMELCMTable = pd.DataFrame(MMCRIMELCMs, columns=['Company', 'LCM'])

# </editor-fold>

# <editor-fold desc="Create Territory Adjustment Dataframes">
if TerrAdjRatebook != "Not found":
    Terrwb = pd.ExcelFile(TerrAdjRatebook)
    bgi = (pd.read_excel(Terrwb, sheet_name='BGI', skiprows=[0, 1, 2, 3, 4, 5, 6], usecols=[0, 1], names=['Grid ID', 'BGI Factor'], engine='openpyxl')
           .sort_values('Grid ID', kind='mergesort').reset_index(drop=True))
    bgii = (pd.read_excel(Terrwb, sheet_name='BGII', skiprows=[0, 1, 2, 3, 4, 5, 6], usecols=[0, 1], names=['Grid ID', 'BGII Factor'], engine='openpyxl')
            .sort_values('Grid ID', kind='mergesort').reset_index(drop=True))
    scol = (pd.read_excel(Terrwb, sheet_name='SCOL', skiprows=[0, 1, 2, 3, 4, 5, 6], usecols=[0, 1], names=['Grid ID', 'SCOL Factor'], engine='openpyxl')
            .sort_values('Grid ID', kind='mergesort').reset_index(drop=True))
    
    GridFac = pd.concat([bgi[['Grid ID']], bgi[['BGI Factor']], bgii[['BGII Factor']], scol[['SCOL Factor']]], axis=1)
    
    TerrCoord = pd.ExcelFile('\\\\Urbdat01.allied.nwie.net\\Actuary\\Actshare\\Com\\Property\\Territory Defs (Independent)\\3.0\\Development\\R Shiny Interface\\RShiny Smoothed Relativities\\' + StateAbb + ' Smoothed Relativities.xlsx')
    Coords = pd.read_excel(TerrCoord, sheet_name='Sheet1', usecols=[1, 2, 3], names=['Grid ID', 'Longitude', 'Latitude']).round(7)
    
    GridFac = pd.merge(GridFac, Coords, how='inner', on='Grid ID')
    GridFac = GridFac[['Grid ID', 'Longitude', 'Latitude', 'BGI Factor', 'BGII Factor', 'SCOL Factor']]
    GridFac = GridFac.rename(columns={"Grid ID": "Territory"})
    GridFac = GridFac.drop_duplicates().reset_index(drop=True)
    
    GridLen = math.ceil(GridFac.shape[0] / 3)
    GridFac1 = GridFac.iloc[:GridLen].reset_index(drop=True)
    GridFac2 = GridFac.iloc[GridLen:2*GridLen].reset_index(drop=True)
    GridFac3 = GridFac.iloc[2*GridLen:].reset_index(drop=True)
    GridFacXL = pd.concat([GridFac1, GridFac2, GridFac3], axis=1).fillna(" ")
    GridFacXL.insert(6, '', '')
    GridFacXL.insert(13, '', '')
    print("territory set up")
else:
    pass

# </editor-fold>

# <editor-fold desc="Create PMF Dataframes">
if MMRatebook != "Not found":
    MMwb = load_workbook(MMRatebook, read_only=False)
    if 'PackageModifierFactorTable_Ext' in MMwb.sheetnames:
        MMPMF = pd.read_excel(MMRatebook, sheet_name='PackageModifierFactorTable_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    else:
        MMPMF = pd.read_excel(NGICRatebook, sheet_name='PackageModifierFactorTable_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    MMPMF = MMPMF.rename(columns={'PackageModifierCode': 'Package Code'})
    MMPMF = MMPMF.drop(labels=[0, 9], axis=0)
else:
    MMwb = Workbook()

if NGICRatebook != "Not found":
    SMPMF = pd.read_excel(NGICRatebook, sheet_name='PackageModifierFactorTable_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
    SMPMF = SMPMF.rename(columns={'PackageModifierCode': 'Package Code'})
    SMPMF = SMPMF.drop(labels=[0, 9], axis=0)

# </editor-fold>

# <editor-fold desc="Create Capping Dataframe">
if v5.get() == 1 and v4.get() == 1:
    SMMinCap = SMMinCapRange.get()
    SMMaxCap = SMMaxCapRange.get()
    
    if NICOFRatebook != "Not found":
        SMCapping = [['Nationwide Affinity Insurance Company of America', SMMinCap, SMMaxCap],
                     ['Nationwide Assurance Company', SMMinCap, SMMaxCap],
                     ['Nationwide Insurance Company of Florida', SMMinCap, SMMaxCap]]
        SMCapTable = pd.DataFrame(SMCapping, columns=['Company', 'Minimum', 'Maximum'])
    
    if NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        SMCapping = [['Nationwide Affinity Insurance Company of America', SMMinCap, SMMaxCap],
                     ['Nationwide Assurance Company', SMMinCap, SMMaxCap]]
        SMCapTable = pd.DataFrame(SMCapping, columns=['Company', 'Minimum', 'Maximum'])
        
    if NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        SMCapping = [['Nationwide Assurance Company', SMMinCap, SMMaxCap]]
        SMCapTable = pd.DataFrame(SMCapping, columns=['Company', 'Minimum', 'Maximum'])

if v10.get() == 1 and v11.get() == 1:
    MMMinCap = MMMinCapRange.get()
    MMMaxCap = MMMaxCapRange.get()
    
    MMCapping = [['Nationwide Insurance Company of America', MMMinCap, MMMaxCap],
                 ['Nationwide Mutual Insurance Company', MMMinCap, MMMaxCap],
                 ['Nationwide Property & Casualty Insurance Company', MMMinCap, MMMaxCap],
                 ['Allied Insurance Company of America', MMMinCap, MMMaxCap]]
    MMCapTable = pd.DataFrame(MMCapping, columns=['Company', 'Minimum', 'Maximum'])

# </editor-fold>

# <editor-fold desc="Create Crime Deductible Dataframe">
CrimeDed = pd.read_excel(CWRatebook, sheet_name='CrimeDeductibleFactor_Ext', skiprows=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
CrimeDed = CrimeDed.rename(columns={'Deductible': 'Deductible Amount'})
# CrimeDed = CrimeDed.drop(labels=[0, 14], axis=0)
# </editor-fold>

# <editor-fold desc="Create AOI Factor Dataframe">
if NGICRatebook != "Not found":
    NGICwb = load_workbook(NGICRatebook, read_only=True)
    if 'BasicGroupILOIFactorBldg' in NGICwb.sheetnames:
        BGIBLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupILOIFactorBldg', skiprows=11)
    else:
        BGIBLOI = pd.read_excel(CWRatebook, sheet_name='BasicGroupILOIFactorBldg', skiprows=11)
    
    if 'BasicGroupIILOIFactorBldg' in NGICwb.sheetnames:
        BGIIBLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupIILOIFactorBldg', skiprows=11)
    else:
        BGIIBLOI = pd.read_excel(CWRatebook, sheet_name='BasicGroupIILOIFactorBldg', skiprows=11)
        
    if 'BroadSpecialLOIFactorBldg' in NGICwb.sheetnames:
        SCOLBLOI = pd.read_excel(NGICRatebook, sheet_name='BroadSpecialLOIFactorBldg', skiprows=11)
    else:
        SCOLBLOI = pd.read_excel(CWRatebook, sheet_name='BroadSpecialLOIFactorBldg', skiprows=11)
        
    if 'BasicGroupILOIFactorPersProp' in NGICwb.sheetnames:
        BGIPPLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupILOIFactorPersProp', skiprows=11)
    else:
        BGIPPLOI = pd.read_excel(CWRatebook, sheet_name='BasicGroupILOIFactorPersProp', skiprows=11)
        
    if 'BasicGroupIILOIFactorPersProp' in NGICwb.sheetnames:
        BGIIPPLOI = pd.read_excel(NGICRatebook, sheet_name='BasicGroupIILOIFactorPersProp', skiprows=11)
    else:
        BGIIPPLOI = pd.read_excel(CWRatebook, sheet_name='BasicGroupIILOIFactorPersProp', skiprows=11)
        
    if 'BroadSpecialLOIFactorPrsn1Prop' in NGICwb.sheetnames:
        SCOLPPLOI = pd.read_excel(NGICRatebook, sheet_name='BroadSpecialLOIFactorPrsn1Prop', skiprows=11)
    else:
        SCOLPPLOI = pd.read_excel(CWRatebook, sheet_name='BroadSpecialLOIFactorPrsn1Prop', skiprows=11)

BGIBLOI = BGIBLOI.pivot(index='Limit', columns='ConstructionCode', values='Factor').reset_index(names=['Limit', 1, 2, 3, 4, 5, 6])
BGIBLOI = BGIBLOI.drop(columns=[2, 3, 5, 6])
BGIBLOI = BGIBLOI.drop(BGIBLOI.tail(1).index)
BGIBLOI['Limit'] = BGIBLOI['Limit'].apply('{:,}'.format)
BGIBLOI.iloc[-1, BGIBLOI.columns.get_loc('Limit')] = BGIBLOI.iloc[-1, BGIBLOI.columns.get_loc('Limit')] + "+"
BGIBLOI = BGIBLOI.rename(columns={1: "Construction Group 1-3", 4: "Construction Group 4-6"})
BGIBLOI = splitdf(BGIBLOI, 2).fillna('')

BGIIBLOI['Limit'] = BGIIBLOI['Limit'].apply('{:,}'.format)
BGIIBLOI = BGIIBLOI.drop(BGIIBLOI.tail(1).index)
BGIIBLOI.iloc[-1, BGIIBLOI.columns.get_loc('Limit')] = BGIIBLOI.iloc[-1, BGIIBLOI.columns.get_loc('Limit')] + "+"
BGIIBLOI = splitdf(BGIIBLOI, 2).fillna('')

SCOLBLOI['Limit'] = SCOLBLOI['Limit'].apply('{:,}'.format)
SCOLBLOI = SCOLBLOI.drop(SCOLBLOI.tail(1).index)
SCOLBLOI.iloc[-1, SCOLBLOI.columns.get_loc('Limit')] = SCOLBLOI.iloc[-1, SCOLBLOI.columns.get_loc('Limit')] + "+"
SCOLBLOI = splitdf(SCOLBLOI, 2).fillna('')

BGIPPLOI = BGIPPLOI.pivot(index='Limit', columns='ConstructionCode', values='Factor').reset_index(names=['Limit', 1, 2, 3, 4, 5, 6])
BGIPPLOI = BGIPPLOI.drop(columns=[2, 3, 5, 6])
BGIPPLOI = BGIPPLOI.drop(BGIPPLOI.tail(1).index)
BGIPPLOI['Limit'] = BGIPPLOI['Limit'].apply('{:,}'.format)
BGIPPLOI.iloc[-1, BGIPPLOI.columns.get_loc('Limit')] = BGIPPLOI.iloc[-1, BGIPPLOI.columns.get_loc('Limit')] + "+"
BGIPPLOI = BGIPPLOI.rename(columns={1: "Construction Group 1-3", 4: "Construction Group 4-6"})
BGIPPLOI = splitdf(BGIPPLOI, 2).fillna('')

BGIIPPLOI['Limit'] = BGIIPPLOI['Limit'].apply('{:,}'.format)
BGIIPPLOI = BGIIPPLOI.drop(BGIIPPLOI.tail(1).index)
BGIIPPLOI.iloc[-1, BGIIPPLOI.columns.get_loc('Limit')] = BGIIPPLOI.iloc[-1, BGIIPPLOI.columns.get_loc('Limit')] + "+"
BGIIPPLOI = splitdf(BGIIPPLOI, 2).fillna('')

SCOLPPLOI['Limit'] = SCOLPPLOI['Limit'].apply('{:,}'.format)
SCOLPPLOI = SCOLPPLOI.drop(SCOLPPLOI.tail(1).index)
SCOLPPLOI.iloc[-1, SCOLPPLOI.columns.get_loc('Limit')] = SCOLPPLOI.iloc[-1, SCOLPPLOI.columns.get_loc('Limit')] + "+"
SCOLPPLOI = splitdf(SCOLPPLOI, 2).fillna('')

# </editor-fold>

# <editor-fold desc="Create Tiering Grade Dataframes">
if 'GroupITierFactor_Ext' in NGICwb.sheetnames:
    BGITier = pd.read_excel(NGICRatebook, sheet_name='GroupITierFactor_Ext', skiprows=11)
else:
    BGITier = pd.read_excel(CWRatebook, sheet_name='GroupITierFactor_Ext', skiprows=11)

if 'GroupIITierFactor_Ext' in NGICwb.sheetnames:
    BGIITier = pd.read_excel(NGICRatebook, sheet_name='GroupIITierFactor_Ext', skiprows=11)
else:
    BGIITier = pd.read_excel(CWRatebook, sheet_name='GroupIITierFactor_Ext', skiprows=11)

if 'SpecialCauseofLossTierFactor_Ex' in NGICwb.sheetnames:
    SCOLTier = pd.read_excel(NGICRatebook, sheet_name='SpecialCauseofLossTierFactor_Ex', skiprows=11)
else:
    SCOLTier = pd.read_excel(CWRatebook, sheet_name='SpecialCauseofLossTierFactor_Ex', skiprows=11)

BGITier = BGITier.pivot(index='TierGradeGroup I', columns='Tiering NAICSGroup', values='Factor')
BGITier = BGITier[[col for col in BGITier.columns if col != 'All Other'] + ['All Other']]
BGITier['Tier Grade'] = BGITier.index
BGITier = BGITier[['Tier Grade'] + [col for col in BGITier.columns if col != 'Tier Grade']]
BGITier.index.name = None
BGITier = BGITier.rename_axis(None, axis=1)
BGITier = BGITier.reset_index(drop=True)

BGIITier = BGIITier.pivot(index='TierGradeGroup II', columns='Tiering NAICSGroup', values='Factor')
BGIITier = BGIITier[[col for col in BGIITier.columns if col != 'All Other'] + ['All Other']]
BGIITier['Tier Grade'] = BGIITier.index
BGIITier = BGIITier[['Tier Grade'] + [col for col in BGIITier.columns if col != 'Tier Grade']]
BGIITier.index.name = None
BGIITier = BGIITier.rename_axis(None, axis=1)
BGIITier = BGIITier.reset_index(drop=True)

SCOLTier = SCOLTier.pivot(index='TierGradeSpecialCOL', columns='Tiering NAICSGroup', values='Factor')
SCOLTier = SCOLTier[[col for col in SCOLTier.columns if col != 'All Other'] + ['All Other']]
SCOLTier['Tier Grade'] = SCOLTier.index
SCOLTier = SCOLTier[['Tier Grade'] + [col for col in SCOLTier.columns if col != 'Tier Grade']]
SCOLTier.index.name = None
SCOLTier = SCOLTier.rename_axis(None, axis=1)
SCOLTier = SCOLTier.reset_index(drop=True)
# </editor-fold>

# <editor-fold desc="Create AOB Dataframes">
if 'AgeOfBuildingFactor_Ext' in NGICwb.sheetnames:
    AOB = pd.read_excel(NGICRatebook, sheet_name='AgeOfBuildingFactor_Ext', skiprows=11)
else:
    AOB = pd.read_excel(CWRatebook, sheet_name='AgeOfBuildingFactor_Ext', skiprows=11)

AOB = AOB.pivot_table(index='AgeOfBuildingFrom', columns=['CoveredObject', 'CauseOfLossGroup'], values='Factor').reset_index(names='Age')
AOB.columns = ["Age", "BGITime", "BGIITime", "BroadTime", "EQ1", "SCOLTime", "BGIContents", "BGIIContents", "PPContents", "EQ2", "SCOLContents", "BGIBldg", "BGIIBldg", "STRBldg", "EQ3", "SCOLBldg"]

BGIAOB = AOB[["Age", "BGIBldg", "BGIContents", "BGITime"]].rename(columns={"BGIBldg": "Building", "BGIContents": "Contents", "BGITime": "Time"})
BGIIAOB = AOB[["Age", "BGIIBldg", "BGIIContents", "BGIITime"]].rename(columns={"BGIIBldg": "Building", "BGIIContents": "Contents", "BGIITime": "Time"})
SCOLAOB = AOB[["Age", "SCOLBldg", "SCOLContents", "SCOLTime"]].rename(columns={"SCOLBldg": "Building", "SCOLContents": "Contents", "SCOLTime": "Time"})

BGIAOB = splitdf(BGIAOB, 3)
BGIIAOB = splitdf(BGIIAOB, 3)
SCOLAOB = splitdf(SCOLAOB, 3)
# </editor-fold>

# <editor-fold desc="Create Deductible Factor Dataframe">
if 'DeductibleFactor' in MMwb.sheetnames:
    DIP1 = pd.read_excel(MMRatebook, sheet_name='DeductibleFactor', skiprows=11)
elif 'DeductibleFactor' in NGICwb.sheetnames:
    DIP1 = pd.read_excel(NGICRatebook, sheet_name='DeductibleFactor', skiprows=11)
else:
    DIP1 = pd.read_excel(CWRatebook, sheet_name='DeductibleFactor', skiprows=11)

if 'Deductible250Factor' in MMwb.sheetnames:
    DIP2 = pd.read_excel(MMRatebook, sheet_name='Deductible250Factor', skiprows=11)
elif 'Deductible250Factor' in NGICwb.sheetnames:
    DIP2 = pd.read_excel(NGICRatebook, sheet_name='Deductible250Factor', skiprows=11)
else:
    DIP2 = pd.read_excel(CWRatebook, sheet_name='Deductible250Factor', skiprows=11)

DIP = DIP1.pivot_table(index=['Deductible', 'Limit'], columns='CauseOfLossDeductible', values='Factor').reset_index(names=['Deductible', 'Limit'])
DIP[['Deductible']] = DIP[['Deductible']].replace({',': ''}, regex=True).astype(float)
DedLim = pd.DataFrame(DIP['Limit'].unique().T, columns=['Limit'])
DedLim = pd.concat([DedLim] * len(DIP2['CauseOfLoss']))
DIP.columns = DIP.columns.str.lower()

Ded250 = np.array(DIP2['CauseOfLoss'])
Ded250F = np.array(DIP2['Factor'])
DedLim['CauseOfLoss'] = Ded250[DedLim.groupby('Limit').cumcount()]
DedLim['Factor'] = Ded250F[DedLim.groupby('Limit').cumcount()]
DedLim = DedLim.assign(Deductible='250')
DedLim = DedLim[['Deductible', 'CauseOfLoss', 'Limit', 'Factor']]

DedLim2 = DedLim.pivot_table(index=['Deductible', 'Limit'], columns='CauseOfLoss', values='Factor').reset_index(names=['Deductible', 'Limit'])
DedLim2 = DedLim2.drop(['Broad'], axis=1).rename(columns={"Basic": "Basic Group I", "All Other": "Basic Group II", "Special": "Other Cause Of Loss"})
DedLim2.columns = DedLim2.columns.str.lower()

DIP = DIP.sort_values(by=['deductible', 'limit'])
DIP['deductible'] = DIP['deductible'].astype(int).apply('{:,}'.format)
DIP = DIP.reset_index(drop=True)
DIP = pd.concat([DedLim2, DIP], ignore_index=True)

DIPI = DIP[['deductible', 'limit', 'basic group i']].rename(columns={"basic group i": "Factor"})
DIPII = DIP[['deductible', 'limit', 'basic group ii']].rename(columns={"basic group ii": "Factor"})
DIPSCOL = DIP[['deductible', 'limit', 'other cause of loss']].rename(columns={"other cause of loss": "Factor"})

DIPI.columns = DIPI.columns.map(str.title)
DIPII.columns = DIPII.columns.map(str.title)
DIPSCOL.columns = DIPSCOL.columns.map(str.title)

DIPILim = DIPI['Limit'].astype(str)
DIPIILim = DIPII['Limit'].astype(str)
DIPSCOLLim = DIPSCOL['Limit'].astype(str)

for i in range(len(DIPI['Limit'])):
    if int(DIPI['Limit'][i]) not in [min(DIPI['Limit'].astype(int)), max(DIPI['Limit'].astype(int))]:
        DIPILim[i] = str(f"{DIPI['Limit'][i-1]:,}") + ' - ' + str(f"{DIPI['Limit'][i]:,}")
    elif int(DIPI['Limit'][i]) == min(DIPI['Limit'].astype(int)):
        DIPILim[i] = str(f"{DIPI['Limit'][i]:,}") + ' or less'
    else:
        DIPILim[i] = 'More than ' + str(f"{DIPI['Limit'][i-1]:,}")

for i in range(len(DIPII['Limit'])):
    if int(DIPII['Limit'][i]) not in [min(DIPII['Limit'].astype(int)), max(DIPII['Limit'].astype(int))]:
        DIPIILim[i] = str(f"{DIPII['Limit'][i-1]:,}") + ' - ' + str(f"{DIPII['Limit'][i]:,}")
    elif int(DIPII['Limit'][i]) == min(DIPII['Limit'].astype(int)):
        DIPIILim[i] = str(f"{DIPII['Limit'][i]:,}") + ' or less'
    else:
        DIPIILim[i] = 'More than ' + str(f"{DIPII['Limit'][i-1]:,}")

for i in range(len(DIPSCOL['Limit'])):
    if int(DIPSCOL['Limit'][i]) not in [min(DIPSCOL['Limit'].astype(int)), max(DIPSCOL['Limit'].astype(int))]:
        DIPSCOLLim[i] = str(f"{DIPSCOL['Limit'][i-1]:,}") + ' - ' + str(f"{DIPSCOL['Limit'][i]:,}")
    elif int(DIPSCOL['Limit'][i]) == min(DIPSCOL['Limit'].astype(int)):
        DIPSCOLLim[i] = str(f"{DIPSCOL['Limit'][i]:,}") + ' or less'
    else:
        DIPSCOLLim[i] = 'More than ' + str(f"{DIPSCOL['Limit'][i-1]:,}")

DIPI['Limit'] = DIPILim
DIPII['Limit'] = DIPIILim
DIPSCOL['Limit'] = DIPSCOLLim

DIPI = splitdf(DIPI, 3)
DIPII = splitdf(DIPII, 3)
DIPSCOL = splitdf(DIPSCOL, 3)
# </editor-fold>

# <editor-fold desc="Create M&S Dataframe">
if 'CrimeTerritoryBaseRate_Ext' in NGICwb.sheetnames:
    CrimeMSDF = pd.read_excel(NGICRatebook, sheet_name='CrimeTerritoryBaseRate_Ext', skiprows=11).rename(columns={"CrimeTerritory": "Territory", "Base Premium": "Premium"})
else:
    CrimeMSDF = pd.read_excel(CWRatebook, sheet_name='CrimeTerritoryBaseRate_Ext', skiprows=11).rename(columns={"CrimeTerritory": "Territory", "Base Premium": "Premium"})

if 'MoneyandSecuritiesOccupancyFact' in NGICwb.sheetnames:
    MSDF = pd.read_excel(NGICRatebook, sheet_name='MoneyandSecuritiesOccupancyFact', skiprows=11)
else:
    MSDF = pd.read_excel(CWRatebook, sheet_name='MoneyandSecuritiesOccupancyFact', skiprows=11)

MSDF = MSDF.pivot_table(index=['Inside', 'Outside'], columns='Occupancy', values='Factor').reset_index()
MSDF = MSDF.rename(columns={'Inside': 'Inside Limit', 'Outside': 'Outside Limit'})
MSDF['Inside Limit'] = MSDF['Inside Limit'].apply('{:,}'.format)
MSDF['Outside Limit'] = MSDF['Outside Limit'].apply('{:,}'.format)
# </editor-fold>

# <editor-fold desc="Create Employee Dishonesty Dataframe">
if 'EmployeeDishonestyLimitsBaseRat' in NGICwb.sheetnames:
    EmployeeDF = pd.read_excel(NGICRatebook, sheet_name='EmployeeDishonestyLimitsBaseRat', skiprows=11)
else:
    EmployeeDF = pd.read_excel(CWRatebook, sheet_name='EmployeeDishonestyLimitsBaseRat', skiprows=11)

if 'EmployeeDishonestyLimitsRateabl' in NGICwb.sheetnames:
    EDLR = pd.read_excel(NGICRatebook, sheet_name='EmployeeDishonestyLimitsRateabl', skiprows=11)
else:
    EDLR = pd.read_excel(CWRatebook, sheet_name='EmployeeDishonestyLimitsRateabl', skiprows=11)

EmployeeDF["Each Add'l Rateable Employee"] = EDLR['Premium Per Rateable Employee >5']
EmployeeDF.rename(columns={'Premium < 5 Rateable Employees': '1-5 Rateable Employees'}, inplace=True)
EmployeeDF['Limit'] = EmployeeDF['Limit'].apply('{:,}'.format)
# </editor-fold>

# <editor-fold desc="Create Fraudulent Impersonation DF">
if 'FraudulentImpersonationEmployee' in NGICwb.sheetnames:
    Fraudone = pd.read_excel(NGICRatebook, sheet_name='FraudulentImpersonationEmployee', skiprows=11)
else:
    Fraudone = pd.read_excel(CWRatebook, sheet_name='FraudulentImpersonationEmployee', skiprows=11)

if 'FraudulentImpersonationEmpl (1)' in NGICwb.sheetnames:
    Fraudtwo = pd.read_excel(NGICRatebook, sheet_name='FraudulentImpersonationEmpl (1)', skiprows=11)
else:
    Fraudtwo = pd.read_excel(CWRatebook, sheet_name='FraudulentImpersonationEmpl (1)', skiprows=11)

FraudDF = Fraudtwo.merge(Fraudone, how='inner', on='Limit').rename(columns={"Premium < 5 Rateable Employees": "1-5 Rateable Employees", "Premium Per Rateable Employee >5": "Each Add'l Rateable Employee"})
FraudDF['Limit'] = FraudDF['Limit'].apply('{:,}'.format)

if 'VerificationFactorEmployeesFact' in NGICwb.sheetnames:
    FraudDF2 = pd.read_excel(NGICRatebook, sheet_name='VerificationFactorEmployeesFact', skiprows=11)
else:
    FraudDF2 = pd.read_excel(CWRatebook, sheet_name='VerificationFactorEmployeesFact', skiprows=11)

FraudDF2['Verification Type'] = FraudDF2['Verification Type'].map({
    'OptionAVerifReqdAllTransInstr': "Verification required for all transfer instructions",
    'OptionBVerifReqdAllTransInstrExcessSpecifiedAmt': "Verification required for all transfer instructions in excess of an amount",
    'OptionCVerifTransferInstrNotReqd': "Verification of transfer instructions not required"
})
# </editor-fold>

# <editor-fold desc="IRPM DF">
if 'Schedule Rating Threshold_Ext' in NGICwb.sheetnames:
    IRPM1 = pd.read_excel(NGICRatebook, sheet_name='Schedule Rating Threshold_Ext', skiprows=11).drop('AbsoluteThreshold', axis=1).rename(columns={'ScheduleEligibilityIndicator': 'Constant'})
else:
    IRPM1 = pd.read_excel(CWRatebook, sheet_name='Schedule Rating Threshold_Ext', skiprows=11).drop('AbsoluteThreshold', axis=1).rename(columns={'ScheduleEligibilityIndicator': 'Constant'})

if 'IRPMMaximumCredit' in NGICwb.sheetnames:
    IRPM2 = pd.read_excel(NGICRatebook, sheet_name='IRPMMaximumCredit', skiprows=11)
else:
    IRPM2 = pd.read_excel(CWRatebook, sheet_name='IRPMMaximumCredit', skiprows=11)

if 'IRPMMaximumDebit' in NGICwb.sheetnames:
    IRPM3 = pd.read_excel(NGICRatebook, sheet_name='IRPMMaximumDebit', skiprows=11)
else:
    IRPM3 = pd.read_excel(CWRatebook, sheet_name='IRPMMaximumDebit', skiprows=11)

IRPM = pd.DataFrame(index=range(3), columns=range(2))
IRPM.iloc[0,0] = "Minimum Eligible Premium"
IRPM.iloc[0,1] = IRPM1.iloc[0,1]
IRPM.iloc[1,0] = "Maximum Credit"
IRPM.iloc[1,1] = IRPM2.iloc[0,1].astype(str) + "%"
IRPM.iloc[2,0] = "Maximum Debit"
IRPM.iloc[2,1] = IRPM3.iloc[0,1].astype(str) + "%"
# </editor-fold>

# <editor-fold desc="Emergency Evacuation DF">
if 'EmergencyEvacuationIncludin (1)' in NGICwb.sheetnames:
    EmergencyDF = pd.read_excel(NGICRatebook, sheet_name='EmergencyEvacuationIncludin (1)', skiprows=11)
else:
    EmergencyDF = pd.read_excel(CWRatebook, sheet_name='EmergencyEvacuationIncludin (1)', skiprows=11)
EmergencyDF['Limit'] = EmergencyDF['Limit'].apply('{:,}'.format)
# </editor-fold>

# <editor-fold desc="Civil Authority DF">
if 'CivilAuthorityIncreasedRadiusCo' in NGICwb.sheetnames:
    CivilDF = pd.read_excel(NGICRatebook, sheet_name='CivilAuthorityIncreasedRadiusCo', skiprows=11).rename(columns={"Radius": "Radius in Miles"})
else:
    CivilDF = pd.read_excel(CWRatebook, sheet_name='CivilAuthorityIncreasedRadiusCo', skiprows=11).rename(columns={"Radius": "Radius in Miles"})
# </editor-fold>

# <editor-fold desc="Computer DFs">
if 'ComputerandFundsTransferFraudBa' in NGICwb.sheetnames:
    ComputerDF = pd.read_excel(NGICRatebook, sheet_name='ComputerandFundsTransferFraudBa', skiprows=11)
else:
    ComputerDF = pd.read_excel(CWRatebook, sheet_name='ComputerandFundsTransferFraudBa', skiprows=11)

if 'ComputerandFundsTransferFraudAn' in NGICwb.sheetnames:
    ComputerAnnualDF = pd.read_excel(NGICRatebook, sheet_name='ComputerandFundsTransferFraudAn', skiprows=11).fillna('0').astype(str)
else:
    ComputerAnnualDF = pd.read_excel(CWRatebook, sheet_name='ComputerandFundsTransferFraudAn', skiprows=11).fillna('0').astype(str)

ComputerAnnualDF['Total Sales Min'] = ComputerAnnualDF['Total Sales Min'].apply(lambda x: "{:,}".format(int(float(x))))
ComputerAnnualDF['Total Sales Max'] = ComputerAnnualDF['Total Sales Max'].apply(lambda x: "{:,}".format(int(float(x))))
ComputerAnnualDF.iloc[0,0] = 0
ComputerAnnualDF['Sales'] = ComputerAnnualDF['Total Sales Min'].astype(str) + ' - ' + ComputerAnnualDF['Total Sales Max'].astype(str)
ComputerAnnualDF = ComputerAnnualDF.drop(columns=['Total Sales Min', 'Total Sales Max'])
ComputerAnnualDF = ComputerAnnualDF[['Sales'] + [col for col in ComputerAnnualDF.columns if col != 'Sales']]
ComputerAnnualDF = ComputerAnnualDF.iloc[0:5]
ComputerAnnualDF.iloc[4,0] = "over 25,000,001"
ComputerAnnualDF.iloc[4,1] = "2.00 + .10 for each addn'l 10,000,000"
# </editor-fold>

# <editor-fold desc="SM Index Sheet">
if v5.get() == 1:
    wb = Workbook()
    wb.active.title = "Index"

    # Initializing Default Settings
    fontName = 'Arial'
    fontSize = 10
    headerFontName = 'Arial'
    headerFontSize = 10
    footerFontName = 'Arial'
    footerFontSize = 10
    leftMargin = 0.5
    rightMargin = 0.5
    topMargin = 2.25
    bottomMargin = 1.5
    headerMargin = 1.5
    footerMargin = 0.5

    # Building Fonts and Formats
    font = Font(name=fontName, size=fontSize)
    fontBold = Font(name=fontName, size=fontSize, bold=True)
    fontBoldUnderline = Font(name=fontName, size=fontSize, bold=True, underline='single')
    fontBoldItalics = Font(name=fontName, size=fontSize, italic=True, bold=True)
    fontBlue = Font(name=fontName, size=fontSize, color='0000FF')
    headerFont = headerFontName + ', Bold'
    footerFont = footerFontName + ', Bold'
    rateFormat = '#,##0.000'
    rateFormat2 = '#,##0.00'
    codeFormat = '#,##0'
    currencyFormat = '$#,##0'
    percentFormat = '0.0%'
# </editor-fold>

# <editor-fold desc="SM LCM Excel Sheet">
if v5.get() == 1 and v1.get() == 1:
    ws = wb.create_sheet(title='LCM')
    formatwkstSM(wkstname=ws, titlerows='3', A1title='BASE RATE CALCULATION', A2title='Loss Cost Multiplier', dfname=SMLCMTable, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            char = get_column_letter(col)
            cell = ws[char + str(row)]
            ws.column_dimensions[char].bestFit = True
            
            if row > 3 and cell.value is not None:
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            
            if row < 2:
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.alignment = Alignment(horizontal='left', vertical='bottom', wrap_text=False)
            elif row == 4:
                cell.number_format = rateFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            else:
                cell.number_format = rateFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    if NICOFRatebook != "Not found":
        ws.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws.oddFooter.left.text = "\nNationwide General Insurance Company"

    ws.oddFooter.left.size = footerFontSize
    ws.oddFooter.left.font = footerFont
    ws.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws.oddFooter.center.size = footerFontSize
    ws.oddFooter.center.font = footerFont
# </editor-fold>

# <editor-fold desc="SM PMF Excel Sheet">
if v5.get() == 1 and v2.get() == 1:
    ws2 = wb.create_sheet(title='PMF')
    formatwkstSM(wkstname=ws2, titlerows='3', A1title='PACKAGE MODIFICATION FACTOR', A2title='Blank', dfname=SMPMF, statename=State, stabb=StateAbb, effdate=EffectiveDate)
    
    for row in range(1, ws2.max_row + 1):
        for col in range(1, ws2.max_column + 1):
            char = get_column_letter(col)
            cell = ws2[char + str(row)]
            ws2.column_dimensions[char].bestFit = True
            
            if row > 2 and cell.value is not None:
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            
            if row < 2:
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif col == 1:
                cell.number_format = codeFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            else:
                cell.number_format = rateFormat2
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)

    for col in ws2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws2.column_dimensions[column].width = adjusted_width

    if NICOFRatebook != "Not found":
        ws2.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws2.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws2.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws2.oddFooter.left.text = "\nNationwide General Insurance Company"

    ws2.oddFooter.left.size = footerFontSize
    ws2.oddFooter.left.font = footerFont
    ws2.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws2.oddFooter.center.size = footerFontSize
    ws2.oddFooter.center.font = footerFont
# </editor-fold>

# <editor-fold desc="SM Crime Endorsement Excel Sheet">
if v5.get() == 1 and v7.get() == 1:
    ws3 = wb.create_sheet(title='Crime')
    ws3['A1'] = 'CRIME ENDORSEMENT'
    ws3['A2'] = 'Blank'
    
    for r in dataframe_to_rows(CrimeDed, False, True):
        if False & len(list(r))==1:
            ws3['B4'] = list(r)[0]
            continue
        ws3.append(r)
        
    ws3['A19'] = 'Blank'
    ws3['A20'] = 'Blank'
    ws3['A21'] = 'Blank'
    
    for r in dataframe_to_rows(SMCRIMELCMTable, False, True):
        ws3.append(r)
        
    ws3.page_setup.orientation = 'portrait'
    ws3.page_setup.blackAndWhite = False
    ws3.page_setup.firstPageNumber = 1
    ws3.page_setup.useFirstPageNumber = True
    ws3.sheet_view.showGridLines = False
    ws3.print_title_rows = '1:3'
    ws3.page_margins.left = leftMargin
    ws3.page_margins.right = rightMargin
    ws3.page_margins.top = topMargin
    ws3.page_margins.bottom = bottomMargin
    ws3.page_margins.header = headerMargin
    ws3.page_margins.footer = footerMargin
    ws3.print_options.horizontalCentered = True

    if State == 'Florida':
        ws3.oddHeader.left.text = "Commercial Lines Manual: Commercial Property Non-Residential"
    else:
        ws3.oddHeader.left.text = "Commercial Lines Manual: Commercial Property"

    ws3.oddHeader.left.size = headerFontSize
    ws3.oddHeader.left.font = headerFont
    ws3.oddHeader.center.text = "\n\n" + State + " Rate Pages"
    ws3.oddHeader.center.size = headerFontSize
    ws3.oddHeader.center.font = headerFont
    ws3.oddHeader.right.text = "Effective Date: " + EffectiveDate
    ws3.oddHeader.right.size = headerFontSize
    ws3.oddHeader.right.font = headerFont
    
    if NICOFRatebook != "Not found":
        ws3.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide Insurance Company of Florida \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook != "Not found":
        ws3.oddFooter.left.text = "\nNationwide Affinity Insurance Company of America \nNationwide Assurance Company \nNationwide General Insurance Company"
    elif NICOFRatebook == "Not found" and NAFFRatebook == "Not found" and NACORatebook != "Not found":
        ws3.oddFooter.left.text = "\nNationwide Assurance Company \nNationwide General Insurance Company"
    else:
        ws3.oddFooter.left.text = "Nationwide General Insurance Company"

    ws3.oddFooter.left.size = footerFontSize
    ws3.oddFooter.left.font = footerFont
    ws3.oddFooter.center.text = StateAbb + " - SRP &[Tab] - &P"
    ws3.oddFooter.center.size = footerFontSize
    ws3.oddFooter.center.font = footerFont

    for row in range(1, ws3.max_row + 1):
        for col in range(1, ws3.max_column + 1):
            char = get_column_letter(col)
            cell = ws3[char + str(row)]
            ws3.column_dimensions[char].bestFit = True
            
            if 3 < row < 19 and cell.value is not None:
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))
            if row < 2:
                cell.font = fontBoldUnderline
            elif row == 2:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif col == 1:
                cell.number_format = codeFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            elif col == 2 and row < 17:
                cell.number_format = rateFormat2
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            else:
                cell.number_format = rateFormat
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='bottom', wrap_text=False)
            
            if 18 < row < 22:
                cell.font = Font(name=fontName, size=fontSize, color='FFFFFFFF')
            elif row > 22 and cell.value is not None:
                cell.border = Border(left=Side(border_style='thin', color='00000000'),
                                     right=Side(border_style='thin', color='00000000'),
                                     top=Side(border_style='thin', color='00000000'),
                                     bottom=Side(border_style='thin', color='00000000'))

    for col in ws3.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws3.column_dimensions[column].width = adjusted_width
# </editor-fold>

# <editor-fold desc="Save SM Workbook">
if v5.get() == 1:
    sheetNames = wb.sheetnames
    for i in range(1, len(sheetNames)):
        if wb[sheetNames[i]]['A3'].value == '*':
            wb["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames[i]}'!E1")
        else:
            wb["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames[i]}'!A1")
        wb["Index"]['A' + str(i)].value = wb[sheetNames[i]]['A1'].value
        wb["Index"]['A' + str(i)].font = fontBlue

    wb.save(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Small Market Rate Pages.xlsx'))
    wb = xw.Book(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Small Market Rate Pages.xlsx'))
    current_label = wb.api.SensitivityLabel.GetLabel()
    
    if current_label.LabelName == "":
        labelinfo = wb.api.SensitivityLabel.CreateLabelInfo()
        labelinfo.AssignmentMethod = 2
        labelinfo.Justification = 'init'
        labelinfo.LabelId = 'fbefcacb-1e54-47c6-a321-a2f3e970fe0d'
        labelinfo.LabelName = 'Restricted'
        wb.api.SensitivityLabel.SetLabel(labelinfo, labelinfo)
    wb.save()
# </editor-fold>

# <editor-fold desc="Save MM Workbook">
if v10.get() == 1:
    sheetNames2 = wb2.sheetnames
    for i in range(1, len(sheetNames2)):
        if wb2[sheetNames2[i]]['A3'].value == '*':
            wb2["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames2[i]}'!E1")
        else:
            wb2["Index"]['A' + str(i)].hyperlink = (f"#'{sheetNames2[i]}'!A1")
        wb2["Index"]['A' + str(i)].value = wb2[sheetNames2[i]]['A1'].value
        wb2["Index"]['A' + str(i)].font = fontBlue

    wb2.save(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Middle Market Rate Pages.xlsx'))
    wb2 = xw.Book(os.path.join(folder_selected, State + ' ' + EffectiveDate + ' Middle Market Rate Pages.xlsx'))
    current_label = wb2.api.SensitivityLabel.GetLabel()
    
    if current_label.LabelName == "":
        labelinfo = wb2.api.SensitivityLabel.CreateLabelInfo()
        labelinfo.AssignmentMethod = 2
        labelinfo.Justification = 'init'
        labelinfo.LabelId = 'fbefcacb-1e54-47c6-a321-a2f3e970fe0d'
        labelinfo.LabelName = 'Restricted'
        wb2.api.SensitivityLabel.SetLabel(labelinfo, labelinfo)
    wb2.save()
# </editor-fold>

# <editor-fold desc="Time taken">
print(time.time() - start_time)
# </editor-fold>