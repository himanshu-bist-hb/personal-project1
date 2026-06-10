"""
FA/FARates.py
=============
Farm Auto rate pages.

Farm Auto inherits EVERYTHING from Business Auto. The only thing this file
does right now is expose the `Auto` class under the FA package so that
FARatePages.py can import it as `FA.Auto`.

HOW FA INHERITANCE WORKS
-------------------------
`class Auto(_BABase)` means:
  - All data-build methods (buildExpenseConstant, buildBaseRates, ...)  ← inherited
  - All page-rule methods (_page_rule_208, _page_rule_222b, ...)        ← inherited
  - All format methods (format222B, formatBaseRates, ...)               ← inherited
  - buildBAPages() and buildFAPages()                                   ← inherited / defined here

THREE SITUATIONS WHEN EDITING FA RULES
----------------------------------------

Situation 1: Rule is IDENTICAL in FA → do nothing. It is already inherited.

Situation 2: Rule DATA changes in FA (different numbers):
  Override the build method here, e.g.:
    def buildExpenseConstant(self, company):
        # FA-specific data logic
        ...

Situation 3: Entirely NEW FA-only rule:
  Step A – Write buildFAXxx(company) in Section A below.
  Step B – Write _page_rule_fa_xxx(RatePages) in Section B below.
  Step C – Add self._page_rule_fa_xxx(RatePages) inside buildFAPages() in Section C.

HOW TO REMOVE A RULE FOR FA
-----------------------------
In buildFAPages() below, comment out the line for that rule:
    # self._page_rule_208(RatePages)   ← commented out = rule skipped for FA

HOW TO REORDER RULES
----------------------
In buildFAPages(), move the rule's line up or down.
"""

import warnings
import pandas as pd

from BA.BARates import Auto as _BABase
from config.constants import BA_INPUT_FILE

# Suppress noise we don't care about at import time
warnings.simplefilter("ignore", DeprecationWarning)
warnings.simplefilter("ignore", FutureWarning)


class Auto(_BABase):
    """
    Farm Auto rate page generator.
    Inherits all rules and formatting from Business Auto.

    FA HIERARCHY DIFFERENCES vs BA
    --------------------------------
    BA:  Company ratebook → NGIC (state-level, mandatory) → BA CW ratebook
    FA:  Company ratebook → NWAG (state-level, mandatory) → FA CW ratebook

    These two class attributes override the BA defaults so that every method
    that references the state-level company (compareCompanies, nesting) uses
    NWAG automatically — without duplicating any method code.
    """

    # ── FA hierarchy: NWAG is the state-level company (replaces BA's NGIC) ──
    _STATE_LEVEL_COMPANY = "NWAG"
    _COMPANIES_CHECK = [
        "NWAG", "NACO", "NAFF", "CCMIC", "HICNJ",
        "NICOF", "NMIC", "AICOA", "NICOA", "NPCIC",
    ]

    # =========================================================================
    # FA nesting override
    # =========================================================================

    def nesting(self):
        """
        FA nesting — identical logic to BA but with NWAG as the Level 2
        (state-level) ratebook instead of NGIC.

        Level 1: individual FA company ratebook (NACO, NAFF, etc.)
        Level 2: NWAG — mandatory FA state-level ratebook
        Level 3: FA CW ratebook (loaded under the "CW" key by FARatePages.run)

        NWAG MUST be last in ratebook_names so it is fully built (own tables +
        CW cascade) before any other company's Level 2 cascade uses it.
        """
        ratebook_names = [
            "NAFF", "NACO", "NICOF", "CCMIC", "HICNJ",
            "NICOA", "AICOA", "NPCIC", "NMIC",
            "NWAG",   # ← Level 2 for FA; must be last
        ]

        available_names = [
            name for name in ratebook_names
            if self.rateTables.get(name) not in (None, "Not found")
        ]

        for name in available_names:
            name, LEVEL1 = self.process_ratebook(name, self.rateTables)
            self.rateTables[name] = LEVEL1

        # FA equivalent of BA's MM suppression: if NMIC (MM) is provided,
        # suppress NWAG pages (same logic BA uses to suppress NGIC with NMIC).
        if (self.rateTables.get("NMIC") is not None) or (self.rateTables.get("CCMIC") is not None):
            self.rateTables["NWAG"] = None

    # =========================================================================
    # Section A: FA-only BUILD methods
    # Add a method here when FA needs data tables that BA does not have.
    # Pattern:
    #   def buildFAXxx(self, company):
    #       data = self.rateTables[company]["SomeNewFATable_Ext"]
    #       # ... build and return a pandas DataFrame
    # =========================================================================

    def buildFAPolicyMinimumPremium(self, company):
        raw = self.rateTables[company].get("PolicyMinimumPremium_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Description", "Minimum Premium"])

        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        df = df[["MinimumPremiumType", "MinimumPremium"]].copy()

        type_display = {
            "All Other":                "All other policies:",
            "Hired and Non-Owned Only": "Policies providing hired auto and/or "
                                        "nonowned auto coverage only:",
        }
        df["MinimumPremiumType"] = (
            df["MinimumPremiumType"]
            .map(type_display)
            .fillna(df["MinimumPremiumType"])
        )
        df["MinimumPremium"] = pd.to_numeric(df["MinimumPremium"], errors="coerce")
        df.columns = ["Description", "Minimum Premium"]
        return df

    def buildFALayupFactors(self, company):
        # FA uses FarmLayUpFactor_Ext (2-column table: Months Laid Up | Factor).
        # BA uses AutoLayUpFactor_Ext (many vehicle-type columns).
        raw = self.rateTables[company].get("FarmLayUpFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Months Laid Up", "Factor"])

        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)

        # Ratebook stores "7-12"; rate page shows "7+"
        df["Months Laid Up"] = df["Months Laid Up"].replace("7-12", "7+")

        # Sort into display order: 0-2, 3, 4, 5, 6, 7+
        _order = {"0-2": 0, "3": 1, "4": 2, "5": 3, "6": 4, "7+": 5}
        df["_sort"] = df["Months Laid Up"].map(_order)
        df = df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

        df["Factor"] = df["Factor"].astype(float).map(lambda x: f"{x:.3f}")
        return df

    def buildFAPollutionBaseRate(self, company):
        # FA Rule 303 — FarmPollutionLiabilityBaseRate_Ext
        # Ratebook: Constant | BaseRate  →  Rate page: Base Rate: | Per $25,000 Limit of Liability
        raw = self.rateTables[company].get("FarmPollutionLiabilityBaseRate_Ext")
        if raw is None:
            return pd.DataFrame(columns=["", "Per $25,000 Limit of Liability"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        base_rate = pd.to_numeric(df["BaseRate"].iloc[0], errors="coerce")
        return pd.DataFrame({
            "": ["Base Rate:"],
            "Per $25,000 Limit of Liability": [f"${base_rate:,.2f}"],
        })

    def buildFAPollutionFleetFactor(self, company):
        # FA Rule 303 — FarmPollutionLiabilityFleetFactor_Ext
        # Ratebook: PowerUnitMinimum | Factor  →  Rate page: Fleet Size (ranges) | Fleet Factor
        raw = self.rateTables[company].get("FarmPollutionLiabilityFleetFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Fleet Size", "Fleet Factor"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        _range_map = {0: "0-4", 5: "5-9", 10: "10-19", 20: "20-29", 30: "30+"}
        df["PowerUnitMinimum"] = (
            pd.to_numeric(df["PowerUnitMinimum"], errors="coerce")
            .apply(lambda v: _range_map.get(int(v), str(int(v))) if pd.notna(v) else v)
        )
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce").map(lambda x: f"{x:.2f}")
        df = df.rename(columns={"PowerUnitMinimum": "Fleet Size", "Factor": "Fleet Factor"})
        return df

    def buildFATieringFactor(self, company):
        # FA Rule 420 — TieringFactor_Ext
        # Ratebook: Loss Grade | Factor  →  Rate page: Grade | Loss Factor
        raw = self.rateTables[company].get("TieringFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Grade", "Loss Factor"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce").map(lambda x: f"{x:.4f}")
        df = df.rename(columns={"Loss Grade": "Grade", "Factor": "Loss Factor"})
        return df[["Grade", "Loss Factor"]]

    def buildFAMitigationFactor(self, company):
        # FA Rule 420 — TieringMitigationFactor_Ext
        # Ratebook: Power Units | Factor  →  Rate page: Number of Vehicles | Factor
        # Row 0 is a boundary marker (dropped). Row 1 → "1-24". Last row (30) → "30+".
        raw = self.rateTables[company].get("TieringMitigationFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Number of Vehicles", "Factor"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        df["Power Units"] = pd.to_numeric(df["Power Units"], errors="coerce")
        df = df[df["Power Units"] != 0].reset_index(drop=True)
        df["Power Units"] = df["Power Units"].apply(
            lambda v: "1-24" if v == 1 else ("30+" if v == 30 else str(int(v)))
        )
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce").map(lambda x: f"{x:.3f}")
        df = df.rename(columns={"Power Units": "Number of Vehicles"})
        return df[["Number of Vehicles", "Factor"]]

    def buildFATieringFactorMatrix(self, company):
        # FA Rule 420 exception states (CA, MT, NY, UT, WA).
        # Cross-product: TieringFactor_Ext (rows=Loss Grade) × RetentionFactor_Ext (cols=Retention Grade).
        # Each cell = TieringFactor × RetentionFactor, formatted to 4 decimal places.
        raw_tier = self.rateTables[company].get("TieringFactor_Ext")
        raw_ret  = self.rateTables[company].get("RetentionFactor_Ext")
        if raw_tier is None or raw_ret is None:
            return pd.DataFrame(columns=["Loss Grade"])

        df_tier = pd.DataFrame(raw_tier[1:], columns=raw_tier[0])
        df_tier = df_tier.dropna(how="all").reset_index(drop=True)
        df_tier["Factor"] = pd.to_numeric(df_tier["Factor"], errors="coerce")

        df_ret = pd.DataFrame(raw_ret[1:], columns=raw_ret[0])
        df_ret = df_ret.dropna(how="all").reset_index(drop=True)
        df_ret["Factor"] = pd.to_numeric(df_ret["Factor"], errors="coerce")

        ret_grades  = df_ret["Retention Grade"].tolist()
        ret_factors = df_ret["Factor"].tolist()

        data = {"Loss Grade": df_tier["Loss Grade"].tolist()}
        for rg, rf in zip(ret_grades, ret_factors):
            data[rg] = [
                f"{tf * rf:.4f}" if pd.notna(tf) and pd.notna(rf) else ""
                for tf in df_tier["Factor"]
            ]
        return pd.DataFrame(data)

    def buildFAMultiplePolicyDiscount(self, company):
        # FA Rule 440 — MultiplePolicyDiscountFactor_Ext
        # Ratebook: MPD Applies | Factor → Rate page shows only the TRUE row factor.
        raw = self.rateTables[company].get("MultiplePolicyDiscountFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Description", "Factor"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        true_row = df[df["MPD Applies"].astype(str).str.upper() == "TRUE"]
        if true_row.empty:
            factor = ""
        else:
            val = pd.to_numeric(true_row["Factor"].iloc[0], errors="coerce")
            factor = f"{val:.2f}"
        return pd.DataFrame({
            "Description": ["Multiply the Eligible Policy Premiums by the Following Factor:"],
            "Factor": [factor],
        })

    def buildFARiskScoreFactor(self, company):
        # FA Rule 455 — RiskScoreFactor_Ext
        # Ratebook: Risk Score Reporting Class | Customer Type | Factor
        # Rate page: Financial Reporting Class | Factor (CountryChoice rows only)
        raw = self.rateTables[company].get("RiskScoreFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Financial Reporting Class", "Factor"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        df = df[df["Customer Type"].astype(str).str.strip() == "CountryChoice"].copy()
        df = df.drop_duplicates(subset=["Risk Score Reporting Class"]).reset_index(drop=True)
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce").map(lambda x: f"{x:.2f}")
        df = df.rename(columns={"Risk Score Reporting Class": "Financial Reporting Class"})
        return df[["Financial Reporting Class", "Factor"]]

    def buildFARule284(self, company):
        # FA Rule 284 — reads from motorcycle factor tables instead of BA's hardcoded data.
        # Liability  → SpecialTypesMotorcycleLiabilityFactor, engine size 501-800
        # Medical    → SpecialTypesMotorcycleFactor MedicalPayments
        # OTC        → SpecialTypesMotorcycleFactor PhysDamOTCACV (ACV chosen over SA)
        # Collision  → SpecialTypesMotorcycleFactor PhysDamCollACV (ACV chosen over SA)
        # UM         → SpecialTypesMotorcycleFactor UM/UIM
        _empty = pd.DataFrame(columns=["Coverage", "All-terrain Vehicles", "Utility Task Vehicles"])
        raw_factor = self.rateTables[company].get("SpecialTypesMotorcycleFactor")
        raw_liab   = self.rateTables[company].get("SpecialTypesMotorcycleLiabilityFactor")
        if raw_factor is None and raw_liab is None:
            return _empty

        df_f = pd.DataFrame(raw_factor[1:], columns=raw_factor[0])
        df_f = df_f.dropna(how="all").reset_index(drop=True)

        def _get(tc):
            row = df_f[df_f["TypeCoverage"] == tc]
            return pd.to_numeric(row["Factor"].iloc[0], errors="coerce") if not row.empty else None

        med_f  = _get("MedicalPayments")
        otc_f  = _get("PhysDamOTCACV")
        coll_f = _get("PhysDamCollACV")
        um_f   = _get("UM/UIM")

        liab_f = None
        if raw_liab is not None:
            df_l = pd.DataFrame(raw_liab[1:], columns=raw_liab[0])
            df_l = df_l.dropna(how="all").reset_index(drop=True)
            row_l = df_l[df_l["EngineSize"] == "501-800"]
            if not row_l.empty:
                liab_f = pd.to_numeric(row_l["Factor"].iloc[0], errors="coerce")

        fmt = lambda v: f"{v:.2f}" if v is not None else ""
        factors = [fmt(liab_f), fmt(med_f), fmt(otc_f), fmt(coll_f), fmt(um_f)]

        return pd.DataFrame({
            "Coverage":              ["Liability", "Medical", "Other than Collision", "Collision", "UM"],
            "All-terrain Vehicles":  factors,
            "Utility Task Vehicles": factors,
        })

    def buildFARule222D(self, company):
        # FA Rule 222.D — UnlicensedFarmTruckFactor_Ext
        # Ratebook: CoverageGroup | Factor
        # Blank CoverageGroup → "All other"; display order matches the manual.
        raw = self.rateTables[company].get("UnlicensedFarmTruckFactor_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Coverage", "Factor"])

        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)

        df["CoverageGroup"] = (
            df["CoverageGroup"].fillna("All other").astype(str).str.strip()
            .replace("", "All other")
        )
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce").map(
            lambda x: f"{x:.2f}" if pd.notna(x) else ""
        )
        df = df.rename(columns={"CoverageGroup": "Coverage"})

        _order = ["Liability", "Medical", "PIP", "UM/UIM",
                  "Collision", "Other Than Collision", "All other"]
        df["_sort"] = df["Coverage"].apply(
            lambda v: _order.index(v) if v in _order else 999
        )
        df = df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)
        return df[["Coverage", "Factor"]]

    def buildFARule231cFarm(self, company):
        # FA Rule 231c supplement — PrivatePassengerClassFactorFarm_Ext (Liability rows only).
        # Pivots Use Class (rows) × Body Style (columns) with Factor to 2 decimal places; NA if missing.
        raw = self.rateTables[company].get("PrivatePassengerClassFactorFarm_Ext")
        if raw is None:
            return pd.DataFrame(columns=["Use Class"])
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        df = df[df["Coverage Type"].astype(str).str.strip() == "Liability"].copy()
        if df.empty:
            return pd.DataFrame(columns=["Use Class"])
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce")
        pivot = df.pivot_table(
            index="Use Class", columns="Body Style", values="Factor", aggfunc="first"
        ).reset_index()
        pivot.columns.name = None
        _row_order = [
            "Not Driven to Work or School",
            "Driven to Work or School",
            "Business Use & All Other",
            "Farm Use",
        ]
        pivot["_sort"] = pivot["Use Class"].apply(
            lambda v: _row_order.index(v) if v in _row_order else 999
        )
        pivot = pivot.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)
        _col_order = [
            "Use Class",
            "Private Passenger or Station Wagon",
            "Utility Vehicle or Sports Utility Vehicle",
            "Pickup",
            "Van",
        ]
        ordered_cols = [c for c in _col_order if c in pivot.columns]
        remaining    = [c for c in pivot.columns if c not in ordered_cols]
        pivot = pivot[ordered_cols + remaining]
        for col in pivot.columns:
            if col != "Use Class":
                pivot[col] = pivot[col].apply(
                    lambda v: f"{v:.2f}" if pd.notna(v) else "NA"
                )
        return pivot

    def buildFARule223c2(self, company):
        # FA Rule 223.C.2 — TruckSecondaryClassFactorFarm_Ext
        # Logic:
        #   Truck Is A Trailer = No  → Liability, Other Than Collision, Collision columns
        #   Truck Is A Trailer = Yes, Coverage Type = Collision → Trailers Collision column
        # Secondary Class display text and Primary Class grouping come from the "FA Rule223" sheet in BA Input File.xlsx.
        _COLS = [
            "Primary Class", "Secondary Class",
            "4th-5th Digits of\nClass Code",
            "Liability", "OTC", "Collision", "Trailers\nCollision",
        ]
        _EMPTY = pd.DataFrame(columns=_COLS)

        try:
            mapping = pd.read_excel(BA_INPUT_FILE, sheet_name="FA_Rule223C2", dtype=str)
        except Exception:
            return _EMPTY

        mapping["Class Code"] = mapping["Class Code"].astype(str).str.strip().str.zfill(2)

        raw = self.rateTables[company].get("TruckSecondaryClassFactorFarm_Ext")
        if raw is None:
            return _EMPTY

        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)

        df["Secondary Class Numeric"] = pd.to_numeric(
            df["Secondary Class Numeric"], errors="coerce"
        ).apply(lambda v: str(int(v)).zfill(2) if pd.notna(v) else "")
        df["Truck Is A Trailer"] = df["Truck Is A Trailer"].fillna("").astype(str).str.strip().str.upper()
        df["Coverage Type"]      = df["Coverage Type"].fillna("").astype(str).str.strip()
        df["Factor"]             = pd.to_numeric(df["Factor"], errors="coerce")

        def _factor(code, trailer, coverage):
            sub = df[
                (df["Secondary Class Numeric"] == code) &
                (df["Truck Is A Trailer"] == trailer) &
                (df["Coverage Type"] == coverage)
            ]
            return sub["Factor"].iloc[0] if not sub.empty else None

        fmt = lambda v: f"{v:.3f}" if v is not None and pd.notna(v) else ""

        rows = []
        for _, m in mapping.iterrows():
            code = m["Class Code"]
            rows.append({
                "Primary Class":              m["Primary Class"],
                "Secondary Class":            m["Secondary Class"],
                "4th-5th Digits of\nClass Code": code,
                "Liability":                  fmt(_factor(code, "NO",  "Liability")),
                "OTC":                        fmt(_factor(code, "NO",  "Other Than Collision")),
                "Collision":                  fmt(_factor(code, "NO",  "Collision")),
                "Trailers\nCollision":        fmt(_factor(code, "YES", "Collision")),
            })

        return pd.DataFrame(rows, columns=_COLS)

    def _build_fa450_age_rows(self, company, gender_code, coverage_key):
        # Returns [(age_label, factor_str)] for one gender/coverage pair.
        # Age 0 → "17 and under"; consecutive missing ages extend the group;
        # last age → "X+"; always appends Neutral Factor and No Hit = 1.00.
        raw = self.rateTables[company].get(coverage_key)
        if raw is None:
            return [("Neutral Factor", "1.00"), ("No Hit", "1.00")]
        df = pd.DataFrame(raw[1:], columns=raw[0])
        df = df.dropna(how="all").reset_index(drop=True)
        df["GenderCode"] = df["GenderCode"].fillna("").astype(str).str.strip()
        df = df[df["GenderCode"] == gender_code].copy()
        if df.empty:
            return [("Neutral Factor", "1.00"), ("No Hit", "1.00")]
        df["_age"] = pd.to_numeric(df["Age or Experience"], errors="coerce")
        df = df.dropna(subset=["_age"]).sort_values("_age").reset_index(drop=True)
        df["Factor"] = pd.to_numeric(df["Factor"], errors="coerce")
        ages    = df["_age"].astype(int).tolist()
        factors = df["Factor"].tolist()
        rows = []
        for i, (age, factor) in enumerate(zip(ages, factors)):
            if age == 0:
                label = "17 and under"
            elif i + 1 < len(ages):
                next_age = ages[i + 1]
                label = str(age) if next_age == age + 1 else f"{age}-{next_age - 1}"
            else:
                label = f"{age}+"
            rows.append((label, f"{factor:.2f}" if pd.notna(factor) else "1.00"))
        rows.append(("Neutral Factor", "1.00"))
        rows.append(("No Hit", "1.00"))
        return rows

    def buildFARule450Violation(self, company):
        # Returns [(violations_label, liab_str, coll_str)] for the violation factor table.
        # Violations 0, 1, 2 shown individually; 3+ groups all remaining at violation=3's factor.
        def _extract(key):
            raw = self.rateTables[company].get(key)
            if raw is None:
                return {}
            df = pd.DataFrame(raw[1:], columns=raw[0])
            df = df.dropna(how="all").reset_index(drop=True)
            result = {}
            for _, row in df.iterrows():
                v = pd.to_numeric(row.get("Violations", None), errors="coerce")
                f = pd.to_numeric(row.get("Factor", None), errors="coerce")
                if pd.notna(v) and pd.notna(f):
                    result[int(v)] = f
            return result

        liab = _extract("DriverBasedRatingLiabilityViolationFactor_Ext")
        coll = _extract("DriverBasedRatingCollisionViolationFactor_Ext")
        fmt  = lambda v, d: f"{d[v]:.2f}" if v in d and pd.notna(d[v]) else "1.00"

        rows = [(str(v), fmt(v, liab), fmt(v, coll)) for v in [0, 1, 2]]
        rows.append(("3+", fmt(3, liab), fmt(3, coll)))
        rows.append(("No Hit", "1.00", "1.00"))
        rows.append(("Neutral Factor", "1.00", "1.00"))
        return rows

    # =========================================================================
    # Section B: FA-only PAGE methods
    # Add a method here when FA needs a new rule page that BA does not have.
    # Pattern:
    #   def _page_rule_fa_xxx(self, RatePages):
    #       self.compareCompanies("SomeNewFATable_Ext")
    #       for CompanyTest in self.CompanyListDif:
    #           comp_name = self.extract_company_name(CompanyTest)
    #           self.title_company_name = CompanyTest
    #           if len(self.CompanyListDif) == 1:
    #               self.title_company_name = ""
    #           RatePages.generateWorksheet(
    #               "Rule FA XXX " + self.title_company_name,
    #               "RULE FA XXX. DESCRIPTION " + self.title_company_name,
    #               "FA XXX.B. Subtitle",
    #               self.buildFAXxx(comp_name), False, True,
    #           )
    #           self.overideFooter(
    #               RatePages.getWB()["Rule FA XXX " + self.title_company_name],
    #               CompanyTest,
    #           )
    # =========================================================================

    def _page_rule_208(self, RatePages):
        """FA override — Rule 208 uses PolicyMinimumPremium_Ext instead of BA's table."""
        self.compareCompanies("PolicyMinimumPremium_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name = "Rule 208 " + self.title_company_name
            RatePages.generateWorksheet(
                ws_name,
                "RULE 208. MINIMUM PREMIUMS " + self.title_company_name,
                "208.B. Rate and Premium Computation",
                self.buildFAPolicyMinimumPremium(comp_name),
                False,  # useIndex
                False,  # useHeader — no column header row on the rate page
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)
            self.formatRule208FA(RatePages.getWB()[ws_name])

    def formatRule208FA(self, ws):
        from config.constants import CURRENCY_FORMAT
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 58
        ws.column_dimensions["B"].width = 14

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )

        for row_idx in range(4, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            if cell_a.value is not None:
                cell_a.font      = Font(name="Arial", size=10)
                cell_a.alignment = Alignment(horizontal="left", vertical="center")
                cell_a.border    = border
            cell_b = ws.cell(row=row_idx, column=2)
            if cell_b.value is not None:
                cell_b.font          = Font(name="Arial", size=10)
                cell_b.alignment     = Alignment(horizontal="center", vertical="center")
                cell_b.border        = border
                cell_b.number_format = CURRENCY_FORMAT

    def _page_rule_222e(self, RatePages):
        # FA override — uses FarmLayUpFactor_Ext (2-col) instead of AutoLayUpFactor_Ext (multi-col).
        # Subtitle changes from "Commercial Lay-up Credit" to "Farm Lay-up Credit".
        self.compareCompanies("FarmLayUpFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            tables    = [self.buildFALayupFactors(comp_name)]
            subtitles = ["222.E. Farm Lay-up Credit"]
            RatePages.generateWorksheetTablesX(
                "Rule 222 E" + self.title_company_name,
                "RULE 222.E PREMIUM DEVELOPMENT - TRUCK, TRACTOR, TRAILER TYPES " + self.title_company_name,
                subtitles, tables, False, True,
            )
            self.overideFooter(RatePages.getWB()["Rule 222 E" + self.title_company_name], CompanyTest)
        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 222 E"):
                self.format222E(_wb[_sn])

    def _page_rule_225d(self, RatePages):
        # FA override — layup table changes; fleet-size table is the same.
        self.compareCompanies("FarmLayUpFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            subtitles = [
                "225.D Fleet Size Rating Factors - Zone Rated",
                "225.D. Farm Lay-up Credit",
            ]
            tables = [
                self.buildFleetSizeRatingFactorsZone(comp_name),
                self.buildFALayupFactors(comp_name),
            ]
            title_start = "RULE 225.​​​ PREMIUM DEVELOPMENT - ZONE-RATED AUTOS "
            RatePages.generateWorksheetTablesX(
                "Rule 225.D " + self.title_company_name,
                title_start + self.title_company_name,
                subtitles, tables, False, True,
            )
            self.overideFooter(RatePages.getWB()["Rule 225.D " + self.title_company_name], CompanyTest)
        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 225.D"):
                self.format225D(_wb[_sn])

    def _page_rule_231c(self, RatePages):
        # FA override — same Class Code table as BA, plus a Farm Use Class Factors table below.
        # Extra table: Use Class × Body Style from PrivatePassengerClassFactorFarm_Ext, Liability only.
        self.compareCompanies([
            "PrivatePassengerClassCode",
            "PrivatePassengerTypesClassFactors_Ext",
            "PrivatePassengerClassFactorFarm_Ext",
        ])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name = "Rule 231 C " + self.title_company_name
            RatePages.generateWorksheet(
                ws_name,
                "RULE 231. PRIVATE PASSENGER TYPES " + self.title_company_name,
                "231.C.2.d. Use and Operator Experience Factors",
                self.build231C(comp_name),
                False, True,
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)
            ws = RatePages.getWB()[ws_name]
            self.format31C(ws)
            self.appendFARule231cFarm(ws, self.buildFARule231cFarm(comp_name))

    def appendFARule231cFarm(self, ws, farm_df):
        # Appends Use Class × Body Style data below the existing content, with a non-bold column header row.
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if farm_df.empty or len(farm_df.columns) <= 1:
            return

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )

        row = ws.max_row + 2

        num_cols = len(farm_df.columns)

        # Column header row — non-bold, wrapped so long names show on two lines
        header_row = row
        for col_idx, col_name in enumerate(farm_df.columns, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = col_name
            cell.font = Font(name="Arial", size=10, bold=False)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx > 1 else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[header_row].height = 30
        row += 1

        for _, data_row in farm_df.iterrows():
            for col_idx, val in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value  = val
                cell.font   = Font(name="Arial", size=10)
                cell.border = border
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            row += 1

        # Column widths (Use Class column needs room for long use-class names)
        ws.column_dimensions["A"].width = 28
        for col_idx in range(2, num_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22

    def _page_rule_239d(self, RatePages):
        # FA override — uses FarmLayUpFactor_Ext and Farm subtitle.
        self.compareCompanies("FarmLayUpFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            title_start = "RULE 239.​ PUBLIC AUTO PREMIUM DEVELOPMENT - OTHER THAN ZONE-RATED AUTOS "
            RatePages.generateWorksheet(
                "Rule 239 D " + self.title_company_name,
                title_start + self.title_company_name,
                "239.D. Farm Lay-up Credit",
                self.buildFALayupFactors(comp_name), False, True,
            )
            self.overideFooter(RatePages.getWB()["Rule 239 D " + self.title_company_name], CompanyTest)

    def _page_rule_241(self, RatePages):
        # FA override — layup table and subtitle change; mechanical lift table is unchanged.
        self.compareCompanies(["FarmLayUpFactor_Ext", "MechanicalLiftFactorOtherThanZoneRated"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""

            if self.StateAbb == "VA":
                subtitles = self._VA_RULE_SUBTITLES["Rule 241"]
            else:
                subtitles = [
                    "241.D.1.d Mechanical Lift Factor",
                    "241.E. Farm Lay-up Credit",
                ]

            if self.StateAbb == "VA":
                tables = [
                    self.buildMechanicalLiftFactor(comp_name),
                    self.buildSpecifiecCausesofLossCoverageFactor(comp_name),
                    self.buildFALayupFactors(comp_name),
                ]
            else:
                tables = [
                    self.buildMechanicalLiftFactor(comp_name),
                    self.buildFALayupFactors(comp_name),
                ]

            RatePages.generateWorksheetTablesX(
                "Rule 241 " + self.title_company_name,
                "RULE 241. PUBLIC AUTO PREMIUM DEVELOPMENT - ZONE-RATED AUTOS " + self.title_company_name,
                subtitles, tables, False, True,
            )
            self.overideFooter(RatePages.getWB()["Rule 241 " + self.title_company_name], CompanyTest)
        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 241"):
                self.format241(_wb[_sn])

    def _page_rule_284(self, RatePages):
        # FA override — Rule 284 reads factors from SpecialTypesMotorcycleFactor and
        # SpecialTypesMotorcycleLiabilityFactor instead of BA's hardcoded ATV/UTV data.
        self.compareCompanies(["SpecialTypesMotorcycleFactor",
                               "SpecialTypesMotorcycleLiabilityFactor"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name   = "Rule 284 " + self.title_company_name
            subtitles = ["284.C. Premium Computation"]
            tables    = [self.buildFARule284(comp_name)]
            RatePages.generateWorksheetTablesX(
                ws_name,
                "RULE 284. ALL-TERRAIN VEHICLES AND UTILITY TASK VEHICLES " + self.title_company_name,
                subtitles, tables, False, True,
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)

        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 284"):
                self.formatRule284FA(_wb[_sn])

    def formatRule284FA(self, ws):
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )
        _subtitle_vals = {"284.C. Premium Computation"}
        _header_vals   = {"Coverage", "All-terrain Vehicles", "Utility Task Vehicles"}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val == "":
                    continue
                if cell.row == 1:
                    cell.font = Font(bold=True, name="Arial", size=10)
                elif val in _subtitle_vals:
                    cell.font = Font(italic=True, name="Arial", size=10)
                elif val in _header_vals:
                    cell.font      = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                    cell.border    = border
                else:
                    cell.font   = Font(name="Arial", size=10)
                    cell.border = border
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _page_rule_303(self, RatePages):
        # FA override — Rule 303 uses FarmPollutionLiability tables (different from BA).
        # Not applicable in New York.
        if self.StateAbb == "NY":
            return

        self.compareCompanies(["FarmPollutionLiabilityBaseRate_Ext",
                               "FarmPollutionLiabilityFleetFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""

            ws_name   = "Rule 303 " + self.title_company_name
            subtitles = ["303.B.1 Premium Computation", "Liability Factor:", "Fleet Factor:"]
            tables    = [
                pd.DataFrame(),                              # separator — produces the subtitle row only
                self.buildFAPollutionBaseRate(comp_name),
                self.buildFAPollutionFleetFactor(comp_name),
            ]
            RatePages.generateWorksheetTablesX(
                ws_name,
                "RULE 303. POLLUTION LIABILITY " + self.title_company_name,
                subtitles, tables, False, True,
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)

        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 303"):
                self.formatRule303FA(_wb[_sn])

    def formatRule303FA(self, ws):
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 30

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )
        _subtitle_vals = {"303.B.1 Premium Computation", "Liability Factor:", "Fleet Factor:"}
        _header_vals   = {"Per $25,000 Limit of Liability", "Fleet Size", "Fleet Factor"}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val == "":
                    continue
                if cell.row == 1:
                    cell.font = Font(bold=True, name="Arial", size=10)
                elif val in _subtitle_vals:
                    cell.font = Font(italic=True, name="Arial", size=10)
                elif val in _header_vals:
                    cell.font      = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                    cell.border    = border
                else:
                    cell.font      = Font(name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = border

    def _page_rule_fa_223c2(self, RatePages):
        # FA-only — Rule 223.C.2 Secondary Classification Factors.
        # Reads TruckSecondaryClassFactorFarm_Ext; class groupings from "FA Rule223" sheet in BA Input File.xlsx.
        # Truck Is A Trailer=No  → Liability / OTC / Collision columns.
        # Truck Is A Trailer=Yes, Coverage=Collision → Trailers Collision column.
        self.compareCompanies("TruckSecondaryClassFactorFarm_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name = "Rule 223 C2 " + self.title_company_name
            df = self.buildFARule223c2(comp_name)
            # Blank out the "Primary Class" column header so it doesn't print as a label
            df_display = df.rename(columns={"Primary Class": ""})
            RatePages.generateWorksheet(
                ws_name,
                "RULE 223. TRUCKS, TRACTORS, TRAILERS CLASSIFICATION " + self.title_company_name,
                "223.C.2. Secondary Classification Factors",
                df_display,
                False, True,
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)

        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 223 C2"):
                self.formatRule223c2FA(_wb[_sn])

    def formatRule223c2FA(self, ws):
        # generateWorksheet layout (subtitle != ' '):
        #   row 1 = title, row 2 = subtitle, row 3 = blank, row 4 = headers, row 5+ = data
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 12   # Primary Class — narrower, wraps to multi-line
        ws.column_dimensions["B"].width = 58   # Secondary Class — wider so text fits single line
        ws.column_dimensions["C"].width = 10   # 4th-5th Digits of Class Code
        ws.column_dimensions["D"].width = 12   # Liability
        ws.column_dimensions["E"].width = 11   # OTC
        ws.column_dimensions["F"].width = 12   # Collision
        ws.column_dimensions["G"].width = 13   # Trailers Collision

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )

        ws.row_dimensions[3].height = 8    # small gap between subtitle and header
        ws.row_dimensions[4].height = 28   # column header row — 2-line headers at 10pt bold need ~28pt

        DATA_START = 5
        if ws.max_row < DATA_START:
            return

        import math
        # Col B width is 58 units ≈ 75 chars at Arial 8pt; each line is 11pt tall.
        _CHARS_PER_LINE = 75
        _LINE_HEIGHT_PT = 11

        # First pass: format all data cells and collect Primary Class group ranges
        groups = []
        current_group = None
        group_start   = None

        for row_idx in range(DATA_START, ws.max_row + 1):
            group_val = ws.cell(row=row_idx, column=1).value

            if group_val and group_val != current_group:
                if current_group is not None:
                    groups.append((group_start, row_idx - 1, current_group))
                current_group = group_val
                group_start   = row_idx

            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx == 1:
                    cell.font      = Font(bold=True, name="Arial", size=8)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif col_idx == 2:
                    cell.font      = Font(name="Arial", size=8)
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.font      = Font(name="Arial", size=8)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Dynamic row height: expand when Secondary Class text wraps to multiple lines.
            text_b   = str(ws.cell(row=row_idx, column=2).value or "")
            num_lines = max(1, math.ceil(len(text_b) / _CHARS_PER_LINE))
            ws.row_dimensions[row_idx].height = num_lines * _LINE_HEIGHT_PT

        if current_group is not None:
            groups.append((group_start, ws.max_row, current_group))

        # Second pass: merge Primary Class column for each contiguous group
        for (start_r, end_r, label) in groups:
            if end_r > start_r:
                ws.merge_cells(start_row=start_r, start_column=1,
                               end_row=end_r,   end_column=1)
            top = ws.cell(row=start_r, column=1)
            top.value     = label
            top.font      = Font(bold=True, name="Arial", size=8)
            top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _page_rule_fa_222d(self, RatePages):
        # FA-only — Rule 222.D Unlicensed/Limited Use Farm or Ranch Trucks.
        # Only applicable in MD, PA, NY, VA, WV.
        if self.StateAbb not in {"MD", "PA", "NY", "VA", "WV"}:
            return

        self.compareCompanies("UnlicensedFarmTruckFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name = "Rule 222 D " + self.title_company_name
            RatePages.generateWorksheet(
                ws_name,
                "RULE 222.D. TRUCKS, TRACTORS, TRAILERS CLASSIFICATION " + self.title_company_name,
                "222.D. Unlicensed/Limited Use Farm or Ranch Trucks",
                self.buildFARule222D(comp_name),
                False,  # useIndex
                True,   # useHeader
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)

        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 222 D"):
                self.formatRule222DFA(_wb[_sn])

    def formatRule222DFA(self, ws):
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 14

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )
        _header_vals = {"Coverage", "Factor"}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val == "":
                    continue
                if cell.row == 1:
                    cell.font = Font(bold=True, name="Arial", size=10)
                elif val.startswith("222.D"):
                    cell.font = Font(italic=True, name="Arial", size=10)
                elif val in _header_vals:
                    cell.font      = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                    cell.border    = border
                else:
                    cell.font   = Font(name="Arial", size=10)
                    cell.border = border
                    if cell.column == 1:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _page_rule_fa_420(self, RatePages):
        # FA-only — Rule 420 Segmentation Rating Plan.
        # Exception states (CA, MT, NY, UT, WA): Implementation Factor is a cross-product
        # matrix of TieringFactor_Ext × RetentionFactor_Ext.
        # All other states: single Grade | Loss Factor column (TieringFactor_Ext only).
        # Mitigation Factors table is the same for all states.
        _EXCEPTION_STATES = {"CA", "MT", "NY", "UT", "WA"}

        self.compareCompanies(["TieringFactor_Ext", "TieringMitigationFactor_Ext",
                               "RetentionFactor_Ext"])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""

            ws_name   = "Rule 420 " + self.title_company_name
            subtitles = ["420.C.1. Implementation Factor", "420.C.2. Mitigation Factors"]

            if self.StateAbb in _EXCEPTION_STATES:
                impl_table = self.buildFATieringFactorMatrix(comp_name)
            else:
                impl_table = self.buildFATieringFactor(comp_name)

            tables = [impl_table, self.buildFAMitigationFactor(comp_name)]
            RatePages.generateWorksheetTablesX(
                ws_name,
                "RULE 420. SEGMENTATION RATING PLAN " + self.title_company_name,
                subtitles, tables, False, True,
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)

        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 420"):
                ws = _wb[_sn]
                if ws.max_column > 3:
                    self.formatRule420FAMatrix(ws)
                else:
                    self.formatRule420FA(ws)

    def formatRule420FA(self, ws):
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )
        _subtitle_vals = {"420.C.1. Implementation Factor", "420.C.2. Mitigation Factors"}
        _header_vals   = {"Grade", "Loss Factor", "Number of Vehicles", "Factor"}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val == "":
                    continue
                if cell.row == 1:
                    cell.font = Font(bold=True, name="Arial", size=10)
                elif val in _subtitle_vals:
                    cell.font = Font(italic=True, name="Arial", size=10)
                elif val in _header_vals:
                    cell.font      = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                    cell.border    = border
                else:
                    cell.font      = Font(name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = border

    def formatRule420FAMatrix(self, ws):
        # Formatter for exception-state Rule 420: Loss Grade × Retention Grade matrix.
        # Header rows are detected by checking column A for "Loss Grade" or "Number of Vehicles",
        # so the single-letter Retention Grade column headers (D, I, M, Q, U) are automatically
        # made bold along with their anchor cell without value-based conflicts.
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        ws.column_dimensions["A"].width = 12
        for col_idx in range(2, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 10

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )
        _subtitle_vals = {"420.C.1. Implementation Factor", "420.C.2. Mitigation Factors"}

        # First pass: find header rows by anchor value in column A
        header_rows = set()
        for row in ws.iter_rows():
            if row[0].value in ("Loss Grade", "Number of Vehicles"):
                header_rows.add(row[0].row)

        # Second pass: apply formatting
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val == "":
                    continue
                if cell.row == 1:
                    cell.font = Font(bold=True, name="Arial", size=10)
                elif val in _subtitle_vals:
                    cell.font = Font(italic=True, name="Arial", size=10)
                elif cell.row in header_rows:
                    cell.font      = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                    cell.border    = border
                else:
                    cell.font      = Font(name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = border

    def _page_rule_fa_440(self, RatePages):
        # FA-only — Rule 440 Multiple Policy Discount.
        # Shows the TRUE row factor alongside a fixed description line.
        self.compareCompanies("MultiplePolicyDiscountFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name = "Rule 440 " + self.title_company_name
            RatePages.generateWorksheet(
                ws_name,
                "RULE 440. MULTIPLE POLICY DISCOUNT " + self.title_company_name,
                "440.B.",
                self.buildFAMultiplePolicyDiscount(comp_name),
                False,  # useIndex
                False,  # useHeader
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)
            self.formatRule440FA(RatePages.getWB()[ws_name])

    def formatRule440FA(self, ws):
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 58
        ws.column_dimensions["B"].width = 14

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )

        for row_idx in range(4, ws.max_row + 1):
            cell_a = ws.cell(row=row_idx, column=1)
            if cell_a.value is not None:
                cell_a.font      = Font(name="Arial", size=10)
                cell_a.alignment = Alignment(horizontal="left", vertical="center")
                cell_a.border    = border
            cell_b = ws.cell(row=row_idx, column=2)
            if cell_b.value is not None:
                cell_b.font      = Font(name="Arial", size=10)
                cell_b.alignment = Alignment(horizontal="center", vertical="center")
                cell_b.border    = border

    def _page_rule_fa_455(self, RatePages):
        # FA-only — Rule 455 Risk Score Factor.
        # Not applicable in CA, MT, NY, UT, WA.
        if self.StateAbb in ("CA", "MT", "NY", "UT", "WA"):
            return

        self.compareCompanies("RiskScoreFactor_Ext")
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""
            ws_name = "Rule 455 " + self.title_company_name
            RatePages.generateWorksheet(
                ws_name,
                "RULE 455. RISK SCORE FACTOR " + self.title_company_name,
                "455.A Risk Score Factors",
                self.buildFARiskScoreFactor(comp_name),
                False,  # useIndex
                True,   # useHeader
            )
            self.overideFooter(RatePages.getWB()[ws_name], CompanyTest)

        _wb = RatePages.getWB()
        for _sn in _wb.sheetnames:
            if _sn.startswith("Rule 455"):
                self.formatRule455FA(_wb[_sn])

    def formatRule455FA(self, ws):
        from openpyxl.styles import Font, Alignment, Border, Side

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 14

        border = Border(
            left=Side(border_style="thin", color="C1C1C1"),
            right=Side(border_style="thin", color="C1C1C1"),
            top=Side(border_style="thin", color="C1C1C1"),
            bottom=Side(border_style="thin", color="C1C1C1"),
        )
        _header_vals = {"Financial Reporting Class", "Factor"}

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value)
                if val == "":
                    continue
                if cell.row == 1:
                    cell.font = Font(bold=True, name="Arial", size=10)
                elif val == "455.A Risk Score Factors":
                    cell.font = Font(italic=True, name="Arial", size=10)
                elif val in _header_vals:
                    cell.font      = Font(bold=True, name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="bottom", wrap_text=True)
                    cell.border    = border
                else:
                    cell.font      = Font(name="Arial", size=10)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = border

    def _page_rule_450(self, RatePages):
        # FA override — Driver Based Rating Plan with 5 tables on one sheet.
        # Page 1: Male Liability | Male Collision (side by side)
        # Page 2: Female Liability | Female Collision (side by side) + Violation table below
        # Page break is inserted after the last Male data row.
        self.compareCompanies([
            "DriverBasedRatingLiabilityAgeGenderFactor_Ext",
            "DriverBasedRatingCollisionAgeGenderFactor_Ext",
            "DriverBasedRatingLiabilityViolationFactor_Ext",
            "DriverBasedRatingCollisionViolationFactor_Ext",
        ])
        for CompanyTest in self.CompanyListDif:
            comp_name = self.extract_company_name(CompanyTest)
            self.title_company_name = CompanyTest
            if len(self.CompanyListDif) == 1:
                self.title_company_name = ""

            ws_name = "Rule 450 " + self.title_company_name
            wb  = RatePages.getWB()
            ws  = wb.create_sheet(ws_name)
            RatePages._apply_page_header_footer(ws)

            from openpyxl.styles import Font, Alignment, Border, Side

            thin   = Side(border_style="thin", color="C1C1C1")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _cell(r, c, val="", bold=False, italic=False,
                      center=False, wrap=False, brdr=False, size=10):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = Font(name="Arial", size=size, bold=bold, italic=italic)
                cell.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="center", wrap_text=wrap,
                )
                if brdr:
                    cell.border = border
                return cell

            # ── Rows 1-3: print-title area ─────────────────────────────────
            title = ("RULE 450.  DRIVER BASED RATING PLAN " + self.title_company_name).strip()
            _cell(1, 1, title, bold=True)          # row 1 repeats on every page
            # rows 2-3 blank — collapsed to avoid large gap on page 1
            ws.row_dimensions[2].height = 3
            ws.row_dimensions[3].height = 3

            # ── Rows 4-6: page-1-only intro ────────────────────────────────
            _cell(4, 1, "450.B.5. Premium Computation")
            _cell(5, 1,
                  "Use the neutral factor for driver characteristics for large "
                  "fleets (100+ vehicles)",
                  italic=True)

            # ── Row 7: section heading ──────────────────────────────────────
            _cell(7, 1, "450.B.1.a. Age/Gender/Coverage Factor")

            # ── Row 8: gender header (Male | Male) ─────────────────────────
            _cell(8, 1, "Male", bold=True, center=True)
            ws.merge_cells(start_row=8, start_column=1, end_row=8, end_column=2)
            _cell(8, 4, "Male", bold=True, center=True)
            ws.merge_cells(start_row=8, start_column=4, end_row=8, end_column=5)

            # ── Row 9: column headers ───────────────────────────────────────
            _cell(9, 1, "Age",       bold=True, center=True, brdr=True)
            _cell(9, 2, "Liability", bold=True, center=True, brdr=True)
            _cell(9, 4, "Age",       bold=True, center=True, brdr=True)
            _cell(9, 5, "Collision", bold=True, center=True, brdr=True)

            # ── Rows 10+: Male data ─────────────────────────────────────────
            male_liab = self._build_fa450_age_rows(
                comp_name, "M", "DriverBasedRatingLiabilityAgeGenderFactor_Ext")
            male_coll = self._build_fa450_age_rows(
                comp_name, "M", "DriverBasedRatingCollisionAgeGenderFactor_Ext")

            row = 10
            for i in range(max(len(male_liab), len(male_coll))):
                if i < len(male_liab):
                    _cell(row, 1, male_liab[i][0], center=False, brdr=True, size=7)
                    _cell(row, 2, male_liab[i][1], center=True,  brdr=True, size=7)
                if i < len(male_coll):
                    _cell(row, 4, male_coll[i][0], center=False, brdr=True, size=7)
                    _cell(row, 5, male_coll[i][1], center=True,  brdr=True, size=7)
                ws.row_dimensions[row].height = 11
                row += 1

            # ── Female section (page 2) ──────────────────────────────────────
            # (page break inserted by FApagebreaks._handle_fa_rule_450)
            row += 1   # one blank row separating Male and Female sections

            _cell(row, 1, "450.B.1.a. Age/Gender/Coverage Factor")
            row += 1

            _cell(row, 1, "Female", bold=True, center=True)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            _cell(row, 4, "Female", bold=True, center=True)
            ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=5)
            row += 1

            _cell(row, 1, "Age",       bold=True, center=True, brdr=True)
            _cell(row, 2, "Liability", bold=True, center=True, brdr=True)
            _cell(row, 4, "Age",       bold=True, center=True, brdr=True)
            _cell(row, 5, "Collision", bold=True, center=True, brdr=True)
            row += 1

            female_liab = self._build_fa450_age_rows(
                comp_name, "F", "DriverBasedRatingLiabilityAgeGenderFactor_Ext")
            female_coll = self._build_fa450_age_rows(
                comp_name, "F", "DriverBasedRatingCollisionAgeGenderFactor_Ext")

            for i in range(max(len(female_liab), len(female_coll))):
                if i < len(female_liab):
                    _cell(row, 1, female_liab[i][0], center=False, brdr=True, size=7)
                    _cell(row, 2, female_liab[i][1], center=True,  brdr=True, size=7)
                if i < len(female_coll):
                    _cell(row, 4, female_coll[i][0], center=False, brdr=True, size=7)
                    _cell(row, 5, female_coll[i][1], center=True,  brdr=True, size=7)
                ws.row_dimensions[row].height = 11
                row += 1

            # ── Violation table ─────────────────────────────────────────────
            row += 1   # blank separator

            _cell(row, 1, "450.B.1.b. Violation Factor")
            row += 1

            # "Factor" header spanning Liability and Collision columns
            _cell(row, 2, "Factor", bold=True, center=True)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1

            _cell(row, 1, "Total Violations", bold=True, center=True, brdr=True)
            _cell(row, 2, "Liability",         bold=True, center=True, brdr=True)
            _cell(row, 3, "Collision",         bold=True, center=True, brdr=True)
            row += 1

            for (label, lf, cf) in self.buildFARule450Violation(comp_name):
                _cell(row, 1, label, center=True, brdr=True, size=7)
                _cell(row, 2, lf,    center=True, brdr=True, size=7)
                _cell(row, 3, cf,    center=True, brdr=True, size=7)
                ws.row_dimensions[row].height = 11
                row += 1

            # ── Column widths ───────────────────────────────────────────────
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 14
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 20
            ws.column_dimensions["E"].width = 14

            self.overideFooter(ws, CompanyTest)

    # =========================================================================
    # Section C: Main FA page generator
    # =========================================================================

    def buildFAPages(self):
        """
        Farm Auto rate page sequence.

        This is the master list of rules for FA.
        - To ADD a new FA-only rule:   add self._page_rule_fa_xxx(RatePages) here.
        - To REMOVE a rule for FA:     comment out its line here.
        - To REORDER rules:            move its line up or down.
        - To CHANGE a rule's data:     override the buildXxx() method in Section A.

        Right now FA uses the exact same set of rules as BA, so this method
        delegates to buildBAPages(). Once FA rules diverge, replace this with
        an explicit rule list (see DEVELOPER_GUIDE.md for the full pattern).
        """
        # FA uses its own Excel factory so FA-specific generate/format
        # methods added to ExcelSettingsFA.Excel will be picked up automatically.
        from . import ExcelSettingsFA

        # Build the company list (same logic as BA)
        companies = [
            c for c in self.rateTables.keys()
            if c not in ("CW", "MM") and self.rateTables[c] is not None
        ]

        RatePages = ExcelSettingsFA.Excel(
            StateAbb=self.StateAbb,
            State=self.State,
            nEffective=self.nEffective,
            rEffective=self.rEffective,
            companyList=companies,
        )

        # Run the nesting + LCM protocol
        self.nesting()

        # Load state-specific sheet exceptions
        state_sheet_exceptions = pd.read_excel(
            BA_INPUT_FILE, sheet_name=None, engine="openpyxl"
        )

        # State-specific warnings
        if self.StateAbb == "MT":
            print("Warning: Rule 297 will not be correct.")
        if self.StateAbb == "DC":
            print("Warning: Additional PIP Base Rate Tables for 222, 232, 239 will be absent.")
        if self.StateAbb == "MI":
            print("Warning: Base Rate Formatting incomplete. 298 has special exceptions not yet built.")
        if self.StateAbb == "VA":
            print("Warning: Due to large shifts in manual presentation, this manual is incomplete.")
        if self.StateAbb in ("NY", "CA"):
            print("Warning: Rule 297 for this state was not built out.")

        # shared dict passed to rules 293, 297, and 451 (they share state)
        shared = {}

        # ── Rules ─────────────────────────────────────────────────────────────
        # Comment out any rule not needed for Farm Auto.
        # Add FA-only rules at the bottom (self._page_rule_fa_xxx).
        self._page_rule_vapcd(RatePages, state_sheet_exceptions)
        self._page_rule_208(RatePages)
        self._page_rule_222_ttt_base_rates(RatePages, state_sheet_exceptions)
        self._page_rule_222b(RatePages)
        self._page_rule_222c(RatePages)
        self._page_rule_fa_222d(RatePages)   # FA-only: 222.D Unlicensed/Limited Use Farm or Ranch Trucks (MD, PA, NY, VA, WV only)
        self._page_rule_222e(RatePages)
        self._page_rule_223b5(RatePages)
        self._page_rule_fa_223c2(RatePages)   # FA-only: 223.C.2 Secondary Classification Factors
        self._page_rule_225c2(RatePages)
        self._page_rule_225c3(RatePages)
        self._page_rule_225_zone_br(RatePages)
        self._page_rule_225d(RatePages)
        self._page_rule_231c(RatePages)
        self._page_rule_232_ppt_base_rates(RatePages, state_sheet_exceptions)
        self._page_rule_232b(RatePages)
        self._page_rule_233(RatePages)
        self._page_rule_239_school_bus_br(RatePages, state_sheet_exceptions)
        self._page_rule_239_other_bus_br(RatePages, state_sheet_exceptions)
        self._page_rule_239_van_pool_br(RatePages, state_sheet_exceptions)
        self._page_rule_239_taxi_br(RatePages, state_sheet_exceptions)
        self._page_rule_239c(RatePages)
        self._page_rule_239d(RatePages)
        self._page_rule_240(RatePages)
        self._page_rule_241(RatePages)
        self._page_rule_243(RatePages)
        self._page_rule_255(RatePages)
        self._page_rule_264(RatePages)
        self._page_rule_266(RatePages)
        self._page_rule_267(RatePages)
        self._page_rule_268(RatePages)
        self._page_rule_269(RatePages)
        self._page_rule_271(RatePages)
        self._page_rule_272(RatePages)
        self._page_rule_273(RatePages)
        self._page_rule_274(RatePages)
        self._page_rule_275(RatePages)
        self._page_rule_276(RatePages)
        self._page_rule_277(RatePages)
        self._page_rule_278(RatePages)
        self._page_rule_279(RatePages)
        self._page_rule_280(RatePages)
        self._page_rule_281(RatePages)
        self._page_rule_283(RatePages)
        self._page_rule_284(RatePages)
        self._page_rule_288(RatePages)
        self._page_rule_289(RatePages)
        self._page_rule_290(RatePages)
        self._page_rule_292(RatePages)
        self._page_rule_293(RatePages, shared)
        self._page_rule_294(RatePages)
        self._page_rule_295(RatePages)
        self._page_rule_296(RatePages)
        self._page_rule_297(RatePages, shared)
        self._page_rule_298(RatePages)
        self._page_rule_300(RatePages)
        self._page_rule_301c(RatePages)
        self._page_rule_301d1(RatePages)
        self._page_rule_301d2(RatePages)
        self._page_rule_303(RatePages)
        self._page_rule_305(RatePages)
        self._page_rule_306(RatePages)
        self._page_rule_307(RatePages)
        self._page_rule_309(RatePages)
        self._page_rule_310(RatePages)
        self._page_rule_313(RatePages)
        self._page_rule_315(RatePages)
        self._page_rule_317(RatePages)
        self._page_rule_416(RatePages)
        self._page_rule_417(RatePages)
        self._page_rule_fa_420(RatePages)
        self._page_rule_425(RatePages)
        self._page_rule_426(RatePages)
        self._page_rule_427(RatePages)
        self._page_rule_fa_440(RatePages)
        self._page_rule_450(RatePages)
        # self._page_rule_451(RatePages, shared)
        self._page_rule_452(RatePages)
        self._page_rule_453(RatePages)
        self._page_rule_454(RatePages)
        self._page_rule_fa_455(RatePages)
        # self._page_rule_dp1(RatePages)
        self._page_rule_state_specific(RatePages)

        # ── FA-only rules (add here when Farm Auto has unique rules) ──────────
        # self._page_rule_fa_farm_machinery(RatePages)   # example

        RatePages.createIndex()

        if self.StateAbb == "FL":
            self.overideHeaderFL(RatePages.getWB())

        return RatePages.getWB()