"""
BApagebreaks.py
================
Post-processing for the BA Rate Pages workbook.

Pipeline:
  1. openpyxl: truncate sheet names > 31 chars, move Index to position 0.
  2. Excel COM: open the file, apply per-rule page breaks/print settings,
     activate Index, SaveAs a fresh xlsx, replace the original.

Page breaks and print settings are applied via Excel COM (not openpyxl) so the
final file is in Excel's canonical XML format. This eliminates the
"We found a problem with some content..." popup that appears when openpyxl
writes the page-setup attributes itself.

----------------------------------------------------------------------------
HOW TO ADD A PAGE-BREAK RULE FOR A NEW SHEET
----------------------------------------------------------------------------

Step 1 - Write a handler function. Signature: (ws, app, dest_filename) -> None
         where `ws` is a COM Worksheet, `app` is the Excel.Application,
         and `dest_filename` is the absolute file path.

    def _handle_rule_999(ws, app, dest_filename):
        max_row = ws.UsedRange.Rows.Count
        ws.PageSetup.PrintArea = f"$A$1:$H${max_row}"
        disable_fit_to_page(ws)
        add_break_after(ws, 37)

Step 2 - Register it in SHEET_RULES below (one line):

    ("Rule 999", _handle_rule_999),

Order matters: more-specific prefixes BEFORE less-specific ones
(e.g., "Rule 239 C" must come before "Rule 239 ").

----------------------------------------------------------------------------
HELPERS YOU CAN USE INSIDE A HANDLER
----------------------------------------------------------------------------

  fit_single_page(ws)         entire sheet onto one printed page
  fit_width_only(ws)          fit width to 1 page; height grows with content
  disable_fit_to_page(ws)     turn off fit-to-page; manual breaks rule
  add_break_after(ws, row)    add a horizontal page break AFTER the given row

COM API CHEATSHEET
  ws.PageSetup.PrintArea       = "$A$1:$H$N"
  ws.PageSetup.PrintTitleRows  = "$1:$3"     (or "" to clear)
  ws.PageSetup.Orientation     = XL_LANDSCAPE   (= 2)
  ws.PageSetup.TopMargin       = app.InchesToPoints(1.00)
  ws.PageSetup.CenterHorizontally = False
  ws.UsedRange.Rows.Count                       row count
  ws.Range("A10").Value                         read a cell
  ws.Range(f"A1:A{n}").Value                    bulk read (faster than per-row)
"""

import gc
import os
import time

import openpyxl

try:
    import pythoncom
    import win32com.client
except ImportError:
    pythoncom = None
    win32com = None


# ============================================================================
#  CONSTANTS
# ============================================================================

XL_LANDSCAPE         = 2     # xlLandscape
XL_OPEN_XML_WORKBOOK = 51    # xlOpenXMLWorkbook (.xlsx)
XL_REPAIR_FILE       = 1     # CorruptLoad=1 silently repairs on open


# ============================================================================
#  HELPERS  -  use these inside rule handlers
# ============================================================================

def fit_single_page(ws):
    """Fit the entire content of the sheet onto a single printed page."""
    ws.PageSetup.Zoom = False           # False enables FitToPages mode
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = 1


def fit_width_only(ws):
    """Fit width to 1 page; height grows as needed (manual breaks honored)."""
    ws.PageSetup.Zoom = False
    ws.PageSetup.FitToPagesWide = 1
    ws.PageSetup.FitToPagesTall = False


def disable_fit_to_page(ws):
    """Turn off fit-to-page entirely; manual breaks are honored as-is."""
    ws.PageSetup.Zoom = 100             # any int disables fit-to-page mode


def add_break_after(ws, row):
    """Add a horizontal page break AFTER the given row (1-indexed)."""
    ws.HPageBreaks.Add(ws.Rows(row + 1))


# ============================================================================
#  RULE HANDLERS  (COM)
#  Signature: (ws, app, dest_filename) -> None
# ============================================================================

def _handle_index(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintTitleRows = ""    # no title rows on the index
    ws.PageSetup.PrintArea = f"$A$1:$J${max_row}"
    fit_width_only(ws)


def _handle_rule_222b(ws, app, dest_filename):
    disable_fit_to_page(ws)
    add_break_after(ws, 25)
    add_break_after(ws, 49)


def _handle_rule_222ttt(ws, app, dest_filename):
    fit_single_page(ws)
    ws.PageSetup.PrintTitleRows = "$1:$3"


def _handle_rule_223b5(ws, app, dest_filename):
    ws.PageSetup.Orientation = XL_LANDSCAPE


def _handle_rule_223c(ws, app, dest_filename):
    fit_single_page(ws)


def _handle_rule_225_zone(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$M${max_row}"
    ws.PageSetup.CenterHorizontally = False
    ws.PageSetup.CenterVertically = False
    disable_fit_to_page(ws)
    for row in range(52, max_row, 51):
        add_break_after(ws, row)


def _handle_rule_225c3(ws, app, dest_filename):
    fit_single_page(ws)


def _handle_rule_232ppt(ws, app, dest_filename):
    fit_single_page(ws)
    ws.PageSetup.PrintTitleRows = "$1:$3"


def _handle_rule_239_general(ws, app, dest_filename):
    fit_single_page(ws)
    ws.PageSetup.PrintTitleRows = "$1:$3"


def _handle_rule_239c(ws, app, dest_filename):
    fit_single_page(ws)
    ws.PageSetup.TopMargin = app.InchesToPoints(1.00)


def _handle_rule_240(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    fit_single_page(ws)
    ws.PageSetup.CenterVertically = True
    ws.PageSetup.PrintTitleRows = "$1:$3"
    ws.PageSetup.PrintArea = f"$A$1:$M${max_row}"
    ws.PageSetup.TopMargin = app.InchesToPoints(1.00)


def _handle_rule_255(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$H${max_row}"
    ws.PageSetup.CenterHorizontally = False
    ws.PageSetup.CenterVertically = False
    disable_fit_to_page(ws)
    add_break_after(ws, 37)


def _handle_rule_275(ws, app, dest_filename):
    if ws.Range("A10").Value == "275.B.1.(b). Short Term - Autos Leased by the Hour, Day, or Week":
        ws.PageSetup.PrintTitleRows = "$1:$1"
        fit_single_page(ws)


def _handle_rule_283(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$P${max_row}"
    targets = {
        "283.B Limited Specified Causes of Loss",
        "283.B Comprehensive",
        "283.B Blanket Collision",
    }
    col_a = ws.Range(f"A1:A{max_row}").Value   # bulk read
    for row in range(1, max_row + 1):
        cell_value = str(col_a[row - 1][0])
        if cell_value in targets and row > 3:
            # Break BEFORE this row (= after row-1)
            ws.HPageBreaks.Add(ws.Rows(row))
    fit_width_only(ws)


def _handle_rule_289(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$H${max_row}"
    disable_fit_to_page(ws)
    add_break_after(ws, 37)


def _handle_rule_297(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$P${max_row}"
    disable_fit_to_page(ws)
    col_a = ws.Range(f"A1:A{max_row}").Value
    occurrence_count = 0
    for row in range(1, max_row + 1):
        cell_value = str(col_a[row - 1][0])
        if cell_value.startswith("Single") or cell_value.startswith("Uninsured"):
            occurrence_count += 1
        if (occurrence_count % 3 == 0) and (occurrence_count != 0):
            occurrence_count += 1
            ws.HPageBreaks.Add(ws.Rows(row))


def _handle_rule_298(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$K${max_row}"
    disable_fit_to_page(ws)
    col_a = ws.Range(f"A1:A{max_row}").Value
    occurrence_count = 0
    for row in range(1, max_row + 1):
        cell_value = str(col_a[row - 1][0])
        if cell_value.startswith("298"):
            occurrence_count += 1
        if occurrence_count == 4:
            occurrence_count += 1
            ws.HPageBreaks.Add(ws.Rows(row))
        if occurrence_count == 8:
            break


_VA_VEHICLE_TYPES = {
    "Extra Heavy Truck-Tractor", "Extra-Heavy Truck", "Heavy Truck",
    "Heavy Truck-Tractor", "Light Truck", "Medium Truck",
    "Private Passenger Types", "Semitrailer",
    "Service or Utility Trailer", "Trailer",
}


def _handle_rule_301ab(ws, app, dest_filename):
    if ws.Range("B4").Value in _VA_VEHICLE_TYPES:
        return
    max_row = ws.UsedRange.Rows.Count
    fit_width_only(ws)
    if ws.Name.startswith("Rule 301.B"):
        ws.PageSetup.PrintArea = f"$A$1:$T${max_row}"
    for row in range(46, max_row, 45):
        add_break_after(ws, row)
    ws.PageSetup.CenterHorizontally = False
    ws.PageSetup.CenterVertically = False
    ws.PageSetup.Orientation = XL_LANDSCAPE
    ws.PageSetup.TopMargin = app.InchesToPoints(1.00)


def _handle_rule_301cd(ws, app, dest_filename):
    if ws.Range("B4").Value not in _VA_VEHICLE_TYPES:
        return
    if "FL" not in dest_filename:
        ws.PageSetup.TopMargin = app.InchesToPoints(1.00)
    fit_single_page(ws)


def _handle_rule_306(ws, app, dest_filename):
    fit_width_only(ws)
    ws.PageSetup.PrintTitleRows = "$1:$4"


def _handle_rule_315(ws, app, dest_filename):
    fit_width_only(ws)
    add_break_after(ws, 23)


def _handle_rule_r1(ws, app, dest_filename):
    max_row = ws.UsedRange.Rows.Count
    ws.PageSetup.PrintArea = f"$A$1:$M${max_row}"
    disable_fit_to_page(ws)
    col_a = ws.Range(f"A1:A{max_row}").Value
    occurrence_count = 0
    for row in range(1, max_row + 1):
        cell_value = str(col_a[row - 1][0])
        if cell_value.startswith("R1"):
            occurrence_count += 1
        if occurrence_count in (3, 6):
            occurrence_count += 1
            ws.HPageBreaks.Add(ws.Rows(row))


# ============================================================================
#  RULE REGISTRY
#  Order matters: more-specific prefixes BEFORE less-specific ones.
#  To add a new rule: write the handler above, then add one line here.
# ============================================================================

SHEET_RULES = [
    ("Index",        _handle_index),

    ("Rule 222 B",   _handle_rule_222b),
    ("Rule 222 TTT", _handle_rule_222ttt),

    ("Rule 223 B.5", _handle_rule_223b5),
    ("Rule 223 C",   _handle_rule_223c),

    ("Rule 225 Zone", _handle_rule_225_zone),
    ("Rule 225.C.3", _handle_rule_225c3),

    ("Rule 232 PPT", _handle_rule_232ppt),

    ("Rule 239 C",   _handle_rule_239c),       # specific BEFORE generic
    ("Rule 239 ",    _handle_rule_239_general),

    ("Rule 240 ",    _handle_rule_240),

    ("Rule 255",     _handle_rule_255),

    ("Rule 275",     _handle_rule_275),

    ("Rule 283",     _handle_rule_283),
    ("Rule 289",     _handle_rule_289),
    ("Rule 297",     _handle_rule_297),
    ("Rule 298",     _handle_rule_298),

    ("Rule 301.C",   _handle_rule_301cd),
    ("Rule 301.D",   _handle_rule_301cd),
    ("Rule 301.A",   _handle_rule_301ab),
    ("Rule 301.B",   _handle_rule_301ab),

    ("Rule 306",     _handle_rule_306),
    ("Rule 315",     _handle_rule_315),

    ("Rule R1",      _handle_rule_r1),
]


def _apply_matching_rule(sheet_name, ws, app, dest_filename):
    """Walk SHEET_RULES and run the first handler whose prefix matches."""
    for prefix, handler in SHEET_RULES:
        if sheet_name.startswith(prefix):
            handler(ws, app, dest_filename)
            return True
    return False


# ============================================================================
#  PUBLIC ENTRY POINT
# ============================================================================

def process_pagebreaks(dest_filename1, dest_filename2):
    """
    Apply page breaks / print settings to dest_filename1 and produce a clean
    xlsx (no "Open and Repair" popup on open).

    dest_filename2 is accepted for backward compatibility (was a PDF path).
    """
    print(f"[BApagebreaks] Processing: {dest_filename1}")
    dest_filename1 = os.path.normpath(os.path.abspath(dest_filename1))

    # ── Phase 1: openpyxl bookkeeping ──────────────────────────────────────
    workbook = openpyxl.load_workbook(dest_filename1)

    for original_name in list(workbook.sheetnames):
        if len(original_name) > 31:
            workbook[original_name].title = original_name[:31]

    if "Index" in workbook.sheetnames:
        ws_index = workbook["Index"]
        ws_index.sheet_state = "visible"
        if workbook.sheetnames.index("Index") != 0:
            workbook._sheets.remove(ws_index)
            workbook._sheets.insert(0, ws_index)
        workbook.active = 0

    workbook.save(dest_filename1)
    workbook.close()

    # ── Phase 2: Excel COM applies page breaks and writes a clean file ─────
    print("[BApagebreaks] Applying page breaks via Excel COM...")
    _apply_via_com(dest_filename1)
    print("[BApagebreaks] Done.")


# ============================================================================
#  EXCEL COM IMPLEMENTATION
# ============================================================================

def _apply_via_com(filename):
    """
    Open the workbook through Excel COM, apply page-break rules, and SaveAs
    a clean xlsx that replaces the original.

    Why CorruptLoad=1 + SaveAs (instead of plain Open + Save):
      - openpyxl's xlsx output sometimes triggers Excel's "Open and Repair"
        prompt because of minor XML quirks.
      - CorruptLoad=1 (xlRepairFile) silently fixes those quirks during open.
      - SaveAs writes the workbook from Excel's in-memory model, producing
        canonical XML that opens without any prompt.
      - Page breaks are added AFTER the repair, so they survive (the repair
        pass would have stripped any breaks openpyxl had written).
    """
    if win32com is None:
        print("[BApagebreaks] win32com not available; skipping COM step.")
        return

    if pythoncom:
        pythoncom.CoInitialize()

    xl_app = None
    xl_book = None
    temp_filename = filename + ".clean.xlsx"

    try:
        xl_app = win32com.client.DispatchEx("Excel.Application")
        xl_app.Visible = False
        xl_app.DisplayAlerts = False
        xl_app.AskToUpdateLinks = False

        xl_book = xl_app.Workbooks.Open(
            filename,
            UpdateLinks=0,
            CorruptLoad=XL_REPAIR_FILE,
        )

        for ws_com in xl_book.Sheets:
            sheet_name = ws_com.Name
            # Sensible defaults; rule handler may override.
            ws_com.PageSetup.PrintTitleRows = "$1:$1"
            fit_single_page(ws_com)
            _apply_matching_rule(sheet_name, ws_com, xl_app, filename)

        # Land the user on the Index when the file is opened
        try:
            xl_book.Sheets("Index").Activate()
        except Exception:
            pass

        # SaveAs writes a fresh xlsx from Excel's clean in-memory model.
        # Remove any leftover temp from a previous failed run.
        if os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except Exception:
                pass

        xl_book.SaveAs(temp_filename, FileFormat=XL_OPEN_XML_WORKBOOK)
        xl_book.Close(False)
        xl_book = None

        # Atomically replace the original with the clean file
        os.replace(temp_filename, filename)

    except Exception as exc:
        print(f"[BApagebreaks] COM step failed: {exc}")
        # Clean up temp if it exists so we don't leave junk behind
        if os.path.exists(temp_filename):
            try:
                os.remove(temp_filename)
            except Exception:
                pass
        raise
    finally:
        if xl_book is not None:
            try:
                xl_book.Close(False)
            except Exception:
                pass
        if xl_app is not None:
            try:
                xl_app.Quit()
            except Exception:
                pass
        gc.collect()
        if pythoncom:
            pythoncom.CoUninitialize()
        # Brief pause so Windows fully releases the file handle
        time.sleep(0.5)


# Example usage:
# process_pagebreaks(r"C:\path\to\workbook.xlsx", "ignored.pdf")
